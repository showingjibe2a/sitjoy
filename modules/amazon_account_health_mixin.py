# -*- coding: utf-8 -*-
"""Amazon 账户健康管理 Mixin"""

from urllib.parse import parse_qs
from datetime import datetime
import io
import cgi

try:
    from openpyxl import Workbook, load_workbook
    _openpyxl_import_error = None
except Exception as e:
    Workbook = None
    load_workbook = None
    _openpyxl_import_error = str(e)


class AmazonAccountHealthMixin:
    def _normalize_datetime_text(self, value):
        text = ('' if value is None else str(value)).strip()
        if not text:
            return None
        formats = (
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%dT%H:%M'
        )
        for fmt in formats:
            try:
                dt = datetime.strptime(text, fmt)
                return dt.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                continue
        return None

    def _ensure_amazon_account_health_table(self):
        self._amazon_account_health_ready = True
        self._set_schema_marker_ready('amazon_account_health_v1')

    def handle_amazon_account_health_api(self, environ, method, start_response):
        """Amazon 账户健康管理 API（CRUD + 图表）"""
        try:
            self._ensure_amazon_account_health_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            int_fields = [
                'account_health_rating',
                'suspected_ip_infringement',
                'intellectual_property_complaints',
                'authenticity_customer_complaints',
                'condition_customer_complaints',
                'food_safety_issues',
                'listing_policy_violations',
                'restricted_product_policy_violations',
                'customer_review_policy_violations',
                'other_policy_violations',
                'regulatory_compliance_issues'
            ]
            percent_fields = [
                'order_defect_rate',
                'negative_feedback_rate',
                'a_to_z_rate',
                'chargeback_rate',
                'late_shipment_rate',
                'pre_fulfillment_cancel_rate',
                'valid_tracking_rate',
                'on_time_delivery_rate'
            ]

            if method == 'GET':
                mode = (query_params.get('mode', [''])[0] or '').strip().lower()
                keyword = (query_params.get('q', [''])[0] or '').strip()
                shop_id = self._parse_int((query_params.get('shop_id', [''])[0] or '').strip())
                start_date = self._parse_date_str((query_params.get('start_date', [''])[0] or '').strip())
                end_date = self._parse_date_str((query_params.get('end_date', [''])[0] or '').strip())

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT id FROM platform_types
                            WHERE LOWER(TRIM(name))='amazon'
                            ORDER BY id ASC
                            LIMIT 1
                            """
                        )
                        amazon_platform = cur.fetchone() or {}
                        amazon_platform_id = amazon_platform.get('id')
                        if not amazon_platform_id:
                            return self.send_json({'status': 'success', 'items': []}, start_response)

                        if mode == 'chart':
                            if not shop_id:
                                return self.send_json({'status': 'error', 'message': 'Missing shop_id'}, start_response)
                            cur.execute(
                                "SELECT id FROM shops WHERE id=%s AND platform_type_id=%s",
                                (shop_id, amazon_platform_id)
                            )
                            selected_shop = cur.fetchone()
                            if not selected_shop:
                                return self.send_json({'status': 'error', 'message': 'Shop is not Amazon platform'}, start_response)
                            sql = [
                                """
                                SELECT DATE(a.record_datetime) AS record_date,
                                       ROUND(AVG(a.account_health_rating), 2) AS account_health_rating,
                                       ROUND(AVG(a.order_defect_rate), 4) AS order_defect_rate,
                                       ROUND(AVG(a.late_shipment_rate), 4) AS late_shipment_rate,
                                       ROUND(AVG(a.pre_fulfillment_cancel_rate), 4) AS pre_fulfillment_cancel_rate,
                                       ROUND(AVG(a.valid_tracking_rate), 4) AS valid_tracking_rate,
                                       ROUND(AVG(a.on_time_delivery_rate), 4) AS on_time_delivery_rate
                                FROM amazon_account_health a
                                LEFT JOIN shops s ON s.id = a.shop_id
                                WHERE a.shop_id=%s AND s.platform_type_id=%s
                                """
                            ]
                            params = [shop_id, amazon_platform_id]
                            if start_date:
                                sql.append("AND DATE(a.record_datetime) >= %s")
                                params.append(start_date)
                            if end_date:
                                sql.append("AND DATE(a.record_datetime) <= %s")
                                params.append(end_date)
                            sql.append("GROUP BY DATE(a.record_datetime) ORDER BY DATE(a.record_datetime) ASC")
                            cur.execute("\n".join(sql), params)
                            rows = cur.fetchall()
                            return self.send_json({'status': 'success', 'items': rows}, start_response)

                        sql = [
                            """
                            SELECT a.*, s.shop_name
                            FROM amazon_account_health a
                            LEFT JOIN shops s ON s.id = a.shop_id
                            WHERE s.platform_type_id=%s
                            """
                        ]
                        params = [amazon_platform_id]
                        if shop_id:
                            sql.append("AND a.shop_id=%s")
                            params.append(shop_id)
                        if start_date:
                            sql.append("AND DATE(a.record_datetime) >= %s")
                            params.append(start_date)
                        if end_date:
                            sql.append("AND DATE(a.record_datetime) <= %s")
                            params.append(end_date)
                        if keyword:
                            sql.append("AND (s.shop_name LIKE %s OR a.remark LIKE %s)")
                            params.extend([f"%{keyword}%", f"%{keyword}%"])
                        sql.append("ORDER BY a.record_datetime DESC, a.id DESC")
                        cur.execute("\n".join(sql), params)
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                shop_id = self._parse_int(data.get('shop_id'))
                if not shop_id:
                    return self.send_json({'status': 'error', 'message': 'Missing shop_id'}, start_response)

                values = {}
                for key in int_fields:
                    parsed = self._parse_int(data.get(key))
                    if parsed is None:
                        return self.send_json({'status': 'error', 'message': f'Missing or invalid {key}'}, start_response)
                    values[key] = parsed
                for key in percent_fields:
                    parsed = self._parse_float(data.get(key))
                    if parsed is None:
                        return self.send_json({'status': 'error', 'message': f'Missing or invalid {key}'}, start_response)
                    values[key] = parsed

                record_datetime = self._normalize_datetime_text(data.get('record_datetime')) or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                remark = (data.get('remark') or '').strip()[:500]

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT s.id
                            FROM shops s
                            JOIN platform_types pt ON pt.id = s.platform_type_id
                            WHERE s.id=%s AND LOWER(TRIM(pt.name))='amazon'
                            """,
                            (shop_id,)
                        )
                        allowed_shop = cur.fetchone()
                        if not allowed_shop:
                            return self.send_json({'status': 'error', 'message': 'Only Amazon platform shop is allowed'}, start_response)
                        cur.execute(
                            """
                            INSERT INTO amazon_account_health (
                                shop_id, account_health_rating,
                                suspected_ip_infringement, intellectual_property_complaints,
                                authenticity_customer_complaints, condition_customer_complaints,
                                food_safety_issues, listing_policy_violations,
                                restricted_product_policy_violations, customer_review_policy_violations,
                                other_policy_violations, regulatory_compliance_issues,
                                order_defect_rate, negative_feedback_rate, a_to_z_rate, chargeback_rate,
                                late_shipment_rate, pre_fulfillment_cancel_rate, valid_tracking_rate, on_time_delivery_rate,
                                record_datetime, remark
                            ) VALUES (
                                %s, %s,
                                %s, %s,
                                %s, %s,
                                %s, %s,
                                %s, %s,
                                %s, %s,
                                %s, %s, %s, %s,
                                %s, %s, %s, %s,
                                %s, %s
                            )
                            """,
                            (
                                shop_id, values['account_health_rating'],
                                values['suspected_ip_infringement'], values['intellectual_property_complaints'],
                                values['authenticity_customer_complaints'], values['condition_customer_complaints'],
                                values['food_safety_issues'], values['listing_policy_violations'],
                                values['restricted_product_policy_violations'], values['customer_review_policy_violations'],
                                values['other_policy_violations'], values['regulatory_compliance_issues'],
                                values['order_defect_rate'], values['negative_feedback_rate'], values['a_to_z_rate'], values['chargeback_rate'],
                                values['late_shipment_rate'], values['pre_fulfillment_cancel_rate'], values['valid_tracking_rate'], values['on_time_delivery_rate'],
                                record_datetime, remark
                            )
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                shop_id = self._parse_int(data.get('shop_id'))
                if not item_id or not shop_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id or shop_id'}, start_response)

                values = {}
                for key in int_fields:
                    parsed = self._parse_int(data.get(key))
                    if parsed is None:
                        return self.send_json({'status': 'error', 'message': f'Missing or invalid {key}'}, start_response)
                    values[key] = parsed
                for key in percent_fields:
                    parsed = self._parse_float(data.get(key))
                    if parsed is None:
                        return self.send_json({'status': 'error', 'message': f'Missing or invalid {key}'}, start_response)
                    values[key] = parsed

                record_datetime = self._normalize_datetime_text(data.get('record_datetime')) or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                remark = (data.get('remark') or '').strip()[:500]

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT s.id
                            FROM shops s
                            JOIN platform_types pt ON pt.id = s.platform_type_id
                            WHERE s.id=%s AND LOWER(TRIM(pt.name))='amazon'
                            """,
                            (shop_id,)
                        )
                        allowed_shop = cur.fetchone()
                        if not allowed_shop:
                            return self.send_json({'status': 'error', 'message': 'Only Amazon platform shop is allowed'}, start_response)
                        cur.execute("SELECT id FROM amazon_account_health WHERE id=%s", (item_id,))
                        exists = cur.fetchone()
                        if not exists:
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                        cur.execute(
                            """
                            UPDATE amazon_account_health
                            SET shop_id=%s,
                                account_health_rating=%s,
                                suspected_ip_infringement=%s,
                                intellectual_property_complaints=%s,
                                authenticity_customer_complaints=%s,
                                condition_customer_complaints=%s,
                                food_safety_issues=%s,
                                listing_policy_violations=%s,
                                restricted_product_policy_violations=%s,
                                customer_review_policy_violations=%s,
                                other_policy_violations=%s,
                                regulatory_compliance_issues=%s,
                                order_defect_rate=%s,
                                negative_feedback_rate=%s,
                                a_to_z_rate=%s,
                                chargeback_rate=%s,
                                late_shipment_rate=%s,
                                pre_fulfillment_cancel_rate=%s,
                                valid_tracking_rate=%s,
                                on_time_delivery_rate=%s,
                                record_datetime=%s,
                                remark=%s
                            WHERE id=%s
                            """,
                            (
                                shop_id,
                                values['account_health_rating'],
                                values['suspected_ip_infringement'],
                                values['intellectual_property_complaints'],
                                values['authenticity_customer_complaints'],
                                values['condition_customer_complaints'],
                                values['food_safety_issues'],
                                values['listing_policy_violations'],
                                values['restricted_product_policy_violations'],
                                values['customer_review_policy_violations'],
                                values['other_policy_violations'],
                                values['regulatory_compliance_issues'],
                                values['order_defect_rate'],
                                values['negative_feedback_rate'],
                                values['a_to_z_rate'],
                                values['chargeback_rate'],
                                values['late_shipment_rate'],
                                values['pre_fulfillment_cancel_rate'],
                                values['valid_tracking_rate'],
                                values['on_time_delivery_rate'],
                                record_datetime,
                                remark,
                                item_id
                            )
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_account_health WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            print("AmazonAccountHealth API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_account_health_template_api(self, environ, method, start_response):
        """Amazon 账户健康模板下载"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter

            self._ensure_amazon_account_health_table()
            shop_names = []
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT s.shop_name
                        FROM shops s
                        JOIN platform_types pt ON pt.id = s.platform_type_id
                        WHERE LOWER(TRIM(pt.name))='amazon'
                        ORDER BY s.shop_name
                        """
                    )
                    shop_names = [str(row.get('shop_name') or '').strip() for row in (cur.fetchall() or []) if str(row.get('shop_name') or '').strip()]

            wb = Workbook()
            ws = wb.active
            ws.title = 'amazon_account_health'

            headers = [
                '店铺*', '记录日期时间*', '账户状况评级*',
                '涉嫌侵犯知识产权*', '知识产权投诉*', '商品真实性买家投诉*', '商品状况买家投诉*',
                '食品和商品安全问题*', '上架政策违规*', '违反受限商品政策*', '违反买家商品评论政策*', '其他违反政策*', '监管合规性*',
                '订单缺陷率(%)*', '负面反馈(%)*', 'A-to-z(%)*', '信用卡拒付(%)*',
                '迟发率(%)*', '配送前取消率(%)*', '有效追踪率(%)*', '准时交货率(%)*',
                '备注'
            ]
            ws.append(headers)

            sample_shop = shop_names[0] if shop_names else ''
            ws.append([
                sample_shop, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 260,
                0, 0, 0, 0,
                0, 0, 0, 0, 0, 0,
                0.35, 0.00, 0.00, 0.00,
                1.20, 0.80, 97.50, 95.20,
                '示例行（请勿修改，此行不会导入）'
            ])

            for cell in ws[1]:
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in ws[2]:
                cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                cell.font = Font(italic=True, color='888888')

            widths = [20, 20, 14, 16, 14, 18, 18, 18, 14, 18, 20, 14, 14, 14, 12, 12, 12, 12, 14, 14, 14, 28]
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            options_ws = wb.create_sheet('options')
            options_ws.sheet_state = 'hidden'
            options_ws.cell(row=1, column=1, value='amazon_shop_name')
            for idx, name in enumerate(shop_names, start=2):
                options_ws.cell(row=idx, column=1, value=name)

            if shop_names:
                shop_validation = DataValidation(type='list', formula1=f'=options!$A$2:$A${len(shop_names) + 1}', allow_blank=False)
                ws.add_data_validation(shop_validation)
                for row_idx in range(3, 500):
                    shop_validation.add(f'A{row_idx}')

            ws.freeze_panes = 'A3'
            return self._send_excel_workbook(wb, 'amazon_account_health_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_account_health_import_api(self, environ, method, start_response):
        """Amazon 账户健康批量导入"""
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            headers = [str(cell.value or '').strip() for cell in ws[1]]
            header_map = {name: idx for idx, name in enumerate(headers)}

            def get_cell(row, name):
                idx = header_map.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx].value

            required_headers = [
                '店铺*', '记录日期时间*', '账户状况评级*',
                '涉嫌侵犯知识产权*', '知识产权投诉*', '商品真实性买家投诉*', '商品状况买家投诉*',
                '食品和商品安全问题*', '上架政策违规*', '违反受限商品政策*', '违反买家商品评论政策*', '其他违反政策*', '监管合规性*',
                '订单缺陷率(%)*', '负面反馈(%)*', 'A-to-z(%)*', '信用卡拒付(%)*',
                '迟发率(%)*', '配送前取消率(%)*', '有效追踪率(%)*', '准时交货率(%)*'
            ]
            for col_name in required_headers:
                if col_name not in header_map:
                    return self.send_json({'status': 'error', 'message': f'模板缺少列: {col_name}'}, start_response)

            self._ensure_amazon_account_health_table()
            created = 0
            updated = 0
            unchanged = 0
            errors = []

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT s.id, s.shop_name
                        FROM shops s
                        JOIN platform_types pt ON pt.id = s.platform_type_id
                        WHERE LOWER(TRIM(pt.name))='amazon'
                        """
                    )
                    shop_rows = cur.fetchall() or []
                    shop_map = {str(row.get('shop_name') or '').strip(): int(row.get('id')) for row in shop_rows if row.get('id')}

                for row_idx in range(2, ws.max_row + 1):
                    if row_idx == 2:
                        continue
                    row = ws[row_idx]
                    if not any(cell.value is not None and str(cell.value).strip() for cell in row):
                        continue
                    try:
                        shop_name = str(get_cell(row, '店铺*') or '').strip()
                        shop_id = shop_map.get(shop_name)
                        if not shop_id:
                            raise ValueError(f'店铺不存在或非Amazon平台: {shop_name}')

                        record_datetime = self._normalize_datetime_text(get_cell(row, '记录日期时间*'))
                        if not record_datetime:
                            raise ValueError('记录日期时间格式错误，请使用 YYYY-MM-DD HH:MM:SS 或 YYYY-MM-DDTHH:MM')

                        parsed = {
                            'account_health_rating': self._parse_int(get_cell(row, '账户状况评级*')),
                            'suspected_ip_infringement': self._parse_int(get_cell(row, '涉嫌侵犯知识产权*')),
                            'intellectual_property_complaints': self._parse_int(get_cell(row, '知识产权投诉*')),
                            'authenticity_customer_complaints': self._parse_int(get_cell(row, '商品真实性买家投诉*')),
                            'condition_customer_complaints': self._parse_int(get_cell(row, '商品状况买家投诉*')),
                            'food_safety_issues': self._parse_int(get_cell(row, '食品和商品安全问题*')),
                            'listing_policy_violations': self._parse_int(get_cell(row, '上架政策违规*')),
                            'restricted_product_policy_violations': self._parse_int(get_cell(row, '违反受限商品政策*')),
                            'customer_review_policy_violations': self._parse_int(get_cell(row, '违反买家商品评论政策*')),
                            'other_policy_violations': self._parse_int(get_cell(row, '其他违反政策*')),
                            'regulatory_compliance_issues': self._parse_int(get_cell(row, '监管合规性*')),
                            'order_defect_rate': self._parse_float(get_cell(row, '订单缺陷率(%)*')),
                            'negative_feedback_rate': self._parse_float(get_cell(row, '负面反馈(%)*')),
                            'a_to_z_rate': self._parse_float(get_cell(row, 'A-to-z(%)*')),
                            'chargeback_rate': self._parse_float(get_cell(row, '信用卡拒付(%)*')),
                            'late_shipment_rate': self._parse_float(get_cell(row, '迟发率(%)*')),
                            'pre_fulfillment_cancel_rate': self._parse_float(get_cell(row, '配送前取消率(%)*')),
                            'valid_tracking_rate': self._parse_float(get_cell(row, '有效追踪率(%)*')),
                            'on_time_delivery_rate': self._parse_float(get_cell(row, '准时交货率(%)*')),
                            'remark': str(get_cell(row, '备注') or '').strip()[:500]
                        }

                        for key, value in parsed.items():
                            if key == 'remark':
                                continue
                            if value is None:
                                raise ValueError(f'{key} 为空或格式错误')

                        with conn.cursor() as cur:
                            cur.execute(
                                """
                                SELECT * FROM amazon_account_health
                                WHERE shop_id=%s AND record_datetime=%s
                                ORDER BY id ASC
                                LIMIT 1
                                """,
                                (shop_id, record_datetime)
                            )
                            existing = cur.fetchone()

                            if existing:
                                cur.execute(
                                    """
                                    UPDATE amazon_account_health
                                    SET account_health_rating=%s,
                                        suspected_ip_infringement=%s,
                                        intellectual_property_complaints=%s,
                                        authenticity_customer_complaints=%s,
                                        condition_customer_complaints=%s,
                                        food_safety_issues=%s,
                                        listing_policy_violations=%s,
                                        restricted_product_policy_violations=%s,
                                        customer_review_policy_violations=%s,
                                        other_policy_violations=%s,
                                        regulatory_compliance_issues=%s,
                                        order_defect_rate=%s,
                                        negative_feedback_rate=%s,
                                        a_to_z_rate=%s,
                                        chargeback_rate=%s,
                                        late_shipment_rate=%s,
                                        pre_fulfillment_cancel_rate=%s,
                                        valid_tracking_rate=%s,
                                        on_time_delivery_rate=%s,
                                        remark=%s
                                    WHERE id=%s
                                    """,
                                    (
                                        parsed['account_health_rating'],
                                        parsed['suspected_ip_infringement'],
                                        parsed['intellectual_property_complaints'],
                                        parsed['authenticity_customer_complaints'],
                                        parsed['condition_customer_complaints'],
                                        parsed['food_safety_issues'],
                                        parsed['listing_policy_violations'],
                                        parsed['restricted_product_policy_violations'],
                                        parsed['customer_review_policy_violations'],
                                        parsed['other_policy_violations'],
                                        parsed['regulatory_compliance_issues'],
                                        parsed['order_defect_rate'],
                                        parsed['negative_feedback_rate'],
                                        parsed['a_to_z_rate'],
                                        parsed['chargeback_rate'],
                                        parsed['late_shipment_rate'],
                                        parsed['pre_fulfillment_cancel_rate'],
                                        parsed['valid_tracking_rate'],
                                        parsed['on_time_delivery_rate'],
                                        parsed['remark'],
                                        existing.get('id')
                                    )
                                )
                                if cur.rowcount:
                                    updated += 1
                                else:
                                    unchanged += 1
                            else:
                                cur.execute(
                                    """
                                    INSERT INTO amazon_account_health (
                                        shop_id, account_health_rating,
                                        suspected_ip_infringement, intellectual_property_complaints,
                                        authenticity_customer_complaints, condition_customer_complaints,
                                        food_safety_issues, listing_policy_violations,
                                        restricted_product_policy_violations, customer_review_policy_violations,
                                        other_policy_violations, regulatory_compliance_issues,
                                        order_defect_rate, negative_feedback_rate, a_to_z_rate, chargeback_rate,
                                        late_shipment_rate, pre_fulfillment_cancel_rate, valid_tracking_rate, on_time_delivery_rate,
                                        record_datetime, remark
                                    ) VALUES (
                                        %s, %s,
                                        %s, %s,
                                        %s, %s,
                                        %s, %s,
                                        %s, %s,
                                        %s, %s,
                                        %s, %s, %s, %s,
                                        %s, %s, %s, %s,
                                        %s, %s
                                    )
                                    """,
                                    (
                                        shop_id, parsed['account_health_rating'],
                                        parsed['suspected_ip_infringement'], parsed['intellectual_property_complaints'],
                                        parsed['authenticity_customer_complaints'], parsed['condition_customer_complaints'],
                                        parsed['food_safety_issues'], parsed['listing_policy_violations'],
                                        parsed['restricted_product_policy_violations'], parsed['customer_review_policy_violations'],
                                        parsed['other_policy_violations'], parsed['regulatory_compliance_issues'],
                                        parsed['order_defect_rate'], parsed['negative_feedback_rate'], parsed['a_to_z_rate'], parsed['chargeback_rate'],
                                        parsed['late_shipment_rate'], parsed['pre_fulfillment_cancel_rate'], parsed['valid_tracking_rate'], parsed['on_time_delivery_rate'],
                                        record_datetime, parsed['remark']
                                    )
                                )
                                created += 1
                    except Exception as row_error:
                        errors.append({'row': row_idx, 'error': str(row_error)})

            return self.send_json({
                'status': 'success',
                'created': created,
                'updated': updated,
                'unchanged': unchanged,
                'errors': errors
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
