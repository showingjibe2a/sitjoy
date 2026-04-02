import ast
import csv
import cgi
import io
import json
import mimetypes
import os
import re
from datetime import datetime, date, timedelta
from urllib.parse import parse_qs, quote

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.comments import Comment
    from openpyxl.worksheet.datavalidation import DataValidation
    _openpyxl_import_error = None
except Exception as _e:
    Workbook = None
    load_workbook = None
    PatternFill = None
    Font = None
    Alignment = None
    Border = None
    Side = None
    Comment = None
    DataValidation = None
    _openpyxl_import_error = _e


class LogisticsInTransitMixin:
    pass

    def _get_logistics_link_root_bytes(self):
        resources_root = self._join_resources('')
        resources_parent = os.path.dirname(resources_root)
        return os.path.join(resources_parent, self._safe_fsencode('『物流仓储关联文件』'))

    def _rename_logistics_bl_folder(self, old_no, new_no):
        old_name = (old_no or '').strip()
        new_name = (new_no or '').strip()
        if not old_name or not new_name or old_name == new_name:
            if new_name:
                self._ensure_logistics_bl_folder(new_name)
            return
        root = self._get_logistics_link_root_bytes()
        if not os.path.exists(root):
            os.makedirs(root, exist_ok=True)
        old_path = os.path.join(root, self._safe_fsencode(old_name))
        new_path = os.path.join(root, self._safe_fsencode(new_name))
        if os.path.exists(old_path):
            if os.path.exists(new_path):
                raise RuntimeError(f'目标提单目录已存在: {new_name}')
            os.rename(old_path, new_path)
        else:
            self._ensure_logistics_bl_folder(new_name)

    def _resolve_logistics_doc_folder(self, transit_id, doc_type):
        doc_kind = (doc_type or '').strip().lower()
        if doc_kind not in ('declaration', 'clearance'):
            raise RuntimeError('Invalid doc_type')
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT bill_of_lading_no FROM logistics_in_transit WHERE id=%s LIMIT 1", (transit_id,))
                row = cur.fetchone() or {}
        bill_no = (row.get('bill_of_lading_no') or '').strip()
        if not bill_no:
            raise RuntimeError('请先填写提单号后再操作资料文件')
        self._ensure_logistics_bl_folder(bill_no)
        sub_name = '报关资料' if doc_kind == 'declaration' else '清关资料'
        parent = os.path.join(self._get_logistics_link_root_bytes(), self._safe_fsencode(bill_no))
        folder = os.path.join(parent, self._safe_fsencode(sub_name))
        if not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
        return folder

    def _calc_qty_consistent_from_items(self, items):
        if not isinstance(items, list) or not items:
            return 0
        for entry in items:
            if not isinstance(entry, dict):
                continue
            shipped_qty = self._parse_int(entry.get('shipped_qty')) or 0
            listed_raw = entry.get('listed_qty')
            listed_qty = shipped_qty if listed_raw in (None, '') else (self._parse_int(listed_raw) or 0)
            if shipped_qty != listed_qty:
                return 0
        return 1

    def _refresh_transit_qty_consistent(self, transit_id):
        transit_id = self._parse_int(transit_id)
        if not transit_id:
            return
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT shipped_qty, listed_qty FROM logistics_in_transit_items WHERE transit_id=%s",
                    (transit_id,)
                )
                rows = cur.fetchall() or []
                qty_consistent = 1
                if not rows:
                    qty_consistent = 0
                else:
                    for row in rows:
                        shipped_qty = self._parse_int((row or {}).get('shipped_qty')) or 0
                        listed_qty = self._parse_int((row or {}).get('listed_qty'))
                        listed_qty = shipped_qty if listed_qty is None else listed_qty
                        if shipped_qty != listed_qty:
                            qty_consistent = 0
                            break
                cur.execute("UPDATE logistics_in_transit SET qty_consistent=%s WHERE id=%s", (qty_consistent, transit_id))

    def handle_logistics_in_transit_api(self, environ, method, start_response):
        perf_ctx = self._perf_begin('logistics_in_transit_api', environ, {'entry_method': method})
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            action = (query_params.get('action', [''])[0] or '').strip().lower()
            self._perf_mark(perf_ctx, f'parse_query_action:{action or "none"}')

            if not (method == 'GET' and action == 'options'):
                if not self.__class__._schema_ready_cache.get('logistics'):
                    self._ensure_logistics_tables()
                self._perf_mark(perf_ctx, 'ensure_logistics_tables')

            def _to_bool_flag(value):
                return 1 if str(value or '').strip().lower() in ('1', 'true', 'yes', 'on') else 0

            def _normalize_date(value):
                text = ('' if value is None else str(value)).strip()
                if not text:
                    return None
                for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S'):
                    try:
                        return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
                    except Exception:
                        continue
                return None

            if method == 'GET':
                if action == 'options':
                    scope = (query_params.get('scope', ['all'])[0] or 'all').strip().lower()
                    option_limit = max(100, min(self._parse_int(query_params.get('order_product_limit', ['600'])[0]) or 600, 1200))
                    with self._get_db_connection() as conn:
                        self._perf_mark(perf_ctx, 'db_connected_for_options')
                        with conn.cursor() as cur:
                            cur.execute("SELECT id, factory_name FROM logistics_factories ORDER BY factory_name ASC")
                            factories = cur.fetchall() or []
                            cur.execute("SELECT id, forwarder_name FROM logistics_forwarders ORDER BY forwarder_name ASC")
                            forwarders = cur.fetchall() or []
                            cur.execute("SELECT id, region_name, sort_order FROM logistics_destination_regions ORDER BY sort_order ASC, id ASC")
                            destination_regions = cur.fetchall() or []
                            cur.execute("SELECT id, warehouse_name, destination_region_id FROM logistics_overseas_warehouses WHERE COALESCE(is_enabled,1)=1 ORDER BY warehouse_name ASC")
                            warehouses = cur.fetchall() or []
                            order_products = []
                            if scope in ('all', 'with_order_products'):
                                cur.execute("SELECT id, sku FROM order_products ORDER BY sku ASC LIMIT %s", (option_limit,))
                                order_products = cur.fetchall() or []
                    self._perf_mark(perf_ctx, 'get_options_payload')
                    return self.send_json({
                        'status': 'success',
                        'factories': factories,
                        'forwarders': forwarders,
                        'destination_regions': destination_regions,
                        'warehouses': warehouses,
                        'order_products': order_products
                    }, start_response)

                if action == 'export_details':
                    selected_ids = []
                    for raw in query_params.get('ids', []):
                        for token in re.split(r'[,，;；\s]+', str(raw or '').strip()):
                            if not token:
                                continue
                            tid = self._parse_int(token)
                            if tid and tid not in selected_ids:
                                selected_ids.append(tid)
                    if not selected_ids:
                        return self.send_json({'status': 'error', 'message': '请先选择要导出的记录'}, start_response)

                    field_defs = {
                        'id': ('记录ID', lambda r: r.get('id') or ''),
                        'logistics_box_no': ('箱号', lambda r: r.get('logistics_box_no') or ''),
                        'bill_of_lading_no': ('提单号', lambda r: r.get('bill_of_lading_no') or ''),
                        'factory_name': ('工厂', lambda r: r.get('factory_name') or ''),
                        'forwarder_name': ('货代', lambda r: r.get('forwarder_name') or ''),
                        'destination_region_name': ('目的区域', lambda r: r.get('destination_region_name') or ''),
                        'destination_warehouse_name': ('目的仓库', lambda r: r.get('destination_warehouse_name') or ''),
                        'expected_listed_date_latest': ('预计上架时间', lambda r: (r.get('expected_listed_date_latest').strftime('%Y-%m-%d') if hasattr(r.get('expected_listed_date_latest'), 'strftime') else (r.get('expected_listed_date_latest') or ''))),
                        'inventory_registered': ('已登记上架', lambda r: '是' if self._parse_int(r.get('inventory_registered')) else '否'),
                        'listed_date': ('实际上架日期', lambda r: (r.get('listed_date').strftime('%Y-%m-%d') if hasattr(r.get('listed_date'), 'strftime') else (r.get('listed_date') or ''))),
                        'qty_verified': ('已核对上架数量', lambda r: '是' if self._parse_int(r.get('qty_verified')) else '否'),
                        'qty_consistent': ('数量一致', lambda r: '是' if self._parse_int(r.get('qty_consistent')) else '否'),
                        'remark': ('备注', lambda r: r.get('remark') or ''),
                        'sku': ('下单SKU', lambda r: r.get('sku') or ''),
                        'shipped_qty': ('发货数量', lambda r: r.get('shipped_qty') if r.get('shipped_qty') is not None else ''),
                        'listed_qty': ('上架数量', lambda r: r.get('listed_qty') if r.get('listed_qty') is not None else ''),
                    }

                    requested_fields = []
                    for raw in query_params.get('fields', []):
                        for token in re.split(r'[,，;；\s]+', str(raw or '').strip()):
                            key = (token or '').strip()
                            if key and key in field_defs and key not in requested_fields:
                                requested_fields.append(key)
                    if not requested_fields:
                        requested_fields = [
                            'logistics_box_no', 'bill_of_lading_no', 'factory_name', 'forwarder_name',
                            'destination_region_name', 'destination_warehouse_name', 'expected_listed_date_latest',
                            'sku', 'shipped_qty', 'listed_qty', 'remark'
                        ]

                    placeholders = ','.join(['%s'] * len(selected_ids))
                    sql = f"""
                        SELECT
                            t.id,
                            t.logistics_box_no,
                            t.bill_of_lading_no,
                            t.expected_listed_date_latest,
                            t.inventory_registered,
                            t.listed_date,
                            t.qty_verified,
                            t.qty_consistent,
                            t.remark,
                            f.factory_name,
                            fw.forwarder_name,
                            dr.region_name AS destination_region_name,
                            ow.warehouse_name AS destination_warehouse_name,
                            op.sku,
                            li.shipped_qty,
                            li.listed_qty,
                            li.id AS item_id
                        FROM logistics_in_transit t
                        LEFT JOIN logistics_factories f ON f.id=t.factory_id
                        LEFT JOIN logistics_forwarders fw ON fw.id=t.forwarder_id
                        LEFT JOIN logistics_destination_regions dr ON dr.id=t.destination_region_id
                        LEFT JOIN logistics_overseas_warehouses ow ON ow.id=t.destination_warehouse_id
                        LEFT JOIN logistics_in_transit_items li ON li.transit_id=t.id
                        LEFT JOIN order_products op ON op.id=li.order_product_id
                        WHERE t.id IN ({placeholders})
                        ORDER BY t.id DESC, li.id ASC
                    """
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(sql, selected_ids)
                            export_rows = cur.fetchall() or []

                    if not export_rows:
                        return self.send_json({'status': 'error', 'message': '未找到可导出的数据'}, start_response)

                    output = io.StringIO(newline='')
                    writer = csv.writer(output)
                    writer.writerow([field_defs[k][0] for k in requested_fields])
                    for row in export_rows:
                        writer.writerow([field_defs[k][1](row) for k in requested_fields])
                    content = output.getvalue().encode('utf-8-sig')

                    filename = f"在途物流明细导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                    headers = [
                        ('Content-Type', 'text/csv; charset=utf-8'),
                        ('Content-Disposition', f"attachment; filename*=UTF-8''{quote(filename)}"),
                        ('Content-Length', str(len(content))),
                    ]
                    start_response('200 OK', headers)
                    return [content]

                item_id = self._parse_int(query_params.get('id', [''])[0])
                keyword = (query_params.get('q', [''])[0] or '').strip()
                page = self._parse_int(query_params.get('page', ['1'])[0]) or 1
                page_size = self._parse_int(query_params.get('page_size', ['50'])[0]) or 50
                page = max(1, page)
                page_size = max(20, min(page_size, 200))
                offset = (page - 1) * page_size

                with self._get_db_connection() as conn:
                    self._perf_mark(perf_ctx, 'db_connected_for_get')
                    with conn.cursor() as cur:
                        def _fetch_name_maps_payload():
                            cur.execute("SELECT id, factory_name FROM logistics_factories")
                            factories = cur.fetchall() or []
                            cur.execute("SELECT id, forwarder_name FROM logistics_forwarders")
                            forwarders = cur.fetchall() or []
                            cur.execute("SELECT id, region_name FROM logistics_destination_regions")
                            destination_regions = cur.fetchall() or []
                            cur.execute("SELECT id, warehouse_name FROM logistics_overseas_warehouses")
                            warehouses = cur.fetchall() or []
                            return {
                                'status': 'success',
                                'factory_name_map': {int(r.get('id')): str(r.get('factory_name') or '') for r in factories if r.get('id')},
                                'forwarder_name_map': {int(r.get('id')): str(r.get('forwarder_name') or '') for r in forwarders if r.get('id')},
                                'destination_region_name_map': {int(r.get('id')): str(r.get('region_name') or '') for r in destination_regions if r.get('id')},
                                'warehouse_name_map': {int(r.get('id')): str(r.get('warehouse_name') or '') for r in warehouses if r.get('id')},
                            }

                        name_maps = _fetch_name_maps_payload()
                        factory_name_map = name_maps.get('factory_name_map') or {}
                        forwarder_name_map = name_maps.get('forwarder_name_map') or {}
                        destination_region_name_map = name_maps.get('destination_region_name_map') or {}
                        warehouse_name_map = name_maps.get('warehouse_name_map') or {}

                        sql = """
                            SELECT t.id, t.factory_id, t.factory_ship_date_initial, t.factory_ship_date_previous, t.factory_ship_date_latest,
                                t.forwarder_id, t.logistics_box_no,
                                   t.etd_initial, t.etd_previous, t.etd_latest,
                                   t.eta_initial, t.eta_previous, t.eta_latest,
                                   t.arrival_port_date, t.expected_warehouse_date, t.expected_listed_date_initial, t.expected_listed_date_latest, t.listed_date,
                                   t.shipping_company, t.vessel_voyage, t.bill_of_lading_no,
                                          t.remark,
                                t.declaration_docs_provided, t.inventory_registered, t.clearance_docs_provided, t.qty_verified, t.qty_consistent,
                                              t.port_of_loading, t.port_of_destination, t.destination_region_id, t.destination_warehouse_id,
                                              t.confirmed_boxed_qty, t.inbound_order_no,
                                   t.created_at, t.updated_at
                            FROM logistics_in_transit t
                        """
                        params = []
                        filters = []
                        if item_id:
                            filters.append('t.id=%s')
                            params.append(item_id)
                        if keyword:
                            like = f"%{keyword}%"
                            search_clauses = [
                                't.logistics_box_no LIKE %s',
                                't.bill_of_lading_no LIKE %s',
                            ]
                            params.extend([like, like])

                            keyword_lower = keyword.lower()
                            matched_factory_ids = [int(fid) for fid, name in factory_name_map.items() if keyword_lower in str(name or '').lower()]
                            matched_forwarder_ids = [int(fid) for fid, name in forwarder_name_map.items() if keyword_lower in str(name or '').lower()]
                            matched_region_ids = [int(rid) for rid, name in destination_region_name_map.items() if keyword_lower in str(name or '').lower()]

                            if matched_factory_ids:
                                placeholders = ','.join(['%s'] * len(matched_factory_ids))
                                search_clauses.append(f"t.factory_id IN ({placeholders})")
                                params.extend(matched_factory_ids)
                            if matched_forwarder_ids:
                                placeholders = ','.join(['%s'] * len(matched_forwarder_ids))
                                search_clauses.append(f"t.forwarder_id IN ({placeholders})")
                                params.extend(matched_forwarder_ids)
                            if matched_region_ids:
                                placeholders = ','.join(['%s'] * len(matched_region_ids))
                                search_clauses.append(f"t.destination_region_id IN ({placeholders})")
                                params.extend(matched_region_ids)

                            filters.append('(' + ' OR '.join(search_clauses) + ')')
                        where_sql = (' WHERE ' + ' AND '.join(filters)) if filters else ''

                        total = None
                        if not item_id:
                            count_sql = "SELECT COUNT(*) AS total FROM logistics_in_transit t" + where_sql
                            cur.execute(count_sql, params)
                            total_row = cur.fetchone() or {}
                            total = int(total_row.get('total') or 0)
                            self._perf_mark(perf_ctx, 'list_count_query')
                            cur.execute(sql + where_sql + ' ORDER BY t.id DESC LIMIT %s OFFSET %s', params + [page_size, offset])
                        else:
                            cur.execute(sql + where_sql + ' ORDER BY t.id DESC', params)
                        rows = cur.fetchall() or []
                        self._perf_mark(perf_ctx, 'list_rows_query')

                        if not rows:
                            if item_id:
                                return self.send_json({'status': 'success', 'item': None}, start_response)
                            return self.send_json({'status': 'success', 'items': [], 'page': page, 'page_size': page_size, 'total': 0}, start_response)

                        if item_id:
                            ids = [int(r.get('id')) for r in rows if r.get('id')]
                            item_map = {}
                            if ids:
                                placeholders = ','.join(['%s'] * len(ids))
                                cur.execute(
                                    f"""
                                    SELECT li.transit_id, li.order_product_id, li.shipped_qty, li.listed_qty, op.sku
                                    FROM logistics_in_transit_items li
                                    JOIN order_products op ON op.id = li.order_product_id
                                    WHERE li.transit_id IN ({placeholders})
                                    ORDER BY li.transit_id, li.id
                                    """,
                                    ids
                                )
                                item_rows = cur.fetchall() or []
                                for irow in item_rows:
                                    transit_id = int(irow.get('transit_id') or 0)
                                    if not transit_id:
                                        continue
                                    item_map.setdefault(transit_id, []).append({
                                        'order_product_id': irow.get('order_product_id'),
                                        'sku': irow.get('sku'),
                                        'shipped_qty': irow.get('shipped_qty'),
                                        'listed_qty': irow.get('listed_qty')
                                    })
                                self._perf_mark(perf_ctx, 'load_item_detail_rows')

                            for row in rows:
                                row['items'] = item_map.get(int(row.get('id') or 0), [])
                                row['listed_status'] = '已上架' if row.get('listed_date') else '未上架'

                        for row in rows:
                            fid = self._parse_int(row.get('factory_id'))
                            fwid = self._parse_int(row.get('forwarder_id'))
                            wid = self._parse_int(row.get('destination_warehouse_id'))
                            row['factory_name'] = factory_name_map.get(fid, '') if fid else ''
                            row['forwarder_name'] = forwarder_name_map.get(fwid, '') if fwid else ''
                            region_id = self._parse_int(row.get('destination_region_id'))
                            row['destination_region_name'] = destination_region_name_map.get(region_id, '') if region_id else ''
                            row['destination_warehouse_name'] = warehouse_name_map.get(wid, '') if wid else ''
                        self._perf_mark(perf_ctx, 'rows_transform')

                if item_id:
                    return self.send_json({'status': 'success', 'item': rows[0]}, start_response)
                return self.send_json({'status': 'success', 'items': rows, 'page': page, 'page_size': page_size, 'total': total if total is not None else len(rows)}, start_response)

            data = self._read_json_body(environ)

            if method == 'POST' and action == 'verify_qty':
                item_id = self._parse_int(data.get('id'))
                verify_items = data.get('items') if isinstance(data.get('items'), list) else []
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)

                listed_by_order = {}
                for entry in verify_items:
                    if not isinstance(entry, dict):
                        continue
                    order_product_id = self._parse_int(entry.get('order_product_id'))
                    if not order_product_id:
                        continue
                    listed_qty = self._parse_int(entry.get('listed_qty'))
                    listed_by_order[order_product_id] = max(0, listed_qty or 0)

                with self._get_db_connection() as conn:
                    self._perf_mark(perf_ctx, 'db_connected_for_verify_qty')
                    with conn.cursor() as cur:
                        cur.execute("SELECT id, inventory_registered FROM logistics_in_transit WHERE id=%s LIMIT 1", (item_id,))
                        existing = cur.fetchone()
                        if not existing:
                            return self.send_json({'status': 'error', 'message': '在途物流记录不存在'}, start_response)
                        if not self._parse_int(existing.get('inventory_registered')):
                            return self.send_json({'status': 'error', 'message': '需要先登记上架才能确认已核对上架数量'}, start_response)

                        cur.execute(
                            "SELECT order_product_id, shipped_qty FROM logistics_in_transit_items WHERE transit_id=%s",
                            (item_id,)
                        )
                        rows = cur.fetchall() or []
                        updates = []
                        for row in rows:
                            order_product_id = self._parse_int((row or {}).get('order_product_id'))
                            shipped_qty = self._parse_int((row or {}).get('shipped_qty')) or 0
                            if not order_product_id:
                                continue
                            listed_qty = listed_by_order.get(order_product_id, shipped_qty)
                            updates.append((listed_qty, item_id, order_product_id))

                        if updates:
                            cur.executemany(
                                "UPDATE logistics_in_transit_items SET listed_qty=%s WHERE transit_id=%s AND order_product_id=%s",
                                updates
                            )
                        cur.execute("UPDATE logistics_in_transit SET qty_verified=1 WHERE id=%s", (item_id,))

                self._refresh_transit_qty_consistent(item_id)
                self._perf_mark(perf_ctx, 'verify_qty_write')
                return self.send_json({'status': 'success', 'id': item_id}, start_response)

            if method == 'POST' and action == 'quick_status':
                item_id = self._parse_int(data.get('id'))
                field = (data.get('field') or '').strip()
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)

                allowed_fields = {
                    'declaration_docs_provided',
                    'clearance_docs_provided',
                    'inventory_registered',
                    'qty_verified',
                    'confirmed_boxed_qty'
                }
                if field not in allowed_fields:
                    return self.send_json({'status': 'error', 'message': 'Invalid field'}, start_response)

                with self._get_db_connection() as conn:
                    self._perf_mark(perf_ctx, 'db_connected_for_quick_status')
                    with conn.cursor() as cur:
                        cur.execute("SELECT id, factory_id, inventory_registered, qty_verified, confirmed_boxed_qty FROM logistics_in_transit WHERE id=%s LIMIT 1", (item_id,))
                        existing = cur.fetchone()
                        if not existing:
                            return self.send_json({'status': 'error', 'message': '在途物流记录不存在'}, start_response)

                        bool_value = _to_bool_flag(data.get('value'))
                        if field == 'qty_verified' and bool_value == 1 and not self._parse_int(existing.get('inventory_registered')):
                            return self.send_json({'status': 'error', 'message': '需要先登记上架才能确认已核对上架数量'}, start_response)
                        if field == 'inventory_registered' and bool_value == 0 and self._parse_int(existing.get('qty_verified')):
                            return self.send_json({'status': 'error', 'message': '已核对上架数量为是时，不能将已登记上架改为否'}, start_response)
                        prev_confirmed = 1 if self._parse_int(existing.get('confirmed_boxed_qty')) else 0

                        if field == 'inventory_registered':
                            if bool_value == 1:
                                cur.execute(
                                    "UPDATE logistics_in_transit SET inventory_registered=1, listed_date=COALESCE(listed_date, CURDATE()) WHERE id=%s",
                                    (item_id,)
                                )
                            else:
                                cur.execute(
                                    "UPDATE logistics_in_transit SET inventory_registered=0, listed_date=NULL WHERE id=%s",
                                    (item_id,)
                                )
                        else:
                            cur.execute(
                                f"UPDATE logistics_in_transit SET {field}=%s WHERE id=%s",
                                (bool_value, item_id)
                            )

                        should_deduct = 1 if _to_bool_flag(data.get('apply_deduct_factory_stock')) else 0
                        if field == 'confirmed_boxed_qty' and bool_value == 1 and prev_confirmed == 0 and should_deduct == 1:
                            cur.execute(
                                "SELECT order_product_id, shipped_qty FROM logistics_in_transit_items WHERE transit_id=%s",
                                (item_id,)
                            )
                            item_rows = cur.fetchall() or []
                            factory_id = self._parse_int(existing.get('factory_id'))
                            if factory_id:
                                for item_row in item_rows:
                                    op_id = self._parse_int((item_row or {}).get('order_product_id'))
                                    shipped_qty = self._parse_int((item_row or {}).get('shipped_qty')) or 0
                                    if not op_id or shipped_qty <= 0:
                                        continue
                                    cur.execute(
                                        """
                                        INSERT INTO factory_stock_inventory (order_product_id, factory_id, quantity)
                                        VALUES (%s, %s, %s)
                                        ON DUPLICATE KEY UPDATE quantity=quantity+VALUES(quantity)
                                        """,
                                        (op_id, factory_id, -shipped_qty)
                                    )
                self._perf_mark(perf_ctx, 'quick_status_write')
                self._refresh_transit_qty_consistent(item_id)
                return self.send_json({'status': 'success', 'id': item_id, 'field': field, 'value': bool_value}, start_response)

            if method in ('POST', 'PUT'):
                item_id = self._parse_int(data.get('id'))
                factory_id = self._parse_int(data.get('factory_id'))
                forwarder_id = self._parse_int(data.get('forwarder_id'))
                destination_region_id = self._parse_int(data.get('destination_region_id'))
                logistics_box_no = (data.get('logistics_box_no') or '').strip()
                if not factory_id or not destination_region_id:
                    return self.send_json({'status': 'error', 'message': '工厂和目的区域为必填'}, start_response)
                expected_listed_required = _normalize_date(data.get('expected_listed_date_latest'))
                if not expected_listed_required:
                    return self.send_json({'status': 'error', 'message': '预计上架时间为必填'}, start_response)

                payload = {
                    'factory_id': factory_id,
                    'forwarder_id': forwarder_id or None,
                    'logistics_box_no': logistics_box_no or None,
                    'arrival_port_date': _normalize_date(data.get('arrival_port_date')),
                    'expected_warehouse_date': _normalize_date(data.get('expected_warehouse_date')),
                    'expected_listed_date_latest': expected_listed_required,
                    'listed_date': _normalize_date(data.get('listed_date')),
                    'shipping_company': (data.get('shipping_company') or '').strip() or None,
                    'vessel_voyage': (data.get('vessel_voyage') or '').strip() or None,
                    'bill_of_lading_no': (data.get('bill_of_lading_no') or '').strip() or None,
                    'remark': (data.get('remark') or '').strip() or None,
                    'declaration_docs_provided': _to_bool_flag(data.get('declaration_docs_provided')),
                    'inventory_registered': _to_bool_flag(data.get('inventory_registered')),
                    'clearance_docs_provided': _to_bool_flag(data.get('clearance_docs_provided')),
                    'qty_verified': _to_bool_flag(data.get('qty_verified')),
                    'qty_consistent': 0,
                    'port_of_loading': (data.get('port_of_loading') or '').strip() or None,
                    'port_of_destination': (data.get('port_of_destination') or '').strip() or None,
                    'destination_region_id': destination_region_id,
                    'destination_warehouse_id': self._parse_int(data.get('destination_warehouse_id')),
                    'confirmed_boxed_qty': _to_bool_flag(data.get('confirmed_boxed_qty')),
                    'inbound_order_no': (data.get('inbound_order_no') or '').strip() or None
                }

                if payload.get('inventory_registered'):
                    payload['listed_date'] = payload.get('listed_date') or datetime.now().strftime('%Y-%m-%d')
                else:
                    payload['listed_date'] = None

                factory_ship_latest = _normalize_date(data.get('factory_ship_date_latest'))
                etd_latest = _normalize_date(data.get('etd_latest'))
                eta_latest = _normalize_date(data.get('eta_latest'))
                items = data.get('items') if isinstance(data.get('items'), list) else []
                normalized_items = []
                for entry in items:
                    if not isinstance(entry, dict):
                        continue
                    order_product_id = self._parse_int(entry.get('order_product_id'))
                    if not order_product_id:
                        continue
                    shipped_qty = self._parse_int(entry.get('shipped_qty')) or 0
                    listed_raw = entry.get('listed_qty')
                    listed_qty = shipped_qty if listed_raw in (None, '') else (self._parse_int(listed_raw) or 0)
                    normalized_items.append({
                        'order_product_id': order_product_id,
                        'shipped_qty': shipped_qty,
                        'listed_qty': listed_qty
                    })
                if not normalized_items:
                    return self.send_json({'status': 'error', 'message': 'SKU及发货数量为必填'}, start_response)
                if any((self._parse_int(x.get('shipped_qty')) or 0) <= 0 for x in normalized_items):
                    return self.send_json({'status': 'error', 'message': 'SKU发货数量必须大于0'}, start_response)
                if payload.get('qty_verified') and not payload.get('inventory_registered'):
                    return self.send_json({'status': 'error', 'message': '需要先登记上架才能确认已核对上架数量'}, start_response)

                payload['qty_consistent'] = self._calc_qty_consistent_from_items(normalized_items)

                with self._get_db_connection() as conn:
                    self._perf_mark(perf_ctx, 'db_connected_for_write')
                    old_bl = None
                    prev_confirmed_boxed_qty = 0
                    if method == 'PUT':
                        if not item_id:
                            return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                        with conn.cursor() as cur:
                            cur.execute(
                                "SELECT id, bill_of_lading_no, factory_ship_date_initial, factory_ship_date_latest, expected_listed_date_initial, expected_listed_date_latest, etd_initial, etd_latest, eta_initial, eta_latest, confirmed_boxed_qty FROM logistics_in_transit WHERE id=%s LIMIT 1",
                                (item_id,)
                            )
                            existing = cur.fetchone()
                        if not existing:
                            return self.send_json({'status': 'error', 'message': '在途物流记录不存在'}, start_response)
                        old_bl = (existing.get('bill_of_lading_no') or '').strip() or None
                        prev_confirmed_boxed_qty = 1 if self._parse_int(existing.get('confirmed_boxed_qty')) else 0
                        payload['factory_ship_date_initial'] = existing.get('factory_ship_date_initial') or factory_ship_latest
                        payload['factory_ship_date_previous'] = existing.get('factory_ship_date_latest') if factory_ship_latest and str(existing.get('factory_ship_date_latest') or '') != str(factory_ship_latest) else existing.get('factory_ship_date_previous')
                        payload['factory_ship_date_latest'] = factory_ship_latest or existing.get('factory_ship_date_latest')
                        payload['expected_listed_date_initial'] = existing.get('expected_listed_date_initial') or payload.get('expected_listed_date_latest')
                        payload['expected_listed_date_latest'] = payload.get('expected_listed_date_latest') or existing.get('expected_listed_date_latest')
                        payload['etd_initial'] = existing.get('etd_initial') or etd_latest
                        payload['etd_previous'] = existing.get('etd_latest') if etd_latest and str(existing.get('etd_latest') or '') != str(etd_latest) else existing.get('etd_previous')
                        payload['etd_latest'] = etd_latest or existing.get('etd_latest')
                        payload['eta_initial'] = existing.get('eta_initial') or eta_latest
                        payload['eta_previous'] = existing.get('eta_latest') if eta_latest and str(existing.get('eta_latest') or '') != str(eta_latest) else existing.get('eta_previous')
                        payload['eta_latest'] = eta_latest or existing.get('eta_latest')
                    else:
                        payload['factory_ship_date_initial'] = factory_ship_latest
                        payload['factory_ship_date_previous'] = None
                        payload['factory_ship_date_latest'] = factory_ship_latest
                        payload['expected_listed_date_initial'] = payload.get('expected_listed_date_latest')
                        payload['etd_initial'] = etd_latest
                        payload['etd_previous'] = None
                        payload['etd_latest'] = etd_latest
                        payload['eta_initial'] = eta_latest
                        payload['eta_previous'] = None
                        payload['eta_latest'] = eta_latest

                    with conn.cursor() as cur:
                        if method == 'POST':
                            cur.execute(
                                """
                                INSERT INTO logistics_in_transit (
                                    factory_id, factory_ship_date_initial, factory_ship_date_previous, factory_ship_date_latest,
                                    forwarder_id, logistics_box_no,
                                    etd_initial, etd_previous, etd_latest,
                                    eta_initial, eta_previous, eta_latest,
                                    arrival_port_date, expected_warehouse_date, expected_listed_date_initial, expected_listed_date_latest, listed_date,
                                    shipping_company, vessel_voyage, bill_of_lading_no,
                                    declaration_docs_provided, inventory_registered, clearance_docs_provided, qty_verified, qty_consistent,
                                    port_of_loading, port_of_destination, destination_region_id, destination_warehouse_id,
                                    confirmed_boxed_qty, inbound_order_no, remark
                                ) VALUES (
                                    %(factory_id)s, %(factory_ship_date_initial)s, %(factory_ship_date_previous)s, %(factory_ship_date_latest)s,
                                    %(forwarder_id)s, %(logistics_box_no)s,
                                    %(etd_initial)s, %(etd_previous)s, %(etd_latest)s,
                                    %(eta_initial)s, %(eta_previous)s, %(eta_latest)s,
                                    %(arrival_port_date)s, %(expected_warehouse_date)s, %(expected_listed_date_initial)s, %(expected_listed_date_latest)s, %(listed_date)s,
                                    %(shipping_company)s, %(vessel_voyage)s, %(bill_of_lading_no)s,
                                    %(declaration_docs_provided)s, %(inventory_registered)s, %(clearance_docs_provided)s, %(qty_verified)s, %(qty_consistent)s,
                                    %(port_of_loading)s, %(port_of_destination)s, %(destination_region_id)s, %(destination_warehouse_id)s,
                                    %(confirmed_boxed_qty)s, %(inbound_order_no)s, %(remark)s
                                )
                                """,
                                payload
                            )
                            item_id = cur.lastrowid
                        else:
                            payload['id'] = item_id
                            cur.execute(
                                """
                                UPDATE logistics_in_transit
                                SET factory_id=%(factory_id)s,
                                    factory_ship_date_initial=%(factory_ship_date_initial)s,
                                    factory_ship_date_previous=%(factory_ship_date_previous)s,
                                    factory_ship_date_latest=%(factory_ship_date_latest)s,
                                    forwarder_id=%(forwarder_id)s,
                                    logistics_box_no=%(logistics_box_no)s,
                                    etd_initial=%(etd_initial)s,
                                    etd_previous=%(etd_previous)s,
                                    etd_latest=%(etd_latest)s,
                                    eta_initial=%(eta_initial)s,
                                    eta_previous=%(eta_previous)s,
                                    eta_latest=%(eta_latest)s,
                                    arrival_port_date=%(arrival_port_date)s,
                                    expected_warehouse_date=%(expected_warehouse_date)s,
                                    expected_listed_date_initial=%(expected_listed_date_initial)s,
                                    expected_listed_date_latest=%(expected_listed_date_latest)s,
                                    listed_date=%(listed_date)s,
                                    shipping_company=%(shipping_company)s,
                                    vessel_voyage=%(vessel_voyage)s,
                                    bill_of_lading_no=%(bill_of_lading_no)s,
                                    declaration_docs_provided=%(declaration_docs_provided)s,
                                    inventory_registered=%(inventory_registered)s,
                                    clearance_docs_provided=%(clearance_docs_provided)s,
                                    qty_verified=%(qty_verified)s,
                                    qty_consistent=%(qty_consistent)s,
                                    port_of_loading=%(port_of_loading)s,
                                    port_of_destination=%(port_of_destination)s,
                                    destination_region_id=%(destination_region_id)s,
                                    destination_warehouse_id=%(destination_warehouse_id)s,
                                    confirmed_boxed_qty=%(confirmed_boxed_qty)s,
                                    inbound_order_no=%(inbound_order_no)s,
                                    remark=%(remark)s
                                WHERE id=%(id)s
                                """,
                                payload
                            )

                        cur.execute("DELETE FROM logistics_in_transit_items WHERE transit_id=%s", (item_id,))
                        if normalized_items:
                            cur.executemany(
                                "INSERT INTO logistics_in_transit_items (transit_id, order_product_id, shipped_qty, listed_qty) VALUES (%s, %s, %s, %s)",
                                [(item_id, x['order_product_id'], x['shipped_qty'], x['listed_qty']) for x in normalized_items]
                            )

                        cur.execute("UPDATE logistics_in_transit SET qty_consistent=%s WHERE id=%s", (payload.get('qty_consistent', 0), item_id))

                        should_deduct = 1 if _to_bool_flag(data.get('apply_deduct_factory_stock')) else 0
                        now_confirmed = 1 if self._parse_int(payload.get('confirmed_boxed_qty')) else 0
                        need_deduct = (now_confirmed == 1 and (method == 'POST' or prev_confirmed_boxed_qty == 0) and should_deduct == 1)
                        if need_deduct:
                            for item in normalized_items:
                                op_id = self._parse_int(item.get('order_product_id'))
                                shipped_qty = self._parse_int(item.get('shipped_qty')) or 0
                                if not op_id or shipped_qty <= 0:
                                    continue
                                cur.execute(
                                    """
                                    INSERT INTO factory_stock_inventory (order_product_id, factory_id, quantity)
                                    VALUES (%s, %s, %s)
                                    ON DUPLICATE KEY UPDATE quantity=quantity+VALUES(quantity)
                                    """,
                                    (op_id, factory_id, -shipped_qty)
                                )
                    self._perf_mark(perf_ctx, 'create_or_update_write')

                new_bl = (payload.get('bill_of_lading_no') or '').strip()
                if method == 'POST' and new_bl:
                    self._ensure_logistics_bl_folder(new_bl)
                elif method == 'PUT':
                    old_bl = (old_bl or '').strip()
                    if old_bl != new_bl:
                        if new_bl:
                            self._rename_logistics_bl_folder(old_bl, new_bl)
                    elif new_bl:
                        self._ensure_logistics_bl_folder(new_bl)

                return self.send_json({'status': 'success', 'id': item_id}, start_response)

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    self._perf_mark(perf_ctx, 'db_connected_for_delete')
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM logistics_in_transit WHERE id=%s", (item_id,))
                self._perf_mark(perf_ctx, 'delete_transit')
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        finally:
            self._perf_end(perf_ctx)

    def handle_logistics_in_transit_template_api(self, environ, method, start_response):
        """在途物流模板下载（Sheet1在途信息 + Sheet2 SKU明细）"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.comments import Comment
            from openpyxl.worksheet.datavalidation import DataValidation

            self._ensure_logistics_tables()
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            selected_ids = []
            for raw in query_params.get('ids', []):
                for token in re.split(r'[,，;；\s]+', str(raw or '').strip()):
                    if not token:
                        continue
                    item_id = self._parse_int(token)
                    if item_id and item_id not in selected_ids:
                        selected_ids.append(item_id)

            wb = Workbook()
            ws_info = wb.active
            ws_info.title = '在途信息'
            ws_items = wb.create_sheet('SKU明细')
            ws_opt = wb.create_sheet('下拉选项')

            info_headers = [
                '无箱号时临时索引*',
                '预计上架时间*',
                '工厂*', '目的区域*', '目的仓库', '工厂发货日期（预估）',
                '货代', '船公司', '船名航次', '提单号', '起运港', '目的港', 'ETD', 'ETA', '提供清关资料', '提供报关资料',
                '已确认装箱量', '箱号',
                '到港日期', '预计送仓日期', '实际上架日期', '入库单号', '已登记上架', '已核对上架数量'
            ]
            item_headers = ['箱号或临时索引', '下单SKU*', '发货数量*', '上架数量（上架后才能维护）']

            header_font = Font(bold=True, color='2A2420')
            thin_border = Border(
                left=Side(style='thin', color='B7AEA4'),
                right=Side(style='thin', color='B7AEA4'),
                top=Side(style='thin', color='B7AEA4'),
                bottom=Side(style='thin', color='B7AEA4')
            )

            def _col_letter(index_1_based):
                idx = int(index_1_based)
                text = ''
                while idx > 0:
                    idx, rem = divmod(idx - 1, 26)
                    text = chr(65 + rem) + text
                return text

            groups = [
                ('临时匹配索引', 1, 1),
                ('预计上架时间', 2, 2),
                ('装货需求', 3, 6),
                ('货代发货', 7, 16),
                ('确认装箱', 17, 18),
                ('到港、送仓、上架', 19, 24),
            ]
            header_fill_by_col = ['D3D3D3'] * len(info_headers)
            for idx, (title, start_col, end_col) in enumerate(groups):
                if hasattr(self, '_get_morandi_section_color_pair'):
                    title_color, sub_header_color = self._get_morandi_section_color_pair(idx)
                else:
                    palette = [('A8B9A5', 'DDE7DB'), ('D7C894', 'ECE5CE')]
                    title_color, sub_header_color = palette[idx % len(palette)]
                ws_info.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws_info.cell(row=1, column=start_col, value=title)
                cell.fill = PatternFill(start_color=title_color, end_color=title_color, fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for col in range(start_col, end_col + 1):
                    header_fill_by_col[col - 1] = sub_header_color

            for col, title in enumerate(info_headers, start=1):
                cell = ws_info.cell(row=2, column=col, value=title)
                header_fill = PatternFill(start_color=header_fill_by_col[col - 1], end_color=header_fill_by_col[col - 1], fill_type='solid')
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                ws_info.column_dimensions[_col_letter(col)].width = 18 if '*' in title else 16

            temp_index_col = info_headers.index('无箱号时临时索引*') + 1
            temp_index_comment = Comment(
                '用于在没有箱号信息时与SKU明细Sheet关联，请勿与箱号重复；该字段不写入数据库，仅用于本次上传临时匹配。',
                'SITJOY'
            )
            ws_info.cell(row=2, column=temp_index_col).comment = temp_index_comment

            ws_items.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            ws_items.cell(row=1, column=1, value='SKU明细关联').alignment = Alignment(horizontal='center', vertical='center')
            if hasattr(self, '_get_morandi_section_color_pair'):
                items_title_color, items_sub_color = self._get_morandi_section_color_pair(0)
            else:
                items_title_color, items_sub_color = ('A8B9A5', 'DDE7DB')
            ws_items.cell(row=1, column=1).fill = PatternFill(start_color=items_title_color, end_color=items_title_color, fill_type='solid')
            ws_items.cell(row=1, column=1).font = Font(bold=True, color='2A2420')
            for col, title in enumerate(item_headers, start=1):
                cell = ws_items.cell(row=2, column=col, value=title)
                header_fill = PatternFill(start_color=items_sub_color, end_color=items_sub_color, fill_type='solid')
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                ws_items.column_dimensions[_col_letter(col)].width = 24 if col == 1 else 18

            export_rows = []
            export_items = []

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id, factory_name FROM logistics_factories ORDER BY factory_name")
                    factories = cur.fetchall() or []
                    cur.execute("SELECT id, forwarder_name FROM logistics_forwarders ORDER BY forwarder_name")
                    forwarders = cur.fetchall() or []
                    cur.execute("SELECT id, region_name, sort_order FROM logistics_destination_regions ORDER BY sort_order ASC, id ASC")
                    destination_regions = cur.fetchall() or []
                    cur.execute("SELECT id, warehouse_name FROM logistics_overseas_warehouses WHERE COALESCE(is_enabled,1)=1 ORDER BY warehouse_name")
                    warehouses = cur.fetchall() or []
                    cur.execute("SELECT id, sku FROM order_products ORDER BY sku")
                    products = cur.fetchall() or []

                    if selected_ids:
                        placeholders = ','.join(['%s'] * len(selected_ids))
                        cur.execute(
                            f"""
                            SELECT
                                t.id, t.logistics_box_no,
                                f.factory_name, fw.forwarder_name, dr.region_name AS destination_region_name, w.warehouse_name,
                                t.factory_ship_date_latest, t.etd_latest, t.eta_latest,
                                t.arrival_port_date, t.expected_warehouse_date, t.expected_listed_date_latest,
                                t.listed_date, t.shipping_company, t.vessel_voyage,
                                t.bill_of_lading_no, t.port_of_loading, t.port_of_destination,
                                t.inbound_order_no, t.declaration_docs_provided, t.clearance_docs_provided,
                                t.qty_verified, t.qty_consistent, t.inventory_registered, t.confirmed_boxed_qty
                            FROM logistics_in_transit t
                            LEFT JOIN logistics_factories f ON f.id = t.factory_id
                            LEFT JOIN logistics_forwarders fw ON fw.id = t.forwarder_id
                            LEFT JOIN logistics_destination_regions dr ON dr.id = t.destination_region_id
                            LEFT JOIN logistics_overseas_warehouses w ON w.id = t.destination_warehouse_id
                            WHERE t.id IN ({placeholders})
                            """,
                            tuple(selected_ids)
                        )
                        selected_rows = cur.fetchall() or []
                        order_map = {sid: idx for idx, sid in enumerate(selected_ids)}
                        selected_rows.sort(key=lambda x: order_map.get(x.get('id'), 10 ** 6))
                        export_rows = selected_rows

                        if selected_rows:
                            row_ids = [r.get('id') for r in selected_rows if r.get('id')]
                            if row_ids:
                                ph2 = ','.join(['%s'] * len(row_ids))
                                cur.execute(
                                    f"""
                                    SELECT t.id AS transit_id, t.logistics_box_no, op.sku, li.shipped_qty, li.listed_qty
                                    FROM logistics_in_transit_items li
                                    JOIN logistics_in_transit t ON t.id = li.transit_id
                                    JOIN order_products op ON op.id = li.order_product_id
                                    WHERE li.transit_id IN ({ph2})
                                    ORDER BY li.transit_id, li.id
                                    """,
                                    tuple(row_ids)
                                )
                                export_items = cur.fetchall() or []

            ws_opt.cell(row=1, column=1, value='工厂')
            ws_opt.cell(row=1, column=2, value='货代')
            ws_opt.cell(row=1, column=3, value='目的区域')
            ws_opt.cell(row=1, column=4, value='仓库')
            ws_opt.cell(row=1, column=5, value='SKU')
            max_options = max(len(factories), len(forwarders), len(destination_regions), len(warehouses), len(products), 1)
            for i in range(max_options):
                row_idx = i + 2
                ws_opt.cell(row=row_idx, column=1, value=(factories[i]['factory_name'] if i < len(factories) else None))
                ws_opt.cell(row=row_idx, column=2, value=(forwarders[i]['forwarder_name'] if i < len(forwarders) else None))
                ws_opt.cell(row=row_idx, column=3, value=(destination_regions[i]['region_name'] if i < len(destination_regions) else None))
                ws_opt.cell(row=row_idx, column=4, value=(warehouses[i]['warehouse_name'] if i < len(warehouses) else None))
                ws_opt.cell(row=row_idx, column=5, value=(products[i]['sku'] if i < len(products) else None))
            ws_opt.sheet_state = 'hidden'

            sample_info_row = [
                '示例-TEMP-001', '2026-03-27',
                '示例工厂', '示例目的区域', '示例目的仓库', '2026-03-20',
                '示例货代', '示例船公司', '示例船名航次', '示例提单号', '示例起运港', '示例目的港', '2026-03-18', '2026-03-26',
                '否', '否',
                '否', '示例箱号（可留空，但需填写无箱号时临时索引用于和 SKU明细sheet 关联）',
                '2026-03-25', '2026-03-28', '2026-03-30', '示例入库单号', '否', '否'
            ]
            for col, val in enumerate(sample_info_row, start=1):
                cell = ws_info.cell(row=3, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = PatternFill(start_color='ECECEC', end_color='ECECEC', fill_type='solid')
                if col in (1, 10):
                    cell.font = Font(color='7B8088', italic=True)

            date_cols_info = {2, 6, 13, 14, 19, 20, 21}
            for col in date_cols_info:
                ws_info.cell(row=3, column=col).number_format = 'yyyy-mm-dd'

            sample_item_row = ['示例箱号（或示例-TEMP-001）', '示例SKU', 10, 10]
            for col, val in enumerate(sample_item_row, start=1):
                cell = ws_items.cell(row=3, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = PatternFill(start_color='ECECEC', end_color='ECECEC', fill_type='solid')
                if col in (1, 2):
                    cell.font = Font(color='7B8088', italic=True)

            max_validation_row = 400
            bool_validation = DataValidation(type='list', formula1='"否,是"', allow_blank=True)
            ws_info.add_data_validation(bool_validation)
            for col in date_cols_info:
                letter = _col_letter(col)
                for row in range(3, max_validation_row + 1):
                    ws_info[f'{letter}{row}'].number_format = 'yyyy-mm-dd'
            for col in (15, 16, 17, 23, 24):
                letter = _col_letter(col)
                for row in range(4, max_validation_row + 1):
                    bool_validation.add(f'{letter}{row}')

            def _add_list_validation(ws, col, options_col, count):
                if count <= 0:
                    return
                letter = _col_letter(col)
                opt_letter = _col_letter(options_col)
                formula = f"'下拉选项'!${opt_letter}$2:${opt_letter}${count + 1}"
                dv = DataValidation(type='list', formula1=formula, allow_blank=True)
                ws.add_data_validation(dv)
                for row in range(4, max_validation_row + 1):
                    dv.add(f'{letter}{row}')

            _add_list_validation(ws_info, 3, 1, len(factories))
            _add_list_validation(ws_info, 7, 2, len(forwarders))
            _add_list_validation(ws_info, 4, 3, len(destination_regions))
            _add_list_validation(ws_info, 5, 4, len(warehouses))
            _add_list_validation(ws_items, 2, 5, len(products))

            def _fmt_date(value):
                if value is None:
                    return ''
                if isinstance(value, datetime):
                    return value.strftime('%Y-%m-%d')
                text = str(value).strip()
                if not text:
                    return ''
                for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S'):
                    try:
                        return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
                    except Exception:
                        continue
                return text

            out_row = 4
            for row in export_rows:
                values = [
                    '',
                    _fmt_date(row.get('expected_listed_date_latest')),
                    row.get('factory_name') or '',
                    row.get('destination_region_name') or '',
                    row.get('warehouse_name') or '',
                    _fmt_date(row.get('factory_ship_date_latest')),
                    row.get('forwarder_name') or '',
                    row.get('shipping_company') or '',
                    row.get('vessel_voyage') or '',
                    row.get('bill_of_lading_no') or '',
                    row.get('port_of_loading') or '',
                    row.get('port_of_destination') or '',
                    _fmt_date(row.get('etd_latest')),
                    _fmt_date(row.get('eta_latest')),
                    '是' if str(row.get('clearance_docs_provided') or '0') in ('1', 'True', 'true') else '否',
                    '是' if str(row.get('declaration_docs_provided') or '0') in ('1', 'True', 'true') else '否',
                    '是' if str(row.get('confirmed_boxed_qty') or '0') in ('1', 'True', 'true') else '否',
                    row.get('logistics_box_no') or '',
                    _fmt_date(row.get('arrival_port_date')),
                    _fmt_date(row.get('expected_warehouse_date')),
                    _fmt_date(row.get('listed_date')),
                    row.get('inbound_order_no') or '',
                    '是' if str(row.get('inventory_registered') or '0') in ('1', 'True', 'true') else '否',
                    '是' if str(row.get('qty_verified') or '0') in ('1', 'True', 'true') else '否',
                ]
                for col, val in enumerate(values, start=1):
                    cell = ws_info.cell(row=out_row, column=col, value=val)
                    if col in date_cols_info:
                        cell.number_format = 'yyyy-mm-dd'
                out_row += 1

            item_row = 4
            for row in export_items:
                ws_items.cell(row=item_row, column=1, value=row.get('logistics_box_no') or '')
                ws_items.cell(row=item_row, column=2, value=row.get('sku') or '')
                shipped_qty = self._parse_int(row.get('shipped_qty')) or 0
                listed_qty = self._parse_int(row.get('listed_qty'))
                listed_qty = shipped_qty if listed_qty is None else listed_qty
                ws_items.cell(row=item_row, column=3, value=shipped_qty)
                ws_items.cell(row=item_row, column=4, value=listed_qty)
                item_row += 1

            ws_info.freeze_panes = 'A4'
            ws_items.freeze_panes = 'A4'
            return self._send_excel_workbook(wb, 'logistics_in_transit_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_in_transit_import_api(self, environ, method, start_response):
        """在途物流模板导入（Sheet1在途信息 + Sheet2 SKU明细）"""
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            self._ensure_logistics_tables()
            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            file_bytes = self._sanitize_xlsx_bool_cells(file_bytes)
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws_info = wb['在途信息'] if '在途信息' in wb.sheetnames else wb.worksheets[0]
            ws_items = wb['SKU明细'] if 'SKU明细' in wb.sheetnames else (wb.worksheets[1] if len(wb.worksheets) > 1 else None)
            if ws_items is None:
                return self.send_json({'status': 'error', 'message': '缺少SKU明细工作表'}, start_response)

            def _cell_text(v):
                if v is None:
                    return ''
                if isinstance(v, datetime):
                    return v.strftime('%Y-%m-%d')
                return str(v).strip()

            def _norm_date(v):
                if v is None:
                    return None
                if isinstance(v, datetime):
                    return v.strftime('%Y-%m-%d')
                if isinstance(v, date):
                    return v.strftime('%Y-%m-%d')
                if isinstance(v, (int, float)):
                    try:
                        serial = float(v)
                        if 1 <= serial <= 60000:
                            return (datetime(1899, 12, 30) + timedelta(days=serial)).strftime('%Y-%m-%d')
                    except Exception:
                        pass
                text = str(v).strip()
                if not text:
                    return None
                text = text.replace('年', '-').replace('月', '-').replace('日', '').replace('/', '-').replace('.', '-')
                text = re.sub(r'\s+', ' ', text)
                month_day_match = re.match(r'^(\d{1,2})-(\d{1,2})$', text)
                if month_day_match:
                    try:
                        return datetime(datetime.now().year, int(month_day_match.group(1)), int(month_day_match.group(2))).strftime('%Y-%m-%d')
                    except Exception:
                        pass
                for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M'):
                    try:
                        return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
                    except Exception:
                        continue
                return None

            def _norm_bool(v):
                text = _cell_text(v).lower()
                return 1 if text in ('1', 'true', 'yes', 'on', '是', 'y') else 0

            row1_info = [_cell_text(c.value) for c in ws_info[1]]
            row2_info = [_cell_text(c.value) for c in ws_info[2]]
            header_row_info = 2 if ('预计上架时间*' in row2_info or '工厂*' in row2_info) else 1

            row1_item = [_cell_text(c.value) for c in ws_items[1]]
            row2_item = [_cell_text(c.value) for c in ws_items[2]]
            header_row_item = 2 if ('下单SKU*' in row2_item or '箱号或临时索引' in row2_item) else 1
            headers_info = [_cell_text(c.value) for c in ws_info[header_row_info]]
            headers_item = [_cell_text(c.value) for c in ws_items[header_row_item]]
            idx_info = {h: i for i, h in enumerate(headers_info)}
            idx_item = {h: i for i, h in enumerate(headers_item)}

            def _pick_col(idx_map, *candidates):
                for col_name in candidates:
                    if col_name in idx_map:
                        return col_name
                return None

            col_box_info = _pick_col(idx_info, '箱号', '箱号*', '物流箱号*')
            col_temp_index_info = _pick_col(idx_info, '无箱号时临时索引*', '无箱号时临时索引')
            col_factory = _pick_col(idx_info, '工厂*')
            col_region = _pick_col(idx_info, '目的区域*', '目的区域')
            col_forwarder = _pick_col(idx_info, '货代*', '货代')
            col_warehouse = _pick_col(idx_info, '目的仓库*', '目的仓库')
            col_box_item = _pick_col(idx_item, '箱号', '箱号或临时索引', '箱号*', '物流箱号*')
            col_sku_item = _pick_col(idx_item, '下单SKU*')
            col_shipped_item = _pick_col(idx_item, '发货数量*')
            col_listed_item = _pick_col(idx_item, '实际上架数量', '实际上架数量*')

            if not col_box_info and not col_temp_index_info:
                return self.send_json({'status': 'error', 'message': 'Sheet 在途信息 缺少列: 箱号 或 无箱号时临时索引*'}, start_response)
            for h in (col_factory, col_region):
                if not h:
                    return self.send_json({'status': 'error', 'message': 'Sheet 在途信息 缺少列: 工厂*/目的区域*'}, start_response)
            if not col_box_item or not col_sku_item or not col_shipped_item:
                return self.send_json({'status': 'error', 'message': 'Sheet SKU明细 缺少列: 箱号或临时索引/下单SKU*/发货数量*'}, start_response)

            errors = []
            warnings = []
            info_rows = []
            seen_box = set()

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id, factory_name FROM logistics_factories")
                    factory_map = {str((r.get('factory_name') or '')).strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}
                    cur.execute("SELECT id, forwarder_name FROM logistics_forwarders")
                    forwarder_map = {str((r.get('forwarder_name') or '')).strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}
                    cur.execute("SELECT id, region_name FROM logistics_destination_regions")
                    destination_region_map = {str((r.get('region_name') or '')).strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}
                    cur.execute("SELECT id, warehouse_name FROM logistics_overseas_warehouses WHERE COALESCE(is_enabled,1)=1")
                    warehouse_map = {str((r.get('warehouse_name') or '')).strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}
                    cur.execute("SELECT id, sku FROM order_products")
                    sku_map = {str((r.get('sku') or '')).strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}

                    for row_idx in range(header_row_info + 1, ws_info.max_row + 1):
                        row_values = [ws_info.cell(row=row_idx, column=i + 1).value for i in range(len(headers_info))]
                        if not any(_cell_text(v) for v in row_values):
                            continue
                        row_join_text = '|'.join([_cell_text(v) for v in row_values]).lower()
                        if '示例' in row_join_text and ('勿导入' in row_join_text or '请勿导入' in row_join_text or '示例-' in row_join_text):
                            continue

                        box_no = _cell_text(row_values[idx_info[col_box_info]]) if col_box_info else ''
                        temp_index = _cell_text(row_values[idx_info[col_temp_index_info]]) if col_temp_index_info else ''
                        link_key = box_no or temp_index
                        factory_name = _cell_text(row_values[idx_info[col_factory]])
                        region_name = _cell_text(row_values[idx_info[col_region]])
                        forwarder_name = _cell_text(row_values[idx_info[col_forwarder]]) if col_forwarder else ''
                        warehouse_name = _cell_text(row_values[idx_info[col_warehouse]]) if col_warehouse else ''

                        if not link_key or not factory_name or not region_name:
                            errors.append({'row': f'在途信息!{row_idx}', 'error': '箱号或临时索引/工厂/目的区域为必填'})
                            continue
                        if link_key in seen_box:
                            errors.append({'row': f'在途信息!{row_idx}', 'error': f'箱号或临时索引重复: {link_key}'})
                            continue
                        seen_box.add(link_key)

                        factory_id = factory_map.get(factory_name)
                        destination_region_id = destination_region_map.get(region_name)
                        forwarder_id = forwarder_map.get(forwarder_name) if forwarder_name else None
                        warehouse_id = warehouse_map.get(warehouse_name) if warehouse_name else None
                        if not factory_id:
                            errors.append({'row': f'在途信息!{row_idx}', 'error': f'工厂不存在: {factory_name}'})
                            continue
                        if forwarder_name and not forwarder_id:
                            errors.append({'row': f'在途信息!{row_idx}', 'error': f'货代不存在: {forwarder_name}'})
                            continue
                        if not destination_region_id:
                            errors.append({'row': f'在途信息!{row_idx}', 'error': f'目的区域不存在: {region_name}'})
                            continue
                        if warehouse_name and not warehouse_id:
                            errors.append({'row': f'在途信息!{row_idx}', 'error': f'目的仓库不存在: {warehouse_name}'})
                            continue

                        def _get(name):
                            return row_values[idx_info[name]] if name in idx_info else None

                        expected_listed_val = _norm_date(_get('预计上架时间*') if '预计上架时间*' in idx_info else (_get('预计上架时间') if '预计上架时间' in idx_info else (_get('最新预计上架时间') if '最新预计上架时间' in idx_info else _get('最新预计上架日期'))))
                        if not expected_listed_val:
                            errors.append({'row': f'在途信息!{row_idx}', 'error': '预计上架时间为必填'})
                            continue

                        info_rows.append({
                            'row_idx': row_idx,
                            'link_key': link_key,
                            'logistics_box_no': box_no,
                            'factory_id': factory_id,
                            'forwarder_id': forwarder_id,
                            'destination_region_id': destination_region_id,
                            'destination_warehouse_id': warehouse_id,
                            'factory_ship_date_latest': _norm_date(
                                _get('工厂发货日期（预估）') if '工厂发货日期（预估）' in idx_info else _get('工厂发货最新日期')
                            ),
                            'etd_latest': _norm_date(
                                _get('ETD') if 'ETD' in idx_info else _get('ETD最新日期')
                            ),
                            'eta_latest': _norm_date(
                                _get('ETA') if 'ETA' in idx_info else _get('ETA最新日期')
                            ),
                            'arrival_port_date': _norm_date(_get('到港日期')),
                            'expected_warehouse_date': _norm_date(_get('预计送仓日期')),
                            'expected_listed_date_latest': expected_listed_val,
                            'listed_date': _norm_date(
                                _get('实际上架日期') if '实际上架日期' in idx_info else _get('上架日期')
                            ),
                            'shipping_company': _cell_text(_get('船公司')) or None,
                            'vessel_voyage': _cell_text(_get('船名航次')) or None,
                            'bill_of_lading_no': _cell_text(_get('提单号')) or None,
                            'port_of_loading': _cell_text(_get('起运港')) or None,
                            'port_of_destination': _cell_text(_get('目的港')) or None,
                            'inbound_order_no': _cell_text(_get('入库单号')) or None,
                            'declaration_docs_provided': _norm_bool(
                                _get('提供报关资料') if '提供报关资料' in idx_info
                                else (_get('是否提供报关资料') if '是否提供报关资料' in idx_info else _get('是否上传报关资料'))
                            ),
                            'clearance_docs_provided': _norm_bool(
                                _get('提供清关资料') if '提供清关资料' in idx_info
                                else (_get('是否提供清关资料') if '是否提供清关资料' in idx_info else _get('是否上传清关资料'))
                            ),
                            'qty_verified': _norm_bool(
                                _get('已核对上架数量') if '已核对上架数量' in idx_info
                                else (_get('是否已核对上架数量') if '是否已核对上架数量' in idx_info else _get('是否已核对数量'))
                            ),
                            'confirmed_boxed_qty': _norm_bool(_get('已确认装箱量') if '已确认装箱量' in idx_info else None),
                            'qty_consistent': 0,
                            'inventory_registered': _norm_bool(
                                _get('已登记上架') if '已登记上架' in idx_info
                                else (_get('是否已登记上架') if '是否已登记上架' in idx_info else _get('库存表已登记'))
                            ),
                        })

                    item_rows = {}
                    for row_idx in range(header_row_item + 1, ws_items.max_row + 1):
                        row_values = [ws_items.cell(row=row_idx, column=i + 1).value for i in range(len(headers_item))]
                        if not any(_cell_text(v) for v in row_values):
                            continue
                        row_join_text = '|'.join([_cell_text(v) for v in row_values]).lower()
                        if '示例' in row_join_text and ('勿导入' in row_join_text or '请勿导入' in row_join_text or '示例-' in row_join_text):
                            continue
                        box_no = _cell_text(row_values[idx_item[col_box_item]])
                        sku = _cell_text(row_values[idx_item[col_sku_item]])
                        shipped_qty = self._parse_int(row_values[idx_item[col_shipped_item]])
                        listed_qty = shipped_qty
                        if col_listed_item:
                            listed_raw = row_values[idx_item[col_listed_item]]
                            listed_qty = shipped_qty if listed_raw in (None, '') else (self._parse_int(listed_raw) or 0)
                        if not box_no or not sku or shipped_qty is None:
                            errors.append({'row': f'SKU明细!{row_idx}', 'error': '箱号或临时索引/下单SKU/发货数量为必填'})
                            continue
                        if shipped_qty < 0 or listed_qty < 0:
                            errors.append({'row': f'SKU明细!{row_idx}', 'error': '发货数量/实际上架数量不能为负数'})
                            continue
                        order_product_id = sku_map.get(sku)
                        if not order_product_id:
                            errors.append({'row': f'SKU明细!{row_idx}', 'error': f'SKU不存在: {sku}'})
                            continue
                        item_rows.setdefault(box_no, {})
                        exist_pair = item_rows[box_no].get(order_product_id) or {'shipped_qty': 0, 'listed_qty': 0}
                        exist_pair['shipped_qty'] += shipped_qty
                        exist_pair['listed_qty'] += listed_qty
                        item_rows[box_no][order_product_id] = exist_pair

                    invalid_link_keys = set()
                    for row in info_rows:
                        if int(row.get('qty_verified') or 0) == 1 and int(row.get('inventory_registered') or 0) != 1:
                            errors.append({'row': f"在途信息!{row.get('row_idx')}", 'error': '已核对上架数量=是时，已登记上架必须为是'})
                            invalid_link_keys.add(str(row.get('link_key') or ''))

                    if invalid_link_keys:
                        info_rows = [r for r in info_rows if str(r.get('link_key') or '') not in invalid_link_keys]
                        for key in list(item_rows.keys()):
                            if str(key or '') in invalid_link_keys:
                                item_rows.pop(key, None)

                    if not info_rows and not item_rows:
                        return self.send_json({'status': 'error', 'message': '未检测到可导入数据'}, start_response)

                    created = 0
                    updated = 0
                    item_updated = 0
                    transit_id_by_box = {}
                    inventory_registered_by_link = {
                        str(r.get('link_key') or ''): int(r.get('inventory_registered') or 0)
                        for r in info_rows
                    }

                    # 批查所有 info_rows 对应的箱号，避免 N+1 问题（临时索引仅用于本次Excel内关联）
                    if info_rows:
                        all_box_nos = [row['logistics_box_no'] for row in info_rows if row.get('logistics_box_no')]
                        existing_map = {}
                        if all_box_nos:
                            placeholders = ','.join(['%s'] * len(all_box_nos))
                            with conn.cursor() as cur:
                                cur.execute(
                                    f"""
                                     SELECT id, logistics_box_no, bill_of_lading_no, factory_ship_date_initial, factory_ship_date_previous, factory_ship_date_latest,
                                           expected_listed_date_initial, expected_listed_date_latest, confirmed_boxed_qty,
                                           etd_initial, etd_previous, etd_latest,
                                           eta_initial, eta_previous, eta_latest
                                    FROM logistics_in_transit
                                    WHERE logistics_box_no IN ({placeholders})
                                    """,
                                    tuple(all_box_nos)
                                )
                                for rr in cur.fetchall() or []:
                                    box_key = str(rr.get('logistics_box_no') or '').strip()
                                    if box_key:
                                        existing_map[box_key] = rr

                    for row in info_rows:
                        existing = existing_map.get(row['logistics_box_no']) if row.get('logistics_box_no') else None

                        payload = {
                            'factory_id': row['factory_id'],
                            'forwarder_id': row['forwarder_id'],
                            'logistics_box_no': row['logistics_box_no'],
                            'arrival_port_date': row['arrival_port_date'],
                            'expected_warehouse_date': row['expected_warehouse_date'],
                            'expected_listed_date_latest': row['expected_listed_date_latest'],
                            'listed_date': row['listed_date'],
                            'shipping_company': row['shipping_company'],
                            'vessel_voyage': row['vessel_voyage'],
                            'bill_of_lading_no': row['bill_of_lading_no'],
                            'declaration_docs_provided': row['declaration_docs_provided'],
                            'inventory_registered': row['inventory_registered'],
                            'clearance_docs_provided': row['clearance_docs_provided'],
                            'qty_verified': row['qty_verified'],
                            'confirmed_boxed_qty': row.get('confirmed_boxed_qty') or 0,
                            'qty_consistent': 0,
                            'port_of_loading': row['port_of_loading'],
                            'port_of_destination': row['port_of_destination'],
                            'destination_region_id': row['destination_region_id'],
                            'destination_warehouse_id': row['destination_warehouse_id'],
                            'inbound_order_no': row['inbound_order_no']
                        }

                        factory_ship_latest = row['factory_ship_date_latest']
                        etd_latest = row['etd_latest']
                        eta_latest = row['eta_latest']

                        if existing:
                            payload['factory_ship_date_initial'] = existing.get('factory_ship_date_initial') or factory_ship_latest
                            payload['factory_ship_date_previous'] = existing.get('factory_ship_date_latest') if factory_ship_latest and str(existing.get('factory_ship_date_latest') or '') != str(factory_ship_latest) else existing.get('factory_ship_date_previous')
                            payload['factory_ship_date_latest'] = factory_ship_latest or existing.get('factory_ship_date_latest')
                            payload['expected_listed_date_initial'] = existing.get('expected_listed_date_initial') or payload.get('expected_listed_date_latest')
                            payload['expected_listed_date_latest'] = payload.get('expected_listed_date_latest') or existing.get('expected_listed_date_latest')
                            payload['etd_initial'] = existing.get('etd_initial') or etd_latest
                            payload['etd_previous'] = existing.get('etd_latest') if etd_latest and str(existing.get('etd_latest') or '') != str(etd_latest) else existing.get('etd_previous')
                            payload['etd_latest'] = etd_latest or existing.get('etd_latest')
                            payload['eta_initial'] = existing.get('eta_initial') or eta_latest
                            payload['eta_previous'] = existing.get('eta_latest') if eta_latest and str(existing.get('eta_latest') or '') != str(eta_latest) else existing.get('eta_previous')
                            payload['eta_latest'] = eta_latest or existing.get('eta_latest')
                            payload['id'] = existing['id']
                            with conn.cursor() as cur:
                                cur.execute(
                                    """
                                    UPDATE logistics_in_transit
                                    SET factory_id=%(factory_id)s,
                                        factory_ship_date_initial=%(factory_ship_date_initial)s,
                                        factory_ship_date_previous=%(factory_ship_date_previous)s,
                                        factory_ship_date_latest=%(factory_ship_date_latest)s,
                                        forwarder_id=%(forwarder_id)s,
                                        logistics_box_no=%(logistics_box_no)s,
                                        etd_initial=%(etd_initial)s,
                                        etd_previous=%(etd_previous)s,
                                        etd_latest=%(etd_latest)s,
                                        eta_initial=%(eta_initial)s,
                                        eta_previous=%(eta_previous)s,
                                        eta_latest=%(eta_latest)s,
                                        arrival_port_date=%(arrival_port_date)s,
                                        expected_warehouse_date=%(expected_warehouse_date)s,
                                        expected_listed_date_initial=%(expected_listed_date_initial)s,
                                        expected_listed_date_latest=%(expected_listed_date_latest)s,
                                        listed_date=%(listed_date)s,
                                        shipping_company=%(shipping_company)s,
                                        vessel_voyage=%(vessel_voyage)s,
                                        bill_of_lading_no=%(bill_of_lading_no)s,
                                        declaration_docs_provided=%(declaration_docs_provided)s,
                                        inventory_registered=%(inventory_registered)s,
                                        clearance_docs_provided=%(clearance_docs_provided)s,
                                        qty_verified=%(qty_verified)s,
                                        qty_consistent=%(qty_consistent)s,
                                        port_of_loading=%(port_of_loading)s,
                                        port_of_destination=%(port_of_destination)s,
                                        destination_region_id=%(destination_region_id)s,
                                        destination_warehouse_id=%(destination_warehouse_id)s,
                                        confirmed_boxed_qty=%(confirmed_boxed_qty)s,
                                        inbound_order_no=%(inbound_order_no)s
                                    WHERE id=%(id)s
                                    """,
                                    payload
                                )
                            transit_id_by_box[row['link_key']] = existing['id']
                            updated += 1
                        else:
                            payload['factory_ship_date_initial'] = factory_ship_latest
                            payload['factory_ship_date_previous'] = None
                            payload['factory_ship_date_latest'] = factory_ship_latest
                            payload['expected_listed_date_initial'] = payload.get('expected_listed_date_latest')
                            payload['etd_initial'] = etd_latest
                            payload['etd_previous'] = None
                            payload['etd_latest'] = etd_latest
                            payload['eta_initial'] = eta_latest
                            payload['eta_previous'] = None
                            payload['eta_latest'] = eta_latest
                            with conn.cursor() as cur:
                                cur.execute(
                                    """
                                    INSERT INTO logistics_in_transit (
                                        factory_id, factory_ship_date_initial, factory_ship_date_previous, factory_ship_date_latest,
                                        forwarder_id, logistics_box_no,
                                        etd_initial, etd_previous, etd_latest,
                                        eta_initial, eta_previous, eta_latest,
                                        arrival_port_date, expected_warehouse_date, expected_listed_date_initial, expected_listed_date_latest, listed_date,
                                        shipping_company, vessel_voyage, bill_of_lading_no,
                                        declaration_docs_provided, inventory_registered, clearance_docs_provided, qty_verified, qty_consistent,
                                        port_of_loading, port_of_destination, destination_region_id, destination_warehouse_id, confirmed_boxed_qty, inbound_order_no
                                    ) VALUES (
                                        %(factory_id)s, %(factory_ship_date_initial)s, %(factory_ship_date_previous)s, %(factory_ship_date_latest)s,
                                        %(forwarder_id)s, %(logistics_box_no)s,
                                        %(etd_initial)s, %(etd_previous)s, %(etd_latest)s,
                                        %(eta_initial)s, %(eta_previous)s, %(eta_latest)s,
                                        %(arrival_port_date)s, %(expected_warehouse_date)s, %(expected_listed_date_initial)s, %(expected_listed_date_latest)s, %(listed_date)s,
                                        %(shipping_company)s, %(vessel_voyage)s, %(bill_of_lading_no)s,
                                        %(declaration_docs_provided)s, %(inventory_registered)s, %(clearance_docs_provided)s, %(qty_verified)s, %(qty_consistent)s,
                                        %(port_of_loading)s, %(port_of_destination)s, %(destination_region_id)s, %(destination_warehouse_id)s, %(confirmed_boxed_qty)s, %(inbound_order_no)s
                                    )
                                    """,
                                    payload
                                )
                                transit_id_by_box[row['link_key']] = cur.lastrowid
                            created += 1

                    if item_rows:
                        all_item_boxes = list(item_rows.keys())
                        unresolved_boxes = [b for b in all_item_boxes if b not in transit_id_by_box]
                        if unresolved_boxes:
                            placeholders = ','.join(['%s'] * len(unresolved_boxes))
                            with conn.cursor() as cur:
                                cur.execute(
                                    f"SELECT id, logistics_box_no FROM logistics_in_transit WHERE logistics_box_no IN ({placeholders})",
                                    tuple(unresolved_boxes)
                                )
                                for rr in cur.fetchall() or []:
                                    box_key = str(rr.get('logistics_box_no') or '').strip()
                                    if box_key:
                                        transit_id_by_box[box_key] = rr.get('id')

                        valid_transit_ids = []
                        bulk_item_rows = []
                        for box_no, sku_qty_map in item_rows.items():
                            transit_id = transit_id_by_box.get(box_no)
                            if not transit_id:
                                errors.append({'row': f'SKU明细({box_no})', 'error': '找不到对应在途信息，请先在在途信息Sheet维护该箱号或临时索引'})
                                continue
                            valid_transit_ids.append(int(transit_id))
                            is_registered = int(inventory_registered_by_link.get(str(box_no), 0)) == 1
                            for op_id, qty_pair in sku_qty_map.items():
                                shipped_qty = int((qty_pair or {}).get('shipped_qty') or 0)
                                listed_qty_input = int((qty_pair or {}).get('listed_qty') or 0)
                                final_listed_qty = listed_qty_input
                                if not is_registered:
                                    final_listed_qty = shipped_qty
                                    if listed_qty_input != shipped_qty:
                                        warnings.append({
                                            'row': f'SKU明细({box_no})',
                                            'warning': f'未登记上架，已忽略上架数量输入（SKU:{op_id}）并按发货数量 {shipped_qty} 处理'
                                        })
                                bulk_item_rows.append((
                                    int(transit_id),
                                    int(op_id),
                                    shipped_qty,
                                    final_listed_qty
                                ))
                            item_updated += 1

                        if valid_transit_ids:
                            unique_transit_ids = list(dict.fromkeys(valid_transit_ids))
                            placeholders = ','.join(['%s'] * len(unique_transit_ids))
                            with conn.cursor() as cur:
                                cur.execute(
                                    f"DELETE FROM logistics_in_transit_items WHERE transit_id IN ({placeholders})",
                                    tuple(unique_transit_ids)
                                )
                                if bulk_item_rows:
                                    cur.executemany(
                                        "INSERT INTO logistics_in_transit_items (transit_id, order_product_id, shipped_qty, listed_qty) VALUES (%s, %s, %s, %s)",
                                        bulk_item_rows
                                    )

                            for tid in unique_transit_ids:
                                self._refresh_transit_qty_consistent(tid)

            return self.send_json({
                'status': 'success',
                'created': created,
                'updated': updated,
                'item_updated': item_updated,
                'errors': errors,
                'warnings': warnings
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_in_transit_doc_upload_api(self, environ, start_response):
        try:
            if environ.get('REQUEST_METHOD') != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            self._ensure_logistics_tables()
            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            form = cgi.FieldStorage(fp=environ['wsgi.input'], environ=environ, keep_blank_values=True)
            transit_id = self._parse_int(form.getfirst('transit_id'))
            doc_type = (form.getfirst('doc_type', '') or '').strip().lower()
            if not transit_id:
                return self.send_json({'status': 'error', 'message': 'Missing transit_id'}, start_response)
            if doc_type not in ('declaration', 'clearance'):
                return self.send_json({'status': 'error', 'message': 'Invalid doc_type'}, start_response)

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id, bill_of_lading_no FROM logistics_in_transit WHERE id=%s LIMIT 1", (transit_id,))
                    item = cur.fetchone()
            if not item:
                return self.send_json({'status': 'error', 'message': '在途物流记录不存在'}, start_response)

            bl_no = (item.get('bill_of_lading_no') or '').strip()
            if not bl_no:
                return self.send_json({'status': 'error', 'message': '请先填写提单号后再上传资料'}, start_response)

            self._ensure_logistics_bl_folder(bl_no)
            root = self._get_logistics_link_root_bytes()
            parent = os.path.join(root, self._safe_fsencode(bl_no))
            sub_name = '报关资料' if doc_type == 'declaration' else '清关资料'
            target_dir = os.path.join(parent, self._safe_fsencode(sub_name))
            os.makedirs(target_dir, exist_ok=True)

            parts = getattr(form, 'list', []) or []
            uploads = [p for p in parts if getattr(p, 'filename', None)]
            if not uploads:
                return self.send_json({'status': 'error', 'message': '未检测到上传文件'}, start_response)

            saved = []
            for up in uploads:
                filename = os.path.basename(up.filename or '').strip()
                if not filename:
                    continue
                content = up.file.read() if getattr(up, 'file', None) else b''
                if content is None:
                    content = b''
                stem, ext = os.path.splitext(filename)
                candidate = filename
                idx = 1
                while os.path.exists(os.path.join(target_dir, self._safe_fsencode(candidate))):
                    candidate = f"{stem}_{idx}{ext}"
                    idx += 1
                with open(os.path.join(target_dir, self._safe_fsencode(candidate)), 'wb') as f:
                    f.write(content)
                saved.append(candidate)

            if not saved:
                return self.send_json({'status': 'error', 'message': '没有可保存的文件'}, start_response)

            flag_column = 'declaration_docs_provided' if doc_type == 'declaration' else 'clearance_docs_provided'
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(f"UPDATE logistics_in_transit SET {flag_column}=1 WHERE id=%s", (transit_id,))

            return self.send_json({'status': 'success', 'saved': saved}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_in_transit_doc_files_api(self, environ, method, start_response):
        try:
            self._ensure_logistics_tables()
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            action = (query_params.get('action', ['list'])[0] or 'list').strip().lower()

            def _decode_name(value):
                if value is None:
                    return ''
                if isinstance(value, bytes):
                    for enc in ('utf-8', 'gbk', 'latin-1'):
                        try:
                            return value.decode(enc)
                        except Exception:
                            continue
                    return value.decode('utf-8', errors='ignore')
                text = str(value)
                if text.startswith("b'") or text.startswith('b"'):
                    try:
                        parsed = ast.literal_eval(text)
                        if isinstance(parsed, bytes):
                            return _decode_name(parsed)
                    except Exception:
                        pass
                return text

            if method == 'GET' and action == 'counts':
                transit_id = self._parse_int(query_params.get('transit_id', [''])[0])
                if not transit_id:
                    return self.send_json({'status': 'error', 'message': 'Missing transit_id'}, start_response)
                declaration_dir = self._resolve_logistics_doc_folder(transit_id, 'declaration')
                clearance_dir = self._resolve_logistics_doc_folder(transit_id, 'clearance')
                decl_count = 0
                clear_count = 0
                try:
                    with os.scandir(declaration_dir) as it:
                        decl_count = sum(1 for x in it if x.is_file())
                except Exception:
                    pass
                try:
                    with os.scandir(clearance_dir) as it:
                        clear_count = sum(1 for x in it if x.is_file())
                except Exception:
                    pass
                return self.send_json({'status': 'success', 'declaration_count': decl_count, 'clearance_count': clear_count}, start_response)

            if method == 'GET' and action == 'download':
                transit_id = self._parse_int(query_params.get('transit_id', [''])[0])
                doc_type = (query_params.get('doc_type', [''])[0] or '').strip().lower()
                file_name = _decode_name((query_params.get('name', [''])[0] or '')).strip()
                if not transit_id or not file_name:
                    return self.send_json({'status': 'error', 'message': 'Missing transit_id/name'}, start_response)
                folder = self._resolve_logistics_doc_folder(transit_id, doc_type)
                file_path = os.path.join(folder, self._safe_fsencode(os.path.basename(file_name)))
                if not os.path.exists(file_path):
                    return self.send_json({'status': 'error', 'message': '文件不存在'}, start_response)
                with open(file_path, 'rb') as f:
                    content = f.read()
                content_type, _ = mimetypes.guess_type(os.fsdecode(file_path))
                safe_name = os.path.basename(file_name)
                headers = [
                    ('Content-Type', content_type or 'application/octet-stream'),
                    ('Content-Disposition', f"attachment; filename*=UTF-8''{quote(safe_name)}"),
                    ('Content-Length', str(len(content))),
                ]
                start_response('200 OK', headers)
                return [content]

            if method == 'GET':
                transit_id = self._parse_int(query_params.get('transit_id', [''])[0])
                doc_type = (query_params.get('doc_type', [''])[0] or '').strip().lower()
                if not transit_id:
                    return self.send_json({'status': 'error', 'message': 'Missing transit_id'}, start_response)
                if doc_type not in ('declaration', 'clearance'):
                    return self.send_json({'status': 'error', 'message': 'Invalid doc_type'}, start_response)
                folder = self._resolve_logistics_doc_folder(transit_id, doc_type)
                files = []
                with os.scandir(folder) as it:
                    for entry in it:
                        if not entry.is_file():
                            continue
                        stat = entry.stat()
                        safe_entry_name = _decode_name(entry.name)
                        files.append({
                            'name': safe_entry_name,
                            'size': int(getattr(stat, 'st_size', 0) or 0),
                            'updated_at': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S') if getattr(stat, 'st_mtime', None) else ''
                        })
                files.sort(key=lambda x: (x.get('updated_at') or '', x.get('name') or ''), reverse=True)
                return self.send_json({'status': 'success', 'items': files}, start_response)

            data = self._read_json_body(environ)

            if method == 'PUT':
                transit_id = self._parse_int(data.get('transit_id'))
                doc_type = (data.get('doc_type') or '').strip().lower()
                old_name = os.path.basename(_decode_name((data.get('old_name') or '')).strip())
                new_name = os.path.basename(_decode_name((data.get('new_name') or '')).strip())
                if not transit_id or not old_name or not new_name:
                    return self.send_json({'status': 'error', 'message': 'Missing fields'}, start_response)
                folder = self._resolve_logistics_doc_folder(transit_id, doc_type)
                old_path = os.path.join(folder, self._safe_fsencode(old_name))
                new_path = os.path.join(folder, self._safe_fsencode(new_name))
                if not os.path.exists(old_path):
                    return self.send_json({'status': 'error', 'message': '原文件不存在'}, start_response)
                if os.path.exists(new_path):
                    return self.send_json({'status': 'error', 'message': '新文件名已存在'}, start_response)
                os.rename(old_path, new_path)
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                transit_id = self._parse_int(data.get('transit_id'))
                doc_type = (data.get('doc_type') or '').strip().lower()
                file_name = os.path.basename(_decode_name((data.get('name') or '')).strip())
                if not transit_id or not file_name:
                    return self.send_json({'status': 'error', 'message': 'Missing fields'}, start_response)
                folder = self._resolve_logistics_doc_folder(transit_id, doc_type)
                file_path = os.path.join(folder, self._safe_fsencode(file_name))
                if not os.path.exists(file_path):
                    return self.send_json({'status': 'error', 'message': '文件不存在'}, start_response)
                os.remove(file_path)
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _ensure_logistics_bl_folder(self, bill_of_lading_no):
        name = (bill_of_lading_no or '').strip()
        if not name:
            return
        root = self._get_logistics_link_root_bytes()
        if not os.path.exists(root):
            os.makedirs(root, exist_ok=True)
        folder = os.path.join(root, self._safe_fsencode(name))
        if not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
        for sub in ('报关资料', '清关资料'):
            sub_folder = os.path.join(folder, self._safe_fsencode(sub))
            if not os.path.exists(sub_folder):
                os.makedirs(sub_folder, exist_ok=True)

