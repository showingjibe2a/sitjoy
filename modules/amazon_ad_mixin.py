# -*- coding: utf-8 -*-
"""Amazon 广告管理 Mixin - 包含11个API处理器"""

import cgi
import io
import json
import re
from datetime import datetime, timedelta
from decimal import Decimal
from urllib.parse import parse_qs

try:
    from openpyxl import Workbook, load_workbook
    _openpyxl_import_error = None
except Exception as e:
    Workbook = None
    load_workbook = None
    _openpyxl_import_error = str(e)


class AmazonAdMixin:
    """Amazon 广告管理 API 处理器 - 持有11个API handler方法"""

    _VALID_CAMPAIGN_BID_STRATEGIES = (
        '动态竞价-仅降低',
        '动态竞价-提高和降低',
        '固定竞价',
    )
    _AMAZON_AD_CHILD_IMPORT_MAX_ROWS = 2500
    _AMAZON_AD_CHILD_IMPORT_MAX_ERRORS = 80
    _AMAZON_AD_CHILD_IMPORT_BATCH_SIZE = 250
    _AMAZON_AD_ADJUSTMENT_IMPORT_MAX_ROWS = 10000
    _AMAZON_AD_ADJUSTMENT_IMPORT_BATCH_SIZE = 500

    def _parse_subtype_operation_type_ids(self, data):
        raw = data.get('operation_type_ids') if isinstance(data, dict) else None
        if not isinstance(raw, list):
            return []
        out = []
        for x in raw:
            oid = self._parse_int(x)
            if oid and oid > 0:
                out.append(int(oid))
        return sorted(set(out))

    def _sync_amazon_ad_subtype_operation_types(self, cur, subtype_id, operation_type_ids):
        sid = self._parse_int(subtype_id)
        if not sid:
            return
        cur.execute('DELETE FROM amazon_ad_subtype_operation_types WHERE subtype_id=%s', (sid,))
        for oid in operation_type_ids or []:
            cur.execute(
                'INSERT INTO amazon_ad_subtype_operation_types (subtype_id, operation_type_id) VALUES (%s, %s)',
                (sid, int(oid)),
            )

    def _attach_subtype_operation_type_ids(self, cur, rows):
        if not rows:
            return []
        ids = [self._parse_int(r.get('id')) for r in rows if self._parse_int(r.get('id'))]
        if not ids:
            return [dict(r) for r in rows]
        placeholders = ','.join(['%s'] * len(ids))
        cur.execute(
            f'SELECT subtype_id, operation_type_id FROM amazon_ad_subtype_operation_types '
            f'WHERE subtype_id IN ({placeholders})',
            tuple(ids),
        )
        by_subtype = {}
        for row in cur.fetchall() or []:
            sid = self._parse_int(row.get('subtype_id'))
            oid = self._parse_int(row.get('operation_type_id'))
            if sid and oid:
                by_subtype.setdefault(sid, []).append(oid)
        out = []
        for r in rows:
            item = dict(r)
            sid = self._parse_int(item.get('id'))
            item['operation_type_ids'] = by_subtype.get(sid, [])
            out.append(item)
        return out

    def _parse_subtype_default_targets(self, value):
        if value is None or value == '':
            return []
        if isinstance(value, list):
            items = value
        else:
            text = str(value).strip()
            if not text:
                return []
            try:
                parsed = json.loads(text)
                items = parsed if isinstance(parsed, list) else []
            except Exception:
                items = []
        out = []
        seen = set()
        for raw in items:
            if not isinstance(raw, dict):
                continue
            name = str(raw.get('name') or raw.get('target_desc') or '').strip()
            bid_value = str(raw.get('value') or raw.get('bid_value') or '').strip()
            if not name or not bid_value:
                continue
            key = name.lower()
            if key in seen:
                continue
            seen.add(key)
            out.append({'name': name, 'value': bid_value})
        return out

    def _encode_subtype_default_targets(self, targets):
        cleaned = self._parse_subtype_default_targets(targets)
        return json.dumps(cleaned, ensure_ascii=False) if cleaned else None

    def _serialize_subtype_row(self, row):
        item = dict(row) if row else {}
        item['campaign_default_targets'] = self._parse_subtype_default_targets(
            item.pop('campaign_default_targets', None)
        )
        item['group_default_targets'] = self._parse_subtype_default_targets(
            item.pop('group_default_targets', None)
        )
        return item

    def _fetch_subtype_default_targets(self, cur, subtype_id, ad_level):
        subtype_id = self._parse_int(subtype_id)
        if not subtype_id:
            return []
        column = 'campaign_default_targets' if ad_level == 'campaign' else 'group_default_targets'
        if ad_level not in ('campaign', 'group'):
            return []
        cur.execute(
            f'SELECT {column} AS targets_json FROM amazon_ad_subtypes WHERE id=%s LIMIT 1',
            (subtype_id,),
        )
        row = cur.fetchone() or {}
        return self._parse_subtype_default_targets(row.get('targets_json'))

    def _create_subtype_default_targets_for_ad_item(self, cur, ad_item_id, subtype_id, ad_level):
        ad_item_id = self._parse_int(ad_item_id)
        if not ad_item_id or ad_level not in ('campaign', 'group'):
            return
        targets = self._fetch_subtype_default_targets(cur, subtype_id, ad_level)
        if not targets:
            return
        now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        interval_text, updated_dt, next_dt = self._build_observe_fields(1, now_text)
        updated_dt = updated_dt or now_text
        for item in targets:
            name = item.get('name')
            bid_value = item.get('value')
            if not name or not bid_value:
                continue
            cur.execute(
                """
                SELECT id FROM amazon_ad_targets
                WHERE ad_item_id=%s AND target_desc=%s LIMIT 1
                """,
                (ad_item_id, name),
            )
            if cur.fetchone():
                continue
            cur.execute(
                """
                INSERT INTO amazon_ad_targets (
                    status, ad_item_id, target_desc, bid_value,
                    observe_interval, next_observe_at, updated_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                ('启动', ad_item_id, name, bid_value, interval_text, next_dt, updated_dt),
            )

    def _validate_amazon_ad_subtype_payload(self, cur, description, ad_class, subtype_code, exclude_id=None):
        description = (description or '').strip()
        ad_class = (ad_class or 'SP').strip().upper() or 'SP'
        subtype_code = (subtype_code or '').strip()
        if not description:
            return None, '请填写描述'
        if not subtype_code:
            return None, '请填写细分简称'
        sql = (
            'SELECT id FROM amazon_ad_subtypes WHERE ad_class=%s AND subtype_code=%s'
        )
        params = [ad_class, subtype_code]
        if exclude_id:
            sql += ' AND id<>%s'
            params.append(int(exclude_id))
        sql += ' LIMIT 1'
        cur.execute(sql, tuple(params))
        if cur.fetchone():
            return None, f'该广告大类（{ad_class}）下细分简称「{subtype_code}」已存在'
        return {
            'description': description,
            'ad_class': ad_class,
            'subtype_code': subtype_code,
        }, None

    def handle_amazon_ad_subtype_api(self, environ, method, start_response):
        """Amazon 广告细分类管理 API（CRUD）"""
        try:
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)
            
            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        sql = 'SELECT * FROM amazon_ad_subtypes WHERE 1=1'
                        params = []
                        if keyword:
                            like = f'%{keyword}%'
                            sql += (
                                ' AND (description LIKE %s OR ad_class LIKE %s OR subtype_code LIKE %s'
                                " OR CONCAT(ad_class, '-', subtype_code) LIKE %s)"
                            )
                            params.extend([like, like, like, like])
                        sql += ' ORDER BY id DESC LIMIT 500'
                        cur.execute(sql, tuple(params))
                        rows = self._attach_subtype_operation_type_ids(cur, cur.fetchall() or [])
                        rows = [self._serialize_subtype_row(r) for r in rows]
                return self.send_json({'status': 'success', 'items': rows}, start_response)
                
            if method == 'POST':
                data = self._read_json_body(environ) or {}
                op_ids = self._parse_subtype_operation_type_ids(data)
                campaign_targets = self._encode_subtype_default_targets(
                    data.get('campaign_default_targets')
                )
                group_targets = self._encode_subtype_default_targets(data.get('group_default_targets'))
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        fields, err = self._validate_amazon_ad_subtype_payload(
                            cur,
                            data.get('description'),
                            data.get('ad_class'),
                            data.get('subtype_code'),
                        )
                        if err:
                            return self.send_json({'status': 'error', 'message': err}, start_response)
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_subtypes (
                                description, ad_class, subtype_code,
                                campaign_default_targets, group_default_targets
                            )
                            VALUES (%s, %s, %s, %s, %s)
                            """,
                            (
                                fields['description'], fields['ad_class'], fields['subtype_code'],
                                campaign_targets, group_targets,
                            ),
                        )
                        new_id = cur.lastrowid
                        self._sync_amazon_ad_subtype_operation_types(cur, new_id, op_ids)
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ) or {}
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                op_ids = self._parse_subtype_operation_type_ids(data)
                campaign_targets = self._encode_subtype_default_targets(
                    data.get('campaign_default_targets')
                )
                group_targets = self._encode_subtype_default_targets(data.get('group_default_targets'))
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute('SELECT id FROM amazon_ad_subtypes WHERE id=%s LIMIT 1', (item_id,))
                        if not cur.fetchone():
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                        fields, err = self._validate_amazon_ad_subtype_payload(
                            cur,
                            data.get('description'),
                            data.get('ad_class'),
                            data.get('subtype_code'),
                            exclude_id=item_id,
                        )
                        if err:
                            return self.send_json({'status': 'error', 'message': err}, start_response)
                        cur.execute(
                            """
                            UPDATE amazon_ad_subtypes
                            SET description=%s, ad_class=%s, subtype_code=%s,
                                campaign_default_targets=%s, group_default_targets=%s
                            WHERE id=%s
                            """,
                            (
                                fields['description'],
                                fields['ad_class'],
                                fields['subtype_code'],
                                campaign_targets,
                                group_targets,
                                item_id,
                            ),
                        )
                        self._sync_amazon_ad_subtype_operation_types(cur, item_id, op_ids)
                return self.send_json({'status': 'success', 'id': item_id}, start_response)
            
            if method == 'DELETE':
                data = self._read_json_body(environ) or {}
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute('DELETE FROM amazon_ad_subtypes WHERE id=%s', (item_id,))
                        if cur.rowcount <= 0:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                return self.send_json({'status': 'success'}, start_response)
                
            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            print(f'Amazon ad subtype API error: {str(e)}')
            message = str(e)
            if 'uniq_ad_subtype' in message.lower() or 'duplicate' in message.lower():
                message = '该广告大类下细分简称已存在'
            return self.send_json({'status': 'error', 'message': message}, start_response)

    def _parse_apply_flag(self, value, default=1):
        if value is None or value == '':
            return 1 if default else 0
        if isinstance(value, bool):
            return 1 if value else 0
        text = str(value).strip().lower()
        if text in ('1', 'true', 'yes', 'y', '是'):
            return 1
        if text in ('0', 'false', 'no', 'n', '否'):
            return 0
        parsed = self._parse_int(value)
        return 1 if parsed else 0

    def _parse_operation_type_reason_names(self, value):
        if value is None or value == '':
            return []
        if isinstance(value, list):
            names = value
        else:
            text = str(value).strip()
            if not text:
                return []
            try:
                parsed = json.loads(text)
                names = parsed if isinstance(parsed, list) else []
            except Exception:
                names = [part.strip() for part in text.split('\n') if part.strip()]
        out = []
        seen = set()
        for raw in names:
            name = (raw if isinstance(raw, str) else (raw.get('reason_name') if isinstance(raw, dict) else ''))
            name = str(name or '').strip()
            if not name:
                continue
            key = name.lower()
            if key in seen:
                continue
            seen.add(key)
            out.append(name)
        return out

    def _normalize_operation_type_reasons_payload(self, payload):
        if payload is None:
            return []
        if isinstance(payload, list):
            return self._parse_operation_type_reason_names(payload)
        return []

    def _encode_operation_type_reason_names(self, names):
        cleaned = self._parse_operation_type_reason_names(names)
        return json.dumps(cleaned, ensure_ascii=False) if cleaned else None

    def _serialize_operation_type_row(self, row):
        item = dict(row) if row else {}
        reason_names = self._parse_operation_type_reason_names(item.pop('reason_names', None))
        item['reasons'] = [{'reason_name': name} for name in reason_names]
        for key in ('apply_portfolio', 'apply_campaign', 'apply_group'):
            if key in item and item[key] is not None:
                item[key] = int(item[key])
        if 'sort_order' in item and item['sort_order'] is not None:
            item['sort_order'] = int(item['sort_order'])
        else:
            item['sort_order'] = 0
        return item

    def _fetch_all_operation_types_with_reasons(self, cur):
        cur.execute(
            "SELECT * FROM amazon_ad_operation_types ORDER BY sort_order ASC, id ASC LIMIT 500"
        )
        return [self._serialize_operation_type_row(r) for r in (cur.fetchall() or [])]

    def _next_operation_type_sort_order(self, cur):
        cur.execute("SELECT COALESCE(MAX(sort_order), 0) AS max_sort FROM amazon_ad_operation_types")
        row = cur.fetchone() or {}
        return int(row.get('max_sort') or 0) + 10

    def handle_amazon_ad_operation_type_api(self, environ, method, start_response):
        """Amazon 广告操作类型 API（含操作原因 CRUD）"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            action = (query_params.get('action', [''])[0] or '').strip().lower()

            if method == 'GET':
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        items = self._fetch_all_operation_types_with_reasons(cur)
                return self.send_json({'status': 'success', 'items': items}, start_response)

            data = self._read_json_body(environ)

            if method == 'PUT' and action == 'reorder':
                ordered_ids = data.get('ordered_ids')
                if not isinstance(ordered_ids, list) or not ordered_ids:
                    return self.send_json({'status': 'error', 'message': '缺少 ordered_ids'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        sort_value = 10
                        for raw_id in ordered_ids:
                            item_id = self._parse_int(raw_id)
                            if not item_id:
                                continue
                            cur.execute(
                                "UPDATE amazon_ad_operation_types SET sort_order=%s WHERE id=%s",
                                (sort_value, item_id),
                            )
                            sort_value += 10
                    conn.commit()
                return self.send_json({'status': 'success'}, start_response)

            if method == 'POST':
                name = (data.get('name') or '').strip()
                if not name:
                    return self.send_json({'status': 'error', 'message': '请填写操作名称'}, start_response)
                apply_portfolio = self._parse_apply_flag(data.get('apply_portfolio'), 1)
                apply_campaign = self._parse_apply_flag(data.get('apply_campaign'), 1)
                apply_group = self._parse_apply_flag(data.get('apply_group'), 1)
                reason_names = self._encode_operation_type_reason_names(
                    self._normalize_operation_type_reasons_payload(data.get('reasons'))
                )
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        sort_order = self._next_operation_type_sort_order(cur)
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_operation_types (
                                name, sort_order, apply_portfolio, apply_campaign, apply_group, reason_names
                            ) VALUES (%s, %s, %s, %s, %s, %s)
                            """,
                            (name, sort_order, apply_portfolio, apply_campaign, apply_group, reason_names),
                        )
                        new_id = cur.lastrowid
                    conn.commit()
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                batch_items = data.get('items')
                if isinstance(batch_items, list):
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            for it in batch_items:
                                if not isinstance(it, dict):
                                    continue
                                item_id = self._parse_int(it.get('id'))
                                if not item_id:
                                    continue
                                cur.execute(
                                    """
                                    UPDATE amazon_ad_operation_types
                                    SET apply_portfolio=%s, apply_campaign=%s, apply_group=%s
                                    WHERE id=%s
                                    """,
                                    (
                                        self._parse_apply_flag(it.get('apply_portfolio'), 1),
                                        self._parse_apply_flag(it.get('apply_campaign'), 1),
                                        self._parse_apply_flag(it.get('apply_group'), 1),
                                        item_id,
                                    ),
                                )
                        conn.commit()
                    return self.send_json({'status': 'success'}, start_response)

                item_id = self._parse_int(data.get('id'))
                name = (data.get('name') or '').strip()
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                if not name:
                    return self.send_json({'status': 'error', 'message': '请填写操作名称'}, start_response)
                apply_portfolio = self._parse_apply_flag(data.get('apply_portfolio'), 1)
                apply_campaign = self._parse_apply_flag(data.get('apply_campaign'), 1)
                apply_group = self._parse_apply_flag(data.get('apply_group'), 1)
                reason_names = self._encode_operation_type_reason_names(
                    self._normalize_operation_type_reasons_payload(data.get('reasons'))
                )
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT id FROM amazon_ad_operation_types WHERE id=%s LIMIT 1",
                            (item_id,),
                        )
                        if not cur.fetchone():
                            return self.send_json({'status': 'error', 'message': '操作类型不存在'}, start_response)
                        cur.execute(
                            """
                            UPDATE amazon_ad_operation_types
                            SET name=%s, apply_portfolio=%s, apply_campaign=%s, apply_group=%s,
                                reason_names=%s
                            WHERE id=%s
                            """,
                            (name, apply_portfolio, apply_campaign, apply_group, reason_names, item_id),
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_ad_operation_types WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            print(f'Amazon ad operation type API error: {str(e)}')
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    _AMAZON_AD_ITEM_SELECT = """
        SELECT
            i.id, i.ad_level, i.sku_family_id,
            COALESCE(
                CASE WHEN i.ad_level = 'portfolio' THEN i.shop_id END,
                p.shop_id
            ) AS shop_id,
            i.portfolio_id, i.campaign_id,
            i.strategy_code, i.subtype_id, i.name, i.is_shared_budget, i.status, i.budget,
            i.bid_strategy,
            i.created_at, i.updated_at,
            pf.sku_family,
            sh.shop_name,
            p.name AS portfolio_name,
            c.name AS campaign_name,
            st.description AS subtype_description,
            st.ad_class,
            st.subtype_code
        FROM amazon_ad_items i
        LEFT JOIN amazon_ad_items c ON i.campaign_id = c.id AND c.ad_level = 'campaign'
        LEFT JOIN amazon_ad_items p ON p.id = COALESCE(NULLIF(i.portfolio_id, 0), c.portfolio_id)
            AND p.ad_level = 'portfolio'
        LEFT JOIN amazon_ad_subtypes st ON i.subtype_id = st.id
        LEFT JOIN product_families pf ON i.sku_family_id = pf.id
        LEFT JOIN shops sh ON sh.id = COALESCE(
            CASE WHEN i.ad_level = 'portfolio' THEN i.shop_id END,
            p.shop_id
        )
    """

    def _normalize_shared_budget(self, value):
        if value is None or value == '':
            return None
        if isinstance(value, bool):
            return 1 if value else 0
        text = str(value).strip()
        if text in ('1', '是', 'true', 'True', 'yes', 'Y'):
            return 1
        if text in ('0', '否', 'false', 'False', 'no', 'N'):
            return 0
        parsed = self._parse_int(value)
        return 1 if parsed else 0

    def _parse_budget_value(self, value):
        if value is None or value == '':
            return None
        if isinstance(value, (int, float, Decimal)):
            return float(value)
        text = str(value).strip()
        if not text:
            return None
        return self._parse_float(text)

    def _normalize_campaign_bid_strategy(self, value, *, required=False):
        text = (value or '').strip()
        if not text:
            if required:
                return None, '请选择竞价策略'
            return None, None
        if text not in self._VALID_CAMPAIGN_BID_STRATEGIES:
            return None, f'无效竞价策略: {text}'
        return text, None

    def _serialize_amazon_ad_item(self, row):
        if not row:
            return row
        item = dict(row)
        budget = item.get('budget')
        if isinstance(budget, Decimal):
            item['budget'] = float(budget)
        shared = item.get('is_shared_budget')
        if shared is not None:
            item['is_shared_budget'] = int(shared)
        return item

    def _fetch_amazon_ad_item_by_id(self, cur, item_id):
        cur.execute(self._AMAZON_AD_ITEM_SELECT + ' WHERE i.id=%s LIMIT 1', (item_id,))
        row = cur.fetchone()
        return self._serialize_amazon_ad_item(row) if row else None

    def _validate_amazon_ad_parent_refs(self, cur, ad_level, portfolio_id=None, campaign_id=None):
        if portfolio_id:
            cur.execute(
                "SELECT id FROM amazon_ad_items WHERE id=%s AND ad_level='portfolio' LIMIT 1",
                (portfolio_id,)
            )
            if not cur.fetchone():
                return '归属广告组合不存在'
        if campaign_id:
            cur.execute(
                "SELECT id, portfolio_id FROM amazon_ad_items WHERE id=%s AND ad_level='campaign' LIMIT 1",
                (campaign_id,)
            )
            campaign = cur.fetchone()
            if not campaign:
                return '归属广告活动不存在'
            if portfolio_id and int(campaign.get('portfolio_id') or 0) != int(portfolio_id):
                return '广告活动不属于所选广告组合'
        if ad_level == 'campaign' and not portfolio_id:
            return '请选择归属广告组合'
        if ad_level == 'group' and not campaign_id:
            return '请选择归属广告活动'
        return None

    def _inherit_campaign_fields_from_portfolio(self, cur, fields):
        """广告活动从归属组合继承 sku_family_id / shop_id。"""
        if fields.get('ad_level') != 'campaign':
            return
        portfolio_id = fields.get('portfolio_id')
        if not portfolio_id:
            return
        cur.execute(
            """
            SELECT sku_family_id, shop_id
            FROM amazon_ad_items WHERE id=%s AND ad_level='portfolio' LIMIT 1
            """,
            (portfolio_id,),
        )
        portfolio = cur.fetchone() or {}
        if portfolio.get('sku_family_id'):
            fields['sku_family_id'] = portfolio.get('sku_family_id')
        if portfolio.get('shop_id'):
            fields['shop_id'] = portfolio.get('shop_id')

    def _inherit_campaign_fields_from_portfolio_record(self, fields, portfolio_row):
        if fields.get('ad_level') != 'campaign' or not portfolio_row:
            return
        if portfolio_row.get('sku_family_id'):
            fields['sku_family_id'] = portfolio_row.get('sku_family_id')
        if portfolio_row.get('shop_id'):
            fields['shop_id'] = portfolio_row.get('shop_id')

    def _inherit_group_fields_from_campaign(self, cur, fields):
        """广告组从归属活动继承 portfolio_id / strategy_code / subtype_id / sku_family_id。"""
        if fields.get('ad_level') != 'group':
            return
        campaign_id = fields.get('campaign_id')
        if not campaign_id:
            return
        cur.execute(
            """
            SELECT portfolio_id, strategy_code, subtype_id, sku_family_id, shop_id
            FROM amazon_ad_items WHERE id=%s AND ad_level='campaign' LIMIT 1
            """,
            (campaign_id,),
        )
        campaign = cur.fetchone() or {}
        if campaign.get('portfolio_id'):
            fields['portfolio_id'] = campaign.get('portfolio_id')
        fields['strategy_code'] = campaign.get('strategy_code')
        fields['subtype_id'] = campaign.get('subtype_id')
        if campaign.get('sku_family_id'):
            fields['sku_family_id'] = campaign.get('sku_family_id')
        if campaign.get('shop_id'):
            fields['shop_id'] = campaign.get('shop_id')
        elif campaign.get('portfolio_id'):
            cur.execute(
                """
                SELECT shop_id FROM amazon_ad_items
                WHERE id=%s AND ad_level='portfolio' LIMIT 1
                """,
                (campaign.get('portfolio_id'),),
            )
            portfolio = cur.fetchone() or {}
            if portfolio.get('shop_id'):
                fields['shop_id'] = portfolio.get('shop_id')

    def _inherit_group_fields_from_campaign_record(self, fields, campaign_row):
        """广告组从内存中的活动记录继承字段（导入批处理用）。"""
        if fields.get('ad_level') != 'group' or not campaign_row:
            return
        if campaign_row.get('portfolio_id'):
            fields['portfolio_id'] = campaign_row.get('portfolio_id')
        fields['strategy_code'] = campaign_row.get('strategy_code')
        fields['subtype_id'] = campaign_row.get('subtype_id')
        if campaign_row.get('sku_family_id'):
            fields['sku_family_id'] = campaign_row.get('sku_family_id')
        if campaign_row.get('shop_id'):
            fields['shop_id'] = campaign_row.get('shop_id')

    def _apply_amazon_ad_portfolio_shop_id(self, fields, data, *, is_create=False, existing_shop_id=None):
        if fields.get('ad_level') != 'portfolio':
            return
        shop_id = self._parse_int((data or {}).get('shop_id'))
        if shop_id:
            fields['shop_id'] = shop_id
        elif is_create:
            fields['shop_id'] = 1
        elif existing_shop_id is not None:
            fields['shop_id'] = existing_shop_id

    def _cascade_amazon_ad_portfolio_shop_id(self, cur, portfolio_id, shop_id):
        if not portfolio_id or not shop_id:
            return
        cur.execute(
            """
            UPDATE amazon_ad_items
            SET shop_id=%s
            WHERE portfolio_id=%s AND ad_level='campaign'
            """,
            (shop_id, portfolio_id),
        )
        cur.execute(
            """
            UPDATE amazon_ad_items g
            INNER JOIN amazon_ad_items c ON g.campaign_id = c.id AND c.ad_level = 'campaign'
            SET g.shop_id=%s
            WHERE c.portfolio_id=%s AND g.ad_level='group'
            """,
            (shop_id, portfolio_id),
        )

    def _validate_amazon_ad_shop_ref(self, cur, shop_id):
        if not shop_id:
            return None
        cur.execute("SELECT id FROM shops WHERE id=%s LIMIT 1", (shop_id,))
        if not cur.fetchone():
            return '关联店铺不存在'
        return None

    def _load_amazon_ad_import_context(self, cur):
        """一次性加载导入索引，避免逐行 SELECT。"""
        portfolio_by_name = {}
        portfolio_by_id = {}
        campaign_by_key = {}
        campaign_by_id = {}
        campaign_by_name = {}
        existing_portfolio_by_name = {}
        existing_campaign_by_key = {}
        existing_group_by_key = {}

        shop_by_name = {}
        shop_by_id = {}
        shop_rows = []
        cur.execute("SELECT id, shop_name FROM shops ORDER BY id ASC")
        shop_rows = cur.fetchall() or []
        for row in shop_rows:
            sid = int(self._parse_int(row.get('id')) or 0)
            if sid:
                shop_by_id[sid] = sid
            name = str(row.get('shop_name') or '').strip()
            if name:
                shop_by_name[name] = row['id']
        default_shop_id = shop_rows[0]['id'] if shop_rows else 1

        cur.execute(
            """
            SELECT id, ad_level, name, portfolio_id, campaign_id, shop_id,
                   sku_family_id, strategy_code, subtype_id, is_shared_budget, status, budget, bid_strategy
            FROM amazon_ad_items
            WHERE ad_level IN ('portfolio', 'campaign', 'group')
            """
        )
        ad_rows = cur.fetchall() or []
        for row in ad_rows:
            if row.get('ad_level') != 'portfolio':
                continue
            name = str(row.get('name') or '').strip()
            if not name:
                continue
            sid = int(self._parse_int(row.get('shop_id')) or 0)
            portfolio_by_name[(sid, name)] = row['id']
            portfolio_by_id[int(row['id'])] = row
            existing_portfolio_by_name[(sid, name)] = row
        for row in ad_rows:
            level = row.get('ad_level')
            name = str(row.get('name') or '').strip()
            if not name or level == 'portfolio':
                continue
            if level == 'campaign':
                pid = int(self._parse_int(row.get('portfolio_id')) or 0)
                p_row = portfolio_by_id.get(pid) or {}
                sid = int(self._parse_int(p_row.get('shop_id')) or 0)
                campaign_by_key[(sid, pid, name)] = row
                campaign_by_id[int(row['id'])] = row
                campaign_by_name.setdefault(name, []).append(row)
                existing_campaign_by_key[(sid, pid, name)] = row
            else:
                cid = int(self._parse_int(row.get('campaign_id')) or 0)
                c_row = campaign_by_id.get(cid) or {}
                pid = int(self._parse_int(c_row.get('portfolio_id')) or 0)
                p_row = portfolio_by_id.get(pid) or {}
                sid = int(self._parse_int(p_row.get('shop_id')) or 0)
                existing_group_by_key[(sid, cid, name)] = row

        subtype_by_key = {}
        cur.execute("SELECT id, ad_class, subtype_code, description FROM amazon_ad_subtypes")
        for row in cur.fetchall() or []:
            key = f"{row.get('ad_class')}-{row.get('subtype_code')}"
            subtype_by_key[key] = row['id']
            subtype_by_key[str(row.get('description') or '').strip()] = row['id']
            code = str(row.get('subtype_code') or '').strip()
            if code:
                subtype_by_key[code] = row['id']

        sku_by_family = {}
        cur.execute("SELECT id, sku_family FROM product_families")
        for row in cur.fetchall() or []:
            sku_by_family[str(row.get('sku_family') or '').strip()] = row['id']

        return {
            'portfolio_by_name': portfolio_by_name,
            'portfolio_by_id': portfolio_by_id,
            'campaign_by_key': campaign_by_key,
            'campaign_by_id': campaign_by_id,
            'campaign_by_name': campaign_by_name,
            'existing_portfolio_by_name': existing_portfolio_by_name,
            'existing_campaign_by_key': existing_campaign_by_key,
            'existing_group_by_key': existing_group_by_key,
            'subtype_by_key': subtype_by_key,
            'sku_by_family': sku_by_family,
            'shop_by_name': shop_by_name,
            'shop_by_id': shop_by_id,
            'default_shop_id': default_shop_id or 1,
        }

    def _find_amazon_ad_import_existing(self, ad_level, name, ctx, portfolio_id=None, campaign_id=None, shop_id=None):
        name = (name or '').strip()
        if not name:
            return None
        shop_id = int(self._parse_int(shop_id) or 0)
        if ad_level == 'portfolio':
            return ctx['existing_portfolio_by_name'].get((shop_id, name))
        if ad_level == 'campaign':
            pid = self._parse_int(portfolio_id)
            if not pid or not shop_id:
                return None
            return ctx['existing_campaign_by_key'].get((shop_id, int(pid), name))
        cid = self._parse_int(campaign_id)
        if not cid or not shop_id:
            return None
        return ctx['existing_group_by_key'].get((shop_id, int(cid), name))

    def _register_amazon_ad_import_row(self, ctx, fields, row_id):
        level = fields['ad_level']
        name = fields['name']
        row = {
            'id': row_id,
            'ad_level': level,
            'name': name,
            'portfolio_id': fields.get('portfolio_id'),
            'campaign_id': fields.get('campaign_id'),
            'shop_id': fields.get('shop_id'),
            'sku_family_id': fields.get('sku_family_id'),
            'strategy_code': fields.get('strategy_code'),
            'subtype_id': fields.get('subtype_id'),
            'is_shared_budget': fields.get('is_shared_budget'),
            'status': fields.get('status'),
            'budget': fields.get('budget'),
            'bid_strategy': fields.get('bid_strategy'),
        }
        if level == 'portfolio':
            sid = int(self._parse_int(fields.get('shop_id')) or 0)
            ctx['portfolio_by_name'][(sid, name)] = row_id
            ctx['portfolio_by_id'][int(row_id)] = row
            ctx['existing_portfolio_by_name'][(sid, name)] = row
        elif level == 'campaign':
            pid = int(self._parse_int(fields.get('portfolio_id')) or 0)
            p_row = ctx['portfolio_by_id'].get(pid) or {}
            sid = int(self._parse_int(p_row.get('shop_id')) or fields.get('shop_id') or 0)
            ctx['campaign_by_key'][(sid, pid, name)] = row
            ctx['campaign_by_id'][int(row_id)] = row
            ctx['campaign_by_name'].setdefault(name, []).append(row)
            ctx['existing_campaign_by_key'][(sid, pid, name)] = row
        else:
            cid = int(self._parse_int(fields.get('campaign_id')) or 0)
            c_row = ctx['campaign_by_id'].get(cid) or {}
            pid = int(self._parse_int(c_row.get('portfolio_id')) or fields.get('portfolio_id') or 0)
            p_row = ctx['portfolio_by_id'].get(pid) or {}
            sid = int(self._parse_int(p_row.get('shop_id')) or fields.get('shop_id') or 0)
            ctx['existing_group_by_key'][(sid, cid, name)] = row

    def _resolve_shop_id_for_ad_fields(self, cur, ad_level, fields):
        if ad_level == 'portfolio':
            return self._parse_int(fields.get('shop_id'))
        pid = self._parse_int(fields.get('portfolio_id'))
        if not pid:
            return None
        cur.execute(
            "SELECT shop_id FROM amazon_ad_items WHERE id=%s AND ad_level='portfolio' LIMIT 1",
            (pid,),
        )
        row = cur.fetchone() or {}
        return self._parse_int(row.get('shop_id'))

    def _resolve_import_portfolio_id_from_ctx(self, ctx, portfolio_name, shop_id=None):
        portfolio_name = (portfolio_name or '').strip()
        if not portfolio_name:
            return None, '请填写广告组合'
        shop_id = self._parse_int(shop_id)
        if shop_id:
            pid = ctx['portfolio_by_name'].get((int(shop_id), portfolio_name))
            if not pid:
                return None, f'未找到店铺ID={shop_id}下的广告组合: {portfolio_name}'
            return pid, None
        matches = [
            pid for (sid, name), pid in ctx['portfolio_by_name'].items()
            if name == portfolio_name
        ]
        if len(matches) == 1:
            return matches[0], None
        if len(matches) > 1:
            return None, f'存在多个店铺下的同名组合「{portfolio_name}」，请先在组合行指定店铺'
        return None, f'未找到广告组合: {portfolio_name}'

    def _find_amazon_ad_item_by_scoped_name(
        self, cur, ad_level, name, portfolio_id=None, campaign_id=None, shop_id=None, exclude_id=None
    ):
        """店铺+组合+活动+组四元组在同一层级下不可重复。"""
        name = (name or '').strip()
        if not name:
            return None
        shop_id = self._parse_int(shop_id)
        exclude_sql = ''
        exclude_params = []
        if exclude_id:
            exclude_sql = ' AND {alias}.id<>%s'
            exclude_params.append(int(exclude_id))

        if ad_level == 'portfolio':
            if not shop_id:
                return None
            cur.execute(
                f"SELECT id FROM amazon_ad_items WHERE ad_level='portfolio' AND name=%s AND shop_id=%s"
                + exclude_sql.format(alias='amazon_ad_items')
                + ' LIMIT 1',
                (name, shop_id, *exclude_params),
            )
        elif ad_level == 'campaign':
            pid = self._parse_int(portfolio_id)
            if not pid or not shop_id:
                return None
            cur.execute(
                f"""
                SELECT c.id FROM amazon_ad_items c
                INNER JOIN amazon_ad_items p ON p.id=c.portfolio_id AND p.ad_level='portfolio'
                WHERE c.ad_level='campaign' AND c.name=%s AND c.portfolio_id=%s AND p.shop_id=%s
                {exclude_sql.format(alias='c')}
                LIMIT 1
                """,
                (name, pid, shop_id, *exclude_params),
            )
        else:
            cid = self._parse_int(campaign_id)
            if not cid or not shop_id:
                return None
            cur.execute(
                f"""
                SELECT g.id FROM amazon_ad_items g
                INNER JOIN amazon_ad_items c ON c.id=g.campaign_id AND c.ad_level='campaign'
                INNER JOIN amazon_ad_items p ON p.id=c.portfolio_id AND p.ad_level='portfolio'
                WHERE g.ad_level='group' AND g.name=%s AND g.campaign_id=%s AND p.shop_id=%s
                {exclude_sql.format(alias='g')}
                LIMIT 1
                """,
                (name, cid, shop_id, *exclude_params),
            )
        return cur.fetchone()

    def _validate_amazon_ad_name_unique(
        self, cur, ad_level, name, portfolio_id=None, campaign_id=None, shop_id=None, exclude_id=None
    ):
        row = self._find_amazon_ad_item_by_scoped_name(
            cur, ad_level, name,
            portfolio_id=portfolio_id, campaign_id=campaign_id, shop_id=shop_id, exclude_id=exclude_id,
        )
        if not row:
            return None
        if ad_level == 'portfolio':
            return '该店铺下已存在同名广告组合'
        if ad_level == 'campaign':
            return '该店铺与广告组合下已存在同名广告活动'
        return '该店铺、广告组合与广告活动下已存在同名广告组'

    def _amazon_ad_batch_scope_key(self, ad_level, name, portfolio_id=None, campaign_id=None, shop_id=None):
        name = (name or '').strip()
        shop_id = int(self._parse_int(shop_id) or 0)
        if ad_level == 'portfolio':
            return ('portfolio', shop_id, name)
        if ad_level == 'campaign':
            return ('campaign', shop_id, int(self._parse_int(portfolio_id) or 0), name)
        return ('group', shop_id, int(self._parse_int(campaign_id) or 0), name)

    def _validate_amazon_ad_name_unique_in_batch(
        self, ad_level, name, portfolio_id=None, campaign_id=None, shop_id=None, batch_seen=None
    ):
        key = self._amazon_ad_batch_scope_key(
            ad_level, name, portfolio_id=portfolio_id, campaign_id=campaign_id, shop_id=shop_id
        )
        if key in (batch_seen or set()):
            if ad_level == 'portfolio':
                return '本批导入中该店铺下广告组合名称重复'
            if ad_level == 'campaign':
                return '本批导入中该店铺与广告组合下广告活动名称重复'
            return '本批导入中该店铺、组合与活动下广告组名称重复'
        return None

    def _build_amazon_ad_write_fields(self, data, ad_level):
        name = (data.get('name') or '').strip()
        status = (data.get('status') or '').strip() or '启动'
        if not name:
            return None, '名称不能为空'

        fields = {
            'ad_level': ad_level,
            'name': name,
            'status': status,
            'sku_family_id': None,
            'portfolio_id': None,
            'campaign_id': None,
            'strategy_code': None,
            'subtype_id': None,
            'is_shared_budget': None,
            'budget': None,
            'bid_strategy': None,
        }

        if ad_level == 'portfolio':
            sku_family_id = self._parse_int(data.get('sku_family_id'))
            fields['sku_family_id'] = sku_family_id if sku_family_id else None
            fields['is_shared_budget'] = self._normalize_shared_budget(data.get('is_shared_budget'))
            if fields['is_shared_budget'] is None:
                return None, '请设置是否共享预算'
        elif ad_level == 'campaign':
            fields['portfolio_id'] = self._parse_int(data.get('portfolio_id'))
            fields['strategy_code'] = (data.get('strategy_code') or '').strip() or None
            fields['subtype_id'] = self._parse_int(data.get('subtype_id'))
            fields['budget'] = self._parse_budget_value(data.get('budget'))
            bid_strategy, bid_err = self._normalize_campaign_bid_strategy(
                data.get('bid_strategy'), required=False,
            )
            if bid_err:
                return None, bid_err
            fields['bid_strategy'] = bid_strategy
            if not fields['strategy_code']:
                return None, '请选择策略'
            if not fields['subtype_id']:
                return None, '请选择细分类'
        else:
            fields['portfolio_id'] = self._parse_int(data.get('portfolio_id'))
            fields['campaign_id'] = self._parse_int(data.get('campaign_id'))
            if not fields['campaign_id']:
                return None, '请选择归属广告活动'

        return fields, None

    def handle_amazon_ad_api(self, environ, method, start_response):
        """Amazon 广告 CRUD API（amazon_ad_items）"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))

            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                level = (query_params.get('level', [''])[0] or '').strip().lower()
                item_id = self._parse_int((query_params.get('id', [''])[0] or '').strip())

                sql = self._AMAZON_AD_ITEM_SELECT + ' WHERE 1=1'
                params = []

                if item_id:
                    sql += ' AND i.id=%s'
                    params.append(item_id)
                if level in ('portfolio', 'campaign', 'group'):
                    sql += ' AND i.ad_level=%s'
                    params.append(level)
                if keyword:
                    like = f'%{keyword}%'
                    sql += (
                        ' AND ('
                        'i.name LIKE %s OR pf.sku_family LIKE %s OR p.name LIKE %s OR c.name LIKE %s '
                        'OR st.description LIKE %s OR st.subtype_code LIKE %s '
                        "OR CONCAT(IFNULL(st.ad_class,''), '-', IFNULL(st.subtype_code,'')) LIKE %s"
                        ')'
                    )
                    params.extend([like] * 7)

                sql += ' ORDER BY i.id DESC LIMIT 500'

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(sql, tuple(params))
                        rows = [self._serialize_amazon_ad_item(row) for row in (cur.fetchall() or [])]

                if item_id:
                    return self.send_json({'status': 'success', 'item': rows[0] if rows else None}, start_response)
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            data = self._read_json_body(environ)

            if method == 'PATCH':
                item_id = self._parse_int(data.get('id'))
                status = (data.get('status') or '').strip()
                if not item_id or not status:
                    return self.send_json({'status': 'error', 'message': 'Missing id or status'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "UPDATE amazon_ad_items SET status=%s WHERE id=%s",
                            (status, item_id)
                        )
                        if cur.rowcount <= 0:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                        item = self._fetch_amazon_ad_item_by_id(cur, item_id)
                return self.send_json({'status': 'success', 'item': item}, start_response)

            if method == 'POST':
                ad_level = (data.get('ad_level') or '').strip().lower()
                if ad_level not in ('portfolio', 'campaign', 'group'):
                    return self.send_json({'status': 'error', 'message': '无效的广告类型'}, start_response)
                fields, err = self._build_amazon_ad_write_fields(data, ad_level)
                if err:
                    return self.send_json({'status': 'error', 'message': err}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        parent_err = self._validate_amazon_ad_parent_refs(
                            cur, ad_level, fields.get('portfolio_id'), fields.get('campaign_id')
                        )
                        if parent_err:
                            return self.send_json({'status': 'error', 'message': parent_err}, start_response)

                        if ad_level == 'portfolio':
                            self._apply_amazon_ad_portfolio_shop_id(fields, data, is_create=True)
                        elif ad_level == 'campaign':
                            self._inherit_campaign_fields_from_portfolio(cur, fields)
                        elif ad_level == 'group':
                            self._inherit_group_fields_from_campaign(cur, fields)

                        if ad_level == 'portfolio':
                            shop_err = self._validate_amazon_ad_shop_ref(cur, fields.get('shop_id'))
                            if shop_err:
                                return self.send_json({'status': 'error', 'message': shop_err}, start_response)

                        scope_shop_id = self._resolve_shop_id_for_ad_fields(cur, ad_level, fields)
                        dup_err = self._validate_amazon_ad_name_unique(
                            cur, ad_level, fields['name'],
                            portfolio_id=fields.get('portfolio_id'),
                            campaign_id=fields.get('campaign_id'),
                            shop_id=scope_shop_id,
                        )
                        if dup_err:
                            return self.send_json({'status': 'error', 'message': dup_err}, start_response)

                        cur.execute(
                            """
                            INSERT INTO amazon_ad_items (
                                ad_level, sku_family_id, shop_id, portfolio_id, campaign_id,
                                strategy_code, subtype_id, name, is_shared_budget, status, budget, bid_strategy
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                fields['ad_level'], fields['sku_family_id'], fields.get('shop_id'),
                                fields['portfolio_id'], fields['campaign_id'], fields['strategy_code'],
                                fields['subtype_id'], fields['name'], fields['is_shared_budget'],
                                fields['status'], fields['budget'], fields.get('bid_strategy'),
                            )
                        )
                        new_id = cur.lastrowid
                        if ad_level in ('campaign', 'group'):
                            self._create_subtype_default_targets_for_ad_item(
                                cur, new_id, fields.get('subtype_id'), ad_level,
                            )
                        item = self._fetch_amazon_ad_item_by_id(cur, new_id)
                return self.send_json({'status': 'success', 'id': new_id, 'item': item}, start_response)

            if method == 'PUT':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT id, ad_level, shop_id FROM amazon_ad_items WHERE id=%s LIMIT 1",
                            (item_id,)
                        )
                        existing = cur.fetchone()
                        if not existing:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)

                        ad_level = existing.get('ad_level')
                        fields, err = self._build_amazon_ad_write_fields(data, ad_level)
                        if err:
                            return self.send_json({'status': 'error', 'message': err}, start_response)

                        parent_err = self._validate_amazon_ad_parent_refs(
                            cur, ad_level, fields.get('portfolio_id'), fields.get('campaign_id')
                        )
                        if parent_err:
                            return self.send_json({'status': 'error', 'message': parent_err}, start_response)

                        if ad_level == 'portfolio':
                            self._apply_amazon_ad_portfolio_shop_id(
                                fields, data, is_create=False, existing_shop_id=existing.get('shop_id')
                            )
                        elif ad_level == 'campaign':
                            self._inherit_campaign_fields_from_portfolio(cur, fields)
                        elif ad_level == 'group':
                            self._inherit_group_fields_from_campaign(cur, fields)

                        if ad_level == 'portfolio':
                            shop_err = self._validate_amazon_ad_shop_ref(cur, fields.get('shop_id'))
                            if shop_err:
                                return self.send_json({'status': 'error', 'message': shop_err}, start_response)

                        scope_shop_id = self._resolve_shop_id_for_ad_fields(cur, ad_level, fields)
                        dup_err = self._validate_amazon_ad_name_unique(
                            cur, ad_level, fields['name'],
                            portfolio_id=fields.get('portfolio_id'),
                            campaign_id=fields.get('campaign_id'),
                            shop_id=scope_shop_id,
                            exclude_id=item_id,
                        )
                        if dup_err:
                            return self.send_json({'status': 'error', 'message': dup_err}, start_response)

                        cur.execute(
                            """
                            UPDATE amazon_ad_items SET
                                sku_family_id=%s, shop_id=%s, portfolio_id=%s, campaign_id=%s,
                                strategy_code=%s, subtype_id=%s, name=%s,
                                is_shared_budget=%s, status=%s, budget=%s, bid_strategy=%s
                            WHERE id=%s
                            """,
                            (
                                fields['sku_family_id'], fields.get('shop_id'), fields['portfolio_id'],
                                fields['campaign_id'], fields['strategy_code'], fields['subtype_id'],
                                fields['name'], fields['is_shared_budget'], fields['status'], fields['budget'],
                                fields.get('bid_strategy'), item_id,
                            )
                        )
                        if ad_level == 'portfolio':
                            self._cascade_amazon_ad_portfolio_shop_id(
                                cur, item_id, fields.get('shop_id')
                            )
                        item = self._fetch_amazon_ad_item_by_id(cur, item_id)
                return self.send_json({'status': 'success', 'item': item}, start_response)

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_ad_items WHERE id=%s", (item_id,))
                        if cur.rowcount <= 0:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            print(f'Amazon ad API error: {str(e)}')
            message = str(e)
            if 'foreign key constraint' in message.lower() or '1451' in message:
                message = '无法删除：该广告记录仍被投放/商品/调整记录引用'
            return self.send_json({'status': 'error', 'message': message}, start_response)

    def _amazon_ad_items_template_headers(self):
        """导入模板列顺序：归属两列置于最右侧。"""
        return [
            '广告类型*', '名称*', '状态*',
            '关联货号', '是否共享预算', '关联店铺',
            '策略', '细分类', '预算', '竞价策略',
            '归属广告组合名称', '归属广告活动名称',
        ]

    def _load_amazon_ad_items_template_options(self):
        """模板下拉：货号、细分类、店铺、广告组合（与导入解析键一致）。"""
        sku_families = []
        subtype_labels = []
        shop_names = []
        portfolio_names = []
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT sku_family FROM product_families
                    WHERE sku_family IS NOT NULL AND TRIM(sku_family) <> ''
                    ORDER BY sku_family ASC
                    """
                )
                sku_families = [
                    str(row.get('sku_family') or '').strip()
                    for row in (cur.fetchall() or [])
                    if str(row.get('sku_family') or '').strip()
                ]
                cur.execute(
                    """
                    SELECT ad_class, subtype_code, description
                    FROM amazon_ad_subtypes
                    ORDER BY ad_class ASC, subtype_code ASC, id ASC
                    """
                )
                seen = set()
                for row in cur.fetchall() or []:
                    ad_class = str(row.get('ad_class') or '').strip()
                    code = str(row.get('subtype_code') or '').strip()
                    desc = str(row.get('description') or '').strip()
                    composite = f'{ad_class}-{code}' if ad_class and code else ''
                    for label in (desc, composite, code):
                        if label and label not in seen:
                            seen.add(label)
                            subtype_labels.append(label)
                cur.execute(
                    """
                    SELECT shop_name FROM shops
                    WHERE shop_name IS NOT NULL AND TRIM(shop_name) <> ''
                    ORDER BY id ASC
                    """
                )
                shop_names = [
                    str(row.get('shop_name') or '').strip()
                    for row in (cur.fetchall() or [])
                    if str(row.get('shop_name') or '').strip()
                ]
                cur.execute(
                    """
                    SELECT name FROM amazon_ad_items
                    WHERE ad_level = 'portfolio'
                      AND name IS NOT NULL AND TRIM(name) <> ''
                    ORDER BY name ASC
                    """
                )
                portfolio_names = [
                    str(row.get('name') or '').strip()
                    for row in (cur.fetchall() or [])
                    if str(row.get('name') or '').strip()
                ]
        return sku_families, subtype_labels, shop_names, portfolio_names

    def _style_amazon_ad_items_template_example_rows(self, ws):
        """第 2–4 行为示例行，使用不同底色区分（导入时跳过）。"""
        from openpyxl.styles import Font, PatternFill, Alignment

        example_font = Font(italic=True, color='7B8088')
        example_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        fills = [
            PatternFill(start_color='E8F0E8', end_color='E8F0E8', fill_type='solid'),
            PatternFill(start_color='E8EEF6', end_color='E8EEF6', fill_type='solid'),
            PatternFill(start_color='F6EEE8', end_color='F6EEE8', fill_type='solid'),
        ]
        for row_idx, fill in enumerate(fills, start=2):
            for cell in ws[row_idx]:
                cell.fill = fill
                cell.font = example_font
                cell.alignment = example_align

    def _apply_amazon_ad_items_template_formatting(
        self,
        ws,
        *,
        options_ws=None,
        sku_families=None,
        subtype_labels=None,
        shop_names=None,
        portfolio_names=None,
        first_data_row=5,
        last_row=1000,
    ):
        """下拉预设 + 按广告类型对不适用字段加灰色底纹（条件格式）。"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.worksheet.datavalidation import DataValidation

        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        gray_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color='2A2420')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        first_data_row = max(5, int(first_data_row or 5))
        data_end = max(first_data_row, int(last_row or 1000))
        anchor = first_data_row

        dv_level = DataValidation(type='list', formula1='"组合,活动,组"', allow_blank=False)
        dv_level.error = '请从列表选择：组合 / 活动 / 组'
        dv_level.errorTitle = '广告类型'
        ws.add_data_validation(dv_level)
        dv_level.add(f'A{first_data_row}:A{data_end}')

        dv_status = DataValidation(type='list', formula1='"启动,暂停,存档"', allow_blank=False)
        dv_status.error = '请从列表选择：启动 / 暂停 / 存档'
        dv_status.errorTitle = '状态'
        ws.add_data_validation(dv_status)
        dv_status.add(f'C{first_data_row}:C{data_end}')

        dv_shared = DataValidation(type='list', formula1='"是,否"', allow_blank=True)
        dv_shared.error = '请从列表选择：是 / 否'
        dv_shared.errorTitle = '是否共享预算'
        ws.add_data_validation(dv_shared)
        dv_shared.add(f'E{first_data_row}:E{data_end}')

        dv_strategy = DataValidation(type='list', formula1='"BE,BD,PC"', allow_blank=True)
        dv_strategy.error = '请从列表选择：BE / BD / PC'
        dv_strategy.errorTitle = '策略'
        ws.add_data_validation(dv_strategy)
        dv_strategy.add(f'G{first_data_row}:G{data_end}')

        dv_bid_strategy = DataValidation(
            type='list',
            formula1='"动态竞价-仅降低,动态竞价-提高和降低,固定竞价"',
            allow_blank=True,
        )
        dv_bid_strategy.error = '请从列表选择竞价策略'
        dv_bid_strategy.errorTitle = '竞价策略'
        ws.add_data_validation(dv_bid_strategy)
        dv_bid_strategy.add(f'J{first_data_row}:J{data_end}')

        sku_families = list(sku_families or [])
        subtype_labels = list(subtype_labels or [])
        shop_names = list(shop_names or [])
        portfolio_names = list(portfolio_names or [])
        if options_ws is not None and sku_families:
            opt_end = len(sku_families) + 1
            dv_sku = DataValidation(
                type='list',
                formula1=f"='_options'!$A$2:$A${opt_end}",
                allow_blank=True,
            )
            dv_sku.error = '请从列表选择货号'
            dv_sku.errorTitle = '关联货号'
            ws.add_data_validation(dv_sku)
            dv_sku.add(f'D{first_data_row}:D{data_end}')
        if options_ws is not None and subtype_labels:
            opt_end = len(subtype_labels) + 1
            dv_subtype = DataValidation(
                type='list',
                formula1=f"='_options'!$B$2:$B${opt_end}",
                allow_blank=True,
            )
            dv_subtype.error = '请从列表选择细分类'
            dv_subtype.errorTitle = '细分类'
            ws.add_data_validation(dv_subtype)
            dv_subtype.add(f'H{first_data_row}:H{data_end}')
        if options_ws is not None and shop_names:
            opt_end = len(shop_names) + 1
            dv_shop = DataValidation(
                type='list',
                formula1=f"='_options'!$C$2:$C${opt_end}",
                allow_blank=True,
            )
            dv_shop.error = '请从列表选择店铺'
            dv_shop.errorTitle = '关联店铺'
            ws.add_data_validation(dv_shop)
            dv_shop.add(f'F{first_data_row}:F{data_end}')
        if options_ws is not None and portfolio_names:
            opt_end = len(portfolio_names) + 1
            dv_portfolio = DataValidation(
                type='list',
                formula1=f"='_options'!$D$2:$D${opt_end}",
                allow_blank=True,
            )
            dv_portfolio.error = '请从列表选择广告组合'
            dv_portfolio.errorTitle = '归属广告组合名称'
            ws.add_data_validation(dv_portfolio)
            dv_portfolio.add(f'K{first_data_row}:K{data_end}')

        # 条件格式：公式为真时显示灰色（表示当前行广告类型下该列无需填写）
        rules = [
            ('D', f'OR($A{anchor}="",$A{anchor}<>"组合")'),
            ('E', f'OR($A{anchor}="",$A{anchor}<>"组合")'),
            ('F', f'OR($A{anchor}="",$A{anchor}<>"组合")'),
            ('G', f'OR($A{anchor}="",$A{anchor}<>"活动")'),
            ('H', f'OR($A{anchor}="",$A{anchor}<>"活动")'),
            ('I', f'OR($A{anchor}="",$A{anchor}<>"活动")'),
            ('J', f'OR($A{anchor}="",$A{anchor}<>"活动")'),
            ('K', f'OR($A{anchor}="",$A{anchor}="组合")'),
            ('L', f'OR($A{anchor}="",$A{anchor}<>"组")'),
        ]
        for col, formula in rules:
            ws.conditional_formatting.add(
                f'{col}{first_data_row}:{col}{data_end}',
                FormulaRule(formula=[formula], fill=gray_fill),
            )

        ws.freeze_panes = f'A{first_data_row}'
        widths = {
            'A': 11, 'B': 28, 'C': 9, 'D': 14, 'E': 14,
            'F': 16, 'G': 8, 'H': 12, 'I': 10, 'J': 18, 'K': 22, 'L': 26,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def _is_amazon_ad_template_sample_row(self, row_idx, name):
        if row_idx <= 4:
            return True
        text = str(name or '').strip()
        return ('示例' in text) or ('请勿导入' in text)

    def _build_amazon_ad_items_import_workbook(self):
        from openpyxl import Workbook

        sku_families, subtype_labels, shop_names, portfolio_names = (
            self._load_amazon_ad_items_template_options()
        )
        example_shop = shop_names[0] if shop_names else ''

        wb = Workbook()
        ws = wb.active
        ws.title = '广告信息'
        headers = self._amazon_ad_items_template_headers()
        ws.append(headers)
        ws.append([
            '组合', '示例-Short-SKU01', '启动',
            'SKU01', '是', example_shop,
            '', '', '', '',
            '', '',
        ])
        ws.append([
            '活动', 'BE-示例组合-SP-KW', '启动',
            '', '', '',
            'BE', 'SP-KW', '50', '动态竞价-仅降低',
            '示例-Short-SKU01', '',
        ])
        ws.append([
            '组', 'BE-示例组合-SP-KW', '启动',
            '', '', '',
            '', '', '', '',
            '示例-Short-SKU01', 'BE-示例组合-SP-KW',
        ])
        self._style_amazon_ad_items_template_example_rows(ws)

        options_ws = wb.create_sheet('_options')
        options_ws.sheet_state = 'hidden'
        options_ws.append(['sku_family', 'subtype', 'shop_name', 'portfolio_name'])
        opt_rows = max(len(sku_families), len(subtype_labels), len(shop_names), len(portfolio_names), 1)
        for i in range(opt_rows):
            options_ws.append([
                sku_families[i] if i < len(sku_families) else '',
                subtype_labels[i] if i < len(subtype_labels) else '',
                shop_names[i] if i < len(shop_names) else '',
                portfolio_names[i] if i < len(portfolio_names) else '',
            ])

        self._apply_amazon_ad_items_template_formatting(
            ws,
            options_ws=options_ws,
            sku_families=sku_families,
            subtype_labels=subtype_labels,
            shop_names=shop_names,
            portfolio_names=portfolio_names,
            first_data_row=5,
            last_row=1200,
        )

        guide = wb.create_sheet('填写说明')
        guide.append(['字段', '组合', '活动', '组', '说明'])
        guide_rows = [
            ('广告类型*', '必填', '必填', '必填', '下拉：组合 / 活动 / 组'),
            ('名称*', '必填', '必填', '必填', ''),
            ('状态*', '必填', '必填', '必填', '下拉：启动 / 暂停 / 存档'),
            ('关联货号', '选填', '—', '—', '仅组合可填；下拉为系统货号；灰底表示本行不适用'),
            ('是否共享预算', '必填', '—', '—', '仅组合可填；下拉：是 / 否'),
            ('关联店铺', '选填', '—', '—', '仅组合可填；下拉为系统店铺；留空默认 id=1'),
            ('策略', '—', '必填', '—', '仅活动可填；下拉：BE / BD / PC'),
            ('细分类', '—', '必填', '—', '仅活动可填；下拉为系统细分类'),
            ('预算', '—', '必填', '—', '仅活动可填'),
            ('竞价策略', '—', '选填', '—', '仅活动可填；部分类型可不填；下拉：动态竞价-仅降低 / 动态竞价-提高和降低 / 固定竞价'),
            ('归属广告组合名称', '—', '必填', '必填', '活动、组均须填写；下拉为系统广告组合'),
            ('归属广告活动名称', '—', '—', '必填', '仅组可填；须为对应组合下已存在的活动'),
            ('', '', '', '', '第2–4行为示例（彩色底纹），导入时自动跳过；请从第5行填写'),
        ]
        for row in guide_rows:
            guide.append(list(row))
        guide.column_dimensions['A'].width = 18
        guide.column_dimensions['B'].width = 8
        guide.column_dimensions['C'].width = 8
        guide.column_dimensions['D'].width = 8
        guide.column_dimensions['E'].width = 42
        return wb

    def handle_amazon_ad_template_api(self, environ, method, start_response):
        """Amazon 广告信息批量导入模板下载"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json(
                    {'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'},
                    start_response
                )
            wb = self._build_amazon_ad_items_import_workbook()
            return self._send_excel_workbook(wb, '广告信息导入模板.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_import_api(self, environ, method, start_response):
        """Amazon 广告信息批量导入"""
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json(
                    {'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'},
                    start_response
                )

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            headers = [str(cell.value or '').strip() for cell in ws[1]]
            header_map = {name: idx for idx, name in enumerate(headers)}

            def cell_value(row, name):
                idx = header_map.get(name)
                if idx is None or idx >= len(row):
                    return None
                value = row[idx].value
                return None if value is None else str(value).strip()

            level_map = {
                '广告组合': 'portfolio', 'portfolio': 'portfolio', '组合': 'portfolio',
                '广告活动': 'campaign', 'campaign': 'campaign', '活动': 'campaign',
                '广告组': 'group', 'group': 'group', '组': 'group',
            }

            created = updated = unchanged = 0
            skipped_sample_rows = 0
            errors = []

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    ctx = self._load_amazon_ad_import_context(cur)
                    portfolio_by_name = ctx['portfolio_by_name']
                    campaign_by_key = ctx['campaign_by_key']
                    campaign_by_name = ctx['campaign_by_name']
                    subtype_by_key = ctx['subtype_by_key']
                    sku_by_family = ctx['sku_by_family']
                    shop_by_name = ctx['shop_by_name']
                    default_shop_id = ctx['default_shop_id']
                    batch_seen_names = set()

                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                        name = cell_value(row, '名称*') or ''
                        if self._is_amazon_ad_template_sample_row(row_idx, name):
                            skipped_sample_rows += 1
                            continue
                        level_raw = cell_value(row, '广告类型*') or ''
                        ad_level = level_map.get(level_raw.lower()) or level_map.get(level_raw)
                        status = cell_value(row, '状态*') or '启动'
                        if not level_raw and not name:
                            continue
                        if not ad_level or not name:
                            errors.append({'row': row_idx, 'message': '广告类型或名称为空'})
                            continue

                        payload = {
                            'ad_level': ad_level,
                            'name': name,
                            'status': status,
                        }

                        if ad_level == 'portfolio':
                            sku_family = cell_value(row, '关联货号') or ''
                            if sku_family:
                                payload['sku_family_id'] = sku_by_family.get(sku_family)
                            payload['is_shared_budget'] = cell_value(row, '是否共享预算') or '是'
                            shop_name = cell_value(row, '关联店铺') or ''
                            if shop_name:
                                payload['shop_id'] = shop_by_name.get(shop_name)
                                if not payload['shop_id']:
                                    errors.append({'row': row_idx, 'message': f'未找到店铺: {shop_name}'})
                                    continue
                        elif ad_level == 'campaign':
                            portfolio_name = cell_value(row, '归属广告组合名称') or ''
                            portfolio_id, p_err = self._resolve_import_portfolio_id_from_ctx(
                                ctx, portfolio_name
                            )
                            if p_err:
                                errors.append({'row': row_idx, 'message': p_err})
                                continue
                            payload['portfolio_id'] = portfolio_id
                            payload['strategy_code'] = cell_value(row, '策略') or ''
                            subtype_text = cell_value(row, '细分类') or ''
                            payload['subtype_id'] = subtype_by_key.get(subtype_text)
                            payload['budget'] = cell_value(row, '预算') or ''
                            bid_strategy, bid_err = self._normalize_campaign_bid_strategy(
                                cell_value(row, '竞价策略'), required=False,
                            )
                            if bid_err:
                                errors.append({'row': row_idx, 'message': bid_err})
                                continue
                            payload['bid_strategy'] = bid_strategy
                            if not payload['portfolio_id']:
                                errors.append({'row': row_idx, 'message': f'未找到广告组合: {portfolio_name}'})
                                continue
                            if not payload['subtype_id']:
                                errors.append({'row': row_idx, 'message': f'未找到细分类: {subtype_text}'})
                                continue
                        else:
                            campaign_name = (cell_value(row, '归属广告活动名称') or '').strip()
                            portfolio_name = (cell_value(row, '归属广告组合名称') or '').strip()
                            if not portfolio_name:
                                errors.append({'row': row_idx, 'message': '广告组须填写归属广告组合名称'})
                                continue
                            if not campaign_name:
                                errors.append({'row': row_idx, 'message': '广告组须填写归属广告活动名称'})
                                continue
                            portfolio_id_for_group, p_err = self._resolve_import_portfolio_id_from_ctx(
                                ctx, portfolio_name
                            )
                            if p_err:
                                errors.append({'row': row_idx, 'message': p_err})
                                continue
                            p_row = ctx['portfolio_by_id'].get(int(portfolio_id_for_group)) or {}
                            sid = int(self._parse_int(p_row.get('shop_id')) or 0)
                            campaign = campaign_by_key.get(
                                (sid, int(portfolio_id_for_group), campaign_name)
                            )
                            if not campaign:
                                errors.append({
                                    'row': row_idx,
                                    'message': (
                                        f'在组合「{portfolio_name}」下未找到广告活动: {campaign_name}'
                                    ),
                                })
                                continue
                            payload['campaign_id'] = campaign['id']
                            payload['portfolio_id'] = campaign.get('portfolio_id')

                        fields, err = self._build_amazon_ad_write_fields(payload, ad_level)
                        if err:
                            errors.append({'row': row_idx, 'message': err})
                            continue
                        if ad_level == 'portfolio':
                            if payload.get('shop_id'):
                                fields['shop_id'] = payload['shop_id']
                        elif ad_level == 'campaign':
                            portfolio_name = cell_value(row, '归属广告组合名称') or ''
                            p_row = ctx['portfolio_by_id'].get(int(fields.get('portfolio_id') or 0)) or {}
                            portfolio_row = None
                            if p_row:
                                sid = int(self._parse_int(p_row.get('shop_id')) or 0)
                                portfolio_row = ctx['existing_portfolio_by_name'].get(
                                    (sid, portfolio_name)
                                )
                            self._inherit_campaign_fields_from_portfolio_record(fields, portfolio_row)
                        elif ad_level == 'group':
                            campaign_row = None
                            if fields.get('campaign_id'):
                                campaign_row = ctx['campaign_by_id'].get(int(fields['campaign_id']))
                            self._inherit_group_fields_from_campaign_record(fields, campaign_row)
                            if not fields.get('shop_id'):
                                portfolio_name = (cell_value(row, '归属广告组合名称') or '').strip()
                                p_row = ctx['portfolio_by_id'].get(int(fields.get('portfolio_id') or 0)) or {}
                                sid = int(self._parse_int(p_row.get('shop_id')) or 0)
                                portfolio_row = ctx['existing_portfolio_by_name'].get((sid, portfolio_name))
                                if portfolio_row and portfolio_row.get('shop_id'):
                                    fields['shop_id'] = portfolio_row.get('shop_id')

                        scope_shop_id = fields.get('shop_id')
                        if not scope_shop_id and fields.get('portfolio_id'):
                            p_row = ctx['portfolio_by_id'].get(int(fields.get('portfolio_id'))) or {}
                            scope_shop_id = p_row.get('shop_id')

                        existing = self._find_amazon_ad_import_existing(
                            ad_level, fields['name'], ctx,
                            portfolio_id=fields.get('portfolio_id'),
                            campaign_id=fields.get('campaign_id'),
                            shop_id=scope_shop_id,
                        )
                        if ad_level == 'portfolio' and not fields.get('shop_id'):
                            fields['shop_id'] = (
                                existing.get('shop_id') if existing else None
                            ) or default_shop_id
                        if existing:
                            cur.execute(
                                """
                                UPDATE amazon_ad_items SET
                                    sku_family_id=%s, shop_id=%s, portfolio_id=%s, campaign_id=%s,
                                    strategy_code=%s, subtype_id=%s,
                                    is_shared_budget=%s, status=%s, budget=%s, bid_strategy=%s
                                WHERE id=%s
                                """,
                                (
                                    fields['sku_family_id'], fields.get('shop_id'),
                                    fields['portfolio_id'], fields['campaign_id'],
                                    fields['strategy_code'], fields['subtype_id'],
                                    fields['is_shared_budget'], fields['status'], fields['budget'],
                                    fields.get('bid_strategy'), existing['id'],
                                )
                            )
                            updated += 1
                            self._register_amazon_ad_import_row(ctx, fields, existing['id'])
                        else:
                            batch_dup_err = self._validate_amazon_ad_name_unique_in_batch(
                                ad_level, fields['name'],
                                portfolio_id=fields.get('portfolio_id'),
                                campaign_id=fields.get('campaign_id'),
                                shop_id=scope_shop_id,
                                batch_seen=batch_seen_names,
                            )
                            if batch_dup_err:
                                errors.append({'row': row_idx, 'message': batch_dup_err})
                                continue
                            cur.execute(
                                """
                                INSERT INTO amazon_ad_items (
                                    ad_level, sku_family_id, shop_id, portfolio_id, campaign_id,
                                    strategy_code, subtype_id, name, is_shared_budget, status, budget, bid_strategy
                                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                """,
                                (
                                    fields['ad_level'], fields['sku_family_id'], fields.get('shop_id'),
                                    fields['portfolio_id'], fields['campaign_id'], fields['strategy_code'],
                                    fields['subtype_id'], fields['name'], fields['is_shared_budget'],
                                    fields['status'], fields['budget'], fields.get('bid_strategy'),
                                )
                            )
                            new_id = cur.lastrowid
                            created += 1
                            if ad_level in ('campaign', 'group'):
                                self._create_subtype_default_targets_for_ad_item(
                                    cur, new_id, fields.get('subtype_id'), ad_level,
                                )
                            batch_seen_names.add(self._amazon_ad_batch_scope_key(
                                ad_level, fields['name'],
                                portfolio_id=fields.get('portfolio_id'),
                                campaign_id=fields.get('campaign_id'),
                                shop_id=scope_shop_id,
                            ))
                            self._register_amazon_ad_import_row(ctx, fields, new_id)

            return self.send_json(
                {
                    'status': 'success',
                    'created': created,
                    'updated': updated,
                    'unchanged': unchanged,
                    'skipped_sample_rows': skipped_sample_rows,
                    'errors': errors,
                },
                start_response
            )
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    _VALID_AD_RECORD_STATUS = ('启动', '暂停', '存档')

    _AMAZON_AD_PRODUCT_LIST_SELECT = """
        SELECT p.*, i.name AS ad_name, i.ad_level, sp.platform_sku
        FROM amazon_ad_products p
        INNER JOIN amazon_ad_items i ON i.id = p.ad_item_id
        LEFT JOIN sales_products sp ON sp.id = p.sales_product_id
    """

    _AMAZON_AD_TARGET_LIST_SELECT = """
        SELECT t.*, i.name AS ad_name, i.ad_level
        FROM amazon_ad_targets t
        INNER JOIN amazon_ad_items i ON i.id = t.ad_item_id
    """

    def _normalize_ad_record_status(self, value):
        text = (value or '').strip() or '启动'
        if text not in self._VALID_AD_RECORD_STATUS:
            return None, f'无效状态: {text}'
        return text, None

    def _format_observe_interval_days(self, days):
        parsed = self._parse_int(days)
        if parsed is None:
            return None
        parsed = max(0, int(parsed))
        return f'{parsed}天'

    def _parse_observe_interval_days(self, text):
        raw = str(text or '').strip()
        if not raw:
            return 1
        parsed = self._parse_int(raw)
        if parsed is not None:
            return max(0, int(parsed))
        m = re.search(r'\d+', raw)
        if m:
            return max(0, int(m.group(0)))
        return 1

    def _build_observe_fields(self, observe_days, updated_at, next_observe_at=None):
        days = self._parse_observe_interval_days(observe_days)
        interval_text = self._format_observe_interval_days(days)
        updated_dt = self._parse_datetime_local_value(updated_at)
        next_dt = self._parse_datetime_local_value(next_observe_at)
        if updated_dt and not next_dt and days:
            try:
                base = datetime.fromisoformat(updated_dt.replace(' ', 'T', 1))
                next_dt = (base + timedelta(days=days)).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                next_dt = None
        return interval_text, updated_dt, next_dt

    def _validate_ad_item_for_relation(self, cur, ad_item_id, allowed_levels):
        ad_item_id = self._parse_int(ad_item_id)
        if not ad_item_id:
            return None, '请选择广告关联'
        cur.execute(
            "SELECT id, ad_level FROM amazon_ad_items WHERE id=%s LIMIT 1",
            (ad_item_id,),
        )
        row = cur.fetchone()
        if not row:
            return None, '广告关联不存在'
        if row.get('ad_level') not in allowed_levels:
            labels = {'campaign': '广告活动', 'group': '广告组'}
            expected = '或'.join(labels.get(x, x) for x in allowed_levels)
            return None, f'广告关联须为{expected}'
        return row, None

    def _resolve_sales_product_id_from_import_text(self, cur, sku_text, sku_by_platform=None):
        text = (sku_text or '').strip()
        if not text:
            return None, '请填写投放商品'
        if sku_by_platform and text in sku_by_platform:
            return sku_by_platform[text], None
        parsed_id = self._parse_int(text)
        if parsed_id:
            cur.execute("SELECT id FROM sales_products WHERE id=%s LIMIT 1", (parsed_id,))
            row = cur.fetchone()
            if row:
                return row['id'], None
        cur.execute(
            "SELECT id FROM sales_products WHERE platform_sku=%s LIMIT 1",
            (text,),
        )
        row = cur.fetchone()
        if row:
            return row['id'], None
        return None, f'未找到投放商品: {text}'

    def _resolve_import_ad_item_level(
        self, cur, shop_text, portfolio_name, campaign_name, group_name='', expected_level=None,
    ):
        ad_item_id, ad_level, err = self._resolve_ad_item_by_four_attrs(
            cur, shop_text, portfolio_name, campaign_name, group_name,
        )
        if err:
            return None, None, err
        if expected_level and ad_level != expected_level:
            labels = {'campaign': '广告活动', 'group': '广告组'}
            return None, None, f'须定位到{labels.get(expected_level, expected_level)}层级'
        return ad_item_id, ad_level, None

    def _load_amazon_ad_product_import_context(self, cur):
        ctx = self._load_amazon_ad_import_context(cur)
        cur.execute(
            """
            SELECT id, platform_sku FROM sales_products
            WHERE platform_sku IS NOT NULL AND TRIM(platform_sku) <> ''
            """
        )
        sku_by_platform = {}
        for row in cur.fetchall() or []:
            sku = str(row.get('platform_sku') or '').strip()
            if sku:
                sku_by_platform[sku] = int(row['id'])
        cur.execute("SELECT id, ad_item_id, sales_product_id FROM amazon_ad_products")
        product_by_key = {}
        for row in cur.fetchall() or []:
            key = (int(row['ad_item_id']), int(row['sales_product_id']))
            product_by_key[key] = int(row['id'])
        ctx['sku_by_platform'] = sku_by_platform
        ctx['sales_product_ids'] = set(sku_by_platform.values())
        ctx['product_by_key'] = product_by_key
        return ctx

    def _infer_target_import_ad_level(self, campaign_name, group_name):
        campaign_name = (campaign_name or '').strip()
        group_name = (group_name or '').strip()
        if not campaign_name:
            return None, '请填写广告活动'
        if group_name:
            return 'group', None
        return 'campaign', None

    def _load_amazon_ad_target_import_context(self, cur):
        ctx = self._load_amazon_ad_import_context(cur)
        cur.execute("SELECT id, ad_item_id, target_desc FROM amazon_ad_targets")
        target_by_key = {}
        for row in cur.fetchall() or []:
            key = (int(row['ad_item_id']), str(row.get('target_desc') or '').strip())
            target_by_key[key] = int(row['id'])
        ctx['target_by_key'] = target_by_key
        return ctx

    def _load_amazon_ad_adjustment_import_context(self, cur):
        """预加载广告项与操作类型索引，避免逐行 SELECT。"""
        ctx = self._load_amazon_ad_import_context(cur)

        ad_item_by_id = {}
        cur.execute(
            """
            SELECT id, ad_level, subtype_id, campaign_id
            FROM amazon_ad_items
            WHERE ad_level IN ('portfolio', 'campaign', 'group')
            """
        )
        for row in cur.fetchall() or []:
            rid = int(self._parse_int(row.get('id')) or 0)
            if not rid:
                continue
            ad_item_by_id[rid] = {
                'id': rid,
                'ad_level': str(row.get('ad_level') or '').strip(),
                'subtype_id': self._parse_int(row.get('subtype_id')),
                'campaign_id': self._parse_int(row.get('campaign_id')),
            }
        ctx['ad_item_by_id'] = ad_item_by_id

        all_ops = []
        op_by_id = {}
        cur.execute(
            """
            SELECT id, name, apply_portfolio, apply_campaign, apply_group, reason_names
            FROM amazon_ad_operation_types
            ORDER BY sort_order ASC, id ASC
            """
        )
        for row in cur.fetchall() or []:
            reason_names = self._parse_operation_type_reason_names(row.get('reason_names'))
            op = {
                'id': int(row['id']),
                'name': str(row.get('name') or '').strip(),
                'apply_portfolio': int(row.get('apply_portfolio') or 0),
                'apply_campaign': int(row.get('apply_campaign') or 0),
                'apply_group': int(row.get('apply_group') or 0),
                'reason_names': reason_names,
                'reason_set': set(reason_names),
            }
            all_ops.append(op)
            op_by_id[op['id']] = op

        ops_by_subtype = {}
        cur.execute(
            "SELECT subtype_id, operation_type_id FROM amazon_ad_subtype_operation_types"
        )
        for row in cur.fetchall() or []:
            sid = int(self._parse_int(row.get('subtype_id')) or 0)
            oid = int(self._parse_int(row.get('operation_type_id')) or 0)
            if not sid or oid not in op_by_id:
                continue
            ops_by_subtype.setdefault(sid, []).append(op_by_id[oid])

        ctx['all_ops'] = all_ops
        ctx['op_by_id'] = op_by_id
        ctx['ops_by_subtype'] = ops_by_subtype
        ctx['portfolio_ops'] = [o for o in all_ops if o['apply_portfolio']]
        ctx['campaign_ops'] = [o for o in all_ops if o['apply_campaign']]
        ctx['group_ops'] = [o for o in all_ops if o['apply_group']]
        ctx['allowed_ops_by_ad'] = {}
        return ctx

    def _get_adjustment_allowed_ops_for_ad_ctx(self, ctx, ad_item_id):
        ad_item_id = int(ad_item_id)
        cache = ctx.get('allowed_ops_by_ad') or {}
        if ad_item_id in cache:
            return cache[ad_item_id]

        ad_row = (ctx.get('ad_item_by_id') or {}).get(ad_item_id)
        if not ad_row:
            cache[ad_item_id] = None
            ctx['allowed_ops_by_ad'] = cache
            return None

        level = ad_row.get('ad_level')
        subtype_id = ad_row.get('subtype_id')
        if level == 'group' and ad_row.get('campaign_id'):
            camp = (ctx.get('ad_item_by_id') or {}).get(int(ad_row['campaign_id']))
            if camp and camp.get('subtype_id'):
                subtype_id = camp.get('subtype_id')

        if subtype_id:
            raw_ops = (ctx.get('ops_by_subtype') or {}).get(int(subtype_id), [])
        elif level == 'portfolio':
            raw_ops = ctx.get('portfolio_ops') or []
        elif level == 'campaign':
            raw_ops = ctx.get('campaign_ops') or []
        else:
            raw_ops = ctx.get('group_ops') or []

        allowed = []
        for op in raw_ops:
            if level == 'portfolio' and not op.get('apply_portfolio'):
                continue
            if level == 'campaign' and not op.get('apply_campaign'):
                continue
            if level == 'group' and not op.get('apply_group'):
                continue
            allowed.append(op)

        cache[ad_item_id] = allowed
        ctx['allowed_ops_by_ad'] = cache
        return allowed

    def _validate_adjustment_import_operation_ctx(self, ctx, ad_item_id, operation_name, reason_name):
        operation_name = (operation_name or '').strip()
        if not operation_name:
            return None, None, '请填写操作'
        allowed = self._get_adjustment_allowed_ops_for_ad_ctx(ctx, ad_item_id)
        if allowed is None:
            return None, None, '关联广告不存在'
        op_match = next((x for x in allowed if (x.get('name') or '').strip() == operation_name), None)
        if not op_match:
            return None, None, f'操作「{operation_name}」不适用于该广告'
        reason_name = (reason_name or '').strip() or None
        if reason_name and reason_name not in op_match.get('reason_set', set()):
            return None, None, f'操作原因「{reason_name}」不属于操作「{operation_name}」'
        return op_match['id'], reason_name, None

    def _adjustment_import_dedupe_key(
        self, ad_item_id, op_id, target_object, adjust_date, before_value, after_value, reason_name,
    ):
        def norm_text(value):
            return str(value or '').strip()

        adj = adjust_date
        if isinstance(adj, datetime):
            adj = adj.strftime('%Y-%m-%d %H:%M:%S')
        else:
            adj = norm_text(adj)
        return (
            int(ad_item_id),
            int(op_id),
            norm_text(target_object),
            adj,
            norm_text(before_value),
            norm_text(after_value),
            norm_text(reason_name),
        )

    def _load_adjustment_import_existing_keys(self, cur, ad_item_ids):
        ids = sorted({int(x) for x in (ad_item_ids or []) if int(x or 0) > 0})
        if not ids:
            return set()
        keys = set()
        chunk_size = 500
        for offset in range(0, len(ids), chunk_size):
            part = ids[offset:offset + chunk_size]
            placeholders = ','.join(['%s'] * len(part))
            cur.execute(
                f"""
                SELECT ad_item_id, operation_type_id, target_object, adjust_date,
                       before_value, after_value, reason_name
                FROM amazon_ad_adjustments
                WHERE ad_item_id IN ({placeholders})
                """,
                tuple(part),
            )
            for row in cur.fetchall() or []:
                adj = row.get('adjust_date')
                if isinstance(adj, datetime):
                    adj = adj.strftime('%Y-%m-%d %H:%M:%S')
                keys.add(self._adjustment_import_dedupe_key(
                    row.get('ad_item_id'),
                    row.get('operation_type_id'),
                    row.get('target_object'),
                    adj,
                    row.get('before_value'),
                    row.get('after_value'),
                    row.get('reason_name'),
                ))
        return keys

    def _append_child_import_error(self, errors, row_idx, message):
        if len(errors) >= self._AMAZON_AD_CHILD_IMPORT_MAX_ERRORS:
            return False
        errors.append({'row': row_idx, 'message': message})
        return True

    def _child_import_success_payload(self, created, updated, errors, skipped_sample_rows):
        truncated = len(errors) >= self._AMAZON_AD_CHILD_IMPORT_MAX_ERRORS
        payload = {
            'status': 'success',
            'created': created,
            'updated': updated,
            'unchanged': 0,
            'skipped_sample_rows': skipped_sample_rows,
            'errors': errors,
        }
        if truncated:
            payload['errors_truncated'] = True
            payload['errors_message'] = f'仅展示前 {self._AMAZON_AD_CHILD_IMPORT_MAX_ERRORS} 条错误'
        return payload

    def _executemany_in_chunks(self, cur, sql, rows, chunk_size=None):
        if not rows:
            return
        size = int(chunk_size or self._AMAZON_AD_CHILD_IMPORT_BATCH_SIZE)
        for offset in range(0, len(rows), size):
            cur.executemany(sql, rows[offset:offset + size])

    def _import_sheet_header_map(self, ws):
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}
        return {
            str(name or '').strip(): idx
            for idx, name in enumerate(header_row)
            if str(name or '').strip()
        }

    def _import_cell_text(self, row, header_map, name):
        idx = header_map.get(name)
        if idx is None or idx >= len(row):
            return None
        value = row[idx]
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if float(value) == int(value):
                return str(int(value))
            return str(value)
        return str(value).strip()

    def _resolve_shop_id_from_import_ctx(self, ctx, shop_text):
        text = (shop_text or '').strip()
        if not text:
            return None, '请填写店铺'
        shop_id = ctx.get('shop_by_name', {}).get(text)
        if shop_id:
            return int(shop_id), None
        parsed_id = self._parse_int(text)
        if parsed_id and parsed_id in ctx.get('shop_by_id', {}):
            return int(parsed_id), None
        return None, f'未找到店铺: {text}'

    def _resolve_ad_item_by_four_attrs_ctx(
        self, ctx, shop_text, portfolio_name, campaign_name='', group_name='',
    ):
        shop_id, shop_err = self._resolve_shop_id_from_import_ctx(ctx, shop_text)
        if shop_err:
            return None, None, shop_err
        portfolio_name = (portfolio_name or '').strip()
        campaign_name = (campaign_name or '').strip()
        group_name = (group_name or '').strip()
        if not portfolio_name:
            return None, None, '请填写广告组合'

        portfolio_id = ctx.get('portfolio_by_name', {}).get((int(shop_id), portfolio_name))
        if not portfolio_id:
            return None, None, f'未找到该店铺下的广告组合: {portfolio_name}'
        portfolio_id = int(portfolio_id)

        if group_name:
            if not campaign_name:
                return None, None, '填写广告组时须同时填写广告活动'
            campaign_row = ctx.get('campaign_by_key', {}).get((int(shop_id), portfolio_id, campaign_name))
            if not campaign_row:
                return None, None, f'在组合「{portfolio_name}」下未找到广告活动: {campaign_name}'
            campaign_id = int(campaign_row.get('id') or 0)
            group_row = ctx.get('existing_group_by_key', {}).get((int(shop_id), campaign_id, group_name))
            if not group_row:
                return None, None, f'在活动「{campaign_name}」下未找到广告组: {group_name}'
            return int(group_row.get('id') or 0), 'group', None

        if campaign_name:
            campaign_row = ctx.get('campaign_by_key', {}).get((int(shop_id), portfolio_id, campaign_name))
            if not campaign_row:
                return None, None, f'在组合「{portfolio_name}」下未找到广告活动: {campaign_name}'
            return int(campaign_row.get('id') or 0), 'campaign', None

        return portfolio_id, 'portfolio', None

    def _resolve_import_ad_item_level_ctx(
        self, ctx, shop_text, portfolio_name, campaign_name, group_name='', expected_level=None,
    ):
        ad_item_id, ad_level, err = self._resolve_ad_item_by_four_attrs_ctx(
            ctx, shop_text, portfolio_name, campaign_name, group_name,
        )
        if err:
            return None, None, err
        if expected_level and ad_level != expected_level:
            labels = {'campaign': '广告活动', 'group': '广告组'}
            return None, None, f'须定位到{labels.get(expected_level, expected_level)}层级'
        return ad_item_id, ad_level, None

    def _resolve_sales_product_id_from_import_ctx(self, sku_text, ctx):
        text = (sku_text or '').strip()
        if not text:
            return None, '请填写投放商品'
        sku_by_platform = ctx.get('sku_by_platform') or {}
        if text in sku_by_platform:
            return sku_by_platform[text], None
        parsed_id = self._parse_int(text)
        if parsed_id and int(parsed_id) in (ctx.get('sales_product_ids') or ()):
            return int(parsed_id), None
        return None, f'未找到投放商品: {text}'

    def _read_batch_import_workbook(self, environ):
        if load_workbook is None:
            return None, f'openpyxl not available: {_openpyxl_import_error}'
        content_type = environ.get('CONTENT_TYPE', '')
        if 'multipart/form-data' not in content_type:
            return None, 'Invalid content type'
        content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
        raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
        env_copy = dict(environ)
        env_copy['CONTENT_LENGTH'] = str(len(raw_body))
        form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
        file_item = form['file'] if 'file' in form else None
        if file_item is None or getattr(file_item, 'file', None) is None:
            return None, 'Missing file'
        file_bytes = file_item.file.read() or b''
        if not file_bytes:
            return None, 'Empty file'
        return load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True), None

    def _is_amazon_ad_child_template_sample_row(self, row_idx, marker_text, *, max_example_row=2):
        if row_idx <= max_example_row:
            return True
        text = str(marker_text or '').strip()
        return ('示例' in text)

    def _amazon_ad_product_template_headers(self):
        return [
            '状态*', '店铺*', '广告组合*', '广告活动*', '广告组*', '投放商品*',
            '最后修改时间*', '下次观察时间间隔（天）', '下次观察时间',
        ]

    def _amazon_ad_target_template_headers(self):
        return [
            '状态*', '店铺*', '广告组合*', '广告活动*', '广告组',
            '投放描述*', '竞价*', '最后修改时间*', '下次观察时间间隔（天）', '下次观察时间',
        ]

    def _build_amazon_ad_product_import_workbook(self):
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation

        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                _, _, shop_names, portfolio_names = self._load_amazon_ad_items_template_options()
                ctx = self._load_amazon_ad_product_import_context(cur)
                platform_skus = sorted((ctx.get('sku_by_platform') or {}).keys())

        example_shop = shop_names[0] if shop_names else '示例店铺'
        example_port = portfolio_names[0] if portfolio_names else '示例-Short-SKU01'
        example_sku = platform_skus[0] if platform_skus else '示例-SKU'

        wb = Workbook()
        ws = wb.active
        ws.title = '广告商品'
        headers = self._amazon_ad_product_template_headers()
        ws.append(headers)
        ws.append([
            '启动', example_shop, example_port, 'BE-示例组合-SP-KW', 'BE-示例组合-SP-KW',
            example_sku, '2026-06-08 10:00', '1', '2026-06-09 10:00',
        ])

        example_font = Font(italic=True, color='7B8088')
        example_fill = PatternFill(start_color='E8EEF6', end_color='E8EEF6', fill_type='solid')
        for cell in ws[2]:
            cell.fill = example_fill
            cell.font = example_font

        options_ws = wb.create_sheet('_options')
        options_ws.sheet_state = 'hidden'
        options_ws.cell(row=1, column=1, value='status')
        for idx, status in enumerate(self._VALID_AD_RECORD_STATUS, start=2):
            options_ws.cell(row=idx, column=1, value=status)
        options_ws.cell(row=1, column=2, value='shop_name')
        for idx, name in enumerate(shop_names, start=2):
            options_ws.cell(row=idx, column=2, value=name)
        options_ws.cell(row=1, column=3, value='portfolio_name')
        for idx, name in enumerate(portfolio_names, start=2):
            options_ws.cell(row=idx, column=3, value=name)
        options_ws.cell(row=1, column=4, value='platform_sku')
        for idx, sku in enumerate(platform_skus, start=2):
            options_ws.cell(row=idx, column=4, value=sku)

        first_data_row = 3
        data_end = 1200
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color='2A2420')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        dv_status = DataValidation(type='list', formula1="='_options'!$A$2:$A$4", allow_blank=False)
        ws.add_data_validation(dv_status)
        dv_status.add(f'A{first_data_row}:A{data_end}')
        if shop_names:
            shop_end = len(shop_names) + 1
            dv_shop = DataValidation(
                type='list',
                formula1=f"='_options'!$B$2:$B${shop_end}",
                allow_blank=False,
            )
            ws.add_data_validation(dv_shop)
            dv_shop.add(f'B{first_data_row}:B{data_end}')
        if portfolio_names:
            port_end = len(portfolio_names) + 1
            dv_port = DataValidation(
                type='list',
                formula1=f"='_options'!$C$2:$C${port_end}",
                allow_blank=False,
            )
            ws.add_data_validation(dv_port)
            dv_port.add(f'C{first_data_row}:C{data_end}')
        if platform_skus:
            sku_end = len(platform_skus) + 1
            dv_sku = DataValidation(
                type='list',
                formula1=f"='_options'!$D$2:$D${sku_end}",
                allow_blank=False,
            )
            ws.add_data_validation(dv_sku)
            dv_sku.add(f'F{first_data_row}:F{data_end}')

        guide = wb.create_sheet('填写说明')
        guide.append(['字段', '必填', '说明'])
        for row in [
            ('状态*', '是', '启动 / 暂停 / 存档'),
            ('店铺*', '是', '下拉为店铺名称'),
            ('广告组合* / 广告活动* / 广告组*', '是', '四元组定位广告组（与调整记录一致）'),
            ('投放商品*', '是', '下拉为销售平台 SKU'),
            ('最后修改时间*', '是', '如 2026-06-08 10:00'),
            ('下次观察时间间隔（天）', '否', '默认 1 天'),
            ('下次观察时间', '否', '留空则按间隔自动计算'),
            ('', '', '第2行为示例，导入从第3行填写'),
        ]:
            guide.append(list(row))
        guide.column_dimensions['A'].width = 28
        guide.column_dimensions['B'].width = 10
        guide.column_dimensions['C'].width = 44
        return wb

    def _build_amazon_ad_target_import_workbook(self):
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation

        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                _, _, shop_names, portfolio_names = self._load_amazon_ad_items_template_options()

        example_shop = shop_names[0] if shop_names else '示例店铺'
        example_port = portfolio_names[0] if portfolio_names else '示例-Short-SKU01'
        example_camp = 'BE-示例组合-SP-KW'

        wb = Workbook()
        ws = wb.active
        ws.title = '广告投放'
        headers = self._amazon_ad_target_template_headers()
        ws.append(headers)
        ws.append([
            '启动', example_shop, example_port, example_camp, '',
            '示例投放描述-活动', '0.35', '2026-06-08 10:00', '1', '2026-06-09 10:00',
        ])
        ws.append([
            '启动', example_shop, example_port, example_camp, example_camp,
            '示例投放描述-组', '18%', '2026-06-08 10:00', '3', '2026-06-11 10:00',
        ])

        example_font = Font(italic=True, color='7B8088')
        example_fill = PatternFill(start_color='E8EEF6', end_color='E8EEF6', fill_type='solid')
        for row_idx in (2, 3):
            for cell in ws[row_idx]:
                cell.fill = example_fill
                cell.font = example_font

        options_ws = wb.create_sheet('_options')
        options_ws.sheet_state = 'hidden'
        options_ws.cell(row=1, column=1, value='status')
        for idx, status in enumerate(self._VALID_AD_RECORD_STATUS, start=2):
            options_ws.cell(row=idx, column=1, value=status)
        options_ws.cell(row=1, column=2, value='shop_name')
        for idx, name in enumerate(shop_names, start=2):
            options_ws.cell(row=idx, column=2, value=name)
        options_ws.cell(row=1, column=3, value='portfolio_name')
        for idx, name in enumerate(portfolio_names, start=2):
            options_ws.cell(row=idx, column=3, value=name)

        first_data_row = 4
        data_end = 1200
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color='2A2420')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        dv_status = DataValidation(type='list', formula1="='_options'!$A$2:$A$4", allow_blank=False)
        ws.add_data_validation(dv_status)
        dv_status.add(f'A{first_data_row}:A{data_end}')
        if shop_names:
            shop_end = len(shop_names) + 1
            dv_shop = DataValidation(
                type='list',
                formula1=f"='_options'!$B$2:$B${shop_end}",
                allow_blank=False,
            )
            ws.add_data_validation(dv_shop)
            dv_shop.add(f'B{first_data_row}:B{data_end}')
        if portfolio_names:
            port_end = len(portfolio_names) + 1
            dv_port = DataValidation(
                type='list',
                formula1=f"='_options'!$C$2:$C${port_end}",
                allow_blank=False,
            )
            ws.add_data_validation(dv_port)
            dv_port.add(f'C{first_data_row}:C{data_end}')

        ws.freeze_panes = f'A{first_data_row}'

        guide = wb.create_sheet('填写说明')
        guide.append(['字段', '必填', '说明'])
        for row in [
            ('状态*', '是', '启动 / 暂停 / 存档'),
            ('店铺* / 广告组合* / 广告活动*', '是', '下拉选择店铺与组合；活动名称必填'),
            ('广告组', '组层级', '填写则定位广告组；留空则定位广告活动'),
            ('投放描述* / 竞价*', '是', '竞价支持金额或百分比'),
            ('最后修改时间*', '是', '如 2026-06-08 10:00'),
            ('', '', '第2–3行为示例，导入从第4行填写；冻结窗格在示例行下方'),
        ]:
            guide.append(list(row))
        guide.column_dimensions['A'].width = 28
        guide.column_dimensions['B'].width = 10
        guide.column_dimensions['C'].width = 46
        return wb

    def handle_amazon_ad_target_api(self, environ, method, start_response):
        """Amazon 广告投放（target）API（CRUD）"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                sql = self._AMAZON_AD_TARGET_LIST_SELECT + ' WHERE 1=1'
                params = []
                if keyword:
                    like = f'%{keyword}%'
                    sql += ' AND (i.name LIKE %s OR t.target_desc LIKE %s OR t.bid_value LIKE %s)'
                    params.extend([like, like, like])
                sql += ' ORDER BY t.id DESC LIMIT 500'
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(sql, tuple(params))
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            data = self._read_json_body(environ) or {}

            if method == 'PATCH':
                batch_items = data.get('items')
                if isinstance(batch_items, list) and batch_items:
                    updated = 0
                    errors = []
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            for raw in batch_items:
                                item_id = self._parse_int((raw or {}).get('id'))
                                bid_value = str((raw or {}).get('bid_value') or '').strip()
                                if not item_id:
                                    errors.append({'id': 0, 'error': 'Missing id'})
                                    continue
                                if not bid_value:
                                    errors.append({'id': item_id, 'error': '竞价不能为空'})
                                    continue
                                try:
                                    cur.execute(
                                        'UPDATE amazon_ad_targets SET bid_value=%s WHERE id=%s',
                                        (bid_value, item_id),
                                    )
                                    if cur.rowcount <= 0:
                                        errors.append({'id': item_id, 'error': '记录不存在'})
                                    else:
                                        updated += 1
                                except Exception as ex:
                                    errors.append({'id': item_id, 'error': str(ex)})
                    return self.send_json({
                        'status': 'success',
                        'updated': updated,
                        'errors': errors,
                    }, start_response)

                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                updates = []
                params = []
                if 'status' in data:
                    status, err = self._normalize_ad_record_status(data.get('status'))
                    if err:
                        return self.send_json({'status': 'error', 'message': err}, start_response)
                    updates.append('status=%s')
                    params.append(status)
                if 'bid_value' in data:
                    bid_value = (data.get('bid_value') or '').strip()
                    if not bid_value:
                        return self.send_json({'status': 'error', 'message': '竞价不能为空'}, start_response)
                    updates.append('bid_value=%s')
                    params.append(bid_value)
                if not updates:
                    return self.send_json({'status': 'error', 'message': '无更新字段'}, start_response)
                params.append(item_id)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            f"UPDATE amazon_ad_targets SET {', '.join(updates)} WHERE id=%s",
                            tuple(params),
                        )
                        if cur.rowcount <= 0:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_ad_targets WHERE id=%s", (item_id,))
                        if cur.rowcount <= 0:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                return self.send_json({'status': 'success'}, start_response)

            status, err = self._normalize_ad_record_status(data.get('status'))
            if err:
                return self.send_json({'status': 'error', 'message': err}, start_response)
            target_desc = (data.get('target_desc') or data.get('delivery_desc') or '').strip()
            bid_value = (data.get('bid_value') or '').strip()
            if not target_desc or not bid_value:
                return self.send_json({'status': 'error', 'message': '投放描述与竞价为必填'}, start_response)
            interval_text, updated_dt, next_dt = self._build_observe_fields(
                data.get('observe_days'), data.get('updated_at'), data.get('next_observe_at'),
            )
            if not updated_dt:
                return self.send_json({'status': 'error', 'message': '最后修改时间不能为空'}, start_response)

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    ad_row, ad_err = self._validate_ad_item_for_relation(
                        cur, data.get('ad_item_id'), ('campaign', 'group'),
                    )
                    if ad_err:
                        return self.send_json({'status': 'error', 'message': ad_err}, start_response)

                    if method == 'POST':
                        cur.execute(
                            """
                            SELECT id FROM amazon_ad_targets
                            WHERE ad_item_id=%s AND target_desc=%s LIMIT 1
                            """,
                            (ad_row['id'], target_desc),
                        )
                        if cur.fetchone():
                            return self.send_json(
                                {'status': 'error', 'message': '该广告下已存在相同投放描述'},
                                start_response,
                            )
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_targets (
                                status, ad_item_id, target_desc, bid_value,
                                observe_interval, next_observe_at, updated_at
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                status, ad_row['id'], target_desc, bid_value,
                                interval_text, next_dt, updated_dt,
                            ),
                        )
                        return self.send_json({'status': 'success', 'id': cur.lastrowid}, start_response)

                    if method == 'PUT':
                        item_id = self._parse_int(data.get('id'))
                        if not item_id:
                            return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                        cur.execute(
                            """
                            SELECT id FROM amazon_ad_targets
                            WHERE ad_item_id=%s AND target_desc=%s AND id<>%s LIMIT 1
                            """,
                            (ad_row['id'], target_desc, item_id),
                        )
                        if cur.fetchone():
                            return self.send_json(
                                {'status': 'error', 'message': '该广告下已存在相同投放描述'},
                                start_response,
                            )
                        cur.execute(
                            """
                            UPDATE amazon_ad_targets SET
                                status=%s, ad_item_id=%s, target_desc=%s, bid_value=%s,
                                observe_interval=%s, next_observe_at=%s, updated_at=%s
                            WHERE id=%s
                            """,
                            (
                                status, ad_row['id'], target_desc, bid_value,
                                interval_text, next_dt, updated_dt, item_id,
                            ),
                        )
                        if cur.rowcount <= 0:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                        return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_product_api(self, environ, method, start_response):
        """Amazon 广告商品 API（CRUD）"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                sql = self._AMAZON_AD_PRODUCT_LIST_SELECT + ' WHERE 1=1'
                params = []
                if keyword:
                    like = f'%{keyword}%'
                    sql += ' AND (i.name LIKE %s OR sp.platform_sku LIKE %s)'
                    params.extend([like, like])
                sql += ' ORDER BY p.id DESC LIMIT 500'
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(sql, tuple(params))
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            data = self._read_json_body(environ) or {}

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_ad_products WHERE id=%s", (item_id,))
                        if cur.rowcount <= 0:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                return self.send_json({'status': 'success'}, start_response)

            status, err = self._normalize_ad_record_status(data.get('status'))
            if err:
                return self.send_json({'status': 'error', 'message': err}, start_response)
            sales_product_id = self._parse_int(data.get('sales_product_id'))
            if not sales_product_id:
                return self.send_json({'status': 'error', 'message': '请选择投放商品'}, start_response)
            interval_text, updated_dt, next_dt = self._build_observe_fields(
                data.get('observe_days'), data.get('updated_at'), data.get('next_observe_at'),
            )
            if not updated_dt:
                return self.send_json({'status': 'error', 'message': '最后修改时间不能为空'}, start_response)

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    ad_row, ad_err = self._validate_ad_item_for_relation(cur, data.get('ad_item_id'), ('group',))
                    if ad_err:
                        return self.send_json({'status': 'error', 'message': ad_err}, start_response)
                    cur.execute(
                        "SELECT id FROM sales_products WHERE id=%s LIMIT 1",
                        (sales_product_id,),
                    )
                    if not cur.fetchone():
                        return self.send_json({'status': 'error', 'message': '投放商品不存在'}, start_response)

                    if method == 'POST':
                        cur.execute(
                            """
                            SELECT id FROM amazon_ad_products
                            WHERE ad_item_id=%s AND sales_product_id=%s LIMIT 1
                            """,
                            (ad_row['id'], sales_product_id),
                        )
                        if cur.fetchone():
                            return self.send_json(
                                {'status': 'error', 'message': '该广告组下已存在相同投放商品'},
                                start_response,
                            )
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_products (
                                status, ad_item_id, sales_product_id,
                                observe_interval, next_observe_at, updated_at
                            ) VALUES (%s, %s, %s, %s, %s, %s)
                            """,
                            (
                                status, ad_row['id'], sales_product_id,
                                interval_text, next_dt, updated_dt,
                            ),
                        )
                        return self.send_json({'status': 'success', 'id': cur.lastrowid}, start_response)

                    if method == 'PUT':
                        item_id = self._parse_int(data.get('id'))
                        if not item_id:
                            return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                        cur.execute(
                            """
                            SELECT id FROM amazon_ad_products
                            WHERE ad_item_id=%s AND sales_product_id=%s AND id<>%s LIMIT 1
                            """,
                            (ad_row['id'], sales_product_id, item_id),
                        )
                        if cur.fetchone():
                            return self.send_json(
                                {'status': 'error', 'message': '该广告组下已存在相同投放商品'},
                                start_response,
                            )
                        cur.execute(
                            """
                            UPDATE amazon_ad_products SET
                                status=%s, ad_item_id=%s, sales_product_id=%s,
                                observe_interval=%s, next_observe_at=%s, updated_at=%s
                            WHERE id=%s
                            """,
                            (
                                status, ad_row['id'], sales_product_id,
                                interval_text, next_dt, updated_dt, item_id,
                            ),
                        )
                        if cur.rowcount <= 0:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                        return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_target_template_api(self, environ, method, start_response):
        """广告投放批量导入模板下载"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json(
                    {'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'},
                    start_response,
                )
            wb = self._build_amazon_ad_target_import_workbook()
            return self._send_excel_workbook(wb, '广告投放导入模板.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_target_import_api(self, environ, method, start_response):
        """广告投放批量导入（预加载索引 + 批量写入）"""
        wb = None
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            wb, wb_err = self._read_batch_import_workbook(environ)
            if wb_err:
                return self.send_json({'status': 'error', 'message': wb_err}, start_response)

            ws = wb.active
            header_map = self._import_sheet_header_map(ws)
            if not header_map:
                return self.send_json({'status': 'error', 'message': '模板表头为空'}, start_response)

            created = updated = 0
            skipped_sample_rows = 0
            errors = []
            insert_rows = []
            update_rows = []
            batch_keys = set()
            max_row = min(
                int(ws.max_row or 2),
                2 + self._AMAZON_AD_CHILD_IMPORT_MAX_ROWS,
            )

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    ctx = self._load_amazon_ad_target_import_context(cur)
                    target_by_key = ctx.get('target_by_key') or {}

                    for row_idx, row in enumerate(
                        ws.iter_rows(min_row=2, max_row=max_row, values_only=True),
                        start=2,
                    ):
                        if not any(v is not None and str(v).strip() for v in row):
                            continue

                        def cell(name, _row=row):
                            return self._import_cell_text(_row, header_map, name)

                        desc = cell('投放描述*') or ''
                        if self._is_amazon_ad_child_template_sample_row(row_idx, desc, max_example_row=3):
                            skipped_sample_rows += 1
                            continue
                        if not desc and not (cell('店铺*') or ''):
                            continue

                        status, status_err = self._normalize_ad_record_status(cell('状态*'))
                        if status_err:
                            if not self._append_child_import_error(errors, row_idx, status_err):
                                break
                            continue

                        campaign_name = cell('广告活动*') or ''
                        group_name = cell('广告组') or ''
                        expected_level, level_err = self._infer_target_import_ad_level(campaign_name, group_name)
                        if level_err:
                            if not self._append_child_import_error(errors, row_idx, level_err):
                                break
                            continue

                        ad_item_id, _, ad_err = self._resolve_import_ad_item_level_ctx(
                            ctx,
                            cell('店铺*') or '',
                            cell('广告组合*') or '',
                            campaign_name,
                            group_name,
                            expected_level,
                        )
                        if ad_err:
                            if not self._append_child_import_error(errors, row_idx, ad_err):
                                break
                            continue

                        bid_value = (cell('竞价*') or '').strip()
                        if not bid_value:
                            if not self._append_child_import_error(errors, row_idx, '竞价不能为空'):
                                break
                            continue

                        interval_text, updated_dt, next_dt = self._build_observe_fields(
                            cell('下次观察时间间隔（天）'),
                            cell('最后修改时间*') or '',
                            cell('下次观察时间'),
                        )
                        if not updated_dt:
                            if not self._append_child_import_error(errors, row_idx, '最后修改时间不能为空'):
                                break
                            continue

                        dedupe_key = (int(ad_item_id), desc)
                        if dedupe_key in batch_keys:
                            if not self._append_child_import_error(errors, row_idx, '本批导入中广告关联+投放描述重复'):
                                break
                            continue
                        batch_keys.add(dedupe_key)

                        existing_id = target_by_key.get(dedupe_key)
                        if existing_id:
                            update_rows.append((
                                status, bid_value, interval_text, next_dt, updated_dt, existing_id,
                            ))
                            updated += 1
                        else:
                            insert_rows.append((
                                status, ad_item_id, desc, bid_value, interval_text, next_dt, updated_dt,
                            ))
                            target_by_key[dedupe_key] = None
                            created += 1

                    self._executemany_in_chunks(
                        cur,
                        """
                        INSERT INTO amazon_ad_targets (
                            status, ad_item_id, target_desc, bid_value,
                            observe_interval, next_observe_at, updated_at
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """,
                        insert_rows,
                    )
                    self._executemany_in_chunks(
                        cur,
                        """
                        UPDATE amazon_ad_targets SET
                            status=%s, bid_value=%s,
                            observe_interval=%s, next_observe_at=%s, updated_at=%s
                        WHERE id=%s
                        """,
                        update_rows,
                    )

            return self.send_json(
                self._child_import_success_payload(created, updated, errors, skipped_sample_rows),
                start_response,
            )
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def handle_amazon_ad_product_template_api(self, environ, method, start_response):
        """广告商品批量导入模板下载"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json(
                    {'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'},
                    start_response,
                )
            wb = self._build_amazon_ad_product_import_workbook()
            return self._send_excel_workbook(wb, '广告商品导入模板.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_product_import_api(self, environ, method, start_response):
        """广告商品批量导入（预加载索引 + 批量写入）"""
        wb = None
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            wb, wb_err = self._read_batch_import_workbook(environ)
            if wb_err:
                return self.send_json({'status': 'error', 'message': wb_err}, start_response)

            ws = wb.active
            header_map = self._import_sheet_header_map(ws)
            if not header_map:
                return self.send_json({'status': 'error', 'message': '模板表头为空'}, start_response)

            created = updated = 0
            skipped_sample_rows = 0
            errors = []
            insert_rows = []
            update_rows = []
            batch_keys = set()
            max_row = min(
                int(ws.max_row or 2),
                2 + self._AMAZON_AD_CHILD_IMPORT_MAX_ROWS,
            )

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    ctx = self._load_amazon_ad_product_import_context(cur)
                    product_by_key = ctx.get('product_by_key') or {}

                    for row_idx, row in enumerate(
                        ws.iter_rows(min_row=2, max_row=max_row, values_only=True),
                        start=2,
                    ):
                        if not any(v is not None and str(v).strip() for v in row):
                            continue

                        def cell(name, _row=row):
                            return self._import_cell_text(_row, header_map, name)

                        sku_text = cell('投放商品*') or ''
                        if self._is_amazon_ad_child_template_sample_row(row_idx, sku_text):
                            skipped_sample_rows += 1
                            continue
                        if not sku_text and not (cell('店铺*') or ''):
                            continue

                        status, status_err = self._normalize_ad_record_status(cell('状态*'))
                        if status_err:
                            if not self._append_child_import_error(errors, row_idx, status_err):
                                break
                            continue

                        ad_item_id, _, ad_err = self._resolve_import_ad_item_level_ctx(
                            ctx,
                            cell('店铺*') or '',
                            cell('广告组合*') or '',
                            cell('广告活动*') or '',
                            cell('广告组*') or '',
                            'group',
                        )
                        if ad_err:
                            if not self._append_child_import_error(errors, row_idx, ad_err):
                                break
                            continue

                        sales_product_id, sku_err = self._resolve_sales_product_id_from_import_ctx(sku_text, ctx)
                        if sku_err:
                            if not self._append_child_import_error(errors, row_idx, sku_err):
                                break
                            continue

                        interval_text, updated_dt, next_dt = self._build_observe_fields(
                            cell('下次观察时间间隔（天）'),
                            cell('最后修改时间*') or '',
                            cell('下次观察时间'),
                        )
                        if not updated_dt:
                            if not self._append_child_import_error(errors, row_idx, '最后修改时间不能为空'):
                                break
                            continue

                        dedupe_key = (int(ad_item_id), int(sales_product_id))
                        if dedupe_key in batch_keys:
                            if not self._append_child_import_error(errors, row_idx, '本批导入中广告组+投放商品重复'):
                                break
                            continue
                        batch_keys.add(dedupe_key)

                        existing_id = product_by_key.get(dedupe_key)
                        if existing_id:
                            update_rows.append((
                                status, interval_text, next_dt, updated_dt, existing_id,
                            ))
                            updated += 1
                        else:
                            insert_rows.append((
                                status, ad_item_id, sales_product_id, interval_text, next_dt, updated_dt,
                            ))
                            product_by_key[dedupe_key] = None
                            created += 1

                    self._executemany_in_chunks(
                        cur,
                        """
                        INSERT INTO amazon_ad_products (
                            status, ad_item_id, sales_product_id,
                            observe_interval, next_observe_at, updated_at
                        ) VALUES (%s, %s, %s, %s, %s, %s)
                        """,
                        insert_rows,
                    )
                    self._executemany_in_chunks(
                        cur,
                        """
                        UPDATE amazon_ad_products SET
                            status=%s, observe_interval=%s,
                            next_observe_at=%s, updated_at=%s
                        WHERE id=%s
                        """,
                        update_rows,
                    )

            return self.send_json(
                self._child_import_success_payload(created, updated, errors, skipped_sample_rows),
                start_response,
            )
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def _parse_datetime_local_value(self, value):
        text = (value or '').strip()
        if not text:
            return None
        if 'T' in text and len(text) == 16:
            text = text + ':00'
        return text.replace('T', ' ')

    def _adjustment_patch_text_value(self, value):
        return (value or '').strip() or None

    def _build_adjustment_patch_fields(self, raw):
        if not isinstance(raw, dict):
            return {}
        datetime_fields = {'adjust_date', 'start_time', 'end_time'}
        text_fields = {
            'target_object',
            'impressions', 'clicks', 'cost', 'orders', 'sales',
            'acos', 'cpc', 'ctr', 'cvr', 'top_of_search_is',
            'attribution_orders', 'attribution_sales', 'remark',
        }
        patch = {}
        for field in datetime_fields:
            if field in raw:
                patch[field] = self._parse_datetime_local_value(raw.get(field))
        for field in text_fields:
            if field in raw:
                patch[field] = self._adjustment_patch_text_value(raw.get(field))
        if 'attribution_checked' in raw:
            patch['attribution_checked'] = 1 if str(raw.get('attribution_checked')) in ('1', 'true', 'True') else 0
        elif 'attribution_orders' in raw or 'attribution_sales' in raw:
            orders = patch.get('attribution_orders')
            sales = patch.get('attribution_sales')
            patch['attribution_checked'] = 1 if (orders or sales) else 0
        return patch

    def _amazon_ad_adjustment_batch_patch(self, cur, items, chunk_size=120):
        """批量 PATCH：按列 CASE WHEN 合并 UPDATE，减少数据库往返。"""
        errors = []
        patches = []
        seen_ids = set()
        for raw in items:
            if not isinstance(raw, dict):
                errors.append({'message': '无效数据项'})
                continue
            item_id = self._parse_int(raw.get('id'))
            if not item_id:
                errors.append({'id': raw.get('id'), 'message': '无效 id'})
                continue
            if item_id in seen_ids:
                continue
            seen_ids.add(item_id)
            if 'target_object' in raw and not (raw.get('target_object') or '').strip():
                errors.append({'id': item_id, 'message': '对象不能为空'})
                continue
            if 'adjust_date' in raw and not self._parse_datetime_local_value(raw.get('adjust_date')):
                errors.append({'id': item_id, 'message': '调整日期不能为空'})
                continue
            patch_fields = self._build_adjustment_patch_fields(raw)
            if not patch_fields:
                errors.append({'id': item_id, 'message': '无可更新字段'})
                continue
            patches.append((item_id, patch_fields))

        if not patches:
            return 0, errors

        id_list = [item_id for item_id, _ in patches]
        placeholders = ','.join(['%s'] * len(id_list))
        cur.execute(
            f'SELECT id FROM amazon_ad_adjustments WHERE id IN ({placeholders})',
            tuple(id_list),
        )
        found = {
            self._parse_int(row.get('id'))
            for row in (cur.fetchall() or [])
            if self._parse_int(row.get('id'))
        }
        valid_patches = []
        for item_id, patch_fields in patches:
            if item_id not in found:
                errors.append({'id': item_id, 'message': '记录不存在'})
                continue
            valid_patches.append((item_id, patch_fields))
        if not valid_patches:
            return 0, errors

        updated = 0
        patch_by_id = {item_id: patch_fields for item_id, patch_fields in valid_patches}
        all_ids = list(patch_by_id.keys())
        size = max(1, int(chunk_size or 120))

        for offset in range(0, len(all_ids), size):
            chunk_ids = all_ids[offset:offset + size]
            all_columns = set()
            for rid in chunk_ids:
                all_columns.update(patch_by_id[rid].keys())
            set_parts = []
            params = []
            for col in sorted(all_columns):
                when_parts = []
                for rid in chunk_ids:
                    patch_fields = patch_by_id[rid]
                    if col not in patch_fields:
                        continue
                    when_parts.append('WHEN %s THEN %s')
                    params.extend([rid, patch_fields[col]])
                if when_parts:
                    set_parts.append(f'{col} = CASE id {" ".join(when_parts)} ELSE {col} END')
            if not set_parts:
                continue
            in_ph = ','.join(['%s'] * len(chunk_ids))
            sql = f'UPDATE amazon_ad_adjustments SET {", ".join(set_parts)} WHERE id IN ({in_ph})'
            cur.execute(sql, tuple(params + chunk_ids))
            updated += int(cur.rowcount or 0)

        return updated, errors

    def _adjustment_portfolio_name_from_row(self, row):
        if not row:
            return ''
        text = (row.get('portfolio_name') or '').strip()
        if text:
            return text
        if row.get('ad_level') == 'portfolio':
            return (row.get('name') or '').strip()
        return ''

    def _serialize_adjustment_ad_list_item(self, row):
        level = row.get('ad_level')
        ad_class = row.get('ad_class') or ''
        subtype_code = row.get('subtype_code') or ''
        if ad_class and subtype_code:
            ad_type_text = f'{ad_class}-{subtype_code}'
        else:
            ad_type_text = row.get('subtype_description') or ''
        name = row.get('name') or ''
        portfolio_name = self._adjustment_portfolio_name_from_row(row)
        campaign_name = row.get('campaign_name') or ''
        portfolio_id = self._parse_int(row.get('portfolio_id'))
        campaign_id = self._parse_int(row.get('campaign_id'))
        if level == 'portfolio':
            portfolio_id = self._parse_int(row.get('id'))
        elif level == 'campaign':
            campaign_id = self._parse_int(row.get('id'))
        return {
            'id': row.get('id'),
            'ad_name': name,
            'ad_level': level,
            'status': row.get('status') or '启动',
            'ad_type_text': ad_type_text,
            'portfolio_name': portfolio_name,
            'campaign_name': campaign_name if level == 'group' else (name if level == 'campaign' else ''),
            'group_name': name if level == 'group' else '',
            'portfolio_id': portfolio_id,
            'campaign_id': campaign_id,
        }

    def _fetch_adjustment_ad_info(self, cur, ad_item_id):
        cur.execute(self._AMAZON_AD_ITEM_SELECT + ' WHERE i.id=%s LIMIT 1', (ad_item_id,))
        row = cur.fetchone()
        if not row:
            return None, None
        item = self._serialize_amazon_ad_item(row)
        level = item.get('ad_level')
        ad_class = item.get('ad_class') or ''
        subtype_code = item.get('subtype_code') or ''
        if ad_class and subtype_code:
            ad_type_text = f'{ad_class}-{subtype_code}'
        else:
            ad_type_text = item.get('subtype_description') or ''
        ad_info = {
            'ad_type_text': ad_type_text,
            'portfolio_name': self._adjustment_portfolio_name_from_row(item),
            'campaign_name': item.get('campaign_name') if level == 'group' else (item.get('name') if level == 'campaign' else ''),
            'group_name': item.get('name') if level == 'group' else '',
            'ad_name': item.get('name') or '',
            'status': item.get('status') or '启动',
            'budget': item.get('budget'),
            'bid_strategy': item.get('bid_strategy') or '',
        }
        if level == 'portfolio':
            ad_info['portfolio_name'] = item.get('name') or ad_info['portfolio_name']
            ad_info['campaign_name'] = ''
            ad_info['group_name'] = ''
        return item, ad_info

    def _resolve_ad_item_shop_id(self, cur, ad_item_id):
        ad_item_id = self._parse_int(ad_item_id)
        if not ad_item_id:
            return None
        cur.execute(
            """
            SELECT COALESCE(
                CASE WHEN i.ad_level = 'portfolio' THEN i.shop_id END,
                p.shop_id
            ) AS shop_id
            FROM amazon_ad_items i
            LEFT JOIN amazon_ad_items c ON i.campaign_id = c.id AND c.ad_level = 'campaign'
            LEFT JOIN amazon_ad_items p ON p.id = COALESCE(NULLIF(i.portfolio_id, 0), c.portfolio_id)
                AND p.ad_level = 'portfolio'
            WHERE i.id=%s
            LIMIT 1
            """,
            (ad_item_id,),
        )
        row = cur.fetchone() or {}
        shop_id = self._parse_int(row.get('shop_id'))
        return shop_id or None

    def _fetch_shop_platform_skus(self, cur, shop_id):
        shop_id = self._parse_int(shop_id)
        if not shop_id:
            return set()
        cur.execute(
            """
            SELECT platform_sku
            FROM sales_products
            WHERE shop_id=%s AND platform_sku IS NOT NULL AND TRIM(platform_sku) <> ''
            """,
            (shop_id,),
        )
        out = set()
        for row in cur.fetchall() or []:
            sku = str(row.get('platform_sku') or '').strip()
            if sku:
                out.add(sku)
        return out

    def _fetch_ad_item_default_bid(self, cur, ad_item_id):
        ad_item_id = self._parse_int(ad_item_id)
        if not ad_item_id:
            return ''
        cur.execute(
            """
            SELECT bid_value
            FROM amazon_ad_targets
            WHERE ad_item_id=%s AND bid_value IS NOT NULL AND TRIM(bid_value) <> ''
            ORDER BY id ASC
            LIMIT 1
            """,
            (ad_item_id,),
        )
        row = cur.fetchone() or {}
        bid = str(row.get('bid_value') or '').strip()
        if bid:
            return bid
        cur.execute(
            """
            SELECT i.ad_level, i.campaign_id
            FROM amazon_ad_items i
            WHERE i.id=%s
            LIMIT 1
            """,
            (ad_item_id,),
        )
        ad_row = cur.fetchone() or {}
        campaign_id = self._parse_int(ad_row.get('campaign_id'))
        if str(ad_row.get('ad_level') or '').strip() == 'group' and campaign_id:
            cur.execute(
                """
                SELECT bid_value
                FROM amazon_ad_targets
                WHERE ad_item_id=%s AND bid_value IS NOT NULL AND TRIM(bid_value) <> ''
                ORDER BY id ASC
                LIMIT 1
                """,
                (campaign_id,),
            )
            camp_row = cur.fetchone() or {}
            return str(camp_row.get('bid_value') or '').strip()
        return ''

    def _resolve_sales_product_id_for_ad_shop(self, cur, ad_item_id, platform_sku):
        platform_sku = (platform_sku or '').strip()
        if not platform_sku:
            return None, '请填写操作对象'
        shop_id = self._resolve_ad_item_shop_id(cur, ad_item_id)
        if not shop_id:
            return None, '无法解析广告归属店铺'
        cur.execute(
            """
            SELECT id FROM sales_products
            WHERE shop_id=%s AND platform_sku=%s
            LIMIT 1
            """,
            (shop_id, platform_sku),
        )
        row = cur.fetchone()
        if not row:
            return None, f'未找到店铺销售平台SKU: {platform_sku}'
        return int(row['id']), None

    def _is_numeric_bid_text(self, value):
        text = str(value or '').strip().replace(',', '')
        if not text or not re.fullmatch(r'-?\d+(\.\d+)?', text):
            return False
        try:
            float(text)
            return True
        except (TypeError, ValueError):
            return False

    def _fetch_allowed_operations_for_ad(self, cur, ad_row):
        if not ad_row:
            return []
        level = ad_row.get('ad_level')
        subtype_id = ad_row.get('subtype_id')
        if level == 'group' and ad_row.get('campaign_id'):
            cur.execute(
                "SELECT subtype_id FROM amazon_ad_items WHERE id=%s AND ad_level='campaign' LIMIT 1",
                (ad_row.get('campaign_id'),)
            )
            camp = cur.fetchone() or {}
            subtype_id = camp.get('subtype_id') or subtype_id

        if subtype_id:
            cur.execute(
                """
                SELECT ot.id, ot.name, ot.apply_portfolio, ot.apply_campaign, ot.apply_group, ot.reason_names
                FROM amazon_ad_operation_types ot
                INNER JOIN amazon_ad_subtype_operation_types link ON link.operation_type_id = ot.id
                WHERE link.subtype_id = %s
                ORDER BY ot.sort_order ASC, ot.id ASC
                """,
                (subtype_id,)
            )
        elif level == 'portfolio':
            cur.execute(
                "SELECT id, name, apply_portfolio, apply_campaign, apply_group, reason_names "
                "FROM amazon_ad_operation_types WHERE apply_portfolio=1 ORDER BY sort_order ASC, id ASC"
            )
        elif level == 'campaign':
            cur.execute(
                "SELECT id, name, apply_portfolio, apply_campaign, apply_group, reason_names "
                "FROM amazon_ad_operation_types WHERE apply_campaign=1 ORDER BY sort_order ASC, id ASC"
            )
        else:
            cur.execute(
                "SELECT id, name, apply_portfolio, apply_campaign, apply_group, reason_names "
                "FROM amazon_ad_operation_types WHERE apply_group=1 ORDER BY sort_order ASC, id ASC"
            )
        ops = cur.fetchall() or []
        result = []
        for op in ops:
            if level == 'portfolio' and not int(op.get('apply_portfolio') or 0):
                continue
            if level == 'campaign' and not int(op.get('apply_campaign') or 0):
                continue
            if level == 'group' and not int(op.get('apply_group') or 0):
                continue
            reason_names = self._parse_operation_type_reason_names(op.get('reason_names'))
            result.append({
                'id': op['id'],
                'name': op.get('name') or '',
                'reasons': [{'reason_name': name} for name in reason_names],
            })
        return result

    def _parse_adjustment_datetime(self, value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.replace(second=0, microsecond=0)
        try:
            return datetime.fromisoformat(str(value).replace(' ', 'T', 1)).replace(second=0, microsecond=0)
        except Exception:
            return None

    def _fetch_last_adjustment_datetime(self, cur, ad_item_id=None, target_object=None):
        clauses = []
        params = []
        if ad_item_id is not None:
            clauses.append('ad_item_id=%s')
            params.append(ad_item_id)
        target_object = (target_object or '').strip()
        if target_object and target_object != '-':
            clauses.append('target_object=%s')
            params.append(target_object)
        where_sql = ' AND '.join(clauses) if clauses else '1=1'
        cur.execute(
            f"""
            SELECT adjust_date FROM amazon_ad_adjustments
            WHERE {where_sql}
            ORDER BY id DESC
            LIMIT 1
            """,
            tuple(params),
        )
        row = cur.fetchone() or {}
        return self._parse_adjustment_datetime(row.get('adjust_date'))

    def _adjustment_defaults_for_ad(self, cur, ad_item_id, target_object=None):
        now = datetime.now().replace(second=0, microsecond=0)
        target_object = (target_object or '').strip()
        last_adjust = self._fetch_last_adjustment_datetime(cur, ad_item_id, target_object)
        if not last_adjust and target_object and target_object != '-':
            last_adjust = self._fetch_last_adjustment_datetime(cur, ad_item_id, None)
        if not last_adjust:
            last_adjust = self._fetch_last_adjustment_datetime(cur, None, None)
        if last_adjust:
            start_time = last_adjust.replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            start_time = now - timedelta(days=7)
        end_time = now
        return {
            'adjust_date': now.strftime('%Y-%m-%d %H:%M:%S'),
            'start_time': start_time.strftime('%Y-%m-%d %H:%M:%S'),
            'end_time': end_time.strftime('%Y-%m-%d %H:%M:%S'),
        }

    def _fetch_observe_defaults_for_target_object(self, cur, ad_item_id, target_object, op_name=None):
        target_object = (target_object or '').strip()
        default_days = 1
        if not ad_item_id or not target_object or target_object == '-':
            return {
                'observe_interval': self._format_observe_interval_days(default_days),
                'observe_days': default_days,
            }
        op_name = self._normalize_adjustment_operation_type_name(op_name)
        use_product = bool(op_name and self._is_modify_product_operation(op_name))
        use_target = bool(
            op_name
            and (
                self._is_modify_delivery_target_operation(op_name)
                or self._is_modify_placement_operation(op_name)
            )
        )
        row = None
        if use_product:
            cur.execute(
                """
                SELECT p.observe_interval
                FROM amazon_ad_products p
                LEFT JOIN sales_products sp ON sp.id = p.sales_product_id
                WHERE p.ad_item_id=%s AND sp.platform_sku=%s
                LIMIT 1
                """,
                (ad_item_id, target_object),
            )
            row = cur.fetchone()
        elif use_target:
            cur.execute(
                """
                SELECT observe_interval
                FROM amazon_ad_targets
                WHERE ad_item_id=%s AND target_desc=%s
                LIMIT 1
                """,
                (ad_item_id, target_object),
            )
            row = cur.fetchone()
        else:
            cur.execute(
                """
                SELECT observe_interval
                FROM amazon_ad_targets
                WHERE ad_item_id=%s AND target_desc=%s
                LIMIT 1
                """,
                (ad_item_id, target_object),
            )
            row = cur.fetchone()
            if not row:
                cur.execute(
                    """
                    SELECT p.observe_interval
                    FROM amazon_ad_products p
                    LEFT JOIN sales_products sp ON sp.id = p.sales_product_id
                    WHERE p.ad_item_id=%s AND sp.platform_sku=%s
                    LIMIT 1
                    """,
                    (ad_item_id, target_object),
                )
                row = cur.fetchone()
        if row and row.get('observe_interval'):
            days = self._parse_observe_interval_days(row.get('observe_interval'))
            return {
                'observe_interval': self._format_observe_interval_days(days),
                'observe_days': days,
            }
        return {
            'observe_interval': self._format_observe_interval_days(default_days),
            'observe_days': default_days,
        }

    def _apply_adjustment_observe_sync(
        self, cur, ad_item_id, op_name, target_object, adjust_date, observe_days, next_observe_at,
    ):
        op_name = self._normalize_adjustment_operation_type_name(op_name)
        if not (
            self._is_modify_product_operation(op_name)
            or self._is_modify_delivery_target_operation(op_name)
            or self._is_modify_placement_operation(op_name)
        ):
            return None
        target_object = (target_object or '').strip()
        if not target_object or target_object == '-':
            return None
        interval_text, updated_dt, next_dt = self._build_observe_fields(
            observe_days, adjust_date, next_observe_at,
        )
        updated_dt = updated_dt or adjust_date
        if self._is_modify_product_operation(op_name):
            cur.execute(
                """
                UPDATE amazon_ad_products p
                INNER JOIN sales_products sp ON sp.id = p.sales_product_id
                SET p.observe_interval=%s, p.next_observe_at=%s, p.updated_at=%s
                WHERE p.ad_item_id=%s AND sp.platform_sku=%s
                """,
                (interval_text, next_dt, updated_dt, ad_item_id, target_object),
            )
        else:
            cur.execute(
                """
                UPDATE amazon_ad_targets
                SET observe_interval=%s, next_observe_at=%s, updated_at=%s
                WHERE ad_item_id=%s AND target_desc=%s
                """,
                (interval_text, next_dt, updated_dt, ad_item_id, target_object),
            )
        return None

    def _normalize_adjustment_operation_type_name(self, name):
        return re.sub(r"[『』【】「」]", '', str(name or '')).strip()

    def _is_modify_delivery_target_operation(self, op_name):
        n = self._normalize_adjustment_operation_type_name(op_name)
        return '修改' in n and '投放' in n and '广告位' not in n

    def _is_modify_placement_operation(self, op_name):
        n = self._normalize_adjustment_operation_type_name(op_name)
        return '修改' in n and '广告位' in n

    def _is_modify_product_operation(self, op_name):
        n = self._normalize_adjustment_operation_type_name(op_name)
        return '修改' in n and '商品' in n

    def _adjustment_observe_active_status(self, status):
        return str(status or '启动').strip() == '启动'

    def _fetch_adjustment_ad_item_brief(self, cur, ad_item_id):
        ad_item_id = self._parse_int(ad_item_id)
        if not ad_item_id:
            return None
        cur.execute(
            """
            SELECT id, ad_level, status, portfolio_id, campaign_id, subtype_id
            FROM amazon_ad_items
            WHERE id=%s
            LIMIT 1
            """,
            (ad_item_id,),
        )
        return cur.fetchone()

    def _adjustment_observe_ad_chain_valid(self, cur, ad_item_id):
        row = self._fetch_adjustment_ad_item_brief(cur, ad_item_id)
        if not row or not self._adjustment_observe_active_status(row.get('status')):
            return False
        level = str(row.get('ad_level') or '').strip()
        if level == 'group':
            campaign_id = self._parse_int(row.get('campaign_id'))
            if not campaign_id:
                return False
            campaign = self._fetch_adjustment_ad_item_brief(cur, campaign_id)
            if not campaign or not self._adjustment_observe_active_status(campaign.get('status')):
                return False
            portfolio_id = self._parse_int(campaign.get('portfolio_id'))
            if portfolio_id:
                portfolio = self._fetch_adjustment_ad_item_brief(cur, portfolio_id)
                if not portfolio or not self._adjustment_observe_active_status(portfolio.get('status')):
                    return False
        elif level == 'campaign':
            portfolio_id = self._parse_int(row.get('portfolio_id'))
            if portfolio_id:
                portfolio = self._fetch_adjustment_ad_item_brief(cur, portfolio_id)
                if not portfolio or not self._adjustment_observe_active_status(portfolio.get('status')):
                    return False
        return True

    def _fetch_adjustment_group_ids_for_campaign(self, cur, campaign_id):
        campaign_id = self._parse_int(campaign_id)
        if not campaign_id:
            return []
        cur.execute(
            """
            SELECT id
            FROM amazon_ad_items
            WHERE ad_level='group' AND campaign_id=%s
            ORDER BY id ASC
            """,
            (campaign_id,),
        )
        out = []
        for row in cur.fetchall() or []:
            gid = self._parse_int(row.get('id'))
            if gid and self._adjustment_observe_ad_chain_valid(cur, gid):
                out.append(gid)
        return out

    def _append_adjustment_observe_ad_item(self, ordered, seen, cur, ad_item_id):
        ad_item_id = self._parse_int(ad_item_id)
        if not ad_item_id or ad_item_id in seen:
            return
        if not self._adjustment_observe_ad_chain_valid(cur, ad_item_id):
            return
        seen.add(ad_item_id)
        ordered.append(ad_item_id)

    def _append_adjustment_observe_campaign_scope(
        self, ordered, seen, cur, campaign_id, include_campaign=True, include_groups=True,
    ):
        campaign_id = self._parse_int(campaign_id)
        if not campaign_id:
            return
        if include_campaign:
            self._append_adjustment_observe_ad_item(ordered, seen, cur, campaign_id)
        if include_groups:
            for group_id in self._fetch_adjustment_group_ids_for_campaign(cur, campaign_id):
                self._append_adjustment_observe_ad_item(ordered, seen, cur, group_id)

    def _append_adjustment_observe_portfolio_campaigns(self, ordered, seen, cur, portfolio_id):
        portfolio_id = self._parse_int(portfolio_id)
        if not portfolio_id:
            return
        cur.execute(
            """
            SELECT id
            FROM amazon_ad_items
            WHERE ad_level='campaign' AND portfolio_id=%s AND status='启动'
            ORDER BY id ASC
            """,
            (portfolio_id,),
        )
        for row in cur.fetchall() or []:
            cid = self._parse_int(row.get('id'))
            if cid:
                self._append_adjustment_observe_campaign_scope(
                    ordered, seen, cur, cid, include_campaign=True, include_groups=True,
                )

    def _pick_operation_type_id_for_observe_kind(self, operations, kind):
        for op in operations or []:
            name = self._normalize_adjustment_operation_type_name(op.get('name'))
            if kind == 'product' and self._is_modify_product_operation(name):
                return self._parse_int(op.get('id'))
            if kind == 'placement' and self._is_modify_placement_operation(name):
                return self._parse_int(op.get('id'))
            if kind == 'delivery' and self._is_modify_delivery_target_operation(name):
                return self._parse_int(op.get('id'))
        return None

    def _operation_type_name_from_allowed(self, allowed, operation_type_id):
        operation_type_id = self._parse_int(operation_type_id)
        if not operation_type_id:
            return ''
        for op in allowed or []:
            if self._parse_int(op.get('id')) == operation_type_id:
                return str(op.get('name') or '').strip()
        return ''

    def _classify_target_observe_kind(self, cur, ad_row, target_desc, placement_cache):
        level = str(ad_row.get('ad_level') or '').strip()
        target_desc = str(target_desc or '').strip()
        if not target_desc:
            return 'delivery'
        if level == 'campaign':
            return 'placement'
        if level == 'group':
            placement_names = self._placement_target_names_for_ad_item(cur, ad_row, placement_cache)
            if target_desc.lower() in placement_names:
                return 'placement'
        return 'delivery'

    def _build_adjustment_observe_ad_item_order(self, cur, ad_row):
        if not ad_row:
            return []
        level = str(ad_row.get('ad_level') or '').strip()
        self_id = self._parse_int(ad_row.get('id'))
        portfolio_id = self._parse_int(ad_row.get('portfolio_id'))
        campaign_id = self._parse_int(ad_row.get('campaign_id'))
        ordered = []
        seen = set()

        if level == 'group':
            self._append_adjustment_observe_ad_item(ordered, seen, cur, self_id)
            self._append_adjustment_observe_campaign_scope(
                ordered, seen, cur, campaign_id, include_campaign=True, include_groups=True,
            )
            if campaign_id and portfolio_id:
                cur.execute(
                    """
                    SELECT id
                    FROM amazon_ad_items
                    WHERE ad_level='campaign' AND portfolio_id=%s AND id>%s AND status='启动'
                    ORDER BY id ASC
                    """,
                    (portfolio_id, campaign_id),
                )
                for row in cur.fetchall() or []:
                    cid = self._parse_int(row.get('id'))
                    if cid:
                        self._append_adjustment_observe_campaign_scope(
                            ordered, seen, cur, cid, include_campaign=True, include_groups=True,
                        )
                cur.execute(
                    """
                    SELECT id
                    FROM amazon_ad_items
                    WHERE ad_level='portfolio' AND id>%s AND status='启动'
                    ORDER BY id ASC
                    """,
                    (portfolio_id,),
                )
                for row in cur.fetchall() or []:
                    pid = self._parse_int(row.get('id'))
                    if pid:
                        self._append_adjustment_observe_portfolio_campaigns(ordered, seen, cur, pid)
        elif level == 'campaign':
            self._append_adjustment_observe_ad_item(ordered, seen, cur, self_id)
            self._append_adjustment_observe_campaign_scope(
                ordered, seen, cur, self_id, include_campaign=False, include_groups=True,
            )
            if portfolio_id:
                cur.execute(
                    """
                    SELECT id
                    FROM amazon_ad_items
                    WHERE ad_level='campaign' AND portfolio_id=%s AND id>%s AND status='启动'
                    ORDER BY id ASC
                    """,
                    (portfolio_id, self_id),
                )
                for row in cur.fetchall() or []:
                    cid = self._parse_int(row.get('id'))
                    if cid:
                        self._append_adjustment_observe_campaign_scope(
                            ordered, seen, cur, cid, include_campaign=True, include_groups=True,
                        )
                cur.execute(
                    """
                    SELECT id
                    FROM amazon_ad_items
                    WHERE ad_level='portfolio' AND id>%s AND status='启动'
                    ORDER BY id ASC
                    """,
                    (portfolio_id,),
                )
                for row in cur.fetchall() or []:
                    pid = self._parse_int(row.get('id'))
                    if pid:
                        self._append_adjustment_observe_portfolio_campaigns(ordered, seen, cur, pid)
        elif level == 'portfolio':
            self._append_adjustment_observe_portfolio_campaigns(ordered, seen, cur, self_id)
            cur.execute(
                """
                SELECT id
                FROM amazon_ad_items
                WHERE ad_level='portfolio' AND id>%s AND status='启动'
                ORDER BY id ASC
                """,
                (self_id,),
            )
            for row in cur.fetchall() or []:
                pid = self._parse_int(row.get('id'))
                if pid:
                    self._append_adjustment_observe_portfolio_campaigns(ordered, seen, cur, pid)
        return ordered

    def _placement_target_names_for_ad_item(self, cur, ad_item_row, cache):
        if not ad_item_row:
            return set()
        subtype_id = self._parse_int(ad_item_row.get('subtype_id'))
        ad_level = str(ad_item_row.get('ad_level') or '').strip()
        if ad_level == 'group' and not subtype_id:
            campaign_id = self._parse_int(ad_item_row.get('campaign_id'))
            if campaign_id:
                cur.execute(
                    "SELECT subtype_id FROM amazon_ad_items WHERE id=%s AND ad_level='campaign' LIMIT 1",
                    (campaign_id,),
                )
                camp = cur.fetchone() or {}
                subtype_id = self._parse_int(camp.get('subtype_id')) or subtype_id
        if ad_level not in ('campaign', 'group') or not subtype_id:
            return set()
        key = (subtype_id, ad_level)
        if key not in cache:
            defaults = self._fetch_subtype_default_targets(cur, subtype_id, ad_level)
            cache[key] = {str(item.get('name') or '').strip().lower() for item in defaults if item.get('name')}
        return cache[key]

    def _adjustment_observe_kind_rank(self, kind):
        if kind == 'delivery':
            return 0
        if kind == 'product':
            return 1
        if kind == 'placement':
            return 2
        return 9

    def _collect_adjustment_observe_candidates(self, cur, ad_item_ids, allowed_kinds):
        if not ad_item_ids:
            return []
        placement_cache = {}
        ad_rows = {}
        for ad_item_id in ad_item_ids:
            ad_rows[ad_item_id] = self._fetch_adjustment_ad_item_brief(cur, ad_item_id)

        placeholders = ','.join(['%s'] * len(ad_item_ids))
        candidates = []

        if 'delivery' in allowed_kinds or 'placement' in allowed_kinds:
            cur.execute(
                f"""
                SELECT ad_item_id, id, target_desc, next_observe_at
                FROM amazon_ad_targets
                WHERE ad_item_id IN ({placeholders})
                  AND next_observe_at IS NOT NULL
                  AND next_observe_at <= NOW()
                ORDER BY next_observe_at ASC, target_desc ASC, id ASC
                """,
                tuple(ad_item_ids),
            )
            for row in cur.fetchall() or []:
                ad_item_id = self._parse_int(row.get('ad_item_id'))
                ad_row = ad_rows.get(ad_item_id)
                if not ad_row:
                    continue
                target_desc = str(row.get('target_desc') or '').strip()
                if not target_desc:
                    continue
                kind = self._classify_target_observe_kind(cur, ad_row, target_desc, placement_cache)
                if kind not in allowed_kinds:
                    continue
                candidates.append({
                    'ad_item_id': ad_item_id,
                    'kind': kind,
                    'target_object': target_desc,
                    'next_observe_at': row.get('next_observe_at'),
                    'entity_id': self._parse_int(row.get('id')),
                })

        if 'product' in allowed_kinds:
            cur.execute(
                f"""
                SELECT p.ad_item_id, p.id, sp.platform_sku, p.next_observe_at
                FROM amazon_ad_products p
                LEFT JOIN sales_products sp ON sp.id = p.sales_product_id
                WHERE p.ad_item_id IN ({placeholders})
                  AND p.next_observe_at IS NOT NULL
                  AND p.next_observe_at <= NOW()
                  AND sp.platform_sku IS NOT NULL
                  AND TRIM(sp.platform_sku) <> ''
                ORDER BY p.next_observe_at ASC, sp.platform_sku ASC, p.id ASC
                """,
                tuple(ad_item_ids),
            )
            for row in cur.fetchall() or []:
                ad_item_id = self._parse_int(row.get('ad_item_id'))
                sku = str(row.get('platform_sku') or '').strip()
                if not sku:
                    continue
                candidates.append({
                    'ad_item_id': ad_item_id,
                    'kind': 'product',
                    'target_object': sku,
                    'next_observe_at': row.get('next_observe_at'),
                    'entity_id': self._parse_int(row.get('id')),
                })

        ad_rank = {ad_id: idx for idx, ad_id in enumerate(ad_item_ids)}

        def sort_key(item):
            return (
                ad_rank.get(item.get('ad_item_id'), 999999),
                self._adjustment_observe_kind_rank(item.get('kind')),
                str(item.get('target_object') or ''),
                self._parse_int(item.get('entity_id')) or 0,
            )

        candidates.sort(key=sort_key)
        return candidates

    def _adjustment_observe_ad_rank(self, ad_item_order, ad_item_id):
        ad_item_id = self._parse_int(ad_item_id)
        try:
            return ad_item_order.index(ad_item_id)
        except ValueError:
            return 999999

    def _first_adjustment_observe_candidate_after_ad_rank(self, candidates, ad_item_order, after_rank):
        for item in candidates:
            rank = self._adjustment_observe_ad_rank(ad_item_order, item.get('ad_item_id'))
            if rank > after_rank:
                return item
        return None

    def _infer_observe_kind_from_operation_name(self, op_name):
        op_name = self._normalize_adjustment_operation_type_name(op_name)
        if self._is_modify_product_operation(op_name):
            return 'product'
        if self._is_modify_placement_operation(op_name):
            return 'placement'
        if self._is_modify_delivery_target_operation(op_name):
            return 'delivery'
        return None

    def _find_current_adjustment_observe_match(
        self, candidates, ad_item_id, target_object, preferred_kind=None,
    ):
        target_object = (target_object or '').strip()
        if not target_object or target_object == '-':
            return None
        ad_item_id = self._parse_int(ad_item_id)
        scoped = [
            item for item in candidates
            if self._parse_int(item.get('ad_item_id')) == ad_item_id
            and str(item.get('target_object') or '').strip() == target_object
        ]
        if not scoped:
            return None
        if preferred_kind:
            for item in scoped:
                if item.get('kind') == preferred_kind:
                    return item
        return scoped[0]

    def _serialize_adjustment_observe_pick(self, cur, picked, stay_on_current=False):
        if not picked:
            return None, '已是最后一个待观察项'
        picked_ad_row = self._fetch_adjustment_ad_item_brief(cur, picked.get('ad_item_id'))
        allowed = self._fetch_allowed_operations_for_ad(cur, picked_ad_row)
        operation_type_id = self._pick_operation_type_id_for_observe_kind(allowed, picked.get('kind'))
        if not operation_type_id:
            return None, '无法匹配操作类型'
        operation_type_name = self._operation_type_name_from_allowed(allowed, operation_type_id)
        next_observe_at = picked.get('next_observe_at')
        next_observe_text = ''
        if next_observe_at is not None:
            if isinstance(next_observe_at, datetime):
                next_observe_text = next_observe_at.strftime('%Y-%m-%d %H:%M:%S')
            else:
                next_observe_text = str(next_observe_at)
        return {
            'ad_item_id': picked.get('ad_item_id'),
            'operation_type_id': operation_type_id,
            'operation_type_name': operation_type_name,
            'target_object': picked.get('target_object'),
            'kind': picked.get('kind'),
            'next_observe_at': next_observe_text,
            'stay_on_current': bool(stay_on_current),
        }, None

    def _find_next_adjustment_observe_candidate(self, cur, ad_item_id, operation_type_id, target_object):
        ad_item_id = self._parse_int(ad_item_id)
        if not ad_item_id:
            return None, '请先选择广告'
        ad_row = self._fetch_adjustment_ad_item_brief(cur, ad_item_id)
        if not ad_row:
            return None, '广告不存在'

        allowed_kinds = {'delivery', 'product', 'placement'}
        ad_item_order = self._build_adjustment_observe_ad_item_order(cur, ad_row)
        candidates = self._collect_adjustment_observe_candidates(cur, ad_item_order, allowed_kinds)
        if not candidates:
            return None, '暂无待观察项'

        target_object = (target_object or '').strip()
        preferred_kind = None
        if operation_type_id:
            cur.execute(
                "SELECT name FROM amazon_ad_operation_types WHERE id=%s LIMIT 1",
                (self._parse_int(operation_type_id),),
            )
            op_row = cur.fetchone() or {}
            preferred_kind = self._infer_observe_kind_from_operation_name(op_row.get('name'))

        current_match = self._find_current_adjustment_observe_match(
            candidates, ad_item_id, target_object, preferred_kind,
        )
        if current_match:
            return self._serialize_adjustment_observe_pick(cur, current_match, stay_on_current=True)

        current_ad_candidates = [
            item for item in candidates
            if self._parse_int(item.get('ad_item_id')) == ad_item_id
        ]
        current_ad_rank = self._adjustment_observe_ad_rank(ad_item_order, ad_item_id)

        picked = None
        if target_object and target_object != '-':
            found_idx = -1
            for idx, item in enumerate(current_ad_candidates):
                if str(item.get('target_object') or '').strip() == target_object:
                    found_idx = idx
                    break
            if found_idx >= 0:
                if found_idx + 1 < len(current_ad_candidates):
                    picked = current_ad_candidates[found_idx + 1]
                else:
                    picked = self._first_adjustment_observe_candidate_after_ad_rank(
                        candidates, ad_item_order, current_ad_rank,
                    )
            elif current_ad_candidates:
                picked = current_ad_candidates[0]
            else:
                picked = self._first_adjustment_observe_candidate_after_ad_rank(
                    candidates, ad_item_order, current_ad_rank,
                )
        elif current_ad_candidates:
            picked = current_ad_candidates[0]
        else:
            picked = self._first_adjustment_observe_candidate_after_ad_rank(
                candidates, ad_item_order, current_ad_rank,
            )

        start_idx = 0
        if picked:
            for idx, item in enumerate(candidates):
                if (
                    self._parse_int(item.get('ad_item_id')) == self._parse_int(picked.get('ad_item_id'))
                    and item.get('kind') == picked.get('kind')
                    and str(item.get('target_object') or '').strip() == str(picked.get('target_object') or '').strip()
                    and self._parse_int(item.get('entity_id')) == self._parse_int(picked.get('entity_id'))
                ):
                    start_idx = idx
                    break

        for idx in range(start_idx, len(candidates)):
            result, err = self._serialize_adjustment_observe_pick(cur, candidates[idx], stay_on_current=False)
            if result:
                return result, None
        return None, err or '已是最后一个待观察项'

    def _is_archive_operation(self, op_name):
        return self._normalize_adjustment_operation_type_name(op_name) == '存档'

    def _apply_adjustment_to_ad_item(self, cur, ad_item_id, operation_name, target_object, after_value):
        if not self._is_archive_operation(operation_name):
            return None
        target_object = (target_object or '').strip()
        if target_object != '-':
            return None
        after_value = (after_value or '').strip()
        if not after_value:
            return None
        status, err = self._normalize_ad_record_status(after_value)
        if err:
            return err
        cur.execute(
            "UPDATE amazon_ad_items SET status=%s, updated_at=NOW() WHERE id=%s",
            (status, ad_item_id),
        )
        if cur.rowcount <= 0:
            return '广告状态更新失败'
        return None

    def _insert_amazon_ad_target_row(self, cur, ad_item_id, target_desc, status, bid_value):
        now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        interval_text, updated_dt, next_dt = self._build_observe_fields(1, now_text)
        updated_dt = updated_dt or now_text
        cur.execute(
            """
            INSERT INTO amazon_ad_targets (
                status, ad_item_id, target_desc, bid_value,
                observe_interval, next_observe_at, updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (
                status, ad_item_id, target_desc, bid_value,
                interval_text, next_dt, updated_dt,
            ),
        )
        if cur.rowcount <= 0:
            return '投放写入失败'
        return None

    def _resolve_new_delivery_target_fields(self, after_value):
        after_value = (after_value or '').strip()
        status, status_err = self._normalize_ad_record_status(after_value)
        if not status_err:
            return None, status, None
        if self._is_numeric_bid_text(after_value):
            return after_value, '启动', None
        return None, None, '新建投放须填写合法竞价或状态'

    def _apply_existing_delivery_target_update(self, cur, row_id, after_value):
        status, status_err = self._normalize_ad_record_status(after_value)
        if not status_err:
            cur.execute(
                """
                UPDATE amazon_ad_targets
                SET status=%s, updated_at=NOW()
                WHERE id=%s
                """,
                (status, row_id),
            )
            if cur.rowcount <= 0:
                return '投放更新失败'
            return None
        if self._is_numeric_bid_text(after_value):
            cur.execute(
                """
                UPDATE amazon_ad_targets
                SET bid_value=%s, updated_at=NOW()
                WHERE id=%s
                """,
                (after_value, row_id),
            )
            if cur.rowcount <= 0:
                return '投放更新失败'
            return None
        return status_err

    def _apply_adjustment_to_target(self, cur, ad_item_id, operation_name, target_object, after_value):
        after_value = (after_value or '').strip()
        target_object = (target_object or '').strip()
        if not after_value or not target_object or target_object == '-' or after_value == '-':
            return None
        if not (
            self._is_modify_delivery_target_operation(operation_name)
            or self._is_modify_placement_operation(operation_name)
        ):
            return None
        cur.execute(
            """
            SELECT id FROM amazon_ad_targets
            WHERE ad_item_id=%s AND target_desc=%s
            LIMIT 1
            """,
            (ad_item_id, target_object),
        )
        row = cur.fetchone()
        if self._is_modify_delivery_target_operation(operation_name):
            if row:
                return self._apply_existing_delivery_target_update(cur, row['id'], after_value)
            bid_value, target_status, err = self._resolve_new_delivery_target_fields(after_value)
            if err:
                return err
            return self._insert_amazon_ad_target_row(
                cur, ad_item_id, target_object, target_status, bid_value,
            )

        if row:
            cur.execute(
                """
                UPDATE amazon_ad_targets
                SET bid_value=%s, updated_at=NOW()
                WHERE id=%s
                """,
                (after_value, row['id']),
            )
            if cur.rowcount <= 0:
                return '投放更新失败'
            return None

        bid_value = after_value
        if not bid_value:
            bid_value = self._fetch_ad_item_default_bid(cur, ad_item_id)
        if not bid_value:
            return '新建广告位须填写合法百分比或竞价'
        return self._insert_amazon_ad_target_row(
            cur, ad_item_id, target_object, '启动', bid_value,
        )

    def _apply_adjustment_to_product(self, cur, ad_item_id, operation_name, target_object, after_value):
        after_value = (after_value or '').strip()
        target_object = (target_object or '').strip()
        if not after_value or not target_object or target_object == '-' or after_value == '-':
            return None
        if not self._is_modify_product_operation(operation_name):
            return None

        cur.execute(
            """
            SELECT p.id, p.status
            FROM amazon_ad_products p
            LEFT JOIN sales_products sp ON sp.id = p.sales_product_id
            WHERE p.ad_item_id=%s AND sp.platform_sku=%s
            LIMIT 1
            """,
            (ad_item_id, target_object),
        )
        row = cur.fetchone()
        status, status_err = self._normalize_ad_record_status(after_value)
        if row:
            if status_err:
                return None
            cur.execute(
                """
                UPDATE amazon_ad_products
                SET status=%s, updated_at=NOW()
                WHERE id=%s
                """,
                (status, row['id']),
            )
            if cur.rowcount <= 0:
                return '商品更新失败'
            return None

        sales_product_id, sp_err = self._resolve_sales_product_id_for_ad_shop(
            cur, ad_item_id, target_object,
        )
        if sp_err:
            return sp_err

        if status_err:
            return status_err

        product_status = status
        now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        interval_text, updated_dt, next_dt = self._build_observe_fields(1, now_text)
        updated_dt = updated_dt or now_text
        cur.execute(
            """
            INSERT INTO amazon_ad_products (
                status, ad_item_id, sales_product_id,
                observe_interval, next_observe_at, updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (
                product_status, ad_item_id, sales_product_id,
                interval_text, next_dt, updated_dt,
            ),
        )
        if cur.rowcount <= 0:
            return '商品写入失败'
        return None

    def _apply_adjustment_sync(self, cur, ad_item_id, operation_name, target_object, after_value):
        ad_err = self._apply_adjustment_to_ad_item(
            cur, ad_item_id, operation_name, target_object, after_value,
        )
        if ad_err:
            return ad_err
        product_err = self._apply_adjustment_to_product(
            cur, ad_item_id, operation_name, target_object, after_value,
        )
        if product_err:
            return product_err
        return self._apply_adjustment_to_target(
            cur, ad_item_id, operation_name, target_object, after_value,
        )

    def handle_amazon_ad_adjustment_api(self, environ, method, start_response):
        """Amazon 广告调整 API（广告搜索 / 默认值 / 调整记录 CRUD）"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            action = (query_params.get('action', [''])[0] or '').strip().lower()

            if method == 'GET' and action == 'ad-search':
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            self._AMAZON_AD_ITEM_SELECT
                            + " WHERE i.ad_level IN ('portfolio', 'campaign', 'group') "
                            + " ORDER BY FIELD(i.ad_level, 'portfolio', 'campaign', 'group'), i.id DESC LIMIT 500"
                        )
                        rows = cur.fetchall() or []
                items = [self._serialize_adjustment_ad_list_item(r) for r in rows]
                return self.send_json({'status': 'success', 'items': items}, start_response)

            if method == 'GET' and action == 'target-options':
                ad_item_id = self._parse_int((query_params.get('ad_item_id', [''])[0] or '').strip())
                if not ad_item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing ad_item_id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        default_bid = self._fetch_ad_item_default_bid(cur, ad_item_id)
                        cur.execute(
                            """
                            SELECT id, status, target_desc, bid_value
                            FROM amazon_ad_targets
                            WHERE ad_item_id=%s
                            ORDER BY target_desc ASC, id ASC
                            """,
                            (ad_item_id,),
                        )
                        rows = cur.fetchall() or []
                items = []
                for row in rows:
                    bid_value = str(row.get('bid_value') or '').strip()
                    items.append({
                        'id': row.get('id'),
                        'status': row.get('status') or '启动',
                        'target_desc': row.get('target_desc') or '',
                        'bid_value': bid_value,
                    })
                return self.send_json({
                    'status': 'success',
                    'items': items,
                    'default_bid': default_bid,
                }, start_response)

            if method == 'GET' and action == 'product-options':
                ad_item_id = self._parse_int((query_params.get('ad_item_id', [''])[0] or '').strip())
                if not ad_item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing ad_item_id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        shop_id = self._resolve_ad_item_shop_id(cur, ad_item_id)
                        platform_skus = sorted(self._fetch_shop_platform_skus(cur, shop_id))
                        default_bid = self._fetch_ad_item_default_bid(cur, ad_item_id)
                        cur.execute(
                            """
                            SELECT p.id, p.status, sp.platform_sku
                            FROM amazon_ad_products p
                            LEFT JOIN sales_products sp ON sp.id = p.sales_product_id
                            WHERE p.ad_item_id=%s
                            ORDER BY sp.platform_sku ASC, p.id ASC
                            """,
                            (ad_item_id,),
                        )
                        rows = cur.fetchall() or []
                items = []
                for row in rows:
                    sku = str(row.get('platform_sku') or '').strip()
                    if not sku:
                        continue
                    items.append({
                        'id': row.get('id'),
                        'status': row.get('status') or '启动',
                        'target_desc': sku,
                    })
                return self.send_json({
                    'status': 'success',
                    'items': items,
                    'platform_skus': platform_skus,
                    'default_bid': default_bid,
                }, start_response)

            if method == 'GET' and action == 'next-observe':
                ad_item_id = self._parse_int((query_params.get('ad_item_id', [''])[0] or '').strip())
                if not ad_item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing ad_item_id'}, start_response)
                operation_type_id = self._parse_int((query_params.get('operation_type_id', [''])[0] or '').strip())
                target_object = (query_params.get('target_object', [''])[0] or '').strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        item, err = self._find_next_adjustment_observe_candidate(
                            cur, ad_item_id, operation_type_id, target_object,
                        )
                if err:
                    return self.send_json({'status': 'success', 'item': None, 'message': err}, start_response)
                return self.send_json({'status': 'success', 'item': item}, start_response)

            if method == 'GET' and action == 'defaults':
                ad_item_id = self._parse_int((query_params.get('ad_item_id', [''])[0] or '').strip())
                if not ad_item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing ad_item_id'}, start_response)
                target_object = (query_params.get('target_object', [''])[0] or '').strip()
                operation_type_id = self._parse_int((query_params.get('operation_type_id', [''])[0] or '').strip())
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        ad_row, ad_info = self._fetch_adjustment_ad_info(cur, ad_item_id)
                        if not ad_row:
                            return self.send_json({'status': 'error', 'message': '广告不存在'}, start_response)
                        allowed_operations = self._fetch_allowed_operations_for_ad(cur, ad_row)
                        defaults = self._adjustment_defaults_for_ad(cur, ad_item_id, target_object)
                        op_name = ''
                        if operation_type_id:
                            cur.execute(
                                "SELECT name FROM amazon_ad_operation_types WHERE id=%s LIMIT 1",
                                (operation_type_id,),
                            )
                            op_row = cur.fetchone() or {}
                            op_name = op_row.get('name') or ''
                        observe_defaults = self._fetch_observe_defaults_for_target_object(
                            cur, ad_item_id, target_object, op_name,
                        )
                        defaults.update(observe_defaults)
                return self.send_json({
                    'status': 'success',
                    'ad_info': ad_info,
                    'allowed_operations': allowed_operations,
                    'defaults': defaults,
                }, start_response)

            if method == 'GET':
                ad_item_id = self._parse_int((query_params.get('ad_item_id', [''])[0] or '').strip())
                sql = """
                    SELECT
                        a.*,
                        i.name AS ad_name,
                        i.ad_level,
                        ot.name AS operation_name
                    FROM amazon_ad_adjustments a
                    INNER JOIN amazon_ad_items i ON i.id = a.ad_item_id
                    LEFT JOIN amazon_ad_operation_types ot ON ot.id = a.operation_type_id
                    WHERE 1=1
                """
                params = []
                if ad_item_id:
                    sql += ' AND a.ad_item_id=%s'
                    params.append(ad_item_id)
                sql += ' ORDER BY a.id DESC'
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(sql, tuple(params))
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows, 'total': len(rows)}, start_response)

            data = self._read_json_body(environ)

            if method == 'POST':
                ad_item_id = self._parse_int(data.get('ad_item_id'))
                operation_type_id = self._parse_int(data.get('operation_type_id'))
                reason_name = (data.get('reason_name') or '').strip() or None
                target_object = (data.get('target_object') or '').strip()
                is_quick = int(data.get('is_quick_submit') or 0)
                if not ad_item_id or not operation_type_id or not target_object:
                    return self.send_json({'status': 'error', 'message': '缺少必填字段'}, start_response)
                if not is_quick:
                    for field in ('before_value', 'after_value'):
                        if not (data.get(field) or '').strip():
                            return self.send_json({'status': 'error', 'message': '完整提交请填写修改前/后'}, start_response)

                adjust_date = self._parse_datetime_local_value(data.get('adjust_date')) or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        op_err = self._validate_adjustment_operation_for_ad(
                            cur, ad_item_id, operation_type_id, reason_name
                        )
                        if op_err:
                            return self.send_json({'status': 'error', 'message': op_err}, start_response)
                        cur.execute(
                            "SELECT name FROM amazon_ad_operation_types WHERE id=%s LIMIT 1",
                            (operation_type_id,),
                        )
                        op_row = cur.fetchone() or {}
                        op_name = op_row.get('name') or ''
                        after_value = (data.get('after_value') or '').strip()
                        sync_err = self._apply_adjustment_sync(
                            cur, ad_item_id, op_name, target_object, after_value,
                        )
                        if sync_err:
                            return self.send_json({'status': 'error', 'message': sync_err}, start_response)
                        if not is_quick:
                            self._apply_adjustment_observe_sync(
                                cur,
                                ad_item_id,
                                op_name,
                                target_object,
                                adjust_date,
                                data.get('observe_days'),
                                data.get('next_observe_at'),
                            )
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_adjustments (
                                adjust_date, ad_item_id, operation_type_id, target_object,
                                before_value, after_value, reason_name,
                                start_time, end_time,
                                impressions, clicks, cost, orders, sales,
                                acos, cpc, ctr, cvr, top_of_search_is,
                                attribution_checked, attribution_orders, attribution_sales,
                                remark, is_quick_submit
                            ) VALUES (
                                %s, %s, %s, %s,
                                %s, %s, %s,
                                %s, %s,
                                %s, %s, %s, %s, %s,
                                %s, %s, %s, %s, %s,
                                %s, %s, %s,
                                %s, %s
                            )
                            """,
                            (
                                adjust_date, ad_item_id, operation_type_id, target_object,
                                (data.get('before_value') or '').strip() or None,
                                (data.get('after_value') or '').strip() or None,
                                reason_name,
                                self._parse_datetime_local_value(data.get('start_time')),
                                self._parse_datetime_local_value(data.get('end_time')),
                                (data.get('impressions') or '').strip() or None,
                                (data.get('clicks') or '').strip() or None,
                                (data.get('cost') or '').strip() or None,
                                (data.get('orders') or '').strip() or None,
                                (data.get('sales') or '').strip() or None,
                                (data.get('acos') or '').strip() or None,
                                (data.get('cpc') or '').strip() or None,
                                (data.get('ctr') or '').strip() or None,
                                (data.get('cvr') or '').strip() or None,
                                (data.get('top_of_search_is') or '').strip() or None,
                                1 if str(data.get('attribution_checked') or '0') == '1' else 0,
                                (data.get('attribution_orders') or '').strip() or None,
                                (data.get('attribution_sales') or '').strip() or None,
                                (data.get('remark') or '').strip() or None,
                                is_quick,
                            )
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PATCH':
                items = data.get('items')
                if not isinstance(items, list) or not items:
                    return self.send_json({'status': 'error', 'message': '缺少 items'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        updated, errors = self._amazon_ad_adjustment_batch_patch(cur, items)
                return self.send_json({'status': 'success', 'updated': updated, 'errors': errors}, start_response)

            if method == 'DELETE':
                action = (query_params.get('action', [''])[0] or '').strip().lower()
                if action == 'bulk_delete':
                    raw_ids = data.get('ids') if isinstance(data, dict) else None
                    if not isinstance(raw_ids, list) or not raw_ids:
                        return self.send_json({'status': 'error', 'message': '缺少有效 ids'}, start_response)
                    id_list = []
                    for raw in raw_ids:
                        item_id = self._parse_int(raw)
                        if item_id and item_id not in id_list:
                            id_list.append(item_id)
                    if not id_list:
                        return self.send_json({'status': 'error', 'message': '缺少有效 ids'}, start_response)
                    placeholders = ','.join(['%s'] * len(id_list))
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                f'SELECT id FROM amazon_ad_adjustments WHERE id IN ({placeholders})',
                                tuple(id_list),
                            )
                            found = {self._parse_int(r.get('id')) for r in (cur.fetchall() or [])}
                            missing = [item_id for item_id in id_list if item_id not in found]
                            if missing:
                                return self.send_json(
                                    {'status': 'error', 'message': f'记录不存在: {missing[0]}'},
                                    start_response,
                                )
                            cur.execute(
                                f'DELETE FROM amazon_ad_adjustments WHERE id IN ({placeholders})',
                                tuple(id_list),
                            )
                    return self.send_json({'status': 'success', 'deleted': len(id_list)}, start_response)

                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_ad_adjustments WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            print(f'Amazon ad adjustment API error: {str(e)}')
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _adjustment_import_cell_text(self, cell):
        if cell is None:
            return None
        value = cell.value
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            num = float(value)
            text = str(int(num)) if num.is_integer() else str(value)
            return text.strip()
        return str(value).strip()

    def _load_amazon_ad_adjustment_template_operations(self, cur):
        cur.execute(
            """
            SELECT id, name, reason_names
            FROM amazon_ad_operation_types
            ORDER BY sort_order ASC, id ASC
            """
        )
        operations = []
        for row in cur.fetchall() or []:
            name = str(row.get('name') or '').strip()
            if not name:
                continue
            reason_names = self._parse_operation_type_reason_names(row.get('reason_names'))
            operations.append({
                'id': row['id'],
                'name': name,
                'reason_names': reason_names,
            })
        return operations

    def _amazon_ad_adjustment_template_headers(self):
        return [
            '调整日期*', '店铺ID*', '广告组合*', '广告活动', '广告组',
            '操作*', '操作原因', '对象*', '修改前*', '修改后*',
            '开始时间', '结束时间',
            '曝光', '点击', '花费', '订单', '销售额',
            'ACOS', 'CPC', 'CTR', 'CVR', '首页首位IS',
            '归因检查', '归因订单', '归因销售额', '备注', '提交类型',
        ]

    def _adjustment_template_example_row(
        self, *, shop_id, portfolio, campaign, group, op_name, reason_name,
    ):
        return [
            '2026-06-08 10:00', shop_id, portfolio, campaign, group,
            op_name, reason_name, '关键词A', '1.20', '1.50',
            '2026-06-01 00:00', '2026-06-08 00:00',
            '1000', '50', '30', '5', '200',
            '12%', '0.60', '5%', '10%', '35%',
            '否', '', '', '示例备注', '完整',
        ]

    def _build_amazon_ad_adjustment_import_workbook(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.workbook.defined_name import DefinedName
        from openpyxl.utils import get_column_letter

        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                operations = self._load_amazon_ad_adjustment_template_operations(cur)
                _, _, shop_names, portfolio_names = self._load_amazon_ad_items_template_options()

        example_op = operations[0]['name'] if operations else '示例操作'
        example_reason = ''
        if operations and operations[0].get('reason_names'):
            example_reason = operations[0]['reason_names'][0]
        example_port = portfolio_names[0] if portfolio_names else '示例-Short-SKU01'
        example_camp = 'BE-示例组合-SP-KW'
        example_shop = shop_names[0] if shop_names else '示例店铺'

        wb = Workbook()
        ws = wb.active
        ws.title = '调整记录'
        headers = self._amazon_ad_adjustment_template_headers()
        ws.append(headers)
        ws.append(self._adjustment_template_example_row(
            shop_id=example_shop, portfolio=example_port, campaign='', group='',
            op_name=example_op, reason_name=example_reason,
        ))
        ws.append(self._adjustment_template_example_row(
            shop_id=example_shop, portfolio=example_port, campaign=example_camp, group='',
            op_name=example_op, reason_name=example_reason,
        ))
        ws.append(self._adjustment_template_example_row(
            shop_id=example_shop, portfolio=example_port, campaign=example_camp, group=example_camp,
            op_name=example_op, reason_name=example_reason,
        ))

        example_font = Font(italic=True, color='7B8088')
        example_fill = PatternFill(start_color='E8EEF6', end_color='E8EEF6', fill_type='solid')
        for row_idx in (2, 3, 4):
            for cell in ws[row_idx]:
                cell.fill = example_fill
                cell.font = example_font

        options_ws = wb.create_sheet('_options')
        options_ws.sheet_state = 'hidden'
        options_ws.cell(row=1, column=1, value='operation_name')
        op_names = [op['name'] for op in operations]
        for idx, name in enumerate(op_names, start=2):
            options_ws.cell(row=idx, column=1, value=name)

        for op_idx, op in enumerate(operations, start=1):
            col_idx = op_idx + 1
            col_letter = get_column_letter(col_idx)
            options_ws.cell(row=1, column=col_idx, value=f'reasons_{op_idx}')
            reasons = op.get('reason_names') or []
            if not reasons:
                options_ws.cell(row=2, column=col_idx, value='')
                end_row = 2
            else:
                for r_idx, reason in enumerate(reasons, start=2):
                    options_ws.cell(row=r_idx, column=col_idx, value=reason)
                end_row = len(reasons) + 1
            defined_name = f'adj_op_{op_idx}'
            wb.defined_names.add(DefinedName(
                name=defined_name,
                attr_text=f"'_options'!${col_letter}$2:${col_letter}${max(end_row, 2)}",
            ))

        meta_col = len(operations) + 2
        options_ws.cell(row=1, column=meta_col, value='shop_name')
        for idx, name in enumerate(shop_names, start=2):
            options_ws.cell(row=idx, column=meta_col, value=name)

        port_col = meta_col + 1
        port_letter = get_column_letter(port_col)
        options_ws.cell(row=1, column=port_col, value='portfolio_name')
        for idx, name in enumerate(portfolio_names, start=2):
            options_ws.cell(row=idx, column=port_col, value=name)

        first_data_row = 5
        data_end = 1200
        anchor = first_data_row

        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        gray_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color='2A2420')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        if shop_names:
            shop_end = len(shop_names) + 1
            shop_letter = get_column_letter(meta_col)
            dv_shop = DataValidation(
                type='list',
                formula1=f"='_options'!${shop_letter}$2:${shop_letter}${shop_end}",
                allow_blank=False,
            )
            dv_shop.error = '请从列表选择店铺'
            dv_shop.errorTitle = '店铺'
            ws.add_data_validation(dv_shop)
            dv_shop.add(f'B{first_data_row}:B{data_end}')

        if op_names:
            op_end = len(op_names) + 1
            dv_op = DataValidation(
                type='list',
                formula1=f"='_options'!$A$2:$A${op_end}",
                allow_blank=False,
            )
            dv_op.error = '请从列表选择操作'
            ws.add_data_validation(dv_op)
            dv_op.add(f'F{first_data_row}:F{data_end}')

            dv_reason = DataValidation(
                type='list',
                formula1=f'=INDIRECT("adj_op_"&MATCH($F{anchor},_options!$A$2:$A${op_end},0))',
                allow_blank=True,
            )
            dv_reason.error = '请先选择操作，再从列表选择原因'
            ws.add_data_validation(dv_reason)
            dv_reason.add(f'G{first_data_row}:G{data_end}')

        if portfolio_names:
            port_end = len(portfolio_names) + 1
            dv_port = DataValidation(
                type='list',
                formula1=f"='_options'!${port_letter}$2:${port_letter}${port_end}",
                allow_blank=True,
            )
            ws.add_data_validation(dv_port)
            dv_port.add(f'C{first_data_row}:C{data_end}')

        dv_attr = DataValidation(type='list', formula1='"是,否"', allow_blank=True)
        ws.add_data_validation(dv_attr)
        dv_attr.add(f'W{first_data_row}:W{data_end}')

        dv_submit = DataValidation(type='list', formula1='"完整,快速"', allow_blank=True)
        ws.add_data_validation(dv_submit)
        dv_submit.add(f'AA{first_data_row}:AA{data_end}')

        rules = [
            ('D', f'AND($C{anchor}<>"",$D{anchor}="",$E{anchor}="")'),
            ('E', f'OR(AND($C{anchor}<>"",$D{anchor}="",$E{anchor}=""),AND($C{anchor}<>"",$D{anchor}<>"",$E{anchor}=""))'),
        ]
        for col, formula in rules:
            ws.conditional_formatting.add(
                f'{col}{first_data_row}:{col}{data_end}',
                FormulaRule(formula=[formula], fill=gray_fill),
            )

        ws.freeze_panes = f'A{first_data_row}'
        widths = {
            'A': 18, 'B': 10, 'C': 22, 'D': 22, 'E': 22,
            'F': 14, 'G': 16, 'H': 14, 'I': 10, 'J': 10,
            'K': 18, 'L': 18, 'M': 8, 'N': 8, 'O': 8,
            'P': 8, 'Q': 10, 'R': 8, 'S': 8, 'T': 8,
            'U': 8, 'V': 12, 'W': 10, 'X': 10, 'Y': 12,
            'Z': 16, 'AA': 10,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        guide = wb.create_sheet('填写说明')
        guide.append(['字段', '必填', '说明'])
        for row in [
            ('调整日期*', '是', '如 2026-06-08 10:00'),
            ('店铺ID*', '是', '下拉为系统店铺名称（导入时按名称匹配）'),
            ('广告组合*', '是', '下拉为系统广告组合；组合层级仅填此项'),
            ('广告活动', '活动层级', '调整活动时填写；组合层级留空'),
            ('广告组', '组层级', '调整组时填写；组合/活动层级留空'),
            ('操作*', '是', '下拉为系统操作类型'),
            ('操作原因', '否', '下拉随操作联动（名称管理器 adj_op_N）'),
            ('对象*', '是', '操作对象'),
            ('修改前* / 修改后*', '完整提交', '快速提交可留空'),
            ('开始/结束时间', '否', '效果区间，选填'),
            ('提交类型', '否', '完整 / 快速，默认完整'),
            ('', '', '第2–4行为示例（组合/活动/组），导入从第5行填写'),
            ('', '', '唯一性：店铺+组合+活动+组四元组定位广告，各层级不可重复'),
        ]:
            guide.append(list(row))
        guide.column_dimensions['A'].width = 22
        guide.column_dimensions['B'].width = 12
        guide.column_dimensions['C'].width = 48
        return wb

    def _resolve_shop_id_from_import_text(self, cur, shop_text):
        text = (shop_text or '').strip()
        if not text:
            return None, '请填写店铺'
        parsed_id = self._parse_int(text)
        if parsed_id:
            cur.execute("SELECT id FROM shops WHERE id=%s LIMIT 1", (parsed_id,))
            row = cur.fetchone()
            if row:
                return row['id'], None
        cur.execute(
            "SELECT id FROM shops WHERE shop_name=%s LIMIT 1",
            (text,),
        )
        row = cur.fetchone()
        if row:
            return row['id'], None
        return None, f'未找到店铺: {text}'

    def _resolve_ad_item_by_four_attrs(self, cur, shop_text, portfolio_name, campaign_name='', group_name=''):
        shop_id, shop_err = self._resolve_shop_id_from_import_text(cur, shop_text)
        if shop_err:
            return None, None, shop_err
        portfolio_name = (portfolio_name or '').strip()
        campaign_name = (campaign_name or '').strip()
        group_name = (group_name or '').strip()

        if not portfolio_name:
            return None, None, '请填写广告组合'

        if group_name:
            if not campaign_name:
                return None, None, '填写广告组时须同时填写广告活动'
            ad_level = 'group'
        elif campaign_name:
            ad_level = 'campaign'
        else:
            ad_level = 'portfolio'

        cur.execute(
            """
            SELECT id FROM amazon_ad_items
            WHERE ad_level='portfolio' AND name=%s AND shop_id=%s LIMIT 1
            """,
            (portfolio_name, shop_id),
        )
        portfolio = cur.fetchone()
        if not portfolio:
            return None, None, f'未找到该店铺下的广告组合: {portfolio_name}'
        portfolio_id = portfolio['id']

        if ad_level == 'portfolio':
            return portfolio_id, ad_level, None

        cur.execute(
            """
            SELECT id FROM amazon_ad_items
            WHERE ad_level='campaign' AND name=%s AND portfolio_id=%s LIMIT 1
            """,
            (campaign_name, portfolio_id),
        )
        campaign = cur.fetchone()
        if not campaign:
            return None, None, f'在组合「{portfolio_name}」下未找到广告活动: {campaign_name}'

        if ad_level == 'campaign':
            return campaign['id'], ad_level, None

        cur.execute(
            """
            SELECT id FROM amazon_ad_items
            WHERE ad_level='group' AND name=%s AND campaign_id=%s LIMIT 1
            """,
            (group_name, campaign['id']),
        )
        group = cur.fetchone()
        if not group:
            return None, None, f'在活动「{campaign_name}」下未找到广告组: {group_name}'
        return group['id'], ad_level, None

    def _validate_adjustment_operation_for_ad(self, cur, ad_item_id, operation_type_id, reason_name=None):
        operation_type_id = self._parse_int(operation_type_id)
        if not operation_type_id:
            return '请选择操作'
        cur.execute(
            "SELECT id, ad_level, campaign_id, subtype_id FROM amazon_ad_items WHERE id=%s LIMIT 1",
            (ad_item_id,),
        )
        ad_row = cur.fetchone()
        if not ad_row:
            return '关联广告不存在'
        allowed = self._fetch_allowed_operations_for_ad(cur, ad_row)
        allowed_ids = {int(x['id']) for x in allowed}
        if int(operation_type_id) not in allowed_ids:
            return '该操作不适用于此广告类型'
        reason_name = (reason_name or '').strip() or None
        if reason_name:
            op_match = next((x for x in allowed if int(x['id']) == int(operation_type_id)), None)
            valid_reasons = {r.get('reason_name') for r in (op_match.get('reasons') or [])}
            if reason_name not in valid_reasons:
                return f'操作原因「{reason_name}」不适用于所选操作'
        return None

    def _validate_adjustment_import_operation(self, cur, ad_item_id, operation_name, reason_name):
        operation_name = (operation_name or '').strip()
        if not operation_name:
            return None, None, '请填写操作'
        cur.execute(
            "SELECT id, ad_level FROM amazon_ad_items WHERE id=%s LIMIT 1",
            (ad_item_id,),
        )
        ad_row = cur.fetchone()
        if not ad_row:
            return None, None, '关联广告不存在'
        allowed = self._fetch_allowed_operations_for_ad(cur, ad_row)
        op_match = next((x for x in allowed if (x.get('name') or '').strip() == operation_name), None)
        if not op_match:
            return None, None, f'操作「{operation_name}」不适用于该广告'
        reason_name = (reason_name or '').strip() or None
        if reason_name:
            valid_reasons = {r.get('reason_name') for r in (op_match.get('reasons') or [])}
            if reason_name not in valid_reasons:
                return None, None, f'操作原因「{reason_name}」不属于操作「{operation_name}」'
        return op_match['id'], reason_name, None

    def handle_amazon_ad_adjustment_template_api(self, environ, method, start_response):
        """广告调整记录批量导入模板下载"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json(
                    {'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'},
                    start_response,
                )
            wb = self._build_amazon_ad_adjustment_import_workbook()
            return self._send_excel_workbook(wb, '广告调整记录导入模板.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_adjustment_import_api(self, environ, method, start_response):
        """广告调整记录批量导入（预加载索引 + 批量写入）"""
        wb = None
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            wb, wb_err = self._read_batch_import_workbook(environ)
            if wb_err:
                return self.send_json({'status': 'error', 'message': wb_err}, start_response)

            ws = wb.active
            header_map = self._import_sheet_header_map(ws)
            if not header_map:
                return self.send_json({'status': 'error', 'message': '模板表头为空'}, start_response)

            skipped_sample_rows = 0
            skipped_existing = 0
            errors = []
            pending_rows = []
            ad_item_ids = set()
            insert_rows = []
            max_row = min(
                int(ws.max_row or 2),
                2 + self._AMAZON_AD_ADJUSTMENT_IMPORT_MAX_ROWS,
            )
            insert_sql = """
                INSERT INTO amazon_ad_adjustments (
                    adjust_date, ad_item_id, operation_type_id, target_object,
                    before_value, after_value, reason_name,
                    start_time, end_time,
                    impressions, clicks, cost, orders, sales,
                    acos, cpc, ctr, cvr, top_of_search_is,
                    attribution_checked, attribution_orders, attribution_sales,
                    remark, is_quick_submit
                ) VALUES (
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s
                )
            """

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    ctx = self._load_amazon_ad_adjustment_import_context(cur)

                    for row_idx, row in enumerate(
                        ws.iter_rows(min_row=2, max_row=max_row, values_only=True),
                        start=2,
                    ):
                        if not any(v is not None and str(v).strip() for v in row):
                            continue

                        def cell(name, _row=row):
                            return self._import_cell_text(_row, header_map, name)

                        portfolio_name = cell('广告组合*') or ''
                        if row_idx <= 4 or ('示例' in portfolio_name):
                            skipped_sample_rows += 1
                            continue

                        shop_text = cell('店铺ID*') or ''
                        campaign_name = cell('广告活动') or ''
                        group_name = cell('广告组') or ''
                        if not shop_text and not portfolio_name:
                            continue

                        ad_item_id, _, ad_err = self._resolve_ad_item_by_four_attrs_ctx(
                            ctx, shop_text, portfolio_name, campaign_name, group_name,
                        )
                        if ad_err:
                            if not self._append_child_import_error(errors, row_idx, ad_err):
                                break
                            continue

                        operation_name = cell('操作*') or ''
                        reason_name = cell('操作原因') or ''
                        op_id, reason_name, op_err = self._validate_adjustment_import_operation_ctx(
                            ctx, ad_item_id, operation_name, reason_name,
                        )
                        if op_err:
                            if not self._append_child_import_error(errors, row_idx, op_err):
                                break
                            continue

                        target_object = cell('对象*') or ''
                        if not target_object:
                            if not self._append_child_import_error(errors, row_idx, '对象不能为空'):
                                break
                            continue

                        submit_type = (cell('提交类型') or '完整').strip()
                        is_quick = 1 if submit_type in ('快速', 'quick') else 0
                        before_value = cell('修改前*') or ''
                        after_value = cell('修改后*') or ''
                        start_time = cell('开始时间') or cell('开始时间*') or ''
                        end_time = cell('结束时间') or cell('结束时间*') or ''
                        if not is_quick:
                            missing = []
                            if not before_value:
                                missing.append('修改前')
                            if not after_value:
                                missing.append('修改后')
                            if missing:
                                msg = f'完整提交须填写：{", ".join(missing)}'
                                if not self._append_child_import_error(errors, row_idx, msg):
                                    break
                                continue

                        adjust_date = self._parse_datetime_local_value(cell('调整日期*'))
                        if not adjust_date:
                            adjust_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                        attr_raw = (cell('归因检查') or '否').strip()
                        attribution_checked = 1 if attr_raw in ('是', '1', 'true', 'True', 'yes') else 0

                        dedupe_key = self._adjustment_import_dedupe_key(
                            ad_item_id, op_id, target_object, adjust_date,
                            before_value, after_value, reason_name,
                        )
                        pending_rows.append({
                            'key': dedupe_key,
                            'insert': (
                                adjust_date, ad_item_id, op_id, target_object,
                                before_value or None, after_value or None, reason_name,
                                self._parse_datetime_local_value(start_time) if (not is_quick and start_time) else None,
                                self._parse_datetime_local_value(end_time) if (not is_quick and end_time) else None,
                                cell('曝光') or None,
                                cell('点击') or None,
                                cell('花费') or None,
                                cell('订单') or None,
                                cell('销售额') or None,
                                cell('ACOS') or None,
                                cell('CPC') or None,
                                cell('CTR') or None,
                                cell('CVR') or None,
                                cell('首页首位IS') or None,
                                attribution_checked,
                                cell('归因订单') or None,
                                cell('归因销售额') or None,
                                cell('备注') or None,
                                is_quick,
                            ),
                        })
                        ad_item_ids.add(int(ad_item_id))

                    existing_keys = self._load_adjustment_import_existing_keys(cur, ad_item_ids)
                    batch_keys = set()
                    insert_rows = []
                    for item in pending_rows:
                        key = item['key']
                        if key in existing_keys or key in batch_keys:
                            skipped_existing += 1
                            continue
                        batch_keys.add(key)
                        insert_rows.append(item['insert'])

                    self._executemany_in_chunks(
                        cur,
                        insert_sql,
                        insert_rows,
                        chunk_size=self._AMAZON_AD_ADJUSTMENT_IMPORT_BATCH_SIZE,
                    )

            payload = self._child_import_success_payload(
                len(insert_rows), 0, errors, skipped_sample_rows,
            )
            payload['skipped_existing'] = skipped_existing
            return self.send_json(payload, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def handle_amazon_ad_keyword_api(self, environ, method, start_response):
        """Amazon 广告关键词 API"""
        try:
            if method == 'GET':
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT * FROM amazon_keywords ORDER BY id DESC LIMIT 500")
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)
            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_keyword_template_api(self, environ, method, start_response):
        """Amazon 广告关键词模板 API"""
        try:
            if method == 'GET':
                return self.send_json({'status': 'success', 'items': []}, start_response)
            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_keyword_import_api(self, environ, method, start_response):
        """Amazon 广告关键词导入 API"""
        try:
            if method == 'POST':
                return self.send_json({'status': 'success', 'imported': 0}, start_response)
            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

