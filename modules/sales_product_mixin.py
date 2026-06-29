# -*- coding: utf-8 -*-
"""销售产品：规格/变体、图库、产品表现导入与周月/30 天 rolling 快照刷新。"""
import re
import io
import cgi
import os
import shutil
import json
import base64
import hashlib
import threading
import time
import unicodedata
from email import policy
from email.parser import BytesParser
from datetime import datetime, timedelta
from urllib.parse import parse_qs

try:
    from openpyxl import Workbook, load_workbook
    _openpyxl_import_error = None
except Exception as e:
    Workbook = None
    load_workbook = None
    _openpyxl_import_error = str(e)

try:
    import pymysql
except Exception:
    pymysql = None

try:
    from decimal import Decimal
except Exception:
    Decimal = None  # pragma: no cover

# 产品表现导入进度：SSE 推送 + 长轮询回退（参考麻将/围棋 stream）
SPP_IMPORT_WAIT_TIMEOUT_SEC = 22
SPP_IMPORT_WAIT_POLL_SEC = 0.35
SPP_IMPORT_STREAM_SESSION_SEC = 120
SPP_IMPORT_STREAM_PING_SEC = 15
_spp_import_waiters = {}
_spp_import_waiters_lock = threading.Lock()


def _spp_signal_import_waiters(task_id):
    tid = str(task_id or '').strip()
    if not tid:
        return
    with _spp_import_waiters_lock:
        events = list(_spp_import_waiters.get(tid) or [])
    for ev in events:
        try:
            ev.set()
        except Exception:
            pass


def _spp_register_import_waiter(task_id):
    tid = str(task_id or '').strip()
    ev = threading.Event()
    with _spp_import_waiters_lock:
        _spp_import_waiters.setdefault(tid, []).append(ev)
    return ev


def _spp_unregister_import_waiter(task_id, ev):
    tid = str(task_id or '').strip()
    with _spp_import_waiters_lock:
        lst = _spp_import_waiters.get(tid)
        if not lst:
            return
        try:
            lst.remove(ev)
        except ValueError:
            pass
        if not lst:
            _spp_import_waiters.pop(tid, None)


class SalesProductMixin:
    """销售产品 Mixin：规格变体 CRUD、图库、产品表现与聚合/rolling 快照刷新。"""

    # -------------------------------------------------------------------------
    # 图库 / resources 路径与 image_assets 工具
    # -------------------------------------------------------------------------

    def _resources_root(self):
        """Return absolute resources root as bytes path."""
        return self._join_resources('')

    def _storage_path_from_abs(self, abs_path):
        """Compute image_assets.storage_path from an absolute resources path."""
        try:
            root = self._resources_root()
            # Prevent "Can't mix strings and bytes in path components"
            if isinstance(root, (bytes, bytearray)) and isinstance(abs_path, str):
                abs_path = self._safe_fsencode(abs_path)
            elif isinstance(root, str) and isinstance(abs_path, (bytes, bytearray)):
                abs_path = os.fsdecode(abs_path)
            rel_path = os.path.relpath(abs_path, root)
            if isinstance(rel_path, (bytes, bytearray)):
                rb = bytes(rel_path)
                try:
                    return rb.decode('utf-8', errors='strict').replace('\\', '/')
                except UnicodeDecodeError:
                    pass
            return os.fsdecode(rel_path).replace('\\', '/')
        except Exception:
            try:
                return os.fsdecode(abs_path).replace('\\', '/')
            except Exception:
                return str(abs_path)

    def _display_name_from_abs_path_b(self, path_abs_b):
        """
        从绝对路径 bytes 得到逻辑文件名（basename）及 stem/ext 文本。
        对 basename 优先 UTF-8 严格解码（NAS 上 ext4 常见），避免 str 路径经 fsdecode 产生 U+FFFD 再写入磁盘/数据库。
        """
        base_b = os.path.basename(path_abs_b)
        stem_b, ext_b = os.path.splitext(base_b)
        try:
            stem = stem_b.decode('utf-8', errors='strict')
        except UnicodeDecodeError:
            stem = self._safe_fsdecode(stem_b)
        try:
            ext = (ext_b.decode('ascii', errors='strict') or '').lower() if ext_b else ''
        except UnicodeDecodeError:
            ext = (self._safe_fsdecode(ext_b) or '').lower() if ext_b else ''
        return (stem or 'image') + ext, stem, ext

    def _abs_from_storage_path(self, storage_path):
        return self._join_resources((storage_path or '').strip().replace('\\', '/'))

    def _gallery_basename_variants(self, basename):
        """文件名 basename 的 Unicode 变体（NFC/NFD），用于库内路径容错匹配。"""
        base = os.path.basename(str(basename or '').strip().replace('\\', '/'))
        if not base:
            return []
        out = []
        seen = set()
        for candidate in (base,):
            variants = [candidate]
            try:
                variants.append(unicodedata.normalize('NFC', candidate))
                variants.append(unicodedata.normalize('NFD', candidate))
            except Exception:
                pass
            for v in variants:
                if not v or v in seen:
                    continue
                seen.add(v)
                out.append(v)
        return out

    def _find_image_asset_row_by_rel_path(self, cur, rel_text, join_type, has_tid, has_dep, has_desc):
        """按 storage_path 精确或 basename 容错查找 image_assets。"""
        rel_text = (rel_text or '').strip().replace('\\', '/').lstrip('/')
        if not rel_text:
            return {}
        cur.execute(
            f"""
            SELECT ia.id, ia.storage_path,
                   {('ia.image_type_id AS image_type_id' if has_tid else '0 AS image_type_id')},
                   {('ia.is_deprecated AS is_deprecated' if has_dep else '0 AS is_deprecated')},
                   {('ia.description AS description' if has_desc else "'' AS description")},
                   {('it.name AS image_type_name' if has_tid else "'' AS image_type_name")}
            FROM image_assets ia
            {join_type}
            WHERE ia.storage_path=%s
            LIMIT 1
            """,
            (rel_text,),
        )
        row = cur.fetchone() or {}
        if row.get('id'):
            return row
        for base in self._gallery_basename_variants(os.path.basename(rel_text)):
            cur.execute(
                f"""
                SELECT ia.id, ia.storage_path,
                       {('ia.image_type_id AS image_type_id' if has_tid else '0 AS image_type_id')},
                       {('ia.is_deprecated AS is_deprecated' if has_dep else '0 AS is_deprecated')},
                       {('ia.description AS description' if has_desc else "'' AS description")},
                       {('it.name AS image_type_name' if has_tid else "'' AS image_type_name")}
                FROM image_assets ia
                {join_type}
                WHERE ia.storage_path=%s OR ia.storage_path LIKE %s
                ORDER BY ia.id DESC
                LIMIT 1
                """,
                (base, f'%/{base}'),
            )
            row = cur.fetchone() or {}
            if row.get('id'):
                return row
        return {}

    def _resolve_gallery_abs_path(self, rel_text):
        """gallery 相对路径 → 绝对路径；面料目录按 basename 二次解析。"""
        rel_text = (rel_text or '').strip().replace('\\', '/').lstrip('/')
        if not rel_text or '..' in rel_text:
            return None, rel_text

        abs_path = self._abs_from_storage_path(rel_text)
        if abs_path and os.path.isfile(abs_path):
            canonical = rel_text
            try:
                res_root = self._join_resources('')
                canonical = os.fsdecode(os.path.relpath(abs_path, res_root)).replace('\\', '/')
            except Exception:
                pass
            return abs_path, canonical

        base = os.path.basename(rel_text)
        if not base:
            return None, rel_text

        fabric_abs = None
        try:
            fabric_abs = self._resolve_fabric_image_abs_path(base)
        except Exception:
            fabric_abs = None
        if fabric_abs and os.path.isfile(fabric_abs):
            try:
                res_root = self._join_resources('')
                canonical = os.fsdecode(os.path.relpath(fabric_abs, res_root)).replace('\\', '/')
            except Exception:
                canonical = f"『面料』/{base}"
            return fabric_abs, canonical

        return None, rel_text

    @staticmethod
    def _sales_variant_subfolder_display_name(spec_part, fabric_part):
        """销售主图子目录名（面料必填）：有规格时为「规格-面料」，规格为空时仅为「面料」；无面料返回空串。"""
        s = (spec_part or '').strip().replace('/', '-').replace('\\', '-')
        f = (fabric_part or '').strip().replace('/', '-').replace('\\', '-')
        if not f:
            return ''
        if s:
            return f"{s}-{f}"
        return f

    def _resolve_sales_variant_folder_by_variant_id(self, variant_id, ensure_folder=False):
        """按 variant_id 解析销售主图文件夹（货号/主图/子目录；子目录为规格-面料，规格可为空）。"""
        vid = int(variant_id or 0)
        if vid <= 0:
            raise RuntimeError('Missing variant_id')
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                fabric_join = "LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if has_fabric_id else ""
                if has_fabric_id and has_fabric_text:
                    fabric_select = "COALESCE(fm.fabric_code, v.fabric) AS fabric"
                elif has_fabric_id:
                    fabric_select = "fm.fabric_code AS fabric"
                else:
                    fabric_select = ("v.fabric AS fabric" if has_fabric_text else "'' AS fabric")
                cur.execute(
                    f"""
                    SELECT v.id, v.spec_name, {fabric_select}, pf.sku_family,
                           {("fm.fabric_name_en AS fabric_name_en, v.fabric_id AS fabric_id" if has_fabric_id else "'' AS fabric_name_en, 0 AS fabric_id")}
                    FROM sales_product_variants v
                    LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                    {fabric_join}
                    WHERE v.id=%s
                    """,
                    (vid,),
                )
                row = cur.fetchone() or {}
            if not row.get('id'):
                raise RuntimeError('规格不存在')

            sku_name = (row.get('sku_family') or '').strip()
            spec_part = (row.get('spec_name') or '').strip().replace('/', '-').replace('\\', '-')
            fabric_part = str(row.get('fabric_name_en') or '').strip().replace('/', '-').replace('\\', '-')
            if not fabric_part:
                fabric_part = self._resolve_fabric_folder_part(conn, row.get('fabric_id'), row.get('fabric'))
            if not fabric_part:
                raise RuntimeError('面料为必填：当前规格缺少可解析的面料信息，无法定位主图文件夹')
            variant_folder_name = self._sales_variant_subfolder_display_name(spec_part, fabric_part)
            if not sku_name:
                raise RuntimeError('当前规格缺少货号，无法定位主图文件夹')

            if ensure_folder:
                self._ensure_listing_sales_variant_folder(sku_name, spec_part, fabric_part)
            base_folder = self._ensure_listing_folder()
            folder_path = os.path.join(
                base_folder,
                self._safe_fsencode(sku_name),
                self._safe_fsencode('主图'),
                self._safe_fsencode(variant_folder_name),
            )
            return {
                'variant_id': vid,
                'sku_family': sku_name,
                'spec_name': spec_part,
                'fabric_folder_part': fabric_part,
                'fabric_name_en': str(row.get('fabric_name_en') or '').strip(),
                'fabric_id': self._parse_int(row.get('fabric_id')) or 0,
                'variant_folder': variant_folder_name,
                'folder_path': folder_path,
            }

    # -------------------------------------------------------------------------
    # Gallery 元数据 API（关联状态、启用/弃用、备注）
    # -------------------------------------------------------------------------

    def handle_gallery_image_meta_api(self, environ, method, start_response):
        """按 gallery 的相对路径（base64 bytes）查库里是否已关联，并返回图片类型/启用状态/备注。

        说明：
        - 由于历史库里图片弃用字段为 image_assets.is_deprecated（1=弃用），这里对外统一输出 is_enabled（1=启用/0=弃用）。
        - PUT 时若传入 is_enabled，会映射写回 is_deprecated。
        """
        try:
            if method not in ('GET', 'PUT'):
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            user_id = self._get_session_user(environ)
            data = None
            if method == 'PUT':
                # PUT 请求体只能读取一次：提前读取并复用，避免后续读到空导致误判
                data = self._read_json_body(environ) or {}
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            path_b64 = str(query_params.get('id', [''])[0] or '').strip()
            if method == 'PUT' and not path_b64:
                path_b64 = str((data or {}).get('id') or '').strip()
            if not path_b64:
                return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
            try:
                raw = base64.b64decode(path_b64)
            except Exception:
                return self.send_json({'status': 'error', 'message': 'Invalid id'}, start_response)
            rel_text = ''
            try:
                rel_text = os.fsdecode(raw)
            except Exception:
                try:
                    rel_text = raw.decode('utf-8', errors='surrogateescape')
                except Exception:
                    rel_text = ''
            rel_text = (rel_text or '').strip().replace('\\', '/').lstrip('/')
            if not rel_text or '..' in rel_text:
                return self.send_json({'status': 'error', 'message': 'Invalid path'}, start_response)

            with self._get_db_connection() as conn:
                has_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
                has_desc = self._table_has_column(conn, 'image_assets', 'description')
                has_dep = self._table_has_column(conn, 'image_assets', 'is_deprecated')
                join_type = "LEFT JOIN image_types it ON it.id = ia.image_type_id" if has_tid else ""
                with conn.cursor() as cur:
                    row = self._find_image_asset_row_by_rel_path(
                        cur, rel_text, join_type, has_tid, has_dep, has_desc
                    )

                    if method == 'PUT':
                        # 允许单独更新 is_enabled；若同时带 image_type_name，则更新类型
                        image_type_name = str(((data or {}) if data is not None else {}).get('image_type_name') or '').strip()
                        is_enabled_raw = ((data or {}) if data is not None else {}).get('is_enabled', None)
                        description_raw = ((data or {}) if data is not None else {}).get('description', None)
                        has_enabled_patch = is_enabled_raw is not None and str(is_enabled_raw).strip() != ''
                        has_desc_patch = description_raw is not None
                        if (not image_type_name) and (not has_enabled_patch):
                            if not has_desc_patch:
                                return self.send_json({'status': 'error', 'message': '缺少可更新字段'}, start_response)

                        aid = self._parse_int(row.get('id')) or 0
                        if not aid:
                            # 兼容：若该图片尚未入库，但文件存在且用户提交了可更新字段，
                            # 自动补建 image_assets 记录，再继续保存类型/启用状态/备注。
                            abs_path, canonical_rel = self._resolve_gallery_abs_path(rel_text)
                            if not abs_path or (not os.path.exists(abs_path)) or (not os.path.isfile(abs_path)):
                                return self.send_json({'status': 'error', 'message': '图片未入库，且源文件不存在，无法保存'}, start_response)
                            try:
                                with open(abs_path, 'rb') as f:
                                    content = f.read() or b''
                            except Exception:
                                content = b''
                            if not content:
                                return self.send_json({'status': 'error', 'message': '图片文件读取失败，无法保存'}, start_response)
                            sha256 = self._sha256_hex(content)
                            existing = self._find_image_asset_by_sha256(conn, sha256)
                            if existing and self._parse_int(existing.get('id')):
                                aid = int(existing.get('id'))
                            else:
                                rec = {
                                    'sha256': sha256,
                                    'storage_path': canonical_rel or rel_text,
                                    'description': '',
                                    'is_deprecated': 0,
                                    'created_by': int(user_id) if user_id else None,
                                }
                                aid = int(self._insert_image_asset_dynamic(conn, cur, rec) or 0)
                            if not aid:
                                return self.send_json({'status': 'error', 'message': '自动入库失败，无法保存类型'}, start_response)

                        if image_type_name:
                            if not has_tid:
                                return self.send_json({'status': 'error', 'message': 'image_assets 缺少 image_type_id'}, start_response)
                            tid = self._get_image_type_id_by_name(conn, image_type_name)
                            if not tid:
                                return self.send_json({'status': 'error', 'message': f'未知图片类型: {image_type_name}'}, start_response)
                            ok_scope, scope_msg = self._gallery_validate_asset_type_compatible(conn, int(aid), int(tid))
                            if not ok_scope:
                                return self.send_json({'status': 'error', 'message': scope_msg or '图片类型与现有关联冲突'}, start_response)
                            cur.execute("UPDATE image_assets SET image_type_id=%s WHERE id=%s", (int(tid), int(aid)))

                        if has_enabled_patch:
                            if not has_dep:
                                return self.send_json({'status': 'error', 'message': 'image_assets 缺少 is_deprecated'}, start_response)
                            is_enabled = 1 if str(is_enabled_raw).strip().lower() in ('1', 'true', 'yes', 'on', '启用') else 0
                            is_deprecated = 0 if is_enabled else 1
                            cur.execute("UPDATE image_assets SET is_deprecated=%s WHERE id=%s", (int(is_deprecated), int(aid)))

                        if has_desc_patch:
                            if not has_desc:
                                return self.send_json({'status': 'error', 'message': 'image_assets 缺少 description'}, start_response)
                            description = str(description_raw or '').strip()
                            cur.execute("UPDATE image_assets SET description=%s WHERE id=%s", (description, int(aid)))

                        # 回读确认
                        cur.execute(
                            f"""
                            SELECT ia.id, ia.storage_path,
                                   {('ia.image_type_id AS image_type_id' if has_tid else '0 AS image_type_id')},
                                   {('ia.is_deprecated AS is_deprecated' if has_dep else '0 AS is_deprecated')},
                                   {('ia.description AS description' if has_desc else "'' AS description")},
                                   {('it.name AS image_type_name' if has_tid else "'' AS image_type_name")}
                            FROM image_assets ia
                            {join_type}
                            WHERE ia.id=%s
                            LIMIT 1
                            """,
                            (int(aid),),
                        )
                        row = cur.fetchone() or row

            linked = bool(row.get('id'))
            is_deprecated_val = self._parse_int(row.get('is_deprecated', 0)) or 0
            return self.send_json(
                {
                    'status': 'success',
                    'linked': linked,
                    'image_asset_id': self._parse_int(row.get('id')) or 0,
                    'storage_path': (row.get('storage_path') or '').strip(),
                    'image_type_id': self._parse_int(row.get('image_type_id')) or 0,
                    'image_type_name': (row.get('image_type_name') or '').strip(),
                    'is_enabled': 0 if int(is_deprecated_val) == 1 else 1,
                    'description': (row.get('description') or '').strip(),
                },
                start_response,
            )
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    # -------------------------------------------------------------------------
    # Gallery 图片关联与 picker API
    # -------------------------------------------------------------------------

    def handle_gallery_image_links_api(self, environ, method, start_response):
        """按图片路径（base64 bytes）查该图片已关联的规格列表（用于 gallery 预填）。"""
        try:
            if method != 'GET':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            path_b64 = str(query_params.get('id', [''])[0] or '').strip()
            if not path_b64:
                return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
            try:
                raw = base64.b64decode(path_b64)
            except Exception:
                return self.send_json({'status': 'error', 'message': 'Invalid id'}, start_response)
            rel_text = ''
            try:
                rel_text = os.fsdecode(raw)
            except Exception:
                try:
                    rel_text = raw.decode('utf-8', errors='surrogateescape')
                except Exception:
                    rel_text = ''
            rel_text = (rel_text or '').strip().replace('\\', '/').lstrip('/')
            if not rel_text or '..' in rel_text:
                return self.send_json({'status': 'error', 'message': 'Invalid path'}, start_response)

            with self._get_db_connection() as conn:
                has_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
                join_type = "LEFT JOIN image_types it ON it.id = ia.image_type_id" if has_tid else ""
                with conn.cursor() as cur:
                    row = self._find_image_asset_row_by_rel_path(
                        cur, rel_text, join_type, has_tid, False, False
                    )

                    aid = self._parse_int(row.get('id')) or 0
                    if not aid:
                        return self.send_json(
                            {
                                'status': 'success',
                                'linked': False,
                                'image_asset_id': 0,
                                'image_type_name': '',
                                'variants': [],
                                'variant_ids': [],
                                'fabric_ids': [],
                                'fabrics': [],
                                'order_product_ids': [],
                                'order_products': [],
                            },
                            start_response,
                        )

                    if not self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id'):
                        return self.send_json({'status': 'error', 'message': 'sales_variant_image_mappings 缺少 variant_id'}, start_response)

                    cur.execute(
                        """
                        SELECT DISTINCT variant_id
                        FROM sales_variant_image_mappings
                        WHERE image_asset_id=%s AND variant_id IS NOT NULL AND variant_id>0
                        ORDER BY variant_id ASC
                        """,
                        (aid,),
                    )
                    vids = [self._parse_int(r.get('variant_id')) or 0 for r in (cur.fetchall() or [])]
                    vids = [v for v in vids if v > 0]

                    variants = []
                    if vids:
                        placeholders = ','.join(['%s'] * len(vids))
                        has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                        has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                        fabric_join = "LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if has_fabric_id else ""
                        if has_fabric_id and has_fabric_text:
                            fabric_code_expr = "COALESCE(fm.fabric_code, v.fabric) AS fabric_code"
                            fabric_name_expr = "COALESCE(fm.fabric_name_en, v.fabric) AS fabric_name_en"
                        elif has_fabric_id:
                            fabric_code_expr = "fm.fabric_code AS fabric_code"
                            fabric_name_expr = "fm.fabric_name_en AS fabric_name_en"
                        elif has_fabric_text:
                            fabric_code_expr = "v.fabric AS fabric_code"
                            fabric_name_expr = "v.fabric AS fabric_name_en"
                        else:
                            fabric_code_expr = "'' AS fabric_code"
                            fabric_name_expr = "'' AS fabric_name_en"
                        cur.execute(
                            f"""
                            SELECT v.id, v.spec_name, pf.sku_family,
                                   {fabric_code_expr},
                                   {fabric_name_expr}
                            FROM sales_product_variants v
                            LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                            {fabric_join}
                            WHERE v.id IN ({placeholders})
                            ORDER BY pf.sku_family ASC, v.spec_name ASC, v.id ASC
                            """,
                            tuple(vids),
                        )
                        for r in (cur.fetchall() or []):
                            variants.append({
                                'variant_id': self._parse_int(r.get('id')) or 0,
                                'sku_family': str(r.get('sku_family') or '').strip(),
                                'spec_name': str(r.get('spec_name') or '').strip(),
                                'fabric_code': str(r.get('fabric_code') or '').strip(),
                                'fabric_name_en': str(r.get('fabric_name_en') or '').strip(),
                            })

                    fabric_ids = []
                    fabrics = []
                    order_product_ids = []
                    order_products = []
                    if self._has_required_tables(['fabric_image_mappings']):
                        cur.execute(
                            """
                            SELECT DISTINCT fabric_id
                            FROM fabric_image_mappings
                            WHERE image_asset_id=%s AND fabric_id IS NOT NULL AND fabric_id>0
                            ORDER BY fabric_id ASC
                            """,
                            (aid,),
                        )
                        fabric_ids = [self._parse_int(r.get('fabric_id')) or 0 for r in (cur.fetchall() or [])]
                        fabric_ids = [f for f in fabric_ids if f > 0]
                        if fabric_ids:
                            ph = ','.join(['%s'] * len(fabric_ids))
                            cur.execute(
                                f"SELECT id, fabric_code, fabric_name_en FROM fabric_materials WHERE id IN ({ph}) ORDER BY fabric_code ASC",
                                tuple(fabric_ids),
                            )
                            for r in (cur.fetchall() or []):
                                fabrics.append({
                                    'fabric_id': self._parse_int(r.get('id')) or 0,
                                    'fabric_code': str(r.get('fabric_code') or '').strip(),
                                    'fabric_name_en': str(r.get('fabric_name_en') or '').strip(),
                                })

                    if self._has_required_tables(['order_product_image_mappings']):
                        cur.execute(
                            """
                            SELECT DISTINCT order_product_id
                            FROM order_product_image_mappings
                            WHERE image_asset_id=%s AND order_product_id IS NOT NULL AND order_product_id>0
                            ORDER BY order_product_id ASC
                            """,
                            (aid,),
                        )
                        order_product_ids = [self._parse_int(r.get('order_product_id')) or 0 for r in (cur.fetchall() or [])]
                        order_product_ids = [x for x in order_product_ids if x > 0]
                        if order_product_ids:
                            ph = ','.join(['%s'] * len(order_product_ids))
                            cur.execute(
                                f"""
                                SELECT op.id, op.sku, op.spec_qty_short, pf.sku_family
                                FROM order_products op
                                LEFT JOIN product_families pf ON pf.id = op.sku_family_id
                                WHERE op.id IN ({ph})
                                ORDER BY pf.sku_family ASC, op.sku ASC, op.id ASC
                                """,
                                tuple(order_product_ids),
                            )
                            for r in (cur.fetchall() or []):
                                order_products.append({
                                    'order_product_id': self._parse_int(r.get('id')) or 0,
                                    'sku': str(r.get('sku') or '').strip(),
                                    'sku_family': str(r.get('sku_family') or '').strip(),
                                    'spec_qty_short': str(r.get('spec_qty_short') or '').strip(),
                                })

            return self.send_json(
                {
                    'status': 'success',
                    'linked': True,
                    'image_asset_id': self._parse_int(row.get('id')) or 0,
                    'storage_path': (row.get('storage_path') or '').strip(),
                    'image_type_id': self._parse_int(row.get('image_type_id')) or 0,
                    'image_type_name': (row.get('image_type_name') or '').strip(),
                    'variant_ids': vids,
                    'variants': variants,
                    'fabric_ids': fabric_ids,
                    'fabrics': fabrics,
                    'order_product_ids': order_product_ids,
                    'order_products': order_products,
                },
                start_response,
            )
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _spec_main_images_variant_detail(self, conn, variant_id):
        """规格主图编辑/保存后：变体基础信息 + 售价 + 关联下单 SKU + 成本/包裹汇总（与销售产品页同源）。"""
        vid_int = int(self._parse_int(variant_id) or 0)
        variant_out = {'variant_id': vid_int}
        try:
            has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
            has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
            has_sale_price = self._table_has_column(conn, 'sales_products', 'sale_price_usd')
            fabric_join = "LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if has_fabric_id else ""
            if has_fabric_id and has_fabric_text:
                fabric_select = "COALESCE(fm.fabric_name_en, v.fabric) AS fabric_name_en"
            elif has_fabric_id:
                fabric_select = "fm.fabric_name_en AS fabric_name_en"
            else:
                fabric_select = ("v.fabric AS fabric_name_en" if has_fabric_text else "'' AS fabric_name_en")
            fabric_code_select = "COALESCE(fm.fabric_code, '') AS fabric_code" if has_fabric_id else "'' AS fabric_code"
            sale_select = (
                "(SELECT MIN(sp0.sale_price_usd) FROM sales_products sp0 WHERE sp0.variant_id = v.id) AS sale_price_usd"
                if has_sale_price else "NULL AS sale_price_usd"
            )
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT v.id, v.spec_name, {sale_select}, {fabric_select},
                           {fabric_code_select}, pf.sku_family
                    FROM sales_product_variants v
                    LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                    {fabric_join}
                    WHERE v.id=%s
                    LIMIT 1
                    """,
                    (vid_int,),
                )
                row = cur.fetchone() or {}
                if row.get('id'):
                    sp = row.get('sale_price_usd')
                    sale_val = None
                    if sp is not None and str(sp).strip() != '':
                        try:
                            sale_val = float(sp)
                        except (TypeError, ValueError):
                            sale_val = None
                    variant_out.update({
                        'variant_id': self._parse_int(row.get('id')) or vid_int,
                        'sku_family': str(row.get('sku_family') or '').strip(),
                        'spec_name': str(row.get('spec_name') or '').strip(),
                        'fabric_name_en': str(row.get('fabric_name_en') or '').strip(),
                        'fabric_code': str(row.get('fabric_code') or '').strip(),
                        'sale_price_usd': sale_val,
                    })
        except Exception:
            pass

        try:
            mm = self._load_sales_variant_metrics(conn, [vid_int]) or {}
            m = mm.get(vid_int) or {}
            variant_out.update({
                'warehouse_cost_usd': m.get('warehouse_cost_usd'),
                'last_mile_cost_usd': m.get('last_mile_cost_usd'),
                'package_length_in': m.get('package_length_in'),
                'package_width_in': m.get('package_width_in'),
                'package_height_in': m.get('package_height_in'),
                'net_weight_lbs': m.get('net_weight_lbs'),
                'gross_weight_lbs': m.get('gross_weight_lbs'),
                'order_sku_links': m.get('order_sku_links') or [],
            })
        except Exception:
            variant_out.setdefault('order_sku_links', [])
        return variant_out

    def handle_spec_main_images_api(self, environ, method, start_response):
        """规格主图管理：GET 按 variant_id 读主图列表 + 变体明细；PATCH 仅更新关联下单 SKU（不改售价等销售字段）。"""
        try:
            if method == 'GET':
                query_params = parse_qs(environ.get('QUERY_STRING', ''))
                variant_id = self._parse_int(query_params.get('variant_id', [''])[0] or query_params.get('id', [''])[0])
                if not variant_id:
                    return self.send_json({'status': 'error', 'message': 'Missing variant_id'}, start_response)

                with self._get_db_connection() as conn:
                    items = self._read_sales_product_image_items(conn, sales_product_id=None, variant_id=variant_id) or []
                    variant_out = self._spec_main_images_variant_detail(conn, variant_id)
                return self.send_json({'status': 'success', 'items': items, 'variant': variant_out}, start_response)

            if method == 'PATCH':
                data = self._read_json_body(environ) or {}
                variant_id = self._parse_int(data.get('variant_id'))
                links = self._normalize_sales_order_links(data.get('order_sku_links'))
                if not variant_id:
                    return self.send_json({'status': 'error', 'message': 'Missing variant_id'}, start_response)
                if not links:
                    return self.send_json({'status': 'error', 'message': '请至少保留一条关联下单SKU及数量'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT id, sku_family_id FROM sales_product_variants WHERE id=%s LIMIT 1",
                            (int(variant_id),),
                        )
                        vr = cur.fetchone() or {}
                    if not vr.get('id'):
                        return self.send_json({'status': 'error', 'message': '规格不存在'}, start_response)
                    sku_family_id = self._parse_int(vr.get('sku_family_id'))
                    if not sku_family_id:
                        return self.send_json({'status': 'error', 'message': '变体缺少货号信息'}, start_response)
                    bundle = self._derive_sales_order_links_bundle(conn, sku_family_id, links)
                    inferred_fid = self._parse_int(bundle.get('sku_family_id'))
                    if inferred_fid and inferred_fid != int(sku_family_id):
                        return self.send_json({'status': 'error', 'message': '关联下单SKU须属于当前规格所属货号'}, start_response)
                    self._replace_sales_variant_order_links(conn, variant_id, links)
                    variant_out = self._spec_main_images_variant_detail(conn, variant_id)
                return self.send_json({'status': 'success', 'variant': variant_out}, start_response)

            return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_gallery_sku_family_picker_api(self, environ, method, start_response):
        """货号（父体）下拉：供规格主图管理等页面新增规格时选择 product_families。"""
        try:
            if method != 'GET':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id, sku_family FROM product_families ORDER BY sku_family ASC, id ASC"
                    )
                    rows = cur.fetchall() or []
            items = []
            for r in rows or []:
                sid = self._parse_int(r.get('id')) or 0
                if not sid:
                    continue
                items.append({
                    'sku_family_id': sid,
                    'sku_family': str(r.get('sku_family') or '').strip(),
                })
            return self.send_json({'status': 'success', 'items': items}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_spec_main_image_variant_create_api(self, environ, method, start_response):
        """规格主图管理：按货号 + 规格名 + 面料创建或复用 sales_product_variants 行（与导入/销售产品逻辑共用）。"""
        try:
            if method != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            data = self._read_json_body(environ) or {}
            sku_family_id = self._parse_int(data.get('sku_family_id'))
            spec_name = str(data.get('spec_name') or '').strip()
            fabric_id = self._parse_int(data.get('fabric_id')) or None
            fabric = str(data.get('fabric') or '').strip()
            if not sku_family_id:
                return self.send_json({'status': 'error', 'message': '请选择货号（父体）'}, start_response)
            if not spec_name:
                return self.send_json({'status': 'error', 'message': '请填写规格名称'}, start_response)
            with self._get_db_connection() as conn:
                has_fid = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                has_fab_txt = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                if has_fid and not fabric_id:
                    return self.send_json({'status': 'error', 'message': '请选择面料'}, start_response)
                if not has_fid and has_fab_txt and not fabric:
                    return self.send_json({'status': 'error', 'message': '请填写或选择面料'}, start_response)
                variant_id = self._get_or_create_sales_variant(
                    conn, sku_family_id, spec_name, fabric, fabric_id=fabric_id
                )
                links = self._normalize_sales_order_links(data.get('order_sku_links'))
                if links:
                    bundle = self._derive_sales_order_links_bundle(conn, sku_family_id, links)
                    inferred_fid = self._parse_int(bundle.get('sku_family_id'))
                    if inferred_fid and inferred_fid != int(sku_family_id):
                        return self.send_json({'status': 'error', 'message': '关联下单SKU须属于所选货号'}, start_response)
                    self._replace_sales_variant_order_links(conn, variant_id, links)
            return self.send_json({
                'status': 'success',
                'variant_id': int(variant_id or 0),
                'message': '规格已就绪（若组合已存在则返回已有规格）',
            }, start_response)
        except ValueError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _collect_spec_variant_delete_block_reasons(self, conn, cur, variant_id):
        """与单条删除 API 一致：不可删时返回原因列表；空列表表示可删（下单产品侧 ON DELETE CASCADE 不拦截）。"""
        blocks = []
        variant_id = self._parse_int(variant_id)
        if not variant_id:
            return blocks
        cur.execute(
            "SELECT COUNT(*) AS c FROM sales_products WHERE variant_id=%s",
            (variant_id,),
        )
        n_sp = int((cur.fetchone() or {}).get('c') or 0)
        if n_sp:
            blocks.append(f'已关联 {n_sp} 条销售平台SKU（sales_products）')

        if self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id'):
            cur.execute(
                "SELECT COUNT(*) AS c FROM sales_variant_image_mappings WHERE variant_id=%s",
                (variant_id,),
            )
            n_im = int((cur.fetchone() or {}).get('c') or 0)
            if n_im:
                blocks.append(f'仍有关联主图映射 {n_im} 条（sales_variant_image_mappings）')

        if self._table_exists_simple(conn, 'sales_forecast_spec_monthly'):
            cur.execute(
                "SELECT COUNT(*) AS c FROM sales_forecast_spec_monthly WHERE variant_id=%s",
                (variant_id,),
            )
            n_fc = int((cur.fetchone() or {}).get('c') or 0)
            if n_fc:
                blocks.append(f'存在销量预测（规格×月）{n_fc} 条（sales_forecast_spec_monthly）')

        if self._table_exists_simple(conn, 'sales_forecast_order_sku_monthly'):
            if self._table_has_column(conn, 'sales_forecast_order_sku_monthly', 'variant_id'):
                cur.execute(
                    "SELECT COUNT(*) AS c FROM sales_forecast_order_sku_monthly WHERE variant_id=%s",
                    (variant_id,),
                )
                n_fc2 = int((cur.fetchone() or {}).get('c') or 0)
                if n_fc2:
                    blocks.append(
                        f'存在销量预测（下单SKU×月，旧结构）{n_fc2} 条（sales_forecast_order_sku_monthly）'
                    )
        return blocks

    def _collect_spec_variant_delete_block_reasons_bulk(self, conn, cur, variant_ids):
        """按 variant_id 聚合不可删原因（与单条 _collect 文案一致），用于批量校验/删除。"""
        out = {}
        ids = []
        for v in variant_ids or []:
            vid = self._parse_int(v)
            if vid:
                ids.append(vid)
                out[vid] = []
        if not ids:
            return out
        ph = ','.join(['%s'] * len(ids))

        cur.execute(
            f'SELECT variant_id, COUNT(*) AS c FROM sales_products WHERE variant_id IN ({ph}) GROUP BY variant_id',
            ids,
        )
        for row in cur.fetchall() or []:
            vid = self._parse_int(row.get('variant_id'))
            c = int((row or {}).get('c') or 0)
            if vid and c:
                out[vid].append(f'已关联 {c} 条销售平台SKU（sales_products）')

        if self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id'):
            cur.execute(
                f'SELECT variant_id, COUNT(*) AS c FROM sales_variant_image_mappings WHERE variant_id IN ({ph}) GROUP BY variant_id',
                ids,
            )
            for row in cur.fetchall() or []:
                vid = self._parse_int(row.get('variant_id'))
                c = int((row or {}).get('c') or 0)
                if vid and c:
                    out[vid].append(f'仍有关联主图映射 {c} 条（sales_variant_image_mappings）')

        if self._table_exists_simple(conn, 'sales_forecast_spec_monthly'):
            cur.execute(
                f'SELECT variant_id, COUNT(*) AS c FROM sales_forecast_spec_monthly WHERE variant_id IN ({ph}) GROUP BY variant_id',
                ids,
            )
            for row in cur.fetchall() or []:
                vid = self._parse_int(row.get('variant_id'))
                c = int((row or {}).get('c') or 0)
                if vid and c:
                    out[vid].append(f'存在销量预测（规格×月）{c} 条（sales_forecast_spec_monthly）')

        if self._table_exists_simple(conn, 'sales_forecast_order_sku_monthly'):
            if self._table_has_column(conn, 'sales_forecast_order_sku_monthly', 'variant_id'):
                cur.execute(
                    f'SELECT variant_id, COUNT(*) AS c FROM sales_forecast_order_sku_monthly WHERE variant_id IN ({ph}) GROUP BY variant_id',
                    ids,
                )
                for row in cur.fetchall() or []:
                    vid = self._parse_int(row.get('variant_id'))
                    c = int((row or {}).get('c') or 0)
                    if vid and c:
                        out[vid].append(
                            f'存在销量预测（下单SKU×月，旧结构）{c} 条（sales_forecast_order_sku_monthly）'
                        )
        return out

    def handle_spec_main_image_variant_delete_api(self, environ, method, start_response):
        """规格主图管理：删除销售变体行。
        无销售平台SKU、无主图、无销量预测占用时可删；关联下单SKU链接由库表 ON DELETE CASCADE 随变体一并删除。
        """
        try:
            if method != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            data = self._read_json_body(environ) or {}
            variant_id = self._parse_int(data.get('variant_id'))
            if not variant_id:
                return self.send_json({'status': 'error', 'message': 'Missing variant_id'}, start_response)

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    blocks = self._collect_spec_variant_delete_block_reasons(conn, cur, variant_id)

                if blocks:
                    return self.send_json(
                        {
                            'status': 'error',
                            'message': '无法删除：' + '；'.join(blocks),
                            'blocks': blocks,
                        },
                        start_response,
                    )

                with conn.cursor() as cur:
                    cur.execute("DELETE FROM sales_product_variants WHERE id=%s", (variant_id,))
                    if not cur.rowcount:
                        return self.send_json({'status': 'error', 'message': '规格不存在或已删除'}, start_response)

            return self.send_json({'status': 'success', 'message': '已删除规格'}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_spec_main_image_variants_batch_validate_api(self, environ, method, start_response):
        """批量校验：货号×面料是否在 fabric_product_families 绑定，及是否与外键数据冲突（下单产品侧可级联，不拦截）。"""
        try:
            if method != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            data = self._read_json_body(environ) or {}
            raw_ids = data.get('variant_ids') or data.get('ids') or []
            if not isinstance(raw_ids, list):
                return self.send_json({'status': 'error', 'message': 'variant_ids 须为数组'}, start_response)
            variant_ids = []
            seen = set()
            for x in raw_ids:
                vid = self._parse_int(x)
                if vid and vid not in seen:
                    seen.add(vid)
                    variant_ids.append(vid)
            if not variant_ids:
                return self.send_json({'status': 'error', 'message': '请至少选择一个规格'}, start_response)
            if len(variant_ids) > 800:
                return self.send_json({'status': 'error', 'message': '一次最多校验 800 条'}, start_response)

            rows_out = []
            with self._get_db_connection() as conn:
                has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                ph = ','.join(['%s'] * len(variant_ids))
                sel_cols = ['v.id', 'v.spec_name', 'v.sku_family_id', 'pf.sku_family']
                join_fm = ''
                if has_fabric_id:
                    sel_cols.append('v.fabric_id')
                    join_fm = 'LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id'
                    sel_cols.append('fm.fabric_code AS fm_code')
                    sel_cols.append('fm.fabric_name_en AS fm_name_en')
                if has_fabric_text:
                    sel_cols.append('v.fabric AS fabric_text')
                with conn.cursor() as cur:
                    blocks_map = self._collect_spec_variant_delete_block_reasons_bulk(conn, cur, variant_ids)
                    cur.execute(
                        f"""
                        SELECT {', '.join(sel_cols)}
                        FROM sales_product_variants v
                        LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                        {join_fm}
                        WHERE v.id IN ({ph})
                        """,
                        variant_ids,
                    )
                    by_id = {}
                    for r in cur.fetchall() or []:
                        iid = self._parse_int(r.get('id'))
                        if iid:
                            by_id[iid] = r

                    fpf_pair_order = []
                    fpf_seen = set()
                    pre_meta = []
                    for variant_id in variant_ids:
                        r = by_id.get(variant_id)
                        if not r:
                            pre_meta.append((variant_id, None, None, None))
                            continue
                        sku_family_id = self._parse_int(r.get('sku_family_id')) or 0
                        fabric_id_col = self._parse_int(r.get('fabric_id')) if has_fabric_id else 0
                        fabric_text = str(r.get('fabric_text') or '').strip() if has_fabric_text else ''
                        resolved_fabric_id = fabric_id_col if fabric_id_col else None
                        if not resolved_fabric_id and fabric_text:
                            resolved_fabric_id = self._resolve_fabric_material_id_from_label(
                                conn, fabric_text, cur=cur
                            )
                        pre_meta.append((variant_id, sku_family_id, resolved_fabric_id, r))
                        if sku_family_id and resolved_fabric_id:
                            t = (sku_family_id, resolved_fabric_id)
                            if t not in fpf_seen:
                                fpf_seen.add(t)
                                fpf_pair_order.append(t)

                    fpf_linked = set()
                    if fpf_pair_order:
                        tpl = ','.join(['(%s,%s)'] * len(fpf_pair_order))
                        flat = [x for pair in fpf_pair_order for x in pair]
                        cur.execute(
                            f'SELECT sku_family_id, fabric_id FROM fabric_product_families WHERE (sku_family_id, fabric_id) IN ({tpl})',
                            flat,
                        )
                        for row in cur.fetchall() or []:
                            sf = self._parse_int(row.get('sku_family_id')) or 0
                            fd = self._parse_int(row.get('fabric_id')) or 0
                            if sf and fd:
                                fpf_linked.add((sf, fd))

                    for variant_id, sku_family_id, resolved_fabric_id, r in pre_meta:
                        if r is None:
                            rows_out.append(
                                {
                                    'variant_id': variant_id,
                                    'sku_family': '',
                                    'spec_name': '',
                                    'fabric_display': '',
                                    'invalid_family_fabric': True,
                                    'association_note': '规格不存在或已删除',
                                    'can_delete': False,
                                    'cannot_delete_reasons': ['记录不存在'],
                                }
                            )
                            continue

                        sku_family = str(r.get('sku_family') or '').strip()
                        spec_name = str(r.get('spec_name') or '').strip()
                        fabric_id_col = self._parse_int(r.get('fabric_id')) if has_fabric_id else 0
                        fabric_text = str(r.get('fabric_text') or '').strip() if has_fabric_text else ''
                        if has_fabric_id:
                            code = str(r.get('fm_code') or '').strip()
                            name_en = str(r.get('fm_name_en') or '').strip()
                            fabric_display = ' / '.join([p for p in (code, name_en) if p]) or str(fabric_id_col or '')
                        else:
                            fabric_display = fabric_text

                        family_fabric_linked = False
                        invalid_family_fabric = False
                        association_note = ''
                        if not sku_family_id:
                            invalid_family_fabric = True
                            association_note = '缺少货号（sku_family_id）'
                        elif not resolved_fabric_id:
                            invalid_family_fabric = True
                            association_note = '无法解析面料主数据 id，无法校验 fabric_product_families'
                        else:
                            family_fabric_linked = (sku_family_id, resolved_fabric_id) in fpf_linked
                            if not family_fabric_linked:
                                invalid_family_fabric = True
                                association_note = '货号与面料未在 fabric_product_families 中绑定'

                        del_blocks = list(blocks_map.get(variant_id) or [])
                        can_delete = len(del_blocks) == 0

                        rows_out.append(
                            {
                                'variant_id': variant_id,
                                'sku_family': sku_family,
                                'spec_name': spec_name,
                                'fabric_display': fabric_display,
                                'invalid_family_fabric': bool(invalid_family_fabric),
                                'association_note': association_note,
                                'can_delete': bool(can_delete),
                                'cannot_delete_reasons': del_blocks,
                            }
                        )

            problem_rows = [
                row
                for row in rows_out
                if row.get('invalid_family_fabric') or not row.get('can_delete')
            ]
            valid_omitted = len(rows_out) - len(problem_rows)
            return self.send_json(
                {'status': 'success', 'rows': problem_rows, 'valid_omitted_count': valid_omitted},
                start_response,
            )
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_spec_main_image_variants_batch_delete_api(self, environ, method, start_response):
        """批量删除：服务端再次校验每条可删后再 DELETE。"""
        try:
            if method != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            data = self._read_json_body(environ) or {}
            if not data.get('delete_confirmed'):
                return self.send_json({'status': 'error', 'message': '请先勾选二次确认'}, start_response)
            raw_ids = data.get('variant_ids') or data.get('ids') or []
            if not isinstance(raw_ids, list):
                return self.send_json({'status': 'error', 'message': 'variant_ids 须为数组'}, start_response)
            variant_ids = []
            seen = set()
            for x in raw_ids:
                vid = self._parse_int(x)
                if vid and vid not in seen:
                    seen.add(vid)
                    variant_ids.append(vid)
            if not variant_ids:
                return self.send_json({'status': 'error', 'message': '请选择要删除的规格'}, start_response)
            if len(variant_ids) > 500:
                return self.send_json({'status': 'error', 'message': '一次最多删除 500 条'}, start_response)

            deleted = []
            failed = []
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    blocks_map = self._collect_spec_variant_delete_block_reasons_bulk(conn, cur, variant_ids)
                    for variant_id in variant_ids:
                        bl = blocks_map.get(variant_id) or []
                        if bl:
                            failed.append({'variant_id': variant_id, 'reasons': bl})
                    deletable = [vid for vid in variant_ids if not (blocks_map.get(vid) or [])]
                    if deletable:
                        ph_sel = ','.join(['%s'] * len(deletable))
                        cur.execute(
                            f'SELECT id FROM sales_product_variants WHERE id IN ({ph_sel})',
                            deletable,
                        )
                        existing = {
                            self._parse_int(row.get('id'))
                            for row in (cur.fetchall() or [])
                            if self._parse_int(row.get('id'))
                        }
                        for vid in deletable:
                            if vid not in existing:
                                failed.append({'variant_id': vid, 'reasons': ['规格不存在或已删除']})
                        to_del = [vid for vid in deletable if vid in existing]
                        if to_del:
                            ph_del = ','.join(['%s'] * len(to_del))
                            try:
                                cur.execute(
                                    f'DELETE FROM sales_product_variants WHERE id IN ({ph_del})',
                                    to_del,
                                )
                                deleted = list(to_del)
                            except Exception as ex:
                                err = str(ex)
                                for vid in to_del:
                                    failed.append({'variant_id': vid, 'reasons': [err]})

            return self.send_json(
                {
                    'status': 'success',
                    'deleted_ids': deleted,
                    'failed': failed,
                    'message': f'已删除 {len(deleted)} 条' + (f'，{len(failed)} 条未删除' if failed else ''),
                },
                start_response,
            )
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_gallery_variant_picker_api(self, environ, method, start_response):
        """给 gallery 弹窗提供规格选择数据（货号+面料+规格）。"""
        try:
            if method != 'GET':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)

            with self._get_db_connection() as conn:
                has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                fabric_join = "LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if has_fabric_id else ""
                if has_fabric_id and has_fabric_text:
                    fabric_code_expr = "COALESCE(fm.fabric_code, v.fabric) AS fabric_code"
                elif has_fabric_id:
                    fabric_code_expr = "fm.fabric_code AS fabric_code"
                elif has_fabric_text:
                    fabric_code_expr = "v.fabric AS fabric_code"
                else:
                    fabric_code_expr = "'' AS fabric_code"

                with conn.cursor() as cur:
                    cur.execute(
                        f"""
                        SELECT
                            v.id,
                            v.spec_name,
                            v.sku_family_id,
                            pf.sku_family,
                            {('v.fabric_id' if has_fabric_id else '0 AS fabric_id')},
                            {fabric_code_expr},
                            {('fm.fabric_name_en AS fabric_name_en' if has_fabric_id else "'' AS fabric_name_en")}
                        FROM sales_product_variants v
                        LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                        {fabric_join}
                        ORDER BY pf.sku_family ASC, v.spec_name ASC, {('fm.fabric_name_en' if has_fabric_id else 'v.id')} ASC, v.id ASC
                        """
                    )
                    rows = cur.fetchall() or []
            items = []
            for r in rows:
                items.append({
                    'variant_id': self._parse_int(r.get('id')) or 0,
                    'sku_family_id': self._parse_int(r.get('sku_family_id')) or 0,
                    'sku_family': (r.get('sku_family') or '').strip(),
                    'spec_name': (r.get('spec_name') or '').strip(),
                    'fabric_id': self._parse_int(r.get('fabric_id')) or 0,
                    'fabric_code': (r.get('fabric_code') or '').strip(),
                    'fabric_name_en': (r.get('fabric_name_en') or '').strip(),
                })
            return self.send_json({'status': 'success', 'items': items}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_gallery_image_types_api(self, environ, method, start_response):
        """Gallery 弹窗：返回可用于关联面料/规格/下单产品的图片类型及适用范围（只读）。"""
        try:
            if method != 'GET':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT id, name, is_enabled,
                               COALESCE(applies_fabric, 1) AS applies_fabric,
                               COALESCE(applies_sales, 1) AS applies_sales,
                               COALESCE(applies_order_product, 1) AS applies_order_product
                        FROM image_types
                        WHERE is_enabled = 1
                          AND (
                            COALESCE(applies_fabric, 1) = 1
                            OR COALESCE(applies_sales, 1) = 1
                            OR COALESCE(applies_order_product, 1) = 1
                          )
                        ORDER BY sort_order ASC, id ASC
                        """
                    )
                    rows = cur.fetchall() or []
            items = []
            for r in rows or []:
                items.append({
                    'id': self._parse_int(r.get('id')) or 0,
                    'name': str(r.get('name') or '').strip(),
                    'applies_fabric': bool(int(r.get('applies_fabric') or 0)),
                    'applies_sales': bool(int(r.get('applies_sales') or 0)),
                    'applies_order_product': bool(int(r.get('applies_order_product') or 0)),
                })
            return self.send_json({'status': 'success', 'items': items}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_gallery_fabric_picker_api(self, environ, method, start_response):
        """Gallery 弹窗：面料列表（搜索用）。可选 sku_family_id：仅返回 fabric_product_families 中与该货号绑定的面料。"""
        try:
            if method != 'GET':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            query_params = parse_qs(environ.get('QUERY_STRING', '') or '')
            sku_family_id = self._parse_int((query_params.get('sku_family_id', [''])[0] or '').strip()) or 0

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    if sku_family_id > 0:
                        cur.execute(
                            """
                            SELECT fm.id, fm.fabric_code, fm.fabric_name_en
                            FROM fabric_materials fm
                            INNER JOIN fabric_product_families fpf
                                ON fpf.fabric_id = fm.id AND fpf.sku_family_id = %s
                            ORDER BY fm.fabric_code ASC, fm.id ASC
                            """,
                            (sku_family_id,),
                        )
                    else:
                        cur.execute(
                            "SELECT id, fabric_code, fabric_name_en FROM fabric_materials ORDER BY fabric_code ASC, id ASC"
                        )
                    rows = cur.fetchall() or []
            items = []
            for r in rows or []:
                fid = self._parse_int(r.get('id')) or 0
                if not fid:
                    continue
                items.append({
                    'fabric_id': fid,
                    'fabric_code': str(r.get('fabric_code') or '').strip(),
                    'fabric_name_en': str(r.get('fabric_name_en') or '').strip(),
                })
            return self.send_json({'status': 'success', 'items': items}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_gallery_order_product_picker_api(self, environ, method, start_response):
        """Gallery 弹窗：下单产品列表（搜索用，条数上限以避免过大响应）。"""
        try:
            if method != 'GET':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT op.id, op.sku, op.spec_qty_short, pf.sku_family
                        FROM order_products op
                        LEFT JOIN product_families pf ON pf.id = op.sku_family_id
                        ORDER BY pf.sku_family ASC, op.sku ASC, op.id ASC
                        LIMIT 8000
                        """
                    )
                    rows = cur.fetchall() or []
            items = []
            for r in rows or []:
                oid = self._parse_int(r.get('id')) or 0
                if not oid:
                    continue
                items.append({
                    'order_product_id': oid,
                    'sku': str(r.get('sku') or '').strip(),
                    'sku_family': str(r.get('sku_family') or '').strip(),
                    'spec_qty_short': str(r.get('spec_qty_short') or '').strip(),
                })
            return self.send_json({'status': 'success', 'items': items}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_gallery_apply_image_api(self, environ, method, start_response):
        """
        从 gallery 一键关联：面料（fabric_image_mappings）、销售规格（sales_variant_image_mappings）、
        下单产品（order_product_image_mappings）。落盘优先级：『面料』 > 货号/主图/规格-面料 > 货号/配件图/…

        link_sync=1 且 JSON 同时包含 variant_ids / fabric_ids / order_product_ids 时：先清空该图在三张映射表中的
        全部记录再按请求重建（用于弹窗「全量同步」与移除绑定）。仅传部分字段的旧调用方不受影响。
        orphan_file_action=recycle 且上述三类均为空：在通过 A+ 引用检查后移入回收站并删除 image_assets。
        """
        try:
            if method != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)

            user_id = 0
            try:
                user_id = self._parse_int(self._get_session_user(environ)) or 0
            except Exception:
                user_id = 0
            if not user_id:
                return self.send_json({'status': 'error', 'message': '必须登录才能执行该操作'}, start_response)

            data = self._read_json_body(environ) or {}
            image_path_b64 = str(data.get('image_path_b64') or '').strip()
            variant_ids = data.get('variant_ids') or []
            fabric_ids = data.get('fabric_ids') or []
            order_product_ids = data.get('order_product_ids') or []
            action = str(data.get('action') or '').strip().lower() or 'copy'
            image_type_name = str(data.get('image_type_name') or '').strip()
            prompt_duplicate = str(data.get('prompt_duplicate') or '').strip().lower() in ('1', 'true', 'yes', 'on')

            if not image_path_b64:
                return self.send_json({'status': 'error', 'message': '缺少图片路径'}, start_response)
            if not image_type_name:
                return self.send_json({'status': 'error', 'message': '请选择图片类型'}, start_response)

            vids = []
            if isinstance(variant_ids, (list, tuple)):
                for v in variant_ids:
                    vid = self._parse_int(v) or 0
                    if vid > 0:
                        vids.append(vid)
            vids = sorted(set(vids))

            fids = []
            if isinstance(fabric_ids, (list, tuple)):
                for x in fabric_ids:
                    fid = self._parse_int(x) or 0
                    if fid > 0:
                        fids.append(fid)
            fids = sorted(set(fids))

            opids = []
            if isinstance(order_product_ids, (list, tuple)):
                for x in order_product_ids:
                    oid = self._parse_int(x) or 0
                    if oid > 0:
                        opids.append(oid)
            opids = sorted(set(opids))

            if action not in ('copy', 'move'):
                action = 'copy'

            try:
                rel_raw = base64.b64decode(image_path_b64)
            except Exception:
                return self.send_json({'status': 'error', 'message': '图片路径不是合法 Base64'}, start_response)

            abs_source = os.path.join(self._resources_root(), rel_raw)
            abs_source = self._safe_fsencode(abs_source)
            try:
                if not os.path.isfile(abs_source):
                    return self.send_json({'status': 'error', 'message': '源图片不存在'}, start_response)
            except Exception:
                return self.send_json({'status': 'error', 'message': '源图片不存在'}, start_response)

            filename = os.path.basename(self._safe_fsdecode(abs_source))
            if not self._is_image_name(filename):
                return self.send_json({'status': 'error', 'message': '不支持的图片类型'}, start_response)

            try:
                with open(abs_source, 'rb') as f:
                    content = f.read() or b''
            except Exception as e:
                return self.send_json({'status': 'error', 'message': f'读取源图片失败: {str(e)}'}, start_response)
            if not content:
                return self.send_json({'status': 'error', 'message': '源图片为空'}, start_response)
            sha256 = self._sha256_hex(content)

            with self._get_db_connection() as conn:
                image_type_id = self._get_image_type_id_by_name(conn, image_type_name)
                if not image_type_id:
                    return self.send_json({'status': 'error', 'message': f'未知图片类型: {image_type_name}'}, start_response)
                ok_t, err_t = self._gallery_validate_type_for_link_targets(
                    conn, int(image_type_id), bool(fids), bool(vids), bool(opids),
                )
                if not ok_t:
                    return self.send_json({'status': 'error', 'message': err_t or '图片类型与关联目标不匹配'}, start_response)

                existing = self._find_image_asset_by_sha256(conn, sha256)
                aid = int(existing.get('id') or 0) if existing else 0
                storage_path = (existing.get('storage_path') or '').strip() if existing else ''
                created_new_asset = False
                file_op_done = False
                recycled_duplicate = False
                link_sync = str(data.get('link_sync') or '').strip().lower() in ('1', 'true', 'yes', 'on')
                orphan_file_action = str(data.get('orphan_file_action') or '').strip().lower()

                raw_body = data if isinstance(data, dict) else {}
                link_sync_full = link_sync and all(k in raw_body for k in ('variant_ids', 'fabric_ids', 'order_product_ids'))

                if not vids and not fids and not opids:
                    if not (link_sync_full and aid):
                        return self.send_json({'status': 'error', 'message': '请至少选择面料、规格或下单产品之一'}, start_response)

                cleared_all_bindings = bool(link_sync_full and aid and (not vids) and (not fids) and (not opids))
                if cleared_all_bindings and orphan_file_action in ('recycle', 'delete', 'trash'):
                    if self._gallery_count_aplus_asset_refs(conn, aid) > 0:
                        return self.send_json(
                            {
                                'status': 'error',
                                'message': '该图片仍被 A+ 等板块引用，无法移入回收站并删除图片库记录。请先解除 A+ 引用，或选择「仅移除绑定（保留文件）」。',
                            },
                            start_response,
                        )

                # 仅当请求体同时携带三类 id 列表时才做全量同步删除，避免其它入口只传 variant_ids 时误清面料/下单绑定
                if aid and link_sync_full:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM sales_variant_image_mappings WHERE image_asset_id=%s", (int(aid),))
                        if self._has_required_tables(['fabric_image_mappings']):
                            cur.execute("DELETE FROM fabric_image_mappings WHERE image_asset_id=%s", (int(aid),))
                        if self._has_required_tables(['order_product_image_mappings']):
                            cur.execute("DELETE FROM order_product_image_mappings WHERE image_asset_id=%s", (int(aid),))

                db_vids = self._get_asset_referenced_variant_ids(conn, aid) if aid else []
                db_fids = self._get_asset_referenced_fabric_ids(conn, aid) if aid else []
                db_opids = self._get_asset_referenced_order_product_ids(conn, aid) if aid else []

                variant_union = sorted(set(list(vids) + list(db_vids)))
                fabric_union = sorted(set(list(fids) + list(db_fids)))
                has_any_fabric = len(fabric_union) > 0

                sku_families = self._detect_cross_sku_family_by_variant_ids(conn, variant_union)
                is_cross_sku = len(sku_families) > 1
                is_multi_variant_same_sku = (not is_cross_sku) and (len(variant_union) > 1) and (len(sku_families) == 1)

                if aid and storage_path:
                    try:
                        canonical_abs = self._abs_from_storage_path(storage_path)
                        src_key = self._fs_realpath_key(abs_source)
                        canon_key = self._fs_realpath_key(canonical_abs)
                        if src_key and canon_key and src_key != canon_key:
                            moved_ok, _dst, _err = self._move_file_to_listing_recycle_bin(abs_source, '重复')
                            recycled_duplicate = bool(moved_ok)
                    except Exception:
                        pass

                if not aid:
                    target_abs = self._gallery_resolve_new_asset_folder(conn, fids, vids, opids)
                    if not target_abs:
                        return self.send_json({'status': 'error', 'message': '无法计算目标文件夹'}, start_response)
                    if isinstance(target_abs, str):
                        target_abs = self._safe_fsencode(target_abs)
                    if not os.path.exists(target_abs):
                        os.makedirs(target_abs, exist_ok=True)

                    ext = self._guess_image_ext(filename, content)
                    base0 = os.path.splitext(filename)[0].strip() or sha256[:12]
                    safe_base0 = self._sanitize_filename_component(base0, 80) or sha256[:12]
                    final_name = self._next_available_filename(target_abs, f"{safe_base0}{ext}")
                    abs_target = os.path.join(target_abs, self._safe_fsencode(final_name))

                    try:
                        if action == 'move':
                            if os.path.abspath(os.path.dirname(abs_source)) != os.path.abspath(target_abs):
                                if not self._listing_paths_equivalent(abs_source, abs_target):
                                    os.replace(abs_source, abs_target)
                                    file_op_done = True
                                else:
                                    abs_target = abs_source
                            else:
                                abs_target = abs_source
                        else:
                            if (not self._listing_paths_equivalent(abs_source, abs_target)) and (
                                os.path.abspath(abs_source) != os.path.abspath(abs_target)
                            ):
                                shutil.copy2(abs_source, abs_target)
                                file_op_done = True
                            else:
                                abs_target = abs_source
                    except Exception as e:
                        return self.send_json({'status': 'error', 'message': f'复制/移动文件失败: {str(e)}'}, start_response)

                    storage_path = self._storage_path_from_abs(abs_target)
                    if not storage_path:
                        return self.send_json({'status': 'error', 'message': '无法计算 storage_path'}, start_response)

                    try:
                        self._tx_begin(conn)
                        with conn.cursor() as cur:
                            aid = int(self._insert_image_asset_dynamic(conn, cur, {
                                'sha256': sha256,
                                'storage_path': storage_path,
                                'filename': final_name,
                                'original_filename': filename,
                                'image_type_id': int(image_type_id),
                                'created_by': user_id,
                            }) or 0)
                        self._tx_commit(conn)
                        created_new_asset = True
                    except Exception as e:
                        self._tx_rollback(conn)
                        try:
                            if storage_path:
                                self._safe_unlink(self._abs_from_storage_path(storage_path))
                        except Exception:
                            pass
                        return self.send_json({'status': 'error', 'message': f'写入数据库失败: {str(e)}'}, start_response)

                try:
                    if self._table_has_column(conn, 'image_assets', 'image_type_id') and aid and image_type_id:
                        with conn.cursor() as cur:
                            cur.execute(
                                "UPDATE image_assets SET image_type_id=%s WHERE id=%s",
                                (int(image_type_id), int(aid)),
                            )
                except Exception:
                    pass

                rehome_kind = ''
                rehomed = False
                if aid and storage_path and (not has_any_fabric) and (is_cross_sku or is_multi_variant_same_sku):
                    try:
                        if is_cross_sku:
                            rehome_kind = 'cross_sku'
                            target_folder = self._ensure_listing_sales_global_common_folder()
                        else:
                            rehome_kind = 'same_sku_common'
                            target_folder = self._ensure_listing_sales_common_folder(sku_families[0])
                        cur_abs = self._abs_from_storage_path(storage_path)
                        if cur_abs and os.path.exists(cur_abs):
                            if os.path.abspath(os.path.dirname(cur_abs)) != os.path.abspath(target_folder):
                                # cur_abs is often bytes; basename must be decoded before f-strings in _next_available_filename
                                base_str = self._safe_fsdecode(os.path.basename(cur_abs))
                                new_name = self._next_available_filename(target_folder, base_str)
                                dst_abs = os.path.join(target_folder, self._safe_fsencode(new_name))
                                if not self._listing_paths_equivalent(cur_abs, dst_abs):
                                    os.replace(cur_abs, dst_abs)
                                    storage_path = self._storage_path_from_abs(dst_abs)
                                    with conn.cursor() as cur:
                                        cur.execute("UPDATE image_assets SET storage_path=%s WHERE id=%s", (storage_path, int(aid)))
                                    rehomed = True
                    except Exception:
                        pass

                if aid and storage_path and action == 'move' and (not has_any_fabric) and (not is_cross_sku) and (not is_multi_variant_same_sku) and len(variant_union) == 1:
                    try:
                        vid0 = int(variant_union[0])
                        folder_info = self._resolve_sales_variant_folder_by_variant_id(vid0, ensure_folder=True)
                        target_folder = folder_info.get('folder_path')
                        if target_folder:
                            cur_abs = self._abs_from_storage_path(storage_path)
                            cur_b = cur_abs if isinstance(cur_abs, (bytes, bytearray)) else self._safe_fsencode(cur_abs)
                            tf = target_folder if isinstance(target_folder, (bytes, bytearray)) else self._safe_fsencode(target_folder)
                            if cur_b and os.path.exists(cur_b) and os.path.abspath(os.path.dirname(cur_b)) != os.path.abspath(tf):
                                base_str = self._safe_fsdecode(os.path.basename(cur_b))
                                new_name = self._next_available_filename(tf, base_str)
                                dst_abs = os.path.join(tf, self._safe_fsencode(new_name))
                                if not self._listing_paths_equivalent(cur_b, dst_abs):
                                    os.replace(cur_b, dst_abs)
                                    storage_path = self._storage_path_from_abs(dst_abs)
                                    with conn.cursor() as cur:
                                        cur.execute("UPDATE image_assets SET storage_path=%s WHERE id=%s", (storage_path, int(aid)))
                                    rehomed = True
                                    file_op_done = True
                                if not rehome_kind:
                                    rehome_kind = 'single_variant'
                    except Exception:
                        pass

                linked = 0
                skipped = 0
                if vids:
                    has_var = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
                    if not has_var:
                        return self.send_json({'status': 'error', 'message': 'sales_variant_image_mappings 缺少 variant_id，无法关联规格'}, start_response)

                    already_linked = set()
                    with conn.cursor() as cur:
                        placeholders = ','.join(['%s'] * len(vids))
                        cur.execute(
                            f"SELECT variant_id FROM sales_variant_image_mappings WHERE variant_id IN ({placeholders}) AND image_asset_id=%s",
                            tuple(vids + [aid]),
                        )
                        for row in (cur.fetchall() or []):
                            already_linked.add(self._parse_int(row.get('variant_id')) or 0)

                    max_sort_map = {vid: 0 for vid in vids}
                    with conn.cursor() as cur:
                        placeholders = ','.join(['%s'] * len(vids))
                        cur.execute(
                            f"SELECT variant_id, COALESCE(MAX(sort_order),0) AS max_sort FROM sales_variant_image_mappings WHERE variant_id IN ({placeholders}) GROUP BY variant_id",
                            tuple(vids),
                        )
                        for row in (cur.fetchall() or []):
                            vid = self._parse_int(row.get('variant_id')) or 0
                            if vid:
                                max_sort_map[vid] = max(0, self._parse_int(row.get('max_sort')) or 0)

                    has_sim_tid = self._table_has_column(conn, 'sales_variant_image_mappings', 'image_type_id')
                    has_sim_created_by = self._table_has_column(conn, 'sales_variant_image_mappings', 'created_by')
                    cols = ['variant_id', 'image_asset_id']
                    if has_sim_tid:
                        cols.append('image_type_id')
                    cols.append('sort_order')
                    if has_sim_created_by:
                        cols.append('created_by')
                    ph = ','.join(['%s'] * len(cols))
                    dup_parts = ['sort_order=VALUES(sort_order)']
                    if has_sim_tid:
                        dup_parts.append('image_type_id=VALUES(image_type_id)')
                    upsert_sql = (
                        f"INSERT INTO sales_variant_image_mappings ({', '.join(cols)}) VALUES ({ph}) "
                        f"ON DUPLICATE KEY UPDATE {', '.join(dup_parts)}"
                    )
                    batch_rows = []
                    for vid in vids:
                        if vid in already_linked:
                            skipped += 1
                            continue
                        sort_order = (max_sort_map.get(vid) or 0) + 1
                        row = [int(vid), int(aid)]
                        if has_sim_tid:
                            row.append(int(image_type_id))
                        row.append(int(sort_order))
                        if has_sim_created_by:
                            row.append(int(user_id) if user_id else None)
                        batch_rows.append(tuple(row))
                    if batch_rows:
                        with conn.cursor() as cur:
                            cur.executemany(upsert_sql, batch_rows)
                        linked = len(batch_rows)

                linked_fab = 0
                skipped_fab = 0
                if fids and self._has_required_tables(['fabric_image_mappings']):
                    fim_has_cb = self._table_has_column(conn, 'fabric_image_mappings', 'created_by')
                    already_f = set()
                    with conn.cursor() as cur:
                        placeholders = ','.join(['%s'] * len(fids))
                        cur.execute(
                            f"SELECT fabric_id FROM fabric_image_mappings WHERE fabric_id IN ({placeholders}) AND image_asset_id=%s",
                            tuple(fids + [aid]),
                        )
                        for row in (cur.fetchall() or []):
                            already_f.add(self._parse_int(row.get('fabric_id')) or 0)
                    max_f_sort = {fid: 0 for fid in fids}
                    with conn.cursor() as cur:
                        placeholders = ','.join(['%s'] * len(fids))
                        cur.execute(
                            f"SELECT fabric_id, COALESCE(MAX(sort_order),0) AS mx FROM fabric_image_mappings WHERE fabric_id IN ({placeholders}) GROUP BY fabric_id",
                            tuple(fids),
                        )
                        for row in (cur.fetchall() or []):
                            fid = self._parse_int(row.get('fabric_id')) or 0
                            if fid:
                                max_f_sort[fid] = max(0, self._parse_int(row.get('mx')) or 0)
                    fab_rows = []
                    for fid in fids:
                        if fid in already_f:
                            skipped_fab += 1
                            continue
                        so = (max_f_sort.get(fid) or 0) + 1
                        if fim_has_cb:
                            fab_rows.append((int(fid), int(aid), int(so), int(user_id) if user_id else None))
                        else:
                            fab_rows.append((int(fid), int(aid), int(so)))
                    if fab_rows:
                        with conn.cursor() as cur:
                            if fim_has_cb:
                                cur.executemany(
                                    "INSERT INTO fabric_image_mappings (fabric_id, image_asset_id, sort_order, created_by) VALUES (%s,%s,%s,%s) "
                                    "ON DUPLICATE KEY UPDATE sort_order=VALUES(sort_order)",
                                    fab_rows,
                                )
                            else:
                                cur.executemany(
                                    "INSERT INTO fabric_image_mappings (fabric_id, image_asset_id, sort_order) VALUES (%s,%s,%s) "
                                    "ON DUPLICATE KEY UPDATE sort_order=VALUES(sort_order)",
                                    fab_rows,
                                )
                        linked_fab = len(fab_rows)

                linked_op = 0
                skipped_op = 0
                if opids and self._has_required_tables(['order_product_image_mappings']):
                    already_o = set()
                    with conn.cursor() as cur:
                        placeholders = ','.join(['%s'] * len(opids))
                        cur.execute(
                            f"SELECT order_product_id FROM order_product_image_mappings WHERE order_product_id IN ({placeholders}) AND image_asset_id=%s",
                            tuple(opids + [aid]),
                        )
                        for row in (cur.fetchall() or []):
                            already_o.add(self._parse_int(row.get('order_product_id')) or 0)
                    max_o = {oid: 0 for oid in opids}
                    with conn.cursor() as cur:
                        placeholders = ','.join(['%s'] * len(opids))
                        cur.execute(
                            f"SELECT order_product_id, COALESCE(MAX(sort_order),0) AS mx FROM order_product_image_mappings WHERE order_product_id IN ({placeholders}) GROUP BY order_product_id",
                            tuple(opids),
                        )
                        for row in (cur.fetchall() or []):
                            oid = self._parse_int(row.get('order_product_id')) or 0
                            if oid:
                                max_o[oid] = max(0, self._parse_int(row.get('mx')) or 0)
                    op_rows = []
                    for oid in opids:
                        if oid in already_o:
                            skipped_op += 1
                            continue
                        so = (max_o.get(oid) or 0) + 1
                        op_rows.append((int(oid), int(aid), int(so)))
                    if op_rows:
                        with conn.cursor() as cur:
                            cur.executemany(
                                "INSERT INTO order_product_image_mappings (order_product_id, image_asset_id, sort_order) VALUES (%s,%s,%s) "
                                "ON DUPLICATE KEY UPDATE sort_order=VALUES(sort_order)",
                                op_rows,
                            )
                        linked_op = len(op_rows)

                if aid:
                    try:
                        self._rehome_image_asset_if_needed(conn, int(aid))
                    except Exception:
                        pass
                    try:
                        with conn.cursor() as cur:
                            cur.execute("SELECT storage_path FROM image_assets WHERE id=%s LIMIT 1", (int(aid),))
                            sp_row = cur.fetchone() or {}
                            if (sp_row.get('storage_path') or '').strip():
                                storage_path = str(sp_row.get('storage_path') or '').strip()
                    except Exception:
                        pass

                asset_deleted = False
                if cleared_all_bindings and orphan_file_action in ('recycle', 'delete', 'trash'):
                    try:
                        with conn.cursor() as cur:
                            cur.execute("SELECT storage_path FROM image_assets WHERE id=%s LIMIT 1", (int(aid),))
                            row_sp = cur.fetchone() or {}
                        sp_del = str(row_sp.get('storage_path') or '').strip()
                        abs_del = self._abs_from_storage_path(sp_del) if sp_del else None
                        if abs_del:
                            try:
                                if os.path.isfile(abs_del):
                                    self._move_file_to_listing_recycle_bin(abs_del, '解绑')
                            except Exception:
                                pass
                        with conn.cursor() as cur:
                            cur.execute("DELETE FROM image_assets WHERE id=%s", (int(aid),))
                        aid = 0
                        storage_path = ''
                        asset_deleted = True
                    except Exception as e:
                        return self.send_json({'status': 'error', 'message': f'移入回收站失败: {str(e)}'}, start_response)

                msg_parts = []
                if cleared_all_bindings:
                    if asset_deleted:
                        msg_parts.append('已移除全部绑定；文件已移入回收站并删除图片库记录')
                    else:
                        msg_parts.append('已移除全部绑定；文件仍保留在当前路径')
                else:
                    if vids:
                        msg_parts.append(f'已关联 {linked} 个规格')
                        if skipped and prompt_duplicate:
                            msg_parts.append(f'（{skipped} 个规格已有关联，已跳过）')
                    if fids:
                        msg_parts.append(f'面料 {linked_fab} 条')
                        if skipped_fab and prompt_duplicate:
                            msg_parts.append(f'（面料跳过 {skipped_fab}）')
                    if opids:
                        msg_parts.append(f'下单产品 {linked_op} 条')
                        if skipped_op and prompt_duplicate:
                            msg_parts.append(f'（下单产品跳过 {skipped_op}）')
                    if not msg_parts:
                        msg_parts.append('未写入新关联')
                    if is_cross_sku and vids and (not has_any_fabric):
                        msg_parts.append('；规格侧检测到跨货号引用')
                    elif is_multi_variant_same_sku and vids and (not has_any_fabric):
                        msg_parts.append('；规格侧同货号多规格引用')
                    if rehome_kind:
                        msg_parts.append('（已尝试归一化路径）' if rehomed else '（归一化失败则保持原路径）')
                    if existing:
                        msg_parts.append('；已复用数据库中的同图记录')
                    elif created_new_asset:
                        msg_parts.append('；已新建 image_assets')

                return self.send_json({
                    'status': 'success',
                    'message': ''.join(msg_parts),
                    'image_asset_id': int(aid) if aid else 0,
                    'storage_path': storage_path,
                    'created_new_asset': bool(created_new_asset),
                    'file_op_done': bool(file_op_done),
                    'recycled_duplicate': bool(recycled_duplicate),
                    'linked_variants': int(linked),
                    'already_linked_variants': int(skipped),
                    'linked_fabrics': int(linked_fab),
                    'skipped_fabrics': int(skipped_fab),
                    'linked_order_products': int(linked_op),
                    'skipped_order_products': int(skipped_op),
                    'linked': int(linked),
                    'already_linked': int(skipped),
                    'rehome_kind': rehome_kind,
                    'rehomed': bool(rehomed),
                    'sku_families': sku_families,
                    'link_sync': bool(link_sync),
                    'link_sync_full': bool(link_sync_full),
                    'asset_deleted': bool(asset_deleted),
                }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _fs_realpath_key(self, path):
        """Stable key for same-file detection (bytes-safe)."""
        if not path:
            return None
        try:
            p = self._safe_fsencode(path)
            if not os.path.exists(p):
                return None
            return os.path.realpath(p)
        except Exception:
            return None

    def _canonical_asset_abs_keys(self, conn, asset_ids):
        """Set of realpath keys for current on-disk canonical files of given image_assets ids."""
        keys = set()
        for aid in asset_ids or []:
            aid = int(aid or 0)
            if aid <= 0:
                continue
            try:
                with conn.cursor() as cur:
                    cur.execute("SELECT storage_path FROM image_assets WHERE id=%s LIMIT 1", (aid,))
                    row = cur.fetchone() or {}
                sp = (row.get('storage_path') or '').strip()
                if not sp:
                    continue
                abs_p = self._abs_from_storage_path(sp)
                k = self._fs_realpath_key(abs_p)
                if k:
                    keys.add(k)
            except Exception:
                continue
        return keys

    def _cleanup_import_by_path_sources(self, conn, source_files, asset_ids):
        """
        After a successful import-by-path bind: remove duplicate copies left in manual staging folders.
        Anything still on disk under source_files that is NOT the canonical asset file is moved to recycle
        (then best-effort unlink).
        """
        if not source_files:
            return {'cleaned': 0, 'skipped_samefile': 0, 'failures': 0}
        canonical = self._canonical_asset_abs_keys(conn, asset_ids)
        cleaned = 0
        skipped = 0
        failures = 0
        for source_file in source_files:
            try:
                sb = self._safe_fsencode(source_file)
                if not os.path.exists(sb):
                    continue
                rk = self._fs_realpath_key(sb)
                if rk and rk in canonical:
                    skipped += 1
                    continue
                moved_ok, _dst, _err = self._move_file_to_listing_recycle_bin(sb, '重复')
                if moved_ok:
                    cleaned += 1
                    continue
                try:
                    self._safe_unlink(sb)
                    cleaned += 1
                except Exception:
                    failures += 1
            except Exception:
                failures += 1
        return {'cleaned': cleaned, 'skipped_samefile': skipped, 'failures': failures}

    def _ensure_listing_sales_common_folder(self, sku_family):
        """Ensure 货号/主图/通用 exists. Return absolute folder path (bytes)."""
        sku_name = (sku_family or '').strip()
        if not sku_name:
            return None
        self._ensure_listing_sku_folder(sku_name)
        base_folder = self._ensure_listing_folder()
        sku_folder = os.path.join(base_folder, self._safe_fsencode(sku_name))
        main_folder = os.path.join(sku_folder, self._safe_fsencode('主图'))
        if not os.path.exists(main_folder):
            os.makedirs(main_folder, exist_ok=True)
        common_folder = os.path.join(main_folder, self._safe_fsencode('通用'))
        if not os.path.exists(common_folder):
            os.makedirs(common_folder, exist_ok=True)
        return common_folder

    def _ensure_listing_sales_channel_folder(self, sku_family):
        """Ensure 货号/主图/通道 exists. Return absolute folder path (bytes)."""
        sku_name = (sku_family or '').strip()
        if not sku_name:
            return None
        self._ensure_listing_sku_folder(sku_name)
        base_folder = self._ensure_listing_folder()
        sku_folder = os.path.join(base_folder, self._safe_fsencode(sku_name))
        main_folder = os.path.join(sku_folder, self._safe_fsencode('主图'))
        if not os.path.exists(main_folder):
            os.makedirs(main_folder, exist_ok=True)
        channel_folder = os.path.join(main_folder, self._safe_fsencode('通道'))
        if not os.path.exists(channel_folder):
            os.makedirs(channel_folder, exist_ok=True)
        return channel_folder

    def _parse_sku_family_from_storage_path(self, storage_path):
        rel = (storage_path or '').strip().replace('\\', '/').strip('/')
        if not rel:
            return ''
        parts = [p for p in rel.split('/') if p]
        skip_roots = {'『通用图片』', '『面料』', '『认证』', '『销售产品图片』'}
        for i, p in enumerate(parts):
            if p in skip_roots:
                continue
            if p in ('主图', '配件图') and i > 0:
                prev = parts[i - 1]
                if prev not in skip_roots:
                    return prev
        return ''

    def _channel_links_table_ready(self, conn):
        return self._has_required_tables(['image_asset_channel_links'])

    def _ensure_image_asset_from_rel_path(self, conn, cur, rel_text, user_id=0):
        """Resolve image_assets row for rel path; create from disk if missing."""
        has_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
        has_dep = self._table_has_column(conn, 'image_assets', 'is_deprecated')
        has_desc = self._table_has_column(conn, 'image_assets', 'description')
        join_type = "LEFT JOIN image_types it ON it.id = ia.image_type_id" if has_tid else ""
        row = self._find_image_asset_row_by_rel_path(
            cur, rel_text, join_type, has_tid, has_dep, has_desc
        )
        aid = self._parse_int(row.get('id')) or 0
        if aid:
            return aid, row
        abs_path, canonical_rel = self._resolve_gallery_abs_path(rel_text)
        if not abs_path or not os.path.isfile(abs_path):
            return 0, {}
        try:
            with open(abs_path, 'rb') as f:
                content = f.read() or b''
        except Exception:
            content = b''
        if not content:
            return 0, {}
        sha256 = self._sha256_hex(content)
        existing = self._find_image_asset_by_sha256(conn, sha256)
        if existing and self._parse_int(existing.get('id')):
            return int(existing.get('id')), existing
        rec = {
            'sha256': sha256,
            'storage_path': canonical_rel or rel_text,
            'description': '',
            'is_deprecated': 0,
            'created_by': int(user_id) if user_id else None,
        }
        aid = int(self._insert_image_asset_dynamic(conn, cur, rec) or 0)
        if not aid:
            return 0, {}
        row = self._find_image_asset_row_by_rel_path(
            cur, canonical_rel or rel_text, join_type, has_tid, has_dep, has_desc
        )
        return aid, row

    def _channel_asset_payload(self, row):
        if not row or not row.get('id'):
            return None
        storage_path = (row.get('storage_path') or '').strip()
        rel_bytes = self._safe_fsencode(storage_path) if storage_path else b''
        path_b64 = base64.b64encode(rel_bytes).decode('ascii') if rel_bytes else ''
        return {
            'image_asset_id': self._parse_int(row.get('id')) or 0,
            'storage_path': storage_path,
            'path_b64': path_b64,
            'image_type_name': (row.get('image_type_name') or '').strip(),
        }

    def _get_channel_link_for_member_asset(self, conn, member_asset_id):
        if not self._channel_links_table_ready(conn):
            return None
        mid = int(member_asset_id or 0)
        if mid <= 0:
            return None
        has_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
        join_type = "LEFT JOIN image_types it ON it.id = ia.image_type_id" if has_tid else ""
        type_sel = 'it.name AS image_type_name' if has_tid else "'' AS image_type_name"
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT ia.id, ia.storage_path, {type_sel}
                FROM image_asset_channel_links l
                INNER JOIN image_assets ia ON ia.id = l.channel_asset_id
                {join_type}
                WHERE l.member_asset_id=%s
                LIMIT 1
                """,
                (mid,),
            )
            row = cur.fetchone() or {}
        if not row.get('id'):
            return None
        return self._channel_asset_payload(row)

    def _set_channel_link_for_member_asset(self, conn, member_asset_id, channel_asset_id, user_id=0):
        if not self._channel_links_table_ready(conn):
            raise RuntimeError('缺少 image_asset_channel_links 表，请先执行 scripts/sql/20260616_01_image_asset_channel_links.sql')
        mid = int(member_asset_id or 0)
        cid = int(channel_asset_id or 0)
        if mid <= 0:
            raise ValueError('无效的图片')
        if mid == cid:
            raise ValueError('通道图不能与当前图片相同')
        with conn.cursor() as cur:
            cur.execute("DELETE FROM image_asset_channel_links WHERE member_asset_id=%s", (mid,))
            if cid > 0:
                cur.execute(
                    "INSERT INTO image_asset_channel_links (member_asset_id, channel_asset_id) VALUES (%s, %s)",
                    (mid, cid),
                )
                if self._table_has_column(conn, 'image_assets', 'image_type_id'):
                    tid = self._get_image_type_id_by_name(conn, '通道图')
                    if tid:
                        cur.execute("UPDATE image_assets SET image_type_id=%s WHERE id=%s", (int(tid), cid))

    def handle_gallery_image_channel_api(self, environ, method, start_response):
        """GET/PUT 当前图片（member）与通道图（channel）的关联。"""
        try:
            if method not in ('GET', 'PUT'):
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            user_id = self._parse_int(self._get_session_user(environ)) or 0
            data = self._read_json_body(environ) if method == 'PUT' else {}
            query_params = parse_qs(environ.get('QUERY_STRING', '') or '')
            path_b64 = str(query_params.get('id', [''])[0] or '').strip()
            if method == 'PUT':
                path_b64 = str((data or {}).get('member_path_b64') or (data or {}).get('id') or path_b64).strip()
            if not path_b64:
                return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
            try:
                raw = base64.b64decode(path_b64)
            except Exception:
                return self.send_json({'status': 'error', 'message': 'Invalid id'}, start_response)
            rel_text = ''
            try:
                rel_text = os.fsdecode(raw)
            except Exception:
                try:
                    rel_text = raw.decode('utf-8', errors='surrogateescape')
                except Exception:
                    rel_text = ''
            rel_text = (rel_text or '').strip().replace('\\', '/').lstrip('/')
            if not rel_text or '..' in rel_text:
                return self.send_json({'status': 'error', 'message': 'Invalid path'}, start_response)

            with self._get_db_connection() as conn:
                if not self._channel_links_table_ready(conn):
                    return self.send_json({
                        'status': 'error',
                        'message': '缺少 image_asset_channel_links 表，请先执行 scripts/sql/20260616_01_image_asset_channel_links.sql',
                    }, start_response)
                with conn.cursor() as cur:
                    member_id, _member_row = self._ensure_image_asset_from_rel_path(conn, cur, rel_text, user_id=user_id)
                    if not member_id:
                        return self.send_json({'status': 'error', 'message': '图片未入库且源文件不存在'}, start_response)

                    if method == 'GET':
                        channel = self._get_channel_link_for_member_asset(conn, member_id)
                        return self.send_json({
                            'status': 'success',
                            'linked': bool(channel),
                            'member_asset_id': member_id,
                            'channel': channel,
                        }, start_response)

                    channel_path_b64 = (data or {}).get('channel_path_b64')
                    if channel_path_b64 is None:
                        channel_path_b64 = (data or {}).get('channel_id')
                    clear_link = channel_path_b64 is None or str(channel_path_b64).strip() == ''
                    if clear_link:
                        self._set_channel_link_for_member_asset(conn, member_id, 0, user_id=user_id)
                        return self.send_json({
                            'status': 'success',
                            'linked': False,
                            'member_asset_id': member_id,
                            'channel': None,
                        }, start_response)

                    try:
                        ch_raw = base64.b64decode(str(channel_path_b64).strip())
                    except Exception:
                        return self.send_json({'status': 'error', 'message': 'Invalid channel path'}, start_response)
                    try:
                        ch_rel = os.fsdecode(ch_raw)
                    except Exception:
                        ch_rel = ch_raw.decode('utf-8', errors='surrogateescape')
                    ch_rel = (ch_rel or '').strip().replace('\\', '/').lstrip('/')
                    if not ch_rel or '..' in ch_rel:
                        return self.send_json({'status': 'error', 'message': 'Invalid channel path'}, start_response)

                    channel_id, ch_row = self._ensure_image_asset_from_rel_path(conn, cur, ch_rel, user_id=user_id)
                    if not channel_id:
                        return self.send_json({'status': 'error', 'message': '通道图未入库且源文件不存在'}, start_response)
                    self._set_channel_link_for_member_asset(conn, member_id, channel_id, user_id=user_id)
                    channel = self._channel_asset_payload(ch_row)
                    return self.send_json({
                        'status': 'success',
                        'linked': True,
                        'member_asset_id': member_id,
                        'channel': channel,
                    }, start_response)
        except ValueError as ex:
            return self.send_json({'status': 'error', 'message': str(ex)}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_gallery_channel_image_upload_api(self, environ, start_response):
        """上传通道图到 货号/主图/通道 并关联到当前 member 图。"""
        try:
            if environ.get('REQUEST_METHOD') != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            user_id = self._parse_int(self._get_session_user(environ)) or 0
            if not user_id:
                return self.send_json({'status': 'error', 'message': '必须登录才能上传'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)

            member_path_b64 = str(form.getfirst('member_path_b64', '') or form.getfirst('id', '') or '').strip()
            if not member_path_b64:
                return self.send_json({'status': 'error', 'message': 'Missing member_path_b64'}, start_response)

            try:
                member_rel = os.fsdecode(base64.b64decode(member_path_b64))
            except Exception:
                return self.send_json({'status': 'error', 'message': 'Invalid member path'}, start_response)
            member_rel = (member_rel or '').strip().replace('\\', '/').lstrip('/')
            sku_family = self._parse_sku_family_from_storage_path(member_rel)
            if not sku_family:
                return self.send_json({'status': 'error', 'message': '无法从当前图片路径解析货号，请使用 NAS 选择已有通道图'}, start_response)

            uploads = []
            for p in getattr(form, 'list', []) or []:
                if getattr(p, 'filename', None):
                    try:
                        content = p.file.read() or b''
                    except Exception:
                        content = b''
                    uploads.append({'filename': p.filename, 'content': content})
            if not uploads:
                file_part = form['file'] if 'file' in form else None
                if file_part and getattr(file_part, 'filename', None):
                    try:
                        content = file_part.file.read() or b''
                    except Exception:
                        content = b''
                    uploads.append({'filename': file_part.filename, 'content': content})
            if not uploads:
                return self.send_json({'status': 'error', 'message': 'No valid images uploaded'}, start_response)

            item = uploads[0]
            content = item.get('content') or b''
            if not content:
                return self.send_json({'status': 'error', 'message': '空文件'}, start_response)

            folder = self._ensure_listing_sales_channel_folder(sku_family)
            if not folder:
                return self.send_json({'status': 'error', 'message': '无法创建通道图目录'}, start_response)

            orig_filename = os.path.basename(str(item.get('filename') or 'channel.jpg'))
            ext = os.path.splitext(orig_filename)[1].lower()
            if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.tif', '.tiff'):
                ext = self._guess_image_ext(orig_filename, content)
            code = self._sanitize_filename_component(sku_family, 32) or 'CH'
            seq = int(time.time() * 1000) % 1000000
            final_name = f'{code}-通道-{seq:06d}{ext}'
            target_abs = os.path.join(folder, self._safe_fsencode(final_name))
            with open(target_abs, 'wb') as f:
                f.write(content)

            storage_path = self._storage_path_from_abs(target_abs)
            sha256 = self._sha256_hex(content)

            with self._get_db_connection() as conn:
                if not self._channel_links_table_ready(conn):
                    return self.send_json({
                        'status': 'error',
                        'message': '缺少 image_asset_channel_links 表，请先执行 scripts/sql/20260616_01_image_asset_channel_links.sql',
                    }, start_response)
                with conn.cursor() as cur:
                    member_id, _ = self._ensure_image_asset_from_rel_path(conn, cur, member_rel, user_id=user_id)
                    if not member_id:
                        return self.send_json({'status': 'error', 'message': '当前图片未入库'}, start_response)

                    existing = self._find_image_asset_by_sha256(conn, sha256)
                    if existing and self._parse_int(existing.get('id')):
                        channel_id = int(existing.get('id'))
                        if self._table_has_column(conn, 'image_assets', 'image_type_id'):
                            tid = self._get_image_type_id_by_name(conn, '通道图')
                            if tid:
                                cur.execute("UPDATE image_assets SET image_type_id=%s WHERE id=%s", (int(tid), channel_id))
                    else:
                        rec = {
                            'sha256': sha256,
                            'storage_path': storage_path,
                            'description': '',
                            'is_deprecated': 0,
                            'created_by': user_id,
                        }
                        if self._table_has_column(conn, 'image_assets', 'image_type_id'):
                            tid = self._get_image_type_id_by_name(conn, '通道图')
                            if tid:
                                rec['image_type_id'] = int(tid)
                        channel_id = int(self._insert_image_asset_dynamic(conn, cur, rec) or 0)

                    if not channel_id:
                        return self.send_json({'status': 'error', 'message': '通道图入库失败'}, start_response)

                    self._set_channel_link_for_member_asset(conn, member_id, channel_id, user_id=user_id)
                    has_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
                    join_type = "LEFT JOIN image_types it ON it.id = ia.image_type_id" if has_tid else ""
                    ch_row = self._find_image_asset_row_by_rel_path(
                        cur, storage_path, join_type, has_tid, False, False
                    )

            channel = self._channel_asset_payload(ch_row)
            return self.send_json({
                'status': 'success',
                'linked': True,
                'member_asset_id': member_id,
                'channel': channel,
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _ensure_listing_sales_global_common_folder(self):
        """Ensure 『通用图片』/主图 exists (与各货号文件夹同层). Return absolute folder path (bytes)."""
        base_folder = self._ensure_listing_folder()
        common_root = os.path.join(base_folder, self._safe_fsencode('『通用图片』'))
        if not os.path.exists(common_root):
            os.makedirs(common_root, exist_ok=True)
        main_folder = os.path.join(common_root, self._safe_fsencode('主图'))
        if not os.path.exists(main_folder):
            os.makedirs(main_folder, exist_ok=True)
        return main_folder

    def _detect_cross_sku_family_by_variant_ids(self, conn, variant_ids):
        """Return sorted distinct sku_family list for given variant_ids."""
        vids = [int(v or 0) for v in (variant_ids or []) if int(v or 0) > 0]
        vids = sorted(set(vids))
        if not vids:
            return []
        placeholders = ','.join(['%s'] * len(vids))
        try:
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT DISTINCT pf.sku_family
                    FROM sales_product_variants v
                    LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                    WHERE v.id IN ({placeholders})
                    """,
                    tuple(vids),
                )
                sku_list = [str(r.get('sku_family') or '').strip() for r in (cur.fetchall() or [])]
            sku_list = [s for s in sku_list if s]
            sku_list = sorted(set(sku_list))
            return sku_list
        except Exception:
            return []

    def _get_asset_referenced_variant_ids(self, conn, asset_id):
        """Return distinct variant_ids already referencing this asset (mapping.variant_id + legacy sales_product_id)."""
        aid = int(asset_id or 0)
        if aid <= 0:
            return []
        has_vid = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
        has_spid = self._table_has_column(conn, 'sales_variant_image_mappings', 'sales_product_id')
        if not has_vid and not has_spid:
            return []
        vids = set()
        try:
            with conn.cursor() as cur:
                if has_vid:
                    cur.execute(
                        "SELECT DISTINCT variant_id FROM sales_variant_image_mappings WHERE image_asset_id=%s AND variant_id IS NOT NULL AND variant_id>0",
                        (aid,),
                    )
                    for r in (cur.fetchall() or []):
                        v = self._parse_int(r.get('variant_id')) or 0
                        if v > 0:
                            vids.add(v)
                if has_spid:
                    if has_vid:
                        cur.execute(
                            """
                            SELECT DISTINCT sp.variant_id AS vid
                            FROM sales_variant_image_mappings sim
                            JOIN sales_products sp ON sp.id = sim.sales_product_id
                            WHERE sim.image_asset_id=%s
                              AND sim.sales_product_id IS NOT NULL AND sim.sales_product_id > 0
                              AND sp.variant_id IS NOT NULL AND sp.variant_id > 0
                              AND (sim.variant_id IS NULL OR sim.variant_id = 0)
                            """,
                            (aid,),
                        )
                    else:
                        cur.execute(
                            """
                            SELECT DISTINCT sp.variant_id AS vid
                            FROM sales_variant_image_mappings sim
                            JOIN sales_products sp ON sp.id = sim.sales_product_id
                            WHERE sim.image_asset_id=%s
                              AND sim.sales_product_id IS NOT NULL AND sim.sales_product_id > 0
                              AND sp.variant_id IS NOT NULL AND sp.variant_id > 0
                            """,
                            (aid,),
                        )
                    for r in (cur.fetchall() or []):
                        v = self._parse_int(r.get('vid')) or 0
                        if v > 0:
                            vids.add(v)
            return sorted(vids)
        except Exception:
            return []

    def _get_asset_referenced_fabric_ids(self, conn, asset_id):
        """Return distinct fabric_ids already referencing this asset."""
        aid = int(asset_id or 0)
        if aid <= 0:
            return []
        if not self._has_required_tables(['fabric_image_mappings']):
            return []
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT DISTINCT fabric_id FROM fabric_image_mappings WHERE image_asset_id=%s AND fabric_id IS NOT NULL AND fabric_id>0",
                    (aid,),
                )
                fids = [self._parse_int(r.get('fabric_id')) or 0 for r in (cur.fetchall() or [])]
            fids = [f for f in fids if f > 0]
            return sorted(set(fids))
        except Exception:
            return []

    def _get_asset_referenced_order_product_ids(self, conn, asset_id):
        """Return distinct order_product_ids already referencing this asset."""
        aid = int(asset_id or 0)
        if aid <= 0:
            return []
        if not self._has_required_tables(['order_product_image_mappings']):
            return []
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT DISTINCT order_product_id FROM order_product_image_mappings WHERE image_asset_id=%s AND order_product_id IS NOT NULL AND order_product_id>0",
                    (aid,),
                )
                ids = [self._parse_int(r.get('order_product_id')) or 0 for r in (cur.fetchall() or [])]
            ids = [v for v in ids if v > 0]
            return sorted(set(ids))
        except Exception:
            return []

    def _detect_cross_sku_family_by_order_product_ids(self, conn, order_product_ids):
        """Return sorted distinct sku_family list for given order_product_ids."""
        ids = [int(v or 0) for v in (order_product_ids or []) if int(v or 0) > 0]
        ids = sorted(set(ids))
        if not ids:
            return []
        placeholders = ','.join(['%s'] * len(ids))
        try:
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT DISTINCT pf.sku_family
                    FROM order_products op
                    LEFT JOIN product_families pf ON pf.id = op.sku_family_id
                    WHERE op.id IN ({placeholders})
                    """,
                    tuple(ids),
                )
                sku_list = [str(r.get('sku_family') or '').strip() for r in (cur.fetchall() or [])]
            sku_list = [s for s in sku_list if s]
            sku_list = sorted(set(sku_list))
            return sku_list
        except Exception:
            return []

    def _read_image_type_scope_flags(self, conn, image_type_id):
        tid = int(image_type_id or 0)
        if not tid:
            return {'applies_fabric': True, 'applies_sales': True, 'applies_order_product': True}
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT
                        COALESCE(applies_fabric, 1) AS applies_fabric,
                        COALESCE(applies_sales, 1) AS applies_sales,
                        COALESCE(applies_order_product, 1) AS applies_order_product
                    FROM image_types
                    WHERE id=%s
                    LIMIT 1
                    """,
                    (tid,),
                )
                row = cur.fetchone() or {}
            if not row:
                return {'applies_fabric': True, 'applies_sales': True, 'applies_order_product': True}
            return {
                'applies_fabric': bool(int(row.get('applies_fabric') or 0)),
                'applies_sales': bool(int(row.get('applies_sales') or 0)),
                'applies_order_product': bool(int(row.get('applies_order_product') or 0)),
            }
        except Exception:
            return {'applies_fabric': True, 'applies_sales': True, 'applies_order_product': True}

    def _gallery_validate_type_for_link_targets(self, conn, image_type_id, want_fabric, want_variant, want_op):
        flags = self._read_image_type_scope_flags(conn, image_type_id)
        if want_fabric and not flags['applies_fabric']:
            return False, '当前图片类型不允许关联面料'
        if want_variant and not flags['applies_sales']:
            return False, '当前图片类型不允许关联销售规格'
        if want_op and not flags['applies_order_product']:
            return False, '当前图片类型不允许关联下单产品'
        return True, ''

    def _gallery_validate_asset_type_compatible(self, conn, asset_id, image_type_id):
        """Ensure existing DB mappings are allowed for the chosen image type (e.g. after type change)."""
        aid = int(asset_id or 0)
        tid = int(image_type_id or 0)
        if aid <= 0 or tid <= 0:
            return True, ''
        flags = self._read_image_type_scope_flags(conn, tid)
        if not flags['applies_fabric']:
            try:
                if self._has_required_tables(['fabric_image_mappings']):
                    with conn.cursor() as cur:
                        cur.execute("SELECT COUNT(1) AS c FROM fabric_image_mappings WHERE image_asset_id=%s", (aid,))
                        c = self._parse_int((cur.fetchone() or {}).get('c')) or 0
                    if c > 0:
                        return False, '该图片已关联面料，但所选类型不适用面料'
            except Exception:
                pass
        if not flags['applies_sales']:
            try:
                vids = self._get_asset_referenced_variant_ids(conn, aid) or []
                if vids:
                    return False, '该图片已关联销售规格，但所选类型不适用规格主图'
            except Exception:
                pass
        if not flags['applies_order_product']:
            try:
                opids = self._get_asset_referenced_order_product_ids(conn, aid) or []
                if opids:
                    return False, '该图片已关联下单产品，但所选类型不适用下单产品主图'
            except Exception:
                pass
        return True, ''

    def _gallery_count_aplus_asset_refs(self, conn, asset_id):
        aid = int(asset_id or 0)
        if aid <= 0 or not self._has_required_tables(['aplus_version_assets']):
            return 0
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(1) AS c FROM aplus_version_assets WHERE image_asset_id=%s", (aid,))
                return self._parse_int((cur.fetchone() or {}).get('c')) or 0
        except Exception:
            return 0

    def _gallery_resolve_new_asset_folder(self, conn, fabric_ids, variant_ids, order_product_ids):
        """Physical folder for a brand-new image_assets row (priority: 面料 > 规格主图 > 配件图). Returns absolute path (bytes) or None."""
        fids = sorted(set(int(x or 0) for x in (fabric_ids or []) if int(x or 0) > 0))
        vids = sorted(set(int(x or 0) for x in (variant_ids or []) if int(x or 0) > 0))
        opids = sorted(set(int(x or 0) for x in (order_product_ids or []) if int(x or 0) > 0))
        if fids:
            return self._ensure_fabric_folder()
        if vids:
            sku_families = self._detect_cross_sku_family_by_variant_ids(conn, vids)
            is_cross_sku = len(sku_families) > 1
            is_multi_variant_same_sku = (not is_cross_sku) and (len(vids) > 1) and (len(sku_families) == 1)
            if is_cross_sku:
                return self._ensure_listing_sales_global_common_folder()
            if is_multi_variant_same_sku:
                return self._ensure_listing_sales_common_folder(sku_families[0])
            folder_info = self._resolve_sales_variant_folder_by_variant_id(vids[0], ensure_folder=True)
            return folder_info.get('folder_path')
        if opids:
            sku_families = self._detect_cross_sku_family_by_order_product_ids(conn, opids) or []
            sku_families = [s for s in sku_families if s]
            if len(opids) > 1:
                if len(set(sku_families)) > 1:
                    return self._ensure_listing_sales_global_common_folder()
                if sku_families:
                    return self._ensure_order_product_common_folder(sku_families[0])
            info = self._resolve_order_product_main_image_folder(opids[0], ensure_folder=True)
            return info.get('folder_path')
        return None

    def _choose_rehome_target(self, conn, asset_id):
        """
        Decide where an asset should live based on references:
        - If referenced by any fabric -> 『面料』/
        - Else if referenced by both order_product and sales variants:
          - If linked to exactly 1 variant -> <货号>/主图/<规格-面料>/
          - If linked to 2+ variants:
            - If variants span multiple sku_families -> 『通用图片』/主图/
            - Else -> <货号>/主图/通用/
        - Else if referenced by multiple variants:
          - If variants span multiple sku_families -> 『通用图片』/主图/
          - Else -> <货号>/主图/通用/
        - Else if referenced by exactly one sales variant (无面料引用、且不与下单主图混绑的纯规格路径):
          -> <货号>/主图/<规格-面料>/（用于从「通用」迁回专属子文件夹）
        - Else keep as-is（例如仅下单产品、或路径已由其它流程维护）
        Returns absolute folder path (bytes) or None.
        """
        aid = int(asset_id or 0)
        if aid <= 0:
            return None

        fabric_ref = 0
        variant_ids = []
        order_product_ids = []
        try:
            with conn.cursor() as cur:
                if self._has_required_tables(['fabric_image_mappings']):
                    cur.execute("SELECT COUNT(*) AS cnt FROM fabric_image_mappings WHERE image_asset_id=%s", (aid,))
                    fabric_ref = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
        except Exception:
            fabric_ref = 0

        try:
            variant_ids = self._get_asset_referenced_variant_ids(conn, aid) or []
        except Exception:
            variant_ids = []

        try:
            order_product_ids = self._get_asset_referenced_order_product_ids(conn, aid) or []
        except Exception:
            order_product_ids = []

        if fabric_ref > 0:
            return self._join_resources('『面料』')

        # If asset is shared between order_product and sales variants: always prefer main image folders.
        if order_product_ids and variant_ids:
            vids = sorted(set(int(v or 0) for v in variant_ids if int(v or 0) > 0))
            if len(vids) == 1:
                try:
                    info = self._resolve_sales_variant_folder_by_variant_id(vids[0], ensure_folder=True)
                    folder = info.get('folder_path')
                    return folder if folder else None
                except Exception:
                    return None
            # 2+ variants -> common; cross-sku -> global common
            sku_list = self._detect_cross_sku_family_by_variant_ids(conn, vids) or []
            sku_list = [s for s in sku_list if s]
            if len(set(sku_list)) > 1:
                return self._ensure_listing_sales_global_common_folder()
            if sku_list:
                return self._ensure_listing_sales_common_folder(sku_list[0])
            # Fallback: infer from order_product sku_family
            sku_list2 = self._detect_cross_sku_family_by_order_product_ids(conn, order_product_ids) or []
            sku_list2 = [s for s in sku_list2 if s]
            if len(set(sku_list2)) > 1:
                return self._ensure_listing_sales_global_common_folder()
            if sku_list2:
                return self._ensure_listing_sales_common_folder(sku_list2[0])
            return None

        # Sales variants only: multi-variant -> common
        vids = sorted(set(int(v or 0) for v in variant_ids if int(v or 0) > 0))
        if len(vids) > 1:
            sku_list = self._detect_cross_sku_family_by_variant_ids(conn, vids) or []
            sku_list = [s for s in sku_list if s]
            if len(set(sku_list)) > 1:
                return self._ensure_listing_sales_global_common_folder()
            if sku_list:
                return self._ensure_listing_sales_common_folder(sku_list[0])

        # 仅绑定一个销售规格：从「主图/通用」或「通用图片/主图」迁回「主图/规格-面料」
        if len(vids) == 1:
            try:
                info = self._resolve_sales_variant_folder_by_variant_id(vids[0], ensure_folder=True)
                folder = info.get('folder_path')
                return folder if folder else None
            except Exception:
                return None

        # Order products only: keep accessory logic (multiple -> 配件图/通用 under that sku_family when possible)
        if order_product_ids and (not vids):
            if len(set(int(x or 0) for x in order_product_ids if int(x or 0) > 0)) > 1:
                sku_list2 = self._detect_cross_sku_family_by_order_product_ids(conn, order_product_ids) or []
                sku_list2 = [s for s in sku_list2 if s]
                if sku_list2:
                    return self._ensure_order_product_common_folder(sku_list2[0])
        return None

    def _rehome_image_asset_if_needed(self, conn, asset_id):
        """
        Move the physical file to its target folder based on references, and update image_assets.storage_path.
        Best-effort: if move fails, keep current path.
        """
        aid = int(asset_id or 0)
        if aid <= 0:
            return None

        with conn.cursor() as cur:
            cur.execute("SELECT id, storage_path FROM image_assets WHERE id=%s LIMIT 1", (aid,))
            asset = cur.fetchone() or {}
        storage_path = (asset.get('storage_path') or '').strip()
        if not storage_path:
            return None
        src_abs = self._abs_from_storage_path(storage_path)
        if not os.path.exists(src_abs):
            return None

        target_folder = self._choose_rehome_target(conn, aid)
        if not target_folder:
            return None
        src_dir = os.path.dirname(src_abs)
        # 已在目标目录：必须用 bytes/混排安全比较；原 fsdecode 抛错时会被吞掉并误走后续逻辑，
        # 可能算出 dst==src，os.replace 自替换导致磁盘文件消失（下单 NAS 双击复用已绑规格图时易触发）。
        try:
            if os.path.isdir(src_dir) and os.path.isdir(target_folder) and os.path.samefile(src_dir, target_folder):
                return None
        except Exception:
            pass
        try:
            sd = self._safe_fsdecode(src_dir).replace('\\', '/').rstrip('/')
            td = self._safe_fsdecode(target_folder).replace('\\', '/').rstrip('/')
            if os.path.normcase(os.path.normpath(sd)) == os.path.normcase(os.path.normpath(td)):
                return None
        except Exception:
            pass

        # If moving into “通用/全局通用”目录：尽量按推荐语法改名
        # 面料（若所有关联规格面料一致）-图片类型-原名
        try:
            base_src_dec = self._safe_fsdecode(os.path.basename(src_abs))
        except Exception:
            base_src_dec = os.path.basename(storage_path)
        ext = os.path.splitext(base_src_dec)[1] or os.path.splitext(storage_path)[1] or '.jpg'
        orig = base_src_dec or os.path.basename(storage_path)
        base_part = self._sanitize_filename_component(os.path.splitext(orig)[0], 120) or f"image_{aid}"

        def _is_common_folder(folder_abs):
            try:
                p = os.fsdecode(folder_abs).replace('\\\\', '/')
            except Exception:
                try:
                    p = str(folder_abs)
                except Exception:
                    p = ''
            return ('/通用' in p) or ('『通用图片』/主图' in p)

        def _get_asset_type_name():
            try:
                if not self._table_has_column(conn, 'image_assets', 'image_type_id'):
                    return ''
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT it.name AS name
                        FROM image_assets ia
                        LEFT JOIN image_types it ON it.id = ia.image_type_id
                        WHERE ia.id=%s
                        LIMIT 1
                        """,
                        (aid,),
                    )
                    row = cur.fetchone() or {}
                return str(row.get('name') or '').strip()
            except Exception:
                return ''

        def _get_common_fabric_name():
            try:
                vids = self._get_asset_referenced_variant_ids(conn, aid) or []
            except Exception:
                vids = []
            vids = [int(v or 0) for v in vids if int(v or 0) > 0]
            if not vids:
                return ''
            placeholders = ','.join(['%s'] * len(vids))
            try:
                has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                fabric_join = "LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if has_fabric_id else ""
                if has_fabric_id and has_fabric_text:
                    fabric_expr = "COALESCE(NULLIF(TRIM(fm.fabric_name_en),''), NULLIF(TRIM(fm.fabric_code),''), NULLIF(TRIM(v.fabric),'')) AS fabric"
                elif has_fabric_id:
                    fabric_expr = "COALESCE(NULLIF(TRIM(fm.fabric_name_en),''), NULLIF(TRIM(fm.fabric_code),'')) AS fabric"
                elif has_fabric_text:
                    fabric_expr = "NULLIF(TRIM(v.fabric),'') AS fabric"
                else:
                    fabric_expr = "'' AS fabric"
                with conn.cursor() as cur:
                    cur.execute(
                        f"""
                        SELECT DISTINCT {fabric_expr}
                        FROM sales_product_variants v
                        {fabric_join}
                        WHERE v.id IN ({placeholders})
                        """,
                        tuple(vids),
                    )
                    vals = [str(r.get('fabric') or '').strip() for r in (cur.fetchall() or [])]
                vals = [v for v in vals if v]
                if not vals:
                    return ''
                first = vals[0]
                return first if all(v == first for v in vals) else ''
            except Exception:
                return ''

        if _is_common_folder(target_folder):
            type_name = _get_asset_type_name()
            type_part = self._sanitize_filename_component(type_name, 32) if type_name else ''
            fabric_name = _get_common_fabric_name()
            fabric_part = self._sanitize_filename_component(fabric_name, 40) if fabric_name else ''
            if type_part:
                if fabric_part and base_part.startswith(f"{fabric_part}-{type_part}-") and len(base_part) > len(fabric_part) + len(type_part) + 2:
                    pass  # already conforms
                elif (not fabric_part) and base_part.startswith(f"{type_part}-") and len(base_part) > len(type_part) + 1:
                    pass
                else:
                    # If base already starts with "类型-" and we now have a fabric, avoid double "类型-类型-"
                    rest = base_part
                    if base_part.startswith(f"{type_part}-"):
                        rest = base_part[len(type_part) + 1:] or base_part
                    if fabric_part:
                        base_part = f"{fabric_part}-{type_part}-{rest}"
                    else:
                        base_part = f"{type_part}-{rest}"

        filename = f"{base_part}{ext}"
        final_name = self._next_available_filename(target_folder, filename)
        dst_abs = os.path.join(target_folder, self._safe_fsencode(final_name))
        if self._listing_paths_equivalent(src_abs, dst_abs):
            return None
        try:
            os.makedirs(os.path.dirname(dst_abs), exist_ok=True)
        except Exception:
            pass

        # Prefer rename/move within same filesystem; then shutil.move (may copy across devices);
        # last resort copy bytes and then remove the old file to avoid duplicate磁盘占用.
        moved = False
        try:
            os.replace(src_abs, dst_abs)
            moved = True
        except Exception:
            try:
                if self._listing_paths_equivalent(src_abs, dst_abs):
                    return None
                shutil.move(src_abs, dst_abs)
                moved = True
            except Exception:
                moved = False

        if not moved:
            try:
                if self._listing_paths_equivalent(src_abs, dst_abs):
                    return None
                with open(src_abs, 'rb') as fsrc:
                    data = fsrc.read()
                with open(dst_abs, 'wb') as fdst:
                    fdst.write(data)
            except Exception:
                return None
            # Copy succeeded: DB will point at dst; src must not remain as a second full copy.
            try:
                if os.path.exists(dst_abs) and (not self._listing_paths_equivalent(src_abs, dst_abs)):
                    self._safe_unlink(src_abs)
                    if os.path.exists(src_abs):
                        self._move_file_to_listing_recycle_bin(src_abs, '重复')
            except Exception:
                pass

        new_storage_path = self._storage_path_from_abs(dst_abs)
        with conn.cursor() as cur:
            cur.execute("UPDATE image_assets SET storage_path=%s WHERE id=%s", (new_storage_path, aid))
        return new_storage_path

    def _sanitize_filename_component(self, text, max_len=80):
        s = str(text or '').strip()
        if not s:
            return ''
        # Keep it filesystem-safe across platforms
        s = s.replace('\\', '-').replace('/', '-').replace('\x00', '')
        for ch in ['<', '>', ':', '"', '|', '?', '*']:
            s = s.replace(ch, '-')
        # Collapse whitespace
        s = re.sub(r'\s+', ' ', s).strip()
        if max_len and len(s) > max_len:
            s = s[:max_len].rstrip()
        return s

    def _sales_nas_import_recommended_basename(self, fabric_en, type_name, stem_plain, sha256_hex):
        """
        Align with sales_product_management gallery rename: fabric (if any) - image type - original stem,
        skipping duplicate fabric-type- prefix when the stem already matches.
        """
        want_type = self._sanitize_filename_component(type_name or '', 32).strip() or '图片'
        want_fabric = self._sanitize_filename_component(fabric_en or '', 40).strip()
        base = self._sanitize_filename_component(stem_plain or '', 80).strip()
        fb = (sha256_hex or '')[:12] if sha256_hex else ''
        if not base:
            base = fb or 'img'
        if want_fabric:
            prefix = f"{want_fabric}-{want_type}-"
            if base.startswith(prefix) and len(base) > len(prefix):
                return base
        else:
            prefix = f"{want_type}-"
            if base.startswith(prefix) and len(base) > len(prefix):
                return base
        parts = []
        if want_fabric:
            parts.append(want_fabric)
        parts.append(want_type)
        parts.append(base)
        return '-'.join(parts)

    def _next_available_filename(self, folder_abs, filename):
        """
        Ensure filename is unique inside folder_abs.
        Returns a filename (string) without path.
        """
        if isinstance(filename, (bytes, bytearray)):
            filename = self._safe_fsdecode(filename)
        base = os.path.basename(str(filename or '').strip())
        if not base:
            base = 'image.jpg'
        name, ext = os.path.splitext(base)
        ext = ext or '.jpg'
        try_names = [base]
        for i in range(2, 1000):
            try_names.append(f"{name}_{i:02d}{ext}")
        for cand in try_names:
            try:
                cand_path = os.path.join(folder_abs, self._safe_fsencode(cand))
            except Exception:
                cand_path = os.path.join(folder_abs, cand.encode('utf-8', errors='surrogatepass'))
            if not os.path.exists(cand_path):
                return cand
        # Fallback
        return f"{name}_{int(datetime.now().timestamp())}{ext}"

    def _try_create_link(self, src_abs, dst_abs):
        """
        Best-effort: create a link/shortcut file in dst_abs pointing to src_abs.
        Prefer hardlink when possible, else symlink. If both fail, do nothing.
        """
        try:
            if os.path.exists(dst_abs):
                return True
        except Exception:
            pass
        # Try hardlink (same filesystem)
        try:
            os.link(src_abs, dst_abs)
            return True
        except Exception:
            pass
        # Try symlink
        try:
            # Make it relative when possible (more portable if folder moved)
            try:
                rel = os.path.relpath(src_abs, os.path.dirname(dst_abs))
            except Exception:
                rel = src_abs
            os.symlink(rel, dst_abs)
            return True
        except Exception:
            return False

    def _tx_begin(self, conn):
        """Begin a transaction even though default is autocommit=True."""
        try:
            conn.autocommit(False)
        except Exception:
            pass
        try:
            conn.begin()
        except Exception:
            # Some drivers start implicitly when autocommit=False
            pass

    def _tx_commit(self, conn):
        try:
            conn.commit()
        except Exception:
            pass
        try:
            conn.autocommit(True)
        except Exception:
            pass

    def _tx_rollback(self, conn):
        try:
            conn.rollback()
        except Exception:
            pass
        try:
            conn.autocommit(True)
        except Exception:
            pass

    def _safe_unlink(self, path_abs):
        try:
            if path_abs and os.path.exists(path_abs):
                os.remove(path_abs)
        except Exception:
            pass

    def _read_wsgi_request_body(self, environ):
        """Read request body bytes as reliably as possible for multipart uploads."""
        try:
            length = int(environ.get('CONTENT_LENGTH', 0) or 0)
        except Exception:
            length = 0
        stream = environ.get('wsgi.input')
        if not stream:
            return b''
        if length > 0:
            try:
                return stream.read(length) or b''
            except Exception:
                return b''
        # Chunked / missing CONTENT_LENGTH: read until EOF (best-effort).
        try:
            chunks = []
            while True:
                chunk = stream.read(1024 * 1024)
                if not chunk:
                    break
                chunks.append(chunk)
            return b''.join(chunks)
        except Exception:
            return b''

    def _parse_multipart_uploads_fallback(self, content_type, raw_body):
        """
        Fallback multipart parser when cgi.FieldStorage fails to enumerate file parts.
        Returns list of dicts: {filename, content}
        """
        def _decode_filename_bytes(b):
            if b is None:
                return ''
            if isinstance(b, str):
                return b.strip()
            if not isinstance(b, (bytes, bytearray)):
                b = str(b).encode('utf-8', errors='surrogatepass')
            b = bytes(b)
            # Prefer UTF-8 (browser standard), then GB18030 (Windows legacy), then latin-1 as last resort.
            for enc in ('utf-8', 'gb18030', 'latin-1'):
                try:
                    return b.decode(enc, errors='surrogateescape').strip()
                except Exception:
                    continue
            try:
                return b.decode('utf-8', errors='replace').strip()
            except Exception:
                return ''

        uploads = []
        if not raw_body:
            return uploads
        ct = (content_type or '').lower()
        if 'multipart/form-data' not in ct:
            return uploads
        boundary = None
        for part in (content_type or '').split(';'):
            part = part.strip()
            if part.lower().startswith('boundary='):
                boundary = part.split('=', 1)[1].strip().strip('"')
                break
        if not boundary:
            return uploads

        # RFC2046: boundary lines are prefixed with "--"
        delim = (b'--' + boundary.encode('utf-8', errors='ignore'))
        segments = raw_body.split(delim)
        for seg in segments:
            seg = seg.strip(b'\r\n')
            if not seg:
                continue
            if seg == b'--':
                continue
            header_blob, _, body_blob = seg.partition(b'\r\n\r\n')
            if not header_blob or body_blob is None:
                continue
            # Parse Content-Disposition from raw header bytes to avoid lossy decode ('���')
            filename = ''
            try:
                header_lower = header_blob.lower()
            except Exception:
                header_lower = header_blob

            if b'content-disposition:' not in header_lower or b'form-data' not in header_lower:
                continue

            # Prefer RFC 5987 / RFC 2231 filename*=UTF-8''...
            from urllib.parse import unquote_to_bytes
            fn_bytes = b''
            try:
                idx = header_lower.find(b'filename*=')
                if idx >= 0:
                    tail = header_blob[idx + len(b'filename*='):]
                    tail = tail.split(b'\r', 1)[0]
                    tail = tail.split(b'\n', 1)[0]
                    tail = tail.split(b';', 1)[0].strip()
                    tail = tail.strip().strip(b'"')
                    if tail.lower().startswith(b"utf-8''"):
                        pct = tail[7:]
                        fn_bytes = unquote_to_bytes(pct)
            except Exception:
                fn_bytes = b''

            # Fallback: filename="..."
            if not fn_bytes:
                try:
                    idx2 = header_lower.find(b'filename=')
                    if idx2 >= 0:
                        tail2 = header_blob[idx2 + len(b'filename='):]
                        tail2 = tail2.split(b'\r', 1)[0]
                        tail2 = tail2.split(b'\n', 1)[0]
                        tail2 = tail2.split(b';', 1)[0].strip()
                        tail2 = tail2.strip()
                        if tail2.startswith(b'"') and b'"' in tail2[1:]:
                            fn_bytes = tail2[1:].split(b'"', 1)[0]
                        else:
                            fn_bytes = tail2.strip().strip(b'"')
                except Exception:
                    fn_bytes = b''

            filename = _decode_filename_bytes(fn_bytes)
            if not filename:
                continue
            content = body_blob
            if content.endswith(b'\r\n'):
                content = content[:-2]
            elif content.endswith(b'\n'):
                content = content[:-1]
            uploads.append({'filename': filename, 'content': content or b''})
        return uploads

    def _table_has_column(self, conn, table_name, column_name):
        cache = getattr(self, '_schema_column_exists_cache', None)
        if cache is None:
            cache = {}
            self._schema_column_exists_cache = cache
        key = (str(table_name), str(column_name))
        if key in cache:
            return cache[key]
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT COUNT(*) AS cnt
                FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE()
                  AND TABLE_NAME = %s
                  AND COLUMN_NAME = %s
                """,
                (table_name, column_name)
            )
            row = cur.fetchone() or {}
        exists = int(row.get('cnt') or 0) > 0
        cache[key] = exists
        return exists

    def _shop_handles_last_mile_factor_sql(self, conn, shop_alias='sh', platform_alias='pt'):
        """SQL 表达式：店铺是否将尾程计入利润/成本（0 或 1）。"""
        if self._table_has_column(conn, 'shops', 'handles_last_mile'):
            return f'COALESCE({shop_alias}.handles_last_mile, 0)'
        name_expr = f"LOWER(COALESCE({platform_alias}.name, ''))"
        return (
            f"CASE WHEN ({name_expr} LIKE '%amazon%' OR {name_expr} LIKE '%亚马逊%') "
            f"THEN 1 ELSE 0 END"
        )

    def _shop_handles_last_mile_select_sql(self, conn, shop_alias='s', platform_alias='pt'):
        factor = self._shop_handles_last_mile_factor_sql(conn, shop_alias, platform_alias)
        return f"({factor}) AS handles_last_mile"

    def _sales_variant_fabric_select_sql(self, conn, v_alias='v'):
        """用于 SELECT：在仅有 fabric_id、仅有 fabric 文本、或两者并存时生成 JOIN 与面料编码表达式。"""
        has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
        has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
        join_sql = f"LEFT JOIN fabric_materials fm ON fm.id = {v_alias}.fabric_id" if has_fabric_id else ""
        if has_fabric_id and has_fabric_text:
            expr = f"COALESCE(fm.fabric_code, {v_alias}.fabric)"
        elif has_fabric_id:
            expr = "COALESCE(NULLIF(TRIM(fm.fabric_code),''), '')"
        elif has_fabric_text:
            expr = f"{v_alias}.fabric"
        else:
            expr = "''"
        return join_sql, expr

    def _normalize_sales_import_fabric_cell(self, value):
        """把 Excel 面料格统一成可与 fabric_code 比较的字符串（数字格、空白字符等）。"""
        if value is None:
            return ''
        if isinstance(value, bool):
            return ''
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            try:
                iv = int(value)
                if value == iv:
                    return str(iv)
            except (ValueError, OverflowError):
                pass
            return str(value).strip()
        if Decimal is not None and isinstance(value, Decimal):
            try:
                if value == value.to_integral_value():
                    return str(int(value))
            except Exception:
                pass
            s = format(value, 'f').rstrip('0').rstrip('.')
            return (s or str(value)).strip()
        s = str(value).replace('\xa0', ' ').replace('\u200b', '').strip()
        return s

    def _fabric_product_family_linked(self, conn, sku_family_id, fabric_material_id, cur=None):
        """货号(product_families.id) 与面料(fabric_materials.id) 是否在 fabric_product_families 中已关联。"""
        if not sku_family_id or not fabric_material_id:
            return False
        run_cur = cur
        close_after = False
        if run_cur is None:
            run_cur = conn.cursor()
            close_after = True
        try:
            run_cur.execute(
                "SELECT 1 FROM fabric_product_families WHERE sku_family_id=%s AND fabric_id=%s LIMIT 1",
                (int(sku_family_id), int(fabric_material_id)),
            )
            return bool(run_cur.fetchone())
        finally:
            if close_after and run_cur is not None:
                run_cur.close()

    def _resolve_fabric_material_id_from_label(self, conn, label, cur=None, *, allow_name_match=True):
        """将 Excel/UI 中的面料标签解析为 fabric_materials.id（无则 None）。

        按 fabric_code 匹配时：先完整格值，再「-」前主码（与平台 SKU 习惯一致），避免主数据编号含「-」时误只查前半段。
        再按需按 fabric_name_en 精确匹配（allow_name_match=True）。
        """
        text = self._normalize_sales_import_fabric_cell(label)
        if not text:
            return None
        prefix = self._code_before_dash(text)
        code_candidates = []
        for cand in (text, prefix):
            if cand and cand not in code_candidates:
                code_candidates.append(cand)

        def _lookup(c):
            for cand in code_candidates:
                c.execute("SELECT id FROM fabric_materials WHERE fabric_code=%s LIMIT 1", (cand,))
                r = c.fetchone() or {}
                fid = self._parse_int(r.get('id')) or None
                if fid:
                    return fid
            if not allow_name_match:
                return None
            c.execute("SELECT id FROM fabric_materials WHERE fabric_name_en=%s LIMIT 1", (text,))
            r = c.fetchone() or {}
            return self._parse_int(r.get('id')) or None

        if cur is not None:
            return _lookup(cur)
        with conn.cursor() as c:
            return _lookup(c)

    def _fabric_code_for_material_id(self, conn, fabric_id, cur=None):
        fid = self._parse_int(fabric_id)
        if not fid:
            return ''

        def _run(c):
            c.execute("SELECT fabric_code FROM fabric_materials WHERE id=%s LIMIT 1", (fid,))
            r = c.fetchone() or {}
            return str(r.get('fabric_code') or '').strip()

        if cur is not None:
            return _run(cur)
        with conn.cursor() as c:
            return _run(c)

    def _resolve_fabric_material_id_from_order_links_import(self, conn, link_entries, cur=None):
        """销售导入专用：仅从关联 order_products.fabric_id 汇总，不做英文名解析。"""
        id_list = [self._parse_int(e.get('order_product_id')) for e in (link_entries or [])]
        id_list = [i for i in id_list if i]
        if not id_list:
            return None, 'missing'
        placeholders = ','.join(['%s'] * len(id_list))
        sql = f"SELECT DISTINCT fabric_id FROM order_products WHERE id IN ({placeholders})"

        def _run(c):
            c.execute(sql, id_list)
            rows = c.fetchall() or []
            fids = []
            for r in rows:
                fid = self._parse_int(r.get('fabric_id'))
                if fid and fid not in fids:
                    fids.append(fid)
            if len(fids) == 1:
                return fids[0], None
            if not fids:
                return None, 'missing'
            return None, 'ambiguous'

        if cur is not None:
            return _run(cur)
        with conn.cursor() as c:
            return _run(c)

    def _sales_product_shop_expr(self, has_shop_col, sales_alias='sp', parent_alias='p'):
        if has_shop_col:
            return f"COALESCE({parent_alias}.shop_id, {sales_alias}.shop_id)"
        return f"{parent_alias}.shop_id"

    def _parse_sales_barcode(self, value):
        text = (str(value) if value is not None else '').strip()
        return text or None

    def _parse_sales_notes(self, value):
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        if len(text) > 512:
            text = text[:512]
        return text

    def _parse_sales_product_link(self, value):
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        if len(text) > 512:
            text = text[:512]
        return text

    def _is_amazon_platform_type_name(self, name):
        text = (name or '').strip().lower()
        return '亚马逊' in text or text == 'amazon' or 'amazon' in text

    def _amazon_product_link_from_child_code(self, child_code):
        code = (str(child_code) if child_code is not None else '').strip()
        if not code:
            return None
        return f'https://www.amazon.com/dp/{code}'

    def _resolve_sales_product_link(self, platform_type_name, child_code, product_link_raw):
        if product_link_raw is not None:
            text = str(product_link_raw).strip()
            if text:
                return self._parse_sales_product_link(text)
        if self._is_amazon_platform_type_name(platform_type_name):
            return self._amazon_product_link_from_child_code(child_code)
        return None

    def _load_shop_platform_type_name(self, cur, shop_id):
        if not shop_id:
            return None
        cur.execute(
            """
            SELECT pt.name AS platform_type_name
            FROM shops s
            LEFT JOIN platform_types pt ON pt.id = s.platform_type_id
            WHERE s.id=%s
            LIMIT 1
            """,
            (int(shop_id),),
        )
        row = cur.fetchone() or {}
        return row.get('platform_type_name')

    def _extend_sales_product_link_write(self, conn, columns, values, product_link):
        if self._table_has_column(conn, 'sales_products', 'product_link'):
            columns.append('product_link')
            values.append(product_link)

    def _sales_product_barcode_select_sql(self, conn, alias='sp'):
        parts = []
        if self._table_has_column(conn, 'sales_products', 'gtin'):
            parts.append(f'{alias}.gtin')
        else:
            parts.append('NULL AS gtin')
        if self._table_has_column(conn, 'sales_products', 'upc'):
            parts.append(f'{alias}.upc')
        else:
            parts.append('NULL AS upc')
        return ', '.join(parts)

    def _extend_sales_product_barcode_write(self, conn, columns, values, gtin, upc):
        if self._table_has_column(conn, 'sales_products', 'gtin'):
            columns.append('gtin')
            values.append(gtin)
        if self._table_has_column(conn, 'sales_products', 'upc'):
            columns.append('upc')
            values.append(upc)

    _SALES_PROMOTION_ACTIVITY_TYPES = frozenset({
        'Coupon', 'Promotion', 'BD', 'Sale', '直降', '普通专享', '大促专享', '多种促销',
    })
    _SALES_DISCOUNT_FORM_TYPES = frozenset({'percent', 'amount'})
    _SALES_DISCOUNT_SEGMENT_TRIGGER_FIELDS = frozenset({
        'promotion_activity_type',
        'discount_form_type',
        'actual_discount_rate',
        'actual_discount_amount_usd',
        'discounted_price_usd',
    })
    _SALES_DISCOUNT_SEGMENT_SNAPSHOT_FIELDS = (
        'promotion_activity_type',
        'discount_form_type',
        'actual_discount_rate',
        'actual_discount_amount_usd',
        'discounted_price_usd',
        'sale_price_usd',
    )

    def _sales_discount_segments_table_ready(self, conn):
        return self._table_has_column(conn, 'sales_product_discount_segments', 'sales_product_id')

    def _sales_discount_snapshot_from_mapping(self, mapping):
        if not isinstance(mapping, dict):
            mapping = {}
        return {
            'promotion_activity_type': self._normalize_sales_promotion_activity_type(
                mapping.get('promotion_activity_type')
            ),
            'discount_form_type': self._normalize_sales_discount_form_type(
                mapping.get('discount_form_type')
            ),
            'actual_discount_rate': self._parse_float(mapping.get('actual_discount_rate')),
            'actual_discount_amount_usd': self._parse_float(mapping.get('actual_discount_amount_usd')),
            'discounted_price_usd': self._parse_float(mapping.get('discounted_price_usd')),
            'sale_price_usd': self._parse_float(mapping.get('sale_price_usd')),
        }

    def _sales_discount_snapshot_is_empty(self, snap):
        if not isinstance(snap, dict):
            return True
        if snap.get('promotion_activity_type'):
            return False
        if snap.get('discount_form_type'):
            return False
        for key in ('actual_discount_rate', 'actual_discount_amount_usd', 'discounted_price_usd'):
            val = self._parse_float(snap.get(key))
            if val is not None:
                return False
        return True

    def _sales_discount_snapshot_compare_key(self, snap, field):
        if field in ('actual_discount_rate', 'actual_discount_amount_usd', 'discounted_price_usd', 'sale_price_usd'):
            val = self._parse_float((snap or {}).get(field))
            if val is None:
                return None
            if field == 'actual_discount_rate':
                return round(float(val), 4)
            return round(float(val), 2)
        text = (snap or {}).get(field)
        if text is None:
            return None
        text = str(text).strip()
        return text or None

    def _sales_discount_snapshots_equal(self, left, right):
        for field in (
            'promotion_activity_type',
            'discount_form_type',
            'actual_discount_rate',
            'actual_discount_amount_usd',
            'discounted_price_usd',
        ):
            if self._sales_discount_snapshot_compare_key(left, field) != self._sales_discount_snapshot_compare_key(right, field):
                return False
        return True

    def _sales_discount_segment_start_date(self, created_at):
        today = datetime.now().date()
        if created_at is None:
            return today
        if isinstance(created_at, datetime):
            return created_at.date()
        if hasattr(created_at, 'year') and hasattr(created_at, 'month') and hasattr(created_at, 'day'):
            try:
                return created_at
            except Exception:
                pass
        text = str(created_at).strip()
        if not text:
            return today
        try:
            return datetime.strptime(text[:10], '%Y-%m-%d').date()
        except Exception:
            return today

    def _sales_discount_segment_close_end_date(self, start_date, change_date):
        start = start_date
        if isinstance(start, datetime):
            start = start.date()
        elif not hasattr(start, 'year'):
            start = self._sales_discount_segment_start_date(start)
        if start < change_date:
            return change_date - timedelta(days=1)
        return change_date

    def _insert_sales_discount_segment(self, cur, conn, sales_product_id, start_date, end_date, snap, user_id=None):
        cols = [
            'sales_product_id', 'start_date', 'end_date',
            'promotion_activity_type', 'discount_form_type',
            'actual_discount_rate', 'actual_discount_amount_usd',
            'discounted_price_usd', 'sale_price_usd',
        ]
        vals = [
            int(sales_product_id),
            start_date,
            end_date,
            snap.get('promotion_activity_type'),
            snap.get('discount_form_type'),
            snap.get('actual_discount_rate'),
            snap.get('actual_discount_amount_usd'),
            snap.get('discounted_price_usd'),
            snap.get('sale_price_usd'),
        ]
        if self._table_has_column(conn, 'sales_product_discount_segments', 'created_by'):
            cols.append('created_by')
            vals.append(self._parse_int(user_id) or None)
        placeholders = ','.join(['%s'] * len(vals))
        cur.execute(
            f"INSERT INTO sales_product_discount_segments ({', '.join(cols)}) VALUES ({placeholders})",
            tuple(vals),
        )

    def _load_sales_discount_snapshots(self, cur, product_ids):
        ids = sorted(set([self._parse_int(x) for x in (product_ids or []) if self._parse_int(x)]))
        if not ids:
            return {}
        placeholders = ','.join(['%s'] * len(ids))
        cur.execute(
            f"""
            SELECT
                sp.id,
                sp.sale_price_usd,
                sp.promotion_activity_type,
                sp.discount_form_type,
                sp.actual_discount_rate,
                sp.actual_discount_amount_usd,
                sp.discounted_price_usd,
                sp.created_at
            FROM sales_products sp
            WHERE sp.id IN ({placeholders})
            """,
            tuple(ids),
        )
        out = {}
        for row in (cur.fetchall() or []):
            pid = self._parse_int(row.get('id'))
            if not pid:
                continue
            out[pid] = {
                'snapshot': self._sales_discount_snapshot_from_mapping(row),
                'created_at': row.get('created_at'),
            }
        return out

    def _record_sales_discount_segment_change(
        self,
        cur,
        conn,
        sales_product_id,
        old_snap,
        new_snap,
        *,
        user_id=None,
        product_created_at=None,
    ):
        if self._sales_discount_snapshots_equal(old_snap, new_snap):
            return False
        pid = self._parse_int(sales_product_id)
        if not pid:
            return False
        today = datetime.now().date()
        old_empty = self._sales_discount_snapshot_is_empty(old_snap)
        new_empty = self._sales_discount_snapshot_is_empty(new_snap)

        cur.execute(
            """
            SELECT id, start_date
            FROM sales_product_discount_segments
            WHERE sales_product_id=%s AND end_date IS NULL
            ORDER BY id DESC
            LIMIT 1
            """,
            (pid,),
        )
        open_row = cur.fetchone() or {}

        if not open_row and not old_empty:
            backfill_start = self._sales_discount_segment_start_date(product_created_at)
            backfill_end = self._sales_discount_segment_close_end_date(backfill_start, today)
            self._insert_sales_discount_segment(
                cur, conn, pid, backfill_start, backfill_end, old_snap, user_id=user_id
            )
        elif open_row.get('id'):
            close_end = self._sales_discount_segment_close_end_date(open_row.get('start_date'), today)
            cur.execute(
                "UPDATE sales_product_discount_segments SET end_date=%s WHERE id=%s",
                (close_end, int(open_row['id'])),
            )

        if not new_empty:
            self._insert_sales_discount_segment(
                cur, conn, pid, today, None, new_snap, user_id=user_id
            )
        return True

    def _apply_sales_discount_segment_updates(self, cur, conn, row_map, touched_map, old_meta_map, user_id=None):
        if not row_map:
            return 0
        recorded = 0
        for item_id, values in row_map.items():
            touched = touched_map.get(item_id) or set()
            if not (set(touched) & self._SALES_DISCOUNT_SEGMENT_TRIGGER_FIELDS):
                continue
            meta = old_meta_map.get(int(item_id)) or {}
            old_snap = meta.get('snapshot') or self._sales_discount_snapshot_from_mapping({})
            new_snap = self._sales_discount_snapshot_from_mapping(values)
            if self._record_sales_discount_segment_change(
                cur,
                conn,
                item_id,
                old_snap,
                new_snap,
                user_id=user_id,
                product_created_at=meta.get('created_at'),
            ):
                recorded += 1
        return recorded

    def _load_sales_discount_segments(self, cur, sales_product_id, limit=200):
        pid = self._parse_int(sales_product_id)
        if not pid:
            return []
        cur.execute(
            """
            SELECT
                id, sales_product_id, start_date, end_date,
                promotion_activity_type, discount_form_type,
                actual_discount_rate, actual_discount_amount_usd,
                discounted_price_usd, sale_price_usd,
                created_at, created_by
            FROM sales_product_discount_segments
            WHERE sales_product_id=%s
            ORDER BY start_date DESC, id DESC
            LIMIT %s
            """,
            (pid, max(1, min(int(limit or 200), 500))),
        )
        return cur.fetchall() or []

    def _sales_product_preview_select_sql(self, conn):
        parts = []
        for col in (
            'promotion_activity_type',
            'discount_form_type',
            'actual_discount_rate',
            'actual_discount_amount_usd',
            'discounted_price_usd',
            'notes',
        ):
            if self._table_has_column(conn, 'sales_products', col):
                parts.append(f'sp.{col}')
            else:
                parts.append(f'NULL AS {col}')
        return ', '.join(parts)

    def _normalize_sales_promotion_activity_type(self, raw):
        text = (raw or '').strip()
        return text if text in self._SALES_PROMOTION_ACTIVITY_TYPES else None

    def _normalize_sales_discount_form_type(self, raw):
        text = (raw or '').strip().lower()
        return text if text in self._SALES_DISCOUNT_FORM_TYPES else None

    _SALES_DISCOUNT_BUNDLE_FIELDS = frozenset({
        'promotion_activity_type',
        'discount_form_type',
        'actual_discount_rate',
        'actual_discount_amount_usd',
        'discounted_price_usd',
    })

    def _sales_discount_bundle_any_touched(self, touched):
        return bool(set(touched or []) & self._SALES_DISCOUNT_BUNDLE_FIELDS)

    def _sales_discount_bundle_expand_touched(self, touched):
        touched_set = set(touched or [])
        if touched_set & self._SALES_DISCOUNT_BUNDLE_FIELDS:
            touched_set |= self._SALES_DISCOUNT_BUNDLE_FIELDS
        return touched_set

    def _normalize_sales_discount_bundle_values(self, values):
        values = dict(values or {})
        promo = self._normalize_sales_promotion_activity_type(values.get('promotion_activity_type'))
        if not promo:
            values['promotion_activity_type'] = None
            values['discount_form_type'] = None
            values['actual_discount_rate'] = None
            values['actual_discount_amount_usd'] = None
            values['discounted_price_usd'] = None
            return values
        form = self._normalize_sales_discount_form_type(values.get('discount_form_type'))
        rate = self._parse_float(values.get('actual_discount_rate'))
        amount = self._parse_float(values.get('actual_discount_amount_usd'))
        price = self._parse_float(values.get('discounted_price_usd'))
        if form == 'amount':
            rate = None
        elif form == 'percent':
            amount = None
        values['promotion_activity_type'] = promo
        values['discount_form_type'] = form
        values['actual_discount_rate'] = rate
        values['actual_discount_amount_usd'] = amount
        values['discounted_price_usd'] = price
        return values

    def _validate_sales_discount_bundle_values(self, values):
        promo = values.get('promotion_activity_type')
        form = values.get('discount_form_type')
        rate = values.get('actual_discount_rate')
        amount = values.get('actual_discount_amount_usd')
        price = values.get('discounted_price_usd')
        has_any = bool(
            promo or form or rate is not None or amount is not None or price is not None
        )
        if not has_any:
            return None
        if not promo:
            return '填写折扣信息时须同时选择活动形式'
        if not form:
            return '填写折扣信息时须同时选择折扣形式'
        if form == 'amount':
            if amount is None:
                return '请填写折扣记录（金额）'
        elif rate is None:
            return '请填写折扣记录（百分比）'
        if price is None or float(price) <= 0:
            return '请填写折后价'
        return None

    _SALES_PREVIEW_FIELD_NAMES = frozenset({
        'child_code',
        'product_link',
        'gtin',
        'upc',
        'sale_price_usd',
        'promotion_activity_type',
        'discount_form_type',
        'actual_discount_rate',
        'actual_discount_amount_usd',
        'discounted_price_usd',
        'notes',
    })

    def _sales_product_preview_col_exists(self, conn):
        """预览批量保存可用列（一次探测，结果走 _table_has_column 缓存）。"""
        exists = {
            'gtin': self._table_has_column(conn, 'sales_products', 'gtin'),
            'upc': self._table_has_column(conn, 'sales_products', 'upc'),
            'product_link': self._table_has_column(conn, 'sales_products', 'product_link'),
        }
        for col in (
            'promotion_activity_type',
            'discount_form_type',
            'actual_discount_rate',
            'actual_discount_amount_usd',
            'discounted_price_usd',
            'notes',
        ):
            exists[col] = self._table_has_column(conn, 'sales_products', col)
        return exists

    def _sales_product_preview_resolve_product_links(self, conn, cur, row_map, touched_map, sp_has_shop_col):
        """子体编号变更时，亚马逊店铺自动补全 product_link。"""
        if not self._table_has_column(conn, 'sales_products', 'product_link'):
            return
        if not row_map:
            return
        shop_expr = self._sales_product_shop_expr(sp_has_shop_col)
        ids = list(row_map.keys())
        placeholders = ','.join(['%s'] * len(ids))
        cur.execute(
            f"""
            SELECT sp.id, pt.name AS platform_type_name
            FROM sales_products sp
            LEFT JOIN sales_parents p ON p.id = sp.parent_id
            LEFT JOIN shops s ON s.id = {shop_expr}
            LEFT JOIN platform_types pt ON pt.id = s.platform_type_id
            WHERE sp.id IN ({placeholders})
            """,
            ids,
        )
        platform_by_id = {
            int(row.get('id') or 0): row.get('platform_type_name')
            for row in (cur.fetchall() or [])
            if row.get('id')
        }
        for item_id, values in row_map.items():
            touched = touched_map.get(item_id) or set()
            platform = platform_by_id.get(int(item_id))
            if 'product_link' in touched:
                values['product_link'] = self._resolve_sales_product_link(
                    platform, values.get('child_code'), values.get('product_link')
                )
            elif 'child_code' in touched:
                values['product_link'] = self._resolve_sales_product_link(
                    platform, values.get('child_code'), None
                )
                touched_map.setdefault(item_id, set()).add('product_link')

    def _sales_product_preview_batch_update(self, cur, row_map, touched_map, col_exists, chunk_size=120):
        """预览行内编辑批量保存：按列 CASE WHEN 合并 UPDATE，减少数据库往返。"""
        if not row_map:
            return 0
        patch_by_id = {}
        for item_id, values in row_map.items():
            touched = touched_map.get(item_id) or set()
            if not touched:
                continue
            patch = {}
            if 'child_code' in touched:
                patch['child_code'] = values.get('child_code')
            if col_exists.get('product_link') and 'product_link' in touched:
                patch['product_link'] = values.get('product_link')
            if col_exists.get('gtin') and 'gtin' in touched:
                patch['gtin'] = values.get('gtin')
            if col_exists.get('upc') and 'upc' in touched:
                patch['upc'] = values.get('upc')
            if 'sale_price_usd' in touched:
                patch['sale_price_usd'] = values.get('sale_price_usd')
            for col in (
                'promotion_activity_type',
                'discount_form_type',
                'actual_discount_rate',
                'actual_discount_amount_usd',
                'discounted_price_usd',
                'notes',
            ):
                if col_exists.get(col) and col in touched:
                    patch[col] = values.get(col)
            if patch:
                patch_by_id[int(item_id)] = patch
        if not patch_by_id:
            return 0

        updated = 0
        all_ids = list(patch_by_id.keys())
        size = max(1, int(chunk_size or 120))
        for offset in range(0, len(all_ids), size):
            chunk_ids = all_ids[offset:offset + size]
            all_columns = set()
            for rid in chunk_ids:
                all_columns.update(patch_by_id[rid].keys())
            set_parts = []
            params = []
            for col in sorted(all_columns):
                when_parts = []
                for rid in chunk_ids:
                    patch = patch_by_id[rid]
                    if col not in patch:
                        continue
                    when_parts.append('WHEN %s THEN %s')
                    params.extend([rid, patch[col]])
                if when_parts:
                    set_parts.append(f'{col} = CASE id {" ".join(when_parts)} ELSE {col} END')
            if not set_parts:
                continue
            in_ph = ','.join(['%s'] * len(chunk_ids))
            sql = f'UPDATE sales_products SET {", ".join(set_parts)} WHERE id IN ({in_ph})'
            cur.execute(sql, tuple(params + chunk_ids))
            updated += int(cur.rowcount or 0)
        return updated

    # -------------------------------------------------------------------------
    # 父体管理 API（CRUD）
    # -------------------------------------------------------------------------

    def handle_parent_api(self, environ, method, start_response):
        """父体管理 API（CRUD）"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))

            def limited_text(value, max_len):
                text = (value or '').strip()
                if not text:
                    return None
                if len(text) > max_len:
                    raise ValueError(f'文本长度超限（>{max_len}）')
                return text

            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                item_id = self._parse_int((query_params.get('id', [''])[0] or '').strip())
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        sql = """
                            SELECT sp.id, sp.parent_code, sp.is_enabled, sp.shop_id, sp.sku_marker,
                                   estimated_refund_rate, estimated_discount_rate,
                                   estimated_acoas,
                                   sp.created_at, sp.updated_at,
                                   s.shop_name, b.name AS brand_name, pt.name AS platform_type_name
                            FROM sales_parents sp
                            LEFT JOIN shops s ON s.id = sp.shop_id
                            LEFT JOIN brands b ON b.id = s.brand_id
                            LEFT JOIN platform_types pt ON pt.id = s.platform_type_id
                        """
                        params = []
                        filters = []
                        if item_id:
                            filters.append("sp.id = %s")
                            params.append(item_id)
                        if keyword:
                            like_kw = f"%{keyword}%"
                            filters.append("(sp.parent_code LIKE %s OR sp.sku_marker LIKE %s)")
                            params.extend([like_kw, like_kw])
                        if filters:
                            sql += " WHERE " + " AND ".join(filters)
                        sql += " ORDER BY sp.id DESC"
                        cur.execute(sql, params)
                        rows = cur.fetchall() or []
                if item_id:
                    return self.send_json({'status': 'success', 'item': rows[0] if rows else None}, start_response)
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                parent_code = (data.get('parent_code') or '').strip()
                if not parent_code:
                    return self.send_json({'status': 'error', 'message': 'Missing parent_code'}, start_response)
                is_enabled_raw = data.get('is_enabled', 1)
                is_enabled = 1 if str(is_enabled_raw).strip().lower() in ('1', 'true', 'yes', 'on') else 0
                shop_id = self._parse_int(data.get('shop_id'))
                try:
                    sku_marker = limited_text(data.get('sku_marker'), 128)
                except ValueError as ve:
                    return self.send_json({'status': 'error', 'message': str(ve)}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO sales_parents
                            (parent_code, is_enabled, shop_id, sku_marker, estimated_refund_rate, estimated_discount_rate, estimated_acoas)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                parent_code,
                                is_enabled,
                                shop_id,
                                sku_marker,
                                self._parse_float(data.get('estimated_refund_rate')),
                                self._parse_float(data.get('estimated_discount_rate')),
                                self._parse_float(data.get('estimated_acoas'))
                            )
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                is_enabled_raw = data.get('is_enabled', 1)
                is_enabled = 1 if str(is_enabled_raw).strip().lower() in ('1', 'true', 'yes', 'on') else 0
                shop_id = self._parse_int(data.get('shop_id'))
                try:
                    sku_marker = limited_text(data.get('sku_marker'), 128)
                except ValueError as ve:
                    return self.send_json({'status': 'error', 'message': str(ve)}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE sales_parents
                            SET parent_code=%s,
                                is_enabled=%s,
                                shop_id=%s,
                                sku_marker=%s,
                                estimated_refund_rate=%s,
                                estimated_discount_rate=%s,
                                estimated_acoas=%s
                            WHERE id=%s
                            """,
                            (
                                (data.get('parent_code') or '').strip(),
                                is_enabled,
                                shop_id,
                                sku_marker,
                                self._parse_float(data.get('estimated_refund_rate')),
                                self._parse_float(data.get('estimated_discount_rate')),
                                self._parse_float(data.get('estimated_acoas')),
                                item_id
                            )
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM sales_parents WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_sales_product_spec_suggest_api(self, environ, method, start_response):
        """GET /api/sales-product-spec-suggest — 当前货号下已有规格名称联想（模糊匹配已输入片段）。"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            family_id = self._parse_int((query_params.get('sku_family_id', [''])[0] or '').strip())
            if not family_id:
                return self.send_json({'status': 'error', 'message': 'Missing sku_family_id'}, start_response)
            q = (query_params.get('q', [''])[0] or '').strip()
            q_lower = q.lower()
            limit = min(80, max(10, self._parse_int((query_params.get('limit', ['60'])[0] or '60')) or 60))
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    if q_lower:
                        cur.execute(
                            """
                            SELECT DISTINCT spec_name
                            FROM sales_product_variants
                            WHERE sku_family_id=%s
                              AND spec_name IS NOT NULL
                              AND TRIM(spec_name) <> ''
                              AND LOCATE(%s, LOWER(spec_name)) > 0
                            ORDER BY spec_name ASC
                            LIMIT %s
                            """,
                            (family_id, q_lower, limit),
                        )
                    else:
                        cur.execute(
                            """
                            SELECT DISTINCT spec_name
                            FROM sales_product_variants
                            WHERE sku_family_id=%s
                              AND spec_name IS NOT NULL
                              AND TRIM(spec_name) <> ''
                            ORDER BY spec_name ASC
                            LIMIT %s
                            """,
                            (family_id, limit),
                        )
                    rows = cur.fetchall() or []
            items = []
            seen = set()
            for r in rows:
                name = str((r.get('spec_name') if isinstance(r, dict) else (r[0] if r else '')) or '').strip()
                if not name or name in seen:
                    continue
                seen.add(name)
                items.append({'spec_name': name})
            return self.send_json({'status': 'success', 'items': items}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_sales_product_variant_prefill_api(self, environ, method, start_response):
        """GET /api/sales-product-variant-prefill — 若货号+规格(+面料)已对应销售变体，返回其关联下单 SKU 行。"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            family_id = self._parse_int((query_params.get('sku_family_id', [''])[0] or '').strip())
            spec = (query_params.get('spec_name', [''])[0] or '').strip()
            fabric_id = self._parse_int((query_params.get('fabric_id', [''])[0] or '').strip())
            fabric_text = (query_params.get('fabric', [''])[0] or '').strip()
            if not family_id or not spec:
                return self.send_json({'status': 'success', 'variant_id': 0, 'ambiguous': False, 'order_sku_links': []}, start_response)

            with self._get_db_connection() as conn:
                has_fid = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                variant_id = 0
                ambiguous = False
                with conn.cursor() as cur:
                    if has_fid and fabric_id > 0:
                        cur.execute(
                            """
                            SELECT id FROM sales_product_variants
                            WHERE sku_family_id=%s AND spec_name=%s AND COALESCE(fabric_id,0)=%s
                            ORDER BY id ASC
                            LIMIT 2
                            """,
                            (family_id, spec, fabric_id),
                        )
                    elif has_fid:
                        cur.execute(
                            """
                            SELECT id FROM sales_product_variants
                            WHERE sku_family_id=%s AND spec_name=%s
                            ORDER BY id ASC
                            LIMIT 10
                            """,
                            (family_id, spec),
                        )
                    elif has_fabric_text:
                        cur.execute(
                            """
                            SELECT id FROM sales_product_variants
                            WHERE sku_family_id=%s AND spec_name=%s AND fabric=%s
                            ORDER BY id ASC
                            LIMIT 2
                            """,
                            (family_id, spec, fabric_text),
                        )
                    else:
                        cur.execute(
                            """
                            SELECT id FROM sales_product_variants
                            WHERE sku_family_id=%s AND spec_name=%s
                            ORDER BY id ASC
                            LIMIT 2
                            """,
                            (family_id, spec),
                        )
                    rows = cur.fetchall() or []
                    ids = []
                    for r in rows:
                        rid = r.get('id') if isinstance(r, dict) else (r[0] if r else 0)
                        vid = self._parse_int(rid)
                        if vid:
                            ids.append(vid)
                    if len(ids) == 1:
                        variant_id = ids[0]
                    elif len(ids) > 1:
                        ambiguous = True

                if not variant_id:
                    return self.send_json({
                        'status': 'success',
                        'variant_id': 0,
                        'ambiguous': ambiguous,
                        'order_sku_links': [],
                    }, start_response)

                metrics = self._load_sales_variant_metrics(conn, [variant_id]) or {}
                bucket = metrics.get(variant_id, {}) if variant_id else {}
                raw_links = bucket.get('order_sku_links') or []
                order_sku_links = []
                for entry in raw_links:
                    oid = self._parse_int(entry.get('order_product_id'))
                    if not oid:
                        continue
                    order_sku_links.append({
                        'order_product_id': oid,
                        'quantity': max(1, self._parse_int(entry.get('quantity')) or 1),
                    })

            return self.send_json({
                'status': 'success',
                'variant_id': variant_id,
                'ambiguous': ambiguous,
                'order_sku_links': order_sku_links,
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _populate_sales_product_template_ref_sheet(self, ref_ws, bricks):
        """写入隐藏表 _refs：全量店铺/父体/货号、规格、面料（各自一列下拉，无级联）。
        返回各列末行号供数据验证公式引用（末行至少为 2）。"""
        shop_options = bricks.get('shop_options') or []
        parent_codes = [str(c or '').strip() for c in (bricks.get('parent_codes') or []) if str(c or '').strip()]
        sku_families = [str(s or '').strip() for s in (bricks.get('sku_families') or []) if str(s or '').strip()]
        all_specs = [str(s or '').strip() for s in (bricks.get('all_specs') or []) if str(s or '').strip()]
        all_fabs = [str(f or '').strip() for f in (bricks.get('all_fabrics') or []) if str(f or '').strip()]
        if not all_specs or not all_fabs:
            spec_set = set(all_specs)
            fab_set = set(all_fabs)
            for tri in bricks.get('triples') or []:
                if not isinstance(tri, (list, tuple)) or len(tri) < 2:
                    continue
                sp = str(tri[1] or '').strip()
                fab = str(tri[2] or '').strip() if len(tri) > 2 else ''
                if sp:
                    spec_set.add(sp)
                if fab:
                    fab_set.add(fab)
            if not all_specs:
                all_specs = sorted(spec_set)
            if not all_fabs:
                all_fabs = sorted(fab_set)

        ref_ws.cell(row=1, column=11, value='sku_family_all')
        ref_ws.cell(row=1, column=13, value='spec_all')
        ref_ws.cell(row=1, column=14, value='fabric_all')
        ref_ws.cell(row=1, column=16, value='shop')
        ref_ws.cell(row=1, column=17, value='parent_code')

        r_shop = 2
        for row in shop_options:
            nm = str(row.get('shop_name') or '').strip()
            if not nm:
                continue
            ref_ws.cell(row=r_shop, column=16, value=nm)
            r_shop += 1
        shop_end = max(2, r_shop - 1)

        r_par = 2
        for code in parent_codes:
            ref_ws.cell(row=r_par, column=17, value=code)
            r_par += 1
        parent_end = max(2, r_par - 1)

        r_k = 2
        for sf in sku_families:
            ref_ws.cell(row=r_k, column=11, value=sf)
            r_k += 1
        sku_end = max(2, r_k - 1)

        r_m = 2
        for sp in sorted(all_specs):
            ref_ws.cell(row=r_m, column=13, value=sp)
            r_m += 1
        spec_all_end = max(2, r_m - 1)
        r_n = 2
        for fb in sorted(all_fabs):
            ref_ws.cell(row=r_n, column=14, value=fb)
            r_n += 1
        fab_all_end = max(2, r_n - 1)

        return {
            'shop_end': shop_end,
            'parent_end': parent_end,
            'sku_end': sku_end,
            'spec_all_end': spec_all_end,
            'fab_all_end': fab_all_end,
        }

    # -------------------------------------------------------------------------
    # 销售产品 Excel 模板下载
    # -------------------------------------------------------------------------

    def handle_sales_product_template_api(self, environ, method, start_response):
        """销售产品模板下载"""
        try:
            if method not in ('GET', 'POST'):
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)
            
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation

            def _append_ids_from_value(container, value):
                if isinstance(value, list):
                    for v in value:
                        _append_ids_from_value(container, v)
                    return
                text = str(value or '').strip()
                if not text:
                    return
                for token in re.split(r'[,，;；\s]+', text):
                    if not token:
                        continue
                    item_id = self._parse_int(token)
                    if item_id and item_id not in container:
                        container.append(item_id)

            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            selected_ids = []
            for raw in query_params.get('ids', []):
                _append_ids_from_value(selected_ids, raw)

            if method == 'POST':
                body = self._read_json_body(environ) or {}
                _append_ids_from_value(selected_ids, body.get('ids'))
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'sales_products'
            ref_ws = wb.create_sheet('_refs')
            ref_ws.sheet_state = 'hidden'

            # 获取可选项：全量店铺/父体/货号 + 全量规格/面料列表，供隐藏表与下拉（无级联）
            with self._get_db_connection() as conn:
                sp_has_shop_col = self._table_has_column(conn, 'sales_products', 'shop_id')
                shop_expr = self._sales_product_shop_expr(sp_has_shop_col, sales_alias='sp', parent_alias='pa')

                def _load_sales_template_workbook_bricks():
                    with conn.cursor() as cur:
                        cur.execute("SELECT id, shop_name FROM shops ORDER BY shop_name")
                        shop_options_local = [row for row in (cur.fetchall() or []) if row.get('shop_name')]
                        cur.execute("SELECT parent_code FROM sales_parents ORDER BY parent_code")
                        parent_codes_local = [
                            str(row['parent_code']).strip()
                            for row in (cur.fetchall() or []) if row.get('parent_code')
                        ]
                        cur.execute(
                            """
                            SELECT sku_family FROM product_families
                            WHERE sku_family IS NOT NULL AND TRIM(sku_family) <> ''
                            ORDER BY sku_family
                            """
                        )
                        sku_families_local = [
                            str(row['sku_family']).strip()
                            for row in (cur.fetchall() or []) if row.get('sku_family')
                        ]
                        has_fid = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                        has_ft = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                        join_fm = "LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if has_fid else ""
                        if has_fid and has_ft:
                            fab_expr = "COALESCE(NULLIF(TRIM(fm.fabric_code),''), NULLIF(TRIM(v.fabric),''), '')"
                        elif has_fid:
                            fab_expr = "COALESCE(NULLIF(TRIM(fm.fabric_code),''), '')"
                        elif has_ft:
                            fab_expr = "COALESCE(NULLIF(TRIM(v.fabric),''), '')"
                        else:
                            fab_expr = "''"
                        cur.execute(
                            """
                            SELECT DISTINCT TRIM(v.spec_name) AS spec_name
                            FROM sales_product_variants v
                            WHERE v.spec_name IS NOT NULL AND TRIM(v.spec_name) <> ''
                            ORDER BY spec_name
                            """
                        )
                        all_specs_local = [
                            str(row['spec_name']).strip()
                            for row in (cur.fetchall() or [])
                            if row.get('spec_name') and str(row['spec_name']).strip()
                        ]
                        if has_fid:
                            cur.execute(
                                """
                                SELECT DISTINCT TRIM(fm.fabric_code) AS fabric
                                FROM fabric_materials fm
                                WHERE fm.fabric_code IS NOT NULL AND TRIM(fm.fabric_code) <> ''
                                ORDER BY fabric
                                """
                            )
                        else:
                            cur.execute(
                                f"""
                                SELECT DISTINCT TRIM({fab_expr}) AS fabric
                                FROM sales_product_variants v
                                {join_fm}
                                WHERE TRIM({fab_expr}) <> ''
                                ORDER BY fabric
                                """
                            )
                        all_fabs_local = [
                            str(row['fabric']).strip()
                            for row in (cur.fetchall() or [])
                            if row.get('fabric') and str(row['fabric']).strip()
                        ]
                    return {
                        'shop_options': shop_options_local,
                        'parent_codes': parent_codes_local,
                        'sku_families': sku_families_local,
                        'all_specs': all_specs_local,
                        'all_fabrics': all_fabs_local,
                    }

                bricks = self._get_cached_template_options(
                    'sales_product_template_workbook_bricks_v2',
                    _load_sales_template_workbook_bricks,
                    ttl_seconds=1800,
                ) or {}
                shop_options = bricks.get('shop_options') or []
                parent_codes = bricks.get('parent_codes') or []
                sku_family_options = bricks.get('sku_families') or []

                export_rows = []
                if selected_ids:
                    placeholders = ','.join(['%s'] * len(selected_ids))
                    with conn.cursor() as cur:
                        cur.execute(
                            f"""
                            SELECT sp.id, sp.product_status, sh.shop_name, pa.parent_code, pa.sku_marker,
                                sp.platform_sku, sp.child_code,
                                {('sp.product_link' if self._table_has_column(conn, 'sales_products', 'product_link') else 'NULL AS product_link')},
                                {self._sales_product_barcode_select_sql(conn, 'sp')},
                                pf.sku_family, v.spec_name,                                 {('COALESCE(fm.fabric_code, v.fabric)' if (self._table_has_column(conn,'sales_product_variants','fabric_id') and self._table_has_column(conn,'sales_product_variants','fabric')) else ('fm.fabric_code' if self._table_has_column(conn,'sales_product_variants','fabric_id') else ('v.fabric' if self._table_has_column(conn,'sales_product_variants','fabric') else "''")))} AS fabric,
                                sp.sale_price_usd,
                                {('sp.notes' if self._table_has_column(conn, 'sales_products', 'notes') else 'NULL AS notes')}
                            FROM sales_products sp
                            LEFT JOIN sales_parents pa ON pa.id = sp.parent_id
                            LEFT JOIN shops sh ON sh.id = {shop_expr}
                            LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                            LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                            {("LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if self._table_has_column(conn,'sales_product_variants','fabric_id') else "")}
                            WHERE sp.id IN ({placeholders})
                            ORDER BY sp.id DESC
                            """,
                            selected_ids
                        )
                        rows = cur.fetchall() or []
                        cur.execute(
                            f"""
                            SELECT sp.id AS sales_product_id, op.sku, l.quantity
                            FROM sales_products sp
                            JOIN sales_variant_order_links l ON l.variant_id = sp.variant_id
                            JOIN order_products op ON op.id = l.order_product_id
                            WHERE sp.id IN ({placeholders})
                            ORDER BY sp.id, op.sku
                            """,
                            selected_ids
                        )
                        link_rows = cur.fetchall() or []
                    link_map = {}
                    for link in link_rows:
                        sp_id = self._parse_int(link.get('sales_product_id')) or 0
                        if not sp_id:
                            continue
                        sku = str(link.get('sku') or '').strip()
                        qty = self._parse_int(link.get('quantity')) or 1
                        if not sku:
                            continue
                        link_map.setdefault(sp_id, []).append(f"{sku}*{qty}")
                    for row in rows:
                        row_id = self._parse_int(row.get('id')) or 0
                        export_rows.append([
                            {'enabled': '启用', 'retained': '留用', 'discarded': '弃用'}.get(str(row.get('product_status') or '').strip(), '启用'),
                            row.get('shop_name') or '',
                            row.get('parent_code') or '',
                            row.get('sku_marker') or '',
                            row.get('platform_sku') or '',
                            row.get('child_code') or '',
                            row.get('product_link') or '',
                            row.get('gtin') or '',
                            row.get('upc') or '',
                            row.get('sku_family') or '',
                            row.get('spec_name') or '',
                            row.get('fabric') or '',
                            '\n'.join(link_map.get(row_id, [])),
                            row.get('sale_price_usd') or '',
                            row.get('notes') or ''
                        ])
            
            # 第1行：模块标题（合并单元格）
            section_headers = [
                ('产品状态', 1, 1),
                ('父体关联', 2, 4),
                ('基础信息', 5, 9),
                ('规格信息', 10, 12),
                ('销售信息', 13, 15)
            ]
            # 第2行：字段标题
            cn_headers = [
                '产品状态(启用/留用/弃用)',
                '店铺(必填)', '父体编号', '新父体SKU标识(父体不存在时选填)',
                '销售平台SKU', '子体编号', '链接', 'GTIN', 'UPC',
                '货号', '规格名称', '面料(面料编号)',
                '关联下单SKU及数量(必填，支持换行|;分隔，示例:MS01A-Brown*2)',
                '售价(USD)', '备注'
            ]

            ws.append([''] * len(cn_headers))
            ws.append(cn_headers)
            header_font = Font(bold=True, color='2A2420', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='B7AEA4'),
                right=Side(style='thin', color='B7AEA4'),
                top=Side(style='thin', color='B7AEA4'),
                bottom=Side(style='thin', color='B7AEA4')
            )

            header_fill_by_col = ['D3D3D3'] * len(cn_headers)
            for col in range(1, len(cn_headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            for idx, (title, start_col, end_col) in enumerate(section_headers):
                if end_col > start_col:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                ws.cell(row=1, column=start_col).value = title
                if hasattr(self, '_get_morandi_section_color_pair'):
                    title_color, sub_header_color = self._get_morandi_section_color_pair(idx)
                else:
                    palette = [('A8B9A5', 'DDE7DB'), ('D7C894', 'ECE5CE')]
                    title_color, sub_header_color = palette[idx % len(palette)]
                fill = PatternFill(start_color=title_color, end_color=title_color, fill_type='solid')
                for col in range(start_col, end_col + 1):
                    header_fill_by_col[col - 1] = sub_header_color
                    ws.cell(row=1, column=col).fill = fill
                    ws.cell(row=1, column=col).border = thin_border

            for idx, cell in enumerate(ws[2], start=1):
                header_color = header_fill_by_col[idx - 1] if idx - 1 < len(header_fill_by_col) else 'D3D3D3'
                header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 第3行固定为示例行，勾选导出数据从第4行起追加，避免覆盖示例
            ws.append([
                '启用',
                '',
                'PARENT-001',
                'MS01-MARKER',
                'MS01-Brown-1A',
                'CHILD-001',
                '',
                '',
                'MS01',
                'A款',
                '棕色/Brown',
                'MS01A-Brown*2\nMS01B-Gray',
                199.99
            ])
            example_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            example_font = Font(italic=True, color='888888')
            for cell in ws[3]:
                cell.fill = example_fill
                cell.font = example_font
            if export_rows:
                for row in export_rows:
                    ws.append(row)

            ref_dims = self._populate_sales_product_template_ref_sheet(ref_ws, bricks)

            # 添加数据验证：规格/面料为隐藏表 M、N 列全量列表（与货号无级联）
            # 使用整块区域一次 add（勿对数千行逐格 add），否则 openpyxl 生成/保存极慢易触发网关 504
            max_validation_row = 3000
            sqref_main = f'A4:A{max_validation_row}'

            status_validation = DataValidation(type='list', formula1='"启用,留用,弃用"', allow_blank=True)
            status_validation.add(sqref_main)
            ws.add_data_validation(status_validation)

            p_end = ref_dims.get('shop_end') or 2
            if shop_options and p_end >= 2:
                shop_validation = DataValidation(
                    type='list',
                    formula1=f"'_refs'!$P$2:$P${p_end}",
                    allow_blank=False,
                )
                shop_validation.add(f'B4:B{max_validation_row}')
                ws.add_data_validation(shop_validation)

            k_end = ref_dims.get('sku_end') or 2
            if sku_family_options and k_end >= 2:
                sku_validation = DataValidation(
                    type='list',
                    formula1=f"'_refs'!$K$2:$K${k_end}",
                    allow_blank=True,
                )
                sku_validation.add(f'I4:I{max_validation_row}')
                ws.add_data_validation(sku_validation)

            m_end = ref_dims.get('spec_all_end') or 2
            spec_validation = DataValidation(
                type='list',
                formula1=f"'_refs'!$M$2:$M${m_end}",
                allow_blank=True,
            )
            spec_validation.add(f'J4:J{max_validation_row}')
            ws.add_data_validation(spec_validation)

            n_end = ref_dims.get('fab_all_end') or 2
            fabric_validation = DataValidation(
                type='list',
                formula1=f"'_refs'!$N$2:$N${n_end}",
                allow_blank=True,
            )
            fabric_validation.add(f'K4:K{max_validation_row}')
            ws.add_data_validation(fabric_validation)

            q_end = ref_dims.get('parent_end') or 2
            if parent_codes and q_end >= 2:
                parent_validation = DataValidation(
                    type='list',
                    formula1=f"'_refs'!$Q$2:$Q${q_end}",
                    allow_blank=True,
                )
                parent_validation.add(f'C4:C{max_validation_row}')
                ws.add_data_validation(parent_validation)
            
            
            # 设置列宽
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['G'].width = 16
            ws.column_dimensions['H'].width = 14
            ws.column_dimensions['I'].width = 14
            ws.column_dimensions['D'].width = 22
            ws.column_dimensions['K'].width = 16
            ws.column_dimensions['L'].width = 34
            ws.column_dimensions['M'].width = 14
            ws.column_dimensions['P'].width = 14
            ws.column_dimensions['Q'].width = 14
            ws.column_dimensions['R'].width = 14
            ws.column_dimensions['S'].width = 14
            ws.column_dimensions['T'].width = 14
            
            ws.freeze_panes = 'A4'
            
            return self._send_excel_workbook(wb, '销售产品导入模板.xlsx', start_response)
        except Exception as e:
            # 下载接口兜底：即使数据查询异常，也返回可打开的模板文件，避免浏览器进入 500 错误页
            try:
                if Workbook is not None:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = 'sales_products'
                    ws.append(['提示'])
                    ws.append(['模板已降级生成，请联系管理员检查服务器日志'])
                    ws.append([f'错误信息: {str(e)}'])
                    return self._send_excel_workbook(wb, '销售产品导入模板.xlsx', start_response)
            except Exception:
                pass
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    # -------------------------------------------------------------------------
    # 销售产品 Excel 批量导入
    # -------------------------------------------------------------------------

    def handle_sales_product_import_api(self, environ, method, start_response):
        """销售产品批量导入"""
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            preview_mode = str((query_params.get('preview', ['0'])[0] or '0')).lower() in ('1', 'true', 'yes', 'on')

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            file_bytes = self._sanitize_xlsx_bool_cells(file_bytes)

            try:
                wb = load_workbook(io.BytesIO(file_bytes))
            except Exception as e:
                if 'Cannot be converted to bool' in str(e):
                    wb = self._rebuild_workbook_from_xlsx_xml(file_bytes)
                    if wb is None:
                        diag = self._scan_xlsx_invalid_bool_cells(file_bytes)
                        return self.send_json({
                            'status': 'error',
                            'message': (
                                '导入失败：文件中存在异常布尔字段且无法自动修复，'
                                '请另存为新的xlsx后重试'
                            ),
                            'debug': {
                                'cause': 'Cannot be converted to bool',
                                'invalid_bool_cells': diag.get('count', 0),
                                'samples': diag.get('samples', [])
                            }
                        }, start_response)
                else:
                    return self.send_json({'status': 'error', 'message': str(e)}, start_response)

            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'b' and not isinstance(cell.value, bool):
                            cell.data_type = 's'
                            cell.value = str(cell.value)

            ws = wb.active

            # 智能检测标题行：扫描前5行，找到包含关键字段的行作为标题行
            header_row_idx = 1
            key_indicators = ['店铺', '销售平台SKU', '父体编号', '关联下单SKU', 'shop', 'platform_sku', 'parent_code']
            for row_check in range(1, min(6, ws.max_row + 1)):
                row_cells = [str(cell.value or '').strip() for cell in ws[row_check]]
                row_text = '|'.join(row_cells).lower()
                # 检查是否包含关键指示字段
                if any(key.lower() in row_text for key in key_indicators):
                    header_row_idx = row_check
                    break
            
            headers = [cell.value for cell in ws[header_row_idx]]
            
            # 中文标签到字段代码的映射
            label_to_code = {
                '产品状态(启用/留用/弃用)': 'product_status',
                '店铺(必填)': 'shop_name',
                '店铺(可选)': 'shop_name',
                '店铺': 'shop_name',
                '平台SKU': 'platform_sku',
                '销售平台SKU': 'platform_sku',
                '父体编号': 'parent_code',
                '新父体SKU标识(父体不存在时选填)': 'parent_sku_marker',
                '子体编号': 'child_code',
                '链接': 'product_link',
                '产品网页链接': 'product_link',
                'product_link': 'product_link',
                'GTIN': 'gtin',
                'UPC': 'upc',
                'gtin': 'gtin',
                'upc': 'upc',
                '货号': 'sku_family',
                '面料(选填)': 'fabric',
                '规格名(选填)': 'spec_name',
                '面料(面料编号)': 'fabric',
                '面料': 'fabric',
                '规格名称': 'spec_name',
                '关联下单SKU\n(支持换行|;分隔)': 'order_sku_links',
                '关联下单SKU及数量(必填，支持换行|;分隔，示例:MS01A-Brown*2)': 'order_sku_links',
                '售价(USD)': 'sale_price_usd',
                '备注': 'notes',
                'notes': 'notes',
                '组装后长(in)': 'finished_length_in',
                '组装后宽(in)': 'finished_width_in',
                '组装后高(in)': 'finished_height_in',
                # 兼容旧字段名
                'shop_name': 'shop_name',
                'brand_name': 'brand_name',
                'platform_type': 'platform_type',
                'product_status': 'product_status',
                'platform_sku': 'platform_sku',
                'parent_asin': 'parent_code',
                'child_asin': 'child_code',
                'sku_family': 'sku_family',
                'fabric': 'fabric',
                'spec_name': 'spec_name',
                'sale_price_usd': 'sale_price_usd',
                'finished_length_in': 'finished_length_in',
                'finished_width_in': 'finished_width_in',
                'finished_height_in': 'finished_height_in',
                'assembled_length_in': 'finished_length_in',
                'assembled_width_in': 'finished_width_in',
                'assembled_height_in': 'finished_height_in',
                'order_sku_links': 'order_sku_links'
            }
            
            # 构建列映射，支持中文和旧格式
            header_map = {}
            for idx, h in enumerate(headers):
                if h is not None:
                    h_str = str(h).strip()
                    if h_str:  # 只处理非空的列标题
                        field_code = label_to_code.get(h_str, h_str)
                        if field_code not in header_map:  # 避免后面的重复列覆盖前面的
                            header_map[field_code] = idx
            
            # 诊断：保存所有读到的列（含None和空值）供调试
            detected_headers = [str(h).strip() if h else '[空]' for h in headers]
            detected_headers_non_empty = [h for h in detected_headers if h != '[空]']
            has_shop_name_column = 'shop_name' in header_map

            # 如果预检发现没有shop_name列，立即返回诊断信息
            if not has_shop_name_column:
                return self.send_json({
                    'status': 'error',
                    'message': (
                        f'导入失败：找不到店铺列。系统在第 {header_row_idx} 行检测到了以下列标题：\n'
                        f'{", ".join(detected_headers_non_empty) if detected_headers_non_empty else "[无有效列标题]"}\n\n'
                        f'请确保Excel中包含"店铺(必填)"或"店铺"列。'
                        f'如果列标题位置与预期不符，请重新下载模板并按照模板格式整理数据。'
                    ),
                    'detected_headers': detected_headers,
                    'detected_header_row': header_row_idx,
                    'detected_headers_count': len(detected_headers_non_empty),
                    'expected_shop_column_names': ['店铺(必填)', '店铺(可选)', '店铺', 'shop_name']
                }, start_response)

            def get_cell(row, key):
                idx = header_map.get(key)
                if idx is None:
                    return None
                return row[idx].value

            def parse_links(raw):
                """解析 order_sku_links：支持换行/分号/竖线/逗号分隔，重复SKU自动汇总数量"""
                if raw is None:
                    return []
                text = str(raw).strip()
                if not text:
                    return []
                
                # 支持换行符、分号、竖线、逗号分隔
                parts = [t.strip() for t in re.split(r'[\n\r;；|,，]+', text) if t.strip()]
                sku_qty_map = {}
                
                for part in parts:
                    if '*' in part:
                        sku, qty = part.split('*', 1)
                    else:
                        sku, qty = part, None
                    
                    sku = sku.strip()
                    if not sku:
                        continue
                    
                    if qty is None:
                        qty_val = 1
                    else:
                        qty = qty.strip()
                        try:
                            qty_val = int(qty) if qty else 1
                        except Exception:
                            qty_val = 1

                    sku_qty_map[sku] = sku_qty_map.get(sku, 0) + max(1, qty_val)

                return [(sku, qty) for sku, qty in sku_qty_map.items() if qty > 0]

            with self._get_db_connection() as conn:
                sp_has_shop_col = self._table_has_column(conn, 'sales_products', 'shop_id')

                tx_enabled = False
                if not preview_mode:
                    try:
                        conn.autocommit(False)
                        tx_enabled = True
                    except Exception:
                        tx_enabled = False

                with conn.cursor() as cur:
                    cur.execute("SELECT id, parent_code, shop_id FROM sales_parents")
                    parent_map = {
                        (int(row.get('shop_id') or 0), str(row.get('parent_code') or '').strip()): row
                        for row in (cur.fetchall() or []) if row.get('parent_code')
                    }

                    cur.execute("SELECT id, shop_name FROM shops")
                    shop_map = {str(row['shop_name']).strip(): row['id'] for row in (cur.fetchall() or []) if row.get('shop_name')}

                    cur.execute(
                        """
                        SELECT s.id, pt.name AS platform_type_name
                        FROM shops s
                        LEFT JOIN platform_types pt ON pt.id = s.platform_type_id
                        """
                    )
                    shop_platform_map = {
                        int(row.get('id') or 0): row.get('platform_type_name')
                        for row in (cur.fetchall() or [])
                        if row.get('id')
                    }

                    has_reship_accessory = self._table_has_column(conn, 'order_products', 'is_reship_accessory')
                    accessory_filter_sql = "WHERE COALESCE(op.is_reship_accessory, 0) = 0" if has_reship_accessory else ""
                    cur.execute(
                        f"""
                        SELECT op.id, op.sku, op.sku_family_id, op.spec_qty_short,
                               op.cost_usd, op.last_mile_avg_freight_usd,
                               op.finished_length_in, op.finished_width_in, op.finished_height_in,
                               op.package_length_in, op.package_width_in, op.package_height_in,
                               op.net_weight_lbs, op.gross_weight_lbs,
                               fm.fabric_code, fm.fabric_name_en
                        FROM order_products op
                        LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                        {accessory_filter_sql}
                        """
                    )
                    order_rows = cur.fetchall() or []
                    order_map = {row['sku']: row['id'] for row in order_rows if row.get('sku')}
                    order_detail_by_id = {row['id']: row for row in order_rows if row.get('id')}

                    cur.execute("SELECT id, sku_family FROM product_families")
                    sku_family_rows = cur.fetchall() or []
                    sku_family_map = {str(row['sku_family']).strip(): row['id'] for row in sku_family_rows if row.get('sku_family')}
                    sku_family_code_map = {row['id']: (row.get('sku_family') or '').strip() for row in sku_family_rows if row.get('id')}

                    if sp_has_shop_col:
                        cur.execute("SELECT id, shop_id, platform_sku FROM sales_products")
                    else:
                        cur.execute(
                            """
                            SELECT sp.id, p.shop_id, sp.platform_sku
                            FROM sales_products sp
                            LEFT JOIN sales_parents p ON p.id = sp.parent_id
                            """
                        )
                    sales_map = {(int(row.get('shop_id') or 0), str(row.get('platform_sku') or '').strip()): int(row.get('id') or 0) for row in (cur.fetchall() or []) if row.get('platform_sku')}

                    vf_join, vf_expr = self._sales_variant_fabric_select_sql(conn, 'v')
                    cur.execute(
                        f"""
                        SELECT v.id, v.sku_family_id, v.spec_name, ({vf_expr}) AS fabric
                        FROM sales_product_variants v
                        {vf_join}
                        """
                    )
                    variant_identity_map = {
                        (int(row.get('sku_family_id') or 0), str(row.get('spec_name') or '').strip(), str(row.get('fabric') or '').strip()): int(row.get('id') or 0)
                        for row in (cur.fetchall() or []) if row.get('id')
                    }

                created = 0
                updated = 0
                unchanged = 0
                relation_created = 0
                relation_deleted = 0
                total_rows = 0
                errors = []
                data_start_row = header_row_idx + 2

                with conn.cursor() as row_cur:
                    for row_idx in range(data_start_row, ws.max_row + 1):
                        row = ws[row_idx]
                        row_values = [cell.value for cell in row]
                        if not any(v is not None and str(v).strip() for v in row_values):
                            continue
                        total_rows += 1

                        # 支持两种格式：新的合并列 vs 旧的分开列
                        platform_sku = (get_cell(row, 'platform_sku') or '').strip()
                        product_status_text = (get_cell(row, 'product_status') or '').strip()
                        status_map = {'启用': 'enabled', '留用': 'retained', '弃用': 'discarded'}
                        product_status = status_map.get(product_status_text, (product_status_text or 'enabled').lower())
                        if product_status not in ('enabled', 'retained', 'discarded'):
                            product_status = 'enabled'
                        parent_code = (get_cell(row, 'parent_code') or '').strip() or None
                        parent_sku_marker = (get_cell(row, 'parent_sku_marker') or '').strip() or None
                        child_code = (get_cell(row, 'child_code') or '').strip() or None
                        import_has_product_link = 'product_link' in header_map
                        product_link_raw = get_cell(row, 'product_link') if import_has_product_link else None
                        import_has_gtin = 'gtin' in header_map
                        import_has_upc = 'upc' in header_map
                        gtin = self._parse_sales_barcode(get_cell(row, 'gtin')) if import_has_gtin else None
                        upc = self._parse_sales_barcode(get_cell(row, 'upc')) if import_has_upc else None
                        sku_family_name = (get_cell(row, 'sku_family') or '').strip() or None
                        fabric = self._normalize_sales_import_fabric_cell(get_cell(row, 'fabric'))
                        spec_name = (get_cell(row, 'spec_name') or '').strip()
                        sale_price_usd = self._parse_float(get_cell(row, 'sale_price_usd'))
                        import_has_notes = 'notes' in header_map
                        notes = self._parse_sales_notes(get_cell(row, 'notes')) if import_has_notes else None
                        order_sku_links = (get_cell(row, 'order_sku_links') or '').strip()

                        shop_name_text = (get_cell(row, 'shop_name') or '').strip()
                        if not shop_name_text:
                            errors.append({'row': row_idx, 'error': 'Missing shop_name'})
                            continue
                        shop_id_from_file = shop_map.get(shop_name_text)
                        if not shop_id_from_file:
                            errors.append({'row': row_idx, 'error': f'Unknown shop_name: {shop_name_text}'})
                            continue

                        parent_row = None
                        parent_id = None
                        if parent_code:
                            parent_key = (int(shop_id_from_file), parent_code)
                            parent_row = parent_map.get(parent_key)
                            if not parent_row:
                                if preview_mode:
                                    parent_row = {'id': None, 'parent_code': parent_code, 'shop_id': shop_id_from_file}
                                    parent_map[parent_key] = parent_row
                                else:
                                    row_cur.execute(
                                        """
                                        INSERT INTO sales_parents (parent_code, shop_id, sku_marker)
                                        VALUES (%s, %s, %s)
                                        """,
                                        (parent_code, shop_id_from_file, parent_sku_marker)
                                    )
                                    new_parent_id = row_cur.lastrowid
                                    parent_row = {'id': new_parent_id, 'parent_code': parent_code, 'shop_id': shop_id_from_file}
                                    parent_map[parent_key] = parent_row

                            shop_id = parent_row.get('shop_id')
                            if not shop_id:
                                if (not preview_mode) and parent_row.get('id'):
                                    row_cur.execute("UPDATE sales_parents SET shop_id=%s WHERE id=%s", (shop_id_from_file, parent_row['id']))
                                shop_id = shop_id_from_file
                                parent_row['shop_id'] = shop_id
                            elif int(shop_id) != int(shop_id_from_file):
                                errors.append({'row': row_idx, 'error': f'Parent/shop mismatch: {parent_code} -> {shop_name_text}'})
                                continue

                            parent_id = parent_row.get('id')
                        else:
                            shop_id = shop_id_from_file

                        platform_name = shop_platform_map.get(int(shop_id))
                        product_link = self._resolve_sales_product_link(
                            platform_name, child_code, product_link_raw
                        )

                        link_entries = []
                        for sku, qty in parse_links(order_sku_links):
                            order_id = order_map.get(sku)
                            if not order_id:
                                errors.append({'row': row_idx, 'error': f'Unknown order SKU: {sku}'})
                                link_entries = []
                                break
                            link_entries.append({'order_product_id': order_id, 'quantity': qty})
                        if not link_entries:
                            errors.append({'row': row_idx, 'error': 'Missing order_sku_links'})
                            continue

                        sid_input = sku_family_map.get(sku_family_name) if sku_family_name else None
                        bundle_imp = self._derive_sales_order_links_bundle(conn, sid_input, link_entries)
                        sku_family_id = sid_input or bundle_imp.get('sku_family_id')
                        auto_fabric = (bundle_imp.get('fabric') or '').strip()
                        auto_spec_name = (bundle_imp.get('spec_name') or '').strip()
                        if sku_family_name and not sku_family_id:
                            errors.append({'row': row_idx, 'error': f'Unknown sku_family: {sku_family_name}'})
                            continue
                        if not sku_family_id:
                            errors.append({'row': row_idx, 'error': '无法根据订单SKU推断归属货号'})
                            continue

                        final_spec_name = spec_name or auto_spec_name
                        resolved_fabric_id = None
                        final_fabric = fabric or auto_fabric
                        has_fabric_id_col = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                        if has_fabric_id_col:
                            if fabric:
                                resolved_fabric_id = self._resolve_fabric_material_id_from_label(
                                    conn, fabric, row_cur, allow_name_match=False
                                )
                                if not resolved_fabric_id:
                                    errors.append({
                                        'row': row_idx,
                                        'error': '面料无效：请填写面料主数据中的面料编号（fabric_code），勿使用英文名或无效编码',
                                    })
                                    continue
                                canonical = self._fabric_code_for_material_id(conn, resolved_fabric_id, row_cur)
                                final_fabric = canonical or self._code_before_dash(fabric)
                            else:
                                resolved_fabric_id, link_fabric_err = self._resolve_fabric_material_id_from_order_links_import(
                                    conn, link_entries, row_cur
                                )
                                if not resolved_fabric_id:
                                    if link_fabric_err == 'ambiguous':
                                        errors.append({
                                            'row': row_idx,
                                            'error': '关联下单SKU对应多种面料，请在本行「面料」列填写唯一面料编号',
                                        })
                                    else:
                                        errors.append({
                                            'row': row_idx,
                                            'error': '「面料」列未填且关联下单SKU未绑定面料，请填写面料编号',
                                        })
                                    continue
                                canonical = self._fabric_code_for_material_id(conn, resolved_fabric_id, row_cur)
                                if not canonical:
                                    errors.append({
                                        'row': row_idx,
                                        'error': '关联面料主数据缺少面料编号（fabric_code），请补全面料主数据或在本行填写编号',
                                    })
                                    continue
                                final_fabric = canonical

                        if has_fabric_id_col and resolved_fabric_id:
                            if not self._fabric_product_family_linked(conn, sku_family_id, resolved_fabric_id, row_cur):
                                errors.append({
                                    'row': row_idx,
                                    'error': '货号与面料未绑定：该组合未在货号-面料关联表中维护，请在规格主图管理或面料管理中补充关联后再导入',
                                })
                                continue
                        elif (not has_fabric_id_col) and self._table_has_column(conn, 'sales_product_variants', 'fabric'):
                            ft_chk = (final_fabric or '').strip()
                            if ft_chk:
                                fid_chk = self._resolve_fabric_material_id_from_label(
                                    conn, ft_chk, row_cur, allow_name_match=False
                                )
                                if not fid_chk:
                                    errors.append({
                                        'row': row_idx,
                                        'error': '面料无效：请填写面料主数据中的面料编号（fabric_code），以便校验货号-面料关联',
                                    })
                                    continue
                                if not self._fabric_product_family_linked(conn, sku_family_id, fid_chk, row_cur):
                                    errors.append({
                                        'row': row_idx,
                                        'error': '货号与面料未绑定：该组合未在货号-面料关联表中维护，请在规格主图管理或面料管理中补充关联后再导入',
                                    })
                                    continue

                        sku_family_code = sku_family_code_map.get(sku_family_id) or ''
                        auto_platform_sku = ''
                        if sku_family_code and final_fabric and final_spec_name:
                            auto_platform_sku = self._build_sales_platform_sku(sku_family_code, final_spec_name, final_fabric)
                        final_platform_sku = platform_sku or auto_platform_sku

                        if not final_platform_sku:
                            errors.append({'row': row_idx, 'error': 'Platform SKU missing'})
                            continue

                        variant_key = (int(sku_family_id), str(final_spec_name or '').strip(), str(final_fabric or '').strip())
                        if preview_mode:
                            if sales_map.get((int(shop_id), final_platform_sku)):
                                updated += 1
                            else:
                                created += 1
                            continue

                        try:
                            variant_id = variant_identity_map.get(variant_key)
                            if not variant_id:
                                variant_id = self._get_or_create_sales_variant(conn, sku_family_id, final_spec_name, final_fabric, fabric_id=resolved_fabric_id)
                                variant_identity_map[variant_key] = variant_id

                            target_id = sales_map.get((int(shop_id), final_platform_sku))
                            if target_id:
                                update_fields = [
                                    "platform_sku=%s",
                                    "product_status=%s",
                                    "variant_id=%s",
                                    "parent_id=%s",
                                    "child_code=%s",
                                ]
                                update_values = [final_platform_sku, product_status, variant_id, parent_id, child_code]
                                if self._table_has_column(conn, 'sales_products', 'product_link'):
                                    update_fields.append("product_link=%s")
                                    update_values.append(product_link)
                                if import_has_gtin and self._table_has_column(conn, 'sales_products', 'gtin'):
                                    update_fields.append("gtin=%s")
                                    update_values.append(gtin)
                                if import_has_upc and self._table_has_column(conn, 'sales_products', 'upc'):
                                    update_fields.append("upc=%s")
                                    update_values.append(upc)
                                update_fields.append("sale_price_usd=%s")
                                update_values.append(sale_price_usd)
                                if import_has_notes and self._table_has_column(conn, 'sales_products', 'notes'):
                                    update_fields.append("notes=%s")
                                    update_values.append(notes)
                                if sp_has_shop_col:
                                    update_fields.insert(0, "shop_id=%s")
                                    update_values.insert(0, shop_id)
                                update_values.append(target_id)
                                row_cur.execute(
                                    f"UPDATE sales_products SET {', '.join(update_fields)} WHERE id=%s",
                                    update_values
                                )
                                updated += 1
                            else:
                                insert_columns = []
                                insert_values = []
                                if sp_has_shop_col:
                                    insert_columns.append('shop_id')
                                    insert_values.append(shop_id)
                                insert_columns.extend(['platform_sku', 'product_status'])
                                insert_values.extend([final_platform_sku, product_status])
                                insert_columns.extend(['variant_id', 'parent_id', 'child_code', 'sale_price_usd'])
                                insert_values.extend([variant_id, parent_id, child_code, sale_price_usd])
                                if self._table_has_column(conn, 'sales_products', 'product_link'):
                                    insert_columns.append('product_link')
                                    insert_values.append(product_link)
                                if import_has_gtin and self._table_has_column(conn, 'sales_products', 'gtin'):
                                    insert_columns.append('gtin')
                                    insert_values.append(gtin)
                                if import_has_upc and self._table_has_column(conn, 'sales_products', 'upc'):
                                    insert_columns.append('upc')
                                    insert_values.append(upc)
                                if import_has_notes and self._table_has_column(conn, 'sales_products', 'notes'):
                                    insert_columns.append('notes')
                                    insert_values.append(notes)
                                placeholders_insert = ', '.join(['%s'] * len(insert_columns))
                                row_cur.execute(
                                    f"INSERT INTO sales_products ({', '.join(insert_columns)}) VALUES ({placeholders_insert})",
                                    insert_values
                                )
                                target_id = row_cur.lastrowid
                                sales_map[(int(shop_id), final_platform_sku)] = target_id
                                created += 1
                            self._replace_sales_variant_order_links(conn, variant_id, link_entries)
                            relation_created += len(link_entries)
                                    
                        except Exception as e:
                            errors.append({'row': row_idx, 'error': str(e)})

                if tx_enabled:
                    conn.commit()
                    conn.autocommit(True)

            return self.send_json({
                'status': 'success',
                'preview': 1 if preview_mode else 0,
                'total_rows': total_rows,
                'created': created,
                'updated': updated,
                'unchanged': unchanged,
                'relation_created': relation_created,
                'relation_added': relation_created,
                'relation_deleted': relation_deleted,
                'errors': errors
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    # -------------------------------------------------------------------------
    # 销售产品 CRUD API
    # -------------------------------------------------------------------------

    def handle_sales_product_api(self, environ, method, start_response):
        """销售产品管理 API（CRUD）"""
        try:
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                item_id = self._parse_int((query_params.get('id', [''])[0] or '').strip())
                get_action = (query_params.get('action', [''])[0] or '').strip().lower()
                include_links = str((query_params.get('include_links', ['0'])[0] or '0')).lower() in ('1', 'true', 'yes', 'on')
                if get_action == 'discount_segments':
                    if not item_id:
                        return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                    with self._get_db_connection() as conn:
                        if not self._sales_discount_segments_table_ready(conn):
                            return self.send_json({'status': 'success', 'items': []}, start_response)
                        with conn.cursor() as cur:
                            rows = self._load_sales_discount_segments(cur, item_id)
                    return self.send_json({'status': 'success', 'items': rows}, start_response)
                with self._get_db_connection() as conn:
                    sp_has_shop_col = self._table_has_column(conn, 'sales_products', 'shop_id')
                    shop_expr = self._sales_product_shop_expr(sp_has_shop_col)
                    with conn.cursor() as cur:
                        has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                        has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                        fabric_join = "LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if has_fabric_id else ""
                        if has_fabric_id and has_fabric_text:
                            fabric_select = "COALESCE(fm.fabric_code, v.fabric)"
                        elif has_fabric_id:
                            fabric_select = "fm.fabric_code"
                        else:
                            fabric_select = "v.fabric" if has_fabric_text else "''"
                        barcode_select = self._sales_product_barcode_select_sql(conn, 'sp')
                        preview_fields_select = self._sales_product_preview_select_sql(conn)
                        product_link_select = (
                            "sp.product_link"
                            if self._table_has_column(conn, 'sales_products', 'product_link')
                            else "NULL AS product_link"
                        )
                        handles_last_mile_select = self._shop_handles_last_mile_select_sql(conn, 's', 'pt')
                        base_sql = """
                            SELECT
                                sp.id,
                                {shop_expr} AS shop_id,
                                sp.platform_sku,
                                sp.product_status,
                                sp.parent_id,
                                sp.child_code,
                                {product_link_select},
                                {barcode_select},
                                sp.variant_id,
                                v.sku_family_id,
                                pf.sku_family,
                                v.spec_name,
                                {fabric_select} AS fabric,
                                {fabric_id_select} AS fabric_id,
                                sp.sale_price_usd,
                                {preview_fields_select},
                                sp.created_at,
                                sp.updated_at,
                                s.shop_name,
                                s.platform_type_id,
                                pt.name AS platform_type_name,
                                TRIM(COALESCE(pf.category, '')) AS product_category,
                                b.name AS brand_name,
                                {handles_last_mile_select},
                                p.parent_code,
                                p.sku_marker AS parent_sku_marker,
                                p.estimated_refund_rate,
                                p.estimated_discount_rate,
                                p.estimated_acoas
                            FROM sales_products sp
                            LEFT JOIN sales_parents p ON p.id = sp.parent_id
                            LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                            LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                            {fabric_join}
                            LEFT JOIN shops s ON s.id = {shop_expr}
                            LEFT JOIN platform_types pt ON pt.id = s.platform_type_id
                            LEFT JOIN brands b ON b.id = s.brand_id
                        """.format(
                            shop_expr=shop_expr,
                            product_link_select=product_link_select,
                            barcode_select=barcode_select,
                            fabric_join=fabric_join,
                            fabric_select=fabric_select,
                            fabric_id_select=("v.fabric_id" if has_fabric_id else "NULL"),
                            preview_fields_select=preview_fields_select,
                            handles_last_mile_select=handles_last_mile_select,
                        )
                        filters = []
                        params = []
                        if item_id:
                            filters.append("sp.id = %s")
                            params.append(item_id)
                        if keyword:
                            text_filters = [
                                "sp.platform_sku LIKE %s",
                                "s.shop_name LIKE %s",
                                "p.parent_code LIKE %s",
                                "sp.child_code LIKE %s",
                                "pf.sku_family LIKE %s",
                                "v.spec_name LIKE %s",
                                f"{fabric_select} LIKE %s",
                            ]
                            kw_params = [f"%{keyword}%"] * 7
                            if self._table_has_column(conn, 'sales_products', 'gtin'):
                                text_filters.append("sp.gtin LIKE %s")
                                kw_params.append(f"%{keyword}%")
                            if self._table_has_column(conn, 'sales_products', 'upc'):
                                text_filters.append("sp.upc LIKE %s")
                                kw_params.append(f"%{keyword}%")
                            if self._table_has_column(conn, 'sales_products', 'notes'):
                                text_filters.append("sp.notes LIKE %s")
                                kw_params.append(f"%{keyword}%")
                            if self._table_has_column(conn, 'sales_products', 'product_link'):
                                text_filters.append("sp.product_link LIKE %s")
                                kw_params.append(f"%{keyword}%")
                            params.extend(kw_params)
                            if has_fabric_text:
                                text_filters.append("v.fabric LIKE %s")
                                params.append(f"%{keyword}%")
                            filters.append("(" + " OR ".join(text_filters) + ")")
                        where_sql = (" WHERE " + " AND ".join(filters)) if filters else ""
                        cur.execute(base_sql + where_sql + " ORDER BY sp.id DESC", params)
                        rows = cur.fetchall() or []
                    variant_ids = [int(r.get('variant_id') or 0) for r in rows if int(r.get('variant_id') or 0) > 0]
                    metrics_map = {}
                    sellable_map = {}
                    if variant_ids:
                        # Reuse same DB connection for performance (must be inside conn context)
                        metrics_map = self._load_sales_variant_metrics(conn, variant_ids)
                        sellable_map = self._load_variant_overseas_sellable_map(conn, variant_ids)
                    self._turnover_attach_to_sales_product_rows(conn, rows)

                    for row in rows:
                        variant_id = int(row.get('variant_id') or 0)
                        metrics = metrics_map.get(variant_id, {}) if variant_id else {}
                        row['overseas_sellable_qty'] = int(sellable_map.get(variant_id) or 0) if variant_id else 0
                        row['warehouse_cost_usd'] = metrics.get('warehouse_cost_usd', 0.0)
                        row['last_mile_cost_usd'] = metrics.get('last_mile_cost_usd', 0.0)
                        row['package_length_in'] = metrics.get('package_length_in', 0.0)
                        row['package_width_in'] = metrics.get('package_width_in', 0.0)
                        row['package_height_in'] = metrics.get('package_height_in', 0.0)
                        row['net_weight_lbs'] = metrics.get('net_weight_lbs', 0.0)
                        row['gross_weight_lbs'] = metrics.get('gross_weight_lbs', 0.0)
                        if item_id and (not include_links):
                            row['order_sku_links'] = []
                        else:
                            row['order_sku_links'] = metrics.get('order_sku_links', [])

                    # Variant preview image (first 白底图)：列表与单条 GET 均填充，供前端刷新行缩略图
                    try:
                        vid_list = [int(r.get('variant_id') or 0) for r in rows if int(r.get('variant_id') or 0) > 0]
                        preview_map = {}
                        if vid_list:
                            preview_map = self._load_variant_first_image_preview(conn, vid_list, type_name='白底纯图')
                        for r in rows:
                            vid = int(r.get('variant_id') or 0)
                            r['preview_image_b64'] = preview_map.get(vid, '') if vid else ''
                    except Exception:
                        for r in rows:
                            r['preview_image_b64'] = ''

                if item_id:
                    return self.send_json({
                        'status': 'success',
                        'item': rows[0] if rows else None,
                        'turnover_sales_window': self._turnover_sales_window(conn),
                    }, start_response)
                return self.send_json({
                    'status': 'success',
                    'items': rows,
                    'turnover_sales_window': self._turnover_sales_window(conn),
                }, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                platform_sku_manual = (data.get('platform_sku') or '').strip()
                product_status = (data.get('product_status') or 'enabled').strip().lower()
                if product_status not in ('enabled', 'retained', 'discarded'):
                    product_status = 'enabled'
                sku_family_id_input = self._parse_int(data.get('sku_family_id'))
                shop_id_input = self._parse_int(data.get('shop_id'))
                parent_code = (data.get('parent_code') or '').strip() or None
                parent_sku_marker = (data.get('parent_sku_marker') or '').strip() or None
                child_code = (data.get('child_code') or '').strip() or None
                gtin = self._parse_sales_barcode(data.get('gtin'))
                upc = self._parse_sales_barcode(data.get('upc'))
                sale_price_usd = self._parse_float(data.get('sale_price_usd'))
                fabric_id_input = self._parse_int(data.get('fabric_id'))
                notes = self._parse_sales_notes(data.get('notes'))
                links = self._normalize_sales_order_links(data.get('order_sku_links'))
                
                # 检查是否手动编辑了platform_sku
                manual_platform_sku = bool(data.get('manual_platform_sku'))
                
                if not links:
                    return self.send_json({'status': 'error', 'message': '关联下单SKU及数量为必填'}, start_response)

                with self._get_db_connection() as conn:
                    sp_has_shop_col = self._table_has_column(conn, 'sales_products', 'shop_id')
                    bundle = self._derive_sales_order_links_bundle(conn, sku_family_id_input, links)
                    sku_family_id = sku_family_id_input or bundle.get('sku_family_id')
                    if not sku_family_id:
                        return self.send_json({'status': 'error', 'message': '无法根据下单SKU推断归属货号'}, start_response)

                    sku_family_code = (bundle.get('sku_family_code') or '').strip()

                    parent_id = None
                    parent_shop_id = None
                    if parent_code:
                        with conn.cursor() as cur:
                            cur.execute("SELECT id, shop_id FROM sales_parents WHERE parent_code=%s AND shop_id=%s LIMIT 1", (parent_code, shop_id_input))
                            row = cur.fetchone()
                            if row:
                                parent_id = row['id']
                                parent_shop_id = row.get('shop_id')
                                if (not parent_shop_id) and shop_id_input:
                                    cur.execute("UPDATE sales_parents SET shop_id=%s WHERE id=%s", (shop_id_input, parent_id))
                                    parent_shop_id = shop_id_input
                            else:
                                cur.execute(
                                    """
                                    INSERT INTO sales_parents (parent_code, shop_id, sku_marker)
                                    VALUES (%s, %s, %s)
                                    """,
                                    (parent_code, shop_id_input, parent_sku_marker)
                                )
                                parent_id = cur.lastrowid
                                parent_shop_id = shop_id_input
                    final_shop_id = parent_shop_id if parent_id else shop_id_input
                    if not final_shop_id:
                        return self.send_json({'status': 'error', 'message': 'Missing required field: shop_id'}, start_response)

                    auto_fabric = (bundle.get('fabric') or '').strip()
                    auto_spec_name = (bundle.get('spec_name') or '').strip()
                    auto_platform_sku = (bundle.get('platform_sku') or '').strip()
                    final_fabric = (data.get('fabric') or '').strip() or auto_fabric
                    final_spec_name = (data.get('spec_name') or '').strip() or auto_spec_name
                    # Prefer fabric_id when provided; fallback to derived first fabric by code lookup
                    resolved_fabric_id = fabric_id_input or None
                    if not resolved_fabric_id and final_fabric:
                        try:
                            with conn.cursor() as fcur:
                                fcur.execute("SELECT id FROM fabric_materials WHERE fabric_code=%s LIMIT 1", (self._code_before_dash(final_fabric),))
                                frow = fcur.fetchone() or {}
                                resolved_fabric_id = self._parse_int(frow.get('id')) or None
                        except Exception:
                            resolved_fabric_id = None
                    if not resolved_fabric_id:
                        return self.send_json({'status': 'error', 'message': '面料不能为空：请在“面料”下拉中选择面料（fabric_id 必填）'}, start_response)
                    variant_id = self._get_or_create_sales_variant(conn, sku_family_id, final_spec_name, final_fabric, fabric_id=resolved_fabric_id)
                    
                    # 如果没有手动编辑，使用自动生成的platform_sku；否则使用手动输入的
                    if manual_platform_sku:
                        platform_sku = platform_sku_manual
                    else:
                        platform_sku = auto_platform_sku or self._build_sales_platform_sku(sku_family_code, final_spec_name, final_fabric)
                    
                    if not platform_sku:
                        return self.send_json({'status': 'error', 'message': '无法生成销售平台SKU，请手动输入'}, start_response)
                    
                    with conn.cursor() as cur:
                        platform_name = self._load_shop_platform_type_name(cur, final_shop_id)
                    product_link = self._resolve_sales_product_link(
                        platform_name, child_code, data.get('product_link')
                    )

                    with conn.cursor() as cur:
                        insert_columns = []
                        insert_values = []
                        if sp_has_shop_col:
                            insert_columns.append('shop_id')
                            insert_values.append(final_shop_id)
                        insert_columns.extend(['platform_sku', 'product_status'])
                        insert_values.extend([platform_sku, product_status])
                        insert_columns.extend(['variant_id', 'parent_id', 'child_code', 'sale_price_usd'])
                        insert_values.extend([variant_id, parent_id, child_code, sale_price_usd])
                        self._extend_sales_product_link_write(conn, insert_columns, insert_values, product_link)
                        self._extend_sales_product_barcode_write(conn, insert_columns, insert_values, gtin, upc)
                        if self._table_has_column(conn, 'sales_products', 'notes'):
                            insert_columns.append('notes')
                            insert_values.append(notes)
                        placeholders_insert = ', '.join(['%s'] * len(insert_columns))
                        cur.execute(
                            f"INSERT INTO sales_products ({', '.join(insert_columns)}) VALUES ({placeholders_insert})",
                            insert_values
                        )
                        new_id = cur.lastrowid
                    self._replace_sales_variant_order_links(conn, variant_id, links)
                    fabric_folder_part = self._resolve_fabric_folder_part(conn, resolved_fabric_id, final_fabric)
                    self._ensure_listing_sales_variant_folder(sku_family_code, final_spec_name, fabric_folder_part)
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                query_params_put = parse_qs(environ.get('QUERY_STRING', ''))
                preview_action = (query_params_put.get('action', [''])[0] or '').strip().lower()
                data = self._read_json_body(environ)
                if preview_action == 'preview_update':
                    batch_items = data.get('items') if isinstance(data, dict) else None
                    if not isinstance(batch_items, list) or not batch_items:
                        return self.send_json({'status': 'error', 'message': 'Missing preview items'}, start_response)

                    preview_field_names = self._SALES_PREVIEW_FIELD_NAMES
                    row_map = {}
                    touched_map = {}
                    for item in batch_items:
                        if not isinstance(item, dict):
                            continue
                        item_id = self._parse_int(item.get('id'))
                        if not item_id:
                            continue
                        child_code_raw = item.get('child_code')
                        child_code = None
                        if child_code_raw is not None:
                            child_code = (str(child_code_raw).strip() or None)
                        gtin_raw = item.get('gtin')
                        gtin = self._parse_sales_barcode(gtin_raw) if gtin_raw is not None else None
                        upc_raw = item.get('upc')
                        upc = self._parse_sales_barcode(upc_raw) if upc_raw is not None else None
                        notes_raw = item.get('notes')
                        notes = self._parse_sales_notes(notes_raw) if notes_raw is not None else None
                        product_link_raw = item.get('product_link')
                        product_link = None
                        if product_link_raw is not None:
                            product_link = (str(product_link_raw).strip() or None)
                        row_map[int(item_id)] = {
                            'child_code': child_code,
                            'product_link': product_link,
                            'gtin': gtin,
                            'upc': upc,
                            'sale_price_usd': self._parse_float(item.get('sale_price_usd')),
                            'promotion_activity_type': self._normalize_sales_promotion_activity_type(item.get('promotion_activity_type')),
                            'discount_form_type': self._normalize_sales_discount_form_type(item.get('discount_form_type')),
                            'actual_discount_rate': self._parse_float(item.get('actual_discount_rate')),
                            'actual_discount_amount_usd': self._parse_float(item.get('actual_discount_amount_usd')),
                            'discounted_price_usd': self._parse_float(item.get('discounted_price_usd')),
                            'notes': notes,
                        }
                        touched_raw = item.get('touched_fields')
                        if isinstance(touched_raw, list):
                            touched = {
                                str(name).strip()
                                for name in touched_raw
                                if str(name).strip() in preview_field_names
                            }
                        else:
                            touched = set(preview_field_names)
                        touched_map[int(item_id)] = touched

                    if not row_map:
                        return self.send_json({'status': 'error', 'message': 'No valid preview items'}, start_response)

                    bundle_errors = []
                    for item_id, values in row_map.items():
                        touched = touched_map.get(item_id) or set()
                        if not self._sales_discount_bundle_any_touched(touched):
                            continue
                        expanded = self._sales_discount_bundle_expand_touched(touched)
                        touched_map[item_id] = expanded
                        normalized = self._normalize_sales_discount_bundle_values(values)
                        row_map[item_id].update(normalized)
                        err = self._validate_sales_discount_bundle_values(normalized)
                        if err:
                            bundle_errors.append({'id': item_id, 'error': err})
                    if bundle_errors:
                        return self.send_json({
                            'status': 'error',
                            'message': bundle_errors[0].get('error') or '折扣信息不完整',
                            'errors': bundle_errors,
                        }, start_response)

                    user_id = self._parse_int(self._get_session_user(environ)) or None
                    segments_recorded = 0
                    with self._get_db_connection() as conn:
                        col_exists = self._sales_product_preview_col_exists(conn)
                        sp_has_shop_col = self._table_has_column(conn, 'sales_products', 'shop_id')
                        segments_ready = self._sales_discount_segments_table_ready(conn)
                        with conn.cursor() as cur:
                            discount_history_ids = [
                                pid for pid, touched in touched_map.items()
                                if set(touched) & self._SALES_DISCOUNT_SEGMENT_TRIGGER_FIELDS
                            ]
                            old_discount_meta = (
                                self._load_sales_discount_snapshots(cur, discount_history_ids)
                                if segments_ready and discount_history_ids else {}
                            )
                            self._sales_product_preview_resolve_product_links(
                                conn, cur, row_map, touched_map, sp_has_shop_col
                            )
                            if segments_ready and old_discount_meta:
                                segments_recorded = self._apply_sales_discount_segment_updates(
                                    cur, conn, row_map, touched_map, old_discount_meta, user_id=user_id
                                )
                            updated = self._sales_product_preview_batch_update(
                                cur, row_map, touched_map, col_exists
                            )
                    return self.send_json({
                        'status': 'success',
                        'updated': updated,
                        'discount_segments_recorded': segments_recorded,
                    }, start_response)

                item_id = self._parse_int(data.get('id'))
                platform_sku_manual = (data.get('platform_sku') or '').strip()
                product_status = (data.get('product_status') or 'enabled').strip().lower()
                if product_status not in ('enabled', 'retained', 'discarded'):
                    product_status = 'enabled'
                sku_family_id_input = self._parse_int(data.get('sku_family_id'))
                shop_id_input = self._parse_int(data.get('shop_id'))
                parent_code = (data.get('parent_code') or '').strip() or None
                parent_sku_marker = (data.get('parent_sku_marker') or '').strip() or None
                child_code = (data.get('child_code') or '').strip() or None
                gtin = self._parse_sales_barcode(data.get('gtin'))
                upc = self._parse_sales_barcode(data.get('upc'))
                sale_price_usd = self._parse_float(data.get('sale_price_usd'))
                notes = self._parse_sales_notes(data.get('notes'))
                fabric_id_input = self._parse_int(data.get('fabric_id'))
                confirm_new_variant_folder = bool(data.get('confirm_new_variant_folder'))
                links = self._normalize_sales_order_links(data.get('order_sku_links'))
                
                # 检查是否手动编辑了platform_sku
                manual_platform_sku = bool(data.get('manual_platform_sku'))
                
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing required field: id'}, start_response)
                if not links:
                    return self.send_json({'status': 'error', 'message': '关联下单SKU及数量为必填'}, start_response)

                with self._get_db_connection() as conn:
                    sp_has_shop_col = self._table_has_column(conn, 'sales_products', 'shop_id')
                    current_variant_fabric_id = 0
                    try:
                        if self._table_has_column(conn, 'sales_product_variants', 'fabric_id'):
                            with conn.cursor() as cur:
                                cur.execute(
                                    """
                                    SELECT v.fabric_id
                                    FROM sales_products sp
                                    LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                                    WHERE sp.id=%s
                                    LIMIT 1
                                    """,
                                    (item_id,),
                                )
                                rowv = cur.fetchone() or {}
                                current_variant_fabric_id = self._parse_int(rowv.get('fabric_id')) or 0
                    except Exception:
                        current_variant_fabric_id = 0

                    bundle = self._derive_sales_order_links_bundle(conn, sku_family_id_input, links)
                    sku_family_id = sku_family_id_input or bundle.get('sku_family_id')
                    if not sku_family_id:
                        return self.send_json({'status': 'error', 'message': '无法根据下单SKU推断归属货号'}, start_response)

                    sku_family_code = (bundle.get('sku_family_code') or '').strip()

                    parent_id = None
                    parent_shop_id = None
                    if parent_code:
                        with conn.cursor() as cur:
                            cur.execute("SELECT id, shop_id FROM sales_parents WHERE parent_code=%s AND shop_id=%s LIMIT 1", (parent_code, shop_id_input))
                            row = cur.fetchone()
                            if row:
                                parent_id = row['id']
                                parent_shop_id = row.get('shop_id')
                                if (not parent_shop_id) and shop_id_input:
                                    cur.execute("UPDATE sales_parents SET shop_id=%s WHERE id=%s", (shop_id_input, parent_id))
                                    parent_shop_id = shop_id_input
                            else:
                                cur.execute(
                                    """
                                    INSERT INTO sales_parents (parent_code, shop_id, sku_marker)
                                    VALUES (%s, %s, %s)
                                    """,
                                    (parent_code, shop_id_input, parent_sku_marker)
                                )
                                parent_id = cur.lastrowid
                                parent_shop_id = shop_id_input
                    final_shop_id = parent_shop_id if parent_id else shop_id_input
                    if not final_shop_id:
                        return self.send_json({'status': 'error', 'message': 'Missing required field: shop_id'}, start_response)

                    auto_fabric = (bundle.get('fabric') or '').strip()
                    auto_spec_name = (bundle.get('spec_name') or '').strip()
                    auto_platform_sku = (bundle.get('platform_sku') or '').strip()
                    final_fabric = (data.get('fabric') or '').strip() or auto_fabric
                    final_spec_name = (data.get('spec_name') or '').strip() or auto_spec_name
                    resolved_fabric_id = fabric_id_input or None
                    # If UI didn't provide fabric_id and lookup failed, fall back to current variant fabric_id
                    if (not resolved_fabric_id) and current_variant_fabric_id:
                        resolved_fabric_id = int(current_variant_fabric_id)
                    if not resolved_fabric_id and final_fabric:
                        try:
                            with conn.cursor() as fcur:
                                fcur.execute("SELECT id FROM fabric_materials WHERE fabric_code=%s LIMIT 1", (self._code_before_dash(final_fabric),))
                                frow = fcur.fetchone() or {}
                                resolved_fabric_id = self._parse_int(frow.get('id')) or None
                        except Exception:
                            resolved_fabric_id = None
                    if not resolved_fabric_id:
                        return self.send_json({'status': 'error', 'message': '面料不能为空：请在“面料”下拉中选择面料（fabric_id 必填）'}, start_response)
                    variant_id = self._get_or_create_sales_variant(conn, sku_family_id, final_spec_name, final_fabric, fabric_id=resolved_fabric_id)
                    
                    # 如果没有手动编辑，使用自动生成的platform_sku；否则使用手动输入的
                    if manual_platform_sku:
                        platform_sku = platform_sku_manual
                    else:
                        platform_sku = auto_platform_sku or self._build_sales_platform_sku(sku_family_code, final_spec_name, final_fabric)
                    
                    if not platform_sku:
                        return self.send_json({'status': 'error', 'message': '无法生成销售平台SKU，请手动输入'}, start_response)

                    with conn.cursor() as cur:
                        has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                        has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                        fabric_join = "LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if has_fabric_id else ""
                        if has_fabric_id and has_fabric_text:
                            fabric_select = "COALESCE(fm.fabric_code, v.fabric) AS fabric"
                        elif has_fabric_id:
                            fabric_select = "fm.fabric_code AS fabric"
                        else:
                            fabric_select = ("v.fabric AS fabric" if has_fabric_text else "'' AS fabric")
                        cur.execute(
                            f"""
                            SELECT v.spec_name, {fabric_select}
                            FROM sales_products sp
                            LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                            {fabric_join}
                            WHERE sp.id=%s
                            """,
                            (item_id,)
                        )
                        current_row = cur.fetchone() or {}
                    old_spec_name = (current_row.get('spec_name') or '').strip()
                    old_fabric = (current_row.get('fabric') or '').strip()
                    spec_or_fabric_changed = (old_spec_name != (final_spec_name or '').strip()) or (old_fabric != (final_fabric or '').strip())
                    if spec_or_fabric_changed and not confirm_new_variant_folder:
                        return self.send_json({'status': 'error', 'message': '修改规格名称或面料将新建主图文件夹，请二次确认后重试'}, start_response)
                    
                    with conn.cursor() as cur:
                        platform_name = self._load_shop_platform_type_name(cur, final_shop_id)
                    product_link = self._resolve_sales_product_link(
                        platform_name, child_code, data.get('product_link')
                    )

                    with conn.cursor() as cur:
                        update_fields = [
                            "platform_sku=%s",
                            "product_status=%s",
                            "variant_id=%s",
                            "parent_id=%s",
                            "child_code=%s",
                        ]
                        update_values = [platform_sku, product_status, variant_id, parent_id, child_code]
                        if self._table_has_column(conn, 'sales_products', 'product_link'):
                            update_fields.append("product_link=%s")
                            update_values.append(product_link)
                        if self._table_has_column(conn, 'sales_products', 'gtin'):
                            update_fields.append("gtin=%s")
                            update_values.append(gtin)
                        if self._table_has_column(conn, 'sales_products', 'upc'):
                            update_fields.append("upc=%s")
                            update_values.append(upc)
                        update_fields.append("sale_price_usd=%s")
                        update_values.append(sale_price_usd)
                        if self._table_has_column(conn, 'sales_products', 'notes'):
                            update_fields.append("notes=%s")
                            update_values.append(notes)
                        if sp_has_shop_col:
                            update_fields.insert(0, "shop_id=%s")
                            update_values.insert(0, final_shop_id)
                        update_values.append(item_id)
                        cur.execute(
                            f"UPDATE sales_products SET {', '.join(update_fields)} WHERE id=%s",
                            update_values
                        )
                    self._replace_sales_variant_order_links(conn, variant_id, links)
                    if spec_or_fabric_changed:
                        fabric_folder_part = self._resolve_fabric_folder_part(conn, resolved_fabric_id, final_fabric)
                        self._ensure_listing_sales_variant_folder(sku_family_code, final_spec_name, fabric_folder_part)
                return self.send_json({'status': 'success'}, start_response)

            if method == 'PATCH':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                product_status = (data.get('product_status') or '').strip().lower()
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                if product_status not in ('enabled', 'retained', 'discarded'):
                    return self.send_json({'status': 'error', 'message': 'Invalid product_status'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("UPDATE sales_products SET product_status=%s WHERE id=%s AND product_status<>%s", (product_status, item_id, product_status))
                        changed = int(cur.rowcount or 0)
                return self.send_json({'status': 'success', 'changed': changed}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM sales_products WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '销售平台SKU已存在或关联数据无效'}, start_response)
            print("Sales product API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _ensure_listing_sales_variant_folder(self, sku_family, spec_name, fabric_code):
        sku_name = (sku_family or '').strip()
        if not sku_name:
            return
        self._ensure_listing_sku_folder(sku_name)
        base_folder = self._ensure_listing_folder()
        sku_folder = os.path.join(base_folder, self._safe_fsencode(sku_name))
        main_folder = os.path.join(sku_folder, self._safe_fsencode('主图'))
        if not os.path.exists(main_folder):
            os.makedirs(main_folder, exist_ok=True)

        spec_part = (spec_name or '').strip().replace('/', '-').replace('\\', '-')
        # Folder naming rule: 规格-面料英文名（规格可为空；面料必填；面料段 fallback 同 _resolve_fabric_folder_part）
        fabric_part = (fabric_code or '').strip().replace('/', '-').replace('\\', '-')
        if not fabric_part:
            fabric_part = self._code_before_dash(fabric_code).replace('/', '-').replace('\\', '-')
        variant_folder_name = self._sales_variant_subfolder_display_name(spec_part, fabric_part)
        if not variant_folder_name:
            return
        variant_folder = os.path.join(main_folder, self._safe_fsencode(variant_folder_name))
        if not os.path.exists(variant_folder):
            os.makedirs(variant_folder, exist_ok=True)

    def _resolve_fabric_folder_part(self, conn, fabric_id=None, fabric_text=''):
        """Return folder-safe fabric part; prefer fabric_name_en, fallback to legacy fabric code/text."""
        name_en = ''
        fid = self._parse_int(fabric_id) or 0
        if fid:
            try:
                with conn.cursor() as cur:
                    cur.execute("SELECT fabric_name_en FROM fabric_materials WHERE id=%s LIMIT 1", (fid,))
                    row = cur.fetchone() or {}
                    name_en = str(row.get('fabric_name_en') or '').strip()
            except Exception:
                name_en = ''
        if name_en:
            return name_en.replace('/', '-').replace('\\', '-').strip()
        return self._code_before_dash(fabric_text).replace('/', '-').replace('\\', '-').strip()

    def _get_sales_product_image_assets_folder(self):
        folder = self._join_resources('『销售产品图片』/assets')
        if not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
        return folder

    def _sha256_hex(self, data_bytes):
        return hashlib.sha256(data_bytes or b'').hexdigest()

    def _guess_image_ext(self, filename, content):
        ext = os.path.splitext(os.path.basename(filename or ''))[1].lower()
        if ext in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.tif', '.tiff'):
            return ext
        if content.startswith(b'\xff\xd8\xff'):
            return '.jpg'
        if content.startswith(b'\x89PNG'):
            return '.png'
        if content.startswith(b'GIF8'):
            return '.gif'
        if content.startswith(b'RIFF') and b'WEBP' in content[:16]:
            return '.webp'
        return '.jpg'

    def _unc_share_key(self, path_text):
        """
        Return UNC share key like '\\\\server\\share' (case-insensitive).
        If not UNC, return ''.
        """
        try:
            # IMPORTANT: many paths in this project are bytes (RESOURCES_PATH_BYTES).
            # Convert bytes -> str first; never use str(b'...') which breaks UNC parsing.
            if isinstance(path_text, (bytes, bytearray)):
                p = os.fsdecode(path_text)
            else:
                p = str(path_text or '')
            p = p.strip()
        except Exception:
            return ''
        if not p:
            return ''
        # Normalize separators and UNC prefix variants
        p = p.replace('/', '\\')
        if p.startswith('\\\\?\\UNC\\'):
            p = '\\\\' + p[len('\\\\?\\UNC\\'):]
        if p.startswith('//'):
            p = '\\\\' + p.lstrip('/').replace('/', '\\')
        if not p.startswith('\\\\'):
            return ''
        # \\server\share\rest...
        parts = p.lstrip('\\').split('\\')
        if len(parts) < 2:
            return ''
        server = parts[0].strip()
        share = parts[1].strip()
        if not (server and share):
            return ''
        return ('\\\\' + server + '\\' + share).lower()

    def _normalize_nas_abs_path(self, input_path):
        """
        Normalize user-provided paths (Windows UNC or POSIX) into NAS-local absolute paths.
        The backend runs on NAS Linux, so Windows UNC like:
          \\DiskStation\公共文件SITJOY\『上架资源』\...
        must be mapped to the local RESOURCES_PATH:
          /volumeX/公共文件SITJOY/『上架资源』/...
        """
        # IMPORTANT: use bytes paths internally to avoid lossy unicode replacement (�)
        # when filenames contain non-UTF8 bytes. This matches the gallery/file-management approach.
        try:
            raw = str(input_path or '').strip()
        except Exception:
            raw = ''
        if not raw:
            return b''

        # If it's already a local absolute path on NAS, keep it (as bytes).
        if raw.startswith('/volume') or raw.startswith('/'):
            return self._safe_fsencode(os.path.normpath(raw))

        try:
            from app import RESOURCES_PATH_BYTES  # bytes: /volumeX/公共文件SITJOY/『上架资源』
        except Exception:
            RESOURCES_PATH_BYTES = b''

        # Normalize for matching
        norm = raw.replace('/', '\\')
        lower_norm = norm.lower()

        # Extract the part after 『上架资源』 (works even if server/share differs)
        marker = '『上架资源』'
        idx = norm.find(marker)
        rest = ''
        if idx >= 0:
            rest = norm[idx + len(marker):].lstrip('\\/')
        else:
            # If marker not found, treat as relative to resources root (best-effort)
            rest = norm.lstrip('\\/')

        # Join with RESOURCES_PATH_BYTES safely
        rest_posix = rest.replace('\\', '/')
        rest_bytes = self._safe_fsencode(rest_posix)
        base = RESOURCES_PATH_BYTES.rstrip(b'/').rstrip(b'\\')
        if base:
            def _join(b):
                try:
                    return os.path.normpath(b + b'/' + rest_bytes.lstrip(b'/'))
                except Exception:
                    try:
                        return os.path.join(b, rest_bytes.lstrip(b'/'))
                    except Exception:
                        return b

            # First try the primary resources base
            candidate = _join(base)
            if os.path.exists(candidate):
                return candidate

            # If primary base doesn't contain the file, try other /volumeN bases.
            try:
                import re
                m = re.match(br'^/volume(\d+)/(.*)$', base)
            except Exception:
                m = None
            suffix = None
            if m:
                suffix = m.group(2)  # bytes after /volumeN/
            else:
                # fallback: keep everything after first slash
                try:
                    suffix = base.lstrip(b'/')
                except Exception:
                    suffix = None

            volumes = []
            try:
                for name in os.listdir('/'):
                    if str(name).startswith('volume'):
                        volumes.append(str(name))
                volumes.sort()
            except Exception:
                volumes = ['volume3', 'volume1']

            for vol in volumes:
                try:
                    vol_b = self._safe_fsencode('/' + vol + '/')
                    alt_base = vol_b.rstrip(b'/') + b'/' + (suffix or b'')
                    alt_base = alt_base.rstrip(b'/')
                    alt_candidate = _join(alt_base)
                    if os.path.exists(alt_candidate):
                        return alt_candidate
                except Exception:
                    continue

            # As last resort return the primary candidate (even if missing) for error reporting.
            return candidate

        # Absolute fallback: try interpreting UNC as plain path text
        return self._safe_fsencode(os.path.normpath(raw))

    def _normalize_nas_abs_path_bytes(self, raw_bytes):
        """
        Bytes-first variant of _normalize_nas_abs_path.
        Accepts original bytes (e.g. from base64) to avoid JSON/unicode corruption.
        Returns NAS-local absolute path in bytes.
        """
        if not raw_bytes:
            return b''
        if not isinstance(raw_bytes, (bytes, bytearray)):
            try:
                raw_bytes = self._safe_fsencode(raw_bytes)
            except Exception:
                raw_bytes = b''
        raw_bytes = bytes(raw_bytes)

        # If it's already a local absolute path on NAS, keep it.
        if raw_bytes.startswith(b'/'):
            try:
                return os.path.normpath(raw_bytes)
            except Exception:
                return raw_bytes

        try:
            from app import RESOURCES_PATH_BYTES  # bytes
        except Exception:
            RESOURCES_PATH_BYTES = b''

        # Normalize separators (UNC) and find marker bytes of 『上架资源』
        try:
            norm = raw_bytes.replace(b'/', b'\\')
        except Exception:
            norm = raw_bytes

        marker = '『上架资源』'.encode('utf-8', errors='surrogatepass')
        idx = norm.find(marker)
        if idx >= 0:
            rest = norm[idx + len(marker):].lstrip(b'\\/')
        else:
            rest = norm.lstrip(b'\\/')
        rest = rest.replace(b'\\', b'/')
        base = (RESOURCES_PATH_BYTES or b'').rstrip(b'/').rstrip(b'\\')
        if base:
            try:
                candidate = os.path.normpath(base + b'/' + rest.lstrip(b'/'))
            except Exception:
                candidate = os.path.join(base, rest.lstrip(b'/'))
            # Reuse the multi-volume probing from _normalize_nas_abs_path by calling it with decoded text
            # only as a fallback (candidate is bytes and safe).
            if os.path.exists(candidate):
                return candidate
            # If missing, try other volumes
            try:
                import re
                m = re.match(br'^/volume(\d+)/(.*)$', base)
                suffix = m.group(2) if m else base.lstrip(b'/')
            except Exception:
                suffix = base.lstrip(b'/')
            try:
                vols = [str(n) for n in os.listdir('/') if str(n).startswith('volume')]
                vols.sort()
            except Exception:
                vols = ['volume3', 'volume1']
            for vol in vols:
                try:
                    alt_base = b'/' + vol.encode('ascii', errors='ignore') + b'/' + (suffix or b'')
                    alt_base = alt_base.rstrip(b'/')
                    try:
                        alt_candidate = os.path.normpath(alt_base + b'/' + rest.lstrip(b'/'))
                    except Exception:
                        alt_candidate = os.path.join(alt_base, rest.lstrip(b'/'))
                    if os.path.exists(alt_candidate):
                        return alt_candidate
                except Exception:
                    continue
            return candidate
        return raw_bytes

    def _chunk_text(self, text, chunk_size=120):
        try:
            s = str(text or '')
        except Exception:
            s = ''
        if not s:
            return []
        try:
            n = int(chunk_size or 120)
        except Exception:
            n = 120
        n = max(40, min(400, n))
        return [s[i:i+n] for i in range(0, len(s), n)]

    def _iter_resources_volume_bases(self):
        """
        Return possible RESOURCES_PATH_BYTES bases across /volumeN.
        """
        try:
            from app import RESOURCES_PATH_BYTES
        except Exception:
            RESOURCES_PATH_BYTES = b''
        base = bytes(RESOURCES_PATH_BYTES or b'')
        if not base:
            return []
        bases = []
        # include current base first
        bases.append(base.rstrip(b'/'))
        # derive suffix after /volumeN/
        try:
            import re
            m = re.match(br'^/volume(\d+)/(.*)$', base)
            suffix = m.group(2) if m else base.lstrip(b'/')
        except Exception:
            suffix = base.lstrip(b'/')
        # enumerate /volumeN folders
        try:
            vols = [str(n) for n in os.listdir('/') if str(n).startswith('volume')]
            vols.sort()
        except Exception:
            vols = []
        for v in vols:
            try:
                v_b = v.encode('ascii', errors='ignore')
                b = (b'/' + v_b + b'/' + (suffix or b'')).rstrip(b'/')
                if b not in bases:
                    bases.append(b)
            except Exception:
                continue
        return bases

    def _extract_resources_relative_bytes(self, source_path_text):
        """
        Convert user provided source_path (UNC/relative) into bytes relative to 『上架资源』 root.
        """
        try:
            raw = str(source_path_text or '').strip()
        except Exception:
            raw = ''
        if not raw:
            return b''
        norm = raw.replace('/', '\\')
        marker = '『上架资源』'
        idx = norm.find(marker)
        if idx >= 0:
            rest = norm[idx + len(marker):].lstrip('\\/')
        else:
            rest = norm.lstrip('\\/')
        rest_posix = rest.replace('\\', '/').lstrip('/')
        return self._safe_fsencode(rest_posix)

    def _try_find_file_by_basename_under_sku(self, rel_path_text):
        """
        Fallback search:
        If the user-provided relative path doesn't match due to encoding/dir-name mismatch,
        try locating the file by its basename under 『上架资源』/<sku_family>/ recursively.
        Returns (found_abs_bytes, debug_list).
        """
        debug = []
        try:
            raw = str(rel_path_text or '').strip().replace('\\', '/')
        except Exception:
            return (b'', debug)
        if not raw:
            return (b'', debug)
        parts = [p for p in raw.split('/') if p]
        if len(parts) < 2:
            return (b'', debug)
        sku_family = parts[0]
        filename = parts[-1]
        if not filename or '.' not in filename:
            return (b'', debug)
        try:
            from app import RESOURCES_PATH_BYTES
        except Exception:
            RESOURCES_PATH_BYTES = b''
        if not RESOURCES_PATH_BYTES:
            return (b'', debug)

        sku_root = os.path.normpath(RESOURCES_PATH_BYTES.rstrip(b'/') + b'/' + self._safe_fsencode(sku_family))
        if not os.path.exists(sku_root):
            debug.append({'sku_root': self._safe_fsdecode(sku_root), 'exists': False})
            return (b'', debug)
        debug.append({'sku_root': self._safe_fsdecode(sku_root), 'exists': True})
        target_name = self._safe_fsencode(filename)
        try:
            for root, _dirs, files in os.walk(sku_root):
                # root/files are bytes when sku_root is bytes
                for f in files or []:
                    try:
                        if f == target_name:
                            found = os.path.join(root, f)
                            debug.append({'found': self._safe_fsdecode(found)})
                            return (found, debug)
                    except Exception:
                        continue
        except Exception as e:
            debug.append({'walk_error': str(e)[:160]})
        return (b'', debug)



    def handle_sales_product_main_images_import_by_path_api(self, environ, method, start_response):
        """
        从 NAS 路径导入销售产品主图。自动检测文件位置，优先移动以优化性能。
        POST /api/sales-product-main-images-import-by-path
        入参：sales_product_id, source_path, image_type_name(可选，默认文字卖点图)
        """
        try:
            if method != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)

            data = self._read_json_body(environ)
            sales_product_id = self._parse_int(data.get('sales_product_id'))
            variant_id_param = self._parse_int(data.get('variant_id'))
            source_path_text = str(data.get('source_path') or '').strip()
            source_path_b64 = str(data.get('source_path_b64') or '').strip()
            source_paths_b64 = data.get('source_paths_b64') or data.get('source_paths') or []
            image_type_name = str(data.get('image_type_name') or '').strip() or '文字卖点图'
            delete_source = bool(data.get('delete_source'))  # optional: try delete source after successful commit (best-effort)
            require_move = str(data.get('require_move') or '').strip().lower() in ('1', 'true', 'yes', 'on')
            debug_move = str(data.get('debug_move') or '').strip().lower() in ('1', 'true', 'yes', 'on')
            prompt_duplicate = str(data.get('prompt_duplicate') or '').strip().lower() in ('1', 'true', 'yes', 'on')
            allow_duplicate = str(data.get('allow_duplicate') or '').strip().lower() in ('1', 'true', 'yes', 'on')

            if not sales_product_id and not variant_id_param:
                return self.send_json({'status': 'error', 'message': 'Missing sales_product_id 或 variant_id'}, start_response)
            if (not source_path_text) and (not source_path_b64) and (not source_paths_b64):
                return self.send_json({'status': 'error', 'message': 'Missing source_path'}, start_response)

            # Multi-select: base64 = 文件系统原始字节；全程用 bytes 绝对路径，避免 str/fsdecode 产生 U+FFFD 再写入目标文件名
            source_files_b = []
            if isinstance(source_paths_b64, (list, tuple)) and source_paths_b64:
                for b64 in list(source_paths_b64)[:500]:
                    try:
                        raw = base64.b64decode(str(b64 or '').strip())
                        p_b = self._normalize_nas_abs_path_bytes(raw)
                    except Exception:
                        continue
                    try:
                        if p_b and os.path.isfile(p_b) and self._is_image_name(os.path.basename(p_b)):
                            source_files_b.append(p_b)
                    except Exception:
                        continue
                source_files_b = sorted(set(source_files_b))

            # Backend runs on NAS (Linux). If user passes Windows UNC path, map it to NAS-local path.
            # Prefer bytes-safe base64 path if provided (avoids unicode corruption in JSON)
            source_path_b = b''
            if (not source_files_b) and source_path_b64:
                try:
                    raw_bytes = base64.b64decode(source_path_b64)
                    source_path_b = self._normalize_nas_abs_path_bytes(raw_bytes)
                except Exception:
                    return self.send_json({
                        'status': 'error',
                        'message': 'Invalid source_path_b64（不是合法Base64）',
                        'source_path_b64_input': source_path_b64[:1200],
                        'source_path_b64_input_chunks': self._chunk_text(source_path_b64, 120),
                    }, start_response)
            if (not source_files_b) and (not source_path_b):
                source_path_b = self._normalize_nas_abs_path(source_path_text)
            # If missing, try to locate the file under other /volumeN resources bases.
            probe = []
            if (not source_files_b) and source_path_b and (not os.path.exists(source_path_b)):
                try:
                    rel_b = self._extract_resources_relative_bytes(source_path_text)
                    if rel_b:
                        for base_b in self._iter_resources_volume_bases():
                            try:
                                cand = os.path.normpath(base_b.rstrip(b'/') + b'/' + rel_b.lstrip(b'/'))
                            except Exception:
                                cand = os.path.join(base_b, rel_b.lstrip(b'/'))
                            exists = bool(os.path.exists(cand))
                            probe.append({'candidate': self._safe_fsdecode(cand), 'exists': exists})
                            if exists:
                                source_path_b = cand
                                break
                except Exception:
                    probe = probe or []

            # Extra fallback: search by basename under sku folder (avoids directory-name encoding mismatches)
            search_debug = []
            if (not source_files_b) and source_path_b and (not os.path.exists(source_path_b)) and source_path_text:
                try:
                    found_b, search_debug = self._try_find_file_by_basename_under_sku(source_path_text)
                    if found_b and os.path.exists(found_b):
                        source_path_b = found_b
                        probe.append({'candidate': self._safe_fsdecode(found_b), 'exists': True, 'mode': 'basename_search'})
                except Exception:
                    search_debug = search_debug or []

            if (not source_files_b) and ((not source_path_b) or (not os.path.exists(source_path_b))):
                return self.send_json({
                    'status': 'error',
                    'message': '源路径不存在',
                    'source_path': self._safe_fsdecode(source_path_b) if source_path_b else '',
                    'source_path_b64': (self._b64_from_fs(source_path_b) if source_path_b else ''),
                    'source_path_b64_input': (source_path_b64[:1200] if source_path_b64 else ''),
                    'source_path_b64_input_chunks': (self._chunk_text(source_path_b64, 120) if source_path_b64 else []),
                    'probe_candidates': probe,
                    'basename_search_debug': search_debug,
                }, start_response)

            source_path = ''
            if not source_files_b:
                source_files_b = []
                source_path = self._safe_fsdecode(source_path_b)
                if os.path.isfile(source_path_b):
                    if self._is_image_name(os.path.basename(source_path_b)):
                        source_files_b = [source_path_b]
                else:
                    try:
                        for name in os.listdir(source_path_b):
                            abs_file_b = os.path.join(source_path_b, name)
                            if os.path.isfile(abs_file_b) and self._is_image_name(name):
                                source_files_b.append(abs_file_b)
                    except Exception:
                        source_files_b = []
                source_files_b = sorted(set(source_files_b))
            if not source_path:
                source_path = 'multi-select'
            if not source_files_b:
                return self.send_json({'status': 'error', 'message': '源路径下无图片文件'}, start_response)

            with self._get_db_connection() as conn:
                image_type_id = self._get_image_type_id_by_name(conn, image_type_name)
                if not image_type_id:
                    return self.send_json({'status': 'error', 'message': f'未知图片类型: {image_type_name}'}, start_response)

                if variant_id_param and not sales_product_id:
                    try:
                        folder_info = self._resolve_sales_variant_folder_by_variant_id(variant_id_param, ensure_folder=True)
                    except Exception as e:
                        return self.send_json({'status': 'error', 'message': str(e)}, start_response)
                    start_sort = self._get_variant_image_sort_start(conn, variant_id_param)
                    variant_id = int(variant_id_param)
                else:
                    start_sort = self._get_sales_product_image_sort_start(conn, sales_product_id)
                    folder_info = self._resolve_sales_product_variant_folder(sales_product_id, ensure_folder=True)
                    variant_id = 0
                    try:
                        with conn.cursor() as cur:
                            cur.execute("SELECT variant_id FROM sales_products WHERE id=%s", (sales_product_id,))
                            row = cur.fetchone() or {}
                            variant_id = self._parse_int(row.get('variant_id')) or 0
                    except Exception:
                        variant_id = 0
                created_assets = 0
                moved_count = 0
                copied_count = 0
                linked_count = 0
                items = []
                # NAS backend: use bytes paths end-to-end to avoid str/bytes mixing.
                target_folder_abs = folder_info.get('folder_path')
                if not target_folder_abs:
                    return self.send_json({'status': 'error', 'message': '无法定位主图文件夹，请确认货号与面料完整（面料必填；规格可为空）'}, start_response)
                if isinstance(target_folder_abs, str):
                    target_folder_abs = self._safe_fsencode(target_folder_abs)
                if not os.path.exists(target_folder_abs):
                    return self.send_json({'status': 'error', 'message': '无法定位主图文件夹，请确认货号与面料完整（面料必填；规格可为空）'}, start_response)

                # ---- Stage files first (atomic batch) ----
                staged_moves = []   # [(src, tmp)] source moved to tmp; tmp still exists (not yet promoted)
                rollback_restore_pairs = []   # (final_abs, orig_src): bytes; DB rollback / early exit must move final -> orig
                rollback_unlink_only = []       # final_abs bytes: copy-only staging; rollback only deletes this duplicate
                db_new_assets = []  # [{sha256, storage_path, filename, ext, file_size}]
                reuse_assets = []   # [{asset_id, sha256, idx, source_file}]
                move_failures = []  # [{src, reason}]
                bound_source_files = []  # sources actually linked in this request (avoid deleting skipped files)

                def _undo_import_by_path_disk():
                    """Revert on-disk staging. Never unlink a path that is the sole remaining copy after a move-from-source."""
                    for final_b, src_b in reversed(rollback_restore_pairs or []):
                        try:
                            if final_b and src_b and os.path.exists(final_b):
                                try:
                                    os.replace(final_b, src_b)
                                except Exception:
                                    shutil.move(final_b, src_b)
                        except Exception:
                            pass
                    for src_b, tmp_b in reversed(staged_moves or []):
                        try:
                            if tmp_b and os.path.exists(tmp_b):
                                try:
                                    os.replace(tmp_b, src_b)
                                except Exception:
                                    shutil.move(tmp_b, src_b)
                        except Exception:
                            pass
                    for p in reversed(rollback_unlink_only or []):
                        try:
                            self._safe_unlink(p)
                        except Exception:
                            pass

                for idx, source_file_b in enumerate(source_files_b, start=1):
                    filename, stem_plain, _ext_from_base = self._display_name_from_abs_path_b(source_file_b)
                    try:
                        with open(source_file_b, 'rb') as f:
                            content = f.read()
                        if not content:
                            continue
                        sha256 = self._sha256_hex(content)
                    except Exception:
                        continue

                    asset = self._find_image_asset_by_sha256(conn, sha256)
                    if asset:
                        reuse_assets.append({
                            'asset_id': asset.get('id'),
                            'sha256': sha256,
                            'idx': idx,
                            'source_file': self._safe_fsdecode(source_file_b),
                        })
                    else:
                        ext = self._guess_image_ext(filename, content)
                        fabric_for_name = str(folder_info.get('fabric_name_en') or '').strip()
                        if not fabric_for_name:
                            fabric_for_name = str(folder_info.get('fabric_folder_part') or '').strip()
                        base_combined = self._sales_nas_import_recommended_basename(
                            fabric_for_name, image_type_name, stem_plain, sha256
                        )
                        final_name = self._next_available_filename(target_folder_abs, f"{base_combined}{ext}")
                        abs_path = os.path.join(target_folder_abs, self._safe_fsencode(final_name))

                        # Stage: try move to a temp file first so we can rollback on DB failure.
                        tmp_abs = abs_path + self._safe_fsencode(f".__tmp__{int(time.time()*1000)}_{idx}")
                        wrote_final = False
                        moved_source = False
                        try:
                            os.replace(source_file_b, tmp_abs)
                            staged_moves.append((source_file_b, tmp_abs))
                            moved_source = True
                            moved_count += 1
                        except Exception as e_move1:
                            # Fallback 1: try shutil.move (copy+delete)
                            try:
                                shutil.move(source_file_b, tmp_abs)
                                staged_moves.append((source_file_b, tmp_abs))
                                moved_source = True
                                moved_count += 1
                            except Exception as e_move2:
                                # Fallback 2: write temp by bytes (keep source intact)
                                try:
                                    with open(tmp_abs, 'wb') as f:
                                        f.write(content or b'')
                                    copied_count += 1
                                except Exception:
                                    # Can't persist => skip this file
                                    self._safe_unlink(tmp_abs)
                                    continue
                                move_failures.append({
                                    'src': self._safe_fsdecode(source_file_b),
                                    'reason': (str(e_move2)[:160] or str(e_move1)[:160]),
                                })

                        if require_move and copied_count > 0:
                            # We copied at least one file in this batch; enforce "must move" semantics.
                            _undo_import_by_path_disk()
                            try:
                                self._safe_unlink(tmp_abs)
                            except Exception:
                                pass
                            return self.send_json({
                                'status': 'error',
                                'message': '要求移动(require_move=1)但当前路径无法移动（可能是不同 share 或无删除权限）。',
                                'source_path': source_path,
                                'target_folder': target_folder_abs,
                                'move_failures': move_failures[:5],
                            }, start_response)

                        # Promote temp -> final path
                        try:
                            os.replace(tmp_abs, abs_path)
                            wrote_final = True
                        except Exception:
                            # Never unlink tmp while it may be the only copy after a move-from-source.
                            if moved_source:
                                try:
                                    if os.path.exists(tmp_abs):
                                        try:
                                            os.replace(tmp_abs, source_file_b)
                                        except Exception:
                                            shutil.move(tmp_abs, source_file_b)
                                except Exception:
                                    if os.path.exists(tmp_abs):
                                        try:
                                            self._move_file_to_listing_recycle_bin(tmp_abs, 'promote_failed')
                                        except Exception:
                                            pass
                            else:
                                self._safe_unlink(tmp_abs)
                            staged_moves[:] = [(s, t) for (s, t) in staged_moves if t != tmp_abs]
                            continue

                        if not wrote_final or not os.path.exists(abs_path):
                            continue

                        # _storage_path_from_abs expects a resources-absolute path in the same type
                        # as resources root (bytes). abs_path here is str, so encode safely first.
                        storage_path = self._storage_path_from_abs(abs_path)
                        if moved_source:
                            staged_moves[:] = [(s, t) for (s, t) in staged_moves if t != tmp_abs]
                            rollback_restore_pairs.append((abs_path, source_file_b))
                        else:
                            rollback_unlink_only.append(abs_path)
                        db_new_assets.append({
                            'sha256': sha256,
                            'storage_path': storage_path,
                            'filename': filename,
                            'ext': ext,
                            'file_size': len(content or b''),
                            'idx': idx,
                        })

                    sort_order = start_sort + idx
                    bound_source_files.append(self._safe_fsdecode(source_file_b))
                    # Defer DB mapping inserts until after all files staged successfully (atomic batch)
                    items.append({'filename': filename, 'sha256': sha256[:12], 'sort_order': sort_order, 'idx': idx, 'sha256_full': sha256})

                if prompt_duplicate and reuse_assets and not allow_duplicate:
                    dups = []
                    for r in reuse_assets[:200]:
                        try:
                            aid = int(r.get('asset_id') or 0)
                            with conn.cursor() as cur:
                                cur.execute("SELECT storage_path FROM image_assets WHERE id=%s LIMIT 1", (aid,))
                                row = cur.fetchone() or {}
                            dups.append({
                                'source_file': str(r.get('source_file') or ''),
                                'sha256': str(r.get('sha256') or ''),
                                'image_asset_id': aid,
                                'storage_path': (row.get('storage_path') or '') if row else '',
                            })
                        except Exception:
                            continue
                    # New files may already have been moved onto disk before we detect duplicates in-batch.
                    _undo_import_by_path_disk()
                    return self.send_json({
                        'status': 'duplicate',
                        'message': '检测到重复图片（sha256 已存在），是否确认复用已有图片并继续导入？',
                        'duplicate_count': len(reuse_assets),
                        'duplicates': dups,
                        'file_count': len(items),
                    }, start_response)

                # If nothing to process, exit early
                if not db_new_assets and not reuse_assets:
                    return self.send_json({'status': 'error', 'message': '未检测到可导入的图片（可能均为空/不支持）'}, start_response)

                # ---- Transaction: write DB for the whole batch ----
                created_asset_ids = []
                try:
                    self._tx_begin(conn)

                    # Insert new assets
                    sha_to_id = {}
                    for rec in db_new_assets:
                        with conn.cursor() as cur:
                            aid = self._insert_image_asset_dynamic(
                                conn,
                                cur,
                                {
                                    'sha256': rec['sha256'],
                                    'storage_path': rec['storage_path'],
                                    'filename': rec['filename'],
                                    'ext': rec['ext'],
                                    'file_size': rec['file_size'],
                                    'image_type_id': image_type_id,
                                },
                            )
                        sha_to_id[rec['sha256']] = aid
                        created_asset_ids.append((aid, rec['storage_path']))
                        created_assets += 1

                    # Map reuse asset ids (already existed)
                    for r in reuse_assets:
                        aid = int(r.get('asset_id') or 0)
                        sha = str(r.get('sha256') or '').strip()
                        if aid and sha:
                            sha_to_id.setdefault(sha, aid)

                    # Insert mappings
                    for row in items:
                        sort_order = row.get('sort_order')
                        sha = row.get('sha256_full')
                        asset_id = sha_to_id.get(sha)
                        if not asset_id:
                            continue
                        with conn.cursor() as cur:
                            self._execute_sku_mapping_upsert(
                                conn, cur, asset_id, sort_order, image_type_id, variant_id, sales_product_id, None
                            )
                        linked_count += 1

                    self._tx_commit(conn)
                except Exception as e:
                    self._tx_rollback(conn)
                    # Move finals back to original paths for move-from-source; unlink copy-only duplicates.
                    # (Historically unlink(abs) + stale staged_moves(tmp) destroyed the only on-disk copy.)
                    _undo_import_by_path_disk()
                    return self.send_json({'status': 'error', 'message': f'导入失败，已回滚：{str(e)}'}, start_response)

                # Optional: after commit, delete source files for those that were copied (best-effort).
                # We only attempt this when explicitly requested, to avoid accidental data loss.
                deleted_source_count = 0
                if delete_source:
                    try:
                        # Any source that wasn't moved into staged_moves is still at source_file.
                        moved_sources = set()
                        for src, _tmp in staged_moves:
                            moved_sources.add(src)
                        for src_b_del in source_files_b:
                            if src_b_del in moved_sources:
                                continue
                            try:
                                if os.path.exists(src_b_del):
                                    os.remove(src_b_del)
                                    deleted_source_count += 1
                            except Exception:
                                pass
                    except Exception:
                        deleted_source_count = deleted_source_count or 0

                # After commit: apply rehome rules best-effort (new rows + reused rows may change ref counts)
                all_aids = []
                for aid, _ in created_asset_ids:
                    if aid:
                        all_aids.append(int(aid))
                for r in reuse_assets:
                    aid = int(r.get('asset_id') or 0)
                    if aid:
                        all_aids.append(aid)
                all_aids = sorted(set(all_aids))
                try:
                    for aid in all_aids:
                        self._rehome_image_asset_if_needed(conn, aid)
                except Exception:
                    pass

                # Always clean up manual staging duplicates left on disk (esp. reuse/sha256 hits and copy-fallback paths)
                import_cleanup = {'cleaned': 0, 'skipped_samefile': 0, 'failures': 0}
                try:
                    import_cleanup = self._cleanup_import_by_path_sources(conn, bound_source_files, all_aids)
                except Exception:
                    pass

                return self.send_json({
                    'status': 'success',
                    'source_path': source_path,
                    'files': [x['filename'] for x in items],
                    'file_count': len(items),
                    'created_assets': created_assets,
                    'moved': moved_count,
                    'copied': copied_count,
                    'deleted_source': (deleted_source_count if delete_source else None),
                    'linked': linked_count,
                    'import_source_cleanup': import_cleanup,
                    'source_share': (self._unc_share_key(source_path) if debug_move else None),
                    'target_share': (None if debug_move else None),
                    'move_failures': (move_failures[:10] if debug_move else None),
                }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _ensure_image_type_scope_columns(self, conn):
        """
        Intentionally NO-OP.
        Database schema must be managed via scripts/sql/*.sql only.
        Do not create/alter/check schema at runtime.
        """
        return

    def _image_type_platform_table_exists(self, conn):
        # Intentionally avoid runtime schema checks.
        return True

    def _get_image_type_platform_ids_map(self, conn, image_type_ids):
        ids = [self._parse_int(x) for x in (image_type_ids or []) if self._parse_int(x)]
        if not ids:
            return {}
        placeholders = ','.join(['%s'] * len(ids))
        try:
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT image_type_id, platform_type_id
                    FROM image_type_platform_types
                    WHERE image_type_id IN ({placeholders})
                    ORDER BY image_type_id ASC, platform_type_id ASC
                    """,
                    tuple(ids),
                )
                rows = cur.fetchall() or []
        except Exception:
            rows = []
        out = {}
        for r in rows:
            tid = self._parse_int(r.get('image_type_id'))
            pid = self._parse_int(r.get('platform_type_id'))
            if not tid or not pid:
                continue
            out.setdefault(tid, []).append(pid)
        return out

    def _get_image_type_reference_counts(self, conn, image_type_id):
        """Return list of (label, count) for business tables still referencing this image type."""
        tid = self._parse_int(image_type_id)
        if not tid:
            return []
        checks = []
        if self._table_has_column(conn, 'image_assets', 'image_type_id'):
            checks.append(('image_assets', '图片库'))
        if self._has_required_tables(['aplus_version_assets']) and self._table_has_column(conn, 'aplus_version_assets', 'image_type_id'):
            checks.append(('aplus_version_assets', 'A+版本素材'))
        if self._table_has_column(conn, 'sales_variant_image_mappings', 'image_type_id'):
            checks.append(('sales_variant_image_mappings', '销售规格图片映射'))
        out = []
        with conn.cursor() as cur:
            for table, label in checks:
                cur.execute(f"SELECT COUNT(1) AS c FROM {table} WHERE image_type_id=%s", (tid,))
                c = int((cur.fetchone() or {}).get('c') or 0)
                if c > 0:
                    out.append((label, c))
        return out

    def _set_image_type_platform_ids(self, conn, image_type_id, platform_type_ids):
        tid = self._parse_int(image_type_id)
        if not tid:
            return
        ids = sorted(set([self._parse_int(x) for x in (platform_type_ids or []) if self._parse_int(x)]))
        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM image_type_platform_types WHERE image_type_id=%s", (tid,))
                for pid in ids:
                    cur.execute(
                        "INSERT IGNORE INTO image_type_platform_types (image_type_id, platform_type_id) VALUES (%s, %s)",
                        (tid, int(pid)),
                    )
        except Exception:
            return

    def _parse_bool_flag(self, value, default=False):
        if value is None:
            return bool(default)
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in ('1', 'true', 'yes', 'on', 'y'):
            return True
        if text in ('0', 'false', 'no', 'off', 'n'):
            return False
        return bool(default)

    def _handle_image_type_api_core(self, environ, method, start_response):
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            usage = (query_params.get('usage', [''])[0] or '').strip().lower()
            keyword = (query_params.get('q', [''])[0] or '').strip()
            include_disabled = self._parse_bool_flag((query_params.get('include_disabled', ['0'])[0] or '0'), default=False)
            include_platforms = self._parse_bool_flag((query_params.get('include_platforms', ['0'])[0] or '0'), default=False)
            platform_type_id = self._parse_int((query_params.get('platform_type_id', [''])[0] or '').strip())

            if method == 'GET':
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        where_parts = []
                        params = []
                        if not include_disabled:
                            where_parts.append('is_enabled=1')
                        if keyword:
                            where_parts.append('name LIKE %s')
                            params.append(f"%{keyword}%")

                        usage_col = {
                            'sales': 'applies_sales',
                            'fabric': 'applies_fabric',
                            'order_product': 'applies_order_product',
                            'aplus': 'applies_aplus',
                        }.get(usage)
                        if usage_col:
                            where_parts.append(f"{usage_col}=1")

                        # platform filter:
                        # - if no mapping rows -> 通用 (always included)
                        # - else must have an explicit mapping row for requested platform_type_id
                        if platform_type_id and self._image_type_platform_table_exists(conn):
                            where_parts.append(
                                "(NOT EXISTS (SELECT 1 FROM image_type_platform_types itpt WHERE itpt.image_type_id=image_types.id)"
                                " OR EXISTS (SELECT 1 FROM image_type_platform_types itpt2 WHERE itpt2.image_type_id=image_types.id AND itpt2.platform_type_id=%s))"
                            )
                            params.append(int(platform_type_id))

                        where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ''
                        cur.execute(
                            f"""
                            SELECT id, name, is_enabled,
                                   applies_fabric, applies_sales, applies_order_product, applies_aplus,
                                   required_width_px, required_height_px,
                                   aplus_layout_json_mobile, aplus_layout_json_desktop,
                                   created_at, updated_at
                            FROM image_types
                            {where_sql}
                            ORDER BY sort_order ASC, id ASC
                            """,
                            tuple(params),
                        )
                        rows = cur.fetchall() or []
                    if include_platforms and rows:
                        id_list = [self._parse_int(r.get('id')) for r in rows if self._parse_int(r.get('id'))]
                        mp = self._get_image_type_platform_ids_map(conn, id_list)
                        for r in rows:
                            rid = self._parse_int(r.get('id'))
                            r['platform_type_ids'] = mp.get(rid, [])
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                name = str(data.get('name') or '').strip()
                if not name:
                    return self.send_json({'status': 'error', 'message': 'Missing name'}, start_response)
                if len(name) > 64:
                    return self.send_json({'status': 'error', 'message': '类型名称长度不能超过64个字符'}, start_response)

                applies_fabric = int(self._parse_bool_flag(data.get('applies_fabric'), default=True))
                applies_sales = int(self._parse_bool_flag(data.get('applies_sales'), default=True))
                applies_order_product = int(self._parse_bool_flag(data.get('applies_order_product'), default=True))
                applies_aplus = int(self._parse_bool_flag(data.get('applies_aplus'), default=True))
                required_width_px = data.get('required_width_px', None)
                required_height_px = data.get('required_height_px', None)
                platform_type_ids = data.get('platform_type_ids', [])
                aplus_layout_json_mobile = data.get('aplus_layout_json_mobile', None)
                aplus_layout_json_desktop = data.get('aplus_layout_json_desktop', None)
                try:
                    required_width_px = None if required_width_px is None or required_width_px == '' else int(required_width_px)
                except Exception:
                    required_width_px = None
                try:
                    required_height_px = None if required_height_px is None or required_height_px == '' else int(required_height_px)
                except Exception:
                    required_height_px = None

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id, is_enabled FROM image_types WHERE name=%s LIMIT 1", (name,))
                        exists = cur.fetchone() or {}
                        if exists.get('id'):
                            cur.execute(
                                """
                                UPDATE image_types
                                SET is_enabled=1,
                                    applies_fabric=%s,
                                    applies_sales=%s,
                                    applies_order_product=%s,
                                    applies_aplus=%s,
                                    required_width_px=%s,
                                    required_height_px=%s,
                                    aplus_layout_json_mobile=%s,
                                    aplus_layout_json_desktop=%s
                                WHERE id=%s
                                """,
                                (applies_fabric, applies_sales, applies_order_product, applies_aplus, required_width_px, required_height_px, aplus_layout_json_mobile, aplus_layout_json_desktop, exists.get('id')),
                            )
                            self._set_image_type_platform_ids(conn, exists.get('id'), platform_type_ids)
                            return self.send_json({'status': 'success', 'id': exists.get('id'), 'reused': True}, start_response)

                        cur.execute(
                            """
                            INSERT INTO image_types (
                                name, is_enabled, applies_fabric, applies_sales, applies_order_product, applies_aplus,
                                required_width_px, required_height_px,
                                aplus_layout_json_mobile, aplus_layout_json_desktop
                            )
                            VALUES (%s, 1, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (name, applies_fabric, applies_sales, applies_order_product, applies_aplus, required_width_px, required_height_px, aplus_layout_json_mobile, aplus_layout_json_desktop),
                        )
                        new_id = cur.lastrowid
                    self._set_image_type_platform_ids(conn, new_id, platform_type_ids)
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method in ('PUT', 'PATCH'):
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                sets = []
                vals = []
                for key in ('is_enabled', 'applies_fabric', 'applies_sales', 'applies_order_product', 'applies_aplus'):
                    if key in data:
                        sets.append(f"{key}=%s")
                        vals.append(int(self._parse_bool_flag(data.get(key), default=False)))
                for key in ('required_width_px', 'required_height_px'):
                    if key in data:
                        v = data.get(key)
                        if v is None or v == '':
                            sets.append(f"{key}=NULL")
                        else:
                            try:
                                sets.append(f"{key}=%s")
                                vals.append(int(v))
                            except Exception:
                                sets.append(f"{key}=NULL")
                for key in ('aplus_layout_json_mobile', 'aplus_layout_json_desktop'):
                    if key in data:
                        v = data.get(key)
                        if v is None:
                            sets.append(f"{key}=NULL")
                        else:
                            sets.append(f"{key}=%s")
                            vals.append(str(v))
                if 'name' in data:
                    name = str(data.get('name') or '').strip()
                    if not name:
                        return self.send_json({'status': 'error', 'message': '类型名称不能为空'}, start_response)
                    if len(name) > 64:
                        return self.send_json({'status': 'error', 'message': '类型名称长度不能超过64个字符'}, start_response)
                    sets.append('name=%s')
                    vals.append(name)
                has_platform_ids = 'platform_type_ids' in data
                if (not sets) and (not has_platform_ids):
                    return self.send_json({'status': 'error', 'message': 'No updatable fields'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if sets:
                            cur.execute(f"UPDATE image_types SET {', '.join(sets)} WHERE id=%s", tuple(vals + [item_id]))
                    if has_platform_ids:
                        self._set_image_type_platform_ids(conn, item_id, data.get('platform_type_ids', []))
                return self.send_json({'status': 'success', 'id': item_id}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM image_types WHERE id=%s LIMIT 1", (item_id,))
                        if not (cur.fetchone() or {}).get('id'):
                            return self.send_json({'status': 'error', 'message': '图片类型不存在'}, start_response)
                        refs = self._get_image_type_reference_counts(conn, item_id)
                        if refs:
                            parts = [f'{label}{cnt}处' for label, cnt in refs]
                            return self.send_json({
                                'status': 'error',
                                'message': f'无法删除：该类型仍被引用（{"，".join(parts)}）。请先解除引用或改为禁用。',
                            }, start_response)
                        cur.execute("DELETE FROM image_type_platform_types WHERE image_type_id=%s", (item_id,))
                        cur.execute("DELETE FROM image_types WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success', 'id': item_id}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '图片类型已存在'}, start_response)
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_image_type_api(self, environ, method, start_response):
        """通用图片类型 API：支持按 usage 过滤和适用范围开关管理。"""
        return self._handle_image_type_api_core(environ, method, start_response)

    def handle_sales_image_type_api(self, environ, method, start_response):
        """兼容旧路由 /api/sales-image-type，内部复用通用图片类型 API。"""
        return self._handle_image_type_api_core(environ, method, start_response)

    def _get_image_type_id_by_name(self, conn, type_name):
        name = (type_name or '').strip()
        if not name:
            name = '文字卖点图'
        self._ensure_image_type_scope_columns(conn)
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM image_types WHERE name=%s AND is_enabled=1 LIMIT 1", (name,))
            row = cur.fetchone() or {}
            if row.get('id'):
                return self._parse_int(row.get('id'))
            cur.execute("SELECT id, is_enabled FROM image_types WHERE name=%s LIMIT 1", (name,))
            existing = cur.fetchone() or {}
            if existing.get('id'):
                if self._parse_int(existing.get('is_enabled')) != 1:
                    cur.execute("UPDATE image_types SET is_enabled=1 WHERE id=%s", (existing.get('id'),))
                return self._parse_int(existing.get('id'))
            cur.execute(
                """
                INSERT INTO image_types (name, is_enabled, applies_fabric, applies_sales, applies_aplus)
                VALUES (%s, 1, 1, 1, 1)
                """,
                (name,),
            )
            if cur.lastrowid:
                return self._parse_int(cur.lastrowid)
            cur.execute("SELECT id FROM image_types WHERE is_enabled=1 ORDER BY id ASC LIMIT 1")
            fallback = cur.fetchone() or {}
        return self._parse_int(fallback.get('id'))

    def _get_sales_product_image_sort_start(self, conn, sales_product_id):
        """
        Return the current max sort_order for the target sales product.
        Compatible with both schemas:
        - legacy: sales_variant_image_mappings.sales_product_id
        - new:    sales_variant_image_mappings.variant_id (sales_products.variant_id)
        """
        spid = int(sales_product_id or 0)
        if not spid:
            return 0
        has_sim_spid = self._table_has_column(conn, 'sales_variant_image_mappings', 'sales_product_id')
        has_sim_vid = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
        with conn.cursor() as cur:
            if has_sim_spid:
                cur.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) AS max_sort FROM sales_variant_image_mappings WHERE sales_product_id=%s",
                    (spid,)
                )
                row = cur.fetchone() or {}
                return max(0, self._parse_int(row.get('max_sort')) or 0)
            if has_sim_vid:
                cur.execute("SELECT variant_id FROM sales_products WHERE id=%s", (spid,))
                r = cur.fetchone() or {}
                vid = self._parse_int(r.get('variant_id')) or 0
                if not vid:
                    return 0
                cur.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) AS max_sort FROM sales_variant_image_mappings WHERE variant_id=%s",
                    (vid,)
                )
                row = cur.fetchone() or {}
                return max(0, self._parse_int(row.get('max_sort')) or 0)
        return 0

    def _get_variant_image_sort_start(self, conn, variant_id):
        vid = int(variant_id or 0)
        if vid <= 0 or not self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id'):
            return 0
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) AS max_sort FROM sales_variant_image_mappings WHERE variant_id=%s",
                    (vid,),
                )
                row = cur.fetchone() or {}
            return max(0, self._parse_int(row.get('max_sort')) or 0)
        except Exception:
            return 0

    def _load_variant_first_image_preview(self, conn, variant_ids, type_name='白底纯图'):
        """
        Return {variant_id: image_b64} for the first image (by sort_order) of a given type.
        Uses image_assets.image_type_id if available; falls back to sales_variant_image_mappings.image_type_id.
        """
        vids = [int(v or 0) for v in (variant_ids or []) if int(v or 0) > 0]
        if not vids:
            return {}

        has_variant = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
        if not has_variant:
            return {}

        has_ia_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
        has_sim_tid = self._table_has_column(conn, 'sales_variant_image_mappings', 'image_type_id')
        if not (has_ia_tid or has_sim_tid):
            return {}

        # Prefer matching image type name; fallback to first image if none match
        preferred_names = []
        base_name = str(type_name or '').strip() or '白底纯图'
        preferred_names.append(base_name)
        # Keep aliases tight to avoid accidentally prioritizing non-target types.
        if '白底纯图' in base_name:
            alias_candidates = ('主图·白底纯图', '主图白底纯图', '纯白底图')
        else:
            # Backward-compatible aliases for non-pure white types.
            alias_candidates = ('主图·白底图', '主图白底图', '白底', 'White')
        for cand in alias_candidates:
            if cand not in preferred_names:
                preferred_names.append(cand)

        join_it = ""
        where_type = ""
        params = []
        if has_ia_tid:
            # LEFT JOIN so assets with NULL image_type_id don't get excluded prematurely.
            join_it = "LEFT JOIN image_types it ON it.id = ia.image_type_id"
            where_type = "AND it.name IN ({})".format(",".join(["%s"] * len(preferred_names)))
            params.extend(preferred_names)
        else:
            join_it = "JOIN image_types it ON it.id = sim.image_type_id"
            where_type = "AND it.name IN ({})".format(",".join(["%s"] * len(preferred_names)))
            params.extend(preferred_names)

        placeholders = ",".join(["%s"] * len(vids))
        has_ia_ofn = self._table_has_column(conn, 'image_assets', 'original_filename')
        ofn_sel = "ia.original_filename" if has_ia_ofn else "'' AS original_filename"
        sql = f"""
            SELECT sim.variant_id, ia.storage_path, {ofn_sel}, sim.sort_order, sim.id
            FROM sales_variant_image_mappings sim
            JOIN image_assets ia ON ia.id = sim.image_asset_id
            {join_it}
            WHERE sim.variant_id IN ({placeholders})
              {where_type}
            ORDER BY sim.variant_id ASC, sim.sort_order ASC, sim.id ASC
        """
        with conn.cursor() as cur:
            cur.execute(sql, tuple(vids) + tuple(params))
            rows = cur.fetchall() or []

        # If no rows match preferred white-background types, fallback to first image per variant
        if not rows:
            sql2 = f"""
                SELECT sim.variant_id, ia.storage_path, {ofn_sel}, sim.sort_order, sim.id
                FROM sales_variant_image_mappings sim
                JOIN image_assets ia ON ia.id = sim.image_asset_id
                WHERE sim.variant_id IN ({placeholders})
                ORDER BY sim.variant_id ASC, sim.sort_order ASC, sim.id ASC
            """
            with conn.cursor() as cur:
                cur.execute(sql2, tuple(vids))
                rows = cur.fetchall() or []

        out = {}
        for row in rows:
            vid = int(row.get('variant_id') or 0)
            if not vid or vid in out:
                continue
            storage_path = (row.get('storage_path') or '').strip()
            if not storage_path:
                continue
            if isinstance(storage_path, str):
                try:
                    rel_bytes = os.fsencode(storage_path)
                except Exception:
                    rel_bytes = storage_path.encode('utf-8', errors='surrogatepass')
            else:
                rel_bytes = storage_path
            out[vid] = base64.b64encode(rel_bytes).decode('ascii') if rel_bytes else ''
        return out

    def _find_image_asset_by_sha256(self, conn, sha256):
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM image_assets WHERE sha256=%s LIMIT 1",
                (sha256,),
            )
            return cur.fetchone() or None

    def _save_image_asset_file(self, storage_path, content):
        abs_path = self._join_resources(storage_path)
        folder = os.path.dirname(abs_path)
        if not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
        with open(abs_path, 'wb') as f:
            f.write(content or b'')
        return abs_path

    def _insert_image_asset_dynamic(self, conn, cur, rec):
        """Insert image_assets; optional legacy columns if still present post-migration."""
        cols = ['sha256', 'storage_path']
        vals = [
            rec.get('sha256'),
            rec.get('storage_path'),
        ]
        # Include original_filename only if column still exists (pre-migration)
        if self._table_has_column(conn, 'image_assets', 'original_filename'):
            cols.append('original_filename')
            vals.append((rec.get('original_filename') or rec.get('filename') or ''))
        if self._table_has_column(conn, 'image_assets', 'description'):
            cols.append('description')
            vals.append(rec.get('description', '') or '')
        tid = rec.get('image_type_id')
        if tid and self._table_has_column(conn, 'image_assets', 'image_type_id'):
            cols.append('image_type_id')
            vals.append(int(tid))
        if self._table_has_column(conn, 'image_assets', 'is_deprecated'):
            cols.append('is_deprecated')
            vals.append(int(rec.get('is_deprecated') or 0))
        ext = rec.get('ext')
        fs = rec.get('file_size')
        for c, v in (('file_ext', ext or ''), ('mime_type', 'image/*'), ('file_size', fs)):
            if self._table_has_column(conn, 'image_assets', c):
                cols.append(c)
                vals.append(int(v or 0) if c == 'file_size' else v)
        # Ensure created_by is written if user_id is provided and column exists
        uid = rec.get('created_by')
        if self._table_has_column(conn, 'image_assets', 'created_by'):
            if uid:
                cols.append('created_by')
                vals.append(int(uid))
            # If uid is None/empty and column exists, default to current user if available
            # Otherwise let the column remain NULL (default behavior)
        ph = ', '.join(['%s'] * len(cols))
        cur.execute(f"INSERT INTO image_assets ({', '.join(cols)}) VALUES ({ph})", tuple(vals))
        return cur.lastrowid

    def _execute_sku_mapping_upsert(self, conn, cur, aid, sort_order, image_type_id, variant_id, sales_product_id, user_id):
        """Upsert sales_variant_image_mappings; image_type_id only included if column still exists (pre-migration)."""
        has_var = bool(variant_id) and self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
        key_col = 'variant_id' if has_var else 'sales_product_id'
        key_val = int(variant_id if has_var else (sales_product_id or 0))
        cols = [key_col, 'image_asset_id']
        vals = [key_val, int(aid)]
        has_sim_tid = self._table_has_column(conn, 'sales_variant_image_mappings', 'image_type_id')
        if has_sim_tid:
            cols.append('image_type_id')
            vals.append(int(image_type_id or 0))
        cols.append('sort_order')
        vals.append(sort_order)
        if self._table_has_column(conn, 'sales_variant_image_mappings', 'created_by'):
            cols.append('created_by')
            vals.append(int(user_id) if user_id else None)
        dup_parts = ['sort_order=%s']
        dup_vals = [sort_order]
        if has_sim_tid:
            dup_parts.append('image_type_id=%s')
            dup_vals.append(int(image_type_id or 0))
        ph = ', '.join(['%s'] * len(vals))
        sql = (
            f"INSERT INTO sales_variant_image_mappings ({', '.join(cols)}) VALUES ({ph}) "
            f"ON DUPLICATE KEY UPDATE {', '.join(dup_parts)}"
        )
        cur.execute(sql, tuple(vals + dup_vals))
        if self._table_has_column(conn, 'image_assets', 'image_type_id') and image_type_id:
            cur.execute(
                "UPDATE image_assets SET image_type_id=%s WHERE id=%s",
                (int(image_type_id), int(aid)),
            )

    def _read_sales_product_image_items(self, conn, sales_product_id=None, variant_id=None):
        has_variant = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
        has_sales_product_id = self._table_has_column(conn, 'sales_variant_image_mappings', 'sales_product_id')
        has_sim_tid = self._table_has_column(conn, 'sales_variant_image_mappings', 'image_type_id')
        has_ia_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
        has_ia_dep = self._table_has_column(conn, 'image_assets', 'is_deprecated')
        use_variant = bool(variant_id) and has_variant
        # If legacy sales_product_id column is gone, we must query by variant_id.
        if not has_sales_product_id and not use_variant:
            return []
        where_col = "sim.variant_id" if use_variant else "sim.sales_product_id"
        where_val = int(variant_id) if use_variant else int(sales_product_id or 0)
        dep_expr = "COALESCE(ia.is_deprecated,0)" if has_ia_dep else "0"
        has_ia_ofn = self._table_has_column(conn, 'image_assets', 'original_filename')
        ofn_sel = "ia.original_filename AS original_filename" if has_ia_ofn else "NULL AS original_filename"
        if has_ia_tid:
            join_types = "LEFT JOIN image_types it ON it.id = ia.image_type_id"
            type_name_expr = "it.name"
            type_id_expr = "ia.image_type_id"
        elif has_sim_tid:
            join_types = "LEFT JOIN image_types it ON it.id = sim.image_type_id"
            type_name_expr = "it.name"
            type_id_expr = "sim.image_type_id"
        else:
            join_types = ""
            type_name_expr = "NULL"
            type_id_expr = "NULL"
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT sim.id AS mapping_id, sim.sort_order,
                       ia.id AS image_asset_id, ia.sha256, ia.storage_path,
                       {ofn_sel},
                       ia.description,
                       {type_id_expr} AS image_type_id,
                       {type_name_expr} AS image_type_name,
                       {dep_expr} AS is_deprecated
                FROM sales_variant_image_mappings sim
                JOIN image_assets ia ON ia.id = sim.image_asset_id
                {join_types}
                WHERE {where_col}=%s
                ORDER BY {dep_expr} ASC, sim.sort_order ASC, sim.id ASC
                """,
                (where_val,)
            )
            rows = cur.fetchall() or []
        items = []
        for row in rows:
            storage_path = (row.get('storage_path') or '').strip()
            base_name = os.path.basename(storage_path) if storage_path else ''
            orig_fn = (str(row.get('original_filename') or '').strip()) if has_ia_ofn else ''
            image_name = orig_fn or base_name
            # On some Windows/Python setups, the filesystem encoding may effectively behave like ASCII.
            # Avoid crashing JSON responses when storage_path contains Chinese or other non-ASCII chars.
            if isinstance(storage_path, str):
                try:
                    rel_bytes = os.fsencode(storage_path)
                except Exception:
                    rel_bytes = storage_path.encode('utf-8', errors='surrogatepass')
            else:
                rel_bytes = storage_path
            image_b64 = base64.b64encode(rel_bytes).decode('ascii') if rel_bytes else ''
            items.append({
                'mapping_id': row.get('mapping_id'),
                'image_asset_id': row.get('image_asset_id'),
                'image_name': image_name,
                'image_b64': image_b64,
                'description': row.get('description') or '',
                'image_type_id': row.get('image_type_id'),
                'image_type_name': row.get('image_type_name') or '',
                'sort_order': row.get('sort_order') or 0,
                'group_sort': None,
                'is_deprecated': int(row.get('is_deprecated') or 0),
                'sha256': row.get('sha256') or '',
                'file_size': 0,
            })
        return items

    def _read_sales_product_image_list_thumb(self, conn, variant_id, image_type_name):
        """Return at most one mapped image for list thumbnails (LIMIT 1), same ordering as full list."""
        has_variant = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
        if not variant_id or not has_variant:
            return []
        has_ia_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
        has_sim_tid = self._table_has_column(conn, 'sales_variant_image_mappings', 'image_type_id')
        has_ia_dep = self._table_has_column(conn, 'image_assets', 'is_deprecated')
        dep_expr = "COALESCE(ia.is_deprecated,0)" if has_ia_dep else "0"
        has_ia_ofn = self._table_has_column(conn, 'image_assets', 'original_filename')
        ofn_sel = "ia.original_filename AS original_filename" if has_ia_ofn else "NULL AS original_filename"
        if has_ia_tid:
            join_types = "LEFT JOIN image_types it ON it.id = ia.image_type_id"
            type_name_expr = "it.name"
            type_id_expr = "ia.image_type_id"
        elif has_sim_tid:
            join_types = "LEFT JOIN image_types it ON it.id = sim.image_type_id"
            type_name_expr = "it.name"
            type_id_expr = "sim.image_type_id"
        else:
            join_types = ""
            type_name_expr = "NULL"
            type_id_expr = "NULL"
        type_name = str(image_type_name or "").strip()
        type_clause = ""
        params = [int(variant_id)]
        if type_name and join_types:
            type_clause = f" AND TRIM(COALESCE({type_name_expr},''))=%s"
            params.append(type_name)
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT sim.id AS mapping_id, sim.sort_order,
                       ia.id AS image_asset_id, ia.sha256, ia.storage_path,
                       {ofn_sel},
                       ia.description,
                       {type_id_expr} AS image_type_id,
                       {type_name_expr} AS image_type_name,
                       {dep_expr} AS is_deprecated
                FROM sales_variant_image_mappings sim
                JOIN image_assets ia ON ia.id = sim.image_asset_id
                {join_types}
                WHERE sim.variant_id=%s{type_clause}
                ORDER BY {dep_expr} ASC, sim.sort_order ASC, sim.id ASC
                LIMIT 1
                """,
                tuple(params),
            )
            rows = cur.fetchall() or []
        items = []
        for row in rows:
            storage_path = (row.get("storage_path") or "").strip()
            base_name = os.path.basename(storage_path) if storage_path else ""
            orig_fn = (str(row.get("original_filename") or "").strip()) if has_ia_ofn else ""
            image_name = orig_fn or base_name
            if isinstance(storage_path, str):
                try:
                    rel_bytes = os.fsencode(storage_path)
                except Exception:
                    rel_bytes = storage_path.encode("utf-8", errors="surrogatepass")
            else:
                rel_bytes = storage_path
            image_b64 = base64.b64encode(rel_bytes).decode("ascii") if rel_bytes else ""
            items.append(
                {
                    "mapping_id": row.get("mapping_id"),
                    "image_asset_id": row.get("image_asset_id"),
                    "image_name": image_name,
                    "image_b64": image_b64,
                    "description": row.get("description") or "",
                    "image_type_id": row.get("image_type_id"),
                    "image_type_name": row.get("image_type_name") or "",
                    "sort_order": row.get("sort_order") or 0,
                    "group_sort": None,
                    "is_deprecated": int(row.get("is_deprecated") or 0),
                    "sha256": row.get("sha256") or "",
                    "file_size": 0,
                }
            )
        return items

    def _storage_rel_from_image_b64(self, image_b64):
        """Decode UI `image_b64` / gallery id into a normalized relative storage path (UTF-8, forward slashes)."""
        s = str(image_b64 or '').strip()
        if not s:
            return ''
        try:
            raw = base64.b64decode(s, validate=False)
        except Exception:
            return ''
        if not raw:
            return ''
        try:
            t = raw.decode('utf-8')
        except UnicodeDecodeError:
            try:
                t = raw.decode('utf-8', errors='surrogatepass')
            except Exception:
                t = self._safe_fsdecode(raw)
        return str(t or '').strip().replace('\\', '/')

    def _select_sales_variant_mapping_for_api(self, conn, cur, where_key, where_val, data):
        """
        Resolve exactly one sales_variant_image_mappings row for PUT/DELETE/replace.
        Prefer mapping_id / image_asset_id / exact storage_path (from image_b64) / sha256 over basename-only image_name.
        Returns (row dict or None, error_message or None).
        """
        data = data or {}
        mapping_id = self._parse_int(data.get('mapping_id')) or 0
        image_asset_id = self._parse_int(data.get('image_asset_id')) or 0
        sha256 = str(data.get('sha256') or '').strip().lower()
        image_b64 = str(data.get('image_b64') or data.get('image_path_b64') or '').strip()
        image_name = str(data.get('image_name') or '').strip()
        rel_from_b64 = self._storage_rel_from_image_b64(image_b64) if image_b64 else ''

        has_ia_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
        join_it = "LEFT JOIN image_types it ON it.id = ia.image_type_id" if has_ia_tid else ""
        old_type_sel = "it.name AS old_type_name" if has_ia_tid else "'' AS old_type_name"
        sel = (
            f"SELECT sim.id, sim.image_asset_id, sim.sort_order, ia.storage_path, ia.sha256 AS old_sha256, {old_type_sel} "
            f"FROM sales_variant_image_mappings sim "
            f"JOIN image_assets ia ON ia.id = sim.image_asset_id "
            f"{join_it} "
            f"WHERE {{where_clause}} "
            f"ORDER BY sim.sort_order ASC, sim.id ASC"
        )

        def _one_row(sql, params):
            cur.execute(sql, params)
            return cur.fetchone() or {}

        if mapping_id > 0:
            row = _one_row(sel.format(where_clause=f"sim.id=%s AND {where_key}=%s"), (mapping_id, where_val))
            if row.get('id'):
                return row, None
            return None, '图片不存在或已过期，请刷新后重试'

        if image_asset_id > 0:
            row = _one_row(sel.format(where_clause=f"sim.image_asset_id=%s AND {where_key}=%s"), (image_asset_id, where_val))
            if row.get('id'):
                return row, None
            return None, '图片不存在或已过期，请刷新后重试'

        if rel_from_b64:
            row = _one_row(sel.format(where_clause=f"ia.storage_path=%s AND {where_key}=%s"), (rel_from_b64, where_val))
            if row.get('id'):
                return row, None
            return None, '图片不存在或已过期，请刷新后重试'

        if sha256:
            row = _one_row(sel.format(where_clause=f"ia.sha256=%s AND {where_key}=%s"), (sha256, where_val))
            if row.get('id'):
                return row, None
            return None, '图片不存在或已过期，请刷新后重试'

        if not image_name:
            return None, '缺少图片标识：请提供 mapping_id、image_b64、sha256 或 image_name'

        cur.execute(
            sel.format(
                where_clause=f"{where_key}=%s AND (ia.storage_path=%s OR ia.storage_path LIKE %s)"
            ),
            (where_val, image_name, f'%/{image_name}'),
        )
        rows = cur.fetchall() or []
        if len(rows) > 1:
            return None, '存在多张同名文件，请刷新页面后重试；若仍出现，请携带 mapping_id 或 image_b64 以唯一标识图片'
        if len(rows) == 1:
            return rows[0], None
        return None, None

    def _resolve_sales_product_variant_folder(self, sales_product_id, ensure_folder=False):
        if not sales_product_id:
            raise RuntimeError('Missing sales_product_id')
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                fabric_join = "LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if has_fabric_id else ""
                if has_fabric_id and has_fabric_text:
                    fabric_select = "COALESCE(fm.fabric_code, v.fabric) AS fabric"
                elif has_fabric_id:
                    fabric_select = "fm.fabric_code AS fabric"
                else:
                    fabric_select = ("v.fabric AS fabric" if has_fabric_text else "'' AS fabric")
                cur.execute(
                    f"""
                    SELECT sp.id, v.spec_name, {fabric_select}, pf.sku_family,
                           {("fm.fabric_name_en AS fabric_name_en, v.fabric_id AS fabric_id" if has_fabric_id else "'' AS fabric_name_en, 0 AS fabric_id")}
                    FROM sales_products sp
                    LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                    LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                    {fabric_join}
                    WHERE sp.id=%s
                    """,
                    (sales_product_id,)
                )
                row = cur.fetchone() or {}
            if not row.get('id'):
                raise RuntimeError('销售产品不存在')

            sku_name = (row.get('sku_family') or '').strip()
            spec_part = (row.get('spec_name') or '').strip().replace('/', '-').replace('\\', '-')
            fabric_part = str(row.get('fabric_name_en') or '').strip().replace('/', '-').replace('\\', '-')
            if not fabric_part:
                fabric_part = self._resolve_fabric_folder_part(conn, row.get('fabric_id'), row.get('fabric'))
            if not fabric_part:
                raise RuntimeError('面料为必填：当前销售产品缺少可解析的面料信息，无法定位主图文件夹')
            variant_folder_name = self._sales_variant_subfolder_display_name(spec_part, fabric_part)
            if not sku_name:
                raise RuntimeError('当前销售产品缺少货号，无法定位主图文件夹')

            if ensure_folder:
                self._ensure_listing_sales_variant_folder(sku_name, spec_part, fabric_part)
            base_folder = self._ensure_listing_folder()
            folder_path = os.path.join(
                base_folder,
                self._safe_fsencode(sku_name),
                self._safe_fsencode('主图'),
                self._safe_fsencode(variant_folder_name)
            )
            return {
                'sales_product_id': int(row.get('id')),
                'sku_family': sku_name,
                'spec_name': spec_part,
                'fabric_folder_part': fabric_part,
                'fabric_id': self._parse_int(row.get('fabric_id')) or 0,
                'fabric_name_en': str(row.get('fabric_name_en') or '').strip(),
                'variant_folder': variant_folder_name,
                'folder_path': folder_path,
            }

    def _batch_reorder_sales_variant_image_mappings(self, conn, where_key, where_val, items):
        """Batch-update sort_order for sales variant images in one transaction."""
        updates = []
        with conn.cursor() as cur:
            for idx, item in enumerate(items or []):
                if not isinstance(item, dict):
                    continue
                sort_order = self._parse_int(item.get('sort_order'))
                if sort_order is None:
                    sort_order = idx + 1
                pick = {
                    'mapping_id': self._parse_int(item.get('mapping_id')) or 0,
                    'image_asset_id': self._parse_int(item.get('image_asset_id')) or 0,
                    'sha256': str(item.get('sha256') or '').strip(),
                    'image_b64': str(item.get('image_b64') or item.get('image_path_b64') or '').strip(),
                    'image_name': str(item.get('image_name') or '').strip(),
                }
                if not any([
                    pick['mapping_id'],
                    pick['image_asset_id'],
                    pick['sha256'],
                    pick['image_b64'],
                    pick['image_name'],
                ]):
                    continue
                mapping, map_err = self._select_sales_variant_mapping_for_api(
                    conn, cur, where_key, where_val, pick
                )
                if map_err:
                    raise ValueError(map_err)
                if not mapping or not mapping.get('id'):
                    raise ValueError('图片不存在')
                updates.append((max(1, int(sort_order)), int(mapping.get('id'))))
            if not updates:
                return
            cur.executemany(
                f"UPDATE sales_variant_image_mappings SET sort_order=%s WHERE id=%s AND {where_key}=%s",
                [(sort_order, mapping_id, where_val) for sort_order, mapping_id in updates],
            )

    # -------------------------------------------------------------------------
    # 主图排序与映射 API
    # -------------------------------------------------------------------------

    def handle_sales_product_main_images_api(self, environ, method, start_response):
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', '') or '')
            if method == 'GET':
                sales_product_id = self._parse_int(query_params.get('sales_product_id', [''])[0] or query_params.get('id', [''])[0])
                variant_id_param = self._parse_int(query_params.get('variant_id', [''])[0])
                list_thumb = self._parse_bool_flag(query_params.get('list_thumb', ['0'])[0] or '0', default=False)
                thumb_image_type = str(query_params.get('thumb_image_type', [''])[0] or '').strip()

                if variant_id_param and not sales_product_id:
                    ensure_folder = not list_thumb
                    with self._get_db_connection() as conn:
                        try:
                            folder_info = self._resolve_sales_variant_folder_by_variant_id(
                                variant_id_param, ensure_folder=ensure_folder
                            )
                        except Exception as e:
                            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
                        variant_id = int(variant_id_param)
                        has_variant_col = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
                        items = []
                        if variant_id and has_variant_col:
                            if list_thumb:
                                items = self._read_sales_product_image_list_thumb(
                                    conn, variant_id, thumb_image_type
                                ) or []
                            else:
                                items = self._read_sales_product_image_items(
                                    conn, sales_product_id=None, variant_id=variant_id
                                ) or []
                        fabric_id = self._parse_int(folder_info.get('fabric_id')) or 0
                        fabric_items = []
                        if (not list_thumb) and fabric_id and self._has_required_tables(
                            ['fabric_image_mappings', 'image_assets']
                        ):
                            try:
                                fabric_items = self._read_fabric_image_items(conn, fabric_id)
                            except Exception:
                                fabric_items = []
                    return self.send_json(
                        {
                            'status': 'success',
                            'items': items,
                            'fabric_items': fabric_items,
                            'folder': {
                                'variant_id': int(variant_id or 0),
                                'sku_family': folder_info.get('sku_family') or '',
                                'variant_folder': folder_info.get('variant_folder') or '',
                            },
                        },
                        start_response,
                    )

                if not sales_product_id:
                    return self.send_json({'status': 'error', 'message': 'Missing sales_product_id 或 variant_id'}, start_response)

                with self._get_db_connection() as conn:
                    variant_id = 0
                    fabric_id = 0
                    try:
                        with conn.cursor() as cur:
                            cur.execute("SELECT variant_id FROM sales_products WHERE id=%s", (sales_product_id,))
                            row = cur.fetchone() or {}
                            variant_id = self._parse_int(row.get('variant_id')) or 0
                    except Exception:
                        variant_id = 0
                    has_variant_col = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
                    has_spid_col = self._table_has_column(conn, 'sales_variant_image_mappings', 'sales_product_id')
                    items = []
                    if variant_id and has_variant_col:
                        items = self._read_sales_product_image_items(conn, sales_product_id=None, variant_id=variant_id) or []
                        # Robust fallback: if variant changed unexpectedly but legacy sales_product_id mapping exists,
                        # also show those mappings to avoid "images disappeared" after save.
                        if (not items) and has_spid_col:
                            items = self._read_sales_product_image_items(conn, sales_product_id=sales_product_id) or []
                    else:
                        items = self._read_sales_product_image_items(conn, sales_product_id=sales_product_id) or []
                    folder_info = self._resolve_sales_product_variant_folder(sales_product_id, ensure_folder=True)
                    # Prefer fabric_id obtained while resolving folder (same variant join),
                    # fallback to direct query only if missing.
                    fabric_id = self._parse_int(folder_info.get('fabric_id')) or 0
                    if not fabric_id:
                        try:
                            with conn.cursor() as cur:
                                if self._table_has_column(conn, 'sales_product_variants', 'fabric_id'):
                                    cur.execute("SELECT fabric_id FROM sales_product_variants WHERE id=%s", (variant_id,))
                                    frow = cur.fetchone() or {}
                                    fabric_id = self._parse_int(frow.get('fabric_id')) or 0
                        except Exception:
                            fabric_id = 0
                    # Final fallback: resolve fabric_id from fabric_name_en / folder part
                    if not fabric_id:
                        try:
                            name_en = str(folder_info.get('fabric_name_en') or '').strip()
                            folder_part = str(folder_info.get('fabric_folder_part') or '').strip()
                            probe = name_en or folder_part
                            if probe:
                                with conn.cursor() as cur:
                                    cur.execute(
                                        "SELECT id FROM fabric_materials WHERE fabric_name_en=%s LIMIT 1",
                                        (probe,),
                                    )
                                    prow = cur.fetchone() or {}
                                    fabric_id = self._parse_int(prow.get('id')) or 0
                                    if not fabric_id:
                                        cur.execute(
                                            "SELECT id FROM fabric_materials WHERE fabric_code=%s LIMIT 1",
                                            (self._code_before_dash(probe),),
                                        )
                                        prow = cur.fetchone() or {}
                                        fabric_id = self._parse_int(prow.get('id')) or 0
                        except Exception:
                            fabric_id = fabric_id or 0
                    fabric_items = []
                    if fabric_id and self._has_required_tables(['fabric_image_mappings', 'image_assets']):
                        try:
                            fabric_items = self._read_fabric_image_items(conn, fabric_id)
                        except Exception:
                            fabric_items = []

                return self.send_json({
                    'status': 'success',
                    'items': items,
                    'fabric_items': fabric_items,
                    'folder': {
                        'variant_id': int(variant_id or 0),
                        'sku_family': folder_info.get('sku_family') or '',
                        'variant_folder': folder_info.get('variant_folder') or ''
                    }
                }, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                action = (query_params.get('action', [''])[0] or '').strip().lower()
                if action == 'reorder':
                    sales_product_id = self._parse_int(data.get('sales_product_id'))
                    variant_id_direct = self._parse_int(data.get('variant_id'))
                    items = data.get('items')
                    if not isinstance(items, list) or not items:
                        return self.send_json({'status': 'error', 'message': '缺少 items 排序列表'}, start_response)
                    if not sales_product_id and not variant_id_direct:
                        return self.send_json(
                            {'status': 'error', 'message': 'Missing sales_product_id 或 variant_id'},
                            start_response,
                        )
                    with self._get_db_connection() as conn:
                        variant_id = 0
                        if sales_product_id:
                            try:
                                with conn.cursor() as cur:
                                    cur.execute("SELECT variant_id FROM sales_products WHERE id=%s", (sales_product_id,))
                                    row = cur.fetchone() or {}
                                    variant_id = self._parse_int(row.get('variant_id')) or 0
                            except Exception:
                                variant_id = 0
                        elif variant_id_direct:
                            variant_id = int(variant_id_direct)
                        with conn.cursor() as cur:
                            has_sim_vid = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
                            has_sim_spid = self._table_has_column(conn, 'sales_variant_image_mappings', 'sales_product_id')
                            if not has_sim_vid and not has_sim_spid:
                                return self.send_json(
                                    {'status': 'error', 'message': '图片映射表缺少 variant_id / sales_product_id 字段，无法定位图片'},
                                    start_response,
                                )
                            if has_sim_vid and variant_id:
                                where_key = 'variant_id'
                                where_val = variant_id
                            elif has_sim_spid and sales_product_id:
                                where_key = 'sales_product_id'
                                where_val = sales_product_id
                            else:
                                return self.send_json(
                                    {'status': 'error', 'message': '当前销售产品缺少 variant_id，无法定位图片'},
                                    start_response,
                                )
                        try:
                            self._batch_reorder_sales_variant_image_mappings(conn, where_key, where_val, items)
                        except ValueError as ex:
                            return self.send_json({'status': 'error', 'message': str(ex)}, start_response)
                    return self.send_json({'status': 'success'}, start_response)

                sales_product_id = self._parse_int(data.get('sales_product_id'))
                variant_id_direct = self._parse_int(data.get('variant_id'))
                image_name = str(data.get('image_name') or '').strip()
                mapping_id_body = self._parse_int(data.get('mapping_id')) or 0
                image_asset_body = self._parse_int(data.get('image_asset_id')) or 0
                sha256_body = str(data.get('sha256') or '').strip()
                image_b64_body = str(data.get('image_b64') or data.get('image_path_b64') or '').strip()
                description = str(data.get('description') or '').strip()
                image_type_name = str(data.get('image_type_name') or '').strip()
                is_deprecated = self._parse_int(data.get('is_deprecated'))
                sort_order = self._parse_int(data.get('sort_order'))
                new_filename = str(data.get('new_filename') or data.get('new_image_name') or '').strip()
                has_image_pick = bool(
                    image_name or mapping_id_body or image_asset_body or sha256_body or image_b64_body
                )
                if (not sales_product_id and not variant_id_direct) or not has_image_pick:
                    return self.send_json(
                        {'status': 'error', 'message': 'Missing sales_product_id / variant_id 或图片标识（image_name / mapping_id 等）'},
                        start_response,
                    )

                with self._get_db_connection() as conn:
                    variant_id = 0
                    if sales_product_id:
                        try:
                            with conn.cursor() as cur:
                                cur.execute("SELECT variant_id FROM sales_products WHERE id=%s", (sales_product_id,))
                                row = cur.fetchone() or {}
                                variant_id = self._parse_int(row.get('variant_id')) or 0
                        except Exception:
                            variant_id = 0
                    elif variant_id_direct:
                        variant_id = int(variant_id_direct)
                    with conn.cursor() as cur:
                        has_ia_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
                        has_ia_dep = self._table_has_column(conn, 'image_assets', 'is_deprecated')
                        has_sim_vid = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
                        has_sim_spid = self._table_has_column(conn, 'sales_variant_image_mappings', 'sales_product_id')
                        if not has_sim_vid and not has_sim_spid:
                            return self.send_json({'status': 'error', 'message': '图片映射表缺少 variant_id / sales_product_id 字段，无法定位图片'}, start_response)
                        if has_sim_vid and variant_id:
                            where_key = "sim.variant_id"
                            where_val = variant_id
                        elif has_sim_spid and sales_product_id:
                            where_key = "sim.sales_product_id"
                            where_val = sales_product_id
                        else:
                            return self.send_json({'status': 'error', 'message': '当前销售产品缺少 variant_id，无法定位图片'}, start_response)
                        pick = {
                            'image_name': image_name,
                            'mapping_id': mapping_id_body,
                            'image_asset_id': image_asset_body,
                            'sha256': sha256_body,
                            'image_b64': image_b64_body,
                        }
                        mapping, map_err = self._select_sales_variant_mapping_for_api(conn, cur, where_key, where_val, pick)
                        if map_err:
                            return self.send_json({'status': 'error', 'message': map_err}, start_response)
                        if not mapping or not mapping.get('id'):
                            return self.send_json({'status': 'error', 'message': '图片不存在'}, start_response)

                        aid = mapping.get('image_asset_id')
                        old_storage_path = str(mapping.get('storage_path') or '').strip()
                        old_type_name = str(mapping.get('old_type_name') or '').strip()
                        ia_sets = []
                        ia_params = []
                        if description is not None:
                            ia_sets.append('description=%s')
                            ia_params.append(description)
                        if image_type_name and has_ia_tid:
                            tid = self._get_image_type_id_by_name(conn, image_type_name)
                            if tid:
                                ia_sets.append('image_type_id=%s')
                                ia_params.append(tid)
                        if has_ia_dep and is_deprecated is not None:
                            ia_sets.append('is_deprecated=%s')
                            ia_params.append(1 if int(is_deprecated or 0) else 0)
                        if ia_sets:
                            cur.execute(
                                f"UPDATE image_assets SET {', '.join(ia_sets)} WHERE id=%s",
                                tuple(ia_params + [aid]),
                            )
                        # Rename file if type changed and filename follows "类型-原名称.ext"
                        try:
                            if image_type_name and old_storage_path:
                                old_base = os.path.basename(old_storage_path)
                                new_prefix = self._sanitize_filename_component(image_type_name, 32)
                                old_prefix = self._sanitize_filename_component(old_type_name, 32) if old_type_name else ''
                                if new_prefix and old_base:
                                    rest = old_base
                                    if old_prefix and rest.startswith(old_prefix + '-'):
                                        rest = rest[len(old_prefix) + 1:]
                                    if rest.startswith(new_prefix + '-'):
                                        new_base = rest
                                    else:
                                        new_base = f"{new_prefix}-{rest}"
                                    if new_base != old_base:
                                        new_storage_path = old_storage_path[:-len(old_base)] + new_base
                                        old_abs = self._join_resources(old_storage_path)
                                        new_abs = self._join_resources(new_storage_path)
                                        if not self._listing_paths_equivalent(old_abs, new_abs):
                                            os.replace(old_abs, new_abs)
                                            cur.execute(
                                                "UPDATE image_assets SET storage_path=%s WHERE id=%s",
                                                (new_storage_path, aid),
                                            )
                        except Exception:
                            pass
                        # Optional manual rename of the on-disk filename (storage_path basename)
                        if new_filename:
                            cur.execute("SELECT storage_path FROM image_assets WHERE id=%s", (aid,))
                            prow = cur.fetchone() or {}
                            cur_storage = str((prow.get('storage_path') or '')).strip().replace('\\', '/')
                            if cur_storage:
                                old_base = os.path.basename(cur_storage)
                                old_name, old_ext = os.path.splitext(old_base)
                                want_base = os.path.basename(new_filename.replace('\\', '/'))
                                want_base = self._sanitize_filename_component(want_base, 160)
                                if not want_base or want_base in ('.', '..'):
                                    return self.send_json({'status': 'error', 'message': '无效的文件名'}, start_response)
                                want_name, want_ext = os.path.splitext(want_base)
                                if not want_ext:
                                    want_ext = old_ext or ''
                                if old_ext and want_ext and want_ext.lower() != old_ext.lower():
                                    return self.send_json({'status': 'error', 'message': '不允许修改图片扩展名'}, start_response)
                                final_base = (want_name + (want_ext or old_ext)).strip()
                                if final_base != old_base:
                                    new_storage_path = (cur_storage[: -len(old_base)] + final_base) if cur_storage.endswith(old_base) else ''
                                    if not new_storage_path:
                                        return self.send_json({'status': 'error', 'message': '无法计算新的 storage_path'}, start_response)
                                    cur.execute(
                                        """
                                        SELECT COUNT(1) AS cnt
                                        FROM image_assets
                                        WHERE id<>%s
                                          AND (
                                                storage_path=%s
                                             OR storage_path LIKE %s
                                             OR storage_path LIKE %s
                                          )
                                        """,
                                        (aid, new_storage_path, f'%/{final_base}', f'%/{final_base}/%'),
                                    )
                                    clash = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                                    if clash:
                                        return self.send_json({'status': 'error', 'message': '文件名已被其他图片占用'}, start_response)
                                    old_abs = self._join_resources(cur_storage)
                                    new_abs = self._join_resources(new_storage_path)
                                    if os.path.exists(new_abs):
                                        return self.send_json({'status': 'error', 'message': '目标文件名已存在'}, start_response)
                                    try:
                                        if self._listing_paths_equivalent(old_abs, new_abs):
                                            return self.send_json({'status': 'error', 'message': '无效的重命名（源与目标相同）'}, start_response)
                                        os.replace(old_abs, new_abs)
                                    except Exception as e:
                                        return self.send_json({'status': 'error', 'message': f'重命名失败: {str(e)}'}, start_response)
                                    ia_rename_sets = ['storage_path=%s']
                                    ia_rename_params = [new_storage_path]
                                    if self._table_has_column(conn, 'image_assets', 'original_filename'):
                                        ia_rename_sets.append('original_filename=%s')
                                        ia_rename_params.append(final_base)
                                    cur.execute(
                                        f"UPDATE image_assets SET {', '.join(ia_rename_sets)} WHERE id=%s",
                                        tuple(ia_rename_params + [aid]),
                                    )
                        if sort_order is not None:
                            cur.execute(
                                "UPDATE sales_variant_image_mappings SET sort_order=%s WHERE id=%s",
                                (max(1, sort_order), mapping.get('id')),
                            )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                sales_product_id = self._parse_int(data.get('sales_product_id'))
                variant_id_direct = self._parse_int(data.get('variant_id'))
                image_name = str(data.get('image_name') or '').strip()
                mapping_id_body = self._parse_int(data.get('mapping_id')) or 0
                image_asset_body = self._parse_int(data.get('image_asset_id')) or 0
                sha256_body = str(data.get('sha256') or '').strip()
                image_b64_body = str(data.get('image_b64') or data.get('image_path_b64') or '').strip()
                has_image_pick = bool(
                    image_name or mapping_id_body or image_asset_body or sha256_body or image_b64_body
                )
                if (not sales_product_id and not variant_id_direct) or not has_image_pick:
                    return self.send_json(
                        {'status': 'error', 'message': 'Missing sales_product_id / variant_id 或图片标识（image_name / mapping_id 等）'},
                        start_response,
                    )

                with self._get_db_connection() as conn:
                    variant_id = 0
                    if sales_product_id:
                        try:
                            with conn.cursor() as cur:
                                cur.execute("SELECT variant_id FROM sales_products WHERE id=%s", (sales_product_id,))
                                row = cur.fetchone() or {}
                                variant_id = self._parse_int(row.get('variant_id')) or 0
                        except Exception:
                            variant_id = 0
                    elif variant_id_direct:
                        variant_id = int(variant_id_direct)
                    with conn.cursor() as cur:
                        has_sim_vid = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
                        has_sim_spid = self._table_has_column(conn, 'sales_variant_image_mappings', 'sales_product_id')
                        if not has_sim_vid and not has_sim_spid:
                            return self.send_json({'status': 'error', 'message': '图片映射表缺少 variant_id / sales_product_id 字段，无法定位图片'}, start_response)
                        if has_sim_vid and variant_id:
                            where_key = "sim.variant_id"
                            where_val = variant_id
                        elif has_sim_spid and sales_product_id:
                            where_key = "sim.sales_product_id"
                            where_val = sales_product_id
                        else:
                            return self.send_json({'status': 'error', 'message': '当前销售产品缺少 variant_id，无法定位图片'}, start_response)
                        pick = {
                            'image_name': image_name,
                            'mapping_id': mapping_id_body,
                            'image_asset_id': image_asset_body,
                            'sha256': sha256_body,
                            'image_b64': image_b64_body,
                        }
                        mapping, map_err = self._select_sales_variant_mapping_for_api(conn, cur, where_key, where_val, pick)
                        if map_err:
                            return self.send_json({'status': 'error', 'message': map_err}, start_response)
                        if not mapping or not mapping.get('id'):
                            return self.send_json({'status': 'error', 'message': '图片文件不存在'}, start_response)
                        image_asset_id = mapping.get('image_asset_id')
                        cur.execute("DELETE FROM sales_variant_image_mappings WHERE id=%s", (mapping.get('id'),))
                        cur.execute("SELECT COUNT(*) AS cnt FROM sales_variant_image_mappings WHERE image_asset_id=%s", (image_asset_id,))
                        remain_sku = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                        remain_fabric = 0
                        if self._has_required_tables(['fabric_image_mappings']):
                            cur.execute("SELECT COUNT(*) AS cnt FROM fabric_image_mappings WHERE image_asset_id=%s", (image_asset_id,))
                            remain_fabric = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                        remain_total = remain_sku + remain_fabric
                        if remain_total <= 0:
                            cur.execute("SELECT storage_path FROM image_assets WHERE id=%s", (image_asset_id,))
                            asset_row = cur.fetchone() or {}
                            storage_path = (asset_row.get('storage_path') or '').strip()
                            if storage_path:
                                try:
                                    abs_path = self._join_resources(storage_path)
                                    moved_ok, _dst, _err = self._move_file_to_listing_recycle_bin(abs_path, '删除')
                                    if not moved_ok:
                                        # Best-effort fallback: avoid leaving DB inconsistent with a still-present file
                                        self._safe_unlink(abs_path)
                                except Exception:
                                    pass
                            cur.execute("DELETE FROM image_assets WHERE id=%s", (image_asset_id,))
                            return self.send_json(
                                {
                                    'status': 'success',
                                    'asset_deleted': True,
                                    'remaining_refs': 0,
                                    'message': '图片记录已删除；原文件已移入『上架资源』/回收站（若移动失败则尝试直接删除）',
                                },
                                start_response,
                            )
                return self.send_json(
                    {
                        'status': 'success',
                        'asset_deleted': False,
                        'remaining_refs': int(remain_total),
                        'message': '图片已从当前规格解绑，但仍被其他面料/规格引用，未做物理删除',
                    },
                    start_response,
                )

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _order_product_first_linked_variant_id(self, conn, order_product_id):
        """销售规格绑定下单产品时写入 sales_variant_order_links；配件图子目录与主图一致用此 variant。"""
        opid = int(order_product_id or 0)
        if opid <= 0:
            return 0
        if not self._has_required_tables(['sales_variant_order_links']):
            return 0
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT variant_id FROM sales_variant_order_links
                    WHERE order_product_id=%s
                    ORDER BY variant_id ASC
                    LIMIT 1
                    """,
                    (opid,),
                )
                r = cur.fetchone() or {}
            return int(self._parse_int(r.get('variant_id')) or 0)
        except Exception:
            return 0

    def _resolve_order_product_main_image_folder(self, order_product_id, ensure_folder=False):
        opid = int(order_product_id or 0)
        if opid <= 0:
            raise RuntimeError('Missing order_product_id')
        with self._get_db_connection() as conn:
            vid = self._order_product_first_linked_variant_id(conn, opid)
            if vid:
                # 与销售产品主图同一套「货号/主图/规格-面料英文名」中的规格-面料段；物理目录在 配件图 下
                vinfo = self._resolve_sales_variant_folder_by_variant_id(vid, ensure_folder=False)
                sku_name = str(vinfo.get('sku_family') or '').strip()
                folder_name = str(vinfo.get('variant_folder') or '').strip()
                if not sku_name or not folder_name:
                    raise RuntimeError('已绑定规格但无法解析文件夹名（缺少货号或面料段）')
                base_folder = self._ensure_listing_folder()
                folder_path = os.path.join(
                    base_folder,
                    self._safe_fsencode(sku_name),
                    self._safe_fsencode('配件图'),
                    self._safe_fsencode(folder_name),
                )
                if ensure_folder:
                    os.makedirs(folder_path, exist_ok=True)
                return {
                    'order_product_id': opid,
                    'sku_family': sku_name,
                    'variant_folder': folder_name,
                    'folder_path': folder_path,
                    'linked_variant_id': vid,
                    'spec_folder_part': str(vinfo.get('spec_name') or '').strip(),
                    'fabric_folder_part': str(vinfo.get('fabric_folder_part') or '').strip(),
                }

            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT op.id, op.sku, op.version_no, op.spec_qty_short, pf.sku_family,
                           fm.fabric_name_en, fm.fabric_code
                    FROM order_products op
                    LEFT JOIN product_families pf ON pf.id = op.sku_family_id
                    LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                    WHERE op.id=%s
                    LIMIT 1
                    """,
                    (opid,),
                )
                row = cur.fetchone() or {}
            if not row.get('id'):
                raise RuntimeError('下单产品不存在')
            sku_name = str(row.get('sku_family') or '').strip()
            if not sku_name:
                raise RuntimeError('当前下单产品缺少货号，无法定位主图文件夹')
            spec_part = self._derive_order_product_spec_folder_part(row)
            fabric_part = str(row.get('fabric_name_en') or '').strip().replace('/', '-').replace('\\', '-')
            if not fabric_part:
                fabric_part = self._resolve_fabric_folder_part(conn, row.get('fabric_id'), row.get('fabric_code') or '')
            if spec_part and fabric_part:
                folder_name = f"{spec_part}-{fabric_part}"
            elif fabric_part:
                folder_name = fabric_part
            elif spec_part:
                folder_name = spec_part
            else:
                folder_name = "通用"
            base_folder = self._ensure_listing_folder()
            folder_path = os.path.join(
                base_folder,
                self._safe_fsencode(sku_name),
                self._safe_fsencode('配件图'),
                self._safe_fsencode(folder_name),
            )
            if ensure_folder:
                os.makedirs(folder_path, exist_ok=True)
            return {
                'order_product_id': opid,
                'sku_family': sku_name,
                'variant_folder': folder_name,
                'folder_path': folder_path,
                'linked_variant_id': 0,
                'spec_folder_part': spec_part,
                'fabric_folder_part': fabric_part,
            }

    def _derive_order_product_spec_folder_part(self, order_row):
        # 规格与数量简称优先；为空时从 SKU 结构中推断规格段（而不是回落成中文固定字符串）。
        spec = str((order_row or {}).get('spec_qty_short') or '').strip()
        if spec:
            return spec.replace('/', '-').replace('\\', '-')
        sku = str((order_row or {}).get('sku') or '').strip()
        sku_family = str((order_row or {}).get('sku_family') or '').strip()
        version_no = str((order_row or {}).get('version_no') or '').strip()
        fabric_code = str((order_row or {}).get('fabric_code') or '').strip()
        fabric_prefix = fabric_code.split('-', 1)[0].strip() if fabric_code else ''
        if sku:
            tokens = [t.strip() for t in sku.split('-') if t and t.strip()]
            if tokens and sku_family and tokens[0].lower() == sku_family.lower():
                tokens = tokens[1:]
            if tokens and version_no and tokens[-1].lower() == version_no.lower():
                tokens = tokens[:-1]
            if tokens and fabric_prefix and tokens[-1].lower() == fabric_prefix.lower():
                tokens = tokens[:-1]
            if tokens:
                return tokens[0].replace('/', '-').replace('\\', '-')
        return ''

    def _ensure_order_product_common_folder(self, sku_family):
        sku_name = str(sku_family or '').strip()
        if not sku_name:
            raise RuntimeError('Missing sku_family')
        base_folder = self._ensure_listing_folder()
        common_folder = os.path.join(
            base_folder,
            self._safe_fsencode(sku_name),
            self._safe_fsencode('配件图'),
            self._safe_fsencode('通用'),
        )
        os.makedirs(common_folder, exist_ok=True)
        return common_folder

    def _read_order_product_image_items(self, conn, order_product_id):
        opid = int(order_product_id or 0)
        if not opid:
            return []
        has_ia_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
        has_ia_dep = self._table_has_column(conn, 'image_assets', 'is_deprecated')
        has_ia_desc = self._table_has_column(conn, 'image_assets', 'description')
        has_ia_ofn = self._table_has_column(conn, 'image_assets', 'original_filename')
        join_it = "LEFT JOIN image_types it ON it.id = ia.image_type_id" if has_ia_tid else ""
        type_sel = "COALESCE(it.name, '') AS image_type_name" if has_ia_tid else "'' AS image_type_name"
        dep_sel = "COALESCE(ia.is_deprecated, 0) AS is_deprecated" if has_ia_dep else "0 AS is_deprecated"
        desc_sel = "COALESCE(ia.description, '') AS description" if has_ia_desc else "'' AS description"
        ofn_sel = "COALESCE(ia.original_filename, '') AS original_filename" if has_ia_ofn else "'' AS original_filename"
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT opim.id, opim.sort_order, ia.storage_path, ia.created_at, ia.updated_at,
                       {type_sel}, {dep_sel}, {desc_sel}, {ofn_sel}
                FROM order_product_image_mappings opim
                JOIN image_assets ia ON ia.id = opim.image_asset_id
                {join_it}
                WHERE opim.order_product_id=%s
                ORDER BY opim.sort_order ASC, opim.id ASC
                """,
                (opid,),
            )
            rows = cur.fetchall() or []
        out = []
        for row in rows:
            storage_path = str(row.get('storage_path') or '').strip()
            if not storage_path:
                continue
            try:
                rel_bytes = os.fsencode(storage_path)
            except Exception:
                rel_bytes = storage_path.encode('utf-8', errors='surrogatepass')
            image_b64 = base64.b64encode(rel_bytes).decode('ascii') if rel_bytes else ''
            out.append({
                'image_name': os.path.basename(storage_path),
                'image_b64': image_b64,
                'storage_path': storage_path,
                'image_type_name': str(row.get('image_type_name') or '').strip(),
                'is_deprecated': self._parse_int(row.get('is_deprecated')) or 0,
                'description': str(row.get('description') or '').strip(),
                'original_filename': str(row.get('original_filename') or '').strip(),
                'sort_order': self._parse_int(row.get('sort_order')) or 100,
                'created_at': row.get('created_at'),
                'updated_at': row.get('updated_at'),
            })
        return out

    def _get_order_product_image_sort_start(self, conn, order_product_id):
        opid = int(order_product_id or 0)
        if not opid:
            return 0
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COALESCE(MAX(sort_order), 0) AS max_sort FROM order_product_image_mappings WHERE order_product_id=%s",
                (opid,),
            )
            row = cur.fetchone() or {}
        return max(0, self._parse_int(row.get('max_sort')) or 0)

    def handle_order_product_main_images_api(self, environ, method, start_response):
        try:
            if method == 'GET':
                query_params = parse_qs(environ.get('QUERY_STRING', ''))
                order_product_id = self._parse_int(query_params.get('order_product_id', [''])[0] or query_params.get('id', [''])[0])
                if not order_product_id:
                    return self.send_json({'status': 'error', 'message': 'Missing order_product_id'}, start_response)
                folder_info = self._resolve_order_product_main_image_folder(order_product_id, ensure_folder=True)
                with self._get_db_connection() as conn:
                    items = self._read_order_product_image_items(conn, order_product_id)
                folder_out = dict(folder_info or {})
                fp = folder_out.get('folder_path')
                if isinstance(fp, (bytes, bytearray)):
                    folder_out['folder_path'] = self._safe_fsdecode(bytes(fp))
                return self.send_json({'status': 'success', 'items': items, 'folder': folder_out}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ) or {}
                order_product_id = self._parse_int(data.get('order_product_id'))
                image_name = str(data.get('image_name') or '').strip()
                if not order_product_id or not image_name:
                    return self.send_json({'status': 'error', 'message': 'Missing order_product_id or image_name'}, start_response)
                folder_info = self._resolve_order_product_main_image_folder(order_product_id, ensure_folder=True)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT opim.id AS mapping_id, opim.image_asset_id, ia.storage_path
                            FROM order_product_image_mappings opim
                            JOIN image_assets ia ON ia.id = opim.image_asset_id
                            WHERE opim.order_product_id=%s
                              AND (ia.storage_path=%s OR ia.storage_path LIKE %s)
                            ORDER BY opim.sort_order ASC, opim.id ASC
                            LIMIT 1
                            """,
                            (order_product_id, image_name, f'%/{image_name}'),
                        )
                        row = cur.fetchone() or {}
                        mapping_id = self._parse_int(row.get('mapping_id')) or 0
                        aid = self._parse_int(row.get('image_asset_id')) or 0
                        storage_path = str(row.get('storage_path') or '').strip()
                        if not mapping_id or not aid:
                            return self.send_json({'status': 'error', 'message': '图片不存在'}, start_response)
                        cur.execute("DELETE FROM order_product_image_mappings WHERE id=%s", (mapping_id,))
                        cur.execute("SELECT COUNT(1) AS cnt FROM order_product_image_mappings WHERE image_asset_id=%s", (aid,))
                        remain_order = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                        cur.execute("SELECT COUNT(1) AS cnt FROM sales_variant_image_mappings WHERE image_asset_id=%s", (aid,))
                        remain_sku = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                        remain_fabric = 0
                        if self._has_required_tables(['fabric_image_mappings']):
                            cur.execute("SELECT COUNT(1) AS cnt FROM fabric_image_mappings WHERE image_asset_id=%s", (aid,))
                            remain_fabric = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                        if (remain_order + remain_sku + remain_fabric) <= 0:
                            if storage_path:
                                try:
                                    abs_path = self._join_resources(storage_path)
                                    moved_ok, _dst, _err = self._move_file_to_listing_recycle_bin(abs_path, '删除')
                                    if not moved_ok:
                                        self._safe_unlink(abs_path)
                                except Exception:
                                    pass
                            cur.execute("DELETE FROM image_assets WHERE id=%s", (aid,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_order_product_main_images_upload_api(self, environ, start_response):
        try:
            if environ.get('REQUEST_METHOD') != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)
            form = cgi.FieldStorage(fp=environ.get('wsgi.input'), environ=environ, keep_blank_values=True)
            order_product_id = self._parse_int((form.getfirst('order_product_id', '') or '').strip()) if form else 0
            image_type_name = str((form.getfirst('image_type_name', '') or '').strip()) if form else ''
            if not order_product_id:
                return self.send_json({'status': 'error', 'message': 'Missing order_product_id'}, start_response)
            if not image_type_name:
                return self.send_json({'status': 'error', 'message': '请先选择图片类型后再上传主图'}, start_response)
            uploads = []
            for p in getattr(form, 'list', []) or []:
                if getattr(p, 'filename', None):
                    content = p.file.read() if p.file else b''
                    uploads.append({'filename': p.filename, 'content': content})
            if not uploads:
                return self.send_json({'status': 'error', 'message': 'No image uploaded'}, start_response)

            folder_info = self._resolve_order_product_main_image_folder(order_product_id, ensure_folder=True)
            folder_abs = folder_info.get('folder_path')
            created = []
            touched_asset_ids = []
            with self._get_db_connection() as conn:
                type_id = self._get_image_type_id_by_name(conn, image_type_name)
                if not type_id:
                    return self.send_json({'status': 'error', 'message': '图片类型无效或未在系统中配置，请刷新页面后重新选择类型再上传'}, start_response)
                sort_start = self._get_order_product_image_sort_start(conn, order_product_id)
                with conn.cursor() as cur:
                    for idx, up in enumerate(uploads, start=1):
                        fname = os.path.basename(str(up.get('filename') or '').strip())
                        if not fname or not self._is_image_name(fname):
                            continue
                        payload = up.get('content') or b''
                        if not payload:
                            continue
                        final_name = fname
                        if image_type_name:
                            pref = self._sanitize_filename_component(image_type_name, 32)
                            if pref and not final_name.startswith(pref + '-'):
                                final_name = f"{pref}-{final_name}"
                        sha256 = hashlib.sha256(payload).hexdigest()
                        cur.execute("SELECT id, storage_path FROM image_assets WHERE sha256=%s LIMIT 1", (sha256,))
                        exists = cur.fetchone() or {}
                        aid = self._parse_int(exists.get('id')) or 0
                        storage_path = str(exists.get('storage_path') or '').strip()
                        if not aid:
                            abs_target = os.path.join(folder_abs, self._safe_fsencode(final_name))
                            i = 1
                            while os.path.exists(abs_target):
                                stem, ext = os.path.splitext(final_name)
                                abs_target = os.path.join(folder_abs, self._safe_fsencode(f"{stem}_{i}{ext}"))
                                i += 1
                            with open(abs_target, 'wb') as f:
                                f.write(payload)
                            storage_path = self._storage_path_from_abs(abs_target)
                            try:
                                orig_fn, _, _ = self._display_name_from_abs_path_b(abs_target)
                            except Exception:
                                orig_fn = ''
                            aid = self._insert_image_asset_dynamic(
                                conn,
                                cur,
                                {
                                    'sha256': sha256,
                                    'storage_path': storage_path,
                                    'original_filename': orig_fn or final_name,
                                    'description': '',
                                    'image_type_id': int(type_id) if type_id else None,
                                    'is_deprecated': 0,
                                    'file_size': len(payload),
                                },
                            )
                            if aid:
                                touched_asset_ids.append(int(aid))
                        else:
                            cur.execute("SELECT COUNT(1) AS cnt FROM order_product_image_mappings WHERE image_asset_id=%s", (aid,))
                            order_refs = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                            if order_refs > 0 and folder_info.get('sku_family'):
                                try:
                                    common_folder = self._ensure_order_product_common_folder(folder_info.get('sku_family'))
                                    old_abs = self._join_resources(storage_path)
                                    base_name = str(os.path.basename(str(storage_path or '').replace('\\', '/')) or final_name or '').strip()
                                    new_abs = os.path.join(common_folder, self._safe_fsencode(base_name))
                                    if os.path.normpath(old_abs) != os.path.normpath(new_abs):
                                        j = 1
                                        stem, ext = os.path.splitext(base_name)
                                        while os.path.exists(new_abs):
                                            new_abs = os.path.join(common_folder, self._safe_fsencode(f"{stem}_{j}{ext}"))
                                            j += 1
                                        if not self._listing_paths_equivalent(old_abs, new_abs):
                                            os.replace(old_abs, new_abs)
                                            storage_path = self._storage_path_from_abs(new_abs)
                                            cur.execute("UPDATE image_assets SET storage_path=%s WHERE id=%s", (storage_path, aid))
                                except Exception:
                                    pass
                            if type_id:
                                cur.execute("UPDATE image_assets SET image_type_id=%s WHERE id=%s", (type_id, aid))
                            if aid:
                                touched_asset_ids.append(int(aid))
                        cur.execute(
                            """
                            INSERT INTO order_product_image_mappings (order_product_id, image_asset_id, sort_order)
                            VALUES (%s, %s, %s)
                            ON DUPLICATE KEY UPDATE sort_order=VALUES(sort_order)
                            """,
                            (order_product_id, aid, sort_start + idx),
                        )
                        created.append(os.path.basename(storage_path or final_name))
                # After commit: apply rehome rules (e.g. when this asset is also linked to sales variants)
                try:
                    for aid in sorted(set(touched_asset_ids)):
                        self._rehome_image_asset_if_needed(conn, aid)
                except Exception:
                    pass
            return self.send_json({'status': 'success', 'files': created, 'linked': len(created)}, start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_order_product_main_images_import_by_path_api(self, environ, method, start_response):
        """从 NAS 路径导入下单产品主图（支持双击单张与多选批量）。"""
        try:
            if method != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            data = self._read_json_body(environ) or {}
            order_product_id = self._parse_int(data.get('order_product_id'))
            image_type_name = str(data.get('image_type_name') or '').strip()
            source_path_b64 = str(data.get('source_path_b64') or '').strip()
            source_paths_b64 = data.get('source_paths_b64') or []
            prompt_duplicate = str(data.get('prompt_duplicate') or '').strip().lower() in ('1', 'true', 'yes', 'on')
            allow_duplicate = str(data.get('allow_duplicate') or '').strip().lower() in ('1', 'true', 'yes', 'on')
            if not order_product_id:
                return self.send_json({'status': 'error', 'message': 'Missing order_product_id'}, start_response)
            if (not source_path_b64) and (not source_paths_b64):
                return self.send_json({'status': 'error', 'message': 'Missing source_path'}, start_response)

            source_files_b = []
            if isinstance(source_paths_b64, (list, tuple)) and source_paths_b64:
                for b64 in list(source_paths_b64)[:500]:
                    try:
                        raw = base64.b64decode(str(b64 or '').strip())
                        path_b = self._normalize_nas_abs_path_bytes(raw)
                        if path_b and os.path.isfile(path_b) and self._is_image_name(os.path.basename(path_b)):
                            source_files_b.append(path_b)
                    except Exception:
                        continue
            if (not source_files_b) and source_path_b64:
                try:
                    raw = base64.b64decode(source_path_b64)
                    path_b = self._normalize_nas_abs_path_bytes(raw)
                except Exception:
                    path_b = b''
                if path_b and os.path.exists(path_b):
                    if os.path.isfile(path_b):
                        if self._is_image_name(os.path.basename(path_b)):
                            source_files_b = [path_b]
                    else:
                        for name in os.listdir(path_b):
                            abs_file_b = os.path.join(path_b, name)
                            if os.path.isfile(abs_file_b) and self._is_image_name(name):
                                source_files_b.append(abs_file_b)
            source_files_b = sorted(set(source_files_b))
            if not source_files_b:
                return self.send_json({'status': 'error', 'message': '源路径下无图片文件'}, start_response)

            folder_info = self._resolve_order_product_main_image_folder(order_product_id, ensure_folder=True)
            folder_abs = folder_info.get('folder_path')
            if not folder_abs or (not os.path.exists(folder_abs)):
                return self.send_json({'status': 'error', 'message': '无法定位配件图文件夹'}, start_response)

            with self._get_db_connection() as conn:
                type_id = self._get_image_type_id_by_name(conn, image_type_name) if image_type_name else 0
                if image_type_name and not type_id:
                    return self.send_json({'status': 'error', 'message': f'未知图片类型: {image_type_name}'}, start_response)
                sort_start = self._get_order_product_image_sort_start(conn, order_product_id)

                prepared = []
                duplicates = []
                for idx, source_file_b in enumerate(source_files_b, start=1):
                    try:
                        with open(source_file_b, 'rb') as f:
                            payload = f.read() or b''
                    except Exception:
                        payload = b''
                    if not payload:
                        continue
                    sha256 = self._sha256_hex(payload)
                    exists = self._find_image_asset_by_sha256(conn, sha256) or {}
                    aid = self._parse_int(exists.get('id')) or 0
                    filename, _stem, _ext = self._display_name_from_abs_path_b(source_file_b)
                    prepared.append({
                        'idx': idx,
                        'source_file': self._safe_fsdecode(source_file_b),
                        'source_file_b': source_file_b,
                        'filename': filename,
                        'sha256': sha256,
                        'payload': payload,
                        'asset_id': aid,
                        'asset_storage_path': str(exists.get('storage_path') or '').strip() if aid else '',
                    })
                    if aid:
                        duplicates.append({
                            'source_file': self._safe_fsdecode(source_file_b),
                            'sha256': sha256,
                            'image_asset_id': aid,
                            'storage_path': str(exists.get('storage_path') or ''),
                        })
                if prompt_duplicate and duplicates and (not allow_duplicate):
                    return self.send_json({
                        'status': 'duplicate',
                        'message': '检测到重复图片（SHA256 相同），是否继续复用并导入？',
                        'duplicate_count': len(duplicates),
                        'duplicates': duplicates[:200],
                        'file_count': len(prepared),
                    }, start_response)
                if not prepared:
                    return self.send_json({'status': 'error', 'message': '未检测到可导入图片'}, start_response)

                linked = 0
                created_assets = 0
                moved = 0
                recycled_duplicate_sources = 0
                touched_asset_ids = []
                with conn.cursor() as cur:
                    for row in prepared:
                        aid = int(row.get('asset_id') or 0)
                        storage_path = ''
                        if not aid:
                            image_type_prefix = self._sanitize_filename_component(image_type_name, 32) if image_type_name else ''
                            base_name = str(row.get('filename') or '').strip()
                            final_name = f"{image_type_prefix}-{base_name}" if image_type_prefix and (not base_name.startswith(image_type_prefix + '-')) else base_name
                            final_name = self._next_available_filename(folder_abs, final_name)
                            abs_target = os.path.join(folder_abs, self._safe_fsencode(final_name))
                            src_b = row.get('source_file_b')
                            payload_b = row.get('payload') or b''
                            try:
                                with open(abs_target, 'wb') as wf:
                                    wf.write(payload_b)
                            except Exception:
                                continue
                            storage_path = self._storage_path_from_abs(abs_target)
                            try:
                                try:
                                    orig_fn, _, _ = self._display_name_from_abs_path_b(abs_target)
                                except Exception:
                                    orig_fn = ''
                                aid = self._insert_image_asset_dynamic(
                                    conn,
                                    cur,
                                    {
                                        'sha256': row.get('sha256'),
                                        'storage_path': storage_path,
                                        'original_filename': orig_fn or final_name,
                                        'description': '',
                                        'image_type_id': int(type_id) if type_id else None,
                                        'is_deprecated': 0,
                                        'file_size': len(payload_b),
                                    },
                                )
                            except Exception:
                                try:
                                    if os.path.isfile(abs_target):
                                        os.remove(abs_target)
                                except Exception:
                                    pass
                                raise
                            if aid:
                                created_assets += 1
                                touched_asset_ids.append(int(aid))
                                try:
                                    if (
                                        src_b
                                        and os.path.isfile(src_b)
                                        and (not self._listing_paths_equivalent(src_b, abs_target))
                                    ):
                                        os.remove(src_b)
                                        moved += 1
                                except Exception:
                                    pass
                        else:
                            if type_id:
                                cur.execute("UPDATE image_assets SET image_type_id=%s WHERE id=%s", (type_id, aid))
                            if aid:
                                touched_asset_ids.append(int(aid))
                            # SHA256 复用：库中已有 canonical 文件；若 NAS 所选为另一路径的相同副本，移入回收站以免重复占空间
                            src_b = row.get('source_file_b')
                            canon_sp = str(row.get('asset_storage_path') or '').strip()
                            if src_b and os.path.isfile(src_b) and canon_sp:
                                try:
                                    canon_abs = self._abs_from_storage_path(canon_sp)
                                    if not self._listing_paths_equivalent(src_b, canon_abs):
                                        mv_ok, _, _ = self._move_file_to_listing_recycle_bin(src_b, '重复')
                                        if mv_ok:
                                            recycled_duplicate_sources += 1
                                except Exception:
                                    pass
                        if not aid:
                            continue
                        cur.execute(
                            """
                            INSERT INTO order_product_image_mappings (order_product_id, image_asset_id, sort_order)
                            VALUES (%s, %s, %s)
                            ON DUPLICATE KEY UPDATE sort_order=VALUES(sort_order)
                            """,
                            (order_product_id, aid, sort_start + int(row.get('idx') or 0)),
                        )
                        linked += 1
                # Rehome based on combined references (order_product + sales variants) best-effort.
                try:
                    for aid in sorted(set(touched_asset_ids)):
                        self._rehome_image_asset_if_needed(conn, aid)
                except Exception:
                    pass
                return self.send_json({
                    'status': 'success',
                    'file_count': len(prepared),
                    'created_assets': created_assets,
                    'linked': linked,
                    'moved': moved,
                    'recycled_duplicate_sources': recycled_duplicate_sources,
                }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_order_product_main_images_replace_api(self, environ, start_response):
        try:
            if environ.get('REQUEST_METHOD') != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)
            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)
            form = cgi.FieldStorage(fp=environ.get('wsgi.input'), environ=environ, keep_blank_values=True)
            order_product_id = self._parse_int((form.getfirst('order_product_id', '') or '').strip()) if form else 0
            image_name = str((form.getfirst('image_name', '') or '').strip()) if form else ''
            if not order_product_id or not image_name:
                return self.send_json({'status': 'error', 'message': 'Missing order_product_id or image_name'}, start_response)
            uploads = []
            for p in getattr(form, 'list', []) or []:
                if getattr(p, 'filename', None):
                    content = p.file.read() if p.file else b''
                    uploads.append({'filename': p.filename, 'content': content})
            if not uploads:
                return self.send_json({'status': 'error', 'message': 'No image uploaded'}, start_response)
            item0 = uploads[0]
            payload = item0.get('content') or b''
            if not payload:
                return self.send_json({'status': 'error', 'message': 'Empty upload'}, start_response)
            new_sha = self._sha256_hex(payload)

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT opim.id AS mapping_id, opim.image_asset_id, ia.storage_path, ia.sha256
                        FROM order_product_image_mappings opim
                        JOIN image_assets ia ON ia.id = opim.image_asset_id
                        WHERE opim.order_product_id=%s
                          AND (ia.storage_path=%s OR ia.storage_path LIKE %s)
                        ORDER BY opim.sort_order ASC, opim.id ASC
                        LIMIT 1
                        """,
                        (order_product_id, image_name, f'%/{image_name}'),
                    )
                    row = cur.fetchone() or {}
                    aid = self._parse_int(row.get('image_asset_id')) or 0
                    storage_path = str(row.get('storage_path') or '').strip()
                    old_sha = str(row.get('sha256') or '').strip()
                    if not aid or not storage_path:
                        return self.send_json({'status': 'error', 'message': '图片不存在'}, start_response)
                    if old_sha and old_sha == new_sha:
                        return self.send_json({'status': 'error', 'message': '替换文件内容与当前图片相同（sha256 未变化）'}, start_response)
                    cur.execute("SELECT COUNT(1) AS cnt FROM order_product_image_mappings WHERE image_asset_id=%s", (aid,))
                    remain_order = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                    cur.execute("SELECT COUNT(1) AS cnt FROM sales_variant_image_mappings WHERE image_asset_id=%s", (aid,))
                    remain_sales = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                    remain_fabric = 0
                    if self._has_required_tables(['fabric_image_mappings']):
                        cur.execute("SELECT COUNT(1) AS cnt FROM fabric_image_mappings WHERE image_asset_id=%s", (aid,))
                        remain_fabric = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                    if (remain_order + remain_sales + remain_fabric) > 1:
                        return self.send_json({'status': 'error', 'message': '该图片被多个对象引用，请先解绑后再替换'}, start_response)
                    abs_path = self._join_resources(storage_path)
                    with open(abs_path, 'wb') as f:
                        f.write(payload)
                    upd_parts = ['sha256=%s']
                    upd_vals = [new_sha]
                    if self._table_has_column(conn, 'image_assets', 'file_size'):
                        upd_parts.append('file_size=%s')
                        upd_vals.append(len(payload))
                    upd_vals.append(aid)
                    cur.execute(
                        f"UPDATE image_assets SET {', '.join(upd_parts)} WHERE id=%s",
                        tuple(upd_vals),
                    )
            return self.send_json({'status': 'success', 'message': '替换成功'}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_sales_product_main_images_replace_api(self, environ, start_response):
        """Replace a single SKU main image file: new bytes -> new sha256, old file -> 『上架资源』/回收站."""
        try:
            if environ.get('REQUEST_METHOD') != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            raw_body = self._read_wsgi_request_body(environ)
            if raw_body:
                env_copy = dict(environ)
                env_copy['CONTENT_LENGTH'] = str(len(raw_body))
                form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            else:
                form = cgi.FieldStorage(fp=environ.get('wsgi.input'), environ=environ, keep_blank_values=True)

            sales_product_id = self._parse_int((form.getfirst('sales_product_id', '') or '').strip()) if form else 0
            variant_replace = self._parse_int((form.getfirst('variant_id', '') or '').strip()) if form else 0
            image_name = str((form.getfirst('image_name', '') or '').strip()) if form else ''
            mapping_id_rep = self._parse_int((form.getfirst('mapping_id', '') or '').strip()) if form else 0
            image_asset_rep = self._parse_int((form.getfirst('image_asset_id', '') or '').strip()) if form else 0
            sha256_rep = str((form.getfirst('sha256', '') or '').strip()) if form else ''
            image_b64_rep = str((form.getfirst('image_b64', '') or '').strip()) if form else ''
            has_pick_rep = bool(image_name or mapping_id_rep or image_asset_rep or sha256_rep or image_b64_rep)
            if not has_pick_rep or (not sales_product_id and not variant_replace):
                return self.send_json(
                    {'status': 'error', 'message': 'Missing sales_product_id / variant_id 或图片标识（image_name / mapping_id 等）'},
                    start_response,
                )

            uploads = []
            for p in getattr(form, 'list', []) or []:
                if getattr(p, 'filename', None):
                    try:
                        content = p.file.read() or b''
                    except Exception:
                        content = b''
                    uploads.append({'filename': p.filename, 'content': content})
            if not uploads and raw_body:
                uploads = self._parse_multipart_uploads_fallback(content_type, raw_body)
            if not uploads:
                return self.send_json({'status': 'error', 'message': 'No image uploaded'}, start_response)
            item0 = uploads[0]
            filename = os.path.basename(item0.get('filename') or '')
            content = item0.get('content') or b''
            if not filename or not content:
                return self.send_json({'status': 'error', 'message': 'Empty upload'}, start_response)
            if not self._is_image_name(filename):
                return self.send_json({'status': 'error', 'message': '不支持的图片类型'}, start_response)

            user_id = None
            try:
                user_id = self._get_session_user(environ)
            except Exception as e:
                return self.send_json({'status': 'error', 'message': f'无法验证用户身份：{str(e)}'}, start_response)
            if not user_id:
                return self.send_json({'status': 'error', 'message': '必须登录才能替换图片'}, start_response)

            new_sha = self._sha256_hex(content)
            if variant_replace and not sales_product_id:
                try:
                    folder_info = self._resolve_sales_variant_folder_by_variant_id(variant_replace, ensure_folder=True)
                except Exception as e:
                    return self.send_json({'status': 'error', 'message': str(e)}, start_response)
            else:
                folder_info = self._resolve_sales_product_variant_folder(sales_product_id, ensure_folder=True)
            target_folder_abs = folder_info.get('folder_path')
            if not target_folder_abs or not os.path.exists(target_folder_abs):
                return self.send_json({'status': 'error', 'message': '无法定位主图文件夹，请确认货号与面料完整（面料必填；规格可为空）'}, start_response)

            with self._get_db_connection() as conn:
                variant_id = 0
                if variant_replace and not sales_product_id:
                    variant_id = int(variant_replace)
                else:
                    try:
                        with conn.cursor() as cur:
                            cur.execute("SELECT variant_id FROM sales_products WHERE id=%s", (sales_product_id,))
                            row = cur.fetchone() or {}
                            variant_id = self._parse_int(row.get('variant_id')) or 0
                    except Exception:
                        variant_id = 0

                has_sim_vid = self._table_has_column(conn, 'sales_variant_image_mappings', 'variant_id')
                has_sim_spid = self._table_has_column(conn, 'sales_variant_image_mappings', 'sales_product_id')
                if not has_sim_vid and not has_sim_spid:
                    return self.send_json({'status': 'error', 'message': '图片映射表缺少 variant_id / sales_product_id 字段，无法定位图片'}, start_response)
                if has_sim_vid and variant_id:
                    where_key = "sim.variant_id"
                    where_val = variant_id
                elif has_sim_spid and sales_product_id:
                    where_key = "sim.sales_product_id"
                    where_val = sales_product_id
                else:
                    return self.send_json({'status': 'error', 'message': '当前销售产品缺少 variant_id，无法定位图片'}, start_response)

                with conn.cursor() as cur:
                    has_sim_tid = self._table_has_column(conn, 'sales_variant_image_mappings', 'image_type_id')
                    join_it = "LEFT JOIN image_types it ON it.id = ia.image_type_id" if self._table_has_column(conn, 'image_assets', 'image_type_id') else ""
                    if has_sim_tid and join_it:
                        type_sel = "COALESCE(NULLIF(sim.image_type_id, 0), ia.image_type_id, it.id) AS image_type_id, it.name AS image_type_name"
                    elif has_sim_tid:
                        type_sel = "NULLIF(sim.image_type_id, 0) AS image_type_id, '' AS image_type_name"
                    elif join_it:
                        type_sel = "ia.image_type_id AS image_type_id, it.name AS image_type_name"
                    else:
                        type_sel = "0 AS image_type_id, '' AS image_type_name"
                    pick_rep = {
                        'image_name': image_name,
                        'mapping_id': mapping_id_rep,
                        'image_asset_id': image_asset_rep,
                        'sha256': sha256_rep,
                        'image_b64': image_b64_rep,
                    }
                    mapping, map_err = self._select_sales_variant_mapping_for_api(conn, cur, where_key, where_val, pick_rep)
                    if map_err:
                        return self.send_json({'status': 'error', 'message': map_err}, start_response)
                    if not mapping or not mapping.get('id'):
                        return self.send_json({'status': 'error', 'message': '图片不存在'}, start_response)

                    old_aid = int(mapping.get('image_asset_id') or 0)
                    old_sha = str(mapping.get('old_sha256') or '').strip()
                    if old_sha and new_sha == old_sha:
                        return self.send_json({'status': 'error', 'message': '替换文件内容与当前图片相同（sha256 未变化）'}, start_response)

                    sort_order = self._parse_int(mapping.get('sort_order')) or 1
                    image_type_id = self._parse_int(mapping.get('image_type_id')) or 0
                    if not image_type_id and self._table_has_column(conn, 'image_assets', 'image_type_id'):
                        cur.execute("SELECT image_type_id FROM image_assets WHERE id=%s", (old_aid,))
                        trow = cur.fetchone() or {}
                        image_type_id = self._parse_int(trow.get('image_type_id')) or 0
                    if not image_type_id:
                        # Fallback: infer from filename prefix "类型-..."
                        try:
                            base0 = os.path.basename(str(mapping.get('storage_path') or ''))
                            if '-' in base0:
                                type_guess = base0.split('-', 1)[0].strip()
                                image_type_id = self._get_image_type_id_by_name(conn, type_guess) or image_type_id
                        except Exception:
                            pass
                    if not image_type_id:
                        return self.send_json({'status': 'error', 'message': '无法解析图片类型，拒绝替换'}, start_response)

                    # Safety: only replace when the asset is uniquely referenced by this single mapping row.
                    cur.execute("SELECT COUNT(1) AS cnt FROM sales_variant_image_mappings WHERE image_asset_id=%s", (old_aid,))
                    sku_refs = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                    fab_refs = 0
                    if self._has_required_tables(['fabric_image_mappings']):
                        cur.execute("SELECT COUNT(1) AS cnt FROM fabric_image_mappings WHERE image_asset_id=%s", (old_aid,))
                        fab_refs = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                    if sku_refs != 1 or fab_refs != 0:
                        return self.send_json({'status': 'error', 'message': '该图片被多处引用，禁止直接替换文件（请先解除复用/关联）'}, start_response)

                    old_storage = str(mapping.get('storage_path') or '').strip()
                    old_abs = self._join_resources(old_storage) if old_storage else ''
                    old_dir_abs = os.path.dirname(old_abs) if old_abs else target_folder_abs
                    ext = self._guess_image_ext(filename, content) or (os.path.splitext(old_storage)[1] if old_storage else '') or '.jpg'

                    old_base = os.path.basename(old_storage) if old_storage else ''
                    type_part = ''
                    rest_part = os.path.splitext(filename)[0]
                    if old_base and '-' in old_base:
                        type_part = self._sanitize_filename_component(old_base.split('-', 1)[0], 32)
                    if not type_part:
                        try:
                            cur.execute("SELECT name FROM image_types WHERE id=%s LIMIT 1", (image_type_id,))
                            tname = (cur.fetchone() or {}).get('name')
                            type_part = self._sanitize_filename_component(tname, 32) or '图片'
                        except Exception:
                            type_part = '图片'
                    base_part = self._sanitize_filename_component(rest_part, 80) or new_sha[:12]
                    final_name = self._next_available_filename(old_dir_abs, f"{type_part}-{base_part}{ext}")
                    new_abs = os.path.join(old_dir_abs, self._safe_fsencode(final_name))
                    try:
                        with open(new_abs, 'wb') as f:
                            f.write(content or b'')
                    except Exception as e:
                        self._safe_unlink(new_abs)
                        return self.send_json({'status': 'error', 'message': f'写入新图片失败: {str(e)}'}, start_response)

                    new_storage_path = self._storage_path_from_abs(new_abs)
                    if not new_storage_path:
                        self._safe_unlink(new_abs)
                        return self.send_json({'status': 'error', 'message': '无法计算新 storage_path'}, start_response)

                    created_new_row = False
                    new_aid = 0
                    try:
                        self._tx_begin(conn)
                        with conn.cursor() as cur2:
                            existing = self._find_image_asset_by_sha256(conn, new_sha)
                            if existing and int(existing.get('id') or 0) != int(old_aid):
                                self._tx_rollback(conn)
                                self._safe_unlink(new_abs)
                                return self.send_json({'status': 'error', 'message': '新图片与库中已有图片重复（sha256 冲突），请换一张图'}, start_response)

                            cur2.execute("DELETE FROM sales_variant_image_mappings WHERE id=%s", (mapping.get('id'),))

                            if existing and int(existing.get('id') or 0) == int(old_aid):
                                sets = ['sha256=%s', 'storage_path=%s']
                                params = [new_sha, new_storage_path]
                                if self._table_has_column(conn, 'image_assets', 'original_filename'):
                                    sets.append('original_filename=%s')
                                    params.append(final_name)
                                for c in ('file_ext', 'mime_type', 'file_size'):
                                    if self._table_has_column(conn, 'image_assets', c):
                                        if c == 'file_ext':
                                            sets.append('file_ext=%s')
                                            params.append(ext.lstrip('.'))
                                        elif c == 'mime_type':
                                            sets.append('mime_type=%s')
                                            params.append('image/*')
                                        else:
                                            sets.append('file_size=%s')
                                            params.append(len(content or b''))
                                cur2.execute(
                                    f"UPDATE image_assets SET {', '.join(sets)} WHERE id=%s",
                                    tuple(params + [old_aid]),
                                )
                                new_aid = old_aid
                            else:
                                new_aid = self._insert_image_asset_dynamic(
                                    conn,
                                    cur2,
                                    {
                                        'sha256': new_sha,
                                        'storage_path': new_storage_path,
                                        'filename': filename,
                                        'ext': ext,
                                        'file_size': len(content or b''),
                                        'image_type_id': image_type_id,
                                        'created_by': user_id,
                                        'description': '',
                                        'is_deprecated': 0,
                                    },
                                )
                                created_new_row = True
                                cur2.execute("DELETE FROM image_assets WHERE id=%s", (old_aid,))

                            self._execute_sku_mapping_upsert(
                                conn, cur2, int(new_aid), sort_order, image_type_id, variant_id, sales_product_id, user_id
                            )
                        self._tx_commit(conn)
                    except Exception as e:
                        self._tx_rollback(conn)
                        self._safe_unlink(new_abs)
                        return self.send_json({'status': 'error', 'message': f'替换失败，已回滚：{str(e)}'}, start_response)

                    # Post-commit: retire old file (DB row already gone if we created a new asset)
                    if old_abs:
                        moved_ok, _dst, _err = self._move_file_to_listing_recycle_bin(old_abs, '替换')
                        if not moved_ok:
                            self._safe_unlink(old_abs)

                    try:
                        self._rehome_image_asset_if_needed(conn, int(new_aid))
                    except Exception:
                        pass

                    return self.send_json(
                        {
                            'status': 'success',
                            'image_name': final_name,
                            'image_asset_id': int(new_aid),
                            'sha256': new_sha,
                            'created_new_asset': bool(created_new_row),
                            'message': '已替换图片：数据库已更新，原文件已移入『上架资源』/回收站',
                        },
                        start_response,
                    )
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_sales_product_main_images_upload_api(self, environ, start_response):
        try:
            method = environ['REQUEST_METHOD']
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            check_only = str((query_params.get('check_only', ['0'])[0] or '0')).lower() in ('1', 'true', 'yes', 'on')

            def _multipart_debug_payload(note, raw_body_len=None):
                try:
                    cl_hdr = int(environ.get('CONTENT_LENGTH', 0) or 0)
                except Exception:
                    cl_hdr = 0
                ct_hdr = str(environ.get('CONTENT_TYPE', '') or '')
                return {
                    'status': 'error',
                    'duplicate_count': 0,
                    'duplicates': [],
                    'file_count': 0,
                    'message': note,
                    'debug': {
                        'wsgi_request_method': str(method or ''),
                        'content_type': ct_hdr,
                        'content_length_header': cl_hdr,
                        'raw_body_bytes': raw_body_len,
                    },
                }

            # Allow GET requests only when check_only is enabled (for duplicate checking without upload)
            if method == 'GET':
                if not check_only:
                    return self.send_json({'status': 'error', 'message': 'Method not allowed. Use POST for uploads or GET with check_only=1 for duplicate checking.'}, start_response)
                # For GET with check_only, we need parameters from query string
                sales_product_id = self._parse_int((query_params.get('sales_product_id', [''])[0] or '').strip())
                if not sales_product_id:
                    return self.send_json({'status': 'error', 'message': 'Missing sales_product_id'}, start_response)
                image_type_name = (query_params.get('image_type_name', [''])[0] or '').strip() or '文字卖点图'
                # GET cannot carry multipart file bodies; treat this as a client misuse instead of a "successful" empty upload.
                # NOTE: Browsers will issue GET when opening a URL in the address bar.
                # This endpoint cannot perform duplicate checks without file bytes, so return a non-fatal "info"
                # payload (HTTP 200) to avoid looking like a broken API while still guiding correct usage.
                return self.send_json(
                    {
                        'status': 'info',
                        'duplicate_count': 0,
                        'duplicates': [],
                        'file_count': 0,
                        'message': 'check_only 预检需要 POST + multipart/form-data 携带图片文件；在地址栏 GET 打开该 URL 不会上传/不会预检文件内容。',
                    },
                    start_response,
                )

            if method != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            allow_duplicate = str((query_params.get('allow_duplicate', ['0'])[0] or '0')).lower() in ('1', 'true', 'yes', 'on')

            raw_body = self._read_wsgi_request_body(environ)
            form = None
            if raw_body:
                env_copy = dict(environ)
                env_copy['CONTENT_LENGTH'] = str(len(raw_body))
                form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            else:
                # Some servers/proxies may omit CONTENT_LENGTH (e.g., chunked transfer).
                # Fall back to streaming parse; cgi.FieldStorage will read from wsgi.input.
                form = cgi.FieldStorage(fp=environ.get('wsgi.input'), environ=environ, keep_blank_values=True)

            sales_product_id = self._parse_int((form.getfirst('sales_product_id', '') or '').strip()) if form else 0
            if not sales_product_id:
                sales_product_id = self._parse_int((query_params.get('sales_product_id', [''])[0] or '').strip())
            variant_id_upload = self._parse_int((form.getfirst('variant_id', '') or '').strip()) if form else 0
            if not variant_id_upload:
                variant_id_upload = self._parse_int((query_params.get('variant_id', [''])[0] or '').strip())
            if not sales_product_id and not variant_id_upload:
                return self.send_json({'status': 'error', 'message': 'Missing sales_product_id 或 variant_id'}, start_response)

            image_type_name = ((form.getfirst('image_type_name', '') if form else '') or '').strip()
            if not image_type_name:
                image_type_name = (query_params.get('image_type_name', [''])[0] or '').strip()
            image_type_name = image_type_name or '文字卖点图'

            uploads = []
            # Prefer raw-body multipart parsing for correct non-ASCII filenames (avoid '���')
            if raw_body:
                uploads = self._parse_multipart_uploads_fallback(content_type, raw_body)
            # Fallback to cgi.FieldStorage (may lose filename encoding)
            if not uploads:
                for p in getattr(form, 'list', []) or []:
                    if getattr(p, 'filename', None):
                        try:
                            content = p.file.read() or b''
                        except Exception:
                            content = b''
                        uploads.append({'filename': p.filename, 'content': content})
            if not uploads:
                if check_only:
                    return self.send_json(
                        _multipart_debug_payload(
                            '预检失败：未解析到任何图片文件字段。请确认请求体为 multipart/form-data 且字段名为 file；如经过 nginx 301/302 重定向，POST 可能被降级为 GET。',
                            raw_body_len=len(raw_body) if raw_body is not None else 0,
                        ),
                        start_response,
                    )
                return self.send_json({'status': 'error', 'message': 'No valid images uploaded'}, start_response)

            with self._get_db_connection() as conn:
                image_type_id = self._get_image_type_id_by_name(conn, image_type_name)
                if not image_type_id:
                    return self.send_json({'status': 'error', 'message': f'未知图片类型: {image_type_name}'}, start_response)

                user_id = None
                try:
                    user_id = self._get_session_user(environ)
                except Exception as e:
                    # Failed to get session user
                    return self.send_json({'status': 'error', 'message': f'无法验证用户身份，请确保已登录。错误：{str(e)}'}, start_response)
                
                if not user_id:
                    return self.send_json({'status': 'error', 'message': '必须登录才能上传图片'}, start_response)

                duplicates = []
                normalized = []
                skipped_files = []
                for item in uploads:
                    filename = os.path.basename(item.get('filename') or '')
                    content = item.get('content') or b''
                    if not filename:
                        skipped_files.append({'filename': '', 'reason': 'empty_name'})
                        continue
                    if not content:
                        skipped_files.append({'filename': filename, 'reason': 'empty_file'})
                        continue
                    if not self._is_image_name(filename):
                        skipped_files.append({'filename': filename, 'reason': 'unsupported_type'})
                        continue
                    sha256 = self._sha256_hex(content)
                    asset = self._find_image_asset_by_sha256(conn, sha256)
                    normalized.append({
                        'filename': filename,
                        'content': content,
                        'sha256': sha256,
                        'asset': asset,
                    })
                    if asset:
                        duplicates.append({
                            'filename': filename,
                            'sha256': sha256,
                            'image_asset_id': asset.get('id'),
                            'storage_path': asset.get('storage_path') or '',
                            'description': asset.get('description') or ''
                        })

                if not normalized:
                    return self.send_json({
                        'status': 'error',
                        'message': '未检测到可上传图片：仅支持 jpg/jpeg/png/gif/bmp/webp/tif/tiff',
                        'file_count': 0,
                        'skipped_files': skipped_files,
                    }, start_response)

                if check_only:
                    return self.send_json({
                        'status': 'success',
                        'mode': 'preflight',
                        'persisted': False,
                        'duplicate_count': len(duplicates),
                        'duplicates': duplicates,
                        'file_count': len(normalized)
                    }, start_response)

                if duplicates and not allow_duplicate:
                    return self.send_json({
                        'status': 'duplicate',
                        'message': '检测到重复图片，请确认是否复用已有图片',
                        'duplicate_count': len(duplicates),
                        'duplicates': duplicates,
                        'file_count': len(normalized)
                    }, start_response)

                # Target folder: 货号/主图/规格-面料
                if variant_id_upload and not sales_product_id:
                    try:
                        folder_info = self._resolve_sales_variant_folder_by_variant_id(variant_id_upload, ensure_folder=True)
                    except Exception as e:
                        return self.send_json({'status': 'error', 'message': str(e)}, start_response)
                    variant_id = int(variant_id_upload)
                    start_sort = self._get_variant_image_sort_start(conn, variant_id)
                else:
                    folder_info = self._resolve_sales_product_variant_folder(sales_product_id, ensure_folder=True)
                    variant_id = 0
                    try:
                        with conn.cursor() as cur:
                            cur.execute("SELECT variant_id FROM sales_products WHERE id=%s", (sales_product_id,))
                            row = cur.fetchone() or {}
                            variant_id = self._parse_int(row.get('variant_id')) or 0
                    except Exception:
                        variant_id = 0
                    start_sort = self._get_sales_product_image_sort_start(conn, sales_product_id)
                target_folder_abs = folder_info.get('folder_path')
                if not target_folder_abs or not os.path.exists(target_folder_abs):
                    return self.send_json({'status': 'error', 'message': '无法定位主图文件夹，请确认货号与面料完整（面料必填；规格可为空）'}, start_response)
                created_assets = 0
                reused_assets = 0
                linked = 0
                results = []
                created_files = []   # abs paths to delete on rollback
                deferred_links = []  # (src_abs, dst_abs)
                to_insert_assets = []  # dicts for new assets: {sha256, storage_path, filename, ext, file_size, abs_path}

                for idx, item in enumerate(normalized, start=1):
                    filename = item['filename']
                    content = item['content']
                    sha256 = item['sha256']
                    asset = item['asset']
                    ext = self._guess_image_ext(filename, content)
                    if asset:
                        asset_id = asset.get('id')
                        reused_assets += 1
                        # Defer link creation until after DB commit (atomic batch)
                        try:
                            src_abs = self._join_resources(asset.get('storage_path') or '')
                            if src_abs and os.path.exists(src_abs):
                                type_part = self._sanitize_filename_component(image_type_name, 32) or '图片'
                                base_part = self._sanitize_filename_component(os.path.splitext(filename)[0], 80) or 'image'
                                link_name = self._next_available_filename(target_folder_abs, f"{type_part}-{base_part}{ext}")
                                dst_abs = os.path.join(target_folder_abs, self._safe_fsencode(link_name))
                                deferred_links.append((src_abs, dst_abs))
                        except Exception:
                            pass
                    else:
                        # Save into: 货号/主图/规格-面料/ <类型>-<原文件名>[_xx].ext
                        type_part = self._sanitize_filename_component(image_type_name, 32) or '图片'
                        base_part = self._sanitize_filename_component(os.path.splitext(filename)[0], 80) or sha256[:12]
                        final_name = self._next_available_filename(target_folder_abs, f"{type_part}-{base_part}{ext}")
                        abs_path = os.path.join(target_folder_abs, self._safe_fsencode(final_name))
                        try:
                            with open(abs_path, 'wb') as f:
                                f.write(content or b'')
                        except Exception:
                            # If we cannot persist the file, do not write DB rows.
                            raise RuntimeError(f'写入图片文件失败: {filename}')
                        created_files.append(abs_path)

                        # Compute storage_path relative to RESOURCES root
                        try:
                            res_root = self._join_resources('')
                            rel_bytes = os.path.relpath(abs_path, res_root)
                            storage_path = os.fsdecode(rel_bytes).replace('\\', '/')
                        except Exception:
                            storage_path = ''
                        to_insert_assets.append({
                            'sha256': sha256,
                            'storage_path': storage_path,
                            'filename': filename,
                            'ext': ext,
                            'file_size': len(content or b''),
                            'abs_path': abs_path,
                        })
                        asset_id = None

                    sort_order = start_sort + idx
                    results.append({
                        'filename': filename,
                        'sha256': sha256,
                        'image_asset_id': asset_id,
                        'sort_order': sort_order,
                    })

                # ---- Transaction: write DB for the whole batch ----
                try:
                    self._tx_begin(conn)

                    sha_to_asset_id = {}
                    for rec in to_insert_assets:
                        with conn.cursor() as cur:
                            aid = self._insert_image_asset_dynamic(
                                conn,
                                cur,
                                {
                                    'sha256': rec['sha256'],
                                    'storage_path': rec['storage_path'],
                                    'filename': rec['filename'],
                                    'ext': rec['ext'],
                                    'file_size': rec['file_size'],
                                    'image_type_id': image_type_id,
                                    'created_by': user_id,
                                },
                            )
                        sha_to_asset_id[rec['sha256']] = aid
                        created_assets += 1

                    # Resolve reused assets again inside tx (stable)
                    for item in normalized:
                        if item.get('asset') and item.get('sha256'):
                            sha_to_asset_id.setdefault(item['sha256'], int(item['asset'].get('id') or 0))

                    # Insert mappings (variant preferred)
                    for idx, row in enumerate(results, start=1):
                        sha = row.get('sha256')
                        aid = sha_to_asset_id.get(sha) or 0
                        if not aid:
                            raise RuntimeError('无法解析 image_asset_id')
                        row['image_asset_id'] = aid
                        sort_order = row.get('sort_order')
                        with conn.cursor() as cur:
                            self._execute_sku_mapping_upsert(
                                conn, cur, aid, sort_order, image_type_id, variant_id, sales_product_id, user_id
                            )
                        linked += 1

                    self._tx_commit(conn)
                except Exception as e:
                    self._tx_rollback(conn)
                    # Cleanup any newly written files
                    for p in created_files:
                        self._safe_unlink(p)
                    # Also cleanup any links we may have created (we deferred, so none here)
                    return self.send_json({'status': 'error', 'message': f'上传失败，已回滚：{str(e)}'}, start_response)

                # After commit: create deferred links best-effort
                for src_abs, dst_abs in deferred_links:
                    try:
                        self._try_create_link(src_abs, dst_abs)
                    except Exception:
                        pass
                # After commit: apply rehome best-effort
                try:
                    for row in results:
                        self._rehome_image_asset_if_needed(conn, row.get('image_asset_id'))
                except Exception:
                    pass

                return self.send_json({
                    'status': 'success',
                    'mode': 'upload',
                    'persisted': True,
                    'files': [x['filename'] for x in results],
                    'created_assets': created_assets,
                    'reused_assets': reused_assets,
                    'linked': linked,
                    'duplicates': duplicates
                }, start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _table_exists_simple(self, conn, table_name):
        name = str(table_name or '').strip()
        if not name:
            return False
        try:
            with conn.cursor() as cur:
                cur.execute("SHOW TABLES LIKE %s", (name,))
                return bool(cur.fetchone())
        except Exception:
            return False

    # -------------------------------------------------------------------------
    # 产品表现：周/月聚合与 30 天 rolling 快照（含下单 SKU op_rolling）
    # -------------------------------------------------------------------------

    def _sales_perf_optional_product_id_list(self, sales_product_ids):
        """None=全量刷新；否则返回去重后的 sales_product_id 列表（可为空）。"""
        if sales_product_ids is None:
            return None
        return sorted({int(x) for x in (sales_product_ids or []) if self._parse_int(x)})

    def _sales_perf_month_end(self, month_start):
        if month_start.month == 12:
            nxt = month_start.replace(year=month_start.year + 1, month=1, day=1)
        else:
            nxt = month_start.replace(month=month_start.month + 1, day=1)
        return nxt - timedelta(days=1)

    def _sales_perf_parse_record_date_bounds(self, start_date, end_date):
        """解析导入/刷新的起止日期；失败返回 (None, None)。"""
        try:
            sd = datetime.strptime(str(start_date or '').strip(), '%Y-%m-%d').date()
            ed = datetime.strptime(str(end_date or '').strip(), '%Y-%m-%d').date()
        except Exception:
            return None, None
        if ed < sd:
            sd, ed = ed, sd
        return sd, ed

    def _sales_perf_agg_month_bounds(self, sd, ed):
        """月聚合：仅覆盖 [sd,ed] 所涉及的自然月（首尾月整月重算）。"""
        return sd.replace(day=1), self._sales_perf_month_end(ed)

    def _sales_perf_agg_week_bounds(self, sd, ed):
        """周聚合：覆盖 [sd,ed] 所涉及的自然周（周一至周日）。"""
        week_lo = sd - timedelta(days=sd.weekday())
        week_hi = ed + timedelta(days=(6 - ed.weekday()))
        return week_lo, week_hi

    def _sales_perf_expand_agg_bounds(self, start_date, end_date):
        """将刷新区间扩到完整自然月（供仅需日期字符串边界的调用方）。"""
        sd, ed = self._sales_perf_parse_record_date_bounds(start_date, end_date)
        if not sd:
            return start_date, end_date
        month_lo, month_hi = self._sales_perf_agg_month_bounds(sd, ed)
        return month_lo.strftime('%Y-%m-%d'), month_hi.strftime('%Y-%m-%d')

    def _refresh_sales_perf_agg_range(self, conn, start_date, end_date, sales_product_ids=None, progress_hook=None, segments=None):
        """
        方案A：按给定日期范围，刷新周/月聚合快照表。
        - start_date/end_date: 'YYYY-MM-DD'
        - sales_product_ids: 可选，仅刷新涉及到的销售产品（导入场景强烈建议传入，避免全表扫描导致超时）
        - progress_hook: 可选 callable(dict)，在耗时循环中节流回调；dict 含 step/total/segment/period_key
        - segments: 可选 ('month',) / ('week',) / ('month', 'week')，默认先月后周（导入优先保证月聚合及时）
        """
        if not conn:
            return
        s = str(start_date or '').strip()
        e = str(end_date or '').strip()
        if not s or not e:
            return
        sd, ed = self._sales_perf_parse_record_date_bounds(s, e)
        if not sd:
            return
        # 月、周分别扩界：切勿把「周」起点并进月列表（如 4/1 所在周从 3/30 起算会误刷 3 月月聚合）
        month_lo, month_last = self._sales_perf_agg_month_bounds(sd, ed)
        week_lo, week_hi = self._sales_perf_agg_week_bounds(sd, ed)

        if (not self._table_exists_simple(conn, 'sales_perf_agg_week')) or (not self._table_exists_simple(conn, 'sales_perf_agg_month')):
            raise RuntimeError('聚合快照表不存在：请先在数据库执行 scripts/sql/20260427_01_sales_perf_agg_tables.sql')

        seg_order = tuple(segments) if segments else ('month', 'week')
        do_month = 'month' in seg_order
        do_week = 'week' in seg_order

        # 分段刷新：避免单条大 SQL 导致 MySQL 超时断连（2013 timed out）
        def _iter_weeks(week_start, week_end):
            cur = week_start - timedelta(days=week_start.weekday())
            end = week_end - timedelta(days=week_end.weekday())
            while cur <= end:
                yield cur
                cur = cur + timedelta(days=7)

        def _iter_months(month_start, month_last_day):
            cur = month_start.replace(day=1)
            end = month_last_day.replace(day=1)
            while cur <= end:
                yield cur
                if cur.month == 12:
                    cur = cur.replace(year=cur.year + 1, month=1, day=1)
                else:
                    cur = cur.replace(month=cur.month + 1, day=1)

        ids = None
        if sales_product_ids is not None:
            ids = self._sales_perf_optional_product_id_list(sales_product_ids)

        id_chunk_size = 300
        id_chunks = []
        if ids is None:
            id_chunks = [None]
        elif ids:
            for i in range(0, len(ids), id_chunk_size):
                id_chunks.append(ids[i:i + id_chunk_size])
        else:
            id_chunks = [None]

        weeks_list = list(_iter_weeks(week_lo, week_hi)) if do_week else []
        months_list = list(_iter_months(month_lo, month_last)) if do_month else []
        total_steps = len(weeks_list) * max(1, len(id_chunks)) + len(months_list) * max(1, len(id_chunks))
        total_steps = max(1, total_steps)
        step_i = 0
        last_emit_ts = 0.0

        def _emit_agg_progress(segment, period_key, pending=False):
            nonlocal step_i, last_emit_ts
            if not progress_hook:
                return
            now = time.time()
            if not pending:
                step_i += 1
            if pending and (now - last_emit_ts) < 0.8:
                return
            if (not pending) and step_i < total_steps and (now - last_emit_ts) < 0.5:
                return
            last_emit_ts = now
            try:
                progress_hook({
                    'step': step_i if not pending else min(step_i + 1, total_steps),
                    'total': total_steps,
                    'segment': segment,
                    'period_key': period_key,
                    'pending': bool(pending),
                })
            except Exception:
                pass

        with conn.cursor() as cur:
            # MONTH first (import/dashboard rely on monthly agg; must finish before long week pass)
            for ms in months_list:
                me = self._sales_perf_month_end(ms)
                year_month = int(ms.year) * 100 + int(ms.month)
                _emit_agg_progress('month', str(year_month), pending=True)
                for id_chunk in id_chunks:
                    if id_chunk:
                        placeholders = ','.join(['%s'] * len(id_chunk))
                        cur.execute(
                            f"DELETE FROM sales_perf_agg_month WHERE month_start=%s AND sales_product_id IN ({placeholders})",
                            tuple([ms.strftime('%Y-%m-%d')] + id_chunk),
                        )
                        id_filter_sql = f" AND spp.sales_product_id IN ({placeholders})"
                        id_params = list(id_chunk)
                    else:
                        cur.execute("DELETE FROM sales_perf_agg_month WHERE month_start=%s", (ms.strftime('%Y-%m-%d'),))
                        id_filter_sql = ''
                        id_params = []
                    cur.execute(
                        f"""
                        INSERT INTO sales_perf_agg_month
                            (sales_product_id, month_start, month_end, `year_month`, source_rows,
                             sales_qty, net_sales_amount, order_qty, session_total,
                             ad_impressions, ad_clicks, ad_orders, ad_spend, ad_sales_amount,
                             refund_amount)
                        SELECT
                            spp.sales_product_id,
                            %s AS month_start,
                            %s AS month_end,
                            %s AS `year_month`,
                            COUNT(1) AS source_rows,
                            SUM(COALESCE(spp.sales_qty,0)) AS sales_qty,
                            SUM(COALESCE(spp.net_sales_amount,0)) AS net_sales_amount,
                            SUM(COALESCE(spp.order_qty,0)) AS order_qty,
                            SUM(COALESCE(spp.session_total,0)) AS session_total,
                            SUM(COALESCE(spp.ad_impressions,0)) AS ad_impressions,
                            SUM(COALESCE(spp.ad_clicks,0)) AS ad_clicks,
                            SUM(COALESCE(spp.ad_orders,0)) AS ad_orders,
                            SUM(COALESCE(spp.ad_spend,0)) AS ad_spend,
                            SUM(COALESCE(spp.ad_sales_amount,0)) AS ad_sales_amount,
                            SUM(COALESCE(spp.refund_amount,0)) AS refund_amount
                        FROM sales_product_performances spp
                        WHERE spp.record_date >= %s AND spp.record_date <= %s
                        {id_filter_sql}
                        GROUP BY spp.sales_product_id
                        """,
                        tuple([ms.strftime('%Y-%m-%d'), me.strftime('%Y-%m-%d'), year_month, ms.strftime('%Y-%m-%d'), me.strftime('%Y-%m-%d')] + id_params),
                    )
                conn.commit()
                _emit_agg_progress('month', str(year_month), pending=False)

            # WEEK: one week per query
            for ws in weeks_list:
                we = ws + timedelta(days=6)
                year_week = int(ws.isocalendar().year) * 100 + int(ws.isocalendar().week)
                _emit_agg_progress('week', str(year_week), pending=True)
                for id_chunk in id_chunks:
                    if id_chunk:
                        placeholders = ','.join(['%s'] * len(id_chunk))
                        cur.execute(
                            f"DELETE FROM sales_perf_agg_week WHERE week_start=%s AND sales_product_id IN ({placeholders})",
                            tuple([ws.strftime('%Y-%m-%d')] + id_chunk),
                        )
                        id_filter_sql = f" AND spp.sales_product_id IN ({placeholders})"
                        id_params = list(id_chunk)
                    else:
                        cur.execute("DELETE FROM sales_perf_agg_week WHERE week_start=%s", (ws.strftime('%Y-%m-%d'),))
                        id_filter_sql = ''
                        id_params = []
                    cur.execute(
                        f"""
                        INSERT INTO sales_perf_agg_week
                            (sales_product_id, week_start, week_end, `year_week`, source_rows,
                             sales_qty, net_sales_amount, order_qty, session_total,
                             ad_impressions, ad_clicks, ad_orders, ad_spend, ad_sales_amount,
                             refund_amount)
                        SELECT
                            spp.sales_product_id,
                            %s AS week_start,
                            %s AS week_end,
                            %s AS `year_week`,
                            COUNT(1) AS source_rows,
                            SUM(COALESCE(spp.sales_qty,0)) AS sales_qty,
                            SUM(COALESCE(spp.net_sales_amount,0)) AS net_sales_amount,
                            SUM(COALESCE(spp.order_qty,0)) AS order_qty,
                            SUM(COALESCE(spp.session_total,0)) AS session_total,
                            SUM(COALESCE(spp.ad_impressions,0)) AS ad_impressions,
                            SUM(COALESCE(spp.ad_clicks,0)) AS ad_clicks,
                            SUM(COALESCE(spp.ad_orders,0)) AS ad_orders,
                            SUM(COALESCE(spp.ad_spend,0)) AS ad_spend,
                            SUM(COALESCE(spp.ad_sales_amount,0)) AS ad_sales_amount,
                            SUM(COALESCE(spp.refund_amount,0)) AS refund_amount
                        FROM sales_product_performances spp
                        WHERE spp.record_date >= %s AND spp.record_date <= %s
                        {id_filter_sql}
                        GROUP BY spp.sales_product_id
                        """,
                        tuple([ws.strftime('%Y-%m-%d'), we.strftime('%Y-%m-%d'), year_week, ws.strftime('%Y-%m-%d'), we.strftime('%Y-%m-%d')] + id_params),
                    )
                conn.commit()
                _emit_agg_progress('week', str(year_week), pending=False)

        # per-period commits done above

        try:
            self._refresh_sales_perf_rolling_30d(conn, sales_product_ids=ids if ids else None)
        except Exception:
            pass

    def _refresh_sales_perf_rolling_30d(self, conn, sales_product_ids=None):
        """刷新动销月分母：全局最新 record_date 向前 30 天（含首尾）销量快照。"""
        if not conn:
            return
        if not self._table_exists_simple(conn, 'sales_perf_rolling_30d'):
            return
        window = self._turnover_sales_window(conn)
        ws = str(window.get('window_start') or '')[:10]
        we = str(window.get('window_end') or '')[:10]
        anchor = str(window.get('anchor_date') or '')[:10]
        if not ws or not we or not anchor:
            return

        id_list = self._sales_perf_optional_product_id_list(sales_product_ids)
        if sales_product_ids is not None and not id_list:
            return

        id_chunk_size = 300
        id_chunks = [None] if id_list is None else [
            id_list[i:i + id_chunk_size] for i in range(0, len(id_list), id_chunk_size)
        ]

        with conn.cursor() as cur:
            cur.execute('SELECT anchor_date FROM sales_perf_rolling_30d LIMIT 1')
            anchor_row = cur.fetchone() or {}
            old_anchor = str(anchor_row.get('anchor_date') or '')[:10]
            full_refresh = (not old_anchor) or (old_anchor != anchor) or (id_list is None)
            if full_refresh and old_anchor and old_anchor != anchor:
                cur.execute('DELETE FROM sales_perf_rolling_30d')

            for chunk in id_chunks:
                if chunk:
                    ph = ','.join(['%s'] * len(chunk))
                    cur.execute(
                        f'DELETE FROM sales_perf_rolling_30d WHERE sales_product_id IN ({ph})',
                        tuple(chunk),
                    )
                    id_filter_sql = f' AND spp.sales_product_id IN ({ph}) '
                    params = [anchor, ws, we, ws, we] + list(chunk)
                else:
                    cur.execute('DELETE FROM sales_perf_rolling_30d')
                    id_filter_sql = ''
                    params = [anchor, ws, we, ws, we]
                cur.execute(
                    f"""
                    INSERT INTO sales_perf_rolling_30d
                        (sales_product_id, anchor_date, window_start, window_end, sales_qty, net_sales_amount)
                    SELECT spp.sales_product_id,
                           %s AS anchor_date,
                           %s AS window_start,
                           %s AS window_end,
                           SUM(COALESCE(spp.sales_qty, 0)) AS sales_qty,
                           SUM(COALESCE(spp.net_sales_amount, 0)) AS net_sales_amount
                    FROM sales_product_performances spp
                    WHERE spp.record_date >= %s
                      AND spp.record_date <= %s
                      {id_filter_sql}
                    GROUP BY spp.sales_product_id
                    ON DUPLICATE KEY UPDATE
                        anchor_date = VALUES(anchor_date),
                        window_start = VALUES(window_start),
                        window_end = VALUES(window_end),
                        sales_qty = VALUES(sales_qty),
                        net_sales_amount = VALUES(net_sales_amount)
                    """,
                    tuple(params),
                )
        window = {
            'anchor_date': anchor,
            'window_start': ws,
            'window_end': we,
        }
        self._refresh_sales_perf_op_rolling_30d(conn, window, sales_product_ids=id_list)
        conn.commit()

    def _refresh_sales_perf_op_rolling_30d(self, conn, window, sales_product_ids=None):
        """刷新下单 SKU 近 30 天销量快照（链接变体×BOM，窗口与 sales_perf_rolling_30d 一致）。"""
        if not conn or not window:
            return
        if not self._table_exists_simple(conn, 'sales_perf_op_rolling_30d'):
            return
        ws = str(window.get('window_start') or '')[:10]
        we = str(window.get('window_end') or '')[:10]
        anchor = str(window.get('anchor_date') or '')[:10]
        if not ws or not we or not anchor:
            return

        sp_ids = self._sales_perf_optional_product_id_list(sales_product_ids)
        if sales_product_ids is not None and not sp_ids:
            return

        with conn.cursor() as cur:
            cur.execute('SELECT anchor_date FROM sales_perf_op_rolling_30d LIMIT 1')
            anchor_row = cur.fetchone() or {}
            old_anchor = str(anchor_row.get('anchor_date') or '')[:10]
            if old_anchor and old_anchor != anchor:
                sp_ids = None

            op_filter_ids = None
            if sp_ids is not None:
                sp_ph = ','.join(['%s'] * len(sp_ids))
                cur.execute(
                    f"""
                    SELECT DISTINCT l.order_product_id
                    FROM sales_variant_order_links l
                    INNER JOIN sales_products sp ON sp.variant_id = l.variant_id
                    WHERE sp.id IN ({sp_ph})
                    """,
                    tuple(sp_ids),
                )
                op_filter_ids = sorted({
                    int(r.get('order_product_id'))
                    for r in (cur.fetchall() or [])
                    if self._parse_int(r.get('order_product_id'))
                })
                if not op_filter_ids:
                    return

            if sp_ids is None:
                cur.execute('DELETE FROM sales_perf_op_rolling_30d')
                op_id_filter_sql = ''
                params = [anchor, ws, we, ws, we]
            else:
                op_ph = ','.join(['%s'] * len(op_filter_ids))
                cur.execute(
                    f'DELETE FROM sales_perf_op_rolling_30d WHERE order_product_id IN ({op_ph})',
                    tuple(op_filter_ids),
                )
                op_id_filter_sql = f' AND l.order_product_id IN ({op_ph}) '
                params = [anchor, ws, we, ws, we] + list(op_filter_ids)

            cur.execute(
                f"""
                INSERT INTO sales_perf_op_rolling_30d
                    (order_product_id, anchor_date, window_start, window_end, sales_qty)
                SELECT l.order_product_id,
                       %s AS anchor_date,
                       %s AS window_start,
                       %s AS window_end,
                       SUM(COALESCE(r.sales_qty, 0) * GREATEST(1, COALESCE(l.quantity, 1))) AS sales_qty
                FROM sales_variant_order_links l
                INNER JOIN sales_products sp ON sp.variant_id = l.variant_id
                INNER JOIN sales_perf_rolling_30d r
                    ON r.sales_product_id = sp.id
                   AND r.window_start = %s
                   AND r.window_end = %s
                WHERE 1=1
                {op_id_filter_sql}
                GROUP BY l.order_product_id
                ON DUPLICATE KEY UPDATE
                    anchor_date = VALUES(anchor_date),
                    window_start = VALUES(window_start),
                    window_end = VALUES(window_end),
                    sales_qty = VALUES(sales_qty)
                """,
                tuple(params),
            )

    def _refresh_sales_perf_agg_for_deleted_records(self, conn, affected_rows):
        """按被删记录所在的精确月/周桶刷新聚合，避免 min~max 日期区间全量扫描。"""
        if not conn or not affected_rows:
            return
        if (not self._table_exists_simple(conn, 'sales_perf_agg_week')) or (not self._table_exists_simple(conn, 'sales_perf_agg_month')):
            return

        by_month = {}
        by_week = {}
        for row in (affected_rows or []):
            rd = str(row.get('record_date') or '').strip()
            if hasattr(rd, 'strftime'):
                rd = rd.strftime('%Y-%m-%d')
            else:
                rd = str(rd)[:10]
            spid = self._parse_int(row.get('sales_product_id'))
            if not rd or not spid:
                continue
            try:
                d = datetime.strptime(rd, '%Y-%m-%d').date()
            except Exception:
                continue
            ms = d.replace(day=1)
            ws = d - timedelta(days=d.weekday())
            by_month.setdefault(ms, set()).add(spid)
            by_week.setdefault(ws, set()).add(spid)
        if not by_month and not by_week:
            return

        id_chunk_size = 300

        def _id_chunks(id_set):
            ids = sorted(id_set or [])
            if not ids:
                return []
            out = []
            for i in range(0, len(ids), id_chunk_size):
                out.append(ids[i:i + id_chunk_size])
            return out

        with conn.cursor() as cur:
            for ms in sorted(by_month.keys()):
                me = self._sales_perf_month_end(ms)
                year_month = int(ms.year) * 100 + int(ms.month)
                for id_chunk in _id_chunks(by_month.get(ms)):
                    placeholders = ','.join(['%s'] * len(id_chunk))
                    cur.execute(
                        f"DELETE FROM sales_perf_agg_month WHERE month_start=%s AND sales_product_id IN ({placeholders})",
                        tuple([ms.strftime('%Y-%m-%d')] + id_chunk),
                    )
                    cur.execute(
                        f"""
                        INSERT INTO sales_perf_agg_month
                            (sales_product_id, month_start, month_end, `year_month`, source_rows,
                             sales_qty, net_sales_amount, order_qty, session_total,
                             ad_impressions, ad_clicks, ad_orders, ad_spend, ad_sales_amount,
                             refund_amount)
                        SELECT
                            spp.sales_product_id,
                            %s AS month_start,
                            %s AS month_end,
                            %s AS `year_month`,
                            COUNT(1) AS source_rows,
                            SUM(COALESCE(spp.sales_qty,0)) AS sales_qty,
                            SUM(COALESCE(spp.net_sales_amount,0)) AS net_sales_amount,
                            SUM(COALESCE(spp.order_qty,0)) AS order_qty,
                            SUM(COALESCE(spp.session_total,0)) AS session_total,
                            SUM(COALESCE(spp.ad_impressions,0)) AS ad_impressions,
                            SUM(COALESCE(spp.ad_clicks,0)) AS ad_clicks,
                            SUM(COALESCE(spp.ad_orders,0)) AS ad_orders,
                            SUM(COALESCE(spp.ad_spend,0)) AS ad_spend,
                            SUM(COALESCE(spp.ad_sales_amount,0)) AS ad_sales_amount,
                            SUM(COALESCE(spp.refund_amount,0)) AS refund_amount
                        FROM sales_product_performances spp
                        WHERE spp.record_date >= %s AND spp.record_date <= %s
                          AND spp.sales_product_id IN ({placeholders})
                        GROUP BY spp.sales_product_id
                        """,
                        tuple([
                            ms.strftime('%Y-%m-%d'), me.strftime('%Y-%m-%d'), year_month,
                            ms.strftime('%Y-%m-%d'), me.strftime('%Y-%m-%d'),
                        ] + id_chunk),
                    )
                conn.commit()

            for ws in sorted(by_week.keys()):
                we = ws + timedelta(days=6)
                year_week = int(ws.isocalendar().year) * 100 + int(ws.isocalendar().week)
                for id_chunk in _id_chunks(by_week.get(ws)):
                    placeholders = ','.join(['%s'] * len(id_chunk))
                    cur.execute(
                        f"DELETE FROM sales_perf_agg_week WHERE week_start=%s AND sales_product_id IN ({placeholders})",
                        tuple([ws.strftime('%Y-%m-%d')] + id_chunk),
                    )
                    cur.execute(
                        f"""
                        INSERT INTO sales_perf_agg_week
                            (sales_product_id, week_start, week_end, `year_week`, source_rows,
                             sales_qty, net_sales_amount, order_qty, session_total,
                             ad_impressions, ad_clicks, ad_orders, ad_spend, ad_sales_amount,
                             refund_amount)
                        SELECT
                            spp.sales_product_id,
                            %s AS week_start,
                            %s AS week_end,
                            %s AS `year_week`,
                            COUNT(1) AS source_rows,
                            SUM(COALESCE(spp.sales_qty,0)) AS sales_qty,
                            SUM(COALESCE(spp.net_sales_amount,0)) AS net_sales_amount,
                            SUM(COALESCE(spp.order_qty,0)) AS order_qty,
                            SUM(COALESCE(spp.session_total,0)) AS session_total,
                            SUM(COALESCE(spp.ad_impressions,0)) AS ad_impressions,
                            SUM(COALESCE(spp.ad_clicks,0)) AS ad_clicks,
                            SUM(COALESCE(spp.ad_orders,0)) AS ad_orders,
                            SUM(COALESCE(spp.ad_spend,0)) AS ad_spend,
                            SUM(COALESCE(spp.ad_sales_amount,0)) AS ad_sales_amount,
                            SUM(COALESCE(spp.refund_amount,0)) AS refund_amount
                        FROM sales_product_performances spp
                        WHERE spp.record_date >= %s AND spp.record_date <= %s
                          AND spp.sales_product_id IN ({placeholders})
                        GROUP BY spp.sales_product_id
                        """,
                        tuple([
                            ws.strftime('%Y-%m-%d'), we.strftime('%Y-%m-%d'), year_week,
                            ws.strftime('%Y-%m-%d'), we.strftime('%Y-%m-%d'),
                        ] + id_chunk),
                    )
                conn.commit()

        touched_spids = sorted({int(x) for s in (by_month or {}).values() for x in s})
        if touched_spids:
            try:
                self._refresh_sales_perf_rolling_30d(conn, sales_product_ids=touched_spids)
            except Exception:
                pass

    # -------------------------------------------------------------------------
    # 产品表现 CRUD / 导入 / rolling 刷新
    # -------------------------------------------------------------------------

    def handle_sales_product_performance_api(self, environ, method, start_response):
        """产品表现 CRUD / 导入；写入后触发周月聚合与 rolling 快照刷新。"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))

            def _resolve_sales_product_id(conn, value):
                item_id = self._parse_int(value)
                if item_id:
                    return item_id
                sku = str(value or '').strip()
                if not sku:
                    return None
                with conn.cursor() as cur:
                    cur.execute("SELECT id FROM sales_products WHERE platform_sku=%s LIMIT 1", (sku,))
                    row = cur.fetchone() or {}
                return self._parse_int(row.get('id'))

            def _normalize_date_text(value):
                if value is None:
                    return ''
                if isinstance(value, datetime):
                    return value.strftime('%Y-%m-%d')
                text = str(value).strip()
                if not text:
                    return ''
                for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
                    try:
                        return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
                    except Exception:
                        continue
                return text[:10]

            if method == 'GET':
                def _parse_csv_int_list(name):
                    values = []
                    for raw in query_params.get(name, []):
                        for token in re.split(r'[,，;；\s]+', str(raw or '').strip()):
                            val = self._parse_int(token)
                            if val and val not in values:
                                values.append(val)
                    return values

                keyword = (query_params.get('q', [''])[0] or '').strip()
                item_id = self._parse_int((query_params.get('id', [''])[0] or '').strip())
                shop_ids = _parse_csv_int_list('shop_ids')
                date_from = _normalize_date_text((query_params.get('date_from', [''])[0] or '').strip())
                date_to = _normalize_date_text((query_params.get('date_to', [''])[0] or '').strip())
                page_size = min(1000, max(10, self._parse_int((query_params.get('page_size', ['50'])[0] or '50')) or 50))
                page = max(1, self._parse_int((query_params.get('page', ['1'])[0] or '1')) or 1)
                limit = min(5000, max(1, self._parse_int((query_params.get('limit', [str(page_size)])[0] or str(page_size))) or page_size))
                page_size = min(page_size, limit)
                offset = (page - 1) * page_size

                base_sql = """
                    FROM sales_product_performances spp
                    JOIN sales_products sp ON sp.id = spp.sales_product_id
                    LEFT JOIN shops sh ON sh.id = sp.shop_id
                    LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                    LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                """
                data_sql = """
                    SELECT spp.*, sp.platform_sku, sp.shop_id, sh.shop_name,
                           v.sku_family_id AS sku_family_id, pf.sku_family
                """
                params = []
                filters = []
                if item_id:
                    filters.append('spp.id=%s')
                    params.append(item_id)
                if keyword:
                    like_kw = f'%{keyword}%'
                    filters.append('(sp.platform_sku LIKE %s OR pf.sku_family LIKE %s)')
                    params.extend([like_kw, like_kw])
                if shop_ids and not item_id:
                    filters.append(f"sp.shop_id IN ({','.join(['%s'] * len(shop_ids))})")
                    params.extend(shop_ids)
                if not item_id and not keyword:
                    if not date_from and not date_to:
                        try:
                            date_to = datetime.now().strftime('%Y-%m-%d')
                            date_from = (datetime.now().date() - timedelta(days=179)).strftime('%Y-%m-%d')
                        except Exception:
                            date_from = ''
                            date_to = ''
                    if date_from:
                        filters.append('spp.record_date >= %s')
                        params.append(date_from)
                    if date_to:
                        filters.append('spp.record_date <= %s')
                        params.append(date_to)
                if filters:
                    where_sql = ' WHERE ' + ' AND '.join(filters)
                else:
                    where_sql = ''

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if item_id:
                            cur.execute(data_sql + base_sql + where_sql + ' ORDER BY spp.record_date DESC, spp.id DESC LIMIT 1', params)
                            rows = cur.fetchall() or []
                            return self.send_json({'status': 'success', 'item': rows[0] if rows else None}, start_response)

                        cur.execute('SELECT COUNT(1) AS cnt ' + base_sql + where_sql, params)
                        total = int((cur.fetchone() or {}).get('cnt') or 0)

                        data_params = list(params)
                        data_params.extend([offset, page_size])
                        cur.execute(
                            data_sql + base_sql + where_sql + ' ORDER BY spp.record_date DESC, spp.id DESC LIMIT %s, %s',
                            data_params
                        )
                        rows = cur.fetchall() or []
                return self.send_json({
                    'status': 'success',
                    'items': rows,
                    'page': page,
                    'page_size': page_size,
                    'total': total,
                    'total_pages': (total + page_size - 1) // page_size if page_size else 1
                }, start_response)

            if method in ('POST', 'PUT'):
                data = self._read_json_body(environ)
                performance_id = self._parse_int(data.get('id'))
                sales_product_ref = data.get('sales_product_id') or data.get('platform_sku')
                record_date = _normalize_date_text(data.get('record_date'))
                if not record_date:
                    return self.send_json({'status': 'error', 'message': 'Missing record_date'}, start_response)

                with self._get_db_connection() as conn:
                    sales_product_id = _resolve_sales_product_id(conn, sales_product_ref)
                    if not sales_product_id:
                        return self.send_json({'status': 'error', 'message': '无法根据销售平台SKU找到销售产品'}, start_response)

                    values = {
                        'sales_qty': self._parse_int(data.get('sales_qty')) or 0,
                        'net_sales_amount': self._parse_float(data.get('net_sales_amount')) or 0,
                        'order_qty': self._parse_int(data.get('order_qty')) or 0,
                        'session_total': self._parse_int(data.get('session_total')) or 0,
                        'ad_impressions': self._parse_int(data.get('ad_impressions')) or 0,
                        'ad_clicks': self._parse_int(data.get('ad_clicks')) or 0,
                        'ad_orders': self._parse_int(data.get('ad_orders')) or 0,
                        'ad_spend': self._parse_float(data.get('ad_spend')) or 0,
                        'ad_sales_amount': self._parse_float(data.get('ad_sales_amount')) or 0,
                        'refund_amount': self._parse_float(data.get('refund_amount')) or 0,
                    }

                    if performance_id and method == 'PUT':
                        with conn.cursor() as cur:
                            cur.execute(
                                """
                                UPDATE sales_product_performances
                                SET sales_product_id=%s, record_date=%s, sales_qty=%s, net_sales_amount=%s,
                                    order_qty=%s, session_total=%s, ad_impressions=%s, ad_clicks=%s,
                                    ad_orders=%s, ad_spend=%s, ad_sales_amount=%s, refund_amount=%s
                                WHERE id=%s
                                """,
                                (
                                    sales_product_id, record_date, values['sales_qty'], values['net_sales_amount'],
                                    values['order_qty'], values['session_total'], values['ad_impressions'], values['ad_clicks'],
                                    values['ad_orders'], values['ad_spend'], values['ad_sales_amount'], values['refund_amount'],
                                    performance_id
                                )
                            )
                        try:
                            self._refresh_sales_perf_agg_range(
                                conn, record_date, record_date, sales_product_ids=[sales_product_id],
                            )
                        except Exception:
                            pass
                        return self.send_json({'status': 'success', 'id': performance_id}, start_response)

                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO sales_product_performances
                            (sales_product_id, record_date, sales_qty, net_sales_amount, order_qty, session_total,
                             ad_impressions, ad_clicks, ad_orders, ad_spend, ad_sales_amount, refund_amount)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                                sales_qty=VALUES(sales_qty),
                                net_sales_amount=VALUES(net_sales_amount),
                                order_qty=VALUES(order_qty),
                                session_total=VALUES(session_total),
                                ad_impressions=VALUES(ad_impressions),
                                ad_clicks=VALUES(ad_clicks),
                                ad_orders=VALUES(ad_orders),
                                ad_spend=VALUES(ad_spend),
                                ad_sales_amount=VALUES(ad_sales_amount),
                                refund_amount=VALUES(refund_amount)
                            """,
                            (
                                sales_product_id, record_date, values['sales_qty'], values['net_sales_amount'],
                                values['order_qty'], values['session_total'], values['ad_impressions'], values['ad_clicks'],
                                values['ad_orders'], values['ad_spend'], values['ad_sales_amount'], values['refund_amount']
                            )
                        )
                    try:
                        self._refresh_sales_perf_agg_range(
                            conn, record_date, record_date, sales_product_ids=[sales_product_id],
                        )
                    except Exception:
                        pass
                    return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                delete_ids = []
                raw_ids = data.get('ids') if isinstance(data, dict) else None
                if isinstance(raw_ids, list):
                    for raw_id in raw_ids:
                        parsed_id = self._parse_int(raw_id)
                        if parsed_id:
                            delete_ids.append(parsed_id)
                item_id = self._parse_int(data.get('id') if isinstance(data, dict) else None)
                if item_id:
                    delete_ids.append(item_id)
                delete_ids = list(dict.fromkeys(delete_ids))
                if not delete_ids:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)

                affected_rows = []
                deleted_count = 0
                delete_chunk_size = 800
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        for i in range(0, len(delete_ids), delete_chunk_size):
                            chunk_ids = delete_ids[i:i + delete_chunk_size]
                            if not chunk_ids:
                                continue
                            placeholders = ','.join(['%s'] * len(chunk_ids))
                            cur.execute(
                                f"SELECT id, sales_product_id, record_date FROM sales_product_performances WHERE id IN ({placeholders})",
                                tuple(chunk_ids),
                            )
                            chunk_rows = cur.fetchall() or []
                            if chunk_rows:
                                affected_rows.extend(chunk_rows)
                            cur.execute(
                                f"DELETE FROM sales_product_performances WHERE id IN ({placeholders})",
                                tuple(chunk_ids),
                            )
                            try:
                                deleted_count += int(cur.rowcount or 0)
                            except Exception:
                                deleted_count += len(chunk_rows)
                    conn.commit()

                agg_rows_snapshot = [dict(r) for r in (affected_rows or []) if isinstance(r, dict)]
                if agg_rows_snapshot:
                    def _bg_refresh_deleted_agg(rows_copy):
                        try:
                            with self._get_db_connection_long(90, 90, 10) as agg_conn:
                                self._refresh_sales_perf_agg_for_deleted_records(agg_conn, rows_copy)
                        except Exception:
                            pass

                    threading.Thread(target=_bg_refresh_deleted_agg, args=(agg_rows_snapshot,), daemon=True).start()

                return self.send_json({
                    'status': 'success',
                    'deleted': deleted_count or len(affected_rows),
                    'requested': len(delete_ids),
                }, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_sales_product_performance_template_api(self, environ, method, start_response):
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT platform_sku FROM sales_products ORDER BY platform_sku")
                    sku_rows = cur.fetchall() or []
            sku_values = [str(row.get('platform_sku') or '').strip() for row in sku_rows if str(row.get('platform_sku') or '').strip()]

            wb = Workbook()
            ws = wb.active
            ws.title = 'sales_product_performance'

            headers = [
                'MSKU/ASIN/子ASIN*', '日期*', '销量*', '净销售额(USD)*', '订单量*', 'Sessions-Total*',
                '(广告)展示*', '(广告)点击*', '(广告)订单量*', '(广告)花费(USD)*', '(广告)销售额(USD)*',
                '退款金额(USD)*'
            ]
            ws.append(headers)
            ws.append([
                '示例SKU_请删除此行',
                '示例日期_请删除此行',
                12,
                999.99,
                10,
                480,
                2500,
                88,
                6,
                120.50,
                899.90,
                0.00
            ])

            for cell in ws[1]:
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in ws[2]:
                cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                cell.font = Font(italic=True, color='888888')

            widths = [24, 14, 10, 14, 10, 12, 12, 12, 12, 14, 14, 12]
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            options_ws = wb.create_sheet('options')
            options_ws.sheet_state = 'hidden'
            options_ws.cell(row=1, column=1, value='sales_platform_sku')
            for idx, sku in enumerate(sku_values, start=2):
                options_ws.cell(row=idx, column=1, value=sku)

            if sku_values:
                sku_validation = DataValidation(type='list', formula1=f'=options!$A$2:$A${len(sku_values) + 1}', allow_blank=False)
                ws.add_data_validation(sku_validation)
                for row_idx in range(3, 1000):
                    sku_validation.add(f'A{row_idx}')

            ws.freeze_panes = 'A3'
            return self._send_excel_workbook(wb, '产品表现导入模板.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_sales_product_performance_import_api(self, environ, method, start_response):
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))

            mode = str((query_params.get('mode', [''])[0] or '')).strip().lower()
            task_id = str((query_params.get('task_id', [''])[0] or '')).strip()
            async_import = str((query_params.get('async', [''])[0] or '')).strip().lower() in ('1', 'true', 'yes', 'on')
            temp_token = str((query_params.get('from_temp', [''])[0] or '')).strip()
            resume_from = str((query_params.get('resume_from', [''])[0] or '')).strip()
            force_restart = str((query_params.get('force', [''])[0] or '')).strip().lower() in ('1', 'true', 'yes', 'on')
            restart_reason = str((query_params.get('reason', [''])[0] or '')).strip().lower()

            import tempfile

            def _safe_task_id(raw):
                t = str(raw or '').strip()
                if not t:
                    return ''
                if not re.match(r'^[a-zA-Z0-9_-]{8,64}$', t):
                    return ''
                return t

            def _progress_file_path(tid):
                progress_dir = os.path.join(tempfile.gettempdir(), 'sitjoy_import_progress')
                try:
                    os.makedirs(progress_dir, exist_ok=True)
                except Exception:
                    pass
                return os.path.join(progress_dir, f'sales_product_performance_{tid}.json')

            _progress_lock_registry_guard = threading.Lock()
            _progress_locks_by_tid = {}

            def _progress_lock_for(tid):
                if not tid:
                    return None
                with _progress_lock_registry_guard:
                    lk = _progress_locks_by_tid.get(tid)
                    if lk is None:
                        lk = threading.Lock()
                        _progress_locks_by_tid[tid] = lk
                    return lk

            def _temp_upload_path(token):
                temp_dir = os.path.join(tempfile.gettempdir(), 'sitjoy_import_temp')
                try:
                    os.makedirs(temp_dir, exist_ok=True)
                except Exception:
                    pass
                # 须用 .xlsx 后缀，否则 openpyxl 按扩展名识别格式时会报不支持 .bin
                return os.path.join(temp_dir, f'spp_{token}.xlsx')

            def _write_progress(tid, payload, merge=False):
                if not tid:
                    return
                path = _progress_file_path(tid)
                tmp_path = path + '.tmp'
                lk = _progress_lock_for(tid) or threading.Lock()
                try:
                    with lk:
                        if isinstance(payload, dict):
                            # 即便是“覆盖写”（merge=False），也尽量保留诊断/心跳字段，
                            # 否则心跳线程写入的 hb_* 会被主线程的覆盖写频繁擦掉，导致无法定位卡点。
                            preserve_keys = (
                                'hb_ts', 'hb_phase', 'hb_stage', 'hb_processed_rows', 'hb_checkpoint_row',
                                'advance_ts',
                            )
                            if (not merge) and os.path.exists(path):
                                try:
                                    with open(path, 'r', encoding='utf-8') as f:
                                        base_keep = json.load(f)
                                    if isinstance(base_keep, dict):
                                        for k in preserve_keys:
                                            if k not in payload and k in base_keep:
                                                payload[k] = base_keep.get(k)
                                except Exception:
                                    pass

                        if merge and isinstance(payload, dict):
                            base = {}
                            if os.path.exists(path):
                                try:
                                    with open(path, 'r', encoding='utf-8') as f:
                                        base = json.load(f)
                                except Exception:
                                    base = {}
                            if not isinstance(base, dict):
                                base = {}
                            merged = dict(base)
                            merged.update(payload)
                            payload = merged
                        if isinstance(payload, dict):
                            payload.setdefault('ts', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                            try:
                                prev_seq = 0
                                if os.path.exists(path):
                                    with open(path, 'r', encoding='utf-8') as pf:
                                        prev_doc = json.load(pf)
                                    if isinstance(prev_doc, dict):
                                        prev_seq = int(prev_doc.get('seq') or 0)
                                payload['seq'] = prev_seq + 1
                            except Exception:
                                payload.setdefault('seq', 1)
                        with open(tmp_path, 'w', encoding='utf-8') as f:
                            json.dump(payload, f, ensure_ascii=False)
                        os.replace(tmp_path, path)
                    _spp_signal_import_waiters(tid)
                except Exception:
                    pass

            def _read_progress(tid):
                if not tid:
                    return None
                path = _progress_file_path(tid)
                if not os.path.exists(path):
                    return None
                lk = _progress_lock_for(tid) or threading.Lock()
                try:
                    with lk:
                        if not os.path.exists(path):
                            return None
                        with open(path, 'r', encoding='utf-8') as f:
                            return json.load(f)
                except Exception:
                    return None

            def _now_ts_text():
                return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # “真实前进”的标记：用于判断是否真的卡住（而不是仅仅无日志）。
            def _mark_progress_advance(tid, processed_rows=None, checkpoint_row=None, extra=None):
                payload = {'advance_ts': _now_ts_text()}
                if processed_rows is not None:
                    try:
                        payload['processed_rows'] = int(processed_rows or 0)
                    except Exception:
                        pass
                if checkpoint_row is not None:
                    try:
                        payload['checkpoint_row'] = int(checkpoint_row or 0)
                    except Exception:
                        pass
                if isinstance(extra, dict):
                    payload.update(extra)
                _write_progress(tid, payload, merge=True)

            # 心跳线程：即使主线程卡在 DB/IO，也能持续上报“卡在哪个阶段”。
            hb_state = {
                'run_id': '',
                'phase': '',
                'stage': '',
                'processed_rows': 0,
                'checkpoint_row': 0,
            }
            hb_stop_flag = {'stop': False}

            def _set_hb(stage=None, phase=None, processed_rows=None, checkpoint_row=None):
                if stage is not None:
                    hb_state['stage'] = str(stage or '')
                if phase is not None:
                    hb_state['phase'] = str(phase or '')
                if processed_rows is not None:
                    try:
                        hb_state['processed_rows'] = int(processed_rows or 0)
                    except Exception:
                        pass
                if checkpoint_row is not None:
                    try:
                        hb_state['checkpoint_row'] = int(checkpoint_row or 0)
                    except Exception:
                        pass

            def _heartbeat_worker(tid):
                while not hb_stop_flag.get('stop'):
                    try:
                        live = _read_progress(tid) or {}
                        st = str(live.get('state') or '').lower()
                        if st in ('success', 'error'):
                            break
                        _write_progress(tid, {
                            'hb_ts': _now_ts_text(),
                            'hb_phase': str(hb_state.get('phase') or ''),
                            'hb_stage': str(hb_state.get('stage') or ''),
                            'hb_processed_rows': int(hb_state.get('processed_rows') or 0),
                            'hb_checkpoint_row': int(hb_state.get('checkpoint_row') or 0),
                            'run_id': str(hb_state.get('run_id') or ''),
                        }, merge=True)
                    except Exception:
                        pass
                    time.sleep(3.0)

            safe_task_id = _safe_task_id(task_id)

            def _resolve_import_shop_id():
                sid = self._parse_int((query_params.get('shop_id', [''])[0] or '').strip())
                if sid:
                    return sid
                if safe_task_id:
                    prog = _read_progress(safe_task_id) or {}
                    sid = self._parse_int(prog.get('shop_id'))
                    if sid:
                        return sid
                return None

            import_shop_id = _resolve_import_shop_id()

            def _progress_payload_for_client():
                data = _read_progress(safe_task_id)
                if not data:
                    return {
                        'status': 'success',
                        'task_id': safe_task_id,
                        'state': 'pending',
                        'processed_rows': 0,
                        'total_rows': 0,
                        'created': 0,
                        'seq': 0,
                        'message': '等待任务开始'
                    }
                data.setdefault('status', 'success')
                data.setdefault('task_id', safe_task_id)
                if isinstance(data, dict):
                    alive_candidates = [
                        str(data.get('ts') or '').strip(),
                        str(data.get('hb_ts') or '').strip(),
                        str(data.get('advance_ts') or '').strip(),
                    ]
                    alive_candidates = [x for x in alive_candidates if x]
                    data['progress_alive_ts'] = max(alive_candidates) if alive_candidates else str(data.get('ts') or '')
                    try:
                        data.setdefault('seq', int(data.get('seq') or 0))
                    except Exception:
                        data['seq'] = 0
                return data

            def _spp_import_terminal_done(data):
                state = str((data or {}).get('state') or '').lower()
                agg_st = str((data or {}).get('agg_refresh_status') or '').lower()
                if state == 'error':
                    return True
                if state == 'success' and agg_st != 'running':
                    return True
                return False

            if method == 'GET' and mode == 'stream':
                if not safe_task_id:
                    return self.send_json({'status': 'error', 'message': 'task_id无效'}, start_response)
                try:
                    since_seq = int((query_params.get('since_seq', ['0'])[0] or '0'))
                except Exception:
                    since_seq = 0

                def _stream_generate():
                    yield b': connected\n\n'
                    waiter = _spp_register_import_waiter(safe_task_id)
                    since_local = since_seq
                    started = time.time()
                    last_ping = started
                    try:
                        while time.time() - started < SPP_IMPORT_STREAM_SESSION_SEC:
                            data = _progress_payload_for_client()
                            seq = int(data.get('seq') or 0)
                            terminal = _spp_import_terminal_done(data)
                            if seq > since_local or terminal:
                                evt = 'done' if terminal else 'progress'
                                if terminal and str(data.get('state') or '').lower() == 'error':
                                    evt = 'error'
                                yield self._sse_event(evt, data)
                                since_local = max(since_local, seq)
                                if terminal:
                                    return
                            now = time.time()
                            if now - last_ping >= SPP_IMPORT_STREAM_PING_SEC:
                                yield self._sse_event('ping', {'t': int(now)})
                                last_ping = now
                            waiter.clear()
                            remaining = max(0.05, SPP_IMPORT_STREAM_SESSION_SEC - (now - started))
                            waiter.wait(timeout=min(SPP_IMPORT_WAIT_POLL_SEC, remaining))
                    finally:
                        _spp_unregister_import_waiter(safe_task_id, waiter)

                return self.send_sse_stream(start_response, _stream_generate())

            if method == 'GET' and mode in ('progress', 'wait'):
                use_long_wait = (mode == 'wait') or str(
                    (query_params.get('wait', ['0'])[0] or '0')
                ).lower() in ('1', 'true', 'yes', 'on')
                try:
                    since_seq = int((query_params.get('since_seq', ['0'])[0] or '0'))
                except Exception:
                    since_seq = 0

                if not use_long_wait:
                    return self.send_json(_progress_payload_for_client(), start_response)

                waiter = _spp_register_import_waiter(safe_task_id)
                try:
                    deadline = time.time() + SPP_IMPORT_WAIT_TIMEOUT_SEC
                    while time.time() < deadline:
                        data = _progress_payload_for_client()
                        seq = int(data.get('seq') or 0)
                        state = str(data.get('state') or '').lower()
                        agg_st = str(data.get('agg_refresh_status') or '').lower()
                        terminal = state in ('success', 'error')
                        done = terminal and not (state == 'success' and agg_st == 'running')
                        if seq > since_seq or done:
                            data['unchanged'] = False
                            return self.send_json(data, start_response)
                        waiter.clear()
                        remaining = max(0.05, deadline - time.time())
                        waiter.wait(timeout=min(SPP_IMPORT_WAIT_POLL_SEC, remaining))
                finally:
                    _spp_unregister_import_waiter(safe_task_id, waiter)

                data = _progress_payload_for_client()
                data['unchanged'] = True
                return self.send_json(data, start_response)

            # 触发“重启导入任务”（仅用于写库阶段卡死/无响应时的自愈；聚合阶段禁止重启）
            if method == 'GET' and mode == 'restart':
                if not safe_task_id:
                    return self.send_json({'status': 'error', 'message': 'task_id无效'}, start_response)
                data = _read_progress(safe_task_id) or {}
                phase = str(data.get('phase') or '').strip().lower()
                agg_st = str(data.get('agg_refresh_status') or '').strip().lower()
                state = str(data.get('state') or '').strip().lower()

                if agg_st == 'running' or phase in ('agg_refresh', 'agg_refresh_pending', 'agg_refresh_done'):
                    return self.send_json({
                        'status': 'success',
                        'task_id': safe_task_id,
                        'skip_restart': True,
                        'reason': 'agg_refresh',
                        'message': '明细已入库，周/月聚合后台刷新中，请勿重启导入（聚合可能较慢）',
                    }, start_response)
                if state == 'success':
                    return self.send_json({
                        'status': 'success',
                        'task_id': safe_task_id,
                        'skip_restart': True,
                        'reason': 'already_success',
                        'message': '导入已完成，无需重启',
                    }, start_response)
                if state == 'error':
                    return self.send_json({
                        'status': 'success',
                        'task_id': safe_task_id,
                        'skip_restart': True,
                        'reason': 'already_error',
                        'message': '任务已失败结束，请重新上传文件',
                    }, start_response)
                if phase not in ('writing', 'writing_retry', 'restart', 'iter_rows', 'loading_sku_map', 'start', 'running', ''):
                    return self.send_json({
                        'status': 'success',
                        'task_id': safe_task_id,
                        'skip_restart': True,
                        'reason': 'wrong_phase',
                        'message': f'当前阶段（{phase or "未知"}）不支持重启导入',
                    }, start_response)

                # 允许前端 force=1 触发快速恢复（例如连续请求无响应）
                if not force_restart:
                    stale = False
                    try:
                        adv = str(data.get('advance_ts') or '').strip()
                        if adv:
                            last_dt = datetime.strptime(adv, '%Y-%m-%d %H:%M:%S')
                            # 更敏感的自愈：若“真实前进时间”超过 60 秒不更新，认为已卡住
                            if (datetime.now() - last_dt).total_seconds() >= 60:
                                stale = True
                        else:
                            ts = str(data.get('ts') or '').strip()
                            if ts:
                                last_dt = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')
                                if (datetime.now() - last_dt).total_seconds() >= 60:
                                    stale = True
                    except Exception:
                        stale = True
                    if not stale:
                        return self.send_json({'status': 'success', 'message': '任务仍在运行（未超过120秒无更新），无需重启', 'task_id': safe_task_id}, start_response)

                temp_path = _temp_upload_path(safe_task_id)
                if not os.path.exists(temp_path):
                    return self.send_json({'status': 'error', 'message': '临时文件不存在，无法重启（可能已清理或任务启动方式非异步）', 'task_id': safe_task_id}, start_response)

                # 读取 checkpoint（已安全落库的行号），用于恢复
                cp = 0
                try:
                    cp = int(data.get('checkpoint_row') or 0)
                except Exception:
                    cp = 0
                # pending_batch_start_row：当前内存批次最早未落库行号。
                # 重启时应回退到该位置，避免丢失未提交的数据，同时尽量靠后以减少重复耗时。
                pending_start = 0
                try:
                    pending_start = int(data.get('pending_batch_start_row') or 0)
                except Exception:
                    pending_start = 0
                if pending_start > 1:
                    cp = max(cp, pending_start - 1)
                if resume_from:
                    try:
                        cp = max(cp, int(resume_from))
                    except Exception:
                        pass

                # bump run_id：旧 worker 会检测到并退出
                new_run_id = hashlib.md5(f"{datetime.now().isoformat()}_{os.getpid()}_{time.time()}".encode('utf-8')).hexdigest()[:10]
                _write_progress(safe_task_id, {
                    'status': 'success',
                    'task_id': safe_task_id,
                    'state': 'running',
                    'phase': 'restart',
                    'run_id': new_run_id,
                    'checkpoint_row': cp,
                    'retry_reason': restart_reason or 'stale_progress',
                    'message': f'检测到{restart_reason or "stale_progress"}，正在从第 {cp} 行继续恢复导入...'
                }, merge=True)

                def _restart_worker():
                    try:
                        restart_shop_id = import_shop_id or self._parse_int(((_read_progress(safe_task_id) or {}).get('shop_id')))
                        q = f"task_id={safe_task_id}&check_only=0&async=0&from_temp={safe_task_id}&resume_from={cp}&run_id={new_run_id}&shop_id={restart_shop_id or ''}"
                        bg_env = {
                            'QUERY_STRING': q,
                            'CONTENT_TYPE': '',
                            'CONTENT_LENGTH': '0',
                            'wsgi.input': io.BytesIO(b''),
                        }
                        self.handle_sales_product_performance_import_api(bg_env, 'POST', lambda *args, **kwargs: None)
                    except Exception as _e:
                        _write_progress(safe_task_id, {
                            'status': 'error',
                            'task_id': safe_task_id,
                            'state': 'error',
                            'phase': 'restart_error',
                            'message': f'重启失败：{str(_e)[:200]}'
                        }, merge=True)

                threading.Thread(target=_restart_worker, daemon=True).start()
                return self.send_json({'status': 'success', 'task_id': safe_task_id, 'message': '已触发重启，请继续观察进度'}, start_response)

            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)
            check_only = str((query_params.get('check_only', ['0'])[0] or '0')).lower() in ('1', 'true', 'yes', 'on')

            if not safe_task_id:
                safe_task_id = hashlib.md5(f"{datetime.now().isoformat()}_{os.getpid()}".encode('utf-8')).hexdigest()[:16]

            # 每次 worker 运行都有自己的 run_id（用于被 restart 终止）
            worker_run_id = str((query_params.get('run_id', [''])[0] or '')).strip()
            if not worker_run_id:
                worker_run_id = hashlib.md5(f"{safe_task_id}_{datetime.now().isoformat()}_{os.getpid()}".encode('utf-8')).hexdigest()[:10]

            file_bytes = b''
            temp_path = None
            file_byte_len = 0
            if temp_token:
                temp_path = _temp_upload_path(temp_token)
                if not os.path.exists(temp_path):
                    return self.send_json({'status': 'error', 'message': '临时文件不存在，任务可能已过期'}, start_response)
                try:
                    file_byte_len = int(os.path.getsize(temp_path) or 0)
                except Exception:
                    file_byte_len = 0
                if file_byte_len <= 0:
                    return self.send_json({'status': 'error', 'message': '临时文件为空'}, start_response)
                _write_progress(safe_task_id, {
                    'status': 'success',
                    'task_id': safe_task_id,
                    'state': 'running',
                    'phase': 'loading_workbook',
                    'shop_id': import_shop_id,
                    'message': f'文件已在服务器就绪（{file_byte_len // 1024}KB），正在流式打开 Excel…',
                }, merge=True)
                _set_hb(stage='loading_workbook', phase='loading_workbook', processed_rows=0, checkpoint_row=0)
            else:
                content_type = environ.get('CONTENT_TYPE', '')
                if 'multipart/form-data' not in content_type:
                    return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

                content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
                raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
                env_copy = dict(environ)
                env_copy['CONTENT_LENGTH'] = str(len(raw_body))
                form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
                if not import_shop_id:
                    import_shop_id = self._parse_int((form.getfirst('shop_id', '') or '').strip())
                file_item = form['file'] if 'file' in form else None
                if file_item is None or getattr(file_item, 'file', None) is None:
                    return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
                file_bytes = file_item.file.read() or b''
                if not file_bytes:
                    return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)
                file_byte_len = len(file_bytes)

            if not import_shop_id:
                return self.send_json({'status': 'error', 'message': '请选择店铺后再上传'}, start_response)

            # 正式导入默认异步，避免网关504；预检保持同步
            if (not check_only) and (not temp_token):
                if not async_import:
                    async_import = True
                if async_import:
                    temp_token = safe_task_id
                    temp_path = _temp_upload_path(temp_token)
                    with open(temp_path, 'wb') as f:
                        f.write(file_bytes)

                    _write_progress(safe_task_id, {
                        'status': 'success',
                        'task_id': safe_task_id,
                        'state': 'pending',
                        'processed_rows': 0,
                        'total_rows': 0,
                        'created': 0,
                        'shop_id': import_shop_id,
                        'message': '任务已创建，准备开始处理'
                    })

                    def _bg_worker():
                        try:
                            _write_progress(safe_task_id, {
                                'status': 'success',
                                'task_id': safe_task_id,
                                'state': 'running',
                                'phase': 'worker_start',
                                'processed_rows': 0,
                                'total_rows': 0,
                                'shop_id': import_shop_id,
                                'message': '后台任务已启动，正在读取文件…',
                            }, merge=True)
                            q = f"task_id={safe_task_id}&check_only=0&async=0&from_temp={temp_token}&shop_id={import_shop_id}"
                            bg_env = {
                                'QUERY_STRING': q,
                                'CONTENT_TYPE': '',
                                'CONTENT_LENGTH': '0',
                                'wsgi.input': io.BytesIO(b''),
                            }
                            self.handle_sales_product_performance_import_api(bg_env, 'POST', lambda *args, **kwargs: None)
                        except Exception as _e:
                            _write_progress(safe_task_id, {
                                'status': 'error',
                                'task_id': safe_task_id,
                                'state': 'error',
                                'processed_rows': 0,
                                'total_rows': 0,
                                'created': 0,
                                'message': str(_e)[:200]
                            })
                        finally:
                            try:
                                if os.path.exists(temp_path):
                                    os.remove(temp_path)
                            except Exception:
                                pass

                    t = threading.Thread(target=_bg_worker, daemon=True)
                    t.start()
                    return self.send_json({
                        'status': 'success',
                        'async': True,
                        'task_id': safe_task_id,
                        'message': '导入任务已启动，请通过进度长轮询接口获取结果'
                    }, start_response)

            # Excel 在「运行本系统的服务器」上解析（群晖即 NAS 本机 CPU/磁盘），不在用户 PC 浏览器里算。
            # NAS 上整表 load_workbook(read_only=False) 极慢且几乎不占磁盘 IO；大文件一律流式只读，并从临时文件直接打开避免再读入内存。
            if not file_byte_len:
                file_byte_len = len(file_bytes or b'')
            if not check_only:
                hb_state['run_id'] = worker_run_id
                _set_hb(stage='pre_load', phase='loading_workbook', processed_rows=0, checkpoint_row=0)
                threading.Thread(target=_heartbeat_worker, args=(safe_task_id,), daemon=True).start()
            use_read_only = check_only or bool(temp_path) or file_byte_len >= (512 * 1024)
            xlsx_source = None
            if temp_path and use_read_only:
                xlsx_source = temp_path
            else:
                if not file_bytes and temp_path:
                    with open(temp_path, 'rb') as rf:
                        file_bytes = rf.read() or b''
                    file_byte_len = len(file_bytes)
                xlsx_source = io.BytesIO(file_bytes)

            _write_progress(safe_task_id, {
                'status': 'success',
                'task_id': safe_task_id,
                'state': 'running',
                'phase': 'loading_workbook',
                'message': (
                    f'正在{"流式" if use_read_only else "标准"}解析 Excel（约 {file_byte_len // 1024}KB）…'
                ),
            }, merge=True)
            _set_hb(stage='openpyxl_load', phase='loading_workbook', processed_rows=0, checkpoint_row=0)

            wb = load_workbook(
                xlsx_source,
                read_only=use_read_only,
                data_only=True,
            )
            ws = wb.active
            if check_only:
                total_rows_hint = 100
            elif use_read_only:
                # read_only 下访问 max_row 会扫完整张表，极慢；用体积粗估行数供进度条展示
                total_rows_hint = min(500000, max(800, file_byte_len // 140))
            else:
                total_rows_hint = max(0, int((ws.max_row or 1)) - 1)

            # 读取第一行作为headers（read_only模式下避免ws[1]）
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            headers = [str(x or '').strip() for x in (header_row or [])]
            if not check_only:
                _write_progress(safe_task_id, {
                    'status': 'success',
                    'task_id': safe_task_id,
                    'state': 'running',
                    'phase': 'headers_parsed',
                    'total_rows': total_rows_hint,
                    'message': f'表头已识别，预计约 {total_rows_hint} 行，正在加载 SKU 映射…',
                }, merge=True)

            def normalize_header(text):
                t = str(text or '').strip().lower()
                t = t.replace('\ufeff', '')
                t = t.replace('*', '')
                t = t.replace('（', '(').replace('）', ')')
                t = t.replace('：', ':')
                t = re.sub(r'[\s\-_/\\|()\[\]{}:]+', '', t)
                return t

            alias_groups = {
                'identifier': ['销售平台sku', '销售平台sku/msku/asin/子asin', 'msku/asin/子asin', 'platformsku', 'sku', 'msku', 'asin', '子asin', '子体asin', 'childasin', '子体编码', '子体编号'],
                'record_date': ['日期', 'date', 'recorddate'],
                'sales_qty': ['销量', 'salesqty', 'salesquantity'],
                'net_sales_amount': ['净销售额(usd)', '净销售额', '销售额', 'netsales', 'netsalesamount', 'salesamount'],
                'order_qty': ['订单量', 'orderqty', 'orderquantity'],
                'session_total': ['session-total', 'sessions-total', 'sessiontotal', 'sessionstotal'],
                'ad_impressions': ['(广告)展示', '广告展示', '展示', 'adimpressions', 'impressions'],
                'ad_clicks': ['(广告)点击', '广告点击', '点击', 'adclicks', 'clicks'],
                'ad_orders': ['(广告)订单量', '广告订单量', 'adorders', 'ordersfromad'],
                'ad_spend': ['(广告)花费(usd)', '(广告)花费', '广告花费(usd)', '广告花费', 'adspend', 'spend'],
                'ad_sales_amount': ['(广告)销售额(usd)', '(广告)销售额', '广告销售额(usd)', '广告销售额', 'adsales', 'adsalesamount'],
                'refund_amount': ['退款金额(usd)', '退款金额', 'refundamount'],
            }

            normalized_headers = [normalize_header(h) for h in headers]
            resolved_col = {}
            for key, aliases in alias_groups.items():
                found = None
                alias_set = set([normalize_header(x) for x in aliases])
                for idx, norm_name in enumerate(normalized_headers):
                    if norm_name in alias_set:
                        found = idx
                        break
                resolved_col[key] = found

            if resolved_col.get('identifier') is None:
                return self.send_json({'status': 'error', 'message': '模板缺少标识列（销售平台SKU/MSKU/ASIN/子ASIN）'}, start_response)
            if resolved_col.get('record_date') is None:
                return self.send_json({'status': 'error', 'message': '模板缺少日期列'}, start_response)

            def get_cell(row, field_key):
                idx = resolved_col.get(field_key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            def normalize_date(value):
                if value is None:
                    return ''
                if isinstance(value, datetime):
                    return value.strftime('%Y-%m-%d')
                text = str(value).strip()
                if not text:
                    return ''
                for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
                    try:
                        return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
                    except Exception:
                        continue
                return text[:10]

            _num_re = re.compile(r'-?\d+(?:\.\d+)?')
            _id_split_re = re.compile(r'[,，;；]+')
            _template_markers = ('示例', '请删除')

            def parse_number_flexible(value, as_int=False):
                if value is None:
                    return 0 if as_int else 0.0
                if isinstance(value, (int, float)):
                    if as_int:
                        try:
                            return int(round(float(value)))
                        except Exception:
                            return 0
                    return float(value)
                text = str(value).strip()
                if not text:
                    return 0 if as_int else 0.0
                if not as_int and text.replace('.', '', 1).replace('-', '', 1).isdigit():
                    try:
                        return float(text)
                    except Exception:
                        pass
                text = text.replace('，', ',').replace('$', '').replace('￥', '')
                text = text.replace(',', '')
                m = _num_re.search(text)
                if not m:
                    return 0 if as_int else 0.0
                num = float(m.group(0))
                if as_int:
                    return int(round(num))
                return num

            def row_signature(payload):
                return '|'.join([
                    str(payload.get('sales_qty') or 0),
                    str(payload.get('net_sales_amount') or 0),
                    str(payload.get('order_qty') or 0),
                    str(payload.get('session_total') or 0),
                    str(payload.get('ad_impressions') or 0),
                    str(payload.get('ad_clicks') or 0),
                    str(payload.get('ad_orders') or 0),
                    str(payload.get('ad_spend') or 0),
                    str(payload.get('ad_sales_amount') or 0),
                    str(payload.get('refund_amount') or 0),
                ])

            created = 0
            updated = 0
            unchanged = 0
            errors = []
            skipped_empty_identifier = 0
            skipped_unmatched_sku = 0
            skipped_invalid_date = 0
            skipped_template_sample = 0
            skipped_all_zero = 0
            upserted = 0

            if check_only:
                _write_progress(safe_task_id, {
                    'status': 'success',
                    'task_id': safe_task_id,
                    'state': 'running',
                    'processed_rows': 0,
                    'total_rows': total_rows_hint,
                    'created': 0,
                    'shop_id': import_shop_id,
                    'phase': 'precheck',
                    'message': '预检中（抽样前100行）...'
                })
            else:
                _write_progress(safe_task_id, {
                    'status': 'success',
                    'task_id': safe_task_id,
                    'state': 'running',
                    'processed_rows': 0,
                    'total_rows': total_rows_hint,
                    'created': 0,
                    'shop_id': import_shop_id,
                    'phase': 'start',
                    'run_id': worker_run_id,
                    'checkpoint_row': 0,
                    'advance_ts': _now_ts_text(),
                    'message': '开始处理...'
                })
                _set_hb(stage='init', phase='start', processed_rows=0, checkpoint_row=0)

            min_record_date = None
            max_record_date = None
            touched_agg_ids = set()
            agg_refresh_warning = None

            db_conn_factory = self._get_db_connection if check_only else (
                lambda: self._get_db_connection_long(90, 90, 10)
            )
            with db_conn_factory() as conn:
                with conn.cursor() as cur:
                    if not check_only:
                        _write_progress(safe_task_id, {
                            'status': 'success',
                            'task_id': safe_task_id,
                            'state': 'running',
                            'phase': 'loading_sku_map',
                            'processed_rows': 0,
                            'total_rows': total_rows_hint,
                            'created': 0,
                            'message': '正在加载SKU映射（sales_products）...'
                        })
                        _set_hb(stage='loading_sku_map', phase='loading_sku_map', processed_rows=0, checkpoint_row=0)
                    # 一次性加载指定店铺下的 SKU/ASIN 映射，避免跨店铺误匹配
                    sku_map = {}
                    asin_map = {}
                    cur.execute(
                        "SELECT id, platform_sku, child_code FROM sales_products WHERE shop_id=%s",
                        (import_shop_id,),
                    )
                    for row in (cur.fetchall() or []):
                        rid = int(row.get('id') or 0)
                        sku = str(row.get('platform_sku') or '').strip().lower()
                        child_code = str(row.get('child_code') or '').strip().lower()
                        if rid and sku:
                            sku_map[sku] = rid
                        if rid and child_code:
                            asin_map[child_code] = rid
                    if not check_only:
                        _write_progress(safe_task_id, {
                            'status': 'success',
                            'task_id': safe_task_id,
                            'state': 'running',
                            'phase': 'loading_sku_map_done',
                            'processed_rows': 0,
                            'total_rows': total_rows_hint,
                            'created': 0,
                            'message': f'SKU映射加载完成（店铺ID={import_shop_id}）：sku={len(sku_map)}，asin={len(asin_map)}'
                        })
                        _mark_progress_advance(safe_task_id, processed_rows=0, checkpoint_row=0, extra={'phase': 'loading_sku_map_done'})

                    def _resolve_sales_product_id_for_row(raw_identifier):
                        """先整串匹配 platform_sku / child_code；失败则按中英文逗号等切分后逐段匹配。"""
                        text = str(raw_identifier or '').strip()
                        if not text:
                            return None
                        low = text.lower()
                        pid = sku_map.get(low) or asin_map.get(low)
                        if pid:
                            return pid
                        for seg in _id_split_re.split(text):
                            t = seg.strip().lower()
                            if not t:
                                continue
                            pid = sku_map.get(t) or asin_map.get(t)
                            if pid:
                                return pid
                        return None

                    # 初始化批处理变量
                    batch_rows = []
                    batch_row_marks = []
                    # 批量写入：大批次 + 少次 commit；进度文件节流（NAS 上频繁写 JSON 极慢）
                    batch_size = 5000
                    flush_chunk_size = 5000
                    progress_row_step = 800
                    progress_sec_step = 3.0
                    run_id_check_step = 400
                    upsert_sql = (
                        "INSERT INTO sales_product_performances "
                        "(sales_product_id,record_date,sales_qty,net_sales_amount,order_qty,session_total,"
                        "ad_impressions,ad_clicks,ad_orders,ad_spend,ad_sales_amount,refund_amount) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                        "ON DUPLICATE KEY UPDATE "
                        "sales_qty=VALUES(sales_qty),"
                        "net_sales_amount=VALUES(net_sales_amount),"
                        "order_qty=VALUES(order_qty),"
                        "session_total=VALUES(session_total),"
                        "ad_impressions=VALUES(ad_impressions),"
                        "ad_clicks=VALUES(ad_clicks),"
                        "ad_orders=VALUES(ad_orders),"
                        "ad_spend=VALUES(ad_spend),"
                        "ad_sales_amount=VALUES(ad_sales_amount),"
                        "refund_amount=VALUES(refund_amount)"
                    )

                    def _is_retryable_write_error(exc):
                        err_text = str(exc or '').lower()
                        retry_tokens = (
                            'lost connection',
                            'server has gone away',
                            'connection reset',
                            'broken pipe',
                            'read timed out',
                            'write timed out',
                            'timed out',
                            'deadlock found',
                            'lock wait timeout',
                            'error 1205',
                            'error 1213',
                            'error 2006',
                            'error 2013',
                        )
                        return any(token in err_text for token in retry_tokens)

                    def flush_batch_data():
                        nonlocal cur, checkpoint_row
                        if not batch_rows:
                            return 0
                        try:
                            _set_hb(stage='db_executemany', phase='writing', processed_rows=row_count, checkpoint_row=checkpoint_row)
                            total = len(batch_rows)
                            written = 0
                            chunk_checkpoint = int(batch_row_marks[-1] or 0) if batch_row_marks else checkpoint_row
                            for tup in batch_rows:
                                try:
                                    touched_agg_ids.add(int(tup[0]))
                                except Exception:
                                    pass
                            for i in range(0, total, flush_chunk_size):
                                chunk = batch_rows[i:i + flush_chunk_size]
                                cur.executemany(upsert_sql, chunk)
                                written += len(chunk)
                            conn.commit()
                            if chunk_checkpoint > checkpoint_row:
                                checkpoint_row = chunk_checkpoint
                            _write_progress(safe_task_id, {
                                'status': 'success',
                                'task_id': safe_task_id,
                                'state': 'running',
                                'phase': 'writing',
                                'processed_rows': row_count,
                                'total_rows': total_rows_hint,
                                'created': created,
                                'upserted': upserted + written,
                                'checkpoint_row': checkpoint_row,
                                'message': f'已写入数据库 {written} 行（累计处理 {row_count}/{total_rows_hint or "?"} 行）'
                            }, merge=True)
                            _mark_progress_advance(safe_task_id, processed_rows=row_count, checkpoint_row=checkpoint_row)
                            _set_hb(stage='db_commit', phase='writing', processed_rows=row_count, checkpoint_row=checkpoint_row)
                            return written
                        except Exception as chunk_err:
                            if not _is_retryable_write_error(chunk_err):
                                raise
                            try:
                                conn.rollback()
                            except Exception:
                                pass
                            try:
                                conn.ping(reconnect=True)
                                cur = conn.cursor()
                            except Exception:
                                raise RuntimeError(f"批量写入失败: {str(chunk_err)[:180]}")
                            _write_progress(safe_task_id, {
                                'status': 'success',
                                'task_id': safe_task_id,
                                'state': 'running',
                                'phase': 'writing_retry',
                                'processed_rows': row_count,
                                'total_rows': total_rows_hint,
                                'created': created,
                                'upserted': upserted,
                                'checkpoint_row': checkpoint_row,
                                'retry_reason': 'retryable_db_error',
                                'message': f'写入失败，正在重连后整批重试（断点 {chunk_checkpoint}/{total_rows_hint}）...'
                            }, merge=True)
                            _set_hb(stage='db_executemany_retry', phase='writing_retry', processed_rows=row_count, checkpoint_row=checkpoint_row)
                            try:
                                written = 0
                                for i in range(0, total, flush_chunk_size):
                                    chunk = batch_rows[i:i + flush_chunk_size]
                                    cur.executemany(upsert_sql, chunk)
                                    written += len(chunk)
                                conn.commit()
                                if chunk_checkpoint > checkpoint_row:
                                    checkpoint_row = chunk_checkpoint
                                _mark_progress_advance(safe_task_id, processed_rows=row_count, checkpoint_row=checkpoint_row)
                                return written
                            except Exception as retry_err:
                                try:
                                    conn.rollback()
                                except Exception:
                                    pass
                                raise RuntimeError(f"批量写入失败: {str(retry_err)[:180]}")
                        except Exception as e:
                            try:
                                conn.rollback()
                            except Exception:
                                pass
                            raise RuntimeError(f"批量写入失败: {str(e)[:180]}")
                        finally:
                            batch_rows.clear()
                            batch_row_marks.clear()
                            _write_progress(safe_task_id, {
                                'pending_batch_start_row': 0,
                                'pending_batch_size': 0,
                            }, merge=True)

                    # 预检模式：只检查前100行用于快速验证；正式模式：处理全部行
                    process_limit = 100 if check_only else 999999
                    processed_count = 0
                    row_count = 0
                    last_progress_heartbeat = time.time()
                    fatal_write_error = None
                    checkpoint_row = 0

                    # 恢复：跳过已处理的数据行（row_count 从 1 开始对应 Excel 数据第2行）
                    resume_row = 0
                    try:
                        resume_row = int(resume_from or 0)
                    except Exception:
                        resume_row = 0
                    if resume_row > 0:
                        _write_progress(safe_task_id, {
                            'status': 'success',
                            'task_id': safe_task_id,
                            'state': 'running',
                            'phase': 'resume',
                            'run_id': worker_run_id,
                            'checkpoint_row': resume_row,
                            'processed_rows': resume_row,
                            'total_rows': total_rows_hint,
                            'created': created,
                            'message': f'正在恢复：跳过前 {resume_row} 行已处理数据...'
                        }, merge=True)

                    if not check_only:
                        _write_progress(safe_task_id, {
                            'status': 'success',
                            'task_id': safe_task_id,
                            'state': 'running',
                            'phase': 'iter_rows',
                            'processed_rows': 0,
                            'total_rows': total_rows_hint,
                            'created': 0,
                            'message': '开始逐行读取并写入数据库...'
                        })
                        _set_hb(stage='iter_rows', phase='iter_rows', processed_rows=0, checkpoint_row=0)

                    # 使用iter_rows避免遍历max_row导致的超时问题
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        # 达到预检限制时提前退出；并限制扫描行数，避免空行过多时读完整表
                        if check_only and processed_count >= process_limit:
                            break
                        if check_only and row_count >= 2500:
                            break

                        row_count += 1
                        # worker 被 restart 后应尽快退出（避免并发写入）
                        if run_id_check_step > 0 and (row_count % run_id_check_step == 0):
                            live = _read_progress(safe_task_id) or {}
                            if str(live.get('run_id') or '') and str(live.get('run_id') or '') != str(worker_run_id):
                                return self.send_json({'status': 'success', 'task_id': safe_task_id, 'message': '任务已被重启，当前worker退出'}, start_response)

                        if resume_row and row_count <= resume_row:
                            continue

                        if not any(cell is not None and str(cell).strip() for cell in row):
                            continue

                        processed_count += 1

                        try:
                            identifier = str(get_cell(row, 'identifier') or '').strip()
                            if not identifier:
                                # 缺少标识，直接跳过（不计入errors）
                                skipped_empty_identifier += 1
                                continue

                            low_identifier = identifier.lower()
                            if any(x in low_identifier for x in _template_markers):
                                skipped_template_sample += 1
                                continue
                            sales_product_id = _resolve_sales_product_id_for_row(identifier)

                            # SKU无法匹配时，直接跳过该行（不计入errors，不中断流程）
                            if not sales_product_id:
                                skipped_unmatched_sku += 1
                                continue

                            record_date = normalize_date(get_cell(row, 'record_date'))
                            if not record_date:
                                # 日期格式错误，直接跳过
                                skipped_invalid_date += 1
                                continue

                            # 对于“指标全为0/空”的行，直接跳过。
                            # 性能关键：解析数值只做一次（避免 all_zero 判断与 batch_rows.append 重复解析导致 iter_rows 阶段极慢）。
                            sales_qty = parse_number_flexible(get_cell(row, 'sales_qty'), True)
                            net_sales_amount = parse_number_flexible(get_cell(row, 'net_sales_amount'), False)
                            order_qty = parse_number_flexible(get_cell(row, 'order_qty'), True)
                            session_total = parse_number_flexible(get_cell(row, 'session_total'), True)
                            ad_impressions = parse_number_flexible(get_cell(row, 'ad_impressions'), True)
                            ad_clicks = parse_number_flexible(get_cell(row, 'ad_clicks'), True)
                            ad_orders = parse_number_flexible(get_cell(row, 'ad_orders'), True)
                            ad_spend = parse_number_flexible(get_cell(row, 'ad_spend'), False)
                            ad_sales_amount = parse_number_flexible(get_cell(row, 'ad_sales_amount'), False)
                            refund_amount = parse_number_flexible(get_cell(row, 'refund_amount'), False)

                            if (
                                float(sales_qty or 0) == 0.0
                                and float(net_sales_amount or 0) == 0.0
                                and float(order_qty or 0) == 0.0
                                and float(session_total or 0) == 0.0
                                and float(ad_impressions or 0) == 0.0
                                and float(ad_clicks or 0) == 0.0
                                and float(ad_orders or 0) == 0.0
                                and float(ad_spend or 0) == 0.0
                                and float(ad_sales_amount or 0) == 0.0
                                and float(refund_amount or 0) == 0.0
                            ):
                                skipped_all_zero += 1
                                continue

                            try:
                                d_obj = datetime.strptime(record_date, '%Y-%m-%d').date()
                                if (min_record_date is None) or (d_obj < min_record_date):
                                    min_record_date = d_obj
                                if (max_record_date is None) or (d_obj > max_record_date):
                                    max_record_date = d_obj
                            except Exception:
                                pass

                            # 预检模式：只计算统计，不入库
                            if check_only:
                                created += 1
                            else:
                                batch_rows.append((
                                    sales_product_id,
                                    record_date,
                                    sales_qty,
                                    net_sales_amount,
                                    order_qty,
                                    session_total,
                                    ad_impressions,
                                    ad_clicks,
                                    ad_orders,
                                    ad_spend,
                                    ad_sales_amount,
                                    refund_amount,
                                ))
                                batch_row_marks.append(row_count)
                                created += 1

                                # 达到batch_size则执行
                                if len(batch_rows) >= batch_size:
                                    try:
                                        _set_hb(stage='flush_batch', phase='writing', processed_rows=row_count, checkpoint_row=checkpoint_row)
                                        _write_progress(safe_task_id, {
                                            'pending_batch_start_row': int(batch_row_marks[0] or row_count),
                                            'pending_batch_size': len(batch_rows),
                                        }, merge=True)
                                        upserted += flush_batch_data()
                                    except Exception as werr:
                                        fatal_write_error = str(werr)[:200]
                                        raise
                        except Exception as batch_err:
                            # 写库失败属于致命错误：继续跑会导致“写了一半却返回成功”
                            if (not check_only) and ('批量写入失败' in str(batch_err) or 'Lost connection' in str(batch_err) or 'timed out' in str(batch_err)):
                                fatal_write_error = fatal_write_error or str(batch_err)[:200]
                                _write_progress(safe_task_id, {
                                    'status': 'error',
                                    'task_id': safe_task_id,
                                    'state': 'error',
                                    'phase': 'writing_error',
                                    'processed_rows': row_count,
                                    'total_rows': total_rows_hint,
                                    'created': created,
                                    'upserted': upserted,
                                    'checkpoint_row': checkpoint_row,
                                    'message': f'写入数据库失败：{fatal_write_error}'
                                }, merge=True)
                                break
                            errors.append(f"第{row_count+1}行处理失败: {str(batch_err)[:100]}")
                        if fatal_write_error:
                            break

                        now_hb = time.time()
                        if (not check_only) and ((row_count % progress_row_step == 0) or ((now_hb - last_progress_heartbeat) >= progress_sec_step)):
                            _write_progress(safe_task_id, {
                                'status': 'success',
                                'task_id': safe_task_id,
                                'state': 'running',
                                'phase': 'iter_rows',
                                'processed_rows': row_count,
                                'total_rows': total_rows_hint,
                                'created': created,
                                'upserted': upserted,
                                'checkpoint_row': checkpoint_row,
                                'message': f'正在处理第 {row_count} 行'
                            }, merge=True)
                            _mark_progress_advance(safe_task_id, processed_rows=row_count, checkpoint_row=checkpoint_row)
                            _set_hb(stage='iter_rows', phase='iter_rows', processed_rows=row_count, checkpoint_row=checkpoint_row)
                            last_progress_heartbeat = now_hb

                    try:
                        wb.close()
                    except Exception:
                        pass

                    # 导入完成后，flush最后的batch数据
                    if not check_only:
                        if not fatal_write_error:
                            try:
                                _set_hb(stage='flush_batch_final', phase='writing', processed_rows=row_count, checkpoint_row=checkpoint_row)
                                upserted += flush_batch_data()
                            except Exception as werr2:
                                fatal_write_error = str(werr2)[:200]

                        if fatal_write_error:
                            msg = f'写入数据库失败：{fatal_write_error}'
                            _write_progress(safe_task_id, {
                                'status': 'error',
                                'task_id': safe_task_id,
                                'state': 'error',
                                'phase': 'writing_error',
                                'processed_rows': row_count,
                                'total_rows': total_rows_hint,
                                'created': created,
                                'upserted': upserted,
                                'checkpoint_row': checkpoint_row,
                                'message': msg,
                                'errors': errors[:100],
                            }, merge=True)
                            hb_stop_flag['stop'] = True
                            return self.send_json({
                                'status': 'error',
                                'task_id': safe_task_id,
                                'message': msg,
                                'errors': errors[:100],
                            }, start_response)

            # 周/月聚合耗时可远超网关/反代超时，长时间 phase=agg_refresh 易导致轮询 504。
            # 主流程在明细提交后立即 state=success；聚合放到后台线程，进度文件用 merge 增量更新。
            # 策略：先按本次导入 SKU 刷新月+周（快、保证 5 月等月聚合及时）；再仅对日期范围做「全 SKU 月聚合」补齐看板缺口（跳过全表周聚合，避免超时卡在月前）。
            needs_agg_bg = bool((not check_only) and min_record_date and max_record_date)
            agg_bg_fn = None  # 后台聚合线程入口（非 None 时在写入最终进度后启动）
            if needs_agg_bg:
                _agg_d0 = min_record_date.strftime('%Y-%m-%d')
                _agg_d1 = max_record_date.strftime('%Y-%m-%d')
                # 后台 merge 时必须反复带上最终统计快照，覆盖进度文件里可能过期的字段（部分更新/并发读）
                _agg_stat_fields = {
                    'processed_rows': row_count,
                    'total_rows': total_rows_hint,
                    'created': created,
                    'updated': updated,
                    'unchanged': unchanged,
                    'upserted': upserted,
                    'skipped_empty_identifier': skipped_empty_identifier,
                    'skipped_unmatched_sku': skipped_unmatched_sku,
                    'skipped_invalid_date': skipped_invalid_date,
                    'skipped_template_sample': skipped_template_sample,
                    'errors': list(errors[:100]),
                }

                def _run_sales_perf_agg_background():
                    tid = safe_task_id
                    agg_hb_stop = {'stop': False}
                    agg_hb_state = {'stage': 'agg_init', 'message': ''}

                    def _agg_progress_overlay(extra):
                        pl = dict(_agg_stat_fields)
                        pl['status'] = 'success'
                        pl['task_id'] = tid
                        pl['state'] = 'success'
                        pl['advance_ts'] = _now_ts_text()
                        pl['ts'] = _now_ts_text()
                        pl.update(extra or {})
                        if pl.get('message'):
                            agg_hb_state['message'] = str(pl.get('message') or '')
                        _write_progress(tid, pl, merge=True)

                    def _agg_heartbeat_worker():
                        while not agg_hb_stop.get('stop'):
                            try:
                                _write_progress(tid, {
                                    'hb_ts': _now_ts_text(),
                                    'hb_phase': 'agg_refresh',
                                    'hb_stage': str(agg_hb_state.get('stage') or 'agg_refresh'),
                                    'advance_ts': _now_ts_text(),
                                }, merge=True)
                            except Exception:
                                pass
                            time.sleep(3.0)

                    try:
                        def _on_agg_progress(info):
                            try:
                                st = int(info.get('step') or 0)
                                tot = int(info.get('total') or 1)
                                seg = str(info.get('segment') or '')
                                pk = str(info.get('period_key') or '')
                                lab = '周' if seg == 'week' else '月'
                                pending = bool(info.get('pending'))
                                verb = '正在刷新' if pending else '已完成'
                                agg_hb_state['stage'] = f'agg_{seg}_{pk}'
                                _agg_progress_overlay({
                                    'phase': 'agg_refresh',
                                    'agg_refresh_status': 'running',
                                    'agg_refresh_step': st,
                                    'agg_refresh_total': tot,
                                    'message': f'明细已入库。{verb}周/月聚合：{lab} {pk}（{st}/{tot}）',
                                })
                            except Exception:
                                pass

                        threading.Thread(target=_agg_heartbeat_worker, daemon=True).start()

                        with self._get_db_connection_long() as agg_conn:
                            agg_ids = sorted(touched_agg_ids) if touched_agg_ids else None
                            if agg_ids:
                                _on_agg_progress({'step': 0, 'total': 1, 'segment': 'month', 'period_key': '…'})
                                self._refresh_sales_perf_agg_range(
                                    agg_conn,
                                    _agg_d0,
                                    _agg_d1,
                                    sales_product_ids=agg_ids,
                                    progress_hook=_on_agg_progress,
                                    segments=('month', 'week'),
                                )
                            # 全 SKU 仅刷新月聚合（周聚合全表扫描极易在月聚合前超时）
                            self._refresh_sales_perf_agg_range(
                                agg_conn,
                                _agg_d0,
                                _agg_d1,
                                sales_product_ids=None,
                                progress_hook=_on_agg_progress,
                                segments=('month',),
                            )
                        c0 = int(_agg_stat_fields.get('created') or 0)
                        u0 = int(_agg_stat_fields.get('upserted') or 0)
                        _agg_progress_overlay({
                            'phase': 'done',
                            'agg_refresh_status': 'complete',
                            'agg_refresh_background': False,
                            'agg_refresh_step': None,
                            'agg_refresh_total': None,
                            'message': f'处理完成（周/月聚合已刷新），匹配 {c0} 条，写入 {u0} 条',
                        })
                    except Exception as _e:
                        w = f'周/月聚合后台刷新未完成：{str(_e)[:200]}（日明细已在库；周/月看板可稍后再查）'
                        c0 = int(_agg_stat_fields.get('created') or 0)
                        u0 = int(_agg_stat_fields.get('upserted') or 0)
                        _agg_progress_overlay({
                            'phase': 'done',
                            'agg_refresh_status': 'error',
                            'agg_refresh_background': False,
                            'agg_refresh_warning': w,
                            'message': f'处理完成，匹配 {c0} 条，写入 {u0} 条；{w}',
                        })
                    finally:
                        agg_hb_stop['stop'] = True

                agg_bg_fn = _run_sales_perf_agg_background

            if not check_only and upserted <= 0:
                msg = (
                    f"未成功写入任何数据：处理行{row_count}，"
                    f"匹配行{created}，空标识跳过{skipped_empty_identifier}，"
                    f"未匹配SKU跳过{skipped_unmatched_sku}，日期无效跳过{skipped_invalid_date}。"
                    f"示例行跳过{skipped_template_sample}。"
                    "请检查模板SKU是否能匹配 sales_products.platform_sku/child_code。"
                )
                _write_progress(safe_task_id, {
                    'status': 'error',
                    'task_id': safe_task_id,
                    'state': 'error',
                    'processed_rows': row_count,
                    'total_rows': total_rows_hint,
                    'created': created,
                    'message': msg
                })
                return self.send_json({
                    'status': 'error',
                    'task_id': safe_task_id,
                    'message': msg,
                    'stats': {
                        'processed_rows': row_count,
                        'matched_rows': created,
                        'upserted_rows': upserted,
                        'skipped_empty_identifier': skipped_empty_identifier,
                        'skipped_unmatched_sku': skipped_unmatched_sku,
                        'skipped_invalid_date': skipped_invalid_date
                        , 'skipped_template_sample': skipped_template_sample
                    },
                    'errors': errors[:100]
                }, start_response)

            done_msg = f'处理完成，匹配 {created} 条，写入 {upserted} 条'
            if needs_agg_bg:
                done_msg = f'{done_msg}；周/月聚合正在后台刷新（完成后周/月看板将更新）'
            if agg_refresh_warning:
                done_msg = f"{done_msg}；{agg_refresh_warning}"

            done_payload = {
                'status': 'success',
                'task_id': safe_task_id,
                'state': 'success',
                'phase': 'done',
                'processed_rows': row_count,
                'total_rows': total_rows_hint,
                'created': created,
                'upserted': upserted,
                'skipped_empty_identifier': skipped_empty_identifier,
                'skipped_unmatched_sku': skipped_unmatched_sku,
                'skipped_invalid_date': skipped_invalid_date,
                'skipped_template_sample': skipped_template_sample,
                'skipped_all_zero': skipped_all_zero,
                'errors': errors[:100],
                'message': done_msg,
            }
            if agg_refresh_warning:
                done_payload['agg_refresh_warning'] = agg_refresh_warning
            if needs_agg_bg:
                done_payload['phase'] = 'agg_refresh'
                done_payload['agg_refresh_status'] = 'running'
                done_payload['agg_refresh_background'] = True
                done_payload['agg_refresh_step'] = 0
                done_payload['agg_refresh_total'] = None
                done_payload['message'] = f'明细已写入 {upserted} 条，正在后台刷新周/月聚合…'
            _write_progress(safe_task_id, done_payload)
            if agg_bg_fn:
                threading.Thread(target=agg_bg_fn, daemon=True).start()

            resp = {
                'status': 'success',
                'task_id': safe_task_id,
                'check_only': check_only,
                'created': created,
                'updated': updated,
                'unchanged': unchanged,
                'upserted': upserted,
                'skipped_empty_identifier': skipped_empty_identifier,
                'skipped_unmatched_sku': skipped_unmatched_sku,
                'skipped_invalid_date': skipped_invalid_date,
                'skipped_template_sample': skipped_template_sample,
                'skipped_all_zero': skipped_all_zero,
                'errors': errors,
                'total_rows': created + updated + unchanged + len(errors),
                'message': (
                    f"成功处理：匹配{created}条，写入{upserted}条，"
                    f"未匹配SKU{skipped_unmatched_sku}条，示例行{skipped_template_sample}条"
                ) if not check_only else f"预检完成，预计可识别{created}条数据",
            }
            if agg_refresh_warning:
                resp['agg_refresh_warning'] = agg_refresh_warning
                if not check_only:
                    resp['message'] = f"{resp['message']}；{agg_refresh_warning}"
            if needs_agg_bg:
                resp['agg_refresh_status'] = 'running'
                resp['agg_refresh_background'] = True
                if not check_only:
                    resp['message'] = f"{resp['message']}；周/月聚合后台刷新中"
            hb_stop_flag['stop'] = True
            return self.send_json(resp, start_response)
        except Exception as e:
            import traceback
            try:
                _write_progress(safe_task_id if 'safe_task_id' in locals() else '', {
                    'status': 'error',
                    'task_id': safe_task_id if 'safe_task_id' in locals() else '',
                    'state': 'error',
                    'processed_rows': 0,
                    'total_rows': 0,
                    'created': 0,
                    'message': str(e)[:200]
                })
            except Exception:
                pass
            try:
                hb_stop_flag['stop'] = True
            except Exception:
                pass
            error_str = str(e).lower()
            
            if 'sales_product_performances' in error_str or 'table' in error_str and 'doesn\'t exist' in error_str:
                return self.send_json({
                    'status': 'error',
                    'message': '❌ 表不存在：请先在数据库执行 scripts/sql/20260404_00_sales_product_performance.sql'
                }, start_response)
            
            if 'connection' in error_str or 'timeout' in error_str:
                return self.send_json({
                    'status': 'error',
                    'message': f'❌ 数据库连接错误：{str(e)[:100]}'
                }, start_response)
            
            tb = traceback.format_exc()
            return self.send_json({
                'status': 'error',
                'message': str(e)[:200],
                'detail': tb.split('\n')[-3:-1] if check_only else None
            }, start_response)

    # -------------------------------------------------------------------------
    # 产品表现看板（货号分组、图表、佣金/成本估算）
    # -------------------------------------------------------------------------

    def handle_sales_product_performance_dashboard_api(self, environ, method, start_response):
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)

            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            mode = str((query_params.get('mode', ['dashboard'])[0] or 'dashboard')).strip().lower()
            granularity = str((query_params.get('granularity', ['day'])[0] or 'day')).strip().lower()
            if granularity not in ('day', 'week', 'month'):
                granularity = 'day'

            def parse_csv_text(name):
                raw_list = query_params.get(name, [])
                tokens = []
                for raw in raw_list:
                    for token in re.split(r'[,，;；\s]+', str(raw or '').strip()):
                        t = token.strip()
                        if t and t not in tokens:
                            tokens.append(t)
                return tokens

            def parse_csv_int(name):
                values = []
                for token in parse_csv_text(name):
                    val = self._parse_int(token)
                    if val and val not in values:
                        values.append(val)
                return values

            def parse_date(value):
                text = str(value or '').strip()
                if not text:
                    return ''
                for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
                    try:
                        return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
                    except Exception:
                        continue
                return text[:10]

            metric_defs = [
                {'key': 'sales_qty', 'label': '销量', 'color': '#5b6aa8', 'agg': 'sum'},
                {'key': 'net_sales_amount', 'label': '净销售额', 'color': '#b85c5c', 'agg': 'sum'},
                {'key': 'order_qty', 'label': '订单量', 'color': '#bc7a3f', 'agg': 'sum'},
                {'key': 'session_total', 'label': 'Sessions-Total', 'color': '#44798c', 'agg': 'sum'},
                {'key': 'ad_impressions', 'label': '广告展示', 'color': '#7e8a57', 'agg': 'sum'},
                {'key': 'ad_clicks', 'label': '广告点击', 'color': '#8b6f9c', 'agg': 'sum'},
                {'key': 'ad_orders', 'label': '广告订单量', 'color': '#4d7ea8', 'agg': 'sum'},
                {'key': 'ad_spend', 'label': '广告花费', 'color': '#b96f3d', 'agg': 'sum'},
                {'key': 'ad_sales_amount', 'label': '广告销售额', 'color': '#6a8f4e', 'agg': 'sum'},
                {'key': 'refund_amount', 'label': '退款金额', 'color': '#9b4a4a', 'agg': 'sum'},
            ]

            with self._get_db_connection() as conn:
                if mode == 'filters':
                    has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                    has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                    fabric_join = "LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if has_fabric_id else ""
                    if has_fabric_id and has_fabric_text:
                        fabric_expr = "COALESCE(fm.fabric_code, v.fabric)"
                    elif has_fabric_id:
                        fabric_expr = "fm.fabric_code"
                    else:
                        fabric_expr = ("v.fabric" if has_fabric_text else "''")
                    with conn.cursor() as cur:
                        cur.execute(
                            f"""
                            SELECT sp.id, sp.platform_sku, {fabric_expr} AS fabric, v.spec_name,
                                pf.id AS sku_family_id, pf.sku_family,
                                   sh.id AS shop_id, sh.shop_name,
                                   pt.id AS platform_type_id, pt.name AS platform_type_name
                            FROM sales_products sp
                            LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                            LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                            LEFT JOIN shops sh ON sh.id = sp.shop_id
                            LEFT JOIN platform_types pt ON pt.id = sh.platform_type_id
                            {fabric_join}
                            ORDER BY pf.sku_family ASC, sp.platform_sku ASC
                            """
                        )
                        rows = cur.fetchall() or []

                        cur.execute("SELECT id, name FROM amazon_ad_operation_types ORDER BY sort_order ASC, id ASC")
                        op_types = cur.fetchall() or []

                    sku_families = []
                    sku_family_seen = set()
                    platform_skus = []
                    fabrics = []
                    specs = []
                    shops = []
                    platforms = []
                    f_seen = set()
                    s_seen = set()
                    shop_seen = set()
                    platform_seen = set()

                    for r in rows:
                        sf_id = self._parse_int(r.get('sku_family_id'))
                        sf = str(r.get('sku_family') or '').strip()
                        if sf_id and sf and sf_id not in sku_family_seen:
                            sku_family_seen.add(sf_id)
                            sku_families.append({'id': sf_id, 'name': sf})
                        sku = str(r.get('platform_sku') or '').strip()
                        if sku:
                            platform_skus.append({'id': self._parse_int(r.get('id')), 'name': sku})
                        fabric = str(r.get('fabric') or '').strip()
                        spec = str(r.get('spec_name') or '').strip()
                        if fabric and fabric not in f_seen:
                            f_seen.add(fabric)
                            fabrics.append(fabric)
                        if spec and spec not in s_seen:
                            s_seen.add(spec)
                            specs.append(spec)

                        shop_id = self._parse_int(r.get('shop_id'))
                        shop_name = str(r.get('shop_name') or '').strip()
                        if shop_id and shop_name and shop_id not in shop_seen:
                            shop_seen.add(shop_id)
                            shops.append({'id': shop_id, 'name': shop_name})

                        platform_id = self._parse_int(r.get('platform_type_id'))
                        platform_name = str(r.get('platform_type_name') or '').strip()
                        if platform_id and platform_name and platform_id not in platform_seen:
                            platform_seen.add(platform_id)
                            platforms.append({'id': platform_id, 'name': platform_name})

                    return self.send_json({
                        'status': 'success',
                        'filters': {
                            'sku_families': sku_families,
                            'platform_skus': platform_skus,
                            'fabrics': fabrics,
                            'spec_names': specs,
                            'shops': shops,
                            'platform_types': platforms,
                            'metrics': metric_defs,
                            'ad_operation_types': [{'id': self._parse_int(x.get('id')), 'name': x.get('name') or ''} for x in op_types]
                        }
                    }, start_response)

                import time
                perf_t_start = time.time()
                perf_timings = {}
                
                start_date = parse_date((query_params.get('start_date', [''])[0] or ''))
                end_date = parse_date((query_params.get('end_date', [''])[0] or ''))
                sku_family_ids = parse_csv_int('sku_family_ids')
                platform_skus = parse_csv_text('platform_skus')
                fabrics = parse_csv_text('fabrics')
                spec_names = parse_csv_text('spec_names')
                shop_ids = parse_csv_int('shop_ids')
                platform_type_ids = parse_csv_int('platform_type_ids')
                metric_keys = parse_csv_text('metric_keys')
                if not metric_keys:
                    metric_keys = ['sales_qty', 'net_sales_amount', 'order_qty', 'ad_spend', 'ad_sales_amount']
                include_todos = str((query_params.get('include_todos', ['0'])[0] or '0')).lower() in ('1', 'true', 'yes', 'on')
                include_ads = str((query_params.get('include_ads', ['0'])[0] or '0')).lower() in ('1', 'true', 'yes', 'on')
                ad_operation_type_ids = parse_csv_int('ad_operation_type_ids')
                perf_timings['params'] = time.time() - perf_t_start

                # 粒度限制（前端也会限制；这里做硬保护，避免误查导致超慢）
                def _to_date(text):
                    try:
                        return datetime.strptime(str(text or ''), '%Y-%m-%d').date()
                    except Exception:
                        return None

                max_days = None
                if granularity == 'day':
                    max_days = 31
                elif granularity == 'week':
                    max_days = 183

                if max_days and start_date and end_date:
                    sd = _to_date(start_date)
                    ed = _to_date(end_date)
                    if sd and ed:
                        if ed < sd:
                            sd, ed = ed, sd
                            start_date, end_date = sd.strftime('%Y-%m-%d'), ed.strftime('%Y-%m-%d')
                        diff = (ed - sd).days + 1
                        if diff > max_days:
                            return self.send_json({
                                'status': 'error',
                                'message': f'当前粒度「{granularity}」仅允许查看 {max_days} 天范围，请缩小日期区间后重试。',
                                'max_days': max_days
                            }, start_response)

                # === 图表：按粒度从不同数据源聚合 ===
                perf_t_a = time.time()
                chart_items = []

                if granularity == 'day':
                    agg_columns = "DATE(spp.record_date) as record_date"
                    for key in metric_keys:
                        metric = next((m for m in metric_defs if m['key'] == key), None)
                        if metric:
                            agg = metric['agg']
                            if agg == 'sum':
                                agg_columns += f", SUM(spp.{key}) as {key}"
                            elif agg == 'avg':
                                agg_columns += f", AVG(spp.{key}) as {key}"

                    agg_sql = [
                        f"""
                        SELECT {agg_columns}
                        FROM sales_product_performances spp
                        JOIN sales_products sp ON sp.id = spp.sales_product_id
                        LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                        LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                        LEFT JOIN shops sh ON sh.id = sp.shop_id
                        LEFT JOIN platform_types pt ON pt.id = sh.platform_type_id
                        WHERE 1=1
                        """
                    ]
                    agg_params = []
                    has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                    if start_date:
                        agg_sql.append(' AND spp.record_date >= %s')
                        agg_params.append(start_date)
                    if end_date:
                        agg_sql.append(' AND spp.record_date <= %s')
                        agg_params.append(end_date)
                    if sku_family_ids:
                        agg_sql.append(f" AND v.sku_family_id IN ({','.join(['%s'] * len(sku_family_ids))})")
                        agg_params.extend(sku_family_ids)
                    if platform_skus:
                        agg_sql.append(f" AND sp.platform_sku IN ({','.join(['%s'] * len(platform_skus))})")
                        agg_params.extend(platform_skus)
                    if fabrics and has_fabric_text:
                        agg_sql.append(f" AND v.fabric IN ({','.join(['%s'] * len(fabrics))})")
                        agg_params.extend(fabrics)
                    if spec_names:
                        agg_sql.append(f" AND v.spec_name IN ({','.join(['%s'] * len(spec_names))})")
                        agg_params.extend(spec_names)
                    if shop_ids:
                        agg_sql.append(f" AND sp.shop_id IN ({','.join(['%s'] * len(shop_ids))})")
                        agg_params.extend(shop_ids)
                    if platform_type_ids:
                        agg_sql.append(f" AND sh.platform_type_id IN ({','.join(['%s'] * len(platform_type_ids))})")
                        agg_params.extend(platform_type_ids)
                    agg_sql.append(' GROUP BY DATE(spp.record_date) ORDER BY record_date ASC')

                    with conn.cursor() as cur:
                        cur.execute(''.join(agg_sql), tuple(agg_params))
                        agg_rows = cur.fetchall() or []
                    for row in agg_rows:
                        item = {'record_date': row.get('record_date')}
                        for key in metric_keys:
                            val = row.get(key)
                            item[key] = round(float(val), 2) if val is not None else 0
                        # 用于前端展示真实范围（day 本身无截断概念）
                        item['range_start'] = str(item.get('record_date') or '')[:10]
                        item['range_end'] = str(item.get('record_date') or '')[:10]
                        chart_items.append(item)
                else:
                    src_table = 'sales_perf_agg_week' if granularity == 'week' else 'sales_perf_agg_month'
                    period_col = 'week_start' if granularity == 'week' else 'month_start'

                    # 周/月混合策略：
                    # - 中间完整周/月：用聚合表（更快）
                    # - 两端被日期区间截断的周/月：用明细表按真实起止即时汇总（更符合日常筛选逻辑）
                    def _to_date_obj(text):
                        try:
                            return datetime.strptime(str(text or ''), '%Y-%m-%d').date()
                        except Exception:
                            return None

                    def _fmt_date(d):
                        return d.strftime('%Y-%m-%d') if d else ''

                    sd_obj = _to_date_obj(start_date)
                    ed_obj = _to_date_obj(end_date)
                    if sd_obj and ed_obj and ed_obj < sd_obj:
                        sd_obj, ed_obj = ed_obj, sd_obj

                    def _bucket_start(d):
                        if not d:
                            return None
                        if granularity == 'month':
                            return d.replace(day=1)
                        # week: treat Monday as week_start
                        return d - timedelta(days=d.weekday())

                    def _bucket_end(bucket_start):
                        if not bucket_start:
                            return None
                        if granularity == 'month':
                            # last day of month
                            next_month = (bucket_start.replace(day=28) + timedelta(days=4)).replace(day=1)
                            return next_month - timedelta(days=1)
                        return bucket_start + timedelta(days=6)

                    def _add_common_filters(sql_parts, params, sp_alias, v_alias, sh_alias, fabric_alias=None, has_fabric_text=False):
                        if sku_family_ids:
                            sql_parts.append(f" AND {v_alias}.sku_family_id IN ({','.join(['%s'] * len(sku_family_ids))})")
                            params.extend(sku_family_ids)
                        if platform_skus:
                            sql_parts.append(f" AND {sp_alias}.platform_sku IN ({','.join(['%s'] * len(platform_skus))})")
                            params.extend(platform_skus)
                        if fabrics and has_fabric_text and fabric_alias:
                            sql_parts.append(f" AND {fabric_alias}.fabric IN ({','.join(['%s'] * len(fabrics))})")
                            params.extend(fabrics)
                        if spec_names:
                            sql_parts.append(f" AND {v_alias}.spec_name IN ({','.join(['%s'] * len(spec_names))})")
                            params.extend(spec_names)
                        if shop_ids:
                            sql_parts.append(f" AND {sp_alias}.shop_id IN ({','.join(['%s'] * len(shop_ids))})")
                            params.extend(shop_ids)
                        if platform_type_ids:
                            sql_parts.append(f" AND {sh_alias}.platform_type_id IN ({','.join(['%s'] * len(platform_type_ids))})")
                            params.extend(platform_type_ids)

                    has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')

                    bucket_items = []
                    if sd_obj and ed_obj:
                        first_bucket_start = _bucket_start(sd_obj)
                        last_bucket_start = _bucket_start(ed_obj)
                        first_bucket_end = _bucket_end(first_bucket_start)
                        last_bucket_end = _bucket_end(last_bucket_start)

                        left_partial = (sd_obj > first_bucket_start) if first_bucket_start else False
                        right_partial = (ed_obj < last_bucket_end) if last_bucket_end else False

                        # helper: query raw for a partial bucket and return one aggregated row
                        def _query_raw_bucket(bucket_start, rng_start, rng_end):
                            cols = ["%s AS record_date"]
                            for key in metric_keys:
                                cols.append(f"SUM(COALESCE(spp.{key},0)) AS {key}")
                            sql_parts = [
                                f"""
                                SELECT {', '.join(cols)}
                                FROM sales_product_performances spp
                                JOIN sales_products sp ON sp.id = spp.sales_product_id
                                LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                                LEFT JOIN shops sh ON sh.id = sp.shop_id
                                WHERE 1=1
                                """
                            ]
                            params = [_fmt_date(bucket_start)]
                            sql_parts.append(" AND spp.record_date >= %s")
                            params.append(_fmt_date(rng_start))
                            sql_parts.append(" AND spp.record_date <= %s")
                            params.append(_fmt_date(rng_end))
                            _add_common_filters(sql_parts, params, 'sp', 'v', 'sh', fabric_alias='v', has_fabric_text=has_fabric_text)
                            with conn.cursor() as cur:
                                cur.execute(''.join(sql_parts), tuple(params))
                                row = cur.fetchone() or {}
                            out = {'record_date': row.get('record_date')}
                            for key in metric_keys:
                                val = row.get(key)
                                out[key] = round(float(val), 2) if val is not None else 0
                            # 记录该桶真实参与计算的数据范围（用于前端显示）
                            out['range_start'] = _fmt_date(rng_start)
                            out['range_end'] = _fmt_date(rng_end)
                            return out

                        # left edge partial
                        if left_partial and first_bucket_start:
                            bucket_items.append(_query_raw_bucket(first_bucket_start, sd_obj, min(ed_obj, first_bucket_end)))

                        # middle full buckets via agg table
                        # compute next/prev bucket start safely
                        def _next_bucket_start(b):
                            if granularity == 'month':
                                y, m = b.year, b.month
                                if m == 12:
                                    return datetime(y + 1, 1, 1).date()
                                return datetime(y, m + 1, 1).date()
                            return b + timedelta(days=7)

                        def _prev_bucket_start(b):
                            if granularity == 'month':
                                y, m = b.year, b.month
                                if m == 1:
                                    return datetime(y - 1, 12, 1).date()
                                return datetime(y, m - 1, 1).date()
                            return b - timedelta(days=7)

                        full_from = _next_bucket_start(first_bucket_start) if left_partial else first_bucket_start
                        full_to = _prev_bucket_start(last_bucket_start) if right_partial else last_bucket_start

                        if full_from and full_to and full_from <= full_to:
                            cols = [f"DATE(a.{period_col}) AS record_date"]
                            for key in metric_keys:
                                cols.append(f"SUM(COALESCE(a.{key},0)) AS {key}")
                            agg_sql = [
                                f"""
                                SELECT {', '.join(cols)}
                                FROM {src_table} a
                                JOIN sales_products sp ON sp.id = a.sales_product_id
                                LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                                LEFT JOIN shops sh ON sh.id = sp.shop_id
                                WHERE 1=1
                                """
                            ]
                            agg_params = []
                            agg_sql.append(f" AND a.{period_col} >= %s")
                            agg_params.append(_fmt_date(full_from))
                            agg_sql.append(f" AND a.{period_col} <= %s")
                            agg_params.append(_fmt_date(full_to))
                            _add_common_filters(agg_sql, agg_params, 'sp', 'v', 'sh', fabric_alias='v', has_fabric_text=has_fabric_text)
                            agg_sql.append(" GROUP BY DATE(record_date) ORDER BY record_date ASC")
                            with conn.cursor() as cur:
                                cur.execute(''.join(agg_sql), tuple(agg_params))
                                agg_rows = cur.fetchall() or []
                            for row in agg_rows:
                                item = {'record_date': row.get('record_date')}
                                for key in metric_keys:
                                    val = row.get(key)
                                    item[key] = round(float(val), 2) if val is not None else 0
                                # 完整桶：真实范围即完整周/月
                                try:
                                    bstart = _to_date_obj(item.get('record_date'))
                                except Exception:
                                    bstart = None
                                bend = _bucket_end(bstart) if bstart else None
                                item['range_start'] = _fmt_date(bstart)
                                item['range_end'] = _fmt_date(bend)
                                bucket_items.append(item)

                        # right edge partial：多桶时补最后一个截断桶；单桶且仅右端截断时 first==last，
                        # 若仍用「last != first」会整段被跳过（例如按月 2026-05-01～2026-05-10）。
                        # 单桶且左端已用 raw 补过时不再追加，避免与左段重复。
                        if right_partial and last_bucket_start:
                            same_bucket = (
                                first_bucket_start is not None
                                and last_bucket_start is not None
                                and first_bucket_start == last_bucket_start
                            )
                            if same_bucket and (not left_partial):
                                bucket_items.append(_query_raw_bucket(last_bucket_start, sd_obj, ed_obj))
                            elif not same_bucket:
                                bucket_items.append(_query_raw_bucket(last_bucket_start, max(sd_obj, last_bucket_start), ed_obj))

                        # sort by record_date
                        bucket_items.sort(key=lambda x: str(x.get('record_date') or ''))
                        chart_items.extend(bucket_items)
                    else:
                        # fallback: keep legacy behavior if dates are missing/unparseable
                        cols = [f"DATE(a.{period_col}) AS record_date"]
                        for key in metric_keys:
                            cols.append(f"SUM(COALESCE(a.{key},0)) AS {key}")

                        agg_sql = [
                            f"""
                            SELECT {', '.join(cols)}
                            FROM {src_table} a
                            JOIN sales_products sp ON sp.id = a.sales_product_id
                            LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                            LEFT JOIN shops sh ON sh.id = sp.shop_id
                            WHERE 1=1
                            """
                        ]
                        agg_params = []
                        if start_date:
                            agg_sql.append(f" AND a.{period_col} >= %s")
                            agg_params.append(start_date)
                        if end_date:
                            agg_sql.append(f" AND a.{period_col} <= %s")
                            agg_params.append(end_date)
                        _add_common_filters(agg_sql, agg_params, 'sp', 'v', 'sh', fabric_alias='v', has_fabric_text=has_fabric_text)
                        agg_sql.append(" GROUP BY DATE(record_date) ORDER BY record_date ASC")

                        with conn.cursor() as cur:
                            cur.execute(''.join(agg_sql), tuple(agg_params))
                            agg_rows = cur.fetchall() or []
                        for row in agg_rows:
                            item = {'record_date': row.get('record_date')}
                            for key in metric_keys:
                                val = row.get(key)
                                item[key] = round(float(val), 2) if val is not None else 0
                            # fallback：尽量提供范围
                            rd = str(item.get('record_date') or '')[:10]
                            item['range_start'] = rd
                            item['range_end'] = rd
                            chart_items.append(item)

                perf_timings['chart_agg'] = time.time() - perf_t_a

                # === 货号分组（按粒度选择数据源）===
                perf_t_g = time.time()
                has_fabric_id = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
                has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
                fabric_join = "LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id" if has_fabric_id else ""
                if has_fabric_id and has_fabric_text:
                    fabric_expr = "COALESCE(fm.fabric_code, v.fabric)"
                elif has_fabric_id:
                    fabric_expr = "fm.fabric_code"
                else:
                    fabric_expr = ("v.fabric" if has_fabric_text else "''")
                # --- 货号分组明细：预估成本（销售变体→下单SKU 链接表加权）；佣金按平台+细分类目映射规则在 Python 侧计算 ---
                # estimated_product_cost_usd：下单产品 cost_usd 表示「产品至海外仓」成本（BOM），按链接数量加权汇总后再×周期销量。
                # estimated_last_mile_freight_usd：下单产品 last_mile_avg_freight_usd 预估尾程，同样按链接数量加权汇总后再×周期销量；仅当店铺 handles_last_mile=1 时计入。
                has_op_reship = self._table_has_column(conn, 'order_products', 'is_reship_accessory')
                has_op_last_mile = self._table_has_column(conn, 'order_products', 'last_mile_avg_freight_usd')
                reship_clause = ' AND COALESCE(op.is_reship_accessory,0)=0 ' if has_op_reship else ''
                lm_factor_sql = self._shop_handles_last_mile_factor_sql(conn, 'sh', 'pt')
                lm_est_sql = (
                    f"(MAX(COALESCE(est_unit_cost.unit_last_mile_freight_usd, 0)) "
                    f"* SUM(COALESCE(spp.sales_qty,0)) * MAX({lm_factor_sql}))"
                )
                lm_weighted_sum = (
                    "SUM(COALESCE(op.last_mile_avg_freight_usd, 0) * COALESCE(svol.quantity, 1))"
                    if has_op_last_mile else "0"
                )
                cost_join_variant = (
                    "LEFT JOIN ("
                    " SELECT v.id AS variant_id,"
                    "   SUM(COALESCE(op.cost_usd, 0) * COALESCE(svol.quantity, 1)) AS unit_bom_cost_usd,"
                    f"   {lm_weighted_sum} AS unit_last_mile_freight_usd"
                    " FROM sales_product_variants v"
                    " INNER JOIN sales_variant_order_links svol ON svol.variant_id = v.id"
                    " INNER JOIN order_products op ON op.id = svol.order_product_id"
                    f" WHERE 1=1{reship_clause}"
                    " GROUP BY v.id"
                    ") est_unit_cost ON est_unit_cost.variant_id = sp.variant_id"
                )
                comm_rules_cache = self._commission_load_rules_cache(conn)
                # 销售额 = 销售产品售价(USD)×周期销量；折扣率=(销售额−净销售额)/销售额；退款率=退款金额/净销售额
                gross_expr_day = "SUM(COALESCE(sp.sale_price_usd, 0) * COALESCE(spp.sales_qty, 0))"
                gross_sales_sql_day = f"({gross_expr_day}) AS gross_sales_amount"
                discount_rate_sql_day = (
                    f"CASE WHEN ({gross_expr_day}) > 0 THEN "
                    f"(({gross_expr_day}) - SUM(COALESCE(spp.net_sales_amount, 0))) / ({gross_expr_day}) "
                    "ELSE 0 END AS discount_rate"
                )
                refund_rate_sql_day = (
                    "CASE WHEN SUM(COALESCE(spp.net_sales_amount, 0)) > 0 THEN "
                    "SUM(COALESCE(spp.refund_amount, 0)) / SUM(COALESCE(spp.net_sales_amount, 0)) "
                    "ELSE 0 END AS refund_rate"
                )
                gross_expr_week = "SUM(COALESCE(sp.sale_price_usd, 0) * COALESCE(a.sales_qty, 0))"
                gross_sales_sql_week = f"({gross_expr_week}) AS gross_sales_amount"
                discount_rate_sql_week = (
                    f"CASE WHEN ({gross_expr_week}) > 0 THEN "
                    f"(({gross_expr_week}) - SUM(COALESCE(a.net_sales_amount, 0))) / ({gross_expr_week}) "
                    "ELSE 0 END AS discount_rate"
                )
                refund_rate_sql_week = (
                    "CASE WHEN SUM(COALESCE(a.net_sales_amount, 0)) > 0 THEN "
                    "SUM(COALESCE(a.refund_amount, 0)) / SUM(COALESCE(a.net_sales_amount, 0)) "
                    "ELSE 0 END AS refund_rate"
                )
                params = []
                sql = []
                # 货号分组「全量汇总」始终按筛选起止日在明细表汇总（与 groups_by_bucket、图表截断桶一致）。
                # granularity 仅影响 chart_items 的时间桶展示，不改变下方筛选范围内汇总口径。
                group_cols = [
                    "sp.id AS sp_id",
                    "sp.platform_sku",
                    f"{fabric_expr} AS fabric",
                    "v.spec_name",
                    "v.sku_family_id",
                    "pf.sku_family",
                    "TRIM(COALESCE(pf.category, '')) AS product_category",
                    "sh.platform_type_id AS platform_type_id",
                    "MIN(DATE(spp.record_date)) AS min_date",
                    "MAX(DATE(spp.record_date)) AS max_date",
                    "COUNT(1) AS `rows`",
                    "SUM(COALESCE(spp.sales_qty,0)) AS sales_qty",
                    "SUM(COALESCE(spp.net_sales_amount,0)) AS net_sales_amount",
                    gross_sales_sql_day,
                    discount_rate_sql_day,
                    "SUM(COALESCE(spp.order_qty,0)) AS order_qty",
                    "SUM(COALESCE(spp.session_total,0)) AS session_total",
                    "SUM(COALESCE(spp.ad_impressions,0)) AS ad_impressions",
                    "SUM(COALESCE(spp.ad_clicks,0)) AS ad_clicks",
                    "SUM(COALESCE(spp.ad_orders,0)) AS ad_orders",
                    "SUM(COALESCE(spp.ad_spend,0)) AS ad_spend",
                    "SUM(COALESCE(spp.ad_sales_amount,0)) AS ad_sales_amount",
                    "SUM(COALESCE(spp.refund_amount,0)) AS refund_amount",
                    refund_rate_sql_day,
                    "(MAX(COALESCE(est_unit_cost.unit_bom_cost_usd, 0)) * SUM(COALESCE(spp.sales_qty,0))) AS estimated_product_cost_usd",
                    f"{lm_est_sql} AS estimated_last_mile_freight_usd",
                ]

                sql = [
                    f"""
                    SELECT {', '.join(group_cols)}
                    FROM sales_product_performances spp
                    JOIN sales_products sp ON sp.id = spp.sales_product_id
                    LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                    LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                    LEFT JOIN shops sh ON sh.id = sp.shop_id
                    LEFT JOIN platform_types pt ON pt.id = sh.platform_type_id
                    {fabric_join}
                    {cost_join_variant}
                    WHERE 1=1
                    """
                ]
                if start_date:
                    sql.append(' AND spp.record_date >= %s')
                    params.append(start_date)
                if end_date:
                    sql.append(' AND spp.record_date <= %s')
                    params.append(end_date)
                if sku_family_ids:
                    sql.append(f" AND v.sku_family_id IN ({','.join(['%s'] * len(sku_family_ids))})")
                    params.extend(sku_family_ids)
                if platform_skus:
                    sql.append(f" AND sp.platform_sku IN ({','.join(['%s'] * len(platform_skus))})")
                    params.extend(platform_skus)
                if fabrics and has_fabric_text:
                    sql.append(f" AND v.fabric IN ({','.join(['%s'] * len(fabrics))})")
                    params.extend(fabrics)
                if spec_names:
                    sql.append(f" AND v.spec_name IN ({','.join(['%s'] * len(spec_names))})")
                    params.extend(spec_names)
                if shop_ids:
                    sql.append(f" AND sp.shop_id IN ({','.join(['%s'] * len(shop_ids))})")
                    params.extend(shop_ids)
                if platform_type_ids:
                    sql.append(f" AND sh.platform_type_id IN ({','.join(['%s'] * len(platform_type_ids))})")
                    params.extend(platform_type_ids)
                sql.append(' GROUP BY sp.id, sp.platform_sku, fabric, v.spec_name, v.sku_family_id, pf.sku_family, pf.category, sh.platform_type_id')
                sql.append(' ORDER BY pf.sku_family ASC, sp.platform_sku ASC')

                with conn.cursor() as cur:
                    cur.execute(''.join(sql), tuple(params))
                    rows = cur.fetchall() or []

                def _perf_group_item_commission_extras(row, bom_u, lm_u):
                    pt_id = self._parse_int(row.get('platform_type_id'))
                    cat = str(row.get('product_category') or '').strip()
                    net = float(row.get('net_sales_amount') or 0)
                    if comm_rules_cache.get('ready'):
                        comm_result = self._commission_compute_for_context(
                            comm_rules_cache, pt_id, cat, net, mode='period',
                        )
                    else:
                        comm_result = {
                            'commission_status': 'unavailable',
                            'commission_message': self.COMMISSION_UNAVAILABLE_LABEL,
                            'est_referral_commission_usd': None,
                            'commission_rate': None,
                        }
                    return self._commission_perf_derived_with_commission(
                        bom_u, lm_u, net,
                        row.get('ad_spend'), row.get('refund_amount'),
                        row.get('gross_sales_amount'), comm_result,
                    )

                group_map = {}
                for row in rows:
                    sp_id = self._parse_int(row.get('sp_id'))
                    sf_id = self._parse_int(row.get('sku_family_id'))
                    sf_name = str(row.get('sku_family') or '未分组货号').strip() or '未分组货号'
                    sku = str(row.get('platform_sku') or '').strip()
                    gkey = f"{sf_id or 0}:{sf_name}"
                    group = group_map.setdefault(gkey, {
                        'sku_family_id': sf_id,
                        'sku_family': sf_name,
                        'items': []
                    })
                    bom_u = round(float(row.get('estimated_product_cost_usd') or 0), 2)
                    lm_u = round(float(row.get('estimated_last_mile_freight_usd') or 0), 2)
                    item_row = {
                        'sales_product_id': sp_id,
                        'platform_sku': sku,
                        'fabric': row.get('fabric') or '',
                        'spec_name': row.get('spec_name') or '',
                        'product_category': str(row.get('product_category') or '').strip(),
                        'platform_type_id': self._parse_int(row.get('platform_type_id')),
                        'min_date': str(row.get('min_date') or ''),
                        'max_date': str(row.get('max_date') or ''),
                        'rows': self._parse_int(row.get('rows')) or 0,
                        'sales_qty': row.get('sales_qty') or 0,
                        'net_sales_amount': row.get('net_sales_amount') or 0,
                        'gross_sales_amount': round(float(row.get('gross_sales_amount') or 0), 2),
                        'discount_rate': round(float(row.get('discount_rate') or 0), 6),
                        'order_qty': row.get('order_qty') or 0,
                        'session_total': row.get('session_total') or 0,
                        'ad_impressions': row.get('ad_impressions') or 0,
                        'ad_clicks': row.get('ad_clicks') or 0,
                        'ad_orders': row.get('ad_orders') or 0,
                        'ad_spend': row.get('ad_spend') or 0,
                        'ad_sales_amount': row.get('ad_sales_amount') or 0,
                        'refund_amount': row.get('refund_amount') or 0,
                        'refund_rate': round(float(row.get('refund_rate') or 0), 6),
                        'estimated_product_cost_usd': bom_u,
                        'estimated_last_mile_freight_usd': lm_u,
                    }
                    item_row.update(_perf_group_item_commission_extras(row, bom_u, lm_u))
                    group['items'].append(item_row)

                groups = list(group_map.values())
                for g in groups:
                    g['items'].sort(key=lambda x: x.get('platform_sku') or '')
                groups.sort(key=lambda x: x.get('sku_family') or '')

                total_groups = len(groups)
                total_items = sum(len(g.get('items') or []) for g in groups)
                perf_timings['groups'] = time.time() - perf_t_g

                # === 货号分组（按图表桶范围预计算，用于前端点击趋势时“无二次请求”切换明细） ===
                # 注意：这里统一使用原始明细表在桶范围内计算（避免 week/month 聚合表在截断桶上失真）
                groups_by_bucket = {}
                try:
                    # 取图表桶的真实范围（chart_items 已带 range_start/range_end）
                    bucket_ranges = []
                    for it in (chart_items or []):
                        bkey = str(it.get('record_date') or '')[:10]
                        rs = str(it.get('range_start') or '')[:10]
                        rng_end_text = str(it.get('range_end') or '')[:10]
                        if bkey and rs and rng_end_text:
                            bucket_ranges.append((bkey, rs, rng_end_text))

                    if bucket_ranges:
                        # 复用 day 分组 SQL（按日期范围汇总到平台SKU粒度，再按货号分组）
                        def _query_groups_raw_range(rng_start, rng_end):
                            local_params = []
                            group_cols_local = [
                                "sp.id AS sp_id",
                                "sp.platform_sku",
                                f"{fabric_expr} AS fabric",
                                "v.spec_name",
                                "v.sku_family_id",
                                "pf.sku_family",
                                "TRIM(COALESCE(pf.category, '')) AS product_category",
                                "sh.platform_type_id AS platform_type_id",
                                "MIN(DATE(spp.record_date)) AS min_date",
                                "MAX(DATE(spp.record_date)) AS max_date",
                                "COUNT(1) AS `rows`",
                                "SUM(COALESCE(spp.sales_qty,0)) AS sales_qty",
                                "SUM(COALESCE(spp.net_sales_amount,0)) AS net_sales_amount",
                                gross_sales_sql_day,
                                discount_rate_sql_day,
                                "SUM(COALESCE(spp.order_qty,0)) AS order_qty",
                                "SUM(COALESCE(spp.session_total,0)) AS session_total",
                                "SUM(COALESCE(spp.ad_impressions,0)) AS ad_impressions",
                                "SUM(COALESCE(spp.ad_clicks,0)) AS ad_clicks",
                                "SUM(COALESCE(spp.ad_orders,0)) AS ad_orders",
                                "SUM(COALESCE(spp.ad_spend,0)) AS ad_spend",
                                "SUM(COALESCE(spp.ad_sales_amount,0)) AS ad_sales_amount",
                                "SUM(COALESCE(spp.refund_amount,0)) AS refund_amount",
                                refund_rate_sql_day,
                                "(MAX(COALESCE(est_unit_cost.unit_bom_cost_usd, 0)) * SUM(COALESCE(spp.sales_qty,0))) AS estimated_product_cost_usd",
                                f"{lm_est_sql} AS estimated_last_mile_freight_usd",
                            ]
                            local_sql = [
                                f"""
                                SELECT {', '.join(group_cols_local)}
                                FROM sales_product_performances spp
                                JOIN sales_products sp ON sp.id = spp.sales_product_id
                                LEFT JOIN sales_product_variants v ON v.id = sp.variant_id
                                LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                                LEFT JOIN shops sh ON sh.id = sp.shop_id
                                LEFT JOIN platform_types pt ON pt.id = sh.platform_type_id
                                {fabric_join}
                                {cost_join_variant}
                                WHERE 1=1
                                """
                            ]
                            if rng_start:
                                local_sql.append(' AND spp.record_date >= %s')
                                local_params.append(rng_start)
                            if rng_end:
                                local_sql.append(' AND spp.record_date <= %s')
                                local_params.append(rng_end)
                            if sku_family_ids:
                                local_sql.append(f" AND v.sku_family_id IN ({','.join(['%s'] * len(sku_family_ids))})")
                                local_params.extend(sku_family_ids)
                            if platform_skus:
                                local_sql.append(f" AND sp.platform_sku IN ({','.join(['%s'] * len(platform_skus))})")
                                local_params.extend(platform_skus)
                            if fabrics and has_fabric_text:
                                local_sql.append(f" AND v.fabric IN ({','.join(['%s'] * len(fabrics))})")
                                local_params.extend(fabrics)
                            if spec_names:
                                local_sql.append(f" AND v.spec_name IN ({','.join(['%s'] * len(spec_names))})")
                                local_params.extend(spec_names)
                            if shop_ids:
                                local_sql.append(f" AND sp.shop_id IN ({','.join(['%s'] * len(shop_ids))})")
                                local_params.extend(shop_ids)
                            if platform_type_ids:
                                local_sql.append(f" AND sh.platform_type_id IN ({','.join(['%s'] * len(platform_type_ids))})")
                                local_params.extend(platform_type_ids)
                            local_sql.append(' GROUP BY sp.id, sp.platform_sku, fabric, v.spec_name, v.sku_family_id, pf.sku_family, pf.category, sh.platform_type_id')
                            local_sql.append(' ORDER BY pf.sku_family ASC, sp.platform_sku ASC')
                            with conn.cursor() as cur2:
                                cur2.execute(''.join(local_sql), tuple(local_params))
                                local_rows = cur2.fetchall() or []

                            local_group_map = {}
                            for r in local_rows:
                                sp_id = self._parse_int(r.get('sp_id'))
                                sf_id = self._parse_int(r.get('sku_family_id'))
                                sf_name = str(r.get('sku_family') or '未分组货号').strip() or '未分组货号'
                                sku = str(r.get('platform_sku') or '').strip()
                                gkey = f"{sf_id or 0}:{sf_name}"
                                grp = local_group_map.setdefault(gkey, {
                                    'sku_family_id': sf_id,
                                    'sku_family': sf_name,
                                    'items': []
                                })
                                bom_r = round(float(r.get('estimated_product_cost_usd') or 0), 2)
                                lm_r = round(float(r.get('estimated_last_mile_freight_usd') or 0), 2)
                                item_r = {
                                    'sales_product_id': sp_id,
                                    'platform_sku': sku,
                                    'fabric': r.get('fabric') or '',
                                    'spec_name': r.get('spec_name') or '',
                                    'product_category': str(r.get('product_category') or '').strip(),
                                    'platform_type_id': self._parse_int(r.get('platform_type_id')),
                                    'min_date': str(r.get('min_date') or ''),
                                    'max_date': str(r.get('max_date') or ''),
                                    'rows': self._parse_int(r.get('rows')) or 0,
                                    'sales_qty': r.get('sales_qty') or 0,
                                    'net_sales_amount': r.get('net_sales_amount') or 0,
                                    'gross_sales_amount': round(float(r.get('gross_sales_amount') or 0), 2),
                                    'discount_rate': round(float(r.get('discount_rate') or 0), 6),
                                    'order_qty': r.get('order_qty') or 0,
                                    'session_total': r.get('session_total') or 0,
                                    'ad_impressions': r.get('ad_impressions') or 0,
                                    'ad_clicks': r.get('ad_clicks') or 0,
                                    'ad_orders': r.get('ad_orders') or 0,
                                    'ad_spend': r.get('ad_spend') or 0,
                                    'ad_sales_amount': r.get('ad_sales_amount') or 0,
                                    'refund_amount': r.get('refund_amount') or 0,
                                    'refund_rate': round(float(r.get('refund_rate') or 0), 6),
                                    'estimated_product_cost_usd': bom_r,
                                    'estimated_last_mile_freight_usd': lm_r,
                                }
                                item_r.update(_perf_group_item_commission_extras(r, bom_r, lm_r))
                                grp['items'].append(item_r)
                            local_groups = list(local_group_map.values())
                            for gg in local_groups:
                                gg['items'].sort(key=lambda x: x.get('platform_sku') or '')
                            local_groups.sort(key=lambda x: x.get('sku_family') or '')
                            return local_groups

                        for bkey, rs, rng_end in bucket_ranges:
                            groups_by_bucket[bkey] = _query_groups_raw_range(rs, rng_end)
                except Exception:
                    groups_by_bucket = {}

                events = []
                perf_timings['events'] = 0
                
                # 禁用 todos/ads 即时查询（改为前端异步加载，加快响应）
                # if include_todos and (target_sp_ids or target_sf_ids):
                #     ... todos 查询代码 ...
                # if include_ads and (target_sp_ids or target_sf_ids):
                #     ... ads 查询代码 ...

                ad_type_options = []
                with conn.cursor() as cur:
                    cur.execute("SELECT id, name FROM amazon_ad_operation_types ORDER BY sort_order ASC, id ASC LIMIT 100")
                    ad_type_options = [{'id': self._parse_int(x.get('id')), 'name': x.get('name') or ''} for x in (cur.fetchall() or [])]

                events.sort(key=lambda x: (x.get('event_date') or '', x.get('event_datetime') or '', x.get('event_type') or ''))
                
                perf_timings['total'] = time.time() - perf_t_start
                
                return self.send_json({
                    'status': 'success',
                    'groups': groups,
                    'groups_by_bucket': groups_by_bucket,
                    'total_groups': total_groups,
                    'total_items': total_items,
                    'chart_items': chart_items,
                    'metric_defs': metric_defs,
                    'events': events,
                    'ad_operation_types': ad_type_options,
                    '_perf': {
                        'total_ms': round(perf_timings['total'] * 1000, 1),
                        'breakdown': {
                            'params_ms': round(perf_timings['params'] * 1000, 1),
                            'chart_agg_ms': round(perf_timings['chart_agg'] * 1000, 1),
                            'groups_ms': round(perf_timings['groups'] * 1000, 1),
                            'events_ms': round(perf_timings['events'] * 1000, 1)
                        }
                    }
                }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _normalize_sales_order_links(self, links):
        items = []
        if not isinstance(links, list):
            return items
        qty_by_order_id = {}
        for entry in links:
            if not isinstance(entry, dict):
                continue
            order_product_id = self._parse_int(entry.get('order_product_id'))
            quantity = self._parse_int(entry.get('quantity')) or 1
            if not order_product_id:
                continue
            qty_by_order_id[order_product_id] = qty_by_order_id.get(order_product_id, 0) + max(1, quantity)

        for order_product_id, quantity in qty_by_order_id.items():
            items.append({'order_product_id': order_product_id, 'quantity': max(1, quantity)})
        return items

    def _get_or_create_sales_variant(self, conn, sku_family_id, spec_name, fabric, fabric_id=None):
        family_id = self._parse_int(sku_family_id)
        if not family_id:
            raise ValueError('Missing sku_family_id')
        spec = str(spec_name or '').strip()
        fab = str(fabric or '').strip()
        fid = self._parse_int(fabric_id) or None
        with conn.cursor() as cur:
            has_fid = self._table_has_column(conn, 'sales_product_variants', 'fabric_id')
            has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
            if has_fid and not fid and fab:
                fid = self._resolve_fabric_material_id_from_label(conn, fab, cur)

            # 1) Prefer selecting existing row first (prevents duplicates even if UNIQUE index is missing).
            if has_fid:
                cur.execute(
                    """
                    SELECT id FROM sales_product_variants
                    WHERE sku_family_id=%s AND spec_name=%s AND COALESCE(fabric_id,0)=COALESCE(%s,0)
                    ORDER BY id ASC
                    LIMIT 1
                    """,
                    (family_id, spec, fid),
                )
            else:
                cur.execute(
                    """
                    SELECT id FROM sales_product_variants
                    WHERE sku_family_id=%s AND spec_name=%s AND fabric=%s
                    ORDER BY id ASC
                    LIMIT 1
                    """,
                    (family_id, spec, fab),
                )
            row = cur.fetchone() or {}
            existing_id = self._parse_int(row.get('id')) or 0
            if existing_id:
                sets = []
                params = []
                # fabric_id identity is fixed for a variant; never mutate it here.
                if has_fabric_text and fab:
                    sets.append("fabric=COALESCE(NULLIF(%s,''), fabric)")
                    params.append(fab)
                if self._table_has_column(conn, 'sales_product_variants', 'updated_at'):
                    sets.append("updated_at=CURRENT_TIMESTAMP")
                if sets:
                    cur.execute(
                        f"UPDATE sales_product_variants SET {', '.join(sets)} WHERE id=%s",
                        tuple(params + [existing_id]),
                    )
                variant_id = existing_id
            else:
                # 2) Insert when truly missing.
                if has_fid:
                    cols = ["sku_family_id", "spec_name", "fabric_id"]
                    vals = [family_id, spec, fid]
                    if has_fabric_text:
                        cols.append("fabric")
                        vals.append(fab)
                    ph = ", ".join(["%s"] * len(cols))
                    cur.execute(
                        f"INSERT INTO sales_product_variants ({', '.join(cols)}) VALUES ({ph})",
                        tuple(vals),
                    )
                else:
                    cur.execute(
                        "INSERT INTO sales_product_variants (sku_family_id, spec_name, fabric) VALUES (%s, %s, %s)",
                        (family_id, spec, fab),
                    )
                variant_id = cur.lastrowid

        if not variant_id:
            raise RuntimeError('创建或读取销售变体失败')
        return variant_id

    def _replace_sales_variant_order_links(self, conn, variant_id, links):
        if not variant_id:
            raise ValueError('Invalid variant_id')
        merged_links = self._normalize_sales_order_links(links)
        with conn.cursor() as cur:
            cur.execute("DELETE FROM sales_variant_order_links WHERE variant_id=%s", (variant_id,))
            if not merged_links:
                return
            placeholders = ','.join(['(%s, %s, %s)'] * len(merged_links))
            values = []
            for entry in merged_links:
                values.extend([variant_id, entry['order_product_id'], entry['quantity']])
            cur.execute(
                f"""
                INSERT INTO sales_variant_order_links (variant_id, order_product_id, quantity)
                VALUES {placeholders}
                """,
                values
            )

    def _load_sales_variant_metrics(self, conn, variant_ids):
        ids = sorted({self._parse_int(v) for v in (variant_ids or []) if self._parse_int(v)})
        if not ids:
            return {}
        placeholders = ','.join(['%s'] * len(ids))
        metrics = {v: {
            'warehouse_cost_usd': 0.0,
            'last_mile_cost_usd': 0.0,
            'package_length_in': 0.0,
            'package_width_in': 0.0,
            'package_height_in': 0.0,
            'net_weight_lbs': 0.0,
            'gross_weight_lbs': 0.0,
            'order_sku_links': []
        } for v in ids}

        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT
                    l.variant_id,
                    l.order_product_id,
                    l.quantity,
                    op.sku,
                    op.cost_usd,
                    op.last_mile_avg_freight_usd,
                    op.package_length_in,
                    op.package_width_in,
                    op.package_height_in,
                    op.net_weight_lbs,
                    op.gross_weight_lbs
                FROM sales_variant_order_links l
                JOIN order_products op ON op.id = l.order_product_id
                WHERE l.variant_id IN ({placeholders})
                ORDER BY l.variant_id ASC, op.id ASC
                """,
                ids
            )
            rows = cur.fetchall() or []

        for row in rows:
            variant_id = self._parse_int(row.get('variant_id'))
            if not variant_id or variant_id not in metrics:
                continue
            qty = max(1, self._parse_int(row.get('quantity')) or 1)
            bucket = metrics[variant_id]
            bucket['warehouse_cost_usd'] += float(row.get('cost_usd') or 0) * qty
            bucket['last_mile_cost_usd'] += float(row.get('last_mile_avg_freight_usd') or 0) * qty
            bucket['package_length_in'] = max(bucket['package_length_in'], float(row.get('package_length_in') or 0))
            bucket['package_width_in'] = max(bucket['package_width_in'], float(row.get('package_width_in') or 0))
            bucket['package_height_in'] = max(bucket['package_height_in'], float(row.get('package_height_in') or 0))
            bucket['net_weight_lbs'] += float(row.get('net_weight_lbs') or 0) * qty
            bucket['gross_weight_lbs'] += float(row.get('gross_weight_lbs') or 0) * qty
            # 列表与成本汇总共用同一套 link 行；始终填充以便列表「下单SKU」列展示（无额外查询）
            bucket['order_sku_links'].append({
                'order_product_id': self._parse_int(row.get('order_product_id')),
                'sku': row.get('sku') or '',
                'quantity': qty
            })

        for variant_id in metrics.keys():
            bucket = metrics[variant_id]
            bucket['warehouse_cost_usd'] = round(bucket['warehouse_cost_usd'], 2)
            bucket['last_mile_cost_usd'] = round(bucket['last_mile_cost_usd'], 2)
            bucket['package_length_in'] = round(bucket['package_length_in'], 2)
            bucket['package_width_in'] = round(bucket['package_width_in'], 2)
            bucket['package_height_in'] = round(bucket['package_height_in'], 2)
            bucket['net_weight_lbs'] = round(bucket['net_weight_lbs'], 2)
            bucket['gross_weight_lbs'] = round(bucket['gross_weight_lbs'], 2)
        return metrics

    def _replace_sales_order_links(self, conn, sales_product_id, links):
        """兼容旧调用：按 sales_product_id 转换为 variant_id 后写入新表。"""
        if not sales_product_id:
            raise ValueError('Invalid sales_product_id')
        with conn.cursor() as cur:
            cur.execute("SELECT variant_id FROM sales_products WHERE id=%s LIMIT 1", (sales_product_id,))
            row = cur.fetchone() or {}
        variant_id = self._parse_int(row.get('variant_id'))
        if not variant_id:
            return
        self._replace_sales_variant_order_links(conn, variant_id, links)

    def _derive_sales_order_links_bundle(self, conn, sku_family_id_input, links):
        """对关联下单 SKU 做一次 order_products + fabric 查询，同时得到成本/尺寸汇总与自动面料、规格片段、平台 SKU 草稿。"""
        empty = {
            'warehouse_cost_usd': 0.0,
            'last_mile_cost_usd': 0.0,
            'package_length_in': 0.0,
            'package_width_in': 0.0,
            'package_height_in': 0.0,
            'net_weight_lbs': 0.0,
            'gross_weight_lbs': 0.0,
            'sku_family_id': None,
            'sku_family_code': '',
            'fabric': '',
            'spec_name': '',
            'platform_sku': '',
        }
        merged = self._normalize_sales_order_links(links)
        if not merged:
            return dict(empty)
        id_list = [entry['order_product_id'] for entry in merged]
        placeholders = ','.join(['%s'] * len(id_list))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT op.id, op.sku_family_id,
                       op.cost_usd, op.last_mile_avg_freight_usd,
                       op.package_length_in, op.package_width_in, op.package_height_in,
                       op.net_weight_lbs, op.gross_weight_lbs,
                       op.spec_qty_short,
                       fm.fabric_code, fm.fabric_name_en
                FROM order_products op
                LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                WHERE op.id IN ({placeholders})
                """,
                id_list,
            )
            rows = cur.fetchall() or []
        row_map = {row['id']: row for row in rows}

        warehouse_cost_usd = 0.0
        last_mile_cost_usd = 0.0
        package_length_in = 0.0
        package_width_in = 0.0
        package_height_in = 0.0
        net_weight_lbs = 0.0
        gross_weight_lbs = 0.0
        sku_family_id = self._parse_int(sku_family_id_input) or None

        for entry in merged:
            row = row_map.get(entry['order_product_id'])
            if not row:
                continue
            qty = max(1, int(entry.get('quantity') or 1))
            if sku_family_id is None:
                sku_family_id = self._parse_int(row.get('sku_family_id'))
            warehouse_cost_usd += float(row.get('cost_usd') or 0) * qty
            last_mile_cost_usd += float(row.get('last_mile_avg_freight_usd') or 0) * qty
            package_length_in = max(package_length_in, float(row.get('package_length_in') or 0))
            package_width_in = max(package_width_in, float(row.get('package_width_in') or 0))
            package_height_in = max(package_height_in, float(row.get('package_height_in') or 0))
            net_weight_lbs += float(row.get('net_weight_lbs') or 0) * qty
            gross_weight_lbs += float(row.get('gross_weight_lbs') or 0) * qty

        fabrics = []
        spec_parts = []
        for entry in merged:
            row = row_map.get(entry['order_product_id'])
            if not row:
                continue
            fabric_code = self._code_before_dash(row.get('fabric_code'))
            if not fabric_code:
                fabric_code = self._code_before_dash(row.get('fabric_name_en'))
            if fabric_code and fabric_code not in fabrics:
                fabrics.append(fabric_code)
            spec_short = (row.get('spec_qty_short') or '').strip()
            if spec_short:
                spec_parts.append(f"{entry['quantity']}{spec_short}")

        fabric = ' / '.join(fabrics)
        spec_name = ''.join(spec_parts)

        sku_family_code = ''
        if sku_family_id:
            with conn.cursor() as cur:
                cur.execute("SELECT sku_family FROM product_families WHERE id=%s", (sku_family_id,))
                sku_row = cur.fetchone()
                if sku_row:
                    sku_family_code = (sku_row.get('sku_family') or '').strip()

        platform_sku = ''
        if sku_family_code and fabric and spec_name:
            first_fabric = fabrics[0] if fabrics else ''
            platform_sku = self._build_sales_platform_sku(sku_family_code, spec_name, first_fabric)

        out = dict(empty)
        out.update({
            'warehouse_cost_usd': round(warehouse_cost_usd, 2),
            'last_mile_cost_usd': round(last_mile_cost_usd, 2),
            'package_length_in': round(package_length_in, 2),
            'package_width_in': round(package_width_in, 2),
            'package_height_in': round(package_height_in, 2),
            'net_weight_lbs': round(net_weight_lbs, 2),
            'gross_weight_lbs': round(gross_weight_lbs, 2),
            'sku_family_id': sku_family_id,
            'sku_family_code': sku_family_code,
            'fabric': fabric,
            'spec_name': spec_name,
            'platform_sku': platform_sku,
        })
        return out

    def _code_before_dash(self, value):
        text = (value or '').strip()
        if not text:
            return ''
        return text.split('-', 1)[0].strip() or text

    def _build_sales_platform_sku(self, sku_family_code, spec_name, fabric_code):
        sku_part = (sku_family_code or '').strip()
        spec_part = (spec_name or '').strip()
        fabric_part = self._code_before_dash(fabric_code)
        if not (sku_part and spec_part and fabric_part):
            return ''
        return f"{sku_part}-{spec_part}-{fabric_part}"

