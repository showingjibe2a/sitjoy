# -*- coding: utf-8 -*-
"""订单管理 Mixin - order_product 相关 API"""

import cgi
import io
import re
from urllib.parse import parse_qs

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception as _openpyxl_import_error:
    Workbook = None
    load_workbook = None
    PatternFill = None
    Font = None
    Alignment = None
    get_column_letter = None
    DataValidation = None

class OrderManagementMixin:
    """订单/配送管理 API 处理器"""

    def handle_order_product_api(self, environ, method, start_response):
        """下单产品管理 API - CRUD"""
        try:
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)
            action = (query_params.get('action', [''])[0] or '').strip().lower()

            if action == 'shipping_plans':
                return self._handle_order_product_shipping_plans(environ, method, start_response, query_params)

            if action == 'delete_impact':
                return self._handle_order_product_delete_impact(environ, method, start_response, query_params)

            if action == 'factory_links_template':
                return self._handle_order_product_factory_links_template(environ, method, start_response)

            if action == 'factory_links_import':
                return self._handle_order_product_factory_links_import(environ, method, start_response)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                exclude_reship_accessory = str((query_params.get('exclude_reship_accessory', ['0'])[0] or '0')).strip().lower() in ('1', 'true', 'yes', 'on')
                with self._get_db_connection() as conn:
                    item_id = self._parse_int((query_params.get('id', [''])[0] or '').strip())
                    if item_id:
                        rows = self._load_order_product_rows(conn, keyword=keyword, include_relations=False, limit_rows=1, item_id=item_id, exclude_reship_accessory=exclude_reship_accessory)
                        item = rows[0] if rows else None
                        if item:
                            self._attach_order_product_relations(conn, [item])
                            self._attach_order_product_factory_links(conn, [item])
                        return self.send_json({'status': 'success', 'item': item}, start_response)
                    rows = self._load_order_product_rows(conn, keyword=keyword, exclude_reship_accessory=exclude_reship_accessory)
                    self._attach_order_product_factory_links(conn, rows)
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                sku_family_id = self._parse_int(data.get('sku_family_id'))
                sku = (data.get('sku') or '').strip()
                version_no = (data.get('version_no') or '').strip()
                fabric_id = self._parse_int(data.get('fabric_id'))
                if not sku or not sku_family_id or not fabric_id:
                    return self.send_json({'status': 'error', 'message': 'Missing fields'}, start_response)

                filling_material_ids = self._normalize_id_list(data.get('filling_material_ids'))
                frame_material_ids = self._normalize_id_list(data.get('frame_material_ids'))
                feature_ids = self._normalize_id_list(data.get('feature_ids'))
                certification_ids = self._normalize_id_list(data.get('certification_ids'))
                factory_ids = self._normalize_id_list(data.get('factory_ids'))

                is_iteration = 1 if self._parse_int(data.get('is_iteration')) else 0
                is_on_market = 1 if self._parse_int(data.get('is_on_market')) else 0
                is_reship_accessory = 1 if self._parse_int(data.get('is_reship_accessory')) else 0
                is_dachene_product = 1 if self._parse_int(data.get('is_dachene_product')) else 0

                source_order_product_id = self._parse_int(data.get('source_order_product_id'))
                if not is_iteration:
                    source_order_product_id = None

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO order_products (
                                sku, sku_family_id, version_no, fabric_id,
                                spec_qty_short, contents_desc_en,
                                is_iteration, is_dachene_product, is_on_market, is_reship_accessory,
                                source_order_product_id,
                                finished_length_in, finished_width_in, finished_height_in,
                                net_weight_lbs,
                                package_length_in, package_width_in, package_height_in,
                                gross_weight_lbs,
                                cost_usd, carton_qty, package_size_class, last_mile_avg_freight_usd
                            ) VALUES (
                                %s, %s, %s, %s,
                                %s, %s,
                                %s, %s, %s, %s,
                                %s,
                                %s, %s, %s,
                                %s,
                                %s, %s, %s,
                                %s,
                                %s, %s, %s, %s
                            )
                            """,
                            (
                                sku,
                                sku_family_id,
                                version_no,
                                fabric_id,
                                (data.get('spec_qty_short') or '').strip(),
                                (data.get('contents_desc_en') or '').strip() or None,
                                is_iteration,
                                is_dachene_product,
                                is_on_market,
                                is_reship_accessory,
                                source_order_product_id,
                                self._parse_float(data.get('finished_length_in')),
                                self._parse_float(data.get('finished_width_in')),
                                self._parse_float(data.get('finished_height_in')),
                                self._parse_float(data.get('net_weight_lbs')),
                                self._parse_float(data.get('package_length_in')),
                                self._parse_float(data.get('package_width_in')),
                                self._parse_float(data.get('package_height_in')),
                                self._parse_float(data.get('gross_weight_lbs')),
                                self._parse_float(data.get('cost_usd')),
                                self._parse_int(data.get('carton_qty')),
                                (data.get('package_size_class') or '').strip() or None,
                                self._parse_float(data.get('last_mile_avg_freight_usd')),
                            )
                        )
                        new_id = cur.lastrowid
                    if is_iteration and source_order_product_id and version_no:
                        self._auto_sync_iteration_shipping_plans(
                            conn,
                            new_order_product_id=new_id,
                            source_order_product_id=source_order_product_id,
                            version_no=version_no
                        )
                    self._replace_order_product_relations(
                        conn,
                        new_id,
                        filling_material_ids,
                        frame_material_ids,
                        feature_ids,
                        certification_ids,
                    )
                    self._replace_order_product_factory_links(conn, new_id, factory_ids)
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                preview_action = (query_params.get('action', [''])[0] or '').strip().lower()
                if preview_action == 'preview_update':
                    batch_items = data.get('items') if isinstance(data, dict) else None
                    if not isinstance(batch_items, list) or not batch_items:
                        return self.send_json({'status': 'error', 'message': 'Missing preview items'}, start_response)

                    updates = []
                    for item in batch_items:
                        if not isinstance(item, dict):
                            continue
                        item_id = self._parse_int(item.get('id'))
                        if not item_id:
                            continue
                        updates.append((
                            self._parse_float(item.get('finished_length_in')),
                            self._parse_float(item.get('finished_width_in')),
                            self._parse_float(item.get('finished_height_in')),
                            self._parse_float(item.get('package_length_in')),
                            self._parse_float(item.get('package_width_in')),
                            self._parse_float(item.get('package_height_in')),
                            self._parse_float(item.get('cost_usd')),
                            (item.get('package_size_class') or '').strip() or None,
                            self._parse_int(item.get('carton_qty')),
                            self._parse_float(item.get('last_mile_avg_freight_usd')),
                            1 if self._parse_int(item.get('is_on_market')) else 0,
                            1 if self._parse_int(item.get('is_reship_accessory')) else 0,
                            item_id
                        ))

                    if not updates:
                        return self.send_json({'status': 'error', 'message': 'No valid preview items'}, start_response)

                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            row_map = {}
                            for finished_length_in, finished_width_in, finished_height_in, package_length_in, package_width_in, package_height_in, cost_usd, package_size_class, carton_qty, last_mile_avg_freight_usd, is_on_market, is_reship_accessory, item_id in updates:
                                row_map[int(item_id)] = {
                                    'finished_length_in': finished_length_in,
                                    'finished_width_in': finished_width_in,
                                    'finished_height_in': finished_height_in,
                                    'package_length_in': package_length_in,
                                    'package_width_in': package_width_in,
                                    'package_height_in': package_height_in,
                                    'cost_usd': cost_usd,
                                    'package_size_class': package_size_class,
                                    'carton_qty': carton_qty,
                                    'last_mile_avg_freight_usd': last_mile_avg_freight_usd,
                                    'is_on_market': is_on_market,
                                    'is_reship_accessory': is_reship_accessory,
                                }

                            ids = list(row_map.keys())
                            case_params = []

                            def build_case(field_name):
                                parts = []
                                for rid in ids:
                                    parts.append('WHEN %s THEN %s')
                                    case_params.extend([rid, row_map[rid][field_name]])
                                return f"CASE id {' '.join(parts)} ELSE {field_name} END"

                            set_clause = [
                                f"finished_length_in = {build_case('finished_length_in')}",
                                f"finished_width_in = {build_case('finished_width_in')}",
                                f"finished_height_in = {build_case('finished_height_in')}",
                                f"package_length_in = {build_case('package_length_in')}",
                                f"package_width_in = {build_case('package_width_in')}",
                                f"package_height_in = {build_case('package_height_in')}",
                                f"cost_usd = {build_case('cost_usd')}",
                                f"package_size_class = {build_case('package_size_class')}",
                                f"carton_qty = {build_case('carton_qty')}",
                                f"last_mile_avg_freight_usd = {build_case('last_mile_avg_freight_usd')}",
                                f"is_on_market = {build_case('is_on_market')}",
                                f"is_reship_accessory = {build_case('is_reship_accessory')}",
                            ]
                            where_placeholders = ','.join(['%s'] * len(ids))
                            sql = f"UPDATE order_products SET {', '.join(set_clause)} WHERE id IN ({where_placeholders})"
                            cur.execute(sql, tuple(case_params + ids))
                    return self.send_json({'status': 'success', 'updated': len(updates)}, start_response)

                batch_items = data.get('items') if isinstance(data, dict) else None
                if isinstance(batch_items, list):
                    updates = []
                    for item in batch_items:
                        if not isinstance(item, dict):
                            continue
                        item_id = self._parse_int(item.get('id'))
                        if not item_id:
                            continue
                        updates.append((
                            self._parse_float(item.get('cost_usd')),
                            (item.get('package_size_class') or '').strip() or None,
                            self._parse_int(item.get('carton_qty')),
                            self._parse_float(item.get('last_mile_avg_freight_usd')),
                            item_id
                        ))

                    if not updates:
                        return self.send_json({'status': 'error', 'message': 'Missing fields'}, start_response)

                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.executemany(
                                """
                                UPDATE order_products
                                SET cost_usd=%s,
                                    package_size_class=%s,
                                    carton_qty=%s,
                                    last_mile_avg_freight_usd=%s
                                WHERE id=%s
                                """,
                                updates
                            )
                    return self.send_json({'status': 'success', 'updated': len(updates)}, start_response)

                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)

                updates = []
                params = []

                text_fields = ['sku', 'version_no', 'spec_qty_short', 'contents_desc_en', 'package_size_class']
                for field in text_fields:
                    if field in data:
                        val = (data.get(field) or '').strip()
                        if field in ('contents_desc_en', 'package_size_class'):
                            val = val or None
                        updates.append(f"{field}=%s")
                        params.append(val)

                int_fields = ['sku_family_id', 'fabric_id', 'source_order_product_id', 'carton_qty']
                for field in int_fields:
                    if field in data:
                        updates.append(f"{field}=%s")
                        params.append(self._parse_int(data.get(field)))

                float_fields = [
                    'finished_length_in', 'finished_width_in', 'finished_height_in', 'net_weight_lbs',
                    'package_length_in', 'package_width_in', 'package_height_in', 'gross_weight_lbs',
                    'cost_usd', 'last_mile_avg_freight_usd'
                ]
                for field in float_fields:
                    if field in data:
                        updates.append(f"{field}=%s")
                        params.append(self._parse_float(data.get(field)))

                bool_fields = ['is_iteration', 'is_dachene_product', 'is_on_market', 'is_reship_accessory']
                for field in bool_fields:
                    if field in data:
                        updates.append(f"{field}=%s")
                        params.append(1 if self._parse_int(data.get(field)) else 0)

                if 'is_iteration' in data and not (1 if self._parse_int(data.get('is_iteration')) else 0):
                    updates.append("source_order_product_id=%s")
                    params.append(None)

                has_relation_updates = any(
                    key in data for key in ('filling_material_ids', 'frame_material_ids', 'feature_ids', 'certification_ids', 'factory_ids')
                )

                with self._get_db_connection() as conn:
                    if updates:
                        with conn.cursor() as cur:
                            cur.execute(
                                f"UPDATE order_products SET {', '.join(updates)} WHERE id=%s",
                                tuple(params + [item_id])
                            )

                    if has_relation_updates:
                        filling_material_ids = self._normalize_id_list(data.get('filling_material_ids'))
                        frame_material_ids = self._normalize_id_list(data.get('frame_material_ids'))
                        feature_ids = self._normalize_id_list(data.get('feature_ids'))
                        certification_ids = self._normalize_id_list(data.get('certification_ids'))
                        self._replace_order_product_relations(
                            conn,
                            item_id,
                            filling_material_ids,
                            frame_material_ids,
                            feature_ids,
                            certification_ids,
                        )
                        if 'factory_ids' in data:
                            factory_ids = self._normalize_id_list(data.get('factory_ids'))
                            self._replace_order_product_factory_links(conn, item_id, factory_ids)
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM order_product_shipping_plan_items WHERE substitute_order_product_id=%s", (item_id,))
                        ref_deleted = cur.rowcount or 0
                        cur.execute("DELETE FROM order_product_shipping_plans WHERE order_product_id=%s", (item_id,))
                        plan_deleted = cur.rowcount or 0
                        cur.execute("DELETE FROM order_products WHERE id=%s", (item_id,))
                        sku_deleted = cur.rowcount or 0
                return self.send_json({
                    'status': 'success',
                    'deleted': {
                        'order_product': sku_deleted,
                        'shipping_plans': plan_deleted,
                        'substitute_references': ref_deleted
                    }
                }, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            print(f"Order Product API error: {str(e)}")
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _handle_order_product_delete_impact(self, environ, method, start_response, query_params):
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)

            target = (query_params.get('target', [''])[0] or '').strip().lower()
            target_id = self._parse_int(query_params.get('id', [''])[0])
            if not target or not target_id:
                return self.send_json({'status': 'error', 'message': 'Missing target or id'}, start_response)

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    if target == 'shipping_plan':
                        cur.execute(
                            """
                            SELECT
                                ops.id,
                                ops.plan_name,
                                ops.order_product_id,
                                op.sku AS owner_sku
                            FROM order_product_shipping_plans ops
                            JOIN order_products op ON op.id = ops.order_product_id
                            WHERE ops.id=%s
                            LIMIT 1
                            """,
                            (target_id,)
                        )
                        plan_row = cur.fetchone() or {}
                        if not plan_row:
                            return self.send_json({'status': 'error', 'message': 'Plan not found'}, start_response)

                        cur.execute(
                            """
                            SELECT op.sku AS substitute_order_sku, opsi.quantity
                            FROM order_product_shipping_plan_items opsi
                            JOIN order_products op ON op.id = opsi.substitute_order_product_id
                            WHERE opsi.shipping_plan_id=%s
                            ORDER BY opsi.sort_order ASC, opsi.id ASC
                            """,
                            (target_id,)
                        )
                        items = cur.fetchall() or []
                        return self.send_json({
                            'status': 'success',
                            'impact': {
                                'target': 'shipping_plan',
                                'id': target_id,
                                'plan_name': plan_row.get('plan_name') or '',
                                'owner_sku': plan_row.get('owner_sku') or '',
                                'item_count': len(items),
                                'items': items
                            }
                        }, start_response)

                    if target == 'order_product':
                        cur.execute("SELECT id, sku FROM order_products WHERE id=%s LIMIT 1", (target_id,))
                        product_row = cur.fetchone() or {}
                        if not product_row:
                            return self.send_json({'status': 'error', 'message': 'SKU not found'}, start_response)

                        cur.execute(
                            "SELECT id, plan_name FROM order_product_shipping_plans WHERE order_product_id=%s ORDER BY id DESC LIMIT 10",
                            (target_id,)
                        )
                        own_plans = cur.fetchall() or []

                        cur.execute(
                            """
                            SELECT COUNT(*) AS cnt
                            FROM order_product_shipping_plan_items opsi
                            JOIN order_product_shipping_plans ops ON ops.id = opsi.shipping_plan_id
                            WHERE ops.order_product_id=%s
                            """,
                            (target_id,)
                        )
                        own_plan_item_count = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0

                        cur.execute(
                            """
                            SELECT COUNT(*) AS cnt
                            FROM order_product_shipping_plan_items
                            WHERE substitute_order_product_id=%s
                            """,
                            (target_id,)
                        )
                        referenced_count = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0

                        cur.execute(
                            """
                            SELECT
                                owner_op.sku AS owner_sku,
                                ops.plan_name
                            FROM order_product_shipping_plan_items opsi
                            JOIN order_product_shipping_plans ops ON ops.id = opsi.shipping_plan_id
                            JOIN order_products owner_op ON owner_op.id = ops.order_product_id
                            WHERE opsi.substitute_order_product_id=%s
                            ORDER BY ops.id DESC
                            LIMIT 10
                            """,
                            (target_id,)
                        )
                        referenced_in = cur.fetchall() or []

                        return self.send_json({
                            'status': 'success',
                            'impact': {
                                'target': 'order_product',
                                'id': target_id,
                                'sku': product_row.get('sku') or '',
                                'own_plan_count': len(own_plans),
                                'own_plan_item_count': own_plan_item_count,
                                'own_plans': own_plans,
                                'referenced_count': referenced_count,
                                'referenced_in': referenced_in
                            }
                        }, start_response)

            return self.send_json({'status': 'error', 'message': 'Unsupported target'}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_order_product_carton_calc_api(self, environ, method, start_response):
        """纸箱数量计算 API"""
        try:
            if method != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)

            data = self._read_json_body(environ)
            length = self._parse_float(data.get('length'))
            width = self._parse_float(data.get('width'))
            height = self._parse_float(data.get('height'))

            if not all([length, width, height]):
                return self.send_json({'status': 'error', 'message': 'Missing dimensions'}, start_response)

            qty = self._calc_carton_qty_by_40hq(length, width, height)
            return self.send_json({'status': 'success', 'carton_qty': qty}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)


    def handle_order_product_template_api(self, environ, method, start_response):
        """下单产品模板下载"""
        try:
            if method not in ('GET', 'POST'):
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation

            def _append_ids_from_value(container, value):
                if isinstance(value, list):
                    for v in value:
                        _append_ids_from_value(container, v)
                    return
                text = str(value or '').strip()
                if not text:
                    return
                for token in re.split(r'[,，;；\s]+', text):
                    if not token:
                        continue
                    item_id = self._parse_int(token)
                    if item_id and item_id not in container:
                        container.append(item_id)

            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            selected_ids = []
            for raw in query_params.get('ids', []):
                _append_ids_from_value(selected_ids, raw)

            if method == 'POST':
                body = self._read_json_body(environ) or {}
                _append_ids_from_value(selected_ids, body.get('ids'))

            wb = Workbook()
            ws = wb.active
            ws.title = 'order_products'

            max_multi_columns = {
                'filling_materials': 3,
                'frame_materials': 3,
                'features': 3,
                'certifications': 3,
                'factories': 3,
            }
            export_rows = []
            filling_rel = {}
            frame_rel = {}
            feature_rel = {}
            cert_rel = {}
            factory_rel = {}

            # 获取所有可用的数据用于下拉菜单 + 勾选导出数据
            with self._get_db_connection() as conn:
                def _load_order_template_options():
                    with conn.cursor() as cur:
                        cur.execute("SELECT sku_family FROM product_families ORDER BY sku_family")
                        sku_families_local = [row['sku_family'] for row in cur.fetchall()]

                        cur.execute("SELECT fabric_code FROM fabric_materials ORDER BY fabric_code")
                        fabrics_local = [row['fabric_code'] for row in cur.fetchall()]

                        cur.execute("""
                            SELECT m.name
                            FROM materials m
                            JOIN material_types mt ON m.material_type_id = mt.id
                            WHERE mt.name = '填充'
                            ORDER BY m.name
                        """)
                        filling_local = [row['name'] for row in cur.fetchall()]

                        cur.execute("""
                            SELECT m.name
                            FROM materials m
                            JOIN material_types mt ON m.material_type_id = mt.id
                            WHERE mt.name = '框架'
                            ORDER BY m.name
                        """)
                        frame_local = [row['name'] for row in cur.fetchall()]

                        cur.execute("SELECT name FROM features ORDER BY name")
                        feature_local = [row['name'] for row in cur.fetchall()]

                        cur.execute("SELECT name FROM certifications ORDER BY name")
                        cert_local = [row['name'] for row in cur.fetchall()]
                        cur.execute("SELECT factory_name FROM logistics_factories ORDER BY factory_name")
                        factory_local = [row['factory_name'] for row in cur.fetchall()]
                    return (sku_families_local, fabrics_local, filling_local, frame_local, feature_local, cert_local, factory_local)

                sku_families, fabrics, filling_materials, frame_materials, features, certifications, factories = self._get_cached_template_options(
                    'order_product_template_options_v1',
                    _load_order_template_options,
                    ttl_seconds=180
                )

                with conn.cursor() as cur:

                    if selected_ids:
                        placeholders = ','.join(['%s'] * len(selected_ids))
                        cur.execute(
                            f"""
                            SELECT
                                op.id,
                                op.sku,
                                pf.sku_family,
                                op.version_no,
                                fm.fabric_code,
                                op.spec_qty_short,
                                op.contents_desc_en,
                                op.is_on_market,
                                op.is_reship_accessory,
                                op.is_iteration,
                                op.is_dachene_product,
                                src.sku AS source_sku,
                                op.finished_length_in,
                                op.finished_width_in,
                                op.finished_height_in,
                                op.net_weight_lbs,
                                op.package_length_in,
                                op.package_width_in,
                                op.package_height_in,
                                op.gross_weight_lbs,
                                op.cost_usd,
                                op.carton_qty,
                                op.package_size_class,
                                op.last_mile_avg_freight_usd
                            FROM order_products op
                            LEFT JOIN product_families pf ON op.sku_family_id = pf.id
                            LEFT JOIN fabric_materials fm ON op.fabric_id = fm.id
                            LEFT JOIN order_products src ON op.source_order_product_id = src.id
                            WHERE op.id IN ({placeholders})
                            """,
                            tuple(selected_ids)
                        )
                        selected_rows = cur.fetchall() or []
                        order_map = {sid: idx for idx, sid in enumerate(selected_ids)}
                        selected_rows.sort(key=lambda x: order_map.get(x.get('id'), 10 ** 6))
                        export_rows = selected_rows

                        if selected_rows:
                            row_ids = [row['id'] for row in selected_rows]
                            rel_placeholders = ','.join(['%s'] * len(row_ids))

                            cur.execute(
                                f"""
                                SELECT opm.order_product_id, m.name, mt.name AS type_name
                                FROM order_product_materials opm
                                JOIN materials m ON opm.material_id = m.id
                                JOIN material_types mt ON m.material_type_id = mt.id
                                WHERE opm.order_product_id IN ({rel_placeholders})
                                ORDER BY opm.order_product_id, m.name
                                """,
                                tuple(row_ids)
                            )
                            for rel in cur.fetchall() or []:
                                target = filling_rel if rel.get('type_name') == '填充' else frame_rel
                                target.setdefault(rel['order_product_id'], [])
                                if rel['name'] not in target[rel['order_product_id']]:
                                    target[rel['order_product_id']].append(rel['name'])

                            cur.execute(
                                f"""
                                SELECT opf.order_product_id, f.name
                                FROM order_product_features opf
                                JOIN features f ON opf.feature_id = f.id
                                WHERE opf.order_product_id IN ({rel_placeholders})
                                ORDER BY opf.order_product_id, f.name
                                """,
                                tuple(row_ids)
                            )
                            for rel in cur.fetchall() or []:
                                feature_rel.setdefault(rel['order_product_id'], [])
                                if rel['name'] not in feature_rel[rel['order_product_id']]:
                                    feature_rel[rel['order_product_id']].append(rel['name'])

                            cur.execute(
                                f"""
                                SELECT opc.order_product_id, c.name
                                FROM order_product_certifications opc
                                JOIN certifications c ON opc.certification_id = c.id
                                WHERE opc.order_product_id IN ({rel_placeholders})
                                ORDER BY opc.order_product_id, c.name
                                """,
                                tuple(row_ids)
                            )
                            for rel in cur.fetchall() or []:
                                cert_rel.setdefault(rel['order_product_id'], [])
                                if rel['name'] not in cert_rel[rel['order_product_id']]:
                                    cert_rel[rel['order_product_id']].append(rel['name'])

                            cur.execute(
                                f"""
                                SELECT opl.order_product_id, lf.factory_name
                                FROM order_product_factory_links opl
                                JOIN logistics_factories lf ON lf.id = opl.factory_id
                                WHERE opl.order_product_id IN ({rel_placeholders})
                                ORDER BY opl.order_product_id, lf.factory_name
                                """,
                                tuple(row_ids)
                            )
                            for rel in cur.fetchall() or []:
                                factory_rel.setdefault(rel['order_product_id'], [])
                                if rel['factory_name'] not in factory_rel[rel['order_product_id']]:
                                    factory_rel[rel['order_product_id']].append(rel['factory_name'])

                            for row in selected_rows:
                                rid = row['id']
                                max_multi_columns['filling_materials'] = max(max_multi_columns['filling_materials'], len(filling_rel.get(rid, [])))
                                max_multi_columns['frame_materials'] = max(max_multi_columns['frame_materials'], len(frame_rel.get(rid, [])))
                                max_multi_columns['features'] = max(max_multi_columns['features'], len(feature_rel.get(rid, [])))
                                max_multi_columns['certifications'] = max(max_multi_columns['certifications'], len(cert_rel.get(rid, [])))
                                max_multi_columns['factories'] = max(max_multi_columns['factories'], len(factory_rel.get(rid, [])))
            
            # 定义组件和字段（带中文标签）
            sections = [
                {
                    'title': '迭代款',
                    'bg_color': 'E8DFD4',
                    'fields': [
                        ('is_iteration', '是否迭代款', 'dropdown', ['否', '是']),
                        ('is_dachene_product', '是否为大健云仓产品（在下单SKU处填写大健云仓Item Code）', 'dropdown', ['否', '是']),
                        ('source_sku', '来源下单SKU', 'text', None),
                        ('version_no', '版本号', 'text', None)
                    ]
                },
                {
                    'title': '基础信息',
                    'bg_color': 'F5F1ED',
                    'fields': [
                        ('is_on_market', '是否在市', 'dropdown', ['否', '是']),
                        ('is_reship_accessory', '是否为补发用配件', 'dropdown', ['否', '是']),
                        ('sku', '下单SKU *', 'text', None),
                        ('sku_family', '归属货号 *', 'dropdown', sku_families),
                        ('fabric_code', '面料 *', 'dropdown', fabrics),
                        ('spec_qty_short', '规格与数量简称', 'text', None),
                        ('contents_desc_en', '内含物英文描述', 'text', None)
                    ]
                },
                {
                    'title': '成品尺寸/重量',
                    'bg_color': 'E8DFD4',
                    'fields': [
                        ('finished_length_in', '成品长(inch)', 'number', None),
                        ('finished_width_in', '成品宽(inch)', 'number', None),
                        ('finished_height_in', '成品高(inch)', 'number', None),
                        ('net_weight_lbs', '净重(lbs)', 'number', None)
                    ]
                },
                {
                    'title': '包裹尺寸/重量',
                    'bg_color': 'F5F1ED',
                    'fields': [
                        ('package_length_in', '包裹长(inch)', 'number', None),
                        ('package_width_in', '包裹宽(inch)', 'number', None),
                        ('package_height_in', '包裹高(inch)', 'number', None),
                        ('gross_weight_lbs', '毛重(lbs)', 'number', None),
                        ('carton_qty', '装箱量', 'number', None),
                        ('package_size_class', '包裹大小归类(Fedx)', 'text', None)
                    ]
                },
                {
                    'title': '成本',
                    'bg_color': 'E8DFD4',
                    'fields': [
                        ('cost_usd', '产品成本及发货至海外仓成本估算(USD，不含仓储费)', 'number', None),
                        ('last_mile_avg_freight_usd', '尾程平均运费(美元)', 'number', None)
                    ]
                },
                {
                    'title': '材料与卖点',
                    'bg_color': 'F5F1ED',
                    'fields': [
                        ('filling_materials', '填充材料(可多项)', 'multi_dropdown', filling_materials),
                        ('frame_materials', '框架材料(可多项)', 'multi_dropdown', frame_materials),
                        ('features', '卖点特点(可多项)', 'multi_dropdown', features),
                        ('certifications', '认证(可多项)', 'multi_dropdown', certifications),
                        ('factories', '工厂(可多项)', 'multi_dropdown', factories)
                    ]
                }
            ]
            
            # 建立模块标题行和列名行
            section_headers = []  # 模块名称行
            column_headers = []   # 列名行
            header_to_column = {}  # 用于数据验证时查找列
            col_idx = 0
            field_to_options = {}  # 记录字段对应的可选项
            
            for section in sections:
                section_title = section['title']
                section_start_col = col_idx
                
                for field_info in section['fields']:
                    field_code = field_info[0]
                    field_label = field_info[1]
                    field_type = field_info[2]
                    field_options = field_info[3] if len(field_info) > 3 else None
                    
                    if field_type == 'multi_dropdown':
                        num_cols = max_multi_columns.get(field_code, 3)
                        for i in range(1, num_cols + 1):
                            col_name = f"{field_code}_{i}"
                            column_headers.append(field_label if i == 1 else '')
                            header_to_column[col_name] = col_idx
                            field_to_options[col_name] = field_options
                            col_idx += 1
                    else:
                        column_headers.append(field_label)
                        header_to_column[field_code] = col_idx
                        if field_options:
                            field_to_options[field_code] = field_options
                        col_idx += 1
                
                # 填充模块标题（需要合并的列数）
                section_span = col_idx - section_start_col
                section_headers.append((section_title, section_start_col, section_span, section.get('bg_color') or 'CFC7BD'))
            
            # 第1行：模块标题（合并单元格）
            for i in range(col_idx):
                ws.cell(row=1, column=i+1).value = ''  # 先填充空值
            
            title_font = Font(bold=True, color='2A2420', size=11)
            title_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='B7AEA4'),
                right=Side(style='thin', color='B7AEA4'),
                top=Side(style='thin', color='B7AEA4'),
                bottom=Side(style='thin', color='B7AEA4')
            )
            
            for title, start_col, span, bg_color in section_headers:
                if span > 1:
                    ws.merge_cells(start_row=1, start_column=start_col+1, end_row=1, end_column=start_col+span)
                ws.cell(row=1, column=start_col+1).value = title
                title_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                for col in range(start_col, start_col + span):
                    ws.cell(row=1, column=col+1).fill = title_fill
                    ws.cell(row=1, column=col+1).font = title_font
                    ws.cell(row=1, column=col+1).alignment = title_alignment
                    ws.cell(row=1, column=col+1).border = thin_border
            
            # 第2行：列名
            for idx, header in enumerate(column_headers):
                cell = ws.cell(row=2, column=idx+1)
                cell.value = header
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # 第3行：示例行
            example_row_idx = 3
            example_row_data = []
            
            for col_name in list(header_to_column.keys()):
                field_base = col_name.rsplit('_', 1)[0] if '_' in col_name else col_name
                
                if field_base == 'is_iteration':
                    example_row_data.append(('是否迭代款', 0, '否'))
                elif field_base == 'is_dachene_product':
                    example_row_data.append(('是否为大健云仓产品（在下单SKU处填写大健云仓Item Code）', 0, '否'))
                elif field_base == 'sku':
                    example_row_data.append(('下单SKU', 0, 'MS01A-Brown'))
                elif field_base == 'is_on_market':
                    example_row_data.append(('是否在市', 0, '是'))
                elif field_base == 'is_reship_accessory':
                    example_row_data.append(('是否为补发用配件', 0, '否'))
                elif field_base == 'sku_family':
                    example_row_data.append(('归属货号', 0, 'MS01'))
                elif field_base == 'fabric_code':
                    example_row_data.append(('面料', 0, 'Brown'))
                elif field_base == 'spec_qty_short':
                    example_row_data.append(('规格与数量简称', 0, 'A'))
                elif field_base == 'contents_desc_en':
                    example_row_data.append(('内含物英文描述', 0, 'memory foam + metal frame'))
                elif field_base == 'version_no':
                    example_row_data.append(('版本号', 0, '1'))
                elif field_base == 'source_sku':
                    example_row_data.append(('来源下单SKU', 0, ''))
                elif field_base == 'finished_length_in':
                    example_row_data.append(('成品长(inch)', 0, 30))
                elif field_base == 'finished_width_in':
                    example_row_data.append(('成品宽(inch)', 0, 20))
                elif field_base == 'finished_height_in':
                    example_row_data.append(('成品高(inch)', 0, 10))
                elif field_base == 'net_weight_lbs':
                    example_row_data.append(('净重(lbs)', 0, 5.5))
                elif field_base == 'package_length_in':
                    example_row_data.append(('包裹长(inch)', 0, 32))
                elif field_base == 'package_width_in':
                    example_row_data.append(('包裹宽(inch)', 0, 22))
                elif field_base == 'package_height_in':
                    example_row_data.append(('包裹高(inch)', 0, 12))
                elif field_base == 'gross_weight_lbs':
                    example_row_data.append(('毛重(lbs)', 0, 6.5))
                elif field_base == 'cost_usd':
                    example_row_data.append(('产品成本及发货至海外仓成本估算(USD，不含仓储费)', 0, 25.00))
                elif field_base == 'carton_qty':
                    example_row_data.append(('装箱量', 0, 50))
                elif field_base == 'package_size_class':
                    example_row_data.append(('包裹大小归类(Fedx)', 0, 'Small'))
                elif field_base == 'last_mile_avg_freight_usd':
                    example_row_data.append(('尾程平均运费(美元)', 0, 3.50))
                elif field_base in ['filling_materials', 'frame_materials', 'features', 'certifications', 'factories']:
                    # 多选字段只在第一列填充示例
                    if col_name.endswith('_1'):
                        if field_base == 'filling_materials':
                            example_row_data.append(('填充材料(可多项)', 0, '海绵'))
                        elif field_base == 'frame_materials':
                            example_row_data.append(('框架材料(可多项)', 0, '金属'))
                        elif field_base == 'features':
                            example_row_data.append(('卖点特点(可多项)', 0, '可拆洗'))
                        elif field_base == 'certifications':
                            example_row_data.append(('认证(可多项)', 0, 'CE'))
                        elif field_base == 'factories':
                            example_row_data.append(('工厂(可多项)', 0, '示例工厂A'))
                    else:
                        example_row_data.append(('', 0, None))
                else:
                    example_row_data.append(('', 0, None))
            
            for idx, (label, unused, value) in enumerate(example_row_data):
                cell = ws.cell(row=example_row_idx, column=idx+1)
                cell.value = value
                cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                cell.font = Font(italic=True, color='888888')
            
            # 辅助函数：将列索引转换为Excel列字母
            def col_idx_to_letter(idx):
                """将0-based列索引转换为Excel列字母"""
                result = ''
                while idx >= 0:
                    result = chr(65 + (idx % 26)) + result
                    idx = idx // 26 - 1
                return result
            
            # 添加数据验证
            yes_no_validation = DataValidation(type='list', formula1='"否,是"', allow_blank=True)
            ws.add_data_validation(yes_no_validation)
            max_validation_row = 400
            
            for bool_field in ('is_iteration', 'is_dachene_product', 'is_on_market', 'is_reship_accessory'):
                if bool_field in header_to_column:
                    col_letter = col_idx_to_letter(header_to_column[bool_field])
                    for row in range(4, max_validation_row + 1):
                        yes_no_validation.add(f'{col_letter}{row}')
            
            # 为下拉字段添加验证
            for field_name, options in field_to_options.items():
                if options and field_name in header_to_column:
                    col_idx = header_to_column[field_name]
                    col_letter = col_idx_to_letter(col_idx)
                    
                    validation = DataValidation(type='list', formula1=f'"{",".join(options)}"', allow_blank=True)
                    ws.add_data_validation(validation)

                    for row in range(4, max_validation_row + 1):
                        validation.add(f'{col_letter}{row}')

            if export_rows:
                def set_multi_values(row_idx, field_name, values):
                    values = values or []
                    total = max_multi_columns.get(field_name, 3)
                    for i in range(1, total + 1):
                        key = f'{field_name}_{i}'
                        if key not in header_to_column:
                            continue
                        value = values[i - 1] if i - 1 < len(values) else None
                        ws.cell(row=row_idx, column=header_to_column[key] + 1).value = value

                data_row = 4
                for item in export_rows:
                    row_id = item.get('id')
                    direct_values = {
                        'is_iteration': '是' if str(item.get('is_iteration') or '0') in ('1', 'True', 'true') else '否',
                        'is_dachene_product': '是' if str(item.get('is_dachene_product') or '0') in ('1', 'True', 'true') else '否',
                        'is_on_market': '是' if str(item.get('is_on_market') or '0') in ('1', 'True', 'true') else '否',
                        'is_reship_accessory': '是' if str(item.get('is_reship_accessory') or '0') in ('1', 'True', 'true') else '否',
                        'source_sku': item.get('source_sku') or '',
                        'version_no': item.get('version_no') or '',
                        'sku': item.get('sku') or '',
                        'sku_family': item.get('sku_family') or '',
                        'fabric_code': item.get('fabric_code') or '',
                        'spec_qty_short': item.get('spec_qty_short') or '',
                        'contents_desc_en': item.get('contents_desc_en') or '',
                        'finished_length_in': item.get('finished_length_in'),
                        'finished_width_in': item.get('finished_width_in'),
                        'finished_height_in': item.get('finished_height_in'),
                        'net_weight_lbs': item.get('net_weight_lbs'),
                        'package_length_in': item.get('package_length_in'),
                        'package_width_in': item.get('package_width_in'),
                        'package_height_in': item.get('package_height_in'),
                        'gross_weight_lbs': item.get('gross_weight_lbs'),
                        'cost_usd': item.get('cost_usd'),
                        'carton_qty': item.get('carton_qty'),
                        'package_size_class': item.get('package_size_class') or '',
                        'last_mile_avg_freight_usd': item.get('last_mile_avg_freight_usd'),
                    }
                    for field_name, value in direct_values.items():
                        if field_name not in header_to_column:
                            continue
                        ws.cell(row=data_row, column=header_to_column[field_name] + 1).value = value

                    set_multi_values(data_row, 'filling_materials', filling_rel.get(row_id, []))
                    set_multi_values(data_row, 'frame_materials', frame_rel.get(row_id, []))
                    set_multi_values(data_row, 'features', feature_rel.get(row_id, []))
                    set_multi_values(data_row, 'certifications', cert_rel.get(row_id, []))
                    set_multi_values(data_row, 'factories', factory_rel.get(row_id, []))
                    data_row += 1
            
            # 设置列宽
            for idx, header in enumerate(column_headers):
                col_letter = col_idx_to_letter(idx)
                if '材料' in header or '特点' in header or '认证' in header or '工厂' in header:
                    ws.column_dimensions[col_letter].width = 18
                elif 'SKU' in header:
                    ws.column_dimensions[col_letter].width = 15
                elif '简称' in header:
                    ws.column_dimensions[col_letter].width = 12
                else:
                    ws.column_dimensions[col_letter].width = 14
            
            # 冻结表头
            ws.freeze_panes = 'A4'
            
            return self._send_excel_workbook(wb, 'order_product_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_order_product_import_api(self, environ, method, start_response):
        """下单产品批量导入"""
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            preview_mode = str((query_params.get('preview', ['0'])[0] or '0')).lower() in ('1', 'true', 'yes', 'on')

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            file_bytes = self._sanitize_xlsx_bool_cells(file_bytes)

            # load and sanitize workbook
            try:
                wb = load_workbook(io.BytesIO(file_bytes))
            except Exception as e:
                if 'Cannot be converted to bool' in str(e):
                    wb = self._rebuild_workbook_from_xlsx_xml(file_bytes)
                    if wb is None:
                        diag = self._scan_xlsx_invalid_bool_cells(file_bytes)
                        return self.send_json({
                            'status': 'error',
                            'message': (
                                '导入失败：文件中存在异常布尔字段且无法自动修复，'
                                '请另存为新的xlsx后重试'
                            ),
                            'debug': {
                                'cause': 'Cannot be converted to bool',
                                'invalid_bool_cells': diag.get('count', 0),
                                'samples': diag.get('samples', [])
                            }
                        }, start_response)
                else:
                    return self.send_json({'status': 'error', 'message': str(e)}, start_response)

            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'b' and not isinstance(cell.value, bool):
                            cell.data_type = 's'
                            cell.value = str(cell.value)

            ws = wb.active

            # 支持两种表头格式：新的中文表头（从第2行）或旧的字段代码表头（从第1行）
            header_row_idx = 2 if ws.cell(row=1, column=1).value in ['迭代款', '基础信息', '成品尺寸/重量', '包裹尺寸/重量', '成本与物流', '成本', '材料与卖点'] else 1
            
            headers = [cell.value for cell in ws[header_row_idx]]
            
            # 中文字段标签到字段代码的映射
            label_to_code = {
                '是否迭代款': 'is_iteration',
                '是否为大健云仓产品（在下单SKU处填写大健云仓Item Code）': 'is_dachene_product',
                '是否在市': 'is_on_market',
                '是否为补发用配件': 'is_reship_accessory',
                '来源下单SKU': 'source_sku',
                '版本号': 'version_no',
                '下单SKU *': 'sku',
                '归属货号 *': 'sku_family',
                '面料 *': 'fabric_code',
                '规格与数量简称': 'spec_qty_short',
                '内含物英文描述': 'contents_desc_en',
                '成品长(inch)': 'finished_length_in',
                '成品宽(inch)': 'finished_width_in',
                '成品高(inch)': 'finished_height_in',
                '净重(lbs)': 'net_weight_lbs',
                '包裹长(inch)': 'package_length_in',
                '包裹宽(inch)': 'package_width_in',
                '包裹高(inch)': 'package_height_in',
                '毛重(lbs)': 'gross_weight_lbs',
                '成本价(美元)': 'cost_usd',
                '产品成本及发货至海外仓成本估算(USD，不含仓储费)': 'cost_usd',
                '装箱量': 'carton_qty',
                '包裹大小归类(Fedx)': 'package_size_class',
                '尾程平均运费(美元)': 'last_mile_avg_freight_usd',
                '填充材料(可多项)': 'filling_materials',
                '框架材料(可多项)': 'frame_materials',
                '卖点特点(可多项)': 'features',
                '认证(可多项)': 'certifications',
                '工厂(可多项)': 'factories'
            }

            multi_base_fields = {'filling_materials', 'frame_materials', 'features', 'certifications', 'factories'}
            single_fields = {
                'sku', 'sku_family', 'fabric_code', 'spec_qty_short', 'is_iteration', 'is_dachene_product', 'is_on_market', 'is_reship_accessory', 'source_sku', 'version_no',
                'contents_desc_en',
                'finished_length_in', 'finished_width_in', 'finished_height_in', 'net_weight_lbs',
                'package_length_in', 'package_width_in', 'package_height_in', 'gross_weight_lbs',
                'cost_usd', 'carton_qty', 'package_size_class', 'last_mile_avg_freight_usd'
            }
            
            # 构建列映射，支持中文标签或字段代码
            header_map = {}
            active_multi_base = None
            active_multi_index = 0
            for idx, h in enumerate(headers):
                h_stripped = str(h).strip() if h is not None else ''
                if not h_stripped:
                    if active_multi_base:
                        active_multi_index += 1
                        header_map[f'{active_multi_base}_{active_multi_index}'] = idx
                    continue

                field_code = label_to_code.get(h_stripped, h_stripped)
                base_field = field_code.rsplit('_', 1)[0] if '_' in field_code and field_code[-1].isdigit() else field_code

                if base_field in multi_base_fields:
                    active_multi_base = base_field
                    if '_' in field_code and field_code[-1].isdigit():
                        active_multi_index = int(field_code.rsplit('_', 1)[1])
                    else:
                        active_multi_index = 1
                    header_map[f'{base_field}_{active_multi_index}'] = idx
                    if f'{base_field}_1' not in header_map:
                        header_map[f'{base_field}_1'] = idx
                    if base_field not in header_map:
                        header_map[base_field] = idx
                elif base_field in single_fields:
                    active_multi_base = None
                    active_multi_index = 0
                    header_map[base_field] = idx
                else:
                    active_multi_base = None
                    active_multi_index = 0
                    header_map[field_code] = idx

            def get_cell(row, key):
                idx = header_map.get(key)
                if idx is None:
                    return None
                return row[idx].value

            def parse_bool(raw):
                if raw is None:
                    return 0
                text = str(raw).strip().lower()
                if text in ('1', 'true', 'yes', 'y', '是', '对', 'on', '是否迭代款'):
                    return 1
                return 0

            def normalize_lookup_key(raw):
                text = str(raw or '').strip()
                if not text:
                    return ''
                text = text.replace('\u3000', ' ').replace('\xa0', ' ')
                text = re.sub(r'\s+', ' ', text)
                return text.lower()

            def resolve_option_value(raw, exact_map, normalized_map):
                text = str(raw or '').strip()
                if not text:
                    return None
                val = exact_map.get(text)
                if val:
                    return val
                text_head = text.split('/', 1)[0].strip()
                if text_head:
                    val = exact_map.get(text_head)
                    if val:
                        return val
                return normalized_map.get(normalize_lookup_key(text)) or normalized_map.get(normalize_lookup_key(text_head))
            
            # 支持多列的多选字段收集函数（动态识别所有 _1, _2, _3... 等列）
            def collect_multi_select_values(row, field_base_name, options_map):
                """
                收集某个多选字段的所有列中的值（动态识别 _1, _2, _3, ... 等）
                """
                values = []
                normalized_options_map = {normalize_lookup_key(k): v for k, v in (options_map or {}).items() if normalize_lookup_key(k)}
                # 尝试所有可能的后缀（1-20）
                for i in range(1, 21):
                    col_name = f"{field_base_name}_{i}"
                    cell_value = (get_cell(row, col_name) or '').strip()
                    if cell_value and options_map:
                        val_id = resolve_option_value(cell_value, options_map, normalized_options_map)
                        if val_id:
                            values.append(val_id)
                    elif not cell_value:
                        # 如果某列为空，后续列也可能有值，继续检查
                        continue
                return values

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id, sku_family FROM product_families")
                    sku_map = {row['sku_family']: row['id'] for row in cur.fetchall()}
                    sku_map_norm = {normalize_lookup_key(k): v for k, v in sku_map.items() if normalize_lookup_key(k)}
                    cur.execute("SELECT id, fabric_code, fabric_name_en FROM fabric_materials")
                    fabric_rows = cur.fetchall() or []
                    fabric_map = {}
                    fabric_map_norm = {}
                    for row in fabric_rows:
                        item_id = row.get('id')
                        code = str(row.get('fabric_code') or '').strip()
                        name_en = str(row.get('fabric_name_en') or '').strip()
                        if code:
                            fabric_map[code] = item_id
                            fabric_map_norm[normalize_lookup_key(code)] = item_id
                        if name_en:
                            fabric_map_norm[normalize_lookup_key(name_en)] = item_id
                    cur.execute(
                        """
                        SELECT m.id, m.name, mt.name AS type_name
                        FROM materials m
                        JOIN material_types mt ON m.material_type_id = mt.id
                        """
                    )
                    material_rows = cur.fetchall()
                    filling_map = {row['name']: row['id'] for row in material_rows if row['type_name'] == '填充'}
                    frame_map = {row['name']: row['id'] for row in material_rows if row['type_name'] == '框架'}
                    cur.execute("SELECT id, name FROM features")
                    feature_map = {row['name']: row['id'] for row in cur.fetchall()}
                    cur.execute("SELECT id, name FROM certifications")
                    cert_map = {row['name']: row['id'] for row in cur.fetchall()}
                    cur.execute("SELECT id, factory_name FROM logistics_factories")
                    factory_name_map = {row['factory_name']: row['id'] for row in cur.fetchall()}
                    cur.execute(
                        """
                        SELECT id, sku, sku_family_id, version_no, fabric_id, spec_qty_short,
                               contents_desc_en,
                               is_iteration, is_dachene_product, source_order_product_id,
                               finished_length_in, finished_width_in, finished_height_in,
                               net_weight_lbs, package_length_in, package_width_in, package_height_in,
                               gross_weight_lbs, cost_usd, carton_qty, package_size_class, last_mile_avg_freight_usd
                        FROM order_products
                        """
                    )
                    order_rows = cur.fetchall() or []
                    order_map = {row['sku']: row['id'] for row in order_rows}
                    order_map_norm = {normalize_lookup_key(k): v for k, v in order_map.items() if normalize_lookup_key(k)}
                    order_row_map = {row['id']: row for row in order_rows}

                    cur.execute("SELECT order_product_id, material_id FROM order_product_materials")
                    material_rows = cur.fetchall() or []
                    material_map = {}
                    for mr in material_rows:
                        material_map.setdefault(mr['order_product_id'], set()).add(mr['material_id'])

                    cur.execute("SELECT order_product_id, feature_id FROM order_product_features")
                    feature_rows = cur.fetchall() or []
                    feature_rel_map = {}
                    for fr in feature_rows:
                        feature_rel_map.setdefault(fr['order_product_id'], set()).add(fr['feature_id'])

                    cur.execute("SELECT order_product_id, certification_id FROM order_product_certifications")
                    cert_rows = cur.fetchall() or []
                    cert_rel_map = {}
                    for cr in cert_rows:
                        cert_rel_map.setdefault(cr['order_product_id'], set()).add(cr['certification_id'])

                    cur.execute("SELECT order_product_id, factory_id FROM order_product_factory_links")
                    factory_rows = cur.fetchall() or []
                    factory_rel_map = {}
                    for fr in factory_rows:
                        factory_rel_map.setdefault(fr['order_product_id'], set()).add(fr['factory_id'])

                def _norm(v):
                    if v is None:
                        return None
                    try:
                        if isinstance(v, float):
                            return round(v, 4)
                        return round(float(v), 4)
                    except Exception:
                        return str(v).strip()

                created = 0
                updated = 0
                unchanged = 0
                relation_added = 0
                relation_deleted = 0
                total_rows = 0
                errors = []
                preview_temp_id = -1
                tx_enabled = False
                batch_write_count = 0
                batch_size = 200
                if not preview_mode:
                    try:
                        conn.autocommit(False)
                        tx_enabled = True
                    except Exception:
                        tx_enabled = False
                data_start_row = 4 if header_row_idx == 2 else 2
                
                for row_idx in range(data_start_row, ws.max_row + 1):
                    row = ws[row_idx]
                    row_values = [cell.value for cell in row]
                    if not any(v is not None and str(v).strip() for v in row_values):
                        continue
                    total_rows += 1

                    sku = (get_cell(row, 'sku') or '').strip()
                    sku_family = (get_cell(row, 'sku_family') or '').strip()
                    version_no = (get_cell(row, 'version_no') or '').strip()
                    fabric_code = (get_cell(row, 'fabric_code') or '').strip()
                    spec_qty_short = (get_cell(row, 'spec_qty_short') or '').strip()
                    contents_desc_en = (get_cell(row, 'contents_desc_en') or '').strip()
                    is_iteration = parse_bool(get_cell(row, 'is_iteration'))
                    source_sku = (get_cell(row, 'source_sku') or '').strip()
                    is_dachene_product = parse_bool(get_cell(row, 'is_dachene_product'))
                    is_on_market = parse_bool(get_cell(row, 'is_on_market')) if header_map.get('is_on_market') is not None else 1
                    is_reship_accessory = parse_bool(get_cell(row, 'is_reship_accessory'))

                    if not sku or not sku_family or not fabric_code:
                        errors.append({'row': row_idx, 'error': 'Missing required fields'})
                        continue
                    if is_iteration and not version_no:
                        errors.append({'row': row_idx, 'error': 'Missing version for iteration'})
                        continue

                    sku_family_id = resolve_option_value(sku_family, sku_map, sku_map_norm)
                    fabric_id = resolve_option_value(fabric_code, fabric_map, fabric_map_norm)
                    if not sku_family_id or not fabric_id:
                        errors.append({'row': row_idx, 'error': 'Invalid sku_family or fabric_code'})
                        continue

                    source_order_product_id = None
                    if is_iteration:
                        source_order_product_id = resolve_option_value(source_sku, order_map, order_map_norm)
                        if not source_sku or not source_order_product_id:
                            errors.append({'row': row_idx, 'error': 'Invalid source SKU'})
                            continue

                    payload = {
                        'sku': sku,
                        'sku_family_id': sku_family_id,
                        'version_no': version_no,
                        'fabric_id': fabric_id,
                        'spec_qty_short': spec_qty_short,
                        'contents_desc_en': contents_desc_en or None,
                        'is_iteration': is_iteration,
                        'is_dachene_product': is_dachene_product,
                        'is_on_market': is_on_market,
                        'is_reship_accessory': is_reship_accessory,
                        'source_order_product_id': source_order_product_id,
                        'finished_length_in': self._parse_float(get_cell(row, 'finished_length_in')),
                        'finished_width_in': self._parse_float(get_cell(row, 'finished_width_in')),
                        'finished_height_in': self._parse_float(get_cell(row, 'finished_height_in')),
                        'net_weight_lbs': self._parse_float(get_cell(row, 'net_weight_lbs')),
                        'package_length_in': self._parse_float(get_cell(row, 'package_length_in')),
                        'package_width_in': self._parse_float(get_cell(row, 'package_width_in')),
                        'package_height_in': self._parse_float(get_cell(row, 'package_height_in')),
                        'gross_weight_lbs': self._parse_float(get_cell(row, 'gross_weight_lbs')),
                        'cost_usd': self._parse_float(get_cell(row, 'cost_usd')),
                        'carton_qty': self._parse_int(get_cell(row, 'carton_qty')),
                        'package_size_class': (get_cell(row, 'package_size_class') or '').strip() or None,
                        'last_mile_avg_freight_usd': self._parse_float(get_cell(row, 'last_mile_avg_freight_usd'))
                    }

                    # 支持动态多列多选格式 (field_1, field_2, field_3, ...)
                    filling_ids = collect_multi_select_values(row, 'filling_materials', filling_map)
                    frame_ids = collect_multi_select_values(row, 'frame_materials', frame_map)
                    feature_ids = collect_multi_select_values(row, 'features', feature_map)
                    cert_ids = collect_multi_select_values(row, 'certifications', cert_map)
                    factory_ids = collect_multi_select_values(row, 'factories', factory_name_map)

                    dedup_material_ids = set((filling_ids or []) + (frame_ids or []))
                    dedup_feature_ids = set(feature_ids or [])
                    dedup_cert_ids = set(cert_ids or [])
                    dedup_factory_ids = set(factory_ids or [])

                    target_id = order_map.get(sku)
                    old_material_ids = material_map.get(target_id, set()) if target_id else set()
                    old_feature_ids = feature_rel_map.get(target_id, set()) if target_id else set()
                    old_cert_ids = cert_rel_map.get(target_id, set()) if target_id else set()
                    old_factory_ids = factory_rel_map.get(target_id, set()) if target_id else set()

                    relation_added += len(dedup_material_ids - old_material_ids)
                    relation_added += len(dedup_feature_ids - old_feature_ids)
                    relation_added += len(dedup_cert_ids - old_cert_ids)
                    relation_added += len(dedup_factory_ids - old_factory_ids)
                    relation_deleted += len(old_material_ids - dedup_material_ids)
                    relation_deleted += len(old_feature_ids - dedup_feature_ids)
                    relation_deleted += len(old_cert_ids - dedup_cert_ids)
                    relation_deleted += len(old_factory_ids - dedup_factory_ids)

                    payload_keys = [
                        'sku_family_id', 'version_no', 'fabric_id', 'spec_qty_short', 'contents_desc_en', 'is_iteration', 'is_dachene_product', 'is_on_market', 'is_reship_accessory', 'source_order_product_id',
                        'finished_length_in', 'finished_width_in', 'finished_height_in', 'net_weight_lbs',
                        'package_length_in', 'package_width_in', 'package_height_in', 'gross_weight_lbs',
                        'cost_usd', 'carton_qty', 'package_size_class', 'last_mile_avg_freight_usd'
                    ]
                    is_payload_changed = True
                    if target_id and target_id in order_row_map:
                        old_row = order_row_map[target_id]
                        is_payload_changed = any(_norm(payload.get(k)) != _norm(old_row.get(k)) for k in payload_keys)
                    is_relation_changed = (
                        (dedup_material_ids != old_material_ids)
                        or (dedup_feature_ids != old_feature_ids)
                        or (dedup_cert_ids != old_cert_ids)
                        or (dedup_factory_ids != old_factory_ids)
                    )

                    if target_id and (not is_payload_changed) and (not is_relation_changed):
                        unchanged += 1
                        continue

                    if preview_mode:
                        if target_id:
                            updated += 1
                            order_row_map[target_id] = {**payload, 'id': target_id}
                            material_map[target_id] = dedup_material_ids
                            feature_rel_map[target_id] = dedup_feature_ids
                            cert_rel_map[target_id] = dedup_cert_ids
                            factory_rel_map[target_id] = dedup_factory_ids
                        else:
                            created += 1
                            target_id = preview_temp_id
                            preview_temp_id -= 1
                            order_map[sku] = target_id
                            order_row_map[target_id] = {**payload, 'id': target_id}
                            material_map[target_id] = dedup_material_ids
                            feature_rel_map[target_id] = dedup_feature_ids
                            cert_rel_map[target_id] = dedup_cert_ids
                            factory_rel_map[target_id] = dedup_factory_ids
                        continue

                    try:
                        with conn.cursor() as cur:
                            if target_id:
                                cur.execute(
                                    """
                                    UPDATE order_products
                                    SET sku_family_id=%(sku_family_id)s,
                                        version_no=%(version_no)s,
                                        fabric_id=%(fabric_id)s,
                                        spec_qty_short=%(spec_qty_short)s,
                                        contents_desc_en=%(contents_desc_en)s,
                                        is_iteration=%(is_iteration)s,
                                        is_dachene_product=%(is_dachene_product)s,
                                        is_on_market=%(is_on_market)s,
                                        is_reship_accessory=%(is_reship_accessory)s,
                                        source_order_product_id=%(source_order_product_id)s,
                                        finished_length_in=%(finished_length_in)s,
                                        finished_width_in=%(finished_width_in)s,
                                        finished_height_in=%(finished_height_in)s,
                                        net_weight_lbs=%(net_weight_lbs)s,
                                        package_length_in=%(package_length_in)s,
                                        package_width_in=%(package_width_in)s,
                                        package_height_in=%(package_height_in)s,
                                        gross_weight_lbs=%(gross_weight_lbs)s,
                                        cost_usd=%(cost_usd)s,
                                        carton_qty=%(carton_qty)s,
                                        package_size_class=%(package_size_class)s,
                                        last_mile_avg_freight_usd=%(last_mile_avg_freight_usd)s
                                    WHERE id=%(id)s
                                    """,
                                    {**payload, 'id': target_id}
                                )
                                new_id = target_id
                            else:
                                cur.execute(
                                    """
                                    INSERT INTO order_products (
                                        sku, sku_family_id, version_no, fabric_id, spec_qty_short, contents_desc_en,
                                        is_iteration, is_dachene_product, is_on_market, is_reship_accessory, source_order_product_id,
                                        finished_length_in, finished_width_in, finished_height_in,
                                        net_weight_lbs, package_length_in, package_width_in, package_height_in,
                                        gross_weight_lbs, cost_usd, carton_qty, package_size_class, last_mile_avg_freight_usd
                                    ) VALUES (
                                        %(sku)s, %(sku_family_id)s, %(version_no)s, %(fabric_id)s, %(spec_qty_short)s, %(contents_desc_en)s,
                                        %(is_iteration)s, %(is_dachene_product)s, %(is_on_market)s, %(is_reship_accessory)s, %(source_order_product_id)s,
                                        %(finished_length_in)s, %(finished_width_in)s, %(finished_height_in)s,
                                        %(net_weight_lbs)s, %(package_length_in)s, %(package_width_in)s, %(package_height_in)s,
                                        %(gross_weight_lbs)s, %(cost_usd)s, %(carton_qty)s, %(package_size_class)s, %(last_mile_avg_freight_usd)s
                                    )
                                    """,
                                    payload
                                )
                                new_id = cur.lastrowid
                        if (not target_id) or is_relation_changed:
                            self._replace_order_product_material_ids(conn, new_id, filling_ids, frame_ids)
                            self._replace_order_product_feature_ids(conn, new_id, feature_ids)
                            self._replace_order_product_certification_ids(conn, new_id, cert_ids)
                            self._replace_order_product_factory_links(conn, new_id, factory_ids)

                        material_map[new_id] = dedup_material_ids
                        feature_rel_map[new_id] = dedup_feature_ids
                        cert_rel_map[new_id] = dedup_cert_ids
                        factory_rel_map[new_id] = dedup_factory_ids
                        order_row_map[new_id] = {**payload, 'id': new_id}
                        if target_id:
                            updated += 1
                        else:
                            created += 1
                            order_map[sku] = new_id

                        if tx_enabled:
                            batch_write_count += 1
                            if batch_write_count >= batch_size:
                                conn.commit()
                                batch_write_count = 0
                    except Exception as e:
                        errors.append({'row': row_idx, 'error': str(e)})

                if tx_enabled:
                    if batch_write_count > 0:
                        conn.commit()
                    conn.autocommit(True)

            return self.send_json({
                'status': 'success',
                'preview': 1 if preview_mode else 0,
                'total_rows': total_rows,
                'created': created,
                'updated': updated,
                'unchanged': unchanged,
                'relation_added': relation_added,
                'relation_deleted': relation_deleted,
                'errors': errors
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_parent_api(self, environ, method, start_response):
        """父体管理 API（CRUD）"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))

            def limited_text(value, max_len):
                text = (value or '').strip()
                if not text:
                    return None
                if len(text) > max_len:
                    raise ValueError(f'文本长度超限（>{max_len}）')
                return text

            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                item_id = self._parse_int((query_params.get('id', [''])[0] or '').strip())
                simple_mode = str((query_params.get('simple', ['0'])[0] or '0')).lower() in ('1', 'true', 'yes', 'on')
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if simple_mode:
                            sql = """
                                SELECT sp.id, sp.parent_code, sp.is_enabled, sp.shop_id, sp.sku_marker,
                                       estimated_refund_rate, estimated_discount_rate,
                                       commission_rate, estimated_acoas,
                                       sp.created_at, sp.updated_at,
                                       s.shop_name, b.name AS brand_name, pt.name AS platform_type_name
                                FROM sales_parents sp
                                LEFT JOIN shops s ON s.id = sp.shop_id
                                LEFT JOIN brands b ON b.id = s.brand_id
                                LEFT JOIN platform_types pt ON pt.id = s.platform_type_id
                            """
                        else:
                            sql = """
                                SELECT sp.id, sp.parent_code, sp.is_enabled, sp.shop_id, sp.sku_marker,
                                       estimated_refund_rate, estimated_discount_rate,
                                       commission_rate, estimated_acoas,
                                       sales_title, sales_intro,
                                       sales_bullet_1, sales_bullet_2, sales_bullet_3, sales_bullet_4, sales_bullet_5,
                                    sp.created_at, sp.updated_at,
                                    s.shop_name, b.name AS brand_name, pt.name AS platform_type_name
                                FROM sales_parents sp
                                LEFT JOIN shops s ON s.id = sp.shop_id
                                LEFT JOIN brands b ON b.id = s.brand_id
                                LEFT JOIN platform_types pt ON pt.id = s.platform_type_id
                            """
                        params = []
                        filters = []
                        if item_id:
                            filters.append("sp.id = %s")
                            params.append(item_id)
                        if keyword:
                            like_kw = f"%{keyword}%"
                            if simple_mode:
                                filters.append("(sp.parent_code LIKE %s OR sp.sku_marker LIKE %s)")
                                params.extend([like_kw, like_kw])
                            else:
                                filters.append("(sp.parent_code LIKE %s OR sp.sku_marker LIKE %s OR sp.sales_title LIKE %s OR sp.sales_intro LIKE %s)")
                                params.extend([like_kw, like_kw, like_kw, like_kw])
                        if filters:
                            sql += " WHERE " + " AND ".join(filters)
                        sql += " ORDER BY sp.id DESC"
                        cur.execute(sql, params)
                        rows = cur.fetchall() or []
                if item_id:
                    return self.send_json({'status': 'success', 'item': rows[0] if rows else None}, start_response)
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                parent_code = (data.get('parent_code') or '').strip()
                if not parent_code:
                    return self.send_json({'status': 'error', 'message': 'Missing parent_code'}, start_response)
                is_enabled_raw = data.get('is_enabled', 1)
                is_enabled = 1 if str(is_enabled_raw).strip().lower() in ('1', 'true', 'yes', 'on') else 0
                shop_id = self._parse_int(data.get('shop_id'))
                try:
                    sku_marker = limited_text(data.get('sku_marker'), 128)
                    sales_title = limited_text(data.get('sales_title'), 200)
                    sales_intro = limited_text(data.get('sales_intro'), 500)
                    sales_bullet_1 = limited_text(data.get('sales_bullet_1'), 500)
                    sales_bullet_2 = limited_text(data.get('sales_bullet_2'), 500)
                    sales_bullet_3 = limited_text(data.get('sales_bullet_3'), 500)
                    sales_bullet_4 = limited_text(data.get('sales_bullet_4'), 500)
                    sales_bullet_5 = limited_text(data.get('sales_bullet_5'), 500)
                except ValueError as ve:
                    return self.send_json({'status': 'error', 'message': str(ve)}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO sales_parents
                            (parent_code, is_enabled, shop_id, sku_marker, estimated_refund_rate, estimated_discount_rate, commission_rate, estimated_acoas,
                             sales_title, sales_intro, sales_bullet_1, sales_bullet_2, sales_bullet_3, sales_bullet_4, sales_bullet_5)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                parent_code,
                                is_enabled,
                                shop_id,
                                sku_marker,
                                self._parse_float(data.get('estimated_refund_rate')),
                                self._parse_float(data.get('estimated_discount_rate')),
                                self._parse_float(data.get('commission_rate')),
                                self._parse_float(data.get('estimated_acoas')),
                                sales_title,
                                sales_intro,
                                sales_bullet_1,
                                sales_bullet_2,
                                sales_bullet_3,
                                sales_bullet_4,
                                sales_bullet_5
                            )
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                is_enabled_raw = data.get('is_enabled', 1)
                is_enabled = 1 if str(is_enabled_raw).strip().lower() in ('1', 'true', 'yes', 'on') else 0
                shop_id = self._parse_int(data.get('shop_id'))
                try:
                    sku_marker = limited_text(data.get('sku_marker'), 128)
                    sales_title = limited_text(data.get('sales_title'), 200)
                    sales_intro = limited_text(data.get('sales_intro'), 500)
                    sales_bullet_1 = limited_text(data.get('sales_bullet_1'), 500)
                    sales_bullet_2 = limited_text(data.get('sales_bullet_2'), 500)
                    sales_bullet_3 = limited_text(data.get('sales_bullet_3'), 500)
                    sales_bullet_4 = limited_text(data.get('sales_bullet_4'), 500)
                    sales_bullet_5 = limited_text(data.get('sales_bullet_5'), 500)
                except ValueError as ve:
                    return self.send_json({'status': 'error', 'message': str(ve)}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE sales_parents
                            SET parent_code=%s,
                                is_enabled=%s,
                                shop_id=%s,
                                sku_marker=%s,
                                estimated_refund_rate=%s,
                                estimated_discount_rate=%s,
                                commission_rate=%s,
                                estimated_acoas=%s,
                                sales_title=%s,
                                sales_intro=%s,
                                sales_bullet_1=%s,
                                sales_bullet_2=%s,
                                sales_bullet_3=%s,
                                sales_bullet_4=%s,
                                sales_bullet_5=%s
                            WHERE id=%s
                            """,
                            (
                                (data.get('parent_code') or '').strip(),
                                is_enabled,
                                shop_id,
                                sku_marker,
                                self._parse_float(data.get('estimated_refund_rate')),
                                self._parse_float(data.get('estimated_discount_rate')),
                                self._parse_float(data.get('commission_rate')),
                                self._parse_float(data.get('estimated_acoas')),
                                sales_title,
                                sales_intro,
                                sales_bullet_1,
                                sales_bullet_2,
                                sales_bullet_3,
                                sales_bullet_4,
                                sales_bullet_5,
                                item_id
                            )
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM sales_parents WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)



    def _normalize_id_list(self, value):
        if value is None:
            return []
        items = value if isinstance(value, list) else re.split(r'[\s,，;；]+', str(value))
        out = []
        seen = set()
        for raw in items:
            item_id = self._parse_int(raw)
            if not item_id or item_id in seen:
                continue
            seen.add(item_id)
            out.append(item_id)
        return out

    def _load_order_product_rows(self, conn, keyword='', include_relations=False, limit_rows=1200, item_id=None, exclude_reship_accessory=False):
        max_rows = max(1, self._parse_int(limit_rows) or 1200)
        accessory_filter = " AND COALESCE(op.is_reship_accessory, 0)=0" if exclude_reship_accessory else ""
        with conn.cursor() as cur:
            if item_id:
                cur.execute(
                    f"""
                    SELECT
                        op.*,
                        pf.sku_family,
                        pf.category,
                        fm.fabric_code,
                        fm.fabric_name_en,
                        fm.representative_color
                    FROM order_products op
                    LEFT JOIN product_families pf ON pf.id = op.sku_family_id
                    LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                    WHERE op.id=%s{accessory_filter}
                    LIMIT 1
                    """,
                    (item_id,)
                )
            elif keyword:
                like_val = f"%{keyword}%"
                cur.execute(
                    f"""
                    SELECT
                        op.*,
                        pf.sku_family,
                        pf.category,
                        fm.fabric_code,
                        fm.fabric_name_en,
                        fm.representative_color
                    FROM order_products op
                    LEFT JOIN product_families pf ON pf.id = op.sku_family_id
                    LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                    WHERE op.sku LIKE %s
                       OR pf.sku_family LIKE %s
                       OR fm.fabric_code LIKE %s
                       OR fm.fabric_name_en LIKE %s
                       OR op.version_no LIKE %s
                      {accessory_filter}
                    ORDER BY op.id DESC
                    LIMIT %s
                    """,
                    (like_val, like_val, like_val, like_val, like_val, max_rows)
                )
            else:
                cur.execute(
                    f"""
                    SELECT
                        op.*,
                        pf.sku_family,
                        pf.category,
                        fm.fabric_code,
                        fm.fabric_name_en,
                        fm.representative_color
                    FROM order_products op
                    LEFT JOIN product_families pf ON pf.id = op.sku_family_id
                    LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                    WHERE 1=1 {accessory_filter}
                    ORDER BY op.id DESC
                    LIMIT %s
                    """,
                    (max_rows,)
                )
            rows = cur.fetchall() or []

        if include_relations and rows:
            self._attach_order_product_relations(conn, rows)
        return rows

    def _attach_order_product_relations(self, conn, rows):
        op_ids = [self._parse_int(row.get('id')) for row in rows if self._parse_int(row.get('id'))]
        material_map = {}
        feature_map = {}
        certification_map = {}

        if op_ids:
            placeholders = ','.join(['%s'] * len(op_ids))
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT opm.order_product_id, opm.material_id, m.name, m.name_en,
                           mt.name AS material_type_name
                    FROM order_product_materials opm
                    JOIN materials m ON m.id = opm.material_id
                    LEFT JOIN material_types mt ON mt.id = m.material_type_id
                    WHERE opm.order_product_id IN ({placeholders})
                    ORDER BY opm.order_product_id ASC, opm.material_id ASC
                    """,
                    tuple(op_ids)
                )
                for rel in (cur.fetchall() or []):
                    order_id = self._parse_int(rel.get('order_product_id'))
                    if not order_id:
                        continue
                    material_map.setdefault(order_id, []).append(rel)

                cur.execute(
                    f"""
                    SELECT opf.order_product_id, opf.feature_id, f.name, f.name_en
                    FROM order_product_features opf
                    JOIN features f ON f.id = opf.feature_id
                    WHERE opf.order_product_id IN ({placeholders})
                    ORDER BY opf.order_product_id ASC, opf.feature_id ASC
                    """,
                    tuple(op_ids)
                )
                for rel in (cur.fetchall() or []):
                    order_id = self._parse_int(rel.get('order_product_id'))
                    if not order_id:
                        continue
                    feature_map.setdefault(order_id, []).append(rel)

                cur.execute(
                    f"""
                    SELECT opc.order_product_id, opc.certification_id, c.name
                    FROM order_product_certifications opc
                    JOIN certifications c ON c.id = opc.certification_id
                    WHERE opc.order_product_id IN ({placeholders})
                    ORDER BY opc.order_product_id ASC, opc.certification_id ASC
                    """,
                    tuple(op_ids)
                )
                for rel in (cur.fetchall() or []):
                    order_id = self._parse_int(rel.get('order_product_id'))
                    if not order_id:
                        continue
                    certification_map.setdefault(order_id, []).append(rel)

        for row in rows:
            order_id = self._parse_int(row.get('id'))
            materials = material_map.get(order_id, []) if order_id else []
            features = feature_map.get(order_id, []) if order_id else []
            certifications = certification_map.get(order_id, []) if order_id else []

            filling_ids = []
            frame_ids = []
            filling_names = []
            frame_names = []
            for material in materials:
                material_id = self._parse_int(material.get('material_id'))
                if not material_id:
                    continue
                material_name = f"{material.get('name') or ''} / {material.get('name_en') or ''}".strip(' /')
                type_name = str(material.get('material_type_name') or '').strip()
                if type_name == '填充':
                    filling_ids.append(material_id)
                    if material_name:
                        filling_names.append(material_name)
                elif type_name == '框架':
                    frame_ids.append(material_id)
                    if material_name:
                        frame_names.append(material_name)

            row['filling_material_ids'] = filling_ids
            row['frame_material_ids'] = frame_ids
            row['filling_material_names'] = filling_names
            row['frame_material_names'] = frame_names

            row['feature_ids'] = [self._parse_int(item.get('feature_id')) for item in features if self._parse_int(item.get('feature_id'))]
            row['feature_names'] = [
                f"{item.get('name') or ''} / {item.get('name_en') or ''}".strip(' /')
                for item in features
                if item.get('name') or item.get('name_en')
            ]

            row['certification_ids'] = [self._parse_int(item.get('certification_id')) for item in certifications if self._parse_int(item.get('certification_id'))]
            row['certification_names'] = [item.get('name') for item in certifications if item.get('name')]

    def _attach_order_product_factory_links(self, conn, rows):
        op_ids = [self._parse_int(row.get('id')) for row in rows if self._parse_int(row.get('id'))]
        factory_map = {}
        if op_ids:
            placeholders = ','.join(['%s'] * len(op_ids))
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT opl.order_product_id, opl.factory_id, lf.factory_name
                    FROM order_product_factory_links opl
                    JOIN logistics_factories lf ON lf.id = opl.factory_id
                    WHERE opl.order_product_id IN ({placeholders})
                    ORDER BY opl.order_product_id ASC, lf.factory_name ASC
                    """,
                    tuple(op_ids)
                )
                for rel in (cur.fetchall() or []):
                    order_id = self._parse_int(rel.get('order_product_id'))
                    if not order_id:
                        continue
                    factory_map.setdefault(order_id, []).append(rel)

        for row in rows:
            order_id = self._parse_int(row.get('id'))
            factories = factory_map.get(order_id, []) if order_id else []
            row['factory_ids'] = [self._parse_int(item.get('factory_id')) for item in factories if self._parse_int(item.get('factory_id'))]
            row['factory_names'] = [item.get('factory_name') for item in factories if item.get('factory_name')]

    def _replace_order_product_relations(self, conn, order_product_id, filling_material_ids, frame_material_ids, feature_ids, certification_ids):
        self._replace_order_product_material_ids(conn, order_product_id, filling_material_ids, frame_material_ids)
        self._replace_order_product_feature_ids(conn, order_product_id, feature_ids)
        self._replace_order_product_certification_ids(conn, order_product_id, certification_ids)

    def _replace_order_product_material_ids(self, conn, order_product_id, filling_material_ids, frame_material_ids):
        merged = []
        seen = set()
        for raw_id in list(filling_material_ids or []) + list(frame_material_ids or []):
            material_id = self._parse_int(raw_id)
            if not material_id or material_id in seen:
                continue
            seen.add(material_id)
            merged.append(material_id)

        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_materials WHERE order_product_id=%s", (order_product_id,))
            if merged:
                cur.executemany(
                    "INSERT INTO order_product_materials (order_product_id, material_id) VALUES (%s, %s)",
                    [(order_product_id, material_id) for material_id in merged]
                )

    def _replace_order_product_feature_ids(self, conn, order_product_id, feature_ids):
        valid_ids = []
        seen = set()
        for raw_id in (feature_ids or []):
            feature_id = self._parse_int(raw_id)
            if not feature_id or feature_id in seen:
                continue
            seen.add(feature_id)
            valid_ids.append(feature_id)

        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_features WHERE order_product_id=%s", (order_product_id,))
            if valid_ids:
                cur.executemany(
                    "INSERT INTO order_product_features (order_product_id, feature_id) VALUES (%s, %s)",
                    [(order_product_id, feature_id) for feature_id in valid_ids]
                )

    def _replace_order_product_certification_ids(self, conn, order_product_id, certification_ids):
        valid_ids = []
        seen = set()
        for raw_id in (certification_ids or []):
            certification_id = self._parse_int(raw_id)
            if not certification_id or certification_id in seen:
                continue
            seen.add(certification_id)
            valid_ids.append(certification_id)

        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_certifications WHERE order_product_id=%s", (order_product_id,))
            if valid_ids:
                cur.executemany(
                    "INSERT INTO order_product_certifications (order_product_id, certification_id) VALUES (%s, %s)",
                    [(order_product_id, certification_id) for certification_id in valid_ids]
                )

    def _replace_order_product_factory_links(self, conn, order_product_id, factory_ids):
        valid_ids = []
        seen = set()
        for raw_id in (factory_ids or []):
            factory_id = self._parse_int(raw_id)
            if not factory_id or factory_id in seen:
                continue
            seen.add(factory_id)
            valid_ids.append(factory_id)

        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_factory_links WHERE order_product_id=%s", (order_product_id,))
            if valid_ids:
                cur.executemany(
                    "INSERT INTO order_product_factory_links (order_product_id, factory_id) VALUES (%s, %s)",
                    [(order_product_id, factory_id) for factory_id in valid_ids]
                )

    def _handle_order_product_factory_links_template(self, environ, method, start_response):
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)

            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            conn = self.get_db_connection()
            try:
                # 获取所有SKU列表
                sku_list = []
                with conn.cursor() as cur:
                    cur.execute("SELECT DISTINCT sku FROM order_products WHERE sku IS NOT NULL AND sku != '' ORDER BY sku")
                    sku_list = [row[0] for row in cur.fetchall()]
                
                # 获取所有工厂列表
                factory_list = []
                with conn.cursor() as cur:
                    cur.execute("SELECT DISTINCT factory_name FROM logistics_factories WHERE factory_name IS NOT NULL AND factory_name != '' ORDER BY factory_name")
                    factory_list = [row[0] for row in cur.fetchall()]
            finally:
                if conn:
                    conn.close()

            wb = Workbook()
            ws = wb.active
            ws.title = 'sku_factory_links'

            ws.cell(row=1, column=1).value = 'SKU'
            ws.cell(row=1, column=2).value = '工厂'
            ws.cell(row=2, column=1).value = '示例SKU' if not sku_list else sku_list[0]
            ws.cell(row=2, column=2).value = '示例工厂' if not factory_list else factory_list[0]

            ws.column_dimensions['A'].width = 28
            ws.column_dimensions['B'].width = 28
            ws.freeze_panes = 'A2'

            # 为SKU列(A列)添加数据验证，从第3行开始到第1000行
            if sku_list and DataValidation is not None:
                sku_dv = DataValidation(
                    type='list',
                    formula1='"' + ','.join(sku_list) + '"',
                    allow_blank=True
                )
                sku_dv.error = '请从列表中选择有效的SKU'
                sku_dv.errorTitle = '无效的SKU'
                ws.add_data_validation(sku_dv)
                sku_dv.add(f'A3:A1000')

            # 为工厂列(B列)添加数据验证，从第3行开始到第1000行
            if factory_list and DataValidation is not None:
                factory_dv = DataValidation(
                    type='list',
                    formula1='"' + ','.join(factory_list) + '"',
                    allow_blank=True
                )
                factory_dv.error = '请从列表中选择有效的工厂'
                factory_dv.errorTitle = '无效的工厂'
                ws.add_data_validation(factory_dv)
                factory_dv.add(f'B3:B1000')

            return self._send_excel_workbook(wb, 'order_product_factory_links_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _handle_order_product_factory_links_import(self, environ, method, start_response):
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': 'openpyxl not available'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            header_row = 1
            header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), tuple())
            headers = [str(v or '').strip() for v in header_values]
            header_map = {name: idx for idx, name in enumerate(headers)}
            for required in ('SKU', '工厂'):
                if required not in header_map:
                    return self.send_json({'status': 'error', 'message': f'模板缺少列: {required}'}, start_response)

            def get_cell(row, name):
                idx = header_map.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            created = 0
            unchanged = 0
            errors = []
            user_id = self._get_session_user(environ)
            scope_ids = self._get_user_factory_scope_ids(user_id)

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    sku_scope_ids = self._get_linked_order_product_ids(scope_ids)
                    if sku_scope_ids is None:
                        cur.execute("SELECT id, sku FROM order_products")
                    elif sku_scope_ids:
                        placeholders = ','.join(['%s'] * len(sku_scope_ids))
                        cur.execute(
                            f"SELECT id, sku FROM order_products WHERE id IN ({placeholders})",
                            tuple(sku_scope_ids)
                        )
                    else:
                        cur.execute("SELECT id, sku FROM order_products WHERE 1=0")
                    sku_map = {str(r.get('sku') or '').strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}

                    factory_clause, factory_params = self._factory_scope_clause('id', user_id, prefix='WHERE')
                    cur.execute(
                        f"SELECT id, factory_name FROM logistics_factories{factory_clause}",
                        factory_params
                    )
                    factory_map = {str(r.get('factory_name') or '').strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}

                    pairs = []
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if not any(value is not None and str(value).strip() for value in row):
                            continue
                        try:
                            sku = str(get_cell(row, 'SKU') or '').strip()
                            factory_name = str(get_cell(row, '工厂') or '').strip()
                            if not sku or not factory_name:
                                raise ValueError('SKU 和工厂不能为空')
                            order_product_id = sku_map.get(sku)
                            factory_id = factory_map.get(factory_name)
                            if not order_product_id:
                                raise ValueError(f'未找到SKU: {sku}')
                            if not factory_id:
                                raise ValueError(f'未找到工厂: {factory_name}')
                            if not self._order_product_allowed_for_factory(order_product_id, factory_id):
                                raise ValueError(f'SKU {sku} 未关联工厂 {factory_name}')
                            pairs.append((order_product_id, factory_id))
                        except Exception as row_error:
                            errors.append({'row': row_idx, 'error': str(row_error)})

                    if pairs:
                        cur.executemany(
                            "INSERT IGNORE INTO order_product_factory_links (order_product_id, factory_id) VALUES (%s, %s)",
                            pairs
                        )
                        created = cur.rowcount or len(pairs)

            return self.send_json({'status': 'success', 'created': created, 'unchanged': unchanged, 'errors': errors}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

        return rows

    def _get_or_create_shipping_plan(self, conn, order_product_id, plan_name):
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id FROM order_product_shipping_plans WHERE order_product_id=%s AND plan_name=%s LIMIT 1",
                (order_product_id, plan_name)
            )
            row = cur.fetchone() or {}
            plan_id = self._parse_int(row.get('id'))
            if plan_id:
                return plan_id

            cur.execute(
                "INSERT INTO order_product_shipping_plans (order_product_id, plan_name) VALUES (%s, %s)",
                (order_product_id, plan_name)
            )
            return cur.lastrowid

    def _set_shipping_plan_items(self, conn, plan_id, substitute_order_product_ids):
        ids = []
        seen = set()
        for raw_id in (substitute_order_product_ids or []):
            item_id = self._parse_int(raw_id)
            if not item_id or item_id in seen:
                continue
            seen.add(item_id)
            ids.append(item_id)

        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_shipping_plan_items WHERE shipping_plan_id=%s", (plan_id,))
            if ids:
                cur.executemany(
                    """
                    INSERT INTO order_product_shipping_plan_items
                        (shipping_plan_id, substitute_order_product_id, quantity, sort_order)
                    VALUES (%s, %s, %s, %s)
                    """,
                    [
                        (plan_id, sid, 1, idx + 1)
                        for idx, sid in enumerate(ids)
                    ]
                )

    def _auto_sync_iteration_shipping_plans(self, conn, new_order_product_id, source_order_product_id, version_no):
        new_generation = self._parse_iteration_generation(version_no)
        if not new_generation:
            return

        source_id = self._parse_int(source_order_product_id)
        new_id = self._parse_int(new_order_product_id)
        if not source_id or not new_id:
            return

        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, source_order_product_id, is_iteration, version_no
                FROM order_products
                WHERE id=%s OR source_order_product_id=%s
                ORDER BY id ASC
                """,
                (source_id, source_id)
            )
            family_rows = cur.fetchall() or []

        members = []
        for row in family_rows:
            member_id = self._parse_int(row.get('id'))
            if not member_id:
                continue
            if member_id == source_id and not self._parse_int(row.get('is_iteration')):
                member_generation = 1
            else:
                member_generation = self._parse_iteration_generation(row.get('version_no'))
            if not member_generation:
                continue
            members.append((member_id, member_generation))

        if not members:
            return

        other_members = [(mid, gen) for (mid, gen) in members if mid != new_id]
        if not other_members:
            return

        # 旧款：创建/覆盖“迭代款-新代数”方案，且仅包含新SKU。
        forward_plan_name = f"迭代款-{new_generation}代"
        for owner_id, _ in other_members:
            owner_plan_id = self._get_or_create_shipping_plan(conn, owner_id, forward_plan_name)
            self._set_shipping_plan_items(conn, owner_plan_id, [new_id])

        # 新款：按已有每一代创建对应方案，每个方案包含该代已有SKU。
        generation_map = {}
        for member_id, member_generation in other_members:
            if member_generation <= 0:
                continue
            generation_map.setdefault(member_generation, []).append(member_id)

        for member_generation in sorted(generation_map.keys()):
            reverse_plan_name = f"迭代款-{member_generation}代"
            new_plan_id = self._get_or_create_shipping_plan(conn, new_id, reverse_plan_name)
            self._set_shipping_plan_items(conn, new_plan_id, generation_map.get(member_generation, []))

    def _handle_order_product_shipping_plans(self, environ, method, start_response, query_params):
        try:
            if method == 'GET':
                order_product_id = self._parse_int(query_params.get('order_product_id', [''])[0])
                if not order_product_id:
                    return self.send_json({'status': 'error', 'message': 'Missing order_product_id'}, start_response)

                usage_items = []
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT id, order_product_id, plan_name, created_at, updated_at
                            FROM order_product_shipping_plans
                            WHERE order_product_id=%s
                            ORDER BY id DESC
                            """,
                            (order_product_id,)
                        )
                        plans = cur.fetchall() or []
                        plan_ids = [self._parse_int(row.get('id')) for row in plans if self._parse_int(row.get('id'))]
                        item_map = {}
                        if plan_ids:
                            placeholders = ','.join(['%s'] * len(plan_ids))
                            cur.execute(
                                f"""
                                SELECT
                                    opsi.id,
                                    opsi.shipping_plan_id,
                                    opsi.substitute_order_product_id,
                                    opsi.quantity,
                                    opsi.sort_order,
                                    op.sku AS substitute_order_sku
                                FROM order_product_shipping_plan_items opsi
                                JOIN order_products op ON op.id = opsi.substitute_order_product_id
                                WHERE opsi.shipping_plan_id IN ({placeholders})
                                ORDER BY opsi.shipping_plan_id ASC, opsi.sort_order ASC, opsi.id ASC
                                """,
                                tuple(plan_ids)
                            )
                            for item in (cur.fetchall() or []):
                                plan_id = self._parse_int(item.get('shipping_plan_id'))
                                if not plan_id:
                                    continue
                                item_map.setdefault(plan_id, []).append(item)

                        for plan in plans:
                            pid = self._parse_int(plan.get('id'))
                            plan['items'] = item_map.get(pid, []) if pid else []

                        cur.execute(
                            """
                            SELECT DISTINCT
                                ops.id,
                                ops.plan_name,
                                ops.order_product_id,
                                target_op.sku AS target_order_sku
                            FROM order_product_shipping_plan_items opsi
                            JOIN order_product_shipping_plans ops ON ops.id = opsi.shipping_plan_id
                            JOIN order_products target_op ON target_op.id = ops.order_product_id
                            WHERE opsi.substitute_order_product_id = %s
                            ORDER BY ops.id DESC
                            """,
                            (order_product_id,)
                        )
                        usage_plan_rows = cur.fetchall() or []
                        usage_plan_ids = [self._parse_int(row.get('id')) for row in usage_plan_rows if self._parse_int(row.get('id'))]

                        usage_item_map = {}
                        if usage_plan_ids:
                            usage_placeholders = ','.join(['%s'] * len(usage_plan_ids))
                            cur.execute(
                                f"""
                                SELECT
                                    opsi.id,
                                    opsi.shipping_plan_id,
                                    opsi.substitute_order_product_id,
                                    opsi.quantity,
                                    opsi.sort_order,
                                    op.sku AS substitute_order_sku
                                FROM order_product_shipping_plan_items opsi
                                JOIN order_products op ON op.id = opsi.substitute_order_product_id
                                WHERE opsi.shipping_plan_id IN ({usage_placeholders})
                                ORDER BY opsi.shipping_plan_id ASC, opsi.sort_order ASC, opsi.id ASC
                                """,
                                tuple(usage_plan_ids)
                            )
                            for usage_item in (cur.fetchall() or []):
                                usage_plan_id = self._parse_int(usage_item.get('shipping_plan_id'))
                                if not usage_plan_id:
                                    continue
                                usage_item_map.setdefault(usage_plan_id, []).append(usage_item)

                        current_substitute_ids = set()
                        for items_in_plan in item_map.values():
                            for item in (items_in_plan or []):
                                sid = self._parse_int(item.get('substitute_order_product_id'))
                                if sid:
                                    current_substitute_ids.add(sid)

                        usage_items = []
                        for usage_plan_row in usage_plan_rows:
                            plan_id = self._parse_int(usage_plan_row.get('id'))
                            if not plan_id:
                                continue
                            target_order_product_id = self._parse_int(usage_plan_row.get('order_product_id'))
                            usage_plan_items = usage_item_map.get(plan_id, [])
                            only_contains_current = (
                                len(usage_plan_items) == 1
                                and self._parse_int(usage_plan_items[0].get('substitute_order_product_id')) == order_product_id
                            )
                            can_quick_apply = (
                                bool(target_order_product_id)
                                and only_contains_current
                                and target_order_product_id not in current_substitute_ids
                            )
                            plan_obj = {
                                'id': plan_id,
                                'plan_name': usage_plan_row.get('plan_name') or '',
                                'order_product_id': target_order_product_id,
                                'target_order_sku': usage_plan_row.get('target_order_sku') or '',
                                'items': usage_plan_items,
                                'can_quick_apply': can_quick_apply
                            }
                            usage_items.append(plan_obj)

                return self.send_json({'status': 'success', 'items': plans, 'usage_items': usage_items}, start_response)

            data = self._read_json_body(environ)

            if method == 'POST':
                order_product_id = self._parse_int(data.get('order_product_id'))
                if data.get('quick_apply_from_usage'):
                    target_order_product_id = self._parse_int(data.get('target_order_product_id'))
                    if not order_product_id or not target_order_product_id:
                        return self.send_json({'status': 'error', 'message': 'Missing fields'}, start_response)
                    if order_product_id == target_order_product_id:
                        return self.send_json({'status': 'error', 'message': '不能将当前SKU本身设为替代SKU'}, start_response)

                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute("SELECT sku FROM order_products WHERE id=%s", (target_order_product_id,))
                            target_row = cur.fetchone() or {}
                            target_sku = (target_row.get('sku') or '').strip()
                            if not target_sku:
                                return self.send_json({'status': 'error', 'message': '目标SKU不存在'}, start_response)

                            cur.execute(
                                """
                                SELECT 1
                                FROM order_product_shipping_plan_items opsi
                                JOIN order_product_shipping_plans ops ON ops.id = opsi.shipping_plan_id
                                WHERE ops.order_product_id = %s
                                  AND opsi.substitute_order_product_id = %s
                                LIMIT 1
                                """,
                                (order_product_id, target_order_product_id)
                            )
                            if cur.fetchone():
                                return self.send_json({'status': 'error', 'message': '当前SKU已将该SKU作为替代发货选项'}, start_response)

                            base_plan_name = f"快速替代-{target_sku}"
                            plan_name = base_plan_name
                            suffix = 2
                            while True:
                                cur.execute(
                                    """
                                    SELECT 1
                                    FROM order_product_shipping_plans
                                    WHERE order_product_id=%s AND plan_name=%s
                                    LIMIT 1
                                    """,
                                    (order_product_id, plan_name)
                                )
                                if not cur.fetchone():
                                    break
                                plan_name = f"{base_plan_name}-{suffix}"
                                suffix += 1

                            cur.execute(
                                "INSERT INTO order_product_shipping_plans (order_product_id, plan_name) VALUES (%s, %s)",
                                (order_product_id, plan_name)
                            )
                            new_plan_id = cur.lastrowid
                            cur.execute(
                                """
                                INSERT INTO order_product_shipping_plan_items
                                    (shipping_plan_id, substitute_order_product_id, quantity, sort_order)
                                VALUES (%s, %s, %s, %s)
                                """,
                                (new_plan_id, target_order_product_id, 1, 1)
                            )

                    return self.send_json({'status': 'success', 'id': new_plan_id}, start_response)

                plan_name = (data.get('plan_name') or '').strip()
                items = data.get('items') or []
                if not order_product_id or not plan_name:
                    return self.send_json({'status': 'error', 'message': 'Missing fields'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "INSERT INTO order_product_shipping_plans (order_product_id, plan_name) VALUES (%s, %s)",
                            (order_product_id, plan_name)
                        )
                        plan_id = cur.lastrowid

                        insert_rows = []
                        for idx, item in enumerate(items):
                            substitute_order_product_id = self._parse_int(item.get('substitute_order_product_id'))
                            quantity = max(1, self._parse_int(item.get('quantity')) or 1)
                            sort_order = self._parse_int(item.get('sort_order')) or (idx + 1)
                            if not substitute_order_product_id:
                                continue
                            insert_rows.append((plan_id, substitute_order_product_id, quantity, sort_order))
                        if insert_rows:
                            cur.executemany(
                                """
                                INSERT INTO order_product_shipping_plan_items
                                    (shipping_plan_id, substitute_order_product_id, quantity, sort_order)
                                VALUES (%s, %s, %s, %s)
                                """,
                                insert_rows
                            )
                return self.send_json({'status': 'success', 'id': plan_id}, start_response)

            if method == 'PUT':
                plan_id = self._parse_int(data.get('id'))
                order_product_id = self._parse_int(data.get('order_product_id'))
                plan_name = (data.get('plan_name') or '').strip()
                items = data.get('items') or []
                if not plan_id or not order_product_id or not plan_name:
                    return self.send_json({'status': 'error', 'message': 'Missing fields'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE order_product_shipping_plans
                            SET order_product_id=%s, plan_name=%s
                            WHERE id=%s
                            """,
                            (order_product_id, plan_name, plan_id)
                        )
                        cur.execute("DELETE FROM order_product_shipping_plan_items WHERE shipping_plan_id=%s", (plan_id,))
                        insert_rows = []
                        for idx, item in enumerate(items):
                            substitute_order_product_id = self._parse_int(item.get('substitute_order_product_id'))
                            quantity = max(1, self._parse_int(item.get('quantity')) or 1)
                            sort_order = self._parse_int(item.get('sort_order')) or (idx + 1)
                            if not substitute_order_product_id:
                                continue
                            insert_rows.append((plan_id, substitute_order_product_id, quantity, sort_order))
                        if insert_rows:
                            cur.executemany(
                                """
                                INSERT INTO order_product_shipping_plan_items
                                    (shipping_plan_id, substitute_order_product_id, quantity, sort_order)
                                VALUES (%s, %s, %s, %s)
                                """,
                                insert_rows
                            )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                plan_id = self._parse_int(data.get('id'))
                if not plan_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM order_product_shipping_plans WHERE id=%s", (plan_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
