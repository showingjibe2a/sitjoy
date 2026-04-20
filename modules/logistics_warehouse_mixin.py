import cgi
import csv
import io
import os
import re
from datetime import datetime, date, timedelta
from urllib.parse import parse_qs, quote

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    _openpyxl_import_error = None
except Exception as _e:
    Workbook = None
    load_workbook = None
    PatternFill = None
    Font = None
    Alignment = None
    Border = None
    Side = None
    DataValidation = None
    _openpyxl_import_error = _e


class LogisticsWarehouseMixin:

    def _normalize_id_list_local(self, value):
        if value is None:
            return []
        items = value if isinstance(value, list) else re.split(r'[\s,，;；]+', str(value))
        out = []
        seen = set()
        for raw in items:
            item_id = self._parse_int(raw)
            if not item_id or item_id in seen:
                continue
            seen.add(item_id)
            out.append(item_id)
        return out

    def _replace_factory_order_product_links(self, conn, factory_id, order_product_ids):
        factory_num = self._parse_int(factory_id)
        if not factory_num:
            return
        ids = self._normalize_id_list_local(order_product_ids)
        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_factory_links WHERE factory_id=%s", (factory_num,))
            if ids:
                cur.executemany(
                    "INSERT INTO order_product_factory_links (order_product_id, factory_id) VALUES (%s, %s)",
                    [(item_id, factory_num) for item_id in ids]
                )

    def _attach_factory_order_product_links(self, conn, rows):
        if not rows:
            return
        factory_ids = [self._parse_int(row.get('id')) for row in rows if self._parse_int(row.get('id'))]
        if not factory_ids:
            for row in rows:
                row['order_product_ids'] = []
                row['order_product_skus'] = []
            return
        placeholders = ','.join(['%s'] * len(factory_ids))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT
                    opl.factory_id,
                    op.id AS order_product_id,
                    op.sku
                FROM order_product_factory_links opl
                JOIN order_products op ON op.id = opl.order_product_id
                WHERE opl.factory_id IN ({placeholders})
                ORDER BY op.sku ASC
                """,
                tuple(factory_ids)
            )
            link_rows = cur.fetchall() or []
        bucket = {}
        for item in link_rows:
            factory_id = self._parse_int(item.get('factory_id'))
            order_product_id = self._parse_int(item.get('order_product_id'))
            if not factory_id or not order_product_id:
                continue
            state = bucket.setdefault(factory_id, {'ids': [], 'skus': []})
            state['ids'].append(order_product_id)
            state['skus'].append(str(item.get('sku') or '').strip())
        for row in rows:
            row_id = self._parse_int(row.get('id'))
            state = bucket.get(row_id) or {'ids': [], 'skus': []}
            row['order_product_ids'] = state['ids']
            row['order_product_skus'] = state['skus']

    def handle_factory_stock_api(self, environ, method, start_response):
        """工厂在库库存 CRUD"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            user_id = self._get_session_user(environ)

            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                action = (query_params.get('action', [''])[0] or '').strip().lower()
                if action == '__disabled_filter_options__':
                    column = self._parse_int(query_params.get('column', ['0'])[0])
                    search = (query_params.get('q', [''])[0] or '').strip()
                    exact = _parse_yes_no(query_params.get('exact', ['0'])[0])
                    limit = max(1, min(200, self._parse_int(query_params.get('limit', ['120'])[0]) or 120))
                    filter_map = {
                        1: {
                            'value_expr': "CASE WHEN COALESCE(op.is_on_market, 0) = 1 THEN '1' ELSE '0' END",
                            'label_expr': "CASE WHEN COALESCE(op.is_on_market, 0) = 1 THEN '在市' ELSE '下市' END",
                        },
                        2: {
                            'value_expr': "NULLIF(TRIM(op.sku), '')",
                            'label_expr': "NULLIF(TRIM(op.sku), '')",
                        },
                        3: {
                            'value_expr': "NULLIF(TRIM(COALESCE(fm.representative_color, '')), '')",
                            'label_expr': "NULLIF(TRIM(COALESCE(fm.representative_color, '')), '')",
                        },
                        4: {
                            'value_expr': "NULLIF(TRIM(f.factory_name), '')",
                            'label_expr': "NULLIF(TRIM(f.factory_name), '')",
                        },
                        5: {
                            'value_expr': "NULLIF(TRIM(COALESCE(fc.order_no, '')), '')",
                            'label_expr': "NULLIF(TRIM(COALESCE(fc.order_no, '')), '')",
                        },
                        6: {
                            'value_expr': "NULLIF(TRIM(COALESCE(fc.contract_no, '')), '')",
                            'label_expr': "NULLIF(TRIM(COALESCE(fc.contract_no, '')), '')",
                        },
                        7: {
                            'value_expr': "CAST(fw.quantity AS CHAR)",
                            'label_expr': "CAST(fw.quantity AS CHAR)",
                        },
                        8: {
                            'value_expr': "DATE_FORMAT(fw.expected_completion_date, '%Y-%m-%d')",
                            'label_expr': "DATE_FORMAT(fw.expected_completion_date, '%Y-%m-%d')",
                        },
                        9: {
                            'value_expr': "DATE_FORMAT(fw.initial_expected_completion_date, '%Y-%m-%d')",
                            'label_expr': "DATE_FORMAT(fw.initial_expected_completion_date, '%Y-%m-%d')",
                        },
                        10: {
                            'value_expr': "CASE WHEN COALESCE(fw.is_completed, 0) = 1 THEN '1' ELSE '0' END",
                            'label_expr': "CASE WHEN COALESCE(fw.is_completed, 0) = 1 THEN '是' ELSE '否' END",
                        },
                        11: {
                            'value_expr': "DATE_FORMAT(fw.actual_completion_date, '%Y-%m-%d')",
                            'label_expr': "DATE_FORMAT(fw.actual_completion_date, '%Y-%m-%d')",
                        },
                        12: {
                            'value_expr': "NULLIF(TRIM(COALESCE(fw.notes, '')), '')",
                            'label_expr': "NULLIF(TRIM(COALESCE(fw.notes, '')), '')",
                        },
                        13: {
                            'value_expr': "DATE_FORMAT(fw.created_at, '%Y-%m-%d %H:%i:%s')",
                            'label_expr': "DATE_FORMAT(fw.created_at, '%Y-%m-%d %H:%i:%s')",
                        },
                        14: {
                            'value_expr': "DATE_FORMAT(COALESCE(fw.update_time, fw.updated_at), '%Y-%m-%d %H:%i:%s')",
                            'label_expr': "DATE_FORMAT(COALESCE(fw.update_time, fw.updated_at), '%Y-%m-%d %H:%i:%s')",
                        },
                    }
                    config = filter_map.get(column)
                    if not config:
                        return self.send_json({'status': 'error', 'message': '不支持的筛选列'}, start_response)
                    scope_clause, scope_params = self._factory_scope_clause('f.id', user_id, prefix='AND')
                    base_sql = f"""
                        SELECT {config['value_expr']} AS value, {config['label_expr']} AS label
                        FROM factory_wip_inventory fw
                        JOIN order_products op ON op.id = fw.order_product_id
                        JOIN logistics_factories f ON f.id = fw.factory_id
                        LEFT JOIN factory_contracts fc ON fc.id = fw.contract_id
                        LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                        WHERE 1=1 {scope_clause}
                    """
                    sql = f"""
                        SELECT value, label, COUNT(*) AS count
                        FROM (
                            {base_sql}
                        ) src
                        WHERE value IS NOT NULL AND value != ''
                    """
                    params = list(scope_params)
                    if search:
                        if exact:
                            sql += " AND (value = %s OR label = %s)"
                            params.extend([search, search])
                        else:
                            like = f"%{search}%"
                            sql += " AND (value LIKE %s OR label LIKE %s)"
                            params.extend([like, like])
                    sql += " GROUP BY value, label ORDER BY count DESC, label ASC LIMIT %s"
                    params.append(limit)
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(sql, tuple(params))
                            values = cur.fetchall() or []
                    return self.send_json({'status': 'success', 'column': column, 'values': values}, start_response)
                if action == 'download_all_stock_data':
                    scope_clause, scope_params = self._factory_scope_clause('f.id', user_id, prefix='AND')
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                f"""
                                SELECT f.factory_name, op.sku, fs.quantity, fs.notes, fs.updated_at
                                FROM factory_stock_inventory fs
                                JOIN order_products op ON op.id = fs.order_product_id
                                JOIN logistics_factories f ON f.id = fs.factory_id
                                WHERE 1=1 {scope_clause}
                                ORDER BY f.factory_name ASC, op.sku ASC
                                """,
                                scope_params
                            )
                            rows = cur.fetchall() or []

                    output = io.StringIO(newline='')
                    writer = csv.writer(output)
                    writer.writerow(['工厂', 'SKU', '数量', '备注', '更新时间'])
                    for row in rows:
                        updated_at = row.get('updated_at')
                        updated_text = ''
                        if updated_at:
                            updated_text = str(updated_at).replace('T', ' ')[:19]
                        writer.writerow([
                            row.get('factory_name') or '',
                            row.get('sku') or '',
                            self._parse_int(row.get('quantity')) or 0,
                            row.get('notes') or '',
                            updated_text,
                        ])

                    content = output.getvalue().encode('utf-8-sig')
                    filename = f"工厂在库库存_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                    headers = [
                        ('Content-Type', 'text/csv; charset=utf-8'),
                        ('Content-Disposition', f"attachment; filename*=UTF-8''{quote(filename)}"),
                        ('Content-Length', str(len(content))),
                    ]
                    start_response('200 OK', headers)
                    return [content]
                if action == 'options':
                    scope_ids = self._get_user_factory_scope_ids(user_id)
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            factory_clause, factory_params = self._factory_scope_clause('id', user_id, prefix='WHERE')
                            cur.execute(
                                f"SELECT id, factory_name FROM logistics_factories{factory_clause} ORDER BY factory_name ASC",
                                factory_params
                            )
                            factories = cur.fetchall() or []
                            sku_scope_ids = self._get_linked_order_product_ids(scope_ids)
                            if sku_scope_ids is None:
                                cur.execute("SELECT id, sku FROM order_products ORDER BY sku ASC")
                                order_products = cur.fetchall() or []
                            elif sku_scope_ids:
                                placeholders = ','.join(['%s'] * len(sku_scope_ids))
                                cur.execute(
                                    f"SELECT id, sku FROM order_products WHERE id IN ({placeholders}) ORDER BY sku ASC",
                                    tuple(sku_scope_ids)
                                )
                                order_products = cur.fetchall() or []
                            else:
                                order_products = []

                            factory_ids = [self._parse_int(item.get('id')) for item in factories if self._parse_int(item.get('id'))]
                            order_product_ids = [self._parse_int(item.get('id')) for item in order_products if self._parse_int(item.get('id'))]
                            links = []
                            if factory_ids and order_product_ids:
                                factory_placeholders = ','.join(['%s'] * len(factory_ids))
                                op_placeholders = ','.join(['%s'] * len(order_product_ids))
                                cur.execute(
                                    f"""
                                    SELECT order_product_id, factory_id
                                    FROM order_product_factory_links
                                    WHERE factory_id IN ({factory_placeholders})
                                      AND order_product_id IN ({op_placeholders})
                                    """,
                                    tuple(factory_ids) + tuple(order_product_ids)
                                )
                                links = [
                                    {
                                        'order_product_id': self._parse_int(item.get('order_product_id')),
                                        'factory_id': self._parse_int(item.get('factory_id'))
                                    }
                                    for item in (cur.fetchall() or [])
                                    if self._parse_int(item.get('order_product_id')) and self._parse_int(item.get('factory_id'))
                                ]

                    return self.send_json(
                        {'status': 'success', 'factories': factories, 'order_products': order_products, 'links': links},
                        start_response
                    )
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        scope_clause, scope_params = self._factory_scope_clause('f.id', user_id, prefix='AND')
                        if keyword:
                            cur.execute(
                                f"""
                                SELECT fs.id, fs.order_product_id, fs.factory_id, fs.quantity, fs.notes, fs.updated_at,
                                        op.sku, op.is_on_market, f.factory_name,
                                        fm.representative_color
                                FROM factory_stock_inventory fs
                                JOIN order_products op ON op.id = fs.order_product_id
                                JOIN logistics_factories f ON f.id = fs.factory_id
                                    LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                                WHERE (op.sku LIKE %s OR f.factory_name LIKE %s){scope_clause}
                                ORDER BY op.sku ASC, f.factory_name ASC
                                """,
                                (f"%{keyword}%", f"%{keyword}%") + scope_params
                            )
                        else:
                            cur.execute(
                                f"""
                                SELECT fs.id, fs.order_product_id, fs.factory_id, fs.quantity, fs.notes, fs.updated_at,
                                        op.sku, op.is_on_market, f.factory_name,
                                        fm.representative_color
                                FROM factory_stock_inventory fs
                                JOIN order_products op ON op.id = fs.order_product_id
                                JOIN logistics_factories f ON f.id = fs.factory_id
                                    LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                                WHERE 1=1 {scope_clause}
                                ORDER BY op.sku ASC, f.factory_name ASC
                                """,
                                scope_params
                            )
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            data = self._read_json_body(environ)
            if method == 'POST':
                op_id = self._parse_int(data.get('order_product_id'))
                factory_id = self._parse_int(data.get('factory_id'))
                quantity = max(0, self._parse_int(data.get('quantity')) or 0)
                notes = (data.get('notes') or '').strip() or None
                if not op_id or not factory_id:
                    return self.send_json({'status': 'error', 'message': '缺少 order_product_id 或 factory_id'}, start_response)
                if not self._factory_scope_contains(user_id, factory_id):
                    return self.send_json({'status': 'error', 'message': '无权限操作该工厂数据'}, start_response)
                if not self._order_product_allowed_for_factory(op_id, factory_id):
                    return self.send_json({'status': 'error', 'message': '该 SKU 未关联到该工厂，不可写入'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO factory_stock_inventory (order_product_id, factory_id, quantity, notes)
                            VALUES (%s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE quantity=%s, notes=%s
                            """,
                            (op_id, factory_id, quantity, notes, quantity, notes)
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                action = (query_params.get('action', [''])[0] or '').strip().lower()
                if action == 'bulk_update':
                    items = data.get('items') if isinstance(data, dict) else None
                    if not isinstance(items, list) or not items:
                        return self.send_json({'status': 'error', 'message': '缺少批量更新数据'}, start_response)

                    parsed_items = []
                    seen_ids = set()
                    for item in items:
                        if not isinstance(item, dict):
                            continue
                        item_id = self._parse_int(item.get('id'))
                        if not item_id or item_id in seen_ids:
                            continue
                        seen_ids.add(item_id)
                        quantity = max(0, self._parse_int(item.get('quantity')) or 0)
                        notes_raw = item.get('notes')
                        notes = ('' if notes_raw is None else str(notes_raw)).strip() or None
                        parsed_items.append({'id': item_id, 'quantity': quantity, 'notes': notes})

                    if not parsed_items:
                        return self.send_json({'status': 'error', 'message': '没有有效的批量更新项'}, start_response)

                    id_list = [item['id'] for item in parsed_items]
                    id_placeholders = ','.join(['%s'] * len(id_list))

                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                f"SELECT id, factory_id FROM factory_stock_inventory WHERE id IN ({id_placeholders})",
                                tuple(id_list)
                            )
                            existing_rows = cur.fetchall() or []
                            existing_map = {self._parse_int(row.get('id')): row for row in existing_rows if self._parse_int(row.get('id'))}

                            missing_ids = [item_id for item_id in id_list if item_id not in existing_map]
                            if missing_ids:
                                return self.send_json({'status': 'error', 'message': f'部分记录不存在: {missing_ids[:5]}'}, start_response)

                            denied_ids = []
                            for item_id in id_list:
                                factory_id = self._parse_int((existing_map.get(item_id) or {}).get('factory_id'))
                                if not self._factory_scope_contains(user_id, factory_id):
                                    denied_ids.append(item_id)
                            if denied_ids:
                                return self.send_json({'status': 'error', 'message': f'无权限操作以下记录: {denied_ids[:5]}'}, start_response)

                            qty_case = []
                            notes_case = []
                            sql_params = []
                            for item in parsed_items:
                                qty_case.append('WHEN %s THEN %s')
                                sql_params.extend([item['id'], item['quantity']])
                                notes_case.append('WHEN %s THEN %s')
                                sql_params.extend([item['id'], item['notes']])

                            update_sql = f"""
                                UPDATE factory_stock_inventory
                                SET
                                    quantity = CASE id {' '.join(qty_case)} ELSE quantity END,
                                    notes = CASE id {' '.join(notes_case)} ELSE notes END
                                WHERE id IN ({id_placeholders})
                            """
                            sql_params.extend(id_list)
                            cur.execute(update_sql, tuple(sql_params))

                    return self.send_json({'status': 'success', 'updated': len(parsed_items)}, start_response)

                item_id = self._parse_int(data.get('id'))
                quantity = max(0, self._parse_int(data.get('quantity')) or 0)
                notes = (data.get('notes') or '').strip() or None
                if not item_id:
                    return self.send_json({'status': 'error', 'message': '缺少 id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT factory_id FROM factory_stock_inventory WHERE id=%s", (item_id,))
                        existing = cur.fetchone() or {}
                        if not existing:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                        if not self._factory_scope_contains(user_id, existing.get('factory_id')):
                            return self.send_json({'status': 'error', 'message': '无权限操作该工厂数据'}, start_response)
                        cur.execute(
                            "UPDATE factory_stock_inventory SET quantity=%s, notes=%s WHERE id=%s",
                            (quantity, notes, item_id)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': '缺少 id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT factory_id FROM factory_stock_inventory WHERE id=%s", (item_id,))
                        existing = cur.fetchone() or {}
                        if not existing:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                        if not self._factory_scope_contains(user_id, existing.get('factory_id')):
                            return self.send_json({'status': 'error', 'message': '无权限操作该工厂数据'}, start_response)
                        cur.execute("DELETE FROM factory_stock_inventory WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_factory_wip_api(self, environ, method, start_response):
        """工厂在制库存 CRUD"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            user_id = self._get_session_user(environ)

            def _parse_yes_no(value):
                text = str(value if value is not None else '').strip().lower()
                if text in ('1', 'true', 'yes', 'y', '是'):
                    return 1
                return 0

            def _parse_date_text(value):
                text = (value or '').strip()
                if not text:
                    return None
                for fmt in ('%Y-%m-%d', '%Y/%m/%d'):
                    try:
                        return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
                    except Exception:
                        continue
                return None

            def _normalize_contract_no(value):
                text = str(value or '').strip()
                return text[:128] if text else None

            def _normalize_order_no(value):
                text = str(value or '').strip()
                return text[:128] if text else None

            def _resolve_contract_binding(cur, factory_id, contract_no, order_no):
                factory_id = self._parse_int(factory_id)
                if not factory_id:
                    raise ValueError('缺少工厂信息，无法匹配合同/订单编号')

                normalized_contract = _normalize_contract_no(contract_no)
                normalized_order = _normalize_order_no(order_no)
                if not normalized_contract and not normalized_order:
                    return None, None, None

                by_contract = None
                by_order = None

                if normalized_contract:
                    cur.execute(
                        """
                        SELECT id, contract_no, order_no
                        FROM factory_contracts
                        WHERE factory_id=%s AND contract_no=%s
                        LIMIT 1
                        """,
                        (factory_id, normalized_contract)
                    )
                    by_contract = cur.fetchone() or None

                if normalized_order:
                    cur.execute(
                        """
                        SELECT id, contract_no, order_no
                        FROM factory_contracts
                        WHERE factory_id=%s AND order_no=%s
                        LIMIT 1
                        """,
                        (factory_id, normalized_order)
                    )
                    by_order = cur.fetchone() or None

                if by_contract and by_order and int(by_contract.get('id') or 0) != int(by_order.get('id') or 0):
                    raise ValueError('合同编号与订单编号不匹配，请检查输入')

                existing = by_contract or by_order
                if existing:
                    existing_contract = _normalize_contract_no(existing.get('contract_no'))
                    existing_order = _normalize_order_no(existing.get('order_no'))
                    if normalized_contract and existing_contract and normalized_contract != existing_contract:
                        raise ValueError('合同编号与已存在映射不一致')
                    if normalized_order and existing_order and normalized_order != existing_order:
                        raise ValueError('订单编号与已存在映射不一致')
                    if normalized_contract and not normalized_order and not existing_order:
                        raise ValueError('该合同编号尚未绑定订单编号，请同时填写订单编号')
                    if normalized_order and not normalized_contract and not existing_contract:
                        raise ValueError('该订单编号尚未绑定合同编号，请同时填写合同编号')
                    return (
                        int(existing.get('id')),
                        existing_contract or normalized_contract,
                        existing_order or normalized_order,
                    )

                if not (normalized_contract and normalized_order):
                    raise ValueError('新增合同/订单映射时需同时填写合同编号与订单编号')

                cur.execute(
                    """
                    INSERT INTO factory_contracts (factory_id, contract_no, order_no)
                    VALUES (%s, %s, %s)
                    """,
                    (factory_id, normalized_contract, normalized_order)
                )
                return int(cur.lastrowid), normalized_contract, normalized_order

            def _cleanup_orphan_contracts(cur, contract_ids):
                ids = [self._parse_int(v) for v in (contract_ids or []) if self._parse_int(v)]
                if not ids:
                    return
                placeholders = ','.join(['%s'] * len(ids))
                cur.execute(
                    f"""
                    DELETE fc
                    FROM factory_contracts fc
                    LEFT JOIN factory_wip_inventory fw ON fw.contract_id = fc.id
                    WHERE fc.id IN ({placeholders})
                      AND fw.id IS NULL
                    """,
                    tuple(ids)
                )

            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                action = (query_params.get('action', [''])[0] or '').strip().lower()
                if action == 'download_unfinished_data':
                    scope_clause, scope_params = self._factory_scope_clause('f.id', user_id, prefix='AND')
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                f"""
                                SELECT
                                    f.factory_name,
                                    op.sku,
                                    fw.quantity,
                                    fc.order_no,
                                    fc.contract_no,
                                    fw.expected_completion_date,
                                    fw.initial_expected_completion_date,
                                    fw.notes,
                                    COALESCE(fw.update_time, fw.updated_at) AS update_time
                                FROM factory_wip_inventory fw
                                JOIN order_products op ON op.id = fw.order_product_id
                                JOIN logistics_factories f ON f.id = fw.factory_id
                                LEFT JOIN factory_contracts fc ON fc.id = fw.contract_id
                                WHERE COALESCE(fw.is_completed, 0) = 0 {scope_clause}
                                ORDER BY f.factory_name ASC, op.sku ASC, fw.expected_completion_date ASC
                                """,
                                scope_params
                            )
                            rows = cur.fetchall() or []

                    output = io.StringIO(newline='')
                    writer = csv.writer(output)
                    writer.writerow(['工厂', 'SKU', '数量', '订单号', '合同编号', '预计完工日期', '最初预计完工日期', '备注', '更新时间'])
                    for row in rows:
                        expected_completion = row.get('expected_completion_date')
                        initial_expected = row.get('initial_expected_completion_date')
                        update_time = row.get('update_time')
                        writer.writerow([
                            row.get('factory_name') or '',
                            row.get('sku') or '',
                            self._parse_int(row.get('quantity')) or 0,
                            row.get('order_no') or '',
                            row.get('contract_no') or '',
                            str(expected_completion)[:10] if expected_completion else '',
                            str(initial_expected)[:10] if initial_expected else '',
                            row.get('notes') or '',
                            str(update_time).replace('T', ' ')[:19] if update_time else '',
                        ])

                    content = output.getvalue().encode('utf-8-sig')
                    filename = f"工厂在制未完工库存_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                    headers = [
                        ('Content-Type', 'text/csv; charset=utf-8'),
                        ('Content-Disposition', f"attachment; filename*=UTF-8''{quote(filename)}"),
                        ('Content-Length', str(len(content))),
                    ]
                    start_response('200 OK', headers)
                    return [content]
                if action == 'filter_options':
                    column = self._parse_int(query_params.get('column', ['0'])[0])
                    search = (query_params.get('q', [''])[0] or '').strip()
                    exact = _parse_yes_no(query_params.get('exact', ['0'])[0])
                    limit = max(1, min(200, self._parse_int(query_params.get('limit', ['120'])[0]) or 120))
                    filter_map = {
                        1: {
                            'value_expr': "NULLIF(TRIM(f.factory_name), '')",
                            'label_expr': "NULLIF(TRIM(f.factory_name), '')",
                        },
                        2: {
                            'value_expr': "NULLIF(TRIM(COALESCE(fc.order_no, '')), '')",
                            'label_expr': "NULLIF(TRIM(COALESCE(fc.order_no, '')), '')",
                        },
                        3: {
                            'value_expr': "NULLIF(TRIM(COALESCE(fc.contract_no, '')), '')",
                            'label_expr': "NULLIF(TRIM(COALESCE(fc.contract_no, '')), '')",
                        },
                        4: {
                            'value_expr': "NULLIF(TRIM(COALESCE(fm.representative_color, '')), '')",
                            'label_expr': "NULLIF(TRIM(COALESCE(fm.representative_color, '')), '')",
                        },
                        5: {
                            'value_expr': "NULLIF(TRIM(op.sku), '')",
                            'label_expr': "NULLIF(TRIM(op.sku), '')",
                        },
                        6: {
                            'value_expr': "CASE WHEN COALESCE(op.is_on_market, 0) = 1 THEN '1' ELSE '0' END",
                            'label_expr': "CASE WHEN COALESCE(op.is_on_market, 0) = 1 THEN '在市' ELSE '下市' END",
                        },
                        7: {
                            'value_expr': "CAST(fw.quantity AS CHAR)",
                            'label_expr': "CAST(fw.quantity AS CHAR)",
                        },
                        8: {
                            'value_expr': "DATE_FORMAT(fw.expected_completion_date, '%%Y-%%m-%%d')",
                            'label_expr': "DATE_FORMAT(fw.expected_completion_date, '%%Y-%%m-%%d')",
                        },
                        9: {
                            'value_expr': "DATE_FORMAT(fw.initial_expected_completion_date, '%%Y-%%m-%%d')",
                            'label_expr': "DATE_FORMAT(fw.initial_expected_completion_date, '%%Y-%%m-%%d')",
                        },
                        10: {
                            'value_expr': "CASE WHEN COALESCE(fw.is_completed, 0) = 1 THEN '1' ELSE '0' END",
                            'label_expr': "CASE WHEN COALESCE(fw.is_completed, 0) = 1 THEN '是' ELSE '否' END",
                        },
                        11: {
                            'value_expr': "DATE_FORMAT(fw.actual_completion_date, '%%Y-%%m-%%d')",
                            'label_expr': "DATE_FORMAT(fw.actual_completion_date, '%%Y-%%m-%%d')",
                        },
                        12: {
                            'value_expr': "NULLIF(TRIM(COALESCE(fw.notes, '')), '')",
                            'label_expr': "NULLIF(TRIM(COALESCE(fw.notes, '')), '')",
                        },
                        13: {
                            'value_expr': "DATE_FORMAT(fw.created_at, '%%Y-%%m-%%d %%H:%%i:%%s')",
                            'label_expr': "DATE_FORMAT(fw.created_at, '%%Y-%%m-%%d %%H:%%i:%%s')",
                        },
                        14: {
                            'value_expr': "DATE_FORMAT(COALESCE(fw.update_time, fw.updated_at), '%%Y-%%m-%%d %%H:%%i:%%s')",
                            'label_expr': "DATE_FORMAT(COALESCE(fw.update_time, fw.updated_at), '%%Y-%%m-%%d %%H:%%i:%%s')",
                        },
                    }
                    config = filter_map.get(column)
                    if not config:
                        return self.send_json({'status': 'error', 'message': '不支持的筛选列'}, start_response)
                    scope_clause, scope_params = self._factory_scope_clause('f.id', user_id, prefix='AND')
                    base_sql = f"""
                        SELECT {config['value_expr']} AS value, {config['label_expr']} AS label
                        FROM factory_wip_inventory fw
                        JOIN order_products op ON op.id = fw.order_product_id
                        JOIN logistics_factories f ON f.id = fw.factory_id
                        LEFT JOIN factory_contracts fc ON fc.id = fw.contract_id
                        LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                        WHERE 1=1 {scope_clause}
                    """
                    sql = f"""
                        SELECT value, label, COUNT(*) AS count
                        FROM (
                            {base_sql}
                        ) src
                        WHERE value IS NOT NULL AND value != ''
                    """
                    params = list(scope_params)
                    if search:
                        if exact:
                            sql += " AND (value = %s OR label = %s)"
                            params.extend([search, search])
                        else:
                            like = f"%{search}%"
                            sql += " AND (value LIKE %s OR label LIKE %s)"
                            params.extend([like, like])
                    sql += " GROUP BY value, label ORDER BY count DESC, label ASC LIMIT %s"
                    params.append(limit)
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(sql, tuple(params))
                            values = cur.fetchall() or []
                    return self.send_json({'status': 'success', 'column': column, 'values': values}, start_response)
                if action == 'options':
                    scope_ids = self._get_user_factory_scope_ids(user_id)
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            factory_clause, factory_params = self._factory_scope_clause('id', user_id, prefix='WHERE')
                            cur.execute(
                                f"SELECT id, factory_name FROM logistics_factories{factory_clause} ORDER BY factory_name ASC",
                                factory_params
                            )
                            factories = cur.fetchall() or []
                            sku_scope_ids = self._get_linked_order_product_ids(scope_ids)
                            if sku_scope_ids is None:
                                cur.execute("SELECT id, sku FROM order_products ORDER BY sku ASC")
                                order_products = cur.fetchall() or []
                            elif sku_scope_ids:
                                placeholders = ','.join(['%s'] * len(sku_scope_ids))
                                cur.execute(
                                    f"SELECT id, sku FROM order_products WHERE id IN ({placeholders}) ORDER BY sku ASC",
                                    tuple(sku_scope_ids)
                                )
                                order_products = cur.fetchall() or []
                            else:
                                order_products = []

                            factory_ids = [self._parse_int(item.get('id')) for item in factories if self._parse_int(item.get('id'))]
                            order_product_ids = [self._parse_int(item.get('id')) for item in order_products if self._parse_int(item.get('id'))]
                            links = []
                            if factory_ids and order_product_ids:
                                factory_placeholders = ','.join(['%s'] * len(factory_ids))
                                op_placeholders = ','.join(['%s'] * len(order_product_ids))
                                cur.execute(
                                    f"""
                                    SELECT order_product_id, factory_id
                                    FROM order_product_factory_links
                                    WHERE factory_id IN ({factory_placeholders})
                                      AND order_product_id IN ({op_placeholders})
                                    """,
                                    tuple(factory_ids) + tuple(order_product_ids)
                                )
                                links = [
                                    {
                                        'order_product_id': self._parse_int(item.get('order_product_id')),
                                        'factory_id': self._parse_int(item.get('factory_id'))
                                    }
                                    for item in (cur.fetchall() or [])
                                    if self._parse_int(item.get('order_product_id')) and self._parse_int(item.get('factory_id'))
                                ]

                            scope_clause, scope_params = self._factory_scope_clause('f.id', user_id, prefix='WHERE')
                            cur.execute(
                                f"""
                                SELECT fc.id, fc.factory_id, fc.contract_no, fc.order_no, f.factory_name
                                FROM factory_contracts fc
                                JOIN logistics_factories f ON f.id = fc.factory_id
                                {scope_clause}
                                ORDER BY f.factory_name ASC, fc.contract_no ASC
                                """,
                                scope_params
                            )
                            contracts = cur.fetchall() or []

                    return self.send_json(
                        {
                            'status': 'success',
                            'factories': factories,
                            'order_products': order_products,
                            'links': links,
                            'contracts': contracts
                        },
                        start_response
                    )
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        scope_clause, scope_params = self._factory_scope_clause('f.id', user_id, prefix='AND')
                        if keyword:
                            cur.execute(
                                f"""
                                SELECT fw.id, fw.order_product_id, fw.factory_id, fw.contract_id, fw.quantity,
                                        fw.expected_completion_date, fw.is_completed, fw.actual_completion_date,
                                        fw.initial_expected_completion_date,
                                        fw.notes, fw.created_at, fw.updated_at, fw.update_time,
                                    fc.order_no,
                                        fc.contract_no,
                                        op.sku, op.is_on_market, f.factory_name,
                                        fm.representative_color
                                FROM factory_wip_inventory fw
                                JOIN order_products op ON op.id = fw.order_product_id
                                JOIN logistics_factories f ON f.id = fw.factory_id
                                    LEFT JOIN factory_contracts fc ON fc.id = fw.contract_id
                                    LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                                WHERE (op.sku LIKE %s OR f.factory_name LIKE %s){scope_clause}
                                ORDER BY op.sku ASC, f.factory_name ASC, fw.expected_completion_date ASC
                                """,
                                (f"%{keyword}%", f"%{keyword}%") + scope_params
                            )
                        else:
                            cur.execute(
                                f"""
                                SELECT fw.id, fw.order_product_id, fw.factory_id, fw.contract_id, fw.quantity,
                                        fw.expected_completion_date, fw.is_completed, fw.actual_completion_date,
                                        fw.initial_expected_completion_date,
                                        fw.notes, fw.created_at, fw.updated_at, fw.update_time,
                                    fc.order_no,
                                        fc.contract_no,
                                        op.sku, op.is_on_market, f.factory_name,
                                        fm.representative_color
                                FROM factory_wip_inventory fw
                                JOIN order_products op ON op.id = fw.order_product_id
                                JOIN logistics_factories f ON f.id = fw.factory_id
                                    LEFT JOIN factory_contracts fc ON fc.id = fw.contract_id
                                    LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                                WHERE 1=1 {scope_clause}
                                ORDER BY op.sku ASC, f.factory_name ASC, fw.expected_completion_date ASC
                                """,
                                scope_params
                            )
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            data = self._read_json_body(environ)
            if method == 'POST':
                op_id = self._parse_int(data.get('order_product_id'))
                factory_id = self._parse_int(data.get('factory_id'))
                quantity = max(0, self._parse_int(data.get('quantity')) or 0)
                order_no = _normalize_order_no(data.get('order_no'))
                notes = (data.get('notes') or '').strip() or None
                contract_no = _normalize_contract_no(data.get('contract_no'))
                expected_date = _parse_date_text(data.get('expected_completion_date'))
                is_completed = _parse_yes_no(data.get('is_completed'))
                actual_completion_date = _parse_date_text(data.get('actual_completion_date'))
                if is_completed and not actual_completion_date:
                    actual_completion_date = datetime.now().strftime('%Y-%m-%d')
                if not is_completed:
                    actual_completion_date = None
                if not op_id or not factory_id:
                    return self.send_json({'status': 'error', 'message': '缺少 order_product_id 或 factory_id'}, start_response)
                if not self._factory_scope_contains(user_id, factory_id):
                    return self.send_json({'status': 'error', 'message': '无权限操作该工厂数据'}, start_response)
                if not self._order_product_allowed_for_factory(op_id, factory_id):
                    return self.send_json({'status': 'error', 'message': '该 SKU 未关联到该工厂，不可写入'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        contract_id, contract_no, order_no = _resolve_contract_binding(cur, factory_id, contract_no, order_no)
                        cur.execute(
                            """
                            INSERT INTO factory_wip_inventory (
                                order_product_id, factory_id, contract_id, quantity,
                                expected_completion_date, initial_expected_completion_date,
                                is_completed, actual_completion_date, notes
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                op_id,
                                factory_id,
                                contract_id,
                                quantity,
                                expected_date,
                                expected_date,
                                is_completed,
                                actual_completion_date,
                                notes
                            )
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                action = (query_params.get('action', [''])[0] or '').strip().lower()
                if action == 'bulk_update':
                    items = data.get('items') if isinstance(data, dict) else None
                    if not isinstance(items, list) or not items:
                        return self.send_json({'status': 'error', 'message': '缺少批量更新数据'}, start_response)

                    parsed_items = []
                    seen_ids = set()
                    for item in items:
                        if not isinstance(item, dict):
                            continue
                        item_id = self._parse_int(item.get('id'))
                        if not item_id or item_id in seen_ids:
                            continue
                        seen_ids.add(item_id)
                        quantity = max(0, self._parse_int(item.get('quantity')) or 0)
                        expected_date = _parse_date_text(item.get('expected_completion_date'))
                        is_completed = _parse_yes_no(item.get('is_completed'))
                        order_no = _normalize_order_no(item.get('order_no')) if ('order_no' in item) else None
                        has_order_no = 'order_no' in item
                        contract_no = _normalize_contract_no(item.get('contract_no')) if ('contract_no' in item) else None
                        has_contract_no = 'contract_no' in item
                        actual_completion_date = _parse_date_text(item.get('actual_completion_date'))
                        notes_raw = item.get('notes')
                        notes = ('' if notes_raw is None else str(notes_raw)).strip() or None
                        if is_completed and not actual_completion_date:
                            actual_completion_date = datetime.now().strftime('%Y-%m-%d')
                        if not is_completed:
                            actual_completion_date = None
                        parsed_items.append({
                            'id': item_id,
                            'quantity': quantity,
                            'order_no': order_no,
                            'has_order_no': has_order_no,
                            'contract_no': contract_no,
                            'has_contract_no': has_contract_no,
                            'expected_completion_date': expected_date,
                            'is_completed': is_completed,
                            'actual_completion_date': actual_completion_date,
                            'notes': notes,
                        })

                    if not parsed_items:
                        return self.send_json({'status': 'error', 'message': '没有有效的批量更新项'}, start_response)

                    id_list = [item['id'] for item in parsed_items]
                    id_placeholders = ','.join(['%s'] * len(id_list))

                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                f"SELECT id, factory_id, contract_id FROM factory_wip_inventory WHERE id IN ({id_placeholders})",
                                tuple(id_list)
                            )
                            existing_rows = cur.fetchall() or []
                            existing_map = {self._parse_int(row.get('id')): row for row in existing_rows if self._parse_int(row.get('id'))}

                            missing_ids = [item_id for item_id in id_list if item_id not in existing_map]
                            if missing_ids:
                                return self.send_json({'status': 'error', 'message': f'记录不存在: {missing_ids[0]}'}, start_response)

                            for item in parsed_items:
                                existing_row = existing_map.get(item['id']) or {}
                                if not self._factory_scope_contains(user_id, existing_row.get('factory_id')):
                                    return self.send_json({'status': 'error', 'message': '无权限操作该工厂数据'}, start_response)

                            for item in parsed_items:
                                contract_id = None
                                if item.get('has_contract_no') or item.get('has_order_no'):
                                    existing_row = existing_map.get(item['id']) or {}
                                    contract_id, _, _ = _resolve_contract_binding(
                                        cur,
                                        existing_row.get('factory_id'),
                                        item.get('contract_no') if item.get('has_contract_no') else None,
                                        item.get('order_no') if item.get('has_order_no') else None,
                                    )
                                else:
                                    existing_row = existing_map.get(item['id']) or {}
                                    contract_id = self._parse_int(existing_row.get('contract_id'))
                                cur.execute(
                                    """
                                    UPDATE factory_wip_inventory
                                    SET
                                        quantity=%s,
                                        contract_id=%s,
                                        expected_completion_date=%s,
                                        is_completed=%s,
                                        actual_completion_date=%s,
                                        notes=%s
                                    WHERE id=%s
                                    """,
                                    (
                                        item['quantity'],
                                        contract_id,
                                        item['expected_completion_date'],
                                        item['is_completed'],
                                        item['actual_completion_date'],
                                        item['notes'],
                                        item['id'],
                                    )
                                )

                    return self.send_json({'status': 'success', 'updated': len(parsed_items)}, start_response)

                item_id = self._parse_int(data.get('id'))
                quantity = max(0, self._parse_int(data.get('quantity')) or 0)
                order_no = _normalize_order_no(data.get('order_no'))
                notes = (data.get('notes') or '').strip() or None
                contract_no = _normalize_contract_no(data.get('contract_no'))
                expected_date = _parse_date_text(data.get('expected_completion_date'))
                is_completed = _parse_yes_no(data.get('is_completed'))
                add_to_factory_stock = _parse_yes_no(data.get('add_to_factory_stock'))
                actual_completion_date = _parse_date_text(data.get('actual_completion_date'))
                if is_completed and not actual_completion_date:
                    actual_completion_date = datetime.now().strftime('%Y-%m-%d')
                if not is_completed:
                    actual_completion_date = None
                if not item_id:
                    return self.send_json({'status': 'error', 'message': '缺少 id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT order_product_id, factory_id, quantity, is_completed
                            FROM factory_wip_inventory
                            WHERE id=%s
                            """,
                            (item_id,)
                        )
                        existing = cur.fetchone() or {}
                        if not existing:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                        if not self._factory_scope_contains(user_id, existing.get('factory_id')):
                            return self.send_json({'status': 'error', 'message': '无权限操作该工厂数据'}, start_response)

                        previous_completed = int(existing.get('is_completed') or 0)
                        if add_to_factory_stock:
                            if not is_completed:
                                return self.send_json({'status': 'error', 'message': '仅完工后可新增到工厂在库'}, start_response)
                            if previous_completed == 1:
                                return self.send_json({'status': 'error', 'message': '该记录已完工，已阻止重复新增到工厂在库'}, start_response)

                        cur.execute(
                            """
                            UPDATE factory_wip_inventory
                            SET quantity=%s, contract_id=%s, expected_completion_date=%s, is_completed=%s, actual_completion_date=%s, notes=%s
                            WHERE id=%s
                            """,
                            (
                                quantity,
                                _resolve_contract_binding(cur, existing.get('factory_id'), contract_no, order_no)[0],
                                expected_date,
                                is_completed,
                                actual_completion_date,
                                notes,
                                item_id
                            )
                        )

                        if add_to_factory_stock:
                            transfer_qty = max(0, quantity)
                            if transfer_qty <= 0:
                                return self.send_json({'status': 'error', 'message': '数量为0，无法新增到工厂在库'}, start_response)
                            order_product_id = self._parse_int(existing.get('order_product_id'))
                            factory_id = self._parse_int(existing.get('factory_id'))
                            if not order_product_id or not factory_id:
                                return self.send_json({'status': 'error', 'message': '缺少下单SKU或工厂，无法新增到工厂在库'}, start_response)
                            if not self._order_product_allowed_for_factory(order_product_id, factory_id):
                                return self.send_json({'status': 'error', 'message': '该 SKU 未关联到该工厂，不可新增到工厂在库'}, start_response)
                            cur.execute(
                                """
                                INSERT INTO factory_stock_inventory (order_product_id, factory_id, quantity, notes)
                                VALUES (%s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE
                                  quantity = quantity + VALUES(quantity),
                                  notes = COALESCE(VALUES(notes), notes)
                                """,
                                (order_product_id, factory_id, transfer_qty, notes)
                            )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                action = (query_params.get('action', [''])[0] or '').strip().lower()
                if action == 'bulk_delete':
                    raw_ids = data.get('ids') if isinstance(data, dict) else None
                    id_list = self._normalize_id_list_local(raw_ids)
                    if not id_list:
                        return self.send_json({'status': 'error', 'message': '缺少有效 ids'}, start_response)
                    placeholders = ','.join(['%s'] * len(id_list))
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                f"SELECT id, factory_id, contract_id FROM factory_wip_inventory WHERE id IN ({placeholders})",
                                tuple(id_list)
                            )
                            rows = cur.fetchall() or []
                            row_map = {self._parse_int(r.get('id')): r for r in rows if self._parse_int(r.get('id'))}
                            missing = [item_id for item_id in id_list if item_id not in row_map]
                            if missing:
                                return self.send_json({'status': 'error', 'message': f'记录不存在: {missing[0]}'}, start_response)
                            denied = []
                            for item_id in id_list:
                                row = row_map.get(item_id) or {}
                                if not self._factory_scope_contains(user_id, row.get('factory_id')):
                                    denied.append(item_id)
                            if denied:
                                return self.send_json({'status': 'error', 'message': f'无权限操作以下记录: {denied[:5]}'}, start_response)

                            contract_ids = [self._parse_int((row_map.get(item_id) or {}).get('contract_id')) for item_id in id_list]
                            cur.execute(f"DELETE FROM factory_wip_inventory WHERE id IN ({placeholders})", tuple(id_list))
                            _cleanup_orphan_contracts(cur, contract_ids)
                    return self.send_json({'status': 'success', 'deleted': len(id_list)}, start_response)

                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': '缺少 id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT factory_id, contract_id FROM factory_wip_inventory WHERE id=%s", (item_id,))
                        existing = cur.fetchone() or {}
                        if not existing:
                            return self.send_json({'status': 'error', 'message': '记录不存在'}, start_response)
                        if not self._factory_scope_contains(user_id, existing.get('factory_id')):
                            return self.send_json({'status': 'error', 'message': '无权限操作该工厂数据'}, start_response)
                        cur.execute("DELETE FROM factory_wip_inventory WHERE id=%s", (item_id,))
                        _cleanup_orphan_contracts(cur, [existing.get('contract_id')])
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_factory_stock_template_api(self, environ, method, start_response):
        try:
            user_id = self._get_session_user(environ)
            scope_ids = self._get_user_factory_scope_ids(user_id)
            query_params = parse_qs(environ.get('QUERY_STRING', ''))

            def _append_ids_from_value(container, value):
                if isinstance(value, list):
                    for v in value:
                        _append_ids_from_value(container, v)
                    return
                text = str(value or '').strip()
                if not text:
                    return
                for token in re.split(r'[\s,，;；]+', text):
                    item_id = self._parse_int(token)
                    if item_id and item_id not in container:
                        container.append(item_id)

            selected_ids = []
            _append_ids_from_value(selected_ids, query_params.get('ids', [''])[0])
            if method == 'POST':
                body = self._read_json_body(environ) or {}
                _append_ids_from_value(selected_ids, body.get('ids'))

            if method not in ('GET', 'POST'):
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = 'factory_stock'

            headers = ['SKU', '工厂', '数量', '备注']
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value='工厂在库库存导入模板')
            title_cell.fill = PatternFill(start_color='A8B9A5', end_color='A8B9A5', fill_type='solid')
            title_cell.font = Font(bold=True, color='2A2420')
            title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.append(headers)
            for cell in ws[2]:
                cell.fill = PatternFill(start_color='DDE7DB', end_color='DDE7DB', fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            sample_row = ['示例SKU（请勿导入）', '示例工厂（请勿导入）', 100, '示例行（请勿导入，此行仅演示格式）']
            ws.append(sample_row)
            for cell in ws[3]:
                cell.fill = PatternFill(start_color='ECECEC', end_color='ECECEC', fill_type='solid')
                cell.font = Font(italic=True, color='7B8088')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            widths = [24, 24, 10, 28]
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            option_sheet = wb.create_sheet('_options')
            option_sheet.sheet_state = 'hidden'
            option_sheet.append(['factory_name', 'sku'])

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    factory_clause, factory_params = self._factory_scope_clause('id', user_id, prefix='WHERE')
                    cur.execute(
                        f"SELECT factory_name FROM logistics_factories{factory_clause} ORDER BY factory_name ASC",
                        factory_params
                    )
                    factories = [str(r.get('factory_name') or '').strip() for r in (cur.fetchall() or []) if r.get('factory_name')]
                    sku_scope_ids = self._get_linked_order_product_ids(scope_ids)
                    if sku_scope_ids is None:
                        cur.execute("SELECT sku FROM order_products ORDER BY sku ASC")
                        skus = [str(r.get('sku') or '').strip() for r in (cur.fetchall() or []) if r.get('sku')]
                    elif sku_scope_ids:
                        placeholders = ','.join(['%s'] * len(sku_scope_ids))
                        cur.execute(
                            f"SELECT sku FROM order_products WHERE id IN ({placeholders}) ORDER BY sku ASC",
                            tuple(sku_scope_ids)
                        )
                        skus = [str(r.get('sku') or '').strip() for r in (cur.fetchall() or []) if r.get('sku')]
                    else:
                        skus = []

                    selected_rows = []
                    if selected_ids:
                        placeholders = ','.join(['%s'] * len(selected_ids))
                        scope_clause, scope_params = self._factory_scope_clause('f.id', user_id, prefix='AND')
                        cur.execute(
                            f"""
                            SELECT
                                fs.id,
                                op.sku,
                                f.factory_name,
                                fs.quantity,
                                fs.notes
                            FROM factory_stock_inventory fs
                            JOIN order_products op ON op.id = fs.order_product_id
                            JOIN logistics_factories f ON f.id = fs.factory_id
                            WHERE fs.id IN ({placeholders}) {scope_clause}
                            ORDER BY op.sku ASC, f.factory_name ASC
                            """,
                            tuple(selected_ids) + scope_params
                        )
                        selected_rows = cur.fetchall() or []

            max_len = max(len(factories), len(skus), 1)
            for i in range(max_len):
                option_sheet.append([
                    factories[i] if i < len(factories) else '',
                    skus[i] if i < len(skus) else ''
                ])

            max_row = 400
            for row in range(3, max_row + 1):
                ws[f'D{row}'].number_format = 'yyyy-mm-dd'
                ws[f'F{row}'].number_format = 'yyyy-mm-dd'
            if factories:
                dv_factory = DataValidation(type='list', formula1=f"='_options'!$A$2:$A${len(factories) + 1}", allow_blank=False)
                ws.add_data_validation(dv_factory)
                for row in range(4, max_row + 1):
                    dv_factory.add(f'B{row}')
            if skus:
                dv_sku = DataValidation(type='list', formula1=f"='_options'!$B$2:$B${len(skus) + 1}", allow_blank=False)
                ws.add_data_validation(dv_sku)
                for row in range(4, max_row + 1):
                    dv_sku.add(f'A{row}')

            if selected_rows:
                write_row = 4
                for item in selected_rows:
                    ws.cell(row=write_row, column=1, value=str(item.get('sku') or '').strip())
                    ws.cell(row=write_row, column=2, value=str(item.get('factory_name') or '').strip())
                    ws.cell(row=write_row, column=3, value=int(item.get('quantity') or 0))
                    ws.cell(row=write_row, column=4, value=str(item.get('notes') or '').strip())
                    write_row += 1

            ws.freeze_panes = 'A4'
            return self._send_excel_workbook(wb, 'factory_stock_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_factory_stock_import_api(self, environ, method, start_response):
        try:
            user_id = self._get_session_user(environ)
            scope_ids = self._get_user_factory_scope_ids(user_id)
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            header_row = 2 if str(ws.cell(row=1, column=1).value or '').strip().startswith('工厂在库库存导入模板') else 1
            header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), tuple())
            headers = [str(value or '').strip() for value in header_values]
            header_map = {name: idx for idx, name in enumerate(headers)}

            for required in ('SKU', '工厂', '数量'):
                if required not in header_map:
                    return self.send_json({'status': 'error', 'message': f'模板缺少列: {required}'}, start_response)

            def get_cell(row, name):
                idx = header_map.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            created = 0
            updated = 0
            unchanged = 0
            errors = []

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    sku_scope_ids = self._get_linked_order_product_ids(scope_ids)
                    if sku_scope_ids is None:
                        cur.execute("SELECT id, sku FROM order_products")
                        sku_map = {str(r.get('sku') or '').strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}
                    elif sku_scope_ids:
                        placeholders = ','.join(['%s'] * len(sku_scope_ids))
                        cur.execute(
                            f"SELECT id, sku FROM order_products WHERE id IN ({placeholders})",
                            tuple(sku_scope_ids)
                        )
                        sku_map = {str(r.get('sku') or '').strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}
                    else:
                        sku_map = {}
                    factory_clause, factory_params = self._factory_scope_clause('id', user_id, prefix='WHERE')
                    cur.execute(
                        f"SELECT id, factory_name FROM logistics_factories{factory_clause}",
                        factory_params
                    )
                    factory_map = {str(r.get('factory_name') or '').strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}

                    normalized_rows = []
                    pair_keys = set()
                    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                        if not any(value is not None and str(value).strip() for value in row):
                            continue
                        row_join = '|'.join([str(v or '').strip() for v in row])
                        if '示例' in row_join and '勿导入' in row_join:
                            continue
                        try:
                            sku = str(get_cell(row, 'SKU') or '').strip()
                            factory_name = str(get_cell(row, '工厂') or '').strip()
                            quantity = self._parse_int(get_cell(row, '数量'))
                            notes = str(get_cell(row, '备注') or '').strip() or None
                            if not sku or not factory_name or quantity is None:
                                raise ValueError('SKU/工厂/数量不能为空，且数量需为整数')
                            order_product_id = sku_map.get(sku)
                            factory_id = factory_map.get(factory_name)
                            if not order_product_id:
                                raise ValueError(f'未找到SKU: {sku}')
                            if not factory_id:
                                raise ValueError(f'未找到工厂: {factory_name}')
                            if not self._order_product_allowed_for_factory(order_product_id, factory_id):
                                raise ValueError(f'SKU {sku} 未关联工厂 {factory_name}')
                            quantity = max(0, int(quantity))
                            normalized_rows.append((order_product_id, factory_id, quantity, notes))
                            pair_keys.add((order_product_id, factory_id))
                        except Exception as row_error:
                            errors.append({'row': row_idx, 'error': str(row_error)})

                    existing_map = {}
                    if pair_keys:
                        op_ids = sorted({k[0] for k in pair_keys})
                        factory_ids = sorted({k[1] for k in pair_keys})
                        op_placeholders = ','.join(['%s'] * len(op_ids))
                        fa_placeholders = ','.join(['%s'] * len(factory_ids))
                        cur.execute(
                            f"""
                            SELECT id, order_product_id, factory_id, quantity, notes
                            FROM factory_stock_inventory
                            WHERE order_product_id IN ({op_placeholders})
                              AND factory_id IN ({fa_placeholders})
                            """,
                            tuple(op_ids) + tuple(factory_ids)
                        )
                        for ex in (cur.fetchall() or []):
                            key = (int(ex.get('order_product_id')), int(ex.get('factory_id')))
                            existing_map[key] = {
                                'id': int(ex.get('id')),
                                'quantity': int(ex.get('quantity') or 0),
                                'notes': (ex.get('notes') or '').strip() or None
                            }

                    for order_product_id, factory_id, quantity, notes in normalized_rows:
                        key = (order_product_id, factory_id)
                        ex = existing_map.get(key)
                        if ex:
                            if ex.get('quantity') == quantity and (ex.get('notes') or None) == (notes or None):
                                unchanged += 1
                                continue
                            cur.execute(
                                "UPDATE factory_stock_inventory SET quantity=%s, notes=%s WHERE id=%s",
                                (quantity, notes, ex['id'])
                            )
                            updated += 1
                        else:
                            cur.execute(
                                "INSERT INTO factory_stock_inventory (order_product_id, factory_id, quantity, notes) VALUES (%s, %s, %s, %s)",
                                (order_product_id, factory_id, quantity, notes)
                            )
                            created += 1

            return self.send_json({'status': 'success', 'created': created, 'updated': updated, 'unchanged': unchanged, 'errors': errors}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_factory_wip_template_api(self, environ, method, start_response):
        try:
            user_id = self._get_session_user(environ)
            scope_ids = self._get_user_factory_scope_ids(user_id)
            query_params = parse_qs(environ.get('QUERY_STRING', ''))

            def _append_ids_from_value(container, value):
                if isinstance(value, list):
                    for v in value:
                        _append_ids_from_value(container, v)
                    return
                text = str(value or '').strip()
                if not text:
                    return
                for token in re.split(r'[\s,，;；]+', text):
                    item_id = self._parse_int(token)
                    if item_id and item_id not in container:
                        container.append(item_id)

            selected_ids = []
            _append_ids_from_value(selected_ids, query_params.get('ids', [''])[0])
            if method == 'POST':
                body = self._read_json_body(environ) or {}
                _append_ids_from_value(selected_ids, body.get('ids'))

            if method not in ('GET', 'POST'):
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = 'factory_wip'

            headers = ['工厂', '订单号', '合同编号', 'SKU', '数量', '预计完工日期', '是否完工(是/否)', '实际完工时间', '备注']
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value='工厂在制库存导入模板')
            title_cell.fill = PatternFill(start_color='A8B9A5', end_color='A8B9A5', fill_type='solid')
            title_cell.font = Font(bold=True, color='2A2420')
            title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.append(headers)
            for cell in ws[2]:
                cell.fill = PatternFill(start_color='DDE7DB', end_color='DDE7DB', fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            sample_row = ['示例工厂（请勿导入）', 'OD-2026-001', 'CT-2026-001', '示例SKU（请勿导入）', 50, '2026-03-31', '否', '', '示例行（请勿导入，此行仅演示格式）']
            ws.append(sample_row)
            for cell in ws[3]:
                cell.fill = PatternFill(start_color='ECECEC', end_color='ECECEC', fill_type='solid')
                cell.font = Font(italic=True, color='7B8088')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            widths = [24, 18, 18, 24, 10, 16, 14, 16, 28]
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            option_sheet = wb.create_sheet('_options')
            option_sheet.sheet_state = 'hidden'
            option_sheet.append(['factory_name', 'sku'])

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    factory_clause, factory_params = self._factory_scope_clause('id', user_id, prefix='WHERE')
                    cur.execute(
                        f"SELECT factory_name FROM logistics_factories{factory_clause} ORDER BY factory_name ASC",
                        factory_params
                    )
                    factories = [str(r.get('factory_name') or '').strip() for r in (cur.fetchall() or []) if r.get('factory_name')]
                    sku_scope_ids = self._get_linked_order_product_ids(scope_ids)
                    if sku_scope_ids is None:
                        cur.execute("SELECT sku FROM order_products ORDER BY sku ASC")
                        skus = [str(r.get('sku') or '').strip() for r in (cur.fetchall() or []) if r.get('sku')]
                    elif sku_scope_ids:
                        placeholders = ','.join(['%s'] * len(sku_scope_ids))
                        cur.execute(
                            f"SELECT sku FROM order_products WHERE id IN ({placeholders}) ORDER BY sku ASC",
                            tuple(sku_scope_ids)
                        )
                        skus = [str(r.get('sku') or '').strip() for r in (cur.fetchall() or []) if r.get('sku')]
                    else:
                        skus = []

                    selected_rows = []
                    if selected_ids:
                        placeholders = ','.join(['%s'] * len(selected_ids))
                        scope_clause, scope_params = self._factory_scope_clause('f.id', user_id, prefix='AND')
                        cur.execute(
                            f"""
                            SELECT
                                fw.id,
                                op.sku,
                                f.factory_name,
                                fw.quantity,
                                fw.expected_completion_date,
                                fc.order_no,
                                fc.contract_no,
                                fw.is_completed,
                                fw.actual_completion_date,
                                fw.notes,
                                fw.initial_expected_completion_date
                            FROM factory_wip_inventory fw
                            JOIN order_products op ON op.id = fw.order_product_id
                            JOIN logistics_factories f ON f.id = fw.factory_id
                            LEFT JOIN factory_contracts fc ON fc.id = fw.contract_id
                            WHERE fw.id IN ({placeholders}) {scope_clause}
                            ORDER BY op.sku ASC, f.factory_name ASC
                            """,
                            tuple(selected_ids) + scope_params
                        )
                        selected_rows = cur.fetchall() or []

            max_len = max(len(factories), len(skus), 1)
            for i in range(max_len):
                option_sheet.append([
                    factories[i] if i < len(factories) else '',
                    skus[i] if i < len(skus) else ''
                ])

            max_row = 400
            if factories:
                dv_factory = DataValidation(type='list', formula1=f"='_options'!$A$2:$A${len(factories) + 1}", allow_blank=False)
                ws.add_data_validation(dv_factory)
                for row in range(4, max_row + 1):
                    dv_factory.add(f'A{row}')
            if skus:
                dv_sku = DataValidation(type='list', formula1=f"='_options'!$B$2:$B${len(skus) + 1}", allow_blank=False)
                ws.add_data_validation(dv_sku)
                for row in range(4, max_row + 1):
                    dv_sku.add(f'D{row}')
            dv_completed = DataValidation(type='list', formula1='"否,是"', allow_blank=True)
            ws.add_data_validation(dv_completed)
            for row in range(4, max_row + 1):
                dv_completed.add(f'G{row}')

            if selected_rows:
                write_row = 4
                for item in selected_rows:
                    ws.cell(row=write_row, column=1, value=str(item.get('factory_name') or '').strip())
                    ws.cell(row=write_row, column=2, value=str(item.get('order_no') or '').strip())
                    ws.cell(row=write_row, column=3, value=str(item.get('contract_no') or '').strip())
                    ws.cell(row=write_row, column=4, value=str(item.get('sku') or '').strip())
                    ws.cell(row=write_row, column=5, value=int(item.get('quantity') or 0))
                    ws.cell(row=write_row, column=6, value=item.get('expected_completion_date') or None)
                    ws.cell(row=write_row, column=7, value='是' if int(item.get('is_completed') or 0) else '否')
                    ws.cell(row=write_row, column=8, value=item.get('actual_completion_date') or None)
                    ws.cell(row=write_row, column=9, value=str(item.get('notes') or '').strip())
                    write_row += 1

            ws.freeze_panes = 'A4'
            return self._send_excel_workbook(wb, 'factory_wip_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_factory_wip_import_api(self, environ, method, start_response):
        try:
            user_id = self._get_session_user(environ)
            scope_ids = self._get_user_factory_scope_ids(user_id)
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            header_row = 2 if str(ws.cell(row=1, column=1).value or '').strip().startswith('工厂在制库存导入模板') else 1
            header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), tuple())
            headers = [str(value or '').strip() for value in header_values]
            header_map = {name: idx for idx, name in enumerate(headers)}

            for required in ('工厂', 'SKU', '数量'):
                if required not in header_map:
                    return self.send_json({'status': 'error', 'message': f'模板缺少列: {required}'}, start_response)

            def get_cell(row, name):
                idx = header_map.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            def parse_date(value):
                if value is None:
                    return None
                if isinstance(value, datetime):
                    return value.strftime('%Y-%m-%d')
                if isinstance(value, date):
                    return value.strftime('%Y-%m-%d')
                if isinstance(value, (int, float)):
                    try:
                        serial = float(value)
                        if 1 <= serial <= 60000:
                            return (datetime(1899, 12, 30) + timedelta(days=serial)).strftime('%Y-%m-%d')
                    except Exception:
                        pass
                text = str(value or '').strip()
                if not text:
                    return None
                text = text.replace('年', '-').replace('月', '-').replace('日', '').replace('/', '-').replace('.', '-')
                month_day_match = re.match(r'^(\d{1,2})-(\d{1,2})$', text)
                if month_day_match:
                    try:
                        return datetime(datetime.now().year, int(month_day_match.group(1)), int(month_day_match.group(2))).strftime('%Y-%m-%d')
                    except Exception:
                        pass
                for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M'):
                    try:
                        return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
                    except Exception:
                        continue
                return None

            def parse_yes_no(value):
                text = str(value or '').strip().lower()
                return 1 if text in ('1', 'true', 'yes', 'y', '是') else 0

            def parse_contract_no(value):
                text = str(value or '').strip()
                return text[:128] if text else None

            def parse_order_no(value):
                text = str(value or '').strip()
                return text[:128] if text else None

            def resolve_contract_binding(cur, factory_id, contract_no, order_no):
                factory_id = self._parse_int(factory_id)
                if not factory_id:
                    raise ValueError('缺少工厂信息，无法匹配合同/订单编号')

                normalized_contract = parse_contract_no(contract_no)
                normalized_order = parse_order_no(order_no)
                if not normalized_contract and not normalized_order:
                    return None

                by_contract = None
                by_order = None
                if normalized_contract:
                    cur.execute(
                        """
                        SELECT id, contract_no, order_no
                        FROM factory_contracts
                        WHERE factory_id=%s AND contract_no=%s
                        LIMIT 1
                        """,
                        (factory_id, normalized_contract)
                    )
                    by_contract = cur.fetchone() or None
                if normalized_order:
                    cur.execute(
                        """
                        SELECT id, contract_no, order_no
                        FROM factory_contracts
                        WHERE factory_id=%s AND order_no=%s
                        LIMIT 1
                        """,
                        (factory_id, normalized_order)
                    )
                    by_order = cur.fetchone() or None

                if by_contract and by_order and int(by_contract.get('id') or 0) != int(by_order.get('id') or 0):
                    raise ValueError('合同编号与订单编号不匹配')

                existing = by_contract or by_order
                if existing:
                    existing_contract = parse_contract_no(existing.get('contract_no'))
                    existing_order = parse_order_no(existing.get('order_no'))
                    if normalized_contract and existing_contract and normalized_contract != existing_contract:
                        raise ValueError('合同编号与已存在映射不一致')
                    if normalized_order and existing_order and normalized_order != existing_order:
                        raise ValueError('订单编号与已存在映射不一致')
                    if normalized_contract and not normalized_order and not existing_order:
                        raise ValueError('该合同编号尚未绑定订单编号，请同时填写订单编号')
                    if normalized_order and not normalized_contract and not existing_contract:
                        raise ValueError('该订单编号尚未绑定合同编号，请同时填写合同编号')
                    return int(existing.get('id'))

                if not (normalized_contract and normalized_order):
                    raise ValueError('新增映射时需同时填写合同编号和订单编号')

                cur.execute(
                    """
                    INSERT INTO factory_contracts (factory_id, contract_no, order_no)
                    VALUES (%s, %s, %s)
                    """,
                    (factory_id, normalized_contract, normalized_order)
                )
                return int(cur.lastrowid)

            created = 0
            updated = 0
            unchanged = 0
            errors = []

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    if scope_ids is None:
                        cur.execute("SELECT id, sku FROM order_products")
                        sku_map = {str(r.get('sku') or '').strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}
                    elif scope_ids:
                        placeholders = ','.join(['%s'] * len(scope_ids))
                        cur.execute(
                            f"""
                            SELECT DISTINCT op.id, op.sku
                            FROM order_products op
                            JOIN (
                                SELECT order_product_id FROM factory_stock_inventory WHERE factory_id IN ({placeholders})
                                UNION
                                SELECT order_product_id FROM factory_wip_inventory WHERE factory_id IN ({placeholders})
                            ) x ON x.order_product_id = op.id
                            """,
                            tuple(scope_ids) + tuple(scope_ids)
                        )
                        sku_map = {str(r.get('sku') or '').strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}
                    else:
                        sku_map = {}
                    factory_clause, factory_params = self._factory_scope_clause('id', user_id, prefix='WHERE')
                    cur.execute(
                        f"SELECT id, factory_name FROM logistics_factories{factory_clause}",
                        factory_params
                    )
                    factory_map = {str(r.get('factory_name') or '').strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}

                    normalized_rows = []
                    pair_keys = set()
                    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                        if not any(value is not None and str(value).strip() for value in row):
                            continue
                        row_join = '|'.join([str(v or '').strip() for v in row])
                        if '示例' in row_join and '勿导入' in row_join:
                            continue
                        try:
                            factory_name = str(get_cell(row, '工厂') or '').strip()
                            sku = str(get_cell(row, 'SKU') or '').strip()
                            quantity = self._parse_int(get_cell(row, '数量'))
                            expected_completion_date = parse_date(get_cell(row, '预计完工日期'))
                            order_no = parse_order_no(get_cell(row, '订单号'))
                            contract_no = parse_contract_no(get_cell(row, '合同编号'))
                            is_completed = parse_yes_no(get_cell(row, '是否完工(是/否)'))
                            actual_completion_date = parse_date(get_cell(row, '实际完工时间'))
                            notes = str(get_cell(row, '备注') or '').strip() or None
                            if not sku or not factory_name or quantity is None:
                                raise ValueError('SKU/工厂/数量不能为空，且数量需为整数')
                            order_product_id = sku_map.get(sku)
                            factory_id = factory_map.get(factory_name)
                            if not order_product_id:
                                raise ValueError(f'未找到SKU: {sku}')
                            if not factory_id:
                                raise ValueError(f'未找到工厂: {factory_name}')
                            if not self._order_product_allowed_for_factory(order_product_id, factory_id):
                                raise ValueError(f'SKU {sku} 未关联工厂 {factory_name}')
                            if is_completed and not actual_completion_date:
                                actual_completion_date = datetime.now().strftime('%Y-%m-%d')
                            if not is_completed:
                                actual_completion_date = None
                            quantity = max(0, int(quantity))
                            normalized_rows.append((order_product_id, factory_id, quantity, expected_completion_date, order_no, contract_no, is_completed, actual_completion_date, notes))
                            pair_keys.add((order_product_id, factory_id))
                        except Exception as row_error:
                            errors.append({'row': row_idx, 'error': str(row_error)})

                    existing_map = {}
                    if pair_keys:
                        op_ids = sorted({k[0] for k in pair_keys})
                        factory_ids = sorted({k[1] for k in pair_keys})
                        op_placeholders = ','.join(['%s'] * len(op_ids))
                        fa_placeholders = ','.join(['%s'] * len(factory_ids))
                        cur.execute(
                            f"""
                            SELECT id, order_product_id, factory_id, contract_id, quantity,
                                expected_completion_date, initial_expected_completion_date,
                                is_completed, actual_completion_date, notes
                            FROM factory_wip_inventory
                            WHERE order_product_id IN ({op_placeholders})
                              AND factory_id IN ({fa_placeholders})
                            ORDER BY id DESC
                            """,
                            tuple(op_ids) + tuple(factory_ids)
                        )
                        for ex in (cur.fetchall() or []):
                            key = (int(ex.get('order_product_id')), int(ex.get('factory_id')))
                            if key in existing_map:
                                continue
                            existing_map[key] = {
                                'id': int(ex.get('id')),
                                'contract_id': self._parse_int(ex.get('contract_id')),
                                'quantity': int(ex.get('quantity') or 0),
                                'expected_completion_date': (str(ex.get('expected_completion_date') or '').strip() or None),
                                'initial_expected_completion_date': (str(ex.get('initial_expected_completion_date') or '').strip() or None),
                                'is_completed': int(ex.get('is_completed') or 0),
                                'actual_completion_date': (str(ex.get('actual_completion_date') or '').strip() or None),
                                'notes': (ex.get('notes') or '').strip() or None
                            }

                    for order_product_id, factory_id, quantity, expected_completion_date, order_no, contract_no, is_completed, actual_completion_date, notes in normalized_rows:
                        key = (order_product_id, factory_id)
                        ex = existing_map.get(key)
                        contract_id = resolve_contract_binding(cur, factory_id, contract_no, order_no)
                        if ex:
                            same = (
                                ex.get('quantity') == quantity and
                                (self._parse_int(ex.get('contract_id')) or None) == (contract_id or None) and
                                (ex.get('expected_completion_date') or None) == (expected_completion_date or None) and
                                int(ex.get('is_completed') or 0) == int(is_completed or 0) and
                                (ex.get('actual_completion_date') or None) == (actual_completion_date or None) and
                                (ex.get('notes') or None) == (notes or None)
                            )
                            if same:
                                unchanged += 1
                                continue
                            cur.execute(
                                """
                                UPDATE factory_wip_inventory
                                                                SET quantity=%s, contract_id=%s, expected_completion_date=%s, is_completed=%s, actual_completion_date=%s, notes=%s
                                WHERE id=%s
                                """,
                                                                (quantity, contract_id, expected_completion_date, is_completed, actual_completion_date, notes, ex['id'])
                            )
                            updated += 1
                        else:
                            cur.execute(
                                """
                                INSERT INTO factory_wip_inventory
                                                                    (order_product_id, factory_id, contract_id, quantity, expected_completion_date, initial_expected_completion_date, is_completed, actual_completion_date, notes)
                                                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                                """,
                                                                (
                                                                        order_product_id,
                                                                        factory_id,
                                                                        contract_id,
                                                                        quantity,
                                                                        expected_completion_date,
                                                                        expected_completion_date,
                                                                        is_completed,
                                                                        actual_completion_date,
                                                                        notes
                                                                )
                            )
                            created += 1

            return self.send_json({'status': 'success', 'created': created, 'updated': updated, 'unchanged': unchanged, 'errors': errors}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_factory_api(self, environ, method, start_response):
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            action = (query_params.get('action', [''])[0] or '').strip().lower()
            user_id = self._get_session_user(environ)
            actor_record = self._get_user_permission_record(user_id) if user_id else None
            can_manage_factory_master = bool(actor_record and actor_record.get('is_admin'))
            if method == 'GET':
                if action == 'order_product_options':
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            # 显示所有可选的 SKU（不按工厂绑定关系筛选）
                            cur.execute("SELECT id, sku FROM order_products ORDER BY sku ASC")
                            options = cur.fetchall() or []
                    return self.send_json({'status': 'success', 'items': options}, start_response)

                keyword = (query_params.get('q', [''])[0] or '').strip()
                item_id = self._parse_int((query_params.get('id', [''])[0] or '').strip())
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        scope_clause, scope_params = self._factory_scope_clause('id', user_id, prefix='AND')
                        if item_id:
                            cur.execute(
                                f"SELECT id, factory_name, created_at, updated_at FROM logistics_factories WHERE id=%s {scope_clause} LIMIT 1",
                                (item_id,) + scope_params
                            )
                            rows = cur.fetchall() or []
                        elif keyword:
                            cur.execute(
                                f"SELECT id, factory_name, created_at, updated_at FROM logistics_factories WHERE factory_name LIKE %s{scope_clause} ORDER BY id DESC",
                                (f"%{keyword}%",) + scope_params
                            )
                            rows = cur.fetchall() or []
                        else:
                            cur.execute(
                                f"SELECT id, factory_name, created_at, updated_at FROM logistics_factories WHERE 1=1 {scope_clause} ORDER BY id DESC",
                                scope_params
                            )
                            rows = cur.fetchall() or []
                    self._attach_factory_order_product_links(conn, rows)
                if item_id:
                    return self.send_json({'status': 'success', 'item': rows[0] if rows else None}, start_response)
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            data = self._read_json_body(environ)
            if method == 'POST':
                if not can_manage_factory_master:
                    return self.send_json({'status': 'error', 'message': '仅管理员可维护工厂主数据'}, start_response)
                name = (data.get('factory_name') or '').strip()
                order_product_ids = self._normalize_id_list_local(data.get('order_product_ids'))
                if not name:
                    return self.send_json({'status': 'error', 'message': 'Missing factory_name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("INSERT INTO logistics_factories (factory_name) VALUES (%s)", (name,))
                        new_id = cur.lastrowid
                    self._replace_factory_order_product_links(conn, new_id, order_product_ids)
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                if not can_manage_factory_master:
                    return self.send_json({'status': 'error', 'message': '仅管理员可维护工厂主数据'}, start_response)
                item_id = self._parse_int(data.get('id'))
                name = (data.get('factory_name') or '').strip()
                order_product_ids = self._normalize_id_list_local(data.get('order_product_ids'))
                if not item_id or not name:
                    return self.send_json({'status': 'error', 'message': 'Missing id or factory_name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("UPDATE logistics_factories SET factory_name=%s WHERE id=%s", (name, item_id))
                    self._replace_factory_order_product_links(conn, item_id, order_product_ids)
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                if not can_manage_factory_master:
                    return self.send_json({'status': 'error', 'message': '仅管理员可维护工厂主数据'}, start_response)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM logistics_factories WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_forwarder_api(self, environ, method, start_response):
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                "SELECT id, forwarder_name, created_at, updated_at FROM logistics_forwarders WHERE forwarder_name LIKE %s ORDER BY id DESC",
                                (f"%{keyword}%",)
                            )
                        else:
                            cur.execute("SELECT id, forwarder_name, created_at, updated_at FROM logistics_forwarders ORDER BY id DESC")
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            data = self._read_json_body(environ)
            if method == 'POST':
                name = (data.get('forwarder_name') or '').strip()
                if not name:
                    return self.send_json({'status': 'error', 'message': 'Missing forwarder_name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("INSERT INTO logistics_forwarders (forwarder_name) VALUES (%s)", (name,))
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                item_id = self._parse_int(data.get('id'))
                name = (data.get('forwarder_name') or '').strip()
                if not item_id or not name:
                    return self.send_json({'status': 'error', 'message': 'Missing id or forwarder_name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("UPDATE logistics_forwarders SET forwarder_name=%s WHERE id=%s", (name, item_id))
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM logistics_forwarders WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_supplier_api(self, environ, method, start_response):
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                "SELECT id, supplier_name, created_at, updated_at FROM logistics_suppliers WHERE supplier_name LIKE %s ORDER BY id DESC",
                                (f"%{keyword}%",)
                            )
                        else:
                            cur.execute("SELECT id, supplier_name, created_at, updated_at FROM logistics_suppliers ORDER BY id DESC")
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            data = self._read_json_body(environ)
            if method == 'POST':
                name = (data.get('supplier_name') or '').strip()
                if not name:
                    return self.send_json({'status': 'error', 'message': 'Missing supplier_name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("INSERT INTO logistics_suppliers (supplier_name) VALUES (%s)", (name,))
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                item_id = self._parse_int(data.get('id'))
                name = (data.get('supplier_name') or '').strip()
                if not item_id or not name:
                    return self.send_json({'status': 'error', 'message': 'Missing id or supplier_name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("UPDATE logistics_suppliers SET supplier_name=%s WHERE id=%s", (name, item_id))
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM logistics_suppliers WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_destination_region_api(self, environ, method, start_response):
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            action = (query_params.get('action', [''])[0] or '').strip().lower()
            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                "SELECT id, region_name, sort_order, created_at, updated_at FROM logistics_destination_regions WHERE region_name LIKE %s ORDER BY sort_order ASC, id ASC",
                                (f"%{keyword}%",)
                            )
                        else:
                            cur.execute("SELECT id, region_name, sort_order, created_at, updated_at FROM logistics_destination_regions ORDER BY sort_order ASC, id ASC")
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            data = self._read_json_body(environ)
            if method == 'PUT' and action == 'reorder':
                ordered_ids = data.get('ordered_ids') if isinstance(data.get('ordered_ids'), list) else []
                ids = [self._parse_int(x) for x in ordered_ids]
                ids = [x for x in ids if x]
                if not ids:
                    return self.send_json({'status': 'error', 'message': 'ordered_ids 不能为空'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM logistics_destination_regions")
                        existing = {self._parse_int((r or {}).get('id')) for r in (cur.fetchall() or [])}
                        if any(x not in existing for x in ids):
                            return self.send_json({'status': 'error', 'message': 'ordered_ids 包含不存在的区域ID'}, start_response)
                        sort_value = 10
                        for rid in ids:
                            cur.execute("UPDATE logistics_destination_regions SET sort_order=%s WHERE id=%s", (sort_value, rid))
                            sort_value += 10
                        remain_ids = [x for x in sorted(existing) if x and x not in ids]
                        for rid in remain_ids:
                            cur.execute("UPDATE logistics_destination_regions SET sort_order=%s WHERE id=%s", (sort_value, rid))
                            sort_value += 10
                return self.send_json({'status': 'success'}, start_response)

            if method == 'POST':
                name = (data.get('region_name') or '').strip()
                if not name:
                    return self.send_json({'status': 'error', 'message': 'Missing region_name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT COALESCE(MAX(sort_order),0) AS max_sort FROM logistics_destination_regions")
                        max_sort = self._parse_int((cur.fetchone() or {}).get('max_sort')) or 0
                        cur.execute("INSERT INTO logistics_destination_regions (region_name, sort_order) VALUES (%s, %s)", (name, max_sort + 10))
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                item_id = self._parse_int(data.get('id'))
                name = (data.get('region_name') or '').strip()
                if not item_id or not name:
                    return self.send_json({'status': 'error', 'message': 'Missing id or region_name'}, start_response)
                sort_order = self._parse_int(data.get('sort_order'))
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if sort_order and sort_order > 0:
                            cur.execute("UPDATE logistics_destination_regions SET region_name=%s, sort_order=%s WHERE id=%s", (name, sort_order, item_id))
                        else:
                            cur.execute("UPDATE logistics_destination_regions SET region_name=%s WHERE id=%s", (name, item_id))
                        cur.execute("UPDATE logistics_overseas_warehouses SET region=%s WHERE destination_region_id=%s", (name, item_id))
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT COUNT(*) AS cnt FROM logistics_overseas_warehouses WHERE destination_region_id=%s", (item_id,))
                        used_cnt = self._parse_int((cur.fetchone() or {}).get('cnt')) or 0
                        if used_cnt > 0:
                            return self.send_json({'status': 'error', 'message': '该目的区域已被仓库使用，无法删除'}, start_response)
                        cur.execute("DELETE FROM logistics_destination_regions WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_warehouse_api(self, environ, method, start_response):
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))

            def _resolve_region(conn, destination_region_id, region_name):
                region_id = self._parse_int(destination_region_id)
                text = (region_name or '').strip()
                with conn.cursor() as cur:
                    if region_id:
                        cur.execute("SELECT id, region_name FROM logistics_destination_regions WHERE id=%s LIMIT 1", (region_id,))
                        row = cur.fetchone()
                        if row:
                            return int(row.get('id')), str(row.get('region_name') or '').strip(), None
                    if text:
                        cur.execute("SELECT id, region_name FROM logistics_destination_regions WHERE region_name=%s LIMIT 1", (text,))
                        row = cur.fetchone()
                        if row:
                            return int(row.get('id')), str(row.get('region_name') or '').strip(), None
                    return None, None, '目的区域不存在或未填写'

            def _resolve_name_short(conn, supplier_id, warehouse_name, warehouse_short_name):
                with conn.cursor() as cur:
                    cur.execute("SELECT supplier_name FROM logistics_suppliers WHERE id=%s", (supplier_id,))
                    supplier = cur.fetchone()
                if not supplier:
                    return None, None, '供应商不存在'
                supplier_name = (supplier.get('supplier_name') or '').strip()
                name = (warehouse_name or '').strip()
                short_name = (warehouse_short_name or '').strip()

                if name and not short_name:
                    if name.startswith(supplier_name + ' '):
                        short_name = name[len(supplier_name) + 1:].strip()
                    elif ' ' in name:
                        parts = [p for p in name.split(' ') if p]
                        if len(parts) >= 2:
                            short_name = ' '.join(parts[1:]).strip() if parts[0] == supplier_name else parts[-1].strip()
                if short_name and not name:
                    name = f"{supplier_name} {short_name}".strip()

                if not name or not short_name:
                    return None, None, '仓库名称和仓库简称至少需要一个完整可推导'
                return name, short_name, None

            if method == 'GET':
                action = (query_params.get('action', [''])[0] or '').strip().lower()
                if action == 'options':
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute("SELECT id, supplier_name FROM logistics_suppliers ORDER BY id DESC")
                            suppliers = cur.fetchall() or []
                            cur.execute("SELECT id, region_name, sort_order FROM logistics_destination_regions ORDER BY sort_order ASC, id ASC")
                            regions = cur.fetchall() or []
                    return self.send_json({'status': 'success', 'suppliers': suppliers, 'destination_regions': regions}, start_response)

                keyword = (query_params.get('q', [''])[0] or '').strip()
                supplier_id = self._parse_int(query_params.get('supplier_id', [''])[0])
                destination_region_id = self._parse_int(query_params.get('destination_region_id', [''])[0])
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        sql = """
                            SELECT w.id, w.warehouse_name, w.supplier_id, w.warehouse_short_name, w.is_enabled,
                                COALESCE(dr.region_name, w.region) AS region,
                                w.destination_region_id,
                                dr.region_name AS destination_region_name,
                                   w.created_at, w.updated_at, s.supplier_name
                            FROM logistics_overseas_warehouses w
                            JOIN logistics_suppliers s ON s.id = w.supplier_id
                            LEFT JOIN logistics_destination_regions dr ON dr.id = w.destination_region_id
                        """
                        filters = []
                        params = []
                        if supplier_id:
                            filters.append("w.supplier_id=%s")
                            params.append(supplier_id)
                        if destination_region_id:
                            filters.append("w.destination_region_id=%s")
                            params.append(destination_region_id)
                        if keyword:
                            like = f"%{keyword}%"
                            filters.append("(w.warehouse_name LIKE %s OR s.supplier_name LIKE %s OR w.warehouse_short_name LIKE %s OR COALESCE(dr.region_name, w.region) LIKE %s)")
                            params.extend([like, like, like, like])
                        where_sql = (' WHERE ' + ' AND '.join(filters)) if filters else ''
                        cur.execute(sql + where_sql + " ORDER BY w.id DESC", params)
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            data = self._read_json_body(environ)
            action = (query_params.get('action', [''])[0] or '').strip().lower()

            if method == 'PUT' and action == 'toggle_enabled':
                item_id = self._parse_int(data.get('id'))
                is_enabled = 1 if self._parse_int(data.get('is_enabled')) else 0
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("UPDATE logistics_overseas_warehouses SET is_enabled=%s WHERE id=%s", (is_enabled, item_id))
                return self.send_json({'status': 'success'}, start_response)

            if method in ('POST', 'PUT'):
                supplier_id = self._parse_int(data.get('supplier_id'))
                destination_region_id = self._parse_int(data.get('destination_region_id'))
                region = (data.get('region') or '').strip()
                is_enabled = 1 if self._parse_int(data.get('is_enabled', 1)) else 0
                if not supplier_id:
                    return self.send_json({'status': 'error', 'message': '供应商和目的区域必填'}, start_response)
                with self._get_db_connection() as conn:
                    destination_region_id, resolved_region_name, region_err = _resolve_region(conn, destination_region_id, region)
                    if region_err:
                        return self.send_json({'status': 'error', 'message': region_err}, start_response)
                    name, short_name, err = _resolve_name_short(conn, supplier_id, data.get('warehouse_name'), data.get('warehouse_short_name'))
                    if err:
                        return self.send_json({'status': 'error', 'message': err}, start_response)
                    with conn.cursor() as cur:
                        if method == 'POST':
                            cur.execute(
                                """
                                INSERT INTO logistics_overseas_warehouses
                                (warehouse_name, supplier_id, warehouse_short_name, is_enabled, region, destination_region_id)
                                VALUES (%s, %s, %s, %s, %s, %s)
                                """,
                                (name, supplier_id, short_name, is_enabled, resolved_region_name, destination_region_id)
                            )
                            return self.send_json({'status': 'success', 'id': cur.lastrowid}, start_response)
                        item_id = self._parse_int(data.get('id'))
                        if not item_id:
                            return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                        cur.execute(
                            """
                            UPDATE logistics_overseas_warehouses
                            SET warehouse_name=%s, supplier_id=%s, warehouse_short_name=%s, is_enabled=%s, region=%s, destination_region_id=%s
                            WHERE id=%s
                            """,
                            (name, supplier_id, short_name, is_enabled, resolved_region_name, destination_region_id, item_id)
                        )
                        return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM logistics_overseas_warehouses WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_warehouse_template_api(self, environ, method, start_response):
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.datavalidation import DataValidation

            supplier_names = []
            destination_region_names = []
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT supplier_name FROM logistics_suppliers ORDER BY supplier_name ASC")
                    supplier_names = [str(r.get('supplier_name') or '').strip() for r in (cur.fetchall() or []) if str(r.get('supplier_name') or '').strip()]
                    cur.execute("SELECT region_name FROM logistics_destination_regions ORDER BY region_name ASC")
                    destination_region_names = [str(r.get('region_name') or '').strip() for r in (cur.fetchall() or []) if str(r.get('region_name') or '').strip()]

            wb = Workbook()
            ws = wb.active
            ws.title = 'warehouse_import'
            headers = ['仓库名称', '供应商', '仓库简称', '区域']
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value='海外仓导入模板')
            title_cell.fill = PatternFill(start_color='A8B9A5', end_color='A8B9A5', fill_type='solid')
            title_cell.font = Font(bold=True, color='2A2420')
            title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.append(headers)

            for cell in ws[2]:
                cell.fill = PatternFill(start_color='DDE7DB', end_color='DDE7DB', fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.append(['示例仓库（请勿导入）', '示例供应商（请勿导入）', '示例简称', '示例区域'])
            for cell in ws[3]:
                cell.fill = PatternFill(start_color='ECECEC', end_color='ECECEC', fill_type='solid')
                cell.font = Font(italic=True, color='7B8088')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            widths = [28, 22, 18, 12]
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width
            ws.freeze_panes = 'A4'

            option_ws = wb.create_sheet('_options')
            option_ws.append(['supplier_options', 'region_options'])
            max_len = max(len(supplier_names), len(destination_region_names), 1)
            for i in range(max_len):
                option_ws.append([
                    supplier_names[i] if i < len(supplier_names) else None,
                    destination_region_names[i] if i < len(destination_region_names) else None
                ])
            option_ws.sheet_state = 'hidden'

            if supplier_names:
                supplier_end_row = 1 + len(supplier_names)
                dv_supplier = DataValidation(type='list', formula1=f"='_options'!$A$2:$A${supplier_end_row}", allow_blank=False)
                ws.add_data_validation(dv_supplier)
                dv_supplier.add('B4:B1000')

            if destination_region_names:
                region_end_row = 1 + len(destination_region_names)
                dv_region = DataValidation(type='list', formula1=f"='_options'!$B$2:$B${region_end_row}", allow_blank=False)
                ws.add_data_validation(dv_region)
                dv_region.add('D4:D1000')

            return self._send_excel_workbook(wb, 'logistics_warehouse_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_warehouse_import_api(self, environ, method, start_response):
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)

            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            header_row = 2 if str(ws.cell(row=1, column=1).value or '').strip().startswith('海外仓导入模板') else 1
            headers = [str(cell.value or '').strip() for cell in ws[header_row]]
            header_map = {name: idx for idx, name in enumerate(headers)}
            required_headers = ['仓库名称', '供应商', '仓库简称', '区域']
            for col_name in required_headers:
                if col_name not in header_map:
                    return self.send_json({'status': 'error', 'message': f'模板缺少列: {col_name}'}, start_response)

            created = 0
            updated = 0
            unchanged = 0
            errors = []

            def get_cell(row, name):
                idx = header_map.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx].value

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id, supplier_name FROM logistics_suppliers")
                    supplier_rows = cur.fetchall() or []
                    supplier_map = {str(r.get('supplier_name') or '').strip(): int(r.get('id')) for r in supplier_rows if r.get('id') and str(r.get('supplier_name') or '').strip()}
                    cur.execute("SELECT id, region_name FROM logistics_destination_regions")
                    region_rows = cur.fetchall() or []
                    region_map = {str(r.get('region_name') or '').strip(): int(r.get('id')) for r in region_rows if r.get('id') and str(r.get('region_name') or '').strip()}

                for row_idx in range(header_row + 1, ws.max_row + 1):
                    row = ws[row_idx]
                    if not any(cell.value is not None and str(cell.value).strip() for cell in row):
                        continue
                    row_join = '|'.join([str(cell.value or '').strip() for cell in row])
                    if '示例' in row_join and '勿导入' in row_join:
                        continue
                    try:
                        warehouse_name = str(get_cell(row, '仓库名称') or '').strip()
                        supplier_name = str(get_cell(row, '供应商') or '').strip()
                        warehouse_short_name = str(get_cell(row, '仓库简称') or '').strip()
                        region = str(get_cell(row, '区域') or '').strip()

                        if supplier_name not in supplier_map:
                            raise ValueError(f'供应商必须为系统可选项: {supplier_name or "[空]"}')
                        if region not in region_map:
                            raise ValueError(f'目的区域必须为系统可选项: {region or "[空]"}')

                        supplier_id = supplier_map.get(supplier_name)
                        destination_region_id = region_map.get(region)
                        if not warehouse_name and warehouse_short_name:
                            warehouse_name = f"{supplier_name} {warehouse_short_name}".strip()
                        if warehouse_name and not warehouse_short_name:
                            if warehouse_name.startswith(supplier_name + ' '):
                                warehouse_short_name = warehouse_name[len(supplier_name) + 1:].strip()
                        if not warehouse_name or not warehouse_short_name:
                            raise ValueError('仓库名称/仓库简称无效，至少需形成可推导的完整名称')

                        with conn.cursor() as cur:
                            cur.execute(
                                "SELECT id, supplier_id, warehouse_short_name, region, destination_region_id FROM logistics_overseas_warehouses WHERE warehouse_name=%s LIMIT 1",
                                (warehouse_name,)
                            )
                            existing = cur.fetchone()
                            if existing:
                                if int(existing.get('supplier_id') or 0) == int(supplier_id or 0) and str(existing.get('warehouse_short_name') or '').strip() == warehouse_short_name and int(existing.get('destination_region_id') or 0) == int(destination_region_id or 0):
                                    unchanged += 1
                                else:
                                    cur.execute(
                                        "UPDATE logistics_overseas_warehouses SET supplier_id=%s, warehouse_short_name=%s, region=%s, destination_region_id=%s WHERE id=%s",
                                        (supplier_id, warehouse_short_name, region, destination_region_id, existing.get('id'))
                                    )
                                    updated += 1
                            else:
                                cur.execute(
                                    "INSERT INTO logistics_overseas_warehouses (warehouse_name, supplier_id, warehouse_short_name, region, destination_region_id) VALUES (%s, %s, %s, %s, %s)",
                                    (warehouse_name, supplier_id, warehouse_short_name, region, destination_region_id)
                                )
                                created += 1
                    except Exception as row_error:
                        errors.append({'row': row_idx, 'error': str(row_error)})

            return self.send_json({
                'status': 'success',
                'created': created,
                'updated': updated,
                'unchanged': unchanged,
                'errors': errors
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_warehouse_inventory_api(self, environ, method, start_response):
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))

            if method == 'GET':
                action = (query_params.get('action', [''])[0] or '').strip().lower()
                if action == 'options':
                    option_limit = max(100, min(self._parse_int(query_params.get('order_product_limit', ['800'])[0]) or 800, 2000))
                    include_order_products = str(query_params.get('include_order_products', ['1'])[0]).strip().lower() in ('1', 'true', 'yes')

                    def _load_options_payload():
                        with self._get_db_connection() as conn:
                            with conn.cursor() as cur:
                                cur.execute("SELECT id, warehouse_name FROM logistics_overseas_warehouses WHERE COALESCE(is_enabled,1)=1 ORDER BY warehouse_name ASC")
                                warehouses = cur.fetchall() or []
                                order_products = []
                                if include_order_products:
                                    cur.execute("SELECT id, sku FROM order_products ORDER BY sku ASC LIMIT %s", (option_limit,))
                                    order_products = cur.fetchall() or []
                        return {'status': 'success', 'warehouses': warehouses, 'order_products': order_products}

                    cache_key = f'logistics_wh_inventory_options_{1 if include_order_products else 0}_{option_limit}'
                    payload = self._get_cached_template_options(cache_key, _load_options_payload, ttl_seconds=1800)
                    return self.send_json(payload, start_response)

                if action == 'download_stock_summary':
                    keyword = (query_params.get('q', [''])[0] or '').strip()
                    warehouse_id = self._parse_int(query_params.get('warehouse_id', [''])[0])
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            sql = """
                                SELECT op.sku, SUM(COALESCE(i.available_qty, 0)) AS total_available_qty
                                FROM logistics_overseas_inventory i
                                JOIN logistics_overseas_warehouses w ON w.id = i.warehouse_id
                                JOIN order_products op ON op.id = i.order_product_id
                            """
                            filters = ["COALESCE(w.is_enabled,1)=1"]
                            params = []
                            if warehouse_id:
                                filters.append("i.warehouse_id=%s")
                                params.append(warehouse_id)
                            if keyword:
                                filters.append("(op.sku LIKE %s OR w.warehouse_name LIKE %s)")
                                like = f"%{keyword}%"
                                params.extend([like, like])
                            where_sql = (' WHERE ' + ' AND '.join(filters)) if filters else ''
                            cur.execute(sql + where_sql + " GROUP BY op.sku ORDER BY op.sku ASC", params)
                            rows = cur.fetchall() or []

                    output = io.StringIO(newline='')
                    writer = csv.writer(output)
                    writer.writerow(['SKU', '在库数量'])
                    for row in rows:
                        writer.writerow([
                            row.get('sku') or '',
                            self._parse_int(row.get('total_available_qty')) or 0,
                        ])

                    content = output.getvalue().encode('utf-8-sig')
                    filename = f"海外仓在库SKU汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                    headers = [
                        ('Content-Type', 'text/csv; charset=utf-8'),
                        ('Content-Disposition', f"attachment; filename*=UTF-8''{quote(filename)}"),
                        ('Content-Length', str(len(content))),
                    ]
                    start_response('200 OK', headers)
                    return [content]

                keyword = (query_params.get('q', [''])[0] or '').strip()
                warehouse_id = self._parse_int(query_params.get('warehouse_id', [''])[0])
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        sql = """
                            SELECT i.id, i.warehouse_id, i.order_product_id, i.available_qty,
                                   i.updated_at, w.warehouse_name, op.sku
                            FROM logistics_overseas_inventory i
                            JOIN logistics_overseas_warehouses w ON w.id = i.warehouse_id
                            JOIN order_products op ON op.id = i.order_product_id
                        """
                        filters = ["COALESCE(w.is_enabled,1)=1"]
                        params = []
                        if warehouse_id:
                            filters.append("i.warehouse_id=%s")
                            params.append(warehouse_id)
                        if keyword:
                            filters.append("(op.sku LIKE %s OR w.warehouse_name LIKE %s)")
                            like = f"%{keyword}%"
                            params.extend([like, like])
                        where_sql = (' WHERE ' + ' AND '.join(filters)) if filters else ''
                        cur.execute(sql + where_sql + " ORDER BY i.id DESC", params)
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            data = self._read_json_body(environ)
            if method == 'POST':
                warehouse_id = self._parse_int(data.get('warehouse_id'))
                order_product_id = self._parse_int(data.get('order_product_id'))
                available_qty = self._parse_int(data.get('available_qty'))
                if not warehouse_id or not order_product_id or available_qty is None:
                    return self.send_json({'status': 'error', 'message': 'Missing warehouse_id/order_product_id/available_qty'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO logistics_overseas_inventory (warehouse_id, order_product_id, available_qty, in_transit_qty)
                            VALUES (%s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                                available_qty=VALUES(available_qty)
                            """,
                            (warehouse_id, order_product_id, available_qty, 0)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'PUT':
                item_id = self._parse_int(data.get('id'))
                warehouse_id = self._parse_int(data.get('warehouse_id'))
                order_product_id = self._parse_int(data.get('order_product_id'))
                available_qty = self._parse_int(data.get('available_qty'))
                if not item_id or not warehouse_id or not order_product_id or available_qty is None:
                    return self.send_json({'status': 'error', 'message': 'Missing required fields'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE logistics_overseas_inventory
                            SET warehouse_id=%s, order_product_id=%s, available_qty=%s
                            WHERE id=%s
                            """,
                            (warehouse_id, order_product_id, available_qty, item_id)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM logistics_overseas_inventory WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_warehouse_inventory_template_api(self, environ, method, start_response):
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = 'warehouse_inventory'
            headers = ['SKU', '仓库', '可用量']
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value='海外仓库存导入模板')
            title_cell.fill = PatternFill(start_color='A8B9A5', end_color='A8B9A5', fill_type='solid')
            title_cell.font = Font(bold=True, color='2A2420')
            title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.append(headers)
            for cell in ws[2]:
                cell.fill = PatternFill(start_color='DDE7DB', end_color='DDE7DB', fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.append(['示例SKU（请勿导入）', '示例仓库（请勿导入）', 0])
            for cell in ws[3]:
                cell.fill = PatternFill(start_color='ECECEC', end_color='ECECEC', fill_type='solid')
                cell.font = Font(italic=True, color='7B8088')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            widths = [24, 28, 12]
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width
            ws.freeze_panes = 'A4'
            return self._send_excel_workbook(wb, 'warehouse_inventory_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_warehouse_inventory_import_api(self, environ, method, start_response):
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            import_mode = (query_params.get('mode', ['partial'])[0] or 'partial').strip().lower()
            if import_mode not in ('partial', 'replace_all'):
                import_mode = 'partial'
            ws = wb.active
            header_row = 2 if str(ws.cell(row=1, column=1).value or '').strip().startswith('海外仓库存导入模板') else 1
            header_rows = ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
            header_values = next(header_rows, tuple())
            headers = [str(value or '').strip() for value in header_values]
            header_map = {name: idx for idx, name in enumerate(headers)}
            required_headers = ['SKU', '仓库', '可用量']
            for col_name in required_headers:
                if col_name not in header_map:
                    return self.send_json({'status': 'error', 'message': f'模板缺少列: {col_name}'}, start_response)

            def get_cell(row, name):
                idx = header_map.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            created = 0
            updated = 0
            unchanged = 0
            errors = []

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    raw_rows = []
                    sku_names = set()
                    warehouse_names = set()
                    warehouse_ids = set()
                    order_product_ids = set()
                    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                        if not any(value is not None and str(value).strip() for value in row):
                            continue
                        row_join = '|'.join([str(v or '').strip() for v in row])
                        if '示例' in row_join and '勿导入' in row_join:
                            continue
                        try:
                            sku = str(get_cell(row, 'SKU') or '').strip()
                            warehouse_name = str(get_cell(row, '仓库') or '').strip()
                            available_qty = self._parse_int(get_cell(row, '可用量'))
                            if not sku or not warehouse_name or available_qty is None:
                                raise ValueError('SKU、仓库、可用量不能为空且可用量需为整数')
                            raw_rows.append((row_idx, sku, warehouse_name, available_qty))
                            sku_names.add(sku)
                            warehouse_names.add(warehouse_name)
                        except Exception as row_error:
                            errors.append({'row': row_idx, 'error': str(row_error)})

                    sku_map = {}
                    wh_map = {}
                    if sku_names:
                        sku_list = sorted(sku_names)
                        sku_placeholders = ','.join(['%s'] * len(sku_list))
                        cur.execute(
                            f"SELECT id, sku FROM order_products WHERE sku IN ({sku_placeholders})",
                            tuple(sku_list)
                        )
                        sku_map = {str(r.get('sku') or '').strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}
                    if warehouse_names:
                        wh_list = sorted(warehouse_names)
                        wh_placeholders = ','.join(['%s'] * len(wh_list))
                        cur.execute(
                            f"SELECT id, warehouse_name FROM logistics_overseas_warehouses WHERE COALESCE(is_enabled,1)=1 AND warehouse_name IN ({wh_placeholders})",
                            tuple(wh_list)
                        )
                        wh_map = {str(r.get('warehouse_name') or '').strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}

                    normalized_map = {}
                    for row_idx, sku, warehouse_name, available_qty in raw_rows:
                        order_product_id = sku_map.get(sku)
                        warehouse_id = wh_map.get(warehouse_name)
                        if not order_product_id:
                            errors.append({'row': row_idx, 'error': f'未找到SKU: {sku}'})
                            continue
                        if not warehouse_id:
                            errors.append({'row': row_idx, 'error': f'未找到仓库: {warehouse_name}'})
                            continue
                        warehouse_ids.add(warehouse_id)
                        order_product_ids.add(order_product_id)
                        normalized_map[(warehouse_id, order_product_id)] = available_qty

                    if import_mode == 'replace_all':
                        cur.execute("UPDATE logistics_overseas_inventory SET available_qty=0 WHERE available_qty<>0")

                    existing_map = {}
                    if warehouse_ids and order_product_ids:
                        wh_placeholders = ','.join(['%s'] * len(warehouse_ids))
                        op_placeholders = ','.join(['%s'] * len(order_product_ids))
                        cur.execute(
                            f"""
                            SELECT id, warehouse_id, order_product_id, available_qty
                            FROM logistics_overseas_inventory
                            WHERE warehouse_id IN ({wh_placeholders})
                              AND order_product_id IN ({op_placeholders})
                            """,
                            tuple(warehouse_ids) + tuple(order_product_ids)
                        )
                        for existing in (cur.fetchall() or []):
                            key = (self._parse_int(existing.get('warehouse_id')), self._parse_int(existing.get('order_product_id')))
                            if key[0] and key[1]:
                                existing_map[key] = {
                                    'id': self._parse_int(existing.get('id')),
                                    'available_qty': self._parse_int(existing.get('available_qty')),
                                }

                    to_upsert = []
                    to_insert = []
                    for (warehouse_id, order_product_id), available_qty in normalized_map.items():
                        key = (warehouse_id, order_product_id)
                        existing = existing_map.get(key)
                        if existing:
                            if (existing.get('available_qty') or 0) != available_qty:
                                to_upsert.append((warehouse_id, order_product_id, available_qty, 0))
                            else:
                                unchanged += 1
                        else:
                            to_insert.append((warehouse_id, order_product_id, available_qty, 0))
                            to_upsert.append((warehouse_id, order_product_id, available_qty, 0))

                    if to_upsert:
                        cur.executemany(
                            """
                            INSERT INTO logistics_overseas_inventory (warehouse_id, order_product_id, available_qty, in_transit_qty)
                            VALUES (%s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE available_qty=VALUES(available_qty)
                            """,
                            to_upsert
                        )
                    updated += len([1 for t in to_upsert if (t[0], t[1]) in existing_map and (existing_map[(t[0], t[1])].get('available_qty') or 0) != t[2]])
                    created += len(to_insert)

            return self.send_json({
                'status': 'success',
                'created': created,
                'updated': updated,
                'unchanged': unchanged,
                'mode': import_mode,
                'errors': errors
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_logistics_warehouse_dashboard_api(self, environ, method, start_response):
        """仓储看板：SKU库存聚合 + 仓库分列 + 在途角标（只读，工厂库存从独立表读取）"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            action = (query_params.get('action', [''])[0] or '').strip().lower()

            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)

            region_order = {}

            def _region_rank(region_name):
                text = str(region_name or '').strip()
                for key, rank in region_order.items():
                    if key in text:
                        return rank
                return 99

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id, region_name, sort_order FROM logistics_destination_regions ORDER BY sort_order ASC, id ASC")
                    destination_region_rows = cur.fetchall() or []
                    for idx, rr in enumerate(destination_region_rows, start=1):
                        rname = str((rr or {}).get('region_name') or '').strip()
                        if rname and rname not in region_order:
                            region_order[rname] = idx

                    cur.execute(
                        """
                        SELECT w.id, w.warehouse_name, w.warehouse_short_name,
                               COALESCE(dr.region_name, w.region) AS region,
                               s.supplier_name
                        FROM logistics_overseas_warehouses w
                        JOIN logistics_suppliers s ON s.id = w.supplier_id
                        LEFT JOIN logistics_destination_regions dr ON dr.id = w.destination_region_id
                        WHERE COALESCE(w.is_enabled,1)=1
                        """
                    )
                    warehouse_rows = cur.fetchall() or []

                    cur.execute(
                        """
                        SELECT op.id, op.sku, pf.sku_family, op.is_iteration, op.is_on_market, op.source_order_product_id,
                               fm.fabric_name_en, fm.representative_color
                        FROM order_products op
                        LEFT JOIN product_families pf ON pf.id = op.sku_family_id
                        LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                        ORDER BY op.sku DESC
                        """
                    )
                    sku_rows = cur.fetchall() or []

                    cur.execute(
                        """
                        SELECT order_product_id, warehouse_id, SUM(available_qty) AS qty
                        FROM logistics_overseas_inventory
                        GROUP BY order_product_id, warehouse_id
                        """
                    )
                    inv_rows = cur.fetchall() or []

                    cur.execute(
                        """
                        SELECT
                            li.order_product_id,
                            t.destination_warehouse_id AS warehouse_id,
                            COALESCE(drt.region_name, dr.region_name, w.region) AS region,
                            t.logistics_box_no,
                            COALESCE(t.expected_warehouse_date, t.eta_latest, t.expected_listed_date_latest) AS expected_arrival_date,
                            SUM(li.shipped_qty) AS qty
                        FROM logistics_in_transit_items li
                        JOIN logistics_in_transit t ON t.id = li.transit_id
                        LEFT JOIN logistics_overseas_warehouses w ON w.id = t.destination_warehouse_id
                        LEFT JOIN logistics_destination_regions drt ON drt.id = t.destination_region_id
                        LEFT JOIN logistics_destination_regions dr ON dr.id = w.destination_region_id
                        WHERE t.listed_date IS NULL
                                                    AND COALESCE(w.is_enabled,1)=1
                        GROUP BY li.order_product_id, t.destination_warehouse_id, COALESCE(drt.region_name, dr.region_name, w.region), t.logistics_box_no, COALESCE(t.expected_warehouse_date, t.eta_latest, t.expected_listed_date_latest)
                        """
                    )
                    transit_rows = cur.fetchall() or []

                    cur.execute("SELECT id, factory_name FROM logistics_factories ORDER BY factory_name ASC")
                    factory_rows = cur.fetchall() or []

                    cur.execute(
                        """
                        SELECT order_product_id, factory_id, quantity
                        FROM factory_stock_inventory
                        """
                    )
                    fstock_rows = cur.fetchall() or []

                    cur.execute(
                        """
                        SELECT order_product_id, factory_id, SUM(quantity) AS quantity,
                               MIN(expected_completion_date) AS earliest_date
                        FROM factory_wip_inventory
                        WHERE COALESCE(is_completed, 0) = 0
                        GROUP BY order_product_id, factory_id
                        """
                    )
                    fwip_rows = cur.fetchall() or []

                    cur.execute(
                        """
                        SELECT id, order_product_id, factory_id, quantity, expected_completion_date
                        FROM factory_wip_inventory
                        WHERE COALESCE(is_completed, 0) = 0
                        ORDER BY expected_completion_date ASC, id ASC
                        """
                    )
                    fwip_detail_rows = cur.fetchall() or []

            factories = [
                {'id': int(r['id']), 'factory_name': r['factory_name'] or ''}
                for r in factory_rows if r.get('id')
            ]
            factory_ids = [f['id'] for f in factories]

            warehouses = sorted(
                [
                    {
                        'id': int(w.get('id')),
                        'warehouse_name': w.get('warehouse_name') or '',
                        'warehouse_short_name': w.get('warehouse_short_name') or '',
                        'region': w.get('region') or '',
                        'supplier_name': w.get('supplier_name') or ''
                    }
                    for w in warehouse_rows if w.get('id')
                ],
                key=lambda x: (_region_rank(x.get('region')), str(x.get('warehouse_name') or ''))
            )

            inv_map = {}
            for row in inv_rows:
                op_id = self._parse_int(row.get('order_product_id'))
                wh_id = self._parse_int(row.get('warehouse_id'))
                qty = self._parse_int(row.get('qty')) or 0
                if not op_id or not wh_id:
                    continue
                inv_map[(op_id, wh_id)] = qty

            transit_total_map = {}
            transit_wh_map = {}
            transit_region_map = {}
            transit_tip_map = {}
            transit_wh_tip_map = {}
            transit_region_tip_map = {}
            for row in transit_rows:
                op_id = self._parse_int(row.get('order_product_id'))
                wh_id = self._parse_int(row.get('warehouse_id'))
                region = (row.get('region') or '').strip()
                box_no = (row.get('logistics_box_no') or '').strip()
                arrival_date = row.get('expected_arrival_date')
                arrival_text = str(arrival_date)[:10] if arrival_date else '未知'
                qty = self._parse_int(row.get('qty')) or 0
                if not op_id or qty <= 0:
                    continue
                transit_total_map[op_id] = transit_total_map.get(op_id, 0) + qty
                tip_label = box_no or '批次'
                tip_text = f"{tip_label}: {qty}（预计到货 {arrival_text}）"
                transit_tip_map.setdefault(op_id, []).append(tip_text)
                if wh_id:
                    transit_wh_map[(op_id, wh_id)] = transit_wh_map.get((op_id, wh_id), 0) + qty
                    transit_wh_tip_map.setdefault((op_id, wh_id), []).append(tip_text)
                if region:
                    transit_region_map.setdefault(op_id, {})
                    transit_region_map[op_id][region] = transit_region_map[op_id].get(region, 0) + qty
                    transit_region_tip_map.setdefault(op_id, {})
                    transit_region_tip_map[op_id].setdefault(region, [])
                    transit_region_tip_map[op_id][region].append(tip_text)

            fstock_map = {}
            for row in fstock_rows:
                op_id = self._parse_int(row.get('order_product_id'))
                f_id = self._parse_int(row.get('factory_id'))
                qty = self._parse_int(row.get('quantity')) or 0
                if op_id and f_id:
                    fstock_map.setdefault(op_id, {})[f_id] = qty

            fwip_map = {}
            for row in fwip_rows:
                op_id = self._parse_int(row.get('order_product_id'))
                f_id = self._parse_int(row.get('factory_id'))
                qty = self._parse_int(row.get('quantity')) or 0
                date_val = row.get('earliest_date')
                if op_id and f_id:
                    fwip_map.setdefault(op_id, {})[f_id] = {
                        'quantity': qty,
                        'earliest_date': str(date_val) if date_val else None
                    }

            fwip_tip_map = {}
            for row in fwip_detail_rows:
                op_id = self._parse_int(row.get('order_product_id'))
                f_id = self._parse_int(row.get('factory_id'))
                row_id = self._parse_int(row.get('id'))
                qty = self._parse_int(row.get('quantity')) or 0
                date_val = row.get('expected_completion_date')
                if not op_id or not f_id or qty <= 0:
                    continue
                date_text = str(date_val)[:10] if date_val else '未知'
                tip_text = f"批次#{row_id}: {qty}（预计完工 {date_text}）"
                fwip_tip_map.setdefault((op_id, f_id), []).append(tip_text)

            raw_row_map = {}
            for row in sku_rows:
                op_id = self._parse_int(row.get('id'))
                if not op_id:
                    continue
                warehouse_qty = {}
                available_total = 0
                for wh in warehouses:
                    wh_id = wh['id']
                    qty = inv_map.get((op_id, wh_id), 0)
                    warehouse_qty[str(wh_id)] = qty
                    available_total += qty
                op_fstock = fstock_map.get(op_id, {})
                op_fwip = fwip_map.get(op_id, {})
                raw_row_map[op_id] = {
                    'order_product_id': op_id,
                    'sku_family': row.get('sku_family') or '',
                    'sku': row.get('sku') or '',
                    'fabric_name_en': row.get('fabric_name_en') or '',
                    'representative_color': row.get('representative_color') or '',
                    'available_total': available_total,
                    'in_transit_total': transit_total_map.get(op_id, 0),
                    'factory_stock_total': sum(op_fstock.values()),
                    'factory_stock_by_factory': {str(fid): qty for fid, qty in op_fstock.items()},
                    'factory_wip_total': sum(v['quantity'] for v in op_fwip.values()),
                    'factory_wip_by_factory': {
                        str(fid): {'quantity': v['quantity'], 'earliest_date': v['earliest_date']}
                        for fid, v in op_fwip.items()
                    },
                    'in_transit_by_region': transit_region_map.get(op_id, {}),
                    'in_transit_tip_by_region': {
                        str(rname): '\n'.join(tips)
                        for rname, tips in (transit_region_tip_map.get(op_id, {}) or {}).items()
                    },
                    'is_iteration': 1 if self._parse_int(row.get('is_iteration')) else 0,
                    'is_on_market': 1 if self._parse_int(row.get('is_on_market', 1)) else 0,
                    'source_order_product_id': self._parse_int(row.get('source_order_product_id')),
                    'warehouse_qty': warehouse_qty,
                    'in_transit_by_warehouse': {
                        str(wh['id']): transit_wh_map.get((op_id, wh['id']), 0) for wh in warehouses
                    },
                    'in_transit_tip_by_warehouse': {
                        str(wh['id']): '\n'.join(transit_wh_tip_map.get((op_id, wh['id']), [])) for wh in warehouses
                    },
                    'factory_wip_tip_by_factory': {
                        str(fid): '\n'.join(fwip_tip_map.get((op_id, fid), [])) for fid in factory_ids
                    },
                    'in_transit_tip': '\n'.join(transit_tip_map.get(op_id, [])),
                    'iteration_children': []
                }

            def _resolve_source_root(op_id):
                visited = set()
                current_id = op_id
                while current_id and current_id not in visited:
                    visited.add(current_id)
                    current = raw_row_map.get(current_id)
                    if not current:
                        return op_id
                    source_id = self._parse_int(current.get('source_order_product_id'))
                    if not current.get('is_iteration') or not source_id or source_id == current_id or source_id not in raw_row_map:
                        return current_id
                    current_id = source_id
                return op_id

            top_level_ids = []
            seen_top_level = set()
            for row in sku_rows:
                op_id = self._parse_int(row.get('id'))
                if not op_id or op_id not in raw_row_map:
                    continue
                root_id = _resolve_source_root(op_id)
                if root_id == op_id:
                    if op_id not in seen_top_level:
                        seen_top_level.add(op_id)
                        top_level_ids.append(op_id)
                    continue
                parent = raw_row_map.get(root_id)
                child = raw_row_map.get(op_id)
                if not parent or not child:
                    if op_id not in seen_top_level:
                        seen_top_level.add(op_id)
                        top_level_ids.append(op_id)
                    continue
                parent['available_total'] += child.get('available_total', 0)
                parent['in_transit_total'] += child.get('in_transit_total', 0)
                parent['factory_stock_total'] += child.get('factory_stock_total', 0)
                parent['factory_wip_total'] += child.get('factory_wip_total', 0)
                for region_key, rqty in (child.get('in_transit_by_region') or {}).items():
                    parent['in_transit_by_region'][region_key] = parent['in_transit_by_region'].get(region_key, 0) + rqty
                for region_key, rtip in (child.get('in_transit_tip_by_region') or {}).items():
                    parent_region_tip = (parent.get('in_transit_tip_by_region') or {}).get(region_key) or ''
                    merged_region_tip = '\n'.join([text for text in [parent_region_tip, rtip] if text])
                    parent.setdefault('in_transit_tip_by_region', {})[region_key] = merged_region_tip
                for fid_str, qty in (child.get('factory_stock_by_factory') or {}).items():
                    parent['factory_stock_by_factory'][fid_str] = parent['factory_stock_by_factory'].get(fid_str, 0) + qty
                for fid_str, wip_val in (child.get('factory_wip_by_factory') or {}).items():
                    if fid_str in parent['factory_wip_by_factory']:
                        parent['factory_wip_by_factory'][fid_str]['quantity'] += wip_val.get('quantity', 0)
                    else:
                        parent['factory_wip_by_factory'][fid_str] = dict(wip_val)
                if child.get('in_transit_tip'):
                    parent['in_transit_tip'] = '\n'.join(filter(None, [parent.get('in_transit_tip') or '', child.get('in_transit_tip') or '']))
                for wh in warehouses:
                    wh_key = str(wh['id'])
                    parent['warehouse_qty'][wh_key] = (parent['warehouse_qty'].get(wh_key) or 0) + (child.get('warehouse_qty', {}).get(wh_key) or 0)
                    parent['in_transit_by_warehouse'][wh_key] = (parent['in_transit_by_warehouse'].get(wh_key) or 0) + (child.get('in_transit_by_warehouse', {}).get(wh_key) or 0)
                    child_tip = (child.get('in_transit_tip_by_warehouse', {}) or {}).get(wh_key) or ''
                    parent_tip = (parent.get('in_transit_tip_by_warehouse', {}) or {}).get(wh_key) or ''
                    merged_tip = '\n'.join([text for text in [parent_tip, child_tip] if text])
                    parent.setdefault('in_transit_tip_by_warehouse', {})[wh_key] = merged_tip
                for fid in factory_ids:
                    f_key = str(fid)
                    child_tip = (child.get('factory_wip_tip_by_factory', {}) or {}).get(f_key) or ''
                    parent_tip = (parent.get('factory_wip_tip_by_factory', {}) or {}).get(f_key) or ''
                    merged_tip = '\n'.join([text for text in [parent_tip, child_tip] if text])
                    parent.setdefault('factory_wip_tip_by_factory', {})[f_key] = merged_tip
                parent['iteration_children'].append({
                    'order_product_id': child.get('order_product_id'),
                    'sku': child.get('sku') or '',
                    'fabric_name_en': child.get('fabric_name_en') or '',
                    'representative_color': child.get('representative_color') or '',
                    'is_on_market': 1 if self._parse_int(child.get('is_on_market', 1)) else 0,
                    'available_total': child.get('available_total', 0) or 0,
                    'in_transit_total': child.get('in_transit_total', 0) or 0,
                    'warehouse_qty': dict(child.get('warehouse_qty') or {}),
                    'in_transit_by_warehouse': dict(child.get('in_transit_by_warehouse') or {}),
                    'in_transit_tip': child.get('in_transit_tip') or '',
                    'factory_stock_total': child.get('factory_stock_total', 0) or 0,
                    'factory_wip_total': child.get('factory_wip_total', 0) or 0,
                    'in_transit_tip_by_warehouse': dict(child.get('in_transit_tip_by_warehouse') or {}),
                    'in_transit_tip_by_region': dict(child.get('in_transit_tip_by_region') or {}),
                    'factory_wip_tip_by_factory': dict(child.get('factory_wip_tip_by_factory') or {}),
                    'factory_stock_by_factory': dict(child.get('factory_stock_by_factory') or {}),
                    'factory_wip_by_factory': {
                        str(fid): {
                            'quantity': int((val or {}).get('quantity') or 0),
                            'earliest_date': (val or {}).get('earliest_date')
                        }
                        for fid, val in (child.get('factory_wip_by_factory') or {}).items()
                    }
                })

            rows = []
            for op_id in top_level_ids:
                item = raw_row_map.get(op_id)
                if not item:
                    continue
                item['iteration_children'] = sorted(item.get('iteration_children') or [], key=lambda x: str(x.get('sku') or ''))
                rows.append(item)

            ordered_region_names = [
                str((rr or {}).get('region_name') or '').strip()
                for rr in (destination_region_rows or [])
                if str((rr or {}).get('region_name') or '').strip()
            ]

            if action == 'export':
                if Workbook is None:
                    return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)
                wb = Workbook()
                ws = wb.active
                ws.title = 'warehouse_dashboard'
                factory_headers = [f['factory_name'] for f in factories]
                headers = (
                    ['SKU', '现货库存', '在途数量']
                    + [w['warehouse_short_name'] or w['warehouse_name'] for w in warehouses]
                    + [f'工厂在库-{fn}' for fn in factory_headers]
                    + [f'工厂在制-{fn}' for fn in factory_headers]
                )
                for idx, title in enumerate(headers, start=1):
                    ws.cell(row=1, column=idx, value=title)
                line = 2
                for item in rows:
                    data_line = [
                        item.get('sku') or '',
                        item.get('available_total') or 0,
                        item.get('in_transit_total') or 0,
                    ]
                    for wh in warehouses:
                        data_line.append((item.get('warehouse_qty') or {}).get(str(wh['id']), 0))
                    fstock_bf = item.get('factory_stock_by_factory') or {}
                    for f in factories:
                        data_line.append(fstock_bf.get(str(f['id']), 0))
                    fwip_bf = item.get('factory_wip_by_factory') or {}
                    for f in factories:
                        wip_val = fwip_bf.get(str(f['id'])) or {}
                        data_line.append(wip_val.get('quantity', 0))
                    for col, value in enumerate(data_line, start=1):
                        ws.cell(row=line, column=col, value=value)
                    line += 1
                return self._send_excel_workbook(wb, 'logistics_warehouse_dashboard.xlsx', start_response)

            return self.send_json({'status': 'success', 'warehouses': warehouses, 'factories': factories, 'region_order': ordered_region_names, 'items': rows}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)



