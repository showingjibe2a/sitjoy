# -*- coding: utf-8 -*-
"""销售管理：订单登记、销售预测/看板、动销月（30 天窗口销量与库存覆盖）。"""
import io
import cgi
import re
import calendar
from datetime import date, datetime, timedelta
from urllib.parse import parse_qs


def _sf_effective_wsgi_query_string(environ):
    """与 file_management_mixin 一致：代理环境下 QUERY_STRING 可能为空。"""
    if not environ:
        return ''
    raw = environ.get('QUERY_STRING')
    if raw is not None and str(raw).strip():
        return str(raw)
    for key in ('REDIRECT_QUERY_STRING', 'HTTP_X_ORIGINAL_QUERY_STRING', 'HTTP_X_QUERY_STRING'):
        val = environ.get(key)
        if val is not None and str(val).strip():
            return str(val)
    for uri_key in ('REQUEST_URI', 'RAW_URI', 'UNENCODED_URL'):
        uri = environ.get(uri_key)
        if uri is None:
            continue
        if isinstance(uri, (bytes, bytearray)):
            try:
                uri = uri.decode('latin-1', errors='replace')
            except Exception:
                continue
        uri = str(uri)
        q = uri.find('?')
        if q >= 0 and q < len(uri) - 1:
            return uri[q + 1 :]
    return ''


def _sf_parse_qs(qs):
    if not qs:
        return {}
    try:
        return parse_qs(qs, separator='&', keep_blank_values=False)
    except TypeError:
        return parse_qs(qs, keep_blank_values=False)


try:
    from openpyxl import Workbook, load_workbook
    _openpyxl_import_error = None
except Exception as e:
    Workbook = None
    load_workbook = None
    _openpyxl_import_error = str(e)


class SalesManagementMixin:
    """销售管理 Mixin：订单登记、预测看板、动销月计算与仓储看板挂载。"""

    def _registration_get_replacement_options(self, conn, base_order_product_ids):
        """按基础发货 SKU 加载替代方案及方案明细。"""
        result = {}
        ids = [self._parse_int(x) for x in (base_order_product_ids or []) if self._parse_int(x)]
        if not ids:
            return result

        placeholders = ','.join(['%s'] * len(ids))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT ops.id, ops.order_product_id, ops.plan_name
                FROM order_product_shipping_plans ops
                WHERE ops.order_product_id IN ({placeholders})
                ORDER BY ops.order_product_id ASC, ops.plan_name ASC, ops.id ASC
                """,
                tuple(ids)
            )
            plan_rows = cur.fetchall() or []

            plan_ids = [self._parse_int(x.get('id')) for x in plan_rows if self._parse_int(x.get('id'))]
            plan_item_map = {}
            if plan_ids:
                plan_placeholders = ','.join(['%s'] * len(plan_ids))
                cur.execute(
                    f"""
                    SELECT
                        opsi.shipping_plan_id,
                        opsi.substitute_order_product_id,
                        opsi.quantity,
                        opsi.sort_order,
                        op.sku
                    FROM order_product_shipping_plan_items opsi
                    JOIN order_products op ON op.id = opsi.substitute_order_product_id
                    WHERE opsi.shipping_plan_id IN ({plan_placeholders})
                    ORDER BY opsi.shipping_plan_id ASC, opsi.sort_order ASC, opsi.id ASC
                    """,
                    tuple(plan_ids)
                )
                for row in (cur.fetchall() or []):
                    pid = self._parse_int(row.get('shipping_plan_id'))
                    if not pid:
                        continue
                    plan_item_map.setdefault(pid, []).append({
                        'order_product_id': self._parse_int(row.get('substitute_order_product_id')),
                        'order_sku': (row.get('sku') or '').strip(),
                        'quantity': max(1, self._parse_int(row.get('quantity')) or 1)
                    })

            for row in plan_rows:
                base_id = self._parse_int(row.get('order_product_id'))
                plan_id = self._parse_int(row.get('id'))
                if not base_id or not plan_id:
                    continue
                result.setdefault(base_id, []).append({
                    'plan_id': plan_id,
                    'plan_name': (row.get('plan_name') or '').strip(),
                    'items': plan_item_map.get(plan_id, [])
                })

        return result

    def _bool_from_any(self, value, default=0):
        if value is None:
            return 1 if default else 0
        text = str(value).strip().lower()
        if text in ('1', 'true', 'yes', 'y', '是', 'on'):
            return 1
        if text in ('0', 'false', 'no', 'n', '否', 'off'):
            return 0
        return 1 if default else 0

    def _validate_us_phone_zip(self, phone, zip_code):
        phone_text = (phone or '').strip()
        zip_text = (zip_code or '').strip()
        if phone_text:
            digits = re.sub(r'\D+', '', phone_text)
            if len(digits) == 11 and digits.startswith('1'):
                digits = digits[1:]
            if len(digits) != 10:
                raise ValueError('电话格式无效，请填写美国电话（10位数字，可含+1）')
        if zip_text and not re.match(r'^\d{5}(-\d{4})?$', zip_text):
            raise ValueError('邮编格式无效，请填写美国邮编（5位或5+4）')

    def _normalize_registration_platform_items(self, items):
        normalized = []
        if not isinstance(items, list):
            return normalized
        for entry in items:
            if not isinstance(entry, dict):
                continue
            platform_sku = (entry.get('platform_sku') or '').strip()
            sales_product_id = self._parse_int(entry.get('sales_product_id'))
            quantity = self._parse_int(entry.get('quantity')) or 1
            shipping_plan_id = self._parse_int(entry.get('shipping_plan_id'))
            if not platform_sku and not sales_product_id:
                continue
            normalized.append({
                'sales_product_id': sales_product_id,
                'platform_sku': platform_sku,
                'quantity': max(1, quantity),
                'shipping_plan_id': shipping_plan_id
            })
        return normalized

    def _normalize_registration_shipment_items(self, items):
        normalized = []
        if not isinstance(items, list):
            return normalized
        for entry in items:
            if not isinstance(entry, dict):
                continue
            order_product_id = self._parse_int(entry.get('order_product_id'))
            order_sku = (entry.get('order_sku') or '').strip()
            quantity = self._parse_int(entry.get('quantity')) or 1
            shipping_plan_id = self._parse_int(entry.get('shipping_plan_id'))
            source_type = (entry.get('source_type') or 'manual').strip().lower()
            if source_type not in ('manual', 'auto', 'plan'):
                source_type = 'manual'
            if not order_product_id and not order_sku:
                continue
            normalized.append({
                'order_product_id': order_product_id,
                'order_sku': order_sku,
                'quantity': max(1, quantity),
                'source_type': source_type,
                'shipping_plan_id': shipping_plan_id
            })
        return normalized

    def _normalize_registration_logistics_items(self, items):
        normalized = []
        if not isinstance(items, list):
            return normalized
        for index, entry in enumerate(items, start=1):
            if not isinstance(entry, dict):
                continue
            shipping_carrier = (entry.get('shipping_carrier') or '').strip()
            tracking_no = (entry.get('tracking_no') or '').strip()
            sort_order = self._parse_int(entry.get('sort_order')) or index
            if not shipping_carrier and not tracking_no:
                continue
            normalized.append({
                'shipping_carrier': shipping_carrier or None,
                'tracking_no': tracking_no or None,
                'sort_order': max(1, sort_order)
            })
        return normalized

    def _resolve_registration_auto_shipments(self, conn, platform_items):
        aggregate = {}
        if not platform_items:
            return []

        with conn.cursor() as cur:
            for item in platform_items:
                qty = max(1, self._parse_int(item.get('quantity')) or 1)
                shipping_plan_id = self._parse_int(item.get('shipping_plan_id'))

                if shipping_plan_id:
                    cur.execute(
                        """
                        SELECT op.id AS order_product_id, op.sku, opsi.quantity
                        FROM order_product_shipping_plan_items opsi
                        JOIN order_products op ON op.id = opsi.substitute_order_product_id
                        WHERE opsi.shipping_plan_id=%s
                        ORDER BY opsi.sort_order ASC, opsi.id ASC
                        """,
                        (shipping_plan_id,)
                    )
                    rels = cur.fetchall() or []
                    for rel in rels:
                        key = int(rel.get('order_product_id'))
                        aggregate.setdefault(key, {
                            'order_product_id': key,
                            'order_sku': rel.get('sku') or '',
                            'quantity': 0,
                            'source_type': 'plan',
                            'shipping_plan_id': shipping_plan_id
                        })
                        aggregate[key]['quantity'] += qty * (self._parse_int(rel.get('quantity')) or 1)
                    continue

                sales_product_id = self._parse_int(item.get('sales_product_id'))
                platform_sku = (item.get('platform_sku') or '').strip()
                if not sales_product_id and platform_sku:
                    cur.execute("SELECT id FROM sales_products WHERE platform_sku=%s LIMIT 1", (platform_sku,))
                    row = cur.fetchone() or {}
                    sales_product_id = self._parse_int(row.get('id'))
                if not sales_product_id:
                    continue

                cur.execute(
                    """
                    SELECT op.id AS order_product_id, op.sku, svol.quantity
                    FROM sales_products sp
                    JOIN sales_variant_order_links svol ON svol.variant_id = sp.variant_id
                    JOIN order_products op ON op.id = svol.order_product_id
                    WHERE sp.id=%s
                    """,
                    (sales_product_id,)
                )
                rels = cur.fetchall() or []
                for rel in rels:
                    key = int(rel.get('order_product_id'))
                    aggregate.setdefault(key, {
                        'order_product_id': key,
                        'order_sku': rel.get('sku') or '',
                        'quantity': 0,
                        'source_type': 'auto',
                        'shipping_plan_id': None
                    })
                    aggregate[key]['quantity'] += qty * (self._parse_int(rel.get('quantity')) or 1)

        items = list(aggregate.values())
        items.sort(key=lambda x: (x.get('order_sku') or '', x.get('order_product_id') or 0))
        return items

    def _registration_parse_date(self, value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')
        text = str(value).strip()
        if not text:
            return None
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
            except Exception:
                continue
        return None

    def _registration_parse_item_text(self, raw_text):
        items = []
        if raw_text is None:
            return items
        text = str(raw_text).strip()
        if not text:
            return items
        for token in re.split(r'[\n;；|]+', text):
            token = (token or '').strip()
            if not token:
                continue
            if '*' in token:
                left, right = token.split('*', 1)
                sku = (left or '').strip()
                qty = self._parse_int(right)
                qty = max(1, qty or 1)
            else:
                sku = token
                qty = 1
            if not sku:
                continue
            items.append({'sku': sku, 'quantity': qty})
        return items

    def _registration_parse_logistics_text(self, raw_text):
        items = []
        if raw_text is None:
            return items
        text = str(raw_text).strip()
        if not text:
            return items
        for index, token in enumerate(re.split(r'[\n;；|]+', text), start=1):
            token = (token or '').strip()
            if not token:
                continue
            if ':' in token:
                carrier, tracking = token.split(':', 1)
            else:
                carrier, tracking = '', token
            items.append({
                'shipping_carrier': (carrier or '').strip() or None,
                'tracking_no': (tracking or '').strip() or None,
                'sort_order': index
            })
        return items

    def _registration_save_children(self, conn, registration_id, platform_items, shipment_items, logistics_items):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM sales_order_registration_platform_items WHERE registration_id=%s", (registration_id,))
            cur.execute("DELETE FROM sales_order_registration_shipment_items WHERE registration_id=%s", (registration_id,))
            cur.execute("DELETE FROM sales_order_registration_logistics_items WHERE registration_id=%s", (registration_id,))

            if platform_items:
                cur.executemany(
                    """
                    INSERT INTO sales_order_registration_platform_items
                        (registration_id, sales_product_id, platform_sku, quantity, shipping_plan_id)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    [
                        (
                            registration_id,
                            self._parse_int(item.get('sales_product_id')),
                            (item.get('platform_sku') or '').strip(),
                            max(1, self._parse_int(item.get('quantity')) or 1),
                            self._parse_int(item.get('shipping_plan_id'))
                        )
                        for item in platform_items
                    ]
                )

            if shipment_items:
                cur.executemany(
                    """
                    INSERT INTO sales_order_registration_shipment_items
                        (registration_id, order_product_id, order_sku, quantity, source_type, shipping_plan_id)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    [
                        (
                            registration_id,
                            self._parse_int(item.get('order_product_id')),
                            (item.get('order_sku') or '').strip(),
                            max(1, self._parse_int(item.get('quantity')) or 1),
                            (item.get('source_type') or 'manual').strip().lower() if (item.get('source_type') or '').strip().lower() in ('manual', 'auto', 'plan') else 'manual',
                            self._parse_int(item.get('shipping_plan_id'))
                        )
                        for item in shipment_items
                    ]
                )

            if logistics_items:
                cur.executemany(
                    """
                    INSERT INTO sales_order_registration_logistics_items
                        (registration_id, shipping_carrier, tracking_no, sort_order)
                    VALUES (%s, %s, %s, %s)
                    """,
                    [
                        (
                            registration_id,
                            (item.get('shipping_carrier') or '').strip() or None,
                            (item.get('tracking_no') or '').strip() or None,
                            max(1, self._parse_int(item.get('sort_order')) or 1)
                        )
                        for item in logistics_items
                    ]
                )

    def _registration_fill_item_ids(self, conn, platform_items, shipment_items):
        with conn.cursor() as cur:
            sku_to_sales_id = {}
            if platform_items:
                platform_skus = sorted({(x.get('platform_sku') or '').strip() for x in platform_items if (x.get('platform_sku') or '').strip()})
                if platform_skus:
                    placeholders = ','.join(['%s'] * len(platform_skus))
                    cur.execute(f"SELECT id, platform_sku FROM sales_products WHERE platform_sku IN ({placeholders})", tuple(platform_skus))
                    for row in (cur.fetchall() or []):
                        sku_to_sales_id[(row.get('platform_sku') or '').strip()] = self._parse_int(row.get('id'))
                for item in platform_items:
                    if not self._parse_int(item.get('sales_product_id')):
                        item['sales_product_id'] = sku_to_sales_id.get((item.get('platform_sku') or '').strip())

            sku_to_order_id = {}
            if shipment_items:
                order_skus = sorted({(x.get('order_sku') or '').strip() for x in shipment_items if (x.get('order_sku') or '').strip()})
                if order_skus:
                    placeholders = ','.join(['%s'] * len(order_skus))
                    cur.execute(f"SELECT id, sku FROM order_products WHERE sku IN ({placeholders})", tuple(order_skus))
                    for row in (cur.fetchall() or []):
                        sku_to_order_id[(row.get('sku') or '').strip()] = self._parse_int(row.get('id'))
                for item in shipment_items:
                    if not self._parse_int(item.get('order_product_id')):
                        item['order_product_id'] = sku_to_order_id.get((item.get('order_sku') or '').strip())

    def _registration_fetch_detail(self, conn, item_id):
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT r.*, s.shop_name
                FROM sales_order_registrations r
                LEFT JOIN shops s ON s.id = r.shop_id
                WHERE r.id=%s
                LIMIT 1
                """,
                (item_id,)
            )
            row = cur.fetchone() or None
            if not row:
                return None

            cur.execute(
                """
                SELECT id, sales_product_id, platform_sku, quantity, shipping_plan_id
                FROM sales_order_registration_platform_items
                WHERE registration_id=%s
                ORDER BY id ASC
                """,
                (item_id,)
            )
            row['platform_items'] = cur.fetchall() or []

            cur.execute(
                """
                SELECT id, order_product_id, order_sku, quantity, source_type, shipping_plan_id
                FROM sales_order_registration_shipment_items
                WHERE registration_id=%s
                ORDER BY id ASC
                """,
                (item_id,)
            )
            row['shipment_items'] = cur.fetchall() or []

            cur.execute(
                """
                SELECT id, shipping_carrier, tracking_no, sort_order
                FROM sales_order_registration_logistics_items
                WHERE registration_id=%s
                ORDER BY sort_order ASC, id ASC
                """,
                (item_id,)
            )
            row['logistics_items'] = cur.fetchall() or []

            return row

    def handle_sales_order_registration_api(self, environ, method, start_response):
        perf_ctx = self._perf_begin('sales_order_registration_api', environ, {'entry_method': method})
        try:
            self._perf_mark(perf_ctx, 'ensure_sales_order_registration_tables')
            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            action = (query_params.get('action', [''])[0] or '').strip().lower()
            self._perf_mark(perf_ctx, f'parse_query_action:{action or "none"}')

            if method == 'GET' and action == 'options':
                scope = (query_params.get('scope', ['all'])[0] or 'all').strip().lower()
                limit = max(50, min(self._parse_int(query_params.get('limit', ['300'])[0]) or 300, 1000))
                def _load_options_payload():
                    payload = {'status': 'success'}
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            if scope in ('all', 'shops'):
                                cur.execute("SELECT id, shop_name FROM shops ORDER BY shop_name ASC LIMIT %s", (limit,))
                                payload['shops'] = cur.fetchall() or []
                            if scope == 'all':
                                cur.execute("SELECT id, platform_sku FROM sales_products ORDER BY platform_sku ASC LIMIT %s", (limit,))
                                payload['sales_products'] = cur.fetchall() or []
                                cur.execute("SELECT id, sku FROM order_products WHERE COALESCE(is_reship_accessory, 0)=0 ORDER BY sku ASC LIMIT %s", (limit,))
                                payload['order_products'] = cur.fetchall() or []
                                try:
                                    cur.execute(
                                        """
                                        SELECT ops.id, op.sku AS order_sku, ops.plan_name
                                        FROM order_product_shipping_plans ops
                                        JOIN order_products op ON op.id = ops.order_product_id
                                        WHERE COALESCE(op.is_reship_accessory, 0)=0
                                        ORDER BY op.sku ASC, ops.plan_name ASC
                                        LIMIT %s
                                        """,
                                        (limit,)
                                    )
                                    payload['shipping_plans'] = cur.fetchall() or []
                                except Exception:
                                    payload['shipping_plans'] = []
                    return payload
                payload = self._get_cached_template_options(f'sales_order_registration_options_{scope}_{limit}', _load_options_payload, ttl_seconds=1800)
                self._perf_mark(perf_ctx, 'get_options_payload')
                return self.send_json(payload, start_response)

            if method == 'POST' and action == 'shipment_preview':
                data = self._read_json_body(environ)
                platform_items = self._normalize_registration_platform_items(data.get('platform_items'))
                if not platform_items:
                    return self.send_json({'status': 'success', 'auto_shipments': [], 'replacement_options': {}}, start_response)

                with self._get_db_connection() as conn:
                    self._registration_fill_item_ids(conn, platform_items, [])
                    auto_shipments = self._resolve_registration_auto_shipments(conn, platform_items)
                    base_ids = [self._parse_int(x.get('order_product_id')) for x in auto_shipments if self._parse_int(x.get('order_product_id'))]
                    replacement_options = self._registration_get_replacement_options(conn, base_ids)

                return self.send_json({
                    'status': 'success',
                    'auto_shipments': auto_shipments,
                    'replacement_options': replacement_options
                }, start_response)

            if method == 'POST' and action == 'quick_update':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)

                updatable = {}
                if 'shipping_status' in data:
                    status = (data.get('shipping_status') or 'pending').strip().lower()
                    if status not in ('pending', 'unshipped', 'shipped', 'cancelled'):
                        return self.send_json({'status': 'error', 'message': '无效运输状态'}, start_response)
                    updatable['shipping_status'] = status
                if 'is_review_invited' in data:
                    updatable['is_review_invited'] = self._bool_from_any(data.get('is_review_invited'))
                if 'is_logistics_emailed' in data:
                    updatable['is_logistics_emailed'] = self._bool_from_any(data.get('is_logistics_emailed'))
                if 'compensation_action' in data:
                    updatable['compensation_action'] = (data.get('compensation_action') or '').strip() or None
                if 'remark' in data:
                    updatable['remark'] = (data.get('remark') or '').strip() or None

                if not updatable:
                    return self.send_json({'status': 'error', 'message': 'No fields to update'}, start_response)

                set_sql = []
                params = []
                for key, val in updatable.items():
                    set_sql.append(f"{key}=%s")
                    params.append(val)
                set_sql.append("updated_at=CURRENT_TIMESTAMP")
                params.append(item_id)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "UPDATE sales_order_registrations SET " + ', '.join(set_sql) + " WHERE id=%s",
                            tuple(params)
                        )
                        if cur.rowcount <= 0:
                            return self.send_json({'status': 'error', 'message': '记录不存在或未变更'}, start_response)

                return self.send_json({'status': 'success', 'id': item_id}, start_response)

            if method == 'GET' and action == 'summaries':
                ids = []
                for raw in query_params.get('ids', []):
                    for token in re.split(r'[,，;；\s]+', str(raw or '').strip()):
                        val = self._parse_int(token)
                        if val:
                            ids.append(val)
                ids = list(dict.fromkeys(ids))
                if not ids:
                    return self.send_json({'status': 'success', 'items': []}, start_response)

                with self._get_db_connection() as conn:
                    self._perf_mark(perf_ctx, 'db_connected_for_summaries')
                    with conn.cursor() as cur:
                        placeholders = ','.join(['%s'] * len(ids))
                        cur.execute(
                            f"""
                            SELECT registration_id, platform_sku, quantity
                            FROM sales_order_registration_platform_items
                            WHERE registration_id IN ({placeholders})
                            ORDER BY registration_id ASC, id ASC
                            """,
                            tuple(ids)
                        )
                        platform_rows = cur.fetchall() or []

                        cur.execute(
                            f"""
                            SELECT registration_id, order_sku, quantity
                            FROM sales_order_registration_shipment_items
                            WHERE registration_id IN ({placeholders})
                            ORDER BY registration_id ASC, id ASC
                            """,
                            tuple(ids)
                        )
                        shipment_rows = cur.fetchall() or []

                        cur.execute(
                            f"""
                            SELECT registration_id, shipping_carrier, tracking_no
                            FROM sales_order_registration_logistics_items
                            WHERE registration_id IN ({placeholders})
                            ORDER BY registration_id ASC, sort_order ASC, id ASC
                            """,
                            tuple(ids)
                        )
                        logistics_rows = cur.fetchall() or []
                    self._perf_mark(perf_ctx, 'load_summaries_rows')

                out = {int(i): {'id': int(i), 'platform_summary': [], 'shipment_summary': [], 'logistics_summary': []} for i in ids}
                for row in platform_rows:
                    rid = self._parse_int(row.get('registration_id'))
                    if rid in out:
                        out[rid]['platform_summary'].append(f"{(row.get('platform_sku') or '').strip()}*{self._parse_int(row.get('quantity')) or 1}")
                for row in shipment_rows:
                    rid = self._parse_int(row.get('registration_id'))
                    if rid in out:
                        out[rid]['shipment_summary'].append(f"{(row.get('order_sku') or '').strip()}*{self._parse_int(row.get('quantity')) or 1}")
                for row in logistics_rows:
                    rid = self._parse_int(row.get('registration_id'))
                    if rid in out:
                        carrier = (row.get('shipping_carrier') or '').strip()
                        tracking = (row.get('tracking_no') or '').strip()
                        out[rid]['logistics_summary'].append(f"{carrier}:{tracking}" if carrier else tracking)

                return self.send_json({'status': 'success', 'items': [out[x] for x in ids]}, start_response)

            if method == 'GET':
                item_id = self._parse_int(query_params.get('id', [''])[0])
                keyword = (query_params.get('q', [''])[0] or '').strip()
                include_summaries = str(query_params.get('include_summaries', ['1'])[0] or '1').strip().lower() not in ('0', 'false', 'no', 'off')
                page = max(1, self._parse_int(query_params.get('page', ['1'])[0]) or 1)
                page_size = max(20, min(self._parse_int(query_params.get('page_size', ['50'])[0]) or 50, 200))
                offset = (page - 1) * page_size

                with self._get_db_connection() as conn:
                    self._perf_mark(perf_ctx, 'db_connected_for_get')
                    if item_id:
                        item = self._registration_fetch_detail(conn, item_id)
                        self._perf_mark(perf_ctx, 'fetch_detail')
                        return self.send_json({'status': 'success', 'item': item}, start_response)

                    with conn.cursor() as cur:
                        filters = []
                        params = []
                        if keyword:
                            like = f"%{keyword}%"
                            filters.append("(r.order_no LIKE %s OR r.customer_name LIKE %s OR r.phone LIKE %s OR s.shop_name LIKE %s)")
                            params.extend([like, like, like, like])
                        where_sql = (' WHERE ' + ' AND '.join(filters)) if filters else ''

                        cur.execute(
                            "SELECT COUNT(*) AS total FROM sales_order_registrations r LEFT JOIN shops s ON s.id = r.shop_id" + where_sql,
                            tuple(params)
                        )
                        total = int((cur.fetchone() or {}).get('total') or 0)
                        self._perf_mark(perf_ctx, 'list_count_query')

                        if include_summaries:
                            cur.execute(
                                """
                                SELECT r.id, r.shop_id, r.order_no, r.order_date, r.customer_name, r.shipping_status,
                                        r.is_review_invited, r.is_logistics_emailed, r.compensation_action, r.remark,
                                       r.created_at, r.updated_at, s.shop_name, s.shop_name AS shop_display_name,
                                       (
                                           SELECT GROUP_CONCAT(CONCAT(COALESCE(p.platform_sku, ''), '*', COALESCE(p.quantity, 1)) ORDER BY p.id ASC SEPARATOR '\n')
                                           FROM sales_order_registration_platform_items p
                                           WHERE p.registration_id = r.id
                                       ) AS platform_summary_text,
                                       (
                                           SELECT GROUP_CONCAT(CONCAT(COALESCE(si.order_sku, ''), '*', COALESCE(si.quantity, 1)) ORDER BY si.id ASC SEPARATOR '\n')
                                           FROM sales_order_registration_shipment_items si
                                           WHERE si.registration_id = r.id
                                       ) AS shipment_summary_text,
                                       (
                                           SELECT GROUP_CONCAT(
                                               CASE
                                                   WHEN COALESCE(li.shipping_carrier, '') <> ''
                                                   THEN CONCAT(li.shipping_carrier, ':', COALESCE(li.tracking_no, ''))
                                                   ELSE COALESCE(li.tracking_no, '')
                                               END
                                               ORDER BY li.sort_order ASC, li.id ASC SEPARATOR '\n'
                                           )
                                           FROM sales_order_registration_logistics_items li
                                           WHERE li.registration_id = r.id
                                       ) AS logistics_summary_text
                                FROM sales_order_registrations r
                                LEFT JOIN shops s ON s.id = r.shop_id
                                """ + where_sql + " ORDER BY r.id DESC LIMIT %s OFFSET %s",
                                tuple(params + [page_size, offset])
                            )
                        else:
                            cur.execute(
                                """
                                SELECT r.id, r.shop_id, r.order_no, r.order_date, r.customer_name, r.shipping_status,
                                        r.is_review_invited, r.is_logistics_emailed, r.compensation_action, r.remark,
                                       r.created_at, r.updated_at, s.shop_name, s.shop_name AS shop_display_name
                                FROM sales_order_registrations r
                                LEFT JOIN shops s ON s.id = r.shop_id
                                """ + where_sql + " ORDER BY r.id DESC LIMIT %s OFFSET %s",
                                tuple(params + [page_size, offset])
                            )
                        rows = cur.fetchall() or []
                        self._perf_mark(perf_ctx, 'list_rows_query')

                for row in rows:
                    if include_summaries:
                        platform_text = (row.get('platform_summary_text') or '').strip()
                        shipment_text = (row.get('shipment_summary_text') or '').strip()
                        logistics_text = (row.get('logistics_summary_text') or '').strip()
                        row['platform_summary'] = [x for x in platform_text.split('\n') if x] if platform_text else []
                        row['shipment_summary'] = [x for x in shipment_text.split('\n') if x] if shipment_text else []
                        row['logistics_summary'] = [x for x in logistics_text.split('\n') if x] if logistics_text else []
                        row.pop('platform_summary_text', None)
                        row.pop('shipment_summary_text', None)
                        row.pop('logistics_summary_text', None)
                    else:
                        row['platform_summary'] = []
                        row['shipment_summary'] = []
                        row['logistics_summary'] = []
                self._perf_mark(perf_ctx, 'list_rows_transform')

                return self.send_json(
                    {'status': 'success', 'items': rows, 'page': page, 'page_size': page_size, 'total': total},
                    start_response
                )

            data = self._read_json_body(environ)

            if method in ('POST', 'PUT'):
                if method == 'PUT':
                    item_id = self._parse_int(data.get('id'))
                    if not item_id:
                        return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                else:
                    item_id = None

                order_no = (data.get('order_no') or '').strip()
                if not order_no:
                    return self.send_json({'status': 'error', 'message': '订单号必填'}, start_response)

                shop_id = self._parse_int(data.get('shop_id'))
                order_date = self._registration_parse_date(data.get('order_date'))
                customer_name = (data.get('customer_name') or '').strip() or None
                phone = (data.get('phone') or '').strip() or None
                zip_code = (data.get('zip_code') or '').strip() or None
                address = (data.get('address') or '').strip() or None
                city = (data.get('city') or '').strip() or None
                state = (data.get('state') or '').strip() or None
                shipping_status = (data.get('shipping_status') or 'pending').strip() or 'pending'
                compensation_action = (data.get('compensation_action') or '').strip() or None
                remark = (data.get('remark') or '').strip() or None

                self._validate_us_phone_zip(phone, zip_code)

                platform_items = self._normalize_registration_platform_items(data.get('platform_items'))
                if not platform_items:
                    return self.send_json({'status': 'error', 'message': '请至少填写1条销售平台SKU'}, start_response)
                shipment_items = self._normalize_registration_shipment_items(data.get('shipment_items'))
                logistics_items = self._normalize_registration_logistics_items(data.get('logistics_items'))

                with self._get_db_connection() as conn:
                    self._perf_mark(perf_ctx, 'db_connected_for_write')
                    self._registration_fill_item_ids(conn, platform_items, shipment_items)
                    if not shipment_items:
                        shipment_items = self._resolve_registration_auto_shipments(conn, platform_items)
                    self._perf_mark(perf_ctx, 'prepare_write_payload')

                    with conn.cursor() as cur:
                        if method == 'POST':
                            cur.execute(
                                """
                                INSERT INTO sales_order_registrations
                                    (shop_id, order_no, order_date, customer_name, phone, zip_code, address, city, state,
                                     shipping_status, is_review_invited, is_logistics_emailed, compensation_action, remark)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                """,
                                (
                                    shop_id,
                                    order_no,
                                    order_date,
                                    customer_name,
                                    phone,
                                    zip_code,
                                    address,
                                    city,
                                    state,
                                    shipping_status,
                                    self._bool_from_any(data.get('is_review_invited')),
                                    self._bool_from_any(data.get('is_logistics_emailed')),
                                    compensation_action,
                                    remark
                                )
                            )
                            item_id = cur.lastrowid
                        else:
                            cur.execute(
                                """
                                UPDATE sales_order_registrations
                                SET shop_id=%s,
                                    order_no=%s,
                                    order_date=%s,
                                    customer_name=%s,
                                    phone=%s,
                                    zip_code=%s,
                                    address=%s,
                                    city=%s,
                                    state=%s,
                                    shipping_status=%s,
                                    is_review_invited=%s,
                                    is_logistics_emailed=%s,
                                    compensation_action=%s,
                                    remark=%s
                                WHERE id=%s
                                """,
                                (
                                    shop_id,
                                    order_no,
                                    order_date,
                                    customer_name,
                                    phone,
                                    zip_code,
                                    address,
                                    city,
                                    state,
                                    shipping_status,
                                    self._bool_from_any(data.get('is_review_invited')),
                                    self._bool_from_any(data.get('is_logistics_emailed')),
                                    compensation_action,
                                    remark,
                                    item_id
                                )
                            )

                    self._registration_save_children(conn, item_id, platform_items, shipment_items, logistics_items)
                    self._perf_mark(perf_ctx, 'write_registration_and_children')

                return self.send_json({'status': 'success', 'id': item_id}, start_response)

            if method == 'DELETE':
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM sales_order_registrations WHERE id=%s", (item_id,))
                self._perf_mark(perf_ctx, 'delete_registration')
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except ValueError as ve:
            return self.send_json({'status': 'error', 'message': str(ve)}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        finally:
            self._perf_end(perf_ctx)

    def handle_sales_order_registration_template_api(self, environ, method, start_response):
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = 'orders'

            headers = [
                '店铺', '订单号', '订单日期', '订单状态（原运输状态）',
                '销售平台SKU', '实际发货SKU',
                '姓名', '电话(US)', '地址', '城市', '州', '邮编(US)', '跟踪号（可多条）',
                '是否已发物流邮件（是/否）', '是否邀评（是/否）', '补偿措施', '备注'
            ]
            groups = [
                ('订单基础信息', 1, 4),
                ('SKU信息', 5, 6),
                ('物流信息', 7, 13),
                ('售后状态', 14, 17),
            ]
            group_colors = [
                ('A8B9A5', 'DDE7DB'),
                ('D7C894', 'ECE5CE'),
                ('B8C3D6', 'E3E8F2'),
                ('D8B7C5', 'F0E3E8'),
            ]

            header_font = Font(bold=True, color='2A2420')
            example_font = Font(italic=True, color='7B8088')
            thin_border = Border(
                left=Side(style='thin', color='B7AEA4'),
                right=Side(style='thin', color='B7AEA4'),
                top=Side(style='thin', color='B7AEA4'),
                bottom=Side(style='thin', color='B7AEA4')
            )

            header_fill_by_col = ['DDE7DB'] * len(headers)
            for idx, (title, start_col, end_col) in enumerate(groups):
                title_color, sub_header_color = group_colors[idx % len(group_colors)]
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws.cell(row=1, column=start_col, value=title)
                cell.fill = PatternFill(start_color=title_color, end_color=title_color, fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for col in range(start_col, end_col + 1):
                    header_fill_by_col[col - 1] = sub_header_color

            for col, title in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col, value=title)
                cell.fill = PatternFill(start_color=header_fill_by_col[col - 1], end_color=header_fill_by_col[col - 1], fill_type='solid')
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                width = 18
                if col in (1, 2):
                    width = 16
                if col in (5, 6, 13):
                    width = 24
                if col in (9, 17):
                    width = 28
                ws.column_dimensions[get_column_letter(col)].width = width

            ws.append([
                '示例店铺', 'SO-20250101-001', '2025-01-01', 'pending',
                'MS01-Brown-1A*1|MS01-Gray-1A*1', '',
                'John Doe', '+1 415-888-9999', '123 Main St', 'Los Angeles', 'CA', '90001', 'UPS:1Z123|FedEx:999',
                '否', '是', '示例补偿', '示例备注'
            ])
            for cell in ws[3]:
                cell.fill = PatternFill(start_color='ECECEC', end_color='ECECEC', fill_type='solid')
                cell.font = example_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            dv_yes_no = DataValidation(type='list', formula1='"是,否"', allow_blank=True)
            ws.add_data_validation(dv_yes_no)
            for row in range(4, 1201):
                dv_yes_no.add(f'N{row}')
                dv_yes_no.add(f'O{row}')

            ws.freeze_panes = 'A4'
            return self._send_excel_workbook(wb, '订单登记导入模板.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_sales_order_registration_import_api(self, environ, method, start_response):
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            first_row = [str(x.value or '').strip() for x in ws[1]]
            module_titles = {'订单基础信息', 'SKU信息', '物流信息', '售后状态'}
            header_row_idx = 2 if any(cell in module_titles for cell in first_row) else 1
            data_start_row = 4 if header_row_idx == 2 else 2
            headers = [str(x.value or '').strip() for x in ws[header_row_idx]]
            index_map = {name: idx for idx, name in enumerate(headers) if name}

            alias_map = {
                'shop_name': ['店铺'],
                'order_no': ['订单号'],
                'order_date': ['订单日期', '下单日期(YYYY-MM-DD)'],
                'customer_name': ['姓名', '客户姓名'],
                'phone': ['电话(US)', '电话'],
                'zip_code': ['邮编(US)', '邮编'],
                'address': ['地址'],
                'city': ['城市'],
                'state': ['州'],
                'shipping_status': ['订单状态（原运输状态）', '发货状态(pending/shipped/delivered/cancelled)'],
                'is_logistics_emailed': ['是否已发物流邮件（是/否）', '是否已发物流邮件', '物流已邮件(0/1)'],
                'is_review_invited': ['是否邀评（是/否）', '是否邀评', '邀评(0/1)'],
                'compensation_action': ['补偿措施', '赔偿处理'],
                'remark': ['备注'],
                'platform_text': ['销售平台SKU', '平台SKU明细(平台SKU*数量, 用|分隔)'],
                'shipment_text': ['实际发货SKU', '发货SKU明细(下单SKU*数量, 用|分隔, 留空则自动计算)'],
                'logistics_text': ['跟踪号（可多条）', '跟踪号（原物流行，可多条）', '物流明细(承运商:单号, 用|分隔)']
            }

            def get_val(row_values, aliases, default=None):
                for key in aliases:
                    idx = index_map.get(key)
                    if idx is None or idx >= len(row_values):
                        continue
                    value = row_values[idx]
                    if value is None:
                        continue
                    if str(value).strip() == '':
                        continue
                    return value
                return default

            required_aliases = [
                ('订单号', alias_map['order_no']),
                ('销售平台SKU', alias_map['platform_text'])
            ]
            for label, aliases in required_aliases:
                if not any(alias in index_map for alias in aliases):
                    return self.send_json({'status': 'error', 'message': f'模板缺少列: {label}'}, start_response)

            def _aggregate_sku_items(parsed_items):
                qty_by_sku = {}
                sku_order = []
                for item in (parsed_items or []):
                    sku = str((item or {}).get('sku') or '').strip()
                    if not sku:
                        continue
                    qty = max(1, self._parse_int((item or {}).get('quantity')) or 1)
                    if sku not in qty_by_sku:
                        qty_by_sku[sku] = 0
                        sku_order.append(sku)
                    qty_by_sku[sku] += qty
                return [{'sku': sku, 'quantity': qty_by_sku[sku]} for sku in sku_order]

            created = 0
            updated = 0
            errors = []

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id, shop_name FROM shops")
                    shop_map = {str(x.get('shop_name') or '').strip(): self._parse_int(x.get('id')) for x in (cur.fetchall() or [])}

                staged_map = {}
                staged_keys = []

                for row_idx in range(data_start_row, ws.max_row + 1):
                    row_values = [cell.value for cell in ws[row_idx]]
                    if not any(v is not None and str(v).strip() for v in row_values):
                        continue
                    try:
                        shop_name = str(get_val(row_values, alias_map['shop_name']) or '').strip()
                        order_no = str(get_val(row_values, alias_map['order_no']) or '').strip()
                        if not order_no:
                            raise ValueError('订单号必填')

                        shop_id = shop_map.get(shop_name) if shop_name else None
                        order_date = self._registration_parse_date(get_val(row_values, alias_map['order_date']))
                        customer_name = str(get_val(row_values, alias_map['customer_name']) or '').strip() or None
                        phone = str(get_val(row_values, alias_map['phone']) or '').strip() or None
                        zip_code = str(get_val(row_values, alias_map['zip_code']) or '').strip() or None
                        address = str(get_val(row_values, alias_map['address']) or '').strip() or None
                        city = str(get_val(row_values, alias_map['city']) or '').strip() or None
                        state = str(get_val(row_values, alias_map['state']) or '').strip() or None
                        shipping_status = str(get_val(row_values, alias_map['shipping_status'], 'pending') or 'pending').strip() or 'pending'
                        is_review_invited = self._bool_from_any(get_val(row_values, alias_map['is_review_invited']))
                        is_logistics_emailed = self._bool_from_any(get_val(row_values, alias_map['is_logistics_emailed']))
                        compensation_action = str(get_val(row_values, alias_map['compensation_action']) or '').strip() or None
                        remark = str(get_val(row_values, alias_map['remark']) or '').strip() or None

                        self._validate_us_phone_zip(phone, zip_code)

                        platform_text = get_val(row_values, alias_map['platform_text'])
                        shipment_text = get_val(row_values, alias_map['shipment_text'])
                        logistics_text = get_val(row_values, alias_map['logistics_text'])

                        parsed_platform = self._registration_parse_item_text(platform_text)
                        if not parsed_platform:
                            raise ValueError('平台SKU明细必填')
                        parsed_shipment = self._registration_parse_item_text(shipment_text)
                        parsed_logistics = self._registration_parse_logistics_text(logistics_text)

                        group_key = (
                            shop_id or 0,
                            order_no,
                            order_date or '',
                            customer_name or '',
                            phone or '',
                            zip_code or '',
                            address or '',
                            city or '',
                            state or '',
                            shipping_status or 'pending',
                            int(is_review_invited or 0),
                            int(is_logistics_emailed or 0),
                            compensation_action or '',
                            remark or ''
                        )

                        if group_key not in staged_map:
                            staged_map[group_key] = {
                                'row': row_idx,
                                'shop_id': shop_id,
                                'order_no': order_no,
                                'order_date': order_date,
                                'customer_name': customer_name,
                                'phone': phone,
                                'zip_code': zip_code,
                                'address': address,
                                'city': city,
                                'state': state,
                                'shipping_status': shipping_status,
                                'is_review_invited': is_review_invited,
                                'is_logistics_emailed': is_logistics_emailed,
                                'compensation_action': compensation_action,
                                'remark': remark,
                                'platform_parsed': [],
                                'shipment_parsed': [],
                                'logistics_items': []
                            }
                            staged_keys.append(group_key)

                        staged_map[group_key]['platform_parsed'].extend(parsed_platform)
                        staged_map[group_key]['shipment_parsed'].extend(parsed_shipment)
                        staged_map[group_key]['logistics_items'].extend(parsed_logistics)
                    except Exception as row_error:
                        errors.append({'row': row_idx, 'error': str(row_error)})

                for key in staged_keys:
                    staged = staged_map.get(key) or {}
                    try:
                        platform_agg = _aggregate_sku_items(staged.get('platform_parsed'))
                        if not platform_agg:
                            raise ValueError('平台SKU明细必填')
                        shipment_agg = _aggregate_sku_items(staged.get('shipment_parsed'))

                        platform_items = [
                            {
                                'sales_product_id': None,
                                'platform_sku': x['sku'],
                                'quantity': x['quantity'],
                                'shipping_plan_id': None
                            }
                            for x in platform_agg
                        ]
                        shipment_items = [
                            {
                                'order_product_id': None,
                                'order_sku': x['sku'],
                                'quantity': x['quantity'],
                                'source_type': 'manual',
                                'shipping_plan_id': None
                            }
                            for x in shipment_agg
                        ]

                        logistics_items = []
                        for item in (staged.get('logistics_items') or []):
                            shipping_carrier = (item.get('shipping_carrier') or '').strip()
                            tracking_no = (item.get('tracking_no') or '').strip()
                            if not shipping_carrier and not tracking_no:
                                continue
                            logistics_items.append({
                                'shipping_carrier': shipping_carrier or None,
                                'tracking_no': tracking_no or None,
                                'sort_order': len(logistics_items) + 1
                            })

                        self._registration_fill_item_ids(conn, platform_items, shipment_items)
                        if not shipment_items:
                            shipment_items = self._resolve_registration_auto_shipments(conn, platform_items)

                        with conn.cursor() as cur:
                            if staged.get('shop_id'):
                                cur.execute(
                                    "SELECT id FROM sales_order_registrations WHERE order_no=%s AND shop_id=%s LIMIT 1",
                                    (staged.get('order_no'), staged.get('shop_id'))
                                )
                            else:
                                cur.execute(
                                    "SELECT id FROM sales_order_registrations WHERE order_no=%s AND shop_id IS NULL LIMIT 1",
                                    (staged.get('order_no'),)
                                )
                            existing = cur.fetchone() or {}
                            existing_id = self._parse_int(existing.get('id'))

                            if existing_id:
                                cur.execute(
                                    """
                                    UPDATE sales_order_registrations
                                    SET shop_id=%s, order_date=%s, customer_name=%s, phone=%s, zip_code=%s, address=%s, city=%s, state=%s,
                                        shipping_status=%s, is_review_invited=%s, is_logistics_emailed=%s, compensation_action=%s, remark=%s
                                    WHERE id=%s
                                    """,
                                    (
                                        staged.get('shop_id'), staged.get('order_date'), staged.get('customer_name'), staged.get('phone'),
                                        staged.get('zip_code'), staged.get('address'), staged.get('city'), staged.get('state'),
                                        staged.get('shipping_status'), staged.get('is_review_invited'), staged.get('is_logistics_emailed'),
                                        staged.get('compensation_action'), staged.get('remark'), existing_id
                                    )
                                )
                                registration_id = existing_id
                                updated += 1
                            else:
                                cur.execute(
                                    """
                                    INSERT INTO sales_order_registrations
                                        (shop_id, order_no, order_date, customer_name, phone, zip_code, address, city, state,
                                         shipping_status, is_review_invited, is_logistics_emailed, compensation_action, remark)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                    """,
                                    (
                                        staged.get('shop_id'), staged.get('order_no'), staged.get('order_date'), staged.get('customer_name'),
                                        staged.get('phone'), staged.get('zip_code'), staged.get('address'), staged.get('city'), staged.get('state'),
                                        staged.get('shipping_status'), staged.get('is_review_invited'), staged.get('is_logistics_emailed'),
                                        staged.get('compensation_action'), staged.get('remark')
                                    )
                                )
                                registration_id = cur.lastrowid
                                created += 1

                        self._registration_save_children(conn, registration_id, platform_items, shipment_items, logistics_items)
                    except Exception as group_error:
                        errors.append({'row': staged.get('row') or 0, 'error': str(group_error)})

            return self.send_json({'status': 'success', 'created': created, 'updated': updated, 'errors': errors}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    # ==============================================
    # 销量预测（Sales Forecast）
    # ==============================================

    FORECAST_MODES = ('platform', 'spec', 'order')
    FORECAST_PRODUCT_STATUSES = ('enabled', 'retained', 'discarded')

    def _forecast_normalize_month(self, value):
        """将输入归一化为本月首日的 YYYY-MM-01 字符串。支持 YYYY-MM、YYYY-MM-DD、YYYY/MM。"""
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        candidates = ('%Y-%m', '%Y/%m', '%Y-%m-%d', '%Y/%m/%d')
        for fmt in candidates:
            try:
                dt = datetime.strptime(text, fmt)
                return dt.strftime('%Y-%m-01')
            except Exception:
                continue
        match = re.match(r'^(\d{4})\D+(\d{1,2})', text)
        if match:
            try:
                year = int(match.group(1))
                month = int(match.group(2))
                if 1 <= month <= 12:
                    return f'{year:04d}-{month:02d}-01'
            except Exception:
                pass
        return None

    def _forecast_parse_months(self, raw):
        """解析 months=YYYY-MM,YYYY-MM,... 形式参数，去重并按时间顺序排序。"""
        results = []
        seen = set()
        if raw is None:
            return results
        if isinstance(raw, (list, tuple, set)):
            tokens = []
            for item in raw:
                if item is None:
                    continue
                tokens.extend(str(item).split(','))
        else:
            tokens = str(raw).split(',')
        for token in tokens:
            month_str = self._forecast_normalize_month(token)
            if month_str and month_str not in seen:
                seen.add(month_str)
                results.append(month_str)
        results.sort()
        return results

    def _forecast_default_future_months(self, count=6):
        """默认返回从“本月”起的 count 个月（含本月）。"""
        today = datetime.now()
        months = []
        year = today.year
        month = today.month
        for _ in range(max(1, int(count or 0))):
            months.append(f'{year:04d}-{month:02d}-01')
            month += 1
            if month > 12:
                month = 1
                year += 1
        return months

    def _forecast_history_month_range(self, raw_start, raw_end, default_months=12):
        """将历史月份范围解析为 [start_month, end_month]，闭区间（每个值是月首日）。

        默认：以“上个月”作为 end，向前回溯 default_months 个月作为 start。"""
        end_month = self._forecast_normalize_month(raw_end)
        start_month = self._forecast_normalize_month(raw_start)
        if not end_month:
            # 默认包含「本月」：历史月范围截止到当前月（数据可能仍不完整）
            today = datetime.now()
            end_month = f'{today.year:04d}-{today.month:02d}-01'
        if not start_month:
            try:
                end_dt = datetime.strptime(end_month, '%Y-%m-%d')
            except Exception:
                end_dt = datetime.now()
            year = end_dt.year
            month = end_dt.month - (default_months - 1)
            while month < 1:
                month += 12
                year -= 1
            start_month = f'{year:04d}-{month:02d}-01'
        if start_month > end_month:
            start_month, end_month = end_month, start_month
        return start_month, end_month

    def _forecast_iter_months(self, start_month, end_month):
        if not start_month or not end_month:
            return []
        try:
            sd = datetime.strptime(start_month, '%Y-%m-%d')
            ed = datetime.strptime(end_month, '%Y-%m-%d')
        except Exception:
            return []
        months = []
        y, m = sd.year, sd.month
        while True:
            current = f'{y:04d}-{m:02d}-01'
            months.append(current)
            if y > ed.year or (y == ed.year and m >= ed.month):
                break
            m += 1
            if m > 12:
                m = 1
                y += 1
        return months

    def _forecast_format_dt(self, value):
        if value is None:
            return None
        if hasattr(value, 'strftime'):
            try:
                return value.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                return str(value)
        return str(value)

    def _forecast_month_to_str(self, value, fallback=None):
        if hasattr(value, 'strftime'):
            try:
                return value.strftime('%Y-%m-01')
            except Exception:
                pass
        if value:
            text = str(value)[:7]
            if len(text) == 7:
                return text + '-01'
        return fallback

    def _forecast_normalize_mode(self, raw):
        text = str(raw or '').strip().lower()
        if text in self.FORECAST_MODES:
            return text
        if text in ('platform_sku', 'platformsku', 'sku'):
            return 'platform'
        if text in ('spec_fabric', 'specfabric', 'variant'):
            return 'spec'
        if text in ('order_sku', 'ordersku', 'order_product'):
            return 'order'
        return 'spec'

    def _forecast_load_variants(self, conn, query_params=None):
        """加载预测页所需的规格列表（含货号/规格/面料/关联下单SKU）。"""
        sku_keyword = ''
        spec_keyword = ''
        fabric_keyword = ''
        if query_params:
            sku_keyword = (query_params.get('sku', [''])[0] or '').strip()
            spec_keyword = (query_params.get('spec', [''])[0] or '').strip()
            fabric_keyword = (query_params.get('fabric', [''])[0] or '').strip()

        clauses = ["1=1"]
        params = []
        if sku_keyword:
            clauses.append("pf.sku_family LIKE %s")
            params.append(f"%{sku_keyword}%")
        if spec_keyword:
            clauses.append("v.spec_name LIKE %s")
            params.append(f"%{spec_keyword}%")
        if fabric_keyword:
            clauses.append("(fm.fabric_code LIKE %s OR fm.fabric_name_en LIKE %s)")
            params.append(f"%{fabric_keyword}%")
            params.append(f"%{fabric_keyword}%")

        where_sql = ' AND '.join(clauses)
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT v.id AS variant_id,
                       v.sku_family_id,
                       v.spec_name,
                       v.fabric_id,
                       pf.sku_family,
                       fm.fabric_code,
                       fm.fabric_name_en,
                       fm.representative_color
                FROM sales_product_variants v
                LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id
                WHERE {where_sql}
                ORDER BY pf.sku_family ASC, v.spec_name ASC, v.id ASC
                """,
                tuple(params)
            )
            variants = cur.fetchall() or []

            variant_ids = [self._parse_int(v.get('variant_id')) for v in variants if self._parse_int(v.get('variant_id'))]
            order_links_by_variant = {}
            order_products_map = {}
            if variant_ids:
                placeholders = ','.join(['%s'] * len(variant_ids))
                cur.execute(
                    f"""
                    SELECT l.variant_id, l.order_product_id, l.quantity,
                           op.sku, op.spec_qty_short, op.contents_desc_en, op.is_on_market
                    FROM sales_variant_order_links l
                    JOIN order_products op ON op.id = l.order_product_id
                    WHERE l.variant_id IN ({placeholders})
                    ORDER BY l.variant_id ASC, op.sku ASC
                    """,
                    tuple(variant_ids)
                )
                for row in (cur.fetchall() or []):
                    vid = self._parse_int(row.get('variant_id'))
                    op_id = self._parse_int(row.get('order_product_id'))
                    qty = max(1, self._parse_int(row.get('quantity')) or 1)
                    if not vid or not op_id:
                        continue
                    order_links_by_variant.setdefault(vid, []).append({
                        'order_product_id': op_id,
                        'quantity': qty,
                        'sku': row.get('sku') or '',
                    })
                    order_products_map[op_id] = {
                        'order_product_id': op_id,
                        'sku': row.get('sku') or '',
                        'spec_qty_short': row.get('spec_qty_short') or '',
                        'contents_desc_en': row.get('contents_desc_en') or '',
                        'is_on_market': self._parse_int(row.get('is_on_market')) or 0,
                    }

        for v in variants:
            vid = self._parse_int(v.get('variant_id'))
            v['order_links'] = order_links_by_variant.get(vid, [])
        return variants, order_products_map

    def _forecast_history_end_exclusive(self, end_month):
        try:
            ed_dt = datetime.strptime(end_month, '%Y-%m-%d')
            year = ed_dt.year
            month = ed_dt.month + 1
            if month > 12:
                month = 1
                year += 1
            return f'{year:04d}-{month:02d}-01'
        except Exception:
            return None

    def _forecast_apply_sf_group_platform(self, sf_group, clauses, params):
        """与前端 groupKey 一致：shop||family，各自与 COALESCE(TRIM(...), '-') 比较。"""
        if not sf_group:
            return
        raw = str(sf_group).strip()
        if not raw:
            return
        parts = raw.split('||', 1)
        shop_g = parts[0].strip() if parts else '-'
        fam_g = parts[1].strip() if len(parts) > 1 else '-'
        clauses.append("COALESCE(NULLIF(TRIM(sh.shop_name),''), '-') = %s")
        clauses.append("COALESCE(NULLIF(TRIM(pf.sku_family),''), '-') = %s")
        params.extend([shop_g or '-', fam_g or '-'])

    def _forecast_parse_sf_status_filter(self, query_params):
        """平台/规格维度产品状态筛选；默认仅 enabled。"""
        raw = ''
        if query_params:
            raw = (query_params.get('sf_status', [''])[0] or '').strip().lower()
        if not raw:
            return ['enabled']
        parts = [p.strip() for p in raw.split(',') if p.strip()]
        valid = [p for p in parts if p in self.FORECAST_PRODUCT_STATUSES]
        return valid if valid else ['enabled']

    def _forecast_apply_platform_status_filter(self, clauses, params, statuses):
        if not statuses or set(statuses) >= set(self.FORECAST_PRODUCT_STATUSES):
            return
        ph = ','.join(['%s'] * len(statuses))
        clauses.append(f"COALESCE(NULLIF(TRIM(sp.product_status),''), 'enabled') IN ({ph})")
        params.extend(statuses)

    def _forecast_apply_variant_linked_status_filter(self, clauses, params, statuses):
        """规格维度：至少有一条关联平台 SKU 的状态落在所选集合内。"""
        if not statuses or set(statuses) >= set(self.FORECAST_PRODUCT_STATUSES):
            return
        ph = ','.join(['%s'] * len(statuses))
        clauses.append(
            f"""EXISTS (
                SELECT 1 FROM sales_products sp_sf
                WHERE sp_sf.variant_id = v.id
                  AND COALESCE(NULLIF(TRIM(sp_sf.product_status),''), 'enabled') IN ({ph})
            )"""
        )
        params.extend(statuses)

    def _forecast_apply_spec_sku_keyword_filter(self, sku_keyword, clauses, params, variant_alias='v'):
        """规格维度 SKU 筛选：货号或关联平台 SKU（与筛选框「货号/平台SKU」一致）。"""
        if not sku_keyword:
            return
        like_val = f"%{sku_keyword}%"
        clauses.append(
            f"""(pf.sku_family LIKE %s OR EXISTS (
                SELECT 1 FROM sales_products sp_sf
                WHERE sp_sf.variant_id = {variant_alias}.id
                  AND sp_sf.platform_sku LIKE %s
            ))"""
        )
        params.extend([like_val, like_val])

    @staticmethod
    def _forecast_order_has_on_market_substitute_exists_sql(owner_alias='op'):
        """任一发货方案（不限默认 MIN 方案）是否含在市替代 SKU。"""
        return f"""EXISTS (
            SELECT 1
            FROM order_product_shipping_plans ops0
            INNER JOIN order_product_shipping_plan_items i0 ON i0.shipping_plan_id = ops0.id
            INNER JOIN order_products sub0 ON sub0.id = i0.substitute_order_product_id
            WHERE ops0.order_product_id = {owner_alias}.id
              AND COALESCE(sub0.is_on_market, 0) = 1
        )"""

    @staticmethod
    def _forecast_order_dim_visibility_sql():
        """下单 SKU 维度可见性：仅在市 SKU；下市且任一方案均无在市替代时展示并标异常。"""
        has_on_sub = SalesManagementMixin._forecast_order_has_on_market_substitute_exists_sql('op')
        return f"""(
            COALESCE(op.is_on_market, 0) = 1
            OR (
                COALESCE(op.is_on_market, 0) = 0
                AND NOT ({has_on_sub})
            )
        )"""

    def _forecast_load_order_product_brief_map(self, conn, order_product_ids):
        ids = sorted({int(x) for x in (order_product_ids or []) if self._parse_int(x)})
        out = {}
        if not ids:
            return out
        ph = ','.join(['%s'] * len(ids))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT id, sku, COALESCE(is_on_market, 0) AS is_on_market
                FROM order_products
                WHERE id IN ({ph})
                """,
                tuple(ids),
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('id'))
                if oid:
                    out[oid] = {
                        'sku': rr.get('sku') or '',
                        'is_on_market': self._parse_int(rr.get('is_on_market')) or 0,
                    }
        return out

    def _forecast_load_delisted_owner_refs_by_substitute_ids(self, conn, substitute_ids):
        """下市 owner 的默认替代方案引用到 substitute_ids 时，供备注列使用。"""
        ids = sorted({int(x) for x in (substitute_ids or []) if self._parse_int(x)})
        out = {i: [] for i in ids}
        if not ids:
            return out
        ph = ','.join(['%s'] * len(ids))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT owner.id AS owner_id,
                       owner.sku AS owner_sku,
                       sub.id AS sub_id
                FROM order_products owner
                INNER JOIN order_product_shipping_plans ops ON ops.order_product_id = owner.id
                INNER JOIN order_product_shipping_plan_items i ON i.shipping_plan_id = ops.id
                INNER JOIN order_products sub ON sub.id = i.substitute_order_product_id
                WHERE COALESCE(owner.is_on_market, 0) = 0
                  AND sub.id IN ({ph})
                ORDER BY owner.id ASC, sub.id ASC
                """,
                tuple(ids),
            )
            rows = cur.fetchall() or []
        owner_ids = sorted({
            self._parse_int(rr.get('owner_id'))
            for rr in rows
            if self._parse_int(rr.get('owner_id'))
        })
        plan_by_owner = self._forecast_load_default_substitute_plan_items_by_owner(conn, owner_ids)
        all_ids = set(owner_ids)
        for _oid, items in plan_by_owner.items():
            for sid, _m in items:
                if sid:
                    all_ids.add(int(sid))
        brief = self._forecast_load_order_product_brief_map(conn, sorted(all_ids))
        for rr in rows:
            sub_id = self._parse_int(rr.get('sub_id'))
            owner_id = self._parse_int(rr.get('owner_id'))
            if not sub_id or not owner_id:
                continue
            items = plan_by_owner.get(owner_id) or []
            on_market_sub_skus = []
            for sid, _m in items:
                if (self._parse_int((brief.get(sid) or {}).get('is_on_market')) or 0) == 1:
                    on_market_sub_skus.append((brief.get(sid) or {}).get('sku') or f'#{sid}')
            qty_in_plan = 1
            for sid, m in items:
                if int(sid) == int(sub_id):
                    qty_in_plan = max(1, int(m))
                    break
            out.setdefault(sub_id, []).append({
                'owner_id': owner_id,
                'owner_sku': rr.get('owner_sku') or (brief.get(owner_id) or {}).get('sku') or '',
                'qty_in_plan': qty_in_plan,
                'on_market_sub_count': len(on_market_sub_skus),
                'on_market_sub_skus': on_market_sub_skus,
            })
        return out

    def _forecast_attach_remarks_to_rows(self, conn, rows, forecast_mode):
        if not rows:
            return
        for row in rows:
            row['remarks'] = []

        if forecast_mode == 'platform':
            status_labels = {'retained': '留用', 'discarded': '弃用'}
            for row in rows:
                ps = str((row.get('labels') or {}).get('product_status') or 'enabled').strip()
                if ps in status_labels:
                    row['remarks'].append(f"平台SKU状态：{status_labels[ps]}")
            return

        if forecast_mode == 'spec':
            variant_ids = sorted({
                self._parse_int((r.get('labels') or {}).get('variant_id'))
                for r in rows
                if self._parse_int((r.get('labels') or {}).get('variant_id'))
            })
            platforms_by_vid = self._forecast_load_variant_platform_skus(conn, variant_ids)
            for row in rows:
                vid = self._parse_int((row.get('labels') or {}).get('variant_id'))
                skus = platforms_by_vid.get(vid) or []
                if not skus:
                    row['remarks'].append('未关联任何平台SKU')
                    continue
                enabled = [
                    s for s in skus
                    if str(s.get('product_status') or 'enabled').strip() == 'enabled'
                ]
                if not enabled:
                    statuses = {str(s.get('product_status') or 'enabled').strip() for s in skus}
                    if statuses == {'discarded'}:
                        row['remarks'].append('关联平台SKU均已弃用')
                    else:
                        row['remarks'].append('关联平台SKU均无启用状态')
            return

        if forecast_mode != 'order':
            return

        op_ids = sorted({
            self._parse_int((r.get('labels') or {}).get('order_product_id'))
            for r in rows
            if self._parse_int((r.get('labels') or {}).get('order_product_id'))
        })
        if not op_ids:
            return
        plan_by_owner = self._forecast_load_default_substitute_plan_items_by_owner(conn, op_ids)
        sub_refs = self._forecast_load_delisted_owner_refs_by_substitute_ids(conn, op_ids)
        brief_ids = set(op_ids)
        for oid, items in plan_by_owner.items():
            brief_ids.add(int(oid))
            for sid, _m in items:
                if sid:
                    brief_ids.add(int(sid))
        brief = self._forecast_load_order_product_brief_map(conn, sorted(brief_ids))

        for row in rows:
            labels = row.get('labels') or {}
            oid = self._parse_int(labels.get('order_product_id'))
            if not oid:
                continue
            is_on = self._parse_int(labels.get('is_on_market')) or 0
            links = labels.get('variant_links') or []
            plan_items = plan_by_owner.get(oid) or []

            if not links:
                row['remarks'].append('未关联销售规格')

            if is_on == 0:
                row['remarks'].append('已下市且无可用在市替代发货方案')
                if not plan_items:
                    row['remarks'].append('无替代发货方案')
                else:
                    on_market_subs = [
                        sid for sid, _m in plan_items
                        if (self._parse_int((brief.get(sid) or {}).get('is_on_market')) or 0) == 1
                    ]
                    if not on_market_subs:
                        row['remarks'].append('替代方案SKU均在市外')
                continue

            on_market_subs_in_plan = [
                sid for sid, _m in plan_items
                if (self._parse_int((brief.get(sid) or {}).get('is_on_market')) or 0) == 1
            ]
            if plan_items and not on_market_subs_in_plan:
                row['remarks'].append('替代发货方案均在市外，按本体SKU计库存')

            for ref in sub_refs.get(oid) or []:
                if int(ref.get('on_market_sub_count') or 0) > 1:
                    owner_sku = ref.get('owner_sku') or ''
                    sub_skus = ref.get('on_market_sub_skus') or []
                    joined = '、'.join([s for s in sub_skus if s])
                    row['remarks'].append(
                        f"异常：下市下单SKU「{owner_sku}」的替代方案含多个在市SKU（{joined}）"
                    )

    def _forecast_load_platform_sku_dim(self, conn, query_params=None, sf_group=None):
        """加载平台SKU维度行：sales_products + variant + family + fabric + 在售标识。"""
        sku_keyword = ''
        spec_keyword = ''
        fabric_keyword = ''
        if query_params:
            sku_keyword = (query_params.get('sku', [''])[0] or '').strip()
            spec_keyword = (query_params.get('spec', [''])[0] or '').strip()
            fabric_keyword = (query_params.get('fabric', [''])[0] or '').strip()

        clauses = ["1=1"]
        params = []
        if sku_keyword:
            clauses.append("(sp.platform_sku LIKE %s OR pf.sku_family LIKE %s)")
            like_val = f"%{sku_keyword}%"
            params.extend([like_val, like_val])
        if spec_keyword:
            clauses.append("v.spec_name LIKE %s")
            params.append(f"%{spec_keyword}%")
        if fabric_keyword:
            clauses.append("(fm.fabric_code LIKE %s OR fm.fabric_name_en LIKE %s)")
            like_val = f"%{fabric_keyword}%"
            params.extend([like_val, like_val])
        self._forecast_apply_platform_status_filter(
            clauses, params, self._forecast_parse_sf_status_filter(query_params)
        )
        self._forecast_apply_sf_group_platform(sf_group, clauses, params)

        where_sql = ' AND '.join(clauses)
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT sp.id AS sales_product_id,
                       sp.platform_sku,
                       sp.product_status,
                       sp.shop_id,
                       sh.shop_name,
                       sh.platform_type_id,
                       v.id AS variant_id,
                       v.sku_family_id,
                       v.spec_name,
                       v.fabric_id,
                       pf.sku_family,
                       fm.fabric_code,
                       fm.fabric_name_en,
                       fm.representative_color
                FROM sales_products sp
                JOIN sales_product_variants v ON v.id = sp.variant_id
                LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id
                LEFT JOIN shops sh ON sh.id = sp.shop_id
                WHERE {where_sql}
                ORDER BY pf.sku_family ASC, v.spec_name ASC, sp.platform_sku ASC, sp.id ASC
                """,
                tuple(params)
            )
            return [dict(row) for row in (cur.fetchall() or [])]

    def _forecast_list_platform_groups(self, conn, query_params=None):
        """轻量：仅返回店铺/货号分组及行数（与前端 buildGroups 的 key 一致）。"""
        sku_keyword = ''
        spec_keyword = ''
        fabric_keyword = ''
        if query_params:
            sku_keyword = (query_params.get('sku', [''])[0] or '').strip()
            spec_keyword = (query_params.get('spec', [''])[0] or '').strip()
            fabric_keyword = (query_params.get('fabric', [''])[0] or '').strip()
        clauses = ["1=1"]
        params = []
        if sku_keyword:
            clauses.append("(sp.platform_sku LIKE %s OR pf.sku_family LIKE %s)")
            like_val = f"%{sku_keyword}%"
            params.extend([like_val, like_val])
        if spec_keyword:
            clauses.append("v.spec_name LIKE %s")
            params.append(f"%{spec_keyword}%")
        if fabric_keyword:
            clauses.append("(fm.fabric_code LIKE %s OR fm.fabric_name_en LIKE %s)")
            like_val = f"%{fabric_keyword}%"
            params.extend([like_val, like_val])
        self._forecast_apply_platform_status_filter(
            clauses, params, self._forecast_parse_sf_status_filter(query_params)
        )
        where_sql = ' AND '.join(clauses)
        out = []
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT COALESCE(NULLIF(TRIM(sh.shop_name),''), '-') AS shop_g,
                       COALESCE(NULLIF(TRIM(pf.sku_family),''), '-') AS fam_g,
                       COUNT(*) AS cnt
                FROM sales_products sp
                JOIN sales_product_variants v ON v.id = sp.variant_id
                LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id
                LEFT JOIN shops sh ON sh.id = sp.shop_id
                WHERE {where_sql}
                GROUP BY COALESCE(NULLIF(TRIM(sh.shop_name),''), '-'),
                         COALESCE(NULLIF(TRIM(pf.sku_family),''), '-')
                ORDER BY shop_g ASC, fam_g ASC
                """,
                tuple(params)
            )
            for row in (cur.fetchall() or []):
                shop_g = str(row.get('shop_g') or '-').strip() or '-'
                fam_g = str(row.get('fam_g') or '-').strip() or '-'
                key = f"{shop_g}||{fam_g}"
                title = f"{shop_g} / {fam_g}"
                cnt = int(self._parse_int(row.get('cnt')) or 0)
                out.append({'key': key, 'title': title, 'item_count': cnt})
        return out

    def _forecast_load_variant_dim(self, conn, query_params=None, sf_group=None):
        """加载规格+面料 维度行（variant 行）。"""
        sku_keyword = ''
        spec_keyword = ''
        fabric_keyword = ''
        if query_params:
            sku_keyword = (query_params.get('sku', [''])[0] or '').strip()
            spec_keyword = (query_params.get('spec', [''])[0] or '').strip()
            fabric_keyword = (query_params.get('fabric', [''])[0] or '').strip()

        clauses = ["1=1"]
        params = []
        self._forecast_apply_spec_sku_keyword_filter(sku_keyword, clauses, params)
        if spec_keyword:
            clauses.append("v.spec_name LIKE %s")
            params.append(f"%{spec_keyword}%")
        if fabric_keyword:
            clauses.append("(fm.fabric_code LIKE %s OR fm.fabric_name_en LIKE %s)")
            like_val = f"%{fabric_keyword}%"
            params.extend([like_val, like_val])
        if sf_group:
            gk = str(sf_group).strip()
            if gk:
                clauses.append("COALESCE(NULLIF(TRIM(pf.sku_family),''), '-') = %s")
                params.append(gk)
        self._forecast_apply_variant_linked_status_filter(
            clauses, params, self._forecast_parse_sf_status_filter(query_params)
        )

        where_sql = ' AND '.join(clauses)
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT v.id AS variant_id,
                       v.sku_family_id,
                       v.spec_name,
                       v.fabric_id,
                       pf.sku_family,
                       fm.fabric_code,
                       fm.fabric_name_en,
                       fm.representative_color
                FROM sales_product_variants v
                LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id
                WHERE {where_sql}
                ORDER BY pf.sku_family ASC, v.spec_name ASC, v.id ASC
                """,
                tuple(params)
            )
            return [dict(row) for row in (cur.fetchall() or [])]

    def _forecast_list_spec_groups(self, conn, query_params=None):
        """轻量：货号分组及变体行数。"""
        sku_keyword = ''
        spec_keyword = ''
        fabric_keyword = ''
        if query_params:
            sku_keyword = (query_params.get('sku', [''])[0] or '').strip()
            spec_keyword = (query_params.get('spec', [''])[0] or '').strip()
            fabric_keyword = (query_params.get('fabric', [''])[0] or '').strip()
        clauses = ["1=1"]
        params = []
        self._forecast_apply_spec_sku_keyword_filter(sku_keyword, clauses, params)
        if spec_keyword:
            clauses.append("v.spec_name LIKE %s")
            params.append(f"%{spec_keyword}%")
        if fabric_keyword:
            clauses.append("(fm.fabric_code LIKE %s OR fm.fabric_name_en LIKE %s)")
            like_val = f"%{fabric_keyword}%"
            params.extend([like_val, like_val])
        self._forecast_apply_variant_linked_status_filter(
            clauses, params, self._forecast_parse_sf_status_filter(query_params)
        )
        where_sql = ' AND '.join(clauses)
        out = []
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT COALESCE(NULLIF(TRIM(pf.sku_family),''), '-') AS gkey,
                       COUNT(*) AS cnt
                FROM sales_product_variants v
                LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id
                WHERE {where_sql}
                GROUP BY COALESCE(NULLIF(TRIM(pf.sku_family),''), '-')
                ORDER BY gkey ASC
                """,
                tuple(params)
            )
            for row in (cur.fetchall() or []):
                gkey = str(row.get('gkey') or '-').strip() or '-'
                cnt = int(self._parse_int(row.get('cnt')) or 0)
                out.append({'key': gkey, 'title': gkey, 'item_count': cnt})
        return out

    def _forecast_load_variant_platform_skus(self, conn, variant_ids):
        """variant_id -> [{sales_product_id, platform_sku, shop_name}]"""
        out = {}
        if not variant_ids:
            return out
        placeholders = ','.join(['%s'] * len(variant_ids))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT sp.id AS sales_product_id,
                       sp.platform_sku,
                       sp.product_status,
                       sp.variant_id,
                       sh.shop_name
                FROM sales_products sp
                LEFT JOIN shops sh ON sh.id = sp.shop_id
                WHERE sp.variant_id IN ({placeholders})
                ORDER BY sp.platform_sku ASC, sp.id ASC
                """,
                tuple(variant_ids)
            )
            for row in (cur.fetchall() or []):
                vid = self._parse_int(row.get('variant_id'))
                if not vid:
                    continue
                out.setdefault(vid, []).append({
                    'sales_product_id': self._parse_int(row.get('sales_product_id')),
                    'platform_sku': row.get('platform_sku') or '',
                    'shop_name': row.get('shop_name') or '',
                    'product_status': row.get('product_status') or 'enabled',
                })
        return out

    def _forecast_load_order_dim(self, conn, query_params=None, sf_group=None):
        """加载下单SKU 维度行 + 关联 variant/规格信息。"""
        sku_keyword = ''
        spec_keyword = ''
        fabric_keyword = ''
        if query_params:
            sku_keyword = (query_params.get('sku', [''])[0] or '').strip()
            spec_keyword = (query_params.get('spec', [''])[0] or '').strip()
            fabric_keyword = (query_params.get('fabric', [''])[0] or '').strip()

        clauses = ["1=1"]
        params = []
        if sku_keyword:
            clauses.append("(op.sku LIKE %s OR pf.sku_family LIKE %s)")
            like_val = f"%{sku_keyword}%"
            params.extend([like_val, like_val])
        if spec_keyword:
            clauses.append("op.spec_qty_short LIKE %s")
            params.append(f"%{spec_keyword}%")
        if fabric_keyword:
            clauses.append("(fm.fabric_code LIKE %s OR fm.fabric_name_en LIKE %s)")
            like_val = f"%{fabric_keyword}%"
            params.extend([like_val, like_val])
        if sf_group:
            gk = str(sf_group).strip()
            if gk:
                clauses.append("COALESCE(NULLIF(TRIM(pf.sku_family),''), '-') = %s")
                params.append(gk)

        clauses.append(self._forecast_order_dim_visibility_sql())
        where_sql = ' AND '.join(clauses)
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT op.id AS order_product_id,
                       op.sku,
                       op.sku_family_id,
                       op.spec_qty_short,
                       op.contents_desc_en,
                       op.is_on_market,
                       op.is_iteration,
                       op.fabric_id,
                       pf.sku_family,
                       fm.fabric_code,
                       fm.fabric_name_en,
                       fm.representative_color
                FROM order_products op
                LEFT JOIN product_families pf ON pf.id = op.sku_family_id
                LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                WHERE {where_sql}
                  AND COALESCE(op.is_reship_accessory, 0) = 0
                ORDER BY pf.sku_family ASC, op.sku ASC, op.id ASC
                """,
                tuple(params)
            )
            rows = [dict(row) for row in (cur.fetchall() or [])]

            order_ids = [self._parse_int(r.get('order_product_id')) for r in rows if self._parse_int(r.get('order_product_id'))]
            links_by_op = {}
            if order_ids:
                placeholders = ','.join(['%s'] * len(order_ids))
                cur.execute(
                    f"""
                    SELECT l.order_product_id,
                           l.variant_id,
                           l.quantity,
                           v.spec_name,
                           pf.sku_family
                    FROM sales_variant_order_links l
                    LEFT JOIN sales_product_variants v ON v.id = l.variant_id
                    LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                    WHERE l.order_product_id IN ({placeholders})
                    """,
                    tuple(order_ids)
                )
                for link in (cur.fetchall() or []):
                    op_id = self._parse_int(link.get('order_product_id'))
                    vid = self._parse_int(link.get('variant_id'))
                    if not op_id or not vid:
                        continue
                    links_by_op.setdefault(op_id, []).append({
                        'variant_id': vid,
                        'quantity': max(1, self._parse_int(link.get('quantity')) or 1),
                        'spec_name': link.get('spec_name') or '',
                        'sku_family': link.get('sku_family') or '',
                    })
        for r in rows:
            op_id = self._parse_int(r.get('order_product_id'))
            r['variant_links'] = links_by_op.get(op_id, [])
        return rows

    def _forecast_list_order_groups(self, conn, query_params=None):
        """轻量：下单 SKU 按货号分组及行数。"""
        sku_keyword = ''
        spec_keyword = ''
        fabric_keyword = ''
        if query_params:
            sku_keyword = (query_params.get('sku', [''])[0] or '').strip()
            spec_keyword = (query_params.get('spec', [''])[0] or '').strip()
            fabric_keyword = (query_params.get('fabric', [''])[0] or '').strip()
        clauses = ["1=1", "COALESCE(op.is_reship_accessory, 0) = 0", self._forecast_order_dim_visibility_sql()]
        params = []
        if sku_keyword:
            clauses.append("(op.sku LIKE %s OR pf.sku_family LIKE %s)")
            like_val = f"%{sku_keyword}%"
            params.extend([like_val, like_val])
        if spec_keyword:
            clauses.append("op.spec_qty_short LIKE %s")
            params.append(f"%{spec_keyword}%")
        if fabric_keyword:
            clauses.append("(fm.fabric_code LIKE %s OR fm.fabric_name_en LIKE %s)")
            like_val = f"%{fabric_keyword}%"
            params.extend([like_val, like_val])
        where_sql = ' AND '.join(clauses)
        out = []
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT COALESCE(NULLIF(TRIM(pf.sku_family),''), '-') AS gkey,
                       COUNT(*) AS cnt
                FROM order_products op
                LEFT JOIN product_families pf ON pf.id = op.sku_family_id
                LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                WHERE {where_sql}
                GROUP BY COALESCE(NULLIF(TRIM(pf.sku_family),''), '-')
                ORDER BY gkey ASC
                """,
                tuple(params)
            )
            for row in (cur.fetchall() or []):
                gkey = str(row.get('gkey') or '-').strip() or '-'
                cnt = int(self._parse_int(row.get('cnt')) or 0)
                out.append({'key': gkey, 'title': gkey, 'item_count': cnt})
        return out

    def _forecast_load_variant_thumb_b64(self, conn, variant_ids):
        """variant_id -> 预览图 b64；仅使用图片类型为「白底纯图」的映射（无则不回退）。"""
        out = {}
        ids = [self._parse_int(x) for x in (variant_ids or []) if self._parse_int(x)]
        if not ids:
            return out
        placeholders = ','.join(['%s'] * len(ids))
        has_sim_tid = self._table_has_column(conn, 'sales_variant_image_mappings', 'image_type_id')
        has_ia_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
        if has_sim_tid and has_ia_tid:
            type_expr = "COALESCE(NULLIF(TRIM(it_sim.name),''), NULLIF(TRIM(it_ia.name),''), '') AS image_type_name"
            join_types = """
                LEFT JOIN image_types it_ia ON it_ia.id = ia.image_type_id
                LEFT JOIN image_types it_sim ON it_sim.id = sim.image_type_id
            """
        elif has_sim_tid:
            type_expr = "COALESCE(NULLIF(TRIM(it_sim.name),''), '') AS image_type_name"
            join_types = "LEFT JOIN image_types it_sim ON it_sim.id = sim.image_type_id"
        elif has_ia_tid:
            type_expr = "COALESCE(NULLIF(TRIM(it_ia.name),''), '') AS image_type_name"
            join_types = "LEFT JOIN image_types it_ia ON it_ia.id = ia.image_type_id"
        else:
            type_expr = "'' AS image_type_name"
            join_types = ""
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT sim.variant_id,
                       ia.storage_path,
                       sim.sort_order,
                       sim.id AS sim_id,
                       {type_expr}
                FROM sales_variant_image_mappings sim
                JOIN image_assets ia ON ia.id = sim.image_asset_id
                {join_types}
                WHERE sim.variant_id IN ({placeholders})
                ORDER BY sim.variant_id ASC, sim.sort_order ASC, sim.id ASC
                """,
                tuple(ids)
            )
            rows = cur.fetchall() or []

        best = {}
        for row in rows:
            vid = self._parse_int(row.get('variant_id'))
            if not vid:
                continue
            storage_path = str(row.get('storage_path') or '').strip()
            if not storage_path:
                continue
            tname = str(row.get('image_type_name') or '').strip()
            if tname != '白底纯图':
                continue
            sort_order = self._parse_int(row.get('sort_order')) or 0
            sim_id = self._parse_int(row.get('sim_id')) or 0
            key = (sort_order, sim_id)
            if vid not in best or key < best[vid][0]:
                b64 = self._b64_from_fs(storage_path.replace('\\', '/').lstrip('/'))
                best[vid] = (key, b64)

        for vid, pair in best.items():
            out[vid] = pair[1]
        return out

    def _forecast_load_order_product_thumb_b64(self, conn, order_product_ids):
        """order_product_id -> 预览图 b64；仅「白底纯图」类型（order_product_image_mappings）。"""
        out = {}
        if not self._has_required_tables(['order_product_image_mappings', 'image_assets']):
            return out
        ids = [self._parse_int(x) for x in (order_product_ids or []) if self._parse_int(x)]
        if not ids:
            return out
        placeholders = ','.join(['%s'] * len(ids))
        has_ia_tid = self._table_has_column(conn, 'image_assets', 'image_type_id')
        if has_ia_tid:
            type_expr = "COALESCE(NULLIF(TRIM(it_ia.name),''), '') AS image_type_name"
            join_types = "LEFT JOIN image_types it_ia ON it_ia.id = ia.image_type_id"
        else:
            type_expr = "'' AS image_type_name"
            join_types = ""
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT opim.order_product_id,
                       ia.storage_path,
                       opim.sort_order,
                       opim.id AS opim_id,
                       {type_expr}
                FROM order_product_image_mappings opim
                JOIN image_assets ia ON ia.id = opim.image_asset_id
                {join_types}
                WHERE opim.order_product_id IN ({placeholders})
                ORDER BY opim.order_product_id ASC, opim.sort_order ASC, opim.id ASC
                """,
                tuple(ids)
            )
            rows = cur.fetchall() or []

        best = {}
        for row in rows:
            opid = self._parse_int(row.get('order_product_id'))
            if not opid:
                continue
            storage_path = str(row.get('storage_path') or '').strip()
            if not storage_path:
                continue
            tname = str(row.get('image_type_name') or '').strip()
            if tname != '白底纯图':
                continue
            sort_order = self._parse_int(row.get('sort_order')) or 0
            rid = self._parse_int(row.get('opim_id')) or 0
            key = (sort_order, rid)
            if opid not in best or key < best[opid][0]:
                b64 = self._b64_from_fs(storage_path.replace('\\', '/').lstrip('/'))
                best[opid] = (key, b64)

        for opid, pair in best.items():
            out[opid] = pair[1]
        return out

    def _forecast_load_platform_cells(self, conn, sales_product_ids, months):
        if not sales_product_ids or not months:
            return {}
        placeholders_p = ','.join(['%s'] * len(sales_product_ids))
        placeholders_m = ','.join(['%s'] * len(months))
        params = list(sales_product_ids) + list(months)
        out = {}
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT id, sales_product_id, forecast_month,
                       initial_qty, prev_qty, latest_qty,
                       created_at, prev_updated_at, latest_updated_at
                FROM sales_forecast_platform_sku_monthly
                WHERE sales_product_id IN ({placeholders_p})
                  AND forecast_month IN ({placeholders_m})
                """,
                tuple(params)
            )
            for row in (cur.fetchall() or []):
                spid = self._parse_int(row.get('sales_product_id'))
                month_str = self._forecast_month_to_str(row.get('forecast_month'))
                if not spid or not month_str:
                    continue
                out[(spid, month_str)] = self._forecast_serialize_cell(row, month_str, extra_keys={'sales_product_id': spid})
        return out

    def _forecast_load_spec_cells(self, conn, variant_ids, months):
        if not variant_ids or not months:
            return {}
        placeholders_v = ','.join(['%s'] * len(variant_ids))
        placeholders_m = ','.join(['%s'] * len(months))
        params = list(variant_ids) + list(months)
        out = {}
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT id, variant_id, forecast_month,
                       initial_qty, prev_qty, latest_qty,
                       created_at, prev_updated_at, latest_updated_at
                FROM sales_forecast_spec_monthly
                WHERE variant_id IN ({placeholders_v})
                  AND forecast_month IN ({placeholders_m})
                """,
                tuple(params)
            )
            for row in (cur.fetchall() or []):
                vid = self._parse_int(row.get('variant_id'))
                month_str = self._forecast_month_to_str(row.get('forecast_month'))
                if not vid or not month_str:
                    continue
                out[(vid, month_str)] = self._forecast_serialize_cell(row, month_str, extra_keys={'variant_id': vid})
        return out

    def _forecast_load_order_cells(self, conn, order_product_ids, months):
        if not order_product_ids or not months:
            return {}
        placeholders_o = ','.join(['%s'] * len(order_product_ids))
        placeholders_m = ','.join(['%s'] * len(months))
        params = list(order_product_ids) + list(months)
        out = {}
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT id, order_product_id, forecast_month,
                       initial_qty, prev_qty, latest_qty,
                       created_at, prev_updated_at, latest_updated_at
                FROM sales_forecast_order_sku_monthly
                WHERE order_product_id IN ({placeholders_o})
                  AND forecast_month IN ({placeholders_m})
                """,
                tuple(params)
            )
            for row in (cur.fetchall() or []):
                op_id = self._parse_int(row.get('order_product_id'))
                month_str = self._forecast_month_to_str(row.get('forecast_month'))
                if not op_id or not month_str:
                    continue
                out[(op_id, month_str)] = self._forecast_serialize_cell(row, month_str, extra_keys={'order_product_id': op_id})
        return out

    def _forecast_serialize_cell(self, row, month_str, extra_keys=None):
        payload = {
            'id': self._parse_int(row.get('id')),
            'forecast_month': month_str,
            'initial_qty': self._parse_int(row.get('initial_qty')) or 0,
            'prev_qty': None if row.get('prev_qty') is None else self._parse_int(row.get('prev_qty')),
            'latest_qty': self._parse_int(row.get('latest_qty')) or 0,
            'created_at': self._forecast_format_dt(row.get('created_at')),
            'prev_updated_at': self._forecast_format_dt(row.get('prev_updated_at')),
            'latest_updated_at': self._forecast_format_dt(row.get('latest_updated_at')),
        }
        if extra_keys:
            payload.update(extra_keys)
        return payload

    def _forecast_perf_history_shop_joins_sql(self):
        return (
            'LEFT JOIN shops sh ON sh.id = sp.shop_id '
            'LEFT JOIN platform_types pt ON pt.id = sh.platform_type_id'
        )

    def _forecast_perf_history_cost_join_sql(self, conn):
        """与产品表现看板货号分组（周/月粒度）一致的变体级单位成本子查询，用于预测历史月单元格。"""
        has_op_reship = self._table_has_column(conn, 'order_products', 'is_reship_accessory')
        has_op_last_mile = self._table_has_column(conn, 'order_products', 'last_mile_avg_freight_usd')
        reship_clause = ' AND COALESCE(op.is_reship_accessory,0)=0 ' if has_op_reship else ''
        lm_weighted_sum = (
            'SUM(COALESCE(op.last_mile_avg_freight_usd, 0) * COALESCE(svol.quantity, 1))'
            if has_op_last_mile else '0'
        )
        return (
            'LEFT JOIN ('
            ' SELECT v.id AS variant_id,'
            '   SUM(COALESCE(op.cost_usd, 0) * COALESCE(svol.quantity, 1)) AS unit_bom_cost_usd,'
            f'   {lm_weighted_sum} AS unit_last_mile_freight_usd'
            ' FROM sales_product_variants v'
            ' INNER JOIN sales_variant_order_links svol ON svol.variant_id = v.id'
            ' INNER JOIN order_products op ON op.id = svol.order_product_id'
            f' WHERE 1=1{reship_clause}'
            ' GROUP BY v.id'
            ') est_unit_cost ON est_unit_cost.variant_id = sp.variant_id'
        )

    def _forecast_finalize_history_perf_payload(
        self, raw, *, platform_type_id=None, product_category=None, rules_cache=None, conn=None,
    ):
        """将 sales_perf_agg_month 一行规范为看板货号分组内层 SKU 行指标；佣金按平台+细分类目规则计算。"""
        net = float(raw.get('net_sales_amount') or 0)
        gross = float(raw.get('gross_sales_amount') or 0)
        ref = float(raw.get('refund_amount') or 0)
        ad_spend = float(raw.get('ad_spend') or 0)
        bom = round(float(raw.get('estimated_product_cost_usd') or 0), 2)
        lm = round(float(raw.get('estimated_last_mile_freight_usd') or 0), 2)
        total_cost = round(bom + lm, 2)
        discount_rate = round((gross - net) / gross, 6) if gross > 1e-12 else 0.0
        refund_rate = round(ref / net, 6) if net > 1e-12 else 0.0

        comm_result = None
        if conn is not None and platform_type_id and product_category:
            cache = rules_cache if rules_cache is not None else self._commission_load_rules_cache(conn)
            comm_result = self._commission_compute_for_context(
                cache, platform_type_id, product_category, net, mode='period',
            )
        elif conn is not None and rules_cache is not None:
            comm_result = {
                'commission_status': 'unavailable',
                'commission_message': self.COMMISSION_UNAVAILABLE_LABEL,
                'est_referral_commission_usd': None,
                'commission_rate': None,
            }

        if comm_result:
            comm = comm_result.get('est_referral_commission_usd')
            commission_rate = comm_result.get('commission_rate')
            if comm_result.get('commission_status') == 'ok' and comm is not None:
                profit = round(net - float(comm) - total_cost - ad_spend - ref, 2)
                net_margin_rate = round(profit / gross, 6) if gross > 1e-12 else 0.0
            else:
                profit = None
                net_margin_rate = None
        else:
            comm = round(min(net, 200.0) * 0.15 + max(net - 200.0, 0.0) * 0.10, 2)
            commission_rate = round(comm / net, 6) if net > 1e-12 else 0.0
            profit = round(net - comm - total_cost - ad_spend - ref, 2)
            net_margin_rate = round(profit / gross, 6) if gross > 1e-12 else 0.0
            comm_result = {
                'commission_status': 'legacy',
                'commission_message': None,
            }

        payload = {
            'rows': int(self._parse_int(raw.get('source_rows')) or 0),
            'sales_qty': float(raw.get('sales_qty') or 0),
            'net_sales_amount': round(net, 2),
            'gross_sales_amount': round(gross, 2),
            'discount_rate': discount_rate,
            'order_qty': float(raw.get('order_qty') or 0),
            'session_total': float(raw.get('session_total') or 0),
            'ad_impressions': float(raw.get('ad_impressions') or 0),
            'ad_clicks': float(raw.get('ad_clicks') or 0),
            'ad_orders': float(raw.get('ad_orders') or 0),
            'ad_spend': round(float(raw.get('ad_spend') or 0), 2),
            'ad_sales_amount': round(float(raw.get('ad_sales_amount') or 0), 2),
            'refund_amount': round(ref, 2),
            'refund_rate': refund_rate,
            'estimated_product_cost_usd': bom,
            'estimated_last_mile_freight_usd': lm,
            'estimated_total_cost_usd': total_cost,
            'est_referral_commission_usd': round(float(comm), 2) if comm is not None else None,
            'commission_rate': commission_rate,
            'estimated_net_profit_usd': profit,
            'net_margin_rate': net_margin_rate,
            'commission_status': (comm_result or {}).get('commission_status'),
            'commission_message': (comm_result or {}).get('commission_message'),
            'commission_group': (comm_result or {}).get('commission_group'),
        }
        return payload

    def _forecast_empty_history_perf_payload(self):
        return self._forecast_finalize_history_perf_payload({
            'source_rows': 0,
            'sales_qty': 0,
            'net_sales_amount': 0,
            'gross_sales_amount': 0,
            'order_qty': 0,
            'session_total': 0,
            'ad_impressions': 0,
            'ad_clicks': 0,
            'ad_orders': 0,
            'ad_spend': 0,
            'ad_sales_amount': 0,
            'refund_amount': 0,
            'estimated_product_cost_usd': 0,
            'estimated_last_mile_freight_usd': 0,
        })

    def _forecast_parse_sf_shop_ids(self, query_params):
        """解析 GET 参数 sf_shop_ids（逗号分隔店铺 id）；空或未传表示不按店铺过滤。"""
        if not query_params:
            return None
        raw = (query_params.get('sf_shop_ids', [''])[0] or '').strip()
        if not raw:
            return None
        out = []
        for part in raw.split(','):
            sid = self._parse_int(str(part).strip())
            if sid and sid > 0:
                out.append(int(sid))
        return sorted(set(out)) if out else None

    def _forecast_query_flag_off(self, query_params, key):
        if not query_params:
            return False
        raw = (query_params.get(key, [''])[0] or '').strip().lower()
        return raw in ('0', 'false', 'no', 'off')

    def _forecast_is_inventory_patch(self, query_params):
        if not query_params:
            return False
        raw = (query_params.get('patch', [''])[0] or '').strip().lower()
        return raw == 'inventory'

    def _forecast_list_shops_for_forecast(self, conn):
        """销量预测页店铺筛选下拉数据。"""
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT id, shop_name FROM shops ORDER BY shop_name ASC")
                rows = cur.fetchall() or []
            return [
                {'id': self._parse_int(r.get('id')), 'shop_name': str(r.get('shop_name') or '').strip()}
                for r in rows
                if self._parse_int(r.get('id'))
            ]
        except Exception:
            return []

    def _forecast_load_history_by_sales_product(self, conn, sales_product_ids, start_month, end_month, shop_ids=None):
        out = {}
        if not sales_product_ids or not start_month or not end_month:
            return out
        end_exclusive = self._forecast_history_end_exclusive(end_month)
        if not end_exclusive:
            return out
        placeholders = ','.join(['%s'] * len(sales_product_ids))
        shop_frag = ''
        params = list(sales_product_ids)
        if shop_ids:
            sph = ','.join(['%s'] * len(shop_ids))
            shop_frag = f' AND sp.shop_id IN ({sph}) '
            params.extend(int(x) for x in shop_ids)
        params.extend([start_month, end_exclusive])
        cost_join = self._forecast_perf_history_cost_join_sql(conn)
        shop_joins = self._forecast_perf_history_shop_joins_sql()
        lm_factor = self._shop_handles_last_mile_factor_sql(conn, 'sh', 'pt')
        rules_cache = self._commission_load_rules_cache(conn)
        sp_ctx = self._commission_load_sp_context_map(conn, sales_product_ids)
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT m.sales_product_id,
                       m.month_start,
                       m.source_rows,
                       m.sales_qty,
                       m.net_sales_amount,
                       m.order_qty,
                       m.session_total,
                       m.ad_impressions,
                       m.ad_clicks,
                       m.ad_orders,
                       m.ad_spend,
                       m.ad_sales_amount,
                       m.refund_amount,
                       (COALESCE(sp.sale_price_usd, 0) * COALESCE(m.sales_qty, 0)) AS gross_sales_amount,
                       (COALESCE(est_unit_cost.unit_bom_cost_usd, 0) * COALESCE(m.sales_qty, 0)) AS estimated_product_cost_usd,
                       (COALESCE(est_unit_cost.unit_last_mile_freight_usd, 0) * COALESCE(m.sales_qty, 0) * ({lm_factor})) AS estimated_last_mile_freight_usd
                FROM sales_perf_agg_month m
                JOIN sales_products sp ON sp.id = m.sales_product_id
                {shop_joins}
                {cost_join}
                WHERE m.sales_product_id IN ({placeholders})
                  AND m.month_start >= %s
                  AND m.month_start < %s
                  {shop_frag}
                """,
                tuple(params)
            )
            for row in (cur.fetchall() or []):
                spid = self._parse_int(row.get('sales_product_id'))
                month_str = self._forecast_month_to_str(row.get('month_start'))
                if not spid or not month_str:
                    continue
                ctx = sp_ctx.get(int(spid)) or {}
                out[(spid, month_str)] = self._forecast_finalize_history_perf_payload(
                    dict(row),
                    platform_type_id=ctx.get('platform_type_id'),
                    product_category=ctx.get('product_category'),
                    rules_cache=rules_cache,
                    conn=conn,
                )
        return out

    def _forecast_load_history_by_variant(self, conn, variant_ids, start_month, end_month, shop_ids=None):
        out = {}
        if not variant_ids or not start_month or not end_month:
            return out
        end_exclusive = self._forecast_history_end_exclusive(end_month)
        if not end_exclusive:
            return out
        placeholders = ','.join(['%s'] * len(variant_ids))
        shop_frag = ''
        params = list(variant_ids)
        if shop_ids:
            sph = ','.join(['%s'] * len(shop_ids))
            shop_frag = f' AND sp.shop_id IN ({sph}) '
            params.extend(int(x) for x in shop_ids)
        params.extend([start_month, end_exclusive])
        cost_join = self._forecast_perf_history_cost_join_sql(conn)
        shop_joins = self._forecast_perf_history_shop_joins_sql()
        lm_factor = self._shop_handles_last_mile_factor_sql(conn, 'sh', 'pt')
        rules_cache = (
            self._commission_load_rules_cache(conn)
            if self._commission_rules_tables_ready(conn) else None
        )
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT sp.variant_id AS variant_id,
                       m.month_start AS month_start,
                       SUM(COALESCE(m.source_rows, 0)) AS source_rows,
                       SUM(COALESCE(m.sales_qty, 0)) AS sales_qty,
                       SUM(COALESCE(m.net_sales_amount, 0)) AS net_sales_amount,
                       SUM(COALESCE(m.order_qty, 0)) AS order_qty,
                       SUM(COALESCE(m.session_total, 0)) AS session_total,
                       SUM(COALESCE(m.ad_impressions, 0)) AS ad_impressions,
                       SUM(COALESCE(m.ad_clicks, 0)) AS ad_clicks,
                       SUM(COALESCE(m.ad_orders, 0)) AS ad_orders,
                       SUM(COALESCE(m.ad_spend, 0)) AS ad_spend,
                       SUM(COALESCE(m.ad_sales_amount, 0)) AS ad_sales_amount,
                       SUM(COALESCE(m.refund_amount, 0)) AS refund_amount,
                       SUM(COALESCE(sp.sale_price_usd, 0) * COALESCE(m.sales_qty, 0)) AS gross_sales_amount,
                       (MAX(COALESCE(est_unit_cost.unit_bom_cost_usd, 0)) * SUM(COALESCE(m.sales_qty, 0))) AS estimated_product_cost_usd,
                       (MAX(COALESCE(est_unit_cost.unit_last_mile_freight_usd, 0) * ({lm_factor})) * SUM(COALESCE(m.sales_qty, 0))) AS estimated_last_mile_freight_usd
                FROM sales_perf_agg_month m
                JOIN sales_products sp ON sp.id = m.sales_product_id
                {shop_joins}
                {cost_join}
                WHERE sp.variant_id IN ({placeholders})
                  AND m.month_start >= %s
                  AND m.month_start < %s
                  {shop_frag}
                GROUP BY sp.variant_id, m.month_start
                """,
                tuple(params)
            )
            for row in (cur.fetchall() or []):
                vid = self._parse_int(row.get('variant_id'))
                month_str = self._forecast_month_to_str(row.get('month_start'))
                if not vid or not month_str:
                    continue
                out[(vid, month_str)] = self._forecast_finalize_history_perf_payload(
                    dict(row),
                    rules_cache=rules_cache if self._commission_rules_tables_ready(conn) else None,
                    conn=conn,
                )
        return out

    def _forecast_current_month_key(self):
        today = datetime.now()
        return f'{today.year:04d}-{today.month:02d}-01'

    def _forecast_perf_max_record_date(self, conn):
        """产品表现日表的全局最新业务日期（按日，精确到日历日）。"""
        try:
            with conn.cursor() as cur:
                cur.execute('SELECT MAX(record_date) AS mx FROM sales_product_performances')
                return (cur.fetchone() or {}).get('mx')
        except Exception:
            return None

    def _forecast_format_date_only(self, value):
        if value is None:
            return None
        if hasattr(value, 'strftime'):
            try:
                return value.strftime('%Y-%m-%d')
            except Exception:
                return str(value)[:10]
        s = str(value or '')
        return s[:10] if len(s) >= 10 else (s or None)

    def _forecast_month_add_years(self, month_key, delta_years):
        try:
            d = datetime.strptime(str(month_key)[:10], '%Y-%m-%d')
        except Exception:
            return None
        y = d.year + int(delta_years)
        return f'{y:04d}-{d.month:02d}-01'

    def _forecast_mtd_time_context(self, record_date_max, cur_month_key, today=None):
        """本月时间进度 T = ref_day / days_in_month。
        ref_day 来自 sales_product_performances 的全局 MAX(record_date)：
        若该日期落在本月，则取「该日号」表示该日整天已计入进度；若早于本月则为 0；
        若晚于本月末则视为已满。无数据时退回「今天在当月内的日号」。"""
        today = today or datetime.now()
        try:
            d0 = datetime.strptime(str(cur_month_key)[:10], '%Y-%m-%d')
        except Exception:
            d0 = today
        year, mon = d0.year, d0.month
        dim = calendar.monthrange(year, mon)[1]
        month_start = date(year, mon, 1)
        month_end = date(year, mon, dim)

        d_max = None
        if record_date_max is not None:
            try:
                if isinstance(record_date_max, datetime):
                    d_max = record_date_max.date()
                elif isinstance(record_date_max, date):
                    d_max = record_date_max
                else:
                    d_max = datetime.strptime(str(record_date_max)[:10], '%Y-%m-%d').date()
            except Exception:
                d_max = None

        ref_day = 0
        if d_max is not None:
            if d_max < month_start:
                ref_day = 0
            elif d_max > month_end:
                ref_day = dim
            elif d_max.year == year and d_max.month == mon:
                ref_day = int(d_max.day)
            else:
                ref_day = dim
        else:
            if today.year == year and today.month == mon:
                ref_day = min(int(today.day), dim)
            else:
                ref_day = dim

        ref_day = max(0, min(int(ref_day), dim))
        tp = (ref_day / float(dim)) if dim else 1.0
        if tp <= 0 and dim:
            tp = 1.0 / float(dim)
        tp = min(1.0, max(tp, (1.0 / float(dim)) if dim else 0.02))
        return {
            'current_month_key': f'{year:04d}-{mon:02d}-01',
            'days_in_month': dim,
            'ref_day': ref_day,
            'time_progress': tp,
        }

    def _forecast_row_mtd_ratio_band(self, ratio, tp):
        band = 'neutral'
        completion_pct = None
        bar_pct = 0.0
        if ratio is None:
            return band, completion_pct, bar_pct
        completion_pct = round(ratio * 100.0, 1)
        bar_pct = min(100.0, max(0.0, ratio * 100.0))
        r, t = ratio, float(tp or 0.05)
        if r < t - 0.10:
            band = 'red'
        elif r > t + 0.30:
            band = 'purple'
        elif r > t:
            band = 'green'
        elif r >= t - 0.10:
            band = 'yellow'
        else:
            band = 'red'
        return band, completion_pct, bar_pct

    def _forecast_row_mtd_completion(self, history, forecasts, mtd_ctx):
        """本月完成率：本月历史销量（左）÷ 本月预测整月量（右，与表格预测列同源）。"""
        if not mtd_ctx:
            return {
                'applicable': False,
                'band': 'neutral',
            }
        cur_key = mtd_ctx.get('current_month_key') or ''
        tp = float(mtd_ctx.get('time_progress') or 0.05)
        cell = (history or {}).get(cur_key) or {}
        mtd_sales = float(cell.get('sales_qty') or 0)
        fc = (forecasts or {}).get(cur_key) or {}
        denom = float(fc.get('value_qty') or 0)
        ratio = (mtd_sales / denom) if denom > 1e-9 else None
        band, completion_pct, bar_pct = self._forecast_row_mtd_ratio_band(ratio, tp)
        expected_curve = round(denom * tp, 4) if denom > 1e-9 else None
        return {
            'applicable': True,
            'mtd_sales_qty': mtd_sales,
            'forecast_month_qty': denom,
            'expected_mtd_qty': expected_curve,
            'time_progress': round(tp, 4),
            'completion_ratio': None if ratio is None else round(ratio, 4),
            'completion_pct': completion_pct,
            'bar_pct': bar_pct,
            'band': band,
            'ref_day': mtd_ctx.get('ref_day'),
            'days_in_month': mtd_ctx.get('days_in_month'),
        }

    def _forecast_inventory_zero(self):
        return {'overseas_qty': 0, 'transit_qty': 0, 'factory_stock_qty': 0, 'wip_qty': 0}

    def _forecast_load_inventory_by_order_product(self, conn, order_product_ids):
        """按 order_product_id 汇总：海外仓、在途、在库、在制（未完工）。
        在途：仅统计「已登记上架=否」的在途单（logistics_in_transit.inventory_registered=0）下的发货数量；
        未登记前明细里 listed_qty 常与 shipped_qty 相同，故不再用 shipped-listed 以免恒为 0。"""
        ids = sorted({int(x) for x in (order_product_ids or []) if self._parse_int(x)})
        out = {i: self._forecast_inventory_zero() for i in ids}
        if not ids:
            return out
        ph = ','.join(['%s'] * len(ids))
        tpl = tuple(ids)
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT oi.order_product_id, COALESCE(SUM(oi.available_qty), 0) AS q
                FROM logistics_overseas_inventory oi
                INNER JOIN logistics_overseas_warehouses w ON w.id = oi.warehouse_id
                WHERE oi.order_product_id IN ({ph})
                  AND COALESCE(w.is_enabled, 1) = 1
                GROUP BY oi.order_product_id
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                if oid in out:
                    out[oid]['overseas_qty'] = int(float(rr.get('q') or 0))

            cur.execute(
                f"""
                SELECT li.order_product_id, COALESCE(SUM(li.shipped_qty), 0) AS q
                FROM logistics_in_transit_items li
                INNER JOIN logistics_in_transit t ON t.id = li.transit_id
                WHERE li.order_product_id IN ({ph})
                  AND COALESCE(t.inventory_registered, 0) = 0
                GROUP BY li.order_product_id
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                if oid in out:
                    out[oid]['transit_qty'] = int(float(rr.get('q') or 0))

            cur.execute(
                f"""
                SELECT order_product_id, COALESCE(SUM(quantity), 0) AS q
                FROM factory_stock_inventory
                WHERE order_product_id IN ({ph})
                GROUP BY order_product_id
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                if oid in out:
                    out[oid]['factory_stock_qty'] = int(float(rr.get('q') or 0))

            cur.execute(
                f"""
                SELECT order_product_id, COALESCE(SUM(quantity), 0) AS q
                FROM factory_wip_inventory
                WHERE order_product_id IN ({ph}) AND COALESCE(is_completed, 0) = 0
                GROUP BY order_product_id
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                if oid in out:
                    out[oid]['wip_qty'] = int(float(rr.get('q') or 0))
        return out

    def _forecast_load_variant_order_links_for_inventory(self, conn, variant_ids):
        """variant_id -> [(order_product_id, qty_per_unit), ...]；同一变体同一下单 SKU 合并件数。"""
        ids = sorted({int(x) for x in (variant_ids or []) if self._parse_int(x)})
        out = {}
        if not ids:
            return out
        ph = ','.join(['%s'] * len(ids))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT variant_id, order_product_id, GREATEST(1, COALESCE(quantity, 1)) AS quantity
                FROM sales_variant_order_links
                WHERE variant_id IN ({ph})
                ORDER BY variant_id ASC, order_product_id ASC
                """,
                tuple(ids),
            )
            for rr in cur.fetchall() or []:
                vid = self._parse_int(rr.get('variant_id'))
                oid = self._parse_int(rr.get('order_product_id'))
                qp = max(1, self._parse_int(rr.get('quantity')) or 1)
                if not vid or not oid:
                    continue
                bucket = out.setdefault(vid, {})
                bucket[oid] = int(bucket.get(oid) or 0) + qp
        return {vid: sorted(links.items(), key=lambda x: x[0]) for vid, links in out.items()}

    def _load_variant_tier_sellable_map(self, conn, variant_ids):
        """variant_id -> {overseas, transit, factory_stock, wip} 成套瓶颈（BOM floor 后取 min）。"""
        ids = sorted({int(x) for x in (variant_ids or []) if self._parse_int(x)})
        empty = {'overseas': 0, 'transit': 0, 'factory_stock': 0, 'wip': 0}
        out = {i: dict(empty) for i in ids}
        if not ids:
            return out
        links_by_vid = self._forecast_load_variant_order_links_for_inventory(conn, ids)
        op_ids = sorted({
            int(oid)
            for links in links_by_vid.values()
            for oid, _ in (links or [])
            if self._parse_int(oid)
        })
        inv_by_op = self._forecast_load_inventory_by_order_product(conn, op_ids) if op_ids else {}
        tier_keys = (
            ('overseas', 'overseas_qty'),
            ('transit', 'transit_qty'),
            ('factory_stock', 'factory_stock_qty'),
            ('wip', 'wip_qty'),
        )
        for vid in ids:
            links = links_by_vid.get(vid) or []
            if not links:
                continue
            tier_parts = {k: [] for k, _ in tier_keys}
            for oid, qp in links:
                oid = int(oid)
                qp = max(1, int(qp or 1))
                inv = inv_by_op.get(oid) or {}
                for tier_k, inv_k in tier_keys:
                    tier_parts[tier_k].append(int(int(inv.get(inv_k) or 0) // qp))
            for tier_k, _ in tier_keys:
                parts = tier_parts.get(tier_k) or []
                if parts:
                    out[vid][tier_k] = min(parts)
        return out

    def _load_variant_overseas_sellable_map(self, conn, variant_ids):
        """variant_id -> 海外仓可售套数（仅海外仓；BOM 各件 floor 后取 min）。"""
        tier_map = self._load_variant_tier_sellable_map(conn, variant_ids)
        ids = sorted({int(x) for x in (variant_ids or []) if self._parse_int(x)})
        return {vid: int((tier_map.get(vid) or {}).get('overseas') or 0) for vid in ids}

    def _forecast_load_all_substitute_plans_by_owner(self, conn, owner_order_product_ids):
        """owner -> [{plan_id, plan_name, items: [(substitute_order_product_id, qty), ...]}, ...]

        含该下单 SKU 下全部有明细的替代发货方案（非仅 id 最小的一条）。
        """
        ids = sorted({int(x) for x in (owner_order_product_ids or []) if self._parse_int(x)})
        out = {i: [] for i in ids}
        if not ids:
            return out
        ph = ','.join(['%s'] * len(ids))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT ops.order_product_id AS owner_id,
                       ops.id AS plan_id,
                       ops.plan_name
                FROM order_product_shipping_plans ops
                WHERE ops.order_product_id IN ({ph})
                  AND EXISTS (
                      SELECT 1 FROM order_product_shipping_plan_items i
                      WHERE i.shipping_plan_id = ops.id
                  )
                ORDER BY ops.order_product_id ASC, ops.id ASC
                """,
                tuple(ids),
            )
            plan_rows = cur.fetchall() or []
            plan_meta = {}
            plan_ids = []
            plans_by_owner = {i: [] for i in ids}
            for rr in plan_rows:
                oid = self._parse_int(rr.get('owner_id'))
                pid = self._parse_int(rr.get('plan_id'))
                if not oid or not pid:
                    continue
                plan_meta[pid] = {
                    'owner_id': oid,
                    'plan_name': (rr.get('plan_name') or '').strip() or ('方案#' + str(pid)),
                }
                plan_ids.append(pid)
                plans_by_owner[oid].append({
                    'plan_id': pid,
                    'plan_name': plan_meta[pid]['plan_name'],
                    'items': [],
                })
            if not plan_ids:
                return out
            ph2 = ','.join(['%s'] * len(plan_ids))
            cur.execute(
                f"""
                SELECT shipping_plan_id,
                       substitute_order_product_id,
                       GREATEST(1, COALESCE(quantity, 1)) AS quantity
                FROM order_product_shipping_plan_items
                WHERE shipping_plan_id IN ({ph2})
                ORDER BY shipping_plan_id ASC, sort_order ASC, id ASC
                """,
                tuple(plan_ids),
            )
            items_by_plan = {pid: [] for pid in plan_ids}
            for rr in cur.fetchall() or []:
                pid = self._parse_int(rr.get('shipping_plan_id'))
                sid = self._parse_int(rr.get('substitute_order_product_id'))
                q = max(1, self._parse_int(rr.get('quantity')) or 1)
                if not pid or not sid:
                    continue
                items_by_plan.setdefault(pid, []).append((sid, q))
            for oid, plan_list in plans_by_owner.items():
                filled = []
                for grp in plan_list:
                    pid = grp['plan_id']
                    grp['items'] = list(items_by_plan.get(pid) or [])
                    if grp['items']:
                        filled.append(grp)
                out[oid] = filled
        return out

    @staticmethod
    def _forecast_merge_substitute_items_all_plans(plan_groups):
        """全部方案的替代 SKU 并集；同一替代 SKU 在多个方案中出现时配比累加。

        仅用于销量继承等非库存场景。库存请用 _forecast_inventory_assembled_from_substitute_plans，
        按方案分别成套后相加，避免多方案重复累加配比导致库存偏小。
        """
        merged = {}
        for grp in plan_groups or []:
            for sid, q in grp.get('items') or []:
                sid = int(sid)
                if not sid:
                    continue
                merged[sid] = int(merged.get(sid) or 0) + max(1, int(q))
        return sorted(((int(s), int(m)) for s, m in merged.items()), key=lambda x: x[0])

    @staticmethod
    def _forecast_substitute_ids_from_plan_groups(plan_groups):
        ids = set()
        for grp in plan_groups or []:
            for sid, _q in grp.get('items') or []:
                sid = int(sid) if sid else 0
                if sid:
                    ids.add(sid)
        return sorted(ids)

    def _forecast_load_all_substitute_plan_items_by_owner(self, conn, owner_order_product_ids):
        """owner -> [(substitute_order_product_id, qty_per_owner_unit), ...]（全部方案合并）。"""
        groups = self._forecast_load_all_substitute_plans_by_owner(conn, owner_order_product_ids)
        ids = sorted({int(x) for x in (owner_order_product_ids or []) if self._parse_int(x)})
        return {oid: self._forecast_merge_substitute_items_all_plans(groups.get(oid) or []) for oid in ids}

    def _forecast_load_default_substitute_plan_items_by_owner(self, conn, owner_order_product_ids):
        """owner_order_product_id -> [(substitute_order_product_id, qty_per_owner_unit), ...]

        取该下单 SKU 下 id 最小、且至少含一条 order_product_shipping_plan_items 的发货方案；
        用于销量继承、下市迁移等需单一默认方案的场景。库存折算请用 _forecast_load_all_substitute_plan_items_by_owner。
        无方案或无明细时返回空列表，库存仍按 sales_variant_order_links 中的本体 SKU 计算。
        """
        ids = sorted({int(x) for x in (owner_order_product_ids or []) if self._parse_int(x)})
        out = {i: [] for i in ids}
        if not ids:
            return out
        ph = ','.join(['%s'] * len(ids))
        tpl = tuple(ids)
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT ops.order_product_id AS owner_id, MIN(ops.id) AS plan_id
                FROM order_product_shipping_plans ops
                WHERE ops.order_product_id IN ({ph})
                  AND EXISTS (
                      SELECT 1 FROM order_product_shipping_plan_items i
                      WHERE i.shipping_plan_id = ops.id
                  )
                GROUP BY ops.order_product_id
                """,
                tpl,
            )
            plan_map = {}
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('owner_id'))
                pid = self._parse_int(rr.get('plan_id'))
                if oid and pid:
                    plan_map[oid] = pid
            if not plan_map:
                return out
            plan_ids = sorted(set(plan_map.values()))
            ph2 = ','.join(['%s'] * len(plan_ids))
            cur.execute(
                f"""
                SELECT shipping_plan_id,
                       substitute_order_product_id,
                       GREATEST(1, COALESCE(quantity, 1)) AS quantity
                FROM order_product_shipping_plan_items
                WHERE shipping_plan_id IN ({ph2})
                ORDER BY shipping_plan_id ASC, sort_order ASC, id ASC
                """,
                tuple(plan_ids),
            )
            items_by_plan = {}
            for rr in cur.fetchall() or []:
                pid = self._parse_int(rr.get('shipping_plan_id'))
                sid = self._parse_int(rr.get('substitute_order_product_id'))
                q = max(1, self._parse_int(rr.get('quantity')) or 1)
                if not pid or not sid:
                    continue
                items_by_plan.setdefault(pid, []).append((sid, q))
            for oid, pid in plan_map.items():
                out[oid] = list(items_by_plan.get(pid) or [])
        return out

    @staticmethod
    def _forecast_expand_owner_pairs_with_substitute_plans(owner_pairs, plan_items_by_owner):
        """将 [(owner_oid, link_weight), ...] 按各 owner 的默认替代方案展开为替代下单 SKU 的等效件数。"""
        merged = {}
        for oid, w in owner_pairs or []:
            oid = int(oid)
            w = max(1, int(w))
            subs = plan_items_by_owner.get(oid) if plan_items_by_owner else None
            if not subs:
                merged[oid] = int(merged.get(oid) or 0) + w
                continue
            for sid, mult in subs:
                sid = int(sid)
                mult = max(1, int(mult))
                merged[sid] = int(merged.get(sid) or 0) + w * mult
        return sorted(merged.items(), key=lambda x: x[0])

    def _forecast_bom_tier_assembled_counts(self, inv_by_op, bom_links):
        """按 BOM 成套计算各阶可售套数；上一阶用剩的散件并入下一阶再算 min 瓶颈。"""
        z = self._forecast_inventory_zero()
        if not bom_links:
            return z
        links = [(int(oid), max(1, int(qp))) for oid, qp in bom_links]
        tier_keys = ['overseas_qty', 'transit_qty', 'factory_stock_qty', 'wip_qty']
        carry = {}
        out = {}
        for tier in tier_keys:
            pool = {}
            for oid, qp in links:
                base = int((inv_by_op.get(oid) or {}).get(tier) or 0)
                pool[oid] = base + int(carry.get(oid) or 0)
            assembled = min(int(pool.get(oid, 0)) // qp for oid, qp in links)
            out[tier] = assembled
            for oid, qp in links:
                carry[oid] = int(pool.get(oid, 0)) - assembled * qp
        return out

    def _forecast_inventory_tier_sum_assembled(self, inv_by_op, source_pairs):
        """单条 BOM 链接列表的各阶可售套数（min 瓶颈）。"""
        return self._forecast_bom_tier_assembled_counts(inv_by_op, source_pairs)

    def _forecast_add_inv_agg(self, base, extra):
        """各阶库存件数相加（本体 1:1 + 替代方案成套结果）。"""
        out = self._forecast_inventory_zero()
        b = base or out
        e = extra or out
        for k in out:
            out[k] = int(b.get(k) or 0) + int(e.get(k) or 0)
        return out

    def _forecast_owner_inv_agg(self, inv_by_op, oid):
        oid = self._parse_int(oid)
        if not oid:
            return self._forecast_inventory_zero()
        return dict(inv_by_op.get(int(oid)) or self._forecast_inventory_zero())

    def _forecast_merge_surplus_detail_add(self, a, b):
        """合并两份盈余明细（相加），用于本体库存 + 替代方案折算库存。"""
        left = a or self._forecast_empty_surplus_detail()
        right = b or self._forecast_empty_surplus_detail()
        merged = self._forecast_empty_surplus_detail()
        regions = set(left.get('overseas_by_region') or {}) | set(right.get('overseas_by_region') or {})
        for rname in regions:
            rname = str(rname or '').strip() or '-'
            merged['overseas_by_region'][rname] = (
                int((left.get('overseas_by_region') or {}).get(rname) or 0)
                + int((right.get('overseas_by_region') or {}).get(rname) or 0)
            )
        merged['transit_batches'] = list(left.get('transit_batches') or []) + list(right.get('transit_batches') or [])
        merged['factory_stock_qty'] = int(left.get('factory_stock_qty') or 0) + int(right.get('factory_stock_qty') or 0)
        merged['wip_batches'] = list(left.get('wip_batches') or []) + list(right.get('wip_batches') or [])
        merged['overseas_updated_at'] = self._forecast_date_max_iso(
            left.get('overseas_updated_at'), right.get('overseas_updated_at')
        )
        merged['transit_updated_at'] = self._forecast_date_max_iso(
            left.get('transit_updated_at'), right.get('transit_updated_at')
        )
        merged['factory_stock_updated_at'] = self._forecast_date_max_iso(
            left.get('factory_stock_updated_at'), right.get('factory_stock_updated_at')
        )
        merged['wip_updated_at'] = self._forecast_date_max_iso(
            left.get('wip_updated_at'), right.get('wip_updated_at')
        )
        return merged

    def _forecast_inventory_assembled_from_substitute_plans(self, inv_by_op, plan_groups):
        """各替代发货方案独立成套后相加（方案之间为并集，同 SKU 跨方案不累加配比）。"""
        out = self._forecast_inventory_zero()
        for grp in plan_groups or []:
            links = [
                (int(self._parse_int(sid)), max(1, self._parse_int(mult) or 1))
                for sid, mult in (grp.get('items') or [])
                if self._parse_int(sid)
            ]
            if not links:
                continue
            part = self._forecast_bom_tier_assembled_counts(inv_by_op, links)
            for k in out:
                out[k] += int(part.get(k) or 0)
        return out

    def _forecast_build_inventory_stock_sources(self, oid, brief, plan_groups):
        """面板展示用：本体 + 各替代方案中的替代 SKU（去重）。"""
        oid = self._parse_int(oid)
        if not oid:
            return []
        brief = brief or {}
        out = [{
            'order_product_id': int(oid),
            'sku': ((brief.get(int(oid)) or {}).get('sku') or '').strip() or ('#' + str(oid)),
            'role': 'self',
            'plan_name': '本体',
            'quantity': 1,
        }]
        seen = {int(oid)}
        for grp in plan_groups or []:
            pname = (grp.get('plan_name') or '').strip() or '替代方案'
            for sid, mult in grp.get('items') or []:
                sid = self._parse_int(sid)
                if not sid or sid in seen:
                    continue
                seen.add(sid)
                out.append({
                    'order_product_id': int(sid),
                    'sku': ((brief.get(sid) or {}).get('sku') or '').strip() or ('#' + str(sid)),
                    'role': 'substitute',
                    'plan_name': pname,
                    'quantity': max(1, self._parse_int(mult) or 1),
                })
        return out

    def _forecast_build_inventory_source_lines(self, oid, is_on_market, plan_groups, brief):
        """替代 SKU + 本体（展示用：替代同 SKU 取各方案配比最大值）。"""
        oid = self._parse_int(oid)
        if not oid:
            return []
        brief = brief or {}
        lines = []
        seen_sid = set()

        def _push(sid, mult, role):
            sid = self._parse_int(sid)
            if not sid or sid in seen_sid:
                return
            seen_sid.add(sid)
            mult = max(1, int(mult or 1))
            b = brief.get(sid) or {}
            lines.append({
                'order_product_id': sid,
                'sku': (b.get('sku') or '').strip() or ('#' + str(sid)),
                'qty_per_unit': mult,
                'role': role,
                'is_on_market': self._parse_int(b.get('is_on_market')) or 0,
            })

        sub_by_sid = {}
        for grp in plan_groups or []:
            for sid, mult in grp.get('items') or []:
                sid = self._parse_int(sid)
                if not sid or sid == oid:
                    continue
                mult = max(1, int(mult or 1))
                sub_by_sid[sid] = max(int(sub_by_sid.get(sid) or 0), mult)
        for sid in sorted(sub_by_sid.keys()):
            _push(sid, sub_by_sid[sid], 'substitute')

        owner_b = brief.get(int(oid)) or {}
        owner_line = {
            'order_product_id': int(oid),
            'sku': (owner_b.get('sku') or '').strip() or ('#' + str(oid)),
            'qty_per_unit': 1,
            'role': 'self',
            'is_on_market': self._parse_int(is_on_market) or self._parse_int(owner_b.get('is_on_market')) or 0,
        }
        return [owner_line] + lines

    def _forecast_build_inventory_tier_breakdown_from_plans(self, inv_by_op, oid, plan_groups, brief):
        """按方案分别折算后，同替代 SKU 的「计入」数量跨方案相加。"""
        oid = self._parse_int(oid)
        brief = brief or {}
        tier_meta = (
            ('overseas_qty', '海外仓'),
            ('transit_qty', '在途'),
            ('factory_stock_qty', '在库'),
            ('wip_qty', '在制'),
        )
        acc = {}
        for grp in plan_groups or []:
            links = [
                (int(self._parse_int(sid)), max(1, self._parse_int(mult) or 1))
                for sid, mult in (grp.get('items') or [])
                if self._parse_int(sid) and self._parse_int(sid) != oid
            ]
            if not links:
                continue
            plan_asm = self._forecast_bom_tier_assembled_counts(inv_by_op, links)
            for sid, mult in links:
                b = brief.get(sid) or {}
                for tier_key, tier_label in tier_meta:
                    bucket = acc.setdefault(sid, {}).setdefault(tier_key, {
                        'order_product_id': sid,
                        'sku': (b.get('sku') or '').strip() or ('#' + str(sid)),
                        'qty_per_unit': mult,
                        'role': 'substitute',
                        'is_on_market': self._parse_int(b.get('is_on_market')) or 0,
                        'tier_key': tier_key,
                        'tier_label': tier_label,
                        'raw_qty': int((inv_by_op.get(sid) or {}).get(tier_key) or 0),
                        'assembled_qty': 0,
                    })
                    bucket['assembled_qty'] += int(plan_asm.get(tier_key) or 0)
                    bucket['qty_per_unit'] = max(bucket['qty_per_unit'], mult)
        breakdown = {tk: [] for tk, _ in tier_meta}
        if oid:
            owner_b = brief.get(int(oid)) or {}
            for tier_key, tier_label in tier_meta:
                raw = int((inv_by_op.get(int(oid)) or {}).get(tier_key) or 0)
                breakdown[tier_key].append({
                    'order_product_id': int(oid),
                    'sku': (owner_b.get('sku') or '').strip() or ('#' + str(oid)),
                    'qty_per_unit': 1,
                    'role': 'self',
                    'is_on_market': self._parse_int(owner_b.get('is_on_market')) or 0,
                    'tier_key': tier_key,
                    'tier_label': tier_label,
                    'raw_qty': raw,
                    'assembled_qty': raw,
                })
        for sid in sorted(acc.keys()):
            for tier_key, tier_label in tier_meta:
                entry = acc.get(sid, {}).get(tier_key)
                if entry:
                    breakdown[tier_key].append(entry)
        return breakdown

    def _forecast_inventory_source_lines_to_pairs(self, source_lines):
        return [
            (int(self._parse_int(x.get('order_product_id'))), max(1, self._parse_int(x.get('qty_per_unit')) or 1))
            for x in (source_lines or [])
            if self._parse_int(x.get('order_product_id'))
        ]

    def _forecast_build_inventory_tier_breakdown(self, inv_by_op, source_lines):
        tier_meta = (
            ('overseas_qty', '海外仓'),
            ('transit_qty', '在途'),
            ('factory_stock_qty', '在库'),
            ('wip_qty', '在制'),
        )
        breakdown = {}
        for tier_key, tier_label in tier_meta:
            entries = []
            for ln in source_lines or []:
                sid = self._parse_int(ln.get('order_product_id'))
                mult = max(1, self._parse_int(ln.get('qty_per_unit')) or 1)
                if not sid:
                    continue
                raw = int((inv_by_op.get(sid) or {}).get(tier_key) or 0)
                assembled = int(
                    self._forecast_bom_tier_assembled_counts(inv_by_op, [(sid, mult)]).get(tier_key) or 0
                )
                entries.append({
                    'order_product_id': sid,
                    'sku': ln.get('sku') or '',
                    'qty_per_unit': mult,
                    'role': ln.get('role') or 'substitute',
                    'is_on_market': self._parse_int(ln.get('is_on_market')) or 0,
                    'tier_key': tier_key,
                    'tier_label': tier_label,
                    'raw_qty': raw,
                    'assembled_qty': assembled,
                })
            breakdown[tier_key] = entries
        return breakdown

    def _forecast_format_substitute_plans_for_api(self, plans, brief):
        brief = brief or {}
        out = []
        for pl in plans or []:
            items = []
            for sid, q in pl.get('items') or []:
                sid = self._parse_int(sid)
                if not sid:
                    continue
                items.append({
                    'substitute_order_product_id': sid,
                    'sku': (brief.get(sid) or {}).get('sku') or '',
                    'quantity': max(1, self._parse_int(q) or 1),
                })
            out.append({
                'plan_id': pl.get('plan_id'),
                'plan_name': pl.get('plan_name') or '',
                'items': items,
            })
        return out

    def _forecast_inventory_composition_meta(
        self, source_lines, row_sku='', substitute_plans=None, brief=None,
    ):
        lines = list(source_lines or [])
        sub_n = sum(1 for x in lines if str(x.get('role') or '') != 'self')
        has_indicator = sub_n > 0
        kind = 'multi_substitute' if sub_n > 1 else ('substitute_only' if sub_n == 1 else 'no_substitute')
        return {
            'has_indicator': bool(has_indicator),
            'kind': kind,
            'row_sku': (row_sku or '').strip(),
            'source_lines': lines,
            'substitute_plans': self._forecast_format_substitute_plans_for_api(
                substitute_plans, brief,
            ) if substitute_plans is not None else [],
        }

    def _forecast_order_row_inventory_aggregate(
        self, inv_by_op, oid, is_on_market, plan_groups, brief, substitute_plans=None,
    ):
        plans = list(substitute_plans if substitute_plans is not None else plan_groups or [])
        source_lines = self._forecast_build_inventory_source_lines(
            oid, is_on_market, plans, brief,
        )
        pairs = self._forecast_inventory_source_lines_to_pairs(source_lines)
        owner_agg = self._forecast_owner_inv_agg(inv_by_op, oid)
        if plans:
            sub_agg = self._forecast_inventory_assembled_from_substitute_plans(inv_by_op, plans)
            agg = self._forecast_add_inv_agg(owner_agg, sub_agg)
        else:
            agg = owner_agg
        sku = (brief.get(int(oid)) or {}).get('sku') or ''
        comp = self._forecast_inventory_composition_meta(
            source_lines, sku, substitute_plans=substitute_plans, brief=brief,
        )
        comp['tier_breakdown'] = self._forecast_build_inventory_tier_breakdown_from_plans(
            inv_by_op, oid, plans, brief,
        )
        comp['pairs'] = pairs
        comp['owner_order_product_id'] = int(oid) if oid else 0
        comp['stock_sources'] = self._forecast_build_inventory_stock_sources(oid, brief, plans)
        return agg, comp

    def _forecast_row_history_sales_avg_tail(self, row, history_months, tail=3):
        """取 history_months 末尾最多 tail 个月的月均销量（用于动销月列）。"""
        if not history_months:
            return None
        keys = list(history_months)[-tail:] if len(history_months) > tail else list(history_months)
        if not keys:
            return None
        total = 0.0
        for k in keys:
            h = (row.get('history') or {}).get(k) or {}
            total += float(h.get('sales_qty') or 0)
        avg = total / float(len(keys))
        return avg if avg > 1e-9 else None

    # -------------------------------------------------------------------------
    # 动销月：30 天窗口销量与库存覆盖月数
    # -------------------------------------------------------------------------

    def _turnover_window_bounds(self, window):
        """从窗口 dict 解析 window_start / window_end（YYYY-MM-DD）。"""
        if not window:
            return '', ''
        ws = str(window.get('window_start') or '')[:10]
        we = str(window.get('window_end') or '')[:10]
        return ws, we

    def _turnover_merge_rolling_daily_maps(self, conn, ids, window, shop_ids, rolling_loader, daily_loader):
        """优先 rolling 快照，零销量 id 再查日表兜底。"""
        out = rolling_loader(conn, ids, window, shop_ids=shop_ids)
        missing = [i for i in ids if float(out.get(i) or 0) <= 1e-9]
        if missing:
            daily = daily_loader(conn, missing, window, shop_ids=shop_ids)
            for entity_id, sales in (daily or {}).items():
                if float(sales or 0) > 1e-9:
                    out[int(entity_id)] = float(sales)
        return out

    def _turnover_sales_window(self, conn):
        """动销月分母窗口：全局最新 record_date 为终点，向前 30 个自然日（含首尾）。"""
        anchor_raw = self._forecast_perf_max_record_date(conn)
        if anchor_raw is not None:
            anchor_text = self._forecast_format_date_only(anchor_raw)
            try:
                anchor_dt = datetime.strptime(anchor_text, '%Y-%m-%d').date()
            except Exception:
                anchor_dt = datetime.now().date()
        else:
            anchor_dt = datetime.now().date()
        window_end = anchor_dt
        window_start = window_end - timedelta(days=29)
        return {
            'anchor_date': window_end.strftime('%Y-%m-%d'),
            'window_start': window_start.strftime('%Y-%m-%d'),
            'window_end': window_end.strftime('%Y-%m-%d'),
            'days': 30,
        }

    def _turnover_coverage(self, numerator, denominator_sales):
        if denominator_sales is None or float(denominator_sales) <= 1e-9:
            return None
        num = int(numerator or 0)
        if num <= 0:
            return 0
        return round(float(num) / float(denominator_sales), 2)

    def _turnover_shop_filter_sql(self, shop_ids, sp_alias='sp'):
        if not shop_ids:
            return '', []
        ids = [int(x) for x in shop_ids if self._parse_int(x)]
        if not ids:
            return '', []
        ph = ','.join(['%s'] * len(ids))
        return f' AND {sp_alias}.shop_id IN ({ph}) ', ids

    def _turnover_load_variant_window_sales_from_rolling(self, conn, variant_ids, window, shop_ids=None):
        ids = sorted({int(x) for x in (variant_ids or []) if self._parse_int(x)})
        out = {}
        if not ids or not window:
            return out
        if not self._table_exists_simple(conn, 'sales_perf_rolling_30d'):
            return out
        ws, we = self._turnover_window_bounds(window)
        if not ws or not we:
            return out
        shop_frag, shop_params = self._turnover_shop_filter_sql(shop_ids, 'sp')
        chunk_size = 400
        with conn.cursor() as cur:
            for i in range(0, len(ids), chunk_size):
                chunk = ids[i:i + chunk_size]
                ph = ','.join(['%s'] * len(chunk))
                params = [ws, we] + list(chunk) + shop_params
                cur.execute(
                    f"""
                    SELECT sp.variant_id AS variant_id,
                           SUM(COALESCE(r.sales_qty, 0)) AS sales_qty
                    FROM sales_products sp
                    INNER JOIN sales_perf_rolling_30d r
                        ON r.sales_product_id = sp.id
                       AND r.window_start = %s
                       AND r.window_end = %s
                    WHERE sp.variant_id IN ({ph})
                    {shop_frag}
                    GROUP BY sp.variant_id
                    """,
                    tuple(params),
                )
                for row in cur.fetchall() or []:
                    vid = self._parse_int(row.get('variant_id'))
                    if vid:
                        out[vid] = float(row.get('sales_qty') or 0)
        return out

    def _turnover_load_variant_window_sales_from_daily(self, conn, variant_ids, window, shop_ids=None):
        ids = sorted({int(x) for x in (variant_ids or []) if self._parse_int(x)})
        out = {}
        if not ids or not window:
            return out
        ws, we = self._turnover_window_bounds(window)
        if not ws or not we:
            return out
        shop_frag, shop_params = self._turnover_shop_filter_sql(shop_ids, 'sp')
        chunk_size = 400
        with conn.cursor() as cur:
            for i in range(0, len(ids), chunk_size):
                chunk = ids[i:i + chunk_size]
                ph = ','.join(['%s'] * len(chunk))
                params = [ws, we] + list(chunk) + shop_params
                cur.execute(
                    f"""
                    SELECT sp.variant_id AS variant_id,
                           SUM(COALESCE(spp.sales_qty, 0)) AS sales_qty
                    FROM sales_products sp
                    INNER JOIN sales_product_performances spp
                        ON spp.sales_product_id = sp.id
                       AND spp.record_date >= %s
                       AND spp.record_date <= %s
                    WHERE sp.variant_id IN ({ph})
                    {shop_frag}
                    GROUP BY sp.variant_id
                    """,
                    tuple(params),
                )
                for row in cur.fetchall() or []:
                    vid = self._parse_int(row.get('variant_id'))
                    if vid:
                        out[vid] = float(row.get('sales_qty') or 0)
        return out

    def _turnover_load_variant_window_sales_map(self, conn, variant_ids, window=None, shop_ids=None):
        """变体在动销月窗口内的 sales_qty 汇总（优先 rolling_30d 快照，日表兜底）。"""
        window = window or self._turnover_sales_window(conn)
        ids = sorted({int(x) for x in (variant_ids or []) if self._parse_int(x)})
        return self._turnover_merge_rolling_daily_maps(
            conn, ids, window, shop_ids,
            self._turnover_load_variant_window_sales_from_rolling,
            self._turnover_load_variant_window_sales_from_daily,
        )

    def _turnover_load_op_window_sales_from_daily(self, conn, order_product_ids, window, shop_ids=None):
        ids = sorted({int(x) for x in (order_product_ids or []) if self._parse_int(x)})
        out = {i: 0.0 for i in ids}
        if not ids or not window:
            return out
        ws, we = self._turnover_window_bounds(window)
        if not ws or not we:
            return out
        shop_frag, shop_params = self._turnover_shop_filter_sql(shop_ids, 'sp')
        chunk_size = 400
        with conn.cursor() as cur:
            for i in range(0, len(ids), chunk_size):
                chunk = ids[i:i + chunk_size]
                ph = ','.join(['%s'] * len(chunk))
                params = [ws, we] + list(chunk) + shop_params
                cur.execute(
                    f"""
                    SELECT l.order_product_id,
                           SUM(COALESCE(spp.sales_qty, 0) * GREATEST(1, COALESCE(l.quantity, 1))) AS sales_qty
                    FROM sales_variant_order_links l
                    INNER JOIN sales_products sp ON sp.variant_id = l.variant_id
                    INNER JOIN sales_product_performances spp
                        ON spp.sales_product_id = sp.id
                       AND spp.record_date >= %s
                       AND spp.record_date <= %s
                    WHERE l.order_product_id IN ({ph})
                    {shop_frag}
                    GROUP BY l.order_product_id
                    """,
                    tuple(params),
                )
                for row in cur.fetchall() or []:
                    op_id = self._parse_int(row.get('order_product_id'))
                    if op_id:
                        out[int(op_id)] = float(row.get('sales_qty') or 0)
        return out

    def _turnover_op_sales_via_variant_links(self, conn, order_product_ids, window, shop_ids=None):
        """下单 SKU 窗口销量：变体窗口销量 × BOM 件数累加。"""
        ids = sorted({int(x) for x in (order_product_ids or []) if self._parse_int(x)})
        out = {i: 0.0 for i in ids}
        if not ids or not window:
            return out
        links_by_op = self._forecast_load_order_variant_links_by_order_product_ids(conn, ids)
        all_vids = sorted({
            self._parse_int(vl.get('variant_id'))
            for links in links_by_op.values()
            for vl in (links or [])
            if self._parse_int(vl.get('variant_id'))
        })
        if not all_vids:
            return out
        variant_sales = self._turnover_load_variant_window_sales_map(
            conn, all_vids, window=window, shop_ids=shop_ids,
        )
        for op_id in ids:
            total = 0.0
            for vl in links_by_op.get(op_id) or []:
                vid = self._parse_int(vl.get('variant_id'))
                q = max(1, self._parse_int(vl.get('quantity')) or 1)
                if vid:
                    total += float(variant_sales.get(vid) or 0) * float(q)
            if total > 1e-9:
                out[int(op_id)] = total
        return out

    def _turnover_load_sales_product_window_sales_from_rolling(self, conn, sales_product_ids, window, shop_ids=None):
        ids = sorted({int(x) for x in (sales_product_ids or []) if self._parse_int(x)})
        out = {}
        if not ids or not window:
            return out
        if not self._table_exists_simple(conn, 'sales_perf_rolling_30d'):
            return out
        ws, we = self._turnover_window_bounds(window)
        if not ws or not we:
            return out
        shop_frag, shop_params = self._turnover_shop_filter_sql(shop_ids, 'sp')
        chunk_size = 400
        with conn.cursor() as cur:
            for i in range(0, len(ids), chunk_size):
                chunk = ids[i:i + chunk_size]
                ph = ','.join(['%s'] * len(chunk))
                params = [ws, we] + list(chunk) + shop_params
                cur.execute(
                    f"""
                    SELECT sp.id AS sales_product_id,
                           SUM(COALESCE(r.sales_qty, 0)) AS sales_qty
                    FROM sales_products sp
                    INNER JOIN sales_perf_rolling_30d r
                        ON r.sales_product_id = sp.id
                       AND r.window_start = %s
                       AND r.window_end = %s
                    WHERE sp.id IN ({ph})
                    {shop_frag}
                    GROUP BY sp.id
                    """,
                    tuple(params),
                )
                for row in cur.fetchall() or []:
                    spid = self._parse_int(row.get('sales_product_id'))
                    if spid:
                        out[spid] = float(row.get('sales_qty') or 0)
        return out

    def _turnover_load_sales_product_window_sales_from_daily(self, conn, sales_product_ids, window, shop_ids=None):
        ids = sorted({int(x) for x in (sales_product_ids or []) if self._parse_int(x)})
        out = {}
        if not ids or not window:
            return out
        ws, we = self._turnover_window_bounds(window)
        if not ws or not we:
            return out
        shop_frag, shop_params = self._turnover_shop_filter_sql(shop_ids, 'sp')
        chunk_size = 400
        with conn.cursor() as cur:
            for i in range(0, len(ids), chunk_size):
                chunk = ids[i:i + chunk_size]
                ph = ','.join(['%s'] * len(chunk))
                params = [ws, we] + list(chunk) + shop_params
                cur.execute(
                    f"""
                    SELECT sp.id AS sales_product_id,
                           SUM(COALESCE(spp.sales_qty, 0)) AS sales_qty
                    FROM sales_products sp
                    INNER JOIN sales_product_performances spp
                        ON spp.sales_product_id = sp.id
                       AND spp.record_date >= %s
                       AND spp.record_date <= %s
                    WHERE sp.id IN ({ph})
                    {shop_frag}
                    GROUP BY sp.id
                    """,
                    tuple(params),
                )
                for row in cur.fetchall() or []:
                    spid = self._parse_int(row.get('sales_product_id'))
                    if spid:
                        out[spid] = float(row.get('sales_qty') or 0)
        return out

    def _turnover_load_sales_product_window_sales_map(self, conn, sales_product_ids, window=None, shop_ids=None):
        """销售产品在动销月窗口内的 sales_qty 汇总（优先 rolling_30d，日表兜底）。"""
        window = window or self._turnover_sales_window(conn)
        ids = sorted({int(x) for x in (sales_product_ids or []) if self._parse_int(x)})
        return self._turnover_merge_rolling_daily_maps(
            conn, ids, window, shop_ids,
            self._turnover_load_sales_product_window_sales_from_rolling,
            self._turnover_load_sales_product_window_sales_from_daily,
        )

    def _turnover_load_op_window_sales_from_rolling(self, conn, order_product_ids, window, shop_ids=None):
        """下单 SKU 窗口销量（优先 sales_perf_op_rolling_30d 快照）。"""
        if shop_ids:
            return {}
        ids = sorted({int(x) for x in (order_product_ids or []) if self._parse_int(x)})
        out = {i: 0.0 for i in ids}
        if not ids or not window:
            return out
        if not self._table_exists_simple(conn, 'sales_perf_op_rolling_30d'):
            return out
        ws, we = self._turnover_window_bounds(window)
        if not ws or not we:
            return out
        chunk_size = 400
        with conn.cursor() as cur:
            for i in range(0, len(ids), chunk_size):
                chunk = ids[i:i + chunk_size]
                ph = ','.join(['%s'] * len(chunk))
                params = [ws, we] + list(chunk)
                cur.execute(
                    f"""
                    SELECT order_product_id, COALESCE(sales_qty, 0) AS sales_qty
                    FROM sales_perf_op_rolling_30d
                    WHERE window_start = %s
                      AND window_end = %s
                      AND order_product_id IN ({ph})
                    """,
                    tuple(params),
                )
                for row in cur.fetchall() or []:
                    op_id = self._parse_int(row.get('order_product_id'))
                    if op_id:
                        out[int(op_id)] = float(row.get('sales_qty') or 0)
        return out

    def _turnover_load_op_window_sales_map(self, conn, order_product_ids, window=None, shop_ids=None):
        """下单 SKU -> 窗口 sales_qty（优先 op rolling 快照，日表与变体链接兜底）。"""
        ids = sorted({int(x) for x in (order_product_ids or []) if self._parse_int(x)})
        if not ids:
            return {}
        window = window or self._turnover_sales_window(conn)
        if shop_ids:
            out = {i: 0.0 for i in ids}
            missing = ids
        else:
            out = self._turnover_load_op_window_sales_from_rolling(conn, ids, window, shop_ids=shop_ids)
            missing = [i for i in ids if float(out.get(i) or 0) <= 1e-9]
        if missing:
            daily = self._turnover_load_op_window_sales_from_daily(conn, missing, window, shop_ids=shop_ids)
            for op_id, sales in (daily or {}).items():
                if float(sales or 0) > 1e-9:
                    out[int(op_id)] = float(sales)
            missing = [i for i in missing if float(out.get(i) or 0) <= 1e-9]
        if missing:
            link_sales = self._turnover_op_sales_via_variant_links(
                conn, missing, window, shop_ids=shop_ids,
            )
            for op_id, sales in (link_sales or {}).items():
                if float(sales or 0) > 1e-9:
                    out[int(op_id)] = float(sales)
        return out

    def _turnover_assembled_sales_from_substitute_plans(self, plan_groups, sales_by_op, owner_id=None):
        """各替代发货方案按瓶颈折算 owner 等效窗口销量后相加（与 _forecast_inventory_assembled_from_substitute_plans 对称）。"""
        owner_id = int(owner_id) if self._parse_int(owner_id) else 0
        total = 0.0
        for grp in plan_groups or []:
            vals = []
            for sid, mult in grp.get('items') or []:
                sid = self._parse_int(sid)
                if not sid or (owner_id and int(sid) == owner_id):
                    continue
                mult = max(1, self._parse_int(mult) or 1)
                s = float((sales_by_op or {}).get(int(sid)) or 0)
                if s > 1e-9:
                    vals.append(s / float(mult))
            if vals:
                total += min(vals)
        return total

    def _turnover_owner_effective_window_sales(self, owner_id, plan_groups, sales_by_op):
        """本体窗口销量 + 全部替代发货方案折算销量（owner 等效件数）。"""
        oid = self._parse_int(owner_id)
        if not oid:
            return 0.0
        owner_sales = float((sales_by_op or {}).get(int(oid)) or 0)
        sub_sales = self._turnover_assembled_sales_from_substitute_plans(
            plan_groups, sales_by_op, owner_id=oid,
        )
        total = owner_sales + sub_sales
        return total if total > 1e-9 else 0.0

    def _turnover_load_op_effective_window_sales_map(self, conn, order_product_ids, window=None, shop_ids=None):
        """下单 SKU -> 近30天等效窗口销量（本体 + 全部替代发货方案折算）。"""
        ids = sorted({int(x) for x in (order_product_ids or []) if self._parse_int(x)})
        if not ids:
            return {}
        window = window or self._turnover_sales_window(conn)
        substitute_plans = self._forecast_load_all_substitute_plans_by_owner(conn, ids)
        extra_subs = []
        for oid in ids:
            for sid in self._forecast_substitute_ids_from_plan_groups(substitute_plans.get(oid) or []):
                extra_subs.append(sid)
        load_ids = sorted(set(ids).union(int(x) for x in extra_subs if self._parse_int(x)))
        sales_map = self._turnover_load_op_window_sales_map(conn, load_ids, window=window, shop_ids=shop_ids)
        still_missing = [i for i in ids if float(sales_map.get(i) or 0) <= 1e-9]
        if still_missing:
            link_sales = self._turnover_op_sales_via_variant_links(
                conn, still_missing, window, shop_ids=shop_ids,
            )
            for op_id, sales in (link_sales or {}).items():
                if float(sales or 0) > 1e-9:
                    sales_map[int(op_id)] = float(sales)
        return {
            oid: self._turnover_owner_effective_window_sales(
                oid, substitute_plans.get(oid) or [], sales_map,
            )
            for oid in ids
        }

    def _turnover_spec_coverage_from_bom_pairs(self, pairs, op_sales_map, inv_by_op, substitute_plans_by_owner):
        """规格行动销月三列：各下单 SKU（含替代方案库存）取最小值。"""
        vals_o, vals_ot, vals_tot = [], [], []
        for oid, qp in pairs or []:
            oid = int(oid)
            qp = max(1, int(qp or 1))
            if not oid:
                continue
            plans = (substitute_plans_by_owner or {}).get(oid) or []
            avg_op = self._turnover_owner_effective_window_sales(oid, plans, op_sales_map)
            if avg_op <= 1e-9:
                continue
            ai = self._forecast_inventory_assembled_from_substitute_plans(inv_by_op, plans)
            inv = inv_by_op.get(oid) or {}
            ai_sum = sum(int(ai.get(k) or 0) for k in ('overseas_qty', 'transit_qty', 'factory_stock_qty', 'wip_qty'))
            raw_sum = sum(int(inv.get(k) or 0) for k in ('overseas_qty', 'transit_qty', 'factory_stock_qty', 'wip_qty'))
            if ai_sum <= 0 and raw_sum > 0:
                ai = {
                    'overseas_qty': int(inv.get('overseas_qty') or 0),
                    'transit_qty': int(inv.get('transit_qty') or 0),
                    'factory_stock_qty': int(inv.get('factory_stock_qty') or 0),
                    'wip_qty': int(inv.get('wip_qty') or 0),
                }
            oi = int(int(ai.get('overseas_qty') or 0) // qp)
            ti = int(int(ai.get('transit_qty') or 0) // qp)
            sti = int(int(ai.get('factory_stock_qty') or 0) // qp)
            wpi = int(int(ai.get('wip_qty') or 0) // qp)
            if oi > 0:
                vals_o.append(round(float(oi) / avg_op, 2))
            if oi + ti > 0:
                vals_ot.append(round(float(oi + ti) / avg_op, 2))
            if oi + ti + sti + wpi > 0:
                vals_tot.append(round(float(oi + ti + sti + wpi) / avg_op, 2))
        return {
            'months_cover_overseas': min(vals_o) if vals_o else None,
            'months_cover_overseas_transit': min(vals_ot) if vals_ot else None,
            'months_cover_total': min(vals_tot) if vals_tot else None,
        }

    def _turnover_spec_coverage_from_bom_pairs_raw(self, pairs, op_sales_map, inv_by_op):
        """规格行动销月三列：本体库存（不含替代方案），与海外仓可售口径一致。"""
        vals_o, vals_ot, vals_tot = [], [], []
        for oid, qp in pairs or []:
            oid = int(oid)
            qp = max(1, int(qp or 1))
            if not oid:
                continue
            avg_op = float(op_sales_map.get(oid) or 0)
            if avg_op <= 1e-9:
                continue
            inv = inv_by_op.get(oid) or {}
            oi = int(int(inv.get('overseas_qty') or 0) // qp)
            ti = int(int(inv.get('transit_qty') or 0) // qp)
            sti = int(int(inv.get('factory_stock_qty') or 0) // qp)
            wpi = int(int(inv.get('wip_qty') or 0) // qp)
            if oi > 0:
                vals_o.append(round(float(oi) / avg_op, 2))
            if oi + ti > 0:
                vals_ot.append(round(float(oi + ti) / avg_op, 2))
            if oi + ti + sti + wpi > 0:
                vals_tot.append(round(float(oi + ti + sti + wpi) / avg_op, 2))
        return {
            'months_cover_overseas': min(vals_o) if vals_o else None,
            'months_cover_overseas_transit': min(vals_ot) if vals_ot else None,
            'months_cover_total': min(vals_tot) if vals_tot else None,
        }

    def _turnover_cov_all_none(self, cov):
        return all((cov or {}).get(k) is None for k in (
            'months_cover_overseas', 'months_cover_overseas_transit', 'months_cover_total',
        ))

    def _turnover_compute_order_turnover_map(self, conn, order_product_ids, shop_ids=None):
        """下单 SKU -> 动销月三列（全渠道库存含全部替代方案；分母含本体与替代 SKU 折算销量）。"""
        ids = sorted({int(x) for x in (order_product_ids or []) if self._parse_int(x)})
        window = self._turnover_sales_window(conn)
        empty = {
            'turnover_sales_window': window,
            'months_cover_overseas': None,
            'months_cover_overseas_transit': None,
            'months_cover_total': None,
            'window_sales_qty': None,
        }
        out = {i: dict(empty) for i in ids}
        if not ids:
            return out
        effective_sales_map = self._turnover_load_op_effective_window_sales_map(
            conn, ids, window=window, shop_ids=shop_ids,
        )
        substitute_plans = self._forecast_load_all_substitute_plans_by_owner(conn, ids)
        extra_subs = []
        for oid in ids:
            for sid in self._forecast_substitute_ids_from_plan_groups(substitute_plans.get(oid) or []):
                extra_subs.append(sid)
        load_ids = sorted(set(ids).union(int(x) for x in extra_subs if self._parse_int(x)))
        brief_by_op = self._forecast_load_order_product_brief_map(conn, load_ids)
        inv_by_op = self._forecast_load_inventory_by_order_product(conn, load_ids)
        for op_id in ids:
            brief = brief_by_op.get(op_id) or {}
            is_on = self._parse_int(brief.get('is_on_market')) or 0
            plans = substitute_plans.get(op_id) or []
            agg, _ = self._forecast_order_row_inventory_aggregate(
                inv_by_op, op_id, is_on, plans, brief_by_op, plans,
            )
            sales = float(effective_sales_map.get(op_id) or 0)
            o = int(agg.get('overseas_qty') or 0)
            t = int(agg.get('transit_qty') or 0)
            st = int(agg.get('factory_stock_qty') or 0)
            wp = int(agg.get('wip_qty') or 0)
            out[op_id] = {
                'turnover_sales_window': window,
                'months_cover_overseas': self._turnover_coverage(o, sales),
                'months_cover_overseas_transit': self._turnover_coverage(o + t, sales),
                'months_cover_total': self._turnover_coverage(o + t + st + wp, sales),
                'window_sales_qty': round(sales, 2) if sales > 1e-9 else None,
            }
        return out

    def _turnover_apply_map_to_row(self, item, turnover_map):
        op_id = self._parse_int(item.get('order_product_id'))
        t = turnover_map.get(int(op_id)) if op_id else {}
        if not t:
            t = {}
        item['turnover_sales_window'] = t.get('turnover_sales_window')
        item['months_cover_overseas'] = t.get('months_cover_overseas')
        item['months_cover_overseas_transit'] = t.get('months_cover_overseas_transit')
        item['months_cover_total'] = t.get('months_cover_total')

    def _turnover_resolve_dashboard_row_sales(self, item, sales_map):
        """合并展示行：父行库存含迭代子 SKU 时，分母累加同族各下单 SKU 窗口销量。"""
        op_id = self._parse_int(item.get('order_product_id'))
        total = 0.0
        if op_id is not None:
            total += float(sales_map.get(op_id) or 0)
        for ch in item.get('iteration_children') or []:
            cid = self._parse_int(ch.get('order_product_id'))
            if cid is not None:
                total += float(sales_map.get(cid) or 0)
        return total if total > 1e-9 else None

    def _turnover_fill_warehouse_dashboard_row(self, item, sales_map, window, inv_override=None):
        sales = self._turnover_resolve_dashboard_row_sales(item, sales_map)
        if inv_override:
            o, t, st, wp = inv_override
        else:
            o = int(item.get('available_total') or 0)
            t = int(item.get('in_transit_total') or 0)
            st = int(item.get('factory_stock_total') or 0)
            wp = int(item.get('factory_wip_total') or 0)
        item['turnover_sales_window'] = window
        item['months_cover_overseas'] = self._turnover_coverage(o, sales)
        item['months_cover_overseas_transit'] = self._turnover_coverage(o + t, sales)
        item['months_cover_total'] = self._turnover_coverage(o + t + st + wp, sales)

    def _turnover_compute_variant_turnover_map(self, conn, variant_ids, shop_ids=None):
        """variant_id -> 动销月三列 + turnover_sales_window（与销量预测规格行逻辑一致）。"""
        ids = sorted({int(x) for x in (variant_ids or []) if self._parse_int(x)})
        window = self._turnover_sales_window(conn)
        empty = {
            'turnover_sales_window': window,
            'months_cover_overseas': None,
            'months_cover_overseas_transit': None,
            'months_cover_total': None,
        }
        out = {i: dict(empty) for i in ids}
        if not ids:
            return out
        variant_sales = self._turnover_load_variant_window_sales_map(conn, ids, window=window, shop_ids=shop_ids)
        tier_map = self._load_variant_tier_sellable_map(conn, ids)
        links_by_vid = self._forecast_load_variant_order_links_for_inventory(conn, ids)
        all_ops = sorted({
            int(o)
            for pairs in links_by_vid.values()
            for o, _q in (pairs or [])
            if self._parse_int(o)
        })

        def _fill_from_variant_sales(vid):
            vs = float(variant_sales.get(vid) or 0)
            if vs <= 1e-9:
                return None
            tiers = tier_map.get(vid) or {}
            o = int(tiers.get('overseas') or 0)
            t = int(tiers.get('transit') or 0)
            st = int(tiers.get('factory_stock') or 0)
            wp = int(tiers.get('wip') or 0)
            return {
                'turnover_sales_window': window,
                'months_cover_overseas': self._turnover_coverage(o, vs),
                'months_cover_overseas_transit': self._turnover_coverage(o + t, vs),
                'months_cover_total': self._turnover_coverage(o + t + st + wp, vs),
            }

        if not all_ops:
            for vid in ids:
                filled = _fill_from_variant_sales(vid)
                if filled:
                    out[vid] = filled
                else:
                    out[vid]['turnover_sales_window'] = window
            return out
        substitute_plans = self._forecast_load_all_substitute_plans_by_owner(conn, all_ops)
        extra_subs = []
        for oid in all_ops:
            for sid in self._forecast_substitute_ids_from_plan_groups(substitute_plans.get(oid) or []):
                extra_subs.append(sid)
        load_ids = sorted(set(all_ops).union(int(x) for x in extra_subs if self._parse_int(x)))
        inv_by_op = self._forecast_load_inventory_by_order_product(conn, load_ids)
        op_sales = self._turnover_load_op_window_sales_map(conn, load_ids, window=window, shop_ids=shop_ids)
        for vid in ids:
            pairs = links_by_vid.get(vid) or []
            cov = self._turnover_spec_coverage_from_bom_pairs(
                pairs, op_sales, inv_by_op, substitute_plans,
            )
            if self._turnover_cov_all_none(cov) and pairs:
                cov = self._turnover_spec_coverage_from_bom_pairs_raw(pairs, op_sales, inv_by_op)
            if self._turnover_cov_all_none(cov):
                filled = _fill_from_variant_sales(vid)
                if filled:
                    out[vid] = filled
                    continue
            out[vid] = {
                'turnover_sales_window': window,
                'months_cover_overseas': cov.get('months_cover_overseas'),
                'months_cover_overseas_transit': cov.get('months_cover_overseas_transit'),
                'months_cover_total': cov.get('months_cover_total'),
            }
        return out

    def _turnover_attach_to_sales_product_rows(self, conn, rows, shop_ids=None):
        variant_ids = sorted({int(r.get('variant_id') or 0) for r in (rows or []) if int(r.get('variant_id') or 0) > 0})
        try:
            turnover_map = self._turnover_compute_variant_turnover_map(conn, variant_ids, shop_ids=shop_ids)
        except Exception as e:
            print(f'Sales product turnover attach error: {type(e).__name__}: {e!r}')
            window = None
            try:
                window = self._turnover_sales_window(conn)
            except Exception:
                pass
            turnover_map = {
                vid: {
                    'turnover_sales_window': window,
                    'months_cover_overseas': None,
                    'months_cover_overseas_transit': None,
                    'months_cover_total': None,
                }
                for vid in variant_ids
            }
        for row in rows or []:
            vid = int(row.get('variant_id') or 0)
            t = turnover_map.get(vid) or {}
            row['turnover_sales_window'] = t.get('turnover_sales_window')
            row['months_cover_overseas'] = t.get('months_cover_overseas')
            row['months_cover_overseas_transit'] = t.get('months_cover_overseas_transit')
            row['months_cover_total'] = t.get('months_cover_total')

    def _turnover_attach_to_warehouse_dashboard_rows(self, conn, rows, shop_ids=None):
        op_ids = []
        for item in rows or []:
            oid = self._parse_int(item.get('order_product_id'))
            if oid:
                op_ids.append(oid)
            for child in item.get('iteration_children') or []:
                cid = self._parse_int(child.get('order_product_id'))
                if cid:
                    op_ids.append(cid)
        window = self._turnover_sales_window(conn)
        sales_map = self._turnover_load_op_window_sales_map(conn, op_ids, window=window, shop_ids=shop_ids)
        for item in rows or []:
            self._turnover_fill_warehouse_dashboard_row(item, sales_map, window)
            for child in item.get('iteration_children') or []:
                self._turnover_fill_warehouse_dashboard_row(child, sales_map, window)

    def _turnover_attach_to_warehouse_dashboard_rows_safe(self, conn, rows, shop_ids=None):
        """仓储看板动销月：失败时不阻断整页加载，返回 None 或错误信息。"""
        try:
            self._turnover_attach_to_warehouse_dashboard_rows(conn, rows, shop_ids=shop_ids)
            return None
        except Exception as e:
            err = f'{type(e).__name__}: {e}'
            print(f'Warehouse dashboard turnover attach error: {err!r}')
            window = None
            try:
                window = self._turnover_sales_window(conn)
            except Exception:
                pass
            for item in rows or []:
                item.setdefault('turnover_sales_window', window)
                item.setdefault('months_cover_overseas', None)
                item.setdefault('months_cover_overseas_transit', None)
                item.setdefault('months_cover_total', None)
                for child in item.get('iteration_children') or []:
                    child.setdefault('turnover_sales_window', window)
                    child.setdefault('months_cover_overseas', None)
                    child.setdefault('months_cover_overseas_transit', None)
                    child.setdefault('months_cover_total', None)
            return err

    def _forecast_grow_linked_orders_variants_bidir(self, conn, touch_orders, touch_variants):
        """沿 sales_variant_order_links 双向扩闭包：与当前订单/变体通过任一边相连的 order_product_id、variant_id。

        用于下单 SKU 继承预测：下市本体 B 仍挂在某规格上时，需把该规格预估值经替代方案折算到在市替代 SKU（如 A×2）。"""
        orders = {int(x) for x in (touch_orders or []) if self._parse_int(x)}
        variants = {int(x) for x in (touch_variants or []) if self._parse_int(x)}
        if not orders and not variants:
            return orders, variants
        max_pass = 16
        with conn.cursor() as cur:
            for _ in range(max_pass):
                grew = False
                if variants:
                    ph = ','.join(['%s'] * len(variants))
                    cur.execute(
                        f"""
                        SELECT DISTINCT order_product_id, variant_id
                        FROM sales_variant_order_links
                        WHERE variant_id IN ({ph})
                        """,
                        tuple(sorted(variants)),
                    )
                    for rr in cur.fetchall() or []:
                        oid = self._parse_int(rr.get('order_product_id'))
                        vid = self._parse_int(rr.get('variant_id'))
                        if oid and oid not in orders:
                            orders.add(oid)
                            grew = True
                        if vid and vid not in variants:
                            variants.add(vid)
                            grew = True
                if orders:
                    ph = ','.join(['%s'] * len(orders))
                    cur.execute(
                        f"""
                        SELECT DISTINCT order_product_id, variant_id
                        FROM sales_variant_order_links
                        WHERE order_product_id IN ({ph})
                        """,
                        tuple(sorted(orders)),
                    )
                    for rr in cur.fetchall() or []:
                        oid = self._parse_int(rr.get('order_product_id'))
                        vid = self._parse_int(rr.get('variant_id'))
                        if oid and oid not in orders:
                            orders.add(oid)
                            grew = True
                        if vid and vid not in variants:
                            variants.add(vid)
                            grew = True
                if not grew:
                    break
        return orders, variants

    def _forecast_load_order_variant_links_by_order_product_ids(self, conn, order_product_ids):
        """order_product_id -> [{ variant_id, quantity }, ...]（与下单 SKU 行 variant_links 同源）。"""
        ids = sorted({int(x) for x in (order_product_ids or []) if self._parse_int(x)})
        out = {}
        if not ids:
            return out
        ph = ','.join(['%s'] * len(ids))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT l.order_product_id,
                       l.variant_id,
                       GREATEST(1, COALESCE(l.quantity, 1)) AS quantity
                FROM sales_variant_order_links l
                WHERE l.order_product_id IN ({ph})
                ORDER BY l.order_product_id ASC, l.variant_id ASC
                """,
                tuple(ids),
            )
            for rr in cur.fetchall() or []:
                op = self._parse_int(rr.get('order_product_id'))
                vid = self._parse_int(rr.get('variant_id'))
                qp = max(1, self._parse_int(rr.get('quantity')) or 1)
                if not op or not vid:
                    continue
                out.setdefault(op, []).append({'variant_id': vid, 'quantity': qp})
        return out

    def _forecast_variant_hist_weight_sum_for_variants(self, conn, variant_ids):
        """各 variant 在 sales_variant_order_links 上的全局权重分母（与下单 SKU 行历史分摊一致）。"""
        ids = sorted({int(x) for x in (variant_ids or []) if self._parse_int(x)})
        out = {}
        if not ids:
            return out
        ph = ','.join(['%s'] * len(ids))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT variant_id,
                       SUM(GREATEST(1, COALESCE(quantity, 1))) AS w
                FROM sales_variant_order_links
                WHERE variant_id IN ({ph})
                GROUP BY variant_id
                """,
                tuple(ids),
            )
            for rr in (cur.fetchall() or []):
                vid = self._parse_int(rr.get('variant_id'))
                if vid:
                    out[vid] = float(rr.get('w') or 0)
        return out

    def _forecast_build_order_like_history_for_links(self, links, history_by_variant, variant_hist_weight_sum, hist_month_keys):
        """与 _forecast_build_order_rows 中下单 SKU 行 history 相同的按月聚合（用于动销月分母）。"""
        empty_hist = {'sales_qty': 0, 'net_sales_amount': 0, 'order_qty': 0, 'session_total': 0, 'refund_amount': 0}
        history = {}
        for hm in hist_month_keys or []:
            agg = {k: 0.0 for k in empty_hist}
            for vl in links or []:
                vid = self._parse_int(vl.get('variant_id'))
                if not vid:
                    continue
                h = history_by_variant.get((vid, hm))
                if not h:
                    continue
                q = float(max(1, self._parse_int(vl.get('quantity')) or 1))
                agg['sales_qty'] += float(h.get('sales_qty') or 0) * q
                denom = float(variant_hist_weight_sum.get(vid) or 0)
                if denom <= 0:
                    continue
                share = q / denom
                agg['net_sales_amount'] += float(h.get('net_sales_amount') or 0) * share
                agg['order_qty'] += float(h.get('order_qty') or 0) * share
                agg['session_total'] += float(h.get('session_total') or 0) * share
                agg['refund_amount'] += float(h.get('refund_amount') or 0) * share
            history[hm] = {
                'sales_qty': agg['sales_qty'],
                'net_sales_amount': agg['net_sales_amount'],
                'order_qty': agg['order_qty'],
                'session_total': agg['session_total'],
                'refund_amount': agg['refund_amount'],
            }
        return history

    def _forecast_load_variant_display_map(self, conn, variant_ids):
        """variant_id -> { spec_name, fabric_code, sku_family }"""
        ids = sorted({int(x) for x in (variant_ids or []) if self._parse_int(x)})
        out = {}
        if not ids:
            return out
        has_fabric_text = self._table_has_column(conn, 'sales_product_variants', 'fabric')
        fabric_join = 'LEFT JOIN fabric_materials fm ON fm.id = v.fabric_id' if self._table_has_column(conn, 'sales_product_variants', 'fabric_id') else ''
        if self._table_has_column(conn, 'sales_product_variants', 'fabric_id'):
            fabric_expr = 'COALESCE(fm.fabric_code, v.fabric)' if has_fabric_text else 'fm.fabric_code'
        elif has_fabric_text:
            fabric_expr = 'v.fabric'
        else:
            fabric_expr = "''"
        ph = ','.join(['%s'] * len(ids))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT v.id AS variant_id,
                       v.spec_name,
                       {fabric_expr} AS fabric_code,
                       pf.sku_family
                FROM sales_product_variants v
                LEFT JOIN product_families pf ON pf.id = v.sku_family_id
                {fabric_join}
                WHERE v.id IN ({ph})
                """,
                tuple(ids),
            )
            for rr in cur.fetchall() or []:
                vid = self._parse_int(rr.get('variant_id'))
                if not vid:
                    continue
                out[vid] = {
                    'spec_name': (rr.get('spec_name') or '').strip(),
                    'fabric_code': (rr.get('fabric_code') or '').strip(),
                    'sku_family': (rr.get('sku_family') or '').strip(),
                }
        return out

    def _forecast_build_order_history_sources_for_links(
        self, links, month, history_by_variant, history_by_sp, platforms_by_variant, variant_display,
    ):
        """下单 SKU 单月历史销量来源：按销售平台 SKU / 规格+面料 拆解（计入 = 月销 × BOM 链接数量）。"""
        by_platform = []
        spec_agg = {}

        for vl in links or []:
            vid = self._parse_int(vl.get('variant_id'))
            if not vid:
                continue
            q_link = max(1, self._parse_int(vl.get('quantity')) or 1)
            vd = variant_display.get(vid) or {}
            spec = (vl.get('spec_name') or vd.get('spec_name') or '').strip() or '—'
            fabric = (vd.get('fabric_code') or '').strip() or '—'
            sf = (vl.get('sku_family') or vd.get('sku_family') or '').strip() or '—'

            h_var = history_by_variant.get((vid, month)) or {}
            var_sales = float(h_var.get('sales_qty') or 0)
            contrib_spec = var_sales * float(q_link)

            sk = (spec, fabric)
            bucket = spec_agg.setdefault(sk, {
                'spec_name': spec,
                'fabric_code': fabric,
                'sku_family': sf,
                'link_qty': 0,
                'variant_sales_qty': var_sales,
                'contributed_sales_qty': 0.0,
            })
            bucket['link_qty'] += int(q_link)
            bucket['variant_sales_qty'] = var_sales
            bucket['contributed_sales_qty'] += contrib_spec

            for ps in platforms_by_variant.get(vid) or []:
                spid = self._parse_int(ps.get('sales_product_id'))
                if not spid:
                    continue
                h_sp = history_by_sp.get((spid, month)) or {}
                sp_sales = float(h_sp.get('sales_qty') or 0)
                contrib = sp_sales * float(q_link)
                if contrib <= 0 and sp_sales <= 0:
                    continue
                by_platform.append({
                    'platform_sku': (ps.get('platform_sku') or '').strip() or '—',
                    'shop_name': (ps.get('shop_name') or '').strip() or '—',
                    'spec_name': spec,
                    'fabric_code': fabric,
                    'link_qty': int(q_link),
                    'platform_sales_qty': sp_sales,
                    'contributed_sales_qty': contrib,
                })

        by_spec = sorted(
            spec_agg.values(),
            key=lambda x: (-float(x.get('contributed_sales_qty') or 0), str(x.get('spec_name') or '')),
        )
        by_platform.sort(
            key=lambda x: (-float(x.get('contributed_sales_qty') or 0), str(x.get('platform_sku') or '')),
        )
        total = sum(float(x.get('contributed_sales_qty') or 0) for x in by_spec)
        return {
            'month': month,
            'total_sales_qty': total,
            'by_platform': by_platform,
            'by_spec_fabric': by_spec,
        }

    def _forecast_ensure_row_history_for_inventory(self, conn, rows, forecast_mode, hist_start, hist_end, shop_ids=None):
        """懒加载库存 patch 等场景下行上 history 为空，需补齐历史月销量供动销月分母计算。"""
        if not rows or not hist_start or not hist_end:
            return
        hist_month_keys = list(self._forecast_iter_months(hist_start, hist_end))
        if not hist_month_keys:
            return
        hist_empty = self._forecast_empty_history_perf_payload()
        need_rows = [r for r in rows if not (r.get('history') or {})]
        if not need_rows:
            return

        if forecast_mode == 'order':
            links_by_op = {}
            all_vids = set()
            for r in need_rows:
                oid = self._parse_int((r.get('labels') or {}).get('order_product_id'))
                if not oid:
                    continue
                links = (r.get('labels') or {}).get('variant_links') or []
                if not links:
                    links = self._forecast_load_order_variant_links_by_order_product_ids(conn, [oid]).get(oid) or []
                links_by_op[oid] = links
                for vl in links or []:
                    vid = self._parse_int(vl.get('variant_id'))
                    if vid:
                        all_vids.add(vid)
            if not links_by_op:
                return
            vlist = sorted(all_vids)
            history_by_variant = (
                self._forecast_load_history_by_variant(conn, vlist, hist_start, hist_end, shop_ids=shop_ids)
                if vlist else {}
            )
            wsum = self._forecast_variant_hist_weight_sum_for_variants(conn, vlist)
            for r in need_rows:
                oid = self._parse_int((r.get('labels') or {}).get('order_product_id'))
                if not oid:
                    continue
                links = links_by_op.get(oid) or []
                r['history'] = self._forecast_build_order_like_history_for_links(
                    links, history_by_variant, wsum, hist_month_keys
                )
            return

        if forecast_mode == 'platform':
            sp_ids = sorted({
                self._parse_int((r.get('labels') or {}).get('sales_product_id'))
                for r in need_rows
                if self._parse_int((r.get('labels') or {}).get('sales_product_id'))
            })
            if not sp_ids:
                return
            history = self._forecast_load_history_by_sales_product(
                conn, sp_ids, hist_start, hist_end, shop_ids=shop_ids
            )
            for r in need_rows:
                spid = self._parse_int((r.get('labels') or {}).get('sales_product_id'))
                if not spid:
                    continue
                row_hist = {}
                for hm in hist_month_keys:
                    h = history.get((spid, hm))
                    row_hist[hm] = h if h else hist_empty.copy()
                r['history'] = row_hist
            return

        vids = sorted({
            self._parse_int((r.get('labels') or {}).get('variant_id'))
            for r in need_rows
            if self._parse_int((r.get('labels') or {}).get('variant_id'))
        })
        if not vids:
            return
        history = self._forecast_load_history_by_variant(
            conn, vids, hist_start, hist_end, shop_ids=shop_ids
        )
        for r in need_rows:
            vid = self._parse_int((r.get('labels') or {}).get('variant_id'))
            if not vid:
                continue
            row_hist = {}
            for hm in hist_month_keys:
                h = history.get((vid, hm))
                row_hist[hm] = h if h else hist_empty.copy()
            r['history'] = row_hist

    def _forecast_attach_inventory_to_rows(self, conn, rows, forecast_mode, history_months, hist_start=None, hist_end=None, shop_ids=None):
        """按各下单 SKU：本体库存 + 全部替代发货方案替代 SKU 库存（海外/在途/在库/在制），再与变体 BOM 成套汇总。
        规格/平台维度下：库存为变体整套瓶颈；动销月三列 = 所链接各下单 SKU（按替代方案展开后单独成套）的动销月再取最小值；
        每个下单 SKU 的动销月分母为其自身窗口销量（全局最新 record_date 向前 30 天，含首尾）。"""
        if not rows:
            return
        window = self._turnover_sales_window(conn)
        hm_list = list(history_months or [])
        if hist_start is None and hm_list:
            hist_start = hm_list[0]
        if hist_end is None and hm_list:
            hist_end = hm_list[-1]
        self._forecast_ensure_row_history_for_inventory(
            conn, rows, forecast_mode, hist_start, hist_end, shop_ids=shop_ids,
        )
        inv_by_op = {}
        variant_bom_links = {}
        variant_expanded = {}
        plan_items_by_owner = {}
        substitute_plans_by_owner = {}

        brief_by_op = {}
        if forecast_mode == 'order':
            op_ids = [self._parse_int((r.get('labels') or {}).get('order_product_id')) for r in rows]
            op_ids = [x for x in op_ids if x]
            substitute_plans_by_owner = self._forecast_load_all_substitute_plans_by_owner(conn, op_ids)
            extra_subs = []
            for oid in op_ids:
                for sid in self._forecast_substitute_ids_from_plan_groups(
                    substitute_plans_by_owner.get(oid) or []
                ):
                    extra_subs.append(sid)
            load_ids = sorted(set(op_ids).union(extra_subs))
            brief_by_op = self._forecast_load_order_product_brief_map(conn, load_ids)
            inv_by_op = self._forecast_load_inventory_by_order_product(conn, load_ids)
        else:
            variant_ids = sorted({
                self._parse_int((r.get('labels') or {}).get('variant_id'))
                for r in rows
                if self._parse_int((r.get('labels') or {}).get('variant_id'))
            })
            variant_bom_links = self._forecast_load_variant_order_links_for_inventory(conn, variant_ids)
            all_owners = []
            for pairs in variant_bom_links.values():
                for oid, _qp in pairs:
                    if oid:
                        all_owners.append(oid)
            substitute_plans_by_owner = self._forecast_load_all_substitute_plans_by_owner(conn, all_owners)
            plan_items_by_owner = {
                oid: self._forecast_merge_substitute_items_all_plans(substitute_plans_by_owner.get(oid) or [])
                for oid in sorted(set(all_owners))
            }
            extra_subs = []
            for oid in set(all_owners):
                for sid in self._forecast_substitute_ids_from_plan_groups(
                    substitute_plans_by_owner.get(oid) or []
                ):
                    extra_subs.append(sid)
            for vid, pairs in variant_bom_links.items():
                variant_expanded[vid] = self._forecast_expand_owner_pairs_with_substitute_plans(pairs, plan_items_by_owner)
            all_ops = []
            for pairs in variant_bom_links.values():
                for o, _q in pairs:
                    if o:
                        all_ops.append(int(o))
            load_spec_ids = sorted(set(all_ops).union(extra_subs))
            brief_by_op = self._forecast_load_order_product_brief_map(conn, load_spec_ids)
            inv_by_op = self._forecast_load_inventory_by_order_product(conn, load_spec_ids)

        op_window_sales_by_op = {}
        variant_window_sales = {}
        sp_window_sales = {}
        if window:
            if forecast_mode == 'order':
                order_op_ids = sorted({
                    self._parse_int((r.get('labels') or {}).get('order_product_id'))
                    for r in rows
                    if self._parse_int((r.get('labels') or {}).get('order_product_id'))
                })
                if order_op_ids:
                    op_window_sales_by_op = self._turnover_load_op_window_sales_map(
                        conn, order_op_ids, window=window, shop_ids=shop_ids,
                    )
            elif forecast_mode == 'platform':
                sp_ids = sorted({
                    self._parse_int((r.get('labels') or {}).get('sales_product_id'))
                    for r in rows
                    if self._parse_int((r.get('labels') or {}).get('sales_product_id'))
                })
                if sp_ids:
                    sp_window_sales = self._turnover_load_sales_product_window_sales_map(
                        conn, sp_ids, window=window, shop_ids=shop_ids,
                    )
            elif variant_expanded:
                all_op_ids_flat = sorted({
                    int(o)
                    for pairs in variant_expanded.values()
                    for o, _q in pairs
                    if self._parse_int(o)
                })
                if all_op_ids_flat:
                    op_window_sales_by_op = self._turnover_load_op_window_sales_map(
                        conn, all_op_ids_flat, window=window, shop_ids=shop_ids,
                    )
                spec_vids = sorted({
                    self._parse_int((r.get('labels') or {}).get('variant_id'))
                    for r in rows
                    if self._parse_int((r.get('labels') or {}).get('variant_id'))
                })
                if spec_vids:
                    variant_window_sales = self._turnover_load_variant_window_sales_map(
                        conn, spec_vids, window=window, shop_ids=shop_ids,
                    )

        for row in rows:
            labels = row.get('labels') or {}
            agg = self._forecast_inventory_zero()
            if forecast_mode == 'order':
                oid = self._parse_int(labels.get('order_product_id'))
                comp = {}
                if oid:
                    is_on = self._parse_int(labels.get('is_on_market')) or 0
                    agg, comp = self._forecast_order_row_inventory_aggregate(
                        inv_by_op,
                        oid,
                        is_on,
                        substitute_plans_by_owner.get(oid) or [],
                        brief_by_op,
                        substitute_plans_by_owner.get(oid) or [],
                    )
                row['inventory_composition'] = comp
            else:
                vid = self._parse_int(labels.get('variant_id'))
                pairs_vid = variant_bom_links.get(vid) or []
                if pairs_vid:
                    per_link = []
                    for oid, qp in pairs_vid:
                        oid = int(oid)
                        qp = max(1, int(qp or 1))
                        if not oid:
                            continue
                        bref = brief_by_op.get(oid) or {}
                        is_on_op = self._parse_int(bref.get('is_on_market')) or 0
                        owner_plans = substitute_plans_by_owner.get(oid) or []
                        ai = self._forecast_inventory_assembled_from_substitute_plans(inv_by_op, owner_plans)
                        per_link.append({
                            k: int(int(ai.get(k) or 0) // qp)
                            for k in ('overseas_qty', 'transit_qty', 'factory_stock_qty', 'wip_qty')
                        })
                    if per_link:
                        agg = {
                            k: min(int(p.get(k) or 0) for p in per_link)
                            for k in ('overseas_qty', 'transit_qty', 'factory_stock_qty', 'wip_qty')
                        }
                else:
                    bom = variant_expanded.get(vid) or []
                    agg = self._forecast_bom_tier_assembled_counts(inv_by_op, bom)

            pairs = []
            if forecast_mode != 'order':
                pairs = variant_bom_links.get(self._parse_int(labels.get('variant_id'))) or []

            o = int(agg['overseas_qty'])
            t = int(agg['transit_qty'])
            st = int(agg['factory_stock_qty'])
            wp = int(agg['wip_qty'])
            vid = self._parse_int(labels.get('variant_id'))
            spid = self._parse_int(labels.get('sales_product_id'))
            if forecast_mode == 'platform' and spid:
                avg_sales = float(sp_window_sales.get(spid) or 0)
            elif vid:
                avg_sales = float(variant_window_sales.get(vid) or 0)
            else:
                avg_sales = 0.0

            def _cov(num, denom):
                d = float(denom or 0)
                if d <= 1e-9 or int(num) <= 0:
                    return None
                return round(float(num) / d, 2)

            if forecast_mode == 'order':
                oid = self._parse_int(labels.get('order_product_id'))
                order_sales = float(op_window_sales_by_op.get(oid) or 0) if oid else 0.0
                mo = _cov(o, order_sales)
                mot = _cov(o + t, order_sales)
                mtot = _cov(o + t + st + wp, order_sales)
                effective_sales = order_sales
                turnover_bom_breakdown = []
            elif pairs:
                vals_o, vals_ot, vals_tot = [], [], []
                turnover_bom_breakdown = []
                for oid, qp in pairs:
                    oid = int(oid)
                    qp = max(1, int(qp or 1))
                    if not oid:
                        continue
                    avg_op = op_window_sales_by_op.get(oid)
                    if avg_op is None or float(avg_op) <= 1e-9:
                        if avg_sales and len(pairs) == 1 and qp == 1:
                            avg_op = avg_sales
                        else:
                            continue
                    avg_op = float(avg_op)
                    bref = brief_by_op.get(oid) or {}
                    is_on_op = self._parse_int(bref.get('is_on_market')) or 0
                    owner_plans = substitute_plans_by_owner.get(oid) or []
                    ai = self._forecast_inventory_assembled_from_substitute_plans(inv_by_op, owner_plans)
                    inv_raw = inv_by_op.get(oid) or {}
                    ai_sum = sum(int(ai.get(k) or 0) for k in ('overseas_qty', 'transit_qty', 'factory_stock_qty', 'wip_qty'))
                    raw_sum = sum(int(inv_raw.get(k) or 0) for k in ('overseas_qty', 'transit_qty', 'factory_stock_qty', 'wip_qty'))
                    if ai_sum <= 0 and raw_sum > 0:
                        ai = {
                            'overseas_qty': int(inv_raw.get('overseas_qty') or 0),
                            'transit_qty': int(inv_raw.get('transit_qty') or 0),
                            'factory_stock_qty': int(inv_raw.get('factory_stock_qty') or 0),
                            'wip_qty': int(inv_raw.get('wip_qty') or 0),
                        }
                    for tier_k in ai:
                        ai[tier_k] = int(int(ai[tier_k] or 0) // max(1, qp))
                    oi = int(ai['overseas_qty'])
                    ti = int(ai['transit_qty'])
                    sti = int(ai['factory_stock_qty'])
                    wpi = int(ai['wip_qty'])
                    if oi > 0:
                        vals_o.append(round(float(oi) / float(avg_op), 2))
                    if oi + ti > 0:
                        vals_ot.append(round(float(oi + ti) / float(avg_op), 2))
                    if oi + ti + sti + wpi > 0:
                        vals_tot.append(round(float(oi + ti + sti + wpi) / float(avg_op), 2))
                    turnover_bom_breakdown.append({
                        'order_product_id': oid,
                        'sku': (brief.get('sku') or '').strip(),
                        'quantity': qp,
                        'overseas_qty': oi,
                        'transit_qty': ti,
                        'factory_stock_qty': sti,
                        'wip_qty': wpi,
                        'window_sales_qty': round(float(avg_op), 2),
                        'months_cover_overseas': round(float(oi) / float(avg_op), 2) if oi > 0 else None,
                        'months_cover_overseas_transit': round(float(oi + ti) / float(avg_op), 2) if oi + ti > 0 else None,
                        'months_cover_total': round(float(oi + ti + sti + wpi) / float(avg_op), 2) if oi + ti + sti + wpi > 0 else None,
                    })
                mo = min(vals_o) if vals_o else None
                mot = min(vals_ot) if vals_ot else None
                mtot = min(vals_tot) if vals_tot else None
                effective_sales = float(avg_sales or 0)
                if mo is None and mot is None and mtot is None and avg_sales and float(avg_sales) > 1e-9:
                    mo = _cov(o, avg_sales)
                    mot = _cov(o + t, avg_sales)
                    mtot = _cov(o + t + st + wp, avg_sales)
            else:
                mo = _cov(o, avg_sales)
                mot = _cov(o + t, avg_sales)
                mtot = _cov(o + t + st + wp, avg_sales)
                effective_sales = float(avg_sales or 0)
                turnover_bom_breakdown = []

            row['inventory'] = {
                'months_cover_overseas': mo,
                'months_cover_overseas_transit': mot,
                'months_cover_total': mtot,
                'overseas_qty': o,
                'transit_qty': t,
                'in_stock_qty': st,
                'wip_qty': wp,
                'turnover_sales_window': window,
                'turnover_window_sales_qty': effective_sales if effective_sales > 1e-9 else None,
                'turnover_bom_breakdown': turnover_bom_breakdown,
            }
            if forecast_mode != 'order':
                pairs_comp = [
                    (int(o), max(1, int(q or 1)))
                    for o, q in (pairs or [])
                    if self._parse_int(o)
                ]
                row['inventory_composition'] = {
                    'pairs': pairs_comp,
                    'row_sku': (labels.get('sku') or labels.get('platform_sku') or '').strip(),
                    'has_indicator': len(pairs_comp) > 1,
                }

    def _forecast_date_max_iso(self, a, b):
        a = self._forecast_format_date_only(a)
        b = self._forecast_format_date_only(b)
        if not a:
            return b
        if not b:
            return a
        return a if a >= b else b

    def _forecast_empty_surplus_detail(self):
        return {
            'overseas_by_region': {},
            'transit_batches': [],
            'factory_stock_qty': 0,
            'wip_batches': [],
            'overseas_updated_at': None,
            'transit_updated_at': None,
            'factory_stock_updated_at': None,
            'wip_updated_at': None,
        }

    def _forecast_list_destination_regions(self, conn):
        out = []
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, region_name, sort_order FROM logistics_destination_regions ORDER BY sort_order ASC, id ASC"
            )
            for rr in cur.fetchall() or []:
                rid = self._parse_int(rr.get('id'))
                name = str(rr.get('region_name') or '').strip()
                if rid and name:
                    out.append({
                        'id': rid,
                        'region_name': name,
                        'sort_order': self._parse_int(rr.get('sort_order')) or 0,
                    })
        return out

    def _forecast_load_inventory_surplus_detail_by_order_product(self, conn, order_product_ids):
        """下单 SKU 盈余估算：按目的区域拆分海外仓/在途批次，及工厂在库/在制明细。"""
        ids = sorted({int(x) for x in (order_product_ids or []) if self._parse_int(x)})
        out = {i: self._forecast_empty_surplus_detail() for i in ids}
        if not ids:
            return out
        ph = ','.join(['%s'] * len(ids))
        tpl = tuple(ids)
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT oi.order_product_id,
                       COALESCE(NULLIF(TRIM(dr.region_name),''), NULLIF(TRIM(w.region),''), '-') AS region_name,
                       COALESCE(SUM(oi.available_qty), 0) AS q
                FROM logistics_overseas_inventory oi
                INNER JOIN logistics_overseas_warehouses w ON w.id = oi.warehouse_id
                LEFT JOIN logistics_destination_regions dr ON dr.id = w.destination_region_id
                WHERE oi.order_product_id IN ({ph})
                  AND COALESCE(w.is_enabled, 1) = 1
                GROUP BY oi.order_product_id,
                         COALESCE(NULLIF(TRIM(dr.region_name),''), NULLIF(TRIM(w.region),''), '-')
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                rname = str(rr.get('region_name') or '-').strip() or '-'
                if oid in out:
                    out[oid]['overseas_by_region'][rname] = int(float(rr.get('q') or 0))

            cur.execute(
                f"""
                SELECT li.order_product_id,
                       COALESCE(NULLIF(TRIM(drt.region_name),''), NULLIF(TRIM(dr.region_name),''), NULLIF(TRIM(w.region),''), '-') AS region_name,
                       COALESCE(t.expected_listed_date_latest, t.expected_warehouse_date, t.eta_latest) AS expected_listed_date,
                       COALESCE(SUM(li.shipped_qty), 0) AS q
                FROM logistics_in_transit_items li
                INNER JOIN logistics_in_transit t ON t.id = li.transit_id
                LEFT JOIN logistics_overseas_warehouses w ON w.id = t.destination_warehouse_id
                LEFT JOIN logistics_destination_regions drt ON drt.id = t.destination_region_id
                LEFT JOIN logistics_destination_regions dr ON dr.id = w.destination_region_id
                WHERE li.order_product_id IN ({ph})
                  AND COALESCE(t.inventory_registered, 0) = 0
                GROUP BY li.order_product_id,
                         COALESCE(NULLIF(TRIM(drt.region_name),''), NULLIF(TRIM(dr.region_name),''), NULLIF(TRIM(w.region),''), '-'),
                         COALESCE(t.expected_listed_date_latest, t.expected_warehouse_date, t.eta_latest)
                HAVING COALESCE(SUM(li.shipped_qty), 0) > 0
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                if oid not in out:
                    continue
                qty = int(float(rr.get('q') or 0))
                if qty <= 0:
                    continue
                out[oid]['transit_batches'].append({
                    'region_name': str(rr.get('region_name') or '-').strip() or '-',
                    'expected_listed_date': self._forecast_format_date_only(rr.get('expected_listed_date')),
                    'qty': qty,
                })

            cur.execute(
                f"""
                SELECT order_product_id, COALESCE(SUM(quantity), 0) AS q
                FROM factory_stock_inventory
                WHERE order_product_id IN ({ph})
                GROUP BY order_product_id
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                if oid in out:
                    out[oid]['factory_stock_qty'] = int(float(rr.get('q') or 0))

            cur.execute(
                f"""
                SELECT order_product_id, quantity, expected_completion_date
                FROM factory_wip_inventory
                WHERE order_product_id IN ({ph}) AND COALESCE(is_completed, 0) = 0
                ORDER BY order_product_id ASC, expected_completion_date ASC, id ASC
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                qty = int(float(rr.get('quantity') or 0))
                if oid not in out or qty <= 0:
                    continue
                out[oid]['wip_batches'].append({
                    'qty': qty,
                    'expected_completion_date': self._forecast_format_date_only(rr.get('expected_completion_date')),
                })

            cur.execute(
                f"""
                SELECT order_product_id, MAX(updated_at) AS dt
                FROM logistics_overseas_inventory
                WHERE order_product_id IN ({ph})
                GROUP BY order_product_id
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                if oid in out:
                    out[oid]['overseas_updated_at'] = self._forecast_format_date_only(rr.get('dt'))

            cur.execute(
                f"""
                SELECT li.order_product_id, MAX(t.updated_at) AS dt
                FROM logistics_in_transit_items li
                INNER JOIN logistics_in_transit t ON t.id = li.transit_id
                WHERE li.order_product_id IN ({ph})
                  AND COALESCE(t.inventory_registered, 0) = 0
                GROUP BY li.order_product_id
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                if oid in out:
                    out[oid]['transit_updated_at'] = self._forecast_format_date_only(rr.get('dt'))

            cur.execute(
                f"""
                SELECT order_product_id, MAX(updated_at) AS dt
                FROM factory_stock_inventory
                WHERE order_product_id IN ({ph})
                GROUP BY order_product_id
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                if oid in out:
                    out[oid]['factory_stock_updated_at'] = self._forecast_format_date_only(rr.get('dt'))

            cur.execute(
                f"""
                SELECT order_product_id, MAX(updated_at) AS dt
                FROM factory_wip_inventory
                WHERE order_product_id IN ({ph}) AND COALESCE(is_completed, 0) = 0
                GROUP BY order_product_id
                """,
                tpl,
            )
            for rr in cur.fetchall() or []:
                oid = self._parse_int(rr.get('order_product_id'))
                if oid in out:
                    out[oid]['wip_updated_at'] = self._forecast_format_date_only(rr.get('dt'))
        return out

    def _forecast_merge_surplus_detail_for_plan_groups(self, plan_groups, detail_by_op):
        merged = self._forecast_empty_surplus_detail()
        for grp in plan_groups or []:
            links = [
                (int(self._parse_int(sid)), max(1, self._parse_int(mult) or 1))
                for sid, mult in (grp.get('items') or [])
                if self._parse_int(sid)
            ]
            if not links:
                continue
            part = self._forecast_merge_surplus_detail_for_pairs(links, detail_by_op)
            for rname, qty in (part.get('overseas_by_region') or {}).items():
                rname = str(rname or '').strip() or '-'
                merged['overseas_by_region'][rname] = int(merged['overseas_by_region'].get(rname) or 0) + int(qty or 0)
            merged['transit_batches'].extend(part.get('transit_batches') or [])
            merged['factory_stock_qty'] += int(part.get('factory_stock_qty') or 0)
            merged['wip_batches'].extend(part.get('wip_batches') or [])
            merged['overseas_updated_at'] = self._forecast_date_max_iso(
                merged.get('overseas_updated_at'), part.get('overseas_updated_at')
            )
            merged['transit_updated_at'] = self._forecast_date_max_iso(
                merged.get('transit_updated_at'), part.get('transit_updated_at')
            )
            merged['factory_stock_updated_at'] = self._forecast_date_max_iso(
                merged.get('factory_stock_updated_at'), part.get('factory_stock_updated_at')
            )
            merged['wip_updated_at'] = self._forecast_date_max_iso(
                merged.get('wip_updated_at'), part.get('wip_updated_at')
            )
        return merged

    def _forecast_surplus_detail_to_inv_scalars(self, detail):
        """将盈余明细折算为与库存列一致的各阶件数（用于方案内 BOM 成套）。"""
        d = detail or self._forecast_empty_surplus_detail()
        overseas_qty = 0
        for _k, v in (d.get('overseas_by_region') or {}).items():
            overseas_qty += max(0, int(float(v or 0)))
        return {
            'overseas_qty': overseas_qty,
            'transit_qty': sum(
                max(0, int(float(b.get('qty') or 0)))
                for b in (d.get('transit_batches') or [])
            ),
            'factory_stock_qty': max(0, int(float(d.get('factory_stock_qty') or 0))),
            'wip_qty': sum(
                max(0, int(float(w.get('qty') or 0)))
                for w in (d.get('wip_batches') or [])
            ),
        }

    def _forecast_merge_surplus_detail_for_pairs(self, pairs, detail_by_op):
        """单条替代方案内按 BOM 瓶颈成套，与库存列 _forecast_bom_tier_assembled_counts 一致。"""
        merged = self._forecast_empty_surplus_detail()
        links = [
            (int(self._parse_int(sid)), max(1, self._parse_int(mult) or 1))
            for sid, mult in (pairs or [])
            if self._parse_int(sid)
        ]
        if not links:
            return merged

        inv_by_op = {}
        for sid, _mult in links:
            d = detail_by_op.get(sid) or self._forecast_empty_surplus_detail()
            inv_by_op[sid] = self._forecast_surplus_detail_to_inv_scalars(d)
        asm = self._forecast_bom_tier_assembled_counts(inv_by_op, links)

        region_names = set()
        for sid, _mult in links:
            d = detail_by_op.get(sid) or self._forecast_empty_surplus_detail()
            for rname in (d.get('overseas_by_region') or {}):
                region_names.add(str(rname or '').strip() or '-')

        for rname in sorted(region_names, key=lambda x: str(x)):
            per_sid = []
            for sid, mult in links:
                d = detail_by_op.get(sid) or self._forecast_empty_surplus_detail()
                raw = int((d.get('overseas_by_region') or {}).get(rname) or 0)
                per_sid.append(int(raw // max(1, mult)))
            qty = min(per_sid) if per_sid else 0
            if qty > 0:
                merged['overseas_by_region'][rname] = qty

        merged['factory_stock_qty'] = int(asm.get('factory_stock_qty') or 0)

        transit_keys = set()
        for sid, _mult in links:
            d = detail_by_op.get(sid) or self._forecast_empty_surplus_detail()
            for batch in d.get('transit_batches') or []:
                rname = str(batch.get('region_name') or '-').strip() or '-'
                edate = batch.get('expected_listed_date')
                transit_keys.add((rname, edate))

        for rname, edate in sorted(transit_keys, key=lambda x: (str(x[0]), str(x[1] or ''))):
            per_sid = []
            for sid, mult in links:
                d = detail_by_op.get(sid) or self._forecast_empty_surplus_detail()
                raw = 0
                for batch in d.get('transit_batches') or []:
                    br = str(batch.get('region_name') or '-').strip() or '-'
                    if br == rname and batch.get('expected_listed_date') == edate:
                        raw += int(float(batch.get('qty') or 0))
                per_sid.append(int(raw // max(1, mult)))
            qty = min(per_sid) if per_sid else 0
            if qty > 0:
                merged['transit_batches'].append({
                    'region_name': rname,
                    'expected_listed_date': edate,
                    'qty': qty,
                })

        wip_keys = set()
        for sid, _mult in links:
            d = detail_by_op.get(sid) or self._forecast_empty_surplus_detail()
            for wb in d.get('wip_batches') or []:
                comp = wb.get('expected_completion_date')
                if comp:
                    wip_keys.add(comp)

        for comp in sorted(wip_keys, key=lambda x: str(x or '')):
            per_sid = []
            for sid, mult in links:
                d = detail_by_op.get(sid) or self._forecast_empty_surplus_detail()
                raw = sum(
                    int(float(w.get('qty') or 0))
                    for w in (d.get('wip_batches') or [])
                    if w.get('expected_completion_date') == comp
                )
                per_sid.append(int(raw // max(1, mult)))
            qty = min(per_sid) if per_sid else 0
            if qty > 0:
                merged['wip_batches'].append({
                    'qty': qty,
                    'expected_completion_date': comp,
                })

        for sid, _mult in links:
            d = detail_by_op.get(sid) or self._forecast_empty_surplus_detail()
            merged['overseas_updated_at'] = self._forecast_date_max_iso(
                merged.get('overseas_updated_at'), d.get('overseas_updated_at')
            )
            merged['transit_updated_at'] = self._forecast_date_max_iso(
                merged.get('transit_updated_at'), d.get('transit_updated_at')
            )
            merged['factory_stock_updated_at'] = self._forecast_date_max_iso(
                merged.get('factory_stock_updated_at'), d.get('factory_stock_updated_at')
            )
            merged['wip_updated_at'] = self._forecast_date_max_iso(
                merged.get('wip_updated_at'), d.get('wip_updated_at')
            )
        return merged

    def _forecast_attach_surplus_detail_to_order_rows(self, conn, rows):
        if not rows:
            return
        op_ids = sorted({
            self._parse_int((r.get('labels') or {}).get('order_product_id'))
            for r in rows
            if self._parse_int((r.get('labels') or {}).get('order_product_id'))
        })
        if not op_ids:
            return
        substitute_plans_by_owner = self._forecast_load_all_substitute_plans_by_owner(conn, op_ids)
        extra_subs = []
        for oid in op_ids:
            for sid in self._forecast_substitute_ids_from_plan_groups(
                substitute_plans_by_owner.get(oid) or []
            ):
                extra_subs.append(int(sid))
        load_ids = sorted(set(op_ids).union(extra_subs))
        brief_by_op = self._forecast_load_order_product_brief_map(conn, load_ids)
        detail_by_op = self._forecast_load_inventory_surplus_detail_by_order_product(conn, load_ids)
        for row in rows:
            oid = self._parse_int((row.get('labels') or {}).get('order_product_id'))
            if not oid:
                row['inventory_surplus_detail'] = self._forecast_empty_surplus_detail()
                row['inventory_stock_sources'] = []
                row['inventory_surplus_by_op'] = {}
                continue
            plans = substitute_plans_by_owner.get(oid) or []
            sources = self._forecast_build_inventory_stock_sources(oid, brief_by_op, plans)
            row['inventory_stock_sources'] = sources
            row['inventory_surplus_by_op'] = {
                str(int(s['order_product_id'])): (
                    detail_by_op.get(int(s['order_product_id'])) or self._forecast_empty_surplus_detail()
                )
                for s in sources
            }
            owner_detail = detail_by_op.get(oid) or self._forecast_empty_surplus_detail()
            if plans:
                plan_detail = self._forecast_merge_surplus_detail_for_plan_groups(
                    plans, detail_by_op,
                )
                row['inventory_surplus_detail'] = self._forecast_merge_surplus_detail_add(
                    owner_detail, plan_detail,
                )
            else:
                row['inventory_surplus_detail'] = owner_detail

    def _forecast_build_spec_inventory_stub_rows(self, conn, query_params, sf_group=None):
        """仅含 row_key / variant_id，供 patch=inventory 附加库存，避免重复拉历史与预测。"""
        dim_rows = self._forecast_load_variant_dim(conn, query_params, sf_group=sf_group)
        out = []
        for r in dim_rows:
            vid = self._parse_int(r.get('variant_id'))
            if not vid:
                continue
            out.append({
                'row_key': str(vid),
                'forecast_mode': 'spec',
                'labels': {
                    'variant_id': vid,
                    'sku_family': r.get('sku_family') or '',
                },
                'history': {},
                'forecasts': {},
            })
        return out

    def _forecast_build_platform_inventory_stub_rows(self, conn, query_params, sf_group=None):
        dim_rows = self._forecast_load_platform_sku_dim(conn, query_params, sf_group=sf_group)
        out = []
        for r in dim_rows:
            spid = self._parse_int(r.get('sales_product_id'))
            if not spid:
                continue
            out.append({
                'row_key': str(spid),
                'forecast_mode': 'platform',
                'labels': {
                    'sales_product_id': spid,
                    'variant_id': self._parse_int(r.get('variant_id')),
                },
                'history': {},
                'forecasts': {},
            })
        return out

    def _forecast_build_order_inventory_stub_rows(self, conn, query_params, sf_group=None):
        dim_rows = self._forecast_load_order_dim(conn, query_params, sf_group=sf_group)
        out = []
        for r in dim_rows:
            op_id = self._parse_int(r.get('order_product_id'))
            if not op_id:
                continue
            links = r.get('variant_links') or []
            out.append({
                'row_key': str(op_id),
                'forecast_mode': 'order',
                'labels': {
                    'order_product_id': op_id,
                    'sku': r.get('sku') or '',
                    'is_on_market': self._parse_int(r.get('is_on_market')) or 0,
                    'variant_links': [
                        {
                            'variant_id': vl.get('variant_id'),
                            'sku_family': vl.get('sku_family'),
                            'spec_name': vl.get('spec_name'),
                            'quantity': vl.get('quantity'),
                        }
                        for vl in links
                    ],
                },
                'history': {},
                'forecasts': {},
            })
        return out

    def handle_sales_forecast_api(self, environ, method, start_response):
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)

            query_params = _sf_parse_qs(_sf_effective_wsgi_query_string(environ))
            forecast_mode = self._forecast_normalize_mode((query_params.get('forecast_mode', [''])[0] or ''))

            months_raw = query_params.get('months') or []
            months = self._forecast_parse_months(months_raw)
            if not months:
                months = self._forecast_default_future_months(6)

            hist_start_raw = (query_params.get('history_start', [''])[0] or '').strip()
            hist_end_raw = (query_params.get('history_end', [''])[0] or '').strip()
            hist_start, hist_end = self._forecast_history_month_range(hist_start_raw, hist_end_raw, default_months=12)
            history_months = self._forecast_iter_months(hist_start, hist_end)
            cur_key = self._forecast_current_month_key()
            perf_max_record_date = None
            mtd_ctx = None
            mtd_in = False
            rows = []
            groups_meta = None
            lazy_raw = (query_params.get('lazy', [''])[0] or '').strip().lower()
            lazy_groups_only = lazy_raw in ('1', 'true', 'yes', 'on')
            inv_raw = (query_params.get('inventory', [''])[0] or '').strip().lower()
            skip_inventory = inv_raw in ('0', 'false', 'no', 'off')
            inventory_patch = self._forecast_is_inventory_patch(query_params)
            sf_group = (query_params.get('sf_group', [''])[0] or '').strip()
            sf_shop_ids = self._forecast_parse_sf_shop_ids(query_params)
            shops_list = []
            destination_regions = []

            with self._get_db_connection() as conn:
                shops_list = self._forecast_list_shops_for_forecast(conn)
                if forecast_mode == 'order':
                    destination_regions = self._forecast_list_destination_regions(conn)
                perf_max_record_date = self._forecast_perf_max_record_date(conn)
                mtd_in = cur_key in set(history_months or [])
                mtd_ctx = self._forecast_mtd_time_context(perf_max_record_date, cur_key) if mtd_in else None
                if lazy_groups_only and not sf_group:
                    if forecast_mode == 'platform':
                        groups_meta = self._forecast_list_platform_groups(conn, query_params)
                    elif forecast_mode == 'order':
                        groups_meta = self._forecast_list_order_groups(conn, query_params)
                    else:
                        groups_meta = self._forecast_list_spec_groups(conn, query_params)
                    rows = []
                elif inventory_patch and sf_group:
                    sf_filter = sf_group
                    if forecast_mode == 'platform':
                        rows = self._forecast_build_platform_inventory_stub_rows(
                            conn, query_params, sf_group=sf_filter
                        )
                    elif forecast_mode == 'order':
                        rows = self._forecast_build_order_inventory_stub_rows(
                            conn, query_params, sf_group=sf_filter
                        )
                    else:
                        rows = self._forecast_build_spec_inventory_stub_rows(
                            conn, query_params, sf_group=sf_filter
                        )
                    self._forecast_attach_inventory_to_rows(
                        conn, rows, forecast_mode, history_months,
                        hist_start=hist_start, hist_end=hist_end, shop_ids=sf_shop_ids,
                    )
                    if forecast_mode == 'order':
                        self._forecast_attach_surplus_detail_to_order_rows(conn, rows)
                else:
                    sf_filter = sf_group if sf_group else None
                    if forecast_mode == 'platform':
                        rows = self._forecast_build_platform_rows(
                            conn, query_params, months, hist_start, hist_end,
                            sf_group=sf_filter, shop_ids=sf_shop_ids,
                        )
                    elif forecast_mode == 'order':
                        rows = self._forecast_build_order_rows(
                            conn, query_params, months, hist_start, hist_end,
                            sf_group=sf_filter, shop_ids=sf_shop_ids,
                        )
                    else:
                        rows = self._forecast_build_spec_rows(
                            conn, query_params, months, hist_start, hist_end,
                            sf_group=sf_filter, shop_ids=sf_shop_ids,
                        )
                    if not skip_inventory:
                        self._forecast_attach_inventory_to_rows(
                            conn, rows, forecast_mode, history_months,
                            hist_start=hist_start, hist_end=hist_end, shop_ids=sf_shop_ids,
                        )
                        self._forecast_attach_surplus_detail_to_order_rows(conn, rows)
                    self._forecast_attach_remarks_to_rows(conn, rows, forecast_mode)
                    for row in rows:
                        row['mtd_completion'] = self._forecast_row_mtd_completion(
                            row.get('history') or {}, row.get('forecasts') or {}, mtd_ctx
                        )

            mtd_payload = {
                'current_month_key': cur_key,
                'in_history_range': mtd_in,
                'perf_max_record_date': self._forecast_format_date_only(perf_max_record_date),
            }
            if mtd_ctx:
                mtd_payload.update({
                    'days_in_month': mtd_ctx.get('days_in_month'),
                    'ref_day': mtd_ctx.get('ref_day'),
                    'time_progress': mtd_ctx.get('time_progress'),
                })

            payload = {
                'status': 'success',
                'forecast_mode': forecast_mode,
                'months': months,
                'history_months': history_months,
                'history_range': {'start': hist_start, 'end': hist_end},
                'turnover_sales_window': self._turnover_sales_window(conn),
                'mtd': mtd_payload,
                'rows': rows,
                'shops': shops_list,
            }
            if forecast_mode == 'order':
                payload['destination_regions'] = destination_regions
            if lazy_groups_only and not sf_group:
                payload['lazy'] = True
                payload['groups'] = groups_meta or []
            elif sf_group:
                payload['lazy'] = True
                payload['sf_group'] = sf_group
            if inventory_patch and sf_group:
                payload['patch'] = 'inventory'
            return self.send_json(payload, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    # ----- 三种模式的行装配 -----

    def _forecast_build_platform_rows(self, conn, query_params, months, hist_start, hist_end, sf_group=None, shop_ids=None):
        dim_rows = self._forecast_load_platform_sku_dim(conn, query_params, sf_group=sf_group)
        sales_product_ids = [self._parse_int(r.get('sales_product_id')) for r in dim_rows if self._parse_int(r.get('sales_product_id'))]
        variant_ids = [self._parse_int(r.get('variant_id')) for r in dim_rows if self._parse_int(r.get('variant_id'))]
        skip_thumbs = self._forecast_query_flag_off(query_params, 'thumbs')
        thumb_by_variant = {} if skip_thumbs else self._forecast_load_variant_thumb_b64(conn, variant_ids)
        cells = self._forecast_load_platform_cells(conn, sales_product_ids, months)
        history = self._forecast_load_history_by_sales_product(
            conn, sales_product_ids, hist_start, hist_end, shop_ids=shop_ids
        )

        hist_empty = self._forecast_empty_history_perf_payload()
        out = []
        for r in dim_rows:
            spid = self._parse_int(r.get('sales_product_id'))
            if not spid:
                continue
            row = {
                'row_key': str(spid),
                'forecast_mode': 'platform',
                'labels': {
                    'sales_product_id': spid,
                    'variant_id': self._parse_int(r.get('variant_id')),
                    'platform_sku': r.get('platform_sku') or '',
                    'shop_name': r.get('shop_name') or '',
                    'sku_family': r.get('sku_family') or '',
                    'spec_name': r.get('spec_name') or '',
                    'fabric_code': r.get('fabric_code') or '',
                    'fabric_name_en': r.get('fabric_name_en') or '',
                    'representative_color': r.get('representative_color') or '',
                    'variant_thumb_b64': thumb_by_variant.get(self._parse_int(r.get('variant_id')) or 0, ''),
                    'product_status': r.get('product_status') or '',
                },
                'history': {},
                'forecasts': {},
            }
            for m in months:
                cell = cells.get((spid, m))
                stored_qty = cell.get('latest_qty') if cell else None
                value_qty = stored_qty or 0
                row['forecasts'][m] = {
                    'stored_qty': stored_qty,
                    'value_qty': value_qty,
                    'is_inherited': False,
                    'source': None,
                    'cell_meta': cell,
                }
            for hm in self._forecast_iter_months(hist_start, hist_end):
                h = history.get((spid, hm))
                row['history'][hm] = h if h else hist_empty.copy()
            out.append(row)
        return out

    def _forecast_build_spec_rows(self, conn, query_params, months, hist_start, hist_end, sf_group=None, shop_ids=None):
        dim_rows = self._forecast_load_variant_dim(conn, query_params, sf_group=sf_group)
        variant_ids = [self._parse_int(r.get('variant_id')) for r in dim_rows if self._parse_int(r.get('variant_id'))]
        skip_thumbs = self._forecast_query_flag_off(query_params, 'thumbs')
        thumb_by_variant = {} if skip_thumbs else self._forecast_load_variant_thumb_b64(conn, variant_ids)
        spec_cells = self._forecast_load_spec_cells(conn, variant_ids, months)
        history = self._forecast_load_history_by_variant(conn, variant_ids, hist_start, hist_end, shop_ids=shop_ids)

        hist_empty = self._forecast_empty_history_perf_payload()
        platforms_by_variant = self._forecast_load_variant_platform_skus(conn, variant_ids)
        all_sp_ids = sorted({
            sp.get('sales_product_id')
            for ps in platforms_by_variant.values()
            for sp in ps
            if sp.get('sales_product_id')
        })
        platform_cells = self._forecast_load_platform_cells(conn, all_sp_ids, months)

        out = []
        for r in dim_rows:
            vid = self._parse_int(r.get('variant_id'))
            if not vid:
                continue
            platform_skus = platforms_by_variant.get(vid, [])
            row = {
                'row_key': str(vid),
                'forecast_mode': 'spec',
                'labels': {
                    'variant_id': vid,
                    'sku_family': r.get('sku_family') or '',
                    'spec_name': r.get('spec_name') or '',
                    'fabric_code': r.get('fabric_code') or '',
                    'fabric_name_en': r.get('fabric_name_en') or '',
                    'representative_color': r.get('representative_color') or '',
                    'variant_thumb_b64': thumb_by_variant.get(vid, ''),
                    'platform_skus': platform_skus,
                },
                'history': {},
                'forecasts': {},
            }
            for m in months:
                cell = spec_cells.get((vid, m))
                stored_qty = cell.get('latest_qty') if cell else None

                inherited_items = []
                inherited_total = 0
                for ps in platform_skus:
                    spid = ps.get('sales_product_id')
                    if not spid:
                        continue
                    pc = platform_cells.get((spid, m))
                    pq = pc.get('latest_qty') if pc else 0
                    if pq:
                        inherited_items.append({
                            'id': spid,
                            'label': ps.get('platform_sku') or f'SP#{spid}',
                            'qty': pq,
                        })
                        inherited_total += pq

                if stored_qty is None and inherited_total > 0:
                    row['forecasts'][m] = {
                        'stored_qty': None,
                        'value_qty': inherited_total,
                        'is_inherited': True,
                        'source': {
                            'type': 'platform_sku',
                            'items': inherited_items,
                            'total_qty': inherited_total,
                        },
                        'cell_meta': None,
                    }
                else:
                    row['forecasts'][m] = {
                        'stored_qty': stored_qty,
                        'value_qty': stored_qty or 0,
                        'is_inherited': False,
                        'source': {
                            'type': 'platform_sku',
                            'items': inherited_items,
                            'total_qty': inherited_total,
                        } if inherited_items else None,
                        'cell_meta': cell,
                    }
            for hm in self._forecast_iter_months(hist_start, hist_end):
                h = history.get((vid, hm))
                row['history'][hm] = h if h else hist_empty.copy()
            out.append(row)
        return out

    def _forecast_build_order_rows(self, conn, query_params, months, hist_start, hist_end, sf_group=None, shop_ids=None):
        dim_rows = self._forecast_load_order_dim(conn, query_params, sf_group=sf_group)
        order_ids = [self._parse_int(r.get('order_product_id')) for r in dim_rows if self._parse_int(r.get('order_product_id'))]
        order_cells = self._forecast_load_order_cells(conn, order_ids, months)
        skip_thumbs = self._forecast_query_flag_off(query_params, 'thumbs')
        thumb_by_order = {} if skip_thumbs else self._forecast_load_order_product_thumb_b64(conn, order_ids)

        # spec 预测的覆盖（包含继承自 platform 的逻辑），用于推导默认值。
        # 沿 sales_variant_order_links 扩订单/变体闭包：下市本体仍挂在某规格上时，继承预测需经替代方案折算到在市 SKU。
        links_by_dim = self._forecast_load_order_variant_links_by_order_product_ids(conn, order_ids)
        seed_variants = {
            self._parse_int(vl.get('variant_id'))
            for oid in order_ids
            for vl in (links_by_dim.get(oid) or [])
            if self._parse_int(vl.get('variant_id'))
        }
        touch_orders, touch_variants = self._forecast_grow_linked_orders_variants_bidir(conn, order_ids, seed_variants)
        for oid in order_ids:
            oi = self._parse_int(oid)
            if oi:
                touch_orders.add(int(oi))

        all_variant_ids = sorted(touch_variants)
        links_by_op_all = self._forecast_load_order_variant_links_by_order_product_ids(conn, sorted(touch_orders))
        plan_items_by_owner = self._forecast_load_default_substitute_plan_items_by_owner(conn, sorted(touch_orders))
        brief_by_op = self._forecast_load_order_product_brief_map(conn, sorted(touch_orders))

        spec_cells = self._forecast_load_spec_cells(conn, all_variant_ids, months)
        platforms_by_variant = self._forecast_load_variant_platform_skus(conn, all_variant_ids)
        all_sp_ids = sorted({
            sp.get('sales_product_id')
            for ps in platforms_by_variant.values()
            for sp in ps
            if sp.get('sales_product_id')
        })
        platform_cells = self._forecast_load_platform_cells(conn, all_sp_ids, months)

        # 历史月销量：聚合为变体粒度。下单 SKU 行「销量」= 各关联变体月销 × sales_variant_order_links.quantity 之和（BOM 件数）；
        # 净销售额/订单/会话/退款仍按链接数量占该变体全局链接权重比例分摊，避免金额重复全额记入每个下单 SKU。
        variant_hist_ids = sorted(touch_variants)
        history_by_variant = (
            self._forecast_load_history_by_variant(conn, variant_hist_ids, hist_start, hist_end, shop_ids=shop_ids)
            if variant_hist_ids else {}
        )
        history_by_sp = (
            self._forecast_load_history_by_sales_product(conn, all_sp_ids, hist_start, hist_end, shop_ids=shop_ids)
            if all_sp_ids else {}
        )
        variant_display = self._forecast_load_variant_display_map(conn, variant_hist_ids)
        variant_hist_weight_sum = self._forecast_variant_hist_weight_sum_for_variants(conn, variant_hist_ids)
        hist_month_keys = self._forecast_iter_months(hist_start, hist_end)

        def spec_effective_qty(vid, month):
            cell = spec_cells.get((vid, month))
            if cell:
                return cell.get('latest_qty') or 0
            total = 0
            for ps in platforms_by_variant.get(vid, []):
                spid = ps.get('sales_product_id')
                pc = platform_cells.get((spid, month)) if spid else None
                total += (pc.get('latest_qty') or 0) if pc else 0
            return total

        # 下市本体 SKU：规格预估值 × BOM 经默认发货替代方案折算到各在市 substitute（与库存展开逻辑一致）
        subst_addon_total = {}
        subst_addon_items = {}
        for m in months:
            for owner_oid in touch_orders:
                oowner = int(owner_oid)
                obrief = brief_by_op.get(oowner) or {}
                if self._parse_int(obrief.get('is_on_market')) != 0:
                    continue
                plan = plan_items_by_owner.get(oowner) or []
                if not plan:
                    continue
                owner_skul = (obrief.get('sku') or '').strip() or f'#{oowner}'
                for vl in links_by_op_all.get(oowner) or []:
                    vid = self._parse_int(vl.get('variant_id'))
                    if not vid:
                        continue
                    qty_per = max(1, self._parse_int(vl.get('quantity')) or 1)
                    spec_qty = spec_effective_qty(vid, m)
                    if not spec_qty:
                        continue
                    expanded = self._forecast_expand_owner_pairs_with_substitute_plans(
                        [(oowner, qty_per)], plan_items_by_owner
                    )
                    for sid, w_eff in expanded:
                        sid = self._parse_int(sid)
                        w_eff = int(w_eff) if w_eff else 0
                        if not sid or sid == oowner:
                            continue
                        sub = int(spec_qty) * w_eff
                        if sub <= 0:
                            continue
                        key = (sid, m)
                        subst_addon_total[key] = int(subst_addon_total.get(key) or 0) + sub
                        lab_var = f"{vl.get('sku_family') or ''} {vl.get('spec_name') or ''}".strip() or f'V#{vid}'
                        subst_addon_items.setdefault(key, []).append({
                            'id': vid,
                            'kind': 'substitute',
                            'owner_order_product_id': oowner,
                            'owner_sku': owner_skul,
                            'label': f"{lab_var}（下市 {owner_skul} 替代→×{w_eff}）",
                            'qty': sub,
                            'spec_qty': spec_qty,
                            'ratio': w_eff,
                        })

        out = []
        for r in dim_rows:
            op_id = self._parse_int(r.get('order_product_id'))
            if not op_id:
                continue
            links = r.get('variant_links') or []
            row = {
                'row_key': str(op_id),
                'forecast_mode': 'order',
                'labels': {
                    'order_product_id': op_id,
                    'sku': r.get('sku') or '',
                    'sku_family': r.get('sku_family') or '',
                    'spec_qty_short': r.get('spec_qty_short') or '',
                    'contents_desc_en': r.get('contents_desc_en') or '',
                    'fabric_code': r.get('fabric_code') or '',
                    'fabric_name_en': r.get('fabric_name_en') or '',
                    'representative_color': r.get('representative_color') or '',
                    'is_on_market': self._parse_int(r.get('is_on_market')) or 0,
                    'order_thumb_b64': thumb_by_order.get(op_id, ''),
                    'variant_links': [
                        {
                            'variant_id': vl.get('variant_id'),
                            'sku_family': vl.get('sku_family'),
                            'spec_name': vl.get('spec_name'),
                            'quantity': vl.get('quantity'),
                        }
                        for vl in links
                    ],
                },
                'history': {},
                'forecasts': {},
            }
            for m in months:
                cell = order_cells.get((op_id, m))
                stored_qty = cell.get('latest_qty') if cell else None

                inherited_items = []
                inherited_total = 0
                for vl in links:
                    vid = vl.get('variant_id')
                    qty_per = max(1, vl.get('quantity') or 1)
                    spec_qty = spec_effective_qty(vid, m)
                    if spec_qty:
                        sub = spec_qty * qty_per
                        inherited_items.append({
                            'id': vid,
                            'label': f"{vl.get('sku_family') or ''} {vl.get('spec_name') or ''}".strip() or f'V#{vid}',
                            'qty': sub,
                            'spec_qty': spec_qty,
                            'ratio': qty_per,
                        })
                        inherited_total += sub

                sk = (op_id, m)
                extra = int(subst_addon_total.get(sk) or 0)
                if extra:
                    inherited_total += extra
                    inherited_items.extend(subst_addon_items.get(sk) or [])

                if stored_qty is None and inherited_total > 0:
                    row['forecasts'][m] = {
                        'stored_qty': None,
                        'value_qty': inherited_total,
                        'is_inherited': True,
                        'source': {
                            'type': 'spec',
                            'items': inherited_items,
                            'total_qty': inherited_total,
                        },
                        'cell_meta': None,
                    }
                else:
                    row['forecasts'][m] = {
                        'stored_qty': stored_qty,
                        'value_qty': stored_qty or 0,
                        'is_inherited': False,
                        'source': {
                            'type': 'spec',
                            'items': inherited_items,
                            'total_qty': inherited_total,
                        } if inherited_items else None,
                        'cell_meta': cell,
                    }
            row['history'] = self._forecast_build_order_like_history_for_links(
                links, history_by_variant, variant_hist_weight_sum, hist_month_keys
            )
            row['history_source'] = {}
            for hm in hist_month_keys:
                row['history_source'][hm] = self._forecast_build_order_history_sources_for_links(
                    links,
                    hm,
                    history_by_variant,
                    history_by_sp,
                    platforms_by_variant,
                    variant_display,
                )
            out.append(row)
        return out

    def _forecast_upsert_cell(self, cur, table, key_cols, key_values, latest_qty):
        """通用 upsert：先 SELECT 看有无记录，再 UPDATE/INSERT。"""
        where_sql = ' AND '.join([f"{c}=%s" for c in key_cols])
        cur.execute(
            f"SELECT id, latest_qty, latest_updated_at FROM {table} WHERE {where_sql} LIMIT 1",
            tuple(key_values)
        )
        row = cur.fetchone() or None
        if row:
            cur.execute(
                f"""
                UPDATE {table}
                SET prev_qty = latest_qty,
                    prev_updated_at = latest_updated_at,
                    latest_qty = %s,
                    latest_updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                """,
                (latest_qty, self._parse_int(row.get('id')))
            )
            return self._parse_int(row.get('id')), False
        col_list = ','.join(key_cols + ['initial_qty', 'prev_qty', 'latest_qty', 'created_at', 'prev_updated_at', 'latest_updated_at'])
        placeholder_count = len(key_cols)
        placeholders = ','.join(['%s'] * placeholder_count) + ', %s, NULL, %s, CURRENT_TIMESTAMP, NULL, CURRENT_TIMESTAMP'
        cur.execute(
            f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})",
            tuple(list(key_values) + [latest_qty, latest_qty])
        )
        return cur.lastrowid, True

    def _forecast_select_cell(self, cur, table, key_cols, key_values):
        where_sql = ' AND '.join([f"{c}=%s" for c in key_cols])
        select_cols = ['id'] + [c for c in key_cols if c != 'forecast_month'] + [
            'forecast_month', 'initial_qty', 'prev_qty', 'latest_qty',
            'created_at', 'prev_updated_at', 'latest_updated_at'
        ]
        cur.execute(
            f"SELECT {','.join(select_cols)} FROM {table} WHERE {where_sql} LIMIT 1",
            tuple(key_values)
        )
        return cur.fetchone() or None

    def _forecast_dedupe_bulk_cells(self, parsed):
        """同一 (row_key, forecast_month) 仅保留最后一条，避免单条 INSERT 多 VALUES 内重复唯一键。"""
        by_key = {}
        for cell in parsed or []:
            if not isinstance(cell, dict):
                continue
            rk = self._parse_int(cell.get('row_key'))
            ms = self._forecast_normalize_month(cell.get('forecast_month'))
            if not rk or not ms:
                continue
            by_key[(rk, ms)] = cell
        out = list(by_key.values())
        out.sort(key=lambda c: (c.get('forecast_month') or '', c.get('row_key') or 0))
        return out

    def _forecast_bulk_upsert_cells_chunked(self, cur, table, id_col_name, parsed, chunk_size=450):
        """多行 INSERT … ON DUPLICATE KEY UPDATE，显著减少大批量保存的往返次数。"""
        if not parsed:
            return
        col_list = (
            f'{id_col_name}, forecast_month, initial_qty, prev_qty, latest_qty, '
            f'created_at, prev_updated_at, latest_updated_at'
        )
        dup_sql = (
            'prev_qty = latest_qty, prev_updated_at = latest_updated_at, '
            'latest_qty = VALUES(latest_qty), latest_updated_at = CURRENT_TIMESTAMP'
        )
        for i in range(0, len(parsed), chunk_size):
            chunk = parsed[i : i + chunk_size]
            placeholders = []
            params = []
            for cell in chunk:
                rid = int(cell['row_key'])
                month = cell['forecast_month']
                q = max(0, int(cell['latest_qty'] or 0))
                placeholders.append('(%s, %s, %s, NULL, %s, CURRENT_TIMESTAMP, NULL, CURRENT_TIMESTAMP)')
                params.extend([rid, month, q, q])
            sql = (
                f'INSERT INTO {table} ({col_list}) VALUES '
                + ','.join(placeholders)
                + f' ON DUPLICATE KEY UPDATE {dup_sql}'
            )
            cur.execute(sql, tuple(params))

    def _forecast_bulk_fetch_cells_meta(self, cur, table, id_col_name, key_cols, parsed, fetch_chunk=500):
        """按本批 keys 批量读回单元格，供前端刷新 cell_meta。"""
        if not parsed:
            return []
        refreshed = []
        keys = [(int(c['row_key']), c['forecast_month']) for c in parsed]
        sel_cols = [
            'id',
            id_col_name,
            'forecast_month',
            'initial_qty',
            'prev_qty',
            'latest_qty',
            'created_at',
            'prev_updated_at',
            'latest_updated_at',
        ]
        sel_sql = ','.join(sel_cols)
        k0 = id_col_name
        for i in range(0, len(keys), fetch_chunk):
            part = keys[i : i + fetch_chunk]
            in_ph = ','.join(['(%s,%s)'] * len(part))
            flat = []
            for rk, ms in part:
                flat.extend([rk, ms])
            cur.execute(
                f'SELECT {sel_sql} FROM {table} WHERE ({k0}, forecast_month) IN ({in_ph})',
                tuple(flat),
            )
            for row in cur.fetchall() or []:
                row_key = self._parse_int(row.get(k0))
                month_str = self._forecast_month_to_str(row.get('forecast_month'))
                if not row_key or not month_str:
                    continue
                refreshed.append({
                    'row_key': str(row_key),
                    'forecast_month': month_str,
                    'cell_meta': self._forecast_serialize_cell(
                        row, month_str, extra_keys={key_cols[0]: row_key}
                    ),
                })
        return refreshed

    def handle_sales_forecast_bulk_update_api(self, environ, method, start_response):
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            data = self._read_json_body(environ) or {}
            forecast_mode = self._forecast_normalize_mode(data.get('forecast_mode') if isinstance(data, dict) else None)
            cells = data.get('cells') if isinstance(data, dict) else None

            parsed = []
            for item in (cells or []):
                if not isinstance(item, dict):
                    continue
                row_key = self._parse_int(item.get('row_key'))
                month_str = self._forecast_normalize_month(item.get('forecast_month'))
                latest_qty = self._parse_int(item.get('latest_qty'))
                if not row_key or not month_str:
                    continue
                if latest_qty is None:
                    latest_qty = 0
                parsed.append({
                    'row_key': row_key,
                    'forecast_month': month_str,
                    'latest_qty': max(0, latest_qty),
                })

            parsed = self._forecast_dedupe_bulk_cells(parsed)
            if not parsed:
                return self.send_json({'status': 'error', 'message': '没有有效的更新数据'}, start_response)

            if forecast_mode == 'platform':
                table = 'sales_forecast_platform_sku_monthly'
                key_cols = ['sales_product_id', 'forecast_month']
                id_col = 'sales_product_id'
            elif forecast_mode == 'order':
                table = 'sales_forecast_order_sku_monthly'
                key_cols = ['order_product_id', 'forecast_month']
                id_col = 'order_product_id'
            else:
                forecast_mode = 'spec'
                table = 'sales_forecast_spec_monthly'
                key_cols = ['variant_id', 'forecast_month']
                id_col = 'variant_id'

            refreshed = []
            with self._get_db_connection() as conn:
                try:
                    with conn.cursor() as cur:
                        self._forecast_bulk_upsert_cells_chunked(cur, table, id_col, parsed)
                        refreshed = self._forecast_bulk_fetch_cells_meta(
                            cur, table, id_col, key_cols, parsed
                        )
                    conn.commit()
                except Exception as inner:
                    try:
                        conn.rollback()
                    except Exception:
                        pass
                    return self.send_json({'status': 'error', 'message': str(inner)}, start_response)

            return self.send_json({
                'status': 'success',
                'forecast_mode': forecast_mode,
                'cells': refreshed,
                'saved_count': len(parsed),
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
