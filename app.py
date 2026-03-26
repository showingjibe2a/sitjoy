#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WSGI 应用 - 用于 Synology Web Station
兼容 Apache + mod_wsgi
"""

import sys
import os

# 强制设置所有I/O为UTF-8
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

os.environ['PYTHONIOENCODING'] = 'utf-8'

from urllib.parse import urlparse, parse_qs, quote
import json
import ast
import re
from datetime import datetime, timedelta
import calendar
import mimetypes
import base64
import io
from pathlib import Path
import time
import cgi
import tempfile
import zipfile
import xml.etree.ElementTree as ET
import hmac
import hashlib
import secrets
import unicodedata
import threading
try:
    from PIL import Image
    _pillow_import_error = None
except Exception as e:
    Image = None
    _pillow_import_error = str(e)
try:
    from openpyxl import Workbook, load_workbook
    _openpyxl_import_error = None
except Exception as e:
    Workbook = None
    load_workbook = None
    _openpyxl_import_error = str(e)
try:
    import pymysql
    _pymysql_import_error = None
except Exception as e:
    pymysql = None
    _pymysql_import_error = str(e)

# 导入业务逻辑 mixin 模块
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'modules'))
try:
    from auth_employee_mixin import AuthEmployeeMixin
    from core_app_mixin import CoreAppMixin
    from db_schema_basics_mixin import DbSchemaBasicsMixin
    from excel_tools_mixin import ExcelToolsMixin
    from file_management_mixin import FileManagementMixin
    from request_routing_mixin import RequestRoutingMixin
    from logistics_warehouse_mixin import LogisticsWarehouseMixin
    from logistics_in_transit_mixin import LogisticsInTransitMixin
    from sales_product_mixin import SalesProductMixin
    from sales_management_mixin import SalesManagementMixin
    from app_entry_mixin import AppEntryMixin
    from page_permission_mixin import PagePermissionMixin
    from logistics_schema_mixin import LogisticsSchemaMixin
    from sales_schema_mixin import SalesSchemaMixin
    _mixin_import_error = None
except Exception as e:
    _mixin_import_error = str(e)
    # 定义空的 mixin 类以防导入失败
    class AuthEmployeeMixin: pass
    class CoreAppMixin: pass
    class DbSchemaBasicsMixin: pass
    class ExcelToolsMixin: pass
    class FileManagementMixin: pass
    class RequestRoutingMixin: pass
    class LogisticsWarehouseMixin: pass
    class LogisticsInTransitMixin: pass
    class SalesProductMixin: pass
    class SalesManagementMixin: pass
    class AppEntryMixin: pass
    class PagePermissionMixin: pass
    class LogisticsSchemaMixin: pass
    class SalesSchemaMixin: pass

# 外部文件夹路径
# 使用 Base64 的子目录名，避免手动输入特殊字符出错
_RESOURCES_PARENT = '/volume1/公共文件SITJOY'
_RESOURCES_CHILD_B64 = '44CO5LiK5p626LWE5rqQ44CP'
_RESOURCES_PARENT_BYTES = _RESOURCES_PARENT.encode('utf-8', errors='surrogatepass')
_RESOURCES_CHILD_BYTES = base64.b64decode(_RESOURCES_CHILD_B64)
RESOURCES_PATH_BYTES = os.path.join(_RESOURCES_PARENT_BYTES, _RESOURCES_CHILD_BYTES)
RESOURCES_PATH = os.fsdecode(RESOURCES_PATH_BYTES)

PAGE_PERMISSION_DEFINITIONS = [
    ('home', '首页', '/', 'templates/index.html'),
    ('about', '关于', '/about', 'templates/about.html'),
    ('gallery', '图片管理', '/gallery', 'templates/gallery.html'),
    ('shop_brand_management', '店铺/品牌管理', '/shop-brand-management', 'templates/shop_brand_management.html'),
    ('amazon_account_health_management', 'Amazon账户健康', '/amazon-account-health-management', 'templates/amazon_account_health_management.html'),
    ('product_management', '品类/货号管理', '/product-management', 'templates/product_management.html'),
    ('fabric_management', '面料管理', '/fabric-management', 'templates/fabric_management.html'),
    ('feature_management', '卖点管理', '/feature-management', 'templates/feature_management.html'),
    ('material_management', '材料管理', '/material-management', 'templates/material_management.html'),
    ('certification_management', '认证管理', '/certification-management', 'templates/certification_management.html'),
    ('order_product_management', '下单产品管理', '/order-product-management', 'templates/order_product_management.html'),
    ('logistics_factory_management', '工厂管理', '/logistics-factory-management', 'templates/logistics_factory_management.html'),
    ('logistics_forwarder_management', '货代管理', '/logistics-forwarder-management', 'templates/logistics_forwarder_management.html'),
    ('logistics_warehouse_management', '海外仓仓库管理', '/logistics-warehouse-management', 'templates/logistics_warehouse_management.html'),
    ('logistics_warehouse_inventory_management', '海外仓库存管理', '/logistics-warehouse-inventory-management', 'templates/logistics_warehouse_inventory_management.html'),
    ('logistics_in_transit_management', '在途物流库存管理', '/logistics-in-transit-management', 'templates/logistics_in_transit_management.html'),
    ('logistics_warehouse_dashboard', '仓储看板', '/logistics-warehouse-dashboard', 'templates/logistics_warehouse_dashboard.html'),
    ('sales_product_management', '销售产品管理', '/sales-product-management', 'templates/sales_product_management.html'),
    ('sales_order_registration_management', '订单登记管理', '/sales-order-registration-management', 'templates/sales_order_registration_management.html'),
    ('parent_management', '父体管理', '/parent-management', 'templates/parent_management.html'),
    ('amazon_ad_adjustment_management', '广告调整', '/amazon-ad-adjustment-management', 'templates/amazon_ad_adjustment_management.html'),
    ('amazon_ad_keyword_management', 'Amazon关键词管理', '/amazon-ad-keyword-management', 'templates/amazon_ad_keyword_management.html'),
    ('amazon_ad_management', '广告信息管理', '/amazon-ad-management', 'templates/amazon_ad_management.html'),
    ('amazon_ad_subtype_management', '广告信息分类管理', '/amazon-ad-subtype-management', 'templates/amazon_ad_subtype_management.html'),
    ('amazon_ad_delivery_management', '广告投放管理', '/amazon-ad-delivery-management', 'templates/amazon_ad_delivery_management.html'),
    ('amazon_ad_product_management', '广告商品管理', '/amazon-ad-product-management', 'templates/amazon_ad_product_management.html'),
    ('factory_stock_management', '工厂在库库存管理', '/factory-stock-management', 'templates/factory_stock_management.html'),
    ('factory_wip_management', '工厂在制库存管理', '/factory-wip-management', 'templates/factory_wip_management.html'),
]

PAGE_PERMISSION_KEYS = [item[0] for item in PAGE_PERMISSION_DEFINITIONS]
PAGE_PERMISSION_LABELS = {item[0]: item[1] for item in PAGE_PERMISSION_DEFINITIONS}
PAGE_TEMPLATE_MAP = {
    '/about': ('templates/about.html', 'about'),
    '/about.html': ('templates/about.html', 'about'),
    '/gallery': ('templates/gallery.html', 'gallery'),
    '/product-management': ('templates/product_management.html', 'product_management'),
    '/fabric-management': ('templates/fabric_management.html', 'fabric_management'),
    '/feature-management': ('templates/feature_management.html', 'feature_management'),
    '/material-management': ('templates/material_management.html', 'material_management'),
    '/certification-management': ('templates/certification_management.html', 'certification_management'),
    '/order-product-management': ('templates/order_product_management.html', 'order_product_management'),
    '/logistics-factory-management': ('templates/logistics_factory_management.html', 'logistics_factory_management'),
    '/logistics-forwarder-management': ('templates/logistics_forwarder_management.html', 'logistics_forwarder_management'),
    '/logistics-warehouse-management': ('templates/logistics_warehouse_management.html', 'logistics_warehouse_management'),
    '/logistics-warehouse-inventory-management': ('templates/logistics_warehouse_inventory_management.html', 'logistics_warehouse_inventory_management'),
    '/logistics-warehouse-dashboard': ('templates/logistics_warehouse_dashboard.html', 'logistics_warehouse_dashboard'),
    '/logistics-in-transit-management': ('templates/logistics_in_transit_management.html', 'logistics_in_transit_management'),
    '/logistics-in-transit-doc-files': ('templates/logistics_in_transit_doc_files.html', 'logistics_in_transit_management'),
    '/shop-brand-management': ('templates/shop_brand_management.html', 'shop_brand_management'),
    '/amazon-account-health-management': ('templates/amazon_account_health_management.html', 'amazon_account_health_management'),
    '/sales-product-management': ('templates/sales_product_management.html', 'sales_product_management'),
    '/sales-order-registration-management': ('templates/sales_order_registration_management.html', 'sales_order_registration_management'),
    '/parent-management': ('templates/parent_management.html', 'parent_management'),
    '/amazon-ad-management': ('templates/amazon_ad_management.html', 'amazon_ad_management'),
    '/amazon-ad-delivery-management': ('templates/amazon_ad_delivery_management.html', 'amazon_ad_delivery_management'),
    '/amazon-ad-product-management': ('templates/amazon_ad_product_management.html', 'amazon_ad_product_management'),
    '/amazon-ad-adjustment-management': ('templates/amazon_ad_adjustment_management.html', 'amazon_ad_adjustment_management'),
    '/amazon-ad-subtype-management': ('templates/amazon_ad_subtype_management.html', 'amazon_ad_subtype_management'),
    '/amazon-ad-keyword-management': ('templates/amazon_ad_keyword_management.html', 'amazon_ad_keyword_management'),
    '/factory-stock-management': ('templates/factory_stock_management.html', 'factory_stock_management'),
    '/factory-wip-management': ('templates/factory_wip_management.html', 'factory_wip_management'),
}
API_PERMISSION_MAP = {
    '/api/employee': 'home',
    '/api/todo': 'home',
    '/api/calendar': 'home',
    '/api/images': 'gallery',
    '/api/browse': 'gallery',
    '/api/image-preview': 'gallery',
    '/api/rename': 'gallery',
    '/api/move': 'gallery',
    '/api/upload': 'gallery',
    '/api/download-zip': 'gallery',
    '/api/sku': 'product_management',
    '/api/category': 'product_management',
    '/api/fabric': 'fabric_management',
    '/api/fabric-images': 'fabric_management',
    '/api/fabric-attach': 'fabric_management',
    '/api/fabric-upload': 'fabric_management',
    '/api/fabric-image-delete': 'fabric_management',
    '/api/feature': 'feature_management',
    '/api/material': 'material_management',
    '/api/material-type': 'material_management',
    '/api/platform-type': 'shop_brand_management',
    '/api/brand': 'shop_brand_management',
    '/api/shop': 'shop_brand_management',
    '/api/amazon-account-health': 'amazon_account_health_management',
    '/api/amazon-account-health-template': 'amazon_account_health_management',
    '/api/amazon-account-health-import': 'amazon_account_health_management',
    '/api/amazon-ad-subtype': 'amazon_ad_subtype_management',
    '/api/amazon-ad-operation-type': 'amazon_ad_subtype_management',
    '/api/amazon-ad': 'amazon_ad_management',
    '/api/amazon-ad-template': 'amazon_ad_management',
    '/api/amazon-ad-import': 'amazon_ad_management',
    '/api/amazon-ad-delivery': 'amazon_ad_delivery_management',
    '/api/amazon-ad-product': 'amazon_ad_product_management',
    '/api/amazon-ad-adjustment': 'amazon_ad_adjustment_management',
    '/api/amazon-ad-keyword': 'amazon_ad_keyword_management',
    '/api/amazon-ad-keyword-template': 'amazon_ad_keyword_management',
    '/api/amazon-ad-keyword-import': 'amazon_ad_keyword_management',
    '/api/certification': 'certification_management',
    '/api/certification-images': 'certification_management',
    '/api/order-product': 'order_product_management',
    '/api/order-product-template': 'order_product_management',
    '/api/order-product-import': 'order_product_management',
    '/api/order-product-carton-calc': 'order_product_management',
    '/api/logistics-factory': 'logistics_factory_management',
    '/api/logistics-forwarder': 'logistics_forwarder_management',
    '/api/logistics-supplier': 'logistics_warehouse_management',
    '/api/logistics-warehouse': 'logistics_warehouse_management',
    '/api/logistics-warehouse-template': 'logistics_warehouse_management',
    '/api/logistics-warehouse-import': 'logistics_warehouse_management',
    '/api/logistics-warehouse-inventory': 'logistics_warehouse_inventory_management',
    '/api/logistics-warehouse-inventory-template': 'logistics_warehouse_inventory_management',
    '/api/logistics-warehouse-inventory-import': 'logistics_warehouse_inventory_management',
    '/api/logistics-warehouse-dashboard': 'logistics_warehouse_dashboard',
    '/api/factory-stock': 'factory_stock_management',
    '/api/factory-wip': 'factory_wip_management',
    '/api/logistics-in-transit': 'logistics_in_transit_management',
    '/api/logistics-in-transit-template': 'logistics_in_transit_management',
    '/api/logistics-in-transit-import': 'logistics_in_transit_management',
    '/api/logistics-in-transit-doc-upload': 'logistics_in_transit_management',
    '/api/logistics-in-transit-doc-files': 'logistics_in_transit_management',
    '/api/sales-product': 'sales_product_management',
    '/api/sales-product-template': 'sales_product_management',
    '/api/sales-product-import': 'sales_product_management',
    '/api/sales-order-registration': 'sales_order_registration_management',
    '/api/sales-order-registration-template': 'sales_order_registration_management',
    '/api/sales-order-registration-import': 'sales_order_registration_management',
    '/api/parent': 'parent_management',
}

class WSGIApp(AppEntryMixin, PagePermissionMixin, AuthEmployeeMixin, DbSchemaBasicsMixin, CoreAppMixin, ExcelToolsMixin, FileManagementMixin, RequestRoutingMixin, LogisticsWarehouseMixin, LogisticsInTransitMixin, SalesProductMixin, SalesManagementMixin):
    """WSGI 应用处理器 - 通过继承各类 mixin 提供综合功能"""
    PAGE_PERMISSION_KEYS = tuple(PAGE_PERMISSION_KEYS)
    _schema_ready_cache = {
        'certification': False,
        'sales_parent': False,
        'sales_order_registration': False,
        'logistics': False,
        'factory_inventory': False,
    }
    
    def __init__(self):
        self.base_path = os.path.dirname(os.path.abspath(__file__))
        self._db_ready = False
        self._order_product_ready = False
        self._material_types_ready = False
        self._materials_ready = False
        self._platform_types_ready = False
        self._brands_ready = False
        self._shops_ready = False
        self._amazon_account_health_ready = False
        self._amazon_ad_ready = False
        self._amazon_ad_delivery_ready = False
        self._amazon_ad_product_ready = False
        self._amazon_ad_adjustment_ready = False
        self._amazon_ad_subtypes_ready = False
        self._amazon_ad_operation_types_ready = False
        self._amazon_keyword_ready = False
        self._sales_parent_ready = bool(self.__class__._schema_ready_cache.get('sales_parent'))
        self._sales_product_ready = False
        self._sales_order_registration_ready = bool(self.__class__._schema_ready_cache.get('sales_order_registration'))
        self._logistics_ready = bool(self.__class__._schema_ready_cache.get('logistics'))
        self._factory_inventory_ready = bool(self.__class__._schema_ready_cache.get('factory_inventory'))
        self._certification_ready = bool(self.__class__._schema_ready_cache.get('certification'))
        self._todo_ready = False
        self._todo_schema_migrated = False
        self._todo_ensure_lock = threading.Lock()
        self._schema_ensure_lock = threading.Lock()
        self._user_session = {}
        self._template_options_cache = {}

    def _get_session_id(self, environ):
        """从 cookie 获取 session_id"""
        cookie = environ.get('HTTP_COOKIE', '')
        pairs = [p.strip().split('=', 1) for p in cookie.split(';') if '=' in p]
        return next((v for k, v in pairs if k == 'session_id'), None)

    def _get_cookie_value(self, environ, name):
        cookie = environ.get('HTTP_COOKIE', '')
        pairs = [p.strip().split('=', 1) for p in cookie.split(';') if '=' in p]
        return next((v for k, v in pairs if k == name), None)

    def _get_auth_secret(self):
        # Stable secret derived from env or db config, avoids cross-worker mismatch
        env_secret = os.environ.get('SITJOY_AUTH_SECRET')
        if env_secret:
            return env_secret.encode('utf-8', errors='ignore')
        cfg = self._get_db_config() or {}
        seed = f"{cfg.get('host','')}|{cfg.get('user','')}|{cfg.get('password','')}|{cfg.get('database','')}"
        return hashlib.sha256(seed.encode('utf-8', errors='ignore')).digest()

    def _b64url_encode(self, raw):
        return base64.urlsafe_b64encode(raw).decode('ascii').rstrip('=')

    def _b64url_decode(self, text):
        pad = '=' * (-len(text) % 4)
        return base64.urlsafe_b64decode((text + pad).encode('ascii'))

    def _make_stateless_token(self, user_id, ttl_seconds=7 * 24 * 3600):
        exp = int(time.time()) + int(ttl_seconds)
        payload = f"{user_id}|{exp}".encode('utf-8', errors='surrogatepass')
        sig = hmac.new(self._get_auth_secret(), payload, hashlib.sha256).hexdigest().encode('ascii')
        return self._b64url_encode(payload + b'|' + sig)

    def _verify_stateless_token(self, token):
        if not token:
            return None
        try:
            raw = self._b64url_decode(token)
            parts = raw.split(b'|')
            if len(parts) != 3:
                return None
            user_id_b, exp_b, sig_b = parts
            payload = user_id_b + b'|' + exp_b
            expected = hmac.new(self._get_auth_secret(), payload, hashlib.sha256).hexdigest().encode('ascii')
            if not hmac.compare_digest(sig_b, expected):
                return None
            exp = int(exp_b.decode('utf-8', errors='ignore') or '0')
            if exp < int(time.time()):
                return None
            return int(user_id_b.decode('utf-8', errors='ignore'))
        except Exception:
            return None

    def _get_session_user(self, environ):
        """从cookie读取登录用户ID"""
        session_id = self._get_session_id(environ)
        token = self._get_cookie_value(environ, 'session_token')
        token_user = self._verify_stateless_token(token)
        if token_user:
            if session_id and session_id not in self._user_session:
                self._user_session[session_id] = token_user
            return token_user
        if not session_id:
            # stateless fallback for environments where DB sessions fail
            if token_user:
                return token_user
        if session_id:
            # 先检查内存缓存
            if session_id in self._user_session:
                return self._user_session[session_id]
            # 回退到数据库查询（支持多进程部署）
            try:
                cfg = self._get_db_config()
                if cfg:
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                "SELECT employee_id FROM sessions WHERE session_id=%s AND (expires_at IS NULL OR expires_at>NOW())",
                                (session_id,)
                            )
                            row = cur.fetchone()
                            if row and row.get('employee_id'):
                                self._user_session[session_id] = row['employee_id']
                                return row['employee_id']
            except Exception as e:
                print(f"Session DB lookup failed: {type(e).__name__}: {e}")
            # session_id 无效时，尝试 stateless token 作为回退
            if token_user:
                return token_user
        return None

    def _set_session_user(self, user_id):
        """创建session并返回session_id"""
        import uuid
        session_id = str(uuid.uuid4())
        # 写入内存缓存
        self._user_session[session_id] = user_id
        # 尝试写入数据库以便在多进程下共享会话
        try:
            cfg = self._get_db_config()
            if cfg:
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        # 过期时间设为 7 天
                        cur.execute(
                            "REPLACE INTO sessions (session_id, employee_id, expires_at) VALUES (%s, %s, DATE_ADD(NOW(), INTERVAL 7 DAY))",
                            (session_id, user_id)
                        )
        except Exception as e:
            print(f"Session DB write failed: {type(e).__name__}: {e}")
        return session_id

    def _b64_from_fs(self, value):
        """将文件系统路径/名称转为 Base64（保留原始字节）"""
        try:
            raw = self._safe_fsencode(value)
        except Exception:
            raw = str(value).encode('utf-8', errors='surrogatepass')
        return base64.b64encode(raw).decode('ascii')

    def _fs_from_b64(self, value):
        """从 Base64 还原文件系统路径/名称"""
        raw = base64.b64decode(value)
        return os.fsdecode(raw)

    def _join_resources(self, rel_path):
        """拼接资源目录（返回 bytes 路径）"""
        if not rel_path:
            return RESOURCES_PATH_BYTES
        try:
            rel_bytes = self._safe_fsencode(rel_path)
        except Exception:
            rel_bytes = str(rel_path).encode('utf-8', errors='surrogatepass')
        return os.path.join(RESOURCES_PATH_BYTES, rel_bytes)

    def _safe_fsencode(self, value):
        if isinstance(value, (bytes, bytearray)):
            return bytes(value)
        try:
            return os.fsencode(value)
        except Exception:
            return str(value).encode('utf-8', errors='surrogatepass')

    def _safe_fsdecode(self, value):
        if isinstance(value, str):
            return value
        try:
            return os.fsdecode(value)
        except Exception:
            return bytes(value).decode('utf-8', errors='surrogatepass')

    def _add_name_and_b64_variants(self, bound_name_map, bound_b64_map, raw_name, fabric_id):
        """Add normalized string variants and base64-of-bytes variants for a given image name into maps."""
        if not raw_name:
            return
        try:
            base = raw_name.split('/')[-1].strip()
        except Exception:
            base = raw_name
        if not base:
            return
        try:
            nfc = unicodedata.normalize('NFC', base)
        except Exception:
            nfc = base
        try:
            nfd = unicodedata.normalize('NFD', base)
        except Exception:
            nfd = nfc

        for key in (nfc, nfc.lower(), nfd, nfd.lower()):
            if not key:
                continue
            if key not in bound_name_map:
                bound_name_map[key] = set()
            if fabric_id is not None:
                bound_name_map[key].add(int(fabric_id))

        # Add multiple byte-encoding variants for more robust matching across
        # filesystem encodings and database-stored strings. Try fs encoding first,
        # then fall back to several common encodings with surrogatepass so that
        # round-trip surrogate bytes are preserved when present on the NAS.
        for variant in (nfc, nfd):
            if not variant:
                continue
            encodings_to_try = []
            try:
                encodings_to_try.append(os.fsencode(variant))
            except Exception:
                pass
            try:
                encodings_to_try.append(variant.encode('utf-8', errors='surrogatepass'))
            except Exception:
                pass
            try:
                encodings_to_try.append(variant.encode('gb18030', errors='surrogatepass'))
            except Exception:
                pass
            try:
                encodings_to_try.append(variant.encode('latin-1', errors='surrogatepass'))
            except Exception:
                pass

            # de-duplicate byte variants
            seen = set()
            for b in encodings_to_try:
                if not isinstance(b, (bytes, bytearray)):
                    continue
                if b in seen:
                    continue
                seen.add(b)
                try:
                    b64 = base64.b64encode(b).decode('ascii')
                    if b64 not in bound_b64_map:
                        bound_b64_map[b64] = set()
                    if fabric_id is not None:
                        bound_b64_map[b64].add(int(fabric_id))
                except Exception:
                    continue

    def _is_image_name(self, name):
        """判断是否为图片文件名（兼容 bytes/str）"""
        if isinstance(name, (bytes, bytearray)):
            try:
                name = os.fsdecode(name)
            except Exception:
                name = name.decode('utf-8', errors='ignore')
        return str(name).lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'))

    def _to_int(self, value, default=None):
        try:
            return int(value)
        except Exception:
            return default

    def _normalize_fabric_remark(self, remark):
        value = (remark or '').strip()
        allowed = {
            '原图',
            '主图·Swatch',
            '主图·卖点',
            'A+·电脑端',
            'A+·手机端',
            'A+·通用',
        }
        if value in allowed:
            return value
        if value in ('平面原图', '褶皱原图'):
            return '原图'
        if '卖点' in value:
            return '主图·卖点'
        if 'Swatch' in value or 'swatch' in value:
            return '主图·Swatch'
        if 'A+' in value or value.startswith('A＋'):
            if '电脑' in value:
                return 'A+·电脑端'
            if '手机' in value:
                return 'A+·手机端'
            return 'A+·通用'
        return '原图'

    def _build_fabric_image_plan(self, images, fabric_code):
        """为面料图片生成重命名计划和最终入库数据"""
        folder = self._ensure_fabric_folder()
        remark_counters = {}
        planned_images = []
        rename_pairs = []
        missing = []
        not_ready = []

        for idx, img in enumerate(images):
            old_name = (img.get('image_name') or '').strip()
            if not old_name:
                continue

            src_path = os.path.join(folder, self._safe_fsencode(old_name))
            if not os.path.exists(src_path):
                missing.append(old_name)
                continue
            try:
                if os.path.getsize(src_path) <= 0:
                    not_ready.append(old_name)
                    continue
            except Exception:
                not_ready.append(old_name)
                continue

            remark = self._normalize_fabric_remark(img.get('remark'))
            remark_counters[remark] = remark_counters.get(remark, 0) + 1
            index_in_remark = remark_counters[remark]
            new_name = self._rename_fabric_image_with_remark(old_name, fabric_code, remark, index_in_remark)

            planned_images.append({
                'image_name': new_name,
                'remark': remark,
                'sort_order': self._to_int(img.get('sort_order'), idx) if isinstance(img, dict) else idx,
                'is_primary': bool(img.get('is_primary', idx == 0)) if isinstance(img, dict) else (idx == 0),
            })

            if new_name != old_name:
                rename_pairs.append((old_name, new_name))

        return {
            'planned_images': planned_images,
            'rename_pairs': rename_pairs,
            'missing': missing,
            'not_ready': not_ready,
        }

    def _execute_fabric_rename_pairs(self, rename_pairs):
        """安全执行批量重命名，避免目标名冲突（两阶段：先临时名，再目标名）"""
        if not rename_pairs:
            return {'status': 'success', 'rollback_pairs': []}

        folder = self._ensure_fabric_folder()
        normalized = []
        seen_src = set()
        seen_dst = set()
        for src_name, dst_name in rename_pairs:
            src = (src_name or '').strip()
            dst = (dst_name or '').strip()
            if not src or not dst or src == dst:
                continue
            if src in seen_src:
                return {'status': 'error', 'message': f'重复源文件: {src}'}
            if dst in seen_dst:
                return {'status': 'error', 'message': f'目标文件名冲突: {dst}'}
            seen_src.add(src)
            seen_dst.add(dst)
            normalized.append((src, dst))

        if not normalized:
            return {'status': 'success', 'rollback_pairs': []}

        src_set = {src for src, _ in normalized}
        for src, dst in normalized:
            src_path = os.path.join(folder, self._safe_fsencode(src))
            dst_path = os.path.join(folder, self._safe_fsencode(dst))
            if not os.path.exists(src_path):
                return {'status': 'error', 'message': f'源文件不存在: {src}'}
            if dst not in src_set and os.path.exists(dst_path):
                return {'status': 'error', 'message': f'目标文件已存在: {dst}'}

        temp_pairs = []
        for index, (src, dst) in enumerate(normalized):
            token = secrets.token_hex(6)
            temp_name = f".__sitjoy_tmp__{token}_{index}"
            while os.path.exists(os.path.join(folder, self._safe_fsencode(temp_name))):
                token = secrets.token_hex(6)
                temp_name = f".__sitjoy_tmp__{token}_{index}"
            temp_pairs.append((src, temp_name, dst))

        moved_to_temp = []
        moved_to_final = []
        try:
            for src, temp_name, _ in temp_pairs:
                src_path = os.path.join(folder, self._safe_fsencode(src))
                temp_path = os.path.join(folder, self._safe_fsencode(temp_name))
                os.rename(src_path, temp_path)
                moved_to_temp.append((src, temp_name))

            for src, temp_name, dst in temp_pairs:
                temp_path = os.path.join(folder, self._safe_fsencode(temp_name))
                dst_path = os.path.join(folder, self._safe_fsencode(dst))
                os.rename(temp_path, dst_path)
                moved_to_final.append((src, dst))

            rollback_pairs = [(dst, src) for src, dst in reversed(moved_to_final)]
            return {'status': 'success', 'rollback_pairs': rollback_pairs}
        except Exception as e:
            try:
                final_map = {dst: src for src, dst in moved_to_final}
                for _, dst in reversed(moved_to_final):
                    dst_path = os.path.join(folder, self._safe_fsencode(dst))
                    src = final_map.get(dst)
                    if src and os.path.exists(dst_path):
                        os.rename(dst_path, os.path.join(folder, self._safe_fsencode(src)))
            except Exception:
                pass

            try:
                for src, temp_name in reversed(moved_to_temp):
                    temp_path = os.path.join(folder, self._safe_fsencode(temp_name))
                    src_path = os.path.join(folder, self._safe_fsencode(src))
                    if os.path.exists(temp_path):
                        os.rename(temp_path, src_path)
            except Exception:
                pass

            return {'status': 'error', 'message': f'文件重命名失败: {str(e)}'}
    
    def handle_images_api(self, environ, start_response):
        """获取图片列表（用Base64编码路径避免编码问题）"""
        images = []
        try:
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)
            
            page = int(query_params.get('page', ['1'])[0])
            per_page = min(int(query_params.get('per_page', ['100'])[0]), 200)
            
            # 检查RESOURCES_PATH是否存在
            if not os.path.exists(RESOURCES_PATH_BYTES):
                # 列出/volume1/下的文件夹帮助调试
                try:
                    volume_contents = os.listdir('/volume1') if os.path.exists('/volume1') else []
                    folders_list = [f for f in volume_contents if os.path.isdir(f'/volume1/{f}')]
                    # 用Base64编码文件夹列表以避免编码问题
                    try:
                        folders_b64 = base64.b64encode(str(folders_list).encode('utf-8', errors='surrogatepass')).decode('ascii')
                    except Exception:
                        folders_b64 = base64.b64encode(str(folders_list).encode('utf-8', errors='ignore')).decode('ascii')
                    return self.send_json({
                        'status': 'error', 
                        'message': 'Path not found',
                        'available_folders_b64': folders_b64
                    }, start_response)
                except:
                    return self.send_json({
                        'status': 'error', 
                        'message': f'Path not found and cannot list volume'
                    }, start_response)
            
            # 扫描文件
            count = 0
            for root, dirs, files in os.walk(RESOURCES_PATH_BYTES):
                for file in files:
                    if self._is_image_name(file):
                        try:
                            full_path = os.path.join(root, file)
                            rel_path = os.path.relpath(full_path, RESOURCES_PATH_BYTES)
                            
                            # 用Base64编码所有内容（保留文件系统原始字节）
                            path_b64 = self._b64_from_fs(rel_path)
                            filename_b64 = self._b64_from_fs(file)
                            
                            # folder也编码，完全避免中文
                            folder = os.path.dirname(rel_path) or b'root'
                            folder_b64 = self._b64_from_fs(folder)
                            
                            images.append({
                                'id': path_b64,
                                'filename': filename_b64,
                                'folder': folder_b64
                            })
                            count += 1
                        except Exception as e:
                            print(f"File error: {type(e).__name__}")
                            pass
            
            # 分页
            total = len(images)
            start_idx = (page - 1) * per_page
            end_idx = start_idx + per_page
            paginated = images[start_idx:end_idx]
            
            # 计算总页数
            import math
            total_pages = math.ceil(total / per_page) if total > 0 else 1
            
            # 完全ASCII的响应
            resp = {
                'status': 'success',
                'total': total,
                'page': page,
                'pages': total_pages,
                'count': len(paginated),
                'images': paginated
            }
            return self.send_json(resp, start_response)
        except Exception as e:
            print(f"Exception in handle_images_api: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            # 返回错误时，消息也要清理，不含中文
            return self.send_json({
                'status': 'error', 
                'message': f'Error: {type(e).__name__}'
            }, start_response)
    
    def handle_browse_api(self, environ, start_response):
        """浏览目录API：返回指定目录下的文件夹和图片"""
        try:
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)
            path_b64 = query_params.get('path', [''])[0]
            debug = query_params.get('debug', ['0'])[0] == '1'

            # 解码路径（如果为空则为根目录）
            if path_b64:
                try:
                    rel_path = self._fs_from_b64(path_b64)
                except:
                    return self.send_json({'status': 'error', 'message': 'Invalid path'}, start_response)
            else:
                rel_path = ''

            # 防止路径遍历
            if '..' in rel_path:
                return self.send_json({'status': 'error', 'message': 'Invalid path'}, start_response)

            # 构建完整路径（bytes）
            current_path = self._join_resources(rel_path)

            # 验证路径安全性
            abs_path = os.path.abspath(current_path)
            abs_resources = os.path.abspath(RESOURCES_PATH_BYTES)
            if not abs_path.startswith(abs_resources):
                return self.send_json({'status': 'error', 'message': 'Access denied'}, start_response)

            if not os.path.exists(current_path):
                return self.send_json({'status': 'error', 'message': 'Path not found'}, start_response)

            folders = []
            images = []

            try:
                debug_items = []
                rel_path_bytes = os.fsencode(rel_path) if rel_path else b''
                with os.scandir(current_path) as it:
                    for entry in it:
                        try:
                            item = entry.name
                            item_bytes = item if isinstance(item, (bytes, bytearray)) else os.fsencode(item)

                            # 跳过系统文件夹
                            if item_bytes.startswith(b'@') or item_bytes.startswith(b'.'):
                                if debug:
                                    debug_items.append({
                                        'name': self._b64_from_fs(item),
                                        'skipped': 'system',
                                        'is_dir': entry.is_dir(follow_symlinks=False),
                                        'is_file': entry.is_file(follow_symlinks=False)
                                    })
                                continue

                            if entry.is_dir(follow_symlinks=False):
                                folder_rel_path = os.path.join(rel_path_bytes, item_bytes) if rel_path_bytes else item_bytes
                                folders.append({
                                    'name': self._b64_from_fs(item_bytes),
                                    'path': self._b64_from_fs(folder_rel_path),
                                    'type': 'folder'
                                })
                            elif entry.is_file(follow_symlinks=False):
                                if self._is_image_name(item_bytes):
                                    image_rel_path = os.path.join(rel_path_bytes, item_bytes) if rel_path_bytes else item_bytes
                                    images.append({
                                        'name': self._b64_from_fs(item_bytes),
                                        'path': self._b64_from_fs(image_rel_path),
                                        'type': 'image'
                                    })
                                elif debug:
                                    debug_items.append({
                                        'name': self._b64_from_fs(item_bytes),
                                        'skipped': 'not_image',
                                        'is_dir': False,
                                        'is_file': True
                                    })
                            elif debug:
                                debug_items.append({
                                    'name': self._b64_from_fs(item_bytes),
                                    'skipped': 'unknown_type',
                                    'is_dir': entry.is_dir(follow_symlinks=False),
                                    'is_file': entry.is_file(follow_symlinks=False)
                                })
                        except Exception as e:
                            print(f"Item error: {type(e).__name__}")
                            if debug:
                                debug_items.append({
                                    'name': 'unknown',
                                    'skipped': f'error:{type(e).__name__}'
                                })
                            pass
            except Exception as e:
                return self.send_json({'status': 'error', 'message': f'Cannot read directory: {type(e).__name__}'}, start_response)

            folders.sort(key=lambda x: x['name'])
            images.sort(key=lambda x: x['name'])

            breadcrumbs = []
            if rel_path:
                rel_path_bytes = os.fsencode(rel_path)
                parts = rel_path_bytes.split(b'/')
                current = b''
                for part in parts:
                    current = os.path.join(current, part) if current else part
                    breadcrumbs.append({
                        'name': self._b64_from_fs(part),
                        'path': self._b64_from_fs(current)
                    })

            resp = {
                'status': 'success',
                'current_path': path_b64,
                'breadcrumbs': breadcrumbs,
                'folders': folders,
                'images': images,
                'total_folders': len(folders),
                'total_images': len(images)
            }

            if debug:
                resp['debug_items'] = debug_items

            return self.send_json(resp, start_response)

        except Exception as e:
            print(f"Browse error: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            return self.send_json({'status': 'error', 'message': f'Error: {type(e).__name__}'}, start_response)
    
    def handle_image_preview(self, environ, start_response):
        """获取图片预览（接受Base64编码的路径）"""
        try:
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)
            path_b64 = query_params.get('id', [''])[0]
            mode = (query_params.get('mode', [''])[0] or '').strip().lower()
            max_w = self._to_int(query_params.get('w', [''])[0], 0) or 0
            max_h = self._to_int(query_params.get('h', [''])[0], 0) or 0
            quality = self._to_int(query_params.get('q', [''])[0], 0) or 0
            use_compressed = mode in ('thumb', 'compressed') or max_w > 0 or max_h > 0 or quality > 0
            
            if not path_b64:
                return self.send_error(400, 'Missing id parameter', start_response)
            
            # 解码Base64路径。前端可能对文件名做了 UTF-8 编码再 base64，
            # 也可能对文件系统原始 bytes 做 base64。优先尝试使用原始 bytes 直接拼接路径。
            try:
                raw = base64.b64decode(path_b64)
            except Exception:
                return self.send_error(400, 'Invalid id', start_response)

            full_path = None
            # 1) 尝试将 raw 作为相对 bytes 路径直接拼接并检查
            try:
                candidate = os.path.join(RESOURCES_PATH_BYTES, raw)
                abs_candidate = os.path.abspath(candidate)
                abs_resources = os.path.abspath(RESOURCES_PATH_BYTES)
                if abs_candidate.startswith(abs_resources) and os.path.exists(candidate):
                    full_path = candidate
            except Exception:
                full_path = None

            # 2) 回退：把 raw 解为字符串（filesystem decode）再拼接
            if full_path is None:
                try:
                    rel_path = os.fsdecode(raw)
                except Exception:
                    try:
                        rel_path = raw.decode('utf-8', errors='surrogatepass')
                    except Exception:
                        return self.send_error(400, 'Invalid id', start_response)

                # 防止路径遍历
                if '..' in rel_path or rel_path.startswith('/'):
                    return self.send_error(403, 'Invalid path', start_response)

                full_path = self._join_resources(rel_path)
            
            # 验证路径安全性并存在性
            try:
                abs_path = os.path.abspath(full_path)
                abs_resources = os.path.abspath(RESOURCES_PATH_BYTES)
                if not abs_path.startswith(abs_resources):
                    return self.send_error(403, 'Access denied', start_response)
            except Exception:
                return self.send_error(403, 'Access denied', start_response)

            if not os.path.exists(full_path):
                return self.send_error(404, 'File not found', start_response)
            
            # 读取图片
            mime_path = os.fsdecode(full_path) if isinstance(full_path, (bytes, bytearray)) else full_path
            mime_type, _ = mimetypes.guess_type(mime_path)
            if not mime_type:
                mime_type = 'image/jpeg'

            if use_compressed and Image:
                try:
                    img = Image.open(full_path)
                    if max_w <= 0 and max_h <= 0:
                        max_w, max_h = 360, 360
                    max_w = max(1, max_w) if max_w > 0 else 360
                    max_h = max(1, max_h) if max_h > 0 else 360
                    img.thumbnail((max_w, max_h), Image.Resampling.LANCZOS)

                    if quality <= 0:
                        quality = 72
                    quality = max(35, min(90, quality))

                    output = io.BytesIO()
                    if img.mode not in ('RGB', 'L'):
                        img = img.convert('RGB')
                    img.save(output, format='JPEG', quality=quality, optimize=True)
                    image_data = output.getvalue()
                    mime_type = 'image/jpeg'
                except Exception:
                    with open(full_path, 'rb') as f:
                        image_data = f.read()
            else:
                with open(full_path, 'rb') as f:
                    image_data = f.read()
            
            start_response('200 OK', [
                ('Content-Type', mime_type),
                ('Content-Length', str(len(image_data))),
                ('Cache-Control', 'public, max-age=300')
            ])
            
            return [image_data]
                    
        except Exception as e:
            print("Preview error: " + str(e))
            return self.send_error(500, str(e), start_response)
    
    def handle_rename_api(self, environ, start_response):
        """处理文件重命名（接受Base64编码路径）"""
        try:
            if environ['REQUEST_METHOD'] != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0))
            body = environ['wsgi.input'].read(content_length)
            data = json.loads(body.decode('utf-8'))

            path_b64 = data.get('id', '')
            new_name_b64 = data.get('new_name_b64', '')

            if not path_b64:
                return self.send_error(400, 'Missing parameters', start_response)

            # 解码路径和新名称
            try:
                old_path = self._fs_from_b64(path_b64)
                new_name = self._fs_from_b64(new_name_b64) if new_name_b64 else ''
            except:
                return self.send_error(400, 'Invalid parameters', start_response)

            if '..' in old_path or ('..' in new_name if new_name else False):
                return self.send_error(403, 'Invalid path', start_response)

            full_old_path = self._join_resources(old_path)

            # 验证安全性
            abs_path = os.path.abspath(full_old_path)
            abs_resources = os.path.abspath(RESOURCES_PATH_BYTES)
            if not abs_path.startswith(abs_resources):
                return self.send_error(403, 'Access denied', start_response)

            if not os.path.exists(full_old_path):
                return self.send_error(404, 'File not found', start_response)

            # 获取扩展名
            folder = os.path.dirname(full_old_path)
            ext = os.path.splitext(os.path.basename(full_old_path))[1]
            new_name_bytes = os.fsencode(new_name)
            new_filename = new_name_bytes + ext if not new_name_bytes.endswith(ext) else new_name_bytes
            full_new_path = os.path.join(folder, new_filename)

            # 检查新名称是否已存在
            if os.path.exists(full_new_path):
                return self.send_error(409, 'File already exists', start_response)

            # 重命名
            os.rename(full_old_path, full_new_path)

            resp = {
                'status': 'success',
                'message': 'Renamed',
                'new_name': os.fsdecode(new_filename)
            }
            return self.send_json(resp, start_response)
        except Exception as e:
            print("Rename error: " + str(e))
            return self.send_error(500, str(e), start_response)

    def handle_move_api(self, environ, start_response):
        """处理文件移动+重命名（目标仅允许根目录下）"""
        try:
            if environ['REQUEST_METHOD'] != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0))
            body = environ['wsgi.input'].read(content_length)
            data = json.loads(body.decode('utf-8'))

            path_b64 = data.get('id', '')
            new_name_b64 = data.get('new_name_b64', '')
            target_folder_b64 = data.get('target_folder_b64', '')

            if not path_b64 or not new_name_b64:
                return self.send_error(400, 'Missing parameters', start_response)

            try:
                old_path = self._fs_from_b64(path_b64)
                new_name = self._fs_from_b64(new_name_b64)
            except:
                return self.send_error(400, 'Invalid parameters', start_response)

            if '..' in old_path or '..' in new_name:
                return self.send_error(403, 'Invalid path', start_response)

            if target_folder_b64:
                try:
                    target_folder_bytes = base64.b64decode(target_folder_b64)
                except:
                    return self.send_error(400, 'Invalid target folder', start_response)
            else:
                target_folder_bytes = b''

            # 仅允许资源根目录内路径
            if target_folder_bytes.startswith((b'/', b'\\')):
                return self.send_error(403, 'Target folder not allowed', start_response)
            if b'..' in target_folder_bytes.split(b'/') or b'..' in target_folder_bytes.split(b'\\'):
                return self.send_error(403, 'Target folder not allowed', start_response)

            full_old_path = self._join_resources(old_path)

            abs_old = os.path.abspath(full_old_path)
            abs_resources = os.path.abspath(RESOURCES_PATH_BYTES)
            if not abs_old.startswith(abs_resources):
                return self.send_error(403, 'Access denied', start_response)

            if not os.path.exists(full_old_path):
                return self.send_error(404, 'File not found', start_response)

            dest_dir = os.path.join(RESOURCES_PATH_BYTES, target_folder_bytes) if target_folder_bytes else os.path.dirname(full_old_path)
            if not os.path.exists(dest_dir) or not os.path.isdir(dest_dir):
                return self.send_error(404, 'Target folder not found', start_response)

            old_basename = os.path.basename(full_old_path)
            ext = os.path.splitext(old_basename)[1]
            if new_name:
                new_name_bytes = os.fsencode(new_name)
                new_filename = new_name_bytes + ext if not new_name_bytes.endswith(ext) else new_name_bytes
            else:
                new_filename = old_basename
            full_new_path = os.path.join(dest_dir, new_filename)

            if os.path.abspath(full_new_path) == os.path.abspath(full_old_path):
                return self.send_error(400, 'No changes', start_response)

            if os.path.exists(full_new_path):
                return self.send_error(409, 'File already exists', start_response)

            os.rename(full_old_path, full_new_path)

            resp = {
                'status': 'success',
                'message': 'Moved',
                'new_name': os.fsdecode(new_filename)
            }
            return self.send_json(resp, start_response)
        except Exception as e:
            print("Move error: " + str(e))
            return self.send_error(500, str(e), start_response)

    def handle_upload_api(self, environ, start_response):
        """处理图片上传（multipart/form-data）"""
        try:
            if environ['REQUEST_METHOD'] != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            form = cgi.FieldStorage(fp=environ['wsgi.input'], environ=environ, keep_blank_values=True)
            path_b64 = form.getfirst('path', '')

            if path_b64:
                try:
                    rel_path = self._fs_from_b64(path_b64)
                except Exception:
                    return self.send_json({'status': 'error', 'message': 'Invalid path'}, start_response)
            else:
                rel_path = ''

            if '..' in rel_path:
                return self.send_json({'status': 'error', 'message': 'Invalid path'}, start_response)

            target_dir = self._join_resources(rel_path)

            abs_target = os.path.abspath(target_dir)
            abs_resources = os.path.abspath(RESOURCES_PATH_BYTES)
            if not abs_target.startswith(abs_resources):
                return self.send_json({'status': 'error', 'message': 'Access denied'}, start_response)

            if not os.path.exists(target_dir):
                return self.send_json({'status': 'error', 'message': 'Path not found'}, start_response)

            if 'file' not in form:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)

            files_field = form['file']
            if isinstance(files_field, list):
                files_list = files_field
            else:
                files_list = [files_field]

            saved = []
            skipped = []
            for item in files_list:
                try:
                    if not item.filename:
                        continue

                    filename = os.path.basename(item.filename)
                    if not self._is_image_name(filename):
                        skipped.append({'name': str(filename), 'reason': 'not_image'})
                        continue

                    try:
                        filename_bytes = os.fsencode(filename)
                    except Exception:
                        filename_bytes = str(filename).encode('utf-8', errors='surrogatepass')

                    dest_path = os.path.join(target_dir, filename_bytes)
                    if os.path.exists(dest_path):
                        skipped.append({'name': str(filename), 'reason': 'exists'})
                        continue

                    with open(dest_path, 'wb') as f:
                        while True:
                            chunk = item.file.read(1024 * 1024)
                            if not chunk:
                                break
                            f.write(chunk)

                    saved.append(str(filename))
                except Exception as e:
                    skipped.append({'name': str(getattr(item, 'filename', 'unknown')), 'reason': str(e)})

            return self.send_json({'status': 'success', 'count': len(saved), 'files': saved, 'skipped': skipped}, start_response)
        except Exception as e:
            print("Upload error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_download_zip(self, environ, method, start_response):
        """将选中图片/文件夹打包为 zip 下载"""
        try:
            if method != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)

            data = self._read_json_body(environ)
            items = data.get('items', []) if isinstance(data, dict) else []
            if not items:
                return self.send_json({'status': 'error', 'message': 'No items selected'}, start_response)

            resources_bytes = RESOURCES_PATH_BYTES
            files = set()

            for item in items:
                path_b64 = item.get('path', '') if isinstance(item, dict) else ''
                if not path_b64:
                    continue
                try:
                    rel_path = self._fs_from_b64(path_b64)
                except Exception:
                    continue
                if '..' in rel_path or rel_path.startswith('/'):
                    continue

                full_path = self._join_resources(rel_path)
                abs_path = os.path.abspath(full_path)
                abs_resources = os.path.abspath(RESOURCES_PATH_BYTES)
                if not abs_path.startswith(abs_resources):
                    continue

                if os.path.isdir(full_path):
                    for root, _, filenames in os.walk(full_path):
                        for name in filenames:
                            if not self._is_image_name(name):
                                continue
                            files.add(os.path.join(root, name))
                elif os.path.isfile(full_path):
                    if self._is_image_name(full_path):
                        files.add(full_path)

            if not files:
                return self.send_json({'status': 'error', 'message': 'No images found'}, start_response)

            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
            tmp_path = tmp.name
            tmp.close()

            with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for file_path in files:
                    file_bytes = file_path if isinstance(file_path, (bytes, bytearray)) else os.fsencode(file_path)
                    try:
                        rel_bytes = os.path.relpath(file_bytes, resources_bytes)
                    except Exception:
                        rel_bytes = os.path.basename(file_bytes)
                    if rel_bytes.startswith(b'..'):
                        continue
                    arcname = rel_bytes.decode('utf-8', errors='replace').replace('\\', '/')
                    zf.write(os.fsdecode(file_bytes), arcname)

            with open(tmp_path, 'rb') as f:
                data_bytes = f.read()

            try:
                os.remove(tmp_path)
            except Exception:
                pass

            filename = f"sitjoy_download_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            start_response('200 OK', [
                ('Content-Type', 'application/zip'),
                ('Content-Disposition', f'attachment; filename="{filename}"'),
                ('Content-Length', str(len(data_bytes)))
            ])
            return [data_bytes]
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _read_json_body(self, environ):
        """读取请求 JSON body"""
        content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
        if content_length <= 0:
            return {}
        body = environ['wsgi.input'].read(content_length)
        if not body:
            return {}
        return json.loads(body.decode('utf-8'))

    def _send_excel_workbook(self, workbook, filename, start_response):
        output = io.BytesIO()
        workbook.save(output)
        data = output.getvalue()
        start_response('200 OK', [
            ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Disposition', f'attachment; filename="{filename}"'),
            ('Content-Length', str(len(data)))
        ])
        return [data]

    def _get_db_config(self):
        """从环境变量读取数据库配置"""
        config = {
            'host': os.environ.get('SITJOY_DB_HOST', '127.0.0.1'),
            'user': os.environ.get('SITJOY_DB_USER', 'root'),
            'password': os.environ.get('SITJOY_DB_PASSWORD', ''),
            'database': os.environ.get('SITJOY_DB_NAME', 'sitjoy'),
            'port': int(os.environ.get('SITJOY_DB_PORT', '3306')),
            'charset': 'utf8mb4'
        }
        # 读取本地配置文件（若存在则覆盖）
        file_cfg = self._load_local_db_config()
        if file_cfg:
            for key in ['host', 'user', 'password', 'database', 'port', 'charset']:
                if key in file_cfg and file_cfg[key] not in (None, ''):
                    if key == 'port':
                        try:
                            config[key] = int(file_cfg[key])
                        except Exception:
                            continue
                    else:
                        config[key] = file_cfg[key]
        return config

    def _load_local_db_config(self):
        """读取项目内 db_config.json（可选）"""
        try:
            cfg_path = os.path.join(self.base_path, 'db_config.json')
            if not os.path.exists(cfg_path):
                return None
            with open(cfg_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return None

    def _get_db_connection(self):
        if not pymysql:
            raise RuntimeError(f"PyMySQL not available: {_pymysql_import_error}")
        cfg = self._get_db_config()
        return pymysql.connect(
            host=cfg['host'],
            user=cfg['user'],
            password=cfg['password'],
            database=cfg['database'],
            port=cfg['port'],
            charset=cfg['charset'],
            cursorclass=pymysql.cursors.DictCursor,
            autocommit=True
        )

    def _ensure_product_table(self):
        if self._db_ready:
            return
        create_sql = """
        CREATE TABLE IF NOT EXISTS product_families (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            sku_family VARCHAR(64) NOT NULL UNIQUE,
            category VARCHAR(64) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        try:
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(create_sql)
            self._db_ready = True
        except Exception as e:
            self._db_ready = False
            raise e

    def _ensure_category_table(self):
        create_sql = """
        CREATE TABLE IF NOT EXISTS product_categories (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            category_cn VARCHAR(64) NOT NULL,
            category_en VARCHAR(64) NOT NULL,
            category_en_name VARCHAR(128) NOT NULL DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_category_cn (category_cn),
            UNIQUE KEY uniq_category_en (category_en)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'product_categories'
                      AND COLUMN_NAME = 'category_en_name'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    cur.execute("ALTER TABLE product_categories ADD COLUMN category_en_name VARCHAR(128) NOT NULL DEFAULT ''")

    def _ensure_fabric_table(self):
        self._ensure_materials_table()
        self._ensure_product_table()
        create_sql = """
        CREATE TABLE IF NOT EXISTS fabric_materials (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            fabric_code VARCHAR(64) NOT NULL UNIQUE,
            fabric_name_en VARCHAR(128) NOT NULL,
            representative_color VARCHAR(7) NULL,
            material_id INT UNSIGNED NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_fabric_material (material_id),
            CONSTRAINT fk_fabric_material FOREIGN KEY (material_id)
                REFERENCES materials(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        create_images_sql = """
        CREATE TABLE IF NOT EXISTS fabric_images (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            fabric_id INT UNSIGNED NOT NULL,
            image_name VARCHAR(255) NOT NULL,
            sort_order INT UNSIGNED NOT NULL DEFAULT 0,
            is_primary TINYINT(1) NOT NULL DEFAULT 0,
            remark VARCHAR(50) NULL DEFAULT NULL COMMENT '备注类型：平面原图/褶皱原图/卖点图',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_fabric_images_fabric (fabric_id),
            INDEX idx_fabric_images_primary (fabric_id, is_primary),
            CONSTRAINT fk_fabric_images_fabric FOREIGN KEY (fabric_id)
                REFERENCES fabric_materials(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        create_fabric_sku_relation = """
        CREATE TABLE IF NOT EXISTS fabric_product_families (
            fabric_id INT UNSIGNED NOT NULL,
            sku_family_id INT UNSIGNED NOT NULL,
            PRIMARY KEY (fabric_id, sku_family_id),
            CONSTRAINT fk_fpf_fabric FOREIGN KEY (fabric_id)
                REFERENCES fabric_materials(id) ON DELETE CASCADE,
            CONSTRAINT fk_fpf_sku_family FOREIGN KEY (sku_family_id)
                REFERENCES product_families(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
                cur.execute(create_images_sql)
                cur.execute(create_fabric_sku_relation)
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'fabric_materials'
                      AND COLUMN_NAME = 'material_id'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    cur.execute("ALTER TABLE fabric_materials ADD COLUMN material_id INT UNSIGNED NULL")
                    try:
                        cur.execute("ALTER TABLE fabric_materials ADD INDEX idx_fabric_material (material_id)")
                    except Exception:
                        pass
                    try:
                        cur.execute(
                            """
                            ALTER TABLE fabric_materials
                            ADD CONSTRAINT fk_fabric_material
                            FOREIGN KEY (material_id) REFERENCES materials(id)
                            ON DELETE SET NULL
                            """
                        )
                    except Exception:
                        pass
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'fabric_materials'
                      AND COLUMN_NAME = 'representative_color'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    cur.execute("ALTER TABLE fabric_materials ADD COLUMN representative_color VARCHAR(7) NULL AFTER fabric_name_en")
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'fabric_materials'
                      AND COLUMN_NAME = 'image_name'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) > 0:
                    cur.execute(
                        """
                        INSERT INTO fabric_images (fabric_id, image_name, sort_order, is_primary)
                        SELECT fm.id, fm.image_name, 0, 1
                        FROM fabric_materials fm
                        LEFT JOIN fabric_images fi
                            ON fi.fabric_id = fm.id AND fi.image_name = fm.image_name
                        WHERE fm.image_name IS NOT NULL AND fm.image_name <> ''
                          AND fi.id IS NULL
                        """
                    )
                    try:
                        cur.execute("ALTER TABLE fabric_materials DROP COLUMN image_name")
                    except Exception:
                        pass
                
                # 添加 remark 字段用于图片类型标注
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'fabric_images'
                      AND COLUMN_NAME = 'remark'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    cur.execute(
                        """
                        ALTER TABLE fabric_images
                        ADD COLUMN remark VARCHAR(50) NULL DEFAULT NULL COMMENT '备注类型：平面原图/褶皱原图/卖点图'
                        AFTER is_primary
                        """
                    )
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'fabric_images'
                      AND COLUMN_NAME = 'sort_order'
                    """
                )
                row = cur.fetchone() or {}
                if int(row.get('cnt') or 0) == 0:
                    cur.execute("ALTER TABLE fabric_images ADD COLUMN sort_order INT UNSIGNED NOT NULL DEFAULT 0 AFTER image_name")
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'fabric_images'
                      AND COLUMN_NAME = 'is_primary'
                    """
                )
                row = cur.fetchone() or {}
                if int(row.get('cnt') or 0) == 0:
                    cur.execute("ALTER TABLE fabric_images ADD COLUMN is_primary TINYINT(1) NOT NULL DEFAULT 0 AFTER sort_order")
                try:
                    cur.execute("ALTER TABLE fabric_images ADD INDEX idx_fabric_images_primary (fabric_id, is_primary)")
                except Exception:
                    pass

    def _ensure_material_types_table(self):
        if self._material_types_ready:
            return
        create_sql = """
        CREATE TABLE IF NOT EXISTS material_types (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(64) NOT NULL UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
        self._material_types_ready = True

    def _ensure_materials_table(self):
        if self._materials_ready:
            return
        self._ensure_material_types_table()
        type_map = {
            'fabric': '面料',
            'filling': '填充',
            'frame': '框架',
            'electronics': '电子元器件'
        }
        create_materials = """
        CREATE TABLE IF NOT EXISTS materials (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(128) NOT NULL,
            name_en VARCHAR(128) NOT NULL DEFAULT '',
            material_type_id INT UNSIGNED NOT NULL,
            parent_id INT UNSIGNED NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_material (material_type_id, name),
            INDEX idx_material_type_id (material_type_id),
            INDEX idx_material_parent (parent_id),
            CONSTRAINT fk_material_type FOREIGN KEY (material_type_id)
                REFERENCES material_types(id) ON DELETE RESTRICT,
            CONSTRAINT fk_material_parent FOREIGN KEY (parent_id)
                REFERENCES materials(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_materials)
                cur.execute("SELECT COUNT(*) AS cnt FROM material_types")
                type_count = cur.fetchone()
                if type_count and type_count.get('cnt', 0) == 0:
                    for name in type_map.values():
                        cur.execute("INSERT IGNORE INTO material_types (name) VALUES (%s)", (name,))
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'materials'
                      AND COLUMN_NAME = 'name_en'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    cur.execute("ALTER TABLE materials ADD COLUMN name_en VARCHAR(128) NOT NULL DEFAULT ''")
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'materials'
                      AND COLUMN_NAME = 'material_type_id'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    cur.execute("ALTER TABLE materials ADD COLUMN material_type_id INT UNSIGNED NULL")
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'materials'
                      AND COLUMN_NAME = 'parent_id'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    cur.execute("ALTER TABLE materials ADD COLUMN parent_id INT UNSIGNED NULL")
                    try:
                        cur.execute("ALTER TABLE materials ADD INDEX idx_material_parent (parent_id)")
                    except Exception:
                        pass
                    try:
                        cur.execute(
                            """
                            ALTER TABLE materials
                            ADD CONSTRAINT fk_material_parent
                            FOREIGN KEY (parent_id) REFERENCES materials(id)
                            ON DELETE SET NULL
                            """
                        )
                    except Exception:
                        pass
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'materials'
                      AND COLUMN_NAME = 'material_type'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) > 0:
                    try:
                        for code, name in type_map.items():
                            cur.execute(
                                """
                                UPDATE materials m
                                JOIN material_types mt ON mt.name = %s
                                SET m.material_type_id = mt.id
                                WHERE m.material_type_id IS NULL AND m.material_type = %s
                                """,
                                (name, code)
                            )
                    except Exception:
                        pass
                    cur.execute("SELECT COUNT(*) AS cnt FROM materials WHERE material_type_id IS NULL")
                    missing = cur.fetchone()
                    if missing and missing.get('cnt', 0) == 0:
                        try:
                            cur.execute("ALTER TABLE materials MODIFY material_type_id INT UNSIGNED NOT NULL")
                        except Exception:
                            pass
                        try:
                            cur.execute("ALTER TABLE materials ADD UNIQUE KEY uniq_material (material_type_id, name)")
                        except Exception:
                            pass
                        try:
                            cur.execute("ALTER TABLE materials ADD INDEX idx_material_type_id (material_type_id)")
                        except Exception:
                            pass
                        try:
                            cur.execute(
                                """
                                ALTER TABLE materials
                                ADD CONSTRAINT fk_material_type
                                FOREIGN KEY (material_type_id) REFERENCES material_types(id)
                                ON DELETE RESTRICT
                                """
                            )
                        except Exception:
                            pass
        self._materials_ready = True

    def _ensure_platform_types_table(self):
        if self._platform_types_ready:
            return
        create_sql = """
        CREATE TABLE IF NOT EXISTS platform_types (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(64) NOT NULL UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
        self._platform_types_ready = True

    def _ensure_brands_table(self):
        if self._brands_ready:
            return
        create_sql = """
        CREATE TABLE IF NOT EXISTS brands (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(128) NOT NULL UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
        self._brands_ready = True

    def _ensure_shops_table(self):
        if self._shops_ready:
            return
        self._ensure_platform_types_table()
        self._ensure_brands_table()
        create_sql = """
        CREATE TABLE IF NOT EXISTS shops (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            shop_name VARCHAR(128) NOT NULL,
            platform_type_id INT UNSIGNED NOT NULL,
            brand_id INT UNSIGNED NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_shop (shop_name, platform_type_id, brand_id),
            INDEX idx_shop_platform (platform_type_id),
            INDEX idx_shop_brand (brand_id),
            CONSTRAINT fk_shop_platform_type FOREIGN KEY (platform_type_id)
                REFERENCES platform_types(id) ON DELETE RESTRICT,
            CONSTRAINT fk_shop_brand FOREIGN KEY (brand_id)
                REFERENCES brands(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
        self._shops_ready = True

    def _ensure_amazon_account_health_table(self):
        if self._amazon_account_health_ready:
            return
        self._ensure_shops_table()
        create_sql = """
        CREATE TABLE IF NOT EXISTS amazon_account_health (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            shop_id INT UNSIGNED NOT NULL,
            account_health_rating INT NOT NULL,
            suspected_ip_infringement INT NOT NULL DEFAULT 0,
            intellectual_property_complaints INT NOT NULL DEFAULT 0,
            authenticity_customer_complaints INT NOT NULL DEFAULT 0,
            condition_customer_complaints INT NOT NULL DEFAULT 0,
            food_safety_issues INT NOT NULL DEFAULT 0,
            listing_policy_violations INT NOT NULL DEFAULT 0,
            restricted_product_policy_violations INT NOT NULL DEFAULT 0,
            customer_review_policy_violations INT NOT NULL DEFAULT 0,
            other_policy_violations INT NOT NULL DEFAULT 0,
            regulatory_compliance_issues INT NOT NULL DEFAULT 0,
            order_defect_rate DECIMAL(8,4) NOT NULL DEFAULT 0,
            negative_feedback_rate DECIMAL(8,4) NOT NULL DEFAULT 0,
            a_to_z_rate DECIMAL(8,4) NOT NULL DEFAULT 0,
            chargeback_rate DECIMAL(8,4) NOT NULL DEFAULT 0,
            late_shipment_rate DECIMAL(8,4) NOT NULL DEFAULT 0,
            pre_fulfillment_cancel_rate DECIMAL(8,4) NOT NULL DEFAULT 0,
            valid_tracking_rate DECIMAL(8,4) NOT NULL DEFAULT 0,
            on_time_delivery_rate DECIMAL(8,4) NOT NULL DEFAULT 0,
            record_datetime DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            remark VARCHAR(500) NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_aah_shop_date (shop_id, record_datetime),
            INDEX idx_aah_record_datetime (record_datetime),
            CONSTRAINT fk_aah_shop FOREIGN KEY (shop_id)
                REFERENCES shops(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
        self._amazon_account_health_ready = True

    def _ensure_amazon_ad_subtypes_table(self):
        if self._amazon_ad_subtypes_ready:
            return
        self._ensure_amazon_ad_operation_types_table()
        create_sql = """
        CREATE TABLE IF NOT EXISTS amazon_ad_subtypes (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            description VARCHAR(255) NOT NULL,
            ad_class VARCHAR(8) NOT NULL DEFAULT 'SP',
            subtype_code VARCHAR(64) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_ad_subtype (ad_class, subtype_code)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        relation_sql = """
        CREATE TABLE IF NOT EXISTS amazon_ad_subtype_operation_types (
            subtype_id INT UNSIGNED NOT NULL,
            operation_type_id INT UNSIGNED NOT NULL,
            PRIMARY KEY (subtype_id, operation_type_id),
            CONSTRAINT fk_ad_subtype_op_subtype FOREIGN KEY (subtype_id)
                REFERENCES amazon_ad_subtypes(id) ON DELETE CASCADE,
            CONSTRAINT fk_ad_subtype_op_type FOREIGN KEY (operation_type_id)
                REFERENCES amazon_ad_operation_types(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
                cur.execute(relation_sql)
        self._amazon_ad_subtypes_ready = True

    def _ensure_amazon_ad_operation_types_table(self):
        if self._amazon_ad_operation_types_ready:
            return
        create_sql = """
        CREATE TABLE IF NOT EXISTS amazon_ad_operation_types (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(128) NOT NULL UNIQUE,
            apply_portfolio TINYINT(1) NOT NULL DEFAULT 1,
            apply_campaign TINYINT(1) NOT NULL DEFAULT 1,
            apply_group TINYINT(1) NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        create_reason_sql = """
        CREATE TABLE IF NOT EXISTS amazon_ad_operation_reasons (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            operation_type_id INT UNSIGNED NOT NULL,
            reason_name VARCHAR(255) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_ad_op_reason (operation_type_id, reason_name),
            INDEX idx_ad_op_reason_type (operation_type_id),
            CONSTRAINT fk_ad_op_reason_type FOREIGN KEY (operation_type_id)
                REFERENCES amazon_ad_operation_types(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'amazon_ad_operation_types'
                      AND COLUMN_NAME = 'apply_portfolio'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    try:
                        cur.execute("ALTER TABLE amazon_ad_operation_types ADD COLUMN apply_portfolio TINYINT(1) NOT NULL DEFAULT 1")
                    except Exception as e:
                        if pymysql and isinstance(e, pymysql.err.OperationalError) and getattr(e, 'args', [None])[0] == 1060:
                            pass
                        else:
                            raise
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'amazon_ad_operation_types'
                      AND COLUMN_NAME = 'apply_campaign'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    try:
                        cur.execute("ALTER TABLE amazon_ad_operation_types ADD COLUMN apply_campaign TINYINT(1) NOT NULL DEFAULT 1")
                    except Exception as e:
                        if pymysql and isinstance(e, pymysql.err.OperationalError) and getattr(e, 'args', [None])[0] == 1060:
                            pass
                        else:
                            raise
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'amazon_ad_operation_types'
                      AND COLUMN_NAME = 'apply_group'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    try:
                        cur.execute("ALTER TABLE amazon_ad_operation_types ADD COLUMN apply_group TINYINT(1) NOT NULL DEFAULT 1")
                    except Exception as e:
                        if pymysql and isinstance(e, pymysql.err.OperationalError) and getattr(e, 'args', [None])[0] == 1060:
                            pass
                        else:
                            raise
                cur.execute(create_reason_sql)
        self._amazon_ad_operation_types_ready = True

    def _ensure_amazon_ad_tables(self):
        if self._amazon_ad_ready:
            return
        self._ensure_product_table()
        self._ensure_category_table()
        self._ensure_amazon_ad_subtypes_table()
        create_sql = """
        CREATE TABLE IF NOT EXISTS amazon_ad_items (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            ad_level VARCHAR(16) NOT NULL,
            sku_family_id INT UNSIGNED NULL,
            portfolio_id INT UNSIGNED NULL,
            campaign_id INT UNSIGNED NULL,
            strategy_code VARCHAR(8) NULL,
            subtype_id INT UNSIGNED NULL,
            name VARCHAR(255) NOT NULL,
            is_shared_budget TINYINT(1) NULL,
            status VARCHAR(16) NULL,
            budget DECIMAL(12,2) NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_ad_level (ad_level),
            INDEX idx_ad_sku (sku_family_id),
            INDEX idx_ad_portfolio (portfolio_id),
            INDEX idx_ad_campaign (campaign_id),
            INDEX idx_ad_subtype (subtype_id),
            CONSTRAINT fk_ad_sku FOREIGN KEY (sku_family_id)
                REFERENCES product_families(id) ON DELETE SET NULL,
            CONSTRAINT fk_ad_portfolio FOREIGN KEY (portfolio_id)
                REFERENCES amazon_ad_items(id) ON DELETE CASCADE,
            CONSTRAINT fk_ad_campaign FOREIGN KEY (campaign_id)
                REFERENCES amazon_ad_items(id) ON DELETE CASCADE,
            CONSTRAINT fk_ad_subtype FOREIGN KEY (subtype_id)
                REFERENCES amazon_ad_subtypes(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
        self._amazon_ad_ready = True

    def _ensure_amazon_ad_delivery_table(self):
        if self._amazon_ad_delivery_ready:
            return
        self._ensure_amazon_ad_tables()
        create_sql = """
        CREATE TABLE IF NOT EXISTS amazon_ad_deliveries (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            status VARCHAR(16) NOT NULL DEFAULT '启动',
            ad_item_id INT UNSIGNED NOT NULL,
            delivery_desc VARCHAR(255) NOT NULL,
            bid_value VARCHAR(32) NULL,
            observe_interval VARCHAR(64) NULL,
            next_observe_at DATETIME NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_ad_delivery_item (ad_item_id),
            INDEX idx_ad_delivery_status (status),
            INDEX idx_ad_delivery_next_observe (next_observe_at),
            CONSTRAINT fk_ad_delivery_item FOREIGN KEY (ad_item_id)
                REFERENCES amazon_ad_items(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
        self._amazon_ad_delivery_ready = True

    def _ensure_amazon_ad_product_table(self):
        if self._amazon_ad_product_ready:
            return
        self._ensure_amazon_ad_tables()
        self._ensure_sales_product_tables()
        create_sql = """
        CREATE TABLE IF NOT EXISTS amazon_ad_products (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            status VARCHAR(16) NOT NULL DEFAULT '启动',
            ad_item_id INT UNSIGNED NOT NULL,
            sales_product_id INT UNSIGNED NOT NULL,
            observe_interval VARCHAR(64) NULL,
            next_observe_at DATETIME NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_ad_product_item (ad_item_id),
            INDEX idx_ad_product_sales (sales_product_id),
            INDEX idx_ad_product_status (status),
            INDEX idx_ad_product_next_observe (next_observe_at),
            CONSTRAINT fk_ad_product_item FOREIGN KEY (ad_item_id)
                REFERENCES amazon_ad_items(id) ON DELETE CASCADE,
            CONSTRAINT fk_ad_product_sales FOREIGN KEY (sales_product_id)
                REFERENCES sales_products(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
        self._amazon_ad_product_ready = True

    def _ensure_amazon_ad_adjustment_table(self):
        if self._amazon_ad_adjustment_ready:
            return
        self._ensure_amazon_ad_tables()
        self._ensure_amazon_ad_operation_types_table()
        create_sql = """
        CREATE TABLE IF NOT EXISTS amazon_ad_adjustments (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            adjust_date DATETIME NOT NULL,
            ad_item_id INT UNSIGNED NOT NULL,
            operation_type_id INT UNSIGNED NOT NULL,
            target_object VARCHAR(255) NOT NULL,
            before_value VARCHAR(64) NULL,
            after_value VARCHAR(64) NULL,
            reason_id INT UNSIGNED NULL,
            start_time DATETIME NULL,
            end_time DATETIME NULL,
            impressions VARCHAR(32) NULL,
            clicks VARCHAR(32) NULL,
            cost VARCHAR(32) NULL,
            orders VARCHAR(32) NULL,
            sales VARCHAR(32) NULL,
            acos VARCHAR(32) NULL,
            cpc VARCHAR(32) NULL,
            ctr VARCHAR(32) NULL,
            cvr VARCHAR(32) NULL,
            attribution_checked TINYINT(1) NOT NULL DEFAULT 0,
            attribution_orders VARCHAR(32) NULL,
            attribution_sales VARCHAR(32) NULL,
            remark VARCHAR(255) NULL,
            is_quick_submit TINYINT(1) NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_ad_adjustment_ad_item (ad_item_id),
            INDEX idx_ad_adjustment_operation (operation_type_id),
            INDEX idx_ad_adjustment_reason (reason_id),
            INDEX idx_ad_adjustment_date (adjust_date),
            CONSTRAINT fk_ad_adjustment_item FOREIGN KEY (ad_item_id)
                REFERENCES amazon_ad_items(id) ON DELETE RESTRICT,
            CONSTRAINT fk_ad_adjustment_operation FOREIGN KEY (operation_type_id)
                REFERENCES amazon_ad_operation_types(id) ON DELETE RESTRICT,
            CONSTRAINT fk_ad_adjustment_reason FOREIGN KEY (reason_id)
                REFERENCES amazon_ad_operation_reasons(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sql)
        self._amazon_ad_adjustment_ready = True

    def _ensure_amazon_keyword_tables(self):
        if self._amazon_keyword_ready:
            return
        self._ensure_category_table()
        self._ensure_product_table()

        create_keywords_sql = """
        CREATE TABLE IF NOT EXISTS amazon_keywords (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            category_id INT UNSIGNED NOT NULL,
            user_search_term VARCHAR(255) NOT NULL,
            search_rank INT NULL,
            rank_updated_at DATETIME NULL,
            previous_search_rank INT NULL,
            previous_rank_updated_at DATETIME NULL,
            top_click_asin1 VARCHAR(64) NULL,
            top_click_asin1_click_share VARCHAR(32) NULL,
            top_click_asin1_conversion_share VARCHAR(32) NULL,
            top_click_asin2 VARCHAR(64) NULL,
            top_click_asin2_click_share VARCHAR(32) NULL,
            top_click_asin2_conversion_share VARCHAR(32) NULL,
            top_click_asin3 VARCHAR(64) NULL,
            top_click_asin3_click_share VARCHAR(32) NULL,
            top_click_asin3_conversion_share VARCHAR(32) NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_amazon_keyword_term (user_search_term),
            INDEX idx_amazon_keyword_category (category_id),
            INDEX idx_amazon_keyword_rank_updated (rank_updated_at),
            CONSTRAINT fk_amazon_keyword_category FOREIGN KEY (category_id)
                REFERENCES product_categories(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_tags_sql = """
        CREATE TABLE IF NOT EXISTS amazon_keyword_tags (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            category_id INT UNSIGNED NOT NULL,
            tag_name VARCHAR(64) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_keyword_tag (category_id, tag_name),
            INDEX idx_keyword_tag_category (category_id),
            CONSTRAINT fk_keyword_tag_category FOREIGN KEY (category_id)
                REFERENCES product_categories(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_tag_rel_sql = """
        CREATE TABLE IF NOT EXISTS amazon_keyword_tag_rel (
            keyword_id INT UNSIGNED NOT NULL,
            tag_id INT UNSIGNED NOT NULL,
            PRIMARY KEY (keyword_id, tag_id),
            CONSTRAINT fk_keyword_tag_rel_keyword FOREIGN KEY (keyword_id)
                REFERENCES amazon_keywords(id) ON DELETE CASCADE,
            CONSTRAINT fk_keyword_tag_rel_tag FOREIGN KEY (tag_id)
                REFERENCES amazon_keyword_tags(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_sku_rel_sql = """
        CREATE TABLE IF NOT EXISTS amazon_keyword_sku_rel (
            keyword_id INT UNSIGNED NOT NULL,
            sku_family_id INT UNSIGNED NOT NULL,
            relevance_score TINYINT UNSIGNED NOT NULL DEFAULT 1,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (keyword_id, sku_family_id),
            INDEX idx_keyword_sku_rel_sku (sku_family_id),
            CONSTRAINT fk_keyword_sku_rel_keyword FOREIGN KEY (keyword_id)
                REFERENCES amazon_keywords(id) ON DELETE CASCADE,
            CONSTRAINT fk_keyword_sku_rel_sku FOREIGN KEY (sku_family_id)
                REFERENCES product_families(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_keywords_sql)
                cur.execute(create_tags_sql)
                cur.execute(create_tag_rel_sql)
                cur.execute(create_sku_rel_sql)

        self._amazon_keyword_ready = True

    def _ensure_certification_table(self):
        if getattr(self, '_certification_ready', False):
            return
        with self._schema_ensure_lock:
            if getattr(self, '_certification_ready', False):
                return
            create_sql = """
            CREATE TABLE IF NOT EXISTS certifications (
                id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(128) NOT NULL UNIQUE,
                icon_name VARCHAR(255) NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(create_sql)
                    cur.execute(
                        """
                        SELECT COUNT(*) AS cnt
                        FROM information_schema.COLUMNS
                        WHERE TABLE_SCHEMA = DATABASE()
                          AND TABLE_NAME = 'certifications'
                          AND COLUMN_NAME = 'icon_name'
                        """
                    )
                    row = cur.fetchone() or {}
                    if int(row.get('cnt') or 0) == 0:
                        cur.execute("ALTER TABLE certifications ADD COLUMN icon_name VARCHAR(255) NULL AFTER name")
            self._certification_ready = True
            self.__class__._schema_ready_cache['certification'] = True

    def _ensure_features_table(self):
        self._ensure_category_table()
        create_features = """
        CREATE TABLE IF NOT EXISTS features (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(128) NOT NULL UNIQUE,
            name_en VARCHAR(128) NOT NULL DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_feature_name (name)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        create_feature_categories = """
        CREATE TABLE IF NOT EXISTS feature_categories (
            feature_id INT UNSIGNED NOT NULL,
            category_id INT UNSIGNED NOT NULL,
            PRIMARY KEY (feature_id, category_id),
            CONSTRAINT fk_feature_category_feature FOREIGN KEY (feature_id)
                REFERENCES features(id) ON DELETE CASCADE,
            CONSTRAINT fk_feature_category_category FOREIGN KEY (category_id)
                REFERENCES product_categories(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_features)
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'features'
                      AND COLUMN_NAME = 'name_en'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    cur.execute("ALTER TABLE features ADD COLUMN name_en VARCHAR(128) NOT NULL DEFAULT ''")
                cur.execute(create_feature_categories)

    def _ensure_order_product_tables(self):
        if self._order_product_ready:
            return
        self._ensure_product_table()
        self._ensure_fabric_table()
        self._ensure_category_table()
        self._ensure_certification_table()
        self._ensure_materials_table()

        create_order_products = """
        CREATE TABLE IF NOT EXISTS order_products (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            sku VARCHAR(64) NOT NULL UNIQUE,
            sku_family_id INT UNSIGNED NULL,
            version_no VARCHAR(64) NOT NULL,
            fabric_id INT UNSIGNED NULL,
            spec_qty_short VARCHAR(128) NOT NULL,
            contents_desc_en VARCHAR(255) NULL,
            is_iteration TINYINT(1) NOT NULL DEFAULT 0,
            is_dachene_product TINYINT(1) NOT NULL DEFAULT 0,
            is_on_market TINYINT(1) NOT NULL DEFAULT 1,
            source_order_product_id INT UNSIGNED NULL,
            finished_length_in DECIMAL(10,2) NULL,
            finished_width_in DECIMAL(10,2) NULL,
            finished_height_in DECIMAL(10,2) NULL,
            net_weight_lbs DECIMAL(10,2) NULL,
            package_length_in DECIMAL(10,2) NULL,
            package_width_in DECIMAL(10,2) NULL,
            package_height_in DECIMAL(10,2) NULL,
            gross_weight_lbs DECIMAL(10,2) NULL,
            cost_usd DECIMAL(10,2) NULL,
            carton_qty INT UNSIGNED NULL,
            package_size_class VARCHAR(64) NULL,
            last_mile_avg_freight_usd DECIMAL(10,2) NULL,
            factory_wip_stock INT NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_sku_family (sku_family_id),
            INDEX idx_fabric (fabric_id),
            INDEX idx_source_order_product (source_order_product_id),
            CONSTRAINT fk_order_products_sku_family FOREIGN KEY (sku_family_id)
                REFERENCES product_families(id) ON DELETE SET NULL,
            CONSTRAINT fk_order_products_fabric FOREIGN KEY (fabric_id)
                REFERENCES fabric_materials(id) ON DELETE SET NULL,
            CONSTRAINT fk_order_products_source FOREIGN KEY (source_order_product_id)
                REFERENCES order_products(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_order_product_materials = """
        CREATE TABLE IF NOT EXISTS order_product_materials (
            order_product_id INT UNSIGNED NOT NULL,
            material_id INT UNSIGNED NOT NULL,
            PRIMARY KEY (order_product_id, material_id),
            CONSTRAINT fk_opm_order_product FOREIGN KEY (order_product_id)
                REFERENCES order_products(id) ON DELETE CASCADE,
            CONSTRAINT fk_opm_material FOREIGN KEY (material_id)
                REFERENCES materials(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_features = """
        CREATE TABLE IF NOT EXISTS features (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(128) NOT NULL UNIQUE,
            name_en VARCHAR(128) NOT NULL DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_feature_name (name)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_feature_categories = """
        CREATE TABLE IF NOT EXISTS feature_categories (
            feature_id INT UNSIGNED NOT NULL,
            category_id INT UNSIGNED NOT NULL,
            PRIMARY KEY (feature_id, category_id),
            CONSTRAINT fk_feature_category_feature FOREIGN KEY (feature_id)
                REFERENCES features(id) ON DELETE CASCADE,
            CONSTRAINT fk_feature_category_category FOREIGN KEY (category_id)
                REFERENCES product_categories(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_order_product_features = """
        CREATE TABLE IF NOT EXISTS order_product_features (
            order_product_id INT UNSIGNED NOT NULL,
            feature_id INT UNSIGNED NOT NULL,
            PRIMARY KEY (order_product_id, feature_id),
            CONSTRAINT fk_opf_order_product FOREIGN KEY (order_product_id)
                REFERENCES order_products(id) ON DELETE CASCADE,
            CONSTRAINT fk_opf_feature FOREIGN KEY (feature_id)
                REFERENCES features(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_order_product_certifications = """
        CREATE TABLE IF NOT EXISTS order_product_certifications (
            order_product_id INT UNSIGNED NOT NULL,
            certification_id INT UNSIGNED NOT NULL,
            PRIMARY KEY (order_product_id, certification_id),
            CONSTRAINT fk_opc_order_product FOREIGN KEY (order_product_id)
                REFERENCES order_products(id) ON DELETE CASCADE,
            CONSTRAINT fk_opc_certification FOREIGN KEY (certification_id)
                REFERENCES certifications(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_order_product_shipping_plans = """
        CREATE TABLE IF NOT EXISTS order_product_shipping_plans (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            order_product_id INT UNSIGNED NOT NULL,
            plan_name VARCHAR(128) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY uk_order_plan_name (order_product_id, plan_name),
            INDEX idx_ops_order (order_product_id),
            CONSTRAINT fk_ops_order FOREIGN KEY (order_product_id)
                REFERENCES order_products(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_order_product_shipping_plan_items = """
        CREATE TABLE IF NOT EXISTS order_product_shipping_plan_items (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            shipping_plan_id INT UNSIGNED NOT NULL,
            substitute_order_product_id INT UNSIGNED NOT NULL,
            quantity INT UNSIGNED NOT NULL DEFAULT 1,
            sort_order INT UNSIGNED NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uk_opsi_unique (shipping_plan_id, substitute_order_product_id, sort_order),
            INDEX idx_opsi_plan (shipping_plan_id),
            CONSTRAINT fk_opsi_plan FOREIGN KEY (shipping_plan_id)
                REFERENCES order_product_shipping_plans(id) ON DELETE CASCADE,
            CONSTRAINT fk_opsi_sub_order FOREIGN KEY (substitute_order_product_id)
                REFERENCES order_products(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_order_products)
                cur.execute(create_order_product_materials)
                cur.execute(create_features)
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'features'
                      AND COLUMN_NAME = 'name_en'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    cur.execute("ALTER TABLE features ADD COLUMN name_en VARCHAR(128) NOT NULL DEFAULT ''")
                cur.execute(create_feature_categories)
                cur.execute(create_order_product_features)
                cur.execute(create_order_product_certifications)
                cur.execute(create_order_product_shipping_plans)
                cur.execute(create_order_product_shipping_plan_items)
                try:
                    cur.execute(
                        """
                        SELECT COUNT(*) AS cnt
                        FROM information_schema.COLUMNS
                        WHERE TABLE_SCHEMA = DATABASE()
                          AND TABLE_NAME = 'order_product_shipping_plans'
                          AND COLUMN_NAME = 'is_default'
                        """
                    )
                    row = cur.fetchone() or {}
                    if int(row.get('cnt') or 0) > 0:
                        cur.execute("ALTER TABLE order_product_shipping_plans DROP COLUMN is_default")
                except Exception:
                    pass
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'order_products'
                      AND COLUMN_NAME = 'dachene_yuncang_no'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) > 0:
                    try:
                        cur.execute("ALTER TABLE order_products DROP COLUMN dachene_yuncang_no")
                    except Exception:
                        pass

                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'order_products'
                      AND COLUMN_NAME = 'spec_qty'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) > 0:
                    try:
                        cur.execute("ALTER TABLE order_products DROP COLUMN spec_qty")
                    except Exception:
                        pass

                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'order_products'
                      AND COLUMN_NAME = 'listing_image_b64'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) > 0:
                    try:
                        cur.execute("ALTER TABLE order_products DROP COLUMN listing_image_b64")
                    except Exception:
                        pass

                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'order_products'
                      AND COLUMN_NAME = 'is_iteration'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    try:
                        cur.execute("ALTER TABLE order_products ADD COLUMN is_iteration TINYINT(1) NOT NULL DEFAULT 0")
                    except Exception:
                        pass

                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'order_products'
                      AND COLUMN_NAME = 'is_dachene_product'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    try:
                        cur.execute("ALTER TABLE order_products ADD COLUMN is_dachene_product TINYINT(1) NOT NULL DEFAULT 0 AFTER is_iteration")
                    except Exception:
                        pass

                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'order_products'
                      AND COLUMN_NAME = 'is_on_market'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    try:
                        cur.execute("ALTER TABLE order_products ADD COLUMN is_on_market TINYINT(1) NOT NULL DEFAULT 1 AFTER is_dachene_product")
                    except Exception:
                        pass

                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'order_products'
                      AND COLUMN_NAME = 'contents_desc_en'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    try:
                        cur.execute("ALTER TABLE order_products ADD COLUMN contents_desc_en VARCHAR(255) NULL AFTER spec_qty_short")
                    except Exception:
                        pass

                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'order_products'
                      AND COLUMN_NAME = 'factory_wip_stock'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    try:
                        cur.execute("ALTER TABLE order_products ADD COLUMN factory_wip_stock INT NOT NULL DEFAULT 0 AFTER last_mile_avg_freight_usd")
                    except Exception:
                        pass

                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME = 'order_products'
                      AND COLUMN_NAME = 'source_order_product_id'
                    """
                )
                row = cur.fetchone()
                if row and row.get('cnt', 0) == 0:
                    try:
                        cur.execute("ALTER TABLE order_products ADD COLUMN source_order_product_id INT UNSIGNED NULL")
                    except Exception:
                        pass
                    try:
                        cur.execute("ALTER TABLE order_products ADD INDEX idx_source_order_product (source_order_product_id)")
                    except Exception:
                        pass
                    try:
                        cur.execute(
                            """
                            ALTER TABLE order_products
                            ADD CONSTRAINT fk_order_products_source
                            FOREIGN KEY (source_order_product_id) REFERENCES order_products(id)
                            ON DELETE SET NULL
                            """
                        )
                    except Exception:
                        pass

        self._order_product_ready = True

    def _ensure_sales_parent_tables(self):
        if self._sales_parent_ready:
            return
        with self._schema_ensure_lock:
            if self._sales_parent_ready:
                return
        self._ensure_shops_table()
        create_sales_parents = """
        CREATE TABLE IF NOT EXISTS sales_parents (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            parent_code VARCHAR(64) NOT NULL UNIQUE,
            is_enabled TINYINT(1) NOT NULL DEFAULT 1,
            shop_id INT UNSIGNED NULL,
            sku_marker VARCHAR(128) NULL,
            estimated_refund_rate DECIMAL(8,4) NULL,
            estimated_discount_rate DECIMAL(8,4) NULL,
            commission_rate DECIMAL(8,4) NULL,
            estimated_acoas DECIMAL(8,4) NULL,
            sales_title VARCHAR(200) NULL,
            sales_intro VARCHAR(500) NULL,
            sales_bullet_1 VARCHAR(500) NULL,
            sales_bullet_2 VARCHAR(500) NULL,
            sales_bullet_3 VARCHAR(500) NULL,
            sales_bullet_4 VARCHAR(500) NULL,
            sales_bullet_5 VARCHAR(500) NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_parent_code (parent_code),
            INDEX idx_parent_shop (shop_id),
            CONSTRAINT fk_sales_parents_shop FOREIGN KEY (shop_id)
                REFERENCES shops(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sales_parents)
                migration_columns = [
                    ("is_enabled", "ALTER TABLE sales_parents ADD COLUMN is_enabled TINYINT(1) NOT NULL DEFAULT 1 AFTER parent_code"),
                    ("shop_id", "ALTER TABLE sales_parents ADD COLUMN shop_id INT UNSIGNED NULL AFTER is_enabled"),
                    ("sku_marker", "ALTER TABLE sales_parents ADD COLUMN sku_marker VARCHAR(128) NULL AFTER parent_code"),
                    ("sales_title", "ALTER TABLE sales_parents ADD COLUMN sales_title VARCHAR(200) NULL AFTER estimated_acoas"),
                    ("sales_intro", "ALTER TABLE sales_parents ADD COLUMN sales_intro VARCHAR(500) NULL AFTER sales_title"),
                    ("sales_bullet_1", "ALTER TABLE sales_parents ADD COLUMN sales_bullet_1 VARCHAR(500) NULL AFTER sales_intro"),
                    ("sales_bullet_2", "ALTER TABLE sales_parents ADD COLUMN sales_bullet_2 VARCHAR(500) NULL AFTER sales_bullet_1"),
                    ("sales_bullet_3", "ALTER TABLE sales_parents ADD COLUMN sales_bullet_3 VARCHAR(500) NULL AFTER sales_bullet_2"),
                    ("sales_bullet_4", "ALTER TABLE sales_parents ADD COLUMN sales_bullet_4 VARCHAR(500) NULL AFTER sales_bullet_3"),
                    ("sales_bullet_5", "ALTER TABLE sales_parents ADD COLUMN sales_bullet_5 VARCHAR(500) NULL AFTER sales_bullet_4")
                ]
                for col_name, alter_sql in migration_columns:
                    try:
                        cur.execute(
                            """
                            SELECT COUNT(*) AS cnt
                            FROM information_schema.COLUMNS
                            WHERE TABLE_SCHEMA=DATABASE()
                              AND TABLE_NAME='sales_parents'
                              AND COLUMN_NAME=%s
                            """,
                            (col_name,)
                        )
                        row = cur.fetchone()
                        if row and row.get('cnt', 0) == 0:
                            cur.execute(alter_sql)
                    except Exception:
                        pass
                try:
                    cur.execute("ALTER TABLE sales_parents ADD INDEX idx_parent_shop (shop_id)")
                except Exception:
                    pass
                try:
                    cur.execute(
                        """
                        ALTER TABLE sales_parents
                        ADD CONSTRAINT fk_sales_parents_shop
                        FOREIGN KEY (shop_id) REFERENCES shops(id)
                        ON DELETE SET NULL
                        """
                    )
                except Exception:
                    pass
            self._sales_parent_ready = True
            self.__class__._schema_ready_cache['sales_parent'] = True

    def _ensure_sales_product_tables(self):
        if self._sales_product_ready:
            return
        self._ensure_shops_table()
        self._ensure_sales_parent_tables()
        self._ensure_amazon_ad_tables()
        self._ensure_order_product_tables()

        create_sales_products = """
        CREATE TABLE IF NOT EXISTS sales_products (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            shop_id INT UNSIGNED NULL,
            portfolio_id INT UNSIGNED NOT NULL,
            platform_sku VARCHAR(128) NOT NULL UNIQUE,
            product_status VARCHAR(16) NOT NULL DEFAULT 'enabled',
            sku_family_id INT UNSIGNED NULL,
            parent_id INT UNSIGNED NULL,
            child_code VARCHAR(64) NULL,
            dachene_yuncang_no VARCHAR(128) NULL,
            fabric VARCHAR(255) NULL,
            spec_name VARCHAR(255) NULL,
            sales_title VARCHAR(200) NULL,
            sale_price_usd DECIMAL(10,2) NULL,
            warehouse_cost_usd DECIMAL(10,2) NULL,
            last_mile_cost_usd DECIMAL(10,2) NULL,
            package_length_in DECIMAL(10,2) NULL,
            package_width_in DECIMAL(10,2) NULL,
            package_height_in DECIMAL(10,2) NULL,
            net_weight_lbs DECIMAL(10,2) NULL,
            gross_weight_lbs DECIMAL(10,2) NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_sp_shop (shop_id),
            INDEX idx_sp_sku_family (sku_family_id),
            INDEX idx_sp_parent (parent_id),
            INDEX idx_sp_portfolio (portfolio_id),
            CONSTRAINT fk_sp_shop FOREIGN KEY (shop_id) REFERENCES shops(id) ON DELETE SET NULL,
            CONSTRAINT fk_sp_sku_family FOREIGN KEY (sku_family_id) REFERENCES product_families(id) ON DELETE SET NULL,
            CONSTRAINT fk_sp_parent FOREIGN KEY (parent_id) REFERENCES sales_parents(id) ON DELETE SET NULL,
            CONSTRAINT fk_sp_portfolio FOREIGN KEY (portfolio_id) REFERENCES amazon_ad_items(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_sales_order_links = """
        CREATE TABLE IF NOT EXISTS sales_product_order_links (
            sales_product_id INT UNSIGNED NOT NULL,
            order_product_id INT UNSIGNED NOT NULL,
            quantity INT UNSIGNED NOT NULL DEFAULT 1,
            PRIMARY KEY (sales_product_id, order_product_id),
            CONSTRAINT fk_spol_sales FOREIGN KEY (sales_product_id)
                REFERENCES sales_products(id) ON DELETE CASCADE,
            CONSTRAINT fk_spol_order FOREIGN KEY (order_product_id)
                REFERENCES order_products(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_sales_products)
                cur.execute(create_sales_order_links)
                
                # 删除 portfolio_id 字段的迁移
                try:
                    cur.execute("""
                        SELECT COLUMN_NAME FROM information_schema.COLUMNS
                        WHERE TABLE_SCHEMA=DATABASE()
                        AND TABLE_NAME='sales_products'
                        AND COLUMN_NAME='portfolio_id'
                    """)
                    if cur.fetchone():
                        # 先删除外键约束
                        cur.execute("""
                            SELECT CONSTRAINT_NAME FROM information_schema.KEY_COLUMN_USAGE
                            WHERE TABLE_SCHEMA=DATABASE()
                            AND TABLE_NAME='sales_products'
                            AND COLUMN_NAME='portfolio_id'
                            AND CONSTRAINT_NAME != 'PRIMARY'
                        """)
                        fk_row = cur.fetchone()
                        if fk_row:
                            fk_name = fk_row['CONSTRAINT_NAME']
                            cur.execute(f"ALTER TABLE sales_products DROP FOREIGN KEY {fk_name}")
                        # 删除索引（如果存在）
                        cur.execute("""
                            SELECT INDEX_NAME FROM information_schema.STATISTICS
                            WHERE TABLE_SCHEMA=DATABASE()
                            AND TABLE_NAME='sales_products'
                            AND COLUMN_NAME='portfolio_id'
                            AND INDEX_NAME != 'PRIMARY'
                        """)
                        idx_row = cur.fetchone()
                        if idx_row:
                            idx_name = idx_row['INDEX_NAME']
                            cur.execute(f"ALTER TABLE sales_products DROP INDEX {idx_name}")
                        # 最后删除列
                        cur.execute("ALTER TABLE sales_products DROP COLUMN portfolio_id")
                except Exception:
                    pass

                # 兼容迁移：旧字段重命名/新增
                migration_columns = [
                    ("product_status", "ALTER TABLE sales_products ADD COLUMN product_status VARCHAR(16) NOT NULL DEFAULT 'enabled' AFTER platform_sku"),
                    ("sku_family_id", "ALTER TABLE sales_products ADD COLUMN sku_family_id INT UNSIGNED NULL AFTER platform_sku"),
                    ("parent_id", "ALTER TABLE sales_products ADD COLUMN parent_id INT UNSIGNED NULL AFTER platform_sku"),
                    ("child_code", "ALTER TABLE sales_products ADD COLUMN child_code VARCHAR(64) NULL AFTER parent_id"),
                    ("dachene_yuncang_no", "ALTER TABLE sales_products ADD COLUMN dachene_yuncang_no VARCHAR(128) NULL AFTER child_code"),
                    ("sales_title", "ALTER TABLE sales_products ADD COLUMN sales_title VARCHAR(200) NULL AFTER spec_name"),
                    ("sale_price_usd", "ALTER TABLE sales_products ADD COLUMN sale_price_usd DECIMAL(10,2) NULL AFTER spec_name"),
                    ("warehouse_cost_usd", "ALTER TABLE sales_products ADD COLUMN warehouse_cost_usd DECIMAL(10,2) NULL AFTER sale_price_usd"),
                    ("last_mile_cost_usd", "ALTER TABLE sales_products ADD COLUMN last_mile_cost_usd DECIMAL(10,2) NULL AFTER warehouse_cost_usd"),
                    ("package_length_in", "ALTER TABLE sales_products ADD COLUMN package_length_in DECIMAL(10,2) NULL AFTER last_mile_cost_usd"),
                    ("package_width_in", "ALTER TABLE sales_products ADD COLUMN package_width_in DECIMAL(10,2) NULL AFTER package_length_in"),
                    ("package_height_in", "ALTER TABLE sales_products ADD COLUMN package_height_in DECIMAL(10,2) NULL AFTER package_width_in"),
                    ("net_weight_lbs", "ALTER TABLE sales_products ADD COLUMN net_weight_lbs DECIMAL(10,2) NULL AFTER package_height_in"),
                    ("gross_weight_lbs", "ALTER TABLE sales_products ADD COLUMN gross_weight_lbs DECIMAL(10,2) NULL AFTER net_weight_lbs")
                ]
                for col_name, alter_sql in migration_columns:
                    try:
                        cur.execute(
                            """
                            SELECT COUNT(*) AS cnt
                            FROM information_schema.COLUMNS
                            WHERE TABLE_SCHEMA=DATABASE()
                              AND TABLE_NAME='sales_products'
                              AND COLUMN_NAME=%s
                            """,
                            (col_name,)
                        )
                        row = cur.fetchone()
                        if row and row.get('cnt', 0) == 0:
                            cur.execute(alter_sql)
                    except Exception:
                        pass

                try:
                    cur.execute(
                        """
                        SELECT COUNT(*) AS cnt
                        FROM information_schema.STATISTICS
                        WHERE TABLE_SCHEMA=DATABASE()
                          AND TABLE_NAME='sales_products'
                          AND INDEX_NAME='idx_sp_sku_family'
                        """
                    )
                    idx_row = cur.fetchone()
                    if not idx_row or idx_row.get('cnt', 0) == 0:
                        cur.execute("ALTER TABLE sales_products ADD INDEX idx_sp_sku_family (sku_family_id)")
                except Exception:
                    pass

                try:
                    cur.execute(
                        """
                        SELECT COUNT(*) AS cnt
                        FROM information_schema.KEY_COLUMN_USAGE
                        WHERE TABLE_SCHEMA=DATABASE()
                          AND TABLE_NAME='sales_products'
                          AND CONSTRAINT_NAME='fk_sp_sku_family'
                        """
                    )
                    fk_row = cur.fetchone()
                    if not fk_row or fk_row.get('cnt', 0) == 0:
                        cur.execute(
                            """
                            ALTER TABLE sales_products
                            ADD CONSTRAINT fk_sp_sku_family
                            FOREIGN KEY (sku_family_id) REFERENCES product_families(id)
                            ON DELETE SET NULL
                            """
                        )
                except Exception:
                    pass

                # 旧 parent_asin/child_asin 字段迁移
                try:
                    cur.execute(
                        """
                        SELECT COUNT(*) AS cnt
                        FROM information_schema.COLUMNS
                        WHERE TABLE_SCHEMA=DATABASE()
                          AND TABLE_NAME='sales_products'
                          AND COLUMN_NAME='child_asin'
                        """
                    )
                    row = cur.fetchone()
                    if row and row.get('cnt', 0) > 0:
                        cur.execute("UPDATE sales_products SET child_code = child_asin WHERE child_code IS NULL AND child_asin IS NOT NULL")
                        cur.execute("ALTER TABLE sales_products DROP COLUMN child_asin")
                except Exception:
                    pass

                try:
                    cur.execute(
                        """
                        SELECT COUNT(*) AS cnt
                        FROM information_schema.COLUMNS
                        WHERE TABLE_SCHEMA=DATABASE()
                          AND TABLE_NAME='sales_products'
                          AND COLUMN_NAME='parent_asin'
                        """
                    )
                    row = cur.fetchone()
                    if row and row.get('cnt', 0) > 0:
                        cur.execute(
                            """
                            INSERT IGNORE INTO sales_parents (parent_code)
                            SELECT DISTINCT parent_asin FROM sales_products
                            WHERE parent_asin IS NOT NULL AND parent_asin <> ''
                            """
                        )
                        cur.execute(
                            """
                            UPDATE sales_products sp
                            JOIN sales_parents p ON p.parent_code = sp.parent_asin
                            SET sp.parent_id = p.id
                            WHERE sp.parent_id IS NULL
                            """
                        )
                        cur.execute("ALTER TABLE sales_products DROP COLUMN parent_asin")
                except Exception:
                    pass

                for old_col in ['assembled_length_in', 'assembled_width_in', 'assembled_height_in']:
                    try:
                        cur.execute(
                            """
                            SELECT COUNT(*) AS cnt
                            FROM information_schema.COLUMNS
                            WHERE TABLE_SCHEMA=DATABASE()
                              AND TABLE_NAME='sales_products'
                              AND COLUMN_NAME=%s
                            """,
                            (old_col,)
                        )
                        row = cur.fetchone()
                        if row and row.get('cnt', 0) > 0:
                            cur.execute(f"ALTER TABLE sales_products DROP COLUMN {old_col}")
                    except Exception:
                        pass


                try:
                    cur.execute("ALTER TABLE sales_products ADD INDEX idx_sp_parent (parent_id)")
                except Exception:
                    pass
                try:
                    cur.execute("ALTER TABLE sales_products MODIFY COLUMN shop_id INT UNSIGNED NULL")
                except Exception:
                    pass
                try:
                    cur.execute(
                        """
                        ALTER TABLE sales_products
                        ADD CONSTRAINT fk_sp_parent FOREIGN KEY (parent_id)
                        REFERENCES sales_parents(id) ON DELETE SET NULL
                        """
                    )
                except Exception:
                    pass
        self._sales_product_ready = True

    def _ensure_sales_order_registration_tables(self):
        if self._sales_order_registration_ready:
            return
        with self._schema_ensure_lock:
            if self._sales_order_registration_ready:
                return
        self._ensure_sales_product_tables()
        self._ensure_order_product_tables()
        self._ensure_shops_table()

        create_orders = """
        CREATE TABLE IF NOT EXISTS sales_order_registrations (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            shop_id INT UNSIGNED NULL,
            order_no VARCHAR(128) NOT NULL,
            order_date DATE NULL,
            customer_name VARCHAR(128) NULL,
            phone VARCHAR(64) NULL,
            zip_code VARCHAR(16) NULL,
            address VARCHAR(255) NULL,
            city VARCHAR(64) NULL,
            state VARCHAR(32) NULL,
            shipping_status VARCHAR(32) NOT NULL DEFAULT 'pending',
            is_review_invited TINYINT(1) NOT NULL DEFAULT 0,
            is_logistics_emailed TINYINT(1) NOT NULL DEFAULT 0,
            compensation_action VARCHAR(255) NULL,
            remark TEXT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_sor_shop (shop_id),
            INDEX idx_sor_order_no (order_no),
            INDEX idx_sor_date (order_date),
            CONSTRAINT fk_sor_shop FOREIGN KEY (shop_id)
                REFERENCES shops(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_platform_items = """
        CREATE TABLE IF NOT EXISTS sales_order_registration_platform_items (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            registration_id INT UNSIGNED NOT NULL,
            sales_product_id INT UNSIGNED NULL,
            platform_sku VARCHAR(128) NOT NULL,
            quantity INT UNSIGNED NOT NULL DEFAULT 1,
            shipping_plan_id INT UNSIGNED NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_sorpi_registration (registration_id),
            INDEX idx_sorpi_sales (sales_product_id),
            INDEX idx_sorpi_plan (shipping_plan_id),
            CONSTRAINT fk_sorpi_registration FOREIGN KEY (registration_id)
                REFERENCES sales_order_registrations(id) ON DELETE CASCADE,
            CONSTRAINT fk_sorpi_sales FOREIGN KEY (sales_product_id)
                REFERENCES sales_products(id) ON DELETE SET NULL,
            CONSTRAINT fk_sorpi_plan FOREIGN KEY (shipping_plan_id)
                REFERENCES order_product_shipping_plans(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_shipment_items = """
        CREATE TABLE IF NOT EXISTS sales_order_registration_shipment_items (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            registration_id INT UNSIGNED NOT NULL,
            order_product_id INT UNSIGNED NULL,
            order_sku VARCHAR(64) NOT NULL,
            quantity INT UNSIGNED NOT NULL DEFAULT 1,
            source_type VARCHAR(16) NOT NULL DEFAULT 'manual',
            shipping_plan_id INT UNSIGNED NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_sorsi_registration (registration_id),
            INDEX idx_sorsi_order_product (order_product_id),
            INDEX idx_sorsi_plan (shipping_plan_id),
            CONSTRAINT fk_sorsi_registration FOREIGN KEY (registration_id)
                REFERENCES sales_order_registrations(id) ON DELETE CASCADE,
            CONSTRAINT fk_sorsi_order_product FOREIGN KEY (order_product_id)
                REFERENCES order_products(id) ON DELETE SET NULL,
            CONSTRAINT fk_sorsi_plan FOREIGN KEY (shipping_plan_id)
                REFERENCES order_product_shipping_plans(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        create_logistics_items = """
        CREATE TABLE IF NOT EXISTS sales_order_registration_logistics_items (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            registration_id INT UNSIGNED NOT NULL,
            shipping_carrier VARCHAR(128) NULL,
            tracking_no VARCHAR(255) NULL,
            sort_order INT UNSIGNED NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_sorli_registration (registration_id),
            INDEX idx_sorli_tracking (tracking_no(128)),
            CONSTRAINT fk_sorli_registration FOREIGN KEY (registration_id)
                REFERENCES sales_order_registrations(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """

        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_orders)
                cur.execute(create_platform_items)
                cur.execute(create_shipment_items)
                cur.execute(create_logistics_items)
                try:
                    cur.execute("ALTER TABLE sales_order_registrations ADD INDEX idx_sor_shop_order (shop_id, order_no)")
                except Exception:
                    pass
                try:
                    cur.execute("ALTER TABLE sales_order_registrations ADD INDEX idx_sor_customer_name (customer_name)")
                except Exception:
                    pass
                try:
                    cur.execute("ALTER TABLE sales_order_registrations ADD INDEX idx_sor_phone (phone)")
                except Exception:
                    pass
                try:
                    cur.execute("ALTER TABLE sales_order_registration_logistics_items ADD INDEX idx_sorli_carrier_tracking (shipping_carrier, tracking_no(128))")
                except Exception:
                    pass
                try:
                    cur.execute("ALTER TABLE sales_order_registration_platform_items ADD INDEX idx_sorpi_registration_id_id (registration_id, id)")
                except Exception:
                    pass
                try:
                    cur.execute("ALTER TABLE sales_order_registration_shipment_items ADD INDEX idx_sorsi_registration_id_id (registration_id, id)")
                except Exception:
                    pass
                try:
                    cur.execute("ALTER TABLE sales_order_registration_logistics_items ADD INDEX idx_sorli_registration_sort_id (registration_id, sort_order, id)")
                except Exception:
                    pass
                for old_col in ('is_cancelled', 'shipping_carrier', 'tracking_no'):
                    try:
                        cur.execute(
                            """
                            SELECT COUNT(*) AS cnt
                            FROM information_schema.COLUMNS
                            WHERE TABLE_SCHEMA = DATABASE()
                              AND TABLE_NAME = 'sales_order_registrations'
                              AND COLUMN_NAME = %s
                            """,
                            (old_col,)
                        )
                        row = cur.fetchone() or {}
                        if int(row.get('cnt') or 0) > 0:
                            cur.execute(f"ALTER TABLE sales_order_registrations DROP COLUMN {old_col}")
                    except Exception:
                        pass
            self._sales_order_registration_ready = True
            self.__class__._schema_ready_cache['sales_order_registration'] = True

    def _ensure_todo_tables(self, lightweight=False):
        if self._todo_ready and (lightweight or self._todo_schema_migrated):
            return

        with self._todo_ensure_lock:
            if self._todo_ready and (lightweight or self._todo_schema_migrated):
                return

            create_users = """
            CREATE TABLE IF NOT EXISTS users (
                id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(64) NOT NULL UNIQUE,
                password_hash VARCHAR(255) NOT NULL,
                name VARCHAR(128) NULL,
                phone VARCHAR(64) NULL,
                birthday DATE NULL,
                is_admin TINYINT UNSIGNED NOT NULL DEFAULT 0,
                can_grant_admin TINYINT UNSIGNED NOT NULL DEFAULT 0,
                page_permissions LONGTEXT NULL,
                is_approved TINYINT(1) NOT NULL DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_username (username),
                INDEX idx_birthday (birthday),
                INDEX idx_name (name)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """

            create_todos = """
            CREATE TABLE IF NOT EXISTS todos (
                id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
                title VARCHAR(255) NOT NULL,
                detail TEXT NULL,
                start_date DATE NOT NULL,
                due_date DATE NOT NULL,
                reminder_interval_days INT UNSIGNED NOT NULL DEFAULT 1,
                last_check_time TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                next_check_time TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                is_recurring TINYINT UNSIGNED NOT NULL DEFAULT 0,
                status VARCHAR(16) NOT NULL DEFAULT 'open',
                priority TINYINT UNSIGNED NOT NULL DEFAULT 2,
                created_by INT UNSIGNED NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_due_date (due_date),
                INDEX idx_status (status),
                INDEX idx_created_by (created_by),
                CONSTRAINT fk_todos_created_by FOREIGN KEY (created_by)
                    REFERENCES users(id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """

            create_todo_assignments = """
            CREATE TABLE IF NOT EXISTS todo_assignments (
                id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
                todo_id INT UNSIGNED NOT NULL,
                assignee_id INT UNSIGNED NOT NULL,
                assignment_status VARCHAR(16) NOT NULL DEFAULT 'pending',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uk_todo_assignee (todo_id, assignee_id),
                CONSTRAINT fk_ta_todo FOREIGN KEY (todo_id)
                    REFERENCES todos(id) ON DELETE CASCADE,
                CONSTRAINT fk_ta_assignee FOREIGN KEY (assignee_id)
                    REFERENCES users(id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """

            create_sessions = """
            CREATE TABLE IF NOT EXISTS sessions (
                session_id VARCHAR(128) PRIMARY KEY,
                employee_id INT UNSIGNED NOT NULL,
                expires_at DATETIME NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_emp (employee_id),
                CONSTRAINT fk_sessions_user FOREIGN KEY (employee_id) REFERENCES users(id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(create_users)
                    cur.execute(create_todos)
                    cur.execute(create_todo_assignments)
                    cur.execute(create_sessions)
                    self._todo_ready = True

                    if lightweight:
                        return

                    for col, ddl in (
                        ('name', "ALTER TABLE users ADD COLUMN name VARCHAR(128) NULL"),
                        ('phone', "ALTER TABLE users ADD COLUMN phone VARCHAR(64) NULL"),
                        ('birthday', "ALTER TABLE users ADD COLUMN birthday DATE NULL"),
                        ('is_admin', "ALTER TABLE users ADD COLUMN is_admin TINYINT UNSIGNED NOT NULL DEFAULT 0"),
                        ('can_grant_admin', "ALTER TABLE users ADD COLUMN can_grant_admin TINYINT UNSIGNED NOT NULL DEFAULT 0"),
                        ('page_permissions', "ALTER TABLE users ADD COLUMN page_permissions LONGTEXT NULL"),
                        ('is_approved', "ALTER TABLE users ADD COLUMN is_approved TINYINT(1) NOT NULL DEFAULT 1"),
                    ):
                        cur.execute(
                            """
                            SELECT COUNT(*) AS cnt
                            FROM information_schema.COLUMNS
                            WHERE TABLE_SCHEMA = DATABASE()
                              AND TABLE_NAME = 'users'
                              AND COLUMN_NAME = %s
                            """,
                            (col,)
                        )
                        row = cur.fetchone()
                        if row and row.get('cnt', 0) == 0:
                            cur.execute(ddl)

                    cur.execute(
                        """
                        SELECT COUNT(*) AS cnt
                        FROM information_schema.COLUMNS
                        WHERE TABLE_SCHEMA = DATABASE()
                          AND TABLE_NAME = 'users'
                          AND COLUMN_NAME = 'can_manage_todos'
                        """
                    )
                    can_manage_col = cur.fetchone()
                    if can_manage_col and can_manage_col.get('cnt', 0) > 0:
                        try:
                            cur.execute("ALTER TABLE users DROP COLUMN can_manage_todos")
                        except Exception:
                            pass

                    cur.execute(
                        """
                        SELECT COUNT(*) AS cnt
                        FROM information_schema.COLUMNS
                        WHERE TABLE_SCHEMA = DATABASE()
                          AND TABLE_NAME = 'users'
                          AND COLUMN_NAME = 'employee_id'
                        """
                    )
                    emp_col = cur.fetchone()
                    if emp_col and emp_col.get('cnt', 0) > 0:
                        cur.execute(
                            """
                            SELECT CONSTRAINT_NAME
                            FROM information_schema.KEY_COLUMN_USAGE
                            WHERE TABLE_SCHEMA = DATABASE()
                              AND TABLE_NAME = 'users'
                              AND COLUMN_NAME = 'employee_id'
                              AND REFERENCED_TABLE_NAME IS NOT NULL
                            """
                        )
                        for fk in cur.fetchall() or []:
                            try:
                                cur.execute(f"ALTER TABLE users DROP FOREIGN KEY {fk['CONSTRAINT_NAME']}")
                            except Exception:
                                pass
                        try:
                            cur.execute("ALTER TABLE users MODIFY COLUMN employee_id INT UNSIGNED NULL")
                        except Exception:
                            pass
                        try:
                            cur.execute("ALTER TABLE users DROP COLUMN employee_id")
                        except Exception:
                            pass

                    for table_name in ('users', 'todos', 'todo_assignments', 'sessions'):
                        cur.execute(
                            """
                            SELECT CONSTRAINT_NAME
                            FROM information_schema.KEY_COLUMN_USAGE
                            WHERE TABLE_SCHEMA = DATABASE()
                              AND TABLE_NAME = %s
                              AND REFERENCED_TABLE_NAME = 'employees'
                            """,
                            (table_name,)
                        )
                        for fk in cur.fetchall() or []:
                            try:
                                cur.execute(f"ALTER TABLE {table_name} DROP FOREIGN KEY {fk['CONSTRAINT_NAME']}")
                            except Exception:
                                pass

                    try:
                        cur.execute("DROP TABLE IF EXISTS employees")
                    except Exception:
                        pass

                    cur.execute("SELECT COUNT(*) AS cnt FROM users WHERE is_admin=1")
                    admin_row = cur.fetchone()
                    if admin_row and admin_row.get('cnt', 0) == 0:
                        cur.execute("SELECT id FROM users ORDER BY id ASC LIMIT 1")
                        first_user = cur.fetchone()
                        if first_user and first_user.get('id'):
                            cur.execute(
                                "UPDATE users SET is_admin=1, can_grant_admin=1, is_approved=1, page_permissions=%s WHERE id=%s",
                                (self._serialize_page_permissions(self._default_page_permissions()), first_user['id'])
                            )

                    try:
                        cur.execute(
                            "ALTER TABLE todos ADD CONSTRAINT fk_todos_created_by FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE CASCADE"
                        )
                    except Exception:
                        pass
                    try:
                        cur.execute(
                            "ALTER TABLE todo_assignments ADD CONSTRAINT fk_ta_assignee FOREIGN KEY (assignee_id) REFERENCES users(id) ON DELETE CASCADE"
                        )
                    except Exception:
                        pass
                    try:
                        cur.execute(
                            "ALTER TABLE sessions ADD CONSTRAINT fk_sessions_user FOREIGN KEY (employee_id) REFERENCES users(id) ON DELETE CASCADE"
                        )
                    except Exception:
                        pass
                    try:
                        cur.execute("ALTER TABLE todo_assignments ADD INDEX idx_ta_assignee_todo (assignee_id, todo_id)")
                    except Exception:
                        pass
                    try:
                        cur.execute("ALTER TABLE todos ADD INDEX idx_todos_creator_due_priority (created_by, due_date, priority, id)")
                    except Exception:
                        pass

            self._todo_schema_migrated = True

    def _split_multi_values(self, value):
        if value is None:
            return []
        if isinstance(value, list):
            raw_items = value
        else:
            raw_items = re.split(r'[\n,，;；/]+', str(value))

        seen = set()
        result = []
        for item in raw_items:
            text = str(item).strip()
            if not text:
                continue
            if text in seen:
                continue
            seen.add(text)
            result.append(text)
        return result

    def _parse_float(self, value):
        if value is None:
            return None
        text = str(value).strip()
        if text == '':
            return None
        try:
            return float(text)
        except Exception:
            return None

    def _parse_int(self, value):
        if value is None:
            return None
        text = str(value).strip()
        if text == '':
            return None
        try:
            return int(float(text))
        except Exception:
            return None

    def _calc_carton_qty_by_40hq(self, package_length_in, package_width_in, package_height_in):
        length_in = self._parse_float(package_length_in)
        width_in = self._parse_float(package_width_in)
        height_in = self._parse_float(package_height_in)
        if length_in is None or width_in is None or height_in is None:
            return None
        if length_in <= 0 or width_in <= 0 or height_in <= 0:
            return None
        inch_to_meter = 0.0254
        volume_m3 = length_in * inch_to_meter * width_in * inch_to_meter * height_in * inch_to_meter
        if volume_m3 <= 0:
            return None
        qty = int(69.0 / volume_m3)
        return qty if qty >= 0 else None

    def _sanitize_xlsx_bool_cells(self, file_bytes):
        if not file_bytes:
            return file_bytes
        try:
            zin = zipfile.ZipFile(io.BytesIO(file_bytes), 'r')
        except Exception:
            return file_bytes

        out_buffer = io.BytesIO()
        changed = False
        ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        valid_bool_values = {'0', '1', 'true', 'false'}

        with zin:
            with zipfile.ZipFile(out_buffer, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                for info in zin.infolist():
                    name = info.filename
                    data = zin.read(name)

                    if name.startswith('xl/worksheets/') and name.endswith('.xml'):
                        try:
                            root = ET.fromstring(data)
                            sheet_changed = False
                            for cell in root.findall('.//x:c', ns):
                                if cell.get('t') != 'b':
                                    continue
                                value_node = cell.find('x:v', ns)
                                raw_text = '' if value_node is None or value_node.text is None else str(value_node.text).strip()
                                if raw_text.lower() not in valid_bool_values:
                                    cell.set('t', 'str')
                                    if value_node is None:
                                        value_node = ET.SubElement(cell, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                                    value_node.text = raw_text
                                    sheet_changed = True
                            if sheet_changed:
                                data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                                changed = True
                        except Exception:
                            pass

                    zout.writestr(info, data)

        if changed:
            return out_buffer.getvalue()
        return file_bytes

    def _scan_xlsx_invalid_bool_cells(self, file_bytes, max_samples=8):
        if not file_bytes:
            return {'count': 0, 'samples': []}
        try:
            zin = zipfile.ZipFile(io.BytesIO(file_bytes), 'r')
        except Exception:
            return {'count': 0, 'samples': []}

        ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        valid_bool_values = {'0', '1', 'true', 'false'}
        count = 0
        samples = []

        with zin:
            for name in zin.namelist():
                if not (name.startswith('xl/worksheets/') and name.endswith('.xml')):
                    continue
                try:
                    root = ET.fromstring(zin.read(name))
                except Exception:
                    continue
                for cell in root.findall('.//x:c', ns):
                    if cell.get('t') != 'b':
                        continue
                    value_node = cell.find('x:v', ns)
                    raw_text = '' if value_node is None or value_node.text is None else str(value_node.text).strip()
                    if raw_text.lower() in valid_bool_values:
                        continue
                    count += 1
                    if len(samples) < max_samples:
                        samples.append({
                            'sheet_xml': name,
                            'cell': cell.get('r') or '',
                            'value': raw_text
                        })

        return {'count': count, 'samples': samples}

    def _xlsx_cell_ref_to_rc(self, ref):
        ref_text = (ref or '').strip().upper()
        match = re.match(r'^([A-Z]+)(\d+)$', ref_text)
        if not match:
            return None, None
        letters, row_text = match.group(1), match.group(2)
        col = 0
        for ch in letters:
            col = col * 26 + (ord(ch) - ord('A') + 1)
        try:
            row = int(row_text)
        except Exception:
            return None, None
        return row, col

    def _extract_xlsx_shared_strings(self, zin):
        ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        if 'xl/sharedStrings.xml' not in zin.namelist():
            return []
        try:
            root = ET.fromstring(zin.read('xl/sharedStrings.xml'))
            items = []
            for si in root.findall('.//x:si', ns):
                texts = []
                for t_node in si.findall('.//x:t', ns):
                    texts.append(t_node.text or '')
                items.append(''.join(texts))
            return items
        except Exception:
            return []

    def _rebuild_workbook_from_xlsx_xml(self, file_bytes):
        try:
            zin = zipfile.ZipFile(io.BytesIO(file_bytes), 'r')
        except Exception:
            return None

        ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        with zin:
            sheet_names = [name for name in zin.namelist() if name.startswith('xl/worksheets/') and name.endswith('.xml')]
            if not sheet_names:
                return None
            sheet_name = 'xl/worksheets/sheet1.xml' if 'xl/worksheets/sheet1.xml' in sheet_names else sorted(sheet_names)[0]
            try:
                sheet_root = ET.fromstring(zin.read(sheet_name))
            except Exception:
                return None
            shared_strings = self._extract_xlsx_shared_strings(zin)

        wb = Workbook()
        ws = wb.active

        for row_node in sheet_root.findall('.//x:sheetData/x:row', ns):
            row_index = self._parse_int(row_node.get('r')) or 1
            fallback_col = 1
            for cell_node in row_node.findall('x:c', ns):
                ref = cell_node.get('r')
                parsed_row, parsed_col = self._xlsx_cell_ref_to_rc(ref) if ref else (None, None)
                target_row = parsed_row or row_index
                target_col = parsed_col or fallback_col
                fallback_col = target_col + 1

                cell_type = (cell_node.get('t') or '').strip()
                if cell_type == 'inlineStr':
                    text_parts = []
                    for t_node in cell_node.findall('.//x:t', ns):
                        text_parts.append(t_node.text or '')
                    value = ''.join(text_parts)
                else:
                    v_node = cell_node.find('x:v', ns)
                    raw_text = '' if v_node is None or v_node.text is None else str(v_node.text)
                    if cell_type == 's':
                        idx = self._parse_int(raw_text)
                        if idx is not None and 0 <= idx < len(shared_strings):
                            value = shared_strings[idx]
                        else:
                            value = raw_text
                    elif cell_type == 'b':
                        lowered = raw_text.strip().lower()
                        if lowered in ('1', 'true'):
                            value = '1'
                        elif lowered in ('0', 'false'):
                            value = '0'
                        else:
                            value = raw_text
                    else:
                        value = raw_text

                if value != '':
                    ws.cell(row=target_row, column=target_col, value=value)

        return wb

    def _parse_date_str(self, value):
        if value is None:
            return None
        text = str(value).strip()
        if text == '':
            return None
        try:
            dt = datetime.strptime(text, '%Y-%m-%d')
            return dt.strftime('%Y-%m-%d')
        except Exception:
            return None

    def _normalize_yes_no(self, value):
        text = ('' if value is None else str(value)).strip().lower()
        if text in ('是', 'yes', 'y', 'true', '1'):
            return 1
        if text in ('否', 'no', 'n', 'false', '0'):
            return 0
        return None

    def _normalize_ad_status(self, value):
        text = ('' if value is None else str(value)).strip()
        if text in ('启动', '暂停', '存档'):
            return text
        return None

    def _normalize_observe_interval(self, value):
        text = ('' if value is None else str(value)).strip()
        if not text:
            return None
        return text[:64]

    def _normalize_observe_days(self, value):
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        m = re.search(r'\d+', text)
        if not m:
            return None
        days = self._parse_int(m.group(0))
        if days is None:
            return None
        if days < 0:
            return None
        return days

    def _normalize_datetime_text(self, value):
        text = ('' if value is None else str(value)).strip()
        if not text:
            return None
        formats = (
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%dT%H:%M'
        )
        for fmt in formats:
            try:
                dt = datetime.strptime(text, fmt)
                return dt.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                continue
        return None

    def _normalize_bid_value(self, value):
        text = ('' if value is None else str(value)).strip().replace(' ', '')
        if not text:
            return None
        if not re.match(r'^(?:\d+(?:\.\d+)?|\.\d+)%?$', text):
            return None
        is_percent = text.endswith('%')
        num_text = text[:-1] if is_percent else text
        if num_text.startswith('.'):
            num_text = '0' + num_text
        try:
            num = float(num_text)
        except Exception:
            return None
        normalized = ('%.6f' % num).rstrip('0').rstrip('.')
        if normalized == '':
            normalized = '0'
        return normalized + ('%' if is_percent else '')

    def _get_sku_family_with_category_short(self, conn, sku_family_id):
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT pf.id, pf.sku_family, pf.category,
                       pc.category_en AS category_short
                FROM product_families pf
                LEFT JOIN product_categories pc ON pc.category_cn = pf.category
                WHERE pf.id=%s
                """,
                (sku_family_id,)
            )
            return cur.fetchone()

    def _get_ad_item_by_id(self, conn, item_id):
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, ad_level, name, portfolio_id, campaign_id
                FROM amazon_ad_items
                WHERE id=%s
                """,
                (item_id,)
            )
            return cur.fetchone()

    def _build_portfolio_name(self, conn, sku_family_id):
        sku_row = self._get_sku_family_with_category_short(conn, sku_family_id)
        if not sku_row:
            return None
        short = (sku_row.get('category_short') or sku_row.get('category') or '').strip()
        sku_family = (sku_row.get('sku_family') or '').strip()
        if not short or not sku_family:
            return None
        return f"{short}-{sku_family}"

    def _build_campaign_name(self, conn, strategy_code, portfolio_id, subtype_id):
        with conn.cursor() as cur:
            cur.execute("SELECT id, name FROM amazon_ad_items WHERE id=%s AND ad_level='portfolio'", (portfolio_id,))
            portfolio = cur.fetchone()
            if not portfolio:
                return None
            cur.execute("SELECT id, ad_class, subtype_code FROM amazon_ad_subtypes WHERE id=%s", (subtype_id,))
            subtype = cur.fetchone()
            if not subtype:
                return None
        strategy = (strategy_code or '').strip().upper()
        if strategy not in ('BE', 'BD', 'PC'):
            return None
        return f"{strategy}-{portfolio.get('name') or ''}-{subtype.get('ad_class') or ''}-{subtype.get('subtype_code') or ''}"

    def _get_material_type_id(self, conn, name_or_code):
        if not name_or_code:
            return None
        type_map = {
            'fabric': '面料',
            'filling': '填充',
            'frame': '框架',
            'electronics': '电子元器件'
        }
        name = type_map.get(name_or_code, name_or_code)
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM material_types WHERE name=%s", (name,))
            row = cur.fetchone()
            return row['id'] if row else None

    def _materials_has_type_id(self, conn):
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT COUNT(*) AS cnt
                FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE()
                  AND TABLE_NAME = 'materials'
                  AND COLUMN_NAME = 'material_type_id'
                """
            )
            row = cur.fetchone()
            return bool(row and row.get('cnt', 0) > 0)

    def _materials_has_parent_id(self, conn):
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT COUNT(*) AS cnt
                FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE()
                  AND TABLE_NAME = 'materials'
                  AND COLUMN_NAME = 'parent_id'
                """
            )
            row = cur.fetchone()
            return bool(row and row.get('cnt', 0) > 0)

    def _upsert_material_ids(self, conn, names, material_type_code):
        ids = []
        if not names:
            return ids
        with conn.cursor() as cur:
            material_type_id = self._get_material_type_id(conn, material_type_code)
            if not material_type_id:
                return ids
            for name in names:
                cur.execute(
                    "SELECT id FROM materials WHERE material_type_id=%s AND name=%s",
                    (material_type_id, name)
                )
                row = cur.fetchone()
                if row:
                    ids.append(row['id'])
                    continue
                cur.execute(
                    "INSERT INTO materials (name, material_type_id) VALUES (%s, %s)",
                    (name, material_type_id)
                )
                ids.append(cur.lastrowid)
        return ids

    def _upsert_feature_ids(self, conn, names):
        ids = []
        if not names:
            return ids
        with conn.cursor() as cur:
            for name in names:
                cur.execute("SELECT id FROM features WHERE name=%s", (name,))
                row = cur.fetchone()
                if row:
                    ids.append(row['id'])
                    continue
                cur.execute("INSERT INTO features (name) VALUES (%s)", (name,))
                ids.append(cur.lastrowid)
        return ids

    def _replace_order_product_materials(self, conn, order_product_id, filling_names, frame_names):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_materials WHERE order_product_id=%s", (order_product_id,))

        for material_type, names in (
            ('filling', filling_names),
            ('frame', frame_names)
        ):
            ids = self._upsert_material_ids(conn, names, material_type)
            if not ids:
                continue
            with conn.cursor() as cur:
                for material_id in ids:
                    cur.execute(
                        "INSERT IGNORE INTO order_product_materials (order_product_id, material_id) VALUES (%s, %s)",
                        (order_product_id, material_id)
                    )

    def _replace_order_product_features(self, conn, order_product_id, feature_names):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_features WHERE order_product_id=%s", (order_product_id,))

        feature_ids = self._upsert_feature_ids(conn, feature_names)
        if not feature_ids:
            return
        with conn.cursor() as cur:
            for feature_id in feature_ids:
                cur.execute(
                    "INSERT IGNORE INTO order_product_features (order_product_id, feature_id) VALUES (%s, %s)",
                    (order_product_id, feature_id)
                )

    def _replace_order_product_material_ids(self, conn, order_product_id, filling_ids, frame_ids):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_materials WHERE order_product_id=%s", (order_product_id,))

        material_ids = []
        if filling_ids:
            material_ids.extend(filling_ids)
        if frame_ids:
            material_ids.extend(frame_ids)
        if not material_ids:
            return
        material_ids = sorted(set(material_ids))
        with conn.cursor() as cur:
            cur.executemany(
                "INSERT IGNORE INTO order_product_materials (order_product_id, material_id) VALUES (%s, %s)",
                [(order_product_id, material_id) for material_id in material_ids]
            )

    def _replace_order_product_feature_ids(self, conn, order_product_id, feature_ids):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_features WHERE order_product_id=%s", (order_product_id,))

        if not feature_ids:
            return
        feature_ids = sorted(set(feature_ids))
        with conn.cursor() as cur:
            cur.executemany(
                "INSERT IGNORE INTO order_product_features (order_product_id, feature_id) VALUES (%s, %s)",
                [(order_product_id, feature_id) for feature_id in feature_ids]
            )

    def _replace_order_product_certification_ids(self, conn, order_product_id, certification_ids):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_certifications WHERE order_product_id=%s", (order_product_id,))

        if not certification_ids:
            return
        certification_ids = sorted(set(certification_ids))
        with conn.cursor() as cur:
            cur.executemany(
                "INSERT IGNORE INTO order_product_certifications (order_product_id, certification_id) VALUES (%s, %s)",
                [(order_product_id, certification_id) for certification_id in certification_ids]
            )

    def _replace_feature_categories(self, conn, feature_id, category_ids):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM feature_categories WHERE feature_id=%s", (feature_id,))

        if not category_ids:
            return
        with conn.cursor() as cur:
            for category_id in category_ids:
                cur.execute(
                    "INSERT IGNORE INTO feature_categories (feature_id, category_id) VALUES (%s, %s)",
                    (feature_id, category_id)
                )

    def _replace_fabric_sku_family_ids(self, conn, fabric_id, sku_family_ids):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM fabric_product_families WHERE fabric_id=%s", (fabric_id,))

        if not sku_family_ids:
            return
        with conn.cursor() as cur:
            for sku_family_id in sku_family_ids:
                cur.execute(
                    "INSERT IGNORE INTO fabric_product_families (fabric_id, sku_family_id) VALUES (%s, %s)",
                    (fabric_id, sku_family_id)
                )

    def _replace_sku_family_fabric_ids(self, conn, sku_family_id, fabric_ids):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM fabric_product_families WHERE sku_family_id=%s", (sku_family_id,))

        if not fabric_ids:
            return
        with conn.cursor() as cur:
            for fabric_id in fabric_ids:
                cur.execute(
                    "INSERT IGNORE INTO fabric_product_families (fabric_id, sku_family_id) VALUES (%s, %s)",
                    (fabric_id, sku_family_id)
                )

    def _replace_ad_subtype_operation_type_ids(self, conn, subtype_id, operation_type_ids):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM amazon_ad_subtype_operation_types WHERE subtype_id=%s", (subtype_id,))

        if not operation_type_ids:
            return
        with conn.cursor() as cur:
            rows = []
            seen = set()
            for operation_type_id in operation_type_ids:
                op_id = self._parse_int(operation_type_id)
                if not op_id or op_id in seen:
                    continue
                seen.add(op_id)
                rows.append((subtype_id, op_id))
            if rows:
                cur.executemany(
                    "INSERT IGNORE INTO amazon_ad_subtype_operation_types (subtype_id, operation_type_id) VALUES (%s, %s)",
                    rows
                )

    def _normalize_ad_operation_reasons(self, reasons):
        items = []
        seen = set()
        if not isinstance(reasons, list):
            return items
        for entry in reasons:
            reason_name = ''
            if isinstance(entry, dict):
                reason_name = (entry.get('reason_name') or entry.get('name') or '').strip()
            elif entry is not None:
                reason_name = str(entry).strip()
            if not reason_name:
                continue
            norm = reason_name.lower()
            if norm in seen:
                continue
            seen.add(norm)
            items.append({'reason_name': reason_name})
        return items

    def _replace_ad_operation_type_reasons(self, conn, operation_type_id, reasons):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM amazon_ad_operation_reasons WHERE operation_type_id=%s", (operation_type_id,))

        if not reasons:
            return
        with conn.cursor() as cur:
            rows = []
            seen = set()
            for reason in reasons:
                reason_name = (reason.get('reason_name') or '').strip()
                if not reason_name:
                    continue
                key = reason_name.lower()
                if key in seen:
                    continue
                seen.add(key)
                rows.append((operation_type_id, reason_name))
            if rows:
                cur.executemany(
                    "INSERT IGNORE INTO amazon_ad_operation_reasons (operation_type_id, reason_name) VALUES (%s, %s)",
                    rows
                )

    def _normalize_sales_order_links(self, links):
        items = []
        if not isinstance(links, list):
            return items
        for entry in links:
            if not isinstance(entry, dict):
                continue
            order_product_id = self._parse_int(entry.get('order_product_id'))
            quantity = self._parse_int(entry.get('quantity')) or 1
            if not order_product_id:
                continue
            items.append({'order_product_id': order_product_id, 'quantity': max(1, quantity)})
        return items

    def _normalize_shipping_plan_items(self, items):
        normalized = []
        if not isinstance(items, list):
            return normalized
        for idx, entry in enumerate(items, start=1):
            if not isinstance(entry, dict):
                continue
            substitute_order_product_id = self._parse_int(entry.get('substitute_order_product_id') or entry.get('order_product_id'))
            quantity = self._parse_int(entry.get('quantity')) or 1
            if not substitute_order_product_id:
                continue
            normalized.append({
                'substitute_order_product_id': substitute_order_product_id,
                'quantity': max(1, quantity),
                'sort_order': self._parse_int(entry.get('sort_order')) or idx
            })
        return normalized

    def _has_duplicate_shipping_plan_substitutes(self, items):
        seen = set()
        for item in (items or []):
            sid = self._parse_int(item.get('substitute_order_product_id'))
            if not sid:
                continue
            if sid in seen:
                return True
            seen.add(sid)
        return False

    def _replace_shipping_plan_items(self, conn, plan_id, items):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM order_product_shipping_plan_items WHERE shipping_plan_id=%s", (plan_id,))
            if not items:
                return
            rows = [
                (plan_id, item['substitute_order_product_id'], item['quantity'], item['sort_order'])
                for item in items
            ]
            cur.executemany(
                """
                INSERT INTO order_product_shipping_plan_items
                    (shipping_plan_id, substitute_order_product_id, quantity, sort_order)
                VALUES (%s, %s, %s, %s)
                """,
                rows
            )

    def _ensure_default_iteration_shipping_plans(self, conn, order_product_id):
        target_id = self._parse_int(order_product_id)
        if not target_id:
            return
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS cnt FROM order_product_shipping_plans WHERE order_product_id=%s", (target_id,))
            row = cur.fetchone() or {}
            if int(row.get('cnt') or 0) > 0:
                return

            cur.execute(
                """
                SELECT id, source_order_product_id, version_no, is_iteration
                FROM order_products
                WHERE id=%s
                """,
                (target_id,)
            )
            current = cur.fetchone() or {}
            if int(current.get('is_iteration') or 0) != 1:
                return
            source_id = self._parse_int(current.get('source_order_product_id'))
            if not source_id:
                return

            version_text = str(current.get('version_no') or '').strip()
            plan_name = f"迭代款-{version_text}" if version_text else "迭代款-1"

            cur.execute(
                """
                SELECT id
                FROM order_products
                WHERE (id=%s OR source_order_product_id=%s)
                  AND id<>%s
                  AND is_on_market=1
                ORDER BY CAST(NULLIF(version_no, '') AS UNSIGNED) ASC, id ASC
                """,
                (source_id, source_id, target_id)
            )
            siblings = cur.fetchall() or []

            candidate_ids = []
            for sibling in siblings:
                sibling_id = self._parse_int(sibling.get('id'))
                if sibling_id and sibling_id not in candidate_ids:
                    candidate_ids.append(sibling_id)

            if not candidate_ids:
                return

            cur.execute(
                """
                INSERT INTO order_product_shipping_plans (order_product_id, plan_name)
                VALUES (%s, %s)
                """,
                (target_id, plan_name)
            )
            plan_id = cur.lastrowid
            rows = []
            for idx, substitute_id in enumerate(candidate_ids, start=1):
                rows.append((plan_id, substitute_id, 1, idx))
            cur.executemany(
                """
                INSERT INTO order_product_shipping_plan_items
                    (shipping_plan_id, substitute_order_product_id, quantity, sort_order)
                VALUES (%s, %s, %s, %s)
                """,
                rows
            )

    def _bool_from_any(self, value, default=0):
        if value is None:
            return 1 if default else 0
        text = str(value).strip().lower()
        if text in ('1', 'true', 'yes', 'y', '是', 'on'):
            return 1
        if text in ('0', 'false', 'no', 'n', '否', 'off'):
            return 0
        return 1 if default else 0

    def _validate_us_phone_zip(self, phone, zip_code):
        phone_text = (phone or '').strip()
        zip_text = (zip_code or '').strip()
        if phone_text:
            digits = re.sub(r'\D+', '', phone_text)
            if len(digits) == 11 and digits.startswith('1'):
                digits = digits[1:]
            if len(digits) != 10:
                raise ValueError('电话格式无效，请填写美国电话（10位数字，可含+1）')
        if zip_text and not re.match(r'^\d{5}(-\d{4})?$', zip_text):
            raise ValueError('邮编格式无效，请填写美国邮编（5位或5+4）')

    def _normalize_registration_platform_items(self, items):
        normalized = []
        if not isinstance(items, list):
            return normalized
        for entry in items:
            if not isinstance(entry, dict):
                continue
            platform_sku = (entry.get('platform_sku') or '').strip()
            sales_product_id = self._parse_int(entry.get('sales_product_id'))
            quantity = self._parse_int(entry.get('quantity')) or 1
            shipping_plan_id = self._parse_int(entry.get('shipping_plan_id'))
            if not platform_sku and not sales_product_id:
                continue
            normalized.append({
                'sales_product_id': sales_product_id,
                'platform_sku': platform_sku,
                'quantity': max(1, quantity),
                'shipping_plan_id': shipping_plan_id
            })
        return normalized

    def _normalize_registration_shipment_items(self, items):
        normalized = []
        if not isinstance(items, list):
            return normalized
        for entry in items:
            if not isinstance(entry, dict):
                continue
            order_product_id = self._parse_int(entry.get('order_product_id'))
            order_sku = (entry.get('order_sku') or '').strip()
            quantity = self._parse_int(entry.get('quantity')) or 1
            shipping_plan_id = self._parse_int(entry.get('shipping_plan_id'))
            source_type = (entry.get('source_type') or 'manual').strip().lower()
            if source_type not in ('manual', 'auto', 'plan'):
                source_type = 'manual'
            if not order_product_id and not order_sku:
                continue
            normalized.append({
                'order_product_id': order_product_id,
                'order_sku': order_sku,
                'quantity': max(1, quantity),
                'source_type': source_type,
                'shipping_plan_id': shipping_plan_id
            })
        return normalized

    def _normalize_registration_logistics_items(self, items):
        normalized = []
        if not isinstance(items, list):
            return normalized
        for index, entry in enumerate(items, start=1):
            if not isinstance(entry, dict):
                continue
            shipping_carrier = (entry.get('shipping_carrier') or '').strip()
            tracking_no = (entry.get('tracking_no') or '').strip()
            sort_order = self._parse_int(entry.get('sort_order')) or index
            if not shipping_carrier and not tracking_no:
                continue
            normalized.append({
                'shipping_carrier': shipping_carrier or None,
                'tracking_no': tracking_no or None,
                'sort_order': max(1, sort_order)
            })
        return normalized

    def _resolve_registration_auto_shipments(self, conn, platform_items):
        aggregate = {}
        if not platform_items:
            return []

        with conn.cursor() as cur:
            for item in platform_items:
                qty = max(1, self._parse_int(item.get('quantity')) or 1)
                shipping_plan_id = self._parse_int(item.get('shipping_plan_id'))

                if shipping_plan_id:
                    cur.execute(
                        """
                        SELECT op.id AS order_product_id, op.sku, opsi.quantity
                        FROM order_product_shipping_plan_items opsi
                        JOIN order_products op ON op.id = opsi.substitute_order_product_id
                        WHERE opsi.shipping_plan_id=%s
                        ORDER BY opsi.sort_order ASC, opsi.id ASC
                        """,
                        (shipping_plan_id,)
                    )
                    rels = cur.fetchall() or []
                    for rel in rels:
                        key = int(rel.get('order_product_id'))
                        aggregate.setdefault(key, {
                            'order_product_id': key,
                            'order_sku': rel.get('sku') or '',
                            'quantity': 0,
                            'source_type': 'plan',
                            'shipping_plan_id': shipping_plan_id
                        })
                        aggregate[key]['quantity'] += qty * (self._parse_int(rel.get('quantity')) or 1)
                    continue

                sales_product_id = self._parse_int(item.get('sales_product_id'))
                platform_sku = (item.get('platform_sku') or '').strip()
                if not sales_product_id and platform_sku:
                    cur.execute("SELECT id FROM sales_products WHERE platform_sku=%s LIMIT 1", (platform_sku,))
                    row = cur.fetchone() or {}
                    sales_product_id = self._parse_int(row.get('id'))
                if not sales_product_id:
                    continue

                cur.execute(
                    """
                    SELECT op.id AS order_product_id, op.sku, spol.quantity
                    FROM sales_product_order_links spol
                    JOIN order_products op ON op.id = spol.order_product_id
                    WHERE spol.sales_product_id=%s
                    """,
                    (sales_product_id,)
                )
                rels = cur.fetchall() or []
                for rel in rels:
                    key = int(rel.get('order_product_id'))
                    aggregate.setdefault(key, {
                        'order_product_id': key,
                        'order_sku': rel.get('sku') or '',
                        'quantity': 0,
                        'source_type': 'auto',
                        'shipping_plan_id': None
                    })
                    aggregate[key]['quantity'] += qty * (self._parse_int(rel.get('quantity')) or 1)

        items = list(aggregate.values())
        items.sort(key=lambda x: (x.get('order_sku') or '', x.get('order_product_id') or 0))
        return items

    def _normalize_keyword_tag_names(self, value):
        if value is None:
            return []
        raw_items = []
        if isinstance(value, list):
            raw_items = value
        else:
            raw_items = re.split(r'[\n\r;,；]+', str(value))
        seen = set()
        result = []
        for item in raw_items:
            name = ('' if item is None else str(item)).strip()
            if not name:
                continue
            key = name.lower()
            if key in seen:
                continue
            seen.add(key)
            result.append(name[:64])
        return result

    def _normalize_keyword_sku_ids(self, value):
        items = []
        if value is None:
            return items
        if isinstance(value, list):
            raw = value
        else:
            raw = re.split(r'[\n\r,;；]+', str(value))
        for entry in raw:
            sku_id = self._parse_int(entry)
            if sku_id:
                items.append(sku_id)
        return sorted(set(items))

    def _ensure_keyword_tags(self, conn, category_id, tag_names):
        if not tag_names:
            return []
        with conn.cursor() as cur:
            for name in tag_names:
                cur.execute(
                    """
                    INSERT INTO amazon_keyword_tags (category_id, tag_name)
                    VALUES (%s, %s)
                    ON DUPLICATE KEY UPDATE tag_name=VALUES(tag_name)
                    """,
                    (category_id, name)
                )
            placeholders = ','.join(['%s'] * len(tag_names))
            cur.execute(
                f"""
                SELECT id, tag_name
                FROM amazon_keyword_tags
                WHERE category_id=%s
                  AND tag_name IN ({placeholders})
                """,
                [category_id] + tag_names
            )
            rows = cur.fetchall() or []
        name_to_id = {str(row.get('tag_name')): int(row.get('id')) for row in rows if row.get('id')}
        return [name_to_id[name] for name in tag_names if name in name_to_id]

    def _replace_keyword_tags(self, conn, keyword_id, category_id, tag_names):
        tag_ids = self._ensure_keyword_tags(conn, category_id, tag_names)
        with conn.cursor() as cur:
            cur.execute("DELETE FROM amazon_keyword_tag_rel WHERE keyword_id=%s", (keyword_id,))
            if tag_ids:
                cur.executemany(
                    "INSERT IGNORE INTO amazon_keyword_tag_rel (keyword_id, tag_id) VALUES (%s, %s)",
                    [(keyword_id, tag_id) for tag_id in tag_ids]
                )

    def _replace_keyword_sku_relevance(self, conn, keyword_id, sku_family_ids):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM amazon_keyword_sku_rel WHERE keyword_id=%s", (keyword_id,))
            if sku_family_ids:
                cur.executemany(
                    """
                    INSERT INTO amazon_keyword_sku_rel (keyword_id, sku_family_id, relevance_score)
                    VALUES (%s, %s, 1)
                    """,
                    [(keyword_id, sku_id) for sku_id in sku_family_ids]
                )

    def _replace_sales_order_links(self, conn, sales_product_id, links):
        with conn.cursor() as cur:
            cur.execute("DELETE FROM sales_product_order_links WHERE sales_product_id=%s", (sales_product_id,))
        if not links:
            return
        with conn.cursor() as cur:
            cur.executemany(
                """
                INSERT INTO sales_product_order_links (sales_product_id, order_product_id, quantity)
                VALUES (%s, %s, %s)
                """,
                [(sales_product_id, entry['order_product_id'], entry['quantity']) for entry in links]
            )

    def _derive_sales_fields(self, conn, sku_family_id, links):
        """自动推导销售产品的面料、规格名称和平台SKU"""
        if not links:
            return '', '', ''
        
        # 获取货号系列代码
        sku_family_code = ''
        if sku_family_id:
            with conn.cursor() as cur:
                cur.execute("SELECT sku_family FROM product_families WHERE id=%s", (sku_family_id,))
                row = cur.fetchone()
                if row:
                    sku_family_code = (row.get('sku_family') or '').strip()
        
        # 获取下单产品信息
        id_list = [entry['order_product_id'] for entry in links]
        placeholders = ','.join(['%s'] * len(id_list))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT op.id, op.sku, op.spec_qty_short, fm.fabric_code, fm.fabric_name_en
                FROM order_products op
                LEFT JOIN fabric_materials fm ON fm.id = op.fabric_id
                WHERE op.id IN ({placeholders})
                """,
                id_list
            )
            rows = cur.fetchall() or []
        
        row_map = {row['id']: row for row in rows}
        fabrics = []
        spec_parts = []
        for entry in links:
            row = row_map.get(entry['order_product_id'])
            if not row:
                continue
            fabric_code = self._code_before_dash(row.get('fabric_code'))
            if not fabric_code:
                fabric_code = self._code_before_dash(row.get('fabric_name_en'))
            if fabric_code and fabric_code not in fabrics:
                fabrics.append(fabric_code)
            spec_short = (row.get('spec_qty_short') or '').strip()
            if spec_short:
                spec_parts.append(f"{entry['quantity']}{spec_short}")
        
        fabric = ' / '.join(fabrics)
        spec_name = ''.join(spec_parts)
        
        # 自动生成平台SKU: 货号-规格名称-面料编号
        platform_sku = ''
        if sku_family_code and fabric and spec_name:
            first_fabric = fabrics[0] if fabrics else ''
            platform_sku = self._build_sales_platform_sku(sku_family_code, spec_name, first_fabric)
        
        return fabric, spec_name, platform_sku

    def _derive_sales_cost_size(self, conn, links):
        """根据关联下单SKU推导成本与尺寸重量（包裹长宽高取最大值）"""
        if not links:
            return {
                'warehouse_cost_usd': 0.0,
                'last_mile_cost_usd': 0.0,
                'package_length_in': 0.0,
                'package_width_in': 0.0,
                'package_height_in': 0.0,
                'net_weight_lbs': 0.0,
                'gross_weight_lbs': 0.0,
                'sku_family_id': None
            }

        id_list = [entry['order_product_id'] for entry in links]
        placeholders = ','.join(['%s'] * len(id_list))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT id, sku_family_id,
                       cost_usd, last_mile_avg_freight_usd,
                       package_length_in, package_width_in, package_height_in,
                       net_weight_lbs, gross_weight_lbs
                FROM order_products
                WHERE id IN ({placeholders})
                """,
                id_list
            )
            rows = cur.fetchall() or []

        row_map = {row['id']: row for row in rows}
        warehouse_cost_usd = 0.0
        last_mile_cost_usd = 0.0
        package_length_in = 0.0
        package_width_in = 0.0
        package_height_in = 0.0
        net_weight_lbs = 0.0
        gross_weight_lbs = 0.0
        sku_family_id = None

        for entry in links:
            row = row_map.get(entry['order_product_id'])
            if not row:
                continue
            qty = max(1, int(entry.get('quantity') or 1))
            if sku_family_id is None:
                sku_family_id = row.get('sku_family_id')

            warehouse_cost_usd += float(row.get('cost_usd') or 0) * qty
            last_mile_cost_usd += float(row.get('last_mile_avg_freight_usd') or 0) * qty
            package_length_in = max(package_length_in, float(row.get('package_length_in') or 0))
            package_width_in = max(package_width_in, float(row.get('package_width_in') or 0))
            package_height_in = max(package_height_in, float(row.get('package_height_in') or 0))
            net_weight_lbs += float(row.get('net_weight_lbs') or 0) * qty
            gross_weight_lbs += float(row.get('gross_weight_lbs') or 0) * qty

        return {
            'warehouse_cost_usd': round(warehouse_cost_usd, 2),
            'last_mile_cost_usd': round(last_mile_cost_usd, 2),
            'package_length_in': round(package_length_in, 2),
            'package_width_in': round(package_width_in, 2),
            'package_height_in': round(package_height_in, 2),
            'net_weight_lbs': round(net_weight_lbs, 2),
            'gross_weight_lbs': round(gross_weight_lbs, 2),
            'sku_family_id': sku_family_id
        }

    def _get_fabric_folder_bytes(self):
        return self._join_resources('『面料』')

    def _ensure_fabric_folder(self):
        folder = self._get_fabric_folder_bytes()
        if not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
        return folder

    def _get_listing_folder_bytes(self):
        # RESOURCES_PATH_BYTES already points to the decoded child (上架资源),
        # so listing folder is the resources path itself.
        return RESOURCES_PATH_BYTES

    def _ensure_listing_folder(self):
        folder = self._get_listing_folder_bytes()
        if not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
        return folder

    def _ensure_listing_sku_folder(self, sku_family):
        if not sku_family:
            return
        base_folder = self._ensure_listing_folder()
        try:
            sku_bytes = os.fsencode(sku_family)
        except Exception:
            sku_bytes = str(sku_family).encode('utf-8', errors='surrogatepass')
        target = os.path.join(base_folder, sku_bytes)
        if not os.path.exists(target):
            os.makedirs(target, exist_ok=True)
        # Create standard subfolders for the SKU
        for sub in ('源文件', '主图', 'A+', '关联文件', '视频', '上传模板'):
            try:
                sub_bytes = os.fsencode(sub)
            except Exception:
                sub_bytes = str(sub).encode('utf-8', errors='surrogatepass')
            sub_path = os.path.join(target, sub_bytes)
            if not os.path.exists(sub_path):
                os.makedirs(sub_path, exist_ok=True)

        # Ensure default common folders under 主图 and A+
        for parent_sub in ('主图', 'A+'):
            try:
                parent_sub_bytes = os.fsencode(parent_sub)
            except Exception:
                parent_sub_bytes = str(parent_sub).encode('utf-8', errors='surrogatepass')
            parent_path = os.path.join(target, parent_sub_bytes)
            try:
                common_sub_bytes = os.fsencode('通用')
            except Exception:
                common_sub_bytes = '通用'.encode('utf-8', errors='surrogatepass')
            common_path = os.path.join(parent_path, common_sub_bytes)
            if not os.path.exists(common_path):
                os.makedirs(common_path, exist_ok=True)

    def _rename_listing_sku_folder(self, old_sku_family, new_sku_family):
        old_name = (old_sku_family or '').strip()
        new_name = (new_sku_family or '').strip()
        if (not old_name) or (not new_name) or old_name == new_name:
            return {'status': 'success', 'renamed': False}

        base_folder = self._ensure_listing_folder()
        old_path = os.path.join(base_folder, self._safe_fsencode(old_name))
        new_path = os.path.join(base_folder, self._safe_fsencode(new_name))

        if not os.path.exists(old_path):
            # 旧目录不存在时按新名称补齐目录
            self._ensure_listing_sku_folder(new_name)
            return {'status': 'success', 'renamed': False}

        if os.path.exists(new_path):
            return {'status': 'error', 'message': f'目标目录已存在: {new_name}'}

        try:
            os.rename(old_path, new_path)
            return {'status': 'success', 'renamed': True}
        except Exception as e:
            return {'status': 'error', 'message': f'重命名目录失败: {e}'}

    def _code_before_dash(self, value):
        text = (value or '').strip()
        if not text:
            return ''
        return text.split('-', 1)[0].strip() or text

    def _build_sales_platform_sku(self, sku_family_code, spec_name, fabric_code):
        sku_part = (sku_family_code or '').strip()
        spec_part = (spec_name or '').strip()
        fabric_part = self._code_before_dash(fabric_code)
        if not (sku_part and spec_part and fabric_part):
            return ''
        return f"{sku_part}-{spec_part}-{fabric_part}"

    def _ensure_listing_sales_variant_folder(self, sku_family, spec_name, fabric_code):
        sku_name = (sku_family or '').strip()
        if not sku_name:
            return
        self._ensure_listing_sku_folder(sku_name)
        base_folder = self._ensure_listing_folder()
        sku_folder = os.path.join(base_folder, self._safe_fsencode(sku_name))
        main_folder = os.path.join(sku_folder, self._safe_fsencode('主图'))
        if not os.path.exists(main_folder):
            os.makedirs(main_folder, exist_ok=True)

        spec_part = (spec_name or '').strip().replace('/', '-').replace('\\', '-')
        fabric_part = self._code_before_dash(fabric_code).replace('/', '-').replace('\\', '-')
        if not (spec_part and fabric_part):
            return
        variant_folder_name = f"{spec_part}-{fabric_part}"
        variant_folder = os.path.join(main_folder, self._safe_fsencode(variant_folder_name))
        if not os.path.exists(variant_folder):
            os.makedirs(variant_folder, exist_ok=True)

    def _get_certification_folder_bytes(self):
        return self._join_resources('『认证』')

    def _ensure_certification_folder(self):
        folder = self._get_certification_folder_bytes()
        if not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
        return folder

    def _normalize_fabric_image_names(self, image_names):
        if not image_names:
            return []
        if isinstance(image_names, (str, bytes, bytearray)):
            raw = [image_names]
        else:
            raw = list(image_names)
        seen = set()
        result = []
        for name in raw:
            if isinstance(name, (bytes, bytearray)):
                try:
                    name = os.fsdecode(name)
                except Exception:
                    name = name.decode('utf-8', errors='ignore')
            name = (str(name).strip() if name is not None else '')
            if not name or name in seen:
                continue
            seen.add(name)
            result.append(name)
        return result

    def _parse_fabric_images_payload(self, images_data):
        """解析图片数组数据，支持新旧格式
        新格式: [{'image_name': 'xxx', 'remark': '原图/卖点图', 'sort_order': 0}, ...]
        旧格式: ['image_name1', 'image_name2', ...]
        返回: [{'image_name': str, 'remark': str|None, 'sort_order': int, 'is_primary': bool}, ...]
        """
        if not images_data:
            return []
        
        result = []
        if isinstance(images_data, list):
            for idx, item in enumerate(images_data):
                if isinstance(item, dict):
                    # 新格式
                    result.append({
                        'image_name': (item.get('image_name') or '').strip(),
                        'remark': self._normalize_fabric_remark(item.get('remark')),
                        'sort_order': self._parse_int(item.get('sort_order')) or idx,
                        'is_primary': bool(item.get('is_primary', idx == 0))
                    })
                else:
                    # 旧格式字符串
                    result.append({
                        'image_name': str(item).strip(),
                        'remark': '原图',
                        'sort_order': idx,
                        'is_primary': (idx == 0)
                    })
        return [r for r in result if r['image_name']]
    
    def _rename_fabric_image_with_remark(self, old_name, fabric_code, remark, index):
        """根据新命名规则重命名图片: 面料编号-备注-序号
        Args:
            old_name: 原文件名
            fabric_code: 面料编号
            remark: 备注类型（原图/卖点图）
            index: 在该备注类型下的序号（从1开始）
        Returns:
            新文件名，如果文件名已符合规则且序号正确则返回原文件名
        """
        if not old_name:
            return None
        
        ext = os.path.splitext(old_name)[1] or '.jpg'
        remark_str = remark or '未分类'
        new_name = f"{fabric_code}-{remark_str}-{index:02d}{ext}"
        
        # 如果原文件名已经符合新规则，保持不变
        if old_name == new_name:
            return old_name
        
        return new_name

    def _next_fabric_image_index(self, existing_names, fabric_code):
        max_idx = 0
        prefix = f"{fabric_code}_"
        for name in existing_names:
            if not name:
                continue
            if name.startswith(prefix):
                match = re.match(rf"^{re.escape(prefix)}(\\d+)", name)
                if match:
                    try:
                        max_idx = max(max_idx, int(match.group(1)))
                    except Exception:
                        continue
            elif name.startswith(f"{fabric_code}."):
                max_idx = max(max_idx, 1)
        return max_idx + 1

    def _build_fabric_upload_name(self, fabric_code, filename, existing_names):
        ext = os.path.splitext(filename)[1]
        index = self._next_fabric_image_index(existing_names, fabric_code)
        return f"{fabric_code}_{index:02d}{ext}"

    def handle_sku_api(self, environ, method, start_response):
        """货号管理 API（CRUD）"""
        try:
            self._ensure_fabric_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                """
                                SELECT pf.id, pf.sku_family, pf.category, pf.created_at,
                                    GROUP_CONCAT(DISTINCT fm.id ORDER BY fm.id SEPARATOR ',') AS fabric_ids,
                                    GROUP_CONCAT(DISTINCT fm.fabric_code ORDER BY fm.fabric_code SEPARATOR ' / ') AS fabric_codes
                                FROM product_families pf
                                LEFT JOIN fabric_product_families fpf ON fpf.sku_family_id = pf.id
                                LEFT JOIN fabric_materials fm ON fm.id = fpf.fabric_id
                                WHERE pf.sku_family LIKE %s OR pf.category LIKE %s
                                GROUP BY pf.id, pf.sku_family, pf.category, pf.created_at
                                ORDER BY pf.id DESC
                                """,
                                (f"%{keyword}%", f"%{keyword}%")
                            )
                        else:
                            cur.execute(
                                """
                                SELECT pf.id, pf.sku_family, pf.category, pf.created_at,
                                    GROUP_CONCAT(DISTINCT fm.id ORDER BY fm.id SEPARATOR ',') AS fabric_ids,
                                    GROUP_CONCAT(DISTINCT fm.fabric_code ORDER BY fm.fabric_code SEPARATOR ' / ') AS fabric_codes
                                FROM product_families pf
                                LEFT JOIN fabric_product_families fpf ON fpf.sku_family_id = pf.id
                                LEFT JOIN fabric_materials fm ON fm.id = fpf.fabric_id
                                GROUP BY pf.id, pf.sku_family, pf.category, pf.created_at
                                ORDER BY pf.id DESC
                                """
                            )
                        rows = cur.fetchall()
                for row in rows:
                    fabric_ids = row.get('fabric_ids')
                    if fabric_ids:
                        row['fabric_ids'] = [v for v in fabric_ids.split(',') if v]
                    else:
                        row['fabric_ids'] = []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                sku_family = (data.get('sku_family') or '').strip()
                category = (data.get('category') or '').strip()
                fabric_ids = [self._parse_int(v) for v in (data.get('fabric_ids') or [])]
                fabric_ids = [v for v in fabric_ids if v]
                if not sku_family or not category:
                    return self.send_json({'status': 'error', 'message': 'Missing sku_family or category'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "INSERT INTO product_families (sku_family, category) VALUES (%s, %s)",
                            (sku_family, category)
                        )
                        new_id = cur.lastrowid
                    self._replace_sku_family_fabric_ids(conn, new_id, fabric_ids)
                self._ensure_listing_sku_folder(sku_family)
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                sku_family = (data.get('sku_family') or '').strip()
                category = (data.get('category') or '').strip()
                fabric_ids = [self._parse_int(v) for v in (data.get('fabric_ids') or [])]
                fabric_ids = [v for v in fabric_ids if v]
                if not item_id or not sku_family or not category:
                    return self.send_json({'status': 'error', 'message': 'Missing id or fields'}, start_response)

                old_sku_family = None
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT sku_family FROM product_families WHERE id=%s", (item_id,))
                        row = cur.fetchone()
                        if not row:
                            return self.send_json({'status': 'error', 'message': 'SKU not found'}, start_response)
                        old_sku_family = (row.get('sku_family') or '').strip()

                rename_result = self._rename_listing_sku_folder(old_sku_family, sku_family)
                if rename_result.get('status') != 'success':
                    return self.send_json({'status': 'error', 'message': rename_result.get('message') or '重命名目录失败'}, start_response)

                db_updated = False
                with self._get_db_connection() as conn:
                    try:
                        with conn.cursor() as cur:
                            cur.execute(
                                """
                                UPDATE product_families
                                SET sku_family=%s, category=%s
                                WHERE id=%s
                                """,
                                (sku_family, category, item_id)
                            )
                        self._replace_sku_family_fabric_ids(conn, item_id, fabric_ids)
                        db_updated = True
                    except Exception:
                        if rename_result.get('renamed'):
                            self._rename_listing_sku_folder(sku_family, old_sku_family)
                        raise
                if db_updated:
                    self._ensure_listing_sku_folder(sku_family)
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM product_families WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': 'SKU 已存在'}, start_response)
            print("SKU API error: " + str(e))
            return self.send_error(500, str(e), start_response)

    def handle_category_api(self, environ, method, start_response):
        """品类管理 API（CRUD）"""
        try:
            self._ensure_category_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                """
                                SELECT id, category_cn, category_en, category_en_name, created_at
                                FROM product_categories
                                WHERE category_cn LIKE %s OR category_en LIKE %s OR category_en_name LIKE %s
                                ORDER BY id DESC
                                """,
                                (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%")
                            )
                        else:
                            cur.execute(
                                """
                                SELECT id, category_cn, category_en, category_en_name, created_at
                                FROM product_categories
                                ORDER BY id DESC
                                """
                            )
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                category_cn = (data.get('category_cn') or '').strip()
                category_en = (data.get('category_en') or '').strip()
                category_en_name = (data.get('category_en_name') or '').strip()
                if not category_cn or not category_en or not category_en_name:
                    return self.send_json({'status': 'error', 'message': 'Missing category_cn or category_en or category_en_name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "INSERT INTO product_categories (category_cn, category_en, category_en_name) VALUES (%s, %s, %s)",
                            (category_cn, category_en, category_en_name)
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                category_cn = (data.get('category_cn') or '').strip()
                category_en = (data.get('category_en') or '').strip()
                category_en_name = (data.get('category_en_name') or '').strip()
                if not item_id or not category_cn or not category_en or not category_en_name:
                    return self.send_json({'status': 'error', 'message': 'Missing id or fields'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE product_categories
                            SET category_cn=%s, category_en=%s, category_en_name=%s
                            WHERE id=%s
                            """,
                            (category_cn, category_en, category_en_name, item_id)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                return self.send_json({'status': 'error', 'message': '不允许删除品类，请使用编辑维护'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '品类已存在'}, start_response)
            print("Category API error: " + str(e))
            return self.send_error(500, str(e), start_response)

    def handle_fabric_images_api(self, environ, start_response):
        """列出面料文件夹内图片"""
        try:
            query_params = parse_qs(environ.get('QUERY_STRING', '') or '')
            unbound_only = str(query_params.get('unbound', [''])[0]).strip().lower() in ('1', 'true', 'yes')
            current_fabric_id = None
            try:
                raw_fabric_id = str(query_params.get('fabric_id', [''])[0]).strip()
                if raw_fabric_id:
                    current_fabric_id = int(raw_fabric_id)
            except Exception:
                current_fabric_id = None

            bound_name_to_fabric_ids = {}
            bound_b64_to_fabric_ids = {}
            if unbound_only:
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT fabric_id, image_name FROM fabric_images")
                        db_count = 0
                        for row in (cur.fetchall() or []):
                            image_name = (row.get('image_name') or '').strip().replace('\\', '/')
                            if not image_name:
                                continue
                            fid = row.get('fabric_id')
                            db_count += 1
                            # use helper to add many normalization/b64 variants
                            self._add_name_and_b64_variants(bound_name_to_fabric_ids, bound_b64_to_fabric_ids, image_name, fid)


            folder = self._get_fabric_folder_bytes()
            if not os.path.exists(folder):
                return self.send_json({'status': 'success', 'items': []}, start_response)

            items = []
            with os.scandir(folder) as it:
                for entry in it:
                    if entry.is_file(follow_symlinks=False) and self._is_image_name(entry.name):
                        raw = entry.name
                        # raw may be bytes when scanning a bytes path; ensure we capture original bytes
                        if isinstance(raw, (str,)):
                            try:
                                raw_bytes = os.fsencode(raw)
                            except Exception:
                                raw_bytes = raw.encode('utf-8', errors='surrogatepass')
                        else:
                            raw_bytes = raw

                        # Try best-effort decode for display name, ensuring no surrogates in result
                        display = None
                        try:
                            display = os.fsdecode(raw_bytes)
                            # Clean surrogates if any
                            display = display.encode('utf-8', errors='surrogatepass').decode('utf-8', errors='replace')
                        except Exception:
                            try:
                                display = raw_bytes.decode('utf-8', errors='replace')
                            except Exception:
                                try:
                                    display = raw_bytes.decode('gb18030', errors='replace')
                                except Exception:
                                    display = raw_bytes.decode('latin-1', errors='replace')

                        if unbound_only:
                            normalized_display = (display or '').replace('\\', '/').split('/')[-1].strip()
                            try:
                                nd_nfc = unicodedata.normalize('NFC', normalized_display)
                            except Exception:
                                nd_nfc = normalized_display
                            try:
                                nd_nfd = unicodedata.normalize('NFD', nd_nfc)
                            except Exception:
                                nd_nfd = nd_nfc

                            check_ids = set()
                            # check string variants
                            for variant in (nd_nfc, nd_nfc.lower(), nd_nfd, nd_nfd.lower()):
                                if variant:
                                    ids = bound_name_to_fabric_ids.get(variant, set())
                                    if ids:
                                        check_ids |= ids

                            # check base64 of raw bytes
                            try:
                                b64_display_raw = base64.b64encode(raw_bytes).decode('ascii')
                                ids = bound_b64_to_fabric_ids.get(b64_display_raw, set())
                                if ids:
                                    check_ids |= ids
                            except Exception:
                                pass

                            # also check base64 of normalized variants
                            for variant in (nd_nfc, nd_nfd):
                                try:
                                    vb = os.fsencode(variant)
                                    b64_v = base64.b64encode(vb).decode('ascii')
                                    ids = bound_b64_to_fabric_ids.get(b64_v, set())
                                    if ids:
                                        check_ids |= ids
                                except Exception:
                                    pass

                            if check_ids:
                                if current_fabric_id is None or current_fabric_id not in check_ids:
                                    continue

                        # 返回相对于 resources 的字节路径 base64（包含『面料』子目录），
                        # 以便前端直接传回 /api/image-preview 使用
                        try:
                            folder_bytes = os.fsencode('『面料』')
                        except Exception:
                            folder_bytes = '『面料』'.encode('utf-8', errors='surrogatepass')
                        try:
                            rel_bytes = os.path.join(folder_bytes, raw_bytes)
                        except Exception:
                            # fallback: simple concat with os.sep
                            rel_bytes = folder_bytes + os.sep.encode('utf-8', errors='surrogatepass') + raw_bytes
                        b64 = base64.b64encode(rel_bytes).decode('ascii')
                        name_raw_b64 = base64.b64encode(raw_bytes).decode('ascii')
                        items.append({'name': display, 'name_raw_b64': name_raw_b64, 'b64': b64})

            # 按显示名排序
            try:
                items.sort(key=lambda x: (x.get('name') or '').lower())
            except Exception:
                pass
            return self.send_json({'status': 'success', 'items': items}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_certification_images_api(self, environ, start_response):
        """列出认证文件夹内图片"""
        try:
            folder = self._ensure_certification_folder()

            items = []
            with os.scandir(folder) as it:
                for entry in it:
                    if entry.is_file(follow_symlinks=False) and self._is_image_name(entry.name):
                        raw = entry.name
                        if isinstance(raw, str):
                            try:
                                raw_bytes = os.fsencode(raw)
                            except Exception:
                                raw_bytes = raw.encode('utf-8', errors='surrogatepass')
                        else:
                            raw_bytes = bytes(raw)

                        try:
                            name = os.fsdecode(raw_bytes)
                            name = name.encode('utf-8', errors='surrogatepass').decode('utf-8', errors='replace')
                        except Exception:
                            name = raw_bytes.decode('utf-8', errors='replace')

                        try:
                            folder_bytes = os.fsencode('『认证』')
                        except Exception:
                            folder_bytes = '『认证』'.encode('utf-8', errors='surrogatepass')
                        rel_bytes = os.path.join(folder_bytes, raw_bytes)
                        items.append({
                            'name': name,
                            'name_raw_b64': base64.b64encode(raw_bytes).decode('ascii'),
                            'b64': base64.b64encode(rel_bytes).decode('ascii')
                        })

            try:
                items.sort(key=lambda x: (x.get('name') or '').lower())
            except Exception:
                pass
            return self.send_json({'status': 'success', 'items': items}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_fabric_upload_api(self, environ, start_response):
        """上传面料图片（支持多张）"""
        try:
            if environ['REQUEST_METHOD'] != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            # Read raw body once and parse via FieldStorage on a BytesIO buffer
            t_start = time.time()
            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            t_before_read = time.time()
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            t_after_read = time.time()
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            t_before_parse = time.time()
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            t_after_parse = time.time()

            fabric_code = (form.getfirst('fabric_code', '') or '').strip()
            # debug: log incoming FieldStorage info
            try:
                fs_list = getattr(form, 'list', None)
                print('=== Fabric upload debug: FieldStorage list ===')
                if fs_list:
                    for fi in fs_list:
                        try:
                            print('Field name=', getattr(fi, 'name', None), 'filename=', getattr(fi, 'filename', None), 'type=', getattr(fi, 'type', None))
                        except Exception:
                            print('Field entry repr:', repr(fi))
                else:
                    print('FieldStorage.list is empty or missing')
            except Exception as _e:
                print('Failed to inspect FieldStorage:', str(_e))
            if not fabric_code:
                return self.send_json({'status': 'error', 'message': 'Missing fabric_code'}, start_response)

            # Accept files from any multipart field: FieldStorage.list contains all parts
            all_parts = getattr(form, 'list', []) or []

            # collect diagnostics about raw form items (helpful when filename is missing)
            raw_items_info = []
            for idx, it in enumerate(all_parts):
                try:
                    raw_items_info.append({
                        'index': idx,
                        'field_name': getattr(it, 'name', None),
                        'filename': getattr(it, 'filename', None),
                        'type': getattr(it, 'type', None)
                    })
                except Exception:
                    raw_items_info.append({'index': idx, 'error': 'inspect_failed'})

            uploads = []
            for p in all_parts:
                if getattr(p, 'filename', None):
                    try:
                        content = p.file.read() or b''
                    except Exception:
                        content = b''
                    uploads.append({
                        'filename': p.filename,
                        'type': getattr(p, 'type', None),
                        'content': content
                    })

            debug_info = {
                'raw_body_len': len(raw_body),
                'content_type': env_copy.get('CONTENT_TYPE', ''),
                'parts': raw_items_info,
                'uploads_count': len(uploads),
                'timing': {
                    'total_since_start': round(time.time() - t_start, 3),
                    'read_seconds': round(t_after_read - t_before_read, 3),
                    'parse_seconds': round(t_after_parse - t_before_parse, 3)
                }
            }

            # Fallback: parse multipart via email parser if FieldStorage failed to extract files
            if not uploads and raw_body:
                try:
                    from email.parser import BytesParser
                    from email.policy import default
                    ct = env_copy.get('CONTENT_TYPE', '')
                    t_email_before = time.time()
                    if ct.startswith('multipart/form-data'):
                        mime_bytes = (
                            b'Content-Type: ' + ct.encode('utf-8', errors='ignore') +
                            b'\r\nMIME-Version: 1.0\r\n\r\n' + raw_body
                        )
                        msg = BytesParser(policy=default).parsebytes(mime_bytes)
                        if msg.is_multipart():
                            for part in msg.iter_parts():
                                disp = part.get('Content-Disposition', '') or ''
                                filename = part.get_filename()
                                name = part.get_param('name', header='content-disposition')
                                if 'form-data' in disp and (filename or name == 'file'):
                                    payload = part.get_payload(decode=True) or b''
                                    uploads.append({
                                        'filename': filename or '',
                                        'type': part.get_content_type(),
                                        'content': payload
                                    })
                    t_email_after = time.time()
                    debug_info['timing']['email_parse_seconds'] = round(t_email_after - t_email_before, 3)
                except Exception as e:
                    print('Fabric upload fallback parser error:', str(e))

            if not uploads:
                print('Fabric upload: no valid items found, debug:', debug_info)
                return self.send_json({'status': 'error', 'message': 'No valid images uploaded', 'details': debug_info}, start_response)

            folder = self._ensure_fabric_folder()
            existing = set()
            try:
                with os.scandir(folder) as it:
                    for entry in it:
                        if entry.is_file(follow_symlinks=False):
                            name = entry.name
                            if isinstance(name, (bytes, bytearray)):
                                try:
                                    name = os.fsdecode(name)
                                except Exception:
                                    name = name.decode('utf-8', errors='ignore')
                            existing.add(str(name))
            except Exception:
                existing = set()

            saved_names = []
            file_reports = []
            t_before_write = time.time()
            for item in uploads:
                report = {
                    'orig_filename': '',
                    'content_len': 0,
                    'ext_from_name': '',
                    'ext_from_type': '',
                    'ext_from_magic': '',
                    'saved': False,
                    'reason': ''
                }
                try:
                    # Determine filename and extension; accept images even if filename lacks proper ext
                    orig_filename = os.path.basename(item.get('filename') or '')
                    content = item.get('content') or b''
                    report['orig_filename'] = orig_filename
                    report['content_len'] = len(content) if isinstance(content, (bytes, bytearray)) else 0
                    report['ext_from_name'] = os.path.splitext(orig_filename)[1] if orig_filename else ''

                    if report['content_len'] == 0:
                        report['reason'] = 'empty_content'
                        file_reports.append(report)
                        continue

                    # helper to infer extension from magic bytes
                    def infer_ext_from_bytes(b):
                        if not b or len(b) < 4:
                            return ''
                        if b.startswith(b"\xff\xd8\xff"):
                            return '.jpg'
                        if b.startswith(b"\x89PNG"):
                            return '.png'
                        if b.startswith(b"GIF8"):
                            return '.gif'
                        if b.startswith(b"BM"):
                            return '.bmp'
                        if b[0:4] == b'RIFF' and b[8:12] == b'WEBP':
                            return '.webp'
                        return ''

                    ext = ''
                    # try from original filename
                    if orig_filename and self._is_image_name(orig_filename):
                        try:
                            ext = os.path.splitext(orig_filename)[1]
                        except Exception:
                            ext = ''

                    # try from content-type provided by field
                    if not ext and item.get('type'):
                        t = (item.get('type') or '').lower()
                        if 'jpeg' in t or 'jpg' in t:
                            ext = '.jpg'
                        elif 'png' in t:
                            ext = '.png'
                        elif 'gif' in t:
                            ext = '.gif'
                        elif 'bmp' in t:
                            ext = '.bmp'
                        elif 'webp' in t:
                            ext = '.webp'
                        report['ext_from_type'] = ext

                    # try magic bytes
                    if not ext:
                        ext = infer_ext_from_bytes(content)
                        report['ext_from_magic'] = ext

                    if not ext:
                        report['reason'] = '无法推断图片类型'
                        file_reports.append(report)
                        continue

                    # build a base filename (use original name without ext if available, else fabric_code)
                    base_name = (os.path.splitext(orig_filename)[0] or fabric_code)
                    # ensure target name is unique according to naming scheme, even if existing set is stale
                    max_attempts = 500
                    index = self._next_fabric_image_index(existing, fabric_code)
                    target_name = None
                    dest_path = None
                    for _ in range(max_attempts):
                        target_name = f"{fabric_code}_{index:02d}{ext}"
                        dest_path = os.path.join(folder, os.fsencode(target_name))
                        if target_name not in existing and not os.path.exists(dest_path):
                            break
                        existing.add(target_name)
                        index += 1
                    if not target_name or not dest_path or os.path.exists(dest_path):
                        report['reason'] = 'target_exists'
                        file_reports.append(report)
                        continue
                    existing.add(target_name)

                    with open(dest_path, 'wb') as f:
                        f.write(content)
                    saved_names.append(target_name)
                    report['saved'] = True
                    report['reason'] = 'saved'
                    file_reports.append(report)
                except Exception as e:
                    report['reason'] = f'exception: {str(e)}'
                    file_reports.append(report)
            t_after_write = time.time()
            debug_info['timing']['write_seconds'] = round(t_after_write - t_before_write, 3)
            debug_info['timing']['total_seconds'] = round(time.time() - t_start, 3)

            if not saved_names:
                # return detailed diagnostics to help debugging
                details = file_reports if file_reports else debug_info
                return self.send_json({'status': 'error', 'message': 'No valid images uploaded', 'details': details}, start_response)

            return self.send_json({'status': 'success', 'image_names': saved_names}, start_response)
        except Exception as e:
            print("Fabric upload error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_fabric_image_delete_api(self, environ, method, start_response):
        """永久删除面料图片文件（并移除绑定关系）"""
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)

            data = self._read_json_body(environ)
            image_name = (data.get('image_name') or '').strip()
            raw_b64 = (data.get('image_name_raw_b64') or '').strip()
            current_fabric_id = self._parse_int(data.get('fabric_id'))

            if not image_name and not raw_b64:
                return self.send_json({'status': 'error', 'message': 'Missing image_name'}, start_response)

            raw_bytes = None
            if raw_b64:
                try:
                    raw_bytes = base64.b64decode(raw_b64)
                except Exception:
                    raw_bytes = None

            folder = self._get_fabric_folder_bytes()
            if not os.path.exists(folder):
                return self.send_json({'status': 'error', 'message': '面料图片目录不存在'}, start_response)

            name_variants = set()
            if image_name:
                name_variants.add(image_name)
                name_variants.add(os.path.basename(image_name))
            if raw_bytes is not None:
                try:
                    decoded_name = os.fsdecode(raw_bytes)
                    if decoded_name:
                        name_variants.add(decoded_name)
                        name_variants.add(os.path.basename(decoded_name))
                except Exception:
                    pass

            file_candidates = []
            if raw_bytes is not None:
                file_candidates.append(os.path.join(folder, raw_bytes))
            if image_name:
                file_candidates.append(os.path.join(folder, self._safe_fsencode(os.path.basename(image_name))))

            file_path = None
            for candidate in file_candidates:
                if os.path.exists(candidate):
                    file_path = candidate
                    break

            if file_path is None and image_name:
                target_name = os.path.basename(image_name).strip().lower()
                if target_name:
                    with os.scandir(folder) as it:
                        for entry in it:
                            if not entry.is_file(follow_symlinks=False):
                                continue
                            entry_name = str(entry.name or '').strip().lower()
                            if entry_name == target_name:
                                file_path = entry.path
                                break

            if file_path is None:
                return self.send_json({'status': 'error', 'message': '图片文件不存在'}, start_response)

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    variants = [v for v in name_variants if v]
                    if variants:
                        placeholders = ','.join(['%s'] * len(variants))
                        cur.execute(
                            f"SELECT id, fabric_id, image_name FROM fabric_images WHERE image_name IN ({placeholders})",
                            tuple(variants)
                        )
                        bound_rows = cur.fetchall() or []
                    else:
                        bound_rows = []

                    if current_fabric_id:
                        other_rows = [r for r in bound_rows if self._parse_int(r.get('fabric_id')) != current_fabric_id]
                        if other_rows:
                            return self.send_json({'status': 'error', 'message': '该图片仍被其他面料关联，无法永久删除'}, start_response)

                    if variants:
                        placeholders = ','.join(['%s'] * len(variants))
                        if current_fabric_id:
                            cur.execute(
                                f"DELETE FROM fabric_images WHERE fabric_id=%s AND image_name IN ({placeholders})",
                                tuple([current_fabric_id] + variants)
                            )
                        else:
                            cur.execute(
                                f"DELETE FROM fabric_images WHERE image_name IN ({placeholders})",
                                tuple(variants)
                            )

            try:
                os.remove(file_path)
            except Exception as remove_err:
                return self.send_json({'status': 'error', 'message': f'删除文件失败: {remove_err}'}, start_response)

            return self.send_json({'status': 'success'}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_fabric_attach_api(self, environ, start_response):
        """将已存在的面料图片关联并重命名为面料编号下划线序号形式，返回新文件名列表
        接受 JSON: { fabric_code: 'FAB001', items: [ <base64 of raw filename bytes>, ... ] }
        返回: { status: 'success', items: [ {old_b64:..., new_name:...}, ... ] }
        """
        try:
            if environ['REQUEST_METHOD'] != 'POST':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0))
            body = environ['wsgi.input'].read(content_length)
            data = json.loads(body.decode('utf-8')) if body else {}
            fabric_code = (data.get('fabric_code') or '').strip()
            items = data.get('items') or []
            if not fabric_code or not items:
                return self.send_json({'status': 'error', 'message': 'Missing fabric_code or items'}, start_response)

            folder = self._ensure_fabric_folder()
            # build existing names set (decoded strings)
            existing = set()
            try:
                with os.scandir(folder) as it:
                    for entry in it:
                        if entry.is_file(follow_symlinks=False):
                            name = entry.name
                            if isinstance(name, (bytes, bytearray)):
                                try:
                                    name = os.fsdecode(name)
                                except Exception:
                                    name = name.decode('utf-8', errors='ignore')
                            existing.add(str(name))
            except Exception:
                existing = set()

            results = []
            # collect valid source paths
            to_process = []
            for raw_b64 in items:
                try:
                    raw_bytes = base64.b64decode(raw_b64)
                except Exception:
                    continue

                # attempt to build source path bytes
                src = None
                try:
                    src = os.path.join(folder, raw_bytes)
                except Exception:
                    try:
                        name_str = os.fsdecode(raw_bytes)
                    except Exception:
                        try:
                            name_str = raw_bytes.decode('utf-8', errors='surrogatepass')
                        except Exception:
                            name_str = None
                    if name_str:
                        src = os.path.join(folder, os.fsencode(name_str))

                if not src or not os.path.exists(src):
                    # try alternative decode
                    try:
                        name_str = None
                        try:
                            name_str = os.fsdecode(raw_bytes)
                        except Exception:
                            name_str = raw_bytes.decode('utf-8', errors='ignore')
                        alt = os.path.join(folder, os.fsencode(name_str))
                        if os.path.exists(alt):
                            src = alt
                    except Exception:
                        src = None

                if not src or not os.path.exists(src):
                    continue

                to_process.append({'raw_b64': raw_b64, 'raw_bytes': raw_bytes, 'src': src})

            if not to_process:
                return self.send_json({'status': 'success', 'items': []}, start_response)

            # compute starting index: always start from 1 as requested
            next_idx = 1

            # plan final names ensuring uniqueness
            planned = []
            used = set(existing)

            # First, detect files that already follow the naming convention for this fabric
            import re
            already_assigned = []
            remaining = []
            pattern = re.compile(rf"^{re.escape(fabric_code)}_(\d+)\.(.+)$")
            for item in to_process:
                src = item['src']
                src_basename = os.path.basename(src)
                try:
                    src_basename_str = os.fsdecode(src_basename)
                except Exception:
                    try:
                        src_basename_str = src_basename.decode('utf-8', errors='ignore')
                    except Exception:
                        src_basename_str = ''

                m = pattern.match(src_basename_str or '')
                if m:
                    # file already matches FABCODE_##.ext — treat as already assigned
                    assigned_name = src_basename_str
                    already_assigned.append({'raw_b64': item['raw_b64'], 'new_name': assigned_name})
                    used.add(assigned_name)
                else:
                    remaining.append(item)

            # For remaining files, assign sequential names starting from 1, skipping used
            for item in remaining:
                src = item['src']
                src_basename = os.path.basename(src)
                try:
                    src_basename_str = os.fsdecode(src_basename)
                except Exception:
                    try:
                        src_basename_str = src_basename.decode('utf-8', errors='ignore')
                    except Exception:
                        src_basename_str = 'img'
                ext = os.path.splitext(src_basename_str)[1] or ''

                idx = next_idx
                # avoid building list comprehension inside loop repeatedly
                planned_names = set(p['new_name'] for p in planned)
                while True:
                    candidate = f"{fabric_code}_{idx:02d}{ext}"
                    if candidate not in used and candidate not in planned_names:
                        break
                    idx += 1
                planned.append({'raw_b64': item['raw_b64'], 'src': src, 'new_name': candidate})
                used.add(candidate)
                next_idx = idx + 1

            # two-phase rename: first -> temp names, then temp -> final
            import time
            temp_paths = []
            ts = int(time.time() * 1000)
            for j, p in enumerate(planned):
                src = p['src']
                tmp_name = f".tmp_attach_{ts}_{j}"
                tmp_bytes = os.fsencode(tmp_name)
                tmp_path = os.path.join(folder, tmp_bytes)
                try:
                    # ensure tmp_path does not exist
                    if os.path.exists(tmp_path):
                        os.unlink(tmp_path)
                    os.rename(src, tmp_path)
                    temp_paths.append({'tmp': tmp_path, 'new_name': p['new_name'], 'raw_b64': p['raw_b64']})
                except Exception:
                    # failed to move -> skip this item
                    continue

            # now rename temps to final names, backing up any existing dst first
            for idx_tp, tp in enumerate(temp_paths):
                try:
                    dst = os.path.join(folder, os.fsencode(tp['new_name']))
                    backup = None
                    if os.path.exists(dst):
                        # move existing dst to backup tmp to avoid overwrite
                        bak_name = f".bak_attach_{ts}_{idx_tp}"
                        bak_bytes = os.fsencode(bak_name)
                        backup = os.path.join(folder, bak_bytes)
                        try:
                            if os.path.exists(backup):
                                try:
                                    os.unlink(backup)
                                except Exception:
                                    pass
                            os.rename(dst, backup)
                        except Exception:
                            backup = None
                    os.rename(tp['tmp'], dst)
                    # remove backup if present
                    if backup and os.path.exists(backup):
                        try:
                            os.unlink(backup)
                        except Exception:
                            pass
                    results.append({'old_b64': tp['raw_b64'], 'new_name': tp['new_name']})
                except Exception:
                    # try to move back from tmp to original name (best effort)
                    try:
                        orig_bytes = base64.b64decode(tp['raw_b64'])
                        try:
                            orig_path = os.path.join(folder, orig_bytes)
                        except Exception:
                            try:
                                orig_path = os.path.join(folder, os.fsencode(os.fsdecode(orig_bytes)))
                            except Exception:
                                orig_path = None
                        if orig_path and os.path.exists(tp['tmp']):
                            os.rename(tp['tmp'], orig_path)
                    except Exception:
                        pass

            return self.send_json({'status': 'success', 'items': results}, start_response)
        except Exception as e:
            print('Fabric attach error: ' + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_fabric_api(self, environ, method, start_response):
        """面料管理 API（CRUD）"""
        try:
            self._ensure_fabric_table()
            def _normalize_color(value):
                text = str(value or '').strip().upper()
                if not text:
                    return None
                if re.match(r'^#[0-9A-F]{6}$', text):
                    return text
                return None
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                """
                                    SELECT fm.id, fm.fabric_code, fm.fabric_name_en, fm.representative_color, fm.material_id,
                                            m.name AS material_name, m.name_en AS material_name_en,
                                            GROUP_CONCAT(fi.image_name ORDER BY fi.is_primary DESC, fi.sort_order ASC, fi.id ASC SEPARATOR '||') AS image_names,
                                            SUBSTRING_INDEX(
                                                GROUP_CONCAT(fi.image_name ORDER BY fi.is_primary DESC, fi.sort_order ASC, fi.id ASC SEPARATOR '||'),
                                                '||',
                                                1
                                            ) AS image_name,
                                            GROUP_CONCAT(DISTINCT pf.id ORDER BY pf.id SEPARATOR ',') AS sku_family_ids,
                                            GROUP_CONCAT(DISTINCT pf.sku_family ORDER BY pf.sku_family SEPARATOR ' / ') AS sku_family_names,
                                            fm.created_at
                                    FROM fabric_materials fm
                                    LEFT JOIN materials m ON fm.material_id = m.id
                                    LEFT JOIN fabric_images fi ON fi.fabric_id = fm.id
                                    LEFT JOIN fabric_product_families fpf ON fpf.fabric_id = fm.id
                                    LEFT JOIN product_families pf ON pf.id = fpf.sku_family_id
                                    WHERE fm.fabric_code LIKE %s OR fm.fabric_name_en LIKE %s OR m.name LIKE %s OR m.name_en LIKE %s
                                    GROUP BY fm.id, fm.fabric_code, fm.fabric_name_en, fm.representative_color, fm.material_id, m.name, m.name_en, fm.created_at
                                    ORDER BY fm.id DESC
                                """,
                                (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%")
                            )
                        else:
                            cur.execute(
                                """
                                    SELECT fm.id, fm.fabric_code, fm.fabric_name_en, fm.representative_color, fm.material_id,
                                            m.name AS material_name, m.name_en AS material_name_en,
                                            GROUP_CONCAT(fi.image_name ORDER BY fi.is_primary DESC, fi.sort_order ASC, fi.id ASC SEPARATOR '||') AS image_names,
                                            SUBSTRING_INDEX(
                                                GROUP_CONCAT(fi.image_name ORDER BY fi.is_primary DESC, fi.sort_order ASC, fi.id ASC SEPARATOR '||'),
                                                '||',
                                                1
                                            ) AS image_name,
                                            GROUP_CONCAT(DISTINCT pf.id ORDER BY pf.id SEPARATOR ',') AS sku_family_ids,
                                            GROUP_CONCAT(DISTINCT pf.sku_family ORDER BY pf.sku_family SEPARATOR ' / ') AS sku_family_names,
                                            fm.created_at
                                    FROM fabric_materials fm
                                    LEFT JOIN materials m ON fm.material_id = m.id
                                    LEFT JOIN fabric_images fi ON fi.fabric_id = fm.id
                                    LEFT JOIN fabric_product_families fpf ON fpf.fabric_id = fm.id
                                    LEFT JOIN product_families pf ON pf.id = fpf.sku_family_id
                                    GROUP BY fm.id, fm.fabric_code, fm.fabric_name_en, fm.representative_color, fm.material_id, m.name, m.name_en, fm.created_at
                                    ORDER BY fm.id DESC
                                """
                            )
                        rows = cur.fetchall()
                        
                        # 获取每个面料的图片详细信息（包含 remark）
                        fabric_ids = [row['id'] for row in rows]
                        images_map = {}
                        if fabric_ids:
                            placeholders = ','.join(['%s'] * len(fabric_ids))
                            cur.execute(
                                f"""
                                SELECT id, fabric_id, image_name, sort_order, is_primary, remark
                                FROM fabric_images
                                WHERE fabric_id IN ({placeholders})
                                ORDER BY fabric_id, is_primary DESC, sort_order ASC, id ASC
                                """,
                                fabric_ids
                            )
                            image_rows = cur.fetchall()
                            for img in image_rows:
                                image_name = (img.get('image_name') or '').strip()
                                if not image_name:
                                    continue

                                fid = img['fabric_id']
                                if fid not in images_map:
                                    images_map[fid] = []
                                images_map[fid].append({
                                    'id': img['id'],
                                    'image_name': image_name,
                                    'sort_order': img['sort_order'],
                                    'is_primary': img['is_primary'],
                                    'remark': self._normalize_fabric_remark(img.get('remark'))
                                })
                        
                for row in rows:
                    # 用详细图片信息替换简单的 image_names 列表
                    row['images'] = images_map.get(row['id'], [])
                    # 保留向后兼容的 image_names（优先使用明细表结果，避免多表 JOIN 重复）
                    if row['images']:
                        row['image_names'] = [img.get('image_name') for img in row['images'] if img.get('image_name')]
                    else:
                        names = row.get('image_names')
                        if names:
                            row['image_names'] = [name for name in names.split('||') if name]
                        else:
                            row['image_names'] = []
                    sku_ids = row.get('sku_family_ids')
                    if sku_ids:
                        row['sku_family_ids'] = [v for v in sku_ids.split(',') if v]
                    else:
                        row['sku_family_ids'] = []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                fabric_code = (data.get('fabric_code') or '').strip()
                fabric_name_en = (data.get('fabric_name_en') or '').strip()
                representative_color = _normalize_color(data.get('representative_color'))
                material_id = self._parse_int(data.get('material_id'))
                
                # 支持新旧格式
                images_payload = data.get('images') or data.get('image_names') or data.get('image_name')
                images = self._parse_fabric_images_payload(images_payload)
                
                sku_family_ids = [self._parse_int(v) for v in (data.get('sku_family_ids') or [])]
                sku_family_ids = [v for v in sku_family_ids if v]
                
                if not fabric_code or not fabric_name_en or not material_id:
                    return self.send_json({'status': 'error', 'message': 'Missing fields'}, start_response)

                plan = self._build_fabric_image_plan(images, fabric_code)
                if plan['not_ready']:
                    not_ready_preview = '、'.join(plan['not_ready'][:3])
                    suffix = '...' if len(plan['not_ready']) > 3 else ''
                    return self.send_json({
                        'status': 'error',
                        'message': f"检测到图片仍在上传或文件不完整，请稍后重试：{not_ready_preview}{suffix}"
                    }, start_response)
                if plan['missing']:
                    missing_preview = '、'.join(plan['missing'][:3])
                    suffix = '...' if len(plan['missing']) > 3 else ''
                    return self.send_json({
                        'status': 'error',
                        'message': f"图片文件不存在，已取消保存：{missing_preview}{suffix}"
                    }, start_response)

                rename_result = self._execute_fabric_rename_pairs(plan['rename_pairs'])
                if rename_result.get('status') != 'success':
                    return self.send_json(rename_result, start_response)

                rollback_pairs = rename_result.get('rollback_pairs') or []

                try:
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                """
                                INSERT INTO fabric_materials (fabric_code, fabric_name_en, representative_color, material_id)
                                VALUES (%s, %s, %s, %s)
                                """,
                                (fabric_code, fabric_name_en, representative_color, material_id)
                            )
                            new_id = cur.lastrowid

                            for img in plan['planned_images']:
                                cur.execute(
                                    """
                                    INSERT INTO fabric_images (fabric_id, image_name, sort_order, is_primary, remark)
                                    VALUES (%s, %s, %s, %s, %s)
                                    """,
                                    (new_id, img['image_name'], img['sort_order'], int(img['is_primary']), img['remark'])
                                )
                        self._replace_fabric_sku_family_ids(conn, new_id, sku_family_ids)
                except Exception:
                    rollback_result = self._execute_fabric_rename_pairs(rollback_pairs)
                    if rollback_result.get('status') != 'success':
                        print('Fabric POST rollback failed:', rollback_result.get('message'))
                    raise

                image_names = [img['image_name'] for img in plan['planned_images']]
                return self.send_json({'status': 'success', 'id': new_id, 'image_names': image_names}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                fabric_code = (data.get('fabric_code') or '').strip()
                fabric_name_en = (data.get('fabric_name_en') or '').strip()
                representative_color = _normalize_color(data.get('representative_color'))
                material_id = self._parse_int(data.get('material_id'))
                
                # 支持新旧格式
                images_payload = data.get('images') or data.get('image_names') or data.get('image_name')
                images = self._parse_fabric_images_payload(images_payload)
                
                sku_family_ids = [self._parse_int(v) for v in (data.get('sku_family_ids') or [])]
                sku_family_ids = [v for v in sku_family_ids if v]
                
                if not item_id or not fabric_code or not fabric_name_en or not material_id:
                    return self.send_json({'status': 'error', 'message': 'Missing fields'}, start_response)

                plan = self._build_fabric_image_plan(images, fabric_code)
                if plan['not_ready']:
                    not_ready_preview = '、'.join(plan['not_ready'][:3])
                    suffix = '...' if len(plan['not_ready']) > 3 else ''
                    return self.send_json({
                        'status': 'error',
                        'message': f"检测到图片仍在上传或文件不完整，请稍后重试：{not_ready_preview}{suffix}"
                    }, start_response)
                if plan['missing']:
                    missing_preview = '、'.join(plan['missing'][:3])
                    suffix = '...' if len(plan['missing']) > 3 else ''
                    return self.send_json({
                        'status': 'error',
                        'message': f"图片文件不存在，已取消保存：{missing_preview}{suffix}"
                    }, start_response)

                rename_result = self._execute_fabric_rename_pairs(plan['rename_pairs'])
                if rename_result.get('status') != 'success':
                    return self.send_json(rename_result, start_response)

                rollback_pairs = rename_result.get('rollback_pairs') or []

                try:
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                """
                                UPDATE fabric_materials
                                SET fabric_code=%s, fabric_name_en=%s, representative_color=%s, material_id=%s
                                WHERE id=%s
                                """,
                                (fabric_code, fabric_name_en, representative_color, material_id, item_id)
                            )
                            cur.execute("DELETE FROM fabric_images WHERE fabric_id=%s", (item_id,))

                            for img in plan['planned_images']:
                                cur.execute(
                                    """
                                    INSERT INTO fabric_images (fabric_id, image_name, sort_order, is_primary, remark)
                                    VALUES (%s, %s, %s, %s, %s)
                                    """,
                                    (item_id, img['image_name'], img['sort_order'], int(img['is_primary']), img['remark'])
                                )
                        self._replace_fabric_sku_family_ids(conn, item_id, sku_family_ids)
                except Exception:
                    rollback_result = self._execute_fabric_rename_pairs(rollback_pairs)
                    if rollback_result.get('status') != 'success':
                        print('Fabric PUT rollback failed:', rollback_result.get('message'))
                    raise

                image_names = [img['image_name'] for img in plan['planned_images']]
                return self.send_json({'status': 'success', 'image_names': image_names}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM fabric_materials WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '面料编号已存在'}, start_response)
            print("Fabric API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)


    def handle_todo_api(self, environ, method, start_response):
        """待办事项 API（CRUD，每人独立待办）"""
        try:
            self._ensure_todo_tables(lightweight=True)
            user_id = self._get_session_user(environ)
            if not user_id:
                return self.send_json({'status': 'error', 'message': '未登录'}, start_response)

            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                # 获取当前用户的所有待办（包括分配给他的）
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT x.id, x.title, x.detail, x.start_date, x.due_date,
                                   x.reminder_interval_days, x.last_check_time, x.next_check_time,
                                   x.is_recurring, x.status, x.priority, x.created_by,
                                   x.created_by_name, x.created_at
                            FROM (
                                SELECT t.id, t.title, t.detail, t.start_date, t.due_date,
                                       t.reminder_interval_days, t.last_check_time, t.next_check_time,
                                       t.is_recurring, t.status, t.priority, t.created_by,
                                       COALESCE(NULLIF(u.name, ''), u.username) AS created_by_name, t.created_at
                                FROM todos t
                                JOIN users u ON t.created_by = u.id
                                WHERE t.created_by = %s

                                UNION DISTINCT

                                SELECT t.id, t.title, t.detail, t.start_date, t.due_date,
                                       t.reminder_interval_days, t.last_check_time, t.next_check_time,
                                       t.is_recurring, t.status, t.priority, t.created_by,
                                       COALESCE(NULLIF(u.name, ''), u.username) AS created_by_name, t.created_at
                                FROM todo_assignments ta
                                JOIN todos t ON t.id = ta.todo_id
                                JOIN users u ON t.created_by = u.id
                                WHERE ta.assignee_id = %s
                            ) x
                            ORDER BY x.due_date ASC, x.priority DESC, x.id ASC
                            LIMIT 300
                            """,
                            (user_id, user_id)
                        )
                        rows = cur.fetchall()
                        todo_ids = [row['id'] for row in rows if row.get('id')]
                        assignee_map = {}
                        if todo_ids:
                            placeholders = ','.join(['%s'] * len(todo_ids))
                            cur.execute(
                                f"""
                                    SELECT ta.assignee_id, ta.assignment_status,
                                        ta.todo_id,
                                        COALESCE(NULLIF(u.name, ''), u.username) AS name
                                    FROM todo_assignments ta
                                    JOIN users u ON ta.assignee_id = u.id
                                    WHERE ta.todo_id IN ({placeholders})
                                    ORDER BY ta.todo_id ASC, ta.id ASC
                                """,
                                tuple(todo_ids)
                            )
                            for item in cur.fetchall() or []:
                                todo_id = item.get('todo_id')
                                if todo_id not in assignee_map:
                                    assignee_map[todo_id] = []
                                assignee_map[todo_id].append({
                                    'assignee_id': item.get('assignee_id'),
                                    'assignment_status': item.get('assignment_status'),
                                    'name': item.get('name')
                                })

                        todos = []
                        for row in rows:
                            todo_dict = dict(row)
                            assignees = assignee_map.get(row['id'], [])
                            todo_dict['assignees'] = assignees
                            todos.append(todo_dict)
                return self.send_json({'status': 'success', 'items': todos}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                title = (data.get('title') or '').strip()
                detail = (data.get('detail') or '').strip()
                start_date = self._parse_date_str(data.get('start_date'))
                due_date = self._parse_date_str(data.get('due_date'))
                reminder_interval = self._parse_int(data.get('reminder_interval_days')) or 1
                is_recurring = self._parse_int(data.get('is_recurring')) or 0
                priority = self._parse_int(data.get('priority')) or 2
                status = (data.get('status') or 'open').strip().lower()
                assignee_ids = data.get('assignee_ids') or []
                
                if status not in ('open', 'done', 'hold'):
                    status = 'open'
                if priority not in (1, 2, 3):
                    priority = 2
                if not title or not start_date or not due_date:
                    return self.send_json({'status': 'error', 'message': '缺少必要字段'}, start_response)

                now = datetime.now()
                next_check = datetime.strptime(start_date, '%Y-%m-%d') + timedelta(days=reminder_interval)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO todos 
                            (title, detail, start_date, due_date, reminder_interval_days, 
                             last_check_time, next_check_time, is_recurring, status, priority, created_by)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (title, detail, start_date, due_date, reminder_interval, 
                             now, next_check, is_recurring, status, priority, user_id)
                        )
                        todo_id = cur.lastrowid

                        # 添加分配记录（如果有指定待办人）
                        if assignee_ids:
                            for eid in assignee_ids:
                                eid = self._parse_int(eid)
                                if eid:
                                    try:
                                        cur.execute(
                                            """
                                            INSERT INTO todo_assignments 
                                            (todo_id, assignee_id, assignment_status)
                                            VALUES (%s, %s, %s)
                                            """,
                                            (todo_id, eid, 'pending')
                                        )
                                    except Exception:
                                        pass

                return self.send_json({'status': 'success', 'id': todo_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': '缺少待办ID'}, start_response)

                # 检查权限：只有创建人或分配对象可编辑
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT created_by FROM todos WHERE id=%s
                            """,
                            (item_id,)
                        )
                        row = cur.fetchone()
                        if not row or (row['created_by'] != user_id):
                            # 检查是否是分配对象且权限允许编辑
                            cur.execute(
                                """
                                SELECT assignment_status FROM todo_assignments 
                                WHERE todo_id=%s AND assignee_id=%s
                                """,
                                (item_id, user_id)
                            )
                            if not cur.fetchone():
                                return self.send_json({'status': 'error', 'message': '权限不足'}, start_response)

                updates = []
                params = []
                if 'title' in data:
                    updates.append('title=%s')
                    params.append((data.get('title') or '').strip())
                if 'detail' in data:
                    updates.append('detail=%s')
                    params.append((data.get('detail') or '').strip())
                if 'status' in data:
                    status = (data.get('status') or '').strip().lower()
                    if status in ('open', 'done', 'hold'):
                        updates.append('status=%s')
                        params.append(status)
                if 'priority' in data:
                    priority = self._parse_int(data.get('priority'))
                    if priority in (1, 2, 3):
                        updates.append('priority=%s')
                        params.append(priority)

                if not updates:
                    return self.send_json({'status': 'error', 'message': '无可更新字段'}, start_response)

                params.append(item_id)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            f"UPDATE todos SET {', '.join(updates)} WHERE id=%s",
                            tuple(params)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': '缺少待办ID'}, start_response)

                # 只有创建人可删除
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT created_by FROM todos WHERE id=%s",
                            (item_id,)
                        )
                        row = cur.fetchone()
                        if not row or row['created_by'] != user_id:
                            return self.send_json({'status': 'error', 'message': '只有创建人可删除'}, start_response)

                        cur.execute("DELETE FROM todos WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except Exception as e:
            print('Todo API error: ' + str(e))
            import traceback
            traceback.print_exc()
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_calendar_api(self, environ, method, start_response):
        """日历数据 API（按月汇总待办与生日）"""
        try:
            if method != 'GET':
                return self.send_json({'status': 'error', 'message': 'Method not allowed'}, start_response)

            self._ensure_todo_tables(lightweight=True)
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)
            now = datetime.now()
            year = self._parse_int(query_params.get('year', [now.year])[0]) or now.year
            month = self._parse_int(query_params.get('month', [now.month])[0]) or now.month
            if month < 1 or month > 12:
                return self.send_json({'status': 'error', 'message': 'Invalid month'}, start_response)

            days_in_month = calendar.monthrange(year, month)[1]
            start_date = f"{year:04d}-{month:02d}-01"
            end_date = f"{year:04d}-{month:02d}-{days_in_month:02d}"

            days = {}

            def ensure_day(key):
                if key not in days:
                    days[key] = {'todos': [], 'birthdays': []}

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT id, title, detail, due_date, status, priority
                        FROM todos
                        WHERE due_date BETWEEN %s AND %s
                        ORDER BY due_date ASC, priority DESC, id ASC
                        """,
                        (start_date, end_date)
                    )
                    todo_rows = cur.fetchall()

                    cur.execute(
                        """
                        SELECT id, COALESCE(NULLIF(name, ''), username) AS name, phone, birthday
                        FROM users
                        WHERE birthday IS NOT NULL AND MONTH(birthday)=%s
                        ORDER BY DAY(birthday) ASC, id ASC
                        """,
                        (month,)
                    )
                    employee_rows = cur.fetchall()

            for row in todo_rows:
                due = row.get('due_date')
                if hasattr(due, 'strftime'):
                    key = due.strftime('%Y-%m-%d')
                else:
                    key = str(due)
                ensure_day(key)
                days[key]['todos'].append(row)

            for row in employee_rows:
                bday = row.get('birthday')
                if hasattr(bday, 'strftime'):
                    day_num = int(bday.strftime('%d'))
                    month_num = int(bday.strftime('%m'))
                else:
                    try:
                        parts = str(bday).split('-')
                        month_num = int(parts[1])
                        day_num = int(parts[2])
                    except Exception:
                        continue
                if month_num != month:
                    continue
                key = f"{year:04d}-{month:02d}-{day_num:02d}"
                ensure_day(key)
                days[key]['birthdays'].append(row)

            return self.send_json({
                'status': 'success',
                'year': year,
                'month': month,
                'days': days
            }, start_response)
        except Exception as e:
            print('Calendar API error: ' + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_feature_api(self, environ, method, start_response):
        """卖点管理 API（CRUD）"""
        try:
            self._ensure_features_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                """
                                    SELECT f.id, f.name, f.name_en, f.created_at,
                                        GROUP_CONCAT(DISTINCT c.category_cn ORDER BY c.category_cn SEPARATOR ' / ') AS category_cn,
                                        GROUP_CONCAT(DISTINCT c.category_en ORDER BY c.category_en SEPARATOR ' / ') AS category_en,
                                        GROUP_CONCAT(DISTINCT c.id ORDER BY c.id SEPARATOR ',') AS category_ids
                                FROM features f
                                    LEFT JOIN feature_categories fc ON fc.feature_id = f.id
                                    LEFT JOIN product_categories c ON fc.category_id = c.id
                                    WHERE f.name LIKE %s OR f.name_en LIKE %s OR c.category_cn LIKE %s OR c.category_en LIKE %s
                                    GROUP BY f.id
                                    ORDER BY f.id DESC
                                """,
                                (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%")
                            )
                        else:
                            cur.execute(
                                """
                                    SELECT f.id, f.name, f.name_en, f.created_at,
                                        GROUP_CONCAT(DISTINCT c.category_cn ORDER BY c.category_cn SEPARATOR ' / ') AS category_cn,
                                        GROUP_CONCAT(DISTINCT c.category_en ORDER BY c.category_en SEPARATOR ' / ') AS category_en,
                                        GROUP_CONCAT(DISTINCT c.id ORDER BY c.id SEPARATOR ',') AS category_ids
                                FROM features f
                                    LEFT JOIN feature_categories fc ON fc.feature_id = f.id
                                    LEFT JOIN product_categories c ON fc.category_id = c.id
                                    GROUP BY f.id
                                    ORDER BY f.id DESC
                                """
                            )
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                name = (data.get('name') or '').strip()
                name_en = (data.get('name_en') or '').strip()
                raw_category_ids = data.get('category_ids')
                category_ids = [self._parse_int(cid) for cid in (raw_category_ids or [])]
                category_ids = [cid for cid in category_ids if cid]
                if not name or not name_en or not category_ids:
                    return self.send_json({'status': 'error', 'message': 'Missing name, name_en or category_ids'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "INSERT INTO features (name, name_en) VALUES (%s, %s)",
                            (name, name_en)
                        )
                        new_id = cur.lastrowid
                    self._replace_feature_categories(conn, new_id, category_ids)
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                name = (data.get('name') or '').strip()
                name_en = (data.get('name_en') or '').strip()
                raw_category_ids = data.get('category_ids')
                category_ids = [self._parse_int(cid) for cid in (raw_category_ids or [])]
                category_ids = [cid for cid in category_ids if cid]
                if not item_id or not name or not name_en or not category_ids:
                    return self.send_json({'status': 'error', 'message': 'Missing id or fields'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE features
                            SET name=%s, name_en=%s
                            WHERE id=%s
                            """,
                            (name, name_en, item_id)
                        )
                    self._replace_feature_categories(conn, item_id, category_ids)
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM features WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '卖点已存在'}, start_response)
            print("Feature API error: " + str(e))
            return self.send_error(500, str(e), start_response)

    def handle_material_type_api(self, environ, method, start_response):
        """材料类型管理 API（CRUD）"""
        try:
            self._ensure_material_types_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                """
                                SELECT id, name, created_at
                                FROM material_types
                                WHERE name LIKE %s
                                ORDER BY id DESC
                                """,
                                (f"%{keyword}%",)
                            )
                        else:
                            cur.execute(
                                """
                                SELECT id, name, created_at
                                FROM material_types
                                ORDER BY id ASC
                                """
                            )
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                name = (data.get('name') or '').strip()
                if not name:
                    return self.send_json({'status': 'error', 'message': 'Missing name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "INSERT INTO material_types (name) VALUES (%s)",
                            (name,)
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                name = (data.get('name') or '').strip()
                if not item_id or not name:
                    return self.send_json({'status': 'error', 'message': 'Missing id or name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM material_types WHERE id=%s", (item_id,))
                        row = cur.fetchone()
                        if not row:
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                        cur.execute(
                            """
                            UPDATE material_types
                            SET name=%s
                            WHERE id=%s
                            """,
                            (name, item_id)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM material_types WHERE id=%s", (item_id,))
                        row = cur.fetchone()
                        if not row:
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                        cur.execute("DELETE FROM material_types WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '材料类型已存在或被使用'}, start_response)
            print("MaterialType API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_platform_type_api(self, environ, method, start_response):
        """平台类型管理 API（CRUD）"""
        try:
            self._ensure_platform_types_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                """
                                SELECT id, name, created_at
                                FROM platform_types
                                WHERE name LIKE %s
                                ORDER BY id DESC
                                """,
                                (f"%{keyword}%",)
                            )
                        else:
                            cur.execute(
                                """
                                SELECT id, name, created_at
                                FROM platform_types
                                ORDER BY id ASC
                                """
                            )
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                name = (data.get('name') or '').strip()
                if not name:
                    return self.send_json({'status': 'error', 'message': 'Missing name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "INSERT INTO platform_types (name) VALUES (%s)",
                            (name,)
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                name = (data.get('name') or '').strip()
                if not item_id or not name:
                    return self.send_json({'status': 'error', 'message': 'Missing id or name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM platform_types WHERE id=%s", (item_id,))
                        row = cur.fetchone()
                        if not row:
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                        cur.execute(
                            """
                            UPDATE platform_types
                            SET name=%s
                            WHERE id=%s
                            """,
                            (name, item_id)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM platform_types WHERE id=%s", (item_id,))
                        row = cur.fetchone()
                        if not row:
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                        cur.execute("DELETE FROM platform_types WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '平台类型已存在或被使用'}, start_response)
            print("PlatformType API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_brand_api(self, environ, method, start_response):
        """品牌管理 API（CRUD）"""
        try:
            self._ensure_brands_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                """
                                SELECT id, name, created_at
                                FROM brands
                                WHERE name LIKE %s
                                ORDER BY id DESC
                                """,
                                (f"%{keyword}%",)
                            )
                        else:
                            cur.execute(
                                """
                                SELECT id, name, created_at
                                FROM brands
                                ORDER BY id ASC
                                """
                            )
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                name = (data.get('name') or '').strip()
                if not name:
                    return self.send_json({'status': 'error', 'message': 'Missing name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "INSERT INTO brands (name) VALUES (%s)",
                            (name,)
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                name = (data.get('name') or '').strip()
                if not item_id or not name:
                    return self.send_json({'status': 'error', 'message': 'Missing id or name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM brands WHERE id=%s", (item_id,))
                        row = cur.fetchone()
                        if not row:
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                        cur.execute(
                            """
                            UPDATE brands
                            SET name=%s
                            WHERE id=%s
                            """,
                            (name, item_id)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM brands WHERE id=%s", (item_id,))
                        row = cur.fetchone()
                        if not row:
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                        cur.execute("DELETE FROM brands WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '品牌已存在或被使用'}, start_response)
            print("Brand API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_shop_api(self, environ, method, start_response):
        """店铺管理 API（CRUD）"""
        try:
            self._ensure_shops_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                platform_type_id = self._parse_int(query_params.get('platform_type_id', [''])[0].strip())
                brand_id = self._parse_int(query_params.get('brand_id', [''])[0].strip())
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        base_sql = """
                            SELECT s.id, s.shop_name, s.platform_type_id, s.brand_id,
                                   pt.name AS platform_type_name,
                                   b.name AS brand_name,
                                   s.created_at
                            FROM shops s
                            LEFT JOIN platform_types pt ON s.platform_type_id = pt.id
                            LEFT JOIN brands b ON s.brand_id = b.id
                        """
                        filters = []
                        params = []
                        if platform_type_id:
                            filters.append("s.platform_type_id=%s")
                            params.append(platform_type_id)
                        if brand_id:
                            filters.append("s.brand_id=%s")
                            params.append(brand_id)
                        if keyword:
                            filters.append("(s.shop_name LIKE %s OR pt.name LIKE %s OR b.name LIKE %s)")
                            params.extend([f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"])
                        where_sql = (" WHERE " + " AND ".join(filters)) if filters else ""
                        cur.execute(base_sql + where_sql + " ORDER BY s.id DESC", params)
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                shop_name = (data.get('shop_name') or '').strip()
                platform_type_id = self._parse_int(data.get('platform_type_id'))
                brand_id = self._parse_int(data.get('brand_id'))
                if not shop_name or not platform_type_id or not brand_id:
                    return self.send_json({'status': 'error', 'message': 'Missing fields'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO shops (shop_name, platform_type_id, brand_id)
                            VALUES (%s, %s, %s)
                            """,
                            (shop_name, platform_type_id, brand_id)
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                shop_name = (data.get('shop_name') or '').strip()
                platform_type_id = self._parse_int(data.get('platform_type_id'))
                brand_id = self._parse_int(data.get('brand_id'))
                if not item_id or not shop_name or not platform_type_id or not brand_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id or fields'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE shops
                            SET shop_name=%s, platform_type_id=%s, brand_id=%s
                            WHERE id=%s
                            """,
                            (shop_name, platform_type_id, brand_id, item_id)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM shops WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '店铺已存在'}, start_response)
            print("Shop API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_account_health_api(self, environ, method, start_response):
        """Amazon 账户健康管理 API（CRUD + 图表）"""
        try:
            self._ensure_amazon_account_health_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            int_fields = [
                'account_health_rating',
                'suspected_ip_infringement',
                'intellectual_property_complaints',
                'authenticity_customer_complaints',
                'condition_customer_complaints',
                'food_safety_issues',
                'listing_policy_violations',
                'restricted_product_policy_violations',
                'customer_review_policy_violations',
                'other_policy_violations',
                'regulatory_compliance_issues'
            ]
            percent_fields = [
                'order_defect_rate',
                'negative_feedback_rate',
                'a_to_z_rate',
                'chargeback_rate',
                'late_shipment_rate',
                'pre_fulfillment_cancel_rate',
                'valid_tracking_rate',
                'on_time_delivery_rate'
            ]

            if method == 'GET':
                mode = (query_params.get('mode', [''])[0] or '').strip().lower()
                keyword = (query_params.get('q', [''])[0] or '').strip()
                shop_id = self._parse_int((query_params.get('shop_id', [''])[0] or '').strip())
                start_date = self._parse_date_str((query_params.get('start_date', [''])[0] or '').strip())
                end_date = self._parse_date_str((query_params.get('end_date', [''])[0] or '').strip())

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT id FROM platform_types
                            WHERE LOWER(TRIM(name))='amazon'
                            ORDER BY id ASC
                            LIMIT 1
                            """
                        )
                        amazon_platform = cur.fetchone() or {}
                        amazon_platform_id = amazon_platform.get('id')
                        if not amazon_platform_id:
                            return self.send_json({'status': 'success', 'items': []}, start_response)

                        if mode == 'chart':
                            if not shop_id:
                                return self.send_json({'status': 'error', 'message': 'Missing shop_id'}, start_response)
                            cur.execute(
                                "SELECT id FROM shops WHERE id=%s AND platform_type_id=%s",
                                (shop_id, amazon_platform_id)
                            )
                            selected_shop = cur.fetchone()
                            if not selected_shop:
                                return self.send_json({'status': 'error', 'message': 'Shop is not Amazon platform'}, start_response)
                            sql = [
                                """
                                SELECT DATE(a.record_datetime) AS record_date,
                                       ROUND(AVG(a.account_health_rating), 2) AS account_health_rating,
                                       ROUND(AVG(a.order_defect_rate), 4) AS order_defect_rate,
                                       ROUND(AVG(a.late_shipment_rate), 4) AS late_shipment_rate,
                                       ROUND(AVG(a.pre_fulfillment_cancel_rate), 4) AS pre_fulfillment_cancel_rate,
                                       ROUND(AVG(a.valid_tracking_rate), 4) AS valid_tracking_rate,
                                       ROUND(AVG(a.on_time_delivery_rate), 4) AS on_time_delivery_rate
                                FROM amazon_account_health a
                                LEFT JOIN shops s ON s.id = a.shop_id
                                WHERE a.shop_id=%s AND s.platform_type_id=%s
                                """
                            ]
                            params = [shop_id, amazon_platform_id]
                            if start_date:
                                sql.append("AND DATE(a.record_datetime) >= %s")
                                params.append(start_date)
                            if end_date:
                                sql.append("AND DATE(a.record_datetime) <= %s")
                                params.append(end_date)
                            sql.append("GROUP BY DATE(a.record_datetime) ORDER BY DATE(a.record_datetime) ASC")
                            cur.execute("\n".join(sql), params)
                            rows = cur.fetchall()
                            return self.send_json({'status': 'success', 'items': rows}, start_response)

                        sql = [
                            """
                            SELECT a.*, s.shop_name
                            FROM amazon_account_health a
                            LEFT JOIN shops s ON s.id = a.shop_id
                            WHERE s.platform_type_id=%s
                            """
                        ]
                        params = [amazon_platform_id]
                        if shop_id:
                            sql.append("AND a.shop_id=%s")
                            params.append(shop_id)
                        if start_date:
                            sql.append("AND DATE(a.record_datetime) >= %s")
                            params.append(start_date)
                        if end_date:
                            sql.append("AND DATE(a.record_datetime) <= %s")
                            params.append(end_date)
                        if keyword:
                            sql.append("AND (s.shop_name LIKE %s OR a.remark LIKE %s)")
                            params.extend([f"%{keyword}%", f"%{keyword}%"])
                        sql.append("ORDER BY a.record_datetime DESC, a.id DESC")
                        cur.execute("\n".join(sql), params)
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                shop_id = self._parse_int(data.get('shop_id'))
                if not shop_id:
                    return self.send_json({'status': 'error', 'message': 'Missing shop_id'}, start_response)

                values = {}
                for key in int_fields:
                    parsed = self._parse_int(data.get(key))
                    if parsed is None:
                        return self.send_json({'status': 'error', 'message': f'Missing or invalid {key}'}, start_response)
                    values[key] = parsed
                for key in percent_fields:
                    parsed = self._parse_float(data.get(key))
                    if parsed is None:
                        return self.send_json({'status': 'error', 'message': f'Missing or invalid {key}'}, start_response)
                    values[key] = parsed

                record_datetime = self._normalize_datetime_text(data.get('record_datetime')) or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                remark = (data.get('remark') or '').strip()[:500]

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT s.id
                            FROM shops s
                            JOIN platform_types pt ON pt.id = s.platform_type_id
                            WHERE s.id=%s AND LOWER(TRIM(pt.name))='amazon'
                            """,
                            (shop_id,)
                        )
                        allowed_shop = cur.fetchone()
                        if not allowed_shop:
                            return self.send_json({'status': 'error', 'message': 'Only Amazon platform shop is allowed'}, start_response)
                        cur.execute(
                            """
                            INSERT INTO amazon_account_health (
                                shop_id, account_health_rating,
                                suspected_ip_infringement, intellectual_property_complaints,
                                authenticity_customer_complaints, condition_customer_complaints,
                                food_safety_issues, listing_policy_violations,
                                restricted_product_policy_violations, customer_review_policy_violations,
                                other_policy_violations, regulatory_compliance_issues,
                                order_defect_rate, negative_feedback_rate, a_to_z_rate, chargeback_rate,
                                late_shipment_rate, pre_fulfillment_cancel_rate, valid_tracking_rate, on_time_delivery_rate,
                                record_datetime, remark
                            ) VALUES (
                                %s, %s,
                                %s, %s,
                                %s, %s,
                                %s, %s,
                                %s, %s,
                                %s, %s,
                                %s, %s, %s, %s,
                                %s, %s, %s, %s,
                                %s, %s
                            )
                            """,
                            (
                                shop_id, values['account_health_rating'],
                                values['suspected_ip_infringement'], values['intellectual_property_complaints'],
                                values['authenticity_customer_complaints'], values['condition_customer_complaints'],
                                values['food_safety_issues'], values['listing_policy_violations'],
                                values['restricted_product_policy_violations'], values['customer_review_policy_violations'],
                                values['other_policy_violations'], values['regulatory_compliance_issues'],
                                values['order_defect_rate'], values['negative_feedback_rate'], values['a_to_z_rate'], values['chargeback_rate'],
                                values['late_shipment_rate'], values['pre_fulfillment_cancel_rate'], values['valid_tracking_rate'], values['on_time_delivery_rate'],
                                record_datetime, remark
                            )
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                shop_id = self._parse_int(data.get('shop_id'))
                if not item_id or not shop_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id or shop_id'}, start_response)

                values = {}
                for key in int_fields:
                    parsed = self._parse_int(data.get(key))
                    if parsed is None:
                        return self.send_json({'status': 'error', 'message': f'Missing or invalid {key}'}, start_response)
                    values[key] = parsed
                for key in percent_fields:
                    parsed = self._parse_float(data.get(key))
                    if parsed is None:
                        return self.send_json({'status': 'error', 'message': f'Missing or invalid {key}'}, start_response)
                    values[key] = parsed

                record_datetime = self._normalize_datetime_text(data.get('record_datetime')) or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                remark = (data.get('remark') or '').strip()[:500]

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT s.id
                            FROM shops s
                            JOIN platform_types pt ON pt.id = s.platform_type_id
                            WHERE s.id=%s AND LOWER(TRIM(pt.name))='amazon'
                            """,
                            (shop_id,)
                        )
                        allowed_shop = cur.fetchone()
                        if not allowed_shop:
                            return self.send_json({'status': 'error', 'message': 'Only Amazon platform shop is allowed'}, start_response)
                        cur.execute("SELECT id FROM amazon_account_health WHERE id=%s", (item_id,))
                        exists = cur.fetchone()
                        if not exists:
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                        cur.execute(
                            """
                            UPDATE amazon_account_health
                            SET shop_id=%s,
                                account_health_rating=%s,
                                suspected_ip_infringement=%s,
                                intellectual_property_complaints=%s,
                                authenticity_customer_complaints=%s,
                                condition_customer_complaints=%s,
                                food_safety_issues=%s,
                                listing_policy_violations=%s,
                                restricted_product_policy_violations=%s,
                                customer_review_policy_violations=%s,
                                other_policy_violations=%s,
                                regulatory_compliance_issues=%s,
                                order_defect_rate=%s,
                                negative_feedback_rate=%s,
                                a_to_z_rate=%s,
                                chargeback_rate=%s,
                                late_shipment_rate=%s,
                                pre_fulfillment_cancel_rate=%s,
                                valid_tracking_rate=%s,
                                on_time_delivery_rate=%s,
                                record_datetime=%s,
                                remark=%s
                            WHERE id=%s
                            """,
                            (
                                shop_id,
                                values['account_health_rating'],
                                values['suspected_ip_infringement'],
                                values['intellectual_property_complaints'],
                                values['authenticity_customer_complaints'],
                                values['condition_customer_complaints'],
                                values['food_safety_issues'],
                                values['listing_policy_violations'],
                                values['restricted_product_policy_violations'],
                                values['customer_review_policy_violations'],
                                values['other_policy_violations'],
                                values['regulatory_compliance_issues'],
                                values['order_defect_rate'],
                                values['negative_feedback_rate'],
                                values['a_to_z_rate'],
                                values['chargeback_rate'],
                                values['late_shipment_rate'],
                                values['pre_fulfillment_cancel_rate'],
                                values['valid_tracking_rate'],
                                values['on_time_delivery_rate'],
                                record_datetime,
                                remark,
                                item_id
                            )
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_account_health WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            print("AmazonAccountHealth API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_account_health_template_api(self, environ, method, start_response):
        """Amazon 账户健康模板下载"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter

            self._ensure_amazon_account_health_table()
            shop_names = []
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT s.shop_name
                        FROM shops s
                        JOIN platform_types pt ON pt.id = s.platform_type_id
                        WHERE LOWER(TRIM(pt.name))='amazon'
                        ORDER BY s.shop_name
                        """
                    )
                    shop_names = [str(row.get('shop_name') or '').strip() for row in (cur.fetchall() or []) if str(row.get('shop_name') or '').strip()]

            wb = Workbook()
            ws = wb.active
            ws.title = 'amazon_account_health'

            headers = [
                '店铺*', '记录日期时间*', '账户状况评级*',
                '涉嫌侵犯知识产权*', '知识产权投诉*', '商品真实性买家投诉*', '商品状况买家投诉*',
                '食品和商品安全问题*', '上架政策违规*', '违反受限商品政策*', '违反买家商品评论政策*', '其他违反政策*', '监管合规性*',
                '订单缺陷率(%)*', '负面反馈(%)*', 'A-to-z(%)*', '信用卡拒付(%)*',
                '迟发率(%)*', '配送前取消率(%)*', '有效追踪率(%)*', '准时交货率(%)*',
                '备注'
            ]
            ws.append(headers)

            sample_shop = shop_names[0] if shop_names else ''
            ws.append([
                sample_shop, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 260,
                0, 0, 0, 0,
                0, 0, 0, 0, 0, 0,
                0.35, 0.00, 0.00, 0.00,
                1.20, 0.80, 97.50, 95.20,
                '示例行（请勿修改，此行不会导入）'
            ])

            for cell in ws[1]:
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in ws[2]:
                cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                cell.font = Font(italic=True, color='888888')

            widths = [20, 20, 14, 16, 14, 18, 18, 18, 14, 18, 20, 14, 14, 14, 12, 12, 12, 12, 14, 14, 14, 28]
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            options_ws = wb.create_sheet('options')
            options_ws.sheet_state = 'hidden'
            options_ws.cell(row=1, column=1, value='amazon_shop_name')
            for idx, name in enumerate(shop_names, start=2):
                options_ws.cell(row=idx, column=1, value=name)

            if shop_names:
                shop_validation = DataValidation(type='list', formula1=f'=options!$A$2:$A${len(shop_names) + 1}', allow_blank=False)
                ws.add_data_validation(shop_validation)
                for row_idx in range(3, 500):
                    shop_validation.add(f'A{row_idx}')

            ws.freeze_panes = 'A3'
            return self._send_excel_workbook(wb, 'amazon_account_health_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_account_health_import_api(self, environ, method, start_response):
        """Amazon 账户健康批量导入"""
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            headers = [str(cell.value or '').strip() for cell in ws[1]]
            header_map = {name: idx for idx, name in enumerate(headers)}

            def get_cell(row, name):
                idx = header_map.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx].value

            required_headers = [
                '店铺*', '记录日期时间*', '账户状况评级*',
                '涉嫌侵犯知识产权*', '知识产权投诉*', '商品真实性买家投诉*', '商品状况买家投诉*',
                '食品和商品安全问题*', '上架政策违规*', '违反受限商品政策*', '违反买家商品评论政策*', '其他违反政策*', '监管合规性*',
                '订单缺陷率(%)*', '负面反馈(%)*', 'A-to-z(%)*', '信用卡拒付(%)*',
                '迟发率(%)*', '配送前取消率(%)*', '有效追踪率(%)*', '准时交货率(%)*'
            ]
            for col_name in required_headers:
                if col_name not in header_map:
                    return self.send_json({'status': 'error', 'message': f'模板缺少列: {col_name}'}, start_response)

            self._ensure_amazon_account_health_table()
            created = 0
            updated = 0
            unchanged = 0
            errors = []

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT s.id, s.shop_name
                        FROM shops s
                        JOIN platform_types pt ON pt.id = s.platform_type_id
                        WHERE LOWER(TRIM(pt.name))='amazon'
                        """
                    )
                    shop_rows = cur.fetchall() or []
                    shop_map = {str(row.get('shop_name') or '').strip(): int(row.get('id')) for row in shop_rows if row.get('id')}

                for row_idx in range(2, ws.max_row + 1):
                    if row_idx == 2:
                        continue
                    row = ws[row_idx]
                    if not any(cell.value is not None and str(cell.value).strip() for cell in row):
                        continue
                    try:
                        shop_name = str(get_cell(row, '店铺*') or '').strip()
                        shop_id = shop_map.get(shop_name)
                        if not shop_id:
                            raise ValueError(f'店铺不存在或非Amazon平台: {shop_name}')

                        record_datetime = self._normalize_datetime_text(get_cell(row, '记录日期时间*'))
                        if not record_datetime:
                            raise ValueError('记录日期时间格式错误，请使用 YYYY-MM-DD HH:MM:SS 或 YYYY-MM-DDTHH:MM')

                        parsed = {
                            'account_health_rating': self._parse_int(get_cell(row, '账户状况评级*')),
                            'suspected_ip_infringement': self._parse_int(get_cell(row, '涉嫌侵犯知识产权*')),
                            'intellectual_property_complaints': self._parse_int(get_cell(row, '知识产权投诉*')),
                            'authenticity_customer_complaints': self._parse_int(get_cell(row, '商品真实性买家投诉*')),
                            'condition_customer_complaints': self._parse_int(get_cell(row, '商品状况买家投诉*')),
                            'food_safety_issues': self._parse_int(get_cell(row, '食品和商品安全问题*')),
                            'listing_policy_violations': self._parse_int(get_cell(row, '上架政策违规*')),
                            'restricted_product_policy_violations': self._parse_int(get_cell(row, '违反受限商品政策*')),
                            'customer_review_policy_violations': self._parse_int(get_cell(row, '违反买家商品评论政策*')),
                            'other_policy_violations': self._parse_int(get_cell(row, '其他违反政策*')),
                            'regulatory_compliance_issues': self._parse_int(get_cell(row, '监管合规性*')),
                            'order_defect_rate': self._parse_float(get_cell(row, '订单缺陷率(%)*')),
                            'negative_feedback_rate': self._parse_float(get_cell(row, '负面反馈(%)*')),
                            'a_to_z_rate': self._parse_float(get_cell(row, 'A-to-z(%)*')),
                            'chargeback_rate': self._parse_float(get_cell(row, '信用卡拒付(%)*')),
                            'late_shipment_rate': self._parse_float(get_cell(row, '迟发率(%)*')),
                            'pre_fulfillment_cancel_rate': self._parse_float(get_cell(row, '配送前取消率(%)*')),
                            'valid_tracking_rate': self._parse_float(get_cell(row, '有效追踪率(%)*')),
                            'on_time_delivery_rate': self._parse_float(get_cell(row, '准时交货率(%)*')),
                            'remark': str(get_cell(row, '备注') or '').strip()[:500]
                        }

                        for key, value in parsed.items():
                            if key == 'remark':
                                continue
                            if value is None:
                                raise ValueError(f'{key} 为空或格式错误')

                        with conn.cursor() as cur:
                            cur.execute(
                                """
                                SELECT * FROM amazon_account_health
                                WHERE shop_id=%s AND record_datetime=%s
                                ORDER BY id ASC
                                LIMIT 1
                                """,
                                (shop_id, record_datetime)
                            )
                            existing = cur.fetchone()

                            if existing:
                                cur.execute(
                                    """
                                    UPDATE amazon_account_health
                                    SET account_health_rating=%s,
                                        suspected_ip_infringement=%s,
                                        intellectual_property_complaints=%s,
                                        authenticity_customer_complaints=%s,
                                        condition_customer_complaints=%s,
                                        food_safety_issues=%s,
                                        listing_policy_violations=%s,
                                        restricted_product_policy_violations=%s,
                                        customer_review_policy_violations=%s,
                                        other_policy_violations=%s,
                                        regulatory_compliance_issues=%s,
                                        order_defect_rate=%s,
                                        negative_feedback_rate=%s,
                                        a_to_z_rate=%s,
                                        chargeback_rate=%s,
                                        late_shipment_rate=%s,
                                        pre_fulfillment_cancel_rate=%s,
                                        valid_tracking_rate=%s,
                                        on_time_delivery_rate=%s,
                                        remark=%s
                                    WHERE id=%s
                                    """,
                                    (
                                        parsed['account_health_rating'],
                                        parsed['suspected_ip_infringement'],
                                        parsed['intellectual_property_complaints'],
                                        parsed['authenticity_customer_complaints'],
                                        parsed['condition_customer_complaints'],
                                        parsed['food_safety_issues'],
                                        parsed['listing_policy_violations'],
                                        parsed['restricted_product_policy_violations'],
                                        parsed['customer_review_policy_violations'],
                                        parsed['other_policy_violations'],
                                        parsed['regulatory_compliance_issues'],
                                        parsed['order_defect_rate'],
                                        parsed['negative_feedback_rate'],
                                        parsed['a_to_z_rate'],
                                        parsed['chargeback_rate'],
                                        parsed['late_shipment_rate'],
                                        parsed['pre_fulfillment_cancel_rate'],
                                        parsed['valid_tracking_rate'],
                                        parsed['on_time_delivery_rate'],
                                        parsed['remark'],
                                        existing.get('id')
                                    )
                                )
                                if cur.rowcount:
                                    updated += 1
                                else:
                                    unchanged += 1
                            else:
                                cur.execute(
                                    """
                                    INSERT INTO amazon_account_health (
                                        shop_id, account_health_rating,
                                        suspected_ip_infringement, intellectual_property_complaints,
                                        authenticity_customer_complaints, condition_customer_complaints,
                                        food_safety_issues, listing_policy_violations,
                                        restricted_product_policy_violations, customer_review_policy_violations,
                                        other_policy_violations, regulatory_compliance_issues,
                                        order_defect_rate, negative_feedback_rate, a_to_z_rate, chargeback_rate,
                                        late_shipment_rate, pre_fulfillment_cancel_rate, valid_tracking_rate, on_time_delivery_rate,
                                        record_datetime, remark
                                    ) VALUES (
                                        %s, %s,
                                        %s, %s,
                                        %s, %s,
                                        %s, %s,
                                        %s, %s,
                                        %s, %s,
                                        %s, %s, %s, %s,
                                        %s, %s, %s, %s,
                                        %s, %s
                                    )
                                    """,
                                    (
                                        shop_id, parsed['account_health_rating'],
                                        parsed['suspected_ip_infringement'], parsed['intellectual_property_complaints'],
                                        parsed['authenticity_customer_complaints'], parsed['condition_customer_complaints'],
                                        parsed['food_safety_issues'], parsed['listing_policy_violations'],
                                        parsed['restricted_product_policy_violations'], parsed['customer_review_policy_violations'],
                                        parsed['other_policy_violations'], parsed['regulatory_compliance_issues'],
                                        parsed['order_defect_rate'], parsed['negative_feedback_rate'], parsed['a_to_z_rate'], parsed['chargeback_rate'],
                                        parsed['late_shipment_rate'], parsed['pre_fulfillment_cancel_rate'], parsed['valid_tracking_rate'], parsed['on_time_delivery_rate'],
                                        record_datetime, parsed['remark']
                                    )
                                )
                                created += 1
                    except Exception as row_error:
                        errors.append({'row': row_idx, 'error': str(row_error)})

            return self.send_json({
                'status': 'success',
                'created': created,
                'updated': updated,
                'unchanged': unchanged,
                'errors': errors
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_subtype_api(self, environ, method, start_response):
        """Amazon 广告细分类管理 API（CRUD）"""
        try:
            self._ensure_amazon_ad_subtypes_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                """
                                SELECT s.id, s.description, s.ad_class, s.subtype_code, s.created_at, s.updated_at,
                                       GROUP_CONCAT(t.id ORDER BY t.id) AS operation_type_ids,
                                       GROUP_CONCAT(t.name ORDER BY t.id SEPARATOR ' / ') AS operation_type_names
                                    FROM amazon_ad_subtypes s
                                LEFT JOIN amazon_ad_subtype_operation_types so ON so.subtype_id = s.id
                                LEFT JOIN amazon_ad_operation_types t ON t.id = so.operation_type_id
                                WHERE s.description LIKE %s OR s.ad_class LIKE %s OR s.subtype_code LIKE %s
                                GROUP BY s.id, s.description, s.ad_class, s.subtype_code, s.created_at, s.updated_at
                                ORDER BY s.id DESC
                                """,
                                (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%")
                            )
                        else:
                            cur.execute(
                                """
                                SELECT s.id, s.description, s.ad_class, s.subtype_code, s.created_at, s.updated_at,
                                       GROUP_CONCAT(t.id ORDER BY t.id) AS operation_type_ids,
                                       GROUP_CONCAT(t.name ORDER BY t.id SEPARATOR ' / ') AS operation_type_names
                                FROM amazon_ad_subtypes s
                                LEFT JOIN amazon_ad_subtype_operation_types so ON so.subtype_id = s.id
                                LEFT JOIN amazon_ad_operation_types t ON t.id = so.operation_type_id
                                GROUP BY s.id, s.description, s.ad_class, s.subtype_code, s.created_at, s.updated_at
                                ORDER BY s.id DESC
                                """
                            )
                        rows = cur.fetchall()
                for row in rows:
                    raw_ids = row.get('operation_type_ids') or ''
                    row['operation_type_ids'] = [int(v) for v in raw_ids.split(',') if str(v).strip()] if raw_ids else []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                description = (data.get('description') or '').strip()
                ad_class = (data.get('ad_class') or 'SP').strip().upper()
                subtype_code = (data.get('subtype_code') or '').strip()
                operation_type_ids = [self._parse_int(v) for v in (data.get('operation_type_ids') or [])]
                operation_type_ids = [v for v in operation_type_ids if v]
                if ad_class not in ('SP', 'SB', 'SD'):
                    ad_class = 'SP'
                if not description or not subtype_code:
                    return self.send_json({'status': 'error', 'message': 'Missing description or subtype_code'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_subtypes (description, ad_class, subtype_code)
                            VALUES (%s, %s, %s)
                            """,
                            (description, ad_class, subtype_code)
                        )
                        new_id = cur.lastrowid
                    self._replace_ad_subtype_operation_type_ids(conn, new_id, operation_type_ids)
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                description = (data.get('description') or '').strip()
                ad_class = (data.get('ad_class') or 'SP').strip().upper()
                subtype_code = (data.get('subtype_code') or '').strip()
                operation_type_ids = [self._parse_int(v) for v in (data.get('operation_type_ids') or [])]
                operation_type_ids = [v for v in operation_type_ids if v]
                if ad_class not in ('SP', 'SB', 'SD'):
                    ad_class = 'SP'
                if not item_id or not description or not subtype_code:
                    return self.send_json({'status': 'error', 'message': 'Missing id or fields'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE amazon_ad_subtypes
                            SET description=%s, ad_class=%s, subtype_code=%s
                            WHERE id=%s
                            """,
                            (description, ad_class, subtype_code, item_id)
                        )
                    self._replace_ad_subtype_operation_type_ids(conn, item_id, operation_type_ids)
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_ad_subtypes WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '细分类已存在或被引用'}, start_response)
            print("Amazon ad subtype API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_operation_type_api(self, environ, method, start_response):
        """Amazon 广告操作类型 API（CRUD）"""
        try:
            self._ensure_amazon_ad_operation_types_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                """
                                SELECT id, name, apply_portfolio, apply_campaign, apply_group, created_at, updated_at
                                FROM amazon_ad_operation_types
                                WHERE name LIKE %s
                                ORDER BY id DESC
                                """,
                                (f"%{keyword}%",)
                            )
                        else:
                            cur.execute(
                                """
                                SELECT id, name, apply_portfolio, apply_campaign, apply_group, created_at, updated_at
                                FROM amazon_ad_operation_types
                                ORDER BY id DESC
                                """
                            )
                        rows = cur.fetchall()
                        if rows:
                            operation_type_ids = [int(row.get('id')) for row in rows if row.get('id')]
                            placeholders = ','.join(['%s'] * len(operation_type_ids))
                            cur.execute(
                                f"""
                                SELECT id, operation_type_id, reason_name
                                FROM amazon_ad_operation_reasons
                                WHERE operation_type_id IN ({placeholders})
                                ORDER BY id ASC
                                """,
                                operation_type_ids
                            )
                            reason_rows = cur.fetchall()
                            reason_map = {}
                            for reason in reason_rows:
                                op_id = int(reason.get('operation_type_id'))
                                reason_map.setdefault(op_id, []).append({
                                    'id': reason.get('id'),
                                    'reason_name': reason.get('reason_name') or ''
                                })
                            for row in rows:
                                row['reasons'] = reason_map.get(int(row.get('id')), [])
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                name = (data.get('name') or '').strip()
                apply_portfolio = 1 if self._parse_int(data.get('apply_portfolio')) else 0
                apply_campaign = 1 if self._parse_int(data.get('apply_campaign')) else 0
                apply_group = 1 if self._parse_int(data.get('apply_group')) else 0
                reasons = self._normalize_ad_operation_reasons(data.get('reasons') or [])
                if not name:
                    return self.send_json({'status': 'error', 'message': 'Missing name'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "INSERT INTO amazon_ad_operation_types (name, apply_portfolio, apply_campaign, apply_group) VALUES (%s, %s, %s, %s)",
                            (name, apply_portfolio, apply_campaign, apply_group)
                        )
                        new_id = cur.lastrowid
                    self._replace_ad_operation_type_reasons(conn, new_id, reasons)
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)

                batch_items = data.get('items') if isinstance(data, dict) else None
                if isinstance(batch_items, list):
                    normalized_items = []
                    seen_ids = set()
                    for entry in batch_items:
                        if not isinstance(entry, dict):
                            continue
                        item_id = self._parse_int(entry.get('id'))
                        if not item_id or item_id in seen_ids:
                            continue
                        seen_ids.add(item_id)
                        normalized_items.append({
                            'id': item_id,
                            'apply_portfolio': 1 if self._parse_int(entry.get('apply_portfolio')) else 0,
                            'apply_campaign': 1 if self._parse_int(entry.get('apply_campaign')) else 0,
                            'apply_group': 1 if self._parse_int(entry.get('apply_group')) else 0,
                        })

                    if not normalized_items:
                        return self.send_json({'status': 'error', 'message': 'Missing valid items'}, start_response)

                    with self._get_db_connection() as conn:
                        ids = [item['id'] for item in normalized_items]
                        placeholders = ','.join(['%s'] * len(ids))
                        with conn.cursor() as cur:
                            cur.execute(
                                f"SELECT id FROM amazon_ad_operation_types WHERE id IN ({placeholders})",
                                ids
                            )
                            existing_rows = cur.fetchall() or []
                            existing_ids = {int(row.get('id')) for row in existing_rows if row.get('id')}
                            payload_rows = [
                                (item['apply_portfolio'], item['apply_campaign'], item['apply_group'], item['id'])
                                for item in normalized_items
                                if item['id'] in existing_ids
                            ]
                            if payload_rows:
                                cur.executemany(
                                    "UPDATE amazon_ad_operation_types SET apply_portfolio=%s, apply_campaign=%s, apply_group=%s WHERE id=%s",
                                    payload_rows
                                )

                    return self.send_json({'status': 'success'}, start_response)

                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    current = None
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT id, name, apply_portfolio, apply_campaign, apply_group FROM amazon_ad_operation_types WHERE id=%s",
                            (item_id,)
                        )
                        current = cur.fetchone()
                        if not current:
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)

                        has_name = 'name' in data
                        name = (data.get('name') or '').strip() if has_name else (current.get('name') or '')
                        if has_name and not name:
                            return self.send_json({'status': 'error', 'message': 'Missing name'}, start_response)

                        apply_portfolio = current.get('apply_portfolio')
                        if 'apply_portfolio' in data:
                            apply_portfolio = 1 if self._parse_int(data.get('apply_portfolio')) else 0

                        apply_campaign = current.get('apply_campaign')
                        if 'apply_campaign' in data:
                            apply_campaign = 1 if self._parse_int(data.get('apply_campaign')) else 0

                        apply_group = current.get('apply_group')
                        if 'apply_group' in data:
                            apply_group = 1 if self._parse_int(data.get('apply_group')) else 0

                        cur.execute(
                            "UPDATE amazon_ad_operation_types SET name=%s, apply_portfolio=%s, apply_campaign=%s, apply_group=%s WHERE id=%s",
                            (name, apply_portfolio, apply_campaign, apply_group, item_id)
                        )

                    if 'reasons' in data:
                        reasons = self._normalize_ad_operation_reasons(data.get('reasons') or [])
                        self._replace_ad_operation_type_reasons(conn, item_id, reasons)
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_ad_operation_types WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '操作类型已存在或被引用'}, start_response)
            print("Amazon ad operation type API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_api(self, environ, method, start_response):
        """Amazon 广告信息 API（组合/活动/组）"""
        try:
            self._ensure_amazon_ad_tables()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            def _record_create_adjustment(conn, ad_item_id):
                self._ensure_amazon_ad_adjustment_table()
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT id FROM amazon_ad_operation_types
                        WHERE name=%s
                        LIMIT 1
                        """,
                        ('【新建】',)
                    )
                    op_row = cur.fetchone()
                    operation_type_id = op_row.get('id') if op_row else None
                    if not operation_type_id:
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_operation_types (name, apply_portfolio, apply_campaign, apply_group)
                            VALUES (%s, 1, 1, 1)
                            """,
                            ('【新建】',)
                        )
                        operation_type_id = cur.lastrowid

                    cur.execute(
                        """
                        SELECT id FROM amazon_ad_operation_reasons
                        WHERE operation_type_id=%s AND reason_name=%s
                        LIMIT 1
                        """,
                        (operation_type_id, '-')
                    )
                    reason_row = cur.fetchone()
                    reason_id = reason_row.get('id') if reason_row else None
                    if not reason_id:
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_operation_reasons (operation_type_id, reason_name)
                            VALUES (%s, %s)
                            """,
                            (operation_type_id, '-')
                        )
                        reason_id = cur.lastrowid

                    cur.execute(
                        """
                        INSERT INTO amazon_ad_adjustments (
                            adjust_date, ad_item_id, operation_type_id, target_object,
                            before_value, after_value, reason_id, start_time, end_time,
                            impressions, clicks, cost, orders, sales, acos, cpc, ctr, cvr,
                            attribution_checked, attribution_orders, attribution_sales, remark, is_quick_submit
                        ) VALUES (
                            %s, %s, %s, %s,
                            %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s
                        )
                        """,
                        (
                            datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ad_item_id, operation_type_id, '-',
                            '-', '-', reason_id, None, None,
                            None, None, None, None, None, None, None, None, None,
                            0, None, None, '-', 1
                        )
                    )

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                level = (query_params.get('level', [''])[0] or '').strip().lower()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        sql = """
                            SELECT
                                ai.id, ai.ad_level, ai.sku_family_id, ai.portfolio_id, ai.campaign_id,
                                ai.strategy_code, ai.subtype_id, ai.name, ai.is_shared_budget,
                                ai.status, ai.budget, ai.created_at, ai.updated_at,
                                pf.sku_family,
                                p.name AS portfolio_name,
                                c.name AS campaign_name,
                                st.description AS subtype_description,
                                st.ad_class,
                                st.subtype_code
                            FROM amazon_ad_items ai
                            LEFT JOIN product_families pf ON ai.sku_family_id = pf.id
                            LEFT JOIN amazon_ad_items p ON ai.portfolio_id = p.id
                            LEFT JOIN amazon_ad_items c ON ai.campaign_id = c.id
                            LEFT JOIN amazon_ad_subtypes st ON ai.subtype_id = st.id
                        """
                        filters = []
                        params = []
                        if level in ('portfolio', 'campaign', 'group'):
                            filters.append("ai.ad_level=%s")
                            params.append(level)
                        if keyword:
                            filters.append("(ai.name LIKE %s OR pf.sku_family LIKE %s OR st.description LIKE %s)")
                            params.extend([f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"])
                        where_sql = (" WHERE " + " AND ".join(filters)) if filters else ""
                        cur.execute(sql + where_sql + " ORDER BY ai.id DESC", params)
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                ad_level = (data.get('ad_level') or '').strip().lower()
                if ad_level not in ('portfolio', 'campaign', 'group'):
                    return self.send_json({'status': 'error', 'message': 'Invalid ad_level'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if ad_level == 'portfolio':
                            sku_family_id = self._parse_int(data.get('sku_family_id'))
                            is_shared_budget = self._normalize_yes_no(data.get('is_shared_budget'))
                            status = self._normalize_ad_status(data.get('status'))
                            custom_name = (data.get('name') or '').strip()
                            if is_shared_budget is None or not status:
                                return self.send_json({'status': 'error', 'message': 'Missing is_shared_budget/status'}, start_response)
                            portfolio_name = custom_name or (self._build_portfolio_name(conn, sku_family_id) if sku_family_id else '')
                            if not portfolio_name:
                                return self.send_json({'status': 'error', 'message': 'Missing name'}, start_response)
                            cur.execute(
                                """
                                INSERT INTO amazon_ad_items
                                (ad_level, sku_family_id, name, is_shared_budget, status)
                                VALUES ('portfolio', %s, %s, %s, %s)
                                """,
                                (sku_family_id, portfolio_name, is_shared_budget, status)
                            )
                            new_id = cur.lastrowid
                            _record_create_adjustment(conn, new_id)
                            return self.send_json({'status': 'success', 'id': new_id}, start_response)

                        if ad_level == 'campaign':
                            portfolio_id = self._parse_int(data.get('portfolio_id'))
                            strategy_code = (data.get('strategy_code') or '').strip().upper()
                            subtype_id = self._parse_int(data.get('subtype_id'))
                            status = self._normalize_ad_status(data.get('status'))
                            budget = self._parse_float(data.get('budget'))
                            custom_name = (data.get('name') or '').strip()
                            if not portfolio_id or strategy_code not in ('BE', 'BD', 'PC') or not subtype_id or not status:
                                return self.send_json({'status': 'error', 'message': 'Missing portfolio_id/strategy_code/subtype_id/status'}, start_response)
                            row = self._get_ad_item_by_id(conn, portfolio_id)
                            if not row or row.get('ad_level') != 'portfolio':
                                return self.send_json({'status': 'error', 'message': 'Invalid portfolio_id'}, start_response)
                            auto_name = self._build_campaign_name(conn, strategy_code, portfolio_id, subtype_id)
                            campaign_name = custom_name or auto_name
                            if not campaign_name:
                                return self.send_json({'status': 'error', 'message': 'Unable to build campaign name'}, start_response)
                            cur.execute(
                                """
                                INSERT INTO amazon_ad_items
                                (ad_level, portfolio_id, strategy_code, subtype_id, name, status, budget)
                                VALUES ('campaign', %s, %s, %s, %s, %s, %s)
                                """,
                                (portfolio_id, strategy_code, subtype_id, campaign_name, status, budget)
                            )
                            new_id = cur.lastrowid
                            _record_create_adjustment(conn, new_id)
                            return self.send_json({'status': 'success', 'id': new_id}, start_response)

                        campaign_id = self._parse_int(data.get('campaign_id'))
                        provided_portfolio_id = self._parse_int(data.get('portfolio_id'))
                        status = self._normalize_ad_status(data.get('status'))
                        group_name = (data.get('name') or '').strip()
                        if not campaign_id or not group_name:
                            return self.send_json({'status': 'error', 'message': 'Missing campaign_id or name'}, start_response)
                        row = self._get_ad_item_by_id(conn, campaign_id)
                        if not row or row.get('ad_level') != 'campaign':
                            return self.send_json({'status': 'error', 'message': 'Invalid campaign_id'}, start_response)
                        campaign_portfolio_id = row.get('portfolio_id')
                        if provided_portfolio_id and str(provided_portfolio_id) != str(campaign_portfolio_id or ''):
                            return self.send_json({'status': 'error', 'message': 'portfolio_id does not match campaign'}, start_response)
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_items
                            (ad_level, campaign_id, portfolio_id, name, status)
                            VALUES ('group', %s, %s, %s, %s)
                            """,
                            (campaign_id, campaign_portfolio_id, group_name, status)
                        )
                        new_id = cur.lastrowid
                        _record_create_adjustment(conn, new_id)
                        return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)

                with self._get_db_connection() as conn:
                    current = self._get_ad_item_by_id(conn, item_id)
                    if not current:
                        return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                    ad_level = current.get('ad_level')
                    with conn.cursor() as cur:
                        if ad_level == 'portfolio':
                            sku_family_id = self._parse_int(data.get('sku_family_id'))
                            is_shared_budget = self._normalize_yes_no(data.get('is_shared_budget'))
                            status = self._normalize_ad_status(data.get('status'))
                            custom_name = (data.get('name') or '').strip()
                            if is_shared_budget is None or not status:
                                return self.send_json({'status': 'error', 'message': 'Missing is_shared_budget/status'}, start_response)
                            portfolio_name = custom_name or (self._build_portfolio_name(conn, sku_family_id) if sku_family_id else '')
                            if not portfolio_name:
                                return self.send_json({'status': 'error', 'message': 'Missing name'}, start_response)
                            cur.execute(
                                """
                                UPDATE amazon_ad_items
                                SET sku_family_id=%s, name=%s, is_shared_budget=%s, status=%s
                                WHERE id=%s
                                """,
                                (sku_family_id, portfolio_name, is_shared_budget, status, item_id)
                            )
                            return self.send_json({'status': 'success'}, start_response)

                        if ad_level == 'campaign':
                            portfolio_id = self._parse_int(data.get('portfolio_id'))
                            strategy_code = (data.get('strategy_code') or '').strip().upper()
                            subtype_id = self._parse_int(data.get('subtype_id'))
                            status = self._normalize_ad_status(data.get('status'))
                            budget = self._parse_float(data.get('budget'))
                            custom_name = (data.get('name') or '').strip()
                            if not portfolio_id or strategy_code not in ('BE', 'BD', 'PC') or not subtype_id or not status:
                                return self.send_json({'status': 'error', 'message': 'Missing portfolio_id/strategy_code/subtype_id/status'}, start_response)
                            row = self._get_ad_item_by_id(conn, portfolio_id)
                            if not row or row.get('ad_level') != 'portfolio':
                                return self.send_json({'status': 'error', 'message': 'Invalid portfolio_id'}, start_response)
                            auto_name = self._build_campaign_name(conn, strategy_code, portfolio_id, subtype_id)
                            campaign_name = custom_name or auto_name
                            if not campaign_name:
                                return self.send_json({'status': 'error', 'message': 'Unable to build campaign name'}, start_response)
                            cur.execute(
                                """
                                UPDATE amazon_ad_items
                                SET portfolio_id=%s, strategy_code=%s, subtype_id=%s, name=%s, status=%s, budget=%s
                                WHERE id=%s
                                """,
                                (portfolio_id, strategy_code, subtype_id, campaign_name, status, budget, item_id)
                            )
                            return self.send_json({'status': 'success'}, start_response)

                        campaign_id = self._parse_int(data.get('campaign_id'))
                        provided_portfolio_id = self._parse_int(data.get('portfolio_id'))
                        group_name = (data.get('name') or '').strip()
                        status = self._normalize_ad_status(data.get('status'))
                        if not campaign_id or not group_name:
                            return self.send_json({'status': 'error', 'message': 'Missing campaign_id or name'}, start_response)
                        row = self._get_ad_item_by_id(conn, campaign_id)
                        if not row or row.get('ad_level') != 'campaign':
                            return self.send_json({'status': 'error', 'message': 'Invalid campaign_id'}, start_response)
                        campaign_portfolio_id = row.get('portfolio_id')
                        if provided_portfolio_id and str(provided_portfolio_id) != str(campaign_portfolio_id or ''):
                            return self.send_json({'status': 'error', 'message': 'portfolio_id does not match campaign'}, start_response)
                        cur.execute(
                            """
                            UPDATE amazon_ad_items
                            SET campaign_id=%s, portfolio_id=%s, name=%s, status=%s
                            WHERE id=%s
                            """,
                            (campaign_id, campaign_portfolio_id, group_name, status, item_id)
                        )
                        return self.send_json({'status': 'success'}, start_response)

            if method == 'PATCH':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                status = self._normalize_ad_status(data.get('status'))
                if not item_id or not status:
                    return self.send_json({'status': 'error', 'message': 'Missing id or status'}, start_response)

                with self._get_db_connection() as conn:
                    current = self._get_ad_item_by_id(conn, item_id)
                    if not current:
                        return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE amazon_ad_items
                            SET status=%s
                            WHERE id=%s
                            """,
                            (status, item_id)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_ad_items WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '广告记录已存在或被引用'}, start_response)
            print("Amazon ad API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_template_api(self, environ, method, start_response):
        """Amazon 广告信息模板下载"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import FormulaRule

            self._ensure_amazon_ad_tables()

            sku_values = []
            portfolio_values = []
            campaign_values = []
            subtype_values = []
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT pf.sku_family
                        FROM product_families pf
                        ORDER BY pf.sku_family
                        """
                    )
                    sku_values = [str(r.get('sku_family') or '').strip() for r in (cur.fetchall() or []) if str(r.get('sku_family') or '').strip()]

                    cur.execute("SELECT name FROM amazon_ad_items WHERE ad_level='portfolio' ORDER BY name")
                    portfolio_values = [str(r.get('name') or '').strip() for r in (cur.fetchall() or []) if str(r.get('name') or '').strip()]

                    cur.execute("SELECT name FROM amazon_ad_items WHERE ad_level='campaign' ORDER BY name")
                    campaign_values = [str(r.get('name') or '').strip() for r in (cur.fetchall() or []) if str(r.get('name') or '').strip()]

                    cur.execute(
                        """
                        SELECT ad_class, subtype_code
                        FROM amazon_ad_subtypes
                        ORDER BY id ASC
                        """
                    )
                    subtype_values = [
                        f"{str(r.get('ad_class') or '').strip()}-{str(r.get('subtype_code') or '').strip()}"
                        for r in (cur.fetchall() or [])
                        if str(r.get('ad_class') or '').strip() or str(r.get('subtype_code') or '').strip()
                    ]

            wb = Workbook()
            ws = wb.active
            ws.title = 'amazon_ad_items'
            headers = [
                '广告类型*', '状态*', '广告名称*',
                '关联货号(仅Portfolio可填)', '是否共享预算(仅Portfolio必填)',
                '归属广告组合(仅Campaign/Group必填)', '策略(仅Campaign必填)', '细分类(仅Campaign必填，格式 ad_class-subtype_code)',
                '预算(仅Campaign选填)', '归属广告活动(仅Group必填)',
                '备注(导入忽略)'
            ]
            ws.append(headers)
            ws.append([
                'portfolio', '启动', '示例-广告组合',
                '', '是',
                '', '', '',
                '', '',
                '示例行（请勿修改，此行不会导入）'
            ])

            for cell in ws[1]:
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in ws[2]:
                cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                cell.font = Font(italic=True, color='888888')

            widths = [14, 10, 30, 22, 20, 30, 14, 28, 14, 30, 26]
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            options_ws = wb.create_sheet('options')
            options_ws.sheet_state = 'hidden'
            option_groups = {
                'ad_level': ['portfolio', 'campaign', 'group'],
                'status': ['启动', '暂停', '存档'],
                'yes_no': ['是', '否'],
                'strategy': ['BE', 'BD', 'PC'],
                'sku': sku_values,
                'portfolio': portfolio_values,
                'campaign': campaign_values,
                'subtype': subtype_values,
            }

            col_idx = 1
            option_ranges = {}
            for key, values in option_groups.items():
                options_ws.cell(row=1, column=col_idx, value=key)
                for r, val in enumerate(values, start=2):
                    options_ws.cell(row=r, column=col_idx, value=val)
                if values:
                    letter = get_column_letter(col_idx)
                    option_ranges[key] = f"=options!${letter}$2:${letter}${len(values)+1}"
                col_idx += 1

            def _add_validation(col_letter, formula, allow_blank=True):
                if not formula:
                    return
                validation = DataValidation(type='list', formula1=formula, allow_blank=allow_blank)
                ws.add_data_validation(validation)
                for row_idx in range(3, 501):
                    validation.add(f'{col_letter}{row_idx}')

            _add_validation('A', option_ranges.get('ad_level'), allow_blank=False)
            _add_validation('B', option_ranges.get('status'), allow_blank=False)
            _add_validation('D', option_ranges.get('sku'))
            _add_validation('E', option_ranges.get('yes_no'))
            _add_validation('F', option_ranges.get('portfolio'))
            _add_validation('G', option_ranges.get('strategy'))
            _add_validation('H', option_ranges.get('subtype'))
            _add_validation('J', option_ranges.get('campaign'))

            gray_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            ws.conditional_formatting.add('D3:E500', FormulaRule(formula=['$A3<>"portfolio"'], stopIfTrue=False, fill=gray_fill))
            ws.conditional_formatting.add('F3:F500', FormulaRule(formula=['$A3="portfolio"'], stopIfTrue=False, fill=gray_fill))
            ws.conditional_formatting.add('G3:I500', FormulaRule(formula=['$A3<>"campaign"'], stopIfTrue=False, fill=gray_fill))
            ws.conditional_formatting.add('J3:J500', FormulaRule(formula=['$A3<>"group"'], stopIfTrue=False, fill=gray_fill))

            ws.freeze_panes = 'A3'
            return self._send_excel_workbook(wb, 'amazon_ad_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_import_api(self, environ, method, start_response):
        """Amazon 广告信息批量导入"""
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            headers = [str(cell.value or '').strip() for cell in ws[1]]
            header_map = {name: idx for idx, name in enumerate(headers)}

            def get_cell(row, name):
                idx = header_map.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx].value

            required_headers = ['广告类型*', '状态*', '广告名称*']
            for col_name in required_headers:
                if col_name not in header_map:
                    return self.send_json({'status': 'error', 'message': f'模板缺少列: {col_name}'}, start_response)

            self._ensure_amazon_ad_tables()
            created = 0
            updated = 0
            unchanged = 0
            errors = []

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id, sku_family FROM product_families")
                    sku_map = {str(r.get('sku_family') or '').strip(): int(r.get('id')) for r in (cur.fetchall() or []) if r.get('id')}

                    cur.execute("SELECT id, name FROM amazon_ad_items WHERE ad_level='portfolio' ORDER BY id ASC")
                    portfolio_rows = cur.fetchall() or []
                    portfolio_map = {str(r.get('name') or '').strip(): int(r.get('id')) for r in portfolio_rows if str(r.get('name') or '').strip()}

                    cur.execute("SELECT id, name, portfolio_id FROM amazon_ad_items WHERE ad_level='campaign' ORDER BY id ASC")
                    campaign_rows = cur.fetchall() or []
                    campaign_map = {str(r.get('name') or '').strip(): {'id': int(r.get('id')), 'portfolio_id': r.get('portfolio_id')} for r in campaign_rows if str(r.get('name') or '').strip() and r.get('id')}

                    cur.execute("SELECT id, ad_class, subtype_code FROM amazon_ad_subtypes")
                    subtype_rows = cur.fetchall() or []
                    subtype_map = {
                        f"{str(r.get('ad_class') or '').strip()}-{str(r.get('subtype_code') or '').strip()}": int(r.get('id'))
                        for r in subtype_rows if r.get('id')
                    }

                for row_idx in range(2, ws.max_row + 1):
                    if row_idx == 2:
                        continue
                    row = ws[row_idx]
                    if not any(cell.value is not None and str(cell.value).strip() for cell in row):
                        continue
                    try:
                        raw_level = str(get_cell(row, '广告类型*') or '').strip().lower()
                        level_map = {'广告组合': 'portfolio', '广告活动': 'campaign', '广告组': 'group'}
                        ad_level = level_map.get(raw_level, raw_level)
                        if ad_level not in ('portfolio', 'campaign', 'group'):
                            raise ValueError('广告类型仅支持 portfolio/campaign/group')

                        status = self._normalize_ad_status(get_cell(row, '状态*'))
                        name = str(get_cell(row, '广告名称*') or '').strip()
                        if not status or not name:
                            raise ValueError('状态和广告名称不能为空')

                        if ad_level == 'portfolio':
                            sku_family_text = str(get_cell(row, '关联货号(仅Portfolio可填)') or '').strip()
                            sku_family_id = sku_map.get(sku_family_text) if sku_family_text else None
                            shared_budget = self._normalize_yes_no(get_cell(row, '是否共享预算(仅Portfolio必填)'))
                            if shared_budget is None:
                                raise ValueError('Portfolio 的“是否共享预算”必填，且仅支持 是/否')

                            with conn.cursor() as cur:
                                cur.execute("SELECT id FROM amazon_ad_items WHERE ad_level='portfolio' AND name=%s LIMIT 1", (name,))
                                existing = cur.fetchone()
                                if existing:
                                    cur.execute(
                                        """
                                        UPDATE amazon_ad_items
                                        SET sku_family_id=%s, is_shared_budget=%s, status=%s
                                        WHERE id=%s
                                        """,
                                        (sku_family_id, shared_budget, status, existing.get('id'))
                                    )
                                    if cur.rowcount:
                                        updated += 1
                                    else:
                                        unchanged += 1
                                else:
                                    cur.execute(
                                        """
                                        INSERT INTO amazon_ad_items (ad_level, sku_family_id, name, is_shared_budget, status)
                                        VALUES ('portfolio', %s, %s, %s, %s)
                                        """,
                                        (sku_family_id, name, shared_budget, status)
                                    )
                                    created += 1
                                    portfolio_map[name] = int(cur.lastrowid)

                        elif ad_level == 'campaign':
                            portfolio_name = str(get_cell(row, '归属广告组合(仅Campaign/Group必填)') or '').strip()
                            strategy_code = str(get_cell(row, '策略(仅Campaign必填)') or '').strip().upper()
                            subtype_text = str(get_cell(row, '细分类(仅Campaign必填，格式 ad_class-subtype_code)') or '').strip()
                            budget = self._parse_float(get_cell(row, '预算(仅Campaign选填)'))
                            portfolio_id = portfolio_map.get(portfolio_name)
                            subtype_id = subtype_map.get(subtype_text)
                            if not portfolio_id:
                                raise ValueError(f'无效归属广告组合: {portfolio_name}')
                            if strategy_code not in ('BE', 'BD', 'PC'):
                                raise ValueError('Campaign 的策略仅支持 BE/BD/PC')
                            if not subtype_id:
                                raise ValueError(f'无效细分类: {subtype_text}')

                            with conn.cursor() as cur:
                                cur.execute("SELECT id FROM amazon_ad_items WHERE ad_level='campaign' AND portfolio_id=%s AND name=%s LIMIT 1", (portfolio_id, name))
                                existing = cur.fetchone()
                                if existing:
                                    cur.execute(
                                        """
                                        UPDATE amazon_ad_items
                                        SET strategy_code=%s, subtype_id=%s, status=%s, budget=%s
                                        WHERE id=%s
                                        """,
                                        (strategy_code, subtype_id, status, budget, existing.get('id'))
                                    )
                                    if cur.rowcount:
                                        updated += 1
                                    else:
                                        unchanged += 1
                                else:
                                    cur.execute(
                                        """
                                        INSERT INTO amazon_ad_items (ad_level, portfolio_id, strategy_code, subtype_id, name, status, budget)
                                        VALUES ('campaign', %s, %s, %s, %s, %s, %s)
                                        """,
                                        (portfolio_id, strategy_code, subtype_id, name, status, budget)
                                    )
                                    created += 1
                                    campaign_map[name] = {'id': int(cur.lastrowid), 'portfolio_id': portfolio_id}

                        else:
                            portfolio_name = str(get_cell(row, '归属广告组合(仅Campaign/Group必填)') or '').strip()
                            campaign_name = str(get_cell(row, '归属广告活动(仅Group必填)') or '').strip()
                            campaign_info = campaign_map.get(campaign_name)
                            if not campaign_info:
                                raise ValueError(f'无效归属广告活动: {campaign_name}')
                            campaign_id = int(campaign_info.get('id'))
                            campaign_portfolio_id = campaign_info.get('portfolio_id')
                            if portfolio_name:
                                portfolio_id = portfolio_map.get(portfolio_name)
                                if not portfolio_id:
                                    raise ValueError(f'无效归属广告组合: {portfolio_name}')
                                if str(portfolio_id) != str(campaign_portfolio_id or ''):
                                    raise ValueError('Group 的归属广告组合与归属广告活动不一致')
                            else:
                                portfolio_id = campaign_portfolio_id

                            with conn.cursor() as cur:
                                cur.execute("SELECT id FROM amazon_ad_items WHERE ad_level='group' AND campaign_id=%s AND name=%s LIMIT 1", (campaign_id, name))
                                existing = cur.fetchone()
                                if existing:
                                    cur.execute(
                                        """
                                        UPDATE amazon_ad_items
                                        SET portfolio_id=%s, status=%s
                                        WHERE id=%s
                                        """,
                                        (portfolio_id, status, existing.get('id'))
                                    )
                                    if cur.rowcount:
                                        updated += 1
                                    else:
                                        unchanged += 1
                                else:
                                    cur.execute(
                                        """
                                        INSERT INTO amazon_ad_items (ad_level, campaign_id, portfolio_id, name, status)
                                        VALUES ('group', %s, %s, %s, %s)
                                        """,
                                        (campaign_id, portfolio_id, name, status)
                                    )
                                    created += 1
                    except Exception as row_error:
                        errors.append({'row': row_idx, 'error': str(row_error)})

            return self.send_json({
                'status': 'success',
                'created': created,
                'updated': updated,
                'unchanged': unchanged,
                'errors': errors
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_delivery_api(self, environ, method, start_response):
        """Amazon 广告投放管理 API"""
        try:
            self._ensure_amazon_ad_delivery_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            def _resolve_schedule(data):
                updated_at_text = self._normalize_datetime_text(data.get('updated_at'))
                if not updated_at_text:
                    updated_at_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                observe_days = self._normalize_observe_days(data.get('observe_days'))
                if observe_days is None:
                    observe_days = self._normalize_observe_days(data.get('observe_interval'))
                if observe_days is None:
                    observe_days = 1
                observe_interval = f"{observe_days}天"
                next_observe_at = self._normalize_datetime_text(data.get('next_observe_at'))
                if not next_observe_at:
                    base_dt = datetime.strptime(updated_at_text, '%Y-%m-%d %H:%M:%S')
                    next_observe_at = (base_dt + timedelta(days=observe_days)).strftime('%Y-%m-%d %H:%M:%S')
                return updated_at_text, observe_interval, next_observe_at

            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        sql = """
                            SELECT
                                d.id,
                                d.status,
                                d.ad_item_id,
                                d.delivery_desc,
                                d.bid_value,
                                d.observe_interval,
                                d.next_observe_at,
                                d.created_at,
                                d.updated_at,
                                ai.ad_level,
                                ai.name AS ad_name,
                                p.name AS portfolio_name
                            FROM amazon_ad_deliveries d
                            LEFT JOIN amazon_ad_items ai ON ai.id = d.ad_item_id
                            LEFT JOIN amazon_ad_items p ON p.id = ai.portfolio_id
                        """
                        filters = []
                        params = []
                        if keyword:
                            filters.append("(d.delivery_desc LIKE %s OR ai.name LIKE %s OR p.name LIKE %s OR (CASE WHEN ai.ad_level='group' THEN '广告组' WHEN ai.ad_level='campaign' THEN '广告活动' ELSE '' END) LIKE %s)")
                            params.extend([f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"])
                        where_sql = (' WHERE ' + ' AND '.join(filters)) if filters else ''
                        cur.execute(sql + where_sql + ' ORDER BY d.id DESC', params)
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                status = self._normalize_ad_status(data.get('status') or '启动')
                ad_item_id = self._parse_int(data.get('ad_item_id'))
                delivery_desc = (data.get('delivery_desc') or '').strip()
                bid_value = self._normalize_bid_value(data.get('bid_value'))
                updated_at_text, observe_interval, next_observe_at = _resolve_schedule(data)
                if not status or not ad_item_id or not delivery_desc or not bid_value:
                    return self.send_json({'status': 'error', 'message': 'Missing status/ad_item_id/delivery_desc/bid_value'}, start_response)
                if bid_value is None:
                    return self.send_json({'status': 'error', 'message': 'Invalid bid_value'}, start_response)

                with self._get_db_connection() as conn:
                    ad_item = self._get_ad_item_by_id(conn, ad_item_id)
                    if not ad_item or ad_item.get('ad_level') != 'group':
                        return self.send_json({'status': 'error', 'message': '广告关联仅支持广告组'}, start_response)
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_deliveries
                            (status, ad_item_id, delivery_desc, bid_value, observe_interval, next_observe_at, updated_at)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """,
                            (status, ad_item_id, delivery_desc, bid_value, observe_interval, next_observe_at, updated_at_text)
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)

                status = self._normalize_ad_status(data.get('status'))
                ad_item_id = self._parse_int(data.get('ad_item_id'))
                delivery_desc = (data.get('delivery_desc') or '').strip()
                bid_value = self._normalize_bid_value(data.get('bid_value'))
                updated_at_text, observe_interval, next_observe_at = _resolve_schedule(data)
                if not status or not ad_item_id or not delivery_desc or not bid_value:
                    return self.send_json({'status': 'error', 'message': 'Missing status/ad_item_id/delivery_desc/bid_value'}, start_response)
                if bid_value is None:
                    return self.send_json({'status': 'error', 'message': 'Invalid bid_value'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM amazon_ad_deliveries WHERE id=%s", (item_id,))
                        if not cur.fetchone():
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                    ad_item = self._get_ad_item_by_id(conn, ad_item_id)
                    if not ad_item or ad_item.get('ad_level') != 'group':
                        return self.send_json({'status': 'error', 'message': '广告关联仅支持广告组'}, start_response)
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE amazon_ad_deliveries
                            SET status=%s, ad_item_id=%s, delivery_desc=%s, bid_value=%s, observe_interval=%s, next_observe_at=%s, updated_at=%s
                            WHERE id=%s
                            """,
                            (status, ad_item_id, delivery_desc, bid_value, observe_interval, next_observe_at, updated_at_text, item_id)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'PATCH':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM amazon_ad_deliveries WHERE id=%s", (item_id,))
                        if not cur.fetchone():
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)

                        updates = []
                        params = []
                        if 'status' in data:
                            status = self._normalize_ad_status(data.get('status'))
                            if not status:
                                return self.send_json({'status': 'error', 'message': 'Invalid status'}, start_response)
                            updates.append('status=%s')
                            params.append(status)
                        if 'bid_value' in data:
                            raw_bid = data.get('bid_value')
                            bid_value = self._normalize_bid_value(raw_bid)
                            if bid_value is None:
                                return self.send_json({'status': 'error', 'message': 'Invalid bid_value'}, start_response)
                            updates.append('bid_value=%s')
                            params.append(bid_value)

                        if not updates:
                            return self.send_json({'status': 'error', 'message': 'No fields to update'}, start_response)
                        params.append(item_id)
                        cur.execute(f"UPDATE amazon_ad_deliveries SET {', '.join(updates)} WHERE id=%s", params)
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_ad_deliveries WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '广告投放记录已存在或被引用'}, start_response)
            print("Amazon ad delivery API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_product_api(self, environ, method, start_response):
        """Amazon 广告商品管理 API"""
        try:
            self._ensure_amazon_ad_product_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            def _resolve_schedule(data):
                updated_at_text = self._normalize_datetime_text(data.get('updated_at'))
                if not updated_at_text:
                    updated_at_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                observe_days = self._normalize_observe_days(data.get('observe_days'))
                if observe_days is None:
                    observe_days = self._normalize_observe_days(data.get('observe_interval'))
                if observe_days is None:
                    observe_days = 1
                observe_interval = f"{observe_days}天"
                next_observe_at = self._normalize_datetime_text(data.get('next_observe_at'))
                if not next_observe_at:
                    base_dt = datetime.strptime(updated_at_text, '%Y-%m-%d %H:%M:%S')
                    next_observe_at = (base_dt + timedelta(days=observe_days)).strftime('%Y-%m-%d %H:%M:%S')
                return updated_at_text, observe_interval, next_observe_at

            if method == 'GET':
                keyword = (query_params.get('q', [''])[0] or '').strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        sql = """
                            SELECT
                                p.id,
                                p.status,
                                p.ad_item_id,
                                p.sales_product_id,
                                p.observe_interval,
                                p.next_observe_at,
                                p.created_at,
                                p.updated_at,
                                ai.ad_level,
                                ai.name AS ad_name,
                                adp.name AS portfolio_name,
                                sp.platform_sku
                            FROM amazon_ad_products p
                            LEFT JOIN amazon_ad_items ai ON ai.id = p.ad_item_id
                            LEFT JOIN amazon_ad_items adp ON adp.id = ai.portfolio_id
                            LEFT JOIN sales_products sp ON sp.id = p.sales_product_id
                        """
                        filters = []
                        params = []
                        if keyword:
                            filters.append("(ai.name LIKE %s OR sp.platform_sku LIKE %s OR adp.name LIKE %s OR (CASE WHEN ai.ad_level='group' THEN '广告组' WHEN ai.ad_level='campaign' THEN '广告活动' ELSE '' END) LIKE %s)")
                            params.extend([f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"])
                        where_sql = (' WHERE ' + ' AND '.join(filters)) if filters else ''
                        cur.execute(sql + where_sql + ' ORDER BY p.id DESC', params)
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                status = self._normalize_ad_status(data.get('status') or '启动')
                ad_item_id = self._parse_int(data.get('ad_item_id'))
                sales_product_id = self._parse_int(data.get('sales_product_id'))
                updated_at_text, observe_interval, next_observe_at = _resolve_schedule(data)
                if not status or not ad_item_id or not sales_product_id:
                    return self.send_json({'status': 'error', 'message': 'Missing status/ad_item_id/sales_product_id'}, start_response)

                with self._get_db_connection() as conn:
                    ad_item = self._get_ad_item_by_id(conn, ad_item_id)
                    if not ad_item or ad_item.get('ad_level') not in ('campaign', 'group'):
                        return self.send_json({'status': 'error', 'message': '广告关联仅支持广告活动或广告组'}, start_response)
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM sales_products WHERE id=%s", (sales_product_id,))
                        if not cur.fetchone():
                            return self.send_json({'status': 'error', 'message': 'Invalid sales_product_id'}, start_response)
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_products
                            (status, ad_item_id, sales_product_id, observe_interval, next_observe_at, updated_at)
                            VALUES (%s, %s, %s, %s, %s, %s)
                            """,
                            (status, ad_item_id, sales_product_id, observe_interval, next_observe_at, updated_at_text)
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)

                status = self._normalize_ad_status(data.get('status'))
                ad_item_id = self._parse_int(data.get('ad_item_id'))
                sales_product_id = self._parse_int(data.get('sales_product_id'))
                updated_at_text, observe_interval, next_observe_at = _resolve_schedule(data)
                if not status or not ad_item_id or not sales_product_id:
                    return self.send_json({'status': 'error', 'message': 'Missing status/ad_item_id/sales_product_id'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM amazon_ad_products WHERE id=%s", (item_id,))
                        if not cur.fetchone():
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)
                    ad_item = self._get_ad_item_by_id(conn, ad_item_id)
                    if not ad_item or ad_item.get('ad_level') not in ('campaign', 'group'):
                        return self.send_json({'status': 'error', 'message': '广告关联仅支持广告活动或广告组'}, start_response)
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM sales_products WHERE id=%s", (sales_product_id,))
                        if not cur.fetchone():
                            return self.send_json({'status': 'error', 'message': 'Invalid sales_product_id'}, start_response)
                        cur.execute(
                            """
                            UPDATE amazon_ad_products
                            SET status=%s, ad_item_id=%s, sales_product_id=%s, observe_interval=%s, next_observe_at=%s, updated_at=%s
                            WHERE id=%s
                            """,
                            (status, ad_item_id, sales_product_id, observe_interval, next_observe_at, updated_at_text, item_id)
                        )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_ad_products WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '广告商品记录已存在或被引用'}, start_response)
            print("Amazon ad product API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_adjustment_api(self, environ, method, start_response):
        """Amazon 广告调整 API（搜索广告、获取默认信息、记录调整）"""
        try:
            self._ensure_amazon_ad_adjustment_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            def _normalize_short_text(value, max_len=64):
                text = ('' if value is None else str(value)).strip()
                return text[:max_len] if text else None

            def _normalize_long_text(value, max_len=255):
                text = ('' if value is None else str(value)).strip()
                return text[:max_len] if text else None

            def _normalize_bool(value):
                text = ('' if value is None else str(value)).strip().lower()
                return 1 if text in ('1', 'true', 'yes', 'y', '是') else 0

            def _resolve_ad_info(conn, ad_item_id):
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT
                            ai.id,
                            ai.ad_level,
                            ai.name AS self_name,
                            ai.portfolio_id,
                            ai.campaign_id,
                            p.name AS portfolio_name,
                            c.name AS campaign_name,
                            ai.subtype_id,
                            c.subtype_id AS campaign_subtype_id,
                            st.description AS self_subtype_desc,
                            st.ad_class AS self_ad_class,
                            st.subtype_code AS self_subtype_code,
                            cst.description AS campaign_subtype_desc,
                            cst.ad_class AS campaign_ad_class,
                            cst.subtype_code AS campaign_subtype_code
                        FROM amazon_ad_items ai
                        LEFT JOIN amazon_ad_items p ON p.id = ai.portfolio_id
                        LEFT JOIN amazon_ad_items c ON c.id = ai.campaign_id
                        LEFT JOIN amazon_ad_subtypes st ON st.id = ai.subtype_id
                        LEFT JOIN amazon_ad_subtypes cst ON cst.id = c.subtype_id
                        WHERE ai.id=%s
                        """,
                        (ad_item_id,)
                    )
                    row = cur.fetchone()
                if not row:
                    return None
                effective_subtype_id = row.get('subtype_id') or row.get('campaign_subtype_id')
                ad_class = row.get('self_ad_class') or row.get('campaign_ad_class') or ''
                subtype_code = row.get('self_subtype_code') or row.get('campaign_subtype_code') or ''
                subtype_desc = row.get('self_subtype_desc') or row.get('campaign_subtype_desc') or ''

                ad_level = row.get('ad_level')
                portfolio_name = row.get('portfolio_name') or ''
                campaign_name = row.get('campaign_name') or ''
                group_name = ''
                if ad_level == 'campaign':
                    campaign_name = row.get('self_name') or campaign_name
                elif ad_level == 'group':
                    group_name = row.get('self_name') or ''

                return {
                    'id': row.get('id'),
                    'ad_level': ad_level,
                    'portfolio_name': portfolio_name,
                    'campaign_name': campaign_name,
                    'group_name': group_name,
                    'ad_name': row.get('self_name') or '',
                    'subtype_id': effective_subtype_id,
                    'subtype_description': subtype_desc,
                    'ad_class': ad_class,
                    'subtype_code': subtype_code,
                    'ad_type_text': f"{ad_class}-{subtype_code}" if (ad_class or subtype_code) else ''
                }

            def _fetch_allowed_operations(conn, subtype_id, ad_level):
                if not subtype_id:
                    return []
                level_col = 'apply_group' if ad_level == 'group' else ('apply_campaign' if ad_level == 'campaign' else 'apply_portfolio')
                with conn.cursor() as cur:
                    cur.execute(
                        f"""
                        SELECT t.id, t.name
                        FROM amazon_ad_subtype_operation_types so
                        JOIN amazon_ad_operation_types t ON t.id = so.operation_type_id
                        WHERE so.subtype_id=%s AND t.{level_col}=1 AND t.name<>%s
                        ORDER BY t.id ASC
                        """,
                        (subtype_id, '【新建】')
                    )
                    ops = cur.fetchall() or []
                    if not ops:
                        return []
                    op_ids = [int(row.get('id')) for row in ops if row.get('id')]
                    placeholders = ','.join(['%s'] * len(op_ids))
                    cur.execute(
                        f"""
                        SELECT id, operation_type_id, reason_name
                        FROM amazon_ad_operation_reasons
                        WHERE operation_type_id IN ({placeholders})
                        ORDER BY id ASC
                        """,
                        op_ids
                    )
                    reason_rows = cur.fetchall() or []
                reason_map = {}
                for reason in reason_rows:
                    op_id = int(reason.get('operation_type_id'))
                    reason_map.setdefault(op_id, []).append({
                        'id': reason.get('id'),
                        'reason_name': reason.get('reason_name') or ''
                    })
                for op in ops:
                    op_id = int(op.get('id'))
                    op['reasons'] = reason_map.get(op_id, [])
                return ops

            def _get_default_time_window(conn, ad_item_id):
                us_now = datetime.utcnow() - timedelta(hours=5)
                us_yesterday_date = (us_now - timedelta(days=1)).strftime('%Y-%m-%d')
                end_time = f"{us_yesterday_date} 23:59:59"
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT adjust_date FROM amazon_ad_adjustments WHERE ad_item_id=%s ORDER BY adjust_date DESC LIMIT 1",
                        (ad_item_id,)
                    )
                    last_row = cur.fetchone()
                if last_row and last_row.get('adjust_date'):
                    start_time = datetime.strftime(last_row.get('adjust_date'), '%Y-%m-%d %H:%M:%S')
                else:
                    start_time = f"{us_yesterday_date} 00:00:00"
                return start_time, end_time

            if method == 'GET':
                action = (query_params.get('action', ['records'])[0] or 'records').strip().lower()

                if action == 'ad-search':
                    keyword = (query_params.get('q', [''])[0] or '').strip()
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            sql = """
                                SELECT
                                    ai.id,
                                    ai.ad_level,
                                    ai.name AS ad_name,
                                    p.name AS portfolio_name,
                                    c.name AS campaign_name,
                                    st.description AS self_subtype_desc,
                                    st.ad_class AS self_ad_class,
                                    st.subtype_code AS self_subtype_code,
                                    cst.description AS campaign_subtype_desc,
                                    cst.ad_class AS campaign_ad_class,
                                    cst.subtype_code AS campaign_subtype_code
                                FROM amazon_ad_items ai
                                LEFT JOIN amazon_ad_items p ON p.id = ai.portfolio_id
                                LEFT JOIN amazon_ad_items c ON c.id = ai.campaign_id
                                LEFT JOIN amazon_ad_subtypes st ON st.id = ai.subtype_id
                                LEFT JOIN amazon_ad_subtypes cst ON cst.id = c.subtype_id
                                WHERE ai.ad_level IN ('campaign', 'group')
                            """
                            params = []
                            if keyword:
                                sql += " AND (ai.name LIKE %s OR p.name LIKE %s OR c.name LIKE %s OR st.description LIKE %s OR cst.description LIKE %s)"
                                like = f"%{keyword}%"
                                params.extend([like, like, like, like, like])
                            sql += " ORDER BY ai.id DESC LIMIT 200"
                            cur.execute(sql, params)
                            rows = cur.fetchall() or []
                    items = []
                    for row in rows:
                        ad_level = row.get('ad_level')
                        subtype_desc = row.get('self_subtype_desc') or row.get('campaign_subtype_desc') or ''
                        ad_class = row.get('self_ad_class') or row.get('campaign_ad_class') or ''
                        subtype_code = row.get('self_subtype_code') or row.get('campaign_subtype_code') or ''
                        campaign_name = row.get('campaign_name') or ''
                        group_name = ''
                        if ad_level == 'campaign':
                            campaign_name = row.get('ad_name') or campaign_name
                        elif ad_level == 'group':
                            group_name = row.get('ad_name') or ''
                        items.append({
                            'id': row.get('id'),
                            'ad_level': ad_level,
                            'ad_name': row.get('ad_name') or '',
                            'portfolio_name': row.get('portfolio_name') or '',
                            'campaign_name': campaign_name,
                            'group_name': group_name,
                            'subtype_description': subtype_desc,
                            'ad_type_text': f"{ad_class}-{subtype_code}" if (ad_class or subtype_code) else ''
                        })
                    return self.send_json({'status': 'success', 'items': items}, start_response)

                if action == 'defaults':
                    ad_item_id = self._parse_int(query_params.get('ad_item_id', [''])[0])
                    if not ad_item_id:
                        return self.send_json({'status': 'error', 'message': 'Missing ad_item_id'}, start_response)
                    with self._get_db_connection() as conn:
                        ad_info = _resolve_ad_info(conn, ad_item_id)
                        if not ad_info:
                            return self.send_json({'status': 'error', 'message': '广告信息不存在'}, start_response)
                        operations = _fetch_allowed_operations(conn, ad_info.get('subtype_id'), ad_info.get('ad_level'))
                        start_time, end_time = _get_default_time_window(conn, ad_item_id)
                    return self.send_json({
                        'status': 'success',
                        'ad_info': ad_info,
                        'allowed_operations': operations,
                        'defaults': {
                            'adjust_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'start_time': start_time,
                            'end_time': end_time
                        }
                    }, start_response)

                keyword = (query_params.get('q', [''])[0] or '').strip()
                ad_item_id = self._parse_int(query_params.get('ad_item_id', [''])[0])
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        sql = """
                            SELECT
                                r.id,
                                r.adjust_date,
                                r.ad_item_id,
                                r.operation_type_id,
                                r.target_object,
                                r.before_value,
                                r.after_value,
                                r.reason_id,
                                r.start_time,
                                r.end_time,
                                r.impressions,
                                r.clicks,
                                r.cost,
                                r.orders,
                                r.sales,
                                r.acos,
                                r.cpc,
                                r.ctr,
                                r.cvr,
                                r.attribution_checked,
                                r.attribution_orders,
                                r.attribution_sales,
                                r.remark,
                                r.is_quick_submit,
                                ai.name AS ad_name,
                                ai.ad_level,
                                p.name AS portfolio_name,
                                c.name AS campaign_name,
                                t.name AS operation_name,
                                reason.reason_name
                            FROM amazon_ad_adjustments r
                            JOIN amazon_ad_items ai ON ai.id = r.ad_item_id
                            LEFT JOIN amazon_ad_items p ON p.id = ai.portfolio_id
                            LEFT JOIN amazon_ad_items c ON c.id = ai.campaign_id
                            JOIN amazon_ad_operation_types t ON t.id = r.operation_type_id
                            LEFT JOIN amazon_ad_operation_reasons reason ON reason.id = r.reason_id
                        """
                        filters = []
                        params = []
                        if ad_item_id:
                            filters.append("r.ad_item_id=%s")
                            params.append(ad_item_id)
                        if keyword:
                            filters.append("(ai.name LIKE %s OR t.name LIKE %s OR r.target_object LIKE %s OR reason.reason_name LIKE %s)")
                            like = f"%{keyword}%"
                            params.extend([like, like, like, like])
                        where_sql = (' WHERE ' + ' AND '.join(filters)) if filters else ''
                        cur.execute(sql + where_sql + ' ORDER BY r.adjust_date DESC, r.id DESC LIMIT 500', params)
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                is_quick_submit = _normalize_bool(data.get('is_quick_submit'))
                ad_item_id = self._parse_int(data.get('ad_item_id'))
                operation_type_id = self._parse_int(data.get('operation_type_id'))
                target_object = _normalize_long_text(data.get('target_object'), 255)
                before_value = _normalize_short_text(data.get('before_value'), 64)
                after_value = _normalize_short_text(data.get('after_value'), 64)
                adjust_date = self._normalize_datetime_text(data.get('adjust_date')) or datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                if not ad_item_id or not operation_type_id or not target_object:
                    return self.send_json({'status': 'error', 'message': 'Missing ad_item_id/operation_type_id/target_object'}, start_response)

                with self._get_db_connection() as conn:
                    ad_info = _resolve_ad_info(conn, ad_item_id)
                    if not ad_info:
                        return self.send_json({'status': 'error', 'message': '广告信息不存在'}, start_response)

                    allowed_ops = _fetch_allowed_operations(conn, ad_info.get('subtype_id'), ad_info.get('ad_level'))
                    allowed_op_ids = {int(item.get('id')) for item in allowed_ops if item.get('id')}
                    if operation_type_id not in allowed_op_ids:
                        return self.send_json({'status': 'error', 'message': '该广告类型不允许当前操作类型'}, start_response)

                    reason_id = self._parse_int(data.get('reason_id'))
                    reason_allowed = set()
                    for op in allowed_ops:
                        if int(op.get('id')) == operation_type_id:
                            for reason in op.get('reasons') or []:
                                if reason.get('id'):
                                    reason_allowed.add(int(reason.get('id')))
                            break

                    if reason_id and reason_id not in reason_allowed:
                        return self.send_json({'status': 'error', 'message': '当前操作类型不允许该修改原因'}, start_response)

                    if not is_quick_submit:
                        start_time = self._normalize_datetime_text(data.get('start_time'))
                        end_time = self._normalize_datetime_text(data.get('end_time'))
                        if not reason_id:
                            return self.send_json({'status': 'error', 'message': '提交并记录需要选择修改原因'}, start_response)
                        if not before_value or not after_value:
                            return self.send_json({'status': 'error', 'message': '提交并记录需要填写修改前和修改后'}, start_response)
                        if not start_time or not end_time:
                            return self.send_json({'status': 'error', 'message': '提交并记录需要填写开始时间和结束时间'}, start_response)
                    else:
                        start_time = None
                        end_time = None
                        reason_id = reason_id if reason_id in reason_allowed else None

                    impressions = _normalize_short_text(data.get('impressions'), 32)
                    clicks = _normalize_short_text(data.get('clicks'), 32)
                    cost = _normalize_short_text(data.get('cost'), 32)
                    orders = _normalize_short_text(data.get('orders'), 32)
                    sales = _normalize_short_text(data.get('sales'), 32)
                    acos = _normalize_short_text(data.get('acos'), 32)
                    cpc = _normalize_short_text(data.get('cpc'), 32)
                    ctr = _normalize_short_text(data.get('ctr'), 32)
                    cvr = _normalize_short_text(data.get('cvr'), 32)
                    attribution_checked = _normalize_bool(data.get('attribution_checked'))
                    attribution_orders = _normalize_short_text(data.get('attribution_orders'), 32)
                    attribution_sales = _normalize_short_text(data.get('attribution_sales'), 32)
                    remark = _normalize_long_text(data.get('remark'), 255)

                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO amazon_ad_adjustments (
                                adjust_date, ad_item_id, operation_type_id, target_object,
                                before_value, after_value, reason_id, start_time, end_time,
                                impressions, clicks, cost, orders, sales, acos, cpc, ctr, cvr,
                                attribution_checked, attribution_orders, attribution_sales, remark, is_quick_submit
                            ) VALUES (
                                %s, %s, %s, %s,
                                %s, %s, %s, %s, %s,
                                %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                %s, %s, %s, %s, %s
                            )
                            """,
                            (
                                adjust_date, ad_item_id, operation_type_id, target_object,
                                before_value, after_value, reason_id, start_time, end_time,
                                impressions, clicks, cost, orders, sales, acos, cpc, ctr, cvr,
                                attribution_checked, attribution_orders, attribution_sales, remark, is_quick_submit
                            )
                        )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_ad_adjustments WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '广告调整记录保存失败，存在无效外键'}, start_response)
            print("Amazon ad adjustment API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def _ensure_logistics_tables(self):
        if self._logistics_ready:
            return
        with self._schema_ensure_lock:
            if self._logistics_ready:
                return
        self._ensure_order_product_tables()
        create_factory_sql = """
        CREATE TABLE IF NOT EXISTS logistics_factories (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            factory_name VARCHAR(255) NOT NULL UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        create_forwarder_sql = """
        CREATE TABLE IF NOT EXISTS logistics_forwarders (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            forwarder_name VARCHAR(255) NOT NULL UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        create_supplier_sql = """
        CREATE TABLE IF NOT EXISTS logistics_suppliers (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            supplier_name VARCHAR(255) NOT NULL UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        create_warehouse_sql = """
        CREATE TABLE IF NOT EXISTS logistics_overseas_warehouses (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            warehouse_name VARCHAR(255) NOT NULL,
            supplier_id INT UNSIGNED NOT NULL,
            warehouse_short_name VARCHAR(128) NOT NULL,
            is_enabled TINYINT(1) NOT NULL DEFAULT 1,
            region VARCHAR(32) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_wh_name (warehouse_name),
            UNIQUE KEY uniq_wh_supplier_short (supplier_id, warehouse_short_name),
            INDEX idx_wh_region (region),
            INDEX idx_wh_enabled (is_enabled),
            CONSTRAINT fk_wh_supplier FOREIGN KEY (supplier_id)
                REFERENCES logistics_suppliers(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        create_inventory_sql = """
        CREATE TABLE IF NOT EXISTS logistics_overseas_inventory (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            warehouse_id INT UNSIGNED NOT NULL,
            order_product_id INT UNSIGNED NOT NULL,
            available_qty INT NOT NULL DEFAULT 0,
            in_transit_qty INT NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_wh_order (warehouse_id, order_product_id),
            INDEX idx_inv_warehouse (warehouse_id),
            INDEX idx_inv_order_product (order_product_id),
            CONSTRAINT fk_inv_warehouse FOREIGN KEY (warehouse_id)
                REFERENCES logistics_overseas_warehouses(id) ON DELETE CASCADE,
            CONSTRAINT fk_inv_order_product FOREIGN KEY (order_product_id)
                REFERENCES order_products(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        create_transit_sql = """
        CREATE TABLE IF NOT EXISTS logistics_in_transit (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            factory_id INT UNSIGNED NOT NULL,
            factory_ship_date_initial DATE NULL,
            factory_ship_date_previous DATE NULL,
            factory_ship_date_latest DATE NULL,
            forwarder_id INT UNSIGNED NOT NULL,
            logistics_box_no VARCHAR(128) NOT NULL,
            customs_clearance_no VARCHAR(128) NOT NULL,
            etd_initial DATE NULL,
            etd_previous DATE NULL,
            etd_latest DATE NULL,
            eta_initial DATE NULL,
            eta_previous DATE NULL,
            eta_latest DATE NULL,
            arrival_port_date DATE NULL,
            expected_warehouse_date DATE NULL,
            expected_listed_date_initial DATE NULL,
            expected_listed_date_latest DATE NULL,
            listed_date DATE NULL,
            shipping_company VARCHAR(128) NULL,
            vessel_voyage VARCHAR(128) NULL,
            bill_of_lading_no VARCHAR(128) NULL,
            declaration_docs_provided TINYINT(1) NOT NULL DEFAULT 0,
            inventory_registered TINYINT(1) NOT NULL DEFAULT 0,
            clearance_docs_provided TINYINT(1) NOT NULL DEFAULT 0,
            qty_verified TINYINT(1) NOT NULL DEFAULT 0,
            qty_consistent TINYINT(1) NOT NULL DEFAULT 0,
            port_of_loading VARCHAR(128) NULL,
            port_of_destination VARCHAR(128) NULL,
            destination_warehouse_id INT UNSIGNED NULL,
            inbound_order_no VARCHAR(128) NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_transit_box_no (logistics_box_no),
            UNIQUE KEY uniq_transit_customs_no (customs_clearance_no),
            UNIQUE KEY uniq_transit_bl_no (bill_of_lading_no),
            INDEX idx_transit_factory (factory_id),
            INDEX idx_transit_forwarder (forwarder_id),
            INDEX idx_transit_wh (destination_warehouse_id),
            CONSTRAINT fk_transit_factory FOREIGN KEY (factory_id)
                REFERENCES logistics_factories(id) ON DELETE RESTRICT,
            CONSTRAINT fk_transit_forwarder FOREIGN KEY (forwarder_id)
                REFERENCES logistics_forwarders(id) ON DELETE RESTRICT,
            CONSTRAINT fk_transit_wh FOREIGN KEY (destination_warehouse_id)
                REFERENCES logistics_overseas_warehouses(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        create_transit_items_sql = """
        CREATE TABLE IF NOT EXISTS logistics_in_transit_items (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            transit_id INT UNSIGNED NOT NULL,
            order_product_id INT UNSIGNED NOT NULL,
            shipped_qty INT NOT NULL DEFAULT 0,
            listed_qty INT NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_transit_item (transit_id, order_product_id),
            INDEX idx_transit_item_transit (transit_id),
            INDEX idx_transit_item_order (order_product_id),
            CONSTRAINT fk_transit_item_transit FOREIGN KEY (transit_id)
                REFERENCES logistics_in_transit(id) ON DELETE CASCADE,
            CONSTRAINT fk_transit_item_order FOREIGN KEY (order_product_id)
                REFERENCES order_products(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_factory_sql)
                cur.execute(create_forwarder_sql)
                cur.execute(create_supplier_sql)
                cur.execute(create_warehouse_sql)
                cur.execute(create_inventory_sql)
                cur.execute(create_transit_sql)
                cur.execute(create_transit_items_sql)
                cur.execute("SHOW COLUMNS FROM logistics_overseas_warehouses")
                warehouse_cols = {str((x or {}).get('Field') or '') for x in (cur.fetchall() or [])}
                if 'is_enabled' not in warehouse_cols:
                    cur.execute("ALTER TABLE logistics_overseas_warehouses ADD COLUMN is_enabled TINYINT(1) NOT NULL DEFAULT 1 AFTER warehouse_short_name")
                try:
                    cur.execute("ALTER TABLE logistics_overseas_warehouses ADD INDEX idx_wh_enabled (is_enabled)")
                except Exception:
                    pass
                cur.execute("SHOW COLUMNS FROM logistics_in_transit")
                transit_cols = {str((x or {}).get('Field') or '') for x in (cur.fetchall() or [])}
                if 'customs_clearance_no' not in transit_cols:
                    cur.execute("ALTER TABLE logistics_in_transit ADD COLUMN customs_clearance_no VARCHAR(128) NULL AFTER logistics_box_no")
                if 'qty_verified' not in transit_cols:
                    cur.execute("ALTER TABLE logistics_in_transit ADD COLUMN qty_verified TINYINT(1) NOT NULL DEFAULT 0 AFTER clearance_docs_provided")
                if 'qty_consistent' not in transit_cols:
                    cur.execute("ALTER TABLE logistics_in_transit ADD COLUMN qty_consistent TINYINT(1) NOT NULL DEFAULT 0 AFTER qty_verified")
                if 'expected_listed_date_initial' not in transit_cols:
                    cur.execute("ALTER TABLE logistics_in_transit ADD COLUMN expected_listed_date_initial DATE NULL AFTER expected_warehouse_date")
                if 'expected_listed_date_latest' not in transit_cols:
                    cur.execute("ALTER TABLE logistics_in_transit ADD COLUMN expected_listed_date_latest DATE NULL AFTER expected_listed_date_initial")
                cur.execute("SHOW COLUMNS FROM logistics_in_transit_items")
                transit_item_cols = {str((x or {}).get('Field') or '') for x in (cur.fetchall() or [])}
                if 'listed_qty' not in transit_item_cols:
                    cur.execute("ALTER TABLE logistics_in_transit_items ADD COLUMN listed_qty INT NOT NULL DEFAULT 0 AFTER shipped_qty")
            self._logistics_ready = True
            self.__class__._schema_ready_cache['logistics'] = True

    def _get_logistics_link_root_bytes(self):
        return os.path.join(_RESOURCES_PARENT_BYTES, self._safe_fsencode('『物流仓储关联文件』'))

    def _ensure_logistics_bl_folder(self, bill_of_lading_no):
        name = (bill_of_lading_no or '').strip()
        if not name:
            return
        root = self._get_logistics_link_root_bytes()
        if not os.path.exists(root):
            os.makedirs(root, exist_ok=True)
        folder = os.path.join(root, self._safe_fsencode(name))
        if not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
        for sub in ('报关资料', '清关资料'):
            sub_folder = os.path.join(folder, self._safe_fsencode(sub))
            if not os.path.exists(sub_folder):
                os.makedirs(sub_folder, exist_ok=True)

    def _rename_logistics_bl_folder(self, old_no, new_no):
        old_name = (old_no or '').strip()
        new_name = (new_no or '').strip()
        if not old_name or not new_name or old_name == new_name:
            if new_name:
                self._ensure_logistics_bl_folder(new_name)
            return
        root = self._get_logistics_link_root_bytes()
        if not os.path.exists(root):
            os.makedirs(root, exist_ok=True)
        old_path = os.path.join(root, self._safe_fsencode(old_name))
        new_path = os.path.join(root, self._safe_fsencode(new_name))
        if os.path.exists(old_path):
            if os.path.exists(new_path):
                raise RuntimeError(f'目标提单目录已存在: {new_name}')
            os.rename(old_path, new_path)
        else:
            self._ensure_logistics_bl_folder(new_name)

    def _resolve_logistics_doc_folder(self, transit_id, doc_type):
        doc_kind = (doc_type or '').strip().lower()
        if doc_kind not in ('declaration', 'clearance'):
            raise RuntimeError('Invalid doc_type')
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT bill_of_lading_no FROM logistics_in_transit WHERE id=%s LIMIT 1", (transit_id,))
                row = cur.fetchone() or {}
        bill_no = (row.get('bill_of_lading_no') or '').strip()
        if not bill_no:
            raise RuntimeError('请先填写提单号后再操作资料文件')
        self._ensure_logistics_bl_folder(bill_no)
        sub_name = '报关资料' if doc_kind == 'declaration' else '清关资料'
        parent = os.path.join(self._get_logistics_link_root_bytes(), self._safe_fsencode(bill_no))
        folder = os.path.join(parent, self._safe_fsencode(sub_name))
        if not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
        return folder

    def _ensure_factory_inventory_tables(self):
        if self._factory_inventory_ready:
            return
        with self._schema_ensure_lock:
            if self._factory_inventory_ready:
                return
        self._ensure_logistics_tables()
        create_factory_stock = """
        CREATE TABLE IF NOT EXISTS factory_stock_inventory (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            order_product_id INT UNSIGNED NOT NULL,
            factory_id INT UNSIGNED NOT NULL,
            quantity INT NOT NULL DEFAULT 0,
            notes TEXT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY uniq_fsi_op_factory (order_product_id, factory_id),
            CONSTRAINT fk_fsi_op FOREIGN KEY (order_product_id) REFERENCES order_products(id) ON DELETE CASCADE,
            CONSTRAINT fk_fsi_factory FOREIGN KEY (factory_id) REFERENCES logistics_factories(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        create_factory_wip = """
        CREATE TABLE IF NOT EXISTS factory_wip_inventory (
            id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
            order_product_id INT UNSIGNED NOT NULL,
            factory_id INT UNSIGNED NOT NULL,
            quantity INT NOT NULL DEFAULT 0,
            expected_completion_date DATE NULL,
            is_completed TINYINT(1) NOT NULL DEFAULT 0,
            actual_completion_date DATE NULL,
            notes TEXT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            KEY idx_fwi_op (order_product_id),
            KEY idx_fwi_factory (factory_id),
            CONSTRAINT fk_fwi_op FOREIGN KEY (order_product_id) REFERENCES order_products(id) ON DELETE CASCADE,
            CONSTRAINT fk_fwi_factory FOREIGN KEY (factory_id) REFERENCES logistics_factories(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with self._get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(create_factory_stock)
                cur.execute(create_factory_wip)
                cur.execute("SHOW COLUMNS FROM factory_wip_inventory")
                cols = {str((row.get('Field') or '')).strip().lower() for row in (cur.fetchall() or [])}
                if 'is_completed' not in cols:
                    cur.execute("ALTER TABLE factory_wip_inventory ADD COLUMN is_completed TINYINT(1) NOT NULL DEFAULT 0 AFTER expected_completion_date")
                if 'actual_completion_date' not in cols:
                    cur.execute("ALTER TABLE factory_wip_inventory ADD COLUMN actual_completion_date DATE NULL AFTER is_completed")
        self._factory_inventory_ready = True
        self.__class__._schema_ready_cache['factory_inventory'] = True






















    def handle_material_api(self, environ, method, start_response):
        """材料管理 API（CRUD）"""
        try:
            self._ensure_materials_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                type_code = query_params.get('type', [''])[0].strip()
                type_name = query_params.get('type_name', [''])[0].strip()
                type_id = self._parse_int(query_params.get('type_id', [''])[0].strip())
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        type_map = {
                            'fabric': '面料',
                            'filling': '填充',
                            'frame': '框架',
                            'electronics': '电子元器件'
                        }
                        has_type_id = self._materials_has_type_id(conn)
                        if has_type_id:
                            base_sql = """
                                SELECT
                                    m.id, m.name, m.name_en, m.material_type_id,
                                    m.parent_id, pm.name AS parent_name,
                                    mt.name AS material_type_name,
                                    m.created_at
                                FROM materials m
                                LEFT JOIN materials pm ON m.parent_id = pm.id
                                LEFT JOIN material_types mt ON m.material_type_id = mt.id
                            """
                            filters = []
                            params = []
                            if type_id:
                                filters.append("m.material_type_id=%s")
                                params.append(type_id)
                            elif type_name or type_code:
                                resolved_name = type_name or type_map.get(type_code, type_code)
                                if resolved_name:
                                    filters.append("mt.name=%s")
                                    params.append(resolved_name)
                            if keyword:
                                filters.append("(m.name LIKE %s OR m.name_en LIKE %s OR mt.name LIKE %s)")
                                params.extend([f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"])
                            where_sql = (" WHERE " + " AND ".join(filters)) if filters else ""
                            cur.execute(base_sql + where_sql + " ORDER BY m.id DESC", params)
                            rows = cur.fetchall()
                        else:
                            resolved_name = type_name or type_map.get(type_code, type_code)
                            name_to_code = {v: k for k, v in type_map.items()}
                            legacy_code = name_to_code.get(resolved_name) if resolved_name else None
                            base_sql = """
                                SELECT m.id, m.name, m.name_en, m.material_type, m.parent_id, pm.name AS parent_name, m.created_at
                                FROM materials m
                                LEFT JOIN materials pm ON m.parent_id = pm.id
                            """
                            filters = []
                            params = []
                            if legacy_code:
                                filters.append("material_type=%s")
                                params.append(legacy_code)
                            if keyword:
                                filters.append("(name LIKE %s OR name_en LIKE %s OR material_type LIKE %s)")
                                params.extend([f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"])
                            where_sql = (" WHERE " + " AND ".join(filters)) if filters else ""
                            cur.execute(base_sql + where_sql + " ORDER BY id DESC", params)
                            rows = cur.fetchall()
                            cur.execute("SELECT id, name FROM material_types")
                            type_rows = cur.fetchall() or []
                            type_lookup = {row['name']: row for row in type_rows}
                            for row in rows:
                                code = row.get('material_type')
                                name = type_map.get(code, '')
                                mapped = type_lookup.get(name) or {}
                                row['material_type_id'] = mapped.get('id')
                                row['material_type_name'] = name
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                name = (data.get('name') or '').strip()
                name_en = (data.get('name_en') or '').strip()
                material_type_id = self._parse_int(data.get('material_type_id'))
                material_type_code = (data.get('material_type') or '').strip()
                parent_id = self._parse_int(data.get('parent_id'))
                if not name or not name_en:
                    return self.send_json({'status': 'error', 'message': 'Missing name or name_en'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        has_type_id = self._materials_has_type_id(conn)
                        has_parent_id = self._materials_has_parent_id(conn)
                        has_type_id = self._materials_has_type_id(conn)
                        if not material_type_id and material_type_code:
                            material_type_id = self._get_material_type_id(conn, material_type_code)
                        parent_row = None
                        if parent_id:
                            if has_type_id:
                                cur.execute("SELECT id, material_type_id FROM materials WHERE id=%s", (parent_id,))
                            else:
                                cur.execute("SELECT id, material_type FROM materials WHERE id=%s", (parent_id,))
                            parent_row = cur.fetchone()
                            if not parent_row:
                                return self.send_json({'status': 'error', 'message': 'Invalid parent_id'}, start_response)
                        if has_type_id:
                            if not material_type_id:
                                return self.send_json({'status': 'error', 'message': 'Missing material_type_id'}, start_response)
                            if parent_row and parent_row.get('material_type_id') != material_type_id:
                                return self.send_json({'status': 'error', 'message': 'Parent type mismatch'}, start_response)
                            if has_parent_id:
                                cur.execute(
                                    "INSERT INTO materials (name, name_en, material_type_id, parent_id) VALUES (%s, %s, %s, %s)",
                                    (name, name_en, material_type_id, parent_id)
                                )
                            else:
                                cur.execute(
                                    "INSERT INTO materials (name, name_en, material_type_id) VALUES (%s, %s, %s)",
                                    (name, name_en, material_type_id)
                                )
                        else:
                            if not material_type_code:
                                return self.send_json({'status': 'error', 'message': 'Missing material_type'}, start_response)
                            if parent_row and parent_row.get('material_type') != material_type_code:
                                return self.send_json({'status': 'error', 'message': 'Parent type mismatch'}, start_response)
                            if has_parent_id:
                                cur.execute(
                                    "INSERT INTO materials (name, name_en, material_type, parent_id) VALUES (%s, %s, %s, %s)",
                                    (name, name_en, material_type_code, parent_id)
                                )
                            else:
                                cur.execute(
                                    "INSERT INTO materials (name, name_en, material_type) VALUES (%s, %s, %s)",
                                    (name, name_en, material_type_code)
                                )
                        new_id = cur.lastrowid
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                name = (data.get('name') or '').strip()
                name_en = (data.get('name_en') or '').strip()
                material_type_id = self._parse_int(data.get('material_type_id'))
                material_type_code = (data.get('material_type') or '').strip()
                parent_id = self._parse_int(data.get('parent_id'))
                if not item_id or not name or not name_en:
                    return self.send_json({'status': 'error', 'message': 'Missing id or fields'}, start_response)
                if parent_id and int(parent_id) == int(item_id):
                    return self.send_json({'status': 'error', 'message': 'Invalid parent_id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        has_type_id = self._materials_has_type_id(conn)
                        has_parent_id = self._materials_has_parent_id(conn)
                        has_type_id = self._materials_has_type_id(conn)
                        if not material_type_id and material_type_code:
                            material_type_id = self._get_material_type_id(conn, material_type_code)
                        parent_row = None
                        if parent_id:
                            if has_type_id:
                                cur.execute("SELECT id, material_type_id FROM materials WHERE id=%s", (parent_id,))
                            else:
                                cur.execute("SELECT id, material_type FROM materials WHERE id=%s", (parent_id,))
                            parent_row = cur.fetchone()
                            if not parent_row:
                                return self.send_json({'status': 'error', 'message': 'Invalid parent_id'}, start_response)
                        if has_type_id:
                            if not material_type_id:
                                return self.send_json({'status': 'error', 'message': 'Missing material_type_id'}, start_response)
                            if parent_row and parent_row.get('material_type_id') != material_type_id:
                                return self.send_json({'status': 'error', 'message': 'Parent type mismatch'}, start_response)
                            if has_parent_id:
                                cur.execute(
                                    """
                                    UPDATE materials
                                    SET name=%s, name_en=%s, material_type_id=%s, parent_id=%s
                                    WHERE id=%s
                                    """,
                                    (name, name_en, material_type_id, parent_id, item_id)
                                )
                            else:
                                cur.execute(
                                    """
                                    UPDATE materials
                                    SET name=%s, name_en=%s, material_type_id=%s
                                    WHERE id=%s
                                    """,
                                    (name, name_en, material_type_id, item_id)
                                )
                        else:
                            if not material_type_code:
                                return self.send_json({'status': 'error', 'message': 'Missing material_type'}, start_response)
                            if parent_row and parent_row.get('material_type') != material_type_code:
                                return self.send_json({'status': 'error', 'message': 'Parent type mismatch'}, start_response)
                            if has_parent_id:
                                cur.execute(
                                    """
                                    UPDATE materials
                                    SET name=%s, name_en=%s, material_type=%s, parent_id=%s
                                    WHERE id=%s
                                    """,
                                    (name, name_en, material_type_code, parent_id, item_id)
                                )
                            else:
                                cur.execute(
                                    """
                                    UPDATE materials
                                    SET name=%s, name_en=%s, material_type=%s
                                    WHERE id=%s
                                    """,
                                    (name, name_en, material_type_code, item_id)
                                )
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM materials WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '材料已存在'}, start_response)
            print("Material API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_certification_api(self, environ, method, start_response):
        """认证管理 API（CRUD）"""
        try:
            self._ensure_certification_table()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                if not keyword:
                    def _load_certifications_cached():
                        with self._get_db_connection() as conn:
                            with conn.cursor() as cur:
                                cur.execute(
                                    """
                                    SELECT id, name, icon_name, created_at
                                    FROM certifications
                                    ORDER BY id DESC
                                    """
                                )
                                rows = cur.fetchall() or []
                        return {'status': 'success', 'items': rows}
                    payload = self._get_cached_template_options('certification_list', _load_certifications_cached, ttl_seconds=180)
                    return self.send_json(payload, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT id, name, icon_name, created_at
                            FROM certifications
                            WHERE name LIKE %s
                            ORDER BY id DESC
                            """,
                            (f"%{keyword}%",)
                        )
                        rows = cur.fetchall() or []
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                name = (data.get('name') or '').strip()
                icon_name = (data.get('icon_name') or '').strip()
                if not name:
                    return self.send_json({'status': 'error', 'message': 'Missing name'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO certifications (name, icon_name)
                            VALUES (%s, %s)
                            """,
                            (name, icon_name or None)
                        )
                        new_id = cur.lastrowid
                self._template_options_cache.pop('certification_list', None)
                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                name = (data.get('name') or '').strip()
                icon_name = (data.get('icon_name') or '').strip()
                if not item_id or not name:
                    return self.send_json({'status': 'error', 'message': 'Missing id or name'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE certifications
                            SET name=%s, icon_name=%s
                            WHERE id=%s
                            """,
                            (name, icon_name or None, item_id)
                        )
                self._template_options_cache.pop('certification_list', None)
                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM certifications WHERE id=%s", (item_id,))
                self._template_options_cache.pop('certification_list', None)
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '认证名称已存在'}, start_response)
            print("Certification API error: " + str(e))
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_order_product_api(self, environ, method, start_response):
        """下单产品管理 API（CRUD）"""
        try:
            self._ensure_order_product_tables()
            query_string = environ.get('QUERY_STRING', '')
            query_params = parse_qs(query_string)
            action = (query_params.get('action', [''])[0] or '').strip().lower()

            if action == 'shipping_plans':
                if method == 'GET':
                    target_order_product_id = self._parse_int(query_params.get('order_product_id', [''])[0])
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            sql = """
                                SELECT
                                    ops.id,
                                    ops.order_product_id,
                                    src.sku AS order_sku,
                                    ops.plan_name,
                                    ops.updated_at
                                FROM order_product_shipping_plans ops
                                JOIN order_products src ON src.id = ops.order_product_id
                            """
                            params = []
                            if target_order_product_id:
                                sql += " WHERE ops.order_product_id=%s"
                                params.append(target_order_product_id)
                            sql += " ORDER BY ops.order_product_id ASC, ops.id ASC"
                            cur.execute(sql, params)
                            plans = cur.fetchall() or []

                            plan_ids = [int(row['id']) for row in plans if row.get('id')]
                            item_map = {plan_id: [] for plan_id in plan_ids}
                            if plan_ids:
                                placeholders = ','.join(['%s'] * len(plan_ids))
                                cur.execute(
                                    f"""
                                    SELECT
                                        opsi.shipping_plan_id,
                                        opsi.id,
                                        opsi.substitute_order_product_id,
                                        opsi.quantity,
                                        opsi.sort_order,
                                        op.sku AS substitute_order_sku
                                    FROM order_product_shipping_plan_items opsi
                                    JOIN order_products op ON op.id = opsi.substitute_order_product_id
                                    WHERE opsi.shipping_plan_id IN ({placeholders})
                                    ORDER BY opsi.shipping_plan_id ASC, opsi.sort_order ASC, opsi.id ASC
                                    """,
                                    plan_ids
                                )
                                for rel in cur.fetchall() or []:
                                    pid = int(rel.get('shipping_plan_id'))
                                    item_map.setdefault(pid, []).append({
                                        'id': rel.get('id'),
                                        'substitute_order_product_id': rel.get('substitute_order_product_id'),
                                        'substitute_order_sku': rel.get('substitute_order_sku') or '',
                                        'quantity': self._parse_int(rel.get('quantity')) or 1,
                                        'sort_order': self._parse_int(rel.get('sort_order')) or 1
                                    })

                    for plan in plans:
                        plan['items'] = item_map.get(int(plan.get('id') or 0), [])
                    return self.send_json({'status': 'success', 'items': plans}, start_response)

                data = self._read_json_body(environ)
                if method == 'POST':
                    order_product_id = self._parse_int(data.get('order_product_id'))
                    plan_name = (data.get('plan_name') or '').strip()
                    items = self._normalize_shipping_plan_items(data.get('items'))
                    if not order_product_id or not plan_name:
                        return self.send_json({'status': 'error', 'message': '缺少必填字段：order_product_id/plan_name'}, start_response)
                    if not items:
                        return self.send_json({'status': 'error', 'message': '请至少提供1条替代下单SKU'}, start_response)
                    if self._has_duplicate_shipping_plan_substitutes(items):
                        return self.send_json({'status': 'error', 'message': '替代下单SKU不允许重复'}, start_response)

                    with self._get_db_connection() as conn:
                        try:
                            conn.autocommit(False)
                            with conn.cursor() as cur:
                                cur.execute("SELECT id FROM order_products WHERE id=%s", (order_product_id,))
                                if not cur.fetchone():
                                    return self.send_json({'status': 'error', 'message': '原下单SKU不存在'}, start_response)
                                cur.execute(
                                    "SELECT id FROM order_product_shipping_plans WHERE order_product_id=%s AND plan_name=%s LIMIT 1",
                                    (order_product_id, plan_name)
                                )
                                if cur.fetchone():
                                    return self.send_json({'status': 'error', 'message': '同一SKU下方案名称已存在'}, start_response)
                                cur.execute(
                                    """
                                    INSERT INTO order_product_shipping_plans (order_product_id, plan_name)
                                    VALUES (%s, %s)
                                    """,
                                    (order_product_id, plan_name)
                                )
                                new_id = cur.lastrowid
                            self._replace_shipping_plan_items(conn, new_id, items)
                            conn.commit()
                        except Exception:
                            conn.rollback()
                            raise
                    return self.send_json({'status': 'success', 'id': new_id}, start_response)

                if method == 'PUT':
                    plan_id = self._parse_int(data.get('id'))
                    order_product_id = self._parse_int(data.get('order_product_id'))
                    plan_name = (data.get('plan_name') or '').strip()
                    items = self._normalize_shipping_plan_items(data.get('items'))
                    if not plan_id or not order_product_id or not plan_name:
                        return self.send_json({'status': 'error', 'message': '缺少必填字段：id/order_product_id/plan_name'}, start_response)
                    if not items:
                        return self.send_json({'status': 'error', 'message': '请至少提供1条替代下单SKU'}, start_response)
                    if self._has_duplicate_shipping_plan_substitutes(items):
                        return self.send_json({'status': 'error', 'message': '替代下单SKU不允许重复'}, start_response)

                    with self._get_db_connection() as conn:
                        try:
                            conn.autocommit(False)
                            with conn.cursor() as cur:
                                cur.execute(
                                    "SELECT id FROM order_product_shipping_plans WHERE order_product_id=%s AND plan_name=%s AND id<>%s LIMIT 1",
                                    (order_product_id, plan_name, plan_id)
                                )
                                if cur.fetchone():
                                    return self.send_json({'status': 'error', 'message': '同一SKU下方案名称已存在'}, start_response)
                                cur.execute(
                                    """
                                    UPDATE order_product_shipping_plans
                                    SET order_product_id=%s, plan_name=%s
                                    WHERE id=%s
                                    """,
                                    (order_product_id, plan_name, plan_id)
                                )
                                if cur.rowcount == 0:
                                    return self.send_json({'status': 'error', 'message': '方案不存在'}, start_response)
                            self._replace_shipping_plan_items(conn, plan_id, items)
                            conn.commit()
                        except Exception:
                            conn.rollback()
                            raise
                    return self.send_json({'status': 'success'}, start_response)

                if method == 'DELETE':
                    plan_id = self._parse_int(data.get('id'))
                    if not plan_id:
                        return self.send_json({'status': 'error', 'message': '缺少方案id'}, start_response)
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute("DELETE FROM order_product_shipping_plans WHERE id=%s", (plan_id,))
                    return self.send_json({'status': 'success'}, start_response)

                return self.send_error(405, 'Method not allowed', start_response)

            if method == 'GET':
                keyword = query_params.get('q', [''])[0].strip()
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        if keyword:
                            cur.execute(
                                """
                                SELECT
                                    op.id, op.sku, op.sku_family_id, op.version_no, op.fabric_id,
                                    op.spec_qty_short, op.contents_desc_en, op.is_iteration, op.is_dachene_product, op.is_on_market, op.source_order_product_id,
                                    op.finished_length_in, op.finished_width_in, op.finished_height_in,
                                    op.net_weight_lbs, op.package_length_in, op.package_width_in,
                                    op.package_height_in, op.gross_weight_lbs, op.cost_usd,
                                    op.carton_qty, op.package_size_class, op.last_mile_avg_freight_usd,
                                    op.created_at,
                                    pf.sku_family, pf.category,
                                    fm.fabric_code, fm.fabric_name_en,
                                    src.sku AS source_sku,
                                    GROUP_CONCAT(DISTINCT IF(mt.name='填充', m.name, NULL) ORDER BY m.name SEPARATOR ' / ') AS filling_materials,
                                    GROUP_CONCAT(DISTINCT IF(mt.name='框架', m.name, NULL) ORDER BY m.name SEPARATOR ' / ') AS frame_materials,
                                    GROUP_CONCAT(DISTINCT f.name ORDER BY f.name SEPARATOR ' / ') AS features,
                                    GROUP_CONCAT(DISTINCT IF(mt.name='填充', m.id, NULL) ORDER BY m.id SEPARATOR ',') AS filling_material_ids,
                                    GROUP_CONCAT(DISTINCT IF(mt.name='框架', m.id, NULL) ORDER BY m.id SEPARATOR ',') AS frame_material_ids,
                                    GROUP_CONCAT(DISTINCT f.id ORDER BY f.id SEPARATOR ',') AS feature_ids,
                                    GROUP_CONCAT(DISTINCT cft.name ORDER BY cft.name SEPARATOR ' / ') AS certifications,
                                    GROUP_CONCAT(DISTINCT cft.id ORDER BY cft.id SEPARATOR ',') AS certification_ids
                                FROM order_products op
                                LEFT JOIN product_families pf ON op.sku_family_id = pf.id
                                LEFT JOIN fabric_materials fm ON op.fabric_id = fm.id
                                LEFT JOIN order_product_materials opm ON opm.order_product_id = op.id
                                LEFT JOIN materials m ON opm.material_id = m.id
                                LEFT JOIN material_types mt ON m.material_type_id = mt.id
                                LEFT JOIN order_product_features opf ON opf.order_product_id = op.id
                                LEFT JOIN features f ON opf.feature_id = f.id
                                LEFT JOIN order_product_certifications opc ON opc.order_product_id = op.id
                                LEFT JOIN certifications cft ON cft.id = opc.certification_id
                                LEFT JOIN order_products src ON src.id = op.source_order_product_id
                                WHERE op.sku LIKE %s
                                   OR op.version_no LIKE %s
                                   OR pf.sku_family LIKE %s
                                   OR fm.fabric_code LIKE %s
                                GROUP BY op.id
                                ORDER BY op.id DESC
                                """,
                                (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%")
                            )
                        else:
                            cur.execute(
                                """
                                SELECT
                                    op.id, op.sku, op.sku_family_id, op.version_no, op.fabric_id,
                                    op.spec_qty_short, op.contents_desc_en, op.is_iteration, op.is_dachene_product, op.is_on_market, op.source_order_product_id,
                                    op.finished_length_in, op.finished_width_in, op.finished_height_in,
                                    op.net_weight_lbs, op.package_length_in, op.package_width_in,
                                    op.package_height_in, op.gross_weight_lbs, op.cost_usd,
                                    op.carton_qty, op.package_size_class, op.last_mile_avg_freight_usd,
                                    op.created_at,
                                    pf.sku_family, pf.category,
                                    fm.fabric_code, fm.fabric_name_en,
                                    src.sku AS source_sku,
                                    GROUP_CONCAT(DISTINCT IF(mt.name='填充', m.name, NULL) ORDER BY m.name SEPARATOR ' / ') AS filling_materials,
                                    GROUP_CONCAT(DISTINCT IF(mt.name='框架', m.name, NULL) ORDER BY m.name SEPARATOR ' / ') AS frame_materials,
                                    GROUP_CONCAT(DISTINCT f.name ORDER BY f.name SEPARATOR ' / ') AS features,
                                    GROUP_CONCAT(DISTINCT IF(mt.name='填充', m.id, NULL) ORDER BY m.id SEPARATOR ',') AS filling_material_ids,
                                    GROUP_CONCAT(DISTINCT IF(mt.name='框架', m.id, NULL) ORDER BY m.id SEPARATOR ',') AS frame_material_ids,
                                    GROUP_CONCAT(DISTINCT f.id ORDER BY f.id SEPARATOR ',') AS feature_ids,
                                    GROUP_CONCAT(DISTINCT cft.name ORDER BY cft.name SEPARATOR ' / ') AS certifications,
                                    GROUP_CONCAT(DISTINCT cft.id ORDER BY cft.id SEPARATOR ',') AS certification_ids
                                FROM order_products op
                                LEFT JOIN product_families pf ON op.sku_family_id = pf.id
                                LEFT JOIN fabric_materials fm ON op.fabric_id = fm.id
                                LEFT JOIN order_product_materials opm ON opm.order_product_id = op.id
                                LEFT JOIN materials m ON opm.material_id = m.id
                                LEFT JOIN material_types mt ON m.material_type_id = mt.id
                                LEFT JOIN order_product_features opf ON opf.order_product_id = op.id
                                LEFT JOIN features f ON opf.feature_id = f.id
                                LEFT JOIN order_product_certifications opc ON opc.order_product_id = op.id
                                LEFT JOIN certifications cft ON cft.id = opc.certification_id
                                LEFT JOIN order_products src ON src.id = op.source_order_product_id
                                GROUP BY op.id
                                ORDER BY op.id DESC
                                """
                            )
                        rows = cur.fetchall()
                return self.send_json({'status': 'success', 'items': rows}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                sku = (data.get('sku') or '').strip()
                sku_family_id = data.get('sku_family_id')
                version_no = (data.get('version_no') or '').strip()
                fabric_id = data.get('fabric_id')
                spec_qty_short = (data.get('spec_qty_short') or '').strip()
                contents_desc_en = (data.get('contents_desc_en') or '').strip()
                is_iteration = 1 if str(data.get('is_iteration') or '').lower() in ('1', 'true', 'yes', 'on') else 0
                is_dachene_product = 1 if str(data.get('is_dachene_product') or '').lower() in ('1', 'true', 'yes', 'on') else 0
                is_on_market = 0 if str(data.get('is_on_market') or '1').lower() in ('0', 'false', 'no', 'off') else 1
                source_order_product_id = self._parse_int(data.get('source_order_product_id'))

                if not sku or not sku_family_id or not fabric_id:
                    return self.send_json({'status': 'error', 'message': 'Missing required fields'}, start_response)
                if is_iteration and not source_order_product_id:
                    return self.send_json({'status': 'error', 'message': 'Missing source SKU'}, start_response)
                if is_iteration and not version_no:
                    return self.send_json({'status': 'error', 'message': 'Missing version'}, start_response)
                if not is_iteration:
                    source_order_product_id = None

                payload = {
                    'sku': sku,
                    'sku_family_id': self._parse_int(sku_family_id),
                    'version_no': version_no,
                    'fabric_id': self._parse_int(fabric_id),
                    'spec_qty_short': spec_qty_short,
                    'contents_desc_en': contents_desc_en or None,
                    'is_iteration': is_iteration,
                    'is_dachene_product': is_dachene_product,
                    'is_on_market': is_on_market,
                    'source_order_product_id': source_order_product_id,
                    'finished_length_in': self._parse_float(data.get('finished_length_in')),
                    'finished_width_in': self._parse_float(data.get('finished_width_in')),
                    'finished_height_in': self._parse_float(data.get('finished_height_in')),
                    'net_weight_lbs': self._parse_float(data.get('net_weight_lbs')),
                    'package_length_in': self._parse_float(data.get('package_length_in')),
                    'package_width_in': self._parse_float(data.get('package_width_in')),
                    'package_height_in': self._parse_float(data.get('package_height_in')),
                    'gross_weight_lbs': self._parse_float(data.get('gross_weight_lbs')),
                    'cost_usd': self._parse_float(data.get('cost_usd')),
                    'carton_qty': self._parse_int(data.get('carton_qty')),
                    'package_size_class': (data.get('package_size_class') or '').strip() or None,
                    'last_mile_avg_freight_usd': self._parse_float(data.get('last_mile_avg_freight_usd')),
                }

                filling_material_ids = [self._parse_int(v) for v in (data.get('filling_material_ids') or [])]
                frame_material_ids = [self._parse_int(v) for v in (data.get('frame_material_ids') or [])]
                feature_ids = [self._parse_int(v) for v in (data.get('feature_ids') or [])]
                certification_ids = [self._parse_int(v) for v in (data.get('certification_ids') or [])]
                filling_material_ids = [v for v in filling_material_ids if v]
                frame_material_ids = [v for v in frame_material_ids if v]
                feature_ids = [v for v in feature_ids if v]
                certification_ids = [v for v in certification_ids if v]

                with self._get_db_connection() as conn:
                    try:
                        conn.autocommit(False)
                        with conn.cursor() as cur:
                            if source_order_product_id:
                                cur.execute(
                                    "SELECT id, is_iteration FROM order_products WHERE id=%s",
                                    (source_order_product_id,)
                                )
                                src_row = cur.fetchone() or {}
                                if not src_row:
                                    return self.send_json({'status': 'error', 'message': '来源SKU不存在'}, start_response)
                                if int(src_row.get('is_iteration') or 0) == 1:
                                    return self.send_json({'status': 'error', 'message': '迭代款SKU不能作为来源SKU'}, start_response)

                            if is_iteration and source_order_product_id and version_no:
                                cur.execute(
                                    """
                                    SELECT id FROM order_products
                                    WHERE source_order_product_id=%s AND version_no=%s
                                    LIMIT 1
                                    """,
                                    (source_order_product_id, version_no)
                                )
                                dup_row = cur.fetchone()
                                if dup_row:
                                    return self.send_json({'status': 'error', 'message': '同一来源SKU下版本号已存在'}, start_response)

                            cur.execute(
                                """
                                INSERT INTO order_products (
                                    sku, sku_family_id, version_no, fabric_id, spec_qty_short, contents_desc_en,
                                    is_iteration, is_dachene_product, is_on_market, source_order_product_id,
                                    finished_length_in, finished_width_in, finished_height_in,
                                    net_weight_lbs, package_length_in, package_width_in, package_height_in,
                                    gross_weight_lbs, cost_usd, carton_qty, package_size_class, last_mile_avg_freight_usd
                                ) VALUES (
                                    %(sku)s, %(sku_family_id)s, %(version_no)s, %(fabric_id)s, %(spec_qty_short)s, %(contents_desc_en)s,
                                    %(is_iteration)s, %(is_dachene_product)s, %(is_on_market)s, %(source_order_product_id)s,
                                    %(finished_length_in)s, %(finished_width_in)s, %(finished_height_in)s,
                                    %(net_weight_lbs)s, %(package_length_in)s, %(package_width_in)s, %(package_height_in)s,
                                    %(gross_weight_lbs)s, %(cost_usd)s, %(carton_qty)s, %(package_size_class)s, %(last_mile_avg_freight_usd)s
                                )
                                """,
                                payload
                            )
                            new_id = cur.lastrowid
                            if is_iteration and source_order_product_id:
                                self._ensure_default_iteration_shipping_plans(conn, new_id)

                        self._replace_order_product_material_ids(conn, new_id, filling_material_ids, frame_material_ids)
                        self._replace_order_product_feature_ids(conn, new_id, feature_ids)
                        self._replace_order_product_certification_ids(conn, new_id, certification_ids)
                        conn.commit()
                    except Exception:
                        conn.rollback()
                        raise

                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)

                batch_items = data.get('items') if isinstance(data, dict) else None
                if isinstance(batch_items, list):
                    normalized_rows = []
                    seen_ids = set()
                    for entry in batch_items:
                        if not isinstance(entry, dict):
                            continue
                        item_id = self._parse_int(entry.get('id'))
                        if not item_id or item_id in seen_ids:
                            continue
                        seen_ids.add(item_id)
                        normalized_rows.append({
                            'id': item_id,
                            'cost_usd': self._parse_float(entry.get('cost_usd')),
                            'package_size_class': (entry.get('package_size_class') or '').strip() or None,
                            'carton_qty': self._parse_int(entry.get('carton_qty')),
                            'last_mile_avg_freight_usd': self._parse_float(entry.get('last_mile_avg_freight_usd')),
                        })

                    if not normalized_rows:
                        return self.send_json({'status': 'error', 'message': 'Missing valid items'}, start_response)

                    ids = [row['id'] for row in normalized_rows]
                    placeholders = ','.join(['%s'] * len(ids))
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                f"SELECT id FROM order_products WHERE id IN ({placeholders})",
                                ids
                            )
                            existing_rows = cur.fetchall() or []
                            existing_ids = {int(row.get('id')) for row in existing_rows if row.get('id')}
                            payload_rows = [
                                (
                                    row['cost_usd'],
                                    row['package_size_class'],
                                    row['carton_qty'],
                                    row['last_mile_avg_freight_usd'],
                                    row['id']
                                )
                                for row in normalized_rows
                                if row['id'] in existing_ids
                            ]
                            if payload_rows:
                                cur.executemany(
                                    """
                                    UPDATE order_products
                                    SET cost_usd=%s,
                                        package_size_class=%s,
                                        carton_qty=%s,
                                        last_mile_avg_freight_usd=%s
                                    WHERE id=%s
                                    """,
                                    payload_rows
                                )
                    return self.send_json({'status': 'success', 'updated': len(normalized_rows)}, start_response)

                item_id = data.get('id')
                sku = (data.get('sku') or '').strip()
                sku_family_id = data.get('sku_family_id')
                version_no = (data.get('version_no') or '').strip()
                fabric_id = data.get('fabric_id')
                spec_qty_short = (data.get('spec_qty_short') or '').strip()
                contents_desc_en = (data.get('contents_desc_en') or '').strip()
                is_iteration = 1 if str(data.get('is_iteration') or '').lower() in ('1', 'true', 'yes', 'on') else 0
                is_dachene_product = 1 if str(data.get('is_dachene_product') or '').lower() in ('1', 'true', 'yes', 'on') else 0
                is_on_market = 0 if str(data.get('is_on_market') or '1').lower() in ('0', 'false', 'no', 'off') else 1
                source_order_product_id = self._parse_int(data.get('source_order_product_id'))

                if not item_id or not sku or not sku_family_id or not fabric_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id or fields'}, start_response)
                if is_iteration and not source_order_product_id:
                    return self.send_json({'status': 'error', 'message': 'Missing source SKU'}, start_response)
                if is_iteration and not version_no:
                    return self.send_json({'status': 'error', 'message': 'Missing version'}, start_response)
                if not is_iteration:
                    source_order_product_id = None
                if source_order_product_id and int(source_order_product_id) == int(item_id):
                    return self.send_json({'status': 'error', 'message': 'Source SKU cannot be itself'}, start_response)

                payload = {
                    'id': item_id,
                    'sku': sku,
                    'sku_family_id': self._parse_int(sku_family_id),
                    'version_no': version_no,
                    'fabric_id': self._parse_int(fabric_id),
                    'spec_qty_short': spec_qty_short,
                    'contents_desc_en': contents_desc_en or None,
                    'is_iteration': is_iteration,
                    'is_dachene_product': is_dachene_product,
                    'is_on_market': is_on_market,
                    'source_order_product_id': source_order_product_id,
                    'finished_length_in': self._parse_float(data.get('finished_length_in')),
                    'finished_width_in': self._parse_float(data.get('finished_width_in')),
                    'finished_height_in': self._parse_float(data.get('finished_height_in')),
                    'net_weight_lbs': self._parse_float(data.get('net_weight_lbs')),
                    'package_length_in': self._parse_float(data.get('package_length_in')),
                    'package_width_in': self._parse_float(data.get('package_width_in')),
                    'package_height_in': self._parse_float(data.get('package_height_in')),
                    'gross_weight_lbs': self._parse_float(data.get('gross_weight_lbs')),
                    'cost_usd': self._parse_float(data.get('cost_usd')),
                    'carton_qty': self._parse_int(data.get('carton_qty')),
                    'package_size_class': (data.get('package_size_class') or '').strip() or None,
                    'last_mile_avg_freight_usd': self._parse_float(data.get('last_mile_avg_freight_usd')),
                }

                filling_material_ids = [self._parse_int(v) for v in (data.get('filling_material_ids') or [])]
                frame_material_ids = [self._parse_int(v) for v in (data.get('frame_material_ids') or [])]
                feature_ids = [self._parse_int(v) for v in (data.get('feature_ids') or [])]
                certification_ids = [self._parse_int(v) for v in (data.get('certification_ids') or [])]
                filling_material_ids = [v for v in filling_material_ids if v]
                frame_material_ids = [v for v in frame_material_ids if v]
                feature_ids = [v for v in feature_ids if v]
                certification_ids = [v for v in certification_ids if v]

                with self._get_db_connection() as conn:
                    try:
                        conn.autocommit(False)
                        with conn.cursor() as cur:
                            if source_order_product_id:
                                cur.execute(
                                    "SELECT id, is_iteration FROM order_products WHERE id=%s",
                                    (source_order_product_id,)
                                )
                                src_row = cur.fetchone() or {}
                                if not src_row:
                                    return self.send_json({'status': 'error', 'message': '来源SKU不存在'}, start_response)
                                if int(src_row.get('is_iteration') or 0) == 1:
                                    return self.send_json({'status': 'error', 'message': '迭代款SKU不能作为来源SKU'}, start_response)

                            if is_iteration and source_order_product_id and version_no:
                                cur.execute(
                                    """
                                    SELECT id FROM order_products
                                    WHERE source_order_product_id=%s AND version_no=%s AND id<>%s
                                    LIMIT 1
                                    """,
                                    (source_order_product_id, version_no, item_id)
                                )
                                dup_row = cur.fetchone()
                                if dup_row:
                                    return self.send_json({'status': 'error', 'message': '同一来源SKU下版本号已存在'}, start_response)

                            cur.execute(
                                """
                                UPDATE order_products
                                SET sku=%(sku)s,
                                    sku_family_id=%(sku_family_id)s,
                                    version_no=%(version_no)s,
                                    fabric_id=%(fabric_id)s,
                                    spec_qty_short=%(spec_qty_short)s,
                                    contents_desc_en=%(contents_desc_en)s,
                                    is_iteration=%(is_iteration)s,
                                    is_dachene_product=%(is_dachene_product)s,
                                    is_on_market=%(is_on_market)s,
                                    source_order_product_id=%(source_order_product_id)s,
                                    finished_length_in=%(finished_length_in)s,
                                    finished_width_in=%(finished_width_in)s,
                                    finished_height_in=%(finished_height_in)s,
                                    net_weight_lbs=%(net_weight_lbs)s,
                                    package_length_in=%(package_length_in)s,
                                    package_width_in=%(package_width_in)s,
                                    package_height_in=%(package_height_in)s,
                                    gross_weight_lbs=%(gross_weight_lbs)s,
                                    cost_usd=%(cost_usd)s,
                                    carton_qty=%(carton_qty)s,
                                    package_size_class=%(package_size_class)s,
                                    last_mile_avg_freight_usd=%(last_mile_avg_freight_usd)s
                                WHERE id=%(id)s
                                """,
                                payload
                            )

                        self._replace_order_product_material_ids(conn, item_id, filling_material_ids, frame_material_ids)
                        self._replace_order_product_feature_ids(conn, item_id, feature_ids)
                        self._replace_order_product_certification_ids(conn, item_id, certification_ids)
                        conn.commit()
                    except Exception:
                        conn.rollback()
                        raise

                return self.send_json({'status': 'success'}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = data.get('id')
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM order_products WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': 'SKU 已存在'}, start_response)
            print("Order product API error: " + str(e))
            return self.send_error(500, str(e), start_response)

    def handle_order_product_carton_calc_api(self, environ, method, start_response):
        """批量根据包裹尺寸计算并更新装箱量（40HQ=69m³，向下取整）"""
        try:
            self._ensure_order_product_tables()
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)

            data = self._read_json_body(environ)
            ids = data.get('ids') or []
            normalized_ids = []
            for value in ids:
                parsed = self._parse_int(value)
                if parsed and parsed not in normalized_ids:
                    normalized_ids.append(parsed)
            if not normalized_ids:
                return self.send_json({'status': 'error', 'message': '请先选择需要更新的SKU'}, start_response)

            placeholders = ','.join(['%s'] * len(normalized_ids))
            updated = 0
            skipped = 0
            errors = []

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        f"""
                        SELECT id, sku, package_length_in, package_width_in, package_height_in
                        FROM order_products
                        WHERE id IN ({placeholders})
                        """,
                        normalized_ids
                    )
                    rows = cur.fetchall() or []

                for row in rows:
                    item_id = row.get('id')
                    sku = row.get('sku') or ''
                    carton_qty = self._calc_carton_qty_by_40hq(
                        row.get('package_length_in'),
                        row.get('package_width_in'),
                        row.get('package_height_in')
                    )
                    if carton_qty is None:
                        skipped += 1
                        errors.append({'row': sku or item_id, 'error': '包裹长宽高缺失或无效，无法计算装箱量'})
                        continue
                    with conn.cursor() as cur:
                        cur.execute("UPDATE order_products SET carton_qty=%s WHERE id=%s", (carton_qty, item_id))
                        if cur.rowcount:
                            updated += 1
                        else:
                            skipped += 1

            return self.send_json({
                'status': 'success',
                'updated': updated,
                'skipped': skipped,
                'errors': errors
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_order_product_template_api(self, environ, method, start_response):
        """下单产品模板下载"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation

            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            selected_ids = []
            for raw in query_params.get('ids', []):
                for token in re.split(r'[,，;；\s]+', str(raw or '').strip()):
                    if not token:
                        continue
                    item_id = self._parse_int(token)
                    if item_id and item_id not in selected_ids:
                        selected_ids.append(item_id)

            self._ensure_order_product_tables()
            wb = Workbook()
            ws = wb.active
            ws.title = 'order_products'

            max_multi_columns = {
                'filling_materials': 3,
                'frame_materials': 3,
                'features': 3,
                'certifications': 3,
            }
            export_rows = []
            filling_rel = {}
            frame_rel = {}
            feature_rel = {}
            cert_rel = {}

            # 获取所有可用的数据用于下拉菜单 + 勾选导出数据
            with self._get_db_connection() as conn:
                def _load_order_template_options():
                    with conn.cursor() as cur:
                        cur.execute("SELECT sku_family FROM product_families ORDER BY sku_family")
                        sku_families_local = [row['sku_family'] for row in cur.fetchall()]

                        cur.execute("SELECT fabric_code FROM fabric_materials ORDER BY fabric_code")
                        fabrics_local = [row['fabric_code'] for row in cur.fetchall()]

                        cur.execute("""
                            SELECT m.name
                            FROM materials m
                            JOIN material_types mt ON m.material_type_id = mt.id
                            WHERE mt.name = '填充'
                            ORDER BY m.name
                        """)
                        filling_local = [row['name'] for row in cur.fetchall()]

                        cur.execute("""
                            SELECT m.name
                            FROM materials m
                            JOIN material_types mt ON m.material_type_id = mt.id
                            WHERE mt.name = '框架'
                            ORDER BY m.name
                        """)
                        frame_local = [row['name'] for row in cur.fetchall()]

                        cur.execute("SELECT name FROM features ORDER BY name")
                        feature_local = [row['name'] for row in cur.fetchall()]

                        cur.execute("SELECT name FROM certifications ORDER BY name")
                        cert_local = [row['name'] for row in cur.fetchall()]
                    return (sku_families_local, fabrics_local, filling_local, frame_local, feature_local, cert_local)

                sku_families, fabrics, filling_materials, frame_materials, features, certifications = self._get_cached_template_options(
                    'order_product_template_options_v1',
                    _load_order_template_options,
                    ttl_seconds=180
                )

                with conn.cursor() as cur:

                    if selected_ids:
                        placeholders = ','.join(['%s'] * len(selected_ids))
                        cur.execute(
                            f"""
                            SELECT
                                op.id,
                                op.sku,
                                pf.sku_family,
                                op.version_no,
                                fm.fabric_code,
                                op.spec_qty_short,
                                op.contents_desc_en,
                                op.is_iteration,
                                op.is_dachene_product,
                                src.sku AS source_sku,
                                op.finished_length_in,
                                op.finished_width_in,
                                op.finished_height_in,
                                op.net_weight_lbs,
                                op.package_length_in,
                                op.package_width_in,
                                op.package_height_in,
                                op.gross_weight_lbs,
                                op.cost_usd,
                                op.carton_qty,
                                op.package_size_class,
                                op.last_mile_avg_freight_usd
                            FROM order_products op
                            LEFT JOIN product_families pf ON op.sku_family_id = pf.id
                            LEFT JOIN fabric_materials fm ON op.fabric_id = fm.id
                            LEFT JOIN order_products src ON op.source_order_product_id = src.id
                            WHERE op.id IN ({placeholders})
                            """,
                            tuple(selected_ids)
                        )
                        selected_rows = cur.fetchall() or []
                        order_map = {sid: idx for idx, sid in enumerate(selected_ids)}
                        selected_rows.sort(key=lambda x: order_map.get(x.get('id'), 10 ** 6))
                        export_rows = selected_rows

                        if selected_rows:
                            row_ids = [row['id'] for row in selected_rows]
                            rel_placeholders = ','.join(['%s'] * len(row_ids))

                            cur.execute(
                                f"""
                                SELECT opm.order_product_id, m.name, mt.name AS type_name
                                FROM order_product_materials opm
                                JOIN materials m ON opm.material_id = m.id
                                JOIN material_types mt ON m.material_type_id = mt.id
                                WHERE opm.order_product_id IN ({rel_placeholders})
                                ORDER BY opm.order_product_id, m.name
                                """,
                                tuple(row_ids)
                            )
                            for rel in cur.fetchall() or []:
                                target = filling_rel if rel.get('type_name') == '填充' else frame_rel
                                target.setdefault(rel['order_product_id'], [])
                                if rel['name'] not in target[rel['order_product_id']]:
                                    target[rel['order_product_id']].append(rel['name'])

                            cur.execute(
                                f"""
                                SELECT opf.order_product_id, f.name
                                FROM order_product_features opf
                                JOIN features f ON opf.feature_id = f.id
                                WHERE opf.order_product_id IN ({rel_placeholders})
                                ORDER BY opf.order_product_id, f.name
                                """,
                                tuple(row_ids)
                            )
                            for rel in cur.fetchall() or []:
                                feature_rel.setdefault(rel['order_product_id'], [])
                                if rel['name'] not in feature_rel[rel['order_product_id']]:
                                    feature_rel[rel['order_product_id']].append(rel['name'])

                            cur.execute(
                                f"""
                                SELECT opc.order_product_id, c.name
                                FROM order_product_certifications opc
                                JOIN certifications c ON opc.certification_id = c.id
                                WHERE opc.order_product_id IN ({rel_placeholders})
                                ORDER BY opc.order_product_id, c.name
                                """,
                                tuple(row_ids)
                            )
                            for rel in cur.fetchall() or []:
                                cert_rel.setdefault(rel['order_product_id'], [])
                                if rel['name'] not in cert_rel[rel['order_product_id']]:
                                    cert_rel[rel['order_product_id']].append(rel['name'])

                            for row in selected_rows:
                                rid = row['id']
                                max_multi_columns['filling_materials'] = max(max_multi_columns['filling_materials'], len(filling_rel.get(rid, [])))
                                max_multi_columns['frame_materials'] = max(max_multi_columns['frame_materials'], len(frame_rel.get(rid, [])))
                                max_multi_columns['features'] = max(max_multi_columns['features'], len(feature_rel.get(rid, [])))
                                max_multi_columns['certifications'] = max(max_multi_columns['certifications'], len(cert_rel.get(rid, [])))
            
            # 定义组件和字段（带中文标签）
            sections = [
                {
                    'title': '迭代款',
                    'bg_color': 'E8DFD4',
                    'fields': [
                        ('is_iteration', '是否迭代款', 'dropdown', ['否', '是']),
                        ('is_dachene_product', '是否为大健云仓产品（在下单SKU处填写大健云仓Item Code）', 'dropdown', ['否', '是']),
                        ('source_sku', '来源下单SKU', 'text', None),
                        ('version_no', '版本号', 'text', None)
                    ]
                },
                {
                    'title': '基础信息',
                    'bg_color': 'F5F1ED',
                    'fields': [
                        ('sku', '下单SKU *', 'text', None),
                        ('sku_family', '归属货号 *', 'dropdown', sku_families),
                        ('fabric_code', '面料 *', 'dropdown', fabrics),
                        ('spec_qty_short', '规格与数量简称', 'text', None),
                        ('contents_desc_en', '内含物英文描述', 'text', None)
                    ]
                },
                {
                    'title': '成品尺寸/重量',
                    'bg_color': 'E8DFD4',
                    'fields': [
                        ('finished_length_in', '成品长(inch)', 'number', None),
                        ('finished_width_in', '成品宽(inch)', 'number', None),
                        ('finished_height_in', '成品高(inch)', 'number', None),
                        ('net_weight_lbs', '净重(lbs)', 'number', None)
                    ]
                },
                {
                    'title': '包裹尺寸/重量',
                    'bg_color': 'F5F1ED',
                    'fields': [
                        ('package_length_in', '包裹长(inch)', 'number', None),
                        ('package_width_in', '包裹宽(inch)', 'number', None),
                        ('package_height_in', '包裹高(inch)', 'number', None),
                        ('gross_weight_lbs', '毛重(lbs)', 'number', None),
                        ('carton_qty', '装箱量', 'number', None),
                        ('package_size_class', '包裹大小归类(Fedx)', 'text', None)
                    ]
                },
                {
                    'title': '成本',
                    'bg_color': 'E8DFD4',
                    'fields': [
                        ('cost_usd', '产品成本及发货至海外仓成本估算(USD，不含仓储费)', 'number', None),
                        ('last_mile_avg_freight_usd', '尾程平均运费(美元)', 'number', None)
                    ]
                },
                {
                    'title': '材料与卖点',
                    'bg_color': 'F5F1ED',
                    'fields': [
                        ('filling_materials', '填充材料(可多项)', 'multi_dropdown', filling_materials),
                        ('frame_materials', '框架材料(可多项)', 'multi_dropdown', frame_materials),
                        ('features', '卖点特点(可多项)', 'multi_dropdown', features),
                        ('certifications', '认证(可多项)', 'multi_dropdown', certifications)
                    ]
                }
            ]
            
            # 建立模块标题行和列名行
            section_headers = []  # 模块名称行
            column_headers = []   # 列名行
            header_to_column = {}  # 用于数据验证时查找列
            col_idx = 0
            field_to_options = {}  # 记录字段对应的可选项
            
            for section in sections:
                section_title = section['title']
                section_start_col = col_idx
                
                for field_info in section['fields']:
                    field_code = field_info[0]
                    field_label = field_info[1]
                    field_type = field_info[2]
                    field_options = field_info[3] if len(field_info) > 3 else None
                    
                    if field_type == 'multi_dropdown':
                        num_cols = max_multi_columns.get(field_code, 3)
                        for i in range(1, num_cols + 1):
                            col_name = f"{field_code}_{i}"
                            column_headers.append(field_label if i == 1 else '')
                            header_to_column[col_name] = col_idx
                            field_to_options[col_name] = field_options
                            col_idx += 1
                    else:
                        column_headers.append(field_label)
                        header_to_column[field_code] = col_idx
                        if field_options:
                            field_to_options[field_code] = field_options
                        col_idx += 1
                
                # 填充模块标题（需要合并的列数）
                section_span = col_idx - section_start_col
                section_headers.append((section_title, section_start_col, section_span, section.get('bg_color') or 'CFC7BD'))
            
            # 第1行：模块标题（合并单元格）
            for i in range(col_idx):
                ws.cell(row=1, column=i+1).value = ''  # 先填充空值
            
            title_font = Font(bold=True, color='2A2420', size=11)
            title_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='B7AEA4'),
                right=Side(style='thin', color='B7AEA4'),
                top=Side(style='thin', color='B7AEA4'),
                bottom=Side(style='thin', color='B7AEA4')
            )
            
            for title, start_col, span, bg_color in section_headers:
                if span > 1:
                    ws.merge_cells(start_row=1, start_column=start_col+1, end_row=1, end_column=start_col+span)
                ws.cell(row=1, column=start_col+1).value = title
                title_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                for col in range(start_col, start_col + span):
                    ws.cell(row=1, column=col+1).fill = title_fill
                    ws.cell(row=1, column=col+1).font = title_font
                    ws.cell(row=1, column=col+1).alignment = title_alignment
                    ws.cell(row=1, column=col+1).border = thin_border
            
            # 第2行：列名
            for idx, header in enumerate(column_headers):
                cell = ws.cell(row=2, column=idx+1)
                cell.value = header
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # 第3行：示例行
            example_row_idx = 3
            example_row_data = []
            
            for col_name in list(header_to_column.keys()):
                field_base = col_name.rsplit('_', 1)[0] if '_' in col_name else col_name
                
                if field_base == 'is_iteration':
                    example_row_data.append(('是否迭代款', 0, '否'))
                elif field_base == 'is_dachene_product':
                    example_row_data.append(('是否为大健云仓产品（在下单SKU处填写大健云仓Item Code）', 0, '否'))
                elif field_base == 'sku':
                    example_row_data.append(('下单SKU', 0, 'MS01A-Brown'))
                elif field_base == 'sku_family':
                    example_row_data.append(('归属货号', 0, 'MS01'))
                elif field_base == 'fabric_code':
                    example_row_data.append(('面料', 0, 'Brown'))
                elif field_base == 'spec_qty_short':
                    example_row_data.append(('规格与数量简称', 0, 'A'))
                elif field_base == 'contents_desc_en':
                    example_row_data.append(('内含物英文描述', 0, 'memory foam + metal frame'))
                elif field_base == 'version_no':
                    example_row_data.append(('版本号', 0, '1'))
                elif field_base == 'source_sku':
                    example_row_data.append(('来源下单SKU', 0, ''))
                elif field_base == 'finished_length_in':
                    example_row_data.append(('成品长(inch)', 0, 30))
                elif field_base == 'finished_width_in':
                    example_row_data.append(('成品宽(inch)', 0, 20))
                elif field_base == 'finished_height_in':
                    example_row_data.append(('成品高(inch)', 0, 10))
                elif field_base == 'net_weight_lbs':
                    example_row_data.append(('净重(lbs)', 0, 5.5))
                elif field_base == 'package_length_in':
                    example_row_data.append(('包裹长(inch)', 0, 32))
                elif field_base == 'package_width_in':
                    example_row_data.append(('包裹宽(inch)', 0, 22))
                elif field_base == 'package_height_in':
                    example_row_data.append(('包裹高(inch)', 0, 12))
                elif field_base == 'gross_weight_lbs':
                    example_row_data.append(('毛重(lbs)', 0, 6.5))
                elif field_base == 'cost_usd':
                    example_row_data.append(('产品成本及发货至海外仓成本估算(USD，不含仓储费)', 0, 25.00))
                elif field_base == 'carton_qty':
                    example_row_data.append(('装箱量', 0, 50))
                elif field_base == 'package_size_class':
                    example_row_data.append(('包裹大小归类(Fedx)', 0, 'Small'))
                elif field_base == 'last_mile_avg_freight_usd':
                    example_row_data.append(('尾程平均运费(美元)', 0, 3.50))
                elif field_base in ['filling_materials', 'frame_materials', 'features', 'certifications']:
                    # 多选字段只在第一列填充示例
                    if col_name.endswith('_1'):
                        if field_base == 'filling_materials':
                            example_row_data.append(('填充材料(可多项)', 0, '海绵'))
                        elif field_base == 'frame_materials':
                            example_row_data.append(('框架材料(可多项)', 0, '金属'))
                        elif field_base == 'features':
                            example_row_data.append(('卖点特点(可多项)', 0, '可拆洗'))
                        elif field_base == 'certifications':
                            example_row_data.append(('认证(可多项)', 0, 'CE'))
                    else:
                        example_row_data.append(('', 0, None))
                else:
                    example_row_data.append(('', 0, None))
            
            for idx, (label, unused, value) in enumerate(example_row_data):
                cell = ws.cell(row=example_row_idx, column=idx+1)
                cell.value = value
                cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                cell.font = Font(italic=True, color='888888')
            
            # 辅助函数：将列索引转换为Excel列字母
            def col_idx_to_letter(idx):
                """将0-based列索引转换为Excel列字母"""
                result = ''
                while idx >= 0:
                    result = chr(65 + (idx % 26)) + result
                    idx = idx // 26 - 1
                return result
            
            # 添加数据验证
            yes_no_validation = DataValidation(type='list', formula1='"否,是"', allow_blank=True)
            ws.add_data_validation(yes_no_validation)
            max_validation_row = 400
            
            for bool_field in ('is_iteration', 'is_dachene_product'):
                if bool_field in header_to_column:
                    col_letter = col_idx_to_letter(header_to_column[bool_field])
                    for row in range(4, max_validation_row + 1):
                        yes_no_validation.add(f'{col_letter}{row}')
            
            # 为下拉字段添加验证
            for field_name, options in field_to_options.items():
                if options and field_name in header_to_column:
                    col_idx = header_to_column[field_name]
                    col_letter = col_idx_to_letter(col_idx)
                    
                    validation = DataValidation(type='list', formula1=f'"{",".join(options)}"', allow_blank=True)
                    ws.add_data_validation(validation)

                    for row in range(4, max_validation_row + 1):
                        validation.add(f'{col_letter}{row}')

            if export_rows:
                def set_multi_values(row_idx, field_name, values):
                    values = values or []
                    total = max_multi_columns.get(field_name, 3)
                    for i in range(1, total + 1):
                        key = f'{field_name}_{i}'
                        if key not in header_to_column:
                            continue
                        value = values[i - 1] if i - 1 < len(values) else None
                        ws.cell(row=row_idx, column=header_to_column[key] + 1).value = value

                data_row = 4
                for item in export_rows:
                    row_id = item.get('id')
                    direct_values = {
                        'is_iteration': '是' if str(item.get('is_iteration') or '0') in ('1', 'True', 'true') else '否',
                        'is_dachene_product': '是' if str(item.get('is_dachene_product') or '0') in ('1', 'True', 'true') else '否',
                        'source_sku': item.get('source_sku') or '',
                        'version_no': item.get('version_no') or '',
                        'sku': item.get('sku') or '',
                        'sku_family': item.get('sku_family') or '',
                        'fabric_code': item.get('fabric_code') or '',
                        'spec_qty_short': item.get('spec_qty_short') or '',
                        'contents_desc_en': item.get('contents_desc_en') or '',
                        'finished_length_in': item.get('finished_length_in'),
                        'finished_width_in': item.get('finished_width_in'),
                        'finished_height_in': item.get('finished_height_in'),
                        'net_weight_lbs': item.get('net_weight_lbs'),
                        'package_length_in': item.get('package_length_in'),
                        'package_width_in': item.get('package_width_in'),
                        'package_height_in': item.get('package_height_in'),
                        'gross_weight_lbs': item.get('gross_weight_lbs'),
                        'cost_usd': item.get('cost_usd'),
                        'carton_qty': item.get('carton_qty'),
                        'package_size_class': item.get('package_size_class') or '',
                        'last_mile_avg_freight_usd': item.get('last_mile_avg_freight_usd'),
                    }
                    for field_name, value in direct_values.items():
                        if field_name not in header_to_column:
                            continue
                        ws.cell(row=data_row, column=header_to_column[field_name] + 1).value = value

                    set_multi_values(data_row, 'filling_materials', filling_rel.get(row_id, []))
                    set_multi_values(data_row, 'frame_materials', frame_rel.get(row_id, []))
                    set_multi_values(data_row, 'features', feature_rel.get(row_id, []))
                    set_multi_values(data_row, 'certifications', cert_rel.get(row_id, []))
                    data_row += 1
            
            # 设置列宽
            for idx, header in enumerate(column_headers):
                col_letter = col_idx_to_letter(idx)
                if '材料' in header or '特点' in header or '认证' in header:
                    ws.column_dimensions[col_letter].width = 18
                elif 'SKU' in header:
                    ws.column_dimensions[col_letter].width = 15
                elif '简称' in header:
                    ws.column_dimensions[col_letter].width = 12
                else:
                    ws.column_dimensions[col_letter].width = 14
            
            # 冻结表头
            ws.freeze_panes = 'A4'
            
            return self._send_excel_workbook(wb, 'order_product_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_order_product_import_api(self, environ, method, start_response):
        """下单产品批量导入"""
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            preview_mode = str((query_params.get('preview', ['0'])[0] or '0')).lower() in ('1', 'true', 'yes', 'on')

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            file_bytes = self._sanitize_xlsx_bool_cells(file_bytes)

            # load and sanitize workbook
            try:
                wb = load_workbook(io.BytesIO(file_bytes))
            except Exception as e:
                if 'Cannot be converted to bool' in str(e):
                    wb = self._rebuild_workbook_from_xlsx_xml(file_bytes)
                    if wb is None:
                        diag = self._scan_xlsx_invalid_bool_cells(file_bytes)
                        return self.send_json({
                            'status': 'error',
                            'message': (
                                '导入失败：文件中存在异常布尔字段且无法自动修复，'
                                '请另存为新的xlsx后重试'
                            ),
                            'debug': {
                                'cause': 'Cannot be converted to bool',
                                'invalid_bool_cells': diag.get('count', 0),
                                'samples': diag.get('samples', [])
                            }
                        }, start_response)
                else:
                    return self.send_json({'status': 'error', 'message': str(e)}, start_response)

            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'b' and not isinstance(cell.value, bool):
                            cell.data_type = 's'
                            cell.value = str(cell.value)

            ws = wb.active

            # 支持两种表头格式：新的中文表头（从第2行）或旧的字段代码表头（从第1行）
            header_row_idx = 2 if ws.cell(row=1, column=1).value in ['迭代款', '基础信息', '成品尺寸/重量', '包裹尺寸/重量', '成本与物流', '成本', '材料与卖点'] else 1
            
            headers = [cell.value for cell in ws[header_row_idx]]
            
            # 中文字段标签到字段代码的映射
            label_to_code = {
                '是否迭代款': 'is_iteration',
                '是否为大健云仓产品（在下单SKU处填写大健云仓Item Code）': 'is_dachene_product',
                '来源下单SKU': 'source_sku',
                '版本号': 'version_no',
                '下单SKU *': 'sku',
                '归属货号 *': 'sku_family',
                '面料 *': 'fabric_code',
                '规格与数量简称': 'spec_qty_short',
                '内含物英文描述': 'contents_desc_en',
                '成品长(inch)': 'finished_length_in',
                '成品宽(inch)': 'finished_width_in',
                '成品高(inch)': 'finished_height_in',
                '净重(lbs)': 'net_weight_lbs',
                '包裹长(inch)': 'package_length_in',
                '包裹宽(inch)': 'package_width_in',
                '包裹高(inch)': 'package_height_in',
                '毛重(lbs)': 'gross_weight_lbs',
                '成本价(美元)': 'cost_usd',
                '产品成本及发货至海外仓成本估算(USD，不含仓储费)': 'cost_usd',
                '装箱量': 'carton_qty',
                '包裹大小归类(Fedx)': 'package_size_class',
                '尾程平均运费(美元)': 'last_mile_avg_freight_usd',
                '填充材料(可多项)': 'filling_materials',
                '框架材料(可多项)': 'frame_materials',
                '卖点特点(可多项)': 'features',
                '认证(可多项)': 'certifications'
            }

            multi_base_fields = {'filling_materials', 'frame_materials', 'features', 'certifications'}
            single_fields = {
                'sku', 'sku_family', 'fabric_code', 'spec_qty_short', 'is_iteration', 'is_dachene_product', 'source_sku', 'version_no',
                'contents_desc_en',
                'finished_length_in', 'finished_width_in', 'finished_height_in', 'net_weight_lbs',
                'package_length_in', 'package_width_in', 'package_height_in', 'gross_weight_lbs',
                'cost_usd', 'carton_qty', 'package_size_class', 'last_mile_avg_freight_usd'
            }
            
            # 构建列映射，支持中文标签或字段代码
            header_map = {}
            active_multi_base = None
            active_multi_index = 0
            for idx, h in enumerate(headers):
                h_stripped = str(h).strip() if h is not None else ''
                if not h_stripped:
                    if active_multi_base:
                        active_multi_index += 1
                        header_map[f'{active_multi_base}_{active_multi_index}'] = idx
                    continue

                field_code = label_to_code.get(h_stripped, h_stripped)
                base_field = field_code.rsplit('_', 1)[0] if '_' in field_code and field_code[-1].isdigit() else field_code

                if base_field in multi_base_fields:
                    active_multi_base = base_field
                    if '_' in field_code and field_code[-1].isdigit():
                        active_multi_index = int(field_code.rsplit('_', 1)[1])
                    else:
                        active_multi_index = 1
                    header_map[f'{base_field}_{active_multi_index}'] = idx
                    if f'{base_field}_1' not in header_map:
                        header_map[f'{base_field}_1'] = idx
                    if base_field not in header_map:
                        header_map[base_field] = idx
                elif base_field in single_fields:
                    active_multi_base = None
                    active_multi_index = 0
                    header_map[base_field] = idx
                else:
                    active_multi_base = None
                    active_multi_index = 0
                    header_map[field_code] = idx

            def get_cell(row, key):
                idx = header_map.get(key)
                if idx is None:
                    return None
                return row[idx].value

            def parse_bool(raw):
                if raw is None:
                    return 0
                text = str(raw).strip().lower()
                if text in ('1', 'true', 'yes', 'y', '是', '对', 'on', '是否迭代款'):
                    return 1
                return 0
            
            # 支持多列的多选字段收集函数（动态识别所有 _1, _2, _3... 等列）
            def collect_multi_select_values(row, field_base_name, options_map):
                """
                收集某个多选字段的所有列中的值（动态识别 _1, _2, _3, ... 等）
                """
                values = []
                # 尝试所有可能的后缀（1-20）
                for i in range(1, 21):
                    col_name = f"{field_base_name}_{i}"
                    cell_value = (get_cell(row, col_name) or '').strip()
                    if cell_value and options_map:
                        val_id = options_map.get(cell_value)
                        if val_id:
                            values.append(val_id)
                    elif not cell_value:
                        # 如果某列为空，后续列也可能有值，继续检查
                        continue
                return values

            self._ensure_order_product_tables()
            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id, sku_family FROM product_families")
                    sku_map = {row['sku_family']: row['id'] for row in cur.fetchall()}
                    cur.execute("SELECT id, fabric_code FROM fabric_materials")
                    fabric_map = {row['fabric_code']: row['id'] for row in cur.fetchall()}
                    cur.execute(
                        """
                        SELECT m.id, m.name, mt.name AS type_name
                        FROM materials m
                        JOIN material_types mt ON m.material_type_id = mt.id
                        """
                    )
                    material_rows = cur.fetchall()
                    filling_map = {row['name']: row['id'] for row in material_rows if row['type_name'] == '填充'}
                    frame_map = {row['name']: row['id'] for row in material_rows if row['type_name'] == '框架'}
                    cur.execute("SELECT id, name FROM features")
                    feature_map = {row['name']: row['id'] for row in cur.fetchall()}
                    cur.execute("SELECT id, name FROM certifications")
                    cert_map = {row['name']: row['id'] for row in cur.fetchall()}
                    cur.execute(
                        """
                        SELECT id, sku, sku_family_id, version_no, fabric_id, spec_qty_short,
                               contents_desc_en,
                               is_iteration, is_dachene_product, source_order_product_id,
                               finished_length_in, finished_width_in, finished_height_in,
                               net_weight_lbs, package_length_in, package_width_in, package_height_in,
                               gross_weight_lbs, cost_usd, carton_qty, package_size_class, last_mile_avg_freight_usd
                        FROM order_products
                        """
                    )
                    order_rows = cur.fetchall() or []
                    order_map = {row['sku']: row['id'] for row in order_rows}
                    order_row_map = {row['id']: row for row in order_rows}

                    cur.execute("SELECT order_product_id, material_id FROM order_product_materials")
                    material_rows = cur.fetchall() or []
                    material_map = {}
                    for mr in material_rows:
                        material_map.setdefault(mr['order_product_id'], set()).add(mr['material_id'])

                    cur.execute("SELECT order_product_id, feature_id FROM order_product_features")
                    feature_rows = cur.fetchall() or []
                    feature_rel_map = {}
                    for fr in feature_rows:
                        feature_rel_map.setdefault(fr['order_product_id'], set()).add(fr['feature_id'])

                    cur.execute("SELECT order_product_id, certification_id FROM order_product_certifications")
                    cert_rows = cur.fetchall() or []
                    cert_rel_map = {}
                    for cr in cert_rows:
                        cert_rel_map.setdefault(cr['order_product_id'], set()).add(cr['certification_id'])

                def _norm(v):
                    if v is None:
                        return None
                    try:
                        if isinstance(v, float):
                            return round(v, 4)
                        return round(float(v), 4)
                    except Exception:
                        return str(v).strip()

                created = 0
                updated = 0
                unchanged = 0
                relation_added = 0
                relation_deleted = 0
                total_rows = 0
                errors = []
                preview_temp_id = -1
                tx_enabled = False
                batch_write_count = 0
                batch_size = 200
                if not preview_mode:
                    try:
                        conn.autocommit(False)
                        tx_enabled = True
                    except Exception:
                        tx_enabled = False
                data_start_row = 4 if header_row_idx == 2 else 2
                
                for row_idx in range(data_start_row, ws.max_row + 1):
                    row = ws[row_idx]
                    row_values = [cell.value for cell in row]
                    if not any(v is not None and str(v).strip() for v in row_values):
                        continue
                    total_rows += 1

                    sku = (get_cell(row, 'sku') or '').strip()
                    sku_family = (get_cell(row, 'sku_family') or '').strip()
                    version_no = (get_cell(row, 'version_no') or '').strip()
                    fabric_code = (get_cell(row, 'fabric_code') or '').strip()
                    spec_qty_short = (get_cell(row, 'spec_qty_short') or '').strip()
                    contents_desc_en = (get_cell(row, 'contents_desc_en') or '').strip()
                    is_iteration = parse_bool(get_cell(row, 'is_iteration'))
                    source_sku = (get_cell(row, 'source_sku') or '').strip()
                    is_dachene_product = parse_bool(get_cell(row, 'is_dachene_product'))

                    if not sku or not sku_family or not fabric_code:
                        errors.append({'row': row_idx, 'error': 'Missing required fields'})
                        continue
                    if is_iteration and not version_no:
                        errors.append({'row': row_idx, 'error': 'Missing version for iteration'})
                        continue

                    sku_family_id = sku_map.get(sku_family)
                    fabric_id = fabric_map.get(fabric_code)
                    if not sku_family_id or not fabric_id:
                        errors.append({'row': row_idx, 'error': 'Invalid sku_family or fabric_code'})
                        continue

                    source_order_product_id = None
                    if is_iteration:
                        if not source_sku or source_sku not in order_map:
                            errors.append({'row': row_idx, 'error': 'Invalid source SKU'})
                            continue
                        source_order_product_id = order_map.get(source_sku)

                    payload = {
                        'sku': sku,
                        'sku_family_id': sku_family_id,
                        'version_no': version_no,
                        'fabric_id': fabric_id,
                        'spec_qty_short': spec_qty_short,
                        'contents_desc_en': contents_desc_en or None,
                        'is_iteration': is_iteration,
                        'is_dachene_product': is_dachene_product,
                        'source_order_product_id': source_order_product_id,
                        'finished_length_in': self._parse_float(get_cell(row, 'finished_length_in')),
                        'finished_width_in': self._parse_float(get_cell(row, 'finished_width_in')),
                        'finished_height_in': self._parse_float(get_cell(row, 'finished_height_in')),
                        'net_weight_lbs': self._parse_float(get_cell(row, 'net_weight_lbs')),
                        'package_length_in': self._parse_float(get_cell(row, 'package_length_in')),
                        'package_width_in': self._parse_float(get_cell(row, 'package_width_in')),
                        'package_height_in': self._parse_float(get_cell(row, 'package_height_in')),
                        'gross_weight_lbs': self._parse_float(get_cell(row, 'gross_weight_lbs')),
                        'cost_usd': self._parse_float(get_cell(row, 'cost_usd')),
                        'carton_qty': self._parse_int(get_cell(row, 'carton_qty')),
                        'package_size_class': (get_cell(row, 'package_size_class') or '').strip() or None,
                        'last_mile_avg_freight_usd': self._parse_float(get_cell(row, 'last_mile_avg_freight_usd'))
                    }

                    # 支持动态多列多选格式 (field_1, field_2, field_3, ...)
                    filling_ids = collect_multi_select_values(row, 'filling_materials', filling_map)
                    frame_ids = collect_multi_select_values(row, 'frame_materials', frame_map)
                    feature_ids = collect_multi_select_values(row, 'features', feature_map)
                    cert_ids = collect_multi_select_values(row, 'certifications', cert_map)

                    dedup_material_ids = set((filling_ids or []) + (frame_ids or []))
                    dedup_feature_ids = set(feature_ids or [])
                    dedup_cert_ids = set(cert_ids or [])

                    target_id = order_map.get(sku)
                    old_material_ids = material_map.get(target_id, set()) if target_id else set()
                    old_feature_ids = feature_rel_map.get(target_id, set()) if target_id else set()
                    old_cert_ids = cert_rel_map.get(target_id, set()) if target_id else set()

                    relation_added += len(dedup_material_ids - old_material_ids)
                    relation_added += len(dedup_feature_ids - old_feature_ids)
                    relation_added += len(dedup_cert_ids - old_cert_ids)
                    relation_deleted += len(old_material_ids - dedup_material_ids)
                    relation_deleted += len(old_feature_ids - dedup_feature_ids)
                    relation_deleted += len(old_cert_ids - dedup_cert_ids)

                    payload_keys = [
                        'sku_family_id', 'version_no', 'fabric_id', 'spec_qty_short', 'contents_desc_en', 'is_iteration', 'is_dachene_product', 'source_order_product_id',
                        'finished_length_in', 'finished_width_in', 'finished_height_in', 'net_weight_lbs',
                        'package_length_in', 'package_width_in', 'package_height_in', 'gross_weight_lbs',
                        'cost_usd', 'carton_qty', 'package_size_class', 'last_mile_avg_freight_usd'
                    ]
                    is_payload_changed = True
                    if target_id and target_id in order_row_map:
                        old_row = order_row_map[target_id]
                        is_payload_changed = any(_norm(payload.get(k)) != _norm(old_row.get(k)) for k in payload_keys)
                    is_relation_changed = (dedup_material_ids != old_material_ids) or (dedup_feature_ids != old_feature_ids) or (dedup_cert_ids != old_cert_ids)

                    if target_id and (not is_payload_changed) and (not is_relation_changed):
                        unchanged += 1
                        continue

                    if preview_mode:
                        if target_id:
                            updated += 1
                            order_row_map[target_id] = {**payload, 'id': target_id}
                            material_map[target_id] = dedup_material_ids
                            feature_rel_map[target_id] = dedup_feature_ids
                            cert_rel_map[target_id] = dedup_cert_ids
                        else:
                            created += 1
                            target_id = preview_temp_id
                            preview_temp_id -= 1
                            order_map[sku] = target_id
                            order_row_map[target_id] = {**payload, 'id': target_id}
                            material_map[target_id] = dedup_material_ids
                            feature_rel_map[target_id] = dedup_feature_ids
                            cert_rel_map[target_id] = dedup_cert_ids
                        continue

                    try:
                        with conn.cursor() as cur:
                            if target_id:
                                cur.execute(
                                    """
                                    UPDATE order_products
                                    SET sku_family_id=%(sku_family_id)s,
                                        version_no=%(version_no)s,
                                        fabric_id=%(fabric_id)s,
                                        spec_qty_short=%(spec_qty_short)s,
                                        contents_desc_en=%(contents_desc_en)s,
                                        is_iteration=%(is_iteration)s,
                                        is_dachene_product=%(is_dachene_product)s,
                                        source_order_product_id=%(source_order_product_id)s,
                                        finished_length_in=%(finished_length_in)s,
                                        finished_width_in=%(finished_width_in)s,
                                        finished_height_in=%(finished_height_in)s,
                                        net_weight_lbs=%(net_weight_lbs)s,
                                        package_length_in=%(package_length_in)s,
                                        package_width_in=%(package_width_in)s,
                                        package_height_in=%(package_height_in)s,
                                        gross_weight_lbs=%(gross_weight_lbs)s,
                                        cost_usd=%(cost_usd)s,
                                        carton_qty=%(carton_qty)s,
                                        package_size_class=%(package_size_class)s,
                                        last_mile_avg_freight_usd=%(last_mile_avg_freight_usd)s
                                    WHERE id=%(id)s
                                    """,
                                    {**payload, 'id': target_id}
                                )
                                new_id = target_id
                            else:
                                cur.execute(
                                    """
                                    INSERT INTO order_products (
                                        sku, sku_family_id, version_no, fabric_id, spec_qty_short, contents_desc_en,
                                        is_iteration, is_dachene_product, source_order_product_id,
                                        finished_length_in, finished_width_in, finished_height_in,
                                        net_weight_lbs, package_length_in, package_width_in, package_height_in,
                                        gross_weight_lbs, cost_usd, carton_qty, package_size_class, last_mile_avg_freight_usd
                                    ) VALUES (
                                        %(sku)s, %(sku_family_id)s, %(version_no)s, %(fabric_id)s, %(spec_qty_short)s, %(contents_desc_en)s,
                                        %(is_iteration)s, %(is_dachene_product)s, %(source_order_product_id)s,
                                        %(finished_length_in)s, %(finished_width_in)s, %(finished_height_in)s,
                                        %(net_weight_lbs)s, %(package_length_in)s, %(package_width_in)s, %(package_height_in)s,
                                        %(gross_weight_lbs)s, %(cost_usd)s, %(carton_qty)s, %(package_size_class)s, %(last_mile_avg_freight_usd)s
                                    )
                                    """,
                                    payload
                                )
                                new_id = cur.lastrowid
                        if (not target_id) or is_relation_changed:
                            self._replace_order_product_material_ids(conn, new_id, filling_ids, frame_ids)
                            self._replace_order_product_feature_ids(conn, new_id, feature_ids)
                            self._replace_order_product_certification_ids(conn, new_id, cert_ids)

                        material_map[new_id] = dedup_material_ids
                        feature_rel_map[new_id] = dedup_feature_ids
                        cert_rel_map[new_id] = dedup_cert_ids
                        order_row_map[new_id] = {**payload, 'id': new_id}
                        if target_id:
                            updated += 1
                        else:
                            created += 1
                            order_map[sku] = new_id

                        if tx_enabled:
                            batch_write_count += 1
                            if batch_write_count >= batch_size:
                                conn.commit()
                                batch_write_count = 0
                    except Exception as e:
                        errors.append({'row': row_idx, 'error': str(e)})

                if tx_enabled:
                    if batch_write_count > 0:
                        conn.commit()
                    conn.autocommit(True)

            return self.send_json({
                'status': 'success',
                'preview': 1 if preview_mode else 0,
                'total_rows': total_rows,
                'created': created,
                'updated': updated,
                'unchanged': unchanged,
                'relation_added': relation_added,
                'relation_deleted': relation_deleted,
                'errors': errors
            }, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)








    def handle_amazon_ad_keyword_api(self, environ, method, start_response):
        """Amazon 关键词管理 API"""
        try:
            self._ensure_amazon_keyword_tables()
            query_params = parse_qs(environ.get('QUERY_STRING', ''))

            if method == 'GET':
                action = (query_params.get('action', ['list'])[0] or 'list').strip().lower()
                category_id = self._parse_int(query_params.get('category_id', [''])[0])
                if not category_id:
                    return self.send_json({'status': 'error', 'message': 'Missing category_id'}, start_response)

                if action == 'options':
                    with self._get_db_connection() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                """
                                SELECT id, tag_name
                                FROM amazon_keyword_tags
                                WHERE category_id=%s
                                ORDER BY tag_name ASC
                                """,
                                (category_id,)
                            )
                            tags = cur.fetchall() or []

                            cur.execute(
                                """
                                SELECT pf.id, pf.sku_family
                                FROM product_families pf
                                JOIN product_categories pc ON pc.category_cn = pf.category
                                WHERE pc.id=%s
                                ORDER BY pf.sku_family ASC
                                """,
                                (category_id,)
                            )
                            skus = cur.fetchall() or []
                    return self.send_json({'status': 'success', 'tags': tags, 'sku_families': skus}, start_response)

                keyword = (query_params.get('q', [''])[0] or '').strip()
                tag_filter = (query_params.get('tag', [''])[0] or '').strip()
                current_sku_family_id = self._parse_int(query_params.get('sku_family_id', [''])[0])

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        sql = """
                            SELECT
                                k.id,
                                k.category_id,
                                k.user_search_term,
                                k.search_rank,
                                k.rank_updated_at,
                                k.previous_search_rank,
                                k.previous_rank_updated_at,
                                k.top_click_asin1,
                                k.top_click_asin1_click_share,
                                k.top_click_asin1_conversion_share,
                                k.top_click_asin2,
                                k.top_click_asin2_click_share,
                                k.top_click_asin2_conversion_share,
                                k.top_click_asin3,
                                k.top_click_asin3_click_share,
                                k.top_click_asin3_conversion_share,
                                k.updated_at,
                                GROUP_CONCAT(DISTINCT t.id ORDER BY t.id) AS tag_ids,
                                GROUP_CONCAT(DISTINCT t.tag_name ORDER BY t.tag_name SEPARATOR '\n') AS tag_names,
                                GROUP_CONCAT(DISTINCT r.sku_family_id ORDER BY r.sku_family_id) AS related_sku_family_ids,
                                MAX(CASE WHEN r.sku_family_id=%s THEN r.relevance_score END) AS current_relevance_score
                            FROM amazon_keywords k
                            LEFT JOIN amazon_keyword_tag_rel tr ON tr.keyword_id = k.id
                            LEFT JOIN amazon_keyword_tags t ON t.id = tr.tag_id
                            LEFT JOIN amazon_keyword_sku_rel r ON r.keyword_id = k.id
                        """
                        filters = ["k.category_id=%s"]
                        params = [current_sku_family_id or 0, category_id]
                        if keyword:
                            filters.append("k.user_search_term LIKE %s")
                            params.append(f"%{keyword}%")
                        if tag_filter:
                            filters.append("t.tag_name=%s")
                            params.append(tag_filter)
                        where_sql = ' WHERE ' + ' AND '.join(filters)
                        cur.execute(sql + where_sql + " GROUP BY k.id ORDER BY k.updated_at DESC, k.id DESC", params)
                        rows = cur.fetchall() or []

                items = []
                relevance_label_map = {
                    3: '高相关精准词',
                    2: '中相关普通词',
                    1: '低相关大词',
                    4: '变体相关',
                    0: '不相关',
                }
                for row in rows:
                    raw_tag_ids = (row.get('tag_ids') or '').strip()
                    raw_tag_names = row.get('tag_names') or ''
                    raw_sku_ids = (row.get('related_sku_family_ids') or '').strip()
                    tag_ids = [int(v) for v in raw_tag_ids.split(',') if str(v).strip()] if raw_tag_ids else []
                    tag_names = [v for v in str(raw_tag_names).split('\n') if str(v).strip()] if raw_tag_names else []
                    related_sku_ids = [int(v) for v in raw_sku_ids.split(',') if str(v).strip()] if raw_sku_ids else []
                    current_relevance_score = row.get('current_relevance_score')
                    if current_relevance_score is not None:
                        try:
                            current_relevance_score = int(current_relevance_score)
                        except Exception:
                            current_relevance_score = None
                    row['tag_ids'] = tag_ids
                    row['tag_names'] = tag_names
                    row['related_sku_family_ids'] = related_sku_ids
                    row['current_relevance_score'] = current_relevance_score
                    row['current_relevance_level'] = relevance_label_map.get(current_relevance_score)
                    row['is_related_current_sku'] = 1 if (current_relevance_score is not None and current_relevance_score > 0) else 0
                    items.append(row)

                return self.send_json({'status': 'success', 'items': items}, start_response)

            if method == 'POST':
                data = self._read_json_body(environ)
                category_id = self._parse_int(data.get('category_id'))
                user_search_term = (data.get('user_search_term') or '').strip()
                if not category_id or not user_search_term:
                    return self.send_json({'status': 'error', 'message': 'Missing category_id or user_search_term'}, start_response)

                search_rank = self._parse_int(data.get('search_rank'))
                now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                tag_names = self._normalize_keyword_tag_names(data.get('tag_names'))
                sku_family_ids = self._normalize_keyword_sku_ids(data.get('related_sku_family_ids'))

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO amazon_keywords (
                                category_id, user_search_term,
                                search_rank, rank_updated_at,
                                top_click_asin1, top_click_asin1_click_share, top_click_asin1_conversion_share,
                                top_click_asin2, top_click_asin2_click_share, top_click_asin2_conversion_share,
                                top_click_asin3, top_click_asin3_click_share, top_click_asin3_conversion_share
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                category_id, user_search_term,
                                search_rank, now_text if search_rank is not None else None,
                                (data.get('top_click_asin1') or '').strip() or None,
                                (data.get('top_click_asin1_click_share') or '').strip() or None,
                                (data.get('top_click_asin1_conversion_share') or '').strip() or None,
                                (data.get('top_click_asin2') or '').strip() or None,
                                (data.get('top_click_asin2_click_share') or '').strip() or None,
                                (data.get('top_click_asin2_conversion_share') or '').strip() or None,
                                (data.get('top_click_asin3') or '').strip() or None,
                                (data.get('top_click_asin3_click_share') or '').strip() or None,
                                (data.get('top_click_asin3_conversion_share') or '').strip() or None,
                            )
                        )
                        new_id = cur.lastrowid
                    self._replace_keyword_tags(conn, new_id, category_id, tag_names)
                    self._replace_keyword_sku_relevance(conn, new_id, sku_family_ids)

                return self.send_json({'status': 'success', 'id': new_id}, start_response)

            if method == 'PUT':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT * FROM amazon_keywords WHERE id=%s", (item_id,))
                        current = cur.fetchone()
                        if not current:
                            return self.send_json({'status': 'error', 'message': 'Not found'}, start_response)

                    category_id = self._parse_int(data.get('category_id')) or int(current.get('category_id'))
                    user_search_term = (data.get('user_search_term') or current.get('user_search_term') or '').strip()
                    if not category_id or not user_search_term:
                        return self.send_json({'status': 'error', 'message': 'Missing category_id or user_search_term'}, start_response)

                    new_rank = self._parse_int(data.get('search_rank')) if ('search_rank' in data) else current.get('search_rank')
                    old_rank = current.get('search_rank')
                    old_rank_updated = current.get('rank_updated_at')
                    previous_rank = current.get('previous_search_rank')
                    previous_rank_updated = current.get('previous_rank_updated_at')
                    rank_updated_at = old_rank_updated
                    if ('search_rank' in data) and (new_rank != old_rank):
                        previous_rank = old_rank
                        previous_rank_updated = old_rank_updated
                        rank_updated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            UPDATE amazon_keywords
                            SET category_id=%s,
                                user_search_term=%s,
                                search_rank=%s,
                                rank_updated_at=%s,
                                previous_search_rank=%s,
                                previous_rank_updated_at=%s,
                                top_click_asin1=%s,
                                top_click_asin1_click_share=%s,
                                top_click_asin1_conversion_share=%s,
                                top_click_asin2=%s,
                                top_click_asin2_click_share=%s,
                                top_click_asin2_conversion_share=%s,
                                top_click_asin3=%s,
                                top_click_asin3_click_share=%s,
                                top_click_asin3_conversion_share=%s
                            WHERE id=%s
                            """,
                            (
                                category_id,
                                user_search_term,
                                new_rank,
                                rank_updated_at,
                                previous_rank,
                                previous_rank_updated,
                                (data.get('top_click_asin1') if 'top_click_asin1' in data else current.get('top_click_asin1')) or None,
                                (data.get('top_click_asin1_click_share') if 'top_click_asin1_click_share' in data else current.get('top_click_asin1_click_share')) or None,
                                (data.get('top_click_asin1_conversion_share') if 'top_click_asin1_conversion_share' in data else current.get('top_click_asin1_conversion_share')) or None,
                                (data.get('top_click_asin2') if 'top_click_asin2' in data else current.get('top_click_asin2')) or None,
                                (data.get('top_click_asin2_click_share') if 'top_click_asin2_click_share' in data else current.get('top_click_asin2_click_share')) or None,
                                (data.get('top_click_asin2_conversion_share') if 'top_click_asin2_conversion_share' in data else current.get('top_click_asin2_conversion_share')) or None,
                                (data.get('top_click_asin3') if 'top_click_asin3' in data else current.get('top_click_asin3')) or None,
                                (data.get('top_click_asin3_click_share') if 'top_click_asin3_click_share' in data else current.get('top_click_asin3_click_share')) or None,
                                (data.get('top_click_asin3_conversion_share') if 'top_click_asin3_conversion_share' in data else current.get('top_click_asin3_conversion_share')) or None,
                                item_id
                            )
                        )

                    if 'tag_names' in data:
                        self._replace_keyword_tags(conn, item_id, category_id, self._normalize_keyword_tag_names(data.get('tag_names')))
                    if 'related_sku_family_ids' in data:
                        self._replace_keyword_sku_relevance(conn, item_id, self._normalize_keyword_sku_ids(data.get('related_sku_family_ids')))

                return self.send_json({'status': 'success'}, start_response)

            if method == 'PATCH':
                data = self._read_json_body(environ)
                category_id = self._parse_int(data.get('category_id'))
                keyword_ids = [self._parse_int(v) for v in (data.get('keyword_ids') or [])]
                keyword_ids = [v for v in keyword_ids if v]
                if not category_id or not keyword_ids:
                    return self.send_json({'status': 'error', 'message': 'Missing category_id or keyword_ids'}, start_response)

                tag_names = self._normalize_keyword_tag_names(data.get('tag_names')) if ('tag_names' in data) else None
                current_sku_family_id = self._parse_int(data.get('sku_family_id'))
                set_related = data.get('set_related') if ('set_related' in data) else None
                relevance_level = (data.get('relevance_level') or '').strip()
                relevance_score_map = {
                    '高相关精准词': 3,
                    '中相关普通词': 2,
                    '低相关大词': 1,
                    '变体相关': 4,
                    '不相关': 0,
                }

                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        placeholders = ','.join(['%s'] * len(keyword_ids))
                        cur.execute(
                            f"SELECT id FROM amazon_keywords WHERE category_id=%s AND id IN ({placeholders})",
                            [category_id] + keyword_ids
                        )
                        existing_ids = [int(row.get('id')) for row in (cur.fetchall() or []) if row.get('id')]
                    if not existing_ids:
                        return self.send_json({'status': 'error', 'message': 'No valid keywords in this category'}, start_response)

                    if tag_names is not None:
                        tag_ids = self._ensure_keyword_tags(conn, category_id, tag_names)
                        if tag_ids:
                            with conn.cursor() as cur:
                                rel_values = []
                                for keyword_id in existing_ids:
                                    for tag_id in tag_ids:
                                        rel_values.append((keyword_id, tag_id))
                                cur.executemany(
                                    "INSERT IGNORE INTO amazon_keyword_tag_rel (keyword_id, tag_id) VALUES (%s, %s)",
                                    rel_values
                                )

                    if current_sku_family_id and relevance_level in relevance_score_map:
                        score = relevance_score_map.get(relevance_level)
                        with conn.cursor() as cur:
                            cur.executemany(
                                """
                                INSERT INTO amazon_keyword_sku_rel (keyword_id, sku_family_id, relevance_score)
                                VALUES (%s, %s, %s)
                                ON DUPLICATE KEY UPDATE relevance_score=VALUES(relevance_score)
                                """,
                                [(keyword_id, current_sku_family_id, score) for keyword_id in existing_ids]
                            )
                    elif current_sku_family_id and set_related is not None:
                        with conn.cursor() as cur:
                            if bool(set_related):
                                cur.executemany(
                                    """
                                    INSERT INTO amazon_keyword_sku_rel (keyword_id, sku_family_id, relevance_score)
                                    VALUES (%s, %s, 1)
                                    ON DUPLICATE KEY UPDATE relevance_score=VALUES(relevance_score)
                                    """,
                                    [(keyword_id, current_sku_family_id) for keyword_id in existing_ids]
                                )
                            else:
                                cur.executemany(
                                    "DELETE FROM amazon_keyword_sku_rel WHERE keyword_id=%s AND sku_family_id=%s",
                                    [(keyword_id, current_sku_family_id) for keyword_id in existing_ids]
                                )

                return self.send_json({'status': 'success', 'updated_count': len(existing_ids)}, start_response)

            if method == 'DELETE':
                data = self._read_json_body(environ)
                item_id = self._parse_int(data.get('id'))
                if not item_id:
                    return self.send_json({'status': 'error', 'message': 'Missing id'}, start_response)
                with self._get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM amazon_keywords WHERE id=%s", (item_id,))
                return self.send_json({'status': 'success'}, start_response)

            return self.send_error(405, 'Method not allowed', start_response)
        except RuntimeError as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)
        except Exception as e:
            if pymysql and isinstance(e, pymysql.err.IntegrityError):
                return self.send_json({'status': 'error', 'message': '搜索词已存在或关联数据无效'}, start_response)
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_keyword_template_api(self, environ, method, start_response):
        """Amazon 关键词模板下载"""
        try:
            if method != 'GET':
                return self.send_error(405, 'Method not allowed', start_response)
            if Workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            category_id = self._parse_int(parse_qs(environ.get('QUERY_STRING', '')).get('category_id', [''])[0])
            if not category_id:
                return self.send_json({'status': 'error', 'message': 'Missing category_id'}, start_response)

            from openpyxl.styles import PatternFill, Font, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = 'amazon_keywords'

            headers = [
                '用户搜索词*', '搜索词排名',
                '最高点击ASIN1', '最高点击ASIN1点击占比', '最高点击ASIN1转化占比',
                '最高点击ASIN2', '最高点击ASIN2点击占比', '最高点击ASIN2转化占比',
                '最高点击ASIN3', '最高点击ASIN3点击占比', '最高点击ASIN3转化占比',
                '标签(换行分隔，自动创建新标签)',
                '关联货号(换行分隔，仅当前品类)'
            ]
            ws.append(headers)
            ws.append([
                'recliner sofa', 3,
                'B0XXXXXX01', '22.1%', '12.5%',
                'B0XXXXXX02', '16.7%', '9.8%',
                'B0XXXXXX03', '13.3%', '7.1%',
                '高转化\n核心词',
                'MS01\nMS02'
            ])

            for cell in ws[1]:
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.font = Font(bold=True, color='2A2420')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in ws[2]:
                cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                cell.font = Font(italic=True, color='888888')

            widths = [24, 12, 16, 16, 16, 16, 16, 16, 16, 16, 16, 28, 24]
            for idx, width in enumerate(widths, start=1):
                col = chr(64 + idx) if idx <= 26 else 'A'
                ws.column_dimensions[col].width = width
            ws.freeze_panes = 'A3'

            return self._send_excel_workbook(wb, 'amazon_keyword_template.xlsx', start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def handle_amazon_ad_keyword_import_api(self, environ, method, start_response):
        """Amazon 关键词批量导入"""
        try:
            if method != 'POST':
                return self.send_error(405, 'Method not allowed', start_response)
            if load_workbook is None:
                return self.send_json({'status': 'error', 'message': f'openpyxl not available: {_openpyxl_import_error}'}, start_response)

            query_params = parse_qs(environ.get('QUERY_STRING', ''))
            category_id = self._parse_int(query_params.get('category_id', [''])[0])
            if not category_id:
                return self.send_json({'status': 'error', 'message': 'Missing category_id'}, start_response)

            content_type = environ.get('CONTENT_TYPE', '')
            if 'multipart/form-data' not in content_type:
                return self.send_json({'status': 'error', 'message': 'Invalid content type'}, start_response)

            content_length = int(environ.get('CONTENT_LENGTH', 0) or 0)
            raw_body = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            env_copy = dict(environ)
            env_copy['CONTENT_LENGTH'] = str(len(raw_body))
            form = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=env_copy, keep_blank_values=True)
            file_item = form['file'] if 'file' in form else None
            if file_item is None or getattr(file_item, 'file', None) is None:
                return self.send_json({'status': 'error', 'message': 'Missing file'}, start_response)
            file_bytes = file_item.file.read() or b''
            if not file_bytes:
                return self.send_json({'status': 'error', 'message': 'Empty file'}, start_response)

            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            headers = [str(cell.value or '').strip() for cell in ws[1]]
            header_map = {name: idx for idx, name in enumerate(headers)}

            def get_cell(row, name):
                idx = header_map.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx].value

            self._ensure_amazon_keyword_tables()
            created = 0
            updated = 0
            errors = []

            with self._get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT pf.id, pf.sku_family
                        FROM product_families pf
                        JOIN product_categories pc ON pc.category_cn = pf.category
                        WHERE pc.id=%s
                        """,
                        (category_id,)
                    )
                    sku_map = {str(row.get('sku_family') or '').strip(): int(row.get('id')) for row in (cur.fetchall() or []) if row.get('id')}

                for row_idx in range(2, ws.max_row + 1):
                    row = ws[row_idx]
                    if not any(cell.value is not None and str(cell.value).strip() for cell in row):
                        continue

                    try:
                        term = (get_cell(row, '用户搜索词*') or '').strip()
                        if not term:
                            errors.append({'row': row_idx, 'error': '用户搜索词不能为空'})
                            continue

                        rank = self._parse_int(get_cell(row, '搜索词排名'))
                        tag_names = self._normalize_keyword_tag_names(get_cell(row, '标签(换行分隔，自动创建新标签)'))
                        sku_text = get_cell(row, '关联货号(换行分隔，仅当前品类)')
                        sku_codes = self._normalize_keyword_tag_names(sku_text)
                        sku_ids = []
                        for sku_code in sku_codes:
                            sku_id = sku_map.get(sku_code)
                            if not sku_id:
                                raise ValueError(f'未知货号: {sku_code}')
                            sku_ids.append(sku_id)
                        sku_ids = sorted(set(sku_ids))

                        with conn.cursor() as cur:
                            cur.execute("SELECT * FROM amazon_keywords WHERE user_search_term=%s", (term,))
                            existing = cur.fetchone()

                        now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        if existing:
                            old_rank = existing.get('search_rank')
                            old_rank_updated = existing.get('rank_updated_at')
                            previous_rank = existing.get('previous_search_rank')
                            previous_rank_updated = existing.get('previous_rank_updated_at')
                            rank_updated_at = old_rank_updated
                            if rank != old_rank:
                                previous_rank = old_rank
                                previous_rank_updated = old_rank_updated
                                rank_updated_at = now_text

                            with conn.cursor() as cur:
                                cur.execute(
                                    """
                                    UPDATE amazon_keywords
                                    SET category_id=%s,
                                        search_rank=%s,
                                        rank_updated_at=%s,
                                        previous_search_rank=%s,
                                        previous_rank_updated_at=%s,
                                        top_click_asin1=%s,
                                        top_click_asin1_click_share=%s,
                                        top_click_asin1_conversion_share=%s,
                                        top_click_asin2=%s,
                                        top_click_asin2_click_share=%s,
                                        top_click_asin2_conversion_share=%s,
                                        top_click_asin3=%s,
                                        top_click_asin3_click_share=%s,
                                        top_click_asin3_conversion_share=%s
                                    WHERE id=%s
                                    """,
                                    (
                                        category_id,
                                        rank,
                                        rank_updated_at,
                                        previous_rank,
                                        previous_rank_updated,
                                        (get_cell(row, '最高点击ASIN1') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN1点击占比') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN1转化占比') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN2') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN2点击占比') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN2转化占比') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN3') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN3点击占比') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN3转化占比') or '').strip() or None,
                                        existing.get('id')
                                    )
                                )
                            keyword_id = int(existing.get('id'))
                            updated += 1
                        else:
                            with conn.cursor() as cur:
                                cur.execute(
                                    """
                                    INSERT INTO amazon_keywords (
                                        category_id, user_search_term,
                                        search_rank, rank_updated_at,
                                        top_click_asin1, top_click_asin1_click_share, top_click_asin1_conversion_share,
                                        top_click_asin2, top_click_asin2_click_share, top_click_asin2_conversion_share,
                                        top_click_asin3, top_click_asin3_click_share, top_click_asin3_conversion_share
                                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                    """,
                                    (
                                        category_id,
                                        term,
                                        rank,
                                        now_text if rank is not None else None,
                                        (get_cell(row, '最高点击ASIN1') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN1点击占比') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN1转化占比') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN2') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN2点击占比') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN2转化占比') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN3') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN3点击占比') or '').strip() or None,
                                        (get_cell(row, '最高点击ASIN3转化占比') or '').strip() or None,
                                    )
                                )
                                keyword_id = cur.lastrowid
                            created += 1

                        self._replace_keyword_tags(conn, keyword_id, category_id, tag_names)
                        self._replace_keyword_sku_relevance(conn, keyword_id, sku_ids)
                    except Exception as row_error:
                        errors.append({'row': row_idx, 'error': str(row_error)})

            return self.send_json({'status': 'success', 'created': created, 'updated': updated, 'errors': errors}, start_response)
        except Exception as e:
            return self.send_json({'status': 'error', 'message': str(e)}, start_response)

    def serve_file(self, filepath, content_type, start_response):
        """提供文件"""
        try:
            full_path = os.path.join(self.base_path, filepath)
            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()

            content_bytes = content.encode('utf-8', errors='surrogatepass')
            start_response('200 OK', [
                ('Content-Type', content_type + '; charset=utf-8'),
                ('Content-Length', str(len(content_bytes)))
            ])
            return [content_bytes]
        except FileNotFoundError:
            return self.send_error(404, 'File Not Found', start_response)
        except Exception as e:
            return self.send_error(500, str(e), start_response)

    def serve_static(self, path, start_response):
        """提供静态文件"""
        try:
            filepath = os.path.join(self.base_path, path.lstrip('/'))

            with open(filepath, 'rb') as f:
                content = f.read()

            content_type, _ = mimetypes.guess_type(filepath)
            if content_type is None:
                content_type = 'application/octet-stream'

            start_response('200 OK', [
                ('Content-Type', content_type),
                ('Content-Length', str(len(content)))
            ])
            return [content]
        except FileNotFoundError:
            return self.send_error(404, 'File Not Found', start_response)
        except Exception as e:
            return self.send_error(500, str(e), start_response)

    def send_json(self, data, start_response):
        """发送 JSON 响应（确保完全ASCII编码）"""
        try:
            # Use UTF-8 with surrogatepass to safely encode any filesystem surrogate characters
            text = json.dumps(data, ensure_ascii=False, default=str)
            response = text.encode('utf-8', errors='surrogatepass')
            start_response('200 OK', [
                ('Content-Type', 'application/json; charset=utf-8'),
                ('Content-Length', str(len(response)))
            ])
            return [response]
        except Exception as e:
            print("JSON encoding error: " + str(e))
            try:
                fallback_text = json.dumps({'status': 'error', 'message': 'encoding error'}, ensure_ascii=False)
                fallback = fallback_text.encode('utf-8', errors='surrogatepass')
            except Exception:
                fallback = b'{"status":"error","message":"encoding error"}'
            start_response('200 OK', [
                ('Content-Type', 'application/json; charset=utf-8'),
                ('Content-Length', str(len(fallback)))
            ])
            return [fallback]
    
    def send_error(self, status_code, message, start_response):
        """发送错误响应"""
        status_text = {
            400: 'Bad Request',
            403: 'Forbidden',
            404: 'Not Found',
            405: 'Method Not Allowed',
            409: 'Conflict',
            500: 'Internal Server Error'
        }.get(status_code, 'Error')
        
        status = f'{status_code} {status_text}'
        
        error_html = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <title>错误 {status_code}</title>
            <meta charset="utf-8">
        </head>
        <body>
            <h1>{status}</h1>
            <p>{message}</p>
        </body>
        </html>
        '''.encode('utf-8', errors='surrogatepass')
        
        start_response(status, [
            ('Content-Type', 'text/html; charset=utf-8'),
            ('Content-Length', str(len(error_html)))
        ])
        return [error_html]

# 运行时强制使用 mixin 版本的方法（过渡期）：
# 先确保拆分代码真正生效，再逐步删除 app.py 里的历史同名实现。
WSGIApp.handle_auth_api = AuthEmployeeMixin.handle_auth_api
WSGIApp.handle_employee_api = AuthEmployeeMixin.handle_employee_api

WSGIApp._dispatch_api_request = RequestRoutingMixin._dispatch_api_request
WSGIApp._validate_api_permission = RequestRoutingMixin._validate_api_permission
WSGIApp._dispatch_page_request = RequestRoutingMixin._dispatch_page_request

WSGIApp.handle_logistics_warehouse_api = LogisticsWarehouseMixin.handle_logistics_warehouse_api
WSGIApp.handle_logistics_warehouse_template_api = LogisticsWarehouseMixin.handle_logistics_warehouse_template_api
WSGIApp.handle_logistics_warehouse_import_api = LogisticsWarehouseMixin.handle_logistics_warehouse_import_api
WSGIApp.handle_logistics_warehouse_inventory_api = LogisticsWarehouseMixin.handle_logistics_warehouse_inventory_api
WSGIApp.handle_logistics_warehouse_inventory_template_api = LogisticsWarehouseMixin.handle_logistics_warehouse_inventory_template_api
WSGIApp.handle_logistics_warehouse_inventory_import_api = LogisticsWarehouseMixin.handle_logistics_warehouse_inventory_import_api
WSGIApp.handle_logistics_warehouse_dashboard_api = LogisticsWarehouseMixin.handle_logistics_warehouse_dashboard_api

WSGIApp.handle_logistics_in_transit_api = LogisticsInTransitMixin.handle_logistics_in_transit_api
WSGIApp.handle_logistics_in_transit_template_api = LogisticsInTransitMixin.handle_logistics_in_transit_template_api
WSGIApp.handle_logistics_in_transit_import_api = LogisticsInTransitMixin.handle_logistics_in_transit_import_api
WSGIApp.handle_logistics_in_transit_doc_upload_api = LogisticsInTransitMixin.handle_logistics_in_transit_doc_upload_api
WSGIApp.handle_logistics_in_transit_doc_files_api = LogisticsInTransitMixin.handle_logistics_in_transit_doc_files_api

WSGIApp.handle_parent_api = SalesProductMixin.handle_parent_api
WSGIApp.handle_sales_product_api = SalesProductMixin.handle_sales_product_api
WSGIApp.handle_sales_product_template_api = SalesProductMixin.handle_sales_product_template_api
WSGIApp.handle_sales_product_import_api = SalesProductMixin.handle_sales_product_import_api

WSGIApp.handle_sales_order_registration_api = SalesManagementMixin.handle_sales_order_registration_api
WSGIApp.handle_sales_order_registration_template_api = SalesManagementMixin.handle_sales_order_registration_template_api
WSGIApp.handle_sales_order_registration_import_api = SalesManagementMixin.handle_sales_order_registration_import_api

WSGIApp._ensure_logistics_tables = LogisticsSchemaMixin._ensure_logistics_tables
WSGIApp._ensure_factory_inventory_tables = LogisticsSchemaMixin._ensure_factory_inventory_tables
WSGIApp._ensure_sales_parent_tables = SalesSchemaMixin._ensure_sales_parent_tables
WSGIApp._ensure_sales_product_tables = SalesSchemaMixin._ensure_sales_product_tables
WSGIApp._ensure_sales_order_registration_tables = SalesSchemaMixin._ensure_sales_order_registration_tables

# WSGI 应用实例 - Web Station 会调用这个
application = WSGIApp()
